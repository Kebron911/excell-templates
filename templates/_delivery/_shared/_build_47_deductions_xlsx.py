"""Build the companion Excel checklist for the 47-deductions hero magnet.

Two tabs:
  - Checklist  — one row per deduction; user marks Captured? + logs $ amount;
                 YTD total + count of captured items roll up at the top.
  - About      — instructions, disclaimer, and the working-draft warning.

Source-of-truth content: deductions_47_data.py (same data feeds the PDF).

Output: templates/_delivery/_shared/47-airbnb-tax-deductions-checklist.xlsx

Re-run any time the data changes:
    python templates/_delivery/_shared/_build_47_deductions_xlsx.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, str(Path(__file__).resolve().parent))
from deductions_47_data import ENTRIES, SECTIONS  # noqa: E402

REPO = Path(__file__).resolve().parent.parent.parent.parent
OUT = REPO / "templates" / "_delivery" / "_shared" / "47-airbnb-tax-deductions-checklist.xlsx"

NAVY = "12304E"
PARCHMENT = "F6EFE2"
PARCHMENT_ALT = "EFE5D0"
GOLD = "C9A24B"
GOLD_SOFT = "F2E5C0"
GRAPHITE = "2B2B2B"
MUTED = "6B7280"
WHITE = "FFFFFF"
INPUT_TINT = "FFF7D6"
GREEN_FILL = "C7EFCF"

FONT_HEAD = "Cormorant Garamond"
FONT_BODY = "Inter"
FONT_MONO = "JetBrains Mono"


def _border_bottom(color: str = "CCCCCC") -> Border:
    return Border(bottom=Side(style="thin", color=color))


def build_checklist(wb: Workbook) -> None:
    ws = wb.create_sheet("Checklist")
    ws.sheet_properties.tabColor = GOLD

    # Column layout:
    # A: # | B: Deduction | C: Section | D: Sched line | E: IRS ref
    # F: Typical $ range | G: Captured? Y/N | H: $ amount captured | I: Notes
    widths = [(1, 5), (2, 44), (3, 30), (4, 22), (5, 22), (6, 22), (7, 14), (8, 16), (9, 36)]
    for col, w in widths:
        ws.column_dimensions[get_column_letter(col)].width = w

    # ----- Brand strip (rows 1-3) -----
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "The STR Ledger  ·  47 Airbnb Tax Deductions Most Hosts Miss  ·  Companion Checklist"
    title.font = Font(name=FONT_HEAD, size=18, bold=True, color=NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title.fill = PatternFill("solid", fgColor=PARCHMENT)

    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = ("Working draft — 2026 edition. Confirm every IRS rate against current-year publications "
                  "before acting. Not tax advice — consult your CPA.")
    sub.font = Font(name=FONT_BODY, size=10, italic=True, color=MUTED)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sub.fill = PatternFill("solid", fgColor=PARCHMENT)

    # ----- Roll-up KPIs (rows 4-5) -----
    ws.row_dimensions[4].height = 26
    captured_label = ws["A4"]
    ws.merge_cells("A4:B4")
    captured_label.value = "CAPTURED THIS YEAR"
    captured_label.font = Font(name=FONT_MONO, size=9, bold=True, color=NAVY)
    captured_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    captured_label.fill = PatternFill("solid", fgColor=GOLD_SOFT)

    captured_count = ws["C4"]
    captured_count.value = '=COUNTIF(G8:G54,"Yes")&" of 47"'
    captured_count.font = Font(name=FONT_MONO, size=14, bold=True, color=NAVY)
    captured_count.alignment = Alignment(horizontal="left", vertical="center")
    captured_count.fill = PatternFill("solid", fgColor=GOLD_SOFT)

    ytd_label = ws["D4"]
    ws.merge_cells("D4:E4")
    ytd_label.value = "TOTAL $ DEDUCTIONS YTD"
    ytd_label.font = Font(name=FONT_MONO, size=9, bold=True, color=NAVY)
    ytd_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ytd_label.fill = PatternFill("solid", fgColor=GOLD_SOFT)

    ytd_value = ws["F4"]
    ytd_value.value = "=SUM(H8:H54)"
    ytd_value.font = Font(name=FONT_MONO, size=14, bold=True, color=GOLD)
    ytd_value.alignment = Alignment(horizontal="right", vertical="center")
    ytd_value.fill = PatternFill("solid", fgColor=GOLD_SOFT)
    ytd_value.number_format = '"$"#,##0'

    flagged_label = ws["G4"]
    ws.merge_cells("G4:H4")
    flagged_label.value = "VERIFY FLAGS REMAINING"
    flagged_label.font = Font(name=FONT_MONO, size=9, bold=True, color=NAVY)
    flagged_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    flagged_label.fill = PatternFill("solid", fgColor=GOLD_SOFT)

    flagged_value = ws["I4"]
    flagged_value.value = '=COUNTIF(C8:C54,"*verify*")'
    flagged_value.font = Font(name=FONT_MONO, size=14, bold=True, color=NAVY)
    flagged_value.alignment = Alignment(horizontal="left", vertical="center")
    flagged_value.fill = PatternFill("solid", fgColor=GOLD_SOFT)

    # spacer row 5
    ws.row_dimensions[5].height = 8

    # ----- Header row (row 7) -----
    headers = [
        ("A", "#"),
        ("B", "Deduction"),
        ("C", "Section"),
        ("D", "Schedule line"),
        ("E", "IRS reference"),
        ("F", "Typical $ range"),
        ("G", "Captured?"),
        ("H", "$ amount YTD"),
        ("I", "Notes"),
    ]
    ws.row_dimensions[7].height = 28
    for col, label in headers:
        c = ws[f"{col}7"]
        c.value = label
        c.font = Font(name=FONT_BODY, size=10, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=NAVY)

    # ----- Entry rows (8-54) -----
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    yes_no_dv.prompt = "Captured this year?"
    yes_no_dv.promptTitle = "Captured?"
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add(f"G8:G{8 + len(ENTRIES) - 1}")

    for idx, entry in enumerate(ENTRIES):
        row = 8 + idx
        ws.row_dimensions[row].height = 30
        # alt-row banding
        bg = PARCHMENT if idx % 2 == 0 else PARCHMENT_ALT
        bg_fill = PatternFill("solid", fgColor=bg)
        body_font = Font(name=FONT_BODY, size=10, color=GRAPHITE)
        bold_navy = Font(name=FONT_BODY, size=10.5, bold=True, color=NAVY)
        mono_small = Font(name=FONT_MONO, size=8.5, color=NAVY)
        mono_meta = Font(name=FONT_MONO, size=8.5, color=GRAPHITE)
        gold_text = Font(name=FONT_MONO, size=8.5, bold=True, color=GOLD)

        cells = {
            "A": (entry["n"], Font(name=FONT_HEAD, size=14, bold=True, color=GOLD),
                   Alignment(horizontal="center", vertical="center")),
            "B": (entry["name"], bold_navy,
                   Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)),
            "C": (entry["section"] + (" — verify" if entry["verify"] else ""),
                   body_font if not entry["verify"] else gold_text,
                   Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)),
            "D": (entry["line"], mono_small,
                   Alignment(horizontal="left", vertical="center", indent=1)),
            "E": (entry["irs_ref"], mono_meta,
                   Alignment(horizontal="left", vertical="center", indent=1)),
            "F": (entry["typical"], gold_text,
                   Alignment(horizontal="left", vertical="center", indent=1)),
            "G": (None, body_font,
                   Alignment(horizontal="center", vertical="center")),
            "H": (None, Font(name=FONT_MONO, size=11, color=NAVY),
                   Alignment(horizontal="right", vertical="center", indent=1)),
            "I": (None, body_font,
                   Alignment(horizontal="left", vertical="center", indent=1)),
        }
        for col, (val, font, align) in cells.items():
            c = ws[f"{col}{row}"]
            c.value = val
            c.font = font
            c.alignment = align
            c.fill = bg_fill
            c.border = _border_bottom()
            if col in ("G", "H", "I"):
                # input cells get the yellow tint
                c.fill = PatternFill("solid", fgColor=INPUT_TINT)
            if col == "H":
                c.number_format = '"$"#,##0'

    # ----- Conditional formatting: green fill on row when Captured = Yes -----
    last_row = 8 + len(ENTRIES) - 1
    green_rule = CellIsRule(
        operator="equal",
        formula=['"Yes"'],
        fill=PatternFill("solid", fgColor=GREEN_FILL),
        font=Font(name=FONT_BODY, size=10, bold=True, color=NAVY),
    )
    ws.conditional_formatting.add(f"G8:G{last_row}", green_rule)

    # ----- Freeze panes + print setup -----
    ws.freeze_panes = "A8"
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_about(wb: Workbook) -> None:
    ws = wb.create_sheet("About", 0)  # leftmost
    ws.sheet_properties.tabColor = NAVY

    for col, w in [(1, 100)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title row
    ws.row_dimensions[1].height = 40
    title = ws["A1"]
    title.value = "47 Airbnb Tax Deductions Most Hosts Miss  ·  Companion Excel Checklist"
    title.font = Font(name=FONT_HEAD, size=22, bold=True, color=NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title.fill = PatternFill("solid", fgColor=PARCHMENT)

    rows = [
        (28, "How this works",
         Font(name=FONT_HEAD, size=16, bold=True, color=NAVY),
         PARCHMENT),
        (60,
         "1. Open the Checklist tab.\n"
         "2. As you encounter or log an expense category from the 47, mark column G \"Yes\".\n"
         "3. Enter the dollar amount in column H. The roll-up at the top updates live.\n"
         "4. Use column I for notes (vendor, IRS publication you confirmed against, etc.).\n"
         "5. In March, hand the populated checklist to your CPA. Watch them ask fewer questions.",
         Font(name=FONT_BODY, size=11, color=GRAPHITE),
         PARCHMENT),
        (24, "Verify flags",
         Font(name=FONT_HEAD, size=16, bold=True, color=NAVY),
         PARCHMENT),
        (60,
         "Some entries are marked with \"— verify\" in the Section column. These are deductions where "
         "the underlying IRS rate, threshold, or rule must be confirmed against current-year IRS "
         "publications before acting. Common examples: bonus depreciation percentage (changes annually), "
         "Section 179 cap, IRS standard mileage rate, QBI safe-harbor specifics.",
         Font(name=FONT_BODY, size=11, color=GRAPHITE),
         PARCHMENT),
        (24, "Disclaimer",
         Font(name=FONT_HEAD, size=16, bold=True, color=NAVY),
         PARCHMENT),
        (90,
         "This checklist is for educational purposes. It is not tax advice. Tax law changes annually; "
         "every cited rate, code section, and threshold must be verified against current-year IRS "
         "publications before you act on it. Consult your CPA for your specific situation. "
         "Canonical references: IRS Pub 527 (Residential Rental Property), Pub 463 (Travel/Gift/Car), "
         "Pub 535 (Business Expenses), Pub 946 (Depreciation), Pub 587 (Home Office), and Schedule E "
         "+ Schedule C instructions.",
         Font(name=FONT_BODY, size=11, italic=True, color=GRAPHITE),
         PARCHMENT_ALT),
        (24, "Pair with the templates",
         Font(name=FONT_HEAD, size=16, bold=True, color=NAVY),
         PARCHMENT),
        (60,
         "Several rows in this checklist are auto-populated by templates from The STR Ledger:\n"
         "  · Mileage (#24) — TAX-001 STR Mileage Log\n"
         "  · Cleaning + maintenance, OTA fees, all 17 Schedule E categories — TAX-002 Single-Property P&L Tracker\n"
         "  · 1099-NEC threshold (relevant to entries #1, #14) — TAX-003 1099-NEC Tracker\n"
         "  · Depreciation (#29-34) — TAX-002 (building) and Portfolio Master (asset-by-asset)\n"
         "Full library at thestrledger.com.",
         Font(name=FONT_BODY, size=11, color=GRAPHITE),
         PARCHMENT),
    ]

    r = 3
    for height, text, font, bg in rows:
        ws.row_dimensions[r].height = height
        c = ws.cell(row=r, column=1, value=text)
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        c.fill = PatternFill("solid", fgColor=bg)
        r += 2  # spacer between blocks

    ws.print_area = f"A1:A{r}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_about(wb)
    build_checklist(wb)

    wb.properties.title = "47 Airbnb Tax Deductions — Companion Checklist"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = "Companion Excel checklist for the 47-deductions hero magnet."

    wb.save(OUT)
    size_kb = OUT.stat().st_size / 1024
    print(f"Wrote {OUT.relative_to(REPO)}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
