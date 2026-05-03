"""Build FIN-002 Break-even Occupancy Calculator (v2.2 standard).

Wizard-mode tool — single property, one fill, one answer. Three tabs
(Start / Calculator / Sensitivity). Per brief: no DEMO/BLANK split —
the sample fills act as the demo; the Calculator is fully editable so
the buyer overwrites with their own numbers.

Generates templates/_masters/FIN-002-break-even-occupancy.xlsx.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE, STATE_BAD_FILL,
    COLOR_GRAY_LIGHT,
)

SKU = "FIN-002"
NAME = "break-even-occupancy"
BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "_masters" / f"{SKU}-{NAME}.xlsx"

# ---------------------------------------------------------------------------
# Sample inputs (the workbook ships with these as the demo; buyer overwrites)
# ---------------------------------------------------------------------------

SAMPLE = {
    "property_name":      "Smokies Ridge Cabin",
    "adr":                245,
    "cleaning_charged":   150,
    "avg_los":            3.2,
    "available_nights":   350,
    "platform_pct":       0.15,
    # Annual fixed costs (10 lines)
    "mortgage_interest":  4800,
    "property_tax":       4800,
    "insurance":          2400,
    "hoa":                0,
    "utilities":          2160,    # $180/mo × 12
    "internet":           960,     # $80/mo × 12
    "software":           480,     # $40/mo × 12
    "marketing":          600,
    "other_fixed_1":      0,
    "other_fixed_2":      0,
    # Per-turnover variable
    "cleaning_paid":      150,
    "supplies":           25,
    "maintenance_reserve": 12,    # ~$0.05/night × LOS
    "other_variable":     0,
    # Current occupancy (Start tab input)
    "current_occupancy":  0.62,
}


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Break-even Occupancy Calculator"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "How empty can it get before you're losing money?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: The answer block rows 10-22
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Property name (pulled from Calculator)
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Calculator!C7="","(enter property name on Calculator tab)",Calculator!C7)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # Break-even occupancy headline
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "BREAK-EVEN OCCUPANCY"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = "=Calculator!C47"
    c.font = Font(name=FONT_HEAD, size=48, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 24

    # Equivalent nights
    ws.merge_cells("A16:L16")
    c = ws["A16"]
    c.value = (
        '="= "&ROUND(Calculator!C46,0)&" booked nights/year   '
        '(~"&ROUND(Calculator!C46/12,1)&"/month)"'
    )
    c.font = Font(name=FONT_BODY, size=12, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 22

    # Current occupancy + margin gauge
    ws.merge_cells("A18:F18")
    c = ws["A18"]
    c.value = "Your current occupancy:"
    c.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=18, column=7, value=SAMPLE["current_occupancy"])
    apply_style(cell, input_cell_style())
    cell.number_format = "0%"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G18:H18")

    ws.merge_cells("I18:L18")
    c = ws["I18"]
    c.value = '=IF(G18="","(enter your %)","")'
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[18].height = 24

    # Margin gauge
    ws.merge_cells("A20:L20")
    c = ws["A20"]
    c.value = (
        '=IF(G18="","",IF(G18>=Calculator!C47,'
        '"✅ Margin = "&TEXT(G18-Calculator!C47,"0%")&" above break-even",'
        '"⚠ Below break-even by "&TEXT(Calculator!C47-G18,"0%")))'
    )
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[20].height = 30

    # Footnote (the integrity differentiator)
    ws.merge_cells("A22:L23")
    c = ws["A22"]
    c.value = (
        "Note: this break-even covers operating costs + mortgage interest. "
        "It does NOT cover principal paydown or your time. True cash break-even "
        "is higher — re-run when costs change (insurance bumps, cleaner rate, refi)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[22].height = 22
    ws.row_dimensions[23].height = 22

    # ZONE 3: Pseudo-button nav rows 25-27
    pseudo_button(ws, "A25", "F27", "→  Open Calculator",
                  "'Calculator'!A1", variant="primary")
    pseudo_button(ws, "G25", "L27", "Sensitivity Table",
                  "'Sensitivity'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    # ZONE 4: Upgrade banner row 30
    ws.merge_cells("A30:L30")
    c = ws["A30"]
    c.value = (
        "💡  Want the full P&L picture? Get the Single-Property P&L Tracker "
        f"(TAX-002) at {BRAND_DOMAIN} — Schedule E-ready, $47."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[30].height = 36

    brand_footer(ws, 32,
                 version_line=f"{SKU} · v1.0 · Free updates forever")

    ws.print_area = "A1:L34"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_calculator_tab(wb):
    ws = wb.create_sheet("Calculator")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 42), ("C", 18),
        ("D", 6), ("E", 36),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Calculator",
                        prev_tab="Start", next_tab="Sensitivity")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "All inputs on one page. Yellow = edit; gray = calculated."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    ws.freeze_panes = "A5"

    # ---- REVENUE ASSUMPTIONS (rows 6-13) ----
    _section_band(ws, 6, "REVENUE ASSUMPTIONS")

    rev_fields = [
        (7,  "Property name:",                         SAMPLE["property_name"],   None,
            "(used in Start tab headline)"),
        (8,  "Property name (alias for headline ref):", SAMPLE["property_name"],   None, ""),
        (9,  "ADR — Avg nightly rate ($):",            SAMPLE["adr"],             '"$"#,##0.00',
            "Net of platform commission applied below"),
        (10, "Cleaning fee charged to guest ($):",     SAMPLE["cleaning_charged"], '"$"#,##0.00',
            "What guest pays — the host can keep the spread vs. cleaning_paid"),
        (11, "Avg length of stay (nights):",           SAMPLE["avg_los"],         "0.0",
            "1.5-7 typical; affects per-night variable cost"),
        (12, "Available nights/year:",                  SAMPLE["available_nights"], "0",
            "365 minus blocked (personal use, maintenance, owner stays)"),
        (13, "Platform commission (%):",                SAMPLE["platform_pct"],   "0.0%",
            "Airbnb 15%, VRBO 8%, direct 0-3%"),
    ]
    # Note: row 8 is the property name we reference from the Start tab — set it
    # as a formula-style "carried forward" cell so changing row 7 changes the
    # headline. Actually simpler: just have the Start tab pull from B8 directly,
    # and row 7 doesn't exist — let me consolidate.
    # Re-write rev_fields cleanly without the duplicate row 8.
    rev_fields = [
        (7,  "Property name:",                         SAMPLE["property_name"],   None,
            "(used in Start tab headline)"),
        (9,  "ADR — Avg nightly rate ($):",            SAMPLE["adr"],             '"$"#,##0.00',
            "Gross nightly rate (commission applied below)"),
        (10, "Cleaning fee charged to guest ($):",     SAMPLE["cleaning_charged"], '"$"#,##0.00',
            "What guest pays — host nets the spread vs. cleaning paid"),
        (11, "Avg length of stay (nights):",           SAMPLE["avg_los"],         "0.0",
            "1.5-7 typical; affects per-night variable cost"),
        (12, "Available nights/year:",                  SAMPLE["available_nights"], "0",
            "365 minus blocked (personal use, maintenance, owner stays)"),
        (13, "Platform commission (%):",                SAMPLE["platform_pct"],   "0.0%",
            "Airbnb 15%, VRBO 8%, direct 0-3%"),
    ]
    for row, label, value, fmt, note in rev_fields:
        _input_row(ws, row, label, value, fmt, note)

    # Row 8 alias for headline reference — simply mirrors B7 so Start tab can
    # use Calculator!B8 (matches the original Start formula reference).
    ws.cell(row=8, column=1).value = None  # ensure no leftover
    ws.cell(row=8, column=2, value="(name shown above)").font = Font(
        name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
    )
    cell = ws.cell(row=8, column=3, value="=B7")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[8].height = 16

    # ---- ANNUAL FIXED COSTS (rows 15-26) ----
    _section_band(ws, 15, "ANNUAL FIXED COSTS")

    fixed_fields = [
        (16, "Mortgage interest (annual, $):",  SAMPLE["mortgage_interest"], '"$"#,##0', "Interest only — principal is excluded from break-even"),
        (17, "Property tax (annual, $):",       SAMPLE["property_tax"],      '"$"#,##0', ""),
        (18, "Insurance (annual, $):",          SAMPLE["insurance"],         '"$"#,##0', "Including STR rider"),
        (19, "HOA / Condo fees (annual, $):",   SAMPLE["hoa"],               '"$"#,##0', ""),
        (20, "Utilities (annual, $):",          SAMPLE["utilities"],         '"$"#,##0', "Electric, gas, water, trash combined"),
        (21, "Internet (annual, $):",           SAMPLE["internet"],          '"$"#,##0', ""),
        (22, "PMS / Software (annual, $):",     SAMPLE["software"],          '"$"#,##0', "Hospitable, Hostfully, OwnerRez, etc."),
        (23, "Marketing / photos (annual, $):", SAMPLE["marketing"],         '"$"#,##0', "Photo refresh, listing fees"),
        (24, "Other fixed #1 (annual, $):",     SAMPLE["other_fixed_1"],     '"$"#,##0', "Label this if used"),
        (25, "Other fixed #2 (annual, $):",     SAMPLE["other_fixed_2"],     '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in fixed_fields:
        _input_row(ws, row, label, value, fmt, note)

    # ---- PER-TURNOVER VARIABLE (rows 28-32) ----
    _section_band(ws, 28, "PER-TURNOVER VARIABLE COSTS")

    var_fields = [
        (29, "Cleaning paid to cleaner ($/turnover):", SAMPLE["cleaning_paid"],     '"$"#,##0', ""),
        (30, "Supplies ($/turnover):",                  SAMPLE["supplies"],          '"$"#,##0', "Toiletries, paper, coffee top-ups"),
        (31, "Maintenance reserve ($/turnover):",       SAMPLE["maintenance_reserve"],'"$"#,##0', "Per-booking allocation; 5-8% of revenue typical"),
        (32, "Other variable ($/turnover):",            SAMPLE["other_variable"],   '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in var_fields:
        _input_row(ws, row, label, value, fmt, note)

    # ---- PULL-FROM-TAX-002 CALLOUT (rows 34-37) ----
    ws.merge_cells("A34:L34")
    c = ws["A34"]
    c.value = "📋  PULL FROM YOUR TAX-002 P&L"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[34].height = 20

    ws.merge_cells("A35:L37")
    c = ws["A35"]
    c.value = (
        "If you own TAX-002 Single-Property P&L Tracker, your annual fixed "
        "costs map cleanly: Schedule E lines 13 (mortgage int), 17 (taxes), "
        "10 (insurance), 18 (utilities), 19 (Other — software, marketing, "
        "etc.). Cleaning paid → Schedule E line 8. Supplies → line 16. "
        "Open TAX-002, copy YTD totals from the Schedule E Summary tab, "
        "annualize if mid-year, paste here."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    for r in range(35, 38):
        ws.row_dimensions[r].height = 18

    # ---- OUTPUTS (rows 39-50) ----
    _section_band(ws, 39, "OUTPUTS  —  CALCULATED")

    # C40: Total annual fixed (inputs are in column C, not B)
    _output_row(ws, 40, "Total annual fixed costs:",
                "=SUM(C16:C25)", '"$"#,##0',
                "All fixed annual costs combined")

    # C41: Variable per turnover
    _output_row(ws, 41, "Variable cost per turnover:",
                "=SUM(C29:C32)", '"$"#,##0.00',
                "Cleaning + supplies + reserve + other")

    # C42: Variable per night
    _output_row(ws, 42, "Variable cost per night:",
                "=IFERROR(C41/C11,0)", '"$"#,##0.00',
                "Per-turnover variable ÷ avg LOS")

    # C43: Net revenue per night
    _output_row(ws, 43, "Net revenue per night:",
                "=C9*(1-C13)+C10/C11", '"$"#,##0.00',
                "ADR × (1-commission) + cleaning_charged ÷ LOS")

    # C44: Contribution margin per night
    _output_row(ws, 44, "Contribution margin per night:",
                "=C43-C42", '"$"#,##0.00',
                "Net rev/night − variable/night")

    # C46: Break-even nights
    _output_row(ws, 46, "BREAK-EVEN NIGHTS / YEAR:",
                "=IFERROR(C40/C44,0)", "0",
                "Nights you must book to cover fixed costs",
                emphasize=True)

    # C47: Break-even occupancy %
    _output_row(ws, 47, "BREAK-EVEN OCCUPANCY:",
                "=IFERROR(C46/C12,0)", "0.0%",
                "Break-even nights ÷ available nights",
                emphasize=True)

    # Final brand_footer
    brand_footer(ws, 50,
                 version_line=f"{SKU} · v1.0 · Calculator")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_sensitivity_tab(wb):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 18),
        ("C", 16), ("D", 16), ("E", 16),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Sensitivity",
                        prev_tab="Calculator", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Annual cash flow at every combination of occupancy × ADR. "
        "Red = losing money. Gold = healthy."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Grid header row 6
    headers = [
        ("B6", "Occupancy"),
        ("C6", "ADR -10%"),
        ("D6", "Current ADR"),
        ("E6", "ADR +10%"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # Grid data rows 7-12: occupancy 30/40/50/60/70/80
    occupancies = [0.30, 0.40, 0.50, 0.60, 0.70, 0.80]
    adr_deltas = [0.90, 1.00, 1.10]

    for i, occ in enumerate(occupancies):
        r = 7 + i
        # Col B: occupancy label
        cell = ws.cell(row=r, column=2, value=occ)
        cell.number_format = "0%"
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

        # Cols C/D/E: cash flow at scenario
        # Cash flow = revenue - fixed - variable
        # revenue = booked_nights × ADR_scenario + (booked_turnovers × cleaning_charged) - platform commission
        # For simplicity: revenue ≈ booked_nights × (ADR×(1-comm) + cleaning_charged/LOS)
        #                   variable = booked_turnovers × var_per_turnover
        #                   booked_turnovers = booked_nights / LOS
        # Thus net cash flow = booked_nights × contrib_margin/night - fixed
        # where contrib uses the scenario ADR (adjusts net rev/night)
        for j, mult in enumerate(adr_deltas):
            col = 3 + j  # C, D, E
            cell = ws.cell(row=r, column=col)
            # Booked nights = occ × available
            # Net rev/night with ADR_scenario = ADR×mult×(1-comm) + cleaning/LOS
            # Variable/night = var_per_turnover/LOS  (LOS in Calculator!B11)
            # Margin/night = net_rev_scenario - variable_per_night
            # Cash flow = booked_nights × margin/night - fixed
            # Calculator inputs live in column C (col B holds the labels).
            formula = (
                f"=({occ}*Calculator!$C$12)"
                f"*((Calculator!$C$9*{mult}*(1-Calculator!$C$13))"
                f"+Calculator!$C$10/Calculator!$C$11"
                f"-(SUM(Calculator!$C$29:$C$32)/Calculator!$C$11))"
                f"-SUM(Calculator!$C$16:$C$25)"
            )
            cell.value = formula
            cell.number_format = '"$"#,##0'
            cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Border
            side = Side(style="thin", color=COLOR_GRAY_LIGHT)
            cell.border = Border(top=side, left=side, right=side, bottom=side)

        ws.row_dimensions[r].height = 24

    # Conditional formatting on grid C7:E12
    ws.conditional_formatting.add(
        "C7:E12",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
                   font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "C7:E12",
        CellIsRule(operator="greaterThan", formula=["5000"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
                   font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)),
    )

    # Note row 14
    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = (
        "Read this grid by finding the ADR / occupancy combination that's "
        "realistic for your market. If most of your row is red, your fixed "
        "costs are too high relative to your revenue assumptions — re-check "
        "the Calculator tab for inflated cost lines, or reconsider the property."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[14].height = 22
    ws.row_dimensions[15].height = 22

    brand_footer(ws, 17,
                 version_line=f"{SKU} · v1.0 · Sensitivity")

    ws.print_area = "A1:L20"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt, note):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, note, emphasize=False):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 22 if emphasize else 18


def main():
    wb = Workbook()
    build_start_tab(wb)
    build_calculator_tab(wb)
    build_sensitivity_tab(wb)

    wb.properties.title = "Break-even Occupancy Calculator — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Single-property break-even occupancy calculator for STR hosts — "
        "headline answer + sensitivity grid."
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
