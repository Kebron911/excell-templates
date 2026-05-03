"""Build TAX-010 Cost Segregation DIY Workbook — wizard tool.

Implements templates/_briefs/TAX-010-cost-segregation-diy.md.

Generates BOTH:
  templates/_masters/TAX-010-cost-segregation-diy-DEMO.xlsx
  templates/_masters/TAX-010-cost-segregation-diy-BLANK.xlsx

Tabs:
  0  Start                — wizard hero + 3-card + QS + Get Started + progress
  1  Property Basics      — Section 1 of 5: purchase price + placed-in-service
  2  Land Allocation      — Section 2 of 5: county-assessor ratio method
  3  5-Year Property      — Section 3 of 5: itemized table (12 rows)
  4  7-Year Property      — Section 4 of 5: itemized table (10 rows)
  5  15-Year Property     — Section 5 of 5: itemized table (10 rows)
  6  Y1 Summary           — residual + per-class Y1 + naive comparison + 5-yr schedule
  7  Form 4562 Mirror     — Parts I/II/III for CPA hand-off
  8  Launch               — Y1 cards + audit-defense block + Form 3115 decision
  9  Settings             — active year · bonus % · marginal rate · MACRS Y1 %s

Usage:
    python build_cost_segregation_diy.py
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    COLOR_FORMULA_TINT, COLOR_WHITE,
    STATE_BAD_FILL,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-010"
NAME = "cost-segregation-diy"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
VERSION_LINE = f"{SKU} · v2.3 · Free updates forever"

TOTAL_SECTIONS = 5

# 2026 IRS thresholds (illustrative — update per IRS Rev. Proc. annually)
BONUS_DEPR_2026 = 0.40
MARGINAL_RATE_DEFAULT = 0.24

# MACRS Y1 percentages (half-year convention, 200% DB for 5/7-yr, 150% DB for 15-yr)
MACRS_Y1_5YR = 0.20
MACRS_Y1_7YR = 0.1429
MACRS_Y1_15YR = 0.05

# Audit-defense thresholds
RECLASS_FLAG_PCT = 0.40   # warn if >40% of improvements reclassed to 5/7/15
LAND_LOW_PCT = 0.10       # warn if land < 10% of total purchase
LAND_HIGH_PCT = 0.35      # warn if land > 35% of total purchase

ROW_5YR_CAPACITY = 12
ROW_7YR_CAPACITY = 10
ROW_15YR_CAPACITY = 10

# --- DEMO sample (Amanda's Smokies cabin from blog post 03-str-depreciation.md) ---
SAMPLE_PROPERTY = {
    "purchase_price":   480_000,
    "closing_costs":      8_000,
    "address":          "1247 Wears Valley Rd, Sevierville TN 37862",
    "placed_in_service": date(2026, 4, 15),
    "schedule":         "Schedule E (passive)",
}
SAMPLE_LAND = {
    "county_land":          80_000,
    "county_improvements": 408_000,
}
SAMPLE_5YR = [
    ("Whole-home furniture package (3BR)",                32_000),
    ("Stainless appliance package (fridge/range/DW/W+D)", 12_500),
    ("Hot tub (8-person, installed)",                     14_500),
    ("Game room equipment (pool table + arcade)",          9_800),
    ("Smart TVs + wall mounts × 4",                        4_500),
    ("Window treatments (custom blackouts × 9)",           4_200),
    ("Decorative lighting + ceiling fans",                 3_800),
    ("Smart locks + Ring camera package",                  2_200),
    ("Initial linens + kitchenware package",               1_500),
]
SAMPLE_7YR = []  # Amanda's example combined into 5-yr; BLANK still gets 10 rows
SAMPLE_15YR = [
    ("Hot tub deck + stairs",                  7_000),
    ("Privacy fence (back yard, 220 ft)",      6_500),
    ("Gravel driveway extension + parking",    9_500),
    ("Landscaping (perennials, mulch, wall)",  7_200),
    ("Exterior lighting (path/flood/accent)",  4_800),
]


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def append_callout(ws, row, text):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 38


def append_warning_callout(ws, row, formula_text):
    """Conditional callout — formula returns warning text or empty string."""
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = formula_text
    c.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 30


def append_footer_nav(ws, row, prev_tab, next_tab):
    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = "Next: Y1 Summary →"
        next_target = "'Y1 Summary'!A1"
    pseudo_button(ws, f"A{row}", f"F{row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{row}", f"L{row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 22


# ============================================================================
# Tab 0: Start
# ============================================================================

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- Hero (rows 1-8, navy) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Cost Segregation DIY"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Pull 5 years of depreciation into one. The IRS-acceptable DIY method."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — single declarative answer (suite Theme 1).
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(IFERROR(\'Y1 Summary\'!F12,0)>0,'
        '"✅  Year-1 deduction = "&TEXT(\'Y1 Summary\'!F12,"$#,##0")'
        '&"  ·  Tax savings = "&TEXT(\'Y1 Summary\'!F20,"$#,##0"),'
        '"\U0001F4CA  Fill Property Basics + Land Allocation to see your acceleration.")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- 3-card "What you'll get" (rows 10-20, parchment) ---
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A11:L11")
    c = ws["A11"]; c.value = "What you'll get"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("⚡ MUCH BIGGER YEAR 1",
         "Reclassify 5/7/15-year components out of the default 27.5-year "
         "schedule. Overlay bonus depreciation. Typical STR pulls "
         "$30K–$60K extra into Year 1."),
        ("🛡 AUDIT-DEFENSIBLE",
         "Methodology disclosure block, supporting-document checklist, "
         "and per-class sanity guardrails — built per the IRS Cost "
         "Segregation Audit Techniques Guide."),
        ("📄 CPA HAND-OFF",
         "Print-ready Form 4562 mirror (Parts I / II / III) plus "
         "the per-asset breakdown your CPA needs — including for a "
         "Form 3115 retroactive catch-up."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (ttl, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]; c.value = ttl
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}19")
        c = ws[f"{first}14"]; c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(13, 20):
        ws.row_dimensions[r].height = 22

    # --- Quick Start (rows 22-29, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 30):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A23:L23")
    c = ws["A23"]; c.value = "Quick Start — be done in 90 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24
    quickstart_items = [
        "① Property Basics: purchase + closing + placed-in-service",
        "② Land Allocation: county-assessor ratio method",
        "③ Itemize 5-year property (carpet, appliances, hot tub…)",
        "④ Itemize 7-year (furniture, electronics, kitchen)",
        "⑤ Itemize 15-year exterior (drive, landscape, fence, deck)",
        "⑥ Print Form 4562 mirror — hand to CPA",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + (i if i < 3 else i - 3)
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # --- Get Started button (rows 31-34) ---
    pseudo_button(ws, "A31", "L34",
                   "GET STARTED — PROPERTY BASICS  →",
                   "'Property Basics'!A5", variant="primary")
    for r in range(31, 35):
        ws.row_dimensions[r].height = 22

    # --- Progress dashboard (rows 37+) ---
    ws.merge_cells("A37:F37")
    c = ws["A37"]; c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # COUNTA across input ranges only — skip formula cells (e.g. B10 TOTAL BASIS).
    counta_ranges = [
        "'Property Basics'!B8:B9,'Property Basics'!B11:B13",   # 5 inputs (B10 is formula)
        "'Land Allocation'!B8:B9",                              # 2 inputs (B10-B13 formulas)
        f"'5-Year Property'!C8:C{7+ROW_5YR_CAPACITY}",
        f"'7-Year Property'!C8:C{7+ROW_7YR_CAPACITY}",
        f"'15-Year Property'!C8:C{7+ROW_15YR_CAPACITY}",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in counta_ranges)
    # Total possible = 5 + 2 + 12 + 10 + 10 = 39
    total_possible = 5 + 2 + ROW_5YR_CAPACITY + ROW_7YR_CAPACITY + ROW_15YR_CAPACITY
    ws.merge_cells("G37:L37")
    c = ws["G37"]
    c.value = f'=TEXT(MIN(({counta_sum})/{total_possible},1),"0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Per-section status rows. The "range_" string is what goes inside COUNTA(),
    # so for Property Basics it's two ranges separated by comma to skip B10 formula cell.
    section_rows = [
        ("① Property Basics",   "Property Basics", "'Property Basics'!B8:B9,'Property Basics'!B11:B13", 5),
        ("② Land Allocation",   "Land Allocation", "'Land Allocation'!B8:B9",                            2),
        ("③ 5-Year Property",   "5-Year Property", f"'5-Year Property'!C8:C{7+ROW_5YR_CAPACITY}",        ROW_5YR_CAPACITY),
        ("④ 7-Year Property",   "7-Year Property", f"'7-Year Property'!C8:C{7+ROW_7YR_CAPACITY}",        ROW_7YR_CAPACITY),
        ("⑤ 15-Year Property",  "15-Year Property", f"'15-Year Property'!C8:C{7+ROW_15YR_CAPACITY}",     ROW_15YR_CAPACITY),
    ]
    for i, (label, tab, range_, _expected) in enumerate(section_rows):
        r = 39 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]; c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA({range_})"
        # 7-yr is "optional" — show ✅ if any rows OR if 0 rows (mark as "skipped")
        if tab == "7-Year Property":
            c = ws[f"G{r}"]
            c.value = (f'=IF({ca}=0,"○ Skipped (optional)",'
                       f'"⏳ "&{ca}&" item(s) entered")')
        else:
            c = ws[f"G{r}"]
            c.value = (f'=IF({ca}=0,"⏳ Empty",'
                       f'"⏳ "&{ca}&" of up to {_expected}")')
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    brand_footer(ws, 46, version_line=VERSION_LINE)

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ============================================================================
# Tab 1: Property Basics (Section 1 of 5)
# ============================================================================

def build_property_basics_tab(wb, variant):
    ws = wb.create_sheet("Property Basics")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 1, TOTAL_SECTIONS,
                         "Property Basics",
                         "Purchase price, closing costs, placed-in-service date.",
                         "", "Land Allocation")

    set_col_widths(ws, [
        ("A", 44), ("B", 26),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "Fill the highlighted fields below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22
    card_header(ws, 7, ("A", "L"), "Purchase + placed-in-service")

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    fields = [
        (8,  "Purchase price ($):",                    _val(variant, SAMPLE_PROPERTY["purchase_price"]), '"$"#,##0',  None),
        (9,  "Capitalizable closing/title costs ($):", _val(variant, SAMPLE_PROPERTY["closing_costs"]),  '"$"#,##0',  None),
        (10, "TOTAL BASIS:",                           "=B8+B9",                                          '"$"#,##0',  "formula"),
        (11, "Property address:",                      _val(variant, SAMPLE_PROPERTY["address"]),         "@",         None),
        (12, "Placed-in-service date:",                _val(variant, SAMPLE_PROPERTY["placed_in_service"]), "yyyy-mm-dd", None),
        (13, "Schedule classification:",               _val(variant, SAMPLE_PROPERTY["schedule"]),        "@",         "dropdown"),
    ]
    for r, label, val, fmt, kind in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        if kind == "formula":
            apply_style(c, formula_cell_style())
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        else:
            apply_style(c, input_cell_style())
        c.number_format = fmt
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22

    add_dropdown(ws, "B13",
                 ["Schedule E (passive)", "Schedule C (active)", "Ask my CPA"])

    card_body_fill(ws, 8, 13, ("A", "L"), border=True)

    append_callout(ws, 15, (
        "ⓘ Capitalizable closing costs include title insurance, recording "
        "fees, transfer taxes, attorney fees, and survey costs — NOT "
        "prepaid insurance, prepaid property tax, or financing points "
        "(those are deducted differently). Your settlement statement "
        "(HUD-1 / Closing Disclosure) lists each line item — when in doubt, "
        "ask your CPA which lines are capitalizable."
    ))
    append_callout(ws, 17, (
        "ⓘ Placed-in-service date drives the 27.5-year mid-month convention "
        "for the building residual. Most STRs are placed in service the "
        "day they're first marketed for rental (listing live + photos), "
        "not the closing date. Document this with your first listing screenshot."
    ))
    append_footer_nav(ws, 19, prev_tab="", next_tab="Land Allocation")


# ============================================================================
# Tab 2: Land Allocation (Section 2 of 5)
# ============================================================================

def build_land_allocation_tab(wb, variant):
    ws = wb.create_sheet("Land Allocation")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 2, TOTAL_SECTIONS,
                         "Land Allocation",
                         "County tax assessor ratio method — separates land from depreciable basis.",
                         "Property Basics", "5-Year Property")

    set_col_widths(ws, [
        ("A", 44), ("B", 26),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("Look up your county tax assessor's land + improvements "
               "values for the property — usually free online.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22
    card_header(ws, 7, ("A", "L"), "County tax assessor ratio")

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    fields = [
        (8,  "County tax assessor — LAND value ($):",         _val(variant, SAMPLE_LAND["county_land"]),         '"$"#,##0', None),
        (9,  "County tax assessor — IMPROVEMENTS value ($):", _val(variant, SAMPLE_LAND["county_improvements"]), '"$"#,##0', None),
        (10, "County total (land + improvements):",           "=B8+B9",   '"$"#,##0', "formula"),
        (11, "Land ratio (land ÷ county total):",             "=IFERROR(B8/B10,0)", "0.0%",     "formula"),
        (12, "ALLOCATED LAND ($) — non-depreciable:",         "='Property Basics'!B10*B11", '"$"#,##0', "total"),
        (13, "DEPRECIABLE IMPROVEMENTS BASIS:",               "='Property Basics'!B10-B12", '"$"#,##0', "total"),
    ]
    for r, label, val, fmt, kind in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        if kind in ("formula", "total"):
            apply_style(c, formula_cell_style())
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            if kind == "total":
                c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
                a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        else:
            apply_style(c, input_cell_style())
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.number_format = fmt
        ws.row_dimensions[r].height = 22

    card_body_fill(ws, 8, 13, ("A", "L"), border=True)

    # Conditional warning row 15: land ratio outside typical range
    append_warning_callout(ws, 15, (
        f'=IF(B11=0,"",'
        f'IF(B11<{LAND_LOW_PCT},'
        f'"⚠ Land ratio under {int(LAND_LOW_PCT*100)}% — confirm county assessor values are correct. STR land is rarely below 10%.",'
        f'IF(B11>{LAND_HIGH_PCT},'
        f'"⚠ Land ratio over {int(LAND_HIGH_PCT*100)}% — confirm county assessor values. High land ratio reduces depreciable basis significantly.",'
        f'"")))'
    ))

    append_callout(ws, 17, (
        "ⓘ The county tax assessor ratio method is the default IRS-accepted "
        "approach for DIY land allocation. Alternatives: (a) use the appraisal "
        "report's land value if your purchase appraisal broke it out, or "
        "(b) hire a land appraiser for ~$300 if you believe the county "
        "ratio is unreasonable. Document whichever method you use in the "
        "Launch tab's methodology block."
    ))
    append_callout(ws, 19, (
        "ⓘ Land is NEVER depreciable — this is the single most-audited line "
        "in cost segregation. Auditors compare your land allocation to "
        "county records. If you allocate less to land than the county does, "
        "be prepared to defend why (e.g., appraisal report, comparable sales)."
    ))
    append_footer_nav(ws, 21, prev_tab="Property Basics", next_tab="5-Year Property")


# ============================================================================
# Tabs 3/4/5: 5-Year / 7-Year / 15-Year Property tables
# ============================================================================

def _build_property_class_tab(wb, variant, *, tab_name, section_num, section_subtitle,
                                prev_tab, next_tab,
                                capacity, sample_items, suggested_text, suggested_examples):
    """Shared builder for the three itemized property-class tabs.

    Each tab is structurally identical: header band, instruction, item table
    (description + amount), total row, audit-defense callout, footer nav.

    Args:
        capacity: number of item rows in the table
        sample_items: list of (description, amount) — only used when variant=="demo"
        suggested_text: guidance line for the type of items expected
        suggested_examples: bullet text shown in a small "examples" callout
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, section_num, TOTAL_SECTIONS,
                         tab_name, section_subtitle, prev_tab, next_tab)

    set_col_widths(ws, [
        ("A", 6), ("B", 50), ("C", 16),
        ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = suggested_text
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: column headers
    headers = ["#", "Item description", "Amount ($)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    # Pad header row across 12 cols visually
    for col in range(4, 13):
        cell = ws.cell(row=7, column=col)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 20

    # Item rows
    for i in range(capacity):
        r = 8 + i
        ws.cell(row=r, column=1, value=i + 1).font = Font(
            name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
        )
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        if i < len(sample_items) and variant == "demo":
            desc, amt = sample_items[i]
            ws.cell(row=r, column=2, value=desc)
            ws.cell(row=r, column=3, value=amt)
        for col_idx in [2, 3]:
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, input_cell_style())
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=r, column=3).number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    last_item_row = 7 + capacity

    # Total row
    total_row = last_item_row + 1
    a = ws.cell(row=total_row, column=2, value="TOTAL:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=total_row, column=3, value=f"=SUM(C8:C{last_item_row})")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[total_row].height = 24

    # Examples callout
    append_callout(ws, total_row + 2, suggested_examples)
    # Audit-defense reminder
    append_callout(ws, total_row + 4, (
        "ⓘ For each item, keep supporting documents: vendor invoices, "
        "purchase agreement furniture inventory, photos with date metadata, "
        "or an inspector report itemization. Vague descriptions ('misc "
        "furnishings', 'fixtures') are the #1 audit weak point — be "
        "specific and itemized."
    ))
    append_footer_nav(ws, total_row + 6, prev_tab=prev_tab, next_tab=next_tab)


def build_5yr_property_tab(wb, variant):
    _build_property_class_tab(
        wb, variant,
        tab_name="5-Year Property",
        section_num=3,
        section_subtitle="Carpet, appliances, decorative lighting, hot tub, smart-home gear, etc.",
        prev_tab="Land Allocation",
        next_tab="7-Year Property",
        capacity=ROW_5YR_CAPACITY,
        sample_items=SAMPLE_5YR,
        suggested_text=("5-year MACRS = personal property used in the rental — "
                        "the highest-value reclassification class for STRs."),
        suggested_examples=(
            "ⓘ Common 5-year items: appliances (fridge, range, dishwasher, "
            "W/D), carpet + area rugs, decorative lighting + ceiling fans, "
            "window treatments + blinds, hot tubs + saunas, smart locks + "
            "cameras, smart TVs, game-room equipment, initial linens + "
            "kitchenware. Some 5-year items also qualify for §179 — see "
            "TAX-009 Section 179 Planner."
        ),
    )


def build_7yr_property_tab(wb, variant):
    _build_property_class_tab(
        wb, variant,
        tab_name="7-Year Property",
        section_num=4,
        section_subtitle="Furniture, electronics, kitchen equipment with 7-year recovery.",
        prev_tab="5-Year Property",
        next_tab="15-Year Property",
        capacity=ROW_7YR_CAPACITY,
        sample_items=SAMPLE_7YR,
        suggested_text=("7-year MACRS = office/business furniture and certain "
                        "equipment. Optional — can be combined into 5-year if simpler."),
        suggested_examples=(
            "ⓘ Common 7-year items: office furniture (if you have a workspace "
            "in the rental), some commercial-grade kitchen equipment, agricultural "
            "or specialty equipment. Most STR personal property fits better in "
            "5-year — leave this tab empty if your items are clearly 5-year."
        ),
    )


def build_15yr_property_tab(wb, variant):
    _build_property_class_tab(
        wb, variant,
        tab_name="15-Year Property",
        section_num=5,
        section_subtitle="Land improvements — driveway, landscaping, fencing, exterior lighting.",
        prev_tab="7-Year Property",
        next_tab="Y1 Summary",
        capacity=ROW_15YR_CAPACITY,
        sample_items=SAMPLE_15YR,
        suggested_text=("15-year MACRS = land improvements. NOT the land itself — "
                        "things attached to or done to the land outside the building."),
        suggested_examples=(
            "ⓘ Common 15-year items: driveway + parking pad, landscaping "
            "(plantings, retaining walls, mulch), exterior lighting (path, "
            "flood, accent), fencing, exterior decks + stairs (if separate "
            "from the building), in-ground sprinklers, walkways, exterior "
            "signage. Land improvements are eligible for bonus depreciation "
            "(40% in 2026)."
        ),
    )


# ============================================================================
# Tab 6: Y1 Summary & Schedule
# ============================================================================

def build_y1_summary_tab(wb, variant):
    ws = wb.create_sheet("Y1 Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Y1 Summary",
                         prev_tab="15-Year Property", next_tab="Form 4562 Mirror")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Per-class Year 1 deduction with bonus overlay, vs. naive "
                "27.5-year baseline. Settings drives the bonus % and "
                "marginal rate.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Per-class deduction table — rows 6-12
    set_col_widths(ws, [
        ("A", 22), ("B", 18), ("C", 14), ("D", 14), ("E", 14), ("F", 18),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    headers = ["Class", "Basis ($)", "Bonus %", "Bonus $", "Non-bonus × MACRS Y1 %", "Total Y1 ($)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Last item row references for each property tab
    last_5yr = 7 + ROW_5YR_CAPACITY
    last_7yr = 7 + ROW_7YR_CAPACITY
    last_15yr = 7 + ROW_15YR_CAPACITY
    # The "TOTAL" row in each property tab is at last_item_row + 1
    total_5yr_row = last_5yr + 1
    total_7yr_row = last_7yr + 1
    total_15yr_row = last_15yr + 1

    # Rows 7-10: 5-yr / 7-yr / 15-yr / 27.5-yr residual / Land
    rows = [
        # (label, basis_formula, macrs_y1_setting_cell, has_bonus)
        ("5-year",   f"='5-Year Property'!C{total_5yr_row}",   "Settings!$B$8",  True),
        ("7-year",   f"='7-Year Property'!C{total_7yr_row}",   "Settings!$B$9",  True),
        ("15-year",  f"='15-Year Property'!C{total_15yr_row}", "Settings!$B$10", True),
    ]
    for i, (label, basis_formula, macrs_cell, has_bonus) in enumerate(rows):
        r = 7 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        # Basis
        c = ws.cell(row=r, column=2, value=basis_formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # Bonus %
        c = ws.cell(row=r, column=3, value="=Settings!$B$6")
        apply_style(c, formula_cell_style())
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Bonus $
        c = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # Non-bonus × MACRS Y1
        c = ws.cell(row=r, column=5, value=f"=(B{r}-D{r})*{macrs_cell}")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # Total Y1
        c = ws.cell(row=r, column=6, value=f"=D{r}+E{r}")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 20

    # Row 10: 27.5-year residual (no bonus)
    r = 10
    a = ws.cell(row=r, column=1, value="27.5-yr (residual)")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    # Basis = improvements − 5/7/15
    c = ws.cell(row=r, column=2,
                 value=f"='Land Allocation'!B13-B7-B8-B9")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    # No bonus
    c = ws.cell(row=r, column=3, value="—")
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_FORMULA_TINT)
    c = ws.cell(row=r, column=4, value=0)
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    # MACRS Y1 = mid-month convention based on placed-in-service
    macrs_residual = (
        f"=IFERROR((12.5-MONTH('Property Basics'!B12))/12/27.5,0)"
    )
    c = ws.cell(row=r, column=5, value=macrs_residual)
    apply_style(c, formula_cell_style())
    c.number_format = "0.000%"
    c.alignment = Alignment(horizontal="center", vertical="center")
    # Total Y1 (residual gets only MACRS portion: basis × Y1 fraction)
    c = ws.cell(row=r, column=6,
                 value=f"=B{r}*E{r}")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[r].height = 20

    # Row 11: Land (non-depreciable)
    r = 11
    a = ws.cell(row=r, column=1, value="Land (non-depreciable)")
    a.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c = ws.cell(row=r, column=2, value=f"='Land Allocation'!B12")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    for col in [3, 4, 5, 6]:
        c = ws.cell(row=r, column=col, value="—")
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=COLOR_FORMULA_TINT)
    ws.row_dimensions[r].height = 20

    # Row 12: TOTAL
    r = 12
    a = ws.cell(row=r, column=1, value="TOTAL")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=r, column=2, value=f"='Property Basics'!B10")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in [3, 4, 5]:
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=r, column=6, value=f"=SUM(F7:F10)")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[r].height = 28

    # Reclassification % flag at row 14
    append_warning_callout(ws, 14, (
        f'=IFERROR(IF((B7+B8+B9)/\'Land Allocation\'!B13>{RECLASS_FLAG_PCT},'
        f'"⚠ Reclassification at "&TEXT((B7+B8+B9)/\'Land Allocation\'!B13,"0.0%")&" of improvements basis — above the typical 25–35% defensible range. Consider an engineer-led study (>$750K acquisition) or reduce reclassifications.",'
        f'""),"")'
    ))

    # Naive baseline comparison — rows 16-20
    card_header(ws, 16, ("A", "L"), "Naive 27.5-yr baseline — what you saved")

    bold_lbl = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    naive_rows = [
        (17, "Naive Y1 deduction (everything at 27.5-yr):",
         f"='Land Allocation'!B13*E10"),
        (18, "Accelerated Y1 deduction (this workbook):",
         "=F12"),
        (19, "EXTRA Y1 DEDUCTION:",
         "=F18-F17"),
        (20, "Y1 TAX SAVINGS @ marginal rate:",
         "=F19*Settings!$B$7"),
    ]
    for r, label, formula in naive_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_lbl
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=6, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if "EXTRA" in label or "SAVINGS" in label:
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 22

    card_body_fill(ws, 17, 20, ("A", "L"), border=True)

    # 5-year MACRS schedule — rows 23-30
    card_header(ws, 23, ("A", "L"), "5-year MACRS schedule (non-bonus portions)")

    # Schedule header row
    sched_headers = ["Year", "5-yr", "7-yr", "15-yr", "27.5-yr (bldg)", "Total"]
    for col, h in enumerate(sched_headers, start=1):
        cell = ws.cell(row=24, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[24].height = 20

    # MACRS schedule percentages (200% DB half-year for 5/7-yr; 150% DB for 15-yr)
    macrs_5yr =   [0.2000, 0.3200, 0.1920, 0.1152, 0.1152, 0.0576]
    macrs_7yr =   [0.1429, 0.2449, 0.1749, 0.1249, 0.0893, 0.0892, 0.0893, 0.0446]
    macrs_15yr =  [0.0500, 0.0950, 0.0855, 0.0770, 0.0693, 0.0623, 0.0590,
                    0.0590, 0.0591, 0.0590, 0.0591, 0.0590, 0.0591, 0.0590, 0.0591, 0.0295]

    # Show Y1-Y5 only (first 5 years — most useful)
    for yr_idx in range(5):
        r = 25 + yr_idx
        ws.cell(row=r, column=1, value=f"Year {yr_idx + 1}").font = Font(
            name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
        )
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        # 5-yr: (basis - bonus_$) × macrs_5yr[yr_idx]
        # Same for 7-yr, 15-yr
        # 27.5-yr: basis × Y1_fraction (Y1 only) then full 1/27.5 for Y2+
        formulas = [
            (2, f"=(B7-D7)*{macrs_5yr[yr_idx]}"),
            (3, f"=(B8-D8)*{macrs_7yr[yr_idx]}"),
            (4, f"=(B9-D9)*{macrs_15yr[yr_idx]}"),
            (5, (f"=B10*E10" if yr_idx == 0 else f"=B10/27.5")),
        ]
        for col, formula in formulas:
            c = ws.cell(row=r, column=col, value=formula)
            apply_style(c, formula_cell_style())
            c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        # Total
        c = ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18

    # Y1 row gets bonus added back as a footnote
    append_callout(ws, 31, (
        "ⓘ Year 1 totals above show the NON-BONUS portions only — add the "
        "bonus depreciation row (F12 minus E10×B10 etc.) to get the full Y1 "
        "number from the table at the top. The Form 4562 Mirror tab "
        "consolidates everything into the IRS line items your CPA needs."
    ))

    append_callout(ws, 33, (
        "ⓘ The 27.5-yr (bldg) Year 1 reflects the mid-month convention based "
        "on your placed-in-service month. Years 2+ use the full 1/27.5 = "
        "3.636% annual rate. Bonus depreciation does NOT apply to the 27.5-yr "
        "residual (only to property with recovery period ≤ 20 years)."
    ))

    ws.print_area = "A1:F35"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ============================================================================
# Tab 7: Form 4562 Mirror
# ============================================================================

def build_form_4562_mirror_tab(wb):
    ws = wb.create_sheet("Form 4562 Mirror")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Form 4562 Mirror",
                         prev_tab="Y1 Summary", next_tab="Launch")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("IRS Form 4562 — Parts II (Bonus) and III (MACRS) for "
                "your CPA. Print this and attach to the asset detail.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6), ("B", 50), ("C", 18),
        ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    headers = ["Line", "Description", "Amount"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        apply_style(c, header_row_style())
    ws.row_dimensions[5].height = 20

    lines = [
        ("Part II",  "PART II — SPECIAL DEPRECIATION (BONUS)",       None,                              "header"),
        ("14a",      "Bonus depr — 5-yr property:",                  "='Y1 Summary'!D7",                "money"),
        ("14b",      "Bonus depr — 7-yr property:",                  "='Y1 Summary'!D8",                "money"),
        ("14c",      "Bonus depr — 15-yr property:",                 "='Y1 Summary'!D9",                "money"),
        ("14",       "Total bonus depreciation (sum 14a–14c):",      "='Y1 Summary'!D7+'Y1 Summary'!D8+'Y1 Summary'!D9", "total"),

        ("Part III", "PART III — MACRS DEPRECIATION (CURRENT YEAR)", None,                              "header"),
        ("19a",      "5-year property — non-bonus basis × Y1 %:",    "='Y1 Summary'!E7",                "money"),
        ("19b",      "7-year property — non-bonus basis × Y1 %:",    "='Y1 Summary'!E8",                "money"),
        ("19c",      "15-year property — non-bonus basis × Y1 %:",   "='Y1 Summary'!E9",                "money"),
        ("19h",      "27.5-yr residential (mid-month conv.):",       "='Y1 Summary'!F10",               "total"),
        ("21",       "Total MACRS this year (sum 19a–19h):",         "='Y1 Summary'!E7+'Y1 Summary'!E8+'Y1 Summary'!E9+'Y1 Summary'!F10", "total"),

        ("Part IV",  "PART IV — SUMMARY",                            None,                              "header"),
        ("22",       "GRAND TOTAL Y1 DEDUCTION:",                    "='Y1 Summary'!F12",               "grand_total"),

        ("Routing",  "ROUTING — where this lands on your return",    None,                              "header"),
        ("Sched E",  "If Schedule E filer: Form 4562 → Sched E Line 18", None, "info"),
        ("Sched C",  "If Schedule C filer: Form 4562 → Sched C Line 13", None, "info"),
    ]

    r = 6
    for line_num, label, formula, kind in lines:
        if kind == "header":
            ws.merge_cells(f"A{r}:C{r}")
            c = ws.cell(row=r, column=1, value=f"  {line_num} — {label}")
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
            c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 24
        elif kind == "info":
            ws.merge_cells(f"A{r}:C{r}")
            c = ws.cell(row=r, column=1, value=f"  {label}")
            c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[r].height = 20
        else:
            ws.cell(row=r, column=1, value=line_num).font = Font(
                name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
            )
            ws.cell(row=r, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=r, column=2, value=label).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )
            ws.cell(row=r, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            c = ws.cell(row=r, column=3, value=formula)
            apply_style(c, formula_cell_style())
            c.number_format = '"$"#,##0'
            if kind in ("total", "grand_total"):
                c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                ws.cell(row=r, column=2).font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            if kind == "grand_total":
                c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
                c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
                ws.cell(row=r, column=2).font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
            ws.row_dimensions[r].height = 20
        r += 1

    ws.print_area = f"A1:C{r}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ============================================================================
# Tab 8: Launch
# ============================================================================

def build_launch_tab(wb, variant):
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero strip
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK", "'Form 4562 Mirror'!A1",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]; c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Your cost segregation is ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Print Form 4562 Mirror, complete the methodology block below, and hand both to your CPA."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # 3-card hero (rows 8-13)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 14):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 — Total Y1 deduction
    ws.merge_cells("A9:D9")
    c = ws["A9"]; c.value = "TOTAL Y1 DEDUCTION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]; c.value = "='Y1 Summary'!F12"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A12:D12")
    c = ws["A12"]; c.value = "Bonus + MACRS, all classes"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 — Extra vs naive
    ws.merge_cells("E9:H9")
    c = ws["E9"]; c.value = "EXTRA vs. NAIVE 27.5-YR"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]; c.value = "='Y1 Summary'!F19"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E12:H12")
    c = ws["E12"]; c.value = "extra Y1 deduction this gives you"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 — Tax savings
    ws.merge_cells("I9:L9")
    c = ws["I9"]; c.value = "Y1 TAX SAVINGS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]; c.value = "='Y1 Summary'!F20"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I12:L12")
    c = ws["I12"]; c.value = "at your marginal rate (Settings)"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # Primary action — Print Form 4562
    pseudo_button(ws, "A16", "L19",
                   "📄  PRINT FORM 4562 MIRROR  →",
                   "'Form 4562 Mirror'!A1", variant="primary")
    for r in range(16, 20):
        ws.row_dimensions[r].height = 22

    # Methodology disclosure block (rows 22-33)
    card_header(ws, 22, ("A", "L"), "Audit-defense — Methodology disclosure (sign + date)")

    bold_lbl = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    body = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = ('I used the DIY rule-of-thumb cost segregation method described '
               'in the IRS Cost Segregation Audit Techniques Guide. Allocations '
               'were derived from the sources checked below:')
    c.font = body
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[23].height = 30

    # Source checkboxes — input cells the customer ticks
    sources = [
        (24, "[ ] Purchase agreement / settlement statement (HUD-1 / Closing Disclosure)"),
        (25, "[ ] Appraisal report (provides land + improvements split)"),
        (26, "[ ] County tax assessor records (for land allocation ratio)"),
        (27, "[ ] Pre-purchase inspector report (itemized fixtures + components)"),
        (28, "[ ] Vendor invoices for post-purchase furniture/appliances"),
        (29, "[ ] Photographs with date metadata (before-rental documentation)"),
        (30, "[ ] Industry benchmark percentages (RSMeans, AICPA, etc.)"),
    ]
    for r, label in sources:
        ws.merge_cells(f"A{r}:L{r}")
        c = ws.cell(row=r, column=1, value=label)
        c.font = body
        c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=4)
        ws.row_dimensions[r].height = 18

    # Signature line
    ws.merge_cells("A32:F32")
    c = ws["A32"]; c.value = "Taxpayer signature:"
    c.font = bold_lbl
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells("G32:L32")
    c = ws["G32"]
    apply_style(c, input_cell_style())
    ws.row_dimensions[32].height = 26

    ws.merge_cells("A33:F33")
    c = ws["A33"]; c.value = "Date:"
    c.font = bold_lbl
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells("G33:L33")
    c = ws["G33"]
    apply_style(c, input_cell_style())
    ws.row_dimensions[33].height = 22

    card_body_fill(ws, 23, 33, ("A", "L"), border=True)

    # Decision card: When to hire an engineer (rows 35-39)
    card_header(ws, 35, ("A", "L"), "When to hire an engineer instead")
    ws.merge_cells("A36:L38")
    c = ws["A36"]
    c.value = (
        "Hire an engineer-led study (typical cost $3K–$8K) when ANY of the following:\n"
        "  • Acquisition cost over $750K  •  Multi-building or multi-unit property\n"
        "  • Major land improvements (large parking, complex landscaping)\n"
        "  • You're in a high audit-risk bracket (>$500K AGI) or the deduction will exceed $100K\n"
        "  • You've owned the property 5+ years and want a Form 3115 catch-up — engineer adds rigor"
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    ws.row_dimensions[36].height = 18
    ws.row_dimensions[37].height = 18
    ws.row_dimensions[38].height = 18
    card_body_fill(ws, 36, 38, ("A", "L"), border=True)

    # Decision card: Form 3115 retroactive (rows 40-44)
    card_header(ws, 40, ("A", "L"), "Owned 1+ years? Form 3115 retroactive catch-up")
    ws.merge_cells("A41:L43")
    c = ws["A41"]
    c.value = (
        "If you've already filed returns depreciating this property at the default 27.5-yr, you can apply this cost-seg retroactively via Form 3115 (Application for Change in Accounting Method) with an IRC §481(a) catch-up adjustment.\n"
        "  • The catch-up is the difference between accumulated depreciation under the new method vs. what you've already deducted — claimed in a single year on the year of change.\n"
        "  • This workbook does NOT compute the §481(a) adjustment — the per-asset breakdown above is what your CPA needs to file Form 3115. Hand them both."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    ws.row_dimensions[41].height = 18
    ws.row_dimensions[42].height = 18
    ws.row_dimensions[43].height = 18
    card_body_fill(ws, 41, 43, ("A", "L"), border=True)

    # Upgrade banner
    ws.merge_cells("A46:L46")
    c = ws["A46"]
    c.value = "💡 Upgrade to the Tax Season Bundle ($147) — TAX-001/002/003/009/010 + Schedule E workbook."
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[46].height = 36

    brand_footer(ws, 48, version_line=VERSION_LINE)

    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ============================================================================
# Tab 9: Settings
# ============================================================================

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Settings", prev_tab="Launch", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · bonus depreciation % · marginal rate · MACRS Y1 % per class."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 42), ("B", 18),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # Note: Y1 Summary references Settings!$B$6/$B$7/$B$8/$B$9/$B$10 — keep this layout stable.
    fields = [
        (5,  "Active tax year:",                                2026,                    "0"),
        (6,  "Bonus depreciation % (placed in service Y1):",    BONUS_DEPR_2026,         "0.0%"),
        (7,  "Marginal tax rate (for Y1 savings calc):",        MARGINAL_RATE_DEFAULT,   "0.0%"),
        (8,  "MACRS Y1 % — 5-year property:",                   MACRS_Y1_5YR,            "0.00%"),
        (9,  "MACRS Y1 % — 7-year property:",                   MACRS_Y1_7YR,            "0.00%"),
        (10, "MACRS Y1 % — 15-year property:",                  MACRS_Y1_15YR,           "0.00%"),
    ]
    for r, label, val, fmt in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Freshness stamp (suite Theme 4) — bonus % is statutory + phasing, MACRS
    # tables are stable but customer should still confirm against the IRS Pub
    # 946 they get from their CPA each year.
    ws.merge_cells("A12:F12")
    note = ws["A12"]
    note.value = (
        "📅 Bonus % + MACRS Y1 percentages as of 2026-01-01 — bonus phases to "
        "0% by 2028 (statutory). Confirm against IRS Pub 946 yearly."
    )
    note.font = italic_muted
    note.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[12].height = 28

    # Historical bonus depreciation reference
    a = ws.cell(row=13, column=1, value="Historical bonus depreciation %:")
    a.font = italic_muted
    historical = [
        (14, "2022:", 1.00), (15, "2023:", 0.80), (16, "2024:", 0.60),
        (17, "2025:", 0.40), (18, "2026:", 0.40), (19, "2027:", 0.20),
        (20, "2028:", 0.00),
    ]
    for row, label, val in historical:
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1)
        rc = ws.cell(row=row, column=2, value=val)
        rc.number_format = "0%"
        rc.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)

    # MACRS schedule reference
    ws.row_dimensions[22].height = 8
    a = ws.cell(row=23, column=1, value="MACRS reference (200% DB half-year):")
    a.font = italic_muted
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    macrs_ref = [
        (24, "5-yr Y1:",  0.2000),
        (25, "5-yr Y2:",  0.3200),
        (26, "5-yr Y3:",  0.1920),
        (27, "7-yr Y1:",  0.1429),
        (28, "7-yr Y2:",  0.2449),
        (29, "15-yr Y1:", 0.0500),  # 150% DB
        (30, "15-yr Y2:", 0.0950),
    ]
    for row, label, val in macrs_ref:
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1)
        rc = ws.cell(row=row, column=2, value=val)
        rc.number_format = "0.00%"
        rc.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)


# ============================================================================
# Workbook assembly
# ============================================================================

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_basics_tab(wb, variant)
    build_land_allocation_tab(wb, variant)
    build_5yr_property_tab(wb, variant)
    build_7yr_property_tab(wb, variant)
    build_15yr_property_tab(wb, variant)
    build_y1_summary_tab(wb, variant)
    build_form_4562_mirror_tab(wb)
    build_launch_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Cost Segregation DIY — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "DIY cost segregation workbook for STR hosts — reclassify "
        "5/7/15-year property out of the default 27.5-yr schedule, with "
        "bonus-depreciation overlay and Form 4562 mirror for CPA hand-off."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
