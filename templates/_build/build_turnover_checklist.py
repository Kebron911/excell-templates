"""Build OPS-001 Cleaner Turnover Checklist + Scorecard (v2.2 standard).

Operational-mode tool for a recurring event (per-turnover). Customer
duplicates / prints the Printable Checklist for each turnover, then logs
the result on the Turnover Log so the Scorecard Dashboard can rank
cleaners over time. Implements the v2.2 visual + interaction language set
by the Welcome Book v2.2 and Mileage Log v2.3.

Special case: the Printable Checklist tab is print-optimized (cleaner
hands the printout to the host). Navigation chrome lives ABOVE the print
area so it doesn't waste ink on every printout.

Generates templates/_masters/OPS-001-turnover-checklist.xlsx.
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, Reference

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

OUT = Path(__file__).resolve().parent.parent / "_masters" / "OPS-001-turnover-checklist.xlsx"

# --- Sample / constant data ---

# v2.3 enrichment — adds CO detector check (separate from smoke, required by
# most STR ordinances), high-touch sanitization items per Airbnb's 5-step
# protocol, hot tub chemical check (highest-liability amenity), pest sweep
# (Bed Bug & Flea insurance prerequisite), tripping hazards, HVAC filter
# cadence, and structured damage photo log for AirCover claim defense.
CHECKLIST_ITEMS = [
    ("BEDROOM (6)", [
        "Dust all horizontal surfaces (nightstands, dresser, headboard)",
        "Strip and replace bed linens (crisp hospital corners)",
        "Fresh pillowcases both sides",
        "Vacuum under bed",
        "Empty wastebasket",
        "Check under bed for guest items",
    ]),
    ("BATHROOM (8)", [
        "Scrub toilet inside + outside base",
        "Clean mirror streak-free",
        "Wipe + SANITIZE sink, faucet handles, light switches (high-touch, EPA-approved)",
        "Scrub shower/tub including drain",
        "Replace towels (bath, hand, face)",
        "Restock toilet paper (min 2 rolls)",
        "Empty wastebasket",
        "Check + restock soap, shampoo, body wash to par",
    ]),
    ("KITCHEN (9)", [
        "Wipe all counters",
        "Clean stovetop (all burners + under)",
        "Wipe inside microwave",
        "Run dishwasher if items inside",
        "Wipe fridge exterior + handles",
        "Check fridge interior for guest leftovers (discard)",
        "Empty trash + replace liner",
        "Restock coffee + filters (check par level)",
        "SANITIZE high-touch: cabinet handles, faucet, light switches (EPA-approved)",
    ]),
    ("LIVING (6)", [
        "Vacuum/sweep floors",
        "Dust TV + surfaces",
        "Fluff + realign pillows + throws",
        "Wipe + SANITIZE remote controls, light switches, door knobs (high-touch)",
        "Reset all furniture to original position",
        "Test wifi connects + speed (note router reset if needed)",
    ]),
    ("OUTDOOR (5)", [
        "Sweep porch/deck",
        "Wipe outdoor furniture if present",
        "Empty outdoor trash + take to curb if pickup day",
        "Pest / wasp / rodent sweep (under deck, eaves, trash bins)",
        "Check exterior — no broken bottles, debris, neighbor-facing mess",
    ]),
    ("HOT TUB / POOL (3)", [
        "Hot tub: cover seated, water level OK, chemical test strip in green range",
        "Pool: skimmer cleared, chemical level OK, safety cover/gate latched",
        "Log chemical reading + dose added (if any) — chemical log on roster tab",
    ]),
    ("SUPPLIES (4)", [
        "Coffee pods ≥ 10",
        "Paper towels ≥ 2 rolls",
        "TP ≥ 2 per bathroom",
        "Dish soap + dishwasher pods topped up",
    ]),
    ("SAFETY (5)", [
        "All SMOKE detectors blinking green (test if unsure)",
        "All CARBON MONOXIDE detectors blinking green (every level + near sleeping areas)",
        "All doors + windows locked",
        "Keyless entry code reset (if applicable) — log new code in Host Notes",
        "Tripping hazards / loose railings / exposed wires sweep",
    ]),
    ("MAINTENANCE (2)", [
        "HVAC filter — check + replace per cadence (60-90 days residential filter)",
        "Run all faucets 30s — check for leaks, slow drains, smells",
    ]),
    ("FINAL WALK (4)", [
        "Damage / wear photo log — every room, naming: yyyy-mm-dd_property_room.jpg",
        "Turn thermostat to host vacancy-saver setting (no frozen pipes, no waste)",
        "If damage found: text host + tag photo with arrow note for AirCover claim",
        "Lock up + leave",
    ]),
]

TOTAL_ITEMS = sum(len(items) for _, items in CHECKLIST_ITEMS)  # 52 in v2.3

# Sample scores re-scaled for new max (was /40, now /52). Conditional-format
# bands on the Scorecard Dashboard use 92% / 83% thresholds (≈48 / ≈43).
TURNOVER_LOG_SAMPLES = [
    ("2026-03-14", "Smokies Ridge",  "Sarah — Smokies Clean",          52, "Perfect turnover", "No", 95),
    ("2026-03-16", "Creek Side",     "Miguel — Ridge Housekeeping",    47, "Missed: CO detector test, hot tub strip", "No", 110),
    ("2026-03-18", "Lakehouse A",    "Sarah — Smokies Clean",          50, "Minor: 1 pillowcase off-center", "No", 90),
    ("2026-03-20", "Smokies Ridge",  "Jamie — Solo",                   41, "TP not restocked; shower drain clogged; HVAC filter skipped", "Yes", 75),
    ("2026-03-21", "Lakehouse B",    "Miguel — Ridge Housekeeping",    49, "All good", "No", 105),
    ("2026-03-23", "Smokies Ridge",  "Sarah — Smokies Clean",          49, "Missed outdoor trash + pest sweep", "No", 92),
    ("2026-03-25", "Creek Side",     "Jamie — Solo",                   38, "Coffee pods empty; mirror streaked; high-touch sanitize skipped", "Yes", 65),
    ("2026-03-26", "Mountain View",  "Miguel — Ridge Housekeeping",    48, "", "No", 100),
    ("2026-03-28", "Lakehouse A",    "Sarah — Smokies Clean",          52, "", "No", 88),
    ("2026-03-29", "Downtown Loft",  "Miguel — Ridge Housekeeping",    45, "Dishwasher not run, damage photo log incomplete", "No", 85),
    ("2026-03-30", "Smokies Ridge",  "Sarah — Smokies Clean",          48, "Late start — guest arrived early", "No", 70),
    ("2026-04-01", "Lakehouse B",    "Jamie — Solo",                   44, "Fridge not checked", "No", 72),
    ("2026-04-03", "Mountain View",  "Miguel — Ridge Housekeeping",    47, "", "No", 95),
    ("2026-04-05", "Creek Side",     "Jamie — Solo",                   44, "Minor: thermostat not reset", "No", 68),
    ("2026-04-07", "Lakehouse A",    "Miguel — Ridge Housekeeping",    50, "", "No", 100),
]

CLEANER_ROSTER_SAMPLES = [
    ("Sarah — Smokies Clean",       "(865) 555-0145", "sarah@smokiesclean.com", 45, "2025-06-01", "Flat rate per turnover; reliable"),
    ("Miguel — Ridge Housekeeping", "(865) 555-0177", "miguel@ridgehk.com",      40, "2025-08-15", "Team of 2; can handle back-to-back"),
    ("Jamie — Solo",                "(865) 555-0192", "jamie.cleans@gmail.com",  35, "2026-01-10", "New — still ramping"),
]

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A", "Lakehouse B", "Mountain View", "Downtown Loft"]

SUPPLIES_SAMPLES = [
    ("Smokies Ridge", 12, 4,  8, 30, 12, 3, 3, 20, 0),
    ("Creek Side",    10, 3,  6, 20, 10, 2, 2, 15, 0),
    ("Lakehouse A",   15, 4, 10, 30, 15, 4, 4, 25, 6),
]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def _thin_border_bottom():
    return Border(bottom=Side(style="thin", color="AAAAAA"))


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb):
    """Sheet 0 — Start (operational hero + cards + activity dashboard)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Cleaner Turnover Checklist + Scorecard"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Turnover chaos has a spreadsheet."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "OPS-001 · v2.2"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "A 40-item per-turnover checklist your cleaner runs every time, plus a "
        "host-side log + Scorecard Dashboard that ranks your cleaners over time. "
        "After a few weeks of data, you'll know who your A-team is — by score, "
        "issue rate, and time-on-site — instead of relying on gut feel."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Print the Printable Checklist tab — one per turnover. Hand to your cleaner.",
        "② Cleaner checks off each item, signs, dates, returns to you (text photo OK).",
        "③ Open Turnover Log → add a row: date, property, cleaner, items checked, notes.",
        "④ Open Scorecard Dashboard → see your ranked cleaners + issue rates.",
        "⑤ Cleaner Roster: list all cleaners — drives the Turnover Log dropdown.",
        "⑥ (Optional) Supplies Par Levels: per-property stock targets so par-checks are objective.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "PRINT CHECKLIST" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  PRINT CHECKLIST  (OPEN PRINTABLE)",
                   "'Printable Checklist'!A1", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Turnovers Logged
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "TURNOVERS LOGGED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = "=COUNTA('Turnover Log'!A6:A506)"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "rows on the Turnover Log"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Avg Score
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "AVG SCORE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = '=IFERROR(AVERAGE(\'Turnover Log\'!D6:D506),"—")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0"
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = f"items checked / {TOTAL_ITEMS} max"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Issue Rate
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "ISSUE RATE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = (
        '=IFERROR(COUNTIF(\'Turnover Log\'!F6:F506,"Yes")'
        '/COUNTA(\'Turnover Log\'!A6:A506),"—")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "turnovers with guest complaints"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Turnover Log",
                   "'Turnover Log'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Scorecard Dashboard",
                   "'Scorecard Dashboard'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "Cleaner Roster",
                   "'Cleaner Roster'!A1", variant="secondary")
    pseudo_button(ws, "J39", "L40", "Supplies Par Levels",
                   "'Supplies Par Levels'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Setup callout (rows 42-43) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "🧰 BEFORE FIRST USE: open the Cleaner Roster tab, replace the 3 sample "
        "cleaners with your real ones; open Scorecard Dashboard rows 18-23 and "
        "swap sample property names. Then print the Checklist for your next turnover."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    add_upgrade_banner(ws, 45)

    brand_footer(ws, 47,
                 version_line="OPS-001 · v2.2 · Free updates forever")

    ws.print_area = "A1:L49"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_printable_tab(wb):
    """Sheet 1 — Printable Checklist (cleaner-facing, fits 1 page).

    Special case: print area starts at row 4, so a small back-to-Start
    button on rows 1-2 doesn't print on every cleaner copy.
    """
    ws = wb.create_sheet("Printable Checklist")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    # Tiny nav strip on rows 1-2 (NOT in print area)
    pseudo_button(ws, "A1", "B2", "← START",
                   "'Start'!A1", variant="secondary")
    pseudo_button(ws, "C1", "D2", "Log →",
                   "'Turnover Log'!A1", variant="secondary")

    # --- Print-area starts at row 4 ---
    # Brand title block (printer-friendly, no heavy fills)
    ws["A4"] = "Cleaner Turnover Checklist"
    ws["A4"].font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 26

    ws["A5"] = "Print one per turnover · The STR Ledger"
    ws["A5"].font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    ws.row_dimensions[5].height = 16

    # Property + Date inputs
    ws.row_dimensions[6].height = 20
    ws.cell(row=6, column=1, value="Property:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    prop_cell = ws.cell(row=6, column=2, value="")
    apply_style(prop_cell, input_cell_style())

    ws.cell(row=6, column=3, value="Date:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    date_cell = ws.cell(row=6, column=4, value="")
    apply_style(date_cell, input_cell_style())

    ws.row_dimensions[7].height = 8
    current_row = 8

    for zone_name, items in CHECKLIST_ITEMS:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        zone_cell = ws.cell(row=current_row, column=1, value=zone_name)
        zone_cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        zone_cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        zone_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for item in items:
            chk = ws.cell(row=current_row, column=1, value="☐")
            chk.font = Font(name=FONT_BODY, size=14, color=COLOR_TEXT)
            chk.alignment = Alignment(horizontal="center", vertical="center")

            txt = ws.cell(row=current_row, column=2, value=item)
            txt.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
            txt.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            init = ws.cell(row=current_row, column=3, value="")
            init.border = _thin_border_bottom()
            init.alignment = Alignment(horizontal="center", vertical="center")
            init.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)

            ws.cell(row=current_row, column=4, value="")

            ws.row_dimensions[current_row].height = 16
            current_row += 1

        ws.row_dimensions[current_row].height = 6
        current_row += 1

    # Signature block
    ws.cell(row=current_row, column=1, value="Cleaner name:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    cn_cell = ws.cell(row=current_row, column=2, value="")
    cn_cell.border = _thin_border_bottom()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    ws.cell(row=current_row, column=1, value="Date:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    d_cell = ws.cell(row=current_row, column=2, value="")
    d_cell.border = _thin_border_bottom()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    ws.row_dimensions[current_row].height = 8
    current_row += 1

    ws.cell(row=current_row, column=1, value="Signature:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells(f"B{current_row}:D{current_row}")
    sig_cell = ws.cell(row=current_row, column=2, value="")
    sig_cell.border = _thin_border_bottom()
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    ws.row_dimensions[current_row].height = 8
    current_row += 1

    ws.cell(row=current_row, column=1, value="Time on site (min):").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    t_cell = ws.cell(row=current_row, column=2, value="")
    t_cell.border = _thin_border_bottom()
    ws.row_dimensions[current_row].height = 18

    last_row = current_row

    # Print area excludes the nav row 1-2; starts at row 4
    ws.print_area = f"A4:D{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    set_col_widths(ws, [("A", 4), ("B", 60), ("C", 10), ("D", 15)])


def build_log_tab(wb):
    """Sheet 2 — Turnover Log (one row per turnover, host-facing)."""
    ws = wb.create_sheet("Turnover Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Turnover Log",
                         prev_tab="Printable Checklist",
                         next_tab="Scorecard Dashboard")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per turnover — feeds the Scorecard Dashboard"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Date", "Property", "Cleaner", f"Items Checked (0-{TOTAL_ITEMS})",
        "Notes/Issues", "Guest Complaint?", "Minutes on Site",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 12), ("B", 22), ("C", 22), ("D", 18),
        ("E", 40), ("F", 16), ("G", 12),
    ])

    for i, row_data in enumerate(TURNOVER_LOG_SAMPLES):
        row = 6 + i
        date_val, prop, cleaner, items, notes, complaint, minutes = row_data
        a = ws.cell(row=row, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=row, column=2, value=prop)
        apply_style(b, input_cell_style())

        c = ws.cell(row=row, column=3, value=cleaner)
        apply_style(c, input_cell_style())

        d = ws.cell(row=row, column=4, value=items)
        apply_style(d, input_cell_style())

        e = ws.cell(row=row, column=5, value=notes)
        apply_style(e, input_cell_style())

        f = ws.cell(row=row, column=6, value=complaint)
        apply_style(f, input_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        g = ws.cell(row=row, column=7, value=minutes)
        apply_style(g, input_cell_style())

        ws.row_dimensions[row].height = 16

    # Blank capacity rows
    for row_idx in range(6 + len(TURNOVER_LOG_SAMPLES), 507):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    add_dropdown(ws, "B6:B506", "='Scorecard Dashboard'!$A$18:$A$27")
    add_dropdown(ws, "C6:C506", "='Cleaner Roster'!$A$6:$A$25")
    add_dropdown(ws, "F6:F506", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_scorecard_tab(wb):
    """Sheet 3 — Scorecard Dashboard (rolling metrics per cleaner)."""
    ws = wb.create_sheet("Scorecard Dashboard")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Scorecard Dashboard",
                         prev_tab="Turnover Log",
                         next_tab="Cleaner Roster")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Rolling metrics per cleaner — auto-calculated from the Turnover Log"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = ["Cleaner", "Turnovers", "Avg Score", "Issue Rate", "Rank"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [("A", 28), ("B", 14), ("C", 14), ("D", 14), ("E", 10)])

    # Pre-populate cleaner names rows 6-8
    for idx, roster_row in enumerate(CLEANER_ROSTER_SAMPLES):
        row = 6 + idx
        ws.cell(row=row, column=1, value=roster_row[0]).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )

    # Formulas for rows 6-15 (10 cleaner capacity)
    for i in range(6, 16):
        b_cell = ws.cell(
            row=i, column=2,
            value=f'=IF(A{i}="","",COUNTIF(\'Turnover Log\'!C:C,A{i}))'
        )
        apply_style(b_cell, formula_cell_style())

        c_cell = ws.cell(
            row=i, column=3,
            value=(
                f'=IF(A{i}="","",IFERROR('
                f"AVERAGEIF('Turnover Log'!C:C,A{i},'Turnover Log'!D:D),0))"
            ),
        )
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = "0.0"

        d_cell = ws.cell(
            row=i, column=4,
            value=(
                f'=IF(A{i}="","",IFERROR('
                f"COUNTIFS('Turnover Log'!C:C,A{i},'Turnover Log'!F:F,\"Yes\")"
                f"/COUNTIF('Turnover Log'!C:C,A{i}),0))"
            ),
        )
        apply_style(d_cell, formula_cell_style())
        d_cell.number_format = "0%"

        e_cell = ws.cell(
            row=i, column=5,
            value=f'=IF(OR(A{i}="",B{i}=0),"",RANK.EQ(C{i},$C$6:$C$15,0))'
        )
        apply_style(e_cell, formula_cell_style())

        ws.row_dimensions[i].height = 16

    # Score bands re-scaled for v2.3 (max items = 52). Green ≥92%, Yellow 83-92%, Red <83%.
    green_threshold = round(TOTAL_ITEMS * 0.92, 2)   # 47.84 → 47.84
    yellow_lo = round(TOTAL_ITEMS * 0.83, 2)          # 43.16
    yellow_hi = round(green_threshold - 0.01, 2)      # 47.83
    red_hi = round(yellow_lo - 0.01, 2)               # 43.15
    ws.conditional_formatting.add(
        "C6:C15",
        CellIsRule(operator="greaterThanOrEqual", formula=[str(green_threshold)],
                    fill=PatternFill("solid", fgColor="C7EFCF")),
    )
    ws.conditional_formatting.add(
        "C6:C15",
        CellIsRule(operator="between", formula=[str(yellow_lo), str(yellow_hi)],
                    fill=PatternFill("solid", fgColor="FFF3BF")),
    )
    ws.conditional_formatting.add(
        "C6:C15",
        CellIsRule(operator="between", formula=["0.01", str(red_hi)],
                    fill=PatternFill("solid", fgColor="FFCCCC")),
    )

    ws.row_dimensions[16].height = 10

    prop_hdr = ws.cell(
        row=17, column=1,
        value="Properties (source for Turnover Log dropdown):"
    )
    prop_hdr.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[17].height = 20

    for idx, prop in enumerate(PROPERTIES):
        row = 18 + idx
        cell = ws.cell(row=row, column=1, value=prop)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    # Pad up to row 27 with blank input cells so dropdown source stays valid
    for row in range(18 + len(PROPERTIES), 28):
        cell = ws.cell(row=row, column=1)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    # --- Chart: Avg Score by Cleaner (column, anchored G5) ---
    bar = BarChart()
    bar.type = "col"
    bar.title = "Avg Score by Cleaner"
    bar.y_axis.title = "Score"
    bar.x_axis.title = "Cleaner"
    bar.legend = None
    bar.height = 9
    bar.width = 14
    score_ref = Reference(ws, min_col=3, min_row=5, max_row=15, max_col=3)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=15)
    bar.add_data(score_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "low"
    style_chart(bar)
    ws.add_chart(bar, "G5")

    ws.freeze_panes = "A6"

    ws.print_area = "A1:E27"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_roster_tab(wb):
    """Sheet 4 — Cleaner Roster (team contact list)."""
    ws = wb.create_sheet("Cleaner Roster")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Cleaner Roster",
                         prev_tab="Scorecard Dashboard",
                         next_tab="Supplies Par Levels")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Your team in one place — feeds the Turnover Log dropdown"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = ["Name", "Phone", "Email", "Pay Rate", "Start Date", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 28), ("B", 18), ("C", 28),
        ("D", 12), ("E", 14), ("F", 40),
    ])

    for i, row_data in enumerate(CLEANER_ROSTER_SAMPLES):
        row = 6 + i
        name, phone, email, pay_rate, start_date, notes = row_data
        a = ws.cell(row=row, column=1, value=name)
        apply_style(a, input_cell_style())

        b = ws.cell(row=row, column=2, value=phone)
        apply_style(b, input_cell_style())

        c = ws.cell(row=row, column=3, value=email)
        apply_style(c, input_cell_style())

        d = ws.cell(row=row, column=4, value=pay_rate)
        apply_style(d, input_cell_style())
        d.number_format = '"$"#,##0'

        e = ws.cell(row=row, column=5, value=_parse_date(start_date))
        apply_style(e, input_cell_style())
        e.number_format = "yyyy-mm-dd"

        f = ws.cell(row=row, column=6, value=notes)
        apply_style(f, input_cell_style())

        ws.row_dimensions[row].height = 16

    # Blank capacity rows up to 25 (matches Turnover Log dropdown ref)
    for row in range(6 + len(CLEANER_ROSTER_SAMPLES), 26):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, input_cell_style())
            if col == 4:
                cell.number_format = '"$"#,##0'
            if col == 5:
                cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[row].height = 16

    ws.freeze_panes = "A6"


def build_supplies_tab(wb):
    """Sheet 5 — Supplies Par Levels by Property."""
    ws = wb.create_sheet("Supplies Par Levels")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Supplies Par Levels",
                         prev_tab="Cleaner Roster", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Per-property stock targets — makes par-checks objective"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Property", "Coffee pods", "Paper towels", "TP rolls",
        "Dish pods", "Laundry pods", "Shampoo", "Body wash",
        "Trash bags", "Snacks",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    widths = [("A", 20)] + [(get_column_letter(2 + i), 12) for i in range(9)]
    set_col_widths(ws, widths)

    for i, row_data in enumerate(SUPPLIES_SAMPLES):
        row = 6 + i
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    # Blank capacity rows
    for row in range(6 + len(SUPPLIES_SAMPLES), 16):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    ws.freeze_panes = "B6"


def main():
    wb = Workbook()
    build_start_tab(wb)
    build_printable_tab(wb)
    build_log_tab(wb)
    build_scorecard_tab(wb)
    build_roster_tab(wb)
    build_supplies_tab(wb)

    wb.properties.title = "Cleaner Turnover Checklist + Scorecard — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = "Printable cleaner checklist + rolling cleaner scorecard for STR hosts (v2.2)."

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
