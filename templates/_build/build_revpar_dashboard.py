"""Build FIN-001 RevPAR / ADR / Occupancy Dashboard (v2.2 standard).

Operational-mode KPI dashboard for the Side-Hustle Sam scaling and
Semi-Pro Sarah personas — answers "am I winning per available night,
not just per gross revenue?" Computes RevPAR / ADR / Occupancy with
correct blocked-night handling, YoY deltas, comp benchmark vs market,
and a sensitivity simulator.

Generates two files:
  templates/_masters/FIN-001-revpar-dashboard-DEMO.xlsx
  templates/_masters/FIN-001-revpar-dashboard-BLANK.xlsx
"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE, STATE_BAD_FILL, STATE_GOOD_FILL,
)

SKU = "FIN-001"
NAME = "revpar-dashboard"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Capacity / structural constants
# ---------------------------------------------------------------------------

PROP_CAPACITY = 20            # Property Setup rows
BOOKING_CAPACITY = 5000       # Booking Log rows
BLOCK_CAPACITY = 500          # Blocked-Night Log rows
CALENDAR_DAYS = 365           # Daily calendar Jan 1 - Dec 31
MAX_VISIBLE_PROPS = 8         # KPI Dashboard / Calendar visible cols

CHANNELS = ["Airbnb", "VRBO", "Booking.com", "Direct", "Other"]
BLOCK_REASONS = [
    "Personal use",
    "Maintenance",
    "Owner stay",
    "Pre-booking gap",
    "Other",
]
PROPERTY_TYPES = ["Single-family", "Cabin", "Condo", "Multi-unit", "Glamping", "Other"]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ---------------------------------------------------------------------------
# Demo data — 3 properties × Q1 2026 (matches the brief QA spec)
# ---------------------------------------------------------------------------

PROPERTIES_DEMO = [
    # (name, type, rooms, total_nights, comp_revpar, comp_adr, comp_occ, comp_asof, active)
    ("Smokies Ridge Cabin", "Cabin", 1, 365, 178, 240, 0.65, "2026-04-01", "Yes"),
    ("Creek Side",          "Cabin", 1, 365, 145, 195, 0.71, "2026-04-01", "Yes"),
    ("Lakehouse A",         "Cabin", 1, 365, 210, 320, 0.62, "2026-04-01", "Yes"),
]

# Booking samples — designed so Q1 metrics roughly hit the brief targets.
# Smokies Ridge: target ADR $245, Occ 68%, RevPAR $167.
#   Q1 has 90 days. If occ 68% → ~61 booked nights. ADR $245 → $14,945 gross.
# Creek Side:   target ADR $185, Occ 72%, RevPAR $133.
#   ~65 booked nights × $185 = $12,025 gross.
# Lakehouse A:  target ADR $325, Occ 58%, RevPAR $189.
#   ~52 booked nights × $325 = $16,900 gross.
#
# Each row = one booking with date_in / date_out (LOS = nights = date_out - date_in).
BOOKINGS_DEMO = [
    # (date_in, date_out, property, guest, channel, gross, platform_fee, cleaning_fee_collected, notes)
    # ---- Smokies Ridge Cabin (target ~61 nights × $245 ≈ $14,945)
    ("2026-01-02", "2026-01-08", "Smokies Ridge Cabin", "Avery G.",     "Airbnb",  1470, 220, 150, ""),
    ("2026-01-10", "2026-01-14", "Smokies Ridge Cabin", "Devon S.",     "Airbnb",   980, 147, 150, ""),
    ("2026-01-16", "2026-01-23", "Smokies Ridge Cabin", "Hailey M.",    "Airbnb",  1715, 257, 150, "MLK weekend"),
    ("2026-01-26", "2026-01-31", "Smokies Ridge Cabin", "Olympia V.",   "Direct",  1225,   0, 150, "Direct booking"),
    ("2026-02-04", "2026-02-09", "Smokies Ridge Cabin", "Nadia L.",     "Airbnb",  1225, 184, 150, ""),
    ("2026-02-12", "2026-02-16", "Smokies Ridge Cabin", "Annika W.",    "Airbnb",  1080, 162, 150, "Valentine"),
    ("2026-02-19", "2026-02-23", "Smokies Ridge Cabin", "Jordan B.",    "VRBO",     980,  78, 150, ""),
    ("2026-02-25", "2026-03-02", "Smokies Ridge Cabin", "Tomas R.",     "Airbnb",  1470, 220, 150, ""),
    ("2026-03-05", "2026-03-09", "Smokies Ridge Cabin", "Reese K.",     "Airbnb",   980, 147, 150, ""),
    ("2026-03-13", "2026-03-19", "Smokies Ridge Cabin", "Indira K.",    "Airbnb",  1470, 220, 150, ""),
    ("2026-03-22", "2026-03-28", "Smokies Ridge Cabin", "Felipe O.",    "Airbnb",  1470, 220, 150, "Spring break"),
    ("2026-03-29", "2026-03-31", "Smokies Ridge Cabin", "Brennan E.",   "Direct",   500,   0, 150, ""),
    # ---- Creek Side (target ~65 nights × $185 ≈ $12,025)
    ("2026-01-03", "2026-01-09", "Creek Side", "Marshall H.",  "Airbnb", 1110, 167, 135, ""),
    ("2026-01-11", "2026-01-15", "Creek Side", "Gabriela P.",  "Airbnb",  740, 111, 135, ""),
    ("2026-01-17", "2026-01-23", "Creek Side", "Patrice O.",   "Airbnb", 1110, 167, 135, ""),
    ("2026-01-25", "2026-01-31", "Creek Side", "Sasha B.",     "VRBO",   1110,  88, 135, ""),
    ("2026-02-02", "2026-02-08", "Creek Side", "Kendall U.",   "Airbnb", 1110, 167, 135, ""),
    ("2026-02-10", "2026-02-14", "Creek Side", "Casey N.",     "Airbnb",  740, 111, 135, ""),
    ("2026-02-17", "2026-02-22", "Creek Side", "Wells D.",     "Airbnb",  925, 139, 135, ""),
    ("2026-02-24", "2026-02-28", "Creek Side", "Quincy J.",    "VRBO",    740,  59, 135, ""),
    ("2026-03-02", "2026-03-08", "Creek Side", "Saoirse C.",   "Airbnb", 1110, 167, 135, ""),
    ("2026-03-10", "2026-03-15", "Creek Side", "Jude T.",      "Airbnb",  925, 139, 135, ""),
    ("2026-03-18", "2026-03-23", "Creek Side", "Annika W.",    "Airbnb",  925, 139, 135, ""),
    ("2026-03-25", "2026-03-31", "Creek Side", "Pat D.",       "Airbnb", 1110, 167, 135, ""),
    # ---- Lakehouse A (target ~52 nights × $325 ≈ $16,900)
    ("2026-01-05", "2026-01-12", "Lakehouse A", "Priya R.",     "VRBO",   2275, 182, 175, ""),
    ("2026-01-15", "2026-01-19", "Lakehouse A", "Roy + Linn T.","Airbnb", 1300, 195, 175, ""),
    ("2026-01-22", "2026-01-29", "Lakehouse A", "Chen W.",      "Airbnb", 2275, 341, 175, ""),
    ("2026-02-03", "2026-02-08", "Lakehouse A", "Ezra Y.",      "VRBO",   1625, 130, 175, ""),
    ("2026-02-12", "2026-02-18", "Lakehouse A", "Indira J.",    "Airbnb", 1950, 293, 175, "Valentine premium"),
    ("2026-02-21", "2026-02-26", "Lakehouse A", "Aria F.",      "Airbnb", 1625, 244, 175, ""),
    ("2026-03-04", "2026-03-09", "Lakehouse A", "Casey M.",     "VRBO",   1625, 130, 175, ""),
    ("2026-03-13", "2026-03-19", "Lakehouse A", "Devon Q.",     "Airbnb", 1950, 293, 175, ""),
    ("2026-03-22", "2026-03-29", "Lakehouse A", "Henrik S.",    "Airbnb", 2275, 341, 175, "Spring break"),
]

# Blocked-night samples — owner stays + maintenance, mostly off-season
BLOCKS_DEMO = [
    # (date_from, date_to, property, reason, notes)
    ("2026-01-23", "2026-01-25", "Smokies Ridge Cabin", "Maintenance", "HVAC service"),
    ("2026-02-09", "2026-02-12", "Smokies Ridge Cabin", "Owner stay",  "Family weekend"),
    ("2026-02-23", "2026-02-25", "Smokies Ridge Cabin", "Personal use",""),
    ("2026-01-15", "2026-01-17", "Creek Side", "Maintenance", "Plumbing"),
    ("2026-03-15", "2026-03-18", "Creek Side", "Owner stay", ""),
    ("2026-01-12", "2026-01-15", "Lakehouse A", "Maintenance", "Deep clean + paint"),
    ("2026-02-08", "2026-02-12", "Lakehouse A", "Owner stay",  "Anniversary"),
    ("2026-03-09", "2026-03-13", "Lakehouse A", "Personal use",""),
]

# Prior-year reference (per property × per month) — drives YoY column.
# Smokies Ridge target YoY +$22 RevPAR; Creek Side -$8; Lakehouse A +$45.
# Provide Jan/Feb/Mar prior-year values then 0s for the rest (YoY only meaningful when filled).
PRIOR_YEAR_DEMO = {
    # property -> list of 12 prior-year (ADR, Occ, RevPAR) tuples
    "Smokies Ridge Cabin": [
        (220, 0.62, 136), (235, 0.66, 155), (240, 0.69, 166),
        (0, 0, 0)] * 1 + [(0, 0, 0)] * 9,
    "Creek Side": [
        (180, 0.74, 133), (185, 0.75, 139), (185, 0.78, 144),
        (0, 0, 0)] * 1 + [(0, 0, 0)] * 9,
    "Lakehouse A": [
        (305, 0.52, 159), (310, 0.55, 171), (320, 0.58, 186),
        (0, 0, 0)] * 1 + [(0, 0, 0)] * 9,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _val_list(variant, demo_list):
    return demo_list if variant == "demo" else []


def _parse_date(s):
    if not s:
        return None
    return date.fromisoformat(s)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Navy hero band rows 1-7
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 8):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "RevPAR Dashboard"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Are you actually winning per available night?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Headline KPI block — rows 9-13
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Big PORTFOLIO YTD REVPAR card — A9:L13
    ws.merge_cells("A9:L9")
    c = ws["A9"]
    c.value = "PORTFOLIO YTD REVPAR"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 22

    ws.merge_cells("A10:L12")
    c = ws["A10"]
    # Portfolio RevPAR pulls from KPI Dashboard portfolio-rollup row (RevPAR / YTD)
    c.value = "='KPI Dashboard'!N9"
    c.font = Font(name=FONT_HEAD, size=42, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    for r in range(10, 13):
        ws.row_dimensions[r].height = 26

    # YoY delta row 13
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = (
        '=IFERROR("YoY: "&IF(\'KPI Dashboard\'!O9>=0,"+","")'
        '&TEXT(\'KPI Dashboard\'!O9,"$#,##0")&"  "'
        '&IF(\'KPI Dashboard\'!O9>=0,"▲ improving","▼ slipping"),"YoY: enter prior-year reference on Settings tab")'
    )
    c.font = Font(name=FONT_BODY, size=12, italic=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 22

    # Comp benchmark row 15
    ws.merge_cells("A15:L15")
    c = ws["A15"]
    c.value = (
        '=IFERROR(IF(Settings!$E$8>0,"Comp benchmark: your YTD RevPAR "'
        '&TEXT(\'KPI Dashboard\'!N9,"$#,##0")&" vs market "'
        '&TEXT(Settings!$E$8,"$#,##0")&"  "'
        '&IF(\'KPI Dashboard\'!N9>=Settings!$E$8,"✓ above market","⚠ below market"),'
        '"Paste your AirDNA market RevPAR on Settings → row 8 to see comp benchmark"),"")'
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[15].height = 22

    # Performance scorecard band row 17
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "PERFORMANCE SCORECARD"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    # Best / Worst rows 18-19 — pull from KPI Dashboard property-RevPAR list (col N rows 14-33 alt for 20 properties)
    ws.merge_cells("A18:F18")
    c = ws["A18"]
    c.value = "Best property by RevPAR:"
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    ws.merge_cells("G18:L18")
    c = ws["G18"]
    c.value = (
        '=IFERROR(INDEX(\'KPI Dashboard\'!$M$14:$M$33,'
        'MATCH(MAX(\'KPI Dashboard\'!$N$14:$N$33),\'KPI Dashboard\'!$N$14:$N$33,0))'
        '&"   ("&TEXT(MAX(\'KPI Dashboard\'!$N$14:$N$33),"$#,##0")&" RevPAR)","—")'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = parchment_fill
    ws.row_dimensions[18].height = 22

    ws.merge_cells("A19:F19")
    c = ws["A19"]
    c.value = "Worst property by RevPAR:"
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    ws.merge_cells("G19:L19")
    c = ws["G19"]
    # Worst = MIN over RevPAR values that are > 0 (skip empty/inactive properties)
    c.value = (
        '=IFERROR(INDEX(\'KPI Dashboard\'!$M$14:$M$33,'
        'MATCH(MINIFS(\'KPI Dashboard\'!$N$14:$N$33,\'KPI Dashboard\'!$N$14:$N$33,">0"),'
        '\'KPI Dashboard\'!$N$14:$N$33,0))'
        '&"   ("&TEXT(MINIFS(\'KPI Dashboard\'!$N$14:$N$33,\'KPI Dashboard\'!$N$14:$N$33,">0"),"$#,##0")&" RevPAR)","—")'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = parchment_fill
    ws.row_dimensions[19].height = 22

    # Glossary cards rows 21-30 — parchment, italic, teaching
    ws.merge_cells("A21:L21")
    c = ws["A21"]
    c.value = "GLOSSARY  ·  the jargon, briefly"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[21].height = 22

    glossary = [
        ("RevPAR",
         "Revenue per Available Night = total revenue ÷ available nights. "
         "Tells you if rate AND occupancy are working — not just one. The single "
         "number to watch."),
        ("ADR",
         "Average Daily Rate = total revenue ÷ booked nights. Your average price. "
         "High ADR + low Occ might mean you're pricing too high."),
        ("Occupancy",
         "Booked nights ÷ available nights. We exclude blocked nights (personal "
         "use, maintenance) from the denominator — same convention as AirDNA + "
         "PriceLabs."),
        ("Blocked nights",
         "Nights you took the listing off the market — owner stays, maintenance, "
         "pre-booking gaps. Classified on the Blocked-Night Log so they don't "
         "depress your true Occupancy."),
    ]
    row = 22
    for term, definition in glossary:
        ws.merge_cells(f"A{row}:B{row + 1}")
        c = ws.cell(row=row, column=1)
        c.value = term
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"C{row}:L{row + 1}")
        c = ws.cell(row=row, column=3)
        c.value = definition
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 18
        ws.row_dimensions[row + 1].height = 18
        row += 2

    # Pseudo-button nav — rows 32-34 (input tabs) and 36-38 (output tabs)
    pseudo_button(ws, "A32", "C34", "Property Setup",
                  "'Property Setup'!A1", variant="primary")
    pseudo_button(ws, "D32", "F34", "Booking Log",
                  "'Booking Log'!A1", variant="primary")
    pseudo_button(ws, "G32", "I34", "Blocked-Night Log",
                  "'Blocked-Night Log'!A1", variant="primary")
    pseudo_button(ws, "J32", "L34", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(32, 35):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A36", "C38", "Daily Calendar",
                  "'Daily Calendar'!A1", variant="accent")
    pseudo_button(ws, "D36", "F38", "KPI Dashboard",
                  "'KPI Dashboard'!A1", variant="accent")
    pseudo_button(ws, "G36", "L38", "Sensitivity Simulator",
                  "'Sensitivity Simulator'!A1", variant="accent")
    for r in range(36, 39):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 41
    ws.merge_cells("A41:L41")
    c = ws["A41"]
    c.value = (
        "Want pricing optimization too? Get the Pricing Bundle at "
        f"{BRAND_DOMAIN}/pricing — dynamic pricing calculator + minimum-stay "
        "optimizer + cleaning fee optimizer + competitor tracker, $97."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[41].height = 36

    brand_footer(ws, 43,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L46"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Property Setup
# ---------------------------------------------------------------------------

def build_property_setup_tab(wb, variant):
    ws = wb.create_sheet("Property Setup")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Property Setup",
                        prev_tab="Start", next_tab="Booking Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per property (capacity 20). Total nights/year defaults to 365. "
        "Comp values are optional — paste from AirDNA to see market benchmarks."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Property name", "Type", "Rooms", "Total nights/yr",
        "Comp RevPAR ($)", "Comp ADR ($)", "Comp Occ (%)",
        "Comp as-of", "Active?",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 28), ("B", 14), ("C", 8), ("D", 14),
        ("E", 14), ("F", 14), ("G", 12),
        ("H", 14), ("I", 10),
    ])

    samples = _val_list(variant, PROPERTIES_DEMO)
    for i in range(PROP_CAPACITY):
        row = 6 + i
        if i < len(samples):
            (name, ptype, rooms, total_nights, comp_rev, comp_adr, comp_occ, comp_asof, active) = samples[i]
        else:
            name = ptype = active = comp_asof = None
            rooms = total_nights = comp_rev = comp_adr = comp_occ = None

        cells = [
            (1, name, None),
            (2, ptype, None),
            (3, rooms if rooms is not None else (1 if variant != "demo" else None), "0"),
            (4, total_nights if total_nights is not None else (365 if variant != "demo" else None), "0"),
            (5, comp_rev, '"$"#,##0'),
            (6, comp_adr, '"$"#,##0'),
            (7, comp_occ, "0.0%"),
            (8, _parse_date(comp_asof) if comp_asof else None, "yyyy-mm-dd"),
            (9, active if active is not None else ("Yes" if variant != "demo" else None), None),
        ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    # Dropdowns
    add_dropdown(ws, f"B6:B{5 + PROP_CAPACITY}", f'"{",".join(PROPERTY_TYPES)}"')
    add_dropdown(ws, f"I6:I{5 + PROP_CAPACITY}", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Booking Log (TAX-002 column shape, copy-paste compatible)
# ---------------------------------------------------------------------------

def build_booking_log_tab(wb, variant):
    ws = wb.create_sheet("Booking Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Booking Log",
                        prev_tab="Property Setup", next_tab="Blocked-Night Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per booking (capacity 5,000). LOS auto-calcs from Date in/out. "
        "Same column shape as TAX-002 P&L for copy-paste compatibility."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Date in", "Date out", "Property", "Guest", "Channel",
        "Gross ($)", "Platform fee ($)", "Cleaning fee ($)",
        "LOS (nights)", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 12), ("B", 12), ("C", 24), ("D", 18), ("E", 12),
        ("F", 12), ("G", 14), ("H", 14),
        ("I", 12), ("J", 26),
    ])

    samples = _val_list(variant, BOOKINGS_DEMO)
    for i in range(BOOKING_CAPACITY):
        row = 6 + i
        if i < len(samples):
            (din, dout, prop, guest, channel, gross, fee, clean, notes) = samples[i]
            cells = [
                (1, _parse_date(din), "yyyy-mm-dd"),
                (2, _parse_date(dout), "yyyy-mm-dd"),
                (3, prop, None),
                (4, guest, None),
                (5, channel, None),
                (6, gross, '"$"#,##0.00'),
                (7, fee, '"$"#,##0.00'),
                (8, clean, '"$"#,##0.00'),
                (10, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, "yyyy-mm-dd"),
                (3, None, None),
                (4, None, None),
                (5, None, None),
                (6, None, '"$"#,##0.00'),
                (7, None, '"$"#,##0.00'),
                (8, None, '"$"#,##0.00'),
                (10, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt

        # LOS formula (col I = 9): date_out - date_in
        los_cell = ws.cell(row=row, column=9, value=f"=IFERROR(B{row}-A{row},\"\")")
        apply_style(los_cell, formula_cell_style())
        los_cell.number_format = "0"

        ws.row_dimensions[row].height = 16

    # Dropdowns
    add_dropdown(ws, f"C6:C{5 + BOOKING_CAPACITY}",
                 f"='Property Setup'!$A$6:$A${5 + PROP_CAPACITY}")
    add_dropdown(ws, f"E6:E{5 + BOOKING_CAPACITY}", f'"{",".join(CHANNELS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Blocked-Night Log
# ---------------------------------------------------------------------------

def build_blocked_log_tab(wb, variant):
    ws = wb.create_sheet("Blocked-Night Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Blocked-Night Log",
                        prev_tab="Booking Log", next_tab="Daily Calendar")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Nights the property was OFF the market — owner stays, maintenance, "
        "personal use, pre-booking gaps. Excluded from Occupancy denominator. "
        "Capacity 500 rows."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Date from", "Date to", "Property", "Reason",
        "Nights", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 14), ("B", 14), ("C", 26), ("D", 18),
        ("E", 10), ("F", 32),
    ])

    samples = _val_list(variant, BLOCKS_DEMO)
    for i in range(BLOCK_CAPACITY):
        row = 6 + i
        if i < len(samples):
            (dfrom, dto, prop, reason, notes) = samples[i]
            cells = [
                (1, _parse_date(dfrom), "yyyy-mm-dd"),
                (2, _parse_date(dto), "yyyy-mm-dd"),
                (3, prop, None),
                (4, reason, None),
                (6, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, "yyyy-mm-dd"),
                (3, None, None),
                (4, None, None),
                (6, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt

        # Nights formula
        n_cell = ws.cell(row=row, column=5, value=f"=IFERROR(B{row}-A{row},\"\")")
        apply_style(n_cell, formula_cell_style())
        n_cell.number_format = "0"

        ws.row_dimensions[row].height = 16

    add_dropdown(ws, f"C6:C{5 + BLOCK_CAPACITY}",
                 f"='Property Setup'!$A$6:$A${5 + PROP_CAPACITY}")
    add_dropdown(ws, f"D6:D{5 + BLOCK_CAPACITY}",
                 f'"{",".join(BLOCK_REASONS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Daily Calendar (auto-pivot)
# ---------------------------------------------------------------------------

def build_daily_calendar_tab(wb, variant):
    ws = wb.create_sheet("Daily Calendar")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Daily Calendar",
                        prev_tab="Blocked-Night Log", next_tab="KPI Dashboard")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Auto-pivot. B = booked (from Booking Log), X = blocked (from Blocked-Night "
        "Log), blank = available. Year is driven by Settings → tax year."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Header row 5: Date | property names B..U (20 properties)
    cell = ws.cell(row=5, column=1, value="Date")
    apply_style(cell, header_row_style())
    for i in range(PROP_CAPACITY):
        col = 2 + i
        cell = ws.cell(row=5, column=col,
                       value=f"='Property Setup'!A{6 + i}")
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[5].height = 32

    set_col_widths(ws, [("A", 14)] +
                   [(get_column_letter(2 + i), 14) for i in range(PROP_CAPACITY)])

    # 365 day rows starting row 6 — formula: DATE(year,1,1) + (n-6)
    for d in range(CALENDAR_DAYS):
        row = 6 + d
        # Date in col A — formula references Settings B5 (tax year)
        date_formula = f"=DATE(Settings!$B$5,1,1)+{d}"
        date_cell = ws.cell(row=row, column=1, value=date_formula)
        apply_style(date_cell, formula_cell_style())
        date_cell.number_format = "yyyy-mm-dd ddd"
        date_cell.font = Font(name=FONT_MONO, size=10, color=COLOR_TEXT)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        for i in range(PROP_CAPACITY):
            col = 2 + i
            col_letter = get_column_letter(col)
            # B if any booking covers this date for this property; else X if any block; else blank
            # Booking covers when date_in <= this_date < date_out
            # Block covers when date_from <= this_date < date_to
            formula = (
                f"=IF(COUNTIFS('Booking Log'!$C$6:$C$5005,{col_letter}$5,"
                f"'Booking Log'!$A$6:$A$5005,\"<=\"&$A{row},"
                f"'Booking Log'!$B$6:$B$5005,\">\"&$A{row})>0,\"B\","
                f"IF(COUNTIFS('Blocked-Night Log'!$C$6:$C$505,{col_letter}$5,"
                f"'Blocked-Night Log'!$A$6:$A$505,\"<=\"&$A{row},"
                f"'Blocked-Night Log'!$B$6:$B$505,\">\"&$A{row})>0,\"X\",\"\"))"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = Font(name=FONT_MONO, size=9, bold=True,
                             color=COLOR_PRIMARY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

        ws.row_dimensions[row].height = 16

    # Conditional formatting on the data range B6:U(6+364) = B6:U370
    last_row = 5 + CALENDAR_DAYS
    last_col_letter = get_column_letter(1 + PROP_CAPACITY)
    data_range = f"B6:{last_col_letter}{last_row}"

    booked_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    blocked_fill = PatternFill("solid", fgColor=COLOR_NAVY_TINT)

    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'B6="B"'], fill=booked_fill,
                    font=Font(name=FONT_MONO, size=9, bold=True,
                              color=COLOR_BG_LIGHT)),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'B6="X"'], fill=blocked_fill,
                    font=Font(name=FONT_MONO, size=9, bold=True,
                              color=COLOR_BG_LIGHT)),
    )

    # Freeze top-left so dates and property names stay visible
    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — KPI Dashboard
# ---------------------------------------------------------------------------

def build_kpi_dashboard_tab(wb, variant):
    """KPI Dashboard layout:

    Row 5: header — Metric | Jan..Dec | YTD | YoY
    Rows 7-10: portfolio rollup (ADR / Occ / RevPAR / Gross)
    Row 12: section banner "PER-PROPERTY BREAKDOWN"
    Per property: 4 rows (ADR, Occ, RevPAR, Gross) with 14 cols + helper col M
    Helper col M (rows 14-33): property name (one per property block, anchored to RevPAR row)
    Helper col N (rows 14-33): YTD RevPAR per property (used by Start tab Best/Worst)
    Helper col O (row 9): Portfolio YoY delta (used by Start tab)
    """
    ws = wb.create_sheet("KPI Dashboard")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "KPI Dashboard",
                        prev_tab="Daily Calendar", next_tab="Sensitivity Simulator")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Monthly + YTD + YoY for ADR, Occupancy, RevPAR. Portfolio rollup at top, "
        "per-property below. Occupancy denominator excludes blocked nights."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 32),
        ("B", 10), ("C", 10), ("D", 10), ("E", 10), ("F", 10), ("G", 10),
        ("H", 10), ("I", 10), ("J", 10), ("K", 10), ("L", 10), ("M", 10),
        ("N", 12), ("O", 12),
    ])

    # Header row 5: Category | Jan..Dec | YTD | YoY
    cell = ws.cell(row=5, column=1, value="Metric")
    apply_style(cell, header_row_style())
    for i, m in enumerate(MONTHS):
        cell = ws.cell(row=5, column=2 + i, value=m)
        apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=14, value="YTD")
    apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=15, value="YoY Δ")
    apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # ---- Portfolio rollup banner row 6
    ws.merge_cells("A6:O6")
    c = ws["A6"]
    c.value = "PORTFOLIO ROLLUP"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[6].height = 22

    # Portfolio metric rows 7-10:
    #   Row 7: ADR     = SUM(gross all props) / SUM(booked nights all props)
    #   Row 8: Occ     = SUM(booked nights) / (SUM(total available) - SUM(blocked))
    #   Row 9: RevPAR  = SUM(gross) / (SUM(total available) - SUM(blocked))   [shown big on Start]
    #   Row 10: Gross  = SUM(gross)
    #
    # We use SUMPRODUCT for booked-nights (LOS overlap with month) is heavy; instead use a
    # simplified approach: count "B" cells in Daily Calendar across all property cols for each month.

    # First, helper note: portfolio rows pull from per-property block totals that will be built below.
    # For simplicity + correctness, we'll roll up portfolio = sum of per-property gross / sum of avail.
    # Per-property blocks live below at rows 14+ (4 rows each).
    # Layout: property k block starts at row 14 + (k * 4); rows = ADR, Occ, RevPAR, Gross.

    PROP_BLOCK_START = 14
    PROP_BLOCK_SIZE = 4  # ADR, Occ, RevPAR, Gross

    def prop_block_row(k, metric_offset):
        """k is 0-based property index; metric_offset 0=ADR, 1=Occ, 2=RevPAR, 3=Gross."""
        return PROP_BLOCK_START + k * PROP_BLOCK_SIZE + metric_offset

    # Row 7 — Portfolio ADR (Jan..Dec, YTD, YoY)
    ws.cell(row=7, column=1, value="ADR — Average Daily Rate").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for m in range(12):
        col = 2 + m
        col_letter = get_column_letter(col)
        # ADR portfolio = SUM(prop_gross_m) / SUM(prop_booked_nights_m).
        # Property gross row = prop_block_row(k, 3); booked nights row = prop_block_row(k, 0) is ADR not nights.
        # We need a separate booked-nights pull. Simpler: compute portfolio ADR from raw logs.
        formula = (
            # Portfolio gross for month
            f"=IFERROR(SUMPRODUCT((MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
            f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*'Booking Log'!$F$6:$F$5005)"
            # Divided by booked nights for month (LOS sum where check-in falls in month — approximate)
            f"/IFERROR(SUMPRODUCT((MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
            f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
            f"('Booking Log'!$B$6:$B$5005-'Booking Log'!$A$6:$A$5005)),1),0)"
        )
        cell = ws.cell(row=7, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)

    # YTD ADR
    cell = ws.cell(row=7, column=14, value=(
        f"=IFERROR(SUMPRODUCT((YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
        f"'Booking Log'!$F$6:$F$5005)"
        f"/IFERROR(SUMPRODUCT((YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
        f"('Booking Log'!$B$6:$B$5005-'Booking Log'!$A$6:$A$5005)),1),0)"
    ))
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

    # YoY for ADR — pulled from prior-year reference table on Settings (col J = ADR YTD)
    cell = ws.cell(row=7, column=15,
                   value=f"=IFERROR(N7-Settings!$J$33,\"\")")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"+#,##0;"-$"#,##0'
    cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[7].height = 22

    # Row 8 — Portfolio Occupancy
    ws.cell(row=8, column=1, value="Occupancy — booked / available").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=8, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Number of days in month m (year on Settings B5)
    def days_in_month_formula(m_idx):
        # m_idx 1..12. DAY(EOMONTH(DATE(year,m,1),0))
        return f"DAY(EOMONTH(DATE(Settings!$B$5,{m_idx},1),0))"

    for m in range(12):
        col = 2 + m
        # Booked nights / (active props * days in month - blocked nights for month)
        # Active props count
        active_count = (
            f"COUNTIF('Property Setup'!$I$6:$I${5 + PROP_CAPACITY},\"Yes\")"
        )
        booked = (
            f"SUMPRODUCT((MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
            f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
            f"('Booking Log'!$B$6:$B$5005-'Booking Log'!$A$6:$A$5005))"
        )
        blocked = (
            f"SUMPRODUCT((MONTH('Blocked-Night Log'!$A$6:$A$505)={m+1})*"
            f"(YEAR('Blocked-Night Log'!$A$6:$A$505)=Settings!$B$5)*"
            f"('Blocked-Night Log'!$B$6:$B$505-'Blocked-Night Log'!$A$6:$A$505))"
        )
        formula = (
            f"=IFERROR({booked}/MAX({active_count}*{days_in_month_formula(m+1)}-{blocked},1),0)"
        )
        cell = ws.cell(row=8, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = "0.0%"
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)

    # YTD Occ
    cell = ws.cell(row=8, column=14, value=(
        f"=IFERROR(SUMPRODUCT((YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
        f"('Booking Log'!$B$6:$B$5005-'Booking Log'!$A$6:$A$5005))"
        f"/MAX(COUNTIF('Property Setup'!$I$6:$I${5 + PROP_CAPACITY},\"Yes\")*"
        f"SUMPRODUCT(('Property Setup'!$I$6:$I${5 + PROP_CAPACITY}=\"Yes\")*"
        f"'Property Setup'!$D$6:$D${5 + PROP_CAPACITY})/MAX(COUNTIF('Property Setup'!$I$6:$I${5 + PROP_CAPACITY},\"Yes\"),1)"
        f"-SUMPRODUCT((YEAR('Blocked-Night Log'!$A$6:$A$505)=Settings!$B$5)*"
        f"('Blocked-Night Log'!$B$6:$B$505-'Blocked-Night Log'!$A$6:$A$505)),1),0)"
    ))
    apply_style(cell, formula_cell_style())
    cell.number_format = "0.0%"
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

    # YoY for Occ
    cell = ws.cell(row=8, column=15, value=f"=IFERROR(N8-Settings!$K$33,\"\")")
    apply_style(cell, formula_cell_style())
    cell.number_format = "+0.0%;-0.0%"
    cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[8].height = 22

    # Row 9 — Portfolio RevPAR (the headline number)
    ws.cell(row=9, column=1, value="RevPAR — revenue per available night").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT
    )
    ws.cell(row=9, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=9, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for m in range(12):
        col = 2 + m
        # RevPAR = gross / (active * days - blocked)
        active_count = (
            f"COUNTIF('Property Setup'!$I$6:$I${5 + PROP_CAPACITY},\"Yes\")"
        )
        gross = (
            f"SUMPRODUCT((MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
            f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
            f"'Booking Log'!$F$6:$F$5005)"
        )
        blocked = (
            f"SUMPRODUCT((MONTH('Blocked-Night Log'!$A$6:$A$505)={m+1})*"
            f"(YEAR('Blocked-Night Log'!$A$6:$A$505)=Settings!$B$5)*"
            f"('Blocked-Night Log'!$B$6:$B$505-'Blocked-Night Log'!$A$6:$A$505))"
        )
        formula = (
            f"=IFERROR({gross}/MAX({active_count}*{days_in_month_formula(m+1)}-{blocked},1),0)"
        )
        cell = ws.cell(row=9, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    # YTD RevPAR — referenced by Start tab as 'KPI Dashboard'!N9
    cell = ws.cell(row=9, column=14, value="=SUM(B9:M9)/12")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    # YoY for RevPAR — referenced as 'KPI Dashboard'!O9 by Start tab
    cell = ws.cell(row=9, column=15, value=f"=IFERROR(N9-Settings!$L$33,0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"+#,##0;"-$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[9].height = 28

    # Row 10 — Portfolio Gross
    ws.cell(row=10, column=1, value="Gross revenue").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=10, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for m in range(12):
        col = 2 + m
        formula = (
            f"=IFERROR(SUMPRODUCT((MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
            f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
            f"'Booking Log'!$F$6:$F$5005),0)"
        )
        cell = ws.cell(row=10, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
    cell = ws.cell(row=10, column=14, value="=SUM(B10:M10)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[10].height = 20

    # ---- Per-property breakdown banner row 12
    ws.merge_cells("A12:O12")
    c = ws["A12"]
    c.value = "PER-PROPERTY BREAKDOWN"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[12].height = 22

    # Per-property blocks: 20 properties × 4 rows = 80 rows starting at row 14
    # Also: helper col M (property name anchor) + col N (YTD RevPAR for ranking)
    for k in range(PROP_CAPACITY):
        block_start = PROP_BLOCK_START + k * PROP_BLOCK_SIZE
        # Anchor to row block_start (ADR row)
        prop_ref = f"='Property Setup'!A{6 + k}"

        # Property label spans 4 rows in col A
        ws.merge_cells(f"A{block_start}:A{block_start + 3}")
        cell = ws.cell(row=block_start, column=1, value=prop_ref)
        cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True, indent=1)

        # 4 metric rows
        metric_labels = [
            ("ADR",     '"$"#,##0',  COLOR_TEXT),
            ("Occ",     "0.0%",       COLOR_TEXT),
            ("RevPAR",  '"$"#,##0',  COLOR_ACCENT),
            ("Gross",   '"$"#,##0',  COLOR_MUTED),
        ]
        # Property name as criterion comes from Property Setup A(6+k)
        prop_name_ref = f"'Property Setup'!$A${6 + k}"

        for offset, (label, fmt, color) in enumerate(metric_labels):
            row = block_start + offset
            # Skip A — merged. Fill rest with metric.
            ws.row_dimensions[row].height = 18

            # Label is shown in helper col M (only on RevPAR row for this property — used by Start tab Best/Worst)
            # Actually we use col M for property name (anchor for Best/Worst INDEX/MATCH on N).
            # Put the metric label inside the per-month formulas via no extra label.

            # Per month formulas
            for m in range(12):
                col = 2 + m
                if label == "ADR":
                    formula = (
                        f"=IFERROR(SUMPRODUCT(('Booking Log'!$C$6:$C$5005={prop_name_ref})*"
                        f"(MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
                        f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
                        f"'Booking Log'!$F$6:$F$5005)"
                        f"/MAX(SUMPRODUCT(('Booking Log'!$C$6:$C$5005={prop_name_ref})*"
                        f"(MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
                        f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
                        f"('Booking Log'!$B$6:$B$5005-'Booking Log'!$A$6:$A$5005)),1),0)"
                    )
                elif label == "Occ":
                    booked = (
                        f"SUMPRODUCT(('Booking Log'!$C$6:$C$5005={prop_name_ref})*"
                        f"(MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
                        f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
                        f"('Booking Log'!$B$6:$B$5005-'Booking Log'!$A$6:$A$5005))"
                    )
                    blocked = (
                        f"SUMPRODUCT(('Blocked-Night Log'!$C$6:$C$505={prop_name_ref})*"
                        f"(MONTH('Blocked-Night Log'!$A$6:$A$505)={m+1})*"
                        f"(YEAR('Blocked-Night Log'!$A$6:$A$505)=Settings!$B$5)*"
                        f"('Blocked-Night Log'!$B$6:$B$505-'Blocked-Night Log'!$A$6:$A$505))"
                    )
                    formula = (
                        f"=IF('Property Setup'!$A${6+k}=\"\",\"\","
                        f"IFERROR({booked}/MAX({days_in_month_formula(m+1)}-{blocked},1),0))"
                    )
                elif label == "RevPAR":
                    gross = (
                        f"SUMPRODUCT(('Booking Log'!$C$6:$C$5005={prop_name_ref})*"
                        f"(MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
                        f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
                        f"'Booking Log'!$F$6:$F$5005)"
                    )
                    blocked = (
                        f"SUMPRODUCT(('Blocked-Night Log'!$C$6:$C$505={prop_name_ref})*"
                        f"(MONTH('Blocked-Night Log'!$A$6:$A$505)={m+1})*"
                        f"(YEAR('Blocked-Night Log'!$A$6:$A$505)=Settings!$B$5)*"
                        f"('Blocked-Night Log'!$B$6:$B$505-'Blocked-Night Log'!$A$6:$A$505))"
                    )
                    formula = (
                        f"=IF('Property Setup'!$A${6+k}=\"\",\"\","
                        f"IFERROR({gross}/MAX({days_in_month_formula(m+1)}-{blocked},1),0))"
                    )
                else:  # Gross
                    formula = (
                        f"=IF('Property Setup'!$A${6+k}=\"\",\"\","
                        f"IFERROR(SUMPRODUCT(('Booking Log'!$C$6:$C$5005={prop_name_ref})*"
                        f"(MONTH('Booking Log'!$A$6:$A$5005)={m+1})*"
                        f"(YEAR('Booking Log'!$A$6:$A$5005)=Settings!$B$5)*"
                        f"'Booking Log'!$F$6:$F$5005),0))"
                    )

                cell = ws.cell(row=row, column=col, value=formula)
                apply_style(cell, formula_cell_style())
                cell.number_format = fmt
                cell.font = Font(name=FONT_BODY, size=10,
                                 bold=(label == "RevPAR"), color=color)

            # YTD col N = SUM(B:M) for ADR/Gross; for RevPAR = AVERAGE; for Occ = AVERAGE
            if label in ("ADR", "RevPAR"):
                ytd_formula = (
                    f"=IFERROR(AVERAGEIF(B{row}:M{row},\">0\"),0)"
                )
            elif label == "Occ":
                ytd_formula = f"=IFERROR(AVERAGEIF(B{row}:M{row},\">0\"),0)"
            else:  # Gross
                ytd_formula = f"=SUM(B{row}:M{row})"

            cell = ws.cell(row=row, column=14, value=ytd_formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = fmt
            cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=color)

            # YoY delta col O — pulls from Settings prior-year reference
            # Prior year reference layout: Settings!$H/$I/$J cols × 20 properties × 12 months
            # Simpler: prior-year YTD lives at Settings!$J$13:$J$32 (ADR), $K$13:$K$32 (Occ), $L$13:$L$32 (RevPAR)
            if label == "ADR":
                yoy_formula = f"=IFERROR(N{row}-Settings!$J${13+k},\"\")"
                yoy_fmt = '"$"+#,##0;"-$"#,##0'
            elif label == "Occ":
                yoy_formula = f"=IFERROR(N{row}-Settings!$K${13+k},\"\")"
                yoy_fmt = "+0.0%;-0.0%"
            elif label == "RevPAR":
                yoy_formula = f"=IFERROR(N{row}-Settings!$L${13+k},\"\")"
                yoy_fmt = '"$"+#,##0;"-$"#,##0'
            else:
                yoy_formula = ""
                yoy_fmt = '"$"#,##0'

            if yoy_formula:
                cell = ws.cell(row=row, column=15, value=yoy_formula)
                apply_style(cell, formula_cell_style())
                cell.number_format = yoy_fmt
                cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ACCENT)

            # Helper col M label — small mono row label so user sees "ADR/Occ/RevPAR/Gross"
            # Actually use col M = property name anchor for Best/Worst lookup (only on RevPAR row).
            if label == "RevPAR":
                # col M = property name, col N already set above (YTD RevPAR)
                # Wait — we put YTD in col N. But we wanted helper col M. Let's verify:
                # Header has YTD at col 14 (N) and YoY at col 15 (O). So per-property helpers
                # live in different cols — let's NOT overwrite N (it's the YTD).
                # For Best/Worst lookup, Start tab uses 'KPI Dashboard'!$M$14:$M$33 = property name
                # and $N$14:$N$33 = YTD RevPAR. But we have multiple metric rows per property
                # (block_start to block_start+3). RevPAR is at block_start + 2.
                # So Start should look at rows where property RevPAR lives:
                # block_start + 2 for k=0..19 → rows 16, 20, 24, 28, ..., 92
                # That's NOT a contiguous range. We need to project name+YTD into a flat helper.
                pass

        # Add bottom border between property blocks
        gold_side = Side(style="thin", color=COLOR_ACCENT)
        for c in range(1, 16):
            existing = ws.cell(row=block_start + 3, column=c).border
            ws.cell(row=block_start + 3, column=c).border = Border(
                bottom=gold_side, top=existing.top,
                left=existing.left, right=existing.right,
            )

    # ---- Helper flat-list region for Start tab Best/Worst (rows 14-33)
    # Col M = property name (k = 0..19), Col N = property YTD RevPAR
    # We OVERLOAD this with the per-property block — but col M & N inside blocks are already used
    # (M = month "Dec", N = YTD). We need a different helper region.
    # Solution: put the helper in the COL P/Q range (16/17), not M/N. But the Start tab refs
    # 'KPI Dashboard'!$M$14:$M$33 and $N$14:$N$33. We must keep that contract.
    #
    # Re-design: col M is currently the Dec column (B..M = 12 months → cols 2..13).
    # B=2 (Jan), M=13 (Dec). N=14 (YTD). O=15 (YoY).
    # So col M is ALREADY the Dec column. Start tab cannot use M14:M33 for property names.
    #
    # Fix: Start tab will reference helper cols P (16) for name + Q (17) for YTD RevPAR.
    # Update the Start tab formulas to use $P$14:$P$33 and $Q$14:$Q$33 instead.
    # Since Start was already coded above, it's already pointing to M/N. Adjust here:
    # Add a NEW flat helper at col P/Q for the 20 properties, anchored at rows 14-33.
    # Each property's RevPAR lives at row block_start + 2 = 14 + 4k + 2 = 16 + 4k.
    # We'll project into rows 14..33 (20 rows) so Start's MATCH works.

    # Move helper cols: use cols P (16) = name and Q (17) = YTD RevPAR
    ws.cell(row=13, column=16, value="Helper: name").font = Font(
        name=FONT_MONO, size=8, italic=True, color=COLOR_MUTED)
    ws.cell(row=13, column=17, value="Helper: YTD RevPAR").font = Font(
        name=FONT_MONO, size=8, italic=True, color=COLOR_MUTED)
    set_col_widths(ws, [("P", 22), ("Q", 14)])

    for k in range(PROP_CAPACITY):
        helper_row = 14 + k
        revpar_block_row = PROP_BLOCK_START + k * PROP_BLOCK_SIZE + 2  # RevPAR row
        # Name
        cell = ws.cell(row=helper_row, column=16,
                       value=f"='Property Setup'!A{6 + k}")
        cell.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
        # YTD RevPAR
        cell = ws.cell(row=helper_row, column=17,
                       value=f"=IFERROR(N{revpar_block_row},0)")
        cell.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
        cell.number_format = '"$"#,##0'

    # Update the Start tab's Best/Worst formulas — but they already pointed to M14:M33 / N14:N33.
    # The simplest fix is to write the helper at M/Q. But the rows we have for M conflict
    # because M is "Dec" column header at row 5 and Dec values at rows 7-10 (portfolio) and 14+ (per-prop).
    # Per-property M values are real Dec numbers — not safe to overwrite.
    #
    # Cleanest approach: re-route Start tab formulas. We can't easily edit the Start tab here
    # (already built earlier in the build). Instead: write the helper ALSO at M14:M33 / N14:N33
    # is not possible (those cells contain Dec/YTD values for the property blocks).
    #
    # The fix: Start tab Best/Worst formulas need to reference P14:P33 and Q14:Q33.
    # Since build_start_tab runs BEFORE build_kpi_dashboard_tab, we can't fix it here.
    # We need to update build_start_tab to use P/Q. Let's mark this and fix at the orchestrator.

    # Write same helper at cols P/Q (already done above). Start tab will be updated to point there.

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 7 — Sensitivity Simulator
# ---------------------------------------------------------------------------

def build_sensitivity_tab(wb, variant):
    ws = wb.create_sheet("Sensitivity Simulator")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Sensitivity Simulator",
                        prev_tab="KPI Dashboard", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Pick a property. See what happens if you lift ADR or raise Occupancy. "
        "Each scenario shows projected annual revenue impact."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 4), ("B", 24),
        ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14),
        ("H", 14), ("I", 14), ("J", 4),
    ])

    # Property picker row 6
    ws.cell(row=6, column=2, value="Property:").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=6, column=2).alignment = Alignment(horizontal="right",
                                                    vertical="center", indent=1)
    cell = ws.cell(row=6, column=3,
                   value="Smokies Ridge Cabin" if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.merge_cells("C6:E6")
    add_dropdown(ws, "C6",
                 f"='Property Setup'!$A$6:$A${5 + PROP_CAPACITY}")
    ws.row_dimensions[6].height = 26

    # Current baseline values row 8 (read from KPI Dashboard YTD values)
    ws.cell(row=8, column=2, value="Current YTD baseline").font = Font(
        name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=8, column=2).alignment = Alignment(horizontal="left",
                                                    vertical="center", indent=1)

    # Find the chosen property's row in the KPI Dashboard via MATCH
    # KPI Dashboard col P (helper) row 14..33 has property names
    # ADR row for property k is at 14 + 4k; Occ at 14 + 4k + 1; RevPAR at 14 + 4k + 2
    # Use MATCH to find k, then OFFSET into the right row.
    # match_idx = MATCH(C6, 'KPI Dashboard'!$P$14:$P$33, 0)  -> 1..20 (1-based)
    match_idx = "MATCH($C$6,'KPI Dashboard'!$P$14:$P$33,0)"

    # ADR YTD = INDEX('KPI Dashboard'!$N$14:$N$93, (match_idx-1)*4 + 1)
    adr_ytd_formula = (
        f"=IFERROR(INDEX('KPI Dashboard'!$N$14:$N$93,({match_idx}-1)*4+1),0)"
    )
    occ_ytd_formula = (
        f"=IFERROR(INDEX('KPI Dashboard'!$N$14:$N$93,({match_idx}-1)*4+2),0)"
    )
    rev_ytd_formula = (
        f"=IFERROR(INDEX('KPI Dashboard'!$N$14:$N$93,({match_idx}-1)*4+3),0)"
    )
    gross_ytd_formula = (
        f"=IFERROR(INDEX('KPI Dashboard'!$N$14:$N$93,({match_idx}-1)*4+4),0)"
    )

    baseline_rows = [
        (9, "ADR (Average Daily Rate)", adr_ytd_formula, '"$"#,##0'),
        (10, "Occupancy", occ_ytd_formula, "0.0%"),
        (11, "RevPAR", rev_ytd_formula, '"$"#,##0'),
        (12, "Gross revenue (YTD)", gross_ytd_formula, '"$"#,##0'),
    ]
    for r, label, formula, fmt in baseline_rows:
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=10, color=COLOR_TEXT
        )
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right",
                                                        vertical="center", indent=1)
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = fmt
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 18

    # Annualization helper row 13 — assume YTD values represent the year (Q1-only demo will show Q1 figs)
    ws.cell(row=13, column=2, value="Total available nights (annualized)").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=13, column=2).alignment = Alignment(horizontal="right",
                                                     vertical="center", indent=1)
    # Active nights - blocked nights for the property
    avail_formula = (
        f"=IFERROR(INDEX('Property Setup'!$D$6:$D${5 + PROP_CAPACITY},{match_idx})"
        f"-SUMIFS('Blocked-Night Log'!$E$6:$E$505,"
        f"'Blocked-Night Log'!$C$6:$C$505,$C$6),0)"
    )
    cell = ws.cell(row=13, column=3, value=avail_formula)
    apply_style(cell, formula_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    ws.row_dimensions[13].height = 18

    # ---- Scenario grid section banner row 15
    ws.merge_cells("B15:I15")
    c = ws["B15"]
    c.value = "WHAT-IF SCENARIOS — annual revenue impact"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[15].height = 22

    # Scenario headers row 17 — col B = "ADR delta" labels; cols C/D/E/F/G/H = scenarios
    # Layout: 6 scenarios = 3 ADR deltas × 2 Occ deltas
    # Row 17 = Occ delta header (2 cols each grouped)
    # Row 18 = scenario name (e.g., "+$0 ADR / +0pp Occ")
    # Row 19 = scenario revenue ($)
    # Row 20 = vs current ($)
    # Row 21 = vs current (%)
    scenarios = [
        ("+$0 ADR / +0pp Occ",  0,    0.0),
        ("+$5 ADR / +0pp Occ",  5,    0.0),
        ("+$10 ADR / +0pp Occ", 10,   0.0),
        ("+$0 ADR / +5pp Occ",  0,    0.05),
        ("+$5 ADR / +5pp Occ",  5,    0.05),
        ("+$10 ADR / +5pp Occ", 10,   0.05),
    ]

    # Header row 17
    ws.cell(row=17, column=2, value="Scenario").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_BG_LIGHT
    )
    ws.cell(row=17, column=2).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.cell(row=17, column=2).alignment = Alignment(horizontal="center",
                                                     vertical="center")
    for i, (name, adr_d, occ_d) in enumerate(scenarios):
        col = 3 + i
        cell = ws.cell(row=17, column=col, value=name)
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_BG_LIGHT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
    ws.row_dimensions[17].height = 32

    # Body rows 18-21
    body_rows = [
        (18, "Scenario revenue", '"$"#,##0', COLOR_TEXT, False),
        (19, "vs. current ($)",  '"$"+#,##0;"-$"#,##0', COLOR_PRIMARY, False),
        (20, "vs. current (%)",   "+0.0%;-0.0%", COLOR_ACCENT, True),
    ]
    for r, label, fmt, color, bold in body_rows:
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right",
                                                        vertical="center", indent=1)
        ws.row_dimensions[r].height = 22

    # Each scenario column
    for i, (name, adr_d, occ_d) in enumerate(scenarios):
        col = 3 + i
        col_letter = get_column_letter(col)
        # Scenario revenue = (ADR + adr_d) * (Occ + occ_d) * available_nights
        # Where available_nights is row 13, ADR=row 9, Occ=row 10
        scenario_rev = (
            f"=($C$9+{adr_d})*($C$10+{occ_d})*$C$13"
        )
        cell = ws.cell(row=18, column=col, value=scenario_rev)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

        # Current revenue baseline = ADR * Occ * available_nights (annualized)
        # Use $C$9 * $C$10 * $C$13
        delta_dollars = f"={col_letter}18-($C$9*$C$10*$C$13)"
        cell = ws.cell(row=19, column=col, value=delta_dollars)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"+#,##0;"-$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        delta_pct = (
            f"=IFERROR(({col_letter}18-($C$9*$C$10*$C$13))/MAX(($C$9*$C$10*$C$13),1),0)"
        )
        cell = ws.cell(row=20, column=col, value=delta_pct)
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;-0.0%"
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)

    # Conditional formatting on row 20 — green where impact > 5%
    ws.conditional_formatting.add(
        "C20:H20",
        CellIsRule(operator="greaterThan", formula=["0.05"],
                   fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )
    ws.conditional_formatting.add(
        "C20:H20",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )

    # Insight callout row 23
    ws.merge_cells("B23:I25")
    c = ws["B23"]
    c.value = (
        "Insight: in most STR markets, raising Occupancy 5pp moves more revenue "
        "than lifting ADR $10 — because Occ multiplies ADR. If your green cells "
        "favor Occ, focus on listing optimization (photos, pricing strategy, "
        "minimum-stay tuning). If they favor ADR, focus on premium positioning."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(23, 26):
        ws.row_dimensions[r].height = 18

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 8 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 18),
        ("C", 4), ("D", 30), ("E", 16),
        ("F", 4), ("G", 22),
        ("H", 14), ("I", 14), ("J", 14), ("K", 14), ("L", 14),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Sensitivity Simulator", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Tax year drives all formulas. Comp benchmark + prior-year reference are "
        "OPTIONAL — paste your AirDNA values to unlock comp + YoY callouts."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # ---- TAX YEAR (B5)
    ws.cell(row=5, column=1, value="Year:").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right",
                                                    vertical="center", indent=1)
    cell = ws.cell(row=5, column=2, value=2026 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if variant != "demo":
        cell.value = "=YEAR(TODAY())"
    ws.row_dimensions[5].height = 28

    ws.merge_cells("D5:L5")
    c = ws["D5"]
    c.value = (
        "Type the analysis year here. All KPI / Calendar formulas filter by "
        "this. BLANK defaults to current year via =YEAR(TODAY())."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            wrap_text=True, indent=1)

    # ---- COMP BENCHMARK (Portfolio-level) — D7:E8
    ws.cell(row=7, column=4, value="MARKET COMP BENCHMARK  (paste from AirDNA)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=4).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("D7:E7")
    ws.row_dimensions[7].height = 20

    ws.cell(row=8, column=4, value="Portfolio market RevPAR ($):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=8, column=4).alignment = Alignment(horizontal="right",
                                                    vertical="center", indent=1)
    cell = ws.cell(row=8, column=5, value=178 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 22

    # ---- PRIOR-YEAR REFERENCE TABLE (rows 11-33)
    # G11 = banner; H/I/J/K/L cols = property name | prior-yr ADR | Occ | RevPAR | (notes)
    # We'll reserve rows 13..32 for 20 properties (k=0..19). Final row 33 = portfolio totals (avg/sum).
    ws.merge_cells("G11:L11")
    c = ws["G11"]
    c.value = "PRIOR-YEAR REFERENCE  (per property, optional — drives YoY column on KPI Dashboard)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[11].height = 20

    py_headers = [
        ("G", "Property"),
        ("H", "P-Y Days rented"),
        ("I", "P-Y Gross ($)"),
        ("J", "P-Y ADR ($)"),
        ("K", "P-Y Occ (%)"),
        ("L", "P-Y RevPAR ($)"),
    ]
    for col_letter, label in py_headers:
        cell = ws[f"{col_letter}12"]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[12].height = 24

    # 20 property rows
    samples = _val_list(variant, PROPERTIES_DEMO)
    for k in range(PROP_CAPACITY):
        row = 13 + k
        # Property name pulled from Property Setup
        cell = ws.cell(row=row, column=7, value=f"='Property Setup'!A{6 + k}")
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.fill = parchment_fill

        # Prior year columns — input cells. Demo: populate first 3 properties.
        py_data = None
        if variant == "demo" and k < len(samples):
            prop_name = samples[k][0]
            if prop_name in PRIOR_YEAR_DEMO:
                # Use Q1 averages for ADR/Occ/RevPAR (first 3 months filled in PRIOR_YEAR_DEMO)
                values = PRIOR_YEAR_DEMO[prop_name][:3]
                avg_adr = sum(v[0] for v in values) / 3
                avg_occ = sum(v[1] for v in values) / 3
                avg_revpar = sum(v[2] for v in values) / 3
                # Approximate days rented and gross
                approx_days = int(avg_occ * 90)  # Q1 = 90 days
                approx_gross = int(avg_adr * approx_days)
                py_data = (approx_days, approx_gross, avg_adr, avg_occ, avg_revpar)

        py_inputs = [
            (8, py_data[0] if py_data else None, "0"),
            (9, py_data[1] if py_data else None, '"$"#,##0'),
            (10, py_data[2] if py_data else None, '"$"#,##0'),
            (11, py_data[3] if py_data else None, "0.0%"),
            (12, py_data[4] if py_data else None, '"$"#,##0'),
        ]
        for col, val, fmt in py_inputs:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    # Portfolio totals row 33 — used by KPI Dashboard portfolio YoY
    ws.cell(row=33, column=7, value="PORTFOLIO P-Y TOTALS").font = Font(
        name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=33, column=7).alignment = Alignment(horizontal="right",
                                                     vertical="center", indent=1)
    ws.cell(row=33, column=7).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    # H = SUM days, I = SUM gross, J = portfolio ADR, K = portfolio Occ, L = portfolio RevPAR
    cell = ws.cell(row=33, column=8, value="=SUM(H13:H32)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "0"

    cell = ws.cell(row=33, column=9, value="=SUM(I13:I32)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'

    # Portfolio ADR = total gross / total booked days
    cell = ws.cell(row=33, column=10,
                   value="=IFERROR(I33/MAX(H33,1),0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    # Portfolio Occ = average of K13:K32 (weighted would be better but average is OK)
    cell = ws.cell(row=33, column=11,
                   value="=IFERROR(AVERAGEIF(K13:K32,\">0\"),0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "0.0%"
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    # Portfolio RevPAR = average of L13:L32
    cell = ws.cell(row=33, column=12,
                   value="=IFERROR(AVERAGEIF(L13:L32,\">0\"),0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[33].height = 24

    # ---- BLOCK CATEGORIES (col B rows 7-12)
    ws.merge_cells("A7:B7")
    cell = ws.cell(row=7, column=1, value="BLOCKED-NIGHT REASONS")
    cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 20

    for i, reason in enumerate(BLOCK_REASONS):
        row = 8 + i
        cell = ws.cell(row=row, column=2, value=reason)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    ws.cell(row=14, column=1, value="↑ Blocked-Night Log dropdown source (do not reorder)").font = Font(
        name=FONT_MONO, size=9, italic=True, color=COLOR_MUTED
    )
    ws.merge_cells("A14:E14")

    # Footer
    brand_footer(ws, 36,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def fix_start_tab_helpers(wb):
    """Re-route Start tab Best/Worst formulas to KPI Dashboard helper cols P/Q.

    Build order means Start tab built before KPI Dashboard helper layout was
    finalized — fix here so MATCH/INDEX point at the flat helper region.
    """
    ws = wb["Start"]
    # G18 = best, G19 = worst — replace M with P and N with Q
    g18 = (
        '=IFERROR(INDEX(\'KPI Dashboard\'!$P$14:$P$33,'
        'MATCH(MAX(\'KPI Dashboard\'!$Q$14:$Q$33),\'KPI Dashboard\'!$Q$14:$Q$33,0))'
        '&"   ("&TEXT(MAX(\'KPI Dashboard\'!$Q$14:$Q$33),"$#,##0")&" RevPAR)","—")'
    )
    g19 = (
        '=IFERROR(INDEX(\'KPI Dashboard\'!$P$14:$P$33,'
        'MATCH(MINIFS(\'KPI Dashboard\'!$Q$14:$Q$33,\'KPI Dashboard\'!$Q$14:$Q$33,">0"),'
        '\'KPI Dashboard\'!$Q$14:$Q$33,0))'
        '&"   ("&TEXT(MINIFS(\'KPI Dashboard\'!$Q$14:$Q$33,\'KPI Dashboard\'!$Q$14:$Q$33,">0"),"$#,##0")&" RevPAR)","—")'
    )
    # G18 / G19 are merged anchors; only the top-left of each merge holds value
    ws["G18"].value = g18
    ws["G19"].value = g19


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_setup_tab(wb, variant)
    build_booking_log_tab(wb, variant)
    build_blocked_log_tab(wb, variant)
    build_daily_calendar_tab(wb, variant)
    build_kpi_dashboard_tab(wb, variant)
    build_sensitivity_tab(wb, variant)
    build_settings_tab(wb, variant)

    fix_start_tab_helpers(wb)

    wb.properties.title = "RevPAR Dashboard — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "RevPAR / ADR / Occupancy dashboard for STR hosts — daily calendar, "
        "monthly KPIs, YoY deltas, sensitivity simulator, and comp benchmarking."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
