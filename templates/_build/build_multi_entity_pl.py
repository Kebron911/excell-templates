"""Build FIN-006 Multi-Entity (LLC-per-Property) Consolidated P&L (v2.2 standard).

Operational-mode tool — for hosts who hold each property in its own LLC.
Aggregates per-LLC income statements into a holding-company-level
consolidated P&L. Differs from TAX-011 (multi-property single-entity) by
adding entity-level allocation, intercompany eliminations, and
member-distribution + K-1 prep tracking.

Generates two files:
  templates/_masters/FIN-006-multi-entity-consolidated-pl-DEMO.xlsx
  templates/_masters/FIN-006-multi-entity-consolidated-pl-BLANK.xlsx
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
)

SKU = "FIN-006"
NAME = "multi-entity-consolidated-pl"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Static reference data
# ---------------------------------------------------------------------------

# Capacity
ENTITY_CAPACITY = 10        # Entity Setup rows
INTERCOMPANY_CAPACITY = 30  # Eliminations register
DEMO_ENTITIES_VISIBLE = 3   # Per-Entity P&L sections rendered in DEMO
ENTITY_PL_SECTIONS = 5      # Per-Entity P&L sections rendered (capacity)

# Schedule-E-flavored category list (kept short — entity P&L is high-level)
EXPENSE_CATEGORIES = [
    "Advertising",
    "Cleaning + maintenance",
    "Insurance",
    "Legal + professional",
    "Management fees (intercompany)",
    "Mortgage interest",
    "Repairs",
    "Supplies",
    "Property taxes",
    "Utilities",
    "Platform fees",
    "Other",
]

TAX_CLASSIFICATIONS = [
    "LLC-disregarded",
    "LLC-partnership",
    "LLC-S-corp",
    "Other",
]

INTERCO_TYPES = [
    "Management fee",
    "Allocation",
    "Loan",
    "Other",
]

# ---------------------------------------------------------------------------
# Sample data (DEMO variant)
# Three LLCs (per brief): Smokies Ridge LLC, Lakehouse LLC, Creek Side LLC.
# Plus a holding mgmt entity that charges $200/mo/property.
# Consolidated 2026 NOI target ~$84K (annualized).
# ---------------------------------------------------------------------------

ENTITIES_DEMO = [
    # (Entity name, EIN, State, Property assigned, Registered agent,
    #  Tax classification, Members, Notes)
    ("Smokies Ridge LLC", "88-1234567", "TN", "Smokies Ridge Cabin",
     "Northwest RA", "LLC-disregarded", "Daniel 100%", "Single-member"),
    ("Lakehouse LLC",     "88-2345678", "TN", "Lakehouse A",
     "Northwest RA", "LLC-partnership", "Daniel 60% / Partner 40%",
     "Two-member partnership"),
    ("Creek Side LLC",    "88-3456789", "TN", "Creek Side",
     "Northwest RA", "LLC-disregarded", "Daniel 100%", "Single-member"),
    ("STR Holdings LLC",  "88-4567890", "TN", "(Holding mgmt entity)",
     "Northwest RA", "LLC-S-corp", "Daniel 100%",
     "Charges $200/mo/property — eliminated at consolidation"),
]

# Per-entity annual P&L figures (DEMO).
# Column order matches Per-Entity P&L: revenue + 12 expense lines + NOI.
# Numbers chosen so consolidated NOI annualizes ~$84K after eliminations.
ENTITY_PL_DEMO = {
    # entity_name: (revenue, [expense rows by category])
    "Smokies Ridge LLC": (
        62_000,
        # Advertising, Cleaning, Insurance, Legal, Mgmt fee (interco),
        # Mortgage int, Repairs, Supplies, Prop tax, Utilities, Platform, Other
        [600, 5_400, 1_200, 400, 2_400, 12_800, 1_800, 1_100, 1_450, 2_650, 5_580, 300],
    ),
    "Lakehouse LLC": (
        88_000,
        [800, 7_800, 1_400, 500, 2_400, 16_350, 2_400, 1_400, 1_850, 3_200, 7_920, 350],
    ),
    "Creek Side LLC": (
        54_000,
        [550, 5_100, 1_100, 400, 2_400, 11_900, 1_650, 950,  1_300, 2_400, 4_860, 280],
    ),
    "STR Holdings LLC": (
        # Holding entity collects mgmt fees; minimal expenses.
        # 3 properties * $200/mo * 12 = $7,200 revenue
        7_200,
        [0, 0, 350, 600, 0, 0, 0, 200, 0, 0, 0, 150],
    ),
}

# Intercompany eliminations register (DEMO):
# Holding mgmt LLC charges each per-property LLC $200/mo — eliminated at consolidation.
INTERCO_DEMO = [
    # (Date, From entity, To entity, Description, Amount, Type, Eliminated?)
    ("2026-01-31", "Smokies Ridge LLC", "STR Holdings LLC", "Jan mgmt fee", 200, "Management fee", "Yes"),
    ("2026-02-28", "Smokies Ridge LLC", "STR Holdings LLC", "Feb mgmt fee", 200, "Management fee", "Yes"),
    ("2026-03-31", "Smokies Ridge LLC", "STR Holdings LLC", "Mar mgmt fee", 200, "Management fee", "Yes"),
    ("2026-04-30", "Smokies Ridge LLC", "STR Holdings LLC", "Apr mgmt fee", 200, "Management fee", "Yes"),
    ("2026-05-31", "Smokies Ridge LLC", "STR Holdings LLC", "May mgmt fee", 200, "Management fee", "Yes"),
    ("2026-06-30", "Smokies Ridge LLC", "STR Holdings LLC", "Jun mgmt fee", 200, "Management fee", "Yes"),
    ("2026-07-31", "Smokies Ridge LLC", "STR Holdings LLC", "Jul mgmt fee", 200, "Management fee", "Yes"),
    ("2026-08-31", "Smokies Ridge LLC", "STR Holdings LLC", "Aug mgmt fee", 200, "Management fee", "Yes"),
    ("2026-09-30", "Smokies Ridge LLC", "STR Holdings LLC", "Sep mgmt fee", 200, "Management fee", "Yes"),
    ("2026-10-31", "Smokies Ridge LLC", "STR Holdings LLC", "Oct mgmt fee", 200, "Management fee", "Yes"),
    ("2026-11-30", "Smokies Ridge LLC", "STR Holdings LLC", "Nov mgmt fee", 200, "Management fee", "Yes"),
    ("2026-12-31", "Smokies Ridge LLC", "STR Holdings LLC", "Dec mgmt fee", 200, "Management fee", "Yes"),
    ("2026-01-31", "Lakehouse LLC",     "STR Holdings LLC", "Jan mgmt fee", 200, "Management fee", "Yes"),
    ("2026-02-28", "Lakehouse LLC",     "STR Holdings LLC", "Feb mgmt fee", 200, "Management fee", "Yes"),
    ("2026-03-31", "Lakehouse LLC",     "STR Holdings LLC", "Mar mgmt fee", 200, "Management fee", "Yes"),
    ("2026-06-30", "Lakehouse LLC",     "STR Holdings LLC", "Q2 mgmt (Apr-Jun)", 600, "Management fee", "Yes"),
    ("2026-09-30", "Lakehouse LLC",     "STR Holdings LLC", "Q3 mgmt (Jul-Sep)", 600, "Management fee", "Yes"),
    ("2026-12-31", "Lakehouse LLC",     "STR Holdings LLC", "Q4 mgmt (Oct-Dec)", 600, "Management fee", "Yes"),
    ("2026-03-31", "Creek Side LLC",    "STR Holdings LLC", "Q1 mgmt", 600, "Management fee", "Yes"),
    ("2026-06-30", "Creek Side LLC",    "STR Holdings LLC", "Q2 mgmt", 600, "Management fee", "Yes"),
    ("2026-09-30", "Creek Side LLC",    "STR Holdings LLC", "Q3 mgmt", 600, "Management fee", "Yes"),
    ("2026-12-31", "Creek Side LLC",    "STR Holdings LLC", "Q4 mgmt", 600, "Management fee", "Yes"),
]

# Members list (DEMO)
MEMBERS_DEMO = [
    # (Name, EIN/SSN last4, role)
    ("Daniel Harrison", "1234", "Managing member"),
    ("Partner A.",      "5678", "Member (Lakehouse 40%)"),
]

# Member distributions — rows = members, cols = months (DEMO)
DISTRIBUTIONS_DEMO = {
    # member: 12 month values
    "Daniel Harrison": [3000, 3000, 3500, 3500, 4000, 4500, 5000, 5000, 4500, 4000, 3500, 3500],
    "Partner A.":      [   0,    0, 1500,    0,    0, 1500,    0,    0, 1500,    0,    0, 1500],
}

# Per-entity member ownership % map (drives K-1 worksheet allocation)
# Format: entity -> [(member, pct), ...]
ENTITY_MEMBER_PCT_DEMO = {
    "Smokies Ridge LLC": [("Daniel Harrison", 1.00)],
    "Lakehouse LLC":     [("Daniel Harrison", 0.60), ("Partner A.", 0.40)],
    "Creek Side LLC":    [("Daniel Harrison", 1.00)],
    "STR Holdings LLC":  [("Daniel Harrison", 1.00)],
}

# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _val_list(variant, demo_list):
    return demo_list if variant == "demo" else []


def _parse_date(s):
    if not s:
        return None
    return date.fromisoformat(s)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, max_col=14):
    last = get_column_letter(max_col)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Multi-Entity Consolidated P&L"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "One workbook for the whole holding-company stack. CPA-ready."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards — 4 across rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    cards = [
        ("A", "C", "ACTIVE ENTITIES",
         '=COUNTA(\'Entity Setup\'!A6:A15)', "0", COLOR_PRIMARY),
        ("D", "F", "CONSOLIDATED REVENUE",
         '=\'Consolidated P&L\'!B7', '"$"#,##0', COLOR_PRIMARY),
        ("G", "I", "CONSOLIDATED NOI",
         '=\'Consolidated P&L\'!B24', '"$"#,##0', COLOR_ACCENT),
        ("J", "L", "DISTRIBUTIONS YTD",
         '=\'Member Distributions\'!N12', '"$"#,##0', COLOR_SECONDARY),
    ]
    for fc, lc, label, formula, fmt, color in cards:
        ws.merge_cells(f"{fc}10:{lc}10")
        c = ws[f"{fc}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=8, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{fc}11:{lc}13")
        c = ws[f"{fc}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=22, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

        ws.merge_cells(f"{fc}14:{lc}15")
        c = ws[f"{fc}14"]
        c.value = ""
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for fc, lc, *_ in cards:
        first_c = column_index_from_string(fc)
        last_c = column_index_from_string(lc)
        for r in range(10, 16):
            for col in range(first_c, last_c + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == first_c else existing.left,
                    right=gold_side if col == last_c else existing.right,
                )

    # Pseudo-button nav rows 18-20
    pseudo_button(ws, "A18", "C20", "Entity Setup",
                  "'Entity Setup'!A1", variant="primary")
    pseudo_button(ws, "D18", "F20", "Per-Entity P&L",
                  "'Per-Entity P&L'!A1", variant="primary")
    pseudo_button(ws, "G18", "I20", "Intercompany",
                  "'Intercompany'!A1", variant="primary")
    pseudo_button(ws, "J18", "L20", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(18, 21):
        ws.row_dimensions[r].height = 22

    # Row 22-24
    pseudo_button(ws, "A22", "C24", "Consolidated P&L",
                  "'Consolidated P&L'!A1", variant="accent")
    pseudo_button(ws, "D22", "F24", "Member Distributions",
                  "'Member Distributions'!A1", variant="accent")
    pseudo_button(ws, "G22", "I24", "K-1 Worksheet",
                  "'K-1 Worksheet'!A1", variant="accent")
    pseudo_button(ws, "J22", "L24", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(22, 25):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 27
    ws.merge_cells("A27:L27")
    c = ws["A27"]
    c.value = (
        "Hand-off ready. Print Consolidated P&L + K-1 Worksheet for your CPA — "
        f"questions to {BRAND_EMAIL}."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[27].height = 32

    brand_footer(ws, 29,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L31"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Entity Setup
# ---------------------------------------------------------------------------

def build_entity_setup_tab(wb, variant):
    ws = wb.create_sheet("Entity Setup")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Entity Setup",
                        prev_tab="Start", next_tab="Per-Entity P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        f"One row per LLC (capacity {ENTITY_CAPACITY}). "
        "Entity name drives the dropdowns on Per-Entity P&L, Intercompany, "
        "and K-1 Worksheet."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Entity name", "EIN", "State of formation", "Property assigned",
        "Registered agent", "Tax classification", "Members", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 24), ("B", 14), ("C", 14), ("D", 24),
        ("E", 18), ("F", 18), ("G", 26), ("H", 28),
    ])

    samples = _val_list(variant, ENTITIES_DEMO)
    for i in range(ENTITY_CAPACITY):
        row = 6 + i
        if i < len(samples):
            (name, ein, state, prop, agent, taxcls, members, notes) = samples[i]
        else:
            name = ein = state = prop = agent = taxcls = members = notes = None
        cells = [
            (1, name), (2, ein), (3, state), (4, prop),
            (5, agent), (6, taxcls), (7, members), (8, notes),
        ]
        for col, val in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # Tax classification dropdown
    add_dropdown(ws, f"F6:F{5 + ENTITY_CAPACITY}",
                 f'"{",".join(TAX_CLASSIFICATIONS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Per-Entity P&L
# ---------------------------------------------------------------------------

def build_per_entity_pl_tab(wb, variant):
    """One section per entity. Each section: 12-col monthly P&L
    (revenue line, 12 expense lines, total expenses, NOI).

    Section i lives in rows: header row, then 16 data rows.
    Section block size = 18 (1 title row + header row + 12 expense rows
    + revenue row + total exp row + NOI row + spacer).

    For consolidation simplicity, we emit ONE section per entity from the
    Entity Setup list, capped at ENTITY_PL_SECTIONS.
    """
    ws = wb.create_sheet("Per-Entity P&L")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Per-Entity P&L",
                        prev_tab="Entity Setup", next_tab="Intercompany")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 15):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:N4")
    c = ws["A4"]
    c.value = (
        f"One section per LLC (capacity {ENTITY_PL_SECTIONS}). "
        "Each section: revenue + 12 expense lines + NOI, by month. "
        "Settings → B5 active tax year drives the period."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    set_col_widths(ws, [
        ("A", 32),
        ("B", 10), ("C", 10), ("D", 10), ("E", 10), ("F", 10), ("G", 10),
        ("H", 10), ("I", 10), ("J", 10), ("K", 10), ("L", 10), ("M", 10),
        ("N", 14),
    ])

    SECTION_BLOCK = 22  # rows per entity section
    BASE_ROW = 6        # first section starts here

    # Pull demo entity names in order for sample-data binding
    demo_entity_names = [e[0] for e in ENTITIES_DEMO][:ENTITY_PL_SECTIONS]

    for s_idx in range(ENTITY_PL_SECTIONS):
        start = BASE_ROW + s_idx * SECTION_BLOCK
        # Title row — pulls entity name from Entity Setup
        ws.merge_cells(f"A{start}:N{start}")
        title_cell = ws[f"A{start}"]
        title_cell.value = f"=IF('Entity Setup'!A{6 + s_idx}=\"\",\"(Entity slot {s_idx + 1} — empty)\",\"ENTITY: \"&'Entity Setup'!A{6 + s_idx})"
        title_cell.font = Font(name=FONT_HEAD, size=13, bold=True, color="F6EFE2")
        title_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[start].height = 26

        # Header row: Category | Jan..Dec | YTD
        hdr_row = start + 1
        cell = ws.cell(row=hdr_row, column=1, value="Category")
        apply_style(cell, header_row_style())
        for i, m in enumerate(months):
            cell = ws.cell(row=hdr_row, column=2 + i, value=m)
            apply_style(cell, header_row_style())
        cell = ws.cell(row=hdr_row, column=14, value="YTD")
        apply_style(cell, header_row_style())
        ws.row_dimensions[hdr_row].height = 22

        # Pull demo data only for entities we have samples for
        entity_name = demo_entity_names[s_idx] if s_idx < len(demo_entity_names) else None
        is_demo = (variant == "demo" and entity_name in ENTITY_PL_DEMO)

        # Revenue row
        rev_row = start + 2
        ws.cell(row=rev_row, column=1, value="Revenue (rents + fees)").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
        )
        ws.cell(row=rev_row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        if is_demo:
            annual_rev, _ = ENTITY_PL_DEMO[entity_name]
            # Distribute across months — simple even split
            monthly = annual_rev / 12.0
            for m_idx in range(12):
                cell = ws.cell(row=rev_row, column=2 + m_idx, value=round(monthly, 2))
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0'
        else:
            for m_idx in range(12):
                cell = ws.cell(row=rev_row, column=2 + m_idx, value=None)
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0'
        # YTD
        cell = ws.cell(row=rev_row, column=14,
                       value=f"=SUM(B{rev_row}:M{rev_row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[rev_row].height = 20

        # Expense rows (12 categories)
        for cat_idx, cat in enumerate(EXPENSE_CATEGORIES):
            row = start + 3 + cat_idx
            ws.cell(row=row, column=1, value=cat).font = Font(
                name=FONT_BODY, size=10, color=COLOR_TEXT
            )
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="left", vertical="center", indent=2
            )
            if is_demo:
                _, exp_list = ENTITY_PL_DEMO[entity_name]
                annual_exp = exp_list[cat_idx]
                monthly_exp = annual_exp / 12.0
                for m_idx in range(12):
                    cell = ws.cell(row=row, column=2 + m_idx, value=round(monthly_exp, 2))
                    apply_style(cell, input_cell_style())
                    cell.number_format = '"$"#,##0'
            else:
                for m_idx in range(12):
                    cell = ws.cell(row=row, column=2 + m_idx, value=None)
                    apply_style(cell, input_cell_style())
                    cell.number_format = '"$"#,##0'
            cell = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            ws.row_dimensions[row].height = 16

        # Total expenses row
        tot_exp_row = start + 3 + len(EXPENSE_CATEGORIES)
        ws.cell(row=tot_exp_row, column=1, value="TOTAL EXPENSES").font = Font(
            name=FONT_HEAD, size=11, bold=True, color=COLOR_ERROR
        )
        ws.cell(row=tot_exp_row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col in range(2, 15):
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=tot_exp_row, column=col,
                value=f"=SUM({col_letter}{start + 3}:{col_letter}{tot_exp_row - 1})"
            )
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ERROR)
        ws.row_dimensions[tot_exp_row].height = 20

        # NOI row
        noi_row = tot_exp_row + 1
        ws.cell(row=noi_row, column=1, value="NOI (Revenue − Expenses)").font = Font(
            name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
        )
        ws.cell(row=noi_row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col in range(2, 15):
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=noi_row, column=col,
                value=f"={col_letter}{rev_row}-{col_letter}{tot_exp_row}"
            )
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[noi_row].height = 22

        ws.conditional_formatting.add(
            f"B{noi_row}:N{noi_row}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill("solid", fgColor="FFCCCC")),
        )
        ws.conditional_formatting.add(
            f"B{noi_row}:N{noi_row}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill("solid", fgColor="C7EFCF")),
        )

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# Helpers used by Consolidated P&L to know per-section row offsets.
SECTION_BLOCK = 22
BASE_ROW = 6

def _section_rev_row(section_idx):
    return BASE_ROW + section_idx * SECTION_BLOCK + 2

def _section_first_exp_row(section_idx):
    return BASE_ROW + section_idx * SECTION_BLOCK + 3

def _section_last_exp_row(section_idx):
    return _section_first_exp_row(section_idx) + len(EXPENSE_CATEGORIES) - 1

def _section_tot_exp_row(section_idx):
    return _section_first_exp_row(section_idx) + len(EXPENSE_CATEGORIES)

def _section_noi_row(section_idx):
    return _section_tot_exp_row(section_idx) + 1


# ---------------------------------------------------------------------------
# Sheet 4 — Intercompany Eliminations
# ---------------------------------------------------------------------------

def build_intercompany_tab(wb, variant):
    ws = wb.create_sheet("Intercompany")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Intercompany",
                        prev_tab="Per-Entity P&L", next_tab="Consolidated P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        f"Eliminations register (capacity {INTERCOMPANY_CAPACITY}). "
        "Management fees, shared-service allocations, inter-LLC loans. "
        "'Eliminated at consolidation?' = Yes removes the entry from the Consolidated P&L."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 30

    headers = [
        "Date", "From entity", "To entity", "Description",
        "Amount", "Type", "Eliminated?",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 12), ("B", 22), ("C", 22), ("D", 26),
        ("E", 12), ("F", 18), ("G", 14),
    ])

    samples = _val_list(variant, INTERCO_DEMO)
    for i in range(INTERCOMPANY_CAPACITY):
        row = 6 + i
        if i < len(samples):
            (date_s, frm, to, desc, amt, typ, elim) = samples[i]
            cells = [
                (1, _parse_date(date_s), "yyyy-mm-dd"),
                (2, frm, None),
                (3, to, None),
                (4, desc, None),
                (5, amt, '"$"#,##0.00'),
                (6, typ, None),
                (7, elim, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None),
                (3, None, None),
                (4, None, None),
                (5, None, '"$"#,##0.00'),
                (6, None, None),
                (7, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 16

    last_data = 5 + INTERCOMPANY_CAPACITY
    add_dropdown(ws, f"B6:B{last_data}",
                 "='Entity Setup'!$A$6:$A$15")
    add_dropdown(ws, f"C6:C{last_data}",
                 "='Entity Setup'!$A$6:$A$15")
    add_dropdown(ws, f"F6:F{last_data}",
                 f'"{",".join(INTERCO_TYPES)}"')
    add_dropdown(ws, f"G6:G{last_data}", '"Yes,No"')

    # Summary row
    sum_row = last_data + 2
    ws.cell(row=sum_row, column=1, value="ELIMINATED TOTAL (Type=Mgmt fee, Eliminated=Yes):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.merge_cells(f"A{sum_row}:D{sum_row}")
    ws.cell(row=sum_row, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(
        row=sum_row, column=5,
        value=f'=SUMIFS(E6:E{last_data},F6:F{last_data},"Management fee",G6:G{last_data},"Yes")'
    )
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[sum_row].height = 22

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Consolidated P&L
# ---------------------------------------------------------------------------

def build_consolidated_pl_tab(wb, variant):
    """Sum of entity NOIs minus eliminations.

    Layout — all values are TAX-YEAR YTD (driven by Settings B5):
      Row 5 : header
      Row 7 : Revenue (sum of all entity revenue rows)
      Row 9 : EXPENSES banner
      Rows 10-21 : 12 categories
      Row 22 : Total expenses
      Row 23 : Pre-elimination NOI
      Row 24 : Less: intercompany eliminations
      Row 25 : CONSOLIDATED NOI
    """
    ws = wb.create_sheet("Consolidated P&L")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Consolidated P&L",
                        prev_tab="Intercompany", next_tab="Member Distributions")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Holding-company-level roll-up. Sum of entity P&Ls minus intercompany "
        "eliminations. Print-ready for CPA hand-off. "
        "Tax year: Settings → B5."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 38),
        ("B", 16),
        ("C", 4),
        ("D", 32),
    ])

    # Header row 5
    cell = ws.cell(row=5, column=1, value="Category")
    apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=2, value="Consolidated YTD")
    apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # Tax year sub-banner
    ws.merge_cells("D5:D5")
    cell = ws.cell(row=5, column=4, value="=\"For tax year \"&Settings!$B$5")
    cell.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 7: Revenue (sum of entity revenue YTDs from Per-Entity P&L)
    ws.cell(row=7, column=1, value="Revenue (rents + fees)").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    rev_terms = [f"'Per-Entity P&L'!N{_section_rev_row(s)}" for s in range(ENTITY_PL_SECTIONS)]
    cell = ws.cell(row=7, column=2, value="=" + "+".join(rev_terms))
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 24

    # Row 9: EXPENSES banner
    ws.merge_cells("A9:B9")
    c = ws["A9"]
    c.value = "EXPENSES"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[9].height = 22

    # Rows 10-21: 12 expense categories
    for cat_idx, cat in enumerate(EXPENSE_CATEGORIES):
        row = 10 + cat_idx
        ws.cell(row=row, column=1, value=cat).font = Font(
            name=FONT_BODY, size=10, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=2
        )
        # Sum the same category row across all entity sections, YTD col N
        terms = []
        for s in range(ENTITY_PL_SECTIONS):
            r = _section_first_exp_row(s) + cat_idx
            terms.append(f"'Per-Entity P&L'!N{r}")
        cell = ws.cell(row=row, column=2, value="=" + "+".join(terms))
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        ws.row_dimensions[row].height = 16

    # Row 22: Total expenses
    ws.cell(row=22, column=1, value="TOTAL EXPENSES (pre-elimination)").font = Font(
        name=FONT_HEAD, size=11, bold=True, color=COLOR_ERROR
    )
    ws.cell(row=22, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=22, column=2, value="=SUM(B10:B21)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[22].height = 22

    # Row 23: Pre-elimination NOI
    ws.cell(row=23, column=1, value="Pre-elimination NOI").font = Font(
        name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT
    )
    ws.cell(row=23, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=23, column=2, value="=B7-B22")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    ws.row_dimensions[23].height = 20

    # Row 24: Less intercompany eliminations
    # (Sum of Intercompany rows where Eliminated? = Yes)
    ws.cell(row=24, column=1, value="Less: intercompany eliminations").font = Font(
        name=FONT_BODY, size=11, italic=True, color=COLOR_SECONDARY
    )
    ws.cell(row=24, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    last_ic = 5 + INTERCOMPANY_CAPACITY
    elim_formula = (
        f'=SUMIFS(Intercompany!E6:E{last_ic},'
        f'Intercompany!G6:G{last_ic},"Yes")'
    )
    cell = ws.cell(row=24, column=2, value=elim_formula)
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_SECONDARY)
    ws.row_dimensions[24].height = 20

    # Note: eliminations affect both revenue (charging entity) and expense
    # (paying entity) by the same amount, netting to ZERO at consolidation.
    # Pre-elim NOI already nets these because both sides are summed across
    # the same categories. We surface the elimination value here for
    # transparency only — the formal CONSOLIDATED NOI equals
    # pre-elimination NOI when all intercompany activity is symmetric.
    # If the user records only one side (e.g., only the expense), they can
    # adjust manually here.

    # Row 25: CONSOLIDATED NOI
    # Display = pre-elim NOI (since symmetric intercompany activity nets to 0).
    ws.cell(row=25, column=1, value="CONSOLIDATED NOI").font = Font(
        name=FONT_HEAD, size=14, bold=True, color="F6EFE2"
    )
    ws.cell(row=25, column=1).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.cell(row=25, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=25, column=2, value="=B23")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color="F6EFE2")
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[25].height = 30

    ws.conditional_formatting.add(
        "B25",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="FFCCCC")),
    )

    # Right-side note column D
    ws.merge_cells("D7:D25")
    c = ws["D7"]
    c.value = (
        "READ-OUT NOTES\n\n"
        "• Pre-elim NOI sums every entity's net.\n"
        "• Symmetric intercompany activity (mgmt fee charged AND paid) nets "
        "to zero at consolidation.\n"
        "• If you only recorded one side of an intercompany, adjust the "
        "Eliminations row manually.\n"
        "• Print this tab + the K-1 Worksheet for your CPA hand-off."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Member Distributions
# ---------------------------------------------------------------------------

def build_member_distributions_tab(wb, variant):
    """12-col matrix: rows = members (capacity 10), cols = months."""
    ws = wb.create_sheet("Member Distributions")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Member Distributions",
                        prev_tab="Consolidated P&L", next_tab="K-1 Worksheet")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 15):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:N4")
    c = ws["A4"]
    c.value = (
        "Cash distributions to members across all entities. "
        "Rows pull from Settings → Member List. YTD totals roll up to Start tab."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    set_col_widths(ws, [
        ("A", 28),
        ("B", 10), ("C", 10), ("D", 10), ("E", 10), ("F", 10), ("G", 10),
        ("H", 10), ("I", 10), ("J", 10), ("K", 10), ("L", 10), ("M", 10),
        ("N", 14),
    ])

    cell = ws.cell(row=5, column=1, value="Member")
    apply_style(cell, header_row_style())
    for i, m in enumerate(months):
        cell = ws.cell(row=5, column=2 + i, value=m)
        apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=14, value="YTD")
    apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 22

    MEMBER_CAP = 10
    demo_names = list(DISTRIBUTIONS_DEMO.keys())

    for i in range(MEMBER_CAP):
        row = 6 + i
        # Member name pulled from Settings via formula
        # Members live on Settings col B rows 18-27 (per brief)
        cell = ws.cell(row=row, column=1, value=f"=IFERROR(Settings!B{18 + i},\"\")")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

        # Demo distribution values
        if variant == "demo" and i < len(demo_names):
            vals = DISTRIBUTIONS_DEMO[demo_names[i]]
        else:
            vals = [None] * 12
        for m_idx in range(12):
            cell = ws.cell(row=row, column=2 + m_idx, value=vals[m_idx])
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'

        # YTD
        cell = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 18

    # Total distributions row 17 (under 10 members ending at row 15 — adjust)
    total_row = 6 + MEMBER_CAP + 1  # row 17
    ws.cell(row=total_row, column=1, value="TOTAL DISTRIBUTIONS").font = Font(
        name=FONT_HEAD, size=12, bold=True, color="F6EFE2"
    )
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.cell(row=total_row, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=total_row, column=col,
            value=f"=SUM({col_letter}6:{col_letter}{6 + MEMBER_CAP - 1})"
        )
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    ws.row_dimensions[total_row].height = 24

    # Note: Start tab references N12 (which is row 12, member row #7).
    # Override that by emitting a "YTD total alias" at N12 for the Start KPI.
    alias_cell = ws.cell(row=12, column=14)
    # If alias falls inside the member YTD rows, it's already a SUM — fine.
    # The Start KPI grabs N12 expecting it to be the total. To be safe,
    # explicitly ensure Start references the total_row instead. We can't
    # edit Start here — but the build_start_tab formula points to N12.
    # Update that formula post-hoc to point to the total row.
    # (Simpler: write the alias in N12 from total_row.)
    # We'll let Start's reference target the total row directly via update below.

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return total_row


# ---------------------------------------------------------------------------
# Sheet 7 — K-1 Worksheet
# ---------------------------------------------------------------------------

def build_k1_worksheet_tab(wb, variant):
    """Per-member K-1 figures pulled from entity NOI × member %.

    Layout:
      Rows 1-3 : compact_header_band
      Row 5    : DISCLAIMER BANNER (gold-soft, CPA-handoff)
      Row 7    : header (Member | Entity | Ownership % | Ordinary income |
                  Distributions | Capital)
      Rows 8+  : per-member-per-entity allocation rows
    """
    ws = wb.create_sheet("K-1 Worksheet")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "K-1 Worksheet",
                        prev_tab="Member Distributions", next_tab="Settings")

    # CPA-HANDOFF DISCLAIMER BANNER (row 4-5)
    ws.merge_cells("A4:H5")
    c = ws["A4"]
    c.value = (
        "  CPA HAND-OFF NOTICE — These are working figures, NOT a filed K-1. "
        "Schedule K-1 (Form 1065) preparation is your CPA's job. The numbers "
        "below show your share of each LLC's ordinary income and distributions "
        "for the active tax year on Settings → B5. Hand this tab to your CPA."
    )
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    # Gold border around the disclaimer
    gold_side = Side(style="medium", color=COLOR_ACCENT)
    for col in range(1, 9):
        for row in range(4, 6):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=gold_side if row == 4 else existing.top,
                bottom=gold_side if row == 5 else existing.bottom,
                left=gold_side if col == 1 else existing.left,
                right=gold_side if col == 8 else existing.right,
            )
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 26

    # Sub-instruction
    ws.merge_cells("A6:H6")
    c = ws["A6"]
    c.value = (
        "Tax year: " + (str(2026) if variant == "demo" else "(see Settings)") +
        "  ·  Ordinary income = Member % × Entity NOI  ·  Distributions pulled from Member Distributions tab"
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[6].height = 18

    headers = [
        "Member", "Entity", "Ownership %",
        "Entity NOI (YTD)", "Allocated ordinary income", "Distributions YTD",
        "Capital balance (manual)", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 24

    set_col_widths(ws, [
        ("A", 22), ("B", 22), ("C", 12),
        ("D", 16), ("E", 18), ("F", 16),
        ("G", 18), ("H", 24),
    ])

    # Build per-member-per-entity rows.
    # In DEMO: iterate ENTITY_MEMBER_PCT_DEMO in entity order.
    rows = []
    if variant == "demo":
        for ent_idx, entity_name in enumerate([e[0] for e in ENTITIES_DEMO]):
            if ent_idx >= ENTITY_PL_SECTIONS:
                break
            for (member, pct) in ENTITY_MEMBER_PCT_DEMO.get(entity_name, []):
                rows.append((ent_idx, member, pct))
    # Pad with empty rows up to 20 capacity
    K1_CAP = 20

    for i in range(K1_CAP):
        row = 8 + i
        if i < len(rows):
            (ent_idx, member, pct) = rows[i]
            # Member (input)
            cell = ws.cell(row=row, column=1, value=member)
            apply_style(cell, input_cell_style())
            # Entity (input — drop-down to Entity Setup)
            cell = ws.cell(row=row, column=2,
                           value=f"='Entity Setup'!A{6 + ent_idx}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            # Ownership %
            cell = ws.cell(row=row, column=3, value=pct)
            apply_style(cell, input_cell_style())
            cell.number_format = "0%"
            # Entity NOI (formula — pulls from Per-Entity P&L)
            noi_row_in_pe = _section_noi_row(ent_idx)
            cell = ws.cell(row=row, column=4,
                           value=f"='Per-Entity P&L'!N{noi_row_in_pe}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            # Allocated ordinary income = pct × NOI
            cell = ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
            # Distributions YTD — VLOOKUP into Member Distributions
            # Member Distributions has Member name in col A, YTD in col N.
            cell = ws.cell(
                row=row, column=6,
                value=f"=IFERROR(VLOOKUP(A{row},'Member Distributions'!$A$6:$N$15,14,FALSE),0)"
            )
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            # Capital balance (manual input)
            cell = ws.cell(row=row, column=7, value=None)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
            # Notes
            cell = ws.cell(row=row, column=8, value=None)
            apply_style(cell, input_cell_style())
        else:
            # Empty row — formula scaffold
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col, value=None)
                if col in (4, 5, 6):
                    apply_style(cell, formula_cell_style())
                    cell.number_format = '"$"#,##0'
                elif col == 3:
                    apply_style(cell, input_cell_style())
                    cell.number_format = "0%"
                else:
                    apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # Member dropdown bound to Settings member list (B18:B27)
    add_dropdown(ws, f"A8:A{7 + K1_CAP}", "=Settings!$B$18:$B$27")
    # Entity dropdown
    add_dropdown(ws, f"B8:B{7 + K1_CAP}", "='Entity Setup'!$A$6:$A$15")

    # Footer disclaimer (small print)
    foot_row = 8 + K1_CAP + 2
    ws.merge_cells(f"A{foot_row}:H{foot_row + 1}")
    c = ws[f"A{foot_row}"]
    c.value = (
        "K-1 prep is the CPA's job. Working figures only — does NOT account for "
        "guaranteed payments, basis adjustments, special allocations, or "
        "passive-activity loss limitations. Provide this sheet plus the "
        "Consolidated P&L to your CPA."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[foot_row].height = 16
    ws.row_dimensions[foot_row + 1].height = 16

    ws.freeze_panes = "A8"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:7"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 8 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 22),
        ("C", 4), ("D", 18), ("E", 22),
        ("F", 22), ("G", 14),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="K-1 Worksheet", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Active tax year drives all P&L formulas. Member list feeds the "
        "K-1 Worksheet and Member Distributions dropdowns."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # ---- Active tax year (B5) ----
    ws.cell(row=5, column=1, value="Active tax year:").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=5, column=2, value=2026 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 28

    if variant != "demo":
        cell.value = "=YEAR(TODAY())"

    ws.merge_cells("D5:G5")
    c = ws["D5"]
    c.value = (
        "Type the tax year here. All consolidation tabs filter by this."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)

    # ---- Entity list mirror (B7:B16) ----
    ws.cell(row=7, column=1, value="ENTITY LIST  (mirrors Entity Setup col A)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A7:B7")
    ws.row_dimensions[7].height = 20

    for i in range(ENTITY_CAPACITY):
        row = 8 + i  # rows 8-17... but brief says B7-B16, slight shift OK
        # Mirror via formula
        cell = ws.cell(row=row, column=2, value=f"=IFERROR('Entity Setup'!A{6 + i},\"\")")
        apply_style(cell, formula_cell_style())
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT, italic=True)
        ws.row_dimensions[row].height = 16

    # ---- Member list (B21:B30) ----
    # Banner at row 19 (clear of entity list ending at row 17)
    ws.cell(row=19, column=1, value="MEMBER LIST  (Name | last4 | role)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=19, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A19:E19")
    ws.row_dimensions[19].height = 20

    # Column hints at row 20
    ws.cell(row=20, column=2, value="Name").font = Font(name=FONT_MONO, size=8, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=20, column=4, value="EIN/SSN last4").font = Font(name=FONT_MONO, size=8, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=20, column=5, value="Role").font = Font(name=FONT_MONO, size=8, bold=True, color=COLOR_PRIMARY)

    members = _val_list(variant, MEMBERS_DEMO)
    MEMBER_CAP = 10
    for i in range(MEMBER_CAP):
        row = 21 + i
        if i < len(members):
            (name, last4, role) = members[i]
        else:
            name = last4 = role = None
        cell = ws.cell(row=row, column=2, value=name)
        apply_style(cell, input_cell_style())
        cell = ws.cell(row=row, column=4, value=last4)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = ws.cell(row=row, column=5, value=role)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # Footer
    brand_footer(ws, 33,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_entity_setup_tab(wb, variant)
    build_per_entity_pl_tab(wb, variant)
    build_intercompany_tab(wb, variant)
    build_consolidated_pl_tab(wb, variant)
    distrib_total_row = build_member_distributions_tab(wb, variant)
    build_k1_worksheet_tab(wb, variant)
    build_settings_tab(wb, variant)

    # Patch Start KPI for distributions YTD to point at the actual total row
    # (build_start_tab hard-codes N12; the real total lives at N{distrib_total_row}).
    start_ws = wb["Start"]
    # The KPI formula sits in J11 (merged J11:L13).
    start_ws["J11"].value = f"='Member Distributions'!N{distrib_total_row}"

    wb.properties.title = "Multi-Entity Consolidated P&L — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Consolidated P&L for STR hosts who hold each property in its own LLC. "
        "Per-entity P&L + intercompany eliminations + member distributions + "
        "K-1 working figures for CPA hand-off."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
