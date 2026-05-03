"""Build PAM-003 Commission / Split Calculator for Co-Hosts (v2.2 standard).

Operational-mode tool for Sarah/Pam: manages someone else's STR for a
revenue share (% of NET or % of GROSS), flat per-booking fee, or hybrid.
Computes per-month commission owed, tracks payments, and produces
1099-NEC-ready year-end totals.

Generates two files:
  templates/_masters/PAM-003-commission-split-calculator-DEMO.xlsx
  templates/_masters/PAM-003-commission-split-calculator-BLANK.xlsx
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_INK, COLOR_WHITE,
)

SKU = "PAM-003"
NAME = "commission-split-calculator"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"


# ---------------------------------------------------------------------------
# Sample data — DEMO variant
# 4 properties × 2 owners (per brief)
# ---------------------------------------------------------------------------

AGREEMENT_TYPES = [
    "% of NET",
    "% of GROSS",
    "Flat per booking",
    "Hybrid (Flat + % of NET)",
]

CHANNELS = ["Airbnb", "VRBO", "Booking.com", "Direct", "Other"]
PAYMENT_STATUSES = ["Pending", "Paid", "Disputed"]

# (Property, Owner name, Owner email, Agreement type, Co-host %, Flat $,
#  Co-host pays cleaner?, Co-host pays platform fees?, Reimbursable items,
#  Effective date, Notes)
AGREEMENTS_DEMO = [
    ("Smokies Ridge", "Lakeside Holdings LLC", "owner1@lakeside.example",
     "% of NET", 0.20, 0,
     "No", "No",
     "Repairs > $200, supplies",
     "2025-01-01",
     "20% of NET. Owner pays cleaner directly via Venmo."),
    ("Lakehouse", "Lakeside Holdings LLC", "owner1@lakeside.example",
     "% of NET", 0.15, 0,
     "No", "No",
     "Repairs > $200, supplies",
     "2025-03-01",
     "15% of NET; owner pays cleaner directly."),
    ("Creek Side", "Maddox Trust", "owner2@maddox.example",
     "Flat per booking", 0.0, 75,
     "Yes", "No",
     "Reimburse cleaner from owner monthly",
     "2025-04-15",
     "$75 per booking flat; co-host pays cleaner, reimbursed monthly."),
    ("Ridge Top", "Maddox Trust", "owner2@maddox.example",
     "Hybrid (Flat + % of NET)", 0.08, 50,
     "Yes", "No",
     "Reimburse cleaner from owner monthly",
     "2025-05-01",
     "$50/booking + 8% of NET; co-host fronts cleaner cost."),
]

# Bookings — YTD 2026, ~106 bookings across 4 properties (per brief totals).
# Designed to roughly hit: Smokies Ridge $48K gross / Lakehouse $52K /
# Creek Side 24 bookings / Ridge Top 22 bookings / commission ~$19.9K.
BOOKINGS_DEMO = [
    # Smokies Ridge — 32 bookings, $48K gross, 20% NET, no cleaner deduction (owner pays)
    # date, property, guest, booking_id, channel, gross, cleaning, platform, other, status, notes
    ("2026-01-04", "Smokies Ridge", "Avery G.",     "AB-90011", "Airbnb", 1620, 150, 243, 0,  "Paid",    ""),
    ("2026-01-11", "Smokies Ridge", "Marshall H.",  "AB-90012", "Airbnb", 1480, 150, 222, 0,  "Paid",    ""),
    ("2026-01-18", "Smokies Ridge", "Priya R.",     "VR-50221", "VRBO",   1740, 150, 139, 0,  "Paid",    ""),
    ("2026-01-25", "Smokies Ridge", "Casey N.",     "DR-00031", "Direct", 1380, 150,   0, 0,  "Paid",    "Direct booking"),
    ("2026-02-01", "Smokies Ridge", "Devon S.",     "AB-90013", "Airbnb", 1340, 150, 201, 0,  "Paid",    ""),
    ("2026-02-08", "Smokies Ridge", "Hailey M.",    "AB-90014", "Airbnb", 1480, 150, 222, 0,  "Paid",    ""),
    ("2026-02-15", "Smokies Ridge", "Patrice O.",   "AB-90015", "Airbnb", 1680, 150, 252, 0,  "Paid",    ""),
    ("2026-02-22", "Smokies Ridge", "Chen W.",      "AB-90016", "Airbnb", 1540, 150, 231, 0,  "Paid",    ""),
    ("2026-03-01", "Smokies Ridge", "Tomás A.",     "VR-50222", "VRBO",   1620, 150, 130, 0,  "Paid",    ""),
    ("2026-03-08", "Smokies Ridge", "Reese F.",     "AB-90017", "Airbnb", 1880, 150, 282, 0,  "Paid",    ""),
    ("2026-03-15", "Smokies Ridge", "Olympia V.",   "DR-00032", "Direct", 1740, 150,   0, 0,  "Paid",    ""),
    ("2026-03-22", "Smokies Ridge", "Sasha B.",     "AB-90018", "Airbnb", 1640, 150, 246, 0,  "Paid",    ""),
    ("2026-03-29", "Smokies Ridge", "Wells D.",     "AB-90019", "Airbnb", 1380, 150, 207, 0,  "Paid",    ""),
    ("2026-04-05", "Smokies Ridge", "Indira K.",    "AB-90020", "Airbnb", 1740, 150, 261, 0,  "Paid",    ""),
    ("2026-04-12", "Smokies Ridge", "Brennan E.",   "AB-90021", "Airbnb", 1480, 150, 222, 0,  "Paid",    ""),
    ("2026-04-19", "Smokies Ridge", "Nadia L.",     "VR-50223", "VRBO",   1620, 150, 130, 0,  "Paid",    ""),
    ("2026-04-26", "Smokies Ridge", "Kendall U.",   "AB-90022", "Airbnb", 1380, 150, 207, 0,  "Paid",    ""),
    ("2026-05-03", "Smokies Ridge", "Ezra Y.",      "AB-90023", "Airbnb", 1480, 150, 222, 0,  "Pending", ""),
    ("2026-05-10", "Smokies Ridge", "Saoirse C.",   "AB-90024", "Airbnb", 1620, 150, 243, 0,  "Pending", ""),
    ("2026-05-17", "Smokies Ridge", "Jude T.",      "AB-90025", "Airbnb", 1480, 150, 222, 0,  "Pending", ""),
    ("2026-05-24", "Smokies Ridge", "Annika W.",    "DR-00033", "Direct", 1640, 150,   0, 0,  "Pending", ""),
    ("2026-05-31", "Smokies Ridge", "Felipe O.",    "AB-90026", "Airbnb", 1380, 150, 207, 0,  "Pending", ""),
    ("2026-06-07", "Smokies Ridge", "Jordan P.",    "AB-90027", "Airbnb", 1480, 150, 222, 0,  "Pending", ""),
    ("2026-06-14", "Smokies Ridge", "Bryn N.",      "AB-90028", "Airbnb", 1540, 150, 231, 0,  "Pending", ""),
    ("2026-06-21", "Smokies Ridge", "Mira K.",      "VR-50224", "VRBO",   1620, 150, 130, 0,  "Pending", ""),
    ("2026-06-28", "Smokies Ridge", "Roy T.",       "AB-90029", "Airbnb", 1480, 150, 222, 0,  "Pending", ""),
    ("2026-07-05", "Smokies Ridge", "Quincy J.",    "AB-90030", "Airbnb", 1620, 150, 243, 0,  "Pending", ""),
    ("2026-07-12", "Smokies Ridge", "Patrice C.",   "AB-90031", "Airbnb", 1480, 150, 222, 0,  "Pending", ""),
    ("2026-07-19", "Smokies Ridge", "Indigo S.",    "AB-90032", "Airbnb", 1380, 150, 207, 0,  "Pending", ""),
    ("2026-07-26", "Smokies Ridge", "Avi M.",       "AB-90033", "Airbnb", 1320, 150, 198, 0,  "Pending", ""),
    ("2026-08-02", "Smokies Ridge", "Pia W.",       "VR-50225", "VRBO",   1480, 150, 118, 0,  "Pending", ""),
    ("2026-08-09", "Smokies Ridge", "Halle J.",     "AB-90034", "Airbnb", 1340, 150, 201, 0,  "Pending", ""),

    # Lakehouse — 28 bookings, $52K gross, 15% NET, owner pays cleaner
    ("2026-01-06", "Lakehouse", "Kazuki H.",   "AB-91001", "Airbnb", 2080, 175, 312, 0,  "Paid",    ""),
    ("2026-01-13", "Lakehouse", "Lin Y.",      "AB-91002", "Airbnb", 1880, 175, 282, 0,  "Paid",    ""),
    ("2026-01-20", "Lakehouse", "Marcos G.",   "VR-51001", "VRBO",   2240, 175, 179, 0,  "Paid",    ""),
    ("2026-01-27", "Lakehouse", "Niamh O.",    "AB-91003", "Airbnb", 1740, 175, 261, 0,  "Paid",    ""),
    ("2026-02-03", "Lakehouse", "Otto K.",     "AB-91004", "Airbnb", 1980, 175, 297, 0,  "Paid",    ""),
    ("2026-02-10", "Lakehouse", "Pavla M.",    "DR-01001", "Direct", 1860, 175,   0, 0,  "Paid",    ""),
    ("2026-02-17", "Lakehouse", "Quill R.",    "AB-91005", "Airbnb", 2080, 175, 312, 0,  "Paid",    ""),
    ("2026-02-24", "Lakehouse", "Rosa B.",     "AB-91006", "Airbnb", 1740, 175, 261, 0,  "Paid",    ""),
    ("2026-03-03", "Lakehouse", "Soren D.",    "AB-91007", "Airbnb", 1880, 175, 282, 0,  "Paid",    ""),
    ("2026-03-10", "Lakehouse", "Talia E.",    "VR-51002", "VRBO",   2080, 175, 166, 0,  "Paid",    ""),
    ("2026-03-17", "Lakehouse", "Umar F.",     "AB-91008", "Airbnb", 1980, 175, 297, 0,  "Paid",    ""),
    ("2026-03-24", "Lakehouse", "Vera H.",     "AB-91009", "Airbnb", 1860, 175, 279, 0,  "Paid",    ""),
    ("2026-03-31", "Lakehouse", "Will I.",     "AB-91010", "Airbnb", 2240, 175, 336, 0,  "Paid",    ""),
    ("2026-04-07", "Lakehouse", "Xio L.",      "AB-91011", "Airbnb", 1880, 175, 282, 0,  "Paid",    ""),
    ("2026-04-14", "Lakehouse", "Yuki M.",     "DR-01002", "Direct", 1740, 175,   0, 0,  "Paid",    ""),
    ("2026-04-21", "Lakehouse", "Zane N.",     "AB-91012", "Airbnb", 2080, 175, 312, 0,  "Paid",    ""),
    ("2026-04-28", "Lakehouse", "Ada O.",      "AB-91013", "Airbnb", 1980, 175, 297, 0,  "Paid",    ""),
    ("2026-05-05", "Lakehouse", "Bram P.",     "AB-91014", "Airbnb", 1740, 175, 261, 0,  "Pending", ""),
    ("2026-05-12", "Lakehouse", "Cleo Q.",     "VR-51003", "VRBO",   2080, 175, 166, 0,  "Pending", ""),
    ("2026-05-19", "Lakehouse", "Dilan R.",    "AB-91015", "Airbnb", 1860, 175, 279, 0,  "Pending", ""),
    ("2026-05-26", "Lakehouse", "Esme S.",     "AB-91016", "Airbnb", 1980, 175, 297, 0,  "Pending", ""),
    ("2026-06-02", "Lakehouse", "Fiore T.",    "AB-91017", "Airbnb", 1740, 175, 261, 0,  "Pending", ""),
    ("2026-06-09", "Lakehouse", "Ghita U.",    "AB-91018", "Airbnb", 1880, 175, 282, 0,  "Pending", ""),
    ("2026-06-16", "Lakehouse", "Hadi V.",     "DR-01003", "Direct", 1620, 175,   0, 0,  "Pending", ""),
    ("2026-06-23", "Lakehouse", "Ivo W.",      "AB-91019", "Airbnb", 1740, 175, 261, 0,  "Pending", ""),
    ("2026-06-30", "Lakehouse", "Jara X.",     "AB-91020", "Airbnb", 1880, 175, 282, 0,  "Pending", ""),
    ("2026-07-07", "Lakehouse", "Kit Y.",      "AB-91021", "Airbnb", 1980, 175, 297, 0,  "Pending", ""),
    ("2026-07-14", "Lakehouse", "Lyra Z.",     "VR-51004", "VRBO",   1860, 175, 149, 0,  "Pending", ""),

    # Creek Side — 24 bookings, flat $75/booking. Co-host pays cleaner ($120 each)
    ("2026-01-08", "Creek Side", "Mara A.",   "AB-92001", "Airbnb", 1240, 120, 186, 0,  "Paid", ""),
    ("2026-01-15", "Creek Side", "Nico B.",   "AB-92002", "Airbnb", 1180, 120, 177, 0,  "Paid", ""),
    ("2026-01-22", "Creek Side", "Otis C.",   "VR-52001", "VRBO",   1340, 120, 107, 0,  "Paid", ""),
    ("2026-01-29", "Creek Side", "Pia D.",    "AB-92003", "Airbnb", 1180, 120, 177, 0,  "Paid", ""),
    ("2026-02-05", "Creek Side", "Quinn E.",  "AB-92004", "Airbnb", 1240, 120, 186, 0,  "Paid", ""),
    ("2026-02-12", "Creek Side", "Rae F.",    "DR-02001", "Direct", 1180, 120,   0, 0,  "Paid", ""),
    ("2026-02-19", "Creek Side", "Sky G.",    "AB-92005", "Airbnb", 1340, 120, 201, 0,  "Paid", ""),
    ("2026-02-26", "Creek Side", "Tobi H.",   "AB-92006", "Airbnb", 1080, 120, 162, 0,  "Paid", ""),
    ("2026-03-05", "Creek Side", "Uli I.",    "AB-92007", "Airbnb", 1240, 120, 186, 0,  "Paid", ""),
    ("2026-03-12", "Creek Side", "Vali J.",   "VR-52002", "VRBO",   1180, 120,  94, 0,  "Paid", ""),
    ("2026-03-19", "Creek Side", "Wren K.",   "AB-92008", "Airbnb", 1080, 120, 162, 0,  "Paid", ""),
    ("2026-03-26", "Creek Side", "Xavi L.",   "AB-92009", "Airbnb", 1240, 120, 186, 0,  "Paid", ""),
    ("2026-04-02", "Creek Side", "Yara M.",   "AB-92010", "Airbnb", 1180, 120, 177, 0,  "Paid", ""),
    ("2026-04-09", "Creek Side", "Zev N.",    "DR-02002", "Direct", 1340, 120,   0, 0,  "Paid", ""),
    ("2026-04-16", "Creek Side", "Aki O.",    "AB-92011", "Airbnb", 1180, 120, 177, 0,  "Paid", ""),
    ("2026-04-23", "Creek Side", "Bea P.",    "AB-92012", "Airbnb", 1080, 120, 162, 0,  "Paid", ""),
    ("2026-04-30", "Creek Side", "Cy Q.",     "AB-92013", "Airbnb", 1240, 120, 186, 0,  "Paid", ""),
    ("2026-05-07", "Creek Side", "Dee R.",    "AB-92014", "Airbnb", 1180, 120, 177, 0,  "Pending", ""),
    ("2026-05-14", "Creek Side", "Eli S.",    "VR-52003", "VRBO",   1080, 120,  86, 0,  "Pending", ""),
    ("2026-05-21", "Creek Side", "Fae T.",    "AB-92015", "Airbnb", 1240, 120, 186, 0,  "Pending", ""),
    ("2026-05-28", "Creek Side", "Gio U.",    "AB-92016", "Airbnb", 1180, 120, 177, 0,  "Pending", ""),
    ("2026-06-04", "Creek Side", "Hux V.",    "AB-92017", "Airbnb", 1080, 120, 162, 0,  "Pending", ""),
    ("2026-06-11", "Creek Side", "Ines W.",   "DR-02003", "Direct", 1240, 120,   0, 0,  "Pending", ""),
    ("2026-06-18", "Creek Side", "Jax X.",    "AB-92018", "Airbnb", 1180, 120, 177, 0,  "Pending", ""),

    # Ridge Top — 22 bookings, hybrid $50 + 8% NET. Co-host pays cleaner ($110)
    ("2026-01-09", "Ridge Top", "Kade A.",  "AB-93001", "Airbnb", 1480, 110, 222, 0, "Paid", ""),
    ("2026-01-16", "Ridge Top", "Liv B.",   "AB-93002", "Airbnb", 1620, 110, 243, 0, "Paid", ""),
    ("2026-01-23", "Ridge Top", "Mio C.",   "VR-53001", "VRBO",   1740, 110, 139, 0, "Paid", ""),
    ("2026-01-30", "Ridge Top", "Nia D.",   "AB-93003", "Airbnb", 1480, 110, 222, 0, "Paid", ""),
    ("2026-02-06", "Ridge Top", "Oren E.",  "AB-93004", "Airbnb", 1620, 110, 243, 0, "Paid", ""),
    ("2026-02-13", "Ridge Top", "Pax F.",   "DR-03001", "Direct", 1480, 110,   0, 0, "Paid", ""),
    ("2026-02-20", "Ridge Top", "Quill G.", "AB-93005", "Airbnb", 1740, 110, 261, 0, "Paid", ""),
    ("2026-02-27", "Ridge Top", "Rae H.",   "AB-93006", "Airbnb", 1620, 110, 243, 0, "Paid", ""),
    ("2026-03-06", "Ridge Top", "Sol I.",   "AB-93007", "Airbnb", 1480, 110, 222, 0, "Paid", ""),
    ("2026-03-13", "Ridge Top", "Tia J.",   "VR-53002", "VRBO",   1620, 110, 130, 0, "Paid", ""),
    ("2026-03-20", "Ridge Top", "Uri K.",   "AB-93008", "Airbnb", 1740, 110, 261, 0, "Paid", ""),
    ("2026-03-27", "Ridge Top", "Vex L.",   "AB-93009", "Airbnb", 1480, 110, 222, 0, "Paid", ""),
    ("2026-04-03", "Ridge Top", "Wyn M.",   "AB-93010", "Airbnb", 1620, 110, 243, 0, "Paid", ""),
    ("2026-04-10", "Ridge Top", "Xan N.",   "DR-03002", "Direct", 1480, 110,   0, 0, "Paid", ""),
    ("2026-04-17", "Ridge Top", "Yul O.",   "AB-93011", "Airbnb", 1740, 110, 261, 0, "Paid", ""),
    ("2026-04-24", "Ridge Top", "Zia P.",   "AB-93012", "Airbnb", 1620, 110, 243, 0, "Paid", ""),
    ("2026-05-01", "Ridge Top", "Ava Q.",   "AB-93013", "Airbnb", 1480, 110, 222, 0, "Pending", ""),
    ("2026-05-08", "Ridge Top", "Ben R.",   "VR-53003", "VRBO",   1620, 110, 130, 0, "Pending", ""),
    ("2026-05-15", "Ridge Top", "Cyd S.",   "AB-93014", "Airbnb", 1480, 110, 222, 0, "Pending", ""),
    ("2026-05-22", "Ridge Top", "Dax T.",   "AB-93015", "Airbnb", 1740, 110, 261, 0, "Pending", ""),
    ("2026-05-29", "Ridge Top", "Eve U.",   "AB-93016", "Airbnb", 1620, 110, 243, 0, "Pending", ""),
    ("2026-06-05", "Ridge Top", "Fox V.",   "AB-93017", "Airbnb", 1480, 110, 222, 0, "Pending", ""),
]

PROPERTIES_DEMO = ["Smokies Ridge", "Lakehouse", "Creek Side", "Ridge Top"]
OWNERS_DEMO = ["Lakeside Holdings LLC", "Maddox Trust"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list):
    return demo_list if variant == "demo" else []


def _parse_date(s):
    if not s:
        return None
    return date.fromisoformat(s)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, max_col=12):
    last = get_column_letter(max_col)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Commission / Split Calculator"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Co-host commissions, owed clearly — every booking, every month."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    cards = [
        ("A", "C", "COMMISSION YTD",
         '=SUMPRODUCT((YEAR(\'Bookings Log\'!$A$6:$A$505)=Settings!$B$5)*'
         '\'Bookings Log\'!$L$6:$L$505)',
         '"$"#,##0', COLOR_PRIMARY, "earned this year"),
        ("D", "F", "PROPERTIES MANAGED",
         '=COUNTA(\'Co-Host Agreements\'!A6:A20)',
         "0", COLOR_ACCENT, "active co-host agreements"),
        ("G", "I", "OWNERS PAID",
         '=SUMPRODUCT((\'Bookings Log\'!$N$6:$N$505="Paid")*'
         '(YEAR(\'Bookings Log\'!$A$6:$A$505)=Settings!$B$5)*'
         '\'Bookings Log\'!$M$6:$M$505)',
         '"$"#,##0', COLOR_PRIMARY, "owner share, YTD paid"),
        ("J", "L", "TOP EARNER",
         '=IFERROR(INDEX(\'Year-End Summary\'!$A$8:$A$22,'
         'MATCH(MAX(\'Year-End Summary\'!$F$8:$F$22),'
         '\'Year-End Summary\'!$F$8:$F$22,0)),"—")',
         "@", COLOR_SECONDARY, "highest commission property"),
    ]
    for fc, lc, label, formula, fmt, color, sub in cards:
        ws.merge_cells(f"{fc}10:{lc}10")
        c = ws[f"{fc}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=8, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{fc}11:{lc}13")
        c = ws[f"{fc}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=22, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        ws.merge_cells(f"{fc}14:{lc}15")
        c = ws[f"{fc}14"]
        c.value = sub
        c.font = Font(name=FONT_BODY, size=8, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for fc, lc, *_ in cards:
        first_c = column_index_from_string(fc)
        last_c = column_index_from_string(lc)
        for r in range(10, 16):
            for col in range(first_c, last_c + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == first_c else existing.left,
                    right=gold_side if col == last_c else existing.right,
                )

    # Workflow callout rows 17-22
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "MONTHLY WORKFLOW"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    workflow_steps = [
        "1. Confirm each property's split agreement on Co-Host Agreements (% of NET / GROSS / Flat / Hybrid).",
        "2. Log every booking on Bookings Log — gross, cleaning, platform fee. Commission auto-computes.",
        "3. Open Monthly Statements → pick property + month → review the printable owner statement.",
        "4. Mark each booking Paid once you receive your commission. Pending > 30 days will highlight.",
        "5. At year-end, open Year-End Summary — properties at >= $600 commission flag for 1099-NEC.",
    ]
    for i, step in enumerate(workflow_steps):
        row = 18 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = step
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22

    # Pseudo-button nav rows 25-27 + 29-31
    pseudo_button(ws, "A25", "C27", "Co-Host Agreements",
                  "'Co-Host Agreements'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "-> Bookings Log",
                  "'Bookings Log'!A1", variant="accent")
    pseudo_button(ws, "G25", "I27", "Monthly Statements",
                  "'Monthly Statements'!A1", variant="primary")
    pseudo_button(ws, "J25", "L27", "Year-End Summary",
                  "'Year-End Summary'!A1", variant="primary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "L31", "Settings  -  property list, owner list, active year",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "Bundle: Owner Reporting + Commission Split + Cleaner Tracker = "
        f"the Pro Co-Host Toolkit at {BRAND_DOMAIN}/pro-cohost."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} - v1.0 - {variant.upper()} - Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Co-Host Agreements
# ---------------------------------------------------------------------------

def build_agreements_tab(wb, variant):
    ws = wb.create_sheet("Co-Host Agreements")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Co-Host Agreements",
                        prev_tab="Start", next_tab="Bookings Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per property arrangement. Capacity 15. Commission formulas on "
        "Bookings Log VLOOKUP into this table by Property name."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Property", "Owner name", "Owner email",
        "Agreement type", "Co-host %", "Flat $/booking",
        "Co-host pays cleaner?", "Co-host pays platform fees?",
        "Reimbursable items", "Effective date", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 22), ("B", 26), ("C", 26),
        ("D", 22), ("E", 12), ("F", 14),
        ("G", 14), ("H", 14),
        ("I", 28), ("J", 14), ("K", 32),
    ])

    samples = _val_list(variant, AGREEMENTS_DEMO)
    for i in range(15):
        row = 6 + i
        if i < len(samples):
            (prop, owner, email, atype, pct, flat,
             pays_clean, pays_plat, reimb, eff_date, notes) = samples[i]
            cells = [
                (1, prop, None),
                (2, owner, None),
                (3, email, None),
                (4, atype, None),
                (5, pct, "0.0%"),
                (6, flat, '"$"#,##0'),
                (7, pays_clean, None),
                (8, pays_plat, None),
                (9, reimb, None),
                (10, _parse_date(eff_date) if eff_date else None, "yyyy-mm-dd"),
                (11, notes, None),
            ]
        else:
            cells = [
                (1, None, None), (2, None, None), (3, None, None),
                (4, None, None), (5, None, "0.0%"), (6, None, '"$"#,##0'),
                (7, None, None), (8, None, None), (9, None, None),
                (10, None, "yyyy-mm-dd"), (11, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    add_dropdown(ws, "A6:A20", "=Settings!$B$7:$B$16")
    add_dropdown(ws, "B6:B20", "=Settings!$B$18:$B$27")
    add_dropdown(ws, "D6:D20", f'"{",".join(AGREEMENT_TYPES)}"')
    add_dropdown(ws, "G6:G20", '"Yes,No"')
    add_dropdown(ws, "H6:H20", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Bookings Log
# ---------------------------------------------------------------------------

def build_bookings_log_tab(wb, variant):
    ws = wb.create_sheet("Bookings Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Bookings Log",
                        prev_tab="Co-Host Agreements", next_tab="Monthly Statements")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "All bookings, all properties. Capacity 500. NET = Gross - Cleaning - "
        "Platform - Other. Commission auto-computes from agreement type."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Booking date", "Property", "Guest", "Booking ID", "Channel",
        "Gross revenue", "Cleaning fee", "Platform fee", "Other deductions",
        "NET to property", "Co-host %", "Co-host commission $",
        "Owner share $", "Payment status", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 12), ("B", 18), ("C", 18), ("D", 12), ("E", 11),
        ("F", 12), ("G", 11), ("H", 11), ("I", 11),
        ("J", 12), ("K", 10), ("L", 14),
        ("M", 12), ("N", 12), ("O", 24),
    ])

    samples = _val_list(variant, BOOKINGS_DEMO)
    for i in range(500):
        row = 6 + i
        if i < len(samples):
            (date_s, prop, guest, bid, channel, gross, clean, plat, other, status, notes) = samples[i]
            input_cells = [
                (1, _parse_date(date_s), "yyyy-mm-dd"),
                (2, prop, None),
                (3, guest, None),
                (4, bid, None),
                (5, channel, None),
                (6, gross, '"$"#,##0.00'),
                (7, clean, '"$"#,##0.00'),
                (8, plat, '"$"#,##0.00'),
                (9, other, '"$"#,##0.00'),
                (14, status, None),
                (15, notes, None),
            ]
        else:
            input_cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None), (3, None, None), (4, None, None), (5, None, None),
                (6, None, '"$"#,##0.00'), (7, None, '"$"#,##0.00'),
                (8, None, '"$"#,##0.00'), (9, None, '"$"#,##0.00'),
                (14, None, None), (15, None, None),
            ]
        for col, val, fmt in input_cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt

        # Col J — NET to property = Gross - Cleaning - Platform - Other
        # (cleaning subtracted from NET only when the co-host pays cleaner;
        #  if owner pays cleaner directly, cleaning shouldn't reduce NET.
        #  We compute a lookup-aware NET via the agreement.)
        net_formula = (
            f'=IF(F{row}="","",'
            f'F{row}'
            f'-IF(IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,7,FALSE),"No")="Yes",G{row},0)'
            f'-IF(IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,8,FALSE),"No")="Yes",H{row},0)'
            f'-N(I{row}))'
        )
        net_cell = ws.cell(row=row, column=10, value=net_formula)
        apply_style(net_cell, formula_cell_style())
        net_cell.number_format = '"$"#,##0.00'

        # Col K — Co-host % (lookup from Agreements)
        pct_formula = (
            f'=IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,5,FALSE),0)'
        )
        pct_cell = ws.cell(row=row, column=11, value=pct_formula)
        apply_style(pct_cell, formula_cell_style())
        pct_cell.number_format = "0.0%"

        # Col L — Co-host commission $. Branch on agreement type.
        # % of NET   -> J * K
        # % of GROSS -> F * K
        # Flat       -> Flat $
        # Hybrid     -> Flat $ + (J * K)
        commission_formula = (
            f'=IF(B{row}="","",'
            f'IFS('
            f'IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,4,FALSE),"")="% of NET",J{row}*K{row},'
            f'IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,4,FALSE),"")="% of GROSS",F{row}*K{row},'
            f'IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,4,FALSE),"")="Flat per booking",IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,6,FALSE),0),'
            f'IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,4,FALSE),"")="Hybrid (Flat + % of NET)",IFERROR(VLOOKUP(B{row},\'Co-Host Agreements\'!$A$6:$K$20,6,FALSE),0)+J{row}*K{row},'
            f'TRUE,0))'
        )
        com_cell = ws.cell(row=row, column=12, value=commission_formula)
        apply_style(com_cell, formula_cell_style())
        com_cell.number_format = '"$"#,##0.00'

        # Col M — Owner share $ = NET - commission
        owner_formula = f'=IF(B{row}="","",J{row}-L{row})'
        own_cell = ws.cell(row=row, column=13, value=owner_formula)
        apply_style(own_cell, formula_cell_style())
        own_cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[row].height = 16

    add_dropdown(ws, "B6:B505", "='Co-Host Agreements'!$A$6:$A$20")
    add_dropdown(ws, "E6:E505", f'"{",".join(CHANNELS)}"')
    add_dropdown(ws, "N6:N505", f'"{",".join(PAYMENT_STATUSES)}"')

    # Conditional format: highlight row when Pending > 30 days
    pending_rule = FormulaRule(
        formula=[f'AND($N6="Pending",ISNUMBER($A6),TODAY()-$A6>30)'],
        fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
        font=Font(name=FONT_BODY, size=10, color=COLOR_ERROR, bold=True),
    )
    ws.conditional_formatting.add("A6:O505", pending_rule)

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Monthly Statements (print-ready, per-property)
# ---------------------------------------------------------------------------

def build_monthly_statement_tab(wb, variant):
    ws = wb.create_sheet("Monthly Statements")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 22), ("C", 14), ("D", 14),
        ("E", 14), ("F", 14), ("G", 14), ("H", 4),
    ])

    # Tiny back link in row 1 (NOT in print area)
    pseudo_button(ws, "B1", "C1", "<- back to dashboard",
                  "'Start'!A1", variant="secondary")
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 12

    # Title block rows 3-5
    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = "MONTHLY COMMISSION STATEMENT"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_INK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 32

    rule_side = Side(style="thin", color=COLOR_ACCENT)
    for col in range(2, 8):
        ws.cell(row=4, column=col).border = Border(top=rule_side)
    ws.row_dimensions[4].height = 6

    # Pickers rows 5-7
    ws.cell(row=6, column=2, value="Property:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=6, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells("C6:G6")
    cell = ws["C6"]
    cell.value = "Smokies Ridge" if variant == "demo" else None
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    add_dropdown(ws, "C6", "='Co-Host Agreements'!$A$6:$A$20")
    ws.row_dimensions[6].height = 26

    ws.cell(row=7, column=2, value="Period:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=7, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=7, column=3, value=3 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    add_dropdown(ws, "C7", '"1,2,3,4,5,6,7,8,9,10,11,12"')

    ws.cell(row=7, column=4, value="month").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )

    cell = ws.cell(row=7, column=5, value="=Settings!B5")
    apply_style(cell, formula_cell_style())
    cell.number_format = "0"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

    ws.cell(row=7, column=6, value="year (Settings)").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.row_dimensions[7].height = 24

    # Owner contact row 9
    ws.cell(row=9, column=2, value="Owner:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=9, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells("C9:G9")
    cell = ws["C9"]
    cell.value = '=IFERROR(VLOOKUP($C$6,\'Co-Host Agreements\'!$A$6:$K$20,2,FALSE)&"  -  "&VLOOKUP($C$6,\'Co-Host Agreements\'!$A$6:$K$20,3,FALSE),"")'
    cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[9].height = 20

    # Agreement summary row 10
    ws.cell(row=10, column=2, value="Agreement:").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=10, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells("C10:G10")
    cell = ws["C10"]
    cell.value = '=IFERROR(VLOOKUP($C$6,\'Co-Host Agreements\'!$A$6:$K$20,4,FALSE),"")'
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[10].height = 18

    # BOOKINGS section header row 12
    ws.merge_cells("B12:G12")
    c = ws["B12"]
    c.value = "BOOKINGS THIS MONTH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_INK)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[12].height = 22

    # Sub-header row 13
    sub_headers = [
        ("B13", "Date"), ("C13", "Guest"), ("D13", "Gross"),
        ("E13", "Deductions"), ("F13", "NET"), ("G13", "Commission"),
    ]
    for ref, lbl in sub_headers:
        cell = ws[ref]
        cell.value = lbl
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_INK)
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=COLOR_MUTED))
    ws.row_dimensions[13].height = 22

    # Booking rows 14-28 (15 visible — show first 15 matching rows from Bookings Log)
    # Use SMALL+IF pattern to surface only matching bookings, sorted by date.
    for i in range(15):
        row = 14 + i
        rank = i + 1

        # Hidden index of the i-th matching booking row (1-based offset within Bookings Log A6:A505)
        # We use SMALL((match-array)*ROW, n)
        idx_formula = (
            f'=IFERROR(SMALL(IF('
            f'(MONTH(\'Bookings Log\'!$A$6:$A$505)=$C$7)*'
            f'(YEAR(\'Bookings Log\'!$A$6:$A$505)=$E$7)*'
            f'(\'Bookings Log\'!$B$6:$B$505=$C$6),'
            f'ROW(\'Bookings Log\'!$A$6:$A$505)-5),{rank}),"")'
        )

        # Col B - date (writable cell only since not merged)
        date_cell = ws.cell(row=row, column=2, value=(
            f'=IFERROR(INDEX(\'Bookings Log\'!$A$6:$A$505,{idx_formula[1:]}),"")'
        ))
        date_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        date_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        date_cell.number_format = "yyyy-mm-dd"

        # Col C - guest
        guest_cell = ws.cell(row=row, column=3, value=(
            f'=IFERROR(INDEX(\'Bookings Log\'!$C$6:$C$505,{idx_formula[1:]}),"")'
        ))
        guest_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        guest_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col D - gross
        gross_cell = ws.cell(row=row, column=4, value=(
            f'=IFERROR(INDEX(\'Bookings Log\'!$F$6:$F$505,{idx_formula[1:]}),"")'
        ))
        gross_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        gross_cell.alignment = Alignment(horizontal="right", vertical="center")
        gross_cell.number_format = '"$"#,##0.00'

        # Col E - total deductions = gross - NET
        ded_cell = ws.cell(row=row, column=5, value=(
            f'=IFERROR(INDEX(\'Bookings Log\'!$F$6:$F$505,{idx_formula[1:]})-INDEX(\'Bookings Log\'!$J$6:$J$505,{idx_formula[1:]}),"")'
        ))
        ded_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_ERROR)
        ded_cell.alignment = Alignment(horizontal="right", vertical="center")
        ded_cell.number_format = '"$"#,##0.00'

        # Col F - NET
        net_cell = ws.cell(row=row, column=6, value=(
            f'=IFERROR(INDEX(\'Bookings Log\'!$J$6:$J$505,{idx_formula[1:]}),"")'
        ))
        net_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        net_cell.alignment = Alignment(horizontal="right", vertical="center")
        net_cell.number_format = '"$"#,##0.00'

        # Col G - commission
        com_cell = ws.cell(row=row, column=7, value=(
            f'=IFERROR(INDEX(\'Bookings Log\'!$L$6:$L$505,{idx_formula[1:]}),"")'
        ))
        com_cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_INK)
        com_cell.alignment = Alignment(horizontal="right", vertical="center")
        com_cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[row].height = 16

    # Totals row 30
    ws.cell(row=30, column=2, value="Subtotals:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_INK
    )
    ws.cell(row=30, column=2).alignment = Alignment(horizontal="right", vertical="center")

    # Use SUMPRODUCT against the source log (more accurate than SUM of visible rows)
    for col_letter, src_letter, label, color in [
        ("D", "F", "gross",   COLOR_TEXT),
        ("E", None, None,     COLOR_ERROR),  # = D - F
        ("F", "J", "net",     COLOR_TEXT),
        ("G", "L", "commission", COLOR_INK),
    ]:
        if src_letter:
            f = (
                f'=SUMPRODUCT('
                f'(MONTH(\'Bookings Log\'!$A$6:$A$505)=$C$7)*'
                f'(YEAR(\'Bookings Log\'!$A$6:$A$505)=$E$7)*'
                f'(\'Bookings Log\'!$B$6:$B$505=$C$6)*'
                f'\'Bookings Log\'!${src_letter}$6:${src_letter}$505)'
            )
        else:
            # Deductions = gross subtotal - net subtotal
            f = f'=D30-F30'
        cell = ws.cell(row=30, column=column_index_from_string(col_letter), value=f)
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[30].height = 24

    # Commission detail row 32-34
    ws.merge_cells("B32:G32")
    c = ws["B32"]
    c.value = "COMMISSION DETAIL"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_INK)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[32].height = 22

    ws.merge_cells("B33:F33")
    c = ws["B33"]
    c.value = '="Per agreement: "&IFERROR(VLOOKUP($C$6,\'Co-Host Agreements\'!$A$6:$K$20,4,FALSE),"")&"   |   "&TEXT(IFERROR(VLOOKUP($C$6,\'Co-Host Agreements\'!$A$6:$K$20,5,FALSE),0),"0.0%")&" + $"&TEXT(IFERROR(VLOOKUP($C$6,\'Co-Host Agreements\'!$A$6:$K$20,6,FALSE),0),"0")&" flat per booking"'
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[33].height = 18

    # FINAL ROW row 35-37
    rule_side_med = Side(style="medium", color=COLOR_INK)
    for col in range(2, 8):
        ws.cell(row=35, column=col).border = Border(top=rule_side_med)
    ws.row_dimensions[35].height = 6

    ws.merge_cells("B36:E36")
    c = ws["B36"]
    c.value = "TOTAL TO CO-HOST"
    c.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_INK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=36, column=7, value="=G30")
    cell.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_INK)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[36].height = 28

    ws.merge_cells("B37:E37")
    c = ws["B37"]
    c.value = "TOTAL TO OWNER"
    c.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_INK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=37, column=7, value="=F30-G30")
    cell.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[37].height = 28

    # Payment instructions rows 39-41
    ws.merge_cells("B39:G39")
    c = ws["B39"]
    c.value = "PAYMENT INSTRUCTIONS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_INK)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[39].height = 18

    ws.merge_cells("B40:G41")
    c = ws["B40"]
    c.value = (
        "Please remit the TOTAL TO CO-HOST amount above via your preferred "
        "payment method by the 10th of the following month. Receipts and "
        "supporting documentation available on request."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(40, 42):
        ws.row_dimensions[r].height = 16

    # Footer rows 43-44
    ws.merge_cells("B43:G43")
    c = ws["B43"]
    c.value = '="Generated "&TEXT(TODAY(),"mmmm d, yyyy")'
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[43].height = 16

    # Print area: rows 3-44 (skip back-link nav row)
    ws.print_area = "A3:H44"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Year-End Summary
# ---------------------------------------------------------------------------

def build_year_end_tab(wb, variant):
    ws = wb.create_sheet("Year-End Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Year-End Summary",
                        prev_tab="Monthly Statements", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-property totals for the active tax year (set on Settings B5). "
        "Properties at >= 1099 threshold flag for 1099-NEC issuance."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Year display row 5
    ws.cell(row=5, column=1, value="Tax year (from Settings):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=5, column=2, value="=Settings!B5")
    apply_style(cell, formula_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 22), ("B", 22),
        ("C", 12), ("D", 14), ("E", 14),
        ("F", 14), ("G", 14), ("H", 18),
    ])

    # Headers row 7
    headers = [
        "Property", "Owner", "Bookings", "Total gross",
        "Total deductions", "Total NET", "Commission paid", "1099-NEC required?",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 28

    # Pull from Co-Host Agreements rows (15 properties)
    for i in range(15):
        row = 8 + i
        ag_row = 6 + i

        # Col A — property
        cell = ws.cell(row=row, column=1, value=(
            f'=IF(\'Co-Host Agreements\'!A{ag_row}="","",\'Co-Host Agreements\'!A{ag_row})'
        ))
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col B — owner
        cell = ws.cell(row=row, column=2, value=(
            f'=IF(A{row}="","",\'Co-Host Agreements\'!B{ag_row})'
        ))
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col C — # bookings (active year)
        cell = ws.cell(row=row, column=3, value=(
            f'=IF(A{row}="","",SUMPRODUCT('
            f'(YEAR(\'Bookings Log\'!$A$6:$A$505)=$B$5)*'
            f'(\'Bookings Log\'!$B$6:$B$505=A{row})*1))'
        ))
        apply_style(cell, formula_cell_style())
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Col D — total gross
        cell = ws.cell(row=row, column=4, value=(
            f'=IF(A{row}="","",SUMPRODUCT('
            f'(YEAR(\'Bookings Log\'!$A$6:$A$505)=$B$5)*'
            f'(\'Bookings Log\'!$B$6:$B$505=A{row})*'
            f'\'Bookings Log\'!$F$6:$F$505))'
        ))
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0.00'

        # Col E — total deductions = gross - NET
        cell = ws.cell(row=row, column=5, value=(
            f'=IF(A{row}="","",D{row}-F{row})'
        ))
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_ERROR)

        # Col F — total NET
        cell = ws.cell(row=row, column=6, value=(
            f'=IF(A{row}="","",SUMPRODUCT('
            f'(YEAR(\'Bookings Log\'!$A$6:$A$505)=$B$5)*'
            f'(\'Bookings Log\'!$B$6:$B$505=A{row})*'
            f'\'Bookings Log\'!$J$6:$J$505))'
        ))
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0.00'

        # Col G — commission paid to co-host
        cell = ws.cell(row=row, column=7, value=(
            f'=IF(A{row}="","",SUMPRODUCT('
            f'(YEAR(\'Bookings Log\'!$A$6:$A$505)=$B$5)*'
            f'(\'Bookings Log\'!$B$6:$B$505=A{row})*'
            f'\'Bookings Log\'!$L$6:$L$505))'
        ))
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # Col H — 1099-NEC required?
        cell = ws.cell(row=row, column=8, value=(
            f'=IF(A{row}="","",IF(G{row}>=Settings!$B$29,"Yes","No"))'
        ))
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        ws.row_dimensions[row].height = 20

    # Totals row 24
    ws.cell(row=24, column=1, value="TOTALS").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_INK
    )
    ws.cell(row=24, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col_letter, color in [("C", COLOR_INK), ("D", COLOR_TEXT),
                               ("E", COLOR_ERROR), ("F", COLOR_TEXT),
                               ("G", COLOR_PRIMARY)]:
        cell = ws.cell(
            row=24, column=column_index_from_string(col_letter),
            value=f"=SUM({col_letter}8:{col_letter}22)"
        )
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if col_letter == "C":
            cell.number_format = "0"
        else:
            cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[24].height = 26

    # Conditional formatting on col H — gold-soft when Yes
    ws.conditional_formatting.add(
        "H8:H22",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )

    # Footer note rows 26-29
    ws.merge_cells("A26:H29")
    c = ws["A26"]
    c.value = (
        "1099-NEC reminder: when a payor (the property owner / LLC) pays a "
        "non-corporate co-host >= $600 in commissions during the tax year, "
        "the payor must issue Form 1099-NEC by January 31. The threshold lives "
        "on Settings B29. Adjust if the IRS changes it. Consult your CPA -- "
        "this worksheet is a prep tool, not legal advice."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(26, 30):
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A8"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:7"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 28),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Year-End Summary", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Active tax year drives every YTD formula. Property + Owner lists drive "
        "every dropdown across the workbook."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Active tax year — row 5
    ws.cell(row=5, column=1, value="Active tax year:").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=5, column=2, value=2026 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if variant != "demo":
        cell.value = "=YEAR(TODAY())"
    ws.row_dimensions[5].height = 32

    # PROPERTY LIST rows 7-16
    _section_band(ws, 6, "PROPERTY LIST  (B7:B16, drives dropdowns)")
    properties = PROPERTIES_DEMO if variant == "demo" else []
    for i in range(10):
        row = 7 + i
        val = properties[i] if i < len(properties) else None
        cell = ws.cell(row=row, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # OWNER LIST rows 18-27
    _section_band(ws, 17, "OWNER LIST  (B18:B27, drives dropdowns)")
    owners = OWNERS_DEMO if variant == "demo" else []
    for i in range(10):
        row = 18 + i
        val = owners[i] if i < len(owners) else None
        cell = ws.cell(row=row, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # 1099 threshold row 29
    _section_band(ws, 28, "1099-NEC THRESHOLD")
    ws.cell(row=29, column=1, value="1099 threshold ($):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=29, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=29, column=2, value=600)
    apply_style(cell, input_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[29].height = 22

    # Note rows 31-33
    ws.merge_cells("A31:C33")
    c = ws["A31"]
    c.value = (
        "Tip: type each property exactly as you'll use it in the Bookings Log "
        "(spelling and capitalization must match). The agreement type, %, and "
        "flat amount are configured on the Co-Host Agreements tab."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(31, 34):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 35,
                 version_line=f"{SKU} - v1.0 - {variant.upper()} - Free updates forever")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_agreements_tab(wb, variant)
    build_bookings_log_tab(wb, variant)
    build_monthly_statement_tab(wb, variant)
    build_year_end_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Commission / Split Calculator -- The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Co-host commission and split calculator for STR property managers."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
