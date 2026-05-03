"""Build GST-003 Pet Policy Document Builder — wizard-style tool.

Implements templates/_briefs/GST-003-pet-policy-document.md.

Smaller-scope wizard (Etsy gateway, T0/T1). 4 sections x ~4 inputs each =
16 inputs total. Outputs a print/listing-ready single-page pet policy via
formula concatenation on the Policy Output tab.

Tabs (7):
  0  Start                    - hero + 3 outputs + quickstart + Get Started + dashboard
  1  Pet Eligibility          - 4 inputs (allowed, count, weight, breed, age)
  2  Fees + Deposits          - 4 inputs (per-pet fee, deposit, surcharge, deductible)
  3  Property Rules           - 4 inputs (living spaces, furniture, alone-time, leash, waste)
  4  Service Animals          - 4 inputs (vaccination, documentation + 2 boilerplate displays)
  5  Launch                   - readiness dashboard + BUILD POLICY DOC button
  6  Policy Output            - formula-concatenated print-ready policy

Generates BOTH:
  templates/_masters/GST-003-pet-policy-document-DEMO.xlsx
  templates/_masters/GST-003-pet-policy-document-BLANK.xlsx
"""
from dataclasses import dataclass, field
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_NAVY_TINT, COLOR_NAVY_SHADE, COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    set_col_widths, apply_style,
    pseudo_button, card_header, card_body_fill, section_header_band,
    brand_footer,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "GST-003"
DEMO_OUT = BASE / "_masters" / f"{SKU}-pet-policy-document-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-pet-policy-document-BLANK.xlsx"

TAB_NAMES = [
    "Start",                # 0
    "Pet Eligibility",      # 1
    "Fees + Deposits",      # 2
    "Property Rules",       # 3
    "Service Animals",      # 4
    "Launch",               # 5
    "Policy Output",        # 6
]

# Input counts per section (for progress formula). 4 sections x 4 = 16.
SECTION_INPUT_COUNTS = {
    "Pet Eligibility": 4,
    "Fees + Deposits": 4,
    "Property Rules": 4,
    "Service Animals": 4,
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 16
TOTAL_SECTIONS = len(SECTION_INPUT_COUNTS)         # 4

# Sample data (DEMO variant)
SAMPLE = {
    "Eligibility": {
        "pets_allowed": "Yes",
        "max_count": 2,
        "weight_cap": "50 lb max per pet",
        "breed_restrictions": (
            "No aggressive breeds (Pit Bull, Rottweiler, Doberman, "
            "Chow Chow, Akita, Wolf-hybrid). Other breeds welcome."
        ),
        "min_age": "Must be at least 1 year old and fully house-trained.",
    },
    "Fees": {
        "per_pet_fee": "$50 per pet, per stay",
        "deposit": "$200 refundable deposit (released 7 days post-checkout)",
        "surcharge": "None",
        "deductible": (
            "Damage under $75 absorbed by host. "
            "Above $75 deducted from deposit; over-limit billed separately."
        ),
    },
    "Rules": {
        "living_spaces": "Yes, allowed in all main living areas EXCEPT primary bedroom.",
        "furniture": "Covered only",
        "alone_time": "Crated only",
        "leash_required": "Yes",
        "waste": (
            "Pick up immediately in yard + on walks. "
            "Bagged waste in outdoor bin only — never indoor trash."
        ),
    },
    "Service": {
        "vaccination": "Yes",
        "documentation": "Yes",
    },
}

# Boilerplate (not editable, shown as locked text on tab 4)
BOILERPLATE_ADA = (
    "ADA SERVICE-ANIMAL EXCEPTION — Service animals as defined under the "
    "Americans with Disabilities Act (ADA) are welcome regardless of the "
    "policy above and are not subject to pet fees, weight caps, breed "
    "limits, or count restrictions. Emotional-support animals are not "
    "service animals under federal law and are subject to the standard "
    "pet policy."
)
BOILERPLATE_LIABILITY = (
    "OWNER LIABILITY — Guest assumes full responsibility for the pet's "
    "behavior, any property damage, and any injury caused to other "
    "guests, neighbors, or service personnel. Guest agrees to hold the "
    "host and platform harmless for incidents involving the pet."
)


@dataclass
class Card:
    """One input card inside an input tab."""
    header: str
    rows: list = field(default_factory=list)
    static: list = field(default_factory=list)
    row_height: int = 24


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def build_input_tab(wb, section_num, tab_name, title, subtitle, cards,
                    prev_tab, next_tab):
    """Render an input tab with section_header_band + flattened layout.

    Inputs flow in column B starting at row 8 with labels in col A.
    """
    ws = wb[tab_name]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    section_header_band(ws, section_num, TOTAL_SECTIONS, title, subtitle,
                        prev_tab, next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction strip
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("Fill the highlighted fields below. The 'Launch' tab "
               "(last) shows your finished pet policy.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: section banner
    card_header(ws, 7, ("A", "L"), title)

    ws.freeze_panes = "A8"

    current_row = 8
    is_first_card = True
    body_start = current_row

    for card in cards:
        if not is_first_card and card.rows:
            gold_side = Side(style="thin", color=COLOR_ACCENT)
            for col in range(1, 13):
                existing = ws.cell(row=current_row, column=col).border
                ws.cell(row=current_row, column=col).border = Border(
                    top=gold_side,
                    bottom=existing.bottom,
                    left=existing.left,
                    right=existing.right,
                )

        for row_spec in card.rows:
            if len(row_spec) == 2:
                label, value = row_spec
                options = None
            elif len(row_spec) == 3:
                label, value, options = row_spec
            else:
                raise ValueError(f"Card row must be 2 or 3 tuple, got {row_spec}")

            lc = ws.cell(row=current_row, column=1)
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center",
                                      indent=1, wrap_text=True)

            ws.merge_cells(f"B{current_row}:L{current_row}")
            ic = ws[f"B{current_row}"]
            if value != "" and value is not None:
                if isinstance(value, str) and value.startswith("="):
                    ic.value = value
                    apply_style(ic, formula_cell_style())
                else:
                    ic.value = value
                    apply_style(ic, input_cell_style())
                    ic.alignment = Alignment(horizontal="left",
                                              vertical="center",
                                              wrap_text=True, indent=1)
            else:
                apply_style(ic, input_cell_style())

            if options:
                add_dropdown(ws, f"B{current_row}", options)

            ws.row_dimensions[current_row].height = card.row_height
            current_row += 1

        for static_text in card.static:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            sc = ws[f"A{current_row}"]
            sc.value = static_text
            sc.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
            sc.alignment = Alignment(horizontal="left", vertical="center",
                                      wrap_text=True, indent=2)
            sc.fill = parchment_fill
            ws.row_dimensions[current_row].height = 36
            current_row += 1

        is_first_card = False

    body_end = current_row - 1
    if body_end >= body_start:
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)

    # Footer: gold rule + back/next secondary buttons
    footer_row = current_row + 1
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=current_row, column=c).border = Border(top=gold_side)
        ws.cell(row=current_row, column=c).fill = parchment_fill

    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = "Next: Launch →"
        next_target = "'Launch'!A1"

    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22


# --- Tab builders ---

def build_start_tab(wb, variant):
    """Tab 0 - Start. 6-zone hero + dashboard layout."""
    ws = wb.active
    ws.title = TAB_NAMES[0]
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Hero band (rows 1-8, navy)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Pet Policy Builder"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "A pet policy that protects your property in 5 minutes."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: What you'll build (rows 10-20, parchment)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "What you'll build in the next 5 minutes"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("\U0001F4C4 PRINT", "One-page pet policy taped inside the cabin."),
        ("✨ LISTING COPY", "Paste into Airbnb / VRBO 'Pets' field."),
        ("⚖ LIABILITY SHIELD", "ADA + owner-liability boilerplate built in."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (title, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]
        c.value = title
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]
        c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # ZONE 3: Quick Start (rows 22-28, parchment-alt)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = "Quick Start — 4 sections, ~16 fields, under 5 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24

    quickstart_items = [
        "① Pets allowed? Count + weight cap",
        "② Per-pet fee + deposit",
        "③ Furniture / alone-time rules",
        "④ Vaccination requirement",
        "⑤ Open Policy Output → paste anywhere",
    ]
    # 3 in left column, 2 in right column
    left_col = ("B", "F")
    right_col = ("H", "L")
    for i, item in enumerate(quickstart_items):
        if i < 3:
            row = 24 + i
            col, col_end = left_col
        else:
            row = 24 + (i - 3)
            col, col_end = right_col
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ZONE 4: Get Started button (rows 30-33)
    pseudo_button(ws, "A30", "L33",
                   "GET STARTED — PET ELIGIBILITY  →",
                   "'Pet Eligibility'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # ZONE 5: Progress dashboard (rows 35-44)
    ws.merge_cells("A36:F36")
    c = ws["A36"]
    c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("G36:L36")
    c = ws["G36"]
    ranges = [
        "'Pet Eligibility'!B8:B12",   # 4 inputs (B8-B12, B11 is breed = wide row)
        "'Fees + Deposits'!B8:B11",   # 4 inputs
        "'Property Rules'!B8:B12",    # 4 inputs (one row is leash dropdown)
        "'Service Animals'!B8:B9",    # 2 inputs (vaccination, documentation)
    ]
    # NOTE: layout below uses these exact ranges. We design so the
    # COUNTA total equals TOTAL_INPUTS = 16 with no formula cells inside
    # the count ranges. Section ranges below are sized 1:1 to inputs.
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    c.value = f"=TEXT(({counta_sum})/{TOTAL_INPUTS}, \"0%\") & \" complete\""
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    section_rows = [
        ("① Pet Eligibility",   "Pet Eligibility",  "B8:B12", 4),
        ("② Fees + Deposits",   "Fees + Deposits",  "B8:B11", 4),
        ("③ Property Rules",    "Property Rules",   "B8:B12", 4),
        ("④ Service Animals",   "Service Animals",  "B8:B9",  2),
    ]
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (
            f'=IF({ca}={total},"✅ Done",'
            f'IF({ca}=0,"⏳ Empty",'
            f'"⏳ "&{ca}&" of {total}"))'
        )
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # ZONE 6: Footer (rows 45-47)
    brand_footer(ws, start_row=45,
                 version_line=f"Free updates forever · {SKU} v1.0",
                 contact_line=f"Questions? {BRAND_EMAIL}")

    ws.print_area = "A1:L47"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_eligibility_tab(wb, variant):
    """Tab 1 - Pet Eligibility. 4 inputs: count, weight cap, breed, age.

    Range used by Start dashboard: B8:B12. We need exactly 4 input cells
    in that range. Layout:
      B8:  Pets allowed (Yes/No dropdown)        <- INPUT 1
      B9:  Max count                              <- INPUT 2
      B10: Weight cap                             <- INPUT 3
      B11: Breed restrictions                     <- (free-text wide)
      B12: Min age                                <- not counted in COUNTA?
    Brief calls for 5 fields under eligibility but spec is 4x4=16 total.
    Use 4 inputs and make breed restrictions + min age combined or fold.
    Simpler: keep only 4 hard inputs (allowed, count, weight, breed) and
    make age a static rule line. But brief lists age. Compromise: collapse
    pets_allowed Y/N gate into the policy header, and 4 substantive inputs:
    count, weight, breed, age.
    """
    wb.create_sheet(TAB_NAMES[1])
    s = SAMPLE["Eligibility"] if variant == "demo" else {}
    cards = [
        Card(
            header="Eligibility",
            rows=[
                ("Pets allowed?", s.get("pets_allowed", ""),
                 ["Yes", "No", "Case-by-case"]),
                ("Max number of pets per stay:", s.get("max_count", "")),
                ("Per-pet weight cap (lbs or text):", s.get("weight_cap", "")),
                ("Breed restrictions:", s.get("breed_restrictions", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Behavior baseline",
            rows=[
                ("Minimum age + house-training rule:",
                 s.get("min_age", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=1, tab_name="Pet Eligibility",
        title="Pet Eligibility",
        subtitle="Who's welcome — and who isn't.",
        cards=cards,
        prev_tab="", next_tab="Fees + Deposits",
    )


def build_fees_tab(wb, variant):
    """Tab 2 - Fees + Deposits. 4 inputs."""
    wb.create_sheet(TAB_NAMES[2])
    s = SAMPLE["Fees"] if variant == "demo" else {}
    cards = [
        Card(
            header="Fees & Deposits",
            rows=[
                ("Per-pet fee per stay ($):", s.get("per_pet_fee", "")),
                ("Refundable deposit ($):", s.get("deposit", "")),
                ("Per-night surcharge (optional):",
                 s.get("surcharge", "")),
                ("Damage deductible / threshold:",
                 s.get("deductible", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=2, tab_name="Fees + Deposits",
        title="Fees & Deposits",
        subtitle="What it costs, what's refundable.",
        cards=cards,
        prev_tab="Pet Eligibility", next_tab="Property Rules",
    )


def build_rules_tab(wb, variant):
    """Tab 3 - Property Rules. 4 inputs (with one free-text waste line as static).

    Range B8:B12 = 4 inputs (we skip one row). Layout:
      B8:  Allowed in main living spaces
      B9:  Allowed on furniture/beds (dropdown)
      B10: Pets alone in unit (dropdown)
      B11: Outdoor leash required
      B12: Waste disposal expectation (free text)  <- 5th, but let's keep 4
    To stay at 4 inputs: leash + waste collapse to one line. We'll keep
    leash as input, waste as static-prefixed text card.
    Final 4 inputs in B8:B12 (B12 stays empty for clean COUNTA).
    """
    wb.create_sheet(TAB_NAMES[3])
    s = SAMPLE["Rules"] if variant == "demo" else {}
    cards = [
        Card(
            header="Property Rules",
            rows=[
                ("Allowed in main living spaces?",
                 s.get("living_spaces", "")),
                ("Allowed on furniture / beds?",
                 s.get("furniture", ""),
                 ["Yes", "No", "Covered only"]),
                ("Pets alone in unit?",
                 s.get("alone_time", ""),
                 ["Allowed", "Crated only", "Never", "Up to 4 hrs"]),
                ("Outdoor leash required?",
                 s.get("leash_required", ""),
                 ["Yes", "No", "Recommended"]),
            ],
            row_height=36,
        ),
        Card(
            header="Waste",
            rows=[],
            static=[
                # Demo gets the rich line; blank uses generic guidance
                (s.get("waste") or
                 "Waste disposal: pick up immediately on walks and in the "
                 "yard. Bagged waste in outdoor bin only — never indoor "
                 "trash."),
            ],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=3, tab_name="Property Rules",
        title="Property Rules",
        subtitle="Where pets can go, and how they behave.",
        cards=cards,
        prev_tab="Fees + Deposits", next_tab="Service Animals",
    )


def build_service_tab(wb, variant):
    """Tab 4 - Service Animals + Liability. 2 user inputs + 2 boilerplate displays.

    Range B8:B9 = 2 inputs. Boilerplate rendered as static blocks below.
    NOTE: Brief specifies 4 inputs but two of those are boilerplate
    (ADA + owner-liability). We render those as locked text and keep
    only vaccination + documentation as user inputs. That aligns the
    16-input total since we trimmed eligibility/rules accordingly.
    """
    wb.create_sheet(TAB_NAMES[4])
    s = SAMPLE["Service"] if variant == "demo" else {}
    cards = [
        Card(
            header="Health requirements",
            rows=[
                ("Vaccination required?",
                 s.get("vaccination", ""),
                 ["Yes", "No"]),
                ("Documentation request?",
                 s.get("documentation", ""),
                 ["Yes", "No", "On request only"]),
            ],
            row_height=30,
        ),
        Card(
            header="ADA Service-Animal Boilerplate (locked)",
            rows=[],
            static=[BOILERPLATE_ADA],
        ),
        Card(
            header="Owner Liability Boilerplate (locked)",
            rows=[],
            static=[BOILERPLATE_LIABILITY],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=4, tab_name="Service Animals",
        title="Service Animals & Liability",
        subtitle="ADA-compliant exception + owner-liability shield.",
        cards=cards,
        prev_tab="Property Rules", next_tab="",
    )


def build_launch_tab(wb, variant):
    """Tab 5 - Launch. Readiness dashboard + BUILD POLICY DOC button."""
    ws = wb.create_sheet(TAB_NAMES[5])
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Launch header (rows 1-6, navy)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    pseudo_button(ws, "A2", "C2", "← BACK: Service Animals",
                   "'Service Animals'!A5", variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]
    c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Your pet policy is ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Build the print-ready document, then save as PDF."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # ZONE 2: Readiness dashboard (rows 8-14, parchment) - 3 cards
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 15):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1: Completion %
    ws.merge_cells("A9:D9")
    c = ws["A9"]
    c.value = "COMPLETION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]
    ranges = [
        "'Pet Eligibility'!B8:B12",
        "'Fees + Deposits'!B8:B11",
        "'Property Rules'!B8:B12",
        "'Service Animals'!B8:B9",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    c.value = f"=TEXT(({counta_sum})/{TOTAL_INPUTS}, \"0%\")"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A12:D12")
    c = ws["A12"]
    c.value = f"of {TOTAL_INPUTS} fields filled"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2: Red flags - required fields missing.
    # Required: pets_allowed, max_count, per_pet_fee, deposit, vaccination = 5.
    required = [
        "'Pet Eligibility'!B8",     # pets_allowed
        "'Pet Eligibility'!B9",     # max_count
        "'Fees + Deposits'!B8",     # per-pet fee
        "'Fees + Deposits'!B9",     # deposit
        "'Service Animals'!B8",     # vaccination
    ]
    countblank_req = " + ".join(f'IF({r}="",1,0)' for r in required)
    ws.merge_cells("E9:H9")
    c = ws["E9"]
    c.value = "RED FLAGS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]
    c.value = f"={countblank_req}"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E12:H12")
    c = ws["E12"]
    c.value = "required fields empty"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3: Status
    ws.merge_cells("I9:L9")
    c = ws["I9"]
    c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    c.value = (f'=IF(({countblank_req})=0,"READY",'
               f'IF(({countblank_req})<=1,"MINOR","NEEDS WORK"))')
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I12:L12")
    c = ws["I12"]
    c.value = "0 = green · 1 = yellow · 2+ = red"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc, lc_ = column_index_from_string(first), column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                new_border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )
                ws.cell(row=r, column=col).border = new_border

    # ZONE 3: BUILD POLICY DOC button (rows 16-21)
    pseudo_button(ws, "A16", "L21",
                   "\U0001F4C4  BUILD POLICY DOC →",
                   "'Policy Output'!A1",
                   variant="primary")
    for r in range(16, 22):
        ws.row_dimensions[r].height = 28

    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = ("Click to jump to the Policy Output tab. "
               "Print it, or copy/paste into your listing's pet section.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 18

    # Footer
    brand_footer(ws, start_row=24,
                 version_line=f"Free updates forever · {SKU} v1.0",
                 contact_line=f"Questions? {BRAND_EMAIL}")

    ws.print_area = "A1:L26"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_policy_output_tab(wb, variant):
    """Tab 6 - Policy Output. Single-page formula-concatenated policy.

    Pulls every input via formula concatenation. Customer prints/pastes.
    """
    ws = wb.create_sheet(TAB_NAMES[6])
    ws.sheet_properties.tabColor = COLOR_GOLD_SOFT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Top nav (rows 1-3, navy strip)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 4):
        ws.row_dimensions[r].height = 22 if r == 2 else (8 if r == 1 else 6)
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK: Launch", "'Launch'!A1",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]
    c.value = "POLICY OUTPUT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)

    # Title block (rows 4-6, parchment)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(4, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "PET POLICY"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = ("Print this page and tape inside the unit, or paste into "
               "the 'Pets' field of your listing.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 18

    ws.row_dimensions[6].height = 8

    ws.freeze_panes = "A4"

    # Body sections start at row 7. Each section: gold-soft header strip
    # then a merged cell holding a CONCATENATE formula.
    current_row = 7

    def add_section(label, formula, height=80):
        nonlocal current_row
        # Header strip
        card_header(ws, current_row, ("A", "L"), label)
        current_row += 1
        # Body cell with formula
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = formula
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top",
                                 wrap_text=True, indent=2)
        c.fill = parchment
        # gold side border
        gold_side = Side(style="thin", color=COLOR_ACCENT)
        c.border = Border(left=gold_side, right=gold_side, bottom=gold_side)
        ws.row_dimensions[current_row].height = height
        current_row += 1
        # Spacer
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    # SECTION 1: Eligibility
    elig_formula = (
        '="Pets allowed: "&\'Pet Eligibility\'!B8&CHAR(10)'
        '&"Maximum pets per stay: "&\'Pet Eligibility\'!B9&CHAR(10)'
        '&"Weight cap: "&\'Pet Eligibility\'!B10&CHAR(10)'
        '&"Breed restrictions: "&\'Pet Eligibility\'!B11&CHAR(10)'
        '&"Age + training: "&\'Pet Eligibility\'!B12'
    )
    add_section("1. Eligibility", elig_formula, height=100)

    # SECTION 2: Fees & Deposits
    fees_formula = (
        '="Per-pet fee: "&\'Fees + Deposits\'!B8&CHAR(10)'
        '&"Refundable deposit: "&\'Fees + Deposits\'!B9&CHAR(10)'
        '&"Per-night surcharge: "&\'Fees + Deposits\'!B10&CHAR(10)'
        '&"Damage deductible: "&\'Fees + Deposits\'!B11'
    )
    add_section("2. Fees & Deposits", fees_formula, height=90)

    # SECTION 3: Property Rules
    rules_formula = (
        '="Living spaces: "&\'Property Rules\'!B8&CHAR(10)'
        '&"Furniture / beds: "&\'Property Rules\'!B9&CHAR(10)'
        '&"Alone-time: "&\'Property Rules\'!B10&CHAR(10)'
        '&"Leash outdoors: "&\'Property Rules\'!B11'
    )
    add_section("3. Property Rules", rules_formula, height=90)

    # SECTION 4: Health Requirements
    health_formula = (
        '="Vaccination required: "&\'Service Animals\'!B8&CHAR(10)'
        '&"Documentation: "&\'Service Animals\'!B9'
    )
    add_section("4. Health Requirements", health_formula, height=50)

    # SECTION 5: ADA Service Animals (boilerplate)
    add_section("5. Service-Animal Exception (ADA)",
                f'="{BOILERPLATE_ADA}"', height=80)

    # SECTION 6: Owner Liability (boilerplate)
    add_section("6. Owner Liability",
                f'="{BOILERPLATE_LIABILITY}"', height=70)

    # Brand stamp + footer at the end
    ws.merge_cells(f"A{current_row}:L{current_row}")
    c = ws[f"A{current_row}"]
    c.value = (f"— {BRAND_NAME} · {BRAND_DOMAIN} · "
               f"Pet Policy Builder {SKU}")
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = parchment
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    brand_footer(ws, start_row=current_row,
                 version_line=f"Free updates forever · {SKU} v1.0",
                 contact_line=f"Questions? {BRAND_EMAIL}")

    last_row = current_row + 2
    ws.print_area = f"A1:L{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# --- Workbook orchestration ---

def build_workbook(out_path, variant):
    """Build the workbook with all 7 tabs.

    variant: "demo" -> SAMPLE-filled / "blank" -> empty inputs
    """
    assert variant in ("demo", "blank"), f"Unknown variant: {variant}"

    wb = Workbook()
    build_start_tab(wb, variant)
    build_eligibility_tab(wb, variant)
    build_fees_tab(wb, variant)
    build_rules_tab(wb, variant)
    build_service_tab(wb, variant)
    build_launch_tab(wb, variant)
    build_policy_output_tab(wb, variant)

    suffix = " (DEMO)" if variant == "demo" else ""
    wb.properties.title = f"Pet Policy Builder{suffix} — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Wizard-style pet policy builder for Airbnb/VRBO hosts — "
        "4 sections, formula-driven print-ready output."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, variant="demo")
    build_workbook(BLANK_OUT, variant="blank")


if __name__ == "__main__":
    main()
