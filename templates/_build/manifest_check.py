"""Pre-publish manifest check for every catalog SKU.

Verifies every SKU directory has the files Etsy / Gumroad publish expects:
  - DEMO xlsx + BLANK xlsx in templates/_masters/
  - License PDF + Howto PDF in templates/_delivery/<sku>/
  - Hero thumbnail PNG in templates/_delivery/<sku>/

Plus shared assets:
  - templates/_delivery/_shared/etsy-upgrade-insert.pdf (A13)
  - templates/_delivery/_shared/license.pdf (generic shared)

Each file is checked for presence, file size > 0, and openable. License PDFs
are scanned for the SKU code in the PDF byte stream — fails on mismatch
(catches the "wrong-SKU license accidentally pasted into the wrong folder"
class of error).

Run:
    python templates/_build/manifest_check.py

Exit 0 = green; exit 1 = any failure (suitable for pre-commit / CI gating).

Source of truth for the SKU list: imported from
templates/_delivery/_shared/_build_delivery_pdfs.py:SKU_CATALOG.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

# Windows default console codepage (cp1252) can't encode the Unicode
# box-drawing characters ('──', '───') this script prints in section
# headers. Reconfigure stdout/stderr to UTF-8 with a replace-fallback
# so the hook can never crash before it does any real check.
# Python 3.7+; no-op on TTYs already using UTF-8.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        # Stream may already be wrapped (e.g. by pytest capture) or
        # not reconfigurable; safe to skip — encoding errors will then
        # fall back to the existing codepage.
        pass

REPO = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO / "templates" / "_delivery" / "_shared"))
sys.path.insert(0, str(REPO / "templates" / "_delivery" / "_bundles"))
from _build_delivery_pdfs import SKU_CATALOG  # noqa: E402
from bundles_config import BUNDLES  # noqa: E402

DELIVERY = REPO / "templates" / "_delivery"
MASTERS = REPO / "templates" / "_masters"
SHARED = DELIVERY / "_shared"

# ANSI color helpers — degrade gracefully if stdout is not a TTY.
def _supports_color() -> bool:
    return sys.stdout.isatty()

GREEN = "\033[92m" if _supports_color() else ""
RED = "\033[91m" if _supports_color() else ""
YELLOW = "\033[93m" if _supports_color() else ""
DIM = "\033[2m" if _supports_color() else ""
RESET = "\033[0m" if _supports_color() else ""


class CheckResult:
    """Tracks PASS / FAIL across the run; produces the final summary."""
    def __init__(self) -> None:
        self.passes = 0
        self.fails: list[str] = []
        self.warns: list[str] = []

    def passed(self, label: str) -> None:
        self.passes += 1
        print(f"  {GREEN}PASS{RESET}  {label}")

    def failed(self, label: str, reason: str) -> None:
        self.fails.append(f"{label}: {reason}")
        print(f"  {RED}FAIL{RESET}  {label}  {DIM}— {reason}{RESET}")

    def warned(self, label: str, reason: str) -> None:
        self.warns.append(f"{label}: {reason}")
        print(f"  {YELLOW}WARN{RESET}  {label}  {DIM}— {reason}{RESET}")


def _check_file_exists(path: Path, *, min_bytes: int = 100) -> tuple[bool, str]:
    if not path.exists():
        return False, "file missing"
    size = path.stat().st_size
    if size < min_bytes:
        return False, f"file too small ({size} bytes)"
    return True, ""


def _check_xlsx_opens(path: Path) -> tuple[bool, str]:
    """Quickly verify xlsx is a valid Open Office XML zip + has worksheets."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False, "openpyxl not installed"
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        return False, f"xlsx open failed: {e.__class__.__name__}: {e}"
    if not sheet_names:
        return False, "xlsx has no sheets"
    return True, ""


def _check_pdf_valid(path: Path) -> tuple[bool, str]:
    """Verify PDF magic bytes + EOF marker."""
    try:
        with path.open("rb") as f:
            head = f.read(8)
            if not head.startswith(b"%PDF-"):
                return False, "not a PDF (no %PDF- header)"
            f.seek(-1024, 2)  # last 1KB
            tail = f.read()
            if b"%%EOF" not in tail:
                return False, "no %%EOF marker (likely truncated)"
    except Exception as e:
        return False, f"read failed: {e.__class__.__name__}"
    return True, ""


def _check_license_pdf_carries_sku(pdf_path: Path, sku_code: str) -> tuple[bool, str]:
    """Extract text from the PDF and confirm the SKU code appears.

    Catches the "wrong-SKU file copied to the wrong folder" trap. Playwright-
    rendered PDFs compress text streams, so a real PDF parser is required.

    The brand strip header renders the SKU with letter-spacing CSS, which
    pypdf may extract as "T A X - 0 0 1" (spaced) — we strip whitespace
    before comparing to handle that case.
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        return False, "pypdf not installed (pip install pypdf)"
    try:
        reader = PdfReader(pdf_path)
        full_text = "".join(p.extract_text() or "" for p in reader.pages)
    except Exception as e:
        return False, f"pdf parse failed: {e.__class__.__name__}: {e}"
    haystack = full_text.replace(" ", "").replace(" ", "")
    needle = sku_code.replace(" ", "")
    if needle in haystack:
        return True, ""
    return False, f"SKU code {sku_code!r} not found in extracted PDF text"


def check_shared_assets(result: CheckResult) -> None:
    print(f"\n{DIM}── Shared assets ──{RESET}")

    a13 = SHARED / "etsy-upgrade-insert.pdf"
    ok, why = _check_file_exists(a13)
    if not ok:
        result.failed("Shared A13 buyer PDF", why)
    else:
        ok, why = _check_pdf_valid(a13)
        if ok:
            result.passed("Shared A13 buyer PDF")
        else:
            result.failed("Shared A13 buyer PDF", why)

    shared_license = SHARED / "license.pdf"
    ok, why = _check_file_exists(shared_license)
    if not ok:
        result.failed("Shared generic license PDF", why)
    else:
        ok, why = _check_pdf_valid(shared_license)
        if ok:
            result.passed("Shared generic license PDF")
        else:
            result.failed("Shared generic license PDF", why)


def check_sku(dir_name: str, sku_code: str, sku_name: str,
              result: CheckResult) -> None:
    print(f"\n{DIM}── {sku_code}  {sku_name}  ──{RESET}")
    sku_dir = DELIVERY / dir_name

    # 1. DEMO xlsx
    demo = MASTERS / f"{dir_name}-DEMO.xlsx"
    ok, why = _check_file_exists(demo, min_bytes=10_000)
    if not ok:
        result.failed(f"{sku_code} DEMO xlsx", why)
    else:
        ok, why = _check_xlsx_opens(demo)
        if ok:
            result.passed(f"{sku_code} DEMO xlsx")
        else:
            result.failed(f"{sku_code} DEMO xlsx", why)

    # 2. BLANK xlsx
    blank = MASTERS / f"{dir_name}-BLANK.xlsx"
    ok, why = _check_file_exists(blank, min_bytes=5_000)
    if not ok:
        result.failed(f"{sku_code} BLANK xlsx", why)
    else:
        ok, why = _check_xlsx_opens(blank)
        if ok:
            result.passed(f"{sku_code} BLANK xlsx")
        else:
            result.failed(f"{sku_code} BLANK xlsx", why)

    # 3. License PDF — exists, valid, AND carries this SKU's code.
    license_pdf = sku_dir / f"{sku_code}-license.pdf"
    ok, why = _check_file_exists(license_pdf)
    if not ok:
        result.failed(f"{sku_code} license.pdf", why)
    else:
        ok, why = _check_pdf_valid(license_pdf)
        if not ok:
            result.failed(f"{sku_code} license.pdf", why)
        else:
            ok, why = _check_license_pdf_carries_sku(license_pdf, sku_code)
            if ok:
                result.passed(f"{sku_code} license.pdf")
            else:
                # This is the "wrong-SKU file in folder" trap the manifest
                # check exists to catch — escalate to FAIL, not WARN.
                result.failed(f"{sku_code} license.pdf", why)

    # 4. Howto PDF
    howto_pdf = sku_dir / f"{sku_code}-howto.pdf"
    ok, why = _check_file_exists(howto_pdf)
    if not ok:
        result.failed(f"{sku_code} howto.pdf", why)
    else:
        ok, why = _check_pdf_valid(howto_pdf)
        if ok:
            result.passed(f"{sku_code} howto.pdf")
        else:
            result.failed(f"{sku_code} howto.pdf", why)

    # 5. Hero thumbnail PNG (Etsy needs at least one image to publish)
    hero = sku_dir / "thumb-1.png"
    ok, why = _check_file_exists(hero, min_bytes=1_000)
    if ok:
        # Basic PNG magic-byte check.
        with hero.open("rb") as f:
            magic = f.read(8)
        if magic == b"\x89PNG\r\n\x1a\n":
            result.passed(f"{sku_code} hero thumbnail")
        else:
            result.failed(f"{sku_code} hero thumbnail", "not a valid PNG")
    else:
        result.failed(f"{sku_code} hero thumbnail", why)


def check_bundle(bundle: dict, result: CheckResult) -> None:
    """Verify a bundle's delivery package is complete + valid."""
    print(f"\n{DIM}── {bundle['code']}  {bundle['name']}  ──{RESET}")
    bundles_dir = REPO / "templates" / "_delivery" / "_bundles"
    slug_dir = bundles_dir / f"{bundle['code']}-{bundle['slug']}"
    base = bundle["slug"]

    # 1. BLANK ZIP
    blank_zip = slug_dir / f"{base}-blank.zip"
    ok, why = _check_file_exists(blank_zip, min_bytes=1_000)
    if ok:
        result.passed(f"{bundle['code']} BLANK zip")
    else:
        result.failed(f"{bundle['code']} BLANK zip", why)

    # 2. DEMO ZIP
    demo_zip = slug_dir / f"{base}-demo.zip"
    ok, why = _check_file_exists(demo_zip, min_bytes=1_000)
    if ok:
        result.passed(f"{bundle['code']} DEMO zip")
    else:
        result.failed(f"{bundle['code']} DEMO zip", why)

    # 3. Merged howto PDF
    howto_pdf = slug_dir / f"{base}-howto.pdf"
    ok, why = _check_file_exists(howto_pdf)
    if not ok:
        result.failed(f"{bundle['code']} howto.pdf", why)
    else:
        ok, why = _check_pdf_valid(howto_pdf)
        if ok:
            result.passed(f"{bundle['code']} howto.pdf")
        else:
            result.failed(f"{bundle['code']} howto.pdf", why)

    # 4. README PDF — must contain bundle code in extracted text
    readme_pdf = slug_dir / f"{base}-readme.pdf"
    ok, why = _check_file_exists(readme_pdf)
    if not ok:
        result.failed(f"{bundle['code']} readme.pdf", why)
    else:
        ok, why = _check_pdf_valid(readme_pdf)
        if not ok:
            result.failed(f"{bundle['code']} readme.pdf", why)
        else:
            # Verify the bundle code is in the readme (catches wrong-bundle-folder)
            ok, why = _check_license_pdf_carries_sku(readme_pdf, bundle["code"])
            if ok:
                result.passed(f"{bundle['code']} readme.pdf")
            else:
                result.failed(f"{bundle['code']} readme.pdf", why)


def main() -> int:
    print(f"{DIM}File-delivery manifest check  ·  {len(SKU_CATALOG)} SKUs + "
           f"{len(BUNDLES)} bundles  ·  repo: {REPO.name}{RESET}")
    result = CheckResult()

    check_shared_assets(result)
    for dir_name, sku_code, sku_name in SKU_CATALOG:
        check_sku(dir_name, sku_code, sku_name, result)
    print(f"\n{DIM}─── BUNDLE DELIVERY PACKAGES ───{RESET}")
    for bundle in BUNDLES:
        check_bundle(bundle, result)

    # Summary
    print()
    total_fail = len(result.fails)
    total_warn = len(result.warns)
    if total_fail == 0:
        print(f"{GREEN}╔══════════════════════════════════════════╗")
        print(f"║  MANIFEST CHECK GREEN  ·  {result.passes:>3} checks pass  ║")
        print(f"╚══════════════════════════════════════════╝{RESET}")
        if total_warn:
            print(f"  {YELLOW}{total_warn} warnings{RESET} (non-blocking).")
        return 0
    else:
        print(f"{RED}╔══════════════════════════════════════════╗")
        print(f"║  MANIFEST CHECK FAILED  ·  {total_fail:>2} failures   ║")
        print(f"╚══════════════════════════════════════════╝{RESET}")
        print(f"  {DIM}{result.passes} passed  ·  {total_warn} warnings{RESET}")
        print()
        print(f"  Failures:")
        for fail in result.fails:
            print(f"    {RED}·{RESET} {fail}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
