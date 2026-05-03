"""Build TAX-002 Single-Property P&L Tracker Excel files (v2.2 standard).

Operational-mode tracker: customer logs revenue + expenses week to week,
hands the Schedule E Summary tab to a CPA at year-end. Implements the
v2.2 visual + interaction language set by the Welcome Book v2.2 and
already shipped in the Mileage Log v2.3.

Critical fixes vs v1 (per skill str-tax-context.md §"Active tax year"):
- Active tax year cell on Settings (replaces brittle YEAR(TODAY()) usage)
- Date strings parsed into datetime.date objects so SUMIFS date filters
  actually match. v1 stored dates as text — every monthly aggregate
  silently returned $0.
- Schedule C/E selector on Settings + dependent Schedule routing footnote
- Pre-startup-cost callout on Start tab (IRC §195)

Generates BOTH DEMO and BLANK variants from shared code (per the suite-wide
customer-eye theme: pre-populated demo data converts 5x better than empty cells,
but customers still need a clean copy to drop their own data into).
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    STATE_BAD_FILL, STATE_GOOD_FILL,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-002"
NAME = "pl-single-property"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# --- 17 Schedule E expense categories ---

EXPENSE_CATEGORIES = [
    "Advertising (Line 6)",
    "Auto/travel (Line 7)",
    "Cleaning + maintenance (Line 8)",
    "Commissions (Line 9)",
    "Insurance (Line 10)",
    "Legal + professional (Line 11)",
    "Management fees (Line 12)",
    "Mortgage interest (Line 13)",
    "Other interest (Line 14)",
    "Repairs (Line 15)",
    "Supplies (Line 16)",
    "Taxes (Line 17)",
    "Utilities (Line 18)",
    "Wages (Line 19)",
    "Other — Platform fees (Line 19)",
    "Other — Misc (Line 19)",
    "Depreciation (Line 20) — see note",
]

BOOKING_CHANNELS = ["Airbnb", "VRBO", "Booking.com", "Direct", "Other"]

ACTIVE_TAX_YEAR = 2026

# 10 sample bookings Jan-Mar 2026
# v2.3 added Nights column — drives Avg Stay calc on Schedule E Summary
# (≤7 nights avg unlocks the STR loophole under IRC §469).
SAMPLE_REVENUE = [
    ("2026-01-08", "Airbnb guest #A1092",       "Airbnb", 2400, 240, 210, 4, "4 nights"),
    ("2026-01-22", "Airbnb guest #A1098",       "Airbnb", 1800, 180, 210, 3, "3 nights"),
    ("2026-02-05", "VRBO guest #V4455",         "VRBO",   2100, 180, 210, 3, "3 nights"),
    ("2026-02-14", "Airbnb guest #A1112",       "Airbnb", 2800, 280, 210, 7, "Valentine's week"),
    ("2026-02-27", "Direct booking — Thompson", "Direct", 2200,   0, 210, 4, "Returning guest"),
    ("2026-03-07", "Airbnb guest #A1135",       "Airbnb", 1900, 190, 210, 3, ""),
    ("2026-03-14", "Airbnb guest #A1140",       "Airbnb", 2100, 210, 210, 4, "St Patrick's weekend"),
    ("2026-03-20", "VRBO guest #V4488",         "VRBO",   1800, 160, 210, 3, ""),
    ("2026-03-25", "Airbnb guest #A1147",       "Airbnb", 1600, 160, 210, 3, ""),
    ("2026-03-30", "Direct booking — Miller",   "Direct", 1800,   0, 210, 3, "Month-end"),
]

# 23 sample expenses
SAMPLE_EXPENSES = [
    # 8 cleanings
    ("2026-01-15", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-01-25", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-02-08", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-02-17", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-03-02", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-03-10", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-03-17", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    ("2026-03-28", "Smokies Clean",       EXPENSE_CATEGORIES[2],  450, "Venmo",       "Yes", "Turnover"),
    # 3 supplies
    ("2026-01-02", "Home Depot",          EXPENSE_CATEGORIES[10], 120, "Credit Card", "Yes", "Supplies restock"),
    ("2026-02-14", "Costco",              EXPENSE_CATEGORIES[10], 180, "Credit Card", "Yes", "Bulk supplies"),
    ("2026-03-08", "Target",              EXPENSE_CATEGORIES[10], 120, "Credit Card", "Yes", "Linens replacement"),
    # 3 mortgage interest
    ("2026-01-01", "Wells Fargo",         EXPENSE_CATEGORIES[7],  400, "ACH",         "Yes", "Jan mortgage int"),
    ("2026-02-01", "Wells Fargo",         EXPENSE_CATEGORIES[7],  400, "ACH",         "Yes", "Feb mortgage int"),
    ("2026-03-01", "Wells Fargo",         EXPENSE_CATEGORIES[7],  400, "ACH",         "Yes", "Mar mortgage int"),
    # utilities
    ("2026-01-15", "Spectrum",            EXPENSE_CATEGORIES[12],  90, "ACH",         "Yes", "Internet"),
    ("2026-01-20", "Sevier Electric",     EXPENSE_CATEGORIES[12], 120, "ACH",         "Yes", "Jan utilities"),
    ("2026-02-20", "Sevier Electric",     EXPENSE_CATEGORIES[12], 140, "ACH",         "Yes", "Feb utilities"),
    # 1 repair
    ("2026-02-15", "Quick Plumbing",      EXPENSE_CATEGORIES[9],  800, "Check",       "Yes", "Emergency leak"),
    # 5 platform fees
    ("2026-01-10", "Airbnb Platform Fee", EXPENSE_CATEGORIES[14], 240, "Auto-deduct", "Yes", "Host fees"),
    ("2026-02-10", "Airbnb Platform Fee", EXPENSE_CATEGORIES[14], 460, "Auto-deduct", "Yes", "Host fees"),
    ("2026-03-10", "Airbnb Platform Fee", EXPENSE_CATEGORIES[14], 560, "Auto-deduct", "Yes", "Host fees"),
    ("2026-02-10", "VRBO Platform Fee",   EXPENSE_CATEGORIES[14], 180, "Auto-deduct", "Yes", ""),
    ("2026-03-20", "VRBO Platform Fee",   EXPENSE_CATEGORIES[14], 160, "Auto-deduct", "Yes", ""),
]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    """v1 stored YYYY-MM-DD as text. SUMIFS with DATE() criteria can't
    match string cells — silent zeros across every monthly aggregate.
    Parse to a real datetime.date at write-time."""
    return datetime.strptime(s, "%Y-%m-%d").date()


def _val(variant, demo_value):
    """Return demo_value when building the DEMO variant, None for BLANK."""
    return demo_value if variant == "demo" else None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational-mode hero + verdict + activity dashboard)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Single-Property P&L"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Close your year before April does."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — single headline answer (Theme 1 of suite-wide
    # customer-eye review). Hosts respond to verdicts, not numbers; the math
    # lives in the cards below, the answer lives on top.
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(\'Monthly P&L\'!N7=0,'
        '"\U0001F4CA  Log your first booking to see your verdict.",'
        'IF(\'Monthly P&L\'!N30/\'Monthly P&L\'!N7>=0.25,'
        '"✅  Net margin = "&TEXT(\'Monthly P&L\'!N30/\'Monthly P&L\'!N7,"0%")'
        '&"  —  above the 25% Schedule E benchmark.",'
        '"⚠  Net margin = "&TEXT(\'Monthly P&L\'!N30/\'Monthly P&L\'!N7,"0%")'
        '&"  —  below the 25% Schedule E benchmark; review expenses."))'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"TAX-002 · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Schedule E Part I categories drive every expense row, so the Schedule E "
        "Summary tab is your CPA hand-off doc — not a separate report you build "
        "at year-end. Log revenue + expenses weekly (the IRS contemporaneous-record "
        "standard), and the Monthly P&L + Summary tabs auto-roll up. Reference: "
        "IRS Publication 527 (Residential Rental Property)."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Property Info: fill in property details once (address, purchase, loan, in-service date).",
        "② Revenue Log: log every booking — date, channel, gross, fees. One row per booking.",
        "③ Expense Log: log every expense and pick the Schedule E category from the dropdown.",
        "④ Monthly P&L: 12-column matrix auto-rolls; YTD column on the right.",
        "⑤ Schedule E Summary: line-mapped totals to hand to your CPA at year-end.",
        "⑥ Settings: bump 'Active tax year' each January so the dashboards stay right.",
        "⑦ Log within 7 days — IRS disallows reconstructed records (Pub 527 §contemporaneous).",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 20

    # --- ZONE 4: Primary "OPEN P&L" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  OPEN MONTHLY P&L",
                   "'Monthly P&L'!A1", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): YTD Revenue
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "YTD REVENUE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = "='Monthly P&L'!N7"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "rents + cleaning fees"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): YTD Expenses
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "YTD EXPENSES"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = "='Monthly P&L'!N28"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "deductible Schedule E"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Net Income
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "NET INCOME"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = "='Monthly P&L'!N30"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "Schedule E Line 26"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Gold borders around the 3 cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Revenue Log",
                   "'Revenue Log'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Expense Log",
                   "'Expense Log'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Schedule E Summary",
                   "'Schedule E Summary'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Tax-time / year-end checkpoint callout (rows 42-43) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📄 TAX TIME: print Schedule E Summary + Monthly P&L "
        "(File → Print → 'Print Active Sheets'). "
        "📅 YEAR END: copy YTD totals into Settings → Year-end Archive before Jan 1, "
        "then bump 'Active tax year' on Settings."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Pre-startup costs warning (row 44) ---
    # IRC §195: trips/expenses BEFORE acquiring or placing in service must
    # be capitalized + amortized 15 yr (with up to $5K Year-1 deduction).
    # Don't let customers blend pre-startup with operational expenses.
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "⚠ Pre-startup costs (IRC §195): trips/expenses BEFORE the property is placed "
        "in service must be capitalized over 15 years, not expensed here. "
        "Track those separately and hand to your CPA."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    # --- ZONE 9: Upgrade banner (row 46) — preserved from v1 ---
    add_upgrade_banner(ws, 46)

    # --- ZONE 10: Footer (rows 48-50) ---
    brand_footer(ws, 48,
                 version_line="TAX-002 · v2.3 · Free updates forever")

    # Print setup
    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_property_info_tab(wb, variant):
    """Sheet 1 — Property Info (12 labeled input rows)."""
    ws = wb.create_sheet("Property Info")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Property Info",
                         prev_tab="Start", next_tab="Revenue Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Fill once per property — purchase, loan, in-service date"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 28), ("B", 40)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    right_align = Alignment(horizontal="right", vertical="center")

    rows_data = [
        (5,  "Property name:",        _val(variant, "Smokies Ridge Cabin"),  None),
        (6,  "Street address:",       _val(variant, "123 Mountain Lane"),    None),
        (7,  "City / State / Zip:",   _val(variant, "Gatlinburg, TN 37738"), None),
        (8,  "Property type:",        _val(variant, "Cabin"),                None),
        (9,  "Purchase date:",        _val(variant, _parse_date("2023-08-15")), "yyyy-mm-dd"),
        (10, "Purchase price ($):",   _val(variant, 420000),                 '"$"#,##0'),
        (11, "Closing costs ($):",    _val(variant, 8500),                   '"$"#,##0'),
        (12, "Loan amount ($):",      _val(variant, 336000),                 '"$"#,##0'),
        (13, "Interest rate (%):",    _val(variant, 0.0675),                 "0.000%"),
        (14, "Loan term (years):",    _val(variant, 30),                     None),
        (15, "Business start date:",  _val(variant, _parse_date("2023-10-01")), "yyyy-mm-dd"),
        (16, "Days rented YTD:",      _val(variant, 72),                     None),
        # v2.3 — Personal-use days drive Pub 527's 14-day/10% test:
        # if personal use exceeds the greater of 14 days or 10% of rental
        # days, the property is a "residence" — losses cannot offset other
        # income. Schedule E Summary computes the threshold and flags the
        # deductibility risk.
        (17, "Personal use days YTD:", _val(variant, 0),                     None),
    ]

    for row_num, label, value, fmt in rows_data:
        a = ws.cell(row=row_num, column=1, value=label)
        a.font = bold_right
        a.alignment = right_align
        ws.row_dimensions[row_num].height = 18

        b = ws.cell(row=row_num, column=2, value=value)
        apply_style(b, input_cell_style())
        if fmt:
            b.number_format = fmt

    # Row 18: explainer for personal use (Pub 527)
    ws.merge_cells("A18:B18")
    c18 = ws["A18"]
    c18.value = (
        "Personal use = days you/family used the property NOT charged "
        "fair-market rent. Includes letting friends stay free."
    )
    c18.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c18.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[18].height = 28


def build_revenue_log(wb, variant):
    """Sheet 2 — Revenue Log (10 sample bookings in DEMO + 1000-row capacity)."""
    ws = wb.create_sheet("Revenue Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Revenue Log",
                         prev_tab="Property Info", next_tab="Expense Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Log every booking — date, guest, channel, gross, fees"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 28), ("C", 14), ("D", 12),
        ("E", 12), ("F", 14), ("G", 12), ("H", 8), ("I", 28),
    ])

    headers = [
        "Date", "Guest / Source", "Channel", "Gross",
        "Platform Fee", "Cleaning Collected", "Net", "Nights", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Rows 6+: sample revenue data (DEMO variant only; BLANK starts empty)
    sample_rows = SAMPLE_REVENUE if variant == "demo" else []
    for i, r in enumerate(sample_rows, start=6):
        date_val, guest, channel, gross, platform_fee, cleaning, nights, notes = r

        a = ws.cell(row=i, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=guest)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=channel)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=gross)
        apply_style(d, input_cell_style())
        d.number_format = '"$"#,##0.00'

        e = ws.cell(row=i, column=5, value=platform_fee)
        apply_style(e, input_cell_style())
        e.number_format = '"$"#,##0.00'

        f = ws.cell(row=i, column=6, value=cleaning)
        apply_style(f, input_cell_style())
        f.number_format = '"$"#,##0.00'

        g = ws.cell(row=i, column=7, value=f"=D{i}-E{i}")
        apply_style(g, formula_cell_style())
        g.number_format = '"$"#,##0.00'

        h = ws.cell(row=i, column=8, value=nights)
        apply_style(h, input_cell_style())
        h.alignment = Alignment(horizontal="center", vertical="center")

        i_cell = ws.cell(row=i, column=9, value=notes if notes else None)
        apply_style(i_cell, input_cell_style())

        ws.row_dimensions[i].height = 16

    # Blank capacity rows (continue past sample data through row 1005)
    last_data_row = len(sample_rows) + 5
    for row_idx in range(last_data_row + 1, 1006):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx in (4, 5, 6):
                cell.number_format = '"$"#,##0.00'
            if col_idx == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        g = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}-E{row_idx}")
        apply_style(g, formula_cell_style())
        g.number_format = '"$"#,##0.00'

    add_dropdown(ws, "C6:C1005", '"Airbnb,VRBO,Booking.com,Direct,Other"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_expense_log(wb, variant):
    """Sheet 3 — Expense Log (23 sample expenses in DEMO + 2000-row capacity)."""
    ws = wb.create_sheet("Expense Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Expense Log",
                         prev_tab="Revenue Log", next_tab="Monthly P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per expense — pick a Schedule E category from the dropdown"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 24), ("C", 38), ("D", 12),
        ("E", 16), ("F", 10), ("G", 28),
    ])

    headers = [
        "Date", "Vendor", "Category (Schedule E line)",
        "Amount", "Payment Method", "Receipt?", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Rows 6+: sample expense data (DEMO variant only; BLANK starts empty)
    sample_rows = SAMPLE_EXPENSES if variant == "demo" else []
    for i, e in enumerate(sample_rows, start=6):
        date_val, vendor, category, amount, method, receipt, notes = e

        a = ws.cell(row=i, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=vendor)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=category)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=amount)
        apply_style(d, input_cell_style())
        d.number_format = '"$"#,##0.00'

        e_cell = ws.cell(row=i, column=5, value=method)
        apply_style(e_cell, input_cell_style())

        f = ws.cell(row=i, column=6, value=receipt)
        apply_style(f, input_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        g = ws.cell(row=i, column=7, value=notes if notes else None)
        apply_style(g, input_cell_style())

        ws.row_dimensions[i].height = 16

    # Blank capacity rows (continue past sample data through row 2005)
    last_data_row = len(sample_rows) + 5
    for row_idx in range(last_data_row + 1, 2006):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 4:
                cell.number_format = '"$"#,##0.00'
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    add_dropdown(ws, "C6:C2005", "=Settings!$A$15:$A$31")
    add_dropdown(ws, "E6:E2005", '"Venmo,Zelle,Check,Cash,ACH,Credit Card,Auto-deduct,Other"')
    add_dropdown(ws, "F6:F2005", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_monthly_pl(wb):
    """Sheet 4 — Monthly P&L (12-column matrix; uses Settings!$B$5 for year)."""
    ws = wb.create_sheet("Monthly P&L")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Monthly P&L",
                         prev_tab="Expense Log", next_tab="Schedule E Summary")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Auto-rolled from logs — change 'Active tax year' on Settings to switch years"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    col_widths = [("A", 38)] + [(get_column_letter(2 + i), 10) for i in range(12)] + [("N", 12)]
    set_col_widths(ws, col_widths)

    headers = ["Category"] + MONTHS + ["YTD"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 8

    # Row 7: Revenue
    rev_label = ws.cell(row=7, column=1, value="Rents + cleaning fees collected")
    rev_label.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    rev_label.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 18

    # Active tax year drives every monthly aggregate. Without this cell,
    # YEAR(TODAY()) ties the dashboard to Excel's clock — sample data
    # silently shows $0 when system date is the wrong year.
    for m in range(1, 13):
        col = m + 1
        if m < 12:
            end_year_expr = "Settings!$B$5"
            end_month = m + 1
        else:
            end_year_expr = "Settings!$B$5+1"
            end_month = 1

        formula = (
            f"=SUMIFS('Revenue Log'!$D:$D,"
            f"'Revenue Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{m},1),"
            f"'Revenue Log'!$A:$A,\"<\"&DATE({end_year_expr},{end_month},1))"
            f"+SUMIFS('Revenue Log'!$F:$F,"
            f"'Revenue Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{m},1),"
            f"'Revenue Log'!$A:$A,\"<\"&DATE({end_year_expr},{end_month},1))"
        )
        cell = ws.cell(row=7, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'

    ytd_rev = ws.cell(row=7, column=14, value="=SUM(B7:M7)")
    apply_style(ytd_rev, formula_cell_style())
    ytd_rev.number_format = '"$"#,##0'
    ytd_rev.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)

    ws.row_dimensions[8].height = 8

    exp_hdr = ws.cell(row=9, column=1, value="EXPENSES")
    exp_hdr.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[9].height = 20

    # Rows 10-26: one row per expense category
    for idx, cat in enumerate(EXPENSE_CATEGORIES):
        row = 10 + idx
        a = ws.cell(row=row, column=1, value=cat)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16

        for m in range(1, 13):
            col = m + 1
            if m < 12:
                end_year_expr = "Settings!$B$5"
                end_month = m + 1
            else:
                end_year_expr = "Settings!$B$5+1"
                end_month = 1

            formula = (
                f"=SUMIFS('Expense Log'!$D:$D,"
                f"'Expense Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{m},1),"
                f"'Expense Log'!$A:$A,\"<\"&DATE({end_year_expr},{end_month},1),"
                f"'Expense Log'!$C:$C,A{row})"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'

        ytd = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(ytd, formula_cell_style())
        ytd.number_format = '"$"#,##0'

    last_exp_row = 10 + len(EXPENSE_CATEGORIES) - 1  # = 26
    tot_row = last_exp_row + 2                         # = 28
    ws.row_dimensions[last_exp_row + 1].height = 8

    tot_label = ws.cell(row=tot_row, column=1, value="TOTAL EXPENSES")
    tot_label.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    tot_label.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tot_row].height = 20

    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col, value=f"=SUM({col_letter}10:{col_letter}26)")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_ERROR)

    net_row = tot_row + 2
    ws.row_dimensions[tot_row + 1].height = 8

    net_label = ws.cell(row=net_row, column=1, value="NET INCOME (LOSS)")
    net_label.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    net_label.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[net_row].height = 22

    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=net_row, column=col, value=f"={col_letter}7-{col_letter}{tot_row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)

    ws.conditional_formatting.add(
        f"B{net_row}:N{net_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                    fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        f"B{net_row}:N{net_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                    fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )

    ws.freeze_panes = "B6"

    # Print setup
    ws.print_area = f"A1:N{net_row + 2}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_schedule_e_summary(wb):
    """Sheet 5 — Schedule E Summary (IRS line-mapped YTD totals)."""
    ws = wb.create_sheet("Schedule E Summary")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Schedule E Summary",
                         prev_tab="Monthly P&L", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "IRS Schedule E Part I line totals — hand directly to your CPA"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 50), ("B", 16)])

    bold_body = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    left_align = Alignment(horizontal="left", vertical="center")

    a5 = ws.cell(row=5, column=1, value="Tax Year:")
    a5.font = bold_body
    a5.alignment = left_align
    b5 = ws.cell(row=5, column=2, value="=Settings!$B$5")
    apply_style(b5, formula_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    a6 = ws.cell(row=6, column=1, value="Property:")
    a6.font = bold_body
    a6.alignment = left_align
    b6 = ws.cell(row=6, column=2, value="='Property Info'!B5")
    apply_style(b6, formula_cell_style())
    ws.row_dimensions[6].height = 18

    ws.row_dimensions[7].height = 8

    a8 = ws.cell(row=8, column=1, value="Line 3 — Rents received")
    a8.font = bold_body
    a8.alignment = left_align
    b8 = ws.cell(row=8, column=2, value="='Monthly P&L'!N7")
    apply_style(b8, formula_cell_style())
    b8.number_format = '"$"#,##0'
    b8.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_TEXT)
    ws.row_dimensions[8].height = 18

    a9 = ws.cell(row=9, column=1, value="Line 4 — Royalties")
    a9.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    a9.alignment = left_align
    b9 = ws.cell(row=9, column=2, value=0)
    apply_style(b9, input_cell_style())
    b9.number_format = '"$"#,##0'
    ws.row_dimensions[9].height = 18

    ws.row_dimensions[10].height = 8

    hdr11 = ws.cell(row=11, column=1, value="EXPENSES (Schedule E Part I)")
    hdr11.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    hdr11.alignment = left_align
    ws.row_dimensions[11].height = 20

    for idx, cat in enumerate(EXPENSE_CATEGORIES):
        row = 12 + idx
        monthly_row = 10 + idx

        a = ws.cell(row=row, column=1, value=cat)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = left_align

        b = ws.cell(row=row, column=2, value=f"='Monthly P&L'!N{monthly_row}")
        apply_style(b, formula_cell_style())
        b.number_format = '"$"#,##0'
        ws.row_dimensions[row].height = 16

    last_cat_row = 12 + len(EXPENSE_CATEGORIES) - 1  # = 28
    tot_row = last_cat_row + 2                         # = 30
    ws.row_dimensions[last_cat_row + 1].height = 8

    a30 = ws.cell(row=tot_row, column=1, value="Line 26a — Total expenses")
    a30.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    a30.alignment = left_align
    b30 = ws.cell(row=tot_row, column=2, value=f"=SUM(B12:B{last_cat_row})")
    apply_style(b30, formula_cell_style())
    b30.number_format = '"$"#,##0'
    b30.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_ERROR)
    ws.row_dimensions[tot_row].height = 20

    net_row = tot_row + 2  # = 32
    ws.row_dimensions[tot_row + 1].height = 8

    a32 = ws.cell(row=net_row, column=1, value="Line 26 — Income or (loss)")
    a32.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    a32.alignment = left_align
    b32 = ws.cell(row=net_row, column=2, value=f"=B8+B9-B{tot_row}")
    apply_style(b32, formula_cell_style())
    b32.number_format = '"$"#,##0'
    b32.font = Font(name=FONT_HEAD, size=14, bold=True, italic=True, color=COLOR_PRIMARY)
    ws.row_dimensions[net_row].height = 26

    # --- Schedule routing footnote (after net income) ---
    # Mirrors Mileage Log v2.3 pattern: the chosen schedule classification
    # on Settings drives where this number actually lands on the customer's
    # 1040. Pre-loaded with both common cases so the customer-CPA hand-off
    # is unambiguous.
    ws.row_dimensions[net_row + 1].height = 10
    footnote_row = net_row + 2
    ws.merge_cells(f"A{footnote_row}:B{footnote_row}")
    c = ws.cell(row=footnote_row, column=1)
    c.value = (
        '=IF(Settings!$B$15="Schedule E (passive)",'
        '"→ Schedule E (Form 1040), Line 26 — passive activity rules apply.",'
        'IF(Settings!$B$15="Schedule C (active)",'
        '"→ Schedule C, Line 31 — SE tax (15.3%) applies on net earnings.",'
        '"→ Schedule classification not set — see Settings."))'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[footnote_row].height = 32

    # --- v2.3 Audit-defense decision block (rows footnote_row+2 ... +14) ---
    # Three top-of-mind questions every CPA asks before filing:
    #   (1) STR loophole eligible? (avg stay ≤ 7 nights → IRC §469 escape)
    #   (2) Personal use within 14-day/10% safe harbor? (Pub 527)
    #   (3) QBI safe harbor met? (250+ hrs of service per Rev Proc 2019-38)
    block_start = footnote_row + 2
    sect = ws.cell(row=block_start, column=1, value="Audit-Defense Checks")
    sect.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[block_start].height = 22

    bold_body = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    left_align = Alignment(horizontal="left", vertical="center")

    # (1) Avg stay — drives STR loophole eligibility under IRC §469
    r = block_start + 1
    a = ws.cell(row=r, column=1, value="Avg stay (nights):")
    a.font = bold_body
    a.alignment = left_align
    b = ws.cell(
        row=r, column=2,
        value="=IFERROR(SUM('Revenue Log'!H6:H1005)/COUNTA('Revenue Log'!A6:A1005),0)"
    )
    apply_style(b, formula_cell_style())
    b.number_format = "0.0"
    ws.row_dimensions[r].height = 18

    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    c_eligible = ws.cell(row=r, column=1)
    c_eligible.value = (
        f'=IF(B{r-1}=0,"→ Add bookings to Revenue Log to compute eligibility.",'
        f'IF(B{r-1}<=7,'
        f'"✓ Avg stay ≤7 nights — STR loophole eligible (IRC §469 escape).",'
        f'"⚠ Avg stay >7 nights — passive activity rules apply; losses limited."))'
    )
    c_eligible.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
    c_eligible.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c_eligible.alignment = Alignment(horizontal="left", vertical="center",
                                      wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 28

    # (2) 14-day / 10% personal-use test (Pub 527)
    r += 1
    a = ws.cell(row=r, column=1, value="14-day / 10% threshold:")
    a.font = bold_body
    a.alignment = left_align
    b = ws.cell(
        row=r, column=2,
        value="=MAX(14, 'Property Info'!B16*0.1)"
    )
    apply_style(b, formula_cell_style())
    b.number_format = "0"
    ws.row_dimensions[r].height = 18

    r += 1
    a = ws.cell(row=r, column=1, value="Personal use days YTD:")
    a.font = bold_body
    a.alignment = left_align
    b = ws.cell(row=r, column=2, value="='Property Info'!B17")
    apply_style(b, formula_cell_style())
    b.number_format = "0"
    ws.row_dimensions[r].height = 18

    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    c_resi = ws.cell(row=r, column=1)
    c_resi.value = (
        f'=IF(B{r-1}<=B{r-2},'
        f'"✓ Personal use within safe harbor — full deductibility (Pub 527).",'
        f'"⚠ Personal use exceeds 14 days OR 10% of rental days — property is a '
        f'\\"residence\\"; rental losses cannot offset other income, excess '
        f'expenses carry forward.")'
    )
    c_resi.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
    c_resi.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c_resi.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 36

    # (3) QBI safe-harbor flag (Rev Proc 2019-38)
    r += 1
    a = ws.cell(row=r, column=1, value="QBI §199A safe harbor met (250+ hrs):")
    a.font = bold_body
    a.alignment = left_align
    b = ws.cell(row=r, column=2, value="='Settings'!$B$16")
    apply_style(b, formula_cell_style())
    ws.row_dimensions[r].height = 18

    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    c_qbi = ws.cell(row=r, column=1)
    c_qbi.value = (
        f'=IF(B{r-1}="Yes",'
        f'"✓ Eligible for the 20% QBI deduction (§199A).",'
        f'IF(B{r-1}="No",'
        f'"→ Not eligible — track service hours on Mileage Log + ops tools.",'
        f'"→ Set Settings!B19 once you know your service hours."))'
    )
    c_qbi.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c_qbi.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 24

    # NIIT note (informational; no field to set since it depends on AGI)
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    c_niit = ws.cell(row=r, column=1)
    c_niit.value = (
        "ℹ NIIT (3.8%) applies to passive rental income above $200K single / "
        "$250K MFJ AGI thresholds. Material participation removes this — see "
        "Mileage Log YTD Hours."
    )
    c_niit.font = italic_muted
    c_niit.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 28

    audit_block_end = r

    # --- Chart: Donut — Expense composition (anchored D5) ---
    donut = DoughnutChart()
    donut.title = "Expense Composition (YTD)"
    donut.height = 9
    donut.width = 11
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=2, min_row=11,
                            max_row=last_cat_row, max_col=2)
    donut_cats = Reference(ws, min_col=1, min_row=12,
                            max_row=last_cat_row)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=False, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, "D5")

    ws.print_area = f"A1:B{audit_block_end + 2}"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    """Sheet 6 — Settings (active tax year, Schedule classification, category list, archive)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Schedule E Summary", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · Schedule classification · expense list · year-end archive"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 38), ("B", 16), ("C", 14), ("D", 16),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active tax year (row 5) — drives every monthly SUMIFS on P&L ---
    a5 = ws.cell(row=5, column=1, value="Active tax year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=ACTIVE_TAX_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    # Row 6: explainer
    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = (
        "Bump this each January. Every monthly aggregate on the Monthly P&L tab "
        "filters by this year — without it, dashboards return $0 when the system "
        "clock doesn't match logged dates."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 32

    ws.row_dimensions[7].height = 8

    # --- Tax Classification (rows 8-15) ---
    sect8 = ws.cell(row=8, column=1, value="Tax Classification")
    sect8.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[8].height = 22

    # Home office toggle
    a9 = ws.cell(row=9, column=1,
                  value="Home office is principal place of business:")
    a9.font = bold_right
    a9.alignment = Alignment(horizontal="right", vertical="center")
    b9 = ws.cell(row=9, column=2, value="Not yet established")
    apply_style(b9, input_cell_style())
    add_dropdown(ws, "B9", '"Yes,No,Not yet established"')
    ws.row_dimensions[9].height = 18

    ws.merge_cells("A10:D10")
    c10 = ws["A10"]
    c10.value = (
        "If 'Yes', residence-to-property trips are deductible business travel. "
        "If 'No', they're commuting (non-deductible). Documenting this election "
        "is on you, not the IRS."
    )
    c10.font = italic_muted
    c10.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[10].height = 28

    ws.row_dimensions[11].height = 6

    # Substantial-services callout (read it before picking the schedule)
    ws.merge_cells("A12:D13")
    c12 = ws["A12"]
    c12.value = (
        "⚠ Substantial-services trigger: if you provide cleaning DURING stays, "
        "meals, or concierge, IRS treats this as a hospitality business — "
        "Schedule C is required (15.3% SE tax applies)."
    )
    c12.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c12.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[12].height = 16
    ws.row_dimensions[13].height = 16

    ws.row_dimensions[14].height = 6

    # Schedule classification dropdown at B15 — referenced by the
    # Schedule E Summary footnote and any future routing logic.
    a15 = ws.cell(row=15, column=1, value="Schedule classification:")
    a15.font = bold_right
    a15.alignment = Alignment(horizontal="right", vertical="center")
    b15 = ws.cell(row=15, column=2, value="Ask my CPA")
    apply_style(b15, input_cell_style())
    add_dropdown(ws, "B15",
                  '"Schedule E (passive),Schedule C (active),Ask my CPA"')
    ws.row_dimensions[15].height = 18

    # v2.3 — QBI §199A safe-harbor flag (Rev Proc 2019-38)
    # Read by the Audit-Defense Checks block on Schedule E Summary.
    a16 = ws.cell(row=16, column=1, value="QBI safe harbor met (250+ service hrs):")
    a16.font = bold_right
    a16.alignment = Alignment(horizontal="right", vertical="center")
    b16 = ws.cell(row=16, column=2, value="Not sure")
    apply_style(b16, input_cell_style())
    add_dropdown(ws, "B16", '"Yes,No,Not sure"')
    ws.row_dimensions[16].height = 18

    # --- Expense category source list (rows 17-35) ---
    # Old v1 dropdown referenced A15:A31; v2.2 moves the list to A19:A35
    # so the Schedule classification block above doesn't collide. The
    # build_workbook() main rewires the Expense Log dropdown to match.
    sect17 = ws.cell(row=17, column=1, value="Expense category list (Schedule E)")
    sect17.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[17].height = 22

    cat_help = ws.cell(row=18, column=1,
                        value="(Source for the Expense Log dropdown — don't reorder)")
    cat_help.font = italic_muted
    ws.row_dimensions[18].height = 16

    for idx, cat in enumerate(EXPENSE_CATEGORIES):
        row = 19 + idx
        cell = ws.cell(row=row, column=1, value=cat)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16

    last_cat_row = 19 + len(EXPENSE_CATEGORIES) - 1  # 35

    # --- Year-end Archive (rows 37-45) ---
    sect37 = ws.cell(row=37, column=1, value="Year-end Archive")
    sect37.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[37].height = 22

    ws.merge_cells("A38:D38")
    c38 = ws["A38"]
    c38.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then bump 'Active tax year' above and clear the logs for the new year."
    )
    c38.font = italic_muted
    c38.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[38].height = 28

    archive_headers = ["Year", "Revenue", "Expenses", "Net"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=39, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[39].height = 18

    for idx, year in enumerate(range(2024, 2031), start=40):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[idx].height = 16

    return last_cat_row


# ---------------------------------------------------------------------------
# Main build function (shared by Lite + Full)
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_info_tab(wb, variant)
    build_revenue_log(wb, variant)
    build_expense_log(wb, variant)
    build_monthly_pl(wb)
    build_schedule_e_summary(wb)
    last_cat_row = build_settings_tab(wb)

    # Re-wire Expense Log dropdown to point at the new category range.
    # The old range was A15:A31; v2.2 moves categories to A19:A35 so the
    # Schedule classification block can sit at A15.
    expense_ws = wb["Expense Log"]
    new_dv = DataValidation(
        type="list",
        formula1=f"=Settings!$A$19:$A${last_cat_row}",
        allow_blank=True,
    )
    new_dv.add("C6:C2005")
    # Drop the stale dropdown that points at A15:A31 by replacing the
    # data_validations list. (openpyxl re-emits the full list on save.)
    keep = [dv for dv in expense_ws.data_validations.dataValidation
            if "C6:C2005" not in str(dv.sqref)]
    expense_ws.data_validations.dataValidation = keep
    expense_ws.add_data_validation(new_dv)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Single-Property P&L Tracker{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = "Single-property P&L with Schedule E category mapping (v2.3)."

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
