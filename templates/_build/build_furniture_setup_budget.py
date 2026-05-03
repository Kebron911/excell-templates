"""Build ACQ-007 Furniture / Setup Budget Calculator (operational mode).

Pre-furnishing budget for a new STR. Per-room item register with cost
benchmarks, MACRS class flag (feeds TAX-013), and 60-day operating
reserve so the host doesn't accidentally spend furniture money on
month-1 cleaning.

Tabs (5):
  1. Start         — KPIs (budget, spent, room completion %, $ to first stay)
  2. Room Detail   — per-room item register (capacity 200 items)
  3. Per-Room Roll-up — totals + completion % + bar chart "$ by room"
  4. Operating Reserve — 60-day setup-phase costs
  5. Settings      — property, dates, MACRS class options, year-end archive

Generates two files:
  templates/_masters/ACQ-007-furniture-setup-budget-DEMO.xlsx
  templates/_masters/ACQ-007-furniture-setup-budget-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "ACQ-007"
NAME = "furniture-setup-budget"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Room dropdown
ROOMS = [
    "Living", "Kitchen", "Master BR", "BR2", "BR3",
    "Master Bath", "Bath2", "Patio/Outdoor", "Laundry",
    "Decor", "Tech",
]
ROOM_LIST_FORMULA = '"' + ",".join(ROOMS) + '"'

STATUS_LIST = ["Researching", "Ordered", "In transit", "Installed"]
STATUS_FORMULA = '"' + ",".join(STATUS_LIST) + '"'

MACRS_LIST = ["5-yr", "7-yr", "Decor — expense"]
MACRS_FORMULA = '"' + ",".join(MACRS_LIST) + '"'

ITEM_CAPACITY = 200  # rows in the Room Detail register

# ---------------------------------------------------------------------------
# Sample data — Smokies Ridge 4BR cabin (95 line items, 11 rooms)
# (room, item, description, est, actual, vendor, status, macrs, notes)
# ---------------------------------------------------------------------------
SAMPLE_ROWS = [
    # Living (12)
    ("Living", "Sectional sofa", "Ashley Bladen 3pc",        1850, 1925, "Ashley HomeStore",   "Installed",   "7-yr", ""),
    ("Living", "Coffee table", "Reclaimed wood, 48\"",        420,  385, "Wayfair",            "Installed",   "7-yr", ""),
    ("Living", "Accent chairs (pair)", "Swivel barrel",       780,  840, "Article",            "Installed",   "7-yr", ""),
    ("Living", "End tables (pair)", "Live-edge",              340,  360, "World Market",       "Installed",   "7-yr", ""),
    ("Living", "Floor + table lamps (3)", "Brass + linen",    280,  312, "Target",             "Installed",   "7-yr", ""),
    ("Living", "65\" smart TV", "TCL 65Q",                    520,  498, "Costco",             "Installed",   "5-yr", ""),
    ("Living", "TV media console", "Mid-century walnut",      450,  475, "West Elm",           "Installed",   "7-yr", ""),
    ("Living", "Throw blankets + pillows", "Mountain palette", 220,  245, "HomeGoods",         "Installed",   "Decor — expense", ""),
    ("Living", "Wall art / gallery", "5-piece mountain set",  380,  420, "Etsy",               "Installed",   "Decor — expense", ""),
    ("Living", "Area rug 9x12", "Wool, neutral",              640,  680, "Rugs USA",           "Installed",   "7-yr", ""),
    ("Living", "Curtains + rods", "Linen blackout",           240,  265, "Amazon",             "Installed",   "Decor — expense", ""),
    ("Living", "Board games + books", "Family mix (8 games)", 180,  195, "Target",             "Installed",   "Decor — expense", ""),

    # Kitchen (15)
    ("Kitchen", "Dish set service-12", "Stoneware",            260,  285, "Crate & Barrel",    "Installed",   "Decor — expense", ""),
    ("Kitchen", "Drink + wine glasses", "24 pieces",           140,  155, "IKEA",              "Installed",   "Decor — expense", ""),
    ("Kitchen", "Coffee mugs (12)", "Stoneware",                 70,   72, "Target",             "Installed",   "Decor — expense", ""),
    ("Kitchen", "Silverware service-12", "Brushed steel",       120,  138, "Costco",            "Installed",   "Decor — expense", ""),
    ("Kitchen", "Cookware 10pc set", "Calphalon nonstick",     420,  395, "Williams-Sonoma",   "Installed",   "5-yr", ""),
    ("Kitchen", "Bakeware 6pc", "USA Pan",                      120,  145, "Amazon",            "Installed",   "Decor — expense", ""),
    ("Kitchen", "Knife block + 8 knives", "Henckels",          240,  265, "Costco",            "Installed",   "5-yr", ""),
    ("Kitchen", "Coffee maker + grinder", "Breville Drip",     280,  299, "Best Buy",          "Installed",   "5-yr", ""),
    ("Kitchen", "Toaster + kettle", "Cuisinart",                160,  175, "Target",            "Installed",   "5-yr", ""),
    ("Kitchen", "Microwave countertop", "1.6 cu ft",            180,  195, "Best Buy",          "Installed",   "5-yr", ""),
    ("Kitchen", "Blender / immersion", "Vitamix E310",          340,  329, "Costco",            "Ordered",     "5-yr", ""),
    ("Kitchen", "Storage containers", "Glass + lids",            85,   92, "OXO/Amazon",        "Installed",   "Decor — expense", ""),
    ("Kitchen", "Trash + recycling bins", "Simplehuman",         180,  195, "Bed Bath",          "Installed",   "5-yr", ""),
    ("Kitchen", "Dish towels + mitts", "Bulk pack",               60,   55, "Target",            "Installed",   "Decor — expense", ""),
    ("Kitchen", "Spice + bar starter", "12-pc + shaker",          90,   85, "Amazon",            "Installed",   "Decor — expense", ""),

    # Master BR (10)
    ("Master BR", "King mattress + boxspring", "Tempur-Cloud", 1800, 1950, "Mattress Firm",     "Installed",   "7-yr", ""),
    ("Master BR", "King bed frame + headboard", "Upholstered", 720,  765, "West Elm",          "Installed",   "7-yr", ""),
    ("Master BR", "Sheets + duvet sets (3)", "Brooklinen",     420,  445, "Brooklinen",        "Installed",   "Decor — expense", ""),
    ("Master BR", "Pillows + shams", "Down + decorative",      180,  205, "Pottery Barn",      "Installed",   "Decor — expense", ""),
    ("Master BR", "Nightstands (pair)", "Mid-century walnut",  480,  520, "West Elm",          "Installed",   "7-yr", ""),
    ("Master BR", "Bedside lamps", "Ceramic, pair",            160,  175, "Target",            "Installed",   "7-yr", ""),
    ("Master BR", "Dresser / chest 6-drawer", "Walnut",        680,  720, "World Market",      "Installed",   "7-yr", ""),
    ("Master BR", "Blackout curtains", "Linen, ceiling",       180,  220, "Amazon",            "Installed",   "Decor — expense", ""),
    ("Master BR", "Area rug 8x10", "Wool",                     420,  445, "Rugs USA",          "Installed",   "7-yr", ""),
    ("Master BR", "Wall art + mirror", "Mountain triptych",    260,  285, "Etsy",              "Installed",   "Decor — expense", ""),

    # BR2 (8)
    ("BR2", "Queen mattress", "Saatva Loom",                   950, 1095, "Saatva",            "Installed",   "7-yr", ""),
    ("BR2", "Queen bed frame", "Platform, oak",                380,  410, "Article",           "Installed",   "7-yr", ""),
    ("BR2", "Sheets + duvet (2 sets)", "Cotton sateen",        260,  285, "Target",            "Installed",   "Decor — expense", ""),
    ("BR2", "Nightstands (pair)", "Rattan",                    280,  310, "Wayfair",           "Installed",   "7-yr", ""),
    ("BR2", "Dresser 4-drawer", "Mid-century",                 480,  520, "World Market",      "In transit",  "7-yr", ""),
    ("BR2", "Lamps + curtains", "Pair + blackout",             220,  235, "Target",            "Installed",   "Decor — expense", ""),
    ("BR2", "Area rug 6x9", "Wool",                            260,  285, "Rugs USA",          "Installed",   "7-yr", ""),
    ("BR2", "Wall art + mirror", "Trio",                       180,  195, "Etsy",              "Installed",   "Decor — expense", ""),

    # BR3 (8)
    ("BR3", "Bunk bed (twin/full)", "Solid pine",              720,  785, "Pottery Barn Kids", "Installed",   "7-yr", ""),
    ("BR3", "Twin + full mattresses", "Saatva Youth",          780,  820, "Saatva",            "Installed",   "7-yr", ""),
    ("BR3", "Sheets + duvets (3 sets)", "Cotton",              280,  310, "Target",            "Installed",   "Decor — expense", ""),
    ("BR3", "Nightstand", "Single, rattan",                    180,  195, "Wayfair",           "Installed",   "7-yr", ""),
    ("BR3", "Dresser 3-drawer", "White",                       320,  340, "IKEA",              "Installed",   "7-yr", ""),
    ("BR3", "Lamp + nightlight", "Kid-safe",                    80,   95, "Amazon",            "Installed",   "Decor — expense", ""),
    ("BR3", "Area rug 5x7", "Washable",                        180,  195, "Ruggable",          "Installed",   "7-yr", ""),
    ("BR3", "Wall art / shelves", "Adventure theme",           160,  185, "Etsy",              "Ordered",     "Decor — expense", ""),

    # Master Bath (8)
    ("Master Bath", "Towels (8 sets)", "Turkish cotton",        220,  245, "Brooklinen",        "Installed",   "Decor — expense", ""),
    ("Master Bath", "Bath mats + rugs", "Pair memory foam",      80,   92, "Target",            "Installed",   "Decor — expense", ""),
    ("Master Bath", "Shower curtain + liner", "Linen",          120,  135, "West Elm",          "Installed",   "Decor — expense", ""),
    ("Master Bath", "Soap dispensers + accessories", "Stoneware", 90,  105, "HomeGoods",        "Installed",   "Decor — expense", ""),
    ("Master Bath", "Trash + toilet brush", "Simplehuman",       110,  120, "Costco",            "Installed",   "5-yr", ""),
    ("Master Bath", "Hairdryer + amenity tray", "Salon-grade",  120,  135, "Sephora",           "Installed",   "5-yr", ""),
    ("Master Bath", "Wall art + mirror", "Brass framed",         160,  185, "Etsy",              "Installed",   "Decor — expense", ""),
    ("Master Bath", "Robes (pair) + slippers", "Waffle",         140,  155, "Brooklinen",        "Installed",   "Decor — expense", ""),

    # Bath2 (6)
    ("Bath2", "Towels (6 sets)", "Standard cotton",             140,  155, "Target",            "Installed",   "Decor — expense", ""),
    ("Bath2", "Bath mats", "Memory foam",                        50,   58, "Target",            "Installed",   "Decor — expense", ""),
    ("Bath2", "Shower curtain + liner", "Cotton",                70,   80, "Target",            "Installed",   "Decor — expense", ""),
    ("Bath2", "Accessories set", "Stoneware",                    70,   85, "Bed Bath",          "Installed",   "Decor — expense", ""),
    ("Bath2", "Trash + toilet brush", "Standard",                 60,   68, "Target",            "Installed",   "5-yr", ""),
    ("Bath2", "Wall art + hooks", "Mountain prints",              80,   95, "Etsy",              "Installed",   "Decor — expense", ""),

    # Patio/Outdoor (10)
    ("Patio/Outdoor", "Patio dining set 6-seat", "Teak + cushions", 1450, 1620, "Frontgate",   "Installed",   "7-yr", ""),
    ("Patio/Outdoor", "Lounge chairs (4)", "Adirondack composite", 760,  825, "Polywood",      "Installed",   "7-yr", ""),
    ("Patio/Outdoor", "Patio umbrella + base", "9ft cantilever",   320,  345, "Costco",        "Installed",   "7-yr", ""),
    ("Patio/Outdoor", "Gas grill + tools", "Weber Spirit",         620,  680, "Home Depot",    "Installed",   "5-yr", ""),
    ("Patio/Outdoor", "Fire pit + chairs", "Solo Stove Bonfire",   480,  525, "Solo Stove",    "Installed",   "7-yr", ""),
    ("Patio/Outdoor", "String lights + planters", "Outdoor LED",   220,  245, "Amazon",        "Installed",   "Decor — expense", ""),
    ("Patio/Outdoor", "Hot tub steps + accessories", "Cedar",      240,  280, "Hot Tub Outpost", "Ordered",  "5-yr", ""),
    ("Patio/Outdoor", "Outdoor rugs (2)", "Polypropylene",         180,  195, "Ruggable",      "Installed",   "Decor — expense", ""),
    ("Patio/Outdoor", "Cornhole + outdoor games", "Tournament",    160,  175, "Amazon",        "Installed",   "Decor — expense", ""),
    ("Patio/Outdoor", "Outdoor cushions storage", "Deck box",      180,  195, "Costco",        "Installed",   "7-yr", ""),

    # Laundry (5)
    ("Laundry", "Washer + dryer pair", "LG front-load",          1850, 1995, "Best Buy",       "Installed",   "5-yr", ""),
    ("Laundry", "Detergent + dryer sheets bulk", "HE",            120,  135, "Costco",         "Installed",   "Decor — expense", ""),
    ("Laundry", "Iron + ironing board", "Rowenta",                120,  140, "Bed Bath",       "Installed",   "5-yr", ""),
    ("Laundry", "Drying rack + hampers", "Wood",                   90,  105, "Target",         "Installed",   "7-yr", ""),
    ("Laundry", "Cleaning supplies starter", "Eco brands",        140,  165, "Grove",          "Installed",   "Decor — expense", ""),

    # Decor (6) — accent / staging
    ("Decor", "Welcome basket + signage", "Local goods",          120,  145, "Local",          "Installed",   "Decor — expense", ""),
    ("Decor", "Plants + planters (8)", "Faux + real mix",         260,  295, "IKEA",           "Installed",   "Decor — expense", ""),
    ("Decor", "Vases + table accessories", "Curated",             180,  210, "HomeGoods",      "Installed",   "Decor — expense", ""),
    ("Decor", "Books + magazines", "Coffee-table styling",         80,   95, "Local bookstore","Installed",   "Decor — expense", ""),
    ("Decor", "Candles + diffusers", "Unscented safe",             90,  108, "Target",         "Installed",   "Decor — expense", ""),
    ("Decor", "Welcome book + binder", "Custom printed",           60,   65, "Local print",    "Installed",   "Decor — expense", ""),

    # Tech (7)
    ("Tech", "Smart lock keyless", "Schlage Encode",              260,  285, "Home Depot",     "Installed",   "5-yr", ""),
    ("Tech", "Doorbell camera", "Ring Pro 2",                     220,  245, "Amazon",         "Installed",   "5-yr", ""),
    ("Tech", "Smart thermostat", "Ecobee Premium",                240,  275, "Best Buy",       "Installed",   "5-yr", ""),
    ("Tech", "WiFi mesh router", "Eero Pro 6E (3-pack)",          480,  525, "Amazon",         "Installed",   "5-yr", ""),
    ("Tech", "Streaming devices (3)", "Roku Ultra",               180,  195, "Best Buy",       "Installed",   "5-yr", ""),
    ("Tech", "Bluetooth speaker", "Sonos Era 100",                240,  265, "Sonos",          "Installed",   "5-yr", ""),
    ("Tech", "Noise sensor (Minut)", "STR monitoring",            220,  235, "Minut",          "Ordered",     "5-yr", ""),
]

# Operating reserve sample (8 lines)
RESERVE_SAMPLE = [
    ("Mortgage P&I × 60 days",        2400),
    ("Insurance (60-day proration)",   600),
    ("Utilities setup + 60 days",      700),
    ("Property tax accrual (60 days)", 480),
    ("PMS / pricing software (60d)",   180),
    ("Marketing prep + listing photos",750),
    ("Cleaning + supplies stock-up",   420),
    ("Contingency (5%)",               270),
]

# Settings
SETTINGS_SAMPLE = {
    "property":   "Smokies Ridge 4BR Cabin",
    "start":      "2026-03-01",
    "first_stay": "2026-05-01",
    "reserve_days": 60,
    "active_year": 2026,
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label, span="A:L"):
    first, last = span.split(":")
    ws.merge_cells(f"{first}{row}:{last}{row}")
    c = ws[f"{first}{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    from openpyxl.utils.cell import column_index_from_string
    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


# ---------------------------------------------------------------------------
# Tab 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Furniture / Setup Budget"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Every room, every item — what you'll spend before the first guest."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # What this does card
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 18

    ws.merge_cells("A12:L14")
    c = ws["A12"]
    c.value = (
        "Walks you room-by-room through every essential furnishing item with cost benchmarks. "
        "Tracks estimated vs actual spend, MACRS class per item (5-yr / 7-yr / Decor), and a "
        "60-day operating reserve so you don't run dry the week you launch."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 18
    ws.row_dimensions[14].height = 18

    # KPI grid (rows 17-21)
    alt_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(17, 22):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = alt_fill

    # 4 cards: Budget total, Spent, Room completion %, $ to first stay
    headers = [
        ("A17:C17", "BUDGET (EST)"),
        ("D17:F17", "SPENT (ACTUAL)"),
        ("G17:I17", "INSTALLED %"),
        ("J17:L17", "$ TO FIRST STAY"),
    ]
    for rng, label in headers:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = label
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 16

    # Budget total: SUM of E on Room Detail
    ws.merge_cells("A18:C20")
    c = ws["A18"]
    c.value = f"=SUM('Room Detail'!E6:E{5+ITEM_CAPACITY})"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("D18:F20")
    c = ws["D18"]
    c.value = f"=SUM('Room Detail'!F6:F{5+ITEM_CAPACITY})"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("G18:I20")
    c = ws["G18"]
    c.value = (
        f'=IFERROR(COUNTIF(\'Room Detail\'!H6:H{5+ITEM_CAPACITY},"Installed")/'
        f'COUNTA(\'Room Detail\'!C6:C{5+ITEM_CAPACITY}),0)'
    )
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0%"

    ws.merge_cells("J18:L20")
    c = ws["J18"]
    c.value = (
        f"=SUM('Room Detail'!F6:F{5+ITEM_CAPACITY})"
        f"+'Operating Reserve'!C16"
    )
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.row_dimensions[18].height = 22
    ws.row_dimensions[19].height = 16
    ws.row_dimensions[20].height = 16

    # Pseudo-button nav
    pseudo_button(ws, "A23", "C25", "Room Detail",
                  "'Room Detail'!A1", variant="primary")
    pseudo_button(ws, "D23", "F25", "Per-Room Roll-up",
                  "'Per-Room Roll-up'!A1", variant="primary")
    pseudo_button(ws, "G23", "I25", "Operating Reserve",
                  "'Operating Reserve'!A1", variant="primary")
    pseudo_button(ws, "J23", "L25", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(23, 26):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A27", "L29", "→  ADD A NEW ITEM",
                  "'Room Detail'!B6", variant="accent")
    for r in range(27, 30):
        ws.row_dimensions[r].height = 22

    # How to use
    ws.merge_cells("A32:L32")
    c = ws["A32"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A33:L35")
    c = ws["A33"]
    c.value = (
        "1) On Settings — set property name + first-stay date.   "
        "2) On Room Detail — log every item: room, name, est cost, MACRS class.   "
        "3) As you buy: fill Actual cost + status. KPIs above and Roll-up update live."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(33, 36):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 38,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L40"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 2 — Room Detail (register, capacity 200)
# ---------------------------------------------------------------------------

def build_room_detail_tab(wb, variant):
    ws = wb.create_sheet("Room Detail")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    set_col_widths(ws, [
        ("A", 5),    # Item #
        ("B", 16),   # Room
        ("C", 30),   # Item
        ("D", 28),   # Description
        ("E", 12),   # Estimated cost
        ("F", 12),   # Actual cost
        ("G", 18),   # Vendor
        ("H", 14),   # Status
        ("I", 18),   # MACRS class
        ("J", 20),   # Notes
        ("K", 4),
        ("L", 4),
    ])

    compact_header_band(ws, "Room Detail",
                        prev_tab="Start", next_tab="Per-Room Roll-up")

    # Row 4 spacer / instruction
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-item register · 200 rows · yellow = edit · gray = formula · "
        "Status \"Installed\" counts toward completion %."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 5
    headers = [
        "#", "Room", "Item", "Description / brand",
        "Est cost", "Actual cost", "Vendor / source", "Status",
        "MACRS class", "Notes",
    ]
    for i, label in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # Data rows 6 → 5 + ITEM_CAPACITY
    last_row = 5 + ITEM_CAPACITY

    sample_data = SAMPLE_ROWS if variant == "demo" else []

    for idx in range(ITEM_CAPACITY):
        row = 6 + idx
        # Auto-numbered # column (formula so it persists)
        a = ws.cell(row=row, column=1,
                    value=f'=IF(C{row}="","",ROW()-5)')
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.number_format = "0"

        # Pull sample if available
        if idx < len(sample_data):
            (room, item, desc, est, actual, vendor,
             status, macrs, notes) = sample_data[idx]
        else:
            room = item = desc = vendor = status = macrs = notes = None
            est = actual = None

        # B Room (dropdown)
        b = ws.cell(row=row, column=2, value=room)
        apply_style(b, input_cell_style())
        b.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # C Item
        c_cell = ws.cell(row=row, column=3, value=item)
        apply_style(c_cell, input_cell_style())
        c_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # D Description
        d_cell = ws.cell(row=row, column=4, value=desc)
        apply_style(d_cell, input_cell_style())
        d_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # E Estimated cost
        e_cell = ws.cell(row=row, column=5, value=est)
        apply_style(e_cell, input_cell_style())
        e_cell.alignment = Alignment(horizontal="right", vertical="center")
        e_cell.number_format = '"$"#,##0'

        # F Actual cost
        f_cell = ws.cell(row=row, column=6, value=actual)
        apply_style(f_cell, input_cell_style())
        f_cell.alignment = Alignment(horizontal="right", vertical="center")
        f_cell.number_format = '"$"#,##0'

        # G Vendor
        g_cell = ws.cell(row=row, column=7, value=vendor)
        apply_style(g_cell, input_cell_style())
        g_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # H Status (dropdown)
        h_cell = ws.cell(row=row, column=8, value=status)
        apply_style(h_cell, input_cell_style())
        h_cell.alignment = Alignment(horizontal="center", vertical="center")

        # I MACRS class (dropdown)
        i_cell = ws.cell(row=row, column=9, value=macrs)
        apply_style(i_cell, input_cell_style())
        i_cell.alignment = Alignment(horizontal="center", vertical="center")

        # J Notes
        j_cell = ws.cell(row=row, column=10, value=notes)
        apply_style(j_cell, input_cell_style())
        j_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Alternating row banding via parchment-alt fill on odd rows
        if idx % 2 == 1:
            band = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            # Apply band only where there's no input fill — keep input cells yellow
            ws.cell(row=row, column=1).fill = band
        ws.row_dimensions[row].height = 16

    # Data validations
    dv_room = DataValidation(type="list", formula1=ROOM_LIST_FORMULA, allow_blank=True)
    dv_room.add(f"B6:B{last_row}")
    ws.add_data_validation(dv_room)

    dv_status = DataValidation(type="list", formula1=STATUS_FORMULA, allow_blank=True)
    dv_status.add(f"H6:H{last_row}")
    ws.add_data_validation(dv_status)

    dv_macrs = DataValidation(type="list", formula1=MACRS_FORMULA, allow_blank=True)
    dv_macrs.add(f"I6:I{last_row}")
    ws.add_data_validation(dv_macrs)

    # Freeze header
    ws.freeze_panes = "A6"

    # Print
    ws.print_area = f"A1:J{last_row}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 3 — Per-Room Roll-up
# ---------------------------------------------------------------------------

def build_rollup_tab(wb, variant):
    ws = wb.create_sheet("Per-Room Roll-up")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    set_col_widths(ws, [
        ("A", 22),   # Room
        ("B", 14),   # Est total
        ("C", 14),   # Actual total
        ("D", 12),   # Item count
        ("E", 14),   # Installed count
        ("F", 14),   # Completion %
        ("G", 4),
        ("H", 4), ("I", 4), ("J", 4), ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Per-Room Roll-up",
                        prev_tab="Room Detail", next_tab="Operating Reserve")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Totals by room · bar chart \"$ by room\" anchored right of table."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 5
    headers = ["Room", "Est total", "Actual total", "Items", "Installed", "Completion %"]
    for i, label in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 22

    last_row = 5 + ITEM_CAPACITY

    # One row per room
    for i, room in enumerate(ROOMS):
        row = 6 + i
        a = ws.cell(row=row, column=1, value=room)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        b = ws.cell(row=row, column=2,
                    value=f"=SUMIF('Room Detail'!B6:B{last_row},A{row},'Room Detail'!E6:E{last_row})")
        apply_style(b, formula_cell_style())
        b.number_format = '"$"#,##0'
        b.alignment = Alignment(horizontal="right", vertical="center")

        c_cell = ws.cell(row=row, column=3,
                         value=f"=SUMIF('Room Detail'!B6:B{last_row},A{row},'Room Detail'!F6:F{last_row})")
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = '"$"#,##0'
        c_cell.alignment = Alignment(horizontal="right", vertical="center")

        d = ws.cell(row=row, column=4,
                    value=f"=COUNTIF('Room Detail'!B6:B{last_row},A{row})")
        apply_style(d, formula_cell_style())
        d.number_format = "0"
        d.alignment = Alignment(horizontal="center", vertical="center")

        e_cell = ws.cell(row=row, column=5,
                         value=(f"=COUNTIFS('Room Detail'!B6:B{last_row},A{row},"
                                f"'Room Detail'!H6:H{last_row},\"Installed\")"))
        apply_style(e_cell, formula_cell_style())
        e_cell.number_format = "0"
        e_cell.alignment = Alignment(horizontal="center", vertical="center")

        f_cell = ws.cell(row=row, column=6,
                         value=f"=IFERROR(E{row}/D{row},0)")
        apply_style(f_cell, formula_cell_style())
        f_cell.number_format = "0.0%"
        f_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18

    # Totals row
    total_row = 6 + len(ROOMS)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT
    )
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.cell(row=total_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    for col, col_letter in [(2, "B"), (3, "C"), (4, "D"), (5, "E")]:
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({col_letter}6:{col_letter}{total_row-1})")
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right" if col in (2, 3) else "center",
                                   vertical="center")
        if col in (2, 3):
            cell.number_format = '"$"#,##0'
        else:
            cell.number_format = "0"
    cell = ws.cell(row=total_row, column=6,
                   value=f"=IFERROR(E{total_row}/D{total_row},0)")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0%"
    ws.row_dimensions[total_row].height = 22

    # Bar chart "$ by room" — anchored at H5
    bar = BarChart()
    bar.type = "col"
    bar.title = "$ by Room (Actual)"
    bar.y_axis.title = "Total $"
    bar.x_axis.title = "Room"
    bar.legend = None
    bar.height = 10
    bar.width = 18
    data_ref = Reference(ws, min_col=3, min_row=5,
                         max_row=5 + len(ROOMS), max_col=3)
    cats_ref = Reference(ws, min_col=1, min_row=6,
                         max_row=5 + len(ROOMS))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "low"
    style_chart(bar)
    ws.add_chart(bar, "H5")

    brand_footer(ws, total_row + 4,
                 version_line=f"{SKU} · Per-Room Roll-up")

    # Print
    ws.print_area = f"A1:T{total_row + 8}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 4 — Operating Reserve
# ---------------------------------------------------------------------------

def build_reserve_tab(wb, variant):
    ws = wb.create_sheet("Operating Reserve")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 16),
        ("D", 4), ("E", 38),
        ("F", 4), ("G", 4), ("H", 4), ("I", 4),
        ("J", 4), ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Operating Reserve",
                        prev_tab="Per-Room Roll-up", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "60-day setup-phase costs — fixed costs you'll pay BEFORE bookings stabilize. "
        "This is your \"$ to first stay\" reserve."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    _section_band(ws, 6, "SETUP-PHASE OPERATING COSTS")

    # 8 input rows starting row 7
    for idx, (label, value) in enumerate(RESERVE_SAMPLE):
        row = 7 + idx
        _input_row(ws, row, f"{label}:",
                   _val(variant, value), '"$"#,##0', "")

    # Total row 16
    total_row = 16
    ws.cell(row=total_row, column=2, value="TOTAL OPERATING RESERVE").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=total_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=total_row, column=3, value="=SUM(C7:C14)")
    cell.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[total_row].height = 28

    # Reassurance row
    ws.merge_cells("A18:L19")
    c = ws["A18"]
    c.value = (
        "Why this matters: hosts who skip the operating reserve burn furniture money "
        "on month-1 cleaning + insurance + utilities, then can't replace the broken thing "
        "the first guest reports. Keep this segregated."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            wrap_text=True, indent=2)
    ws.row_dimensions[18].height = 22
    ws.row_dimensions[19].height = 22

    brand_footer(ws, 22,
                 version_line=f"{SKU} · Operating Reserve")

    # Print
    ws.print_area = "A1:L25"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 5 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                        prev_tab="Operating Reserve", next_tab=None)

    set_col_widths(ws, [
        ("A", 4), ("B", 30), ("C", 22),
        ("D", 4), ("E", 28), ("F", 18),
        ("G", 4), ("H", 4), ("I", 4), ("J", 4), ("K", 4), ("L", 4),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Property meta · MACRS class options · year-end archive table"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 5, "PROPERTY")
    _input_row(ws, 7, "Property name:",
               _val(variant, SETTINGS_SAMPLE["property"]), None, "")
    _input_row(ws, 9, "Setup start date:",
               _val(variant, SETTINGS_SAMPLE["start"]), None,
               "Use ISO YYYY-MM-DD")
    _input_row(ws, 11, "First-stay target:",
               _val(variant, SETTINGS_SAMPLE["first_stay"]), None,
               "Drives \"$ to first stay\" countdown")
    _input_row(ws, 13, "Operating reserve days:",
               _val(variant, SETTINGS_SAMPLE["reserve_days"]), "0",
               "Default 60 — adjust if pre-bookings start sooner")
    _input_row(ws, 15, "Active year:",
               _val(variant, SETTINGS_SAMPLE["active_year"]), "0",
               "Used for date-filtered views and archive")

    # MACRS class reference
    _section_band(ws, 18, "MACRS CLASS REFERENCE")
    macrs_ref_rows = [
        ("5-yr",            "Appliances, electronics, computers, smart-home devices"),
        ("7-yr",            "Furniture, mattresses, rugs, fixtures, outdoor durables"),
        ("Decor — expense", "Linens, art, plants, books, candles (de minimis: <$2,500/item)"),
    ]
    for i, (cls, desc) in enumerate(macrs_ref_rows):
        row = 20 + i
        a = ws.cell(row=row, column=2, value=cls)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"C{row}:F{row}")
        b = ws[f"C{row}"]
        b.value = desc
        b.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    # Year-end archive
    _section_band(ws, 25, "YEAR-END ARCHIVE")
    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = (
        "Each January, copy the YTD totals from Start tab into one row below, "
        "then clear Room Detail to start the new property build clean."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[26].height = 16

    archive_headers = ["Year", "Total Budget", "Total Spent", "# Items", "Notes"]
    archive_cols = ["B", "C", "D", "E", "F"]
    for col_letter, label in zip(archive_cols, archive_headers):
        cell = ws[f"{col_letter}28"]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[28].height = 22

    base_year = SETTINGS_SAMPLE["active_year"]
    for offset in range(7):
        row = 29 + offset
        year_cell = ws.cell(row=row, column=2, value=base_year + offset)
        apply_style(year_cell, input_cell_style())
        year_cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        year_cell.alignment = Alignment(horizontal="center", vertical="center")
        year_cell.number_format = "0"

        for col in (3, 4, 5, 6):
            cell = ws.cell(row=row, column=col, value=None)
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (3, 4):
                cell.number_format = '"$"#,##0'
            elif col == 5:
                cell.number_format = "0"

        ws.row_dimensions[row].height = 18

    brand_footer(ws, 38,
                 version_line=f"{SKU} · Settings")

    ws.print_area = "A1:F40"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_room_detail_tab(wb, variant)
    build_rollup_tab(wb, variant)
    build_reserve_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Furniture / Setup Budget — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Pre-furnishing budget for a new STR. Per-room item register with "
        "cost benchmarks, MACRS class flag (5-yr / 7-yr / Decor), 60-day "
        "operating reserve, and bar chart $ by room."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
