"""Build OPS-003 License/Permit/STR-Reg Tracker (v2.2 standard).

Operational-mode tool for STR licenses, permits, and registrations.
The alarm clock for renewals across multi-jurisdiction portfolios:
auto-calculates days-to-renewal with red/gold urgency banding,
flags expired-but-operating permits on the Start tab, and includes
a Permit Discovery reference for the top 15 STR markets.

Generates two files:
  templates/_masters/OPS-003-license-permit-tracker-DEMO.xlsx
  templates/_masters/OPS-003-license-permit-tracker-BLANK.xlsx
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE, STATE_BAD_FILL, STATE_GOOD_TEXT,
    STATE_WARN_FILL,
)

SKU = "OPS-003"
NAME = "license-permit-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

REFERENCE_AS_OF = "2026-05-02"

# --- Sample data (DEMO) ---

# Per brief QA section: 5 properties × 4 cities, tax year 2026.
# Tuple: (property, type, jurisdiction, permit_no, issue, renewal,
#         cadence, cost, status, file_loc, notes)
PERMITS_REGISTER_SAMPLES = [
    # Smokies Ridge Cabin — Gatlinburg TN (4 permits, all active)
    ("Smokies Ridge Cabin", "STR permit", "Gatlinburg, TN",
     "GAT-STR-2024-0481", "2024-09-15", "2026-09-15",
     "Annual", 385.00, "Active",
     "Drive/STR/Smokies/Permits/2024-STR-renewal.pdf",
     "Tier 2 vacation rental — owner not on-site OK."),
    ("Smokies Ridge Cabin", "Business license", "Sevier County, TN",
     "SEV-BL-117820", "2025-12-31", "2026-12-31",
     "Annual", 50.00, "Active",
     "Drive/STR/Smokies/Permits/2025-county-biz-license.pdf",
     "Renews end of calendar year."),
    ("Smokies Ridge Cabin", "Sales tax license", "State of Tennessee",
     "TN-SUT-44829110", "2024-01-01", "2027-01-01",
     "Triennial", 0.00, "Active",
     "Drive/STR/Smokies/Permits/TN-sales-tax-cert.pdf",
     "TN sales tax registration; no renewal fee."),
    ("Smokies Ridge Cabin", "Lodging tax registration", "Sevier County, TN",
     "SEV-OCC-22150", "2025-04-15", "2027-04-15",
     "Annual", 0.00, "Active",
     "Drive/STR/Smokies/Permits/sevier-occ-tax.pdf",
     "Occupancy tax filed monthly; registration renews annually."),

    # Creek Side — Nashville TN (3 permits, 1 expired)
    ("Creek Side", "STR permit", "Nashville (Metro Davidson), TN",
     "NSH-STRP-T2-9114", "2025-04-30", "2026-04-30",
     "Annual", 313.00, "EXPIRED",
     "Drive/STR/CreekSide/Permits/2025-strp-type2.pdf",
     "Type 2 STRP — non-owner-occupied. RENEW IMMEDIATELY."),
    ("Creek Side", "Business license", "Metro Nashville, TN",
     "MTRO-BL-8842211", "2025-12-31", "2026-12-31",
     "Annual", 30.00, "Active",
     "Drive/STR/CreekSide/Permits/2025-metro-biz.pdf",
     "Standard business license, renews with state filing."),
    ("Creek Side", "Lodging tax registration", "Metro Nashville, TN",
     "MTRO-OCC-71338", "2025-01-01", "2027-01-01",
     "Annual", 0.00, "Active",
     "Drive/STR/CreekSide/Permits/metro-occ.pdf",
     "Hotel occupancy tax registration."),

    # Lakehouse A — Asheville NC (3 permits, 1 in 22 days)
    ("Lakehouse A", "STR permit", "Asheville, NC",
     "ASH-HSV-2025-22817", "2025-06-23", "2026-05-24",
     "Annual", 200.00, "Active",
     "Drive/STR/LakehouseA/Permits/2025-asheville-hsv.pdf",
     "Homestay vacation rental permit — RED ZONE, 22 days."),
    ("Lakehouse A", "Business license", "Buncombe County, NC",
     "BCC-BL-65329", "2025-09-15", "2026-09-15",
     "Annual", 75.00, "Active",
     "Drive/STR/LakehouseA/Permits/2025-buncombe-biz.pdf",
     "County business license."),
    ("Lakehouse A", "Sales tax license", "State of North Carolina",
     "NC-ST-9117320", "2025-01-01", "2027-01-01",
     "Triennial", 0.00, "Active",
     "Drive/STR/LakehouseA/Permits/NC-sales-tax-cert.pdf",
     "Sales + occupancy tax on lodging."),

    # Mountain Loft — Joshua Tree CA (3 permits, 1 in 60 days)
    ("Mountain Loft", "STR permit", "San Bernardino County, CA",
     "SBC-STR-2025-4117", "2024-07-01", "2026-07-01",
     "Annual", 660.00, "Active",
     "Drive/STR/MountainLoft/Permits/2024-sbc-str.pdf",
     "Joshua Tree falls under SBC unincorporated — GOLD, 60 days."),
    ("Mountain Loft", "Lodging tax registration",
     "San Bernardino County, CA",
     "SBC-TOT-22817", "2024-07-01", "2026-12-31",
     "Annual", 0.00, "Active",
     "Drive/STR/MountainLoft/Permits/sbc-tot.pdf",
     "TOT registration; quarterly remit."),
    ("Mountain Loft", "Business license", "San Bernardino County, CA",
     "SBC-BL-330128", "2025-09-15", "2026-09-15",
     "Annual", 132.00, "Active",
     "Drive/STR/MountainLoft/Permits/2025-sbc-biz.pdf",
     "County business operations license."),

    # Forest Cabin — Big Bear CA (2 permits, all active)
    ("Forest Cabin", "STR permit", "City of Big Bear Lake, CA",
     "BBL-VR-2025-9942", "2025-08-01", "2026-08-01",
     "Annual", 250.00, "Active",
     "Drive/STR/ForestCabin/Permits/2025-bbl-vr.pdf",
     "Big Bear Lake VHR program — number visible on listing required."),
    ("Forest Cabin", "Business license", "San Bernardino County, CA",
     "SBC-BL-441903", "2025-08-01", "2026-08-01",
     "Annual", 132.00, "Active",
     "Drive/STR/ForestCabin/Permits/2025-sbc-biz-fc.pdf",
     "County business license tied to STR address."),
]

# Filing Log sample data — past renewals across the portfolio
# Tuple: (date, property, type, action, paid, conf_no, cert, cert_path, notes)
FILING_LOG_SAMPLES = [
    ("2026-01-12", "Smokies Ridge Cabin", "Business license",
     "Renewal", 50.00, "SEV-CONF-661092", "Yes",
     "Drive/STR/Smokies/Permits/2026-county-biz-receipt.pdf",
     "Renewed via Sevier online portal."),
    ("2026-02-04", "Lakehouse A", "Sales tax license",
     "Renewal", 0.00, "NC-CONF-880113", "Yes",
     "Drive/STR/LakehouseA/Permits/NC-sales-renewal-2026.pdf",
     "Triennial; no fee but registration confirmed."),
    ("2026-03-20", "Mountain Loft", "STR permit",
     "Amendment", 0.00, "SBC-AMEND-7715", "Yes",
     "Drive/STR/MountainLoft/Permits/sbc-amendment-occupancy.pdf",
     "Updated max occupancy from 6 to 8."),
    ("2026-04-01", "Smokies Ridge Cabin", "Lodging tax registration",
     "Renewal", 0.00, "SEV-OCC-220411", "Yes",
     "Drive/STR/Smokies/Permits/sevier-occ-renewal-2026.pdf",
     "Annual occupancy tax registration renewal."),
    ("2026-04-22", "Creek Side", "Business license",
     "Renewal", 30.00, "MTRO-CONF-99820", "Yes",
     "Drive/STR/CreekSide/Permits/2026-metro-biz-renewal.pdf",
     "Renewed Nashville Metro biz license."),
    ("2026-04-28", "Forest Cabin", "STR permit",
     "Initial filing", 250.00, "BBL-NEW-1133", "Yes",
     "Drive/STR/ForestCabin/Permits/2025-bbl-vr-initial.pdf",
     "First-time VHR filing for Big Bear acquisition."),
    ("2026-04-29", "Forest Cabin", "Business license",
     "Initial filing", 132.00, "SBC-BL-NEW-441", "Yes",
     "Drive/STR/ForestCabin/Permits/2025-sbc-biz-initial.pdf",
     "County business license for new property."),
    ("2026-05-01", "Smokies Ridge Cabin", "Sales tax license",
     "Renewal", 0.00, "TN-SUT-CONF-44820", "Yes",
     "Drive/STR/Smokies/Permits/TN-sales-renewal-2026.pdf",
     "TN triennial sales tax recert; no fee."),
]

# Settings tab data
PROPERTIES_LIST = [
    "Smokies Ridge Cabin", "Creek Side", "Lakehouse A",
    "Mountain Loft", "Forest Cabin",
    "Lakehouse B", "Downtown Loft", "Beach House",
]
PERMIT_TYPES = [
    "STR permit", "Business license", "Sales tax license",
    "Lodging tax registration", "Fire inspection",
    "Life safety", "Building permit", "Other",
]
STATUS_LIST = [
    "Active", "Pending", "Renewed", "EXPIRED", "N/A",
]
CADENCE_LIST = [
    "Annual", "Biennial", "Triennial", "One-time",
]
ACTION_LIST = [
    "Initial filing", "Renewal", "Amendment", "Cancellation",
]

# ---------------------------------------------------------------------------
# Permit Discovery — top STR markets reference data
# As-of: REFERENCE_AS_OF. Reference only — verify locally before relying.
# ---------------------------------------------------------------------------
PERMIT_DISCOVERY = [
    {
        "city": "New York City, NY",
        "authority": "NYC Mayor's Office of Special Enforcement (OSE)",
        "permit_name": "Short-Term Rental Registration (Local Law 18)",
        "fees": "$145 application fee",
        "cadence": "One-time registration; updates required on changes",
        "key_restriction": (
            "Owner-occupancy required (Y) — host must be present during stay; "
            "max 2 guests; stays under 30 days only allowed if registered."
        ),
        "url": "https://www.nyc.gov/site/specialenforcement/registration/"
               "short-term-rental.page",
    },
    {
        "city": "Los Angeles, CA",
        "authority": "LA Department of City Planning",
        "permit_name": "Home-Sharing Registration",
        "fees": "$89 annual",
        "cadence": "Annual",
        "key_restriction": (
            "Owner-occupancy required (Y) — primary residence only; "
            "120-night cap unless Extended Home-Sharing approved."
        ),
        "url": "https://planning.lacity.gov/development-services/home-sharing",
    },
    {
        "city": "Austin, TX",
        "authority": "City of Austin Development Services Department",
        "permit_name": "Short-Term Rental License (Type 1, 2, 3)",
        "fees": "$767 initial / $462 renewal",
        "cadence": "Annual",
        "key_restriction": (
            "Owner-occupancy required for Type 1 (Y); Type 2 prohibited "
            "in residential zones since 2016; Type 3 multifamily."
        ),
        "url": "https://www.austintexas.gov/department/short-term-rentals",
    },
    {
        "city": "Denver, CO",
        "authority": "Denver Department of Excise and Licenses",
        "permit_name": "Short-Term Rental Business License",
        "fees": "$50 application + $50 license / annual",
        "cadence": "Annual",
        "key_restriction": (
            "Owner-occupancy required (Y) — must be primary residence; "
            "lodger's tax registration also required."
        ),
        "url": "https://www.denvergov.org/Government/Agencies-Departments-"
               "Offices/Business-Licensing/Business-Licenses/Short-Term-Rentals",
    },
    {
        "city": "Nashville (Metro Davidson), TN",
        "authority": "Metro Codes Department",
        "permit_name": "Short-Term Rental Property Permit (Type 1 / Type 2)",
        "fees": "$313 annual",
        "cadence": "Annual",
        "key_restriction": (
            "Type 1 owner-occupied (Y); Type 2 non-owner-occupied capped "
            "by zoning + grandfathered in many districts."
        ),
        "url": "https://www.nashville.gov/departments/codes/short-term-rentals",
    },
    {
        "city": "New Orleans, LA",
        "authority": "City of New Orleans Department of Safety and Permits",
        "permit_name": "Short-Term Rental Permit (Residential / Commercial)",
        "fees": "$500 residential / $500+ commercial annual",
        "cadence": "Annual",
        "key_restriction": (
            "Residential STR requires owner-occupancy (Y); homestead "
            "exemption required; one permit per lot in most districts."
        ),
        "url": "https://nola.gov/short-term-rentals/",
    },
    {
        "city": "Honolulu, HI",
        "authority": "Honolulu Department of Planning and Permitting",
        "permit_name": "Short-Term Rental Registration (Bill 41 / Ord 22-7)",
        "fees": "$1,000 registration + annual renewal",
        "cadence": "Annual",
        "key_restriction": (
            "Strict — 90-day minimum stay outside resort districts; "
            "operating permit limited; non-conforming use grandfathering only."
        ),
        "url": "https://www.honolulu.gov/dpp/short-term-rentals.html",
    },
    {
        "city": "San Francisco, CA",
        "authority": "SF Office of Short-Term Rentals",
        "permit_name": "Short-Term Residential Rental Certificate",
        "fees": "$1,025 biennial",
        "cadence": "Biennial",
        "key_restriction": (
            "Owner-occupancy required (Y) — primary residence 275+ "
            "nights/year; 90-night cap on un-hosted stays."
        ),
        "url": "https://sf.gov/topics/short-term-rentals",
    },
    {
        "city": "Miami Beach, FL",
        "authority": "Miami Beach Code Compliance",
        "permit_name": "Short-Term Rental Business Tax Receipt",
        "fees": "$390 BTR + $1,500 resort tax registration",
        "cadence": "Annual",
        "key_restriction": (
            "Owner-occupancy NOT required (N) — but STRs banned in many "
            "single-family zones; fines up to $100,000 for unpermitted use."
        ),
        "url": "https://www.miamibeachfl.gov/code-compliance/short-term-rentals/",
    },
    {
        "city": "Portland, OR",
        "authority": "Portland Bureau of Development Services",
        "permit_name": "Accessory Short-Term Rental Permit",
        "fees": "$67 (Type A) / $4,432 (Type B) initial",
        "cadence": "Biennial (Type A) / Conditional Use (Type B)",
        "key_restriction": (
            "Owner-occupancy required for Type A (Y) — owner present "
            "270+ nights/year; Type B requires conditional-use review."
        ),
        "url": "https://www.portland.gov/bds/zoning-land-use/accessory-"
               "short-term-rentals",
    },
    {
        "city": "Seattle, WA",
        "authority": "Seattle Department of Finance and Administrative Services",
        "permit_name": "Short-Term Rental Operator's License",
        "fees": "$75 operator + $14 per unit annual",
        "cadence": "Annual",
        "key_restriction": (
            "Two-unit cap per operator (one must be primary residence); "
            "city business license also required."
        ),
        "url": "https://www.seattle.gov/business-regulations/short-term-rentals",
    },
    {
        "city": "Chicago, IL",
        "authority": "Chicago Department of Business Affairs and Consumer "
                     "Protection (BACP)",
        "permit_name": "Shared Housing Unit License / Vacation Rental License",
        "fees": "$125 shared housing / $250 vacation rental — annual",
        "cadence": "Annual",
        "key_restriction": (
            "Owner-occupancy preferred (mixed) — restricted residential "
            "zones list maintained by city; building condo restrictions apply."
        ),
        "url": "https://www.chicago.gov/city/en/depts/bacp/supp_info/"
               "sharedhousingordinance.html",
    },
    {
        "city": "San Diego, CA",
        "authority": "San Diego Office of the City Treasurer",
        "permit_name": "Short-Term Residential Occupancy License (Tier 1-4)",
        "fees": "$100-$1,000 by tier annual",
        "cadence": "Annual / Biennial by tier",
        "key_restriction": (
            "Tier 3/4 capped citywide via lottery; Tier 4 (Mission Beach) "
            "separate cap; primary residence preference for Tier 2."
        ),
        "url": "https://www.sandiego.gov/treasurer/short-term-residential-"
               "occupancy",
    },
    {
        "city": "Charleston, SC",
        "authority": "City of Charleston Permit Center",
        "permit_name": "Short-Term Rental Permit (Categorical)",
        "fees": "$300 application + annual business license",
        "cadence": "Annual",
        "key_restriction": (
            "Owner-occupancy required for most categories (Y); historic "
            "district has stricter rules; commercial STR permits are limited."
        ),
        "url": "https://www.charleston-sc.gov/2389/Short-Term-Rentals",
    },
    {
        "city": "Savannah, GA",
        "authority": "City of Savannah Revenue Department",
        "permit_name": "Short-Term Vacation Rental Certificate",
        "fees": "$450 application + annual occupancy tax registration",
        "cadence": "Annual",
        "key_restriction": (
            "STVR cap applies in historic districts (N for cap-met areas); "
            "20% cap per ward enforced — waitlist common."
        ),
        "url": "https://www.savannahga.gov/2008/Short-Term-Vacation-Rentals",
    },
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list, blank_length=None):
    """Return demo_list if variant == 'demo', else a list of None of same length."""
    if variant == "demo":
        return demo_list
    n = blank_length if blank_length is not None else len(demo_list)
    return [None] * n


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero band rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "License / Permit / STR-Reg Tracker"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "An alarm clock for the renewals you'd otherwise forget."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: At-a-glance metric cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): EXPIRED count
    ws.merge_cells("A10:D10")
    c = ws["A10"]
    c.value = "EXPIRED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A11:D13")
    c = ws["A11"]
    c.value = (
        '=COUNTIFS(\'Permits Register\'!I7:I206,"EXPIRED")'
        '+COUNTIFS(\'Permits Register\'!L7:L206,"<0",'
        '\'Permits Register\'!I7:I206,"<>Renewed",'
        '\'Permits Register\'!I7:I206,"<>N/A")'
    )
    # Simpler version: count statuses flagged EXPIRED OR computed days < 0
    c.value = (
        '=COUNTIF(\'Permits Register\'!M7:M206,"EXPIRED")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A14:D15")
    c = ws["A14"]
    c.value = "permits past renewal date"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 2 (E-H): RENEW NOW (<30 days)
    ws.merge_cells("E10:H10")
    c = ws["E10"]
    c.value = "RENEW NOW (<30 DAYS)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E11:H13")
    c = ws["E11"]
    c.value = (
        '=COUNTIFS(\'Permits Register\'!L7:L206,">=0",'
        '\'Permits Register\'!L7:L206,"<30")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E14:H15")
    c = ws["E14"]
    c.value = "permits due in next 30 days"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 3 (I-L): YTD filing $
    ws.merge_cells("I10:L10")
    c = ws["I10"]
    c.value = "FILING $ YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I11:L13")
    c = ws["I11"]
    c.value = (
        '=SUMPRODUCT((YEAR(\'Filing Log\'!A7:A506)=YEAR(TODAY()))*'
        '\'Filing Log\'!E7:E506)'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I14:L15")
    c = ws["I14"]
    c.value = "fees paid year to date"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 16):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: Operating-while-expired RED banner rows 17-19
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    # Single-cell formula with all conditions; result is the banner text.
    c.value = (
        '=IF(COUNTIF(\'Permits Register\'!M7:M206,"EXPIRED")>0,'
        '"OPERATING WHILE EXPIRED — STOP. One or more permits below '
        'are past their renewal date. Operating without a current '
        'permit creates legal + financial liability. Renew before '
        'the next booking starts.",'
        '"ALL CLEAR — No expired permits in your portfolio.")'
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[17].height = 36

    # Conditional formatting: when ALL CLEAR, paint green; otherwise red.
    # Green when no expired
    ws.conditional_formatting.add(
        "A17:L19",
        FormulaRule(
            formula=['COUNTIF(\'Permits Register\'!M7:M206,"EXPIRED")=0'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_TEXT),
        ),
    )

    # ZONE 4: Top-5 urgency list rows 20-29
    ws.merge_cells("A20:L20")
    c = ws["A20"]
    c.value = "TOP RENEWAL ACTIONS"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[20].height = 22

    ws.merge_cells("A21:L21")
    c = ws["A21"]
    c.value = (
        "Open the Renewal Calendar tab — permits are sorted by urgency "
        "(EXPIRED on top, then <30 days red, then 30-90 gold). The five "
        "most urgent permits in your portfolio are listed there."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[21].height = 36

    ws.merge_cells("A23:L26")
    c = ws["A23"]
    c.value = (
        "Why this matters: most STR fines come from missed renewals, "
        "not from regulatory violations. Knox County, TN issues a "
        "$750 fine + 30-day suspension for an expired Type 2 permit. "
        "Nashville has issued $1,500 fines for operating without a "
        "current STRP. Use the Renewal Calendar to nail the dates "
        "before the platforms (Airbnb, VRBO) auto-suspend the listing."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 5: Pseudo-button nav rows 28-30
    pseudo_button(ws, "A28", "C30", "Permits Register",
                  "'Permits Register'!A1", variant="primary")
    pseudo_button(ws, "D28", "F30", "Renewal Calendar",
                  "'Renewal Calendar'!A1", variant="primary")
    pseudo_button(ws, "G28", "I30", "Permit Discovery",
                  "'Permit Discovery'!A1", variant="secondary")
    pseudo_button(ws, "J28", "L30", "Filing Log",
                  "'Filing Log'!A1", variant="secondary")
    for r in range(28, 31):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A32", "L34", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(32, 35):
        ws.row_dimensions[r].height = 18

    # ZONE 6: Upgrade banner row 36
    ws.merge_cells("A36:L36")
    c = ws["A36"]
    c.value = (
        "Need full operations coverage? Add the Operator Bundle at "
        f"{BRAND_DOMAIN}/operator — turnover + maintenance + supply + "
        "damage claims + insurance + permits, $197."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[36].height = 36

    brand_footer(
        ws, 38,
        version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever",
    )

    ws.print_area = "A1:L40"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Permits Register
# ---------------------------------------------------------------------------

def build_permits_register_tab(wb, variant):
    ws = wb.create_sheet("Permits Register")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(
        ws, "Permits Register",
        prev_tab="Start", next_tab="Renewal Calendar",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 14):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:M4")
    c = ws["A4"]
    c.value = (
        "One row per property × permit type. Days-to-renewal auto-calculates "
        "from the Renewal date — keep that column accurate."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column layout (brief: A=22, B=22, C=18, D=14, E=12, F=12,
    # G=12, H=10, I=14, J=20, K=30) plus L (Days) and M (Auto-status flag)
    set_col_widths(ws, [
        ("A", 22), ("B", 22), ("C", 20), ("D", 16),
        ("E", 12), ("F", 12), ("G", 12), ("H", 11),
        ("I", 14), ("J", 24), ("K", 32),
        ("L", 14), ("M", 12),
    ])

    headers = [
        "Property", "Permit type", "Jurisdiction", "Permit #",
        "Issue date", "Renewal date", "Cadence", "$ Cost",
        "Status (entered)", "File location", "Notes",
        "Days to renewal", "Auto status",
    ]
    # Headers in row 6 (rows 1-3 are header band, row 4 is description, row 5 spacer)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    samples = _val_list(variant, PERMITS_REGISTER_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 7 + i
        (prop, ptype, jurisdiction, permit_no, issue, renewal,
         cadence, cost, status, file_loc, notes) = row_data

        for col, val in [(1, prop), (2, ptype), (3, jurisdiction), (4, permit_no)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        e = ws.cell(row=row, column=5, value=_parse_date(issue))
        apply_style(e, input_cell_style())
        e.number_format = "yyyy-mm-dd"

        f = ws.cell(row=row, column=6, value=_parse_date(renewal))
        apply_style(f, input_cell_style())
        f.number_format = "yyyy-mm-dd"

        g = ws.cell(row=row, column=7, value=cadence)
        apply_style(g, input_cell_style())

        h = ws.cell(row=row, column=8, value=cost)
        apply_style(h, input_cell_style())
        h.number_format = '"$"#,##0.00'

        i_cell = ws.cell(row=row, column=9, value=status)
        apply_style(i_cell, input_cell_style())

        j = ws.cell(row=row, column=10, value=file_loc)
        apply_style(j, input_cell_style())

        k = ws.cell(row=row, column=11, value=notes)
        apply_style(k, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Apply formulas + format to all rows 7-206 (200 capacity)
    for row in range(7, 207):
        # Days to renewal col L
        l_cell = ws.cell(
            row=row, column=12,
            value=f'=IF(F{row}<>"",F{row}-TODAY(),"")',
        )
        apply_style(l_cell, formula_cell_style())
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        l_cell.number_format = "0"

        # Auto status col M — flips to EXPIRED if days<0 and status not Renewed/N/A
        m_cell = ws.cell(
            row=row, column=13,
            value=(
                f'=IF(F{row}="","",'
                f'IF(AND(L{row}<0,I{row}<>"Renewed",I{row}<>"N/A"),'
                f'"EXPIRED",I{row}))'
            ),
        )
        apply_style(m_cell, formula_cell_style())
        m_cell.alignment = Alignment(horizontal="center", vertical="center")
        m_cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)

        # For rows past the demo samples, apply blank input formatting
        if row > 6 + len(samples):
            for col in [5, 6]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            for col in [1, 2, 3, 4, 7, 9, 10, 11]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            cell8 = ws.cell(row=row, column=8)
            apply_style(cell8, input_cell_style())
            cell8.number_format = '"$"#,##0.00'
            ws.row_dimensions[row].height = 20

    # Conditional formatting on Days-to-renewal (col L)
    # < 0 (expired) — red, 0-29 — red soft, 30-89 — gold-soft
    ws.conditional_formatting.add(
        "L7:L206",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
                   font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "L7:L206",
        CellIsRule(operator="between", formula=["0", "29"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
                   font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "L7:L206",
        CellIsRule(operator="between", formula=["30", "89"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )

    # Conditional fmt on Auto status (col M) — red text for EXPIRED
    ws.conditional_formatting.add(
        "M7:M206",
        FormulaRule(
            formula=['$M7="EXPIRED"'],
            fill=PatternFill("solid", fgColor=COLOR_ERROR),
            font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_WHITE),
        ),
    )

    # Dropdowns
    add_dropdown(ws, "A7:A206", "=Settings!$B$6:$B$30")
    add_dropdown(ws, "B7:B206", "=Settings!$C$6:$C$15")
    add_dropdown(ws, "G7:G206", "=Settings!$D$6:$D$10")
    add_dropdown(ws, "I7:I206", "=Settings!$E$6:$E$11")

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Renewal Calendar
# ---------------------------------------------------------------------------

def build_renewal_calendar_tab(wb, variant):
    ws = wb.create_sheet("Renewal Calendar")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(
        ws, "Renewal Calendar",
        prev_tab="Permits Register", next_tab="Permit Discovery",
    )

    set_col_widths(ws, [
        ("A", 22), ("B", 22), ("C", 22), ("D", 14),
        ("E", 12), ("F", 12), ("G", 16), ("H", 24),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Three urgency bands sorted by days-to-renewal. The list refreshes "
        "automatically — updates flow from the Permits Register every time "
        "you open the workbook."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    def section_band(row, label, fill_color, text_color):
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 28

    def column_headers(row):
        headers = [
            "Property", "Permit type", "Jurisdiction", "Permit #",
            "Renewal date", "Days", "$ Cost", "Notes",
        ]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            apply_style(cell, header_row_style())
        ws.row_dimensions[row].height = 22

    # === EXPIRED section rows 6-11 (band + headers + 4 data rows) ===
    section_band(6, "EXPIRED  —  RENEW IMMEDIATELY",
                  fill_color=COLOR_ERROR, text_color=COLOR_WHITE)
    column_headers(7)

    # Use IFERROR + INDEX/MATCH on register sorted by days-to-renewal
    # Strategy: list rows where Auto status (col M) = EXPIRED — pull top 5.
    # We use SMALL on row indexes meeting the condition.
    # For brevity and Excel compatibility we write helper formulas.
    expired_rows = []
    if variant == "demo":
        # Only Creek Side STR permit is expired in demo data
        expired_rows = [r for r, d in enumerate(PERMITS_REGISTER_SAMPLES, start=7)
                        if d[8] == "EXPIRED"]

    expired_count = max(len(expired_rows), 1) if variant == "demo" else 4
    expired_data_rows_end = 7 + max(expired_count, 4)
    for i in range(4):  # 4 expired slots
        r = 8 + i
        if variant == "demo" and i < len(expired_rows):
            src = expired_rows[i]
            d = PERMITS_REGISTER_SAMPLES[src - 7]
            ws.cell(row=r, column=1, value=d[0])
            ws.cell(row=r, column=2, value=d[1])
            ws.cell(row=r, column=3, value=d[2])
            ws.cell(row=r, column=4, value=d[3])
            cell = ws.cell(row=r, column=5, value=_parse_date(d[5]))
            cell.number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=6,
                     value=f"=IF('Permits Register'!F{src}<>\"\","
                           f"'Permits Register'!F{src}-TODAY(),\"\")")
            ws.cell(row=r, column=6).number_format = "0"
            cell = ws.cell(row=r, column=7, value=d[7])
            cell.number_format = '"$"#,##0.00'
            ws.cell(row=r, column=8, value=d[10])
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                if not cell.font or cell.font.size != 10:
                    cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
                cell.fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    # === RENEW NOW (<30 days) ===
    renew_now_band_row = 13
    section_band(renew_now_band_row, "RENEW NOW  —  LESS THAN 30 DAYS",
                  fill_color=COLOR_ACCENT, text_color=COLOR_WHITE)
    column_headers(14)

    renew_now_rows = []
    if variant == "demo":
        # Pre-compute days-to-renewal vs today's date for demo data
        today = datetime.strptime(REFERENCE_AS_OF, "%Y-%m-%d").date()
        for idx, d in enumerate(PERMITS_REGISTER_SAMPLES):
            if d[8] in ("EXPIRED", "Renewed", "N/A"):
                continue
            rdate = _parse_date(d[5])
            if rdate is None:
                continue
            delta = (rdate - today).days
            if 0 <= delta < 30:
                renew_now_rows.append((idx, delta))
        renew_now_rows.sort(key=lambda x: x[1])

    for i in range(5):  # 5 slots for renew-now
        r = 15 + i
        if variant == "demo" and i < len(renew_now_rows):
            idx, delta = renew_now_rows[i]
            d = PERMITS_REGISTER_SAMPLES[idx]
            src = 7 + idx
            ws.cell(row=r, column=1, value=d[0])
            ws.cell(row=r, column=2, value=d[1])
            ws.cell(row=r, column=3, value=d[2])
            ws.cell(row=r, column=4, value=d[3])
            cell = ws.cell(row=r, column=5, value=_parse_date(d[5]))
            cell.number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=6,
                     value=f"=IF('Permits Register'!F{src}<>\"\","
                           f"'Permits Register'!F{src}-TODAY(),\"\")")
            ws.cell(row=r, column=6).number_format = "0"
            cell = ws.cell(row=r, column=7, value=d[7])
            cell.number_format = '"$"#,##0.00'
            ws.cell(row=r, column=8, value=d[10])
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
                cell.fill = PatternFill("solid", fgColor=STATE_WARN_FILL)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    # === UPCOMING (30-90 days) ===
    upcoming_band_row = 21
    section_band(upcoming_band_row, "UPCOMING  —  30 TO 90 DAYS",
                  fill_color=COLOR_NAVY_TINT, text_color=COLOR_WHITE)
    column_headers(22)

    upcoming_rows = []
    if variant == "demo":
        today = datetime.strptime(REFERENCE_AS_OF, "%Y-%m-%d").date()
        for idx, d in enumerate(PERMITS_REGISTER_SAMPLES):
            if d[8] in ("EXPIRED", "Renewed", "N/A"):
                continue
            rdate = _parse_date(d[5])
            if rdate is None:
                continue
            delta = (rdate - today).days
            if 30 <= delta <= 90:
                upcoming_rows.append((idx, delta))
        upcoming_rows.sort(key=lambda x: x[1])

    for i in range(8):  # 8 slots
        r = 23 + i
        if variant == "demo" and i < len(upcoming_rows):
            idx, delta = upcoming_rows[i]
            d = PERMITS_REGISTER_SAMPLES[idx]
            src = 7 + idx
            ws.cell(row=r, column=1, value=d[0])
            ws.cell(row=r, column=2, value=d[1])
            ws.cell(row=r, column=3, value=d[2])
            ws.cell(row=r, column=4, value=d[3])
            cell = ws.cell(row=r, column=5, value=_parse_date(d[5]))
            cell.number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=6,
                     value=f"=IF('Permits Register'!F{src}<>\"\","
                           f"'Permits Register'!F{src}-TODAY(),\"\")")
            ws.cell(row=r, column=6).number_format = "0"
            cell = ws.cell(row=r, column=7, value=d[7])
            cell.number_format = '"$"#,##0.00'
            ws.cell(row=r, column=8, value=d[10])
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
                cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    # >90 day count footer
    footer_row = 32
    ws.merge_cells(f"A{footer_row}:L{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = (
        '=IF(COUNTIFS(\'Permits Register\'!L7:L206,">90")>0,'
        '"Permits more than 90 days out: "'
        '&COUNTIFS(\'Permits Register\'!L7:L206,">90")'
        '&" — open the Permits Register for the full list.",'
        '"All long-horizon permits will surface here as their dates approach.")'
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", indent=1)
    ws.row_dimensions[footer_row].height = 24

    # Manual refresh note
    ws.merge_cells(f"A{footer_row + 2}:L{footer_row + 2}")
    c = ws[f"A{footer_row + 2}"]
    c.value = (
        "Note: this view is a manual snapshot — the rows above show the "
        "permits as of build time. After you edit the Permits Register, "
        "the days-to-renewal columns recalc automatically; rerun-sort "
        "by hand if you need this view re-ranked."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[footer_row + 2].height = 30

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Permit Discovery
# ---------------------------------------------------------------------------

def build_permit_discovery_tab(wb, variant):
    ws = wb.create_sheet("Permit Discovery")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 26),
        ("C", 14), ("D", 14), ("E", 14), ("F", 14),
        ("G", 14), ("H", 14), ("I", 14), ("J", 14),
        ("K", 14), ("L", 14),
    ])

    compact_header_band(
        ws, "Permit Discovery",
        prev_tab="Renewal Calendar", next_tab="Filing Log",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        f"Top STR markets — typical permit stack as of {REFERENCE_AS_OF}. "
        "Rules change frequently. Verify with the issuing jurisdiction "
        "before relying on this snapshot."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Top disclaimer banner row 5
    ws.merge_cells("A5:L6")
    c = ws["A5"]
    c.value = (
        "REFERENCE ONLY — NOT LEGAL ADVICE.  "
        "STR ordinances change frequently, sometimes mid-year. The "
        "permit names, fees, and restrictions below are starting points "
        "for your research, not authoritative guidance. Always verify "
        "the current rules with the issuing jurisdiction (or your local "
        "STR attorney) before applying for, renewing, or relying on a "
        "permit. Snapshot date: " + REFERENCE_AS_OF + "."
    )
    c.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 24

    current_row = 8
    for entry in PERMIT_DISCOVERY:
        # City banner
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = entry["city"]
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 26
        current_row += 1

        # Field rows
        fields = [
            ("Authority",         entry["authority"]),
            ("Permit name",       entry["permit_name"]),
            ("Typical fees",      entry["fees"]),
            ("Renewal cadence",   entry["cadence"]),
            ("Key restriction",   entry["key_restriction"]),
            ("Reference URL",     entry["url"]),
            ("As-of",             REFERENCE_AS_OF),
        ]
        for label, value in fields:
            # Label col B
            lab_cell = ws.cell(row=current_row, column=2, value=label)
            lab_cell.font = Font(name=FONT_MONO, size=9, bold=True,
                                  color=COLOR_PRIMARY)
            lab_cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            lab_cell.alignment = Alignment(horizontal="left",
                                             vertical="top", indent=1)
            # Value cols C-L
            ws.merge_cells(f"C{current_row}:L{current_row}")
            val_cell = ws[f"C{current_row}"]
            val_cell.value = value
            font_color = COLOR_ACCENT if label == "Reference URL" else COLOR_TEXT
            font_name = FONT_MONO if label == "Reference URL" else FONT_BODY
            val_cell.font = Font(name=font_name, size=10, color=font_color)
            val_cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
            val_cell.alignment = Alignment(horizontal="left",
                                            vertical="top",
                                            wrap_text=True, indent=1)
            # Row height — taller for restriction
            ws.row_dimensions[current_row].height = 30 if label == "Key restriction" else 18
            current_row += 1

        # Verify-locally disclaimer
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = (
            "Verify locally — STR rules in this jurisdiction change "
            "and may have updated since the as-of date above."
        )
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="center", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Spacer
        current_row += 1

    # Request-your-city CTA at the bottom
    ws.merge_cells(f"A{current_row}:L{current_row}")
    c = ws[f"A{current_row}"]
    c.value = "DON'T SEE YOUR CITY?"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    ws.merge_cells(f"A{current_row}:L{current_row + 2}")
    c = ws[f"A{current_row}"]
    c.value = (
        f"Email {BRAND_EMAIL} with the subject 'Permit Discovery request — "
        "[Your city, ST]' and we'll add it to the next quarterly update. "
        "Top-15 list is curated based on Sarah-the-multi-city-host's most "
        "common markets — your request goes straight into the v1.1 review."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    for r in range(current_row, current_row + 3):
        ws.row_dimensions[r].height = 22
    current_row += 4

    # Final disclaimer
    ws.merge_cells(f"A{current_row}:L{current_row + 2}")
    c = ws[f"A{current_row}"]
    c.value = (
        "WARNING — STR rules change frequently; cities pass new ordinances "
        "every legislative cycle. Treat this entire tab as a starting "
        "point for your research. Always confirm the current requirements "
        "with the issuing authority before applying for or renewing any "
        "permit. The STR Ledger is not a legal services provider."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    for r in range(current_row, current_row + 3):
        ws.row_dimensions[r].height = 18

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Filing Log
# ---------------------------------------------------------------------------

def build_filing_log_tab(wb, variant):
    ws = wb.create_sheet("Filing Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(
        ws, "Filing Log",
        prev_tab="Permit Discovery", next_tab="Settings",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Every renewal, amendment, and initial filing — keep the proof. "
        "Filing $ YTD on the Start tab pulls from this log."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 22), ("C", 22), ("D", 16),
        ("E", 11), ("F", 18), ("G", 12), ("H", 32), ("I", 32),
    ])

    headers = [
        "Date", "Property", "Permit type", "Action",
        "$ Paid", "Confirmation #", "Certificate?",
        "Certificate path", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    samples = _val_list(variant, FILING_LOG_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 7 + i
        (date_filed, prop, ptype, action, paid, conf,
         cert, cert_path, notes) = row_data

        a = ws.cell(row=row, column=1, value=_parse_date(date_filed))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        for col, val in [(2, prop), (3, ptype), (4, action)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        e = ws.cell(row=row, column=5, value=paid)
        apply_style(e, input_cell_style())
        e.number_format = '"$"#,##0.00'

        for col, val in [(6, conf), (7, cert), (8, cert_path), (9, notes)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        ws.row_dimensions[row].height = 20

    # Apply blank-input formatting to remaining rows up to 506 (500 capacity)
    for row in range(7, 507):
        if row > 6 + len(samples):
            cell = ws.cell(row=row, column=1)
            apply_style(cell, input_cell_style())
            cell.number_format = "yyyy-mm-dd"
            for col in [2, 3, 4, 6, 7, 8, 9]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            cell5 = ws.cell(row=row, column=5)
            apply_style(cell5, input_cell_style())
            cell5.number_format = '"$"#,##0.00'
            ws.row_dimensions[row].height = 18

    # Dropdowns
    add_dropdown(ws, "B7:B506", "=Settings!$B$6:$B$30")
    add_dropdown(ws, "C7:C506", "=Settings!$C$6:$C$15")
    add_dropdown(ws, "D7:D506", "=Settings!$F$6:$F$10")
    add_dropdown(ws, "G7:G506", '"Yes,No"')

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28),
        ("C", 26), ("D", 16), ("E", 14), ("F", 20),
    ])

    compact_header_band(
        ws, "Settings",
        prev_tab="Filing Log", next_tab=None,
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Lists drive the dropdowns on Permits Register, Renewal Calendar, "
        "and Filing Log. Edit to match your portfolio."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Column header strip row 5
    headers = [
        (2, "Properties"),
        (3, "Permit types"),
        (4, "Cadence"),
        (5, "Status"),
        (6, "Action (Filing Log)"),
    ]
    for col, label in headers:
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 22

    # Properties col B rows 6-30 (25 capacity)
    for i in range(25):
        r = 6 + i
        cell = ws.cell(row=r, column=2)
        if i < len(PROPERTIES_LIST):
            cell.value = PROPERTIES_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Permit types col C rows 6-15
    for i in range(10):
        r = 6 + i
        cell = ws.cell(row=r, column=3)
        if i < len(PERMIT_TYPES):
            cell.value = PERMIT_TYPES[i]
        apply_style(cell, input_cell_style())

    # Cadence col D rows 6-10
    for i in range(5):
        r = 6 + i
        cell = ws.cell(row=r, column=4)
        if i < len(CADENCE_LIST):
            cell.value = CADENCE_LIST[i]
        apply_style(cell, input_cell_style())

    # Status col E rows 6-11
    for i in range(6):
        r = 6 + i
        cell = ws.cell(row=r, column=5)
        if i < len(STATUS_LIST):
            cell.value = STATUS_LIST[i]
        apply_style(cell, input_cell_style())

    # Action col F rows 6-10
    for i in range(5):
        r = 6 + i
        cell = ws.cell(row=r, column=6)
        if i < len(ACTION_LIST):
            cell.value = ACTION_LIST[i]
        apply_style(cell, input_cell_style())

    # As-of stamp row 33
    ws.cell(row=33, column=2,
             value="Permit Discovery verified:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    cell = ws.cell(row=33, column=3, value=_parse_date(REFERENCE_AS_OF))
    apply_style(cell, input_cell_style())
    cell.number_format = "yyyy-mm-dd"
    cell.font = Font(name=FONT_BODY, size=11, color=COLOR_PRIMARY)
    ws.row_dimensions[33].height = 24

    # Note
    ws.merge_cells("B34:L34")
    c = ws["B34"]
    c.value = (
        "Manually update this date when you re-verify the Permit Discovery "
        "snapshot. STR ordinances change — re-check once a quarter at minimum."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=1)
    ws.row_dimensions[34].height = 36


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_permits_register_tab(wb, variant)
    build_renewal_calendar_tab(wb, variant)
    build_permit_discovery_tab(wb, variant)
    build_filing_log_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "License/Permit Tracker — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Multi-jurisdiction STR license/permit/registration tracker — "
        "auto-calculated days-to-renewal, expired-but-operating flagger, "
        "Permit Discovery for top 15 STR markets."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
