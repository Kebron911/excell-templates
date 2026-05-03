"""Build STR-001 Escape the W2 Planner (v2.2 standard).

Wizard-mode strategic planner — household fills it once per planning
review and gets THE quit date as the headline output. Optimistic vs
Conservative spread becomes the partner-conversation tool.

Generates three files:
  templates/_masters/STR-001-escape-the-w2-planner-DEMO.xlsx   (Full + sample data)
  templates/_masters/STR-001-escape-the-w2-planner-BLANK.xlsx  (Full + empty)
  templates/_lite/STR-001-escape-the-w2-planner-lite.xlsx      (Lite, 4 tabs, sample data)

Brief: templates/_briefs/STR-001-escape-the-w2-planner.md
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
)

SKU = "STR-001"
NAME = "escape-the-w2-planner"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
LITE_OUT = BASE / "_lite" / f"{SKU}-{NAME}-lite.xlsx"

# ---------------------------------------------------------------------------
# Sample data — Side-Hustle Sam scenario (per brief QA section)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Plan anchor
    "plan_start_date":     date(2026, 5, 1),     # today-ish

    # Current state — household
    "w2_gross":            145000,
    "w2_net":              98000,
    "w2_benefits":         18000,    # employer 401k match + health + life
    "partner_w2_net":      72000,
    "household_expenses":  96000,
    "current_liquid":      48000,
    "annual_savings_rate": 42000,    # what they actually save into deployment cash

    # Current STR portfolio
    "props_owned":         2,
    "portfolio_gross":     58000,
    "portfolio_net":       19000,
    "avg_stab_net_per_prop": 14000,  # from #1 prop in Y3

    # Target state
    "target_replacement":  112000,   # net + healthcare bridge
    "buffer_multiple":     1.5,
    "target_reserve":      48000,    # 6 months of expenses
    "target_stab_net_per_prop": 14000,

    # Acquisition pace assumptions
    "opt_props_per_year":  3.0,
    "cons_props_per_year": 1.5,
    "cash_per_property":   95000,    # down + closing + rehab + furnish + reserve
    "y1_net_haircut":      0.65,     # Y1 is ~65% of stabilized

    # Risk plan
    "healthcare_employer_cost":  3600,   # what Sam pays now
    "healthcare_cobra_cost":     14400,  # 102% of full premium, family of 4
    "healthcare_aca_cost":       9600,   # post-subsidy mid-range
    "cobra_months":              18,
    "aca_bridge_months":         24,
    "disability_lost":           1800,
    "life_lost":                 600,
    "partner_risk_share_pct":    0.40,   # partner W2 covers 40% of expense base in transition
    "partner_buyin":             4,      # 1-5

    # Optimistic vs Conservative scenario assumptions
    "opt_stab_occ":        0.68,
    "cons_stab_occ":       0.55,
    "opt_appreciation":    0.04,
    "cons_appreciation":   0.02,
}


SCENARIOS = ["Optimistic", "Conservative"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet 1 — Start (THE quit date)
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant, is_lite):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    title = "Escape the W2 Planner"
    if is_lite:
        title += " — Lite"
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = title
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "When the math says you can."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    label_suffix = "LITE" if is_lite else variant.upper()
    c.value = f"{SKU} · v1.0 · {label_suffix}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ---- THE answer block (rows 8-22) ----
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Optimistic
    ws.merge_cells("A8:L8")
    c = ws["A8"]
    c.value = "OPTIMISTIC QUIT DATE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 18

    ws.merge_cells("A9:L10")
    c = ws["A9"]
    c.value = "='Quit Date Calculator'!C7"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "mmmm d, yyyy"
    ws.row_dimensions[9].height = 32
    ws.row_dimensions[10].height = 18

    # Conservative
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "CONSERVATIVE QUIT DATE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 18

    ws.merge_cells("A12:L13")
    c = ws["A12"]
    c.value = "='Quit Date Calculator'!C8"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "mmmm d, yyyy"
    ws.row_dimensions[12].height = 28
    ws.row_dimensions[13].height = 14

    # Spread
    ws.merge_cells("A14:L14")
    c = ws["A14"]
    c.value = ('="Spread: "&TEXT(\'Quit Date Calculator\'!C9,"0.0")&" years between best- and worst-case."')
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 20

    # 4 critical numbers row 16-19
    ws.merge_cells("A16:C16")
    ws.merge_cells("D16:F16")
    ws.merge_cells("G16:I16")
    ws.merge_cells("J16:L16")
    headings = [
        ("A16", "PROPS NEEDED"),
        ("D16", "CASH NEEDED AT QUIT"),
        ("G16", "HEALTHCARE BRIDGE"),
        ("J16", "Y1 PROJECTED NET"),
    ]
    for ref, txt in headings:
        cell = ws[ref]
        cell.value = txt
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 18

    ws.merge_cells("A17:C18")
    ws.merge_cells("D17:F18")
    ws.merge_cells("G17:I18")
    ws.merge_cells("J17:L18")

    c = ws["A17"]
    c.value = "='Target State'!C20"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0"

    c = ws["D17"]
    c.value = "='Target State'!C22"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    c = ws["G17"]
    if is_lite:
        c.value = '="(Full only)"'
        c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_MUTED)
    else:
        c.value = "='Risk Plan'!C22"
        c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
        c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="center", vertical="center")

    c = ws["J17"]
    c.value = "='Quit Date Calculator'!C19"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.row_dimensions[17].height = 26
    ws.row_dimensions[18].height = 14

    # Narrative paragraph (row 20-22)
    ws.merge_cells("A20:L22")
    c = ws["A20"]
    c.value = (
        '="You\'re "&TEXT(\'Target State\'!C20-\'Current State\'!C16,"0")&" properties and "&'
        'TEXT(MAX(0,\'Target State\'!C22-\'Current State\'!C18),"$#,##0")&" away from the target. '
        'Two paths get you there: "&TEXT(\'Quit Date Calculator\'!C7,"mmm yyyy")&" optimistic, "&'
        'TEXT(\'Quit Date Calculator\'!C8,"mmm yyyy")&" conservative. '
        'The biggest swing factor is acquisition pace. Re-run when you close on a property."'
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(20, 23):
        ws.row_dimensions[r].height = 18

    # ---- Pseudo-button nav (rows 24-26) ----
    pseudo_button(ws, "A24", "C26", "Current State",
                  "'Current State'!A1", variant="primary")
    pseudo_button(ws, "D24", "F26", "Target State",
                  "'Target State'!A1", variant="primary")
    if is_lite:
        pseudo_button(ws, "G24", "I26", "Quit Date",
                      "'Quit Date Calculator'!A1", variant="primary")
        pseudo_button(ws, "J24", "L26", "Upgrade →",
                      "'Start'!A33", variant="accent",
                      external_link=f"https://{BRAND_DOMAIN}/escape-w2")
    else:
        pseudo_button(ws, "G24", "I26", "Acquisitions",
                      "'Acquisition Schedule'!A1", variant="primary")
        pseudo_button(ws, "J24", "L26", "Risk Plan",
                      "'Risk Plan'!A1", variant="primary")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    if not is_lite:
        pseudo_button(ws, "A28", "F30", "→  Quit Date Calculator",
                      "'Quit Date Calculator'!A1", variant="accent")
        pseudo_button(ws, "G28", "L30", "Conversation Doc (printable)",
                      "'Conversation Doc'!A1", variant="secondary")
        for r in range(28, 31):
            ws.row_dimensions[r].height = 22

    # Upgrade banner
    if is_lite:
        cta = (
            "Need healthcare bridge math, partner risk-share, and the printable "
            f"conversation doc? Upgrade to Full at {BRAND_DOMAIN}/escape-w2 — $47."
        )
    else:
        cta = (
            "Underwriting your next acquisition? Get the STR Deal Analyzer "
            f"at {BRAND_DOMAIN} — $47."
        )
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {label_suffix} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Current State
# ---------------------------------------------------------------------------

def build_current_state_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Current State")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Current State",
                        prev_tab="Start", next_tab="Target State")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 1 — what you have today: W2, partner income, STR portfolio, runway."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Plan anchor
    _section_band(ws, 6, "PLAN ANCHOR")
    _input_row(ws, 7, "Plan start date:", _val(variant, SAMPLE["plan_start_date"]),
               "yyyy-mm-dd", "Today's date — when you started this planning round")

    # W2 + household
    _section_band(ws, 9, "HOUSEHOLD INCOME")
    _input_row(ws, 10, "Your W2 gross ($/yr):",     _val(variant, SAMPLE["w2_gross"]),     '"$"#,##0', "")
    _input_row(ws, 11, "Your W2 net ($/yr):",       _val(variant, SAMPLE["w2_net"]),       '"$"#,##0', "After taxes, 401k, benefits")
    _input_row(ws, 12, "W2 benefits value ($/yr):", _val(variant, SAMPLE["w2_benefits"]),  '"$"#,##0', "Employer 401k match + health + life + disability")
    _input_row(ws, 13, "Partner W2 net ($/yr):",    _val(variant, SAMPLE["partner_w2_net"]), '"$"#,##0', "Zero if single-income household")

    # STR portfolio
    _section_band(ws, 15, "CURRENT STR PORTFOLIO")
    _input_row(ws, 16, "# properties owned:",       _val(variant, SAMPLE["props_owned"]),     "0", "")
    _input_row(ws, 17, "Portfolio gross ($/yr):",   _val(variant, SAMPLE["portfolio_gross"]), '"$"#,##0', "Total revenue across all properties YTD-annualized")
    _input_row(ws, 18, "Portfolio net ($/yr):",     _val(variant, SAMPLE["portfolio_net"]),   '"$"#,##0', "Schedule E line 26 — what hits your tax return")
    _input_row(ws, 19, "Avg stab. net per prop ($/yr):", _val(variant, SAMPLE["avg_stab_net_per_prop"]), '"$"#,##0', "Use a Y3 property — Y1/Y2 are depressed")

    # Expenses + runway
    _section_band(ws, 21, "EXPENSES & RUNWAY")
    _input_row(ws, 22, "Annual household expenses ($):", _val(variant, SAMPLE["household_expenses"]), '"$"#,##0', "All in — mortgage, food, kids, etc.")
    _input_row(ws, 23, "Annual savings rate ($):",       _val(variant, SAMPLE["annual_savings_rate"]), '"$"#,##0', "What you actually deploy into properties each year")
    _input_row(ws, 24, "Current liquid runway ($):",     _val(variant, SAMPLE["current_liquid"]),      '"$"#,##0', "Cash after closing on existing properties")

    # Derived rollups (rows 26-30) — Lite + Full both use these
    _section_band(ws, 26, "ROLLUP")
    _output_row(ws, 27, "Total household income (net):",
                "=C11+C13+C18", '"$"#,##0')
    _output_row(ws, 28, "Income replaced by STR today:",
                "=IFERROR(C18/(C11+C12),0)", "0.0%")
    _output_row(ws, 29, "Implied savings rate:",
                "=IFERROR(C23/(C11+C13+C18),0)", "0.0%")
    _output_row(ws, 30, "Years runway covers expenses:",
                "=IFERROR(C24/C22,0)", "0.0")

    brand_footer(ws, 33,
                 version_line=f"{SKU} · Current State")


# ---------------------------------------------------------------------------
# Sheet 3 — Target State
# ---------------------------------------------------------------------------

def build_target_state_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Target State")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    next_tab = "Acquisition Schedule" if not is_lite else "Quit Date Calculator"
    compact_header_band(ws, "Target State",
                        prev_tab="Current State", next_tab=next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 2 — what 'enough' looks like: replacement income, buffer, properties needed."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Replacement income
    _section_band(ws, 6, "REPLACEMENT INCOME")
    _input_row(ws, 7, "Target replacement income ($/yr):", _val(variant, SAMPLE["target_replacement"]), '"$"#,##0',
               "Default = current W2 net + benefits cost + healthcare bridge")
    _input_row(ws, 8, "Buffer multiple (×):",              _val(variant, SAMPLE["buffer_multiple"]),      "0.0",
               "1.5× = handle a bad year without panic")
    _input_row(ws, 9, "Target liquid reserve at quit ($):", _val(variant, SAMPLE["target_reserve"]),       '"$"#,##0',
               "Default 6 months of expenses")
    _input_row(ws, 10, "Target stabilized net per prop ($/yr):", _val(variant, SAMPLE["target_stab_net_per_prop"]), '"$"#,##0',
               "Use the per-property economics you actually achieve")

    # Derived
    _section_band(ws, 12, "WHAT THE MATH SAYS")
    _output_row(ws, 13, "Buffered income required ($/yr):",
                "=C7*C8", '"$"#,##0', emphasize=False)
    _output_row(ws, 14, "Less current STR portfolio net:",
                "='Current State'!C18", '"$"#,##0')
    _output_row(ws, 15, "Net new STR income required:",
                "=MAX(0,C13-C14)", '"$"#,##0')
    _output_row(ws, 16, "Properties to add (rounded up):",
                "=IFERROR(CEILING(C15/C10,1),0)", "0")
    _output_row(ws, 17, "Total properties at quit:",
                "='Current State'!C16+C16", "0")

    # The 4 critical numbers used by Start tab
    _section_band(ws, 19, "HEADLINE NUMBERS  (referenced by Start tab)")
    _output_row(ws, 20, "TOTAL PROPERTIES NEEDED:",
                "=C17", "0", emphasize=True)
    _output_row(ws, 21, "Cash deployment per property ($):",
                f"={SAMPLE['cash_per_property']}", '"$"#,##0')
    _output_row(ws, 22, "TOTAL CASH NEEDED AT QUIT ($):",
                "=C16*C21+C9", '"$"#,##0', emphasize=True)

    brand_footer(ws, 25,
                 version_line=f"{SKU} · Target State")


# ---------------------------------------------------------------------------
# Sheet 4 — Acquisition Schedule (Full only)
# ---------------------------------------------------------------------------

def build_acquisition_schedule_tab(wb, variant):
    ws = wb.create_sheet("Acquisition Schedule")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 8), ("C", 14), ("D", 16),
        ("E", 16), ("F", 16), ("G", 16),
        ("H", 14), ("I", 14), ("J", 14), ("K", 14), ("L", 14),
    ])

    compact_header_band(ws, "Acquisition Schedule",
                        prev_tab="Target State", next_tab="Risk Plan")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 3 — year-by-year property adds and cash deployment."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Pace assumptions
    _section_band(ws, 6, "PACE ASSUMPTIONS")
    _input_row(ws, 7, "Optimistic props/year:",  _val(variant, SAMPLE["opt_props_per_year"]),  "0.0", "Faster acquisition pace, current per-prop economics")
    _input_row(ws, 8, "Conservative props/year:", _val(variant, SAMPLE["cons_props_per_year"]), "0.0", "Slower pace, conservative occupancy")
    _input_row(ws, 9, "Cash per property ($):",   _val(variant, SAMPLE["cash_per_property"]),   '"$"#,##0', "Down + closing + rehab + furnish + reserve")
    _input_row(ws, 10, "Y1 net haircut (%):",      _val(variant, SAMPLE["y1_net_haircut"]),       "0.0%", "Y1 ≈ 65% of stabilized")

    # 7-year table
    _section_band(ws, 12, "7-YEAR SCHEDULE  (Optimistic pace)")
    headers = ["Year", "Plan year", "Props added", "Cash deployed",
               "Y1 net added", "Stabilized net total", "Cumulative net", "Reaches target?"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=13, column=2 + i, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[13].height = 22

    # Anchor — plan start year from Current State!C7
    # Each row references the prior row's cumulative
    target_ref = "'Target State'!C13"  # buffered income required

    for y in range(1, 8):
        r = 13 + y
        # Year label
        ws.cell(row=r, column=2, value=y).number_format = "0"
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        # Plan year (calendar)
        ws.cell(row=r, column=3, value=f"=YEAR('Current State'!C7)+{y - 1}").number_format = "0"
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        # Props added (cease additions once cumulative net hits target)
        ws.cell(row=r, column=4,
                value=f"=IF(H{r-1}>={target_ref},0,$C$7)" if y > 1
                      else f"=$C$7").number_format = "0.0"
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        # Cash deployed
        ws.cell(row=r, column=5, value=f"=D{r}*$C$9").number_format = '"$"#,##0'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
        # Y1 net added (this year's adds × per-prop stab × haircut)
        ws.cell(row=r, column=6,
                value=f"=D{r}*'Target State'!C10*$C$10").number_format = '"$"#,##0'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        # Stabilized net total cumulative — current portfolio + sum of all adds × stabilized
        # (assumes additions stabilize after Y2; simplification — count all adds at full stab)
        if y == 1:
            ws.cell(row=r, column=7,
                    value=f"='Current State'!C18+D{r}*'Target State'!C10*$C$10")
        else:
            ws.cell(row=r, column=7,
                    value=f"=G{r-1}+D{r}*'Target State'!C10*$C$10+SUM($D$14:$D{r-1})*'Target State'!C10*(1-$C$10)/2")
        ws.cell(row=r, column=7).number_format = '"$"#,##0'
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center")
        # Cumulative net — same as G for our simplified rolling stab model
        ws.cell(row=r, column=8, value=f"=G{r}").number_format = '"$"#,##0'
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="center")
        # Reaches target?
        ws.cell(row=r, column=9, value=f'=IF(H{r}>={target_ref},"YES","")')
        ws.cell(row=r, column=9).font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ACCENT)
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 18

    # Conditional fill on row reaching target
    from openpyxl.formatting.rule import FormulaRule
    gold_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    rule = FormulaRule(formula=[f'$I14>="YES"'], fill=gold_fill)
    ws.conditional_formatting.add("B14:I20", FormulaRule(
        formula=[f'$I14="YES"'], fill=gold_fill))

    # Footnote
    ws.merge_cells("A22:L24")
    c = ws["A22"]
    c.value = (
        "Note: schedule shows the optimistic pace. Conservative pace is "
        "modeled on the Quit Date Calculator tab. This table assumes you "
        "don't sell — additions are layered on top of the current portfolio. "
        "The Conversation Doc tab folds this into the partner-facing summary."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(22, 25):
        ws.row_dimensions[r].height = 14

    brand_footer(ws, 26,
                 version_line=f"{SKU} · Acquisition Schedule")


# ---------------------------------------------------------------------------
# Sheet 5 — Risk Plan (Full only)
# ---------------------------------------------------------------------------

def build_risk_plan_tab(wb, variant):
    ws = wb.create_sheet("Risk Plan")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Risk Plan",
                        prev_tab="Acquisition Schedule",
                        next_tab="Quit Date Calculator")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 4 — what no one talks about: healthcare bridge, insurance gap, partner buy-in."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "HEALTHCARE BRIDGE")
    _input_row(ws, 7,  "Current employer cost ($/yr):", _val(variant, SAMPLE["healthcare_employer_cost"]), '"$"#,##0', "Your share of premium today")
    _input_row(ws, 8,  "COBRA cost ($/yr, family):",    _val(variant, SAMPLE["healthcare_cobra_cost"]),    '"$"#,##0', "Typically 102% of full premium — DOL")
    _input_row(ws, 9,  "ACA cost ($/yr, family):",      _val(variant, SAMPLE["healthcare_aca_cost"]),       '"$"#,##0', "Post-subsidy mid-range — KFF benchmark")
    _input_row(ws, 10, "COBRA bridge (months):",         _val(variant, SAMPLE["cobra_months"]),               "0", "Max 18 per federal law")
    _input_row(ws, 11, "ACA bridge (months):",           _val(variant, SAMPLE["aca_bridge_months"]),          "0", "Until STR income covers ACA")

    _section_band(ws, 13, "INSURANCE & BENEFITS LOST")
    _input_row(ws, 14, "Disability insurance lost ($/yr):", _val(variant, SAMPLE["disability_lost"]), '"$"#,##0', "Replacement long-term disability premium")
    _input_row(ws, 15, "Life insurance lost ($/yr):",         _val(variant, SAMPLE["life_lost"]),       '"$"#,##0', "Term life premium to replace employer policy")

    _section_band(ws, 17, "PARTNER RISK-SHARE")
    _input_row(ws, 18, "Partner expense coverage (%):", _val(variant, SAMPLE["partner_risk_share_pct"]), "0.0%", "Share of household expenses partner W2 covers in transition")
    _input_row(ws, 19, "Partner buy-in (1-5):",          _val(variant, SAMPLE["partner_buyin"]),          "0", "1=skeptical, 5=enthusiastic — drives Conservative scenario")

    # Headline numbers — row 21 = section band, row 22+ = values.
    # The Start tab references 'Risk Plan'!C22 as the bridge total (note
    # the row position; cross-tab refs in this build use C22, not C20).
    _section_band(ws, 21, "HEADLINE NUMBERS  (referenced by Start tab as C22)")
    ws.cell(row=22, column=2, value="HEALTHCARE BRIDGE TOTAL ($):").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=22, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=22, column=3, value="=(C8*C10/12)+(C9*C11/12)")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[22].height = 22

    _output_row(ws, 23, "Insurance gap total ($/yr):",
                "=C14+C15", '"$"#,##0')
    _output_row(ws, 24, "Annual partner coverage ($):",
                "=C18*'Current State'!C22", '"$"#,##0')

    add_dropdown(ws, "C19", '"1,2,3,4,5"')

    brand_footer(ws, 27,
                 version_line=f"{SKU} · Risk Plan")


# ---------------------------------------------------------------------------
# Sheet 6 — Quit Date Calculator
# ---------------------------------------------------------------------------

def build_quit_date_calculator_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Quit Date Calculator")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 22),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    next_tab = None if is_lite else "Conversation Doc"
    compact_header_band(ws, "Quit Date Calculator",
                        prev_tab="Risk Plan" if not is_lite else "Target State",
                        next_tab=next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = ("Section " + ("3" if is_lite else "5") +
               " — the math behind the headline date. Adjust assumptions below.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- THE date readouts (rows 6-9) ----
    _section_band(ws, 6, "QUIT DATES")

    # Date math:
    # years_to_target_optimistic = MAX(0, props_to_add / opt_props_per_year)
    # quit_date_optimistic = EDATE(plan_start, years*12)
    # Conservative: same but with cons_props_per_year + buffer adjustment for partner buy-in
    if is_lite:
        # Simplified — no Risk Plan tab to reference
        opt_pace = SAMPLE["opt_props_per_year"]
        cons_pace = SAMPLE["cons_props_per_year"]
        opt_years = f"(MAX(0,'Target State'!C16/{opt_pace}))"
        cons_years = f"(MAX(0,'Target State'!C16/{cons_pace})*1.15)"
    else:
        opt_years = "(MAX(0,'Target State'!C16/C12))"
        # Conservative scales by (6 - partner_buyin)/3 — buy-in 5 = 0.33, buy-in 1 = 1.67 multiplier
        cons_years = "(MAX(0,'Target State'!C16/C13)*(1+(5-'Risk Plan'!C19)/10))"

    # Optimistic quit date — EDATE returns serial; cell formatted as date
    ws.cell(row=7, column=2, value="Optimistic quit date:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    ws.cell(row=7, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=7, column=3,
                   value=f"=EDATE('Current State'!C7,{opt_years}*12)")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "mmmm d, yyyy"
    ws.row_dimensions[7].height = 24

    ws.cell(row=8, column=2, value="Conservative quit date:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    ws.cell(row=8, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=8, column=3,
                   value=f"=EDATE('Current State'!C7,{cons_years}*12)")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "mmmm d, yyyy"
    ws.row_dimensions[8].height = 24

    ws.cell(row=9, column=2, value="Spread (years):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    ws.cell(row=9, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=9, column=3, value="=(C8-C7)/365.25")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0"
    ws.row_dimensions[9].height = 20

    # ---- Pace assumptions table (rows 11-15) ----
    _section_band(ws, 11, "ASSUMPTIONS  (edit to test scenarios)")
    if not is_lite:
        _input_row(ws, 12, "Optimistic props/year:",  _val(variant, SAMPLE["opt_props_per_year"]),  "0.0", "Faster pace + current per-prop economics")
        _input_row(ws, 13, "Conservative props/year:", _val(variant, SAMPLE["cons_props_per_year"]), "0.0", "Slower pace; further adjusted by partner buy-in")
        _input_row(ws, 14, "Optimistic stabilized occ:", _val(variant, SAMPLE["opt_stab_occ"]),     "0.0%", "Used by acquisition modeling")
        _input_row(ws, 15, "Conservative stabilized occ:", _val(variant, SAMPLE["cons_stab_occ"]),  "0.0%", "")
    else:
        # Lite: just show the figures from sample as fixed values
        _input_row(ws, 12, "Optimistic props/year:",  SAMPLE["opt_props_per_year"],  "0.0", "Lite: fixed assumption — upgrade for editable scenarios")
        _input_row(ws, 13, "Conservative props/year:", SAMPLE["cons_props_per_year"], "0.0", "Lite: fixed assumption — upgrade for editable scenarios")

    # ---- Headline numbers used by Start tab ----
    _section_band(ws, 17, "HEADLINE NUMBERS  (referenced by Start tab)")
    # Y1-after-quit projected net = total stabilized portfolio × Y1 haircut
    # under conservative scenario
    _output_row(ws, 18, "Stabilized portfolio at quit ($/yr):",
                "='Target State'!C17*'Target State'!C10", '"$"#,##0')
    _output_row(ws, 19, "Y1 PROJECTED NET ($/yr):",
                f"=C18*{SAMPLE['y1_net_haircut']}", '"$"#,##0', emphasize=True)

    brand_footer(ws, 22,
                 version_line=f"{SKU} · Quit Date Calculator")


# ---------------------------------------------------------------------------
# Sheet 7 — Conversation Doc (Full only, printable)
# ---------------------------------------------------------------------------

def build_conversation_doc_tab(wb):
    ws = wb.create_sheet("Conversation Doc")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 18), ("E", 18),
        ("F", 4), ("G", 4), ("H", 4), ("I", 4),
        ("J", 4), ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Conversation Doc",
                        prev_tab="Quit Date Calculator", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill

    # Header
    ws.merge_cells("A4:E5")
    c = ws["A4"]
    c.value = '="Our Path Off W2 — "&TEXT(TODAY(),"mmmm d, yyyy")'
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 16

    ws.merge_cells("A7:E7")
    c = ws["A7"]
    c.value = "A calm reading of where we are, what we need, and when the math says we can."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 20

    # Quit dates side by side
    _section_band(ws, 9, "QUIT DATE")
    ws.merge_cells("B10:C10")
    c = ws["B10"]
    c.value = "Optimistic"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D10:E10")
    c = ws["D10"]
    c.value = "Conservative"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B11:C12")
    c = ws["B11"]
    c.value = "='Quit Date Calculator'!C7"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "mmm d, yyyy"
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.merge_cells("D11:E12")
    c = ws["D11"]
    c.value = "='Quit Date Calculator'!C8"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "mmm d, yyyy"
    ws.row_dimensions[11].height = 24
    ws.row_dimensions[12].height = 14

    # 4 critical numbers
    _section_band(ws, 14, "THE FOUR NUMBERS")
    crit = [
        (15, "Properties needed:",         "='Target State'!C20",                "0"),
        (16, "Cash needed at quit:",       "='Target State'!C22",                '"$"#,##0'),
        (17, "Healthcare bridge:",         "='Risk Plan'!C22",                   '"$"#,##0'),
        (18, "Y1-after-quit projected net:", "='Quit Date Calculator'!C19",      '"$"#,##0'),
    ]
    for r, label, formula, fmt in crit:
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1)
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Three biggest swing factors
    _section_band(ws, 20, "THREE BIGGEST SWING FACTORS")
    swings = [
        (21, "1. Acquisition pace",
         '="Optimistic adds "&TEXT(\'Quit Date Calculator\'!C12,"0.0")&" props/yr; conservative "&TEXT(\'Quit Date Calculator\'!C13,"0.0")&"/yr."'),
        (22, "2. Healthcare bridge",
         '="Total bridge cost "&TEXT(\'Risk Plan\'!C20,"$#,##0")&" across "&\'Risk Plan\'!C10&"mo COBRA + "&\'Risk Plan\'!C11&"mo ACA."'),
        (23, "3. Partner buy-in",
         '="Buy-in level "&\'Risk Plan\'!C19&"/5. Partner covers "&TEXT(\'Risk Plan\'!C18,"0%")&" of expenses during transition."'),
    ]
    for r, label, formula in swings:
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"C{r}:E{r}")
        cell = ws.cell(row=r, column=3, value=formula)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 22

    # Action plan
    _section_band(ws, 25, "ACTION PLAN — NEXT THREE PROPERTIES")
    headers = ["#", "Plan year", "Cash needed", "Notes"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=26, column=2 + i, value=h)
        apply_style(cell, header_row_style())
    for n in range(1, 4):
        r = 26 + n
        ws.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3,
                value=f"=YEAR('Current State'!C7)+{n - 1}").number_format = "0"
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4,
                value="='Acquisition Schedule'!C9").number_format = '"$"#,##0'
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        cell = ws.cell(row=r, column=5, value="")
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Risk acknowledgments
    _section_band(ws, 31, "RISK ACKNOWLEDGMENTS")
    risks = [
        "We have priced the healthcare bridge into our target replacement income.",
        "We have a partner risk-share plan if STR income disappoints in Y1.",
        "We will re-run this workbook after every property close.",
        "We will not quit until the conservative cash flow figure is positive.",
    ]
    for i, txt in enumerate(risks):
        r = 32 + i
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]
        c.value = "·  " + txt
        c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 18

    # Sign-off
    _section_band(ws, 38, "PARTNER SIGN-OFF")
    for i, (label, anchor) in enumerate([("Signed:", 39), ("Signed:", 41)]):
        ws.cell(row=anchor, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        ws.cell(row=anchor, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1)
        ws.merge_cells(f"C{anchor}:E{anchor}")
        cell = ws.cell(row=anchor, column=3, value="")
        cell.border = Border(bottom=Side(style="thin", color=COLOR_PRIMARY))
        ws.row_dimensions[anchor].height = 22

    ws.cell(row=40, column=2, value="Date:").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    ws.cell(row=40, column=2).alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=42, column=2, value="Date:").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    ws.cell(row=42, column=2).alignment = Alignment(horizontal="right", indent=1)

    # Print setup
    ws.print_area = "A1:E45"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant, is_lite):
    wb = Workbook()
    build_start_tab(wb, variant, is_lite)
    build_current_state_tab(wb, variant, is_lite)
    build_target_state_tab(wb, variant, is_lite)
    if not is_lite:
        build_acquisition_schedule_tab(wb, variant)
        build_risk_plan_tab(wb, variant)
    build_quit_date_calculator_tab(wb, variant, is_lite)
    if not is_lite:
        build_conversation_doc_tab(wb)

    wb.properties.title = "Escape the W2 Planner — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Career-transition planning workbook for STR investors with W2 jobs. "
        "Returns optimistic and conservative quit dates plus a printable "
        "partner-conversation doc."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo", is_lite=False)
    build_workbook(BLANK_OUT, "blank", is_lite=False)
    build_workbook(LITE_OUT, "demo", is_lite=True)


if __name__ == "__main__":
    main()
