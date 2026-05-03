"""Build REV-003 Competitor Rate Tracker Excel files (v2.2 standard).

Operational-mode tracker: host snapshots 6-10 comp listings each Monday, types
nightly rates into a weekly matrix, and watches their position-vs-comp-set drift
across the season. Replaces a $20/mo PriceLabs subscription for hosts who just
want a gut-check, not full dynamic pricing.

Tabs:
    1. Start            — KPI cards + verdict + nav
    2. Comp Set         — 10-row roster (listing, platform, URL, BR/BA, etc.)
    3. Weekly Snapshots — 52-week matrix: My rate | Comp1..Comp10 | Avg | Pos %
    4. Position Trend   — LineChart "My rate vs comp avg" + trailing-4w KPI
    5. Settings         — Property selector, active year (B7), snapshot day

Forks build_pl_single_property.py for matrix pattern + brand v2.2 chrome.

Critical correctness rules honored:
- Active tax year cell on Settings (B7) — referenced where year context matters
- Dates written as datetime.date, not strings, so date filters match
- DEMO + BLANK from one shared codepath (variant flag)
- Never write to a merged cell after merge — every merge() is followed by a
  write to the top-left only.
- All colors come from brand_config; no raw hex literal in this file.
"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import LineChart, Reference

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "REV-003"
NAME = "competitor-rate-tracker"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_TAX_YEAR = 2026
SNAPSHOT_WEEKS = 52     # full-year capacity
COMP_CAPACITY = 10      # 10 comp columns regardless of variant
DEMO_COMPS_USED = 6     # only first 6 are populated in DEMO

# Comp-Set roster (DEMO uses 6; BLANK leaves rows empty).
SAMPLE_COMPS = [
    # name, platform, url, br_ba, sleeps, dist_mi, note, active
    ("Smokies Lookout Cabin",   "Airbnb", "https://airbnb.com/h/smokies-lookout", "3BR/2BA", 8,  1.2, "Newer hot tub, mountain view",      "Y"),
    ("Pigeon Forge Pines",      "VRBO",   "https://vrbo.com/h/pigeon-pines",      "3BR/3BA", 8,  2.4, "Pool table, theater room",          "Y"),
    ("Gatlinburg Hideaway",     "Airbnb", "https://airbnb.com/h/gat-hideaway",    "2BR/2BA", 6,  0.8, "Walk to downtown",                  "Y"),
    ("Wears Valley Retreat",    "Airbnb", "https://airbnb.com/h/wears-retreat",   "3BR/2BA", 8,  3.1, "Quiet road, fire pit",              "Y"),
    ("Bear Cove Lodge",         "VRBO",   "https://vrbo.com/h/bear-cove",         "4BR/3BA", 10, 1.7, "Pet-friendly, larger sleeps",       "Y"),
    ("Smoky Ridge Sunrise",     "Airbnb", "https://airbnb.com/h/smoky-sunrise",   "3BR/2BA", 8,  2.0, "Recently renovated, gas grill",     "Y"),
]

# 12 weeks of DEMO snapshots — Q1 2026 (weeks 1..12)
# Each week: my rate, then 6 comp rates. Numbers chosen so my avg sits ~8% below.
DEMO_SNAPSHOTS = [
    # (week_no, week_start_iso, my_rate, [comp1..comp6])
    ( 1, "2026-01-05", 245, [275, 290, 260, 270, 310, 285]),
    ( 2, "2026-01-12", 240, [270, 285, 255, 265, 305, 280]),
    ( 3, "2026-01-19", 235, [265, 280, 250, 260, 300, 275]),
    ( 4, "2026-01-26", 250, [275, 290, 260, 270, 310, 285]),
    ( 5, "2026-02-02", 260, [285, 300, 270, 280, 320, 295]),
    ( 6, "2026-02-09", 280, [320, 340, 295, 305, 360, 330]),  # Valentine's
    ( 7, "2026-02-16", 270, [305, 320, 285, 295, 340, 315]),
    ( 8, "2026-02-23", 255, [285, 300, 270, 280, 320, 295]),
    ( 9, "2026-03-02", 250, [280, 295, 265, 275, 315, 290]),
    (10, "2026-03-09", 255, [285, 305, 270, 280, 320, 295]),
    (11, "2026-03-16", 265, [295, 315, 280, 290, 335, 305]),  # St Patrick's
    (12, "2026-03-23", 260, [290, 310, 275, 285, 330, 300]),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value for the DEMO build, None for BLANK."""
    return demo_value if variant == "demo" else None


def _parse_date(s):
    """Parse 'YYYY-MM-DD' to a real date. Avoids text-vs-date filter mismatches."""
    parts = s.split("-")
    return date(int(parts[0]), int(parts[1]), int(parts[2]))


def _week_end(start):
    """Sunday-end of a Monday-start week."""
    return start + timedelta(days=6)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _box_gold(ws, top_row, bottom_row, first_col, last_col):
    """Draw a thin gold border around a rectangular range without
    clobbering interior fills. Each cell's existing border is preserved on
    edges that aren't on the perimeter."""
    side = Side(style="thin", color=COLOR_ACCENT)
    fc = column_index_from_string(first_col) if isinstance(first_col, str) else first_col
    lc = column_index_from_string(last_col) if isinstance(last_col, str) else last_col
    for r in range(top_row, bottom_row + 1):
        for c in range(fc, lc + 1):
            existing = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(
                top=side if r == top_row else existing.top,
                bottom=side if r == bottom_row else existing.bottom,
                left=side if c == fc else existing.left,
                right=side if c == lc else existing.right,
            )


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 1 — Start: navy hero + verdict + KPI cards + nav."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: navy hero band (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Competitor Rate Tracker"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Know where your nightly rate sits — without paying $20/mo."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT — single headline reading off the trailing-4-week position.
    # 'Position Trend'!B7 holds the trailing-4-week position % (computed there).
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(COUNT(\'Weekly Snapshots\'!N7:N58)=0,'
        '"\U0001F4CA  Add comp rates this week to see your verdict.",'
        'IF(\'Position Trend\'!B7<-0.05,'
        '"⚠  You\'re "&TEXT(ABS(\'Position Trend\'!B7),"0%")'
        '&" below comp avg (trailing 4 weeks) — room to lift.",'
        'IF(\'Position Trend\'!B7>0.05,'
        '"\U0001F4B0  You\'re "&TEXT(\'Position Trend\'!B7,"0%")'
        '&" above comp avg (trailing 4 weeks) — protect occupancy.",'
        '"✅  Within 5% of comp avg (trailing 4 weeks) — priced in line.")))'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: "What this does" card (rows 10-15)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L14")
    c = ws["A11"]
    c.value = (
        "Pick 6-10 comp listings on Airbnb/VRBO. Each Monday, snapshot their nightly "
        "rate for the next 4 weekends and type the numbers into the Weekly Snapshots "
        "tab. The workbook tracks how your rate is moving relative to the comp set "
        "average — so you can lift before holiday weekends and hold steady when "
        "comps drop. No scraping, no subscription, no API."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 3: How to use card (rows 17-23)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(17, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 20

    steps = [
        "① Comp Set: pick 6-10 lookalike listings (same BR/BA, area, capacity).",
        "② Weekly Snapshots: each Monday, type comp nightly rates and your own.",
        "③ Position Trend: watch your position % vs the comp average across weeks.",
        "④ Lift before holiday weekends when comps move; hold when they drop.",
        "⑤ Settings: bump 'Active year' each January and clear last year's data.",
        "⑥ No automated scraping — manual snapshots only (Airbnb/VRBO ToS).",
    ]
    for i, item in enumerate(steps):
        row = 18 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 18

    # ZONE 4: primary OPEN button (rows 25-28)
    pseudo_button(ws, "A25", "L28",
                   "→  OPEN WEEKLY SNAPSHOTS",
                   "'Weekly Snapshots'!A1", variant="primary")
    for r in range(25, 29):
        ws.row_dimensions[r].height = 22

    # ZONE 5: KPI cards (rows 30-35) — My avg / Comp avg / Position %
    for r in range(30, 36):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): My avg rate (across logged weeks)
    ws.merge_cells("A30:D30")
    c = ws["A30"]
    c.value = "MY AVG RATE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A31:D33")
    c = ws["A31"]
    c.value = ("=IFERROR(AVERAGE('Weekly Snapshots'!C7:C58),0)")
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A34:D34")
    c = ws["A34"]
    c.value = "logged weeks only"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Comp set avg
    ws.merge_cells("E30:H30")
    c = ws["E30"]
    c.value = "COMP SET AVG"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E31:H33")
    c = ws["E31"]
    c.value = ("=IFERROR(AVERAGE('Weekly Snapshots'!N7:N58),0)")
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E34:H34")
    c = ws["E34"]
    c.value = "active comps only"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): My position %
    ws.merge_cells("I30:L30")
    c = ws["I30"]
    c.value = "MY POSITION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I31:L33")
    c = ws["I31"]
    c.value = (
        "=IFERROR(AVERAGE('Weekly Snapshots'!C7:C58)/"
        "AVERAGE('Weekly Snapshots'!N7:N58)-1,0)"
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "+0%;-0%;0%"
    ws.merge_cells("I34:L34")
    c = ws["I34"]
    c.value = "vs comp set, all weeks"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    _box_gold(ws, 30, 34, "A", "D")
    _box_gold(ws, 30, 34, "E", "H")
    _box_gold(ws, 30, 34, "I", "L")

    # ZONE 6: secondary nav buttons (rows 37-38)
    pseudo_button(ws, "A37", "C38", "Comp Set",
                   "'Comp Set'!A1", variant="secondary")
    pseudo_button(ws, "D37", "F38", "Weekly Snapshots",
                   "'Weekly Snapshots'!A1", variant="secondary")
    pseudo_button(ws, "G37", "I38", "\U0001F4C8 Position Trend",
                   "'Position Trend'!A1", variant="accent")
    pseudo_button(ws, "J37", "L38", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[37].height = 22
    ws.row_dimensions[38].height = 22

    # ZONE 7: anti-pattern callout (row 40)
    ws.merge_cells("A40:L40")
    c = ws["A40"]
    c.value = (
        "⚠ No automated scraping. This is a manual snapshot tool — typing rates "
        "weekly keeps you within Airbnb/VRBO ToS and forces a human read of the "
        "market each Monday."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[40].height = 28

    # ZONE 8: upgrade banner + footer
    add_upgrade_banner(ws, 42)
    brand_footer(ws, 44,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    ws.print_area = "A1:L46"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_comp_set_tab(wb, variant):
    """Sheet 2 — Comp Set: 10-row roster of tracked comps."""
    ws = wb.create_sheet("Comp Set")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Comp Set",
                         prev_tab="Start", next_tab="Weekly Snapshots")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Pick 6-10 lookalike listings — same BR/BA, area, sleeps, vibe"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 8),   # Comp #
        ("B", 30),  # Listing name
        ("C", 12),  # Platform
        ("D", 36),  # URL
        ("E", 10),  # BR/BA
        ("F", 8),   # Sleeps
        ("G", 10),  # Distance
        ("H", 32),  # Note
        ("I", 12),  # Active?
    ])

    # Header row at 6
    headers = ["Comp #", "Listing name", "Platform", "Listing URL",
               "BR/BA", "Sleeps", "Dist (mi)", "Note", "Active?"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # 10 data rows (rows 7-16) — DEMO populates first 6
    for i in range(COMP_CAPACITY):
        row = 7 + i
        # Comp #
        n = ws.cell(row=row, column=1, value=f"Comp {i + 1}")
        n.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        n.alignment = Alignment(horizontal="center", vertical="center")

        if variant == "demo" and i < DEMO_COMPS_USED:
            name, platform, url, brba, sleeps, dist, note, active = SAMPLE_COMPS[i]
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=platform)
            ws.cell(row=row, column=4, value=url)
            ws.cell(row=row, column=5, value=brba)
            ws.cell(row=row, column=6, value=sleeps)
            ws.cell(row=row, column=7, value=dist)
            ws.cell(row=row, column=8, value=note)
            ws.cell(row=row, column=9, value=active)
        # Style every cell as input (yellow tint) regardless of variant
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, input_cell_style())
        # Format distance as number
        ws.cell(row=row, column=7).number_format = "0.0"
        ws.cell(row=row, column=6).number_format = "0"
        ws.row_dimensions[row].height = 18

    # Dropdowns
    add_dropdown(ws, "C7:C16", '"Airbnb,VRBO,Booking.com,Direct,Other"')
    add_dropdown(ws, "I7:I16", '"Y,N"')

    # Footnote
    ws.merge_cells("A18:I18")
    c = ws["A18"]
    c.value = (
        "Tip: anonymized names are fine (\"Comp A\", \"Pines #1\"). The URL is for "
        "your own reference — paste a screenshot folder link if Airbnb listing "
        "URLs change between snapshots."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[18].height = 28

    brand_footer(ws, 20, version_line=f"{SKU} · v2.2")

    ws.freeze_panes = "B7"
    ws.print_area = "A1:I22"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_weekly_snapshots(wb, variant):
    """Sheet 3 — Weekly Snapshots matrix.

    Layout (rows 7..58 = 52 weeks):
        A  Week #
        B  Date span (start..end)
        C  My rate
        D..M  Comp1..Comp10 rates
        N  Comp avg (formula)
        O  My position % (=C/N - 1)
    """
    ws = wb.create_sheet("Weekly Snapshots")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Weekly Snapshots",
                         prev_tab="Comp Set", next_tab="Position Trend")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 16):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:O4")
    c4 = ws["A4"]
    c4.value = ("Each Monday, type comp rates for the upcoming weekend. "
                "Comp avg + position % auto-fill.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column widths
    widths = [
        ("A", 8),   # Week #
        ("B", 22),  # Date span
        ("C", 12),  # My rate
    ]
    for i in range(COMP_CAPACITY):
        widths.append((get_column_letter(4 + i), 11))
    widths.append(("N", 12))  # Comp avg
    widths.append(("O", 13))  # Position %
    set_col_widths(ws, widths)

    # Header row 6
    header_labels = ["Week", "Date span", "My rate"]
    for i in range(COMP_CAPACITY):
        header_labels.append(f"Comp {i + 1}")
    header_labels.append("Comp avg")
    header_labels.append("Position %")
    for col_idx, h in enumerate(header_labels, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # Data rows: 52 weeks (rows 7..58)
    # Anchor week 1 to first Monday of ACTIVE_TAX_YEAR for the BLANK template
    # so date spans land plausibly even before user edits.
    anchor = date(ACTIVE_TAX_YEAR, 1, 5)  # first Monday of 2026 ISO calendar
    # Pre-build a quick lookup for DEMO snapshots
    demo_map = {row[0]: row for row in DEMO_SNAPSHOTS}

    for w in range(1, SNAPSHOT_WEEKS + 1):
        row = 6 + w  # week 1 -> row 7
        ws.row_dimensions[row].height = 18

        # Week #
        cell = ws.cell(row=row, column=1, value=w)
        cell.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Date span
        wk_start = anchor + timedelta(weeks=w - 1)
        wk_end = _week_end(wk_start)
        span_str = f"{wk_start.strftime('%b %d')} – {wk_end.strftime('%b %d')}"
        sb = ws.cell(row=row, column=2, value=span_str)
        sb.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
        sb.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # My rate (col C)
        my_cell = ws.cell(row=row, column=3)
        apply_style(my_cell, input_cell_style())
        my_cell.number_format = '"$"#,##0'
        if variant == "demo" and w in demo_map:
            my_cell.value = demo_map[w][2]

        # Comp rates D..M (cols 4..13)
        for i in range(COMP_CAPACITY):
            col = 4 + i
            comp_cell = ws.cell(row=row, column=col)
            apply_style(comp_cell, input_cell_style())
            comp_cell.number_format = '"$"#,##0'
            if variant == "demo" and w in demo_map and i < DEMO_COMPS_USED:
                comp_cell.value = demo_map[w][3][i]

        # Comp avg (col N = 14): AVERAGE of D..M, only over numeric cells
        avg_cell = ws.cell(row=row, column=14)
        avg_cell.value = f"=IFERROR(AVERAGE(D{row}:M{row}),\"\")"
        apply_style(avg_cell, formula_cell_style())
        avg_cell.number_format = '"$"#,##0'

        # Position % (col O = 15) = My/CompAvg - 1
        pos_cell = ws.cell(row=row, column=15)
        pos_cell.value = (
            f"=IFERROR(IF(N{row}=\"\",\"\",C{row}/N{row}-1),\"\")"
        )
        apply_style(pos_cell, formula_cell_style())
        pos_cell.number_format = "+0%;-0%;0%"

    # Freeze on row 7 (data start) and column C (My rate)
    ws.freeze_panes = "C7"

    # Footer
    brand_footer(ws, 60, version_line=f"{SKU} · v2.2")

    # Print
    ws.print_area = f"A1:O{60 + 2}"
    ws.print_title_rows = "1:6"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_position_trend(wb, variant):
    """Sheet 4 — Position Trend: My rate vs comp avg LineChart + KPI."""
    ws = wb.create_sheet("Position Trend")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Position Trend",
                         prev_tab="Weekly Snapshots", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Trailing-4-week position vs comp avg, plus full-year line"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 28), ("B", 16), ("C", 4),
        ("D", 12), ("E", 14), ("F", 14), ("G", 14),
    ])

    # KPI block: trailing-4-week position % at B7
    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    right = Alignment(horizontal="right", vertical="center")

    a6 = ws.cell(row=6, column=1, value="Trailing 4-week position vs comp avg:")
    a6.font = bold_right
    a6.alignment = right
    ws.row_dimensions[6].height = 18

    # Row 7: B7 = trailing-4-week position. Computed as
    # AVERAGE(last 4 my rates) / AVERAGE(last 4 comp avgs) - 1.
    # Falls back to 0 when fewer than 4 weeks logged.
    a7 = ws.cell(row=7, column=1, value="(my last-4-week avg) / (comp last-4-week avg) − 1:")
    a7.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    a7.alignment = right
    ws.row_dimensions[7].height = 22

    # B7 holds the formula — referenced by Start verdict.
    b7 = ws.cell(row=7, column=2)
    b7.value = (
        "=IFERROR("
        "AVERAGE(OFFSET('Weekly Snapshots'!C58,-3,0,4,1))/"
        "AVERAGE(OFFSET('Weekly Snapshots'!N58,-3,0,4,1))-1,"
        "0)"
    )
    apply_style(b7, formula_cell_style())
    b7.number_format = "+0%;-0%;0%"
    b7.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_ACCENT)
    b7.alignment = Alignment(horizontal="center", vertical="center")

    _box_gold(ws, 6, 7, "A", "B")

    # Reading row
    ws.merge_cells("A9:G9")
    c = ws["A9"]
    c.value = (
        "Read it like this: −5% means you’re priced 5% below the comp average "
        "for the trailing 4 weeks — typically a lift opportunity if occupancy is full."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[9].height = 28

    # ------------------------------------------------------------------
    # LineChart "My rate vs comp avg" — anchored at I5
    # Categories: Week column (A7:A58 on Weekly Snapshots)
    # Series 1: My rate  (C7:C58)
    # Series 2: Comp avg (N7:N58)
    # ------------------------------------------------------------------
    src = wb["Weekly Snapshots"]
    chart = LineChart()
    chart.title = "My rate vs comp avg"
    chart.y_axis.title = "Nightly rate ($)"
    chart.x_axis.title = "Week"
    chart.height = 11
    chart.width = 22

    my_ref = Reference(src, min_col=3, min_row=6, max_col=3, max_row=58)
    comp_ref = Reference(src, min_col=14, min_row=6, max_col=14, max_row=58)
    cats_ref = Reference(src, min_col=1, min_row=7, max_col=1, max_row=58)

    chart.add_data(my_ref, titles_from_data=True)
    chart.add_data(comp_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    style_chart(chart)
    ws.add_chart(chart, "A11")

    brand_footer(ws, 34, version_line=f"{SKU} · v2.2")

    ws.print_area = "A1:L36"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 5 — Settings: property selector, active year (B7), snapshot day."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Position Trend", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Property name · active year · snapshot day"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 22)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    right = Alignment(horizontal="right", vertical="center")

    # Row 5 — Property
    a5 = ws.cell(row=5, column=1, value="My property:")
    a5.font = bold_right
    a5.alignment = right
    b5 = ws.cell(row=5, column=2,
                  value=_val(variant, "Smokies Ridge Cabin"))
    apply_style(b5, input_cell_style())
    ws.row_dimensions[5].height = 18

    # Row 6 spacer
    ws.row_dimensions[6].height = 6

    # Row 7 — Active tax year (this is the "active-year cell" — B7)
    a7 = ws.cell(row=7, column=1, value="Active year:")
    a7.font = bold_right
    a7.alignment = right
    b7 = ws.cell(row=7, column=2, value=ACTIVE_TAX_YEAR)
    apply_style(b7, input_cell_style())
    b7.number_format = "0"
    ws.row_dimensions[7].height = 18

    ws.merge_cells("A8:B8")
    c8 = ws["A8"]
    c8.value = ("Bump each January, then clear last year’s snapshot rows.")
    c8.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c8.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[8].height = 18

    # Row 9 — Snapshot day
    a9 = ws.cell(row=9, column=1, value="Snapshot day:")
    a9.font = bold_right
    a9.alignment = right
    b9 = ws.cell(row=9, column=2, value=_val(variant, "Monday"))
    if variant != "demo":
        b9.value = "Monday"  # sensible default for BLANK too
    apply_style(b9, input_cell_style())
    add_dropdown(ws, "B9",
                  '"Monday,Tuesday,Wednesday,Thursday,Friday,Saturday,Sunday"')
    ws.row_dimensions[9].height = 18

    # Row 10 — last-snapshot date (a real datetime.date so date filters work)
    a10 = ws.cell(row=10, column=1, value="Last snapshot date:")
    a10.font = bold_right
    a10.alignment = right
    b10 = ws.cell(row=10, column=2,
                   value=_val(variant, _parse_date("2026-03-23")))
    apply_style(b10, input_cell_style())
    b10.number_format = "yyyy-mm-dd"
    ws.row_dimensions[10].height = 18

    # Section: notes
    ws.row_dimensions[11].height = 8
    a12 = ws.cell(row=12, column=1, value="Notes")
    a12.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[12].height = 22

    ws.merge_cells("A13:B17")
    c = ws["A13"]
    c.value = (
        "• Pick a single weekday and stick to it — same-day comparisons "
        "matter more than total volume.\n"
        "• If a comp listing disappears, swap it on the Comp Set tab; the "
        "matrix column carries forward.\n"
        "• The Position Trend chart reads weeks 1–52 — logged weeks "
        "appear on the line, blank weeks gap.\n"
        "• Year roll-over: copy this workbook to a new file, clear C7:M58 on "
        "Weekly Snapshots, bump Active year above."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=1)

    brand_footer(ws, 19, version_line=f"{SKU} · v2.2")

    ws.print_area = "A1:L21"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_comp_set_tab(wb, variant)
    build_weekly_snapshots(wb, variant)
    build_position_trend(wb, variant)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Competitor Rate Tracker{suffix} — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Manual weekly rate-snapshot tool for STR hosts — my rate vs comp set "
        "average across 52 weeks, with position-trend line chart (v2.2)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
