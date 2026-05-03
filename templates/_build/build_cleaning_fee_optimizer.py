"""Build REV-001 Cleaning Fee Optimizer (v2.2 standard).

Wizard-mode tool — answer the question "should I bundle my cleaning fee
into the nightly rate?" in five tabs. Compares 3 strategies side-by-side,
estimates Airbnb search-rank impact, and (Full only) breaks the answer
out by length-of-stay cohort.

Generates three files:
  templates/_masters/REV-001-cleaning-fee-optimizer-DEMO.xlsx   (Full + sample data)
  templates/_masters/REV-001-cleaning-fee-optimizer-BLANK.xlsx  (Full + empty)
  templates/_lite/REV-001-cleaning-fee-optimizer-lite.xlsx      (Lite, 4 tabs, sample data)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
    STATE_BAD_FILL, STATE_GOOD_TEXT,
)

SKU = "REV-001"
NAME = "cleaning-fee-optimizer"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
LITE_OUT = BASE / "_lite" / f"{SKU}-{NAME}-lite.xlsx"

# ---------------------------------------------------------------------------
# Sample data (per brief QA section — Smokies cabin profile)
# ---------------------------------------------------------------------------

SAMPLE = {
    "property_name":   "Smokies Ridge Cabin",
    "cleaning_paid":   185,     # what host pays cleaner per turnover
    "current_fee":     185,     # what host charges guest currently
    "current_adr":     245,     # current nightly rate
    "avg_los":         3.2,     # average length of stay
    "monthly_bookings": 8,
    "platform_pct":    0.15,    # 15% Airbnb commission

    # LOS distribution (Full only) — must sum to 1.0
    "los_short_pct":   0.40,    # 1-2 night bookings
    "los_mid_pct":     0.45,    # 3-4 night bookings
    "los_long_pct":    0.15,    # 5+ night bookings

    # Representative LOS for each cohort (used in the cohort-level math)
    "los_short_nights": 1.5,
    "los_mid_nights":   3.5,
    "los_long_nights":  6.0,
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (BLANK)."""
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


# ---------------------------------------------------------------------------
# Formula building blocks — referenced by Strategy Compare + LOS Sensitivity
# ---------------------------------------------------------------------------
# Inputs tab cell map (value cells, NOT label cells):
#   C7  = property name
#   C8  = cleaning paid ($/turnover)
#   C9  = current cleaning fee charged ($/turnover)
#   C10 = current ADR ($)
#   C11 = avg LOS (nights)
#   C12 = monthly bookings
#   C13 = platform commission %
#   C16 = LOS short %
#   C17 = LOS mid %
#   C18 = LOS long %
#
# Strategy Compare tab — 3 columns, B/C/D
#   Col B: Strategy A (Full Fee)
#   Col C: Strategy B (Partial Bundle 50%)
#   Col D: Strategy C (Fully Bundled)
#
# Per-strategy formulas (using avg LOS = Inputs!C11):
#   Cleaning fee charged
#     A: =Inputs!C9
#     B: =Inputs!C9*0.5
#     C: =0
#   Per-night rate uplift = (current_fee - new_fee) / LOS
#     A: =Inputs!C10
#     B: =Inputs!C10 + (Inputs!C9 - Inputs!C9*0.5)/Inputs!C11
#     C: =Inputs!C10 + Inputs!C9/Inputs!C11
#   Gross per booking = ADR*LOS + cleaning_fee
#   Platform take = Gross * platform_pct
#   Cleaning paid = Inputs!C8 (a real cost the host pays the cleaner)
#   Net per booking = Gross - platform_take - cleaning_paid
#   Annual net = Net per booking * monthly_bookings * 12

# These return formula strings keyed off the column letter on the Strategy
# Compare tab so we can re-use them for both compare row math and the
# Start tab readouts.


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant, is_lite):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — Hero band rows 1-9
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 10):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    title = "Cleaning Fee Optimizer"
    if is_lite:
        title += " — Lite"
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = title
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "The fee strategy that nets you the most."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    label_suffix = "LITE" if is_lite else variant.upper()
    c.value = f"{SKU} · v1.0 · {label_suffix}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — Recommendation block rows 11-22
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(11, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Inputs!C7="","(enter property on Inputs tab)",Inputs!C7)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # RECOMMENDED label
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "RECOMMENDED STRATEGY"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # The strategy name comes from Recommendation tab (resolved formula there)
    # Recommendation!C5 = recommended strategy name string
    # Recommendation!C6 = annual lift $
    # Recommendation!C7 = visibility lift % (Full only)
    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = "=Recommendation!C5"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 18

    # Annual lift $ — main number
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "ANNUAL REVENUE LIFT"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 16

    ws.merge_cells("A18:L19")
    c = ws["A18"]
    c.value = '=IF(Recommendation!C6=0,"$0 — current strategy is already optimal","+"&TEXT(Recommendation!C6,"$#,##0")&" / year")'
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 28
    ws.row_dimensions[19].height = 16

    # Visibility lift (Full only)
    if not is_lite:
        ws.merge_cells("A21:L21")
        c = ws["A21"]
        c.value = '="+"&TEXT(Recommendation!C7,"0%")&" search visibility est."'
        c.font = Font(name=FONT_HEAD, size=14, italic=True, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[21].height = 20
    else:
        ws.merge_cells("A21:L21")
        c = ws["A21"]
        c.value = "Search-rank impact estimate available in Full version."
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[21].height = 20

    # Math summary 3-line (rows 23-25 ish, but inside parchment band done above)
    # Reuse rows 23-25 for math summary
    parchment_fill_alt = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(25, 30):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill_alt

    ws.merge_cells("A25:L25")
    c = ws["A25"]
    c.value = "MATH SUMMARY"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[25].height = 18

    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = '="Cleaning paid to cleaner: "&TEXT(Inputs!C8,"$#,##0")&" per turnover"'
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[26].height = 18

    ws.merge_cells("A27:L27")
    c = ws["A27"]
    c.value = '="Cleaning charged to guest today: "&TEXT(Inputs!C9,"$#,##0")&" per turnover"'
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[27].height = 18

    ws.merge_cells("A28:L28")
    c = ws["A28"]
    c.value = '="Implied per-night cleaning subsidy: "&TEXT(IFERROR((Inputs!C8-Inputs!C9)/Inputs!C11,0),"$#,##0.00")&" / night"'
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[28].height = 18

    # Bold caveat row 30
    ws.merge_cells("A30:L30")
    c = ws["A30"]
    c.value = "Re-run when your cleaner rate or ADR changes."
    c.font = Font(name=FONT_BODY, size=10, italic=True, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[30].height = 20

    # ZONE 3 — Pseudo-button nav (rows 32-34)
    pseudo_button(ws, "A32", "C34", "Inputs",
                  "'Inputs'!A1", variant="primary")
    pseudo_button(ws, "D32", "F34", "Strategy Compare",
                  "'Strategy Compare'!A1", variant="primary")
    if not is_lite:
        pseudo_button(ws, "G32", "I34", "LOS Sensitivity",
                      "'LOS Sensitivity'!A1", variant="primary")
        pseudo_button(ws, "J32", "L34", "Recommendation",
                      "'Recommendation'!A1", variant="accent")
    else:
        pseudo_button(ws, "G32", "L34", "Recommendation",
                      "'Recommendation'!A1", variant="accent")
    for r in range(32, 35):
        ws.row_dimensions[r].height = 22

    # ZONE 4 — Upgrade banner (row 37)
    if is_lite:
        cta = (
            "Want LOS sensitivity + Airbnb search-rank estimate? Upgrade to "
            f"Full at {BRAND_DOMAIN}/cleaning-optimizer-full — $37."
        )
    else:
        cta = (
            "Optimizing pricing across the portfolio? Get the Pricing "
            f"Bundle at {BRAND_DOMAIN}/pricing — $97."
        )
    ws.merge_cells("A37:L37")
    c = ws["A37"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[37].height = 36

    brand_footer(ws, 39,
                 version_line=f"{SKU} · v1.0 · {label_suffix} · Free updates forever")

    ws.print_area = "A1:L41"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_inputs_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    next_tab = "Strategy Compare"
    compact_header_band(ws, "Inputs",
                        prev_tab="Start", next_tab=next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Six numbers and the answer flips. Fill all of these — the "
        "math runs live."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "PROPERTY SCENARIO")

    fields = [
        (7,  "Property name:",
         _val(variant, SAMPLE["property_name"]),       None,        ""),
        (8,  "Cleaning cost paid to cleaner ($/turn):",
         _val(variant, SAMPLE["cleaning_paid"]),       '"$"#,##0',  "What you pay your cleaner per turnover"),
        (9,  "Current cleaning fee charged ($/turn):",
         _val(variant, SAMPLE["current_fee"]),         '"$"#,##0',  "What the guest pays at checkout today"),
        (10, "Current nightly rate (ADR, $):",
         _val(variant, SAMPLE["current_adr"]),         '"$"#,##0',  "Average nightly rate today"),
        (11, "Avg length of stay (nights):",
         _val(variant, SAMPLE["avg_los"]),             "0.0",       "Pull from Airbnb/Hospitable/PriceLabs reports"),
        (12, "Monthly bookings (avg):",
         _val(variant, SAMPLE["monthly_bookings"]),    "0.0",       "Average bookings per month"),
        (13, "Platform commission (%):",
         _val(variant, SAMPLE["platform_pct"]),        "0.0%",      "Airbnb 15%, VRBO 8%, direct 0-3%"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    if not is_lite:
        _section_band(ws, 15, "LOS DISTRIBUTION  (must sum to 100%)")
        los_fields = [
            (16, "% bookings — short (1-2 nights):",
             _val(variant, SAMPLE["los_short_pct"]),   "0.0%",
             "Highest sensitivity to bundling"),
            (17, "% bookings — mid (3-4 nights):",
             _val(variant, SAMPLE["los_mid_pct"]),     "0.0%",
             "Most STRs cluster here"),
            (18, "% bookings — long (5+ nights):",
             _val(variant, SAMPLE["los_long_pct"]),    "0.0%",
             "Least sensitivity to bundling"),
        ]
        for row, label, value, fmt, note in los_fields:
            _input_row(ws, row, label, value, fmt, note)

        # LOS distribution sum check
        ws.cell(row=19, column=2,
                value="LOS distribution sum:").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=19, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=19, column=3, value="=SUM(C16:C18)")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.0%"
        # Color it red if not 100%
        ws.conditional_formatting.add(
            "C19",
            CellIsRule(operator="notEqual", formula=["1"],
                       fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
                       font=Font(name=FONT_BODY, size=11, bold=True,
                                 color=COLOR_ERROR))
        )
        ws.row_dimensions[19].height = 18

    # Footnote on platform commission application
    note_row = 21 if not is_lite else 16
    ws.merge_cells(f"A{note_row}:L{note_row + 2}")
    c = ws[f"A{note_row}"]
    c.value = (
        "Note: Airbnb (and most platforms) take their commission on TOTAL "
        "revenue including the cleaning fee — so passing through cleaning "
        "1:1 still loses you the platform cut. The compare tab accounts "
        "for that."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(note_row, note_row + 3):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, note_row + 4,
                 version_line=f"{SKU} · Inputs")


def _strategy_formulas(strategy_letter, los_ref="Inputs!C11"):
    """Return per-strategy formula strings.

    strategy_letter: 'A' (Full Fee), 'B' (Partial 50%), or 'C' (Bundled)
    los_ref: cell reference for LOS to use (default avg LOS on Inputs)

    Returns dict of formula strings (no leading =).
    """
    fee_paid = "Inputs!C8"
    fee_charged_today = "Inputs!C9"
    adr = "Inputs!C10"
    bookings_per_month = "Inputs!C12"
    plat = "Inputs!C13"

    if strategy_letter == "A":
        new_fee = fee_charged_today
        new_adr = adr
    elif strategy_letter == "B":
        new_fee = f"({fee_charged_today}*0.5)"
        new_adr = f"({adr}+({fee_charged_today}*0.5)/{los_ref})"
    elif strategy_letter == "C":
        new_fee = "0"
        new_adr = f"({adr}+{fee_charged_today}/{los_ref})"
    else:
        raise ValueError(strategy_letter)

    gross_per_booking = f"(({new_adr}*{los_ref})+{new_fee})"
    platform_take = f"({gross_per_booking}*{plat})"
    net_per_booking = f"({gross_per_booking}-{platform_take}-{fee_paid})"
    annual_net = f"({net_per_booking}*{bookings_per_month}*12)"

    # Cleaning ratio: cleaning fee / 3-night-equivalent total stay revenue
    # (Airbnb's documented heuristic uses ~3-night stays)
    three_night_total = f"(({new_adr}*3)+{new_fee})"
    cleaning_ratio = f"IFERROR({new_fee}/{three_night_total},0)"

    return {
        "fee_charged":      new_fee,
        "new_adr":          new_adr,
        "gross_per_booking": gross_per_booking,
        "platform_take":    platform_take,
        "net_per_booking":  net_per_booking,
        "annual_net":       annual_net,
        "cleaning_ratio":   cleaning_ratio,
    }


def build_strategy_compare_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Strategy Compare")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 18), ("D", 18), ("E", 18),
        ("F", 4),
        ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    next_tab = "LOS Sensitivity" if not is_lite else "Recommendation"
    compact_header_band(ws, "Strategy Compare",
                        prev_tab="Inputs", next_tab=next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Three pricing strategies, side by side. Numbers update from your Inputs."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Headers row 6 — strategy names
    ws.cell(row=6, column=2, value="Metric")
    ws.cell(row=6, column=3, value="A — Full Fee")
    ws.cell(row=6, column=4, value="B — Partial Bundle (50%)")
    ws.cell(row=6, column=5, value="C — Fully Bundled")
    for col in range(2, 6):
        cell = ws.cell(row=6, column=col)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 26

    # Strategy descriptions row 7
    ws.cell(row=7, column=2, value="Description").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=7, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    desc_cells = [
        (3, "Charge cleaning fee in full at checkout"),
        (4, "Cut fee in half, raise nightly rate"),
        (5, "Roll cleaning fully into nightly rate"),
    ]
    for col, txt in desc_cells:
        cell = ws.cell(row=7, column=col, value=txt)
        cell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[7].height = 30

    # Get formulas for all 3 strategies
    fA = _strategy_formulas("A")
    fB = _strategy_formulas("B")
    fC = _strategy_formulas("C")

    # Row layout — using avg LOS from Inputs!C11
    rows = [
        # (label, fmt, key_in_formulas_dict, bold)
        (9,  "Cleaning fee charged ($/turn):",  '"$"#,##0',     "fee_charged",     False),
        (10, "Effective nightly rate ($):",      '"$"#,##0',     "new_adr",         False),
        (11, "Gross per booking ($):",           '"$"#,##0',     "gross_per_booking", False),
        (12, "Platform commission ($/booking):", '"$"#,##0',     "platform_take",   False),
        (13, "Net per booking ($):",             '"$"#,##0',     "net_per_booking", False),
        (14, "ANNUAL NET REVENUE:",              '"$"#,##0',     "annual_net",      True),
    ]
    for row, label, fmt, key, bold in rows:
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(
            name=FONT_HEAD if bold else FONT_BODY,
            size=12 if bold else 11,
            bold=True,
            color=COLOR_PRIMARY if bold else COLOR_TEXT,
        )
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        for col, fdict in [(3, fA), (4, fB), (5, fC)]:
            c = ws.cell(row=row, column=col, value=f"={fdict[key]}")
            apply_style(c, formula_cell_style())
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = fmt
            if bold:
                c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
                c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        ws.row_dimensions[row].height = 22 if bold else 18

    # Delta vs Strategy A (current default)
    ws.cell(row=15, column=2, value="Δ vs current (Strategy A):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_SECONDARY
    )
    ws.cell(row=15, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    delta_cells = [
        (3, '="—"'),
        (4, "=D14-C14"),
        (5, "=E14-C14"),
    ]
    for col, val in delta_cells:
        cell = ws.cell(row=15, column=col, value=val)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col != 3:
            cell.number_format = '"+"$#,##0;"-"$#,##0;"$0"'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[15].height = 20

    # Conditional formatting on row 14 — highlight max
    # Use a FormulaRule that lights up the cell with highest annual net
    gold_fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    for col_letter in ["C", "D", "E"]:
        ws.conditional_formatting.add(
            f"{col_letter}14",
            FormulaRule(
                formula=[f'{col_letter}14=MAX($C$14:$E$14)'],
                fill=gold_fill,
                font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT),
            )
        )

    # Cleaning ratio + search-rank impact (Full only)
    if not is_lite:
        _section_band(ws, 17, "AIRBNB SEARCH-RANK IMPACT  (estimated)")

        # Row 18 — cleaning ratio
        ws.cell(row=18, column=2, value="Cleaning fee / 3-night stay ratio:").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=18, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col, fdict in [(3, fA), (4, fB), (5, fC)]:
            c = ws.cell(row=18, column=col, value=f"={fdict['cleaning_ratio']}")
            apply_style(c, formula_cell_style())
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.0%"
        ws.row_dimensions[18].height = 18

        # Row 19 — impact label
        ws.cell(row=19, column=2, value="Impact estimate:").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=19, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col, fdict in [(3, fA), (4, fB), (5, fC)]:
            ratio = fdict["cleaning_ratio"]
            impact_formula = (
                f'=IF({ratio}>0.30,"Penalty (~10-15% loss)",'
                f'IF({ratio}>0.15,"Mild penalty (~5-8% loss)",'
                f'"No penalty"))'
            )
            c = ws.cell(row=19, column=col, value=impact_formula)
            apply_style(c, formula_cell_style())
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)

            # Conditional color formatting based on cleaning ratio
            ws.conditional_formatting.add(
                f"{get_column_letter(col)}19",
                FormulaRule(
                    formula=[f'{get_column_letter(col)}18>0.30'],
                    font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR),
                )
            )
            ws.conditional_formatting.add(
                f"{get_column_letter(col)}19",
                FormulaRule(
                    formula=[f'AND({get_column_letter(col)}18>0.15,{get_column_letter(col)}18<=0.30)'],
                    font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_SECONDARY),
                )
            )
            ws.conditional_formatting.add(
                f"{get_column_letter(col)}19",
                FormulaRule(
                    formula=[f'{get_column_letter(col)}18<=0.15'],
                    font=Font(name=FONT_BODY, size=10, bold=True, color=STATE_GOOD_TEXT),
                )
            )
        ws.row_dimensions[19].height = 18

    # Guest-perception note row
    note_row = 21 if not is_lite else 17
    ws.cell(row=note_row, column=2, value="Guest perception:").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=note_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    perception_cells = [
        (3, "Sticker shock at checkout"),
        (4, "Balanced — both items moderate"),
        (5, "Higher per-night feels expensive"),
    ]
    for col, txt in perception_cells:
        cell = ws.cell(row=note_row, column=col, value=txt)
        cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[note_row].height = 28

    # Disclaimer footnote (Full only — for the search rank claim)
    if not is_lite:
        disclaimer_row = 23
        ws.merge_cells(f"A{disclaimer_row}:L{disclaimer_row + 2}")
        c = ws[f"A{disclaimer_row}"]
        c.value = (
            "Search-rank estimates: Airbnb's algorithm is opaque and changes "
            "constantly. Ranges are based on host-community A/B testing and "
            "Airbnb's published ranking factors as of 2026. Treat as "
            "directional, not exact."
        )
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True, indent=2)
        for r in range(disclaimer_row, disclaimer_row + 3):
            ws.row_dimensions[r].height = 16

    footer_row = 27 if not is_lite else 20
    brand_footer(ws, footer_row,
                 version_line=f"{SKU} · Strategy Compare")

    ws.print_area = f"A1:L{footer_row + 3}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def _strategy_formulas_with_los(strategy_letter, los_value_ref):
    """Wrapper used by LOS Sensitivity tab. los_value_ref is a literal
    cell reference (Inputs!Csomething) — but we feed it a constant
    representative LOS instead.

    We pass the constant directly as a numeric literal in the formula
    by using a hardcoded number. Returns dict of formulas.
    """
    return _strategy_formulas(strategy_letter, los_ref=los_value_ref)


def build_los_sensitivity_tab(wb, variant, is_lite):
    """Full only. 3 cohort tables + a weighted blend summary at top."""
    ws = wb.create_sheet("LOS Sensitivity")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 16), ("D", 16), ("E", 16),
        ("F", 4),
        ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "LOS Sensitivity",
                        prev_tab="Strategy Compare", next_tab="Recommendation")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Same compare, sliced by length-of-stay cohort. Short stays heavily "
        "favor bundling — long stays barely care."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Three cohort tables — short / mid / long
    # Each cohort uses a representative LOS hardcoded into the formula.
    # Annual net for cohort = per-booking-net × monthly_bookings × 12 × cohort_pct
    cohorts = [
        # (start_row, label, los_constant, cohort_pct_ref)
        (6,  "SHORT  (1-2 nights)  ·  representative LOS = 1.5",
         "1.5",  "Inputs!C16"),
        (16, "MID  (3-4 nights)  ·  representative LOS = 3.5",
         "3.5",  "Inputs!C17"),
        (26, "LONG  (5+ nights)  ·  representative LOS = 6.0",
         "6.0",  "Inputs!C18"),
    ]

    cohort_lift_refs = []  # collected for the weighted-blend summary

    for start_row, label, los_const, cohort_pct_ref in cohorts:
        _section_band(ws, start_row, label)

        # Headers
        ws.cell(row=start_row + 1, column=2, value="Metric")
        ws.cell(row=start_row + 1, column=3, value="A — Full Fee")
        ws.cell(row=start_row + 1, column=4, value="B — Partial 50%")
        ws.cell(row=start_row + 1, column=5, value="C — Bundled")
        for col in range(2, 6):
            cell = ws.cell(row=start_row + 1, column=col)
            apply_style(cell, header_row_style())
        ws.row_dimensions[start_row + 1].height = 22

        fA = _strategy_formulas_with_los("A", los_const)
        fB = _strategy_formulas_with_los("B", los_const)
        fC = _strategy_formulas_with_los("C", los_const)

        # Compact metric set: net per booking, full annual (unweighted), weighted annual
        rows_def = [
            (start_row + 2, "Net per booking ($):",
             '"$"#,##0', "net_per_booking"),
            (start_row + 3, "Annual net (cohort, weighted by % bookings):",
             '"$"#,##0', "annual_net_weighted"),
        ]

        # Annual weighted = annual_net * cohort_pct (cohort_pct is share of bookings)
        for fdict in (fA, fB, fC):
            fdict["annual_net_weighted"] = (
                f"({fdict['annual_net']}*{cohort_pct_ref})"
            )

        for row, lbl, fmt, key in rows_def:
            cell = ws.cell(row=row, column=2, value=lbl)
            cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            for col, fdict in [(3, fA), (4, fB), (5, fC)]:
                c = ws.cell(row=row, column=col, value=f"={fdict[key]}")
                apply_style(c, formula_cell_style())
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = fmt
            ws.row_dimensions[row].height = 18

        # Δ vs Strategy A (best lift in cohort)
        delta_row = start_row + 4
        ws.cell(row=delta_row, column=2,
                value="Δ vs Strategy A:").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_SECONDARY
        )
        ws.cell(row=delta_row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        ws.cell(row=delta_row, column=3, value='="—"').font = Font(
            name=FONT_BODY, size=11, color=COLOR_MUTED
        )
        ws.cell(row=delta_row, column=3).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        for col_letter, col in [("D", 4), ("E", 5)]:
            cell = ws.cell(row=delta_row, column=col,
                           value=f"={col_letter}{start_row + 3}-C{start_row + 3}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"+"$#,##0;"-"$#,##0;"$0"'
            cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[delta_row].height = 18

        # Track best lift per cohort (Strat C lift assumed best)
        cohort_lift_refs.append(f"E{start_row + 3}-C{start_row + 3}")

    # WEIGHTED BLEND row 36 onward
    _section_band(ws, 36, "PORTFOLIO-WEIGHTED RESULT")

    blend_label_row = 37
    ws.cell(row=blend_label_row, column=2,
            value="Blended annual lift (Strategy C vs A):").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=blend_label_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    blend_formula = "=" + "+".join(f"({ref})" for ref in cohort_lift_refs)
    cell = ws.cell(row=blend_label_row, column=3, value=blend_formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"+"$#,##0;"-"$#,##0;"$0"'
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[blend_label_row].height = 26

    # Insight callout
    ws.merge_cells(f"A{blend_label_row + 2}:L{blend_label_row + 4}")
    c = ws[f"A{blend_label_row + 2}"]
    c.value = (
        "Why this matters: short stays carry the cleaning fee across fewer "
        "nights, so the per-night cleaning subsidy looks huge to the guest "
        "— and Airbnb's algorithm reads the high cleaning ratio as a "
        "signal to bury you. Long stays absorb the fee. Your LOS mix "
        "decides the right answer."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(blend_label_row + 2, blend_label_row + 5):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, blend_label_row + 6,
                 version_line=f"{SKU} · LOS Sensitivity")

    ws.print_area = f"A1:L{blend_label_row + 9}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_recommendation_tab(wb, variant, is_lite):
    """Final tab — 1-page action plan. Resolves which strategy to recommend."""
    ws = wb.create_sheet("Recommendation")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 32),
        ("C", 24), ("D", 4),
        ("E", 8), ("F", 8), ("G", 8), ("H", 8),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    prev_tab = "LOS Sensitivity" if not is_lite else "Strategy Compare"
    compact_header_band(ws, "Recommendation",
                        prev_tab=prev_tab, next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill

    # ---- Resolved values block — referenced by Start tab ----
    # C5 = recommended strategy name string
    # C6 = annual lift $
    # C7 = visibility lift % (Full only — 0 in Lite)

    # Compare annual nets across strategies via Strategy Compare!C14, D14, E14
    # Choose the max — recommend that strategy.
    # Cell references on Strategy Compare:
    #   C14 = Strategy A annual net
    #   D14 = Strategy B annual net
    #   E14 = Strategy C annual net

    ws.cell(row=5, column=2, value="Recommended strategy:").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=5, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    rec_strategy_formula = (
        '=IF(\'Strategy Compare\'!E14>=MAX(\'Strategy Compare\'!C14:\'Strategy Compare\'!D14),"Strategy C — Fully Bundled",'
        'IF(\'Strategy Compare\'!D14>\'Strategy Compare\'!C14,"Strategy B — Partial Bundle (50%)",'
        '"Strategy A — Full Fee (current)"))'
    )
    cell = ws.cell(row=5, column=3, value=rec_strategy_formula)
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[5].height = 32

    ws.cell(row=6, column=2, value="Annual revenue lift:").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=6, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    # Best annual - current (Strategy A)
    rec_lift_formula = (
        '=MAX(\'Strategy Compare\'!C14:\'Strategy Compare\'!E14)-\'Strategy Compare\'!C14'
    )
    cell = ws.cell(row=6, column=3, value=rec_lift_formula)
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[6].height = 26

    # Visibility lift % (Full only)
    ws.cell(row=7, column=2, value="Search visibility lift (est.):").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    if not is_lite:
        # If recommended strategy reduces cleaning ratio from >30% to <15%,
        # estimate +12% visibility. Mid-range cases get +6%.
        # Strategy Compare!C18 = ratio for Strategy A (current)
        # Strategy Compare!E18 = ratio for Strategy C (bundled)
        vis_formula = (
            '=IF(\'Strategy Compare\'!C18>0.30,'
            'IF(MIN(\'Strategy Compare\'!D18:\'Strategy Compare\'!E18)<0.15,0.12,0.06),'
            'IF(\'Strategy Compare\'!C18>0.15,'
            'IF(MIN(\'Strategy Compare\'!D18:\'Strategy Compare\'!E18)<0.15,0.06,0.03),'
            '0))'
        )
        cell = ws.cell(row=7, column=3, value=vis_formula)
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0%"
    else:
        cell = ws.cell(row=7, column=3, value=0)
        cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0%"
    ws.row_dimensions[7].height = 26

    # ---- 1-page action plan ----
    _section_band(ws, 10, "ACTION PLAN")

    ws.merge_cells("B11:L11")
    c = ws["B11"]
    c.value = (
        "1.  Open your Airbnb listing → Pricing → Cleaning fee."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[11].height = 20

    ws.merge_cells("B12:L12")
    c = ws["B12"]
    c.value = (
        '="2.  Set new cleaning fee → "'
        '&IF(\'Strategy Compare\'!E14>=MAX(\'Strategy Compare\'!C14:\'Strategy Compare\'!D14),'
        'TEXT(\'Strategy Compare\'!E9,"$#,##0"),'
        'IF(\'Strategy Compare\'!D14>\'Strategy Compare\'!C14,'
        'TEXT(\'Strategy Compare\'!D9,"$#,##0"),'
        'TEXT(\'Strategy Compare\'!C9,"$#,##0")))'
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[12].height = 20

    ws.merge_cells("B13:L13")
    c = ws["B13"]
    c.value = (
        '="3.  Set new nightly base rate → "'
        '&IF(\'Strategy Compare\'!E14>=MAX(\'Strategy Compare\'!C14:\'Strategy Compare\'!D14),'
        'TEXT(\'Strategy Compare\'!E10,"$#,##0"),'
        'IF(\'Strategy Compare\'!D14>\'Strategy Compare\'!C14,'
        'TEXT(\'Strategy Compare\'!D10,"$#,##0"),'
        'TEXT(\'Strategy Compare\'!C10,"$#,##0")))'
        '&"  (or update your dynamic-pricing tool base rate)"'
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[13].height = 20

    ws.merge_cells("B14:L14")
    c = ws["B14"]
    c.value = (
        "4.  Mirror the change on VRBO + your direct site if you list there."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[14].height = 20

    ws.merge_cells("B15:L15")
    c = ws["B15"]
    c.value = (
        "5.  Hold the change for 60 days. Track booking pace + "
        "search-impression count (Airbnb Stats → Conversion)."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[15].height = 20

    # RATIONALE section
    _section_band(ws, 17, "RATIONALE")

    ws.merge_cells("B18:L21")
    c = ws["B18"]
    c.value = (
        "Airbnb takes a commission on the total stay cost — including "
        "the cleaning fee — so passing cleaning through 1:1 doesn't "
        "actually break even on the platform side. Bundling cleaning "
        "into the nightly rate compresses the cleaning-fee-to-nightly "
        "ratio, which is a documented Airbnb ranking factor (high ratio "
        "= search penalty). The compare tab shows the dollar version of "
        "that math."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    for r in range(18, 22):
        ws.row_dimensions[r].height = 16

    # RISKS section
    _section_band(ws, 23, "RISKS  +  CAVEATS")

    risks = [
        "Booking pace may shift in the first 30 days — guests filtering by "
        "nightly rate may not click your listing. Search visibility usually "
        "compensates within 60 days.",
        "If your market is rate-sensitive (budget-traveler markets — Branson, "
        "Pigeon Forge in shoulder season), a 20%+ ADR jump may hurt more "
        "than the cleaning-ratio fix helps. Test in shoulder season, not peak.",
        "These estimates assume your cleaner cost is fixed. If your cleaner "
        "raises rates mid-year, re-run.",
        "Airbnb's algorithm changes. The cleaning-ratio rule has held since "
        "2023 but isn't permanent.",
    ]
    for i, txt in enumerate(risks):
        r = 24 + i
        ws.merge_cells(f"B{r}:L{r}")
        c = ws[f"B{r}"]
        c.value = f"•  {txt}"
        c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 28

    footer_row = 30
    brand_footer(ws, footer_row,
                 version_line=f"{SKU} · Recommendation")

    ws.print_area = f"A1:L{footer_row + 3}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant, is_lite):
    wb = Workbook()
    build_start_tab(wb, variant, is_lite)
    build_inputs_tab(wb, variant, is_lite)
    build_strategy_compare_tab(wb, variant, is_lite)
    if not is_lite:
        build_los_sensitivity_tab(wb, variant, is_lite)
    build_recommendation_tab(wb, variant, is_lite)

    wb.properties.title = "Cleaning Fee Optimizer — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Compare cleaning-fee strategies side-by-side. Three pricing "
        "scenarios + Airbnb search-rank impact estimate + LOS sensitivity "
        "+ a one-page action plan."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo", is_lite=False)
    build_workbook(BLANK_OUT, "blank", is_lite=False)
    build_workbook(LITE_OUT, "demo", is_lite=True)


if __name__ == "__main__":
    main()
