"""Build ACQ-009 BRRRR-to-STR Refi Math (v2.2 standard).

Operational-mode tool — models the BRRRR (Buy, Rehab, Rent, Refinance,
Repeat) strategy specifically for STR. Computes ARV, refinance cash-out
at target LTV, cash-left-in (or returned), post-refi DSCR, and whether
the deal repeats the host's capital.

Generates:
  templates/_masters/ACQ-009-brrrr-to-str-refi-DEMO.xlsx
  templates/_masters/ACQ-009-brrrr-to-str-refi-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
)

SKU = "ACQ-009"
NAME = "brrrr-to-str-refi"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (Smokies Ridge BRRRR per brief)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Settings
    "property_name":   "Smokies Ridge BRRRR",
    "target_ltv":      0.75,
    "refi_rate":       0.0725,
    "refi_term_years": 30,

    # Acquisition + Rehab
    "purchase":             385000,
    "closing_costs":        9625,    # ~2.5%
    "hml_loan":             308000,  # 80% of purchase via hard money
    "down_and_closing":     86625,   # purchase - hml + closing
    "rehab_budget":         52000,
    "furniture_budget":     28000,
    "holding_months":       4,
    "holding_per_month":    2000,    # insurance + utilities + taxes + interest

    # ARV + Refi
    "arv_comp_sales":       585000,
    "arv_caprate_noi":      575000,  # NOI / cap rate
    "arv_persqft":          595000,  # $/sqft × sqft
    "arv_selected":         585000,
    "hml_payoff":           308000,  # principal payoff at refi
    "refi_closing_pct":     0.03,    # ~3%

    # Post-Refi Operating
    "y1_revenue":           96000,
    "y1_opex":              33000,   # operating expenses (no debt)
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (BLANK)."""
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, note="", emphasize=False,
                output_col="C"):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(output_col)
    cell = ws.cell(row=row, column=col_idx, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero band rows 1-9
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "BRRRR-to-STR Refi Math"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Did this BRRRR actually return all your cash?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Property name line + KPI block (rows 10-26)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 27):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!C5="","(enter property on Settings tab)",Settings!C5)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # Headline KPI: % capital returned
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "% CAPITAL RETURNED AT REFI"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = "='ARV + Refi'!C25"
    c.font = Font(name=FONT_HEAD, size=42, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 24

    # Verdict line
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = (
        "=IF('ARV + Refi'!C25>=0.95,\"✅  Full BRRRR — capital recycled\","
        "IF('ARV + Refi'!C25>=0.75,\"⚠  Partial BRRRR — some cash trapped\","
        "\"❌  Cash-heavy deal — most equity stays in\"))"
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 28

    # Three side-by-side KPIs row 19-21
    ws.merge_cells("A19:D19")
    c = ws["A19"]
    c.value = "CASH INVESTED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("E19:H19")
    c = ws["E19"]
    c.value = "CASH RETURNED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("I19:L19")
    c = ws["I19"]
    c.value = "POST-REFI DSCR"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[19].height = 18

    ws.merge_cells("A20:D21")
    c = ws["A20"]
    c.value = "='Acquisition + Rehab'!C18"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("E20:H21")
    c = ws["E20"]
    c.value = "='ARV + Refi'!C23"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("I20:L21")
    c = ws["I20"]
    c.value = "='Post-Refi Operating'!C13"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.00"
    ws.row_dimensions[20].height = 26
    ws.row_dimensions[21].height = 18

    # Cash-left-in callout + DSCR threshold row 23
    ws.merge_cells("A23:F23")
    c = ws["A23"]
    c.value = (
        '="Cash left in deal: "&TEXT(\'ARV + Refi\'!C24,"$#,##0")'
    )
    c.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G23:L23")
    c = ws["G23"]
    c.value = (
        "=IF('Post-Refi Operating'!C13>=1.20,"
        "\"✅  Meets DSCR-loan minimum (1.20)\","
        "\"⚠  Below DSCR-loan minimum (1.20)\")"
    )
    c.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24

    # Methodology footnote rows 25-26
    ws.merge_cells("A25:L26")
    c = ws["A25"]
    c.value = (
        "Note: BRRRR math here covers acquisition + rehab + holding cash, ARV "
        "estimate, refi cash-out, and post-refi DSCR/cash flow. It does NOT "
        "model rehab overruns, appraisal misses, or DSCR-loan seasoning windows. "
        "Re-run the moment any input changes — every $5K rehab overrun is real money."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[25].height = 18
    ws.row_dimensions[26].height = 18

    # Pseudo-button nav rows 28-30
    pseudo_button(ws, "A28", "C30", "Acquisition + Rehab",
                  "'Acquisition + Rehab'!A1", variant="primary")
    pseudo_button(ws, "D28", "F30", "ARV + Refi",
                  "'ARV + Refi'!A1", variant="primary")
    pseudo_button(ws, "G28", "I30", "Post-Refi Operating",
                  "'Post-Refi Operating'!A1", variant="primary")
    pseudo_button(ws, "J28", "L30", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(28, 31):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "\U0001F4A1  Stacking BRRRRs? Get the Acquisition Bundle at "
        f"{BRAND_DOMAIN} — Deal Analyzer + Cost-to-Launch + BRRRR Math + more."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_acquisition_rehab_tab(wb, variant):
    ws = wb.create_sheet("Acquisition + Rehab")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Acquisition + Rehab",
                        prev_tab="Start", next_tab="ARV + Refi")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 1 of 3 — every dollar that goes in before the refi"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "PURCHASE")
    purchase_fields = [
        (7, "Purchase price ($):",
            _val(variant, SAMPLE["purchase"]), '"$"#,##0',
            "Distressed/below-market is the BRRRR target"),
        (8, "Closing costs at purchase ($):",
            _val(variant, SAMPLE["closing_costs"]), '"$"#,##0',
            "Title, recording, lender fees — typically 2-3%"),
        (9, "Hard money / acquisition loan ($):",
            _val(variant, SAMPLE["hml_loan"]), '"$"#,##0',
            "0 if all-cash. HML typically 80% of purchase + 100% rehab"),
        (10, "Down + closing in cash ($):",
            _val(variant, SAMPLE["down_and_closing"]), '"$"#,##0',
            "Purchase − HML loan + closing costs paid out of pocket"),
    ]
    for row, label, value, fmt, note in purchase_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 12, "REHAB + HOLDING")
    rehab_fields = [
        (13, "Rehab budget ($):",
            _val(variant, SAMPLE["rehab_budget"]), '"$"#,##0',
            "Hard costs — labor, materials, permits"),
        (14, "Furniture budget ($):",
            _val(variant, SAMPLE["furniture_budget"]), '"$"#,##0',
            "Often deferred to post-refi — toggle to $0 if so"),
        (15, "Holding period (months):",
            _val(variant, SAMPLE["holding_months"]), "0",
            "Months from close to refi cash-out"),
        (16, "Holding cost per month ($):",
            _val(variant, SAMPLE["holding_per_month"]), '"$"#,##0',
            "Insurance + utilities + tax + HML interest"),
    ]
    for row, label, value, fmt, note in rehab_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 17, "TOTAL CASH INVESTED  —  CALCULATED")

    # Total cash in = down/closing + rehab + furniture + holding (months × monthly)
    _output_row(ws, 18, "Total cash in (pre-refi):",
                "=C10+C13+C14+(C15*C16)", '"$"#,##0',
                "Sum of all out-of-pocket dollars before refi",
                emphasize=True)

    # Helpful breakdown
    _output_row(ws, 20, "Total holding costs:",
                "=C15*C16", '"$"#,##0',
                "Months × monthly carry")
    _output_row(ws, 21, "Hard money outstanding (refi payoff):",
                "=C9", '"$"#,##0',
                "Pulled to ARV + Refi tab as loan payoff")

    brand_footer(ws, 24,
                 version_line=f"{SKU} · Acquisition + Rehab")


def build_arv_refi_tab(wb, variant):
    ws = wb.create_sheet("ARV + Refi")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "ARV + Refi",
                        prev_tab="Acquisition + Rehab",
                        next_tab="Post-Refi Operating")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 2 of 3 — what it appraises for and how much cash comes back"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "ARV ESTIMATE  —  THREE METHODS")

    arv_fields = [
        (7, "ARV — Comp sales method ($):",
            _val(variant, SAMPLE["arv_comp_sales"]), '"$"#,##0',
            "3 nearby sold comps × adjusted $/sqft"),
        (8, "ARV — Cap rate × NOI method ($):",
            _val(variant, SAMPLE["arv_caprate_noi"]), '"$"#,##0',
            "Year-1 NOI ÷ market cap rate"),
        (9, "ARV — $/sq ft method ($):",
            _val(variant, SAMPLE["arv_persqft"]), '"$"#,##0',
            "Renovated comp $/sqft × your sqft"),
    ]
    for row, label, value, fmt, note in arv_fields:
        _input_row(ws, row, label, value, fmt, note)

    _output_row(ws, 10, "Average of three methods:",
                "=AVERAGE(C7:C9)", '"$"#,##0',
                "Reality check on your selected ARV below")

    _input_row(ws, 12, "Selected ARV ($):",
               _val(variant, SAMPLE["arv_selected"]), '"$"#,##0',
               "What you're underwriting — usually conservative")

    _section_band(ws, 14, "REFINANCE")

    # Target LTV + refi rate pulled from Settings
    _output_row(ws, 15, "Target LTV (from Settings):",
                "=Settings!C7", "0.0%",
                "DSCR loans typically 70-75%")
    _output_row(ws, 16, "Refi rate (from Settings):",
                "=Settings!C9", "0.000%",
                "Current DSCR-loan rate")
    _output_row(ws, 17, "Refi term in years (from Settings):",
                "=Settings!C11", "0",
                "30 typical")

    _output_row(ws, 18, "Max refi loan amount:",
                "=C12*C15", '"$"#,##0',
                "Selected ARV × target LTV")

    _input_row(ws, 19, "Less: original loan payoff ($):",
               _val(variant, SAMPLE["hml_payoff"]), '"$"#,##0',
               "HML principal paid off at refi (0 if all-cash purchase)")

    _input_row(ws, 20, "Refi closing cost (% of refi loan):",
               _val(variant, SAMPLE["refi_closing_pct"]), "0.0%",
               "DSCR-loan closing typically 2-4%")

    _output_row(ws, 21, "Refi closing cost ($):",
                "=C18*C20", '"$"#,##0',
                "Refi loan × refi closing %")

    _output_row(ws, 22, "Cash out at refi (gross):",
                "=C18-C19", '"$"#,##0',
                "Refi proceeds after HML payoff (before closing)")

    _output_row(ws, 23, "Net cash returned after refi:",
                "='Acquisition + Rehab'!C18-MAX(0,'Acquisition + Rehab'!C18-(C18-C19-C21))",
                '"$"#,##0',
                "Cash returned to host — capped at total cash in",
                emphasize=True)

    _output_row(ws, 24, "Cash left in deal:",
                "=MAX(0,'Acquisition + Rehab'!C18-(C18-C19-C21))",
                '"$"#,##0',
                "Cash in − cash returned (never negative)",
                emphasize=True)

    _output_row(ws, 25, "% capital returned:",
                "=IFERROR(C23/'Acquisition + Rehab'!C18,0)", "0.0%",
                "Cash returned ÷ cash in (>=95% = full BRRRR)",
                emphasize=True)

    brand_footer(ws, 28,
                 version_line=f"{SKU} · ARV + Refi")


def build_post_refi_operating_tab(wb, variant):
    ws = wb.create_sheet("Post-Refi Operating")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Post-Refi Operating",
                        prev_tab="ARV + Refi", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 3 of 3 — does it cash flow with the new debt service?"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "OPERATING — YEAR 1 STABILIZED")

    op_fields = [
        (7, "Year-1 STR projected revenue ($):",
            _val(variant, SAMPLE["y1_revenue"]), '"$"#,##0',
            "Net of platform commission"),
        (8, "Year-1 operating expenses ($):",
            _val(variant, SAMPLE["y1_opex"]), '"$"#,##0',
            "Excludes debt service (added below)"),
    ]
    for row, label, value, fmt, note in op_fields:
        _input_row(ws, row, label, value, fmt, note)

    _output_row(ws, 9, "NOI (Net Operating Income):",
                "=C7-C8", '"$"#,##0',
                "Revenue − operating expenses")

    _section_band(ws, 11, "NEW DEBT SERVICE  —  CALCULATED")

    # Excel PMT for monthly P&I — use the refinanced loan + rate + term from Settings
    _output_row(ws, 12, "Monthly P&I (post-refi):",
                "=IFERROR(PMT(Settings!C9/12,Settings!C11*12,-'ARV + Refi'!C18),0)",
                '"$"#,##0.00',
                "Excel PMT on refi loan, rate, and term")

    _output_row(ws, 13, "Annual debt service:",
                "=C12*12", '"$"#,##0',
                "Used for DSCR + cash flow below",
                emphasize=False)

    # Reserve C13 emphasis for DSCR — tweak: add DSCR row at row 13 instead
    # (re-purpose: keep annual debt service at row 14 to free row 13 for DSCR)
    # Actually we already wrote row 13 as annual debt service. Adjust: shift.
    # Simpler — overwrite: place DSCR on a new row 14, cash flow on row 15, etc.

    _section_band(ws, 15, "OUTPUTS")

    _output_row(ws, 16, "Cash flow (NOI − debt service):",
                "=C9-C13", '"$"#,##0',
                "Annual cash flow after debt", emphasize=True)

    _output_row(ws, 17, "DSCR (NOI ÷ debt service):",
                "=IFERROR(C9/C13,0)", "0.00",
                "DSCR loans require ≥ 1.20", emphasize=True)

    _output_row(ws, 18, "Cash-on-cash on cash-left-in:",
                "=IFERROR(C16/'ARV + Refi'!C24,0)", "0.0%",
                "Annual cash flow ÷ cash left in deal",
                emphasize=True)

    # NOTE: Start tab references 'Post-Refi Operating'!C13 for DSCR — but we
    # placed DSCR at row 17. Add a clean alias at C13 so Start formulas remain
    # intuitive: write a small DSCR mirror at the band. Simpler — fix Start
    # tab formula to point at C17. We already set Start to C13; adjust Start.
    # (Start already set to Post-Refi Operating!C13. That cell is annual debt
    # service. We need DSCR there for the KPI. Fix: move DSCR to C13.)
    # We'll resolve by writing a clean DSCR mirror at row 13's C cell — but
    # we already wrote annual debt service there. Approach: clear and rewrite
    # this tab's structure so DSCR sits at C13.

    # Conditional formatting: red if DSCR < 1.20 else gold
    ws.conditional_formatting.add(
        "C17",
        CellIsRule(operator="lessThan", formula=["1.2"],
                   fill=PatternFill("solid", fgColor="FFCCCC"),
                   font=Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ERROR)),
    )

    # Methodology callout
    ws.merge_cells("A20:L22")
    c = ws["A20"]
    c.value = (
        "DSCR 1.20 is the standard floor for non-QM DSCR loans. Below that, "
        "your refi may not approve at the target LTV — lenders will haircut "
        "the loan amount or push you to a lower-LTV product. If your DSCR "
        "is borderline, model a 70% LTV scenario in Settings."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(20, 23):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 24,
                 version_line=f"{SKU} · Post-Refi Operating")


# ---- Re-do the Post-Refi Operating tab with DSCR at C13 (so Start KPI works) ----
# The Start tab references 'Post-Refi Operating'!C13 for DSCR. Provide a
# dedicated builder that lays the rows out so DSCR lands at row 13 cleanly.
def build_post_refi_operating_tab(wb, variant):  # noqa: F811 — intentional override
    ws = wb.create_sheet("Post-Refi Operating")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Post-Refi Operating",
                        prev_tab="ARV + Refi", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 3 of 3 — does it cash flow with the new debt service?"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "OPERATING — YEAR 1 STABILIZED")

    op_fields = [
        (7, "Year-1 STR projected revenue ($):",
            _val(variant, SAMPLE["y1_revenue"]), '"$"#,##0',
            "Net of platform commission"),
        (8, "Year-1 operating expenses ($):",
            _val(variant, SAMPLE["y1_opex"]), '"$"#,##0',
            "Excludes debt service (added below)"),
    ]
    for row, label, value, fmt, note in op_fields:
        _input_row(ws, row, label, value, fmt, note)

    _output_row(ws, 9, "NOI (Net Operating Income):",
                "=C7-C8", '"$"#,##0',
                "Revenue − operating expenses")

    _section_band(ws, 11, "NEW DEBT SERVICE  +  DSCR")

    # Excel PMT for monthly P&I (per the brand contract Critical rule #4)
    _output_row(ws, 12, "Annual debt service (Excel PMT × 12):",
                "=IFERROR(PMT(Settings!C9/12,Settings!C11*12,-'ARV + Refi'!C18),0)*12",
                '"$"#,##0',
                "PMT(rate/12, term×12, -refi_loan) × 12")

    # DSCR at row 13 — the cell Start tab references for the headline
    _output_row(ws, 13, "DSCR (NOI ÷ debt service):",
                "=IFERROR(C9/C12,0)", "0.00",
                "DSCR loans require ≥ 1.20",
                emphasize=True)

    _output_row(ws, 14, "Monthly P&I (post-refi):",
                "=IFERROR(PMT(Settings!C9/12,Settings!C11*12,-'ARV + Refi'!C18),0)",
                '"$"#,##0.00',
                "Excel PMT — handy reference")

    _section_band(ws, 16, "OUTPUTS")

    _output_row(ws, 17, "Cash flow (NOI − debt service):",
                "=C9-C12", '"$"#,##0',
                "Annual cash flow after debt", emphasize=True)

    _output_row(ws, 18, "Cash-on-cash on cash-left-in:",
                "=IFERROR(C17/'ARV + Refi'!C24,0)", "0.0%",
                "Annual cash flow ÷ cash left in",
                emphasize=True)

    # Conditional formatting — red DSCR < 1.20
    ws.conditional_formatting.add(
        "C13",
        CellIsRule(operator="lessThan", formula=["1.2"],
                   fill=PatternFill("solid", fgColor="FFCCCC"),
                   font=Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ERROR)),
    )

    # Methodology callout
    ws.merge_cells("A20:L22")
    c = ws["A20"]
    c.value = (
        "DSCR 1.20 is the standard floor for non-QM DSCR loans. Below that, "
        "your refi may not approve at the target LTV — lenders will haircut "
        "the loan amount or push you to a lower-LTV product. If DSCR is "
        "borderline, model a 70% LTV scenario in Settings."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(20, 23):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 24,
                 version_line=f"{SKU} · Post-Refi Operating")


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Post-Refi Operating", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Property name + refi assumptions used across the workbook"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _input_row(ws, 5, "Property name:",
               _val(variant, SAMPLE["property_name"]), None,
               "Shown on Start tab headline")
    _input_row(ws, 7, "Target LTV (refi):",
               _val(variant, SAMPLE["target_ltv"]), "0.0%",
               "75% typical for DSCR; 70% conservative")
    _input_row(ws, 9, "Refi rate (annual):",
               _val(variant, SAMPLE["refi_rate"]), "0.000%",
               "Current DSCR-loan quote — verify with broker")
    _input_row(ws, 11, "Refi term (years):",
               _val(variant, SAMPLE["refi_term_years"]), "0",
               "30 typical; 25 sometimes for DSCR")

    # Methodology footnote
    ws.merge_cells("A14:L16")
    c = ws["A14"]
    c.value = (
        "These four settings drive Max-refi-loan, post-refi P&I, and DSCR. "
        "Re-quote your DSCR rate every 30 days while sourcing — STR DSCR "
        "rates have moved 100-200bp in single quarters since 2023. The "
        "ARV + Refi tab uses the LTV here. The Post-Refi Operating tab "
        "uses the rate + term here in the Excel PMT formula."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(14, 17):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 19,
                 version_line=f"{SKU} · Settings")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_acquisition_rehab_tab(wb, variant)
    build_arv_refi_tab(wb, variant)
    build_post_refi_operating_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "BRRRR-to-STR Refi Math — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "BRRRR strategy calculator for STR — ARV, refi cash-out, "
        "cash-left-in, post-refi DSCR, and capital-returned KPI."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
