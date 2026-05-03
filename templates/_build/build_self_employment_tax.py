"""Build TAX-008 Self-Employment Tax Calculator — wizard tool.

Implements templates/_briefs/TAX-008-self-employment-tax.md.

Generates BOTH:
  templates/_masters/TAX-008-self-employment-tax-DEMO.xlsx
  templates/_masters/TAX-008-self-employment-tax-BLANK.xlsx

Tabs:
  0  Start             — wizard hero + 3-card + Quick Start + Get Started + progress
  1  Net SE Income     — Sched C/F/K-1 inputs (Section 1 of 3)
  2  Wage Base & W-2   — W-2 wages this year (Section 2 of 3)
  3  Adjustments       — filing status, Additional Medicare threshold (Section 3 of 3)
  4  SE Tax Calc       — computed step-by-step (no inputs)
  5  Schedule SE Map   — IRS form mirror
  6  Launch            — total SE tax card + print packet
  7  Settings          — tax year · wage base · brackets

Usage:
    python build_self_employment_tax.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    COLOR_WHITE,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-008"
NAME = "self-employment-tax"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
VERSION_LINE = f"{SKU} · v2.2 · Free updates forever"

TOTAL_SECTIONS = 3
SECTION_INPUT_COUNTS = {
    "Net SE Income":    3,
    "Wage Base & W-2":  2,
    "Adjustments":      1,
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 6

# 2026 constants
SS_WAGE_BASE_2026 = 176100  # IRS announced 2026 wage base
SS_RATE = 0.124
MEDICARE_RATE = 0.029
ADDL_MEDICARE_RATE = 0.009
NESE_FACTOR = 0.9235  # 100% - 7.65% (employer-equivalent)

SAMPLE = {
    "Net SE Income": {
        "sched_c":     43000,
        "sched_f":     0,
        "k1":          0,
    },
    "Wage Base & W-2": {
        "your_w2":     0,
        "spouse_w2":   88000,
    },
    "Adjustments": {
        "filing_status": "MFJ",
    },
}


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def build_input_tab(wb, tab_name, section_num, title, subtitle, fields,
                     prev_tab, next_tab):
    """Generic flat input tab. Returns (ws, last_input_row)."""
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    section_header_band(ws, section_num, TOTAL_SECTIONS, title, subtitle,
                         prev_tab, next_tab)

    set_col_widths(ws, [
        ("A", 44),
        ("B", 18), ("C", 8), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "Fill the highlighted fields below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    card_header(ws, 7, ("A", "L"), title)

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    for i, (label, value, kind, options) in enumerate(fields):
        r = 8 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=value)
        apply_style(c, input_cell_style())
        if kind == "money":
            c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        elif kind == "pct":
            c.number_format = "0.0%"
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        elif kind == "text":
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if options:
            add_dropdown(ws, f"B{r}", options)
        ws.row_dimensions[r].height = 22

    last_input_row = 7 + len(fields)
    card_body_fill(ws, 8, last_input_row, ("A", "L"), border=True)
    return ws, last_input_row


def append_callout(ws, row, text):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 36


def append_footer_nav(ws, row, prev_tab, next_tab, default_next="SE Tax Calc"):
    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = f"Next: {default_next} →"
        next_target = f"'{default_next}'!A1"
    pseudo_button(ws, f"A{row}", f"F{row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{row}", f"L{row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 22


# --- Tab builders ---

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "SE Tax Calculator"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Compute your 15.3% SE tax. Know what you owe."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — single declarative answer (suite Theme 1).
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(IFERROR(\'SE Tax Calc\'!B17,0)>0,'
        '"⚠  SE tax due = "&TEXT(\'SE Tax Calc\'!B17,"$#,##0")'
        '&"  ·  ½ deductible = "&TEXT(\'SE Tax Calc\'!B18,"$#,##0"),'
        '"\U0001F4CA  Enter your Net SE Income to see your SE tax.")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A11:L11")
    c = ws["A11"]; c.value = "What you'll get"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("💰 SE TAX $", "Your total Schedule SE Line 12 with SS + Medicare + Addl Medicare."),
        ("📉 ½ ADJUSTMENT", "Schedule 1 Line 15 — 50% of SE tax deductible above the line."),
        ("📄 SCHEDULE SE", "Print-ready Schedule SE Part I + Part II mirror for your CPA."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (ttl, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]; c.value = ttl
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]; c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A23:L23")
    c = ws["A23"]; c.value = "Quick Start — be done in 5 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24
    quickstart_items = [
        "① Net SE income (Schedule C / F / K-1)",
        "② W-2 wages this year (yours + spouse if MFJ)",
        "③ Filing status (drives Additional Medicare threshold)",
        "④ Review SE Tax Calc — auto-computed",
        "⑤ Print Schedule SE Map for your CPA",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + (i if i < 3 else i - 3)
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    pseudo_button(ws, "A30", "L33",
                   "GET STARTED — ENTER NET SE INCOME  →",
                   "'Net SE Income'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # Progress dashboard
    ws.merge_cells("A36:F36")
    c = ws["A36"]; c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ranges = [
        "'Net SE Income'!B8:B10",      # 3
        "'Wage Base & W-2'!B8:B9",     # 2
        "'Adjustments'!B8",            # 1
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    ws.merge_cells("G36:L36")
    c = ws["G36"]
    c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS},"0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    section_rows = [
        ("① Net SE Income",    "Net SE Income",    "B8:B10",  3),
        ("② Wage Base & W-2",  "Wage Base & W-2",  "B8:B9",   2),
        ("③ Adjustments",      "Adjustments",      "B8",      1),
    ]
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]; c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (f'=IF({ca}={total},"✅ Done",'
                   f'IF({ca}=0,"⏳ Empty","⏳ "&{ca}&" of {total}"))')
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    brand_footer(ws, 45, version_line=VERSION_LINE)

    ws.print_area = "A1:L47"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_net_se_tab(wb, variant):
    s = SAMPLE["Net SE Income"] if variant == "demo" else {}
    fields = [
        ("Schedule C net profit ($):",            _val(variant, s.get("sched_c")), "money", None),
        ("Schedule F net farm profit ($):",       _val(variant, s.get("sched_f")), "money", None),
        ("Partnership SE income (K-1 Box 14A):",  _val(variant, s.get("k1")), "money", None),
    ]
    ws, last = build_input_tab(
        wb, "Net SE Income", 1,
        "Net SE Income",
        "Schedule C / F / K-1 — IRS Schedule SE Lines 1a, 1b, 2.",
        fields, prev_tab="", next_tab="Wage Base & W-2",
    )

    # Computed total at last+2
    total_row = last + 2
    a = ws.cell(row=total_row, column=1, value="Total net SE income (Line 3):")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=total_row, column=2, value="=SUM(B8:B10)")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[total_row].height = 24

    append_callout(ws, total_row + 2, (
        "ⓘ De minimis exception (IRC §1402(a)): if total Net SE Income is < $400, "
        "no SE tax owed. Schedule E rental income is NOT subject to SE tax — only "
        "Schedule C (active business) and Schedule F (farm) income."
    ))
    append_footer_nav(ws, total_row + 4,
                       prev_tab="", next_tab="Wage Base & W-2")


def build_wage_base_tab(wb, variant):
    s = SAMPLE["Wage Base & W-2"] if variant == "demo" else {}
    fields = [
        ("Your W-2 wages this year ($):",       _val(variant, s.get("your_w2")), "money", None),
        ("Spouse W-2 wages this year ($, MFJ):", _val(variant, s.get("spouse_w2")), "money", None),
    ]
    ws, last = build_input_tab(
        wb, "Wage Base & W-2", 2,
        "Wage Base & W-2",
        "W-2 wages count toward the SS wage base — Schedule SE Line 8a.",
        fields, prev_tab="Net SE Income", next_tab="Adjustments",
    )

    # Reference: SS wage base
    ref_row = last + 2
    a = ws.cell(row=ref_row, column=1, value=f"Social Security wage base (Settings):")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=ref_row, column=2, value="=Settings!B6")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    ws.row_dimensions[ref_row].height = 18

    append_callout(ws, ref_row + 2, (
        "ⓘ The Social Security portion of SE tax (12.4%) only applies to income up "
        "to the wage base. Your W-2 wages this year fill that base first; SE income "
        "above what's left of the base owes only the 2.9% Medicare portion. If you've "
        "already maxed your W-2 above the wage base, NO SS portion of SE tax applies."
    ))
    append_footer_nav(ws, ref_row + 4,
                       prev_tab="Net SE Income", next_tab="Adjustments")


def build_adjustments_tab(wb, variant):
    s = SAMPLE["Adjustments"] if variant == "demo" else {}
    fields = [
        ("Filing status:", _val(variant, s.get("filing_status")), "text",
         ["Single", "MFJ", "HoH", "MFS"]),
    ]
    ws, last = build_input_tab(
        wb, "Adjustments", 3,
        "Adjustments",
        "Filing status — drives Additional Medicare 0.9% threshold.",
        fields, prev_tab="Wage Base & W-2", next_tab="SE Tax Calc",
    )

    # Computed threshold display at last+2
    thr_row = last + 2
    a = ws.cell(row=thr_row, column=1, value="Additional Medicare 0.9% threshold:")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=thr_row, column=2)
    c.value = (
        '=IF(B8="MFJ",250000,IF(B8="MFS",125000,200000))'
    )
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    ws.row_dimensions[thr_row].height = 18

    append_callout(ws, thr_row + 2, (
        "ⓘ Additional Medicare tax (IRC §1401(b)(2)): 0.9% applies to Medicare wages + "
        "SE income above $200K (Single/HoH), $250K (MFJ), or $125K (MFS). Combined "
        "with the regular 2.9% Medicare, this stacks to 3.8% on income above the threshold."
    ))
    append_footer_nav(ws, thr_row + 4,
                       prev_tab="Wage Base & W-2", next_tab="SE Tax Calc")


def build_se_tax_calc_tab(wb):
    """Tab 4 — SE Tax Calc. Computed step-by-step, no inputs."""
    ws = wb.create_sheet("SE Tax Calc")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "SE Tax Calc",
                         prev_tab="Adjustments", next_tab="Schedule SE Map")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Step-by-step SE tax computation — Schedule SE Lines 3-12."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 50), ("B", 18),
        ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Net SE income source
    rows = [
        (5, "Net SE income (from Net SE Income tab):",
         "='Net SE Income'!B12", "money"),
        (6, "× 92.35% (employer-equivalent factor):",
         f"=B5*{NESE_FACTOR}", "money"),
        (7, "Net Earnings from SE (NESE) — Line 4a:",
         "=B6", "money"),
        (8, "Your W-2 wages this year:",
         "='Wage Base & W-2'!B8", "money"),
        (9, "Social Security wage base (Settings):",
         "=Settings!B6", "money"),
        (10, "Wage-base headroom (= base − W-2):",
         "=MAX(B9-B8,0)", "money"),
        (11, "SS-portion subject to 12.4%:",
         "=MIN(B7,B10)", "money"),
        (12, "Social Security tax (Line 10):",
         f"=B11*{SS_RATE}", "money"),
        (13, "Medicare tax (Line 11) — 2.9% × NESE:",
         f"=B7*{MEDICARE_RATE}", "money"),
        (14, "Combined NESE + W-2 (for Addl Medicare check):",
         "=B7+B8+'Wage Base & W-2'!B9", "money"),
        (15, "Additional Medicare threshold (filing-status):",
         '=IF(\'Adjustments\'!B8="MFJ",250000,IF(\'Adjustments\'!B8="MFS",125000,200000))',
         "money"),
        (16, "Additional Medicare tax (0.9% × excess):",
         f"=MAX(B14-B15,0)*{ADDL_MEDICARE_RATE}", "money"),
        (17, "TOTAL SE TAX (Line 12) — to Sched 2 Line 4:",
         "=B12+B13+B16", "total"),
        (18, "½ SE tax adjustment (Sched 1 Line 15):",
         "=B17/2", "money"),
    ]

    for r, label, formula, kind in rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0;[Red]("$"#,##0)'
        if kind == "total":
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 18

    # Callout at row 20
    ws.merge_cells("A20:L20")
    c = ws["A20"]
    c.value = (
        "ⓘ Half-SE-tax adjustment is deductible above the line on Schedule 1 Line 15 — "
        "it reduces AGI directly. The other half is what you actually pay. "
        "If you have a spouse with separate SE income (MFJ), each spouse computes "
        "their own SE tax separately — they DO NOT combine."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[20].height = 36

    pseudo_button(ws, "A22", "F23", "← Back: Adjustments",
                   "'Adjustments'!A5", variant="secondary")
    pseudo_button(ws, "G22", "L23", "Next: Schedule SE Map →",
                   "'Schedule SE Map'!A1", variant="secondary")
    ws.row_dimensions[22].height = 22
    ws.row_dimensions[23].height = 22

    ws.print_area = "A1:L24"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_schedule_se_map_tab(wb):
    ws = wb.create_sheet("Schedule SE Map")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Schedule SE Map",
                         prev_tab="SE Tax Calc", next_tab="Launch")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Print this — IRS Schedule SE Part I + Part II mirror."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6), ("B", 50), ("C", 18),
        ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6),
        ("J", 6), ("K", 6), ("L", 6),
    ])

    headers = ["Line", "Description", "Amount"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        apply_style(c, header_row_style())
    ws.row_dimensions[5].height = 20

    lines = [
        ("PART I", "PART I — SELF-EMPLOYMENT TAX", None, "header"),
        ("1a",     "Net farm profit",                          "='Net SE Income'!B9", "money"),
        ("2",      "Net Schedule C profit",                    "='Net SE Income'!B8", "money"),
        ("",       "Partnership SE (K-1 Box 14A)",             "='Net SE Income'!B10", "money"),
        ("3",      "Combined (sum)",                           "=C7+C8+C9", "money"),
        ("4a",     "Multiply Line 3 by 92.35%",                f"=C10*{NESE_FACTOR}", "money"),
        ("8a",     "Total SS wages this year (W-2)",           "='Wage Base & W-2'!B8", "money"),
        ("8d",     "Subtotal (W-2 entered for wage base):",    "=C12", "money"),
        ("9",      "Subtract: SS wage base − W-2:",            "=MAX(Settings!B6-C12,0)", "money"),
        ("10",     "SS portion (smaller of 4a or 9 × 12.4%):", f"=MIN(C11,C14)*{SS_RATE}", "money"),
        ("11",     "Medicare portion (Line 4a × 2.9%):",       f"=C11*{MEDICARE_RATE}", "money"),
        ("12",     "Total SE tax (Line 10 + 11)",              "=C15+C16", "total"),
        ("13",     "½ SE tax adjustment",                      "=C17/2", "money"),
        ("PART II", "PART II — OPTIONAL METHODS",              None, "header"),
        ("",       "(Not commonly used — see Sched SE for farm/non-farm methods)",
                   '""', "text"),
    ]

    r = 6
    for line_num, label, formula, kind in lines:
        if kind == "header":
            ws.merge_cells(f"A{r}:C{r}")
            c = ws.cell(row=r, column=1, value=f"  {line_num} — {label}")
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
            c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 22
        else:
            ws.cell(row=r, column=1, value=line_num).font = Font(
                name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
            )
            ws.cell(row=r, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=r, column=2, value=label).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )
            ws.cell(row=r, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            c = ws.cell(row=r, column=3, value=formula)
            apply_style(c, formula_cell_style())
            if kind in ("money", "total"):
                c.number_format = '"$"#,##0;[Red]("$"#,##0)'
            if kind == "total":
                c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                ws.cell(row=r, column=2).font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            ws.row_dimensions[r].height = 18
        r += 1

    ws.print_area = f"A1:C{r}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_launch_tab(wb, variant):
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK", "'Schedule SE Map'!A1",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]; c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Your SE tax for the year"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Schedule 2 Line 4 + Schedule 1 Line 15."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 14):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 — Total SE tax
    ws.merge_cells("A9:D9")
    c = ws["A9"]; c.value = "TOTAL SE TAX"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]; c.value = "='SE Tax Calc'!B17"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A12:D12")
    c = ws["A12"]; c.value = "to Schedule 2 Line 4"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 — Half adjustment
    ws.merge_cells("E9:H9")
    c = ws["E9"]; c.value = "½ ADJUSTMENT"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]; c.value = "='SE Tax Calc'!B18"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E12:H12")
    c = ws["E12"]; c.value = "to Schedule 1 Line 15"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 — NESE
    ws.merge_cells("I9:L9")
    c = ws["I9"]; c.value = "NESE (Line 4a)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]; c.value = "='SE Tax Calc'!B7"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I12:L12")
    c = ws["I12"]; c.value = "net earnings from SE"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    pseudo_button(ws, "A16", "L20",
                   "📄  PRINT SCHEDULE SE MAP  →",
                   "'Schedule SE Map'!A1", variant="primary")
    for r in range(16, 21):
        ws.row_dimensions[r].height = 24

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = (
        "💡 Upgrade to the Tax Season Bundle ($147)."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[23].height = 36

    brand_footer(ws, 25, version_line=VERSION_LINE)

    ws.print_area = "A1:L27"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Settings", prev_tab="Launch", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · SS wage base · rates."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 38), ("B", 18),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    fields = [
        (5,  "Active tax year:",                          2026,                    "0"),
        (6,  "SS wage base ($):",                         SS_WAGE_BASE_2026,       '"$"#,##0'),
        (7,  "SS rate (12.4%):",                          SS_RATE,                 "0.0%"),
        (8,  "Medicare rate (2.9%):",                     MEDICARE_RATE,           "0.0%"),
        (9,  "Additional Medicare rate (0.9%):",          ADDL_MEDICARE_RATE,      "0.00%"),
        (10, "NESE factor (92.35%):",                     NESE_FACTOR,             "0.00%"),
    ]
    for r, label, val, fmt in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Historical wage bases (rows 12-16)
    ws.row_dimensions[11].height = 8
    a = ws.cell(row=12, column=1, value="Historical wage bases:")
    a.font = italic_muted
    historical = [(13, "2023:", 160200), (14, "2024:", 168600), (15, "2025:", 176100)]
    for row, label, val in historical:
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        rc = ws.cell(row=row, column=2, value=val)
        rc.number_format = '"$"#,##0'
        rc.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)

    # Year-end Archive
    ws.row_dimensions[16].height = 8
    a = ws.cell(row=17, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[17].height = 22
    archive_headers = ["Year", "NESE", "SE Tax", "½ Adjustment"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=18, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[18].height = 18
    for idx, year in enumerate(range(2024, 2031), start=19):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_net_se_tab(wb, variant)
    build_wage_base_tab(wb, variant)
    build_adjustments_tab(wb, variant)
    build_se_tax_calc_tab(wb)
    build_schedule_se_map_tab(wb)
    build_launch_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Self-Employment Tax Calculator — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "IRS Schedule SE calculator (15.3% SE tax) for STR hosts on Schedule C."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
