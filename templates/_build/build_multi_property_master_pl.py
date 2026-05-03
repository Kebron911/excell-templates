"""Build TAX-011 Multi-Property Master P&L (v2.2 standard).

Operational-mode tool — replaces N copies of TAX-002 with a single
workbook serving up to 20 properties via a Property dropdown. Rolls
up by Entity/LLC, supports allocated shared expenses (manual split
via Settings helper), and outputs a multi-column Schedule E ready
for CPA copy-paste.

Generates two files:
  templates/_masters/TAX-011-multi-property-master-pl-DEMO.xlsx
  templates/_masters/TAX-011-multi-property-master-pl-BLANK.xlsx
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE, STATE_BAD_FILL, STATE_GOOD_FILL,
)

SKU = "TAX-011"
NAME = "multi-property-master-pl"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Static reference data
# ---------------------------------------------------------------------------

EXPENSE_CATEGORIES = [
    "Advertising (Line 6)",
    "Auto/travel (Line 7)",
    "Cleaning + maintenance (Line 8)",
    "Commissions (Line 9)",
    "Insurance (Line 10)",
    "Legal + professional (Line 11)",
    "Management fees (Line 12)",
    "Mortgage interest (Line 13)",
    "Other interest (Line 14)",
    "Repairs (Line 15)",
    "Supplies (Line 16)",
    "Taxes (Line 17)",
    "Utilities (Line 18)",
    "Wages (Line 19)",
    "Other — Platform fees (Line 19)",
    "Other — Misc (Line 19)",
    "Depreciation (Line 20)",
]

# Schedule E line numbers in display order
SCHEDULE_E_LINES = [
    ("Line 3 — Rents received",            "rents"),
    ("Line 4 — Royalties",                  "royalties"),
    ("Line 5 — Total income",               "total_income"),
    ("Line 6 — Advertising",                EXPENSE_CATEGORIES[0]),
    ("Line 7 — Auto/travel",                EXPENSE_CATEGORIES[1]),
    ("Line 8 — Cleaning + maintenance",     EXPENSE_CATEGORIES[2]),
    ("Line 9 — Commissions",                EXPENSE_CATEGORIES[3]),
    ("Line 10 — Insurance",                 EXPENSE_CATEGORIES[4]),
    ("Line 11 — Legal + professional",      EXPENSE_CATEGORIES[5]),
    ("Line 12 — Management fees",           EXPENSE_CATEGORIES[6]),
    ("Line 13 — Mortgage interest",         EXPENSE_CATEGORIES[7]),
    ("Line 14 — Other interest",            EXPENSE_CATEGORIES[8]),
    ("Line 15 — Repairs",                   EXPENSE_CATEGORIES[9]),
    ("Line 16 — Supplies",                  EXPENSE_CATEGORIES[10]),
    ("Line 17 — Taxes",                     EXPENSE_CATEGORIES[11]),
    ("Line 18 — Utilities",                 EXPENSE_CATEGORIES[12]),
    ("Line 19 — Wages",                     EXPENSE_CATEGORIES[13]),
    ("Line 19 — Other (platform)",          EXPENSE_CATEGORIES[14]),
    ("Line 19 — Other (misc)",              EXPENSE_CATEGORIES[15]),
    ("Line 20 — Depreciation",              EXPENSE_CATEGORIES[16]),
    ("Line 26a — Total expenses",           "total_exp"),
    ("Line 26 — Income or (loss)",          "net_income"),
]

# ---------------------------------------------------------------------------
# Sample data (DEMO variant, 5 properties × 2 LLCs, Q1 2026)
# ---------------------------------------------------------------------------

PROPERTIES_DEMO = [
    # (name, address, c/s/z, type, entity, ownership %, acq, biz start, days rented YTD, active)
    ("Smokies Ridge Cabin", "123 Mountain Lane",  "Gatlinburg, TN 37738",   "Cabin",   "LLC-A", 1.00, "2023-08-15", "2023-10-01", 72, "Yes"),
    ("Creek Side",          "47 Creek Hollow Rd", "Pigeon Forge, TN 37863", "Cabin",   "LLC-A", 1.00, "2024-03-20", "2024-05-01", 64, "Yes"),
    ("Lakehouse A",         "1280 Lake Shore Dr", "Sevierville, TN 37862",  "Cabin",   "LLC-A", 1.00, "2024-09-05", "2024-11-01", 78, "Yes"),
    ("Mountain Loft",       "612 Pine Ridge Way", "Gatlinburg, TN 37738",   "Condo",   "LLC-B", 1.00, "2025-04-10", "2025-06-01", 58, "Yes"),
    ("Forest Cabin",        "89 Hidden Trail",    "Cosby, TN 37722",        "Cabin",   "LLC-B", 0.50, "2026-02-15", "2026-03-01", 18, "Yes"),
]

# Revenue samples — ~25 bookings across Q1 2026
REVENUE_SAMPLES = [
    # (date, property, guest, channel, gross, platform_fee, cleaning_collected, notes)
    ("2026-01-08", "Smokies Ridge Cabin", "Avery G.",     "Airbnb",    1245, 187, 150, ""),
    ("2026-01-12", "Creek Side",          "Marshall H.",  "Airbnb",     980, 147, 135, ""),
    ("2026-01-15", "Lakehouse A",         "Priya R.",     "VRBO",      1840, 147, 175, ""),
    ("2026-01-18", "Mountain Loft",       "Casey N.",     "Airbnb",     720, 108, 110, ""),
    ("2026-01-22", "Smokies Ridge Cabin", "Devon S.",     "Direct",     820,   0, 150, "5% direct discount"),
    ("2026-01-26", "Creek Side",          "Gabriela P.",  "Airbnb",    1320, 198, 135, ""),
    ("2026-01-30", "Lakehouse A",         "Roy + Linn T.","Airbnb",    2200, 330, 175, ""),
    ("2026-02-03", "Mountain Loft",       "Quincy J.",    "VRBO",       890,  71, 110, ""),
    ("2026-02-07", "Smokies Ridge Cabin", "Hailey M.",    "Airbnb",    1180, 177, 150, ""),
    ("2026-02-10", "Creek Side",          "Patrice O.",   "Airbnb",     875, 131, 135, ""),
    ("2026-02-14", "Lakehouse A",         "Chen W.",      "Airbnb",    2480, 372, 175, "Valentine premium"),
    ("2026-02-17", "Mountain Loft",       "Tomás A.",     "Airbnb",     945, 142, 110, ""),
    ("2026-02-19", "Forest Cabin",        "Reese F.",     "Airbnb",     680, 102, 100, "1st booking after onboard"),
    ("2026-02-22", "Smokies Ridge Cabin", "Olympia V.",   "Direct",     920,   0, 150, ""),
    ("2026-02-25", "Creek Side",          "Sasha B.",     "VRBO",      1100,  88, 135, ""),
    ("2026-02-28", "Lakehouse A",         "Wells D.",     "Airbnb",    1850, 278, 175, ""),
    ("2026-03-04", "Mountain Loft",       "Indira K.",    "Airbnb",     780, 117, 110, ""),
    ("2026-03-07", "Forest Cabin",        "Brennan E.",   "Airbnb",     720, 108, 100, ""),
    ("2026-03-10", "Smokies Ridge Cabin", "Nadia L.",     "Airbnb",    1340, 201, 150, ""),
    ("2026-03-13", "Creek Side",          "Kendall U.",   "Airbnb",     960, 144, 135, ""),
    ("2026-03-17", "Lakehouse A",         "Ezra Y.",      "VRBO",      2050, 164, 175, "St Pat's"),
    ("2026-03-20", "Mountain Loft",       "Saoirse C.",   "Airbnb",     820, 123, 110, ""),
    ("2026-03-24", "Forest Cabin",        "Jude T.",      "Direct",     650,   0, 100, ""),
    ("2026-03-28", "Smokies Ridge Cabin", "Annika W.",    "Airbnb",    1060, 159, 150, ""),
    ("2026-03-31", "Lakehouse A",         "Felipe O.",    "Airbnb",    2120, 318, 175, "Spring break"),
]

# Expense samples — ~30 entries, includes 2 shared (umbrella, accountant) split via the Allocation Helper
# (each shared cost is recorded as N rows, one per property, with the per-property amount)
# Per QA: $2,400 umbrella ÷ 5 = $480 each; $1,800 accountant ÷ 5 = $360 each.
EXPENSE_SAMPLES = [
    # (date, vendor, property, category, amount, payment_method, receipt, notes)
    # Cleaning (turnovers)
    ("2026-01-09", "Sarah — Smokies Clean",  "Smokies Ridge Cabin", EXPENSE_CATEGORIES[2], 185,   "Venmo",  "Yes", ""),
    ("2026-01-13", "Miguel — Ridge HK",       "Creek Side",         EXPENSE_CATEGORIES[2], 165,   "Venmo",  "Yes", ""),
    ("2026-01-16", "Sarah — Smokies Clean",   "Lakehouse A",        EXPENSE_CATEGORIES[2], 220,   "Venmo",  "Yes", ""),
    ("2026-01-19", "Pearl Cleaners",          "Mountain Loft",      EXPENSE_CATEGORIES[2], 130,   "ACH",    "Yes", ""),
    ("2026-01-23", "Sarah — Smokies Clean",   "Smokies Ridge Cabin", EXPENSE_CATEGORIES[2], 185,   "Venmo",  "Yes", ""),
    ("2026-02-04", "Miguel — Ridge HK",       "Creek Side",         EXPENSE_CATEGORIES[2], 165,   "Venmo",  "Yes", ""),
    ("2026-02-08", "Sarah — Smokies Clean",   "Lakehouse A",        EXPENSE_CATEGORIES[2], 220,   "Venmo",  "Yes", ""),
    ("2026-02-15", "Pearl Cleaners",          "Mountain Loft",      EXPENSE_CATEGORIES[2], 130,   "ACH",    "Yes", ""),
    ("2026-02-20", "Forest Local Cleaner",    "Forest Cabin",       EXPENSE_CATEGORIES[2], 145,   "Cash",   "Yes", "Initial deep clean"),
    ("2026-03-05", "Miguel — Ridge HK",       "Creek Side",         EXPENSE_CATEGORIES[2], 165,   "Venmo",  "Yes", ""),
    ("2026-03-11", "Sarah — Smokies Clean",   "Smokies Ridge Cabin", EXPENSE_CATEGORIES[2], 185,   "Venmo",  "Yes", ""),
    # Mortgage interest (monthly proxies)
    ("2026-01-31", "Wells Fargo",             "Smokies Ridge Cabin", EXPENSE_CATEGORIES[7], 1280,  "Auto-deduct", "Yes", "Jan int"),
    ("2026-02-28", "Wells Fargo",             "Smokies Ridge Cabin", EXPENSE_CATEGORIES[7], 1275,  "Auto-deduct", "Yes", "Feb int"),
    ("2026-03-31", "Wells Fargo",             "Smokies Ridge Cabin", EXPENSE_CATEGORIES[7], 1270,  "Auto-deduct", "Yes", "Mar int"),
    ("2026-01-31", "Truist",                  "Lakehouse A",         EXPENSE_CATEGORIES[7], 1640,  "Auto-deduct", "Yes", ""),
    ("2026-02-28", "Truist",                  "Lakehouse A",         EXPENSE_CATEGORIES[7], 1635,  "Auto-deduct", "Yes", ""),
    ("2026-03-31", "Truist",                  "Lakehouse A",         EXPENSE_CATEGORIES[7], 1630,  "Auto-deduct", "Yes", ""),
    # Property taxes
    ("2026-02-15", "Sevier County",           "Smokies Ridge Cabin", EXPENSE_CATEGORIES[11], 1450, "Check",  "Yes", "Q1 2026"),
    ("2026-02-15", "Sevier County",           "Lakehouse A",         EXPENSE_CATEGORIES[11], 1850, "Check",  "Yes", "Q1 2026"),
    # Utilities
    ("2026-01-15", "Sevier Electric",         "Smokies Ridge Cabin", EXPENSE_CATEGORIES[12], 240,  "Auto",   "Yes", ""),
    ("2026-02-15", "Sevier Electric",         "Smokies Ridge Cabin", EXPENSE_CATEGORIES[12], 215,  "Auto",   "Yes", ""),
    ("2026-03-15", "Sevier Electric",         "Smokies Ridge Cabin", EXPENSE_CATEGORIES[12], 195,  "Auto",   "Yes", ""),
    # Repairs
    ("2026-02-08", "Smoky Plumbing",          "Creek Side",          EXPENSE_CATEGORIES[9],  450,  "Check",  "Yes", "Water heater fix"),
    ("2026-03-19", "Mountain HVAC",           "Mountain Loft",       EXPENSE_CATEGORIES[9],  320,  "Card",   "Yes", "Furnace tune-up"),
    # Supplies
    ("2026-02-12", "Costco",                  "Smokies Ridge Cabin", EXPENSE_CATEGORIES[10], 142,  "Card",   "Yes", "Bulk paper + soap"),
    ("2026-03-05", "Amazon",                  "Lakehouse A",         EXPENSE_CATEGORIES[10],  98,  "Card",   "Yes", "Coffee + linens"),
    # Platform fees (rolled up monthly)
    ("2026-01-31", "Airbnb",                  "Smokies Ridge Cabin", EXPENSE_CATEGORIES[14], 564,  "Auto",   "No",  "Platform fees Jan"),
    ("2026-02-28", "Airbnb",                  "Smokies Ridge Cabin", EXPENSE_CATEGORIES[14], 537,  "Auto",   "No",  ""),
    ("2026-03-31", "Airbnb",                  "Smokies Ridge Cabin", EXPENSE_CATEGORIES[14], 561,  "Auto",   "No",  ""),
    # Shared (umbrella insurance) — split across all 5 ($2,400 ÷ 5 = $480)
    ("2026-01-15", "Liberty Mutual",          "Smokies Ridge Cabin", EXPENSE_CATEGORIES[4],  480,  "ACH",    "Yes", "Umbrella ÷ 5 (allocated)"),
    ("2026-01-15", "Liberty Mutual",          "Creek Side",          EXPENSE_CATEGORIES[4],  480,  "ACH",    "Yes", "Umbrella ÷ 5 (allocated)"),
    ("2026-01-15", "Liberty Mutual",          "Lakehouse A",         EXPENSE_CATEGORIES[4],  480,  "ACH",    "Yes", "Umbrella ÷ 5 (allocated)"),
    ("2026-01-15", "Liberty Mutual",          "Mountain Loft",       EXPENSE_CATEGORIES[4],  480,  "ACH",    "Yes", "Umbrella ÷ 5 (allocated)"),
    ("2026-01-15", "Liberty Mutual",          "Forest Cabin",        EXPENSE_CATEGORIES[4],  480,  "ACH",    "Yes", "Umbrella ÷ 5 (allocated)"),
    # Shared (accountant) — $1,800 ÷ 5 = $360
    ("2026-03-15", "Brookline CPA",           "Smokies Ridge Cabin", EXPENSE_CATEGORIES[5],  360,  "Check",  "Yes", "Tax prep ÷ 5 (allocated)"),
    ("2026-03-15", "Brookline CPA",           "Creek Side",          EXPENSE_CATEGORIES[5],  360,  "Check",  "Yes", "Tax prep ÷ 5 (allocated)"),
    ("2026-03-15", "Brookline CPA",           "Lakehouse A",         EXPENSE_CATEGORIES[5],  360,  "Check",  "Yes", "Tax prep ÷ 5 (allocated)"),
    ("2026-03-15", "Brookline CPA",           "Mountain Loft",       EXPENSE_CATEGORIES[5],  360,  "Check",  "Yes", "Tax prep ÷ 5 (allocated)"),
    ("2026-03-15", "Brookline CPA",           "Forest Cabin",        EXPENSE_CATEGORIES[5],  360,  "Check",  "Yes", "Tax prep ÷ 5 (allocated)"),
]

ENTITIES = ["LLC-A", "LLC-B", "LLC-C", "LLC-D", "LLC-E"]
PROPERTY_TYPES = ["Single-family", "Cabin", "Condo", "Multi-unit", "Glamping", "Other"]
CHANNELS = ["Airbnb", "VRBO", "Booking.com", "Direct", "Other"]
PAYMENT_METHODS = ["Venmo", "Zelle", "Check", "Cash", "ACH", "Card", "Auto", "Auto-deduct", "Other"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _val_list(variant, demo_list):
    return demo_list if variant == "demo" else []


def _parse_date(s):
    if not s:
        return None
    return date.fromisoformat(s)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, max_col=12):
    last = get_column_letter(max_col)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Multi-Property Master P&L"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "One workbook for the whole portfolio. CPA-ready."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Portfolio at-a-glance — 4 cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    cards = [
        ("A", "C", "PORTFOLIO YTD NET",    '=\'Consolidated P&L\'!N32',    '"$"#,##0',  COLOR_PRIMARY),
        ("D", "F", "PROPERTIES ACTIVE",    '=COUNTIF(\'Properties Register\'!J6:J25,"Yes")', "0", COLOR_ACCENT),
        ("G", "I", "YTD REVENUE",          '=\'Consolidated P&L\'!N9',     '"$"#,##0',  COLOR_PRIMARY),
        ("J", "L", "YTD EXPENSES",         '=\'Consolidated P&L\'!N30',    '"$"#,##0',  COLOR_SECONDARY),
    ]
    for fc, lc, label, formula, fmt, color in cards:
        ws.merge_cells(f"{fc}10:{lc}10")
        c = ws[f"{fc}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=8, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{fc}11:{lc}13")
        c = ws[f"{fc}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=24, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

        ws.merge_cells(f"{fc}14:{lc}15")
        c = ws[f"{fc}14"]
        c.value = ""
        c.font = Font(name=FONT_BODY, size=8, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for fc, lc, *_ in cards:
        first_c = column_index_from_string(fc)
        last_c = column_index_from_string(lc)
        for r in range(10, 16):
            for col in range(first_c, last_c + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == first_c else existing.left,
                    right=gold_side if col == last_c else existing.right,
                )

    # Best / Worst rows 17-20
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "PERFORMANCE SCORECARD"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    # Best performer
    ws.merge_cells("A18:F18")
    c = ws["A18"]
    c.value = "Best performer:"
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Pull best name + value via INDEX/MATCH on the Per-Property P&L net row
    ws.merge_cells("G18:L18")
    c = ws["G18"]
    c.value = (
        '=IFERROR(INDEX(\'Per-Property P&L\'!$B$5:$I$5,'
        'MATCH(MAX(\'Per-Property P&L\'!$B$32:$I$32),\'Per-Property P&L\'!$B$32:$I$32,0))'
        '&"   ($"&TEXT(MAX(\'Per-Property P&L\'!$B$32:$I$32),"#,##0")&" net)","—")'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[18].height = 22

    # Worst performer
    ws.merge_cells("A19:F19")
    c = ws["A19"]
    c.value = "Worst performer:"
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    ws.merge_cells("G19:L19")
    c = ws["G19"]
    c.value = (
        '=IFERROR(INDEX(\'Per-Property P&L\'!$B$5:$I$5,'
        'MATCH(MIN(\'Per-Property P&L\'!$B$32:$I$32),\'Per-Property P&L\'!$B$32:$I$32,0))'
        '&"   ($"&TEXT(MIN(\'Per-Property P&L\'!$B$32:$I$32),"#,##0")&" net)","—")'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[19].height = 22

    # Pseudo-button nav rows 22-24 + 26-28
    pseudo_button(ws, "A22", "C24", "Properties Register",
                  "'Properties Register'!A1", variant="primary")
    pseudo_button(ws, "D22", "F24", "Revenue Log",
                  "'Revenue Log'!A1", variant="primary")
    pseudo_button(ws, "G22", "I24", "Expense Log",
                  "'Expense Log'!A1", variant="primary")
    pseudo_button(ws, "J22", "L24", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(22, 25):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A26", "C28", "Per-Property P&L",
                  "'Per-Property P&L'!A1", variant="accent")
    pseudo_button(ws, "D26", "F28", "Consolidated P&L",
                  "'Consolidated P&L'!A1", variant="accent")
    pseudo_button(ws, "G26", "I28", "Entity Rollup",
                  "'Entity Rollup'!A1", variant="accent")
    pseudo_button(ws, "J26", "L28", "Schedule E Multi",
                  "'Schedule E Multi'!A1", variant="accent")
    for r in range(26, 29):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 31
    ws.merge_cells("A31:L31")
    c = ws["A31"]
    c.value = (
        "💡  Want depreciation per asset (Form 4562 ready) + cost segregation? "
        f"Get the Portfolio Bundle at {BRAND_DOMAIN}/portfolio-bundle — $397 "
        "(saves $293 vs individual)."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[31].height = 36

    brand_footer(ws, 33,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L35"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Properties Register
# ---------------------------------------------------------------------------

def build_properties_register_tab(wb, variant):
    ws = wb.create_sheet("Properties Register")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Properties Register",
                        prev_tab="Start", next_tab="Revenue Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "One row per property (capacity 20). LLC/Entity column drives the Entity Rollup tab."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Property name", "Street address", "City / State / Zip",
        "Type", "Entity / LLC", "Ownership %", "Acq. date",
        "Biz start", "Days rented YTD", "Active?",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 26), ("B", 28), ("C", 22), ("D", 12),
        ("E", 14), ("F", 10), ("G", 12), ("H", 12),
        ("I", 12), ("J", 10),
    ])

    samples = _val_list(variant, PROPERTIES_DEMO)
    for i in range(20):
        row = 6 + i
        if i < len(samples):
            (name, addr, csz, ptype, entity, own, acq, biz, days, active) = samples[i]
        else:
            name = addr = csz = ptype = entity = acq = biz = active = None
            own = days = None

        cells = [
            (1, name, None),
            (2, addr, None),
            (3, csz, None),
            (4, ptype, None),
            (5, entity, None),
            (6, own, "0%"),
            (7, _parse_date(acq) if acq else None, "yyyy-mm-dd"),
            (8, _parse_date(biz) if biz else None, "yyyy-mm-dd"),
            (9, days, "0"),
            (10, active, None),
        ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    # Dropdowns
    add_dropdown(ws, "D6:D25", f'"{",".join(PROPERTY_TYPES)}"')
    add_dropdown(ws, "E6:E25", f'"{",".join(ENTITIES)}"')
    add_dropdown(ws, "J6:J25", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Revenue Log
# ---------------------------------------------------------------------------

def build_revenue_log_tab(wb, variant):
    ws = wb.create_sheet("Revenue Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Revenue Log",
                        prev_tab="Properties Register", next_tab="Expense Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Single log — all properties via Property dropdown. Capacity 3000 rows. "
        "Net = Gross − Platform fee."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Date", "Property", "Guest / Source", "Channel",
        "Gross", "Platform Fee", "Cleaning Collected", "Net",
        "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 12), ("B", 24), ("C", 22), ("D", 12),
        ("E", 12), ("F", 12), ("G", 14), ("H", 12),
        ("I", 28),
    ])

    samples = _val_list(variant, REVENUE_SAMPLES)
    for i in range(3000):
        row = 6 + i
        if i < len(samples):
            (date_s, prop, guest, channel, gross, fee, clean, notes) = samples[i]
            cells = [
                (1, _parse_date(date_s), "yyyy-mm-dd"),
                (2, prop, None),
                (3, guest, None),
                (4, channel, None),
                (5, gross, '"$"#,##0.00'),
                (6, fee, '"$"#,##0.00'),
                (7, clean, '"$"#,##0.00'),
                (9, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None),
                (3, None, None),
                (4, None, None),
                (5, None, '"$"#,##0.00'),
                (6, None, '"$"#,##0.00'),
                (7, None, '"$"#,##0.00'),
                (9, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt

        # Net column (formula)
        net_cell = ws.cell(row=row, column=8, value=f"=E{row}-F{row}")
        apply_style(net_cell, formula_cell_style())
        net_cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[row].height = 16

    # Dropdowns — Property pulls from Properties Register A6:A25 (active list)
    add_dropdown(ws, "B6:B3005", "='Properties Register'!$A$6:$A$25")
    add_dropdown(ws, "D6:D3005", f'"{",".join(CHANNELS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Expense Log
# ---------------------------------------------------------------------------

def build_expense_log_tab(wb, variant):
    ws = wb.create_sheet("Expense Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Expense Log",
                        prev_tab="Revenue Log", next_tab="Per-Property P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Single log — all properties. Capacity 5000 rows. For shared expenses "
        "(umbrella, accountant, software bundle), use the Allocation Helper on "
        "Settings → enter as N rows, one per allocated property."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Date", "Vendor", "Property", "Category", "Amount",
        "Pmt method", "Receipt?", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 12), ("B", 24), ("C", 24), ("D", 32),
        ("E", 12), ("F", 14), ("G", 10), ("H", 30),
    ])

    samples = _val_list(variant, EXPENSE_SAMPLES)
    for i in range(5000):
        row = 6 + i
        if i < len(samples):
            (date_s, vendor, prop, cat, amount, pmt, receipt, notes) = samples[i]
            cells = [
                (1, _parse_date(date_s), "yyyy-mm-dd"),
                (2, vendor, None),
                (3, prop, None),
                (4, cat, None),
                (5, amount, '"$"#,##0.00'),
                (6, pmt, None),
                (7, receipt, None),
                (8, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None),
                (3, None, None),
                (4, None, None),
                (5, None, '"$"#,##0.00'),
                (6, None, None),
                (7, None, None),
                (8, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 16

    add_dropdown(ws, "C6:C5005", "='Properties Register'!$A$6:$A$25")
    add_dropdown(ws, "D6:D5005", "=Settings!$F$6:$F$22")
    add_dropdown(ws, "F6:F5005", f'"{",".join(PAYMENT_METHODS)}"')
    add_dropdown(ws, "G6:G5005", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Per-Property P&L
# ---------------------------------------------------------------------------

def build_per_property_pl_tab(wb, variant):
    """8 visible property cols + total. Rows = revenue + expenses + net."""
    ws = wb.create_sheet("Per-Property P&L")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Per-Property P&L",
                        prev_tab="Expense Log", next_tab="Consolidated P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Auto. Cols = first 8 properties from Properties Register + total. "
        "All values filtered by tax year (Settings → B5)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 36),
        ("B", 14), ("C", 14), ("D", 14), ("E", 14),
        ("F", 14), ("G", 14), ("H", 14), ("I", 14),
        ("J", 16),
    ])

    # Header row 5: Category | property cols (B..I = 8 properties) | TOTAL (J)
    cell = ws.cell(row=5, column=1, value="Category")
    apply_style(cell, header_row_style())

    for i in range(8):
        col = 2 + i
        # Pull property name from Properties Register row (6+i)
        cell = ws.cell(row=5, column=col, value=f"='Properties Register'!A{6+i}")
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell = ws.cell(row=5, column=10, value="PORTFOLIO TOTAL")
    apply_style(cell, header_row_style())
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 32

    # Row 7: Revenue (Rents + cleaning fees collected per Schedule E Line 3)
    ws.cell(row=7, column=1, value="Rents + cleaning fees collected (Line 3)").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for i in range(8):
        col = 2 + i
        col_letter = get_column_letter(col)
        # SUMIFS over Revenue Log (gross + cleaning collected) by property + tax year
        # property header is in row 5 of this tab
        formula = (
            f"=SUMIFS('Revenue Log'!$E$6:$E$3005,"
            f"'Revenue Log'!$B$6:$B$3005,{col_letter}$5,"
            f"'Revenue Log'!$A$6:$A$3005,\">=\"&DATE(Settings!$B$5,1,1),"
            f"'Revenue Log'!$A$6:$A$3005,\"<\"&DATE(Settings!$B$5+1,1,1))"
            f"+SUMIFS('Revenue Log'!$G$6:$G$3005,"
            f"'Revenue Log'!$B$6:$B$3005,{col_letter}$5,"
            f"'Revenue Log'!$A$6:$A$3005,\">=\"&DATE(Settings!$B$5,1,1),"
            f"'Revenue Log'!$A$6:$A$3005,\"<\"&DATE(Settings!$B$5+1,1,1))"
        )
        cell = ws.cell(row=7, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    # Total col (J)
    cell = ws.cell(row=7, column=10, value="=SUM(B7:I7)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 22

    # Row 9: EXPENSES section banner
    ws.merge_cells("A9:J9")
    c = ws["A9"]
    c.value = "EXPENSES — by Schedule E line"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[9].height = 22

    # Rows 10-26: 17 expense categories
    for i, cat in enumerate(EXPENSE_CATEGORIES):
        row = 10 + i
        ws.cell(row=row, column=1, value=cat).font = Font(
            name=FONT_BODY, size=10, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )

        for j in range(8):
            col = 2 + j
            col_letter = get_column_letter(col)
            formula = (
                f"=SUMIFS('Expense Log'!$E$6:$E$5005,"
                f"'Expense Log'!$C$6:$C$5005,{col_letter}$5,"
                f"'Expense Log'!$D$6:$D$5005,$A{row},"
                f"'Expense Log'!$A$6:$A$5005,\">=\"&DATE(Settings!$B$5,1,1),"
                f"'Expense Log'!$A$6:$A$5005,\"<\"&DATE(Settings!$B$5+1,1,1))"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'

        # Total col (J)
        cell = ws.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 18

    # Row 28: TOTAL EXPENSES
    ws.cell(row=28, column=1, value="TOTAL EXPENSES").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR
    )
    ws.cell(row=28, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 11):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=28, column=col, value=f"=SUM({col_letter}10:{col_letter}26)")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[28].height = 24

    # Row 32 spacer + NET INCOME
    ws.cell(row=32, column=1, value="NET INCOME (LOSS)").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=32, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 11):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=32, column=col, value=f"={col_letter}7-{col_letter}28")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        # Conditional formatting on the net row
    ws.row_dimensions[32].height = 28

    ws.conditional_formatting.add(
        "B32:J32",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        "B32:J32",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Consolidated P&L (12 months × 17 categories)
# ---------------------------------------------------------------------------

def build_consolidated_pl_tab(wb, variant):
    ws = wb.create_sheet("Consolidated P&L")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Consolidated P&L",
                        prev_tab="Per-Property P&L", next_tab="Entity Rollup")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:N4")
    c = ws["A4"]
    c.value = "Portfolio-wide monthly P&L. All properties combined for the tax year on Settings → B5."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Header row 5: Category | Jan..Dec | YTD
    cell = ws.cell(row=5, column=1, value="Category")
    apply_style(cell, header_row_style())
    for i, m in enumerate(months):
        cell = ws.cell(row=5, column=2 + i, value=m)
        apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=14, value="YTD")
    apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 38),
        ("B", 10), ("C", 10), ("D", 10), ("E", 10), ("F", 10), ("G", 10),
        ("H", 10), ("I", 10), ("J", 10), ("K", 10), ("L", 10), ("M", 10),
        ("N", 14),
    ])

    # Row 7: Rents + cleaning fees (Line 3)
    ws.cell(row=7, column=1, value="Rents + cleaning fees collected (Line 3)").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for m_idx in range(12):
        col = 2 + m_idx
        formula = (
            f"=SUMIFS('Revenue Log'!$E$6:$E$3005,"
            f"'Revenue Log'!$A$6:$A$3005,\">=\"&DATE(Settings!$B$5,{m_idx+1},1),"
            f"'Revenue Log'!$A$6:$A$3005,\"<\"&DATE(Settings!$B$5,{m_idx+2},1))"
            f"+SUMIFS('Revenue Log'!$G$6:$G$3005,"
            f"'Revenue Log'!$A$6:$A$3005,\">=\"&DATE(Settings!$B$5,{m_idx+1},1),"
            f"'Revenue Log'!$A$6:$A$3005,\"<\"&DATE(Settings!$B$5,{m_idx+2},1))"
        )
        cell = ws.cell(row=7, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=7, column=14, value="=SUM(B7:M7)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 22

    # Row 9 spacer note: "(also referenced by Start tab card N9)"
    # Row 9: actually used by Start. I want N9 = revenue YTD or row 7's YTD.
    # Re-aim: Start references Consolidated!N9. Easier to put YTD revenue at N9 too.
    # Add a hidden alias at N9.
    cell = ws.cell(row=9, column=14, value="=N7")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    ws.cell(row=9, column=1, value="(YTD revenue — alias for Start tab)").font = Font(
        name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
    )

    # Row 11: EXPENSES banner
    ws.merge_cells("A11:N11")
    c = ws["A11"]
    c.value = "EXPENSES"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[11].height = 22

    # Rows 12-28: 17 categories
    for i, cat in enumerate(EXPENSE_CATEGORIES):
        row = 12 + i
        ws.cell(row=row, column=1, value=cat).font = Font(
            name=FONT_BODY, size=10, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        for m_idx in range(12):
            col = 2 + m_idx
            formula = (
                f"=SUMIFS('Expense Log'!$E$6:$E$5005,"
                f"'Expense Log'!$D$6:$D$5005,$A{row},"
                f"'Expense Log'!$A$6:$A$5005,\">=\"&DATE(Settings!$B$5,{m_idx+1},1),"
                f"'Expense Log'!$A$6:$A$5005,\"<\"&DATE(Settings!$B$5,{m_idx+2},1))"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
        cell = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 18

    # Row 30: TOTAL EXPENSES
    ws.cell(row=30, column=1, value="TOTAL EXPENSES").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR
    )
    ws.cell(row=30, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=30, column=col, value=f"=SUM({col_letter}12:{col_letter}28)")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[30].height = 24

    # Row 32: NET INCOME
    ws.cell(row=32, column=1, value="NET INCOME (LOSS)").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=32, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=32, column=col, value=f"={col_letter}7-{col_letter}30")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[32].height = 28

    ws.conditional_formatting.add(
        "B32:N32",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        "B32:N32",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 7 — Entity Rollup (groups by LLC)
# ---------------------------------------------------------------------------

def build_entity_rollup_tab(wb, variant):
    ws = wb.create_sheet("Entity Rollup")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Entity Rollup",
                        prev_tab="Consolidated P&L", next_tab="Schedule E Multi")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:I4")
    c = ws["A4"]
    c.value = (
        "Sums by LLC / entity. Properties Register Entity column drives this. "
        "Edit entity list on Settings → column G."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 36),
        ("B", 14), ("C", 14), ("D", 14), ("E", 14), ("F", 14),
        ("G", 14),
    ])

    # Header row 5: Category | LLC-A..LLC-E | TOTAL
    cell = ws.cell(row=5, column=1, value="Category")
    apply_style(cell, header_row_style())
    for i in range(5):
        cell = ws.cell(row=5, column=2 + i, value=ENTITIES[i])
        apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=7, value="TOTAL")
    apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # Row 7: Revenue
    ws.cell(row=7, column=1, value="Rents + cleaning fees collected").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for i in range(5):
        col = 2 + i
        col_letter = get_column_letter(col)
        # SUMPRODUCT pattern: for each row in Revenue Log, look up the entity
        # of that row's property and check it equals the column entity.
        # Year filter included.
        formula = (
            f"=SUMPRODUCT("
            f"(YEAR('Revenue Log'!$A$6:$A$3005)=Settings!$B$5)*"
            f"(IFERROR(VLOOKUP('Revenue Log'!$B$6:$B$3005,'Properties Register'!$A$6:$E$25,5,FALSE),\"\")={col_letter}$5)*"
            f"('Revenue Log'!$E$6:$E$3005+'Revenue Log'!$G$6:$G$3005))"
        )
        cell = ws.cell(row=7, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=7, column=7, value="=SUM(B7:F7)")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 22

    # Row 9 spacer
    ws.merge_cells("A9:G9")
    c = ws["A9"]
    c.value = "EXPENSES"
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[9].height = 22

    # Rows 10-26: 17 categories
    for cat_idx, cat in enumerate(EXPENSE_CATEGORIES):
        row = 10 + cat_idx
        ws.cell(row=row, column=1, value=cat).font = Font(
            name=FONT_BODY, size=10, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        for i in range(5):
            col = 2 + i
            col_letter = get_column_letter(col)
            formula = (
                f"=SUMPRODUCT("
                f"(YEAR('Expense Log'!$A$6:$A$5005)=Settings!$B$5)*"
                f"('Expense Log'!$D$6:$D$5005=$A{row})*"
                f"(IFERROR(VLOOKUP('Expense Log'!$C$6:$C$5005,'Properties Register'!$A$6:$E$25,5,FALSE),\"\")={col_letter}$5)*"
                f"'Expense Log'!$E$6:$E$5005)"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
        cell = ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 18

    # Row 28: total expenses
    ws.cell(row=28, column=1, value="TOTAL EXPENSES").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR
    )
    ws.cell(row=28, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=28, column=col, value=f"=SUM({col_letter}10:{col_letter}26)")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[28].height = 24

    # Row 30: NET INCOME
    ws.cell(row=30, column=1, value="NET INCOME (LOSS)").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=30, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=30, column=col, value=f"={col_letter}7-{col_letter}28")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[30].height = 28

    ws.conditional_formatting.add(
        "B30:G30",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        "B30:G30",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 8 — Schedule E Multi (one column per property)
# ---------------------------------------------------------------------------

def build_schedule_e_multi_tab(wb, variant):
    ws = wb.create_sheet("Schedule E Multi")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Schedule E Multi",
                        prev_tab="Entity Rollup", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:K4")
    c = ws["A4"]
    c.value = (
        "Schedule E Part I — one column per property. Print and hand to your CPA "
        "(or copy values directly into Schedule E)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 36),
        ("B", 13), ("C", 13), ("D", 13), ("E", 13),
        ("F", 13), ("G", 13), ("H", 13), ("I", 13),
        ("J", 16),
    ])

    # Header: Schedule E Line | property cols (B..I = 8 props) | TOTAL (J)
    cell = ws.cell(row=5, column=1, value="Schedule E Part I")
    apply_style(cell, header_row_style())
    for i in range(8):
        col = 2 + i
        cell = ws.cell(row=5, column=col, value=f"='Properties Register'!A{6+i}")
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell = ws.cell(row=5, column=10, value="PORTFOLIO TOTAL")
    apply_style(cell, header_row_style())
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 32

    # Tax year + Property line below header
    ws.cell(row=6, column=1, value="Tax year:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=6, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=6, column=2, value="=Settings!B5")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Each Schedule E line maps to a row on the Per-Property P&L tab
    # Row 7 onwards: Schedule E lines
    # Per-Property P&L row map:
    #   Row 7    = Rents + cleaning fees (Line 3)
    #   Rows 10-26 = 17 expense categories (Lines 6-20)
    #   Row 28   = Total expenses (Line 26a)
    #   Row 32   = Net (Line 26)
    #
    # Schedule E lines we display:
    schedule_lines = [
        # (Schedule E label, source row on Per-Property P&L)
        ("Line 3 — Rents received",            7),
        ("Line 4 — Royalties",                 None),  # hardcode 0
        ("Line 6 — Advertising",               10),
        ("Line 7 — Auto/travel",               11),
        ("Line 8 — Cleaning + maintenance",    12),
        ("Line 9 — Commissions",               13),
        ("Line 10 — Insurance",                14),
        ("Line 11 — Legal + professional",     15),
        ("Line 12 — Management fees",          16),
        ("Line 13 — Mortgage interest",        17),
        ("Line 14 — Other interest",           18),
        ("Line 15 — Repairs",                  19),
        ("Line 16 — Supplies",                 20),
        ("Line 17 — Taxes",                    21),
        ("Line 18 — Utilities",                22),
        ("Line 19 — Wages",                    23),
        ("Line 19 — Other (platform)",         24),
        ("Line 19 — Other (misc)",             25),
        ("Line 20 — Depreciation",             26),
        ("Line 26a — Total expenses",          28),
        ("Line 26 — Income or (loss)",         32),
    ]
    for i, (label, src_row) in enumerate(schedule_lines):
        row = 8 + i
        bold = label.startswith(("Line 3", "Line 26a", "Line 26 —"))
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY,
            size=11 if bold else 10,
            bold=bold,
            color=COLOR_PRIMARY if bold else COLOR_TEXT,
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )

        for j in range(8):
            col = 2 + j
            col_letter = get_column_letter(col)
            if src_row is None:
                cell = ws.cell(row=row, column=col, value=0)
            else:
                cell = ws.cell(row=row, column=col, value=f"='Per-Property P&L'!{col_letter}{src_row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            if bold:
                cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

        # Total col J
        if src_row is None:
            cell = ws.cell(row=row, column=10, value=0)
        else:
            cell = ws.cell(row=row, column=10, value=f"='Per-Property P&L'!J{src_row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        if bold:
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 18 if not bold else 22

    # Final disclaimer
    final_row = 8 + len(schedule_lines) + 2
    ws.merge_cells(f"A{final_row}:J{final_row+1}")
    c = ws[f"A{final_row}"]
    c.value = (
        "⚠  Depreciation (Line 20) is a placeholder — this Lite version doesn't "
        "include a per-asset depreciation calculator. For Form 4562 + 5/7/15/27.5/39-yr "
        "schedules, see the Portfolio Bundle ($397). Override values directly here if your "
        "CPA has computed them externally."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[final_row].height = 18
    ws.row_dimensions[final_row + 1].height = 18

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 9 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 18),
        ("C", 4), ("D", 24), ("E", 18),
        ("F", 38), ("G", 14),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Schedule E Multi", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Tax year drives all P&L formulas. Allocation Helper splits a shared "
        "expense — paste the per-property amount into the Expense Log."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # ---- TAX YEAR (B5) ----
    ws.cell(row=5, column=1, value="Tax year:").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=5, column=2, value=2026 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 28

    ws.merge_cells("D5:G5")
    c = ws["D5"]
    c.value = (
        "Type the tax year here. All P&L tabs filter by this. "
        "Default is 2026 (BLANK uses =YEAR(TODAY()) — overwrite for prior years)."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    # Default formula for blank variant
    if variant != "demo":
        cell.value = "=YEAR(TODAY())"

    # ---- ENTITIES (cols A-B header + values rows 7-12) ----
    ws.cell(row=7, column=1, value="ENTITY LIST  (Properties Register dropdown)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A7:B7")
    ws.row_dimensions[7].height = 20

    for i in range(5):
        row = 8 + i
        cell = ws.cell(row=row, column=2, value=ENTITIES[i])
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # ---- EXPENSE CATEGORIES (col F rows 6-22) — 17 categories ----
    # Header at row 25 below the list (row 5 is in the merged tax-year description)
    for i, cat in enumerate(EXPENSE_CATEGORIES):
        row = 6 + i
        cell = ws.cell(row=row, column=6, value=cat)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=24, column=6, value="↑ Schedule E categories — Expense Log dropdown source (do not reorder)").font = Font(
        name=FONT_MONO, size=9, italic=True, color=COLOR_MUTED
    )

    # ---- ALLOCATION HELPER (D rows 14-22) ----
    ws.cell(row=14, column=1, value="ALLOCATION HELPER  (split a shared expense)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=14, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A14:D14")
    ws.row_dimensions[14].height = 20

    helper_fields = [
        (15, "Total amount ($):",            2400, '"$"#,##0.00'),
        (16, "Number of properties:",        5,    "0"),
        (17, "→ Per-property amount ($):",   "=B15/B16", '"$"#,##0.00'),
    ]
    for row, label, value, fmt in helper_fields:
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=row, column=2, value=value if variant == "demo" or row == 17 else None)
        if row == 17:
            apply_style(cell, formula_cell_style())
            cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        else:
            apply_style(cell, input_cell_style())
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    ws.merge_cells("A19:D22")
    c = ws["A19"]
    c.value = (
        "Use this helper for shared expenses (umbrella insurance, accountant, "
        "software bundle). Enter the total and how many properties to split "
        "across. Then on the Expense Log, create N rows — one per property — "
        "each with the per-property amount above. Add '(allocated)' to the "
        "Notes column so you can audit later."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(19, 23):
        ws.row_dimensions[r].height = 16


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_properties_register_tab(wb, variant)
    build_revenue_log_tab(wb, variant)
    build_expense_log_tab(wb, variant)
    build_per_property_pl_tab(wb, variant)
    build_consolidated_pl_tab(wb, variant)
    build_entity_rollup_tab(wb, variant)
    build_schedule_e_multi_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Multi-Property Master P&L — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Portfolio-wide P&L for STR hosts — per-property + consolidated + "
        "entity rollup + multi-column Schedule E ready for CPA handoff."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
