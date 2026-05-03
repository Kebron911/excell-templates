"""Build FIN-003 12-Month Rolling Cash Flow Forecaster Excel files (v2.2 standard).

Operational-mode forecast tool: customer projects 12 months of revenue,
recurring expenses, and one-off expenses across the portfolio. Surfaces
months with negative cash flow BEFORE they happen — solves the
"ran out of operating cash in February" failure mode.

Six tabs:
  1. Start                — KPIs (avg cash flow, lowest month, runway, days-cash) + nav
  2. Bookings + Revenue   — 12-mo matrix per property (occ %, ADR, projected $)
  3. Recurring Expenses   — 12-mo matrix per category (15 categories)
  4. One-Off Expenses     — date/property/category/amount log
  5. Cash Flow Forecast   — combined revenue/expense/net/cumulative + chart
  6. Settings             — active year, starting cash, property + category lists

Generates BOTH DEMO and BLANK variants from shared code.
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "FIN-003"
NAME = "12-month-cash-flow-forecaster"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026
STARTING_CASH = 48000

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# 3 properties for DEMO
PROPERTIES = [
    "Smokies Ridge Cabin",
    "Lakeside Loft",
    "Downtown Studio",
]

# 15 recurring expense categories
RECURRING_CATEGORIES = [
    "Mortgage (P&I)",
    "Property tax (escrow)",
    "Insurance",
    "HOA / community fees",
    "Internet",
    "Electric",
    "Gas",
    "Water / sewer",
    "Trash / recycling",
    "Software (PMS, accounting)",
    "Cleaning (recurring contract)",
    "Supplies restock",
    "Professional fees (CPA, legal)",
    "Marketing / listing boost",
    "Misc / contingency",
]

# DEMO: per-property monthly occupancy %, ADR, and confirmed bookings $
# Pattern: summer peak (Jun-Aug), shoulder Apr/May/Sep, winter trough Dec-Mar
# Smokies Ridge — large cabin, $300+ ADR
# Lakeside Loft — mid-tier, $200 ADR
# Downtown Studio — steady urban, $140 ADR
DEMO_OCC = {
    "Smokies Ridge Cabin": [0.45, 0.40, 0.55, 0.65, 0.75, 0.88, 0.92, 0.90, 0.78, 0.70, 0.60, 0.65],
    "Lakeside Loft":       [0.38, 0.35, 0.48, 0.60, 0.72, 0.85, 0.90, 0.88, 0.72, 0.62, 0.50, 0.55],
    "Downtown Studio":     [0.55, 0.58, 0.62, 0.68, 0.70, 0.72, 0.74, 0.72, 0.70, 0.68, 0.62, 0.65],
}
DEMO_ADR = {
    "Smokies Ridge Cabin": [280, 280, 290, 300, 320, 360, 380, 380, 340, 320, 300, 340],
    "Lakeside Loft":       [180, 180, 190, 200, 210, 240, 260, 260, 220, 210, 200, 220],
    "Downtown Studio":     [130, 130, 140, 145, 150, 160, 165, 165, 155, 150, 145, 150],
}
DEMO_CONFIRMED = {
    # Confirmed booking $ already on the books per month (from PMS export).
    # Lower than projected — host confirms ~30-50% by month-start.
    "Smokies Ridge Cabin": [3800, 3200, 2800, 2400, 1800, 2400, 1800, 1200,  900,  600,  300,  900],
    "Lakeside Loft":       [2000, 1800, 1600, 1400, 1200, 1500, 1100,  800,  600,  400,  200,  600],
    "Downtown Studio":     [2200, 2300, 1800, 1600, 1300, 1100,  900,  700,  500,  400,  300,  500],
}

# DEMO recurring expenses ~ $11K/mo total
DEMO_RECURRING = {
    "Mortgage (P&I)":              [4200] * 12,
    "Property tax (escrow)":       [ 850] * 12,
    "Insurance":                   [ 420] * 12,
    "HOA / community fees":        [ 180] * 12,
    "Internet":                    [ 270] * 12,  # 3 properties × $90
    "Electric":                    [ 380, 380, 320, 280, 260, 320, 420, 440, 360, 290, 280, 360],
    "Gas":                         [ 240, 240, 200, 140, 100,  80,  60,  60,  90, 140, 200, 240],
    "Water / sewer":               [ 180] * 12,
    "Trash / recycling":           [  90] * 12,
    "Software (PMS, accounting)":  [ 240] * 12,
    "Cleaning (recurring contract)":[2400, 2200, 2400, 2600, 2800, 3000, 3200, 3200, 2800, 2600, 2400, 2600],
    "Supplies restock":            [ 180, 160, 200, 220, 240, 260, 280, 280, 240, 220, 200, 220],
    "Professional fees (CPA, legal)":[800, 200, 200, 200, 200, 200, 200, 200, 200, 200, 200, 200],
    "Marketing / listing boost":   [ 150, 150, 200, 200, 150, 100, 100, 100, 150, 200, 200, 200],
    "Misc / contingency":          [ 200] * 12,
}

# DEMO one-off expenses (the surprises that wreck a forecast)
# Q1 estimated tax + October HVAC
DEMO_ONEOFFS = [
    ("2026-01-15", "Smokies Ridge Cabin", "Estimated tax (Q1)",   1500, "Federal Q1"),
    ("2026-01-15", "Lakeside Loft",       "Estimated tax (Q1)",   1500, "Federal Q1"),
    ("2026-01-15", "Downtown Studio",     "Estimated tax (Q1)",   1500, "Federal Q1"),
    ("2026-04-15", "Smokies Ridge Cabin", "Estimated tax (Q2)",   1200, "Federal Q2"),
    ("2026-04-15", "Lakeside Loft",       "Estimated tax (Q2)",   1200, "Federal Q2"),
    ("2026-04-15", "Downtown Studio",     "Estimated tax (Q2)",   1200, "Federal Q2"),
    ("2026-06-15", "Smokies Ridge Cabin", "Estimated tax (Q3)",   1200, "Federal Q3"),
    ("2026-06-15", "Lakeside Loft",       "Estimated tax (Q3)",   1200, "Federal Q3"),
    ("2026-06-15", "Downtown Studio",     "Estimated tax (Q3)",   1200, "Federal Q3"),
    ("2026-09-15", "Smokies Ridge Cabin", "Estimated tax (Q4)",   1500, "Federal Q4"),
    ("2026-09-15", "Lakeside Loft",       "Estimated tax (Q4)",   1500, "Federal Q4"),
    ("2026-09-15", "Downtown Studio",     "Estimated tax (Q4)",   1500, "Federal Q4"),
    ("2026-10-12", "Smokies Ridge Cabin", "HVAC replacement",     7500, "End-of-life HVAC unit"),
    ("2026-03-22", "Lakeside Loft",       "Appliance — fridge",    900, "Replaced after warranty"),
    ("2026-05-08", "Downtown Studio",     "Lawyer — lease review", 450, "Property mgmt contract"),
    ("2026-08-20", "Smokies Ridge Cabin", "Deck restain",         1100, "Annual maintenance"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    return datetime.strptime(s, "%Y-%m-%d").date()


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _band(ws, row_start, row_end, col_start=1, col_end=12, color=COLOR_BG_LIGHT):
    fill = PatternFill("solid", fgColor=color)
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).fill = fill


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational hero + KPI cards + nav)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "12-Month Cash Flow Forecast"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "See your low month before it sees you."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT — lowest projected cash position
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(MIN(\'Cash Flow Forecast\'!B11:M11)<0,'
        '"⚠  Projected cash dips to "&TEXT(MIN(\'Cash Flow Forecast\'!B11:M11),"$#,##0")'
        '&" in "&INDEX({"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"},'
        'MATCH(MIN(\'Cash Flow Forecast\'!B11:M11),\'Cash Flow Forecast\'!B11:M11,0))'
        '&"  —  raise reserve or move expenses.",'
        '"✅  Cash stays positive every month  —  lowest = "'
        '&TEXT(MIN(\'Cash Flow Forecast\'!B11:M11),"$#,##0")&".")'
    )
    c.font = Font(name=FONT_HEAD, size=15, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 30

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: What this does
    _band(ws, 10, 16, color=COLOR_BG_LIGHT)
    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Combines confirmed bookings, projected revenue (occ × ADR × 30), recurring "
        "monthly expenses, and one-off planned costs (renovations, estimated taxes) "
        "into a single 12-month forecast. Cumulative cash starts from your Jan 1 "
        "balance and runs forward — the lowest month is the only one that matters."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 3: How to use
    _band(ws, 18, 25, color=COLOR_PARCHMENT_ALT)
    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    steps = [
        "① Settings: enter your starting cash balance (Jan 1) and active year.",
        "② Bookings + Revenue: paste confirmed $ from your PMS, set occ % + ADR per month per property.",
        "③ Recurring Expenses: enter monthly amounts for mortgage, insurance, utilities, etc.",
        "④ One-Off Expenses: log planned 1-time costs (renovations, estimated taxes) by date.",
        "⑤ Cash Flow Forecast: read the chart. Red bars = months your cumulative cash goes negative.",
        "⑥ Re-run monthly — drag confirmed bookings up as guests book; the low month moves with you.",
    ]
    for i, item in enumerate(steps):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 20

    # ZONE 4: Primary button
    pseudo_button(ws, "A27", "L30",
                   "→  OPEN CASH FLOW FORECAST",
                   "'Cash Flow Forecast'!A1", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # ZONE 5: 4 KPI cards (rows 32-37)
    _band(ws, 32, 37, color=COLOR_BG_LIGHT)

    kpi_cards = [
        # (col_start, col_end, label, value_formula, fmt, label_color, val_color)
        ("A", "C", "AVG MONTHLY NET",
         "=AVERAGE('Cash Flow Forecast'!B10:M10)",
         '"$"#,##0', COLOR_PRIMARY, COLOR_PRIMARY),
        ("D", "F", "LOWEST MONTH",
         "=MIN('Cash Flow Forecast'!B11:M11)",
         '"$"#,##0', COLOR_ERROR, COLOR_ERROR),
        ("G", "I", "RUNWAY (MONTHS)",
         "=IFERROR(IF(AVERAGE('Cash Flow Forecast'!B10:M10)>=0,"
         '"∞",Settings!$B$7/(-1*AVERAGE(\'Cash Flow Forecast\'!B10:M10))),'
         '"—")',
         "0.0", COLOR_PRIMARY, COLOR_PRIMARY),
        ("J", "L", "DAYS-CASH-ON-HAND",
         "=IFERROR(Settings!$B$7/("
         "(SUM('Recurring Expenses'!B7:M21)+SUMPRODUCT(--ISNUMBER('One-Off Expenses'!D6:D105),"
         "'One-Off Expenses'!D6:D105))/365),0)",
         "0", COLOR_ACCENT, COLOR_ACCENT),
    ]
    for first, last, label, formula, fmt, label_color, val_color in kpi_cards:
        ws.merge_cells(f"{first}32:{last}32")
        c = ws[f"{first}32"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=label_color)
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{first}33:{last}34")
        c = ws[f"{first}33"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=22, bold=True, color=val_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        ws.merge_cells(f"{first}35:{last}35")
        c = ws[f"{first}35"]
        c.value = {
            "AVG MONTHLY NET": "revenue − all expenses",
            "LOWEST MONTH": "cumulative cash floor",
            "RUNWAY (MONTHS)": "if avg net is negative",
            "DAYS-CASH-ON-HAND": "starting cash ÷ daily burn",
        }[label]
        c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center")

    # Gold borders around KPI cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 6: Secondary nav
    pseudo_button(ws, "A39", "C40", "Bookings + Revenue",
                   "'Bookings + Revenue'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Recurring Expenses",
                   "'Recurring Expenses'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "One-Off Expenses",
                   "'One-Off Expenses'!A1", variant="secondary")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # ZONE 7: Operational reminder callout
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📅 RE-FORECAST MONTHLY: replace projected booking $ with confirmed $ as guests book. "
        "The low month always shifts — don't trust a stale forecast."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 30

    # ZONE 8: Upgrade banner
    add_upgrade_banner(ws, 44)

    # ZONE 9: Footer
    brand_footer(ws, 46,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_bookings_revenue_tab(wb, variant):
    """Sheet 1 — Bookings + Revenue. Per-property 12-month matrix.

    Layout per property (block of 5 rows + 1 spacer):
      Row N   : property name header band
      Row N+1 : Confirmed bookings $ (input)
      Row N+2 : Projected occupancy % (input)
      Row N+3 : Projected ADR (input)
      Row N+4 : Projected revenue (=Occ × ADR × 30, formula)
      Row N+5 : Total monthly revenue (=Confirmed + Projected, formula, BOLD)
    Last block: Portfolio Total Revenue row.
    """
    ws = wb.create_sheet("Bookings + Revenue")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Bookings + Revenue",
                         prev_tab="Start", next_tab="Recurring Expenses")

    _band(ws, 4, 4, color=COLOR_BG_LIGHT)
    ws.merge_cells("A4:N4")
    c4 = ws["A4"]
    c4.value = ("Per property: confirmed $ + projected (occ % × ADR × 30 days). "
                "Update confirmed monthly as guests book.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32)] + [(get_column_letter(2 + i), 11) for i in range(12)] + [("N", 13)])

    headers = ["Line"] + MONTHS + ["Total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    bold_text = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    left_align = Alignment(horizontal="left", vertical="center", indent=1)

    BLOCK_HEIGHT = 6  # 1 header + 4 detail + 1 total = 6 rows then 1 spacer
    SPACER = 1
    base_row = 7
    property_total_rows = []

    for p_idx, prop in enumerate(PROPERTIES):
        block_top = base_row + p_idx * (BLOCK_HEIGHT + SPACER)

        # Property header band (gold-soft fill)
        ws.merge_cells(f"A{block_top}:N{block_top}")
        c = ws[f"A{block_top}"]
        c.value = f"  {prop.upper()}"
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[block_top].height = 20

        # Row +1: Confirmed bookings $
        confirmed_row = block_top + 1
        a = ws.cell(row=confirmed_row, column=1, value="Confirmed bookings ($)")
        a.font = bold_text
        a.alignment = left_align
        for m in range(12):
            cell = ws.cell(row=confirmed_row, column=2 + m,
                            value=_val(variant, DEMO_CONFIRMED[prop][m]))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        total = ws.cell(row=confirmed_row, column=14,
                         value=f"=SUM(B{confirmed_row}:M{confirmed_row})")
        apply_style(total, formula_cell_style())
        total.number_format = '"$"#,##0'
        ws.row_dimensions[confirmed_row].height = 18

        # Row +2: Projected occupancy %
        occ_row = block_top + 2
        a = ws.cell(row=occ_row, column=1, value="Projected occupancy %")
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = left_align
        for m in range(12):
            cell = ws.cell(row=occ_row, column=2 + m,
                            value=_val(variant, DEMO_OCC[prop][m]))
            apply_style(cell, input_cell_style())
            cell.number_format = "0%"
        total = ws.cell(row=occ_row, column=14,
                         value=f"=AVERAGE(B{occ_row}:M{occ_row})")
        apply_style(total, formula_cell_style())
        total.number_format = "0%"
        ws.row_dimensions[occ_row].height = 16

        # Row +3: Projected ADR
        adr_row = block_top + 3
        a = ws.cell(row=adr_row, column=1, value="Projected ADR ($)")
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = left_align
        for m in range(12):
            cell = ws.cell(row=adr_row, column=2 + m,
                            value=_val(variant, DEMO_ADR[prop][m]))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        total = ws.cell(row=adr_row, column=14,
                         value=f"=AVERAGE(B{adr_row}:M{adr_row})")
        apply_style(total, formula_cell_style())
        total.number_format = '"$"#,##0'
        ws.row_dimensions[adr_row].height = 16

        # Row +4: Projected revenue (formula: Occ × ADR × 30)
        proj_row = block_top + 4
        a = ws.cell(row=proj_row, column=1, value="Projected revenue (occ × ADR × 30)")
        a.font = italic_muted
        a.alignment = left_align
        for m in range(12):
            col_letter = get_column_letter(2 + m)
            cell = ws.cell(row=proj_row, column=2 + m,
                            value=f"={col_letter}{occ_row}*{col_letter}{adr_row}*30")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
        total = ws.cell(row=proj_row, column=14,
                         value=f"=SUM(B{proj_row}:M{proj_row})")
        apply_style(total, formula_cell_style())
        total.number_format = '"$"#,##0'
        ws.row_dimensions[proj_row].height = 16

        # Row +5: Total monthly revenue (Confirmed + Projected) — BOLD
        total_row = block_top + 5
        property_total_rows.append(total_row)
        a = ws.cell(row=total_row, column=1, value="TOTAL — property monthly revenue")
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        a.alignment = left_align
        for m in range(12):
            col_letter = get_column_letter(2 + m)
            cell = ws.cell(row=total_row, column=2 + m,
                            value=f"={col_letter}{confirmed_row}+{col_letter}{proj_row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell.font = Font(name=FONT_BODY, size=11, bold=True,
                              italic=True, color=COLOR_PRIMARY)
        total = ws.cell(row=total_row, column=14,
                         value=f"=SUM(B{total_row}:M{total_row})")
        apply_style(total, formula_cell_style())
        total.number_format = '"$"#,##0'
        total.font = Font(name=FONT_BODY, size=11, bold=True,
                           italic=True, color=COLOR_PRIMARY)
        ws.row_dimensions[total_row].height = 20

    # Portfolio total row (sum of each property's total row)
    portfolio_row = base_row + len(PROPERTIES) * (BLOCK_HEIGHT + SPACER) + 1
    a = ws.cell(row=portfolio_row, column=1, value="PORTFOLIO TOTAL REVENUE")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for m in range(12):
        col_letter = get_column_letter(2 + m)
        parts = "+".join([f"{col_letter}{r}" for r in property_total_rows])
        cell = ws.cell(row=portfolio_row, column=2 + m, value=f"={parts}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    total = ws.cell(row=portfolio_row, column=14,
                     value=f"=SUM(B{portfolio_row}:M{portfolio_row})")
    apply_style(total, formula_cell_style())
    total.number_format = '"$"#,##0'
    total.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[portfolio_row].height = 24

    # Stash the portfolio row reference for the Cash Flow Forecast tab
    ws["P1"] = portfolio_row  # hidden helper cell — referenced by build_cash_flow_forecast
    ws.column_dimensions["P"].hidden = True

    ws.freeze_panes = "B6"

    ws.print_area = f"A1:N{portfolio_row + 2}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return portfolio_row


def build_recurring_expenses_tab(wb, variant):
    """Sheet 2 — Recurring Expenses. 15 categories × 12 months matrix."""
    ws = wb.create_sheet("Recurring Expenses")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Recurring Expenses",
                         prev_tab="Bookings + Revenue", next_tab="One-Off Expenses")

    _band(ws, 4, 4, color=COLOR_BG_LIGHT)
    ws.merge_cells("A4:N4")
    c4 = ws["A4"]
    c4.value = ("One row per category. Same amount every month? Type once, copy across. "
                "Override any month that's different.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32)] + [(get_column_letter(2 + i), 11) for i in range(12)] + [("N", 13)])

    headers = ["Category"] + MONTHS + ["Total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 6

    # Categories start at row 7 — fixed range (B7:M21) referenced by Cash Flow Forecast
    for idx, cat in enumerate(RECURRING_CATEGORIES):
        row = 7 + idx
        a = ws.cell(row=row, column=1, value=cat)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

        for m in range(12):
            cell = ws.cell(row=row, column=2 + m,
                            value=_val(variant, DEMO_RECURRING[cat][m]))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'

        total = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(total, formula_cell_style())
        total.number_format = '"$"#,##0'

    # Total row
    last_cat_row = 7 + len(RECURRING_CATEGORIES) - 1  # 21
    tot_row = last_cat_row + 2  # 23
    ws.row_dimensions[last_cat_row + 1].height = 8

    a = ws.cell(row=tot_row, column=1, value="TOTAL RECURRING EXPENSES")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                        value=f"=SUM({col_letter}7:{col_letter}{last_cat_row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                          italic=True, color=COLOR_ERROR)
    ws.row_dimensions[tot_row].height = 22

    ws.freeze_panes = "B6"

    ws.print_area = f"A1:N{tot_row + 2}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_oneoff_expenses_tab(wb, variant):
    """Sheet 3 — One-Off Expenses log: Date | Property | Category | Amount | Notes."""
    ws = wb.create_sheet("One-Off Expenses")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "One-Off Expenses",
                         prev_tab="Recurring Expenses", next_tab="Cash Flow Forecast")

    _band(ws, 4, 4, color=COLOR_BG_LIGHT)
    ws.merge_cells("A4:G4")
    c4 = ws["A4"]
    c4.value = ("Planned 1-time expenses — renovations, appliance replacement, "
                "estimated tax payments, legal fees. Posted to Cash Flow Forecast by month.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 28), ("C", 28), ("D", 12),
        ("E", 32), ("F", 8), ("G", 8),
    ])

    headers = ["Date", "Property", "Category", "Amount", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    sample_rows = DEMO_ONEOFFS if variant == "demo" else []
    for i, item in enumerate(sample_rows, start=6):
        date_val, prop, cat, amount, notes = item

        a = ws.cell(row=i, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=prop)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=cat)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=amount)
        apply_style(d, input_cell_style())
        d.number_format = '"$"#,##0'

        e = ws.cell(row=i, column=5, value=notes if notes else None)
        apply_style(e, input_cell_style())

        ws.row_dimensions[i].height = 16

    last_data_row = len(sample_rows) + 5
    for row_idx in range(last_data_row + 1, 106):  # capacity to row 105 (100 entries)
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 4:
                cell.number_format = '"$"#,##0'

    # Property dropdown sourced from Settings
    add_dropdown(ws, "B6:B105", "=Settings!$B$10:$B$19")

    ws.freeze_panes = "A6"

    ws.print_area = "A1:E50"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_cash_flow_forecast_tab(wb, portfolio_revenue_row):
    """Sheet 4 — Cash Flow Forecast. Combined 6-row matrix + chart.

    Row 5 : header (Month columns)
    Row 7 : Revenue (= Bookings + Revenue!portfolio_row)
    Row 8 : Recurring expenses (= Recurring Expenses!tot_row)
    Row 9 : One-off expenses (=SUMPRODUCT for that month)
    Row 10: NET monthly (=Revenue - Recurring - One-off)
    Row 11: Cumulative cash (running total starting from Settings!B7)
    """
    ws = wb.create_sheet("Cash Flow Forecast")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    compact_header_band(ws, "Cash Flow Forecast",
                         prev_tab="One-Off Expenses", next_tab="Settings")

    _band(ws, 4, 4, color=COLOR_BG_LIGHT)
    ws.merge_cells("A4:N4")
    c4 = ws["A4"]
    c4.value = ("Combined 12-month forecast. Red bars = months where cumulative cash is "
                "negative. The chart on the right is your decision-support view.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32)] + [(get_column_letter(2 + i), 11) for i in range(12)] + [("N", 13)])

    headers = ["Line"] + MONTHS + ["Total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 6

    bold_text = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    left_align = Alignment(horizontal="left", vertical="center", indent=1)

    # Row 7: Revenue (pulled from portfolio total row on Bookings + Revenue)
    a = ws.cell(row=7, column=1, value="Revenue (portfolio total)")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a.alignment = left_align
    for m in range(12):
        col_letter = get_column_letter(2 + m)
        cell = ws.cell(row=7, column=2 + m,
                        value=f"='Bookings + Revenue'!{col_letter}{portfolio_revenue_row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
    total = ws.cell(row=7, column=14, value="=SUM(B7:M7)")
    apply_style(total, formula_cell_style())
    total.number_format = '"$"#,##0'
    total.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 20

    # Row 8: Recurring expenses (from Recurring Expenses tab total row 23)
    a = ws.cell(row=8, column=1, value="Recurring expenses")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_ERROR)
    a.alignment = left_align
    for m in range(12):
        col_letter = get_column_letter(2 + m)
        cell = ws.cell(row=8, column=2 + m,
                        value=f"='Recurring Expenses'!{col_letter}23")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
    total = ws.cell(row=8, column=14, value="=SUM(B8:M8)")
    apply_style(total, formula_cell_style())
    total.number_format = '"$"#,##0'
    ws.row_dimensions[8].height = 20

    # Row 9: One-off expenses — SUMPRODUCT by month from One-Off Expenses tab
    a = ws.cell(row=9, column=1, value="One-off expenses")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_ERROR)
    a.alignment = left_align
    for m in range(12):
        month_num = m + 1
        formula = (
            f"=SUMPRODUCT("
            f"(MONTH('One-Off Expenses'!$A$6:$A$105)={month_num})*"
            f"(YEAR('One-Off Expenses'!$A$6:$A$105)=Settings!$B$5)*"
            f"IFERROR('One-Off Expenses'!$D$6:$D$105,0))"
        )
        cell = ws.cell(row=9, column=2 + m, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
    total = ws.cell(row=9, column=14, value="=SUM(B9:M9)")
    apply_style(total, formula_cell_style())
    total.number_format = '"$"#,##0'
    ws.row_dimensions[9].height = 20

    # Row 10: NET monthly
    a = ws.cell(row=10, column=1, value="NET monthly cash flow")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = left_align
    for m in range(12):
        col_letter = get_column_letter(2 + m)
        cell = ws.cell(row=10, column=2 + m,
                        value=f"={col_letter}7-{col_letter}8-{col_letter}9")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                          italic=True, color=COLOR_PRIMARY)
    total = ws.cell(row=10, column=14, value="=SUM(B10:M10)")
    apply_style(total, formula_cell_style())
    total.number_format = '"$"#,##0'
    total.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[10].height = 22

    # Row 11: Cumulative cash (running total starting from Settings!B7)
    a = ws.cell(row=11, column=1, value="CUMULATIVE CASH ON HAND")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    a.alignment = left_align
    # B11 = Settings!B7 + B10
    ws.cell(row=11, column=2, value="=Settings!$B$7+B10")
    for m in range(1, 12):
        col_letter = get_column_letter(2 + m)
        prev_letter = get_column_letter(1 + m)
        ws.cell(row=11, column=2 + m,
                 value=f"={prev_letter}11+{col_letter}10")
    for m in range(12):
        cell = ws.cell(row=11, column=2 + m)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                          italic=True, color=COLOR_PRIMARY)
    # Total column shows the year-end cash
    end_cell = ws.cell(row=11, column=14, value="=M11")
    apply_style(end_cell, formula_cell_style())
    end_cell.number_format = '"$"#,##0'
    end_cell.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[11].height = 26

    # Conditional formatting:
    #   Cumulative cash row red where < 0 (per brief — primary risk indicator)
    #   NET monthly row also flagged where < 0
    red_fill = PatternFill("solid", fgColor="FFCCCC")
    green_fill = PatternFill("solid", fgColor="C7EFCF")
    ws.conditional_formatting.add(
        "B11:M11",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        "B10:M10",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        "B10:M10",
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill),
    )

    # ------------------------------------------------------------------
    # Charts: column (revenue + expenses) + line overlay (cumulative cash)
    # Anchored to the right of the data table at column P, row 5.
    # ------------------------------------------------------------------
    # Column chart: Revenue (row 7) + Recurring (row 8) + One-off (row 9)
    bar = BarChart()
    bar.type = "col"
    bar.style = 2
    bar.title = "Monthly Revenue vs Expenses"
    bar.y_axis.title = "Dollars"
    bar.x_axis.title = "Month"
    bar.height = 11
    bar.width = 22

    # Data: 3 series (rev, recurring, one-off) — rows 7-9, cols B-M
    data_ref = Reference(ws, min_col=1, min_row=7, max_row=9, max_col=13)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_col=13, max_row=5)
    bar.add_data(data_ref, titles_from_data=True, from_rows=True)
    bar.set_categories(cats_ref)
    bar.grouping = "clustered"

    # Line overlay: cumulative cash (row 11)
    line = LineChart()
    line_data = Reference(ws, min_col=1, min_row=11, max_row=11, max_col=13)
    line.add_data(line_data, titles_from_data=True, from_rows=True)
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"  # second axis on the right

    # Combine charts (openpyxl: bar.+= overlay via _charts)
    bar += line

    style_chart(bar)
    ws.add_chart(bar, "P5")

    ws.freeze_panes = "B6"

    # Print: just the data table (chart anchored further right; let it print on a second page)
    ws.print_area = "A1:N13"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 5 — Settings: active year, starting cash, property + recurring lists."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Cash Flow Forecast", next_tab=None)

    _band(ws, 4, 4, color=COLOR_BG_LIGHT)
    ws.merge_cells("A4:D4")
    c4 = ws["A4"]
    c4.value = ("Active year · starting cash balance · property list · category list")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 36), ("B", 24), ("C", 4), ("D", 28)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    right_align = Alignment(horizontal="right", vertical="center")

    # B5: Active year
    a5 = ws.cell(row=5, column=1, value="Active year:")
    a5.font = bold_right
    a5.alignment = right_align
    b5 = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = ("Drives the SUMPRODUCT date filter on the One-Off Expenses roll-up. "
                "Bump each January.")
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 24

    # B7: Starting cash balance
    a7 = ws.cell(row=7, column=1, value="Starting cash balance (Jan 1):")
    a7.font = bold_right
    a7.alignment = right_align
    b7 = ws.cell(row=7, column=2, value=_val(variant, STARTING_CASH))
    apply_style(b7, input_cell_style())
    b7.number_format = '"$"#,##0'
    ws.row_dimensions[7].height = 18

    ws.merge_cells("A8:D8")
    c8 = ws["A8"]
    c8.value = ("Operating cash on hand at the start of the year (checking + savings "
                "earmarked for the rentals). The cumulative-cash row builds from here.")
    c8.font = italic_muted
    c8.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[8].height = 28

    # Property list (rows 10-19) — 10-slot capacity, B10:B19
    sect = ws.cell(row=9, column=1, value="Property list (drives One-Off dropdown)")
    sect.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[9].height = 22

    demo_props = (PROPERTIES + [None] * 10)[:10] if variant == "demo" else [None] * 10
    for idx, prop in enumerate(demo_props):
        row = 10 + idx
        cell = ws.cell(row=row, column=2, value=prop)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[20].height = 6

    # Recurring expense category list (rows 21-35)
    sect21 = ws.cell(row=21, column=1, value="Recurring expense categories")
    sect21.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[21].height = 22

    cat_help = ws.cell(row=22, column=1,
                        value="(Reference list — Recurring Expenses tab carries these)")
    cat_help.font = italic_muted
    ws.row_dimensions[22].height = 16

    for idx, cat in enumerate(RECURRING_CATEGORIES):
        row = 23 + idx
        cell = ws.cell(row=row, column=1, value=cat)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16

    last_cat_row = 23 + len(RECURRING_CATEGORIES) - 1  # 37

    # Footer
    brand_footer(ws, last_cat_row + 3,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    ws.print_area = f"A1:D{last_cat_row + 5}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    portfolio_row = build_bookings_revenue_tab(wb, variant)
    build_recurring_expenses_tab(wb, variant)
    build_oneoff_expenses_tab(wb, variant)
    build_cash_flow_forecast_tab(wb, portfolio_row)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"12-Month Cash Flow Forecaster{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "12-month rolling cash flow forecast with confirmed + projected revenue, "
        "recurring + one-off expenses, and cumulative cash runway (v2.2)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
