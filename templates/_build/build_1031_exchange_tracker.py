"""Build ACQ-011 1031 Exchange Tracker (STR -> STR) — operational deadline tracker.

Implements templates/_briefs/ACQ-011-1031-exchange-tracker.md.

Generates:
  templates/_masters/ACQ-011-1031-exchange-tracker-DEMO.xlsx
  templates/_masters/ACQ-011-1031-exchange-tracker-BLANK.xlsx

Tabs (6):
  0  Start                  — hero + 45/180-day countdown banners + nav
  1  Relinquished Property  — sale details, basis, realized gain, deadlines
  2  Identification         — 3-property / 200% / 95% rule paths
  3  Replacement + Closing  — closing details, boot, deferred gain
  4  Carryover Basis        — new property tax basis carryover
  5  Settings               — year, QI info, CPA contact, sale date

Usage:
    python build_1031_exchange_tracker.py
"""
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_ACCENT, COLOR_SECONDARY, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, compact_header_band, apply_brand_header, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
)

SKU = "ACQ-011"
NAME = "1031-exchange-tracker"
VERSION_LINE = f"{SKU} · v2.3 · Free updates forever"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Reference sale date for DEMO (drives all countdowns).
DEMO_SALE_DATE = date(2026, 4, 15)


def _val(variant, demo_value):
    """Return demo_value when building DEMO, None for BLANK."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _apply_countdown_cf(ws, cell_range):
    """Red <7, gold 7-21, parchment >21 conditional formatting on a numeric
    days-remaining cell range."""
    # Red — critical (less than 7 days)
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["7"],
            fill=PatternFill("solid", fgColor=COLOR_ERROR),
            font=Font(name=FONT_HEAD, size=12, bold=True, color="FFFFFF"),
        ),
    )
    # Gold — warning (7 to 21 inclusive)
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=["7", "21"],
            fill=PatternFill("solid", fgColor=COLOR_ACCENT),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY),
        ),
    )
    # Parchment — safe (greater than 21)
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["21"],
            fill=PatternFill("solid", fgColor=COLOR_BG_LIGHT),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY),
        ),
    )


# ---------------------------------------------------------------------------
# Tab 0 — Start
# ---------------------------------------------------------------------------
def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — navy hero (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "1031 Exchange Tracker"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Two deadlines. Zero forgiveness. Don't miss them."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT — closest deadline summary
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(\'Settings\'!B11="",'
        '"→  Enter sale date in Settings to start the countdown.",'
        '"→  ID deadline in "&\'Relinquished Property\'!B16'
        '&" days  ·  Closing deadline in "&\'Relinquished Property\'!B17&" days")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — Countdown banners (rows 10-15) — two side-by-side cards
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card A: 45-DAY ID countdown (cols A-F)
    ws.merge_cells("A10:F10")
    c = ws["A10"]
    c.value = "45-DAY IDENTIFICATION"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:F13")
    c = ws["A11"]
    c.value = "='Relinquished Property'!B16"
    c.font = Font(name=FONT_HEAD, size=42, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '0" days"'
    ws.row_dimensions[11].height = 28
    ws.row_dimensions[12].height = 28
    ws.row_dimensions[13].height = 28

    ws.merge_cells("A14:F14")
    c = ws["A14"]
    c.value = "='Relinquished Property'!B14"
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"due "yyyy-mm-dd'
    ws.row_dimensions[14].height = 18

    ws.merge_cells("A15:F15")
    c = ws["A15"]
    c.value = "Identify replacement properties by this date"
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Card B: 180-DAY closing countdown (cols G-L)
    ws.merge_cells("G10:L10")
    c = ws["G10"]
    c.value = "180-DAY CLOSING"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G11:L13")
    c = ws["G11"]
    c.value = "='Relinquished Property'!B17"
    c.font = Font(name=FONT_HEAD, size=42, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '0" days"'

    ws.merge_cells("G14:L14")
    c = ws["G14"]
    c.value = "='Relinquished Property'!B15"
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"due "yyyy-mm-dd'

    ws.merge_cells("G15:L15")
    c = ws["G15"]
    c.value = "Close on replacement(s) by this date"
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Conditional formatting on countdown days cells (red/gold/parchment fills)
    _apply_countdown_cf(ws, "A11:F13")
    _apply_countdown_cf(ws, "G11:L13")

    # Gold border around each countdown card
    gold_side = Side(style="medium", color=COLOR_ACCENT)
    for first_col, last_col in [(1, 6), (7, 12)]:
        for r in range(10, 16):
            for col in range(first_col, last_col + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == first_col else existing.left,
                    right=gold_side if col == last_col else existing.right,
                )

    # ZONE 3 — Disclaimer banner (rows 17-19)
    error_fill = PatternFill("solid", fgColor=COLOR_ERROR)
    for r in range(17, 20):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = error_fill
    ws.merge_cells("A17:L19")
    c = ws["A17"]
    c.value = (
        "⚠  1031 EXCHANGE RULES ARE STRICT AND UNFORGIVING.\n"
        "This workbook is recordkeeping support only. Your Qualified Intermediary (QI) and "
        "CPA execute the exchange — not this spreadsheet. The exchange is VOID without a QI "
        "in place before sale closing. Miss the 45-day ID or 180-day closing window by ONE "
        "DAY and the entire deferral fails — you owe full tax on the gain."
    )
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[17].height = 22
    ws.row_dimensions[18].height = 22
    ws.row_dimensions[19].height = 22

    # ZONE 4 — How to use (rows 21-27)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(21, 28):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A21:L21")
    c = ws["A21"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[21].height = 20
    quickstart = [
        "① Settings — enter sale date, QI info (REQUIRED), CPA contact.",
        "② Relinquished Property — sale price, basis, depreciation -> realized gain.",
        "③ Identification — pick a rule (3-property / 200% / 95%), list candidates by Day 45.",
        "④ Replacement + Closing — log each closing, boot received, deferred gain.",
        "⑤ Carryover Basis — read the new property's depreciation basis for next year.",
    ]
    for i, item in enumerate(quickstart):
        row = 22 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # ZONE 5 — Primary CTA (rows 29-32)
    pseudo_button(ws, "A29", "L32", "→  GO TO RELINQUISHED PROPERTY  →",
                   "'Relinquished Property'!A5", variant="primary")
    for r in range(29, 33):
        ws.row_dimensions[r].height = 22

    # ZONE 6 — Secondary nav (rows 34-35)
    pseudo_button(ws, "A34", "C35", "Identification",
                   "'Identification'!A1", variant="secondary")
    pseudo_button(ws, "D34", "F35", "Replacement + Closing",
                   "'Replacement + Closing'!A1", variant="secondary")
    pseudo_button(ws, "G34", "I35", "Carryover Basis",
                   "'Carryover Basis'!A1", variant="secondary")
    pseudo_button(ws, "J34", "L35", "Settings",
                   "'Settings'!A1", variant="accent")
    ws.row_dimensions[34].height = 22
    ws.row_dimensions[35].height = 22

    # Footer
    brand_footer(ws, 37, version_line=VERSION_LINE)

    ws.print_area = "A1:L40"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 1 — Relinquished Property
# ---------------------------------------------------------------------------
def build_relinquished_tab(wb, variant):
    ws = wb.create_sheet("Relinquished Property")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Relinquished Property",
                         prev_tab="Start", next_tab="Identification")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Sale of the property you're exchanging out of. Sale date drives both deadlines.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 38), ("B", 22), ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Section 1 — property + sale (rows 5-12)
    a = ws.cell(row=5, column=1, value="PROPERTY + SALE")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A5:L5")
    ws.row_dimensions[5].height = 20

    fields_top = [
        # (row, label, value_for_demo, format, kind)
        (6,  "Property name:",       _val(variant, "Lakehouse on Pine"),       None,        "text"),
        (7,  "Address:",             _val(variant, "412 Pine Cove Dr, Hendersonville, TN"), None, "text"),
        (8,  "Sale date:",           "='Settings'!B11",                         "yyyy-mm-dd", "formula"),
        (9,  "Sale price:",          _val(variant, 625000),                    '"$"#,##0',  "money"),
        (10, "Selling expenses (commissions, closing):", _val(variant, 37500), '"$"#,##0',  "money"),
        (11, "Net sale proceeds:",   "=B9-B10",                                '"$"#,##0',  "formula"),
    ]
    for r, label, val, fmt, kind in fields_top:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        if kind == "formula":
            apply_style(c, formula_cell_style())
        else:
            apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Section 2 — Basis + gain (rows 12-...)
    ws.row_dimensions[12].height = 8

    fields_basis = [
        (13, "Original cost basis:",                _val(variant, 390000),   '"$"#,##0', "money"),
        # rows 14/15/16/17 are deadline cells — see below — defer basis flow
    ]

    # We compute deadlines first because countdown cells need to be at known
    # rows for the Start tab to reference (B14, B15, B16, B17).
    # Layout decided:
    #   Row 13: Original cost basis (input)
    #   Row 14: 45-day ID deadline (formula = B8 + 45)
    #   Row 15: 180-day closing deadline (formula = B8 + 180)
    #   Row 16: Days remaining ID (formula)
    #   Row 17: Days remaining closing (formula)
    #   Row 18+: continued basis/gain calc
    for r, label, val, fmt, kind in fields_basis:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Deadline rows (14, 15) and countdowns (16, 17)
    a = ws.cell(row=14, column=1, value="45-day ID deadline:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=14, column=2, value="=B8+45")
    apply_style(c, formula_cell_style())
    c.number_format = "yyyy-mm-dd"
    ws.row_dimensions[14].height = 18

    a = ws.cell(row=15, column=1, value="180-day closing deadline:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=15, column=2, value="=B8+180")
    apply_style(c, formula_cell_style())
    c.number_format = "yyyy-mm-dd"
    ws.row_dimensions[15].height = 18

    a = ws.cell(row=16, column=1, value="Days remaining — ID deadline:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=16, column=2, value="=B14-TODAY()")
    apply_style(c, formula_cell_style())
    c.number_format = '0" days"'
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[16].height = 22

    a = ws.cell(row=17, column=1, value="Days remaining — closing deadline:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=17, column=2, value="=B15-TODAY()")
    apply_style(c, formula_cell_style())
    c.number_format = '0" days"'
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[17].height = 22

    # CF on the countdown cells
    _apply_countdown_cf(ws, "B16:B17")

    # Section 3 — basis adj + gain calc (rows 19-)
    ws.row_dimensions[18].height = 8
    a = ws.cell(row=19, column=1, value="BASIS + REALIZED GAIN")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A19:L19")
    ws.row_dimensions[19].height = 20

    gain_rows = [
        (20, "Accumulated depreciation:",     _val(variant, 80000),  '"$"#,##0', "money"),
        (21, "Adjusted basis:",                "=B13-B20",           '"$"#,##0', "formula"),
        (22, "Realized gain (sale proceeds − adjusted basis):",
                                               "=B11-B21",           '"$"#,##0', "formula"),
        (23, "Depreciation recapture portion (taxed up to 25%):",
                                               "=MIN(B20,B22)",      '"$"#,##0', "formula"),
        (24, "Capital gain portion (taxed up to 20%+NIIT):",
                                               "=MAX(B22-B23,0)",    '"$"#,##0', "formula"),
    ]
    for r, label, val, fmt, kind in gain_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        if kind == "formula":
            apply_style(c, formula_cell_style())
        else:
            apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        if r == 22:
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[r].height = 18

    # Callout (row 26)
    ws.row_dimensions[25].height = 8
    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = (
        "ⓘ  In a successful 1031 exchange, this realized gain is DEFERRED — you don't pay tax now. "
        "But the gain still exists; it carries into the new property's basis. The IRS gets paid eventually."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[26].height = 32

    ws.print_area = "A1:L28"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 2 — Identification
# ---------------------------------------------------------------------------
def build_identification_tab(wb, variant):
    ws = wb.create_sheet("Identification")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Identification",
                         prev_tab="Relinquished Property",
                         next_tab="Replacement + Closing")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("List candidate replacements by Day 45. Pick ONE rule: 3-property, 200%, or 95%.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6), ("B", 28), ("C", 32), ("D", 16), ("E", 18),
        ("F", 6), ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Rule selector (row 5-6)
    a = ws.cell(row=5, column=1, value="Rule:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    rule_cell = ws.cell(row=5, column=2,
                        value=_val(variant, "3-property rule"))
    apply_style(rule_cell, input_cell_style())
    rule_cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    add_dropdown(ws, "B5",
                  ["3-property rule", "200% rule", "95% rule"])
    ws.row_dimensions[5].height = 22

    ws.merge_cells("C5:E5")
    c = ws["C5"]
    c.value = (
        '=IF(B5="3-property rule","Up to 3 properties; value doesn\'t matter.",'
        'IF(B5="200% rule","Any number; total FMV ≤ 200% of relinquished sale price.",'
        'IF(B5="95% rule","Any number; must close on ≥95% of identified value.","")))'
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    ws.row_dimensions[6].height = 8

    # Header row 7
    headers = ["#", "Property name", "Address", "Asking price", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 20

    # Demo data — 3 properties
    demo_props = [
        ("Smokies Ridge Cabin", "1142 Smoky Top Ln, Gatlinburg, TN", 485000, "Identified"),
        ("Creek Side Cottage",  "78 Whisper Creek Rd, Asheville, NC", 360000, "Under contract"),
        ("Ridge Top Lookout",   "9 Ridge Top Way, Banner Elk, NC",    545000, "Backup"),
    ]

    # Up to 8 rows for properties (rows 8-15) — supports 3-property rule
    # primary 3 plus extras for 200%/95% paths.
    for i in range(8):
        r = 8 + i
        # # column
        n = ws.cell(row=r, column=1, value=i + 1)
        n.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
        n.alignment = Alignment(horizontal="center", vertical="center")
        # name
        if variant == "demo" and i < len(demo_props):
            name_v, addr_v, ask_v, status_v = demo_props[i]
        else:
            name_v = addr_v = ask_v = status_v = None
        c = ws.cell(row=r, column=2, value=name_v)
        apply_style(c, input_cell_style())
        c = ws.cell(row=r, column=3, value=addr_v)
        apply_style(c, input_cell_style())
        c = ws.cell(row=r, column=4, value=ask_v)
        apply_style(c, input_cell_style())
        c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=5, value=status_v)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    add_dropdown(ws, "E8:E15",
                  ["Identified", "Under contract", "Closed", "Backup", "Withdrawn"])

    # Aggregate calcs (rows 17-)
    ws.row_dimensions[16].height = 8
    a = ws.cell(row=17, column=1, value="AGGREGATES + RULE CHECK")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A17:L17")
    ws.row_dimensions[17].height = 20

    agg_rows = [
        (18, "Count of identified properties:",
             '=COUNTA(B8:B15)', "0"),
        (19, "Aggregate asking value:",
             "=SUM(D8:D15)", '"$"#,##0'),
        (20, "Closed value (status=Closed):",
             '=SUMIFS(D8:D15,E8:E15,"Closed")', '"$"#,##0'),
        (21, "Relinquished sale price (reference):",
             "='Relinquished Property'!B9", '"$"#,##0'),
        (22, "200% cap (2 × relinquished sale price):",
             "=B21*2", '"$"#,##0'),
        (23, "95% threshold (of aggregate identified):",
             "=B19*0.95", '"$"#,##0'),
    ]
    for r, label, formula, fmt in agg_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Compliance flag (row 25)
    ws.row_dimensions[24].height = 8
    a = ws.cell(row=25, column=1, value="Rule compliance:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=25, column=2)
    c.value = (
        '=IF(B5="3-property rule",'
        'IF(B18<=3,"✓ OK — within 3-property limit","✗ OVER — reduce to 3 or switch rule"),'
        'IF(B5="200% rule",'
        'IF(B19<=B22,"✓ OK — within 200% cap","✗ OVER — aggregate exceeds 200% cap"),'
        'IF(B5="95% rule",'
        'IF(B20>=B23,"✓ OK — closed value meets 95% threshold","✗ PENDING — must close on ≥95% of identified"),'
        '"Pick a rule above")))'
    )
    apply_style(c, formula_cell_style())
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[25].height = 24

    # Callout (row 27)
    ws.row_dimensions[26].height = 8
    ws.merge_cells("A27:L27")
    c = ws["A27"]
    c.value = (
        "ⓘ  Identification must be in writing, signed, and delivered to your QI by Day 45. "
        "Verbal identification doesn't count. After Day 45, you can ONLY close on identified properties — "
        "no swaps, no substitutions."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[27].height = 32

    ws.freeze_panes = "A8"
    ws.print_area = "A1:E28"
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


# ---------------------------------------------------------------------------
# Tab 3 — Replacement + Closing
# ---------------------------------------------------------------------------
def build_replacement_tab(wb, variant):
    ws = wb.create_sheet("Replacement + Closing")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Replacement + Closing",
                         prev_tab="Identification",
                         next_tab="Carryover Basis")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Closings on selected replacement(s). Boot received = recognized gain; the rest is deferred.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 38), ("B", 22), ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Section header row 5
    a = ws.cell(row=5, column=1, value="REPLACEMENT PROPERTY (combined if multiple)")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A5:L5")
    ws.row_dimensions[5].height = 20

    rep_rows = [
        (6,  "Selected replacement(s):",
             _val(variant, "Smokies Ridge Cabin + Creek Side Cottage"),
             None, "text"),
        (7,  "Total purchase price:",        _val(variant, 845000),  '"$"#,##0', "money"),
        (8,  "Closing date:",                _val(variant, date(2026, 9, 20)), "yyyy-mm-dd", "date"),
        (9,  "Closing costs (paid by buyer):", _val(variant, 18000), '"$"#,##0', "money"),
        (10, "Additional cash paid in (boot paid):",
             _val(variant, 220000), '"$"#,##0', "money"),
        (11, "Mortgage debt assumed on replacement:",
             _val(variant, 0), '"$"#,##0', "money"),
        (12, "Mortgage debt relieved on relinquished:",
             _val(variant, 0), '"$"#,##0', "money"),
    ]
    for r, label, val, fmt, kind in rep_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Boot + deferred gain section (rows 14-)
    ws.row_dimensions[13].height = 8
    a = ws.cell(row=14, column=1, value="BOOT + DEFERRED GAIN")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A14:L14")
    ws.row_dimensions[14].height = 20

    boot_rows = [
        (15, "Net debt relief (relinquished mortgage − assumed mortgage):",
             "=MAX(B12-B11,0)", '"$"#,##0'),
        (16, "Cash boot received (proceeds not reinvested):",
             "=MAX('Relinquished Property'!B11-B7,0)", '"$"#,##0'),
        (17, "Total boot received (cash + debt relief):",
             "=B15+B16", '"$"#,##0'),
        (18, "Realized gain (from Relinquished tab):",
             "='Relinquished Property'!B22", '"$"#,##0'),
        (19, "Recognized gain (= total boot, capped at realized):",
             "=MIN(B17,B18)", '"$"#,##0'),
        (20, "Recapture portion of recognized gain (taxed first, up to 25%):",
             "=MIN(B19,'Relinquished Property'!B23)", '"$"#,##0'),
        (21, "Capital-gain portion of recognized gain (up to 20%+NIIT):",
             "=MAX(B19-B20,0)", '"$"#,##0'),
        (22, "DEFERRED GAIN (rolls into new basis):",
             "=B18-B19", '"$"#,##0'),
    ]
    for r, label, formula, fmt in boot_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = fmt
        if r == 22:
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color="F6EFE2")
            c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        if r == 19:
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[r].height = 18

    # Estimated tax on recognized gain (informational, rows 24-)
    ws.row_dimensions[23].height = 8
    a = ws.cell(row=24, column=1, value="EST. TAX ON RECOGNIZED GAIN (informational)")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A24:L24")
    ws.row_dimensions[24].height = 20

    tax_rows = [
        (25, "Recapture tax (25% × recapture portion):",
             "=ROUND(B20*0.25,0)", '"$"#,##0'),
        (26, "Capital gain tax (20% × capital gain portion):",
             "=ROUND(B21*0.20,0)", '"$"#,##0'),
        (27, "NIIT (3.8% × capital gain portion, if AGI threshold):",
             "=ROUND(B21*0.038,0)", '"$"#,##0'),
        (28, "Total est. federal tax due on boot:",
             "=B25+B26+B27", '"$"#,##0'),
    ]
    for r, label, formula, fmt in tax_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = fmt
        if r == 28:
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[r].height = 18

    # Callout (row 30)
    ws.row_dimensions[29].height = 8
    ws.merge_cells("A30:L30")
    c = ws["A30"]
    c.value = (
        "ⓘ  Boot = anything you receive that isn't like-kind real estate. Cash, debt relief, "
        "personal property. Recapture is taxed FIRST (up to 25%), then capital gain (up to 20% + 3.8% NIIT). "
        "Your CPA confirms the tax — these are estimates only."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[30].height = 36

    ws.print_area = "A1:L31"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 4 — Carryover Basis
# ---------------------------------------------------------------------------
def build_carryover_tab(wb, variant):
    ws = wb.create_sheet("Carryover Basis")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Carryover Basis",
                         prev_tab="Replacement + Closing",
                         next_tab="Settings")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("New property's tax basis carries from the relinquished property — not its purchase price.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 44), ("B", 22), ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Section header row 5
    a = ws.cell(row=5, column=1, value="CARRYOVER BASIS BUILD-UP")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A5:L5")
    ws.row_dimensions[5].height = 20

    cb_rows = [
        (6,  "Relinquished adjusted basis:",
             "='Relinquished Property'!B21", '"$"#,##0'),
        (7,  "+ Boot paid (additional cash in):",
             "='Replacement + Closing'!B10", '"$"#,##0'),
        (8,  "+ Closing costs added to basis:",
             "='Replacement + Closing'!B9", '"$"#,##0'),
        (9,  "+ Net new debt (assumed − relieved):",
             "=MAX('Replacement + Closing'!B11-'Replacement + Closing'!B12,0)",
             '"$"#,##0'),
        (10, "− Boot received (cash + debt relief):",
             "='Replacement + Closing'!B17", '"$"#,##0'),
        (11, "+ Recognized gain (boot recognized now):",
             "='Replacement + Closing'!B19", '"$"#,##0'),
        (12, "= NEW PROPERTY TAX BASIS (for depreciation):",
             "=B6+B7+B8+B9-B10+B11", '"$"#,##0'),
    ]
    for r, label, formula, fmt in cb_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = fmt
        if r == 12:
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color="F6EFE2")
            c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 20

    # Holding period section (rows 14-)
    ws.row_dimensions[13].height = 8
    a = ws.cell(row=14, column=1, value="HOLDING PERIOD")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A14:L14")
    ws.row_dimensions[14].height = 20

    hp_rows = [
        (15, "Relinquished property acquired:",
             _val(variant, date(2018, 6, 12)), "yyyy-mm-dd", "date"),
        (16, "Relinquished sale date:",
             "='Relinquished Property'!B8", "yyyy-mm-dd", "formula"),
        (17, "Replacement closing date:",
             "='Replacement + Closing'!B8", "yyyy-mm-dd", "formula"),
        (18, "Carryover holding period start (= acquired date above):",
             "=B15", "yyyy-mm-dd", "formula"),
        (19, "Years held (combined, as of replacement close):",
             "=ROUND((B17-B15)/365.25,1)", '0.0" yrs"', "formula"),
    ]
    for r, label, val, fmt, kind in hp_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        if kind == "formula":
            apply_style(c, formula_cell_style())
        else:
            apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Depreciation note (rows 21-)
    ws.row_dimensions[20].height = 8
    a = ws.cell(row=21, column=1, value="DEPRECIATION HANDOFF (for next year's P&L)")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A21:L21")
    ws.row_dimensions[21].height = 20

    dep_rows = [
        (22, "Depreciable building portion (typical 80%):",
             _val(variant, 0.80), "0%", "input"),
        (23, "Building basis subject to depreciation:",
             "=B12*B22", '"$"#,##0', "formula"),
        (24, "MACRS recovery (residential rental — 27.5 yr):",
             "=ROUND(B23/27.5,0)", '"$"#,##0" / yr"', "formula"),
    ]
    for r, label, val, fmt, kind in dep_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        if kind == "formula":
            apply_style(c, formula_cell_style())
        else:
            apply_style(c, input_cell_style())
        c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Callout (row 26)
    ws.row_dimensions[25].height = 8
    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = (
        "ⓘ  Your new property's depreciation basis is NOT what you paid — it's the relinquished basis "
        "plus adjustments. This is the IRS price for deferring the gain. Confirm with your CPA before "
        "filing Form 4562 next year."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[26].height = 36

    ws.print_area = "A1:L28"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 5 — Settings
# ---------------------------------------------------------------------------
def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Settings",
                         prev_tab="Carryover Basis", next_tab=None)

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Active year, sale date (drives countdowns), and your QI + CPA contacts.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 38), ("B", 36), ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    fields = [
        # (row, label, value, format)
        (5,  "Active tax year:",                  2026,                                     "0"),
        (6,  "Filing status:",                    "MFJ",                                    None),
        (7,  "Qualified Intermediary (QI) name:", _val(variant, "First American Exchange"), None),
        (8,  "QI contact (phone / email):",       _val(variant, "(866) 472-1031 / qi@firstam.com"), None),
        (9,  "CPA name + contact:",               _val(variant, "Riverstone CPA — (615) 555-0144"), None),
        (10, "CPA email:",                        _val(variant, "tax@riverstonecpa.com"),  None),
        (11, "Sale date (drives all countdowns):", _val(variant, DEMO_SALE_DATE),           "yyyy-mm-dd"),
    ]
    for r, label, val, fmt in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        if r == 11:
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[r].height = 20

    add_dropdown(ws, "B6", ["Single", "MFJ", "HoH", "MFS"])

    # QI WARNING banner (rows 13-15)
    ws.row_dimensions[12].height = 8
    error_fill = PatternFill("solid", fgColor=COLOR_ERROR)
    for r in range(13, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = error_fill
    ws.merge_cells("A13:L15")
    c = ws["A13"]
    c.value = (
        "⚠  REQUIRED: A Qualified Intermediary (QI) must be in place BEFORE the relinquished sale closes. "
        "If you receive sale proceeds directly — even for one day — the exchange is VOID and the entire "
        "gain is taxable. The QI holds proceeds in escrow and transfers them to the replacement closing. "
        "This workbook tracks the timeline; the QI executes the exchange."
    )
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Disclaimer block (rows 17-19)
    ws.row_dimensions[16].height = 8
    parchment_alt = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(17, 20):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_alt
    ws.merge_cells("A17:L19")
    c = ws["A17"]
    c.value = (
        "Disclaimer: This workbook is recordkeeping and timeline support for a 1031 like-kind exchange "
        "under IRC §1031. It is not legal, tax, or financial advice. The IRS requires strict adherence to "
        "the 45-day identification window, the 180-day closing window, the QI safe harbor (Treas. Reg. §1.1031(k)-1), "
        "and the like-kind requirement (real estate held for investment or productive use only since the 2017 TCJA). "
        "Your QI and CPA execute the exchange. Confirm every figure with them before signing closing documents."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # Footer
    brand_footer(ws, 22, version_line=VERSION_LINE)

    ws.print_area = "A1:L25"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------
def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_relinquished_tab(wb, variant)
    build_identification_tab(wb, variant)
    build_replacement_tab(wb, variant)
    build_carryover_tab(wb, variant)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"1031 Exchange Tracker{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Like-kind exchange tracker for STR -> STR swaps under IRC §1031. "
        "Timeline support only — QI and CPA execute the exchange."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
