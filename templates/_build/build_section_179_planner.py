"""Build TAX-009 Section 179 Planner — wizard tool.

Implements templates/_briefs/TAX-009-section-179-planner.md.

Generates BOTH:
  templates/_masters/TAX-009-section-179-planner-DEMO.xlsx
  templates/_masters/TAX-009-section-179-planner-BLANK.xlsx

Tabs:
  0  Start              — wizard hero + 3-card + QS + Get Started + progress
  1  Eligibility        — active business test, classification, listed-property
  2  Asset List         — up to 15 assets table (Section 2 of 4)
  3  Election Per Asset — §179 / Bonus / MACRS toggles + per-asset deduction
  4  Income Limitation  — §179 ≤ biz income; carryover calc
  5  Form 4562 Map      — IRS form mirror
  6  Launch             — total deduction card + print packet
  7  Settings           — tax year · §179 cap · bonus % · brackets

Usage:
    python build_section_179_planner.py
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    COLOR_WHITE,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-009"
NAME = "section-179-planner"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
VERSION_LINE = f"{SKU} · v2.2 · Free updates forever"

TOTAL_SECTIONS = 4
SECTION_INPUT_COUNTS = {
    "Eligibility":        3,
    "Asset List":        45,   # 5 sample × 9 cols (or up to 15 × 9 = 135 cap)
    "Election Per Asset": 5,   # 5 sample × 1 col
    "Income Limitation":  1,
}
TOTAL_INPUTS = 3 + 45 + 5 + 1  # 54

# 2026 IRS thresholds (illustrative — update per IRS Rev. Proc. annually)
S179_CAP_2026 = 1_250_000
S179_PHASEOUT_THRESHOLD_2026 = 3_130_000
BONUS_DEPR_2026 = 0.40  # 40% phase-down (50% in 2025, 60% in 2024, etc.)
HEAVY_VEHICLE_CAP_2026 = 30_500  # SUV/truck cap (>6,000 lb GVWR)

ASSET_CAPACITY = 15  # max number of assets the table supports
SAMPLE_ASSETS = [
    # (description, MACRS_class_yrs, cost, date_placed_in_service, biz_use_pct, vehicle?)
    ("Hot tub (Smokies Ridge)",        5,    4800, date(2026, 4, 15), 1.00, "N"),
    ("Outdoor sofa set (Creek Side)",  5,    1200, date(2026, 5, 2),  1.00, "N"),
    ("Smart TV + stand (Lakehouse A)", 5,     980, date(2026, 3, 20), 1.00, "N"),
    ("Industrial vacuum (Lakehouse A)", 7,    620, date(2026, 2, 10), 1.00, "N"),
    ("Ford F-150 (heavy SUV/truck)",   5,   48000, date(2026, 1, 15), 0.80, "Y"),
]

# Per-asset elected method (DEMO)
SAMPLE_ELECTIONS = [
    "§179",
    "§179",
    "§179",
    "§179",
    "§179",  # but limited by heavy-vehicle cap
]

NET_BIZ_INCOME_DEMO = 43000  # for income limitation


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def append_callout(ws, row, text):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 36


def append_footer_nav(ws, row, prev_tab, next_tab, default_next="Form 4562 Map"):
    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = f"Next: {default_next} →"
        next_target = f"'{default_next}'!A1"
    pseudo_button(ws, f"A{row}", f"F{row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{row}", f"L{row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 22


def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Section 179 Planner"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Front-load this year's depreciation. Legally."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — single declarative answer (suite Theme 1).
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(IFERROR(\'Income Limitation\'!B17,0)>0,'
        '"✅  Year-1 deduction = "&TEXT(\'Income Limitation\'!B17,"$#,##0")'
        '&"  ·  §179 used = "&TEXT(\'Income Limitation\'!B14,"$#,##0"),'
        '"\U0001F4CA  Add an asset on Asset List to see your deduction.")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A11:L11")
    c = ws["A11"]; c.value = "What you'll get"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("⚡ §179 ELECTION", f"Up to ${S179_CAP_2026:,} expensed this year (2026 cap), per-asset choice."),
        ("📅 BONUS DEPR", f"{int(BONUS_DEPR_2026*100)}% bonus depreciation in 2026 — phasing down to 0% by 2028."),
        ("📄 FORM 4562", "Print-ready Form 4562 Part I (§179) + Part II (bonus) mirror."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (ttl, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]; c.value = ttl
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]; c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A23:L23")
    c = ws["A23"]; c.value = "Quick Start — be done in 15 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24
    quickstart_items = [
        "① Eligibility: active biz, schedule, listed-property check",
        "② Asset List: cost, MACRS class, date, business %",
        "③ Per-asset method: §179 / Bonus / MACRS",
        "④ Income limit: net biz income for §179 cap",
        "⑤ Print Form 4562 Map for your CPA",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + (i if i < 3 else i - 3)
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    pseudo_button(ws, "A30", "L33",
                   "GET STARTED — CHECK ELIGIBILITY  →",
                   "'Eligibility'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # Progress dashboard
    ws.merge_cells("A36:F36")
    c = ws["A36"]; c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ranges = [
        "'Eligibility'!B8:B10",       # 3
        f"'Asset List'!B8:F{8+ASSET_CAPACITY-1}",  # up to 75 cells
        f"'Election Per Asset'!B8:B{8+ASSET_CAPACITY-1}",
        "'Income Limitation'!B8",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    ws.merge_cells("G36:L36")
    c = ws["G36"]
    c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS},"0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    section_rows = [
        ("① Eligibility",        "Eligibility",        "B8:B10",   3),
        ("② Asset List",         "Asset List",         "B8:F22",  45),
        ("③ Election Per Asset", "Election Per Asset", "B8:B22",   5),
        ("④ Income Limitation",  "Income Limitation",  "B8",       1),
    ]
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]; c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (f'=IF({ca}>={total},"✅ Done",'
                   f'IF({ca}=0,"⏳ Empty","⏳ "&{ca}&" of {total}"))')
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    brand_footer(ws, 45, version_line=VERSION_LINE)

    ws.print_area = "A1:L47"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_eligibility_tab(wb, variant):
    ws = wb.create_sheet("Eligibility")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 1, TOTAL_SECTIONS,
                         "Eligibility",
                         "Active business test, classification, listed-property check.",
                         "", "Asset List")

    set_col_widths(ws, [
        ("A", 44), ("B", 22),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "Fill the highlighted fields below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22
    card_header(ws, 7, ("A", "L"), "Eligibility")

    fields = [
        ("Active trade/business? (Y/N)", _val(variant, "Y"), ["Y", "N"]),
        ("Schedule classification:",
         _val(variant, "Schedule C (active)"),
         ["Schedule E (passive)", "Schedule C (active)", "Ask my CPA"]),
        ("Listed property used > 50% for business? (vehicles/computers)",
         _val(variant, "Y"), ["Y", "N", "N/A"]),
    ]
    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    for i, (label, value, options) in enumerate(fields):
        r = 8 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=value)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        add_dropdown(ws, f"B{r}", options)
        ws.row_dimensions[r].height = 22

    card_body_fill(ws, 8, 10, ("A", "L"), border=True)

    # Eligibility verdict at row 12
    a = ws.cell(row=12, column=1, value="Verdict:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("B12:L12")
    c = ws.cell(row=12, column=2)
    c.value = (
        '=IF(B8="N","NOT ELIGIBLE — §179 requires an active trade/business.",'
        'IF(B9="Schedule E (passive)",'
        '"PARTIAL — Schedule E real estate is NOT eligible, but personal-property assets used in the rental are.",'
        '"ELIGIBLE — proceed to Asset List."))'
    )
    apply_style(c, formula_cell_style())
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[12].height = 32

    append_callout(ws, 14, (
        "ⓘ §179 is NOT allowed for residential rental REAL ESTATE itself, but IS "
        "allowed for personal-property assets used in an active rental business "
        "(appliances, furniture, hot tubs, vehicles). Listed property (vehicles, "
        "historically computers) requires > 50% business use to qualify for §179."
    ))
    append_footer_nav(ws, 16, prev_tab="", next_tab="Asset List")


def build_asset_list_tab(wb, variant):
    ws = wb.create_sheet("Asset List")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 2, TOTAL_SECTIONS,
                         "Asset List",
                         f"Up to {ASSET_CAPACITY} assets. Cost, class, date, business %.",
                         "Eligibility", "Election Per Asset")

    set_col_widths(ws, [
        ("A", 6), ("B", 32),
        ("C", 10), ("D", 14), ("E", 14), ("F", 12), ("G", 10),
        ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    # Subtitle row 6
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "MACRS classes: 5-yr (computers, vehicles), 7-yr (furniture/equipment), 15-yr (land improvements)."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: column headers
    headers = ["#", "Asset description", "MACRS yr",
               "Cost ($)", "Date placed in service",
               "Biz %", "Vehicle?"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 20

    # Rows 8 to 7+ASSET_CAPACITY: data rows
    for i in range(ASSET_CAPACITY):
        r = 8 + i
        ws.cell(row=r, column=1, value=i + 1).font = Font(
            name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
        )
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        if i < len(SAMPLE_ASSETS) and variant == "demo":
            (desc, macrs, cost, dt, biz_pct, is_vehicle) = SAMPLE_ASSETS[i]
            ws.cell(row=r, column=2, value=desc)
            ws.cell(row=r, column=3, value=macrs)
            ws.cell(row=r, column=4, value=cost)
            ws.cell(row=r, column=5, value=dt)
            ws.cell(row=r, column=6, value=biz_pct)
            ws.cell(row=r, column=7, value=is_vehicle)
        for col_idx in [2, 3, 4, 5, 6, 7]:
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, input_cell_style())
        ws.cell(row=r, column=4).number_format = '"$"#,##0'
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=6).number_format = "0.00%"
        ws.cell(row=r, column=7).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[r].height = 18

    last_asset_row = 7 + ASSET_CAPACITY  # 22

    add_dropdown(ws, f"C8:C{last_asset_row}", [5, 7, 15, 27.5, 39])
    add_dropdown(ws, f"G8:G{last_asset_row}", ["Y", "N"])

    # Total cost row
    total_row = last_asset_row + 1
    a = ws.cell(row=total_row, column=2, value="Total cost:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=total_row, column=4, value=f"=SUM(D8:D{last_asset_row})")
    apply_style(c, formula_cell_style())
    c.number_format = '"$"#,##0'
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[total_row].height = 24

    # Phase-out flag
    flag_row = total_row + 1
    ws.merge_cells(f"B{flag_row}:G{flag_row}")
    c = ws.cell(row=flag_row, column=2)
    c.value = (
        f'=IF(D{total_row}>{S179_PHASEOUT_THRESHOLD_2026},'
        f'"⚠ Phase-out applies — §179 cap reduced $-for-$ over ${S179_PHASEOUT_THRESHOLD_2026:,}.",'
        f'"OK — under §179 phase-out threshold (${S179_PHASEOUT_THRESHOLD_2026:,}).")'
    )
    apply_style(c, formula_cell_style())
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[flag_row].height = 22

    append_callout(ws, flag_row + 2, (
        "ⓘ MACRS recovery period drives the per-year deduction floor. STR personal "
        "property is most often 5-yr (carpets, appliances, technology) or 7-yr "
        "(furniture, kitchen equipment). Land improvements (fences, sidewalks) = "
        "15-yr. Buildings themselves = 27.5/39-yr (use the Schedule E Workbook + "
        "Home Office Allocator for those)."
    ))
    append_footer_nav(ws, flag_row + 4,
                       prev_tab="Eligibility", next_tab="Election Per Asset")


def build_election_per_asset_tab(wb, variant):
    ws = wb.create_sheet("Election Per Asset")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 3, TOTAL_SECTIONS,
                         "Election Per Asset",
                         "Choose §179 / Bonus / MACRS for each asset.",
                         "Asset List", "Income Limitation")

    set_col_widths(ws, [
        ("A", 6), ("B", 32), ("C", 14),
        ("D", 14), ("E", 14), ("F", 14), ("G", 14),
        ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "§179 (full this year) / Bonus (40% this year) / MACRS (over recovery period)."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    headers = ["#", "Asset (from Asset List)", "Method",
               "§179 deduction", "Bonus deduction", "Y1 MACRS deduction", "Total Y1"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 20

    for i in range(ASSET_CAPACITY):
        r = 8 + i
        # Asset # mirror
        c_num = ws.cell(row=r, column=1, value=f"='Asset List'!A{8+i}")
        c_num.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
        c_num.alignment = Alignment(horizontal="center", vertical="center")
        # Asset description mirror
        c_desc = ws.cell(row=r, column=2, value=f"='Asset List'!B{8+i}")
        apply_style(c_desc, formula_cell_style())
        c_desc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Method input
        if i < len(SAMPLE_ELECTIONS) and variant == "demo":
            method_val = SAMPLE_ELECTIONS[i]
        else:
            method_val = None
        c_method = ws.cell(row=r, column=3, value=method_val)
        apply_style(c_method, input_cell_style())
        c_method.alignment = Alignment(horizontal="center", vertical="center")
        # §179 deduction (D) — biz%-adjusted cost if method = §179
        # Heavy-vehicle cap applies if asset is vehicle
        # Asset List row reference: 8+i
        al = 8 + i
        c_179 = ws.cell(row=r, column=4)
        c_179.value = (
            f'=IF(C{r}="§179",'
            f'IF(\'Asset List\'!G{al}="Y",'
            f'MIN(\'Asset List\'!D{al}*\'Asset List\'!F{al},{HEAVY_VEHICLE_CAP_2026}*\'Asset List\'!F{al}),'
            f'\'Asset List\'!D{al}*\'Asset List\'!F{al}),0)'
        )
        apply_style(c_179, formula_cell_style())
        c_179.number_format = '"$"#,##0'

        # Bonus depreciation (E)
        c_bonus = ws.cell(row=r, column=5)
        c_bonus.value = (
            f'=IF(C{r}="Bonus",'
            f"\\'Asset List'!D{al}*\\'Asset List'!F{al}*{BONUS_DEPR_2026},0)"
        ).replace("\\'", "'")
        apply_style(c_bonus, formula_cell_style())
        c_bonus.number_format = '"$"#,##0'

        # Y1 MACRS (F) — half-year convention, simplified to (cost × biz%) / recovery × 0.5
        c_macrs = ws.cell(row=r, column=6)
        c_macrs.value = (
            f'=IF(C{r}="MACRS",'
            f"IFERROR((\\'Asset List'!D{al}*\\'Asset List'!F{al})/\\'Asset List'!C{al}*0.5,0),0)"
        ).replace("\\'", "'")
        apply_style(c_macrs, formula_cell_style())
        c_macrs.number_format = '"$"#,##0'

        # Total Y1 (G)
        c_total = ws.cell(row=r, column=7,
                           value=f"=D{r}+E{r}+F{r}")
        apply_style(c_total, formula_cell_style())
        c_total.number_format = '"$"#,##0'
        c_total.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        ws.row_dimensions[r].height = 18

    last_asset_row = 7 + ASSET_CAPACITY

    add_dropdown(ws, f"C8:C{last_asset_row}", ["§179", "Bonus", "MACRS"])

    # Totals row
    tot_row = last_asset_row + 1
    a = ws.cell(row=tot_row, column=2, value="Totals:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for col in [4, 5, 6, 7]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=tot_row, column=col,
                     value=f"=SUM({col_letter}8:{col_letter}{last_asset_row})")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[tot_row].height = 24

    append_callout(ws, tot_row + 2, (
        "ⓘ §179 capped per-year (currently $1.25M for 2026) AND limited by your "
        "business income (next tab). Bonus depreciation has NO income limit but "
        "the rate is phasing down (40% → 20% → 0% by 2028). MACRS spreads the "
        "deduction over the recovery period — slowest, but no Y1 limitation."
    ))
    append_footer_nav(ws, tot_row + 4,
                       prev_tab="Asset List", next_tab="Income Limitation")


def build_income_limitation_tab(wb, variant):
    ws = wb.create_sheet("Income Limitation")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 4, TOTAL_SECTIONS,
                         "Income Limitation",
                         "§179 deduction ≤ aggregate net business income.",
                         "Election Per Asset", "Form 4562 Map")

    set_col_widths(ws, [
        ("A", 44), ("B", 18),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "§179 cannot exceed aggregate net biz income — excess carries forward."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    card_header(ws, 7, ("A", "L"), "Income Limitation")

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    a = ws.cell(row=8, column=1,
                 value="Aggregate net business income (before §179):")
    a.font = bold
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=8, column=2, value=_val(variant, NET_BIZ_INCOME_DEMO))
    apply_style(c, input_cell_style())
    c.number_format = '"$"#,##0'
    ws.row_dimensions[8].height = 22

    card_body_fill(ws, 8, 8, ("A", "L"), border=True)

    # Computed rows starting row 10
    rows = [
        (10, "§179 elected (from Election Per Asset):",
         "='Election Per Asset'!D23", "money"),
        (11, "§179 cap (Settings):",
         "=Settings!B6", "money"),
        (12, "Phase-out reduction (cost over threshold):",
         f"=MAX('Asset List'!D23-Settings!B7,0)", "money"),
        (13, "Adjusted §179 cap (after phase-out):",
         "=MAX(B11-B12,0)", "money"),
        (14, "Allowed §179 (≤ adjusted cap AND ≤ biz income):",
         "=MIN(B10,B13,B8)", "money"),
        (15, "Bonus depreciation (no income limit):",
         "='Election Per Asset'!E23", "money"),
        (16, "Y1 MACRS:",
         "='Election Per Asset'!F23", "money"),
        (17, "TOTAL Y1 DEDUCTION:",
         "=B14+B15+B16", "total"),
        (18, "§179 carryover (excess deferred to next year):",
         "=MAX(B10-B14,0)", "money"),
    ]
    for r, label, formula, kind in rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        if kind == "total":
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 18

    append_callout(ws, 20, (
        "ⓘ §179 carryover has no expiration — it stays on the books indefinitely "
        "until offset by future business income. Bonus depreciation does NOT carry "
        "over (it's used in full in Y1 regardless of business income). For STR "
        "hosts: aggregate biz income includes Schedule C net profit + W-2 wages "
        "from any active employment + spouse W-2 (MFJ filers)."
    ))
    append_footer_nav(ws, 22,
                       prev_tab="Election Per Asset", next_tab="Form 4562 Map")


def build_form_4562_map_tab(wb):
    ws = wb.create_sheet("Form 4562 Map")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Form 4562 Map",
                         prev_tab="Income Limitation", next_tab="Launch")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "IRS Form 4562 — Part I (§179) and Part II (Bonus) mirror."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6), ("B", 50), ("C", 18),
        ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    headers = ["Line", "Description", "Amount"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        apply_style(c, header_row_style())
    ws.row_dimensions[5].height = 20

    lines = [
        ("Part I", "PART I — §179 ELECTION TO EXPENSE",        None, "header"),
        ("1",      "Maximum amount (Settings cap):",            "=Settings!B6", "money"),
        ("2",      "Total cost of §179 property placed in service this year:",
                   f"='Asset List'!D{7+ASSET_CAPACITY+1}", "money"),
        ("3",      "Threshold cost ($-for-$ phase-out start):", "=Settings!B7", "money"),
        ("4",      "Reduction in cap (Line 2 − Line 3, if positive):",
                   "=MAX(C8-C9,0)", "money"),
        ("5",      "Adjusted cap (Line 1 − Line 4):",
                   "=MAX(C7-C10,0)", "money"),
        ("8",      "Total §179 elected (sum of asset elections):",
                   "='Election Per Asset'!D23", "money"),
        ("9",      "Tentative deduction (smaller of Line 5 or Line 8):",
                   "=MIN(C11,C13)", "money"),
        ("11",     "Business income limitation:",                "='Income Limitation'!B8", "money"),
        ("12",     "§179 deduction allowed (smallest of 9, 11):",
                   "=MIN(C14,C15)", "total"),
        ("13",     "Carryover to 2027 (Line 8 − Line 12):",
                   "=MAX(C13-C16,0)", "money"),
        ("Part II", "PART II — SPECIAL DEPRECIATION (BONUS)",   None, "header"),
        ("14",     "Bonus depreciation (40% × cost × biz %):",
                   "='Election Per Asset'!E23", "total"),
        ("Part III", "PART III — MACRS (regular depreciation)", None, "header"),
        ("19a-i",  "Y1 MACRS deduction (varies by class life):",
                   "='Election Per Asset'!F23", "total"),
        ("Total",  "GRAND TOTAL Y1 DEDUCTION:",
                   "=C16+C19+C21", "grand_total"),
    ]

    r = 6
    for line_num, label, formula, kind in lines:
        if kind == "header":
            ws.merge_cells(f"A{r}:C{r}")
            c = ws.cell(row=r, column=1, value=f"  {line_num} — {label}")
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
            c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 22
        else:
            ws.cell(row=r, column=1, value=line_num).font = Font(
                name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
            )
            ws.cell(row=r, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=r, column=2, value=label).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )
            ws.cell(row=r, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            c = ws.cell(row=r, column=3, value=formula)
            apply_style(c, formula_cell_style())
            c.number_format = '"$"#,##0'
            if kind in ("total", "grand_total"):
                c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                ws.cell(row=r, column=2).font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            if kind == "grand_total":
                c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
                c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
                ws.cell(row=r, column=2).font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
            ws.row_dimensions[r].height = 18
        r += 1

    ws.print_area = f"A1:C{r}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_launch_tab(wb, variant):
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK", "'Form 4562 Map'!A1",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]; c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Your Section 179 deduction"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Print Form 4562 Map and attach to Schedule C / E."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 14):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 — Total Y1 deduction
    ws.merge_cells("A9:D9")
    c = ws["A9"]; c.value = "TOTAL Y1 DEDUCTION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]; c.value = "='Income Limitation'!B17"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A12:D12")
    c = ws["A12"]; c.value = "§179 + Bonus + Y1 MACRS"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 — §179 portion
    ws.merge_cells("E9:H9")
    c = ws["E9"]; c.value = "§179 ALLOWED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]; c.value = "='Income Limitation'!B14"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E12:H12")
    c = ws["E12"]; c.value = "after income + cap limits"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 — Carryover
    ws.merge_cells("I9:L9")
    c = ws["I9"]; c.value = "CARRYOVER"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]; c.value = "='Income Limitation'!B18"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I12:L12")
    c = ws["I12"]; c.value = "deferred to next year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    pseudo_button(ws, "A16", "L20",
                   "📄  PRINT FORM 4562 MAP  →",
                   "'Form 4562 Map'!A1", variant="primary")
    for r in range(16, 21):
        ws.row_dimensions[r].height = 24

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = (
        "💡 Upgrade to the Tax Season Bundle ($147)."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[23].height = 36

    brand_footer(ws, 25, version_line=VERSION_LINE)

    ws.print_area = "A1:L27"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Settings", prev_tab="Launch", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · §179 cap · phase-out · bonus % · vehicle cap."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 38), ("B", 18),
        ("C", 6), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    fields = [
        (5, "Active tax year:",                                2026,                          "0"),
        (6, "§179 maximum deduction (cap):",                   S179_CAP_2026,                 '"$"#,##0'),
        (7, "§179 phase-out threshold:",                       S179_PHASEOUT_THRESHOLD_2026,  '"$"#,##0'),
        (8, "Bonus depreciation % (placed in service Y1):",    BONUS_DEPR_2026,               "0.0%"),
        (9, "Heavy SUV/truck §179 cap (>6,000 lb GVWR):",      HEAVY_VEHICLE_CAP_2026,        '"$"#,##0'),
    ]
    for r, label, val, fmt in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # Freshness stamp (suite Theme 4) — IRS adjusts the §179 cap annually for
    # inflation, and the bonus % is phasing out (40% 2026 → 0% 2028). Customer
    # should bump rows 5-9 + this stamp every January.
    ws.merge_cells("A10:F10")
    note = ws["A10"]
    note.value = (
        "📅 §179 cap, phase-out + bonus % as of 2026-01-01 — IRS adjusts the "
        "cap each Nov; bonus is statutory and phases to 0% by 2028."
    )
    note.font = italic_muted
    note.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[10].height = 28

    # Historical
    a = ws.cell(row=11, column=1, value="Historical reference (bonus depr %):")
    a.font = italic_muted
    historical = [(12, "2023:", 0.80), (13, "2024:", 0.60), (14, "2025:", 0.50),
                   (15, "2026:", 0.40), (16, "2027:", 0.20), (17, "2028:", 0.00)]
    for row, label, val in historical:
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        rc = ws.cell(row=row, column=2, value=val)
        rc.number_format = "0%"
        rc.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)

    # Year-end Archive
    ws.row_dimensions[18].height = 8
    a = ws.cell(row=19, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[19].height = 22
    archive_headers = ["Year", "§179 elected", "§179 allowed", "Bonus", "Y1 MACRS", "Carryover"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=20, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[20].height = 18
    for idx, year in enumerate(range(2024, 2031), start=21):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in range(2, 7):
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_eligibility_tab(wb, variant)
    build_asset_list_tab(wb, variant)
    build_election_per_asset_tab(wb, variant)
    build_income_limitation_tab(wb, variant)
    build_form_4562_map_tab(wb)
    build_launch_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Section 179 Planner — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "IRS §179 / Bonus depreciation planner (Form 4562) for STR hosts."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
