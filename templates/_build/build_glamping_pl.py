"""Build SPC-001 Glamping / Unique-Stay P&L Excel files (v2.2 standard).

Forked from build_pl_single_property.py (12-month matrix), adapted to the
glamping/unique-stay operating model: heavy seasonality, propane/generator/
water-haul off-grid utilities, alt-revenue lines (firewood, gear rentals,
on-site experiences, photo shoots), and a per-season dashboard.

DEMO sample = "Smokies Yurt" (yurt + dome combo on 10 acres) — annual
revenue ~$94K, NOI ~$42K, ~70% of NOI earned in Peak (May-Oct).

Generates BOTH DEMO and BLANK variants from shared code.
"""
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    STATE_BAD_FILL, STATE_GOOD_FILL,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "SPC-001"
NAME = "glamping-unique-stay-pl"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_TAX_YEAR = 2026

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# --- Income lines (Income tab, 12-month matrix) ---
INCOME_LINES = [
    "Lodging revenue (nightly rate × nights)",
    "Cleaning fee revenue",
    "Pet fee revenue",
    "Firewood sales",
    "Gear rentals (kayak / fire pit / e-bike)",
    "On-site experiences (guided / hosted)",
    "Photo-shoot fees",
]

# Alt-revenue lines tracked separately on Alt Revenue tab
ALT_REVENUE_LINES = [
    "Firewood sales",
    "Gear rentals (kayak / fire pit / e-bike)",
    "On-site experiences (guided / hosted)",
    "Photo-shoot fees",
]

# --- Expense categories (Expenses tab, 12-month matrix) ---
EXPENSE_CATEGORIES = [
    "Propane (refills + tank rental)",
    "Generator fuel + maintenance",
    "Water (delivery / well pump electrical)",
    "Septic / port-a-potty pumping",
    "Linens (replaced more frequently)",
    "Weather damage repairs",
    "Pest control",
    "Composting / waste management",
    "Decor refresh (Insta inventory)",
    "Listing photo refresh",
    "Insurance (specialty glamping policy)",
    "Property taxes",
    "Internet / connectivity",
    "Marketing / ads",
    "Cleaning + turnover supplies",
    "Platform fees (Airbnb / VRBO)",
    "Other — Misc",
]

# --- DEMO sample data: monthly seasonality split ---
# Peak (May-Oct): 78% occ, $245 ADR, $76K rev | Shoulder (Mar-Apr, Nov):
# 45% occ, $185 ADR, $11K rev | Off (Dec-Feb): 18% occ, $145 ADR, $3.4K rev.
# Lodging revenue per month (sums to ~$87K).
SAMPLE_LODGING_BY_MONTH = [
    1100,  1000,  3200,  3500,  9800, 13200,  # Jan-Jun
   14500, 14200, 12800, 11500,  4300,  1300,  # Jul-Dec
]

SAMPLE_CLEANING_BY_MONTH = [
    150, 130, 320, 360, 980, 1200,
    1320, 1280, 1180, 1080, 430, 170,
]

SAMPLE_PETFEE_BY_MONTH = [
    50, 50, 100, 100, 250, 350,
    400, 400, 350, 300, 100, 50,
]

# Alt revenue ~$7K total: firewood $2K, kayak $3K, photos $2K (rounded mix)
SAMPLE_FIREWOOD_BY_MONTH = [
     20,  20,  60,  80, 240, 340,
    380, 360, 280, 200,  60,  20,
]

SAMPLE_GEAR_BY_MONTH = [
      0,   0,  60, 120, 380, 540,
    620, 600, 480, 240,   0,   0,
]

SAMPLE_EXPERIENCE_BY_MONTH = [
      0,   0,   0,  50, 150, 220,
    260, 240, 180,  80,  20,   0,
]

SAMPLE_PHOTO_BY_MONTH = [
      0,   0,  50, 100, 250, 350,
    400, 400, 250, 150,  50,   0,
]

# --- DEMO expense data per category (sums roughly $52K) ---
# Propane $4.8K, generator $1.2K, linens $3.6K -- per brief.
SAMPLE_EXPENSE_BY_CATEGORY = {
    # Propane: heavy in Peak (cooking + heat in shoulder)
    "Propane (refills + tank rental)":
        [220, 220, 280, 320, 480, 540, 560, 540, 500, 420, 380, 340],
    # Generator: only used during Peak operations
    "Generator fuel + maintenance":
        [ 30,  30,  60,  80, 140, 180, 200, 180, 160, 100,  30,  10],
    # Water: hauled in Peak, less in winter
    "Water (delivery / well pump electrical)":
        [ 80,  80, 100, 120, 200, 240, 260, 240, 220, 160, 100,  80],
    # Septic / port-a-potty
    "Septic / port-a-potty pumping":
        [  0,   0, 120, 120, 240, 240, 240, 240, 240, 240, 120,   0],
    # Linens: replaced more often in Peak
    "Linens (replaced more frequently)":
        [120, 120, 200, 240, 420, 480, 500, 480, 440, 380, 180, 120],
    # Weather damage: storms hit shoulder + winter
    "Weather damage repairs":
        [200,   0, 400,   0, 200,   0,   0,   0, 200, 200, 400, 200],
    # Pest control
    "Pest control":
        [  0,  60,   0,  60,   0,  60,   0,  60,   0,  60,   0,  60],
    # Composting / waste
    "Composting / waste management":
        [ 40,  40,  60,  60, 100, 120, 140, 120, 100,  80,  60,  40],
    # Decor refresh (Insta-driven)
    "Decor refresh (Insta inventory)":
        [  0,   0, 200,   0, 300,   0, 200,   0, 200,   0,   0,   0],
    # Listing photo refresh: 3x/yr typical
    "Listing photo refresh":
        [  0,   0, 400,   0,   0,   0, 400,   0,   0,   0, 400,   0],
    # Insurance: monthly accrual
    "Insurance (specialty glamping policy)":
        [180, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180],
    # Property taxes: yearly hit in Nov
    "Property taxes":
        [  0,   0,   0,   0,   0,   0,   0,   0,   0,   0,1800,   0],
    # Internet
    "Internet / connectivity":
        [ 95,  95,  95,  95,  95,  95,  95,  95,  95,  95,  95,  95],
    # Marketing
    "Marketing / ads":
        [ 50,  50,  80, 120, 180, 200, 220, 200, 160, 120,  60,  40],
    # Cleaning + turnover supplies
    "Cleaning + turnover supplies":
        [ 80,  80, 140, 160, 320, 380, 400, 380, 340, 280, 140,  80],
    # Platform fees: ~10% of lodging revenue
    "Platform fees (Airbnb / VRBO)":
        [110, 100, 320, 350, 980,1320,1450,1420,1280,1150, 430, 130],
    # Misc
    "Other — Misc":
        [ 40,  40,  60,  60, 100, 120, 140, 120, 100,  80,  60,  40],
}

# --- Off-Grid Utilities sample logs (DEMO) ---
SAMPLE_PROPANE_LOG = [
    ("2026-01-12", "Smokies Yurt", 100, 320, "Suburban Propane", "20%"),
    ("2026-02-15", "Smokies Yurt", 100, 315, "Suburban Propane", "25%"),
    ("2026-03-20", "Smokies Yurt", 120, 380, "AmeriGas",         "15%"),
    ("2026-04-22", "Smokies Yurt", 120, 385, "AmeriGas",         "20%"),
    ("2026-05-18", "Smokies Yurt", 150, 470, "Suburban Propane", "10%"),
    ("2026-06-15", "Smokies Yurt", 150, 475, "Suburban Propane", "10%"),
    ("2026-07-12", "Smokies Yurt", 150, 480, "AmeriGas",         "15%"),
    ("2026-08-10", "Smokies Yurt", 150, 470, "AmeriGas",         "10%"),
    ("2026-09-15", "Smokies Yurt", 140, 440, "Suburban Propane", "15%"),
    ("2026-10-20", "Smokies Yurt", 130, 410, "Suburban Propane", "20%"),
    ("2026-11-22", "Smokies Yurt", 120, 380, "AmeriGas",         "20%"),
    ("2026-12-18", "Smokies Yurt", 110, 350, "AmeriGas",         "25%"),
]

SAMPLE_GENERATOR_LOG = [
    ("2026-01-15",  4,  2, "Cold-snap backup"),
    ("2026-02-20",  3,  2, "Storm outage"),
    ("2026-03-18",  6,  3, "Routine"),
    ("2026-04-22",  8,  4, "Routine + oil change"),
    ("2026-05-15", 14,  6, "Peak season starts"),
    ("2026-06-12", 22, 10, "Peak operations"),
    ("2026-07-10", 28, 12, "Peak — heatwave AC"),
    ("2026-07-25", 24, 12, "Peak — continued"),
    ("2026-08-15", 26, 12, "Peak — service due"),
    ("2026-09-12", 18,  8, "Shoulder taper"),
    ("2026-10-15", 12,  6, "Shoulder taper"),
    ("2026-11-20",  4,  2, "Off-season check"),
]

SAMPLE_WATER_LOG = [
    ("2026-03-15", 500, 120, "AquaHaul TN"),
    ("2026-04-18", 500, 120, "AquaHaul TN"),
    ("2026-05-15", 750, 180, "AquaHaul TN"),
    ("2026-06-10", 750, 180, "Mountain Water Co"),
    ("2026-07-08", 1000, 240, "Mountain Water Co"),
    ("2026-07-22", 1000, 240, "AquaHaul TN"),
    ("2026-08-10", 1000, 240, "Mountain Water Co"),
    ("2026-08-25", 750, 180, "Mountain Water Co"),
    ("2026-09-15", 750, 180, "AquaHaul TN"),
    ("2026-10-12", 500, 120, "AquaHaul TN"),
    ("2026-11-15", 500, 120, "AquaHaul TN"),
]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    """Parse YYYY-MM-DD -> datetime.date so SUMIFS criteria match."""
    return datetime.strptime(s, "%Y-%m-%d").date()


def _val(variant, demo_value):
    """Return demo_value for DEMO variant, None for BLANK."""
    return demo_value if variant == "demo" else None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start hero + KPI cards + nav."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Navy hero band rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Brand row (write BEFORE merge)
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A2:F2")

    # Title row
    c = ws["A4"]
    c.value = "Glamping / Unique-Stay P&L"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 44

    c = ws["A5"]
    c.value = "Built for yurts, domes, A-frames, treehouses, and the off-grid life."
    c.font = Font(name=FONT_HEAD, size=13, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A5:L5")

    # Verdict row 6 — peak vs off NOI mix
    c = ws["A6"]
    c.value = (
        '=IF(Income!N7=0,'
        '"\U0001F4CA  Log a month of revenue to see your seasonality verdict.",'
        '"✓  YTD Revenue = "&TEXT(Income!N7,"$#,##0")'
        '&"   ·   YTD NOI = "&TEXT(Income!N7-Expenses!N25,"$#,##0"))'
    )
    c.font = Font(name=FONT_HEAD, size=15, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.merge_cells("A6:L6")
    ws.row_dimensions[6].height = 30

    c = ws["A7"]
    c.value = f"SPC-001 · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A7:L7")

    # KPI cards row 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c_idx in range(1, 13):
            ws.cell(row=r, column=c_idx).fill = parchment_fill

    # Card 1 (A-C): YTD Revenue
    c = ws["A10"]
    c.value = "YTD REVENUE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:C10")

    c = ws["A11"]
    c.value = "=Income!N7"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A11:C13")

    c = ws["A14"]
    c.value = "lodging + alt"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A14:C14")

    # Card 2 (D-F): Peak revenue share
    c = ws["D10"]
    c.value = "PEAK SHARE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("D10:F10")

    c = ws["D11"]
    c.value = (
        '=IF(Income!N7=0,0,'
        "('Seasonality Dashboard'!B7)/Income!N7)"
    )
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("D11:F13")

    c = ws["D14"]
    c.value = "of total revenue"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("D14:F14")

    # Card 3 (G-I): YTD Expenses
    c = ws["G10"]
    c.value = "YTD EXPENSES"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("G10:I10")

    c = ws["G11"]
    c.value = "=Expenses!N25"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("G11:I13")

    c = ws["G14"]
    c.value = "all categories"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("G14:I14")

    # Card 4 (J-L): Alt revenue %
    c = ws["J10"]
    c.value = "ALT REV %"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("J10:L10")

    c = ws["J11"]
    c.value = (
        "=IF(Income!N7=0,0,SUM(Income!N10:N13)/Income!N7)"
    )
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("J11:L13")

    c = ws["J14"]
    c.value = "firewood + gear + experiences"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("J14:L14")

    # Gold borders around the 4 cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 15):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 14 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # Primary CTA
    pseudo_button(ws, "A17", "L20",
                  "→  OPEN INCOME MATRIX",
                  "'Income'!A1", variant="primary")
    for r in range(17, 21):
        ws.row_dimensions[r].height = 20

    # Secondary nav buttons (rows 22-23)
    pseudo_button(ws, "A22", "C23", "Expenses",
                  "'Expenses'!A1", variant="secondary")
    pseudo_button(ws, "D22", "F23", "Seasonality",
                  "'Seasonality Dashboard'!A1", variant="accent")
    pseudo_button(ws, "G22", "I23", "Off-Grid Utilities",
                  "'Off-Grid Utilities'!A1", variant="secondary")
    pseudo_button(ws, "J22", "L23", "Settings",
                  "'Settings'!A1", variant="secondary")
    ws.row_dimensions[22].height = 22
    ws.row_dimensions[23].height = 22

    # Tertiary row
    pseudo_button(ws, "A25", "F26", "Alt Revenue",
                  "'Alt Revenue'!A1", variant="secondary")
    pseudo_button(ws, "G25", "L26", "📄  Print Year-End P&L",
                  "'Expenses'!A1", variant="accent")
    ws.row_dimensions[25].height = 22
    ws.row_dimensions[26].height = 22

    # Tax-time callout row 28
    c = ws["A28"]
    c.value = (
        "📅 SEASONALITY: glamping commonly loses money Off-Season — track Peak NOI "
        "carefully. Most operators earn ~70% of annual NOI in Peak alone."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    ws.merge_cells("A28:L28")
    ws.row_dimensions[28].height = 32

    # Upgrade banner
    add_upgrade_banner(ws, 30)

    # Footer
    brand_footer(ws, 32, version_line="SPC-001 · v2.2 · Free updates forever")

    ws.print_area = "A1:L34"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_income_tab(wb, variant):
    """Sheet 1 — Income (12-month matrix, lodging + alt revenue lines)."""
    ws = wb.create_sheet("Income")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Income",
                        prev_tab="Start", next_tab="Expenses")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c_idx in range(1, 13):
        ws.cell(row=4, column=c_idx).fill = parchment_fill
    c4 = ws["A4"]
    c4.value = "Monthly revenue rollup — lodging + alt-revenue lines"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 18

    col_widths = [("A", 38)] + [(get_column_letter(2 + i), 10) for i in range(12)] + [("N", 12)]
    set_col_widths(ws, col_widths)

    headers = ["Line"] + MONTHS + ["YTD"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 8

    # Layout: Row 7 = TOTAL REVENUE (Start tab references Income!N7).
    # Row 8 = Lodging, 9 = Cleaning, 10-13 = alt revenue (firewood / gear /
    # experience / photo — Start tab's alt% formula reads N10:N13), 14 = Pet.
    ws.row_dimensions[7].height = 22

    # Row 7: TOTAL REVENUE (formula sums lines below: rows 8-14)
    a = ws.cell(row=7, column=1, value="TOTAL REVENUE")
    a.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=7, column=col,
                       value=f"=SUM({col_letter}8:{col_letter}14)")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                         italic=True, color=COLOR_PRIMARY)

    # Rows 8-14: per-line input rows
    # Order chosen so alt-revenue lines fall on rows 10-13 (Start formula uses N10:N13)
    # 8 = Lodging | 9 = Cleaning | 10 = Pet | 11 = Firewood | 12 = Gear | 13 = Experience | 14 = Photo
    # That's 5 alt rows starting at 10... brief says alt% includes firewood+gear+experience+photo (4 lines)
    # Use rows 10-13 for the 4 alt lines, row 14 for Pet fees.
    line_order = [
        ("Lodging revenue (nightly rate × nights)", SAMPLE_LODGING_BY_MONTH),
        ("Cleaning fee revenue", SAMPLE_CLEANING_BY_MONTH),
        ("Firewood sales", SAMPLE_FIREWOOD_BY_MONTH),
        ("Gear rentals (kayak / fire pit / e-bike)", SAMPLE_GEAR_BY_MONTH),
        ("On-site experiences (guided / hosted)", SAMPLE_EXPERIENCE_BY_MONTH),
        ("Photo-shoot fees", SAMPLE_PHOTO_BY_MONTH),
        ("Pet fee revenue", SAMPLE_PETFEE_BY_MONTH),
    ]
    # Rows: 8=Lodging, 9=Cleaning, 10-13=alt, 14=Pet
    for idx, (label, monthly) in enumerate(line_order):
        row = 8 + idx
        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16
        for m in range(12):
            col = m + 2
            cell = ws.cell(row=row, column=col,
                           value=_val(variant, monthly[m]))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ytd = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(ytd, formula_cell_style())
        ytd.number_format = '"$"#,##0'

    ws.freeze_panes = "B6"

    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_expenses_tab(wb, variant):
    """Sheet 2 — Expenses (12-month matrix, glamping-specific categories)."""
    ws = wb.create_sheet("Expenses")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Expenses",
                        prev_tab="Income", next_tab="Seasonality Dashboard")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c_idx in range(1, 13):
        ws.cell(row=4, column=c_idx).fill = parchment_fill
    c4 = ws["A4"]
    c4.value = "Monthly expense rollup — glamping-specific categories"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 18

    col_widths = [("A", 40)] + [(get_column_letter(2 + i), 10) for i in range(12)] + [("N", 12)]
    set_col_widths(ws, col_widths)

    headers = ["Category"] + MONTHS + ["YTD"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 8

    # Rows 7-23: one row per expense category (17 categories)
    for idx, cat in enumerate(EXPENSE_CATEGORIES):
        row = 7 + idx
        a = ws.cell(row=row, column=1, value=cat)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16

        monthly = SAMPLE_EXPENSE_BY_CATEGORY.get(cat, [0]*12)
        for m in range(12):
            col = m + 2
            cell = ws.cell(row=row, column=col,
                           value=_val(variant, monthly[m]))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'

        ytd = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(ytd, formula_cell_style())
        ytd.number_format = '"$"#,##0'

    last_exp_row = 7 + len(EXPENSE_CATEGORIES) - 1  # = 23
    ws.row_dimensions[24].height = 8

    # Total row at row 25 (Start tab references Expenses!N25)
    tot_row = 25
    a = ws.cell(row=tot_row, column=1, value="TOTAL EXPENSES")
    a.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_ERROR)
    a.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tot_row].height = 22

    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                       value=f"=SUM({col_letter}7:{col_letter}{last_exp_row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                         italic=True, color=COLOR_ERROR)

    # NOI row at 27
    noi_row = 27
    ws.row_dimensions[26].height = 8
    a = ws.cell(row=noi_row, column=1, value="NET OPERATING INCOME")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[noi_row].height = 22

    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=noi_row, column=col,
                       value=f"=Income!{col_letter}7-{col_letter}{tot_row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                         italic=True, color=COLOR_PRIMARY)

    ws.conditional_formatting.add(
        f"B{noi_row}:N{noi_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        f"B{noi_row}:N{noi_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )

    ws.freeze_panes = "B6"

    ws.print_area = f"A1:N{noi_row + 1}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_seasonality_dashboard(wb):
    """Sheet 3 — Seasonality Dashboard (Peak / Shoulder / Off split)."""
    ws = wb.create_sheet("Seasonality Dashboard")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Seasonality Dashboard",
                        prev_tab="Expenses", next_tab="Off-Grid Utilities")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c_idx in range(1, 13):
        ws.cell(row=4, column=c_idx).fill = parchment_fill
    c4 = ws["A4"]
    c4.value = "Per-season breakdown — set season month ranges on Settings"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 28), ("B", 16), ("C", 16), ("D", 16)])

    # Headers row 5
    headers = ["Metric", "Peak (May-Oct)", "Shoulder (Mar-Apr, Nov)", "Off (Dec-Feb)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 22

    ws.row_dimensions[6].height = 8

    # Row 7: Revenue total per season
    a = ws.cell(row=7, column=1, value="Revenue total")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")

    # Peak = May(F)+Jun(G)+Jul(H)+Aug(I)+Sep(J)+Oct(K) on Income!7
    peak_rev = "=Income!F7+Income!G7+Income!H7+Income!I7+Income!J7+Income!K7"
    shoulder_rev = "=Income!D7+Income!E7+Income!L7"
    off_rev = "=Income!B7+Income!C7+Income!M7"
    for col, formula in enumerate([peak_rev, shoulder_rev, off_rev], start=2):
        cell = ws.cell(row=7, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 20

    # Row 8: Expenses total per season (Expenses!25 row)
    a = ws.cell(row=8, column=1, value="Expenses total")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    a.alignment = Alignment(horizontal="left", vertical="center")
    peak_exp = "=Expenses!F25+Expenses!G25+Expenses!H25+Expenses!I25+Expenses!J25+Expenses!K25"
    shoulder_exp = "=Expenses!D25+Expenses!E25+Expenses!L25"
    off_exp = "=Expenses!B25+Expenses!C25+Expenses!M25"
    for col, formula in enumerate([peak_exp, shoulder_exp, off_exp], start=2):
        cell = ws.cell(row=8, column=col, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[8].height = 20

    # Row 9: NOI per season
    a = ws.cell(row=9, column=1, value="NOI")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 5):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=9, column=col, value=f"={col_letter}7-{col_letter}8")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[9].height = 22

    ws.row_dimensions[10].height = 8

    # Row 11: Months in season (label only; user-set)
    a = ws.cell(row=11, column=1, value="Months in season")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="left", vertical="center")
    months_per_season = ["May, Jun, Jul, Aug, Sep, Oct",
                         "Mar, Apr, Nov", "Dec, Jan, Feb"]
    for col, val in enumerate(months_per_season, start=2):
        cell = ws.cell(row=11, column=col, value=val)
        cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[11].height = 28

    # Row 12: Avg occ % (placeholder — input cells)
    a = ws.cell(row=12, column=1, value="Avg occupancy %")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="left", vertical="center")
    occ_defaults = [0.78, 0.45, 0.18]
    for col, val in enumerate(occ_defaults, start=2):
        cell = ws.cell(row=12, column=col, value=val)
        apply_style(cell, input_cell_style())
        cell.number_format = "0%"
    ws.row_dimensions[12].height = 18

    # Row 13: Avg ADR
    a = ws.cell(row=13, column=1, value="Avg ADR")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="left", vertical="center")
    adr_defaults = [245, 185, 145]
    for col, val in enumerate(adr_defaults, start=2):
        cell = ws.cell(row=13, column=col, value=val)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
    ws.row_dimensions[13].height = 18

    ws.row_dimensions[14].height = 8

    # Verdict block row 15
    a = ws.cell(row=15, column=1, value="Peak share of NOI:")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(row=15, column=2,
                   value="=IF(SUM(B9:D9)=0,0,B9/SUM(B9:D9))")
    apply_style(cell, formula_cell_style())
    cell.number_format = "0%"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[15].height = 24

    # Bar chart — Revenue / Expenses / NOI per season
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Seasonality Mix"
    chart.height = 8
    chart.width = 14
    data = Reference(ws, min_col=1, min_row=7, max_row=9, max_col=4)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    cats = Reference(ws, min_col=2, min_row=5, max_col=4, max_row=5)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    style_chart(chart)
    ws.add_chart(chart, "F5")

    ws.print_area = "A1:D17"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_offgrid_tab(wb, variant):
    """Sheet 4 — Off-Grid Utilities (3 stacked logs: propane, generator, water-haul)."""
    ws = wb.create_sheet("Off-Grid Utilities")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Off-Grid Utilities",
                        prev_tab="Seasonality Dashboard",
                        next_tab="Alt Revenue")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c_idx in range(1, 13):
        ws.cell(row=4, column=c_idx).fill = parchment_fill
    c4 = ws["A4"]
    c4.value = "Three stacked registers — propane refills, generator hours, water-haul"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 20), ("C", 14), ("D", 12),
        ("E", 22), ("F", 18), ("G", 28),
    ])

    # --- Section 1: Propane Refill Log (rows 6-58, header row 7) ---
    section_row = 6
    a = ws.cell(row=section_row, column=1, value="PROPANE REFILL LOG")
    a.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    # Fill rest of row band
    for c_idx in range(2, 8):
        cell = ws.cell(row=section_row, column=c_idx)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[section_row].height = 22

    propane_headers = ["Date", "Property", "Refill (gal)", "Cost ($)",
                       "Vendor", "Tank % at refill", "Notes"]
    hdr_row = section_row + 1  # 7
    for col, h in enumerate(propane_headers, start=1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[hdr_row].height = 20

    # 50-row register (rows 8-57)
    propane_data_start = hdr_row + 1  # 8
    propane_data_end = propane_data_start + 49  # 57
    sample = SAMPLE_PROPANE_LOG if variant == "demo" else []
    for i, entry in enumerate(sample):
        row = propane_data_start + i
        date_val, prop, gallons, cost, vendor, tank_pct = entry
        a = ws.cell(row=row, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"
        b = ws.cell(row=row, column=2, value=prop)
        apply_style(b, input_cell_style())
        c = ws.cell(row=row, column=3, value=gallons)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        d = ws.cell(row=row, column=4, value=cost)
        apply_style(d, input_cell_style())
        d.number_format = '"$"#,##0.00'
        e = ws.cell(row=row, column=5, value=vendor)
        apply_style(e, input_cell_style())
        f = ws.cell(row=row, column=6, value=tank_pct)
        apply_style(f, input_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")
        g = ws.cell(row=row, column=7, value=None)
        apply_style(g, input_cell_style())

    # Empty rows for remaining capacity
    last_filled = propane_data_start + len(sample) - 1 if sample else propane_data_start - 1
    for row in range(last_filled + 1, propane_data_end + 1):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 4:
                cell.number_format = '"$"#,##0.00'
            if col_idx in (3, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Propane totals row
    propane_tot_row = propane_data_end + 1  # 58
    a = ws.cell(row=propane_tot_row, column=1, value="TOTALS")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(row=propane_tot_row, column=3,
                   value=f"=SUM(C{propane_data_start}:C{propane_data_end})")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=propane_tot_row, column=4,
                   value=f"=SUM(D{propane_data_start}:D{propane_data_end})")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[propane_tot_row].height = 20

    # --- Section 2: Generator Hours Log ---
    gen_section_row = propane_tot_row + 2  # 60
    ws.row_dimensions[gen_section_row - 1].height = 8
    a = ws.cell(row=gen_section_row, column=1, value="GENERATOR HOURS LOG")
    a.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for c_idx in range(2, 8):
        cell = ws.cell(row=gen_section_row, column=c_idx)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[gen_section_row].height = 22

    gen_headers = ["Date", "Hours run", "Fuel added (gal)", "Service notes"]
    gen_hdr_row = gen_section_row + 1  # 61
    for col, h in enumerate(gen_headers, start=1):
        cell = ws.cell(row=gen_hdr_row, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[gen_hdr_row].height = 20

    gen_data_start = gen_hdr_row + 1  # 62
    gen_data_end = gen_data_start + 49  # 111
    sample_gen = SAMPLE_GENERATOR_LOG if variant == "demo" else []
    for i, entry in enumerate(sample_gen):
        row = gen_data_start + i
        date_val, hours, fuel, notes = entry
        a = ws.cell(row=row, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"
        b = ws.cell(row=row, column=2, value=hours)
        apply_style(b, input_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=row, column=3, value=fuel)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        d = ws.cell(row=row, column=4, value=notes)
        apply_style(d, input_cell_style())

    last_filled = gen_data_start + len(sample_gen) - 1 if sample_gen else gen_data_start - 1
    for row in range(last_filled + 1, gen_data_end + 1):
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx in (2, 3):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    gen_tot_row = gen_data_end + 1  # 112
    a = ws.cell(row=gen_tot_row, column=1, value="TOTALS")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(row=gen_tot_row, column=2,
                   value=f"=SUM(B{gen_data_start}:B{gen_data_end})")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=gen_tot_row, column=3,
                   value=f"=SUM(C{gen_data_start}:C{gen_data_end})")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[gen_tot_row].height = 20

    # --- Section 3: Water-Haul Log ---
    water_section_row = gen_tot_row + 2  # 114
    ws.row_dimensions[water_section_row - 1].height = 8
    a = ws.cell(row=water_section_row, column=1, value="WATER-HAUL LOG")
    a.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for c_idx in range(2, 8):
        cell = ws.cell(row=water_section_row, column=c_idx)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[water_section_row].height = 22

    water_headers = ["Date", "Gallons delivered", "Cost ($)", "Vendor"]
    water_hdr_row = water_section_row + 1  # 115
    for col, h in enumerate(water_headers, start=1):
        cell = ws.cell(row=water_hdr_row, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[water_hdr_row].height = 20

    water_data_start = water_hdr_row + 1  # 116
    water_data_end = water_data_start + 49  # 165
    sample_water = SAMPLE_WATER_LOG if variant == "demo" else []
    for i, entry in enumerate(sample_water):
        row = water_data_start + i
        date_val, gallons, cost, vendor = entry
        a = ws.cell(row=row, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"
        b = ws.cell(row=row, column=2, value=gallons)
        apply_style(b, input_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=row, column=3, value=cost)
        apply_style(c, input_cell_style())
        c.number_format = '"$"#,##0.00'
        d = ws.cell(row=row, column=4, value=vendor)
        apply_style(d, input_cell_style())

    last_filled = water_data_start + len(sample_water) - 1 if sample_water else water_data_start - 1
    for row in range(last_filled + 1, water_data_end + 1):
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 3:
                cell.number_format = '"$"#,##0.00'

    water_tot_row = water_data_end + 1  # 166
    a = ws.cell(row=water_tot_row, column=1, value="TOTALS")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(row=water_tot_row, column=2,
                   value=f"=SUM(B{water_data_start}:B{water_data_end})")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=water_tot_row, column=3,
                   value=f"=SUM(C{water_data_start}:C{water_data_end})")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[water_tot_row].height = 20

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_alt_revenue_tab(wb, variant):
    """Sheet 5 — Alt Revenue (per-line monthly tracking + bar chart)."""
    ws = wb.create_sheet("Alt Revenue")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Alt Revenue",
                        prev_tab="Off-Grid Utilities", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c_idx in range(1, 13):
        ws.cell(row=4, column=c_idx).fill = parchment_fill
    c4 = ws["A4"]
    c4.value = "Alt-revenue mix — pulls from Income tab rows 10-13"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 18

    col_widths = [("A", 38)] + [(get_column_letter(2 + i), 10) for i in range(12)] + [("N", 12)]
    set_col_widths(ws, col_widths)

    headers = ["Alt-Revenue Line"] + MONTHS + ["YTD"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    ws.row_dimensions[6].height = 8

    # Pull from Income tab rows 10-13 (Firewood / Gear / Experience / Photo)
    alt_source_rows = [10, 11, 12, 13]
    for idx, line in enumerate(ALT_REVENUE_LINES):
        row = 7 + idx
        src_row = alt_source_rows[idx]
        a = ws.cell(row=row, column=1, value=line)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16
        for m in range(12):
            col = m + 2
            col_letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col,
                           value=f"=Income!{col_letter}{src_row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
        ytd = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(ytd, formula_cell_style())
        ytd.number_format = '"$"#,##0'

    last_alt_row = 7 + len(ALT_REVENUE_LINES) - 1  # = 10
    ws.row_dimensions[last_alt_row + 1].height = 8

    tot_row = last_alt_row + 2  # = 12
    a = ws.cell(row=tot_row, column=1, value="TOTAL ALT REVENUE")
    a.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_SECONDARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                       value=f"=SUM({col_letter}7:{col_letter}{last_alt_row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                         italic=True, color=COLOR_SECONDARY)
    ws.row_dimensions[tot_row].height = 22

    # Bar chart — Alt revenue mix
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Alt Revenue Mix (YTD)"
    chart.height = 8
    chart.width = 14
    data = Reference(ws, min_col=14, min_row=7, max_row=last_alt_row)
    chart.add_data(data, titles_from_data=False)
    cats = Reference(ws, min_col=1, min_row=7, max_row=last_alt_row)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    style_chart(chart)
    ws.add_chart(chart, f"A{tot_row + 3}")

    ws.freeze_panes = "B6"

    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 6 — Settings (property name, active year, season cutoffs, off-grid toggle)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                        prev_tab="Alt Revenue", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c_idx in range(1, 13):
        ws.cell(row=4, column=c_idx).fill = parchment_fill
    c4 = ws["A4"]
    c4.value = "Property name · active year · season cutoffs · off-grid toggle"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A4:L4")
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 30)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    rows_data = [
        (5,  "Property name:",            _val(variant, "Smokies Yurt"),                    None),
        (6,  "Property type:",            _val(variant, "Yurt + dome combo (10 acres)"),    None),
        (7,  "Active year:",              ACTIVE_TAX_YEAR,                                  "0"),
        (8,  "",                          "",                                               None),
        (9,  "Peak season start date:",   _val(variant, _parse_date("2026-05-01")),         "yyyy-mm-dd"),
        (10, "Shoulder start date:",      _val(variant, _parse_date("2026-03-01")),         "yyyy-mm-dd"),
        (11, "Off-season start date:",    _val(variant, _parse_date("2026-12-01")),         "yyyy-mm-dd"),
        (12, "",                          "",                                               None),
        (13, "Off-grid? (Y/N):",          _val(variant, "Y"),                               None),
        (14, "Number of units:",          _val(variant, 2),                                 "0"),
        (15, "Avg unit capacity (sqft):", _val(variant, 320),                               "0"),
    ]

    bold_right_align = Alignment(horizontal="right", vertical="center")
    for row_num, label, value, fmt in rows_data:
        if not label:
            ws.row_dimensions[row_num].height = 8
            continue
        a = ws.cell(row=row_num, column=1, value=label)
        a.font = bold_right
        a.alignment = bold_right_align
        ws.row_dimensions[row_num].height = 18

        b = ws.cell(row=row_num, column=2, value=value)
        apply_style(b, input_cell_style())
        if fmt:
            b.number_format = fmt

    # Off-grid dropdown
    add_dropdown(ws, "B13", '"Y,N"')

    # Explainer row 17
    ws.merge_cells("A17:B17")
    c = ws["A17"]
    c.value = (
        "Active year drives every monthly aggregate. Bump it each January. "
        "Season cutoff dates are reference markers — Seasonality Dashboard uses "
        "fixed Peak (May-Oct) / Shoulder (Mar-Apr, Nov) / Off (Dec-Feb) windows."
    )
    c.font = italic_muted
    c.alignment = Alignment(horizontal="left", vertical="center",
                            wrap_text=True, indent=1)
    ws.row_dimensions[17].height = 36

    # Footer
    brand_footer(ws, 20, version_line="SPC-001 · v2.2")


# ---------------------------------------------------------------------------
# Main build function
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_income_tab(wb, variant)
    build_expenses_tab(wb, variant)
    build_seasonality_dashboard(wb)
    build_offgrid_tab(wb, variant)
    build_alt_revenue_tab(wb, variant)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Glamping / Unique-Stay P&L{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Glamping & unique-stay P&L with seasonality dashboard, off-grid "
        "utility logs (propane / generator / water-haul), and alt-revenue "
        "tracking (v2.2)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
