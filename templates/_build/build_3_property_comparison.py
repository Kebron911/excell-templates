"""Build ACQ-004 — 3-Property Side-by-Side Comparison (v2.2 standard).

Operational-mode tool — host pastes 3 acquisition candidates side-by-side and
the workbook surfaces a per-metric "winner" highlight + auto recommendation.

Generates two files:
  templates/_masters/ACQ-004-3-property-side-by-side-DEMO.xlsx   (sample data)
  templates/_masters/ACQ-004-3-property-side-by-side-BLANK.xlsx  (empty)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, apply_brand_header,
    brand_footer,
    COLOR_WHITE,
)

SKU = "ACQ-004"
NAME = "3-property-side-by-side"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data — three Smokies-area STR candidates (per brief)
# ---------------------------------------------------------------------------

SAMPLE = {
    "A": {
        "name":           "Smokies Ridge Cabin",
        "address":        "123 Mountain Lane, Gatlinburg TN",
        "beds":           3,
        "baths":          2,
        "sleeps":         8,
        "purchase":       485000,
        "down_pct":       0.25,
        "rate":           0.0700,
        "term_years":     30,
        "rehab":          35000,
        "furniture":      25000,
        "reserves":       12000,
        "occupancy":      0.68,
        "adr":            285,
        "insurance":      3400,
        "prop_tax":       5800,
        "utilities":      2400,
        "mgmt_pct":       0.0,
        "cleaning":       6800,
        "supplies":       1400,
    },
    "B": {
        "name":           "Lakehouse Retreat",
        "address":        "47 Shoreline Dr, Sevierville TN",
        "beds":           4,
        "baths":          3,
        "sleeps":         10,
        "purchase":       625000,
        "down_pct":       0.25,
        "rate":           0.0700,
        "term_years":     30,
        "rehab":          22000,
        "furniture":      32000,
        "reserves":       15000,
        "occupancy":      0.62,
        "adr":            340,
        "insurance":      4100,
        "prop_tax":       7200,
        "utilities":      3000,
        "mgmt_pct":       0.0,
        "cleaning":       8200,
        "supplies":       1700,
    },
    "C": {
        "name":           "Creek Side Cottage",
        "address":        "9 Creekbed Rd, Pigeon Forge TN",
        "beds":           2,
        "baths":          2,
        "sleeps":         6,
        "purchase":       360000,
        "down_pct":       0.25,
        "rate":           0.0700,
        "term_years":     30,
        "rehab":          18000,
        "furniture":      18000,
        "reserves":       9000,
        "occupancy":      0.64,
        "adr":            215,
        "insurance":      2700,
        "prop_tax":       4200,
        "utilities":      1800,
        "mgmt_pct":       0.0,
        "cleaning":       5400,
        "supplies":       1100,
    },
}

# Property column letters on Inputs/Underwriting tabs (B=label, C=A, D=B, E=C)
PROP_COLS = {"A": "C", "B": "D", "C": "E"}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label, span="A:L"):
    first, last = span.split(":")
    ws.merge_cells(f"{first}{row}:{last}{row}")
    c = ws[f"{first}{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _label_cell(ws, row, text, col=2):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)


def _input_cell(ws, row, col_letter, value, fmt=None):
    cell = ws[f"{col_letter}{row}"]
    cell.value = value
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt


def _formula_cell(ws, row, col_letter, formula, fmt=None):
    cell = ws[f"{col_letter}{row}"]
    cell.value = formula
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt


# ---------------------------------------------------------------------------
# Tab 0 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "3-Property Side-by-Side"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Compare 3 acquisition candidates head-to-head. Pick the winner."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — three-card preview (rows 10-19)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 20):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card headers row 10
    headers = [("A10:D10", "PROPERTY A"), ("E10:H10", "PROPERTY B"), ("I10:L10", "PROPERTY C")]
    for rng, lbl in headers:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = lbl
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[10].height = 22

    # Property name row 12
    name_refs = [("A12:D12", "C7"), ("E12:H12", "D7"), ("I12:L12", "E7")]
    for rng, ref in name_refs:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = f'=IF(Inputs!{ref}="","(unnamed)",Inputs!{ref})'
        c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 24

    # Cash flow Y1 row 14
    cf_label = [("A14:D14", "Y1 CASH FLOW"), ("E14:H14", "Y1 CASH FLOW"), ("I14:L14", "Y1 CASH FLOW")]
    for rng, lbl in cf_label:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = lbl
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 16

    cf_refs = [("A15:D16", "C16"), ("E15:H16", "D16"), ("I15:L16", "E16")]
    for rng, ref in cf_refs:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = f"=Underwriting!{ref}"
        c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '"$"#,##0'
    ws.row_dimensions[15].height = 22
    ws.row_dimensions[16].height = 18

    # CoC row 18
    coc_label = [("A18:D18", "CASH-ON-CASH"), ("E18:H18", "CASH-ON-CASH"), ("I18:L18", "CASH-ON-CASH")]
    for rng, lbl in coc_label:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = lbl
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 16

    coc_refs = [("A19:D19", "C12"), ("E19:H19", "D12"), ("I19:L19", "E12")]
    for rng, ref in coc_refs:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = f"=Underwriting!{ref}"
        c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.0%"
    ws.row_dimensions[19].height = 22

    # Recommendation strip rows 22-23
    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = "RECOMMENDATION"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 18

    ws.merge_cells("A23:L24")
    c = ws["A23"]
    c.value = "='Comparison Matrix'!A30"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[23].height = 30
    ws.row_dimensions[24].height = 18

    # ZONE 3 — pseudo-button nav (rows 26-28)
    pseudo_button(ws, "A26", "C28", "Inputs",
                  "'Inputs'!A1", variant="primary")
    pseudo_button(ws, "D26", "F28", "Underwriting",
                  "'Underwriting'!A1", variant="primary")
    pseudo_button(ws, "G26", "I28", "→ Compare",
                  "'Comparison Matrix'!A1", variant="accent")
    pseudo_button(ws, "J26", "L28", "Decision Notes",
                  "'Decision Notes'!A1", variant="secondary")
    for r in range(26, 29):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 31
    cta = (
        f"💡  Need to compare more than 3? The Acquisition Bundle at {BRAND_DOMAIN} "
        "includes deal analyzer + cost-to-launch + arbitrage analyzer."
    )
    ws.merge_cells("A31:L31")
    c = ws["A31"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[31].height = 36

    brand_footer(ws, 33,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L36"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 1 — Inputs
# ---------------------------------------------------------------------------

def build_inputs_tab(wb, variant):
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 3), ("B", 32),
        ("C", 18), ("D", 18), ("E", 18),
        ("F", 4),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Inputs",
                        prev_tab="Start", next_tab="Underwriting")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Three properties side-by-side. Fill A, B, C columns. Underwriting computes automatically."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column headers row 5
    ws.cell(row=5, column=2).value = ""
    for col_letter, prop_letter in [("C", "A"), ("D", "B"), ("E", "C")]:
        cell = ws[f"{col_letter}5"]
        cell.value = f"PROPERTY {prop_letter}"
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # PROPERTY section
    _section_band(ws, 6, "PROPERTY", span="A:E")

    prop_fields = [
        (7,  "Property name:",      "name",     None),
        (8,  "Address:",            "address",  None),
        (9,  "Bedrooms:",           "beds",     "0"),
        (10, "Bathrooms:",          "baths",    "0.0"),
        (11, "Sleeps:",             "sleeps",   "0"),
    ]
    for row, label, key, fmt in prop_fields:
        _label_cell(ws, row, label)
        for col_letter, prop in [("C", "A"), ("D", "B"), ("E", "C")]:
            _input_cell(ws, row, col_letter, _val(variant, SAMPLE[prop][key]), fmt)
        ws.row_dimensions[row].height = 18

    # FINANCING section
    _section_band(ws, 13, "FINANCING", span="A:E")

    fin_fields = [
        (14, "Purchase price ($):",   "purchase",   '"$"#,##0'),
        (15, "Down payment (%):",     "down_pct",   "0.0%"),
        (16, "Interest rate (%):",    "rate",       "0.000%"),
        (17, "Term (years):",         "term_years", "0"),
        (18, "Rehab budget ($):",     "rehab",      '"$"#,##0'),
        (19, "Furniture budget ($):", "furniture",  '"$"#,##0'),
        (20, "Reserves ($):",         "reserves",   '"$"#,##0'),
    ]
    for row, label, key, fmt in fin_fields:
        _label_cell(ws, row, label)
        for col_letter, prop in [("C", "A"), ("D", "B"), ("E", "C")]:
            _input_cell(ws, row, col_letter, _val(variant, SAMPLE[prop][key]), fmt)
        ws.row_dimensions[row].height = 18

    # REVENUE section
    _section_band(ws, 22, "REVENUE", span="A:E")

    rev_fields = [
        (23, "Projected occupancy (%):", "occupancy", "0.0%"),
        (24, "ADR — nightly rate ($):",  "adr",       '"$"#,##0'),
    ]
    for row, label, key, fmt in rev_fields:
        _label_cell(ws, row, label)
        for col_letter, prop in [("C", "A"), ("D", "B"), ("E", "C")]:
            _input_cell(ws, row, col_letter, _val(variant, SAMPLE[prop][key]), fmt)
        ws.row_dimensions[row].height = 18

    # OPERATING EXPENSES section (annual)
    _section_band(ws, 26, "OPERATING EXPENSES (ANNUAL)", span="A:E")

    op_fields = [
        (27, "Insurance ($):",         "insurance",  '"$"#,##0'),
        (28, "Property tax ($):",      "prop_tax",   '"$"#,##0'),
        (29, "Utilities ($):",         "utilities",  '"$"#,##0'),
        (30, "Management fee (%):",    "mgmt_pct",   "0.0%"),
        (31, "Cleaning ($/yr):",       "cleaning",   '"$"#,##0'),
        (32, "Supplies ($/yr):",       "supplies",   '"$"#,##0'),
    ]
    for row, label, key, fmt in op_fields:
        _label_cell(ws, row, label)
        for col_letter, prop in [("C", "A"), ("D", "B"), ("E", "C")]:
            _input_cell(ws, row, col_letter, _val(variant, SAMPLE[prop][key]), fmt)
        ws.row_dimensions[row].height = 18

    brand_footer(ws, 35, version_line=f"{SKU} · Inputs")

    ws.print_area = "A1:L38"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 2 — Underwriting (computed metrics per property)
# ---------------------------------------------------------------------------

def build_underwriting_tab(wb, variant):
    ws = wb.create_sheet("Underwriting")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 3), ("B", 32),
        ("C", 18), ("D", 18), ("E", 18),
        ("F", 4),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Underwriting",
                        prev_tab="Inputs", next_tab="Comparison Matrix")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Computed metrics. Don't edit — these flow from Inputs."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column headers row 5
    for col_letter, prop_letter in [("C", "A"), ("D", "B"), ("E", "C")]:
        cell = ws[f"{col_letter}5"]
        cell.value = f"PROPERTY {prop_letter}"
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # Helper: build per-property formulas referencing Inputs!{col}{row}
    # Inputs row map:
    #   purchase=14, down_pct=15, rate=16, term=17, rehab=18, furniture=19, reserves=20
    #   occ=23, adr=24
    #   insurance=27, prop_tax=28, utilities=29, mgmt_pct=30, cleaning=31, supplies=32
    def f(col):
        # Loan amount
        loan = f"(Inputs!{col}14*(1-Inputs!{col}15))"
        # Annual debt service via PMT × 12
        ds = f"(IFERROR(PMT(Inputs!{col}16/12,Inputs!{col}17*12,-{loan}),0)*12)"
        # Gross revenue
        revenue = f"(Inputs!{col}23*365*Inputs!{col}24)"
        # Operating expenses (fixed + cleaning + supplies + mgmt%)
        opex = (f"(Inputs!{col}27+Inputs!{col}28+Inputs!{col}29"
                f"+Inputs!{col}31+Inputs!{col}32"
                f"+{revenue}*Inputs!{col}30)")
        noi = f"({revenue}-{opex})"
        cash_invested = (f"(Inputs!{col}14*Inputs!{col}15"
                         f"+Inputs!{col}18+Inputs!{col}19+Inputs!{col}20)")
        cf = f"({noi}-{ds})"
        cap_rate = f"({noi}/Inputs!{col}14)"
        coc = f"({cf}/{cash_invested})"
        dscr = f"({noi}/{ds})"
        # Break-even occupancy = (opex_fixed_part + ds) / (365 * adr * (1-mgmt%) - 0)
        # Approximation: BE_occ = (fixed_opex + ds) / (365 * ADR * (1 - mgmt%))
        fixed_opex = (f"(Inputs!{col}27+Inputs!{col}28+Inputs!{col}29"
                      f"+Inputs!{col}31+Inputs!{col}32)")
        be = f"(({fixed_opex}+{ds})/(365*Inputs!{col}24*(1-Inputs!{col}30)))"
        # 5-year IRR rough: assume CF flat for 5 yrs + equity payoff at exit (3% appreciation)
        # Simplified: IRR ≈ (CF*5 + (purchase*1.16 - loan_balance_y5 - cash_invested)) / cash_invested / 5
        # We'll keep it simple — use cumulative-CF + appreciation as proxy:
        irr = (f"((({cf}*5)+(Inputs!{col}14*0.16))/{cash_invested}/5)")
        return {
            "revenue": revenue, "opex": opex, "noi": noi, "ds": ds,
            "cash_invested": cash_invested, "cf": cf,
            "cap_rate": cap_rate, "coc": coc, "dscr": dscr, "be": be,
            "irr": irr,
        }

    fA = f("C")
    fB = f("D")
    fC = f("E")

    # COMPUTED METRICS section
    _section_band(ws, 6, "COMPUTED METRICS", span="A:E")

    rows = [
        # (row, label, key, fmt, emphasize)
        (7,  "Total cash invested:",       "cash_invested", '"$"#,##0', False),
        (8,  "Gross revenue (annual):",    "revenue",       '"$"#,##0', False),
        (9,  "Operating expenses:",        "opex",          '"$"#,##0', False),
        (10, "NOI (annual):",              "noi",           '"$"#,##0', True),
        (11, "Annual debt service:",       "ds",            '"$"#,##0', False),
        (12, "Cash-on-cash:",              "coc",           "0.0%",     True),
        (13, "Cap rate:",                  "cap_rate",      "0.0%",     False),
        (14, "DSCR:",                      "dscr",          "0.00",     False),
        (15, "Break-even occupancy:",      "be",            "0.0%",     False),
        (16, "Annual cash flow (Y1):",     "cf",            '"$"#,##0', True),
        (17, "5-yr IRR (proxy):",          "irr",           "0.0%",     False),
    ]
    for row, label, key, fmt, emphasize in rows:
        _label_cell(ws, row, label)
        if emphasize:
            ws.cell(row=row, column=2).font = Font(
                name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
            )
        for col_letter, formulas in [("C", fA), ("D", fB), ("E", fC)]:
            _formula_cell(ws, row, col_letter, "=" + formulas[key], fmt)
            if emphasize:
                cell = ws[f"{col_letter}{row}"]
                cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
                cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[row].height = 22 if emphasize else 18

    brand_footer(ws, 20, version_line=f"{SKU} · Underwriting")

    ws.print_area = "A1:L23"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 3 — Comparison Matrix (per-row winner highlighted)
# ---------------------------------------------------------------------------

def build_comparison_tab(wb, variant):
    ws = wb.create_sheet("Comparison Matrix")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 3), ("B", 36),
        ("C", 18), ("D", 18), ("E", 18),
        ("F", 14),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Comparison Matrix",
                        prev_tab="Underwriting", next_tab="Decision Notes")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Per-row winner highlighted in gold. Recommendation auto-computes if one wins ≥ 10/16 metrics."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column headers row 5
    headers = [(2, "Metric"), (3, "Property A"), (4, "Property B"), (5, "Property C"), (6, "Direction")]
    for col, label in headers:
        cell = ws.cell(row=5, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # Metrics: (label, formula key from underwriting OR direct, fmt, direction)
    # direction: "max" = higher is better; "min" = lower is better
    # Each row pulls from Underwriting OR computes directly from Inputs.
    # We'll reference Underwriting directly (rows 7-17 there).
    metrics = [
        # (label, A_ref, B_ref, C_ref, fmt, direction)
        ("Purchase price",          "Inputs!C14",         "Inputs!D14",         "Inputs!E14",         '"$"#,##0', "min"),
        ("Total cash invested",     "Underwriting!C7",    "Underwriting!D7",    "Underwriting!E7",    '"$"#,##0', "min"),
        ("Gross revenue (Y1)",      "Underwriting!C8",    "Underwriting!D8",    "Underwriting!E8",    '"$"#,##0', "max"),
        ("NOI",                     "Underwriting!C10",   "Underwriting!D10",   "Underwriting!E10",   '"$"#,##0', "max"),
        ("Annual cash flow (Y1)",   "Underwriting!C16",   "Underwriting!D16",   "Underwriting!E16",   '"$"#,##0', "max"),
        ("Cash-on-cash",            "Underwriting!C12",   "Underwriting!D12",   "Underwriting!E12",   "0.0%",     "max"),
        ("Cap rate",                "Underwriting!C13",   "Underwriting!D13",   "Underwriting!E13",   "0.0%",     "max"),
        ("DSCR",                    "Underwriting!C14",   "Underwriting!D14",   "Underwriting!E14",   "0.00",     "max"),
        ("Break-even occupancy",    "Underwriting!C15",   "Underwriting!D15",   "Underwriting!E15",   "0.0%",     "min"),
        ("5-yr IRR (proxy)",        "Underwriting!C17",   "Underwriting!D17",   "Underwriting!E17",   "0.0%",     "max"),
        ("Operating expenses",      "Underwriting!C9",    "Underwriting!D9",    "Underwriting!E9",    '"$"#,##0', "min"),
        ("Annual debt service",     "Underwriting!C11",   "Underwriting!D11",   "Underwriting!E11",   '"$"#,##0', "min"),
        ("ADR",                     "Inputs!C24",         "Inputs!D24",         "Inputs!E24",         '"$"#,##0', "max"),
        ("Occupancy",               "Inputs!C23",         "Inputs!D23",         "Inputs!E23",         "0.0%",     "max"),
        ("Bedrooms",                "Inputs!C9",          "Inputs!D9",          "Inputs!E9",          "0",        "max"),
        ("Sleeps",                  "Inputs!C11",         "Inputs!D11",         "Inputs!E11",         "0",        "max"),
    ]

    start_row = 6
    for i, (label, a_ref, b_ref, c_ref, fmt, direction) in enumerate(metrics):
        r = start_row + i
        # Label
        cell = ws.cell(row=r, column=2, value=label)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Banding on alt rows
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        # Values
        for col_letter, ref in [("C", a_ref), ("D", b_ref), ("E", c_ref)]:
            cell = ws[f"{col_letter}{r}"]
            cell.value = f"={ref}"
            cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = fmt
        # Direction
        dir_cell = ws.cell(row=r, column=6,
                            value="↑ higher wins" if direction == "max" else "↓ lower wins")
        dir_cell.font = Font(name=FONT_MONO, size=9, italic=True, color=COLOR_MUTED)
        dir_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 20

    end_row = start_row + len(metrics) - 1

    # Conditional formatting: highlight winner per row in each col C/D/E
    # For "max" direction: cell == MAX(C:E) → gold-soft fill
    # For "min" direction: cell == MIN(C:E) → gold-soft fill
    gold_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for i, (label, a_ref, b_ref, c_ref, fmt, direction) in enumerate(metrics):
        r = start_row + i
        for col_letter in ("C", "D", "E"):
            if direction == "max":
                rule_formula = f'AND(${col_letter}${r}<>"",${col_letter}${r}=MAX($C${r}:$E${r}))'
            else:
                rule_formula = f'AND(${col_letter}${r}<>"",${col_letter}${r}=MIN($C${r}:$E${r}))'
            rule = FormulaRule(formula=[rule_formula], fill=gold_fill,
                               font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY))
            ws.conditional_formatting.add(f"{col_letter}{r}", rule)

    # WIN COUNT row (end_row + 2)
    win_row = end_row + 2
    cell = ws.cell(row=win_row, column=2, value="WINS PER PROPERTY")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    # For each column, count rows where the value matches winner direction.
    # We use SUMPRODUCT over the metric range with a per-row direction flag.
    # Simpler: hardcode the formula using nested IF list — but that gets long.
    # Cleanest: per-property tally formula counting MAX/MIN matches per row.
    # We'll express each as:
    #   SUMPRODUCT((<col>=MAX_or_MIN_per_row)*1)
    # by enumerating the direction array. Easiest: build SUMPRODUCT explicitly.

    # Build direction-aware win-count formula per property column
    def win_formula(col_letter):
        terms = []
        for i, (label, _, _, _, _, direction) in enumerate(metrics):
            r = start_row + i
            if direction == "max":
                terms.append(f'(${col_letter}{r}=MAX($C{r}:$E{r}))')
            else:
                terms.append(f'(${col_letter}{r}=MIN($C{r}:$E{r}))')
        return "=" + "+".join(terms)

    for col_letter in ("C", "D", "E"):
        cell = ws[f"{col_letter}{win_row}"]
        cell.value = win_formula(col_letter)
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.number_format = "0"
    ws.row_dimensions[win_row].height = 26

    # Recommendation row (win_row + 4) — referenced from Start tab as 'Comparison Matrix'!A30
    rec_row = win_row + 4  # for 16 metrics: start_row=6, end_row=21, win_row=23, rec_row=27
    # Brief says recommendation if any property wins ≥ 10/16 metrics. We anchor at A30 per Start tab.
    # Make sure rec is at row 30 for the Start tab reference.
    rec_row = 30

    ws.merge_cells(f"A{rec_row}:L{rec_row}")
    c = ws[f"A{rec_row}"]
    c.value = (
        f'=IF(MAX(C{win_row}:E{win_row})>=10,'
        f'"Recommendation: "&INDEX({{"Property A","Property B","Property C"}},MATCH(MAX(C{win_row}:E{win_row}),C{win_row}:E{win_row},0))&" — wins "&MAX(C{win_row}:E{win_row})&"/16 metrics",'
        f'"⚠  Close call — no clear winner. Review Decision Notes.")'
    )
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[rec_row].height = 32

    brand_footer(ws, rec_row + 3, version_line=f"{SKU} · Comparison Matrix")

    ws.freeze_panes = "C6"
    ws.print_area = f"A1:L{rec_row + 5}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 4 — Decision Notes
# ---------------------------------------------------------------------------

def build_notes_tab(wb, variant):
    ws = wb.create_sheet("Decision Notes")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 3), ("B", 18),
        ("C", 28), ("D", 28), ("E", 28),
        ("F", 4),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Decision Notes",
                        prev_tab="Comparison Matrix", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Free-form notes per property. Pros/cons, gut feel, intel from showings."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Property headers row 6
    headers = [("C", "PROPERTY A"), ("D", "PROPERTY B"), ("E", "PROPERTY C")]
    for col, label in headers:
        cell = ws[f"{col}6"]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # Property name row 7 — pulled from Inputs
    refs = [("C", "C7"), ("D", "D7"), ("E", "E7")]
    for col, ref in refs:
        cell = ws[f"{col}7"]
        cell.value = f'=IF(Inputs!{ref}="","(unnamed)",Inputs!{ref})'
        cell.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    ws.row_dimensions[7].height = 22

    # Notes sections
    sections = [
        (9,  "PROS"),
        (15, "CONS"),
        (21, "INTEL FROM SHOWING"),
        (27, "OFFER STRATEGY"),
    ]

    sample_notes = {
        "demo": {
            "C": {
                9:  "Best CoC return; strong amenity stack (hot tub + view).",
                15: "Older HVAC; furnace replacement likely Y3.",
                21: "Seller motivated — relocation. Comp ADR $268 verified.",
                27: "Offer $470K, request 2.5K closing credit.",
            },
            "D": {
                9:  "Highest ADR ceiling; lakefront premium.",
                15: "Highest cash invested; insurance bid still pending.",
                21: "Multiple offers expected. Seller wants 30-day close.",
                27: "Pass unless price drops below $600K.",
            },
            "E": {
                9:  "Lowest cash invested; lowest break-even occupancy.",
                15: "Smaller (sleeps 6) caps revenue ceiling vs A/B.",
                21: "Tenant occupied — vacancy on close not guaranteed.",
                27: "Offer $345K with vacancy contingency.",
            },
        },
        "blank": {"C": {}, "D": {}, "E": {}},
    }

    for header_row, label in sections:
        ws.merge_cells(f"B{header_row}:E{header_row}")
        c = ws[f"B{header_row}"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[header_row].height = 20

        # 5-row notes block per property
        block_start = header_row + 1
        block_end = header_row + 5
        for col_letter in ("C", "D", "E"):
            ws.merge_cells(f"{col_letter}{block_start}:{col_letter}{block_end}")
            cell = ws[f"{col_letter}{block_start}"]
            sample_val = sample_notes[variant][col_letter].get(header_row)
            cell.value = sample_val
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True, indent=1)
        for r in range(block_start, block_end + 1):
            ws.row_dimensions[r].height = 18

    # Final recommendation block row 33-37
    _section_band(ws, 33, "FINAL RECOMMENDATION", span="A:E")

    ws.merge_cells("B34:E34")
    c = ws["B34"]
    c.value = "='Comparison Matrix'!A30"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[34].height = 28

    ws.merge_cells("B35:E37")
    c = ws["B35"]
    note_default = (
        "Property A wins on cash-on-cash, NOI, and break-even. Lake (B) is overpriced for "
        "the revenue. Creek (C) has best entry price but smaller revenue ceiling. Going "
        "with A — submitting offer this week."
        if variant == "demo" else None
    )
    c.value = note_default
    apply_style(c, input_cell_style())
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(35, 38):
        ws.row_dimensions[r].height = 22

    brand_footer(ws, 40, version_line=f"{SKU} · Decision Notes")

    ws.print_area = "A1:L43"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_inputs_tab(wb, variant)
    build_underwriting_tab(wb, variant)
    build_comparison_tab(wb, variant)
    build_notes_tab(wb, variant)

    wb.properties.title = "3-Property Side-by-Side — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Compare three STR acquisition candidates head-to-head — winner-per-metric "
        "highlight, auto recommendation, decision notes."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
