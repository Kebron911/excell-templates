"""Build PAM-001 Owner Reporting Dashboard (v2.2 standard).

Operational-mode tool for Pro Pam — manages 8-50 properties on behalf of
multiple owners, generates white-label printable owner statements, tracks
withholdings (Pam-paid expenses recovered from owner distribution), and
prepares year-end 1099-MISC worksheets.

The Owner Statement tab is the marquee: NO STR Ledger marks, Pam inserts
her own logo + company name. Owner + month pickers drive every formula.

Generates two files:
  templates/_masters/PAM-001-owner-reporting-dashboard-DEMO.xlsx
  templates/_masters/PAM-001-owner-reporting-dashboard-BLANK.xlsx
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_GRAY_LIGHT, COLOR_INK, COLOR_WHITE,
)

SKU = "PAM-001"
NAME = "owner-reporting-dashboard"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data — DEMO variant
# Per brief QA: 8 properties × 6 owners, March 2026 statement period
# ---------------------------------------------------------------------------

OWNERS_DEMO = [
    # (name, address_line, tax_id, 1099_req, email, payment_method, detail_level)
    ("Lakeside Holdings LLC",   "302 Pine Vale Rd, Sevierville TN 37862",  "82-1234567", "Yes", "owner1@lakeside.example",   "ACH",   "Standard"),
    ("Maddox Trust",            "47 Cedar Cove Ln, Gatlinburg TN 37738",   "***-**-1188", "Yes", "owner2@maddox.example",     "ACH",   "Itemized"),
    ("Highland Properties LLC", "1280 Ridge Parkway, Pigeon Forge TN 37863","82-7654321", "Yes", "owner3@highland.example",   "Wire",  "Standard"),
    ("R + L Vasquez",           "612 Hollow Branch, Cosby TN 37722",        "***-**-2204", "No",  "owner4@vasquez.example",    "Check", "Summary"),
    ("Quinn Asset Mgmt",        "89 Clearwater Way, Sevierville TN 37876",  "82-9988776", "Yes", "owner5@quinn.example",      "ACH",   "Itemized"),
    ("Erickson Family Trust",   "440 Smoky View Dr, Gatlinburg TN 37738",   "***-**-3340", "Yes", "owner6@erickson.example",   "ACH",   "Standard"),
]

# 8 properties × owner mapping
PROPERTY_OWNER_MAP = [
    # (property, owner, effective_from, effective_to, ownership_pct)
    ("Smokies Ridge Cabin",   "Lakeside Holdings LLC",   "2024-01-01", None, 1.00),
    ("Creek Side",            "Lakeside Holdings LLC",   "2024-03-01", None, 1.00),
    ("Lakehouse A",           "Lakeside Holdings LLC",   "2024-09-15", None, 1.00),
    ("Mountain Loft",         "Maddox Trust",            "2025-06-01", None, 1.00),
    ("Forest Cabin",          "Highland Properties LLC", "2025-08-01", None, 1.00),
    ("Cedar Hollow",          "R + L Vasquez",           "2024-11-01", None, 1.00),
    ("Riverbend Cottage",     "Quinn Asset Mgmt",        "2025-03-01", None, 1.00),
    ("Smoky View Lodge",      "Erickson Family Trust",   "2024-07-01", None, 1.00),
]

# Mgmt fee calculator — per property, base $ + % of revenue
# (property, model_label, base_dollars, pct_of_revenue, notes)
FEE_STRUCTURES = [
    ("Smokies Ridge Cabin",   "% of revenue",       0,    0.20, ""),
    ("Creek Side",            "% of revenue",       0,    0.20, ""),
    ("Lakehouse A",           "% of revenue",       0,    0.20, ""),
    ("Mountain Loft",         "% of revenue",       0,    0.15, ""),
    ("Forest Cabin",          "% of revenue",       0,    0.15, ""),
    ("Cedar Hollow",          "% of revenue",       0,    0.15, ""),
    ("Riverbend Cottage",     "Hybrid (base + %)",  800,  0.05, "$800 base + 5% revenue"),
    ("Smoky View Lodge",      "Hybrid (base + %)",  800,  0.05, "$800 base + 5% revenue"),
]

# Bookings — March 2026 (~$58K total per QA), plus a few Feb for balance-forward demo
BOOKING_SAMPLES = [
    # (date, property, guest, channel, gross, platform_fee, cleaning_collected, notes)
    # Owner 1 (Lakeside) — 3 properties × 3-4 bookings each
    ("2026-03-02", "Smokies Ridge Cabin", "Avery G.",      "Airbnb",    1880, 282, 150, ""),
    ("2026-03-08", "Smokies Ridge Cabin", "Marshall H.",   "Airbnb",    2340, 351, 150, ""),
    ("2026-03-15", "Smokies Ridge Cabin", "Priya R.",      "VRBO",      2520, 202, 150, ""),
    ("2026-03-22", "Smokies Ridge Cabin", "Casey N.",      "Direct",    2460,   0, 150, "Direct booking 5% off"),
    ("2026-03-28", "Smokies Ridge Cabin", "Devon S.",      "Airbnb",    1340, 201, 150, ""),
    ("2026-03-04", "Creek Side",          "Hailey M.",     "Airbnb",    1640, 246, 135, ""),
    ("2026-03-12", "Creek Side",          "Patrice O.",    "Airbnb",    2100, 315, 135, ""),
    ("2026-03-19", "Creek Side",          "Chen W.",       "Airbnb",    1880, 282, 135, ""),
    ("2026-03-27", "Creek Side",          "Tomás A.",      "VRBO",      2080, 166, 135, ""),
    ("2026-03-05", "Lakehouse A",         "Reese F.",      "Airbnb",    1840, 276, 175, ""),
    ("2026-03-13", "Lakehouse A",         "Olympia V.",    "Direct",    1920,   0, 175, ""),
    ("2026-03-21", "Lakehouse A",         "Sasha B.",      "Airbnb",    2340, 351, 175, ""),
    ("2026-03-29", "Lakehouse A",         "Wells D.",      "Airbnb",    2125, 319, 175, ""),
    # Owner 2 (Maddox)
    ("2026-03-03", "Mountain Loft",       "Indira K.",     "Airbnb",    1680, 252, 110, ""),
    ("2026-03-10", "Mountain Loft",       "Brennan E.",    "Airbnb",    1480, 222, 110, ""),
    ("2026-03-18", "Mountain Loft",       "Nadia L.",      "Airbnb",    1080, 162, 110, ""),
    ("2026-03-26", "Mountain Loft",       "Kendall U.",    "Airbnb",     960, 144, 110, ""),
    # Owner 3 (Highland)
    ("2026-03-04", "Forest Cabin",        "Ezra Y.",       "Airbnb",    1240, 186, 100, ""),
    ("2026-03-12", "Forest Cabin",        "Saoirse C.",    "VRBO",      1480, 118, 100, ""),
    ("2026-03-22", "Forest Cabin",        "Jude T.",       "Airbnb",    1480, 222, 100, ""),
    ("2026-03-30", "Forest Cabin",        "Annika W.",     "Airbnb",     680, 102, 100, ""),
    # Owner 4 (Vasquez) — has carryover negative balance from Feb
    ("2026-03-05", "Cedar Hollow",        "Felipe O.",     "Airbnb",    1380, 207, 145, ""),
    ("2026-03-13", "Cedar Hollow",        "Jordan P.",     "Direct",    1620,   0, 145, ""),
    ("2026-03-22", "Cedar Hollow",        "Bryn N.",       "Airbnb",    1320, 198, 145, ""),
    ("2026-03-29", "Cedar Hollow",        "Mira K.",       "Airbnb",    1180, 177, 145, ""),
    # Owner 5 (Quinn) — hybrid fee
    ("2026-03-02", "Riverbend Cottage",   "Roy + Linn T.", "Airbnb",    2240, 336, 165, ""),
    ("2026-03-11", "Riverbend Cottage",   "Quincy J.",     "VRBO",      1680, 134, 165, ""),
    ("2026-03-19", "Riverbend Cottage",   "Patrice C.",    "Airbnb",    1820, 273, 165, ""),
    ("2026-03-27", "Riverbend Cottage",   "Indigo S.",     "Airbnb",    1660, 249, 165, ""),
    # Owner 6 (Erickson) — hybrid fee
    ("2026-03-01", "Smoky View Lodge",    "Avi M.",        "Airbnb",    2680, 402, 195, ""),
    ("2026-03-09", "Smoky View Lodge",    "Pia W.",        "VRBO",      2240, 179, 195, ""),
    ("2026-03-18", "Smoky View Lodge",    "Halle J.",      "Airbnb",    2540, 381, 195, ""),
    ("2026-03-26", "Smoky View Lodge",    "Kazuki H.",     "Airbnb",    2040, 306, 195, ""),
]

# Expenses — direct (booking-related) + Pam-paid (withholdings)
EXPENSE_SAMPLES = [
    # (date, vendor, property, category, amount, payment, pam_paid, withholding_month, notes)
    # Direct (cleaner per turnover, etc.)
    ("2026-03-03", "Sarah Cleans",        "Smokies Ridge Cabin", "Cleaning", 185, "Venmo", "No",  "", ""),
    ("2026-03-09", "Sarah Cleans",        "Smokies Ridge Cabin", "Cleaning", 185, "Venmo", "No",  "", ""),
    ("2026-03-16", "Sarah Cleans",        "Smokies Ridge Cabin", "Cleaning", 185, "Venmo", "No",  "", ""),
    ("2026-03-23", "Sarah Cleans",        "Smokies Ridge Cabin", "Cleaning", 185, "Venmo", "No",  "", ""),
    ("2026-03-29", "Sarah Cleans",        "Smokies Ridge Cabin", "Cleaning", 185, "Venmo", "No",  "", ""),
    ("2026-03-05", "Miguel HK",           "Creek Side",          "Cleaning", 165, "Venmo", "No",  "", ""),
    ("2026-03-13", "Miguel HK",           "Creek Side",          "Cleaning", 165, "Venmo", "No",  "", ""),
    ("2026-03-20", "Miguel HK",           "Creek Side",          "Cleaning", 165, "Venmo", "No",  "", ""),
    ("2026-03-28", "Miguel HK",           "Creek Side",          "Cleaning", 165, "Venmo", "No",  "", ""),
    ("2026-03-06", "Sarah Cleans",        "Lakehouse A",         "Cleaning", 220, "Venmo", "No",  "", ""),
    ("2026-03-14", "Sarah Cleans",        "Lakehouse A",         "Cleaning", 220, "Venmo", "No",  "", ""),
    ("2026-03-22", "Sarah Cleans",        "Lakehouse A",         "Cleaning", 220, "Venmo", "No",  "", ""),
    # Pam-paid withholdings — recovered Mar 2026
    ("2026-03-09", "Mountain HVAC",       "Creek Side",          "Repairs",  480, "Card",  "Yes", "2026-03", "HVAC unit replacement — Pam-paid"),
    ("2026-03-15", "Costco",              "Lakehouse A",         "Linens",   200, "Card",  "Yes", "2026-03", "Replacement linens — Pam-paid"),
    ("2026-03-19", "Smoky Plumbing",      "Forest Cabin",        "Repairs",  250, "Check", "Yes", "2026-03", "Toilet flapper — Pam-paid"),
    ("2026-03-22", "Mountain HVAC",       "Cedar Hollow",        "Repairs",  190, "Card",  "Yes", "2026-03", "HVAC tune-up — Pam-paid"),
    ("2026-03-25", "Costco",              "Riverbend Cottage",   "Supplies", 120, "Card",  "Yes", "2026-03", "Cleaning supply restock — Pam-paid"),
    # Feb storm damage — created Feb negative balance forward for Owner 4 (Vasquez/Cedar Hollow)
    ("2026-02-14", "Brookline Roofing",   "Cedar Hollow",        "Repairs",  340, "Check", "Yes", "2026-02", "Storm damage — caused Feb negative balance"),
]

# Statement starting balance (carry-forward from prior month)
# Stored on Settings tab — for DEMO, Owner 4 (Vasquez) starts March with -$340
STARTING_BALANCES = {
    "Lakeside Holdings LLC":   0,
    "Maddox Trust":            0,
    "Highland Properties LLC": 0,
    "R + L Vasquez":           -340,
    "Quinn Asset Mgmt":        0,
    "Erickson Family Trust":   0,
}

CHANNELS = ["Airbnb", "VRBO", "Booking.com", "Direct", "Other"]
PAYMENT_METHODS = ["ACH", "Wire", "Check", "Stripe", "PayPal", "Other"]
PAM_PAYMENT_METHODS = ["Venmo", "Zelle", "Check", "Cash", "ACH", "Card", "Auto", "Other"]
DETAIL_LEVELS = ["Summary", "Standard", "Itemized"]
FEE_MODELS = ["Flat $", "% of revenue", "Hybrid (base + %)", "Custom (override)"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list):
    return demo_list if variant == "demo" else []


def _parse_date(s):
    if not s:
        return None
    return date.fromisoformat(s)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, max_col=12):
    last = get_column_letter(max_col)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Owner Reporting Dashboard"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Statements your owners forward to their CPAs."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # At-a-glance cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    cards = [
        ("A", "C", "OWNERS",
         '=COUNTA(\'Owners Register\'!A6:A35)',
         "0", COLOR_PRIMARY, "owners under management"),
        ("D", "F", "PROPERTIES",
         '=COUNTA(\'Property↔Owner Map\'!A6:A55)',
         "0", COLOR_ACCENT, "active in portfolio"),
        ("G", "I", "MTD GROSS REVENUE",
         '=SUMPRODUCT((MONTH(\'Booking Log\'!$A$6:$A$5005)=MONTH(TODAY()))*'
         '(YEAR(\'Booking Log\'!$A$6:$A$5005)=YEAR(TODAY()))*'
         '\'Booking Log\'!$E$6:$E$5005)',
         '"$"#,##0', COLOR_PRIMARY, "this month, all properties"),
        ("J", "L", "WITHHOLDINGS OPEN",
         '=SUMIFS(\'Expense Log\'!$E$6:$E$8005,'
         '\'Expense Log\'!$G$6:$G$8005,"Yes",'
         '\'Expense Log\'!$H$6:$H$8005,TEXT(TODAY(),"yyyy-mm"))',
         '"$"#,##0', COLOR_SECONDARY, "Pam-paid this month"),
    ]
    for fc, lc, label, formula, fmt, color, sub in cards:
        ws.merge_cells(f"{fc}10:{lc}10")
        c = ws[f"{fc}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=8, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{fc}11:{lc}13")
        c = ws[f"{fc}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=24, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        ws.merge_cells(f"{fc}14:{lc}15")
        c = ws[f"{fc}14"]
        c.value = sub
        c.font = Font(name=FONT_BODY, size=8, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for fc, lc, *_ in cards:
        first_c = column_index_from_string(fc)
        last_c = column_index_from_string(lc)
        for r in range(10, 16):
            for col in range(first_c, last_c + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == first_c else existing.left,
                    right=gold_side if col == last_c else existing.right,
                )

    # Workflow callout rows 17-22
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "MONTH-END WORKFLOW"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    workflow_steps = [
        "①  Confirm all bookings + Pam-paid expenses for the month are logged.",
        "②  Open Owner Statement → pick owner + month → review the statement that auto-populates.",
        "③  Print/PDF the statement (white-label — your logo, not ours). Email to owner.",
        "④  After payment, copy the closing balance to next month's starting balance on Settings.",
        "⑤  At year-end, open the 1099 Worksheet — owners with distributions ≥ $600 are flagged.",
    ]
    for i, step in enumerate(workflow_steps):
        row = 18 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = step
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22

    # Pseudo-button nav rows 25-27 + 29-31
    pseudo_button(ws, "A25", "C27", "Owners Register",
                  "'Owners Register'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "Property↔Owner Map",
                  "'Property↔Owner Map'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Booking Log",
                  "'Booking Log'!A1", variant="primary")
    pseudo_button(ws, "J25", "L27", "Expense Log",
                  "'Expense Log'!A1", variant="primary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "C31", "Mgmt Fee Calculator",
                  "'Mgmt Fee Calculator'!A1", variant="secondary")
    pseudo_button(ws, "D29", "F31", "→  Owner Statement",
                  "'Owner Statement'!A1", variant="accent")
    pseudo_button(ws, "G29", "I31", "Year-End 1099",
                  "'1099 Worksheet'!A1", variant="secondary")
    pseudo_button(ws, "J29", "L31", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "💡  Manage cleaners + commissions + multi-owner reporting in one bundle? "
        f"Get the Pro Manager Bundle at {BRAND_DOMAIN}/pro-manager — $797 "
        "(saves $1,000+ vs individual SKUs)."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Owners Register
# ---------------------------------------------------------------------------

def build_owners_register_tab(wb, variant):
    ws = wb.create_sheet("Owners Register")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Owners Register",
                        prev_tab="Start", next_tab="Property↔Owner Map")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per owner. Capacity 30. Tax IDs visible only to you — "
        "the Owner Statement tab does NOT print this register."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Owner / entity name", "Mailing address", "Tax ID (SSN/EIN)",
        "1099 required?", "Email", "Pmt method", "Statement detail",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 28), ("B", 38), ("C", 16), ("D", 14),
        ("E", 28), ("F", 12), ("G", 14),
    ])

    samples = _val_list(variant, OWNERS_DEMO)
    for i in range(30):
        row = 6 + i
        if i < len(samples):
            cells = list(samples[i])
        else:
            cells = [None] * 7
        for col, val in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    add_dropdown(ws, "D6:D35", '"Yes,No"')
    add_dropdown(ws, "F6:F35", f'"{",".join(PAYMENT_METHODS)}"')
    add_dropdown(ws, "G6:G35", f'"{",".join(DETAIL_LEVELS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Property↔Owner Map
# ---------------------------------------------------------------------------

def build_property_owner_map_tab(wb, variant):
    ws = wb.create_sheet("Property↔Owner Map")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Property↔Owner Map",
                        prev_tab="Owners Register", next_tab="Booking Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Maps each property to its owner. Capacity 50. Drives every per-owner "
        "calculation. Property names here MUST match Booking Log + Expense Log."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = ["Property", "Owner", "Effective from", "Effective to", "Ownership %"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 26), ("B", 28), ("C", 14), ("D", 14), ("E", 14),
    ])

    samples = _val_list(variant, PROPERTY_OWNER_MAP)
    for i in range(50):
        row = 6 + i
        if i < len(samples):
            (prop, owner, eff_from, eff_to, own_pct) = samples[i]
        else:
            prop = owner = eff_from = eff_to = None
            own_pct = None

        for col, val, fmt in [
            (1, prop, None),
            (2, owner, None),
            (3, _parse_date(eff_from) if eff_from else None, "yyyy-mm-dd"),
            (4, _parse_date(eff_to) if eff_to else None, "yyyy-mm-dd"),
            (5, own_pct, "0%"),
        ]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    add_dropdown(ws, "B6:B55", "='Owners Register'!$A$6:$A$35")

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Booking Log
# ---------------------------------------------------------------------------

def build_booking_log_tab(wb, variant):
    ws = wb.create_sheet("Booking Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Booking Log",
                        prev_tab="Property↔Owner Map", next_tab="Expense Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "All bookings, all properties. Capacity 5000 rows. Net = Gross − Platform fee."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Date", "Property", "Guest", "Channel",
        "Gross", "Platform Fee", "Cleaning Collected", "Net", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 12), ("B", 24), ("C", 22), ("D", 12),
        ("E", 12), ("F", 12), ("G", 14), ("H", 12),
        ("I", 28),
    ])

    samples = _val_list(variant, BOOKING_SAMPLES)
    for i in range(5000):
        row = 6 + i
        if i < len(samples):
            (date_s, prop, guest, channel, gross, fee, clean, notes) = samples[i]
            cells = [
                (1, _parse_date(date_s), "yyyy-mm-dd"),
                (2, prop, None),
                (3, guest, None),
                (4, channel, None),
                (5, gross, '"$"#,##0.00'),
                (6, fee, '"$"#,##0.00'),
                (7, clean, '"$"#,##0.00'),
                (9, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None),
                (3, None, None),
                (4, None, None),
                (5, None, '"$"#,##0.00'),
                (6, None, '"$"#,##0.00'),
                (7, None, '"$"#,##0.00'),
                (9, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        net_cell = ws.cell(row=row, column=8, value=f"=E{row}-F{row}")
        apply_style(net_cell, formula_cell_style())
        net_cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 16

    add_dropdown(ws, "B6:B5005", "='Property↔Owner Map'!$A$6:$A$55")
    add_dropdown(ws, "D6:D5005", f'"{",".join(CHANNELS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Expense Log
# ---------------------------------------------------------------------------

def build_expense_log_tab(wb, variant):
    ws = wb.create_sheet("Expense Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Expense Log",
                        prev_tab="Booking Log", next_tab="Mgmt Fee Calculator")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "All expenses, all properties. Capacity 8000. Pam-paid? = Yes means "
        "you fronted the cost — it'll be withheld from owner's distribution "
        "in the month set under Withholding-month (yyyy-mm)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Date", "Vendor", "Property", "Category", "Amount",
        "Pmt method", "Pam-paid?", "Withholding month (yyyy-mm)", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 12), ("B", 22), ("C", 24), ("D", 16),
        ("E", 12), ("F", 12), ("G", 12), ("H", 18), ("I", 30),
    ])

    samples = _val_list(variant, EXPENSE_SAMPLES)
    for i in range(8000):
        row = 6 + i
        if i < len(samples):
            (date_s, vendor, prop, cat, amount, pmt, pam_paid, wh_month, notes) = samples[i]
            cells = [
                (1, _parse_date(date_s), "yyyy-mm-dd"),
                (2, vendor, None),
                (3, prop, None),
                (4, cat, None),
                (5, amount, '"$"#,##0.00'),
                (6, pmt, None),
                (7, pam_paid, None),
                (8, wh_month, None),
                (9, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None),
                (3, None, None),
                (4, None, None),
                (5, None, '"$"#,##0.00'),
                (6, None, None),
                (7, None, None),
                (8, None, None),
                (9, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 16

    add_dropdown(ws, "C6:C8005", "='Property↔Owner Map'!$A$6:$A$55")
    add_dropdown(ws, "F6:F8005", f'"{",".join(PAM_PAYMENT_METHODS)}"')
    add_dropdown(ws, "G6:G8005", '"Yes,No"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Mgmt Fee Calculator
# ---------------------------------------------------------------------------

def build_mgmt_fee_calculator_tab(wb, variant):
    ws = wb.create_sheet("Mgmt Fee Calculator")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Mgmt Fee Calculator",
                        prev_tab="Expense Log", next_tab="Owner Statement")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-property fee structure. Total fee per month = Base $ + (% of revenue × monthly revenue). "
        "For pure-flat: set % to 0. For pure-percent: set base to 0. For hybrid: set both."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = ["Property", "Fee model label", "Base $ (monthly)", "% of revenue", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 26), ("B", 22), ("C", 18), ("D", 14), ("E", 38),
    ])

    samples = _val_list(variant, FEE_STRUCTURES)
    for i in range(50):
        row = 6 + i
        if i < len(samples):
            (prop, model, base, pct, notes) = samples[i]
        else:
            prop = model = notes = None
            base = pct = None

        for col, val, fmt in [
            (1, prop, None),
            (2, model, None),
            (3, base, '"$"#,##0'),
            (4, pct, "0.0%"),
            (5, notes, None),
        ]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    add_dropdown(ws, "A6:A55", "='Property↔Owner Map'!$A$6:$A$55")
    add_dropdown(ws, "B6:B55", f'"{",".join(FEE_MODELS)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 7 — Owner Statement (THE marquee tab — WHITE-LABEL, no STR Ledger marks)
# ---------------------------------------------------------------------------

def build_owner_statement_tab(wb, variant):
    ws = wb.create_sheet("Owner Statement")
    # Tab color subtle — clay rose to differentiate but not screaming
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 26), ("C", 14), ("D", 14),
        ("E", 14), ("F", 14), ("G", 14), ("H", 4),
    ])

    # NOTE: this tab has NO compact_header_band, NO brand_footer.
    # White-label layout below.

    # Tiny back link in row 1 (NOT in print area, so it doesn't print on owner copy)
    pseudo_button(ws, "B1", "C1", "← back to dashboard",
                  "'Start'!A1", variant="secondary")
    ws.row_dimensions[1].height = 16

    # Print area starts row 3.
    # Row 2: empty intentional whitespace
    ws.row_dimensions[2].height = 12

    # Logo placeholder rows 3-7 cols A-C
    ws.merge_cells("B3:C7")
    c = ws["B3"]
    c.value = "[ INSERT YOUR LOGO HERE ]"
    c.font = Font(name=FONT_MONO, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    dashed = Side(style="dashed", color=COLOR_GRAY_LIGHT)
    for r in range(3, 8):
        for col in range(2, 4):
            ws.cell(row=r, column=col).border = Border(
                left=dashed, right=dashed, top=dashed, bottom=dashed,
            )
    for r in range(3, 8):
        ws.row_dimensions[r].height = 16

    # Company name (cols D-G) — pulled from Settings (Pam edits there)
    ws.merge_cells("D3:G3")
    c = ws["D3"]
    c.value = "=Settings!B14"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_INK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 28

    ws.merge_cells("D4:G4")
    c = ws["D4"]
    c.value = "=Settings!B15"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("D5:G5")
    c = ws["D5"]
    c.value = "=Settings!B16"
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("D6:G6")
    c = ws["D6"]
    c.value = "=Settings!B17"
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Title
    ws.merge_cells("B9:G9")
    c = ws["B9"]
    c.value = "OWNER STATEMENT"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_INK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 32

    # Gold rule
    rule_side = Side(style="thin", color=COLOR_ACCENT)
    for col in range(2, 8):
        ws.cell(row=10, column=col).border = Border(top=rule_side)
    ws.row_dimensions[10].height = 6

    # Period + Owner pickers row 12-14
    ws.cell(row=12, column=2, value="Statement period:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=12, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=12, column=3, value=3 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    add_dropdown(ws, "C12", '"1,2,3,4,5,6,7,8,9,10,11,12"')

    ws.cell(row=12, column=4, value="month").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )

    cell = ws.cell(row=12, column=5, value=2026 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

    ws.cell(row=12, column=6, value="year").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.row_dimensions[12].height = 24

    # Owner picker
    ws.cell(row=13, column=2, value="Owner:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=13, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)

    ws.merge_cells("C13:G13")
    cell = ws["C13"]
    cell.value = "Lakeside Holdings LLC" if variant == "demo" else None
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    add_dropdown(ws, "C13", "='Owners Register'!$A$6:$A$35")
    ws.row_dimensions[13].height = 26

    # Owner mailing address (auto pull)
    ws.cell(row=14, column=2, value="Mailing to:").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=14, column=2).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells("C14:G14")
    cell = ws["C14"]
    cell.value = (
        '=IFERROR(VLOOKUP(C13,\'Owners Register\'!$A$6:$B$35,2,FALSE),"")'
    )
    cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[14].height = 18

    # Section: PROPERTIES THIS MONTH (rows 17-26)
    ws.merge_cells("B17:G17")
    c = ws["B17"]
    c.value = "PROPERTIES THIS MONTH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_INK)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    # Header row 18
    sub_headers = [
        ("B18", "Property"),
        ("C18", "Bookings"),
        ("D18", "Gross revenue"),
        ("E18", "Cleaning"),
        ("F18", "Mgmt fee"),
        ("G18", "Subtotal"),
    ]
    for ref, lbl in sub_headers:
        cell = ws[ref]
        cell.value = lbl
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_INK)
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=COLOR_MUTED))
    ws.row_dimensions[18].height = 22

    # Property rows 19-26 (8 visible — Property↔Owner Map first 8 entries
    # filtered by the picker via formulas)
    # We list rows from Property↔Owner Map and only show rows where the owner matches
    for i in range(8):
        row = 19 + i
        po_row = 6 + i  # Property↔Owner Map source row

        # Col B: property name (only if it belongs to selected owner)
        formula_b = (
            f'=IF(\'Property↔Owner Map\'!B{po_row}=$C$13,'
            f'\'Property↔Owner Map\'!A{po_row},"")'
        )
        cell = ws.cell(row=row, column=2, value=formula_b)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col C: # bookings this month
        formula_c = (
            f'=IF(B{row}="","",SUMPRODUCT('
            f'(MONTH(\'Booking Log\'!$A$6:$A$5005)=$C$12)*'
            f'(YEAR(\'Booking Log\'!$A$6:$A$5005)=$E$12)*'
            f'(\'Booking Log\'!$B$6:$B$5005=B{row})*1))'
        )
        cell = ws.cell(row=row, column=3, value=formula_c)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"

        # Col D: gross revenue this month
        formula_d = (
            f'=IF(B{row}="","",SUMPRODUCT('
            f'(MONTH(\'Booking Log\'!$A$6:$A$5005)=$C$12)*'
            f'(YEAR(\'Booking Log\'!$A$6:$A$5005)=$E$12)*'
            f'(\'Booking Log\'!$B$6:$B$5005=B{row})*'
            f'\'Booking Log\'!$E$6:$E$5005))'
        )
        cell = ws.cell(row=row, column=4, value=formula_d)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0.00'

        # Col E: cleaning collected
        formula_e = (
            f'=IF(B{row}="","",SUMPRODUCT('
            f'(MONTH(\'Booking Log\'!$A$6:$A$5005)=$C$12)*'
            f'(YEAR(\'Booking Log\'!$A$6:$A$5005)=$E$12)*'
            f'(\'Booking Log\'!$B$6:$B$5005=B{row})*'
            f'\'Booking Log\'!$G$6:$G$5005))'
        )
        cell = ws.cell(row=row, column=5, value=formula_e)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0.00'

        # Col F: mgmt fee = base + pct × revenue (lookup in Mgmt Fee Calculator)
        formula_f = (
            f'=IF(B{row}="","",-(IFERROR(VLOOKUP(B{row},'
            f'\'Mgmt Fee Calculator\'!$A$6:$D$55,3,FALSE),0)'
            f'+IFERROR(VLOOKUP(B{row},'
            f'\'Mgmt Fee Calculator\'!$A$6:$D$55,4,FALSE),0)*D{row}))'
        )
        cell = ws.cell(row=row, column=6, value=formula_f)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_ERROR)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0.00'

        # Col G: subtotal = gross + cleaning + mgmt_fee (mgmt is negative)
        formula_g = (
            f'=IF(B{row}="","",D{row}+E{row}+F{row})'
        )
        cell = ws.cell(row=row, column=7, value=formula_g)
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_INK)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[row].height = 18

    # Properties total row 28
    ws.cell(row=28, column=2, value="Properties subtotal:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_INK
    )
    ws.cell(row=28, column=2).alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(row=28, column=4, value="=SUM(D19:D26)")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_INK)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell = ws.cell(row=28, column=5, value="=SUM(E19:E26)")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_INK)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell = ws.cell(row=28, column=6, value="=SUM(F19:F26)")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell = ws.cell(row=28, column=7, value="=SUM(G19:G26)")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_INK)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[28].height = 24

    # WITHHOLDINGS section rows 30-37
    ws.merge_cells("B30:G30")
    c = ws["B30"]
    c.value = "WITHHOLDINGS THIS MONTH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_INK)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[30].height = 22

    ws.merge_cells("B31:G31")
    c = ws["B31"]
    c.value = (
        "Expenses I (your manager) paid on your behalf this month, to be "
        "deducted from your distribution. Receipts available on request."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[31].height = 18

    # Single sum row 33
    ws.cell(row=33, column=2, value="Total withheld:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_INK
    )
    ws.cell(row=33, column=2).alignment = Alignment(horizontal="right", vertical="center")

    # Withholding formula: SUMPRODUCT over Expense Log where:
    # - Pam-paid? = "Yes"
    # - Withholding-month = TEXT(DATE(year, month, 1),"yyyy-mm")
    # - Property maps to selected owner (via VLOOKUP into Property↔Owner Map)
    formula = (
        '=-SUMPRODUCT('
        '(\'Expense Log\'!$G$6:$G$8005="Yes")*'
        '(\'Expense Log\'!$H$6:$H$8005=TEXT(DATE($E$12,$C$12,1),"yyyy-mm"))*'
        '(IFERROR(VLOOKUP(\'Expense Log\'!$C$6:$C$8005,'
        '\'Property↔Owner Map\'!$A$6:$B$55,2,FALSE),"")=$C$13)*'
        '\'Expense Log\'!$E$6:$E$8005)'
    )
    cell = ws.cell(row=33, column=7, value=formula)
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[33].height = 22

    # BALANCE FORWARD section row 35-37
    ws.merge_cells("B35:G35")
    c = ws["B35"]
    c.value = "BALANCE FORWARD  +  CLOSING"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_INK)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[35].height = 22

    ws.cell(row=36, column=2, value="Balance forward (from prior period):").font = Font(
        name=FONT_BODY, size=11, color=COLOR_TEXT
    )
    ws.cell(row=36, column=2).alignment = Alignment(horizontal="right", vertical="center")
    # Pull from Settings starting balances table — VLOOKUP owner name
    formula = '=IFERROR(VLOOKUP($C$13,Settings!$A$23:$B$52,2,FALSE),0)'
    cell = ws.cell(row=36, column=7, value=formula)
    cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[36].height = 20

    # FINAL DISTRIBUTION row 38-40
    rule_side = Side(style="medium", color=COLOR_INK)
    for col in range(2, 8):
        ws.cell(row=38, column=col).border = Border(top=rule_side)
    ws.row_dimensions[38].height = 6

    ws.merge_cells("B39:F39")
    c = ws["B39"]
    c.value = "DISTRIBUTION TO OWNER"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_INK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=39, column=7, value="=G28+G33+G36")
    cell.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_INK)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[39].height = 36

    # Footer rows 42-43 (white-label — NO STR Ledger marks)
    ws.merge_cells("B42:G42")
    c = ws["B42"]
    c.value = '="Questions? "&Settings!B19'
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[42].height = 18

    ws.merge_cells("B43:G43")
    c = ws["B43"]
    c.value = '="Generated "&TEXT(TODAY(),"mmmm d, yyyy")'
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[43].height = 18

    # Print area covers rows 3-43 (skips row 1 nav)
    ws.print_area = "A3:H44"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 8 — Year-End 1099 Worksheet
# ---------------------------------------------------------------------------

def build_1099_worksheet_tab(wb, variant):
    ws = wb.create_sheet("1099 Worksheet")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Year-End 1099 Worksheet",
                        prev_tab="Owner Statement", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Annual distributions per owner. IRS requires 1099-MISC for non-corporate "
        "owners receiving > $600/yr. Set the year below; rest auto-pulls."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Year selector
    ws.cell(row=5, column=1, value="Tax year:").font = Font(
        name=FONT_BODY, size=12, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=5, column=2, value=2026 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if variant != "demo":
        cell.value = "=YEAR(TODAY())-1"
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 28), ("B", 32), ("C", 18), ("D", 14), ("E", 14), ("F", 14),
        ("G", 38),
    ])

    # Header row 7
    headers = ["Owner", "Mailing address", "Tax ID",
               "Annual distributions", "Required?", "Status", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 24

    # 30 rows pulling from Owners Register
    for i in range(30):
        row = 8 + i
        or_row = 6 + i  # Owners Register row

        # Col A: owner name (formula)
        formula = f'=IF(\'Owners Register\'!A{or_row}="","",\'Owners Register\'!A{or_row})'
        cell = ws.cell(row=row, column=1, value=formula)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col B: address
        formula = f'=IF(A{row}="","",\'Owners Register\'!B{or_row})'
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        # Col C: tax ID
        formula = f'=IF(A{row}="","",\'Owners Register\'!C{or_row})'
        cell = ws.cell(row=row, column=3, value=formula)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Col D: annual distributions (sum across 12 months × all owner's properties)
        # = sum of (gross + cleaning - mgmt fees - withholdings) for all owner's
        # properties across the tax year. For simplicity in v1: use total
        # gross + cleaning collected (the IRS-relevant gross figure for 1099 reporting).
        # Manager-of-record reports the gross received on owner's behalf.
        formula = (
            f'=IF(A{row}="","",SUMPRODUCT('
            f'(YEAR(\'Booking Log\'!$A$6:$A$5005)=$B$5)*'
            f'(IFERROR(VLOOKUP(\'Booking Log\'!$B$6:$B$5005,'
            f'\'Property↔Owner Map\'!$A$6:$B$55,2,FALSE),"")=A{row})*'
            f'(\'Booking Log\'!$E$6:$E$5005+\'Booking Log\'!$G$6:$G$5005)))'
        )
        cell = ws.cell(row=row, column=4, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0.00'

        # Col E: 1099 required? (Yes if distributions ≥ 600 and owner.1099_req = Yes)
        formula = (
            f'=IF(A{row}="","",IF(AND(D{row}>=600,'
            f'\'Owners Register\'!D{or_row}="Yes"),"Yes","No"))'
        )
        cell = ws.cell(row=row, column=5, value=formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # Col F: status (Pam tracks: Pending / Filed / N/A)
        cell = ws.cell(row=row, column=6, value=None)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        add_dropdown(ws, f"F{row}", '"Pending,Filed,N/A"')

        # Col G: notes
        cell = ws.cell(row=row, column=7, value=None)
        apply_style(cell, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Conditional formatting on col E (Required?) — gold-soft when Yes
    ws.conditional_formatting.add(
        "E8:E37",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )

    # Footer note
    ws.merge_cells("A39:G41")
    c = ws["A39"]
    c.value = (
        "IRS rule: a 1099-MISC (Box 1 — Rents) is required when you, as "
        "property manager, pay ≥ $600 in rents to a non-corporate owner during "
        "the tax year. Corporate owners (LLC taxed as C-corp, S-corp, partnership "
        "with corporate election) are exempt. Verify each owner's tax classification "
        "via Form W-9 before filing. Consult your CPA — this worksheet is a prep "
        "tool, not legal advice."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(39, 42):
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A8"

    ws.print_area = "A1:G42"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:7"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 9 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 30), ("C", 36),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="1099 Worksheet", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Your company info appears on every Owner Statement. Starting balances persist month-to-month."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # YOUR COMPANY (rows 6-12)
    _section_band(ws, 6, "YOUR COMPANY (appears on Owner Statement)")

    company_fields = [
        # (row, label, demo_value, blank_default)
        (7,  "Company name:",     "Smoky Mountain Hosting Co.",     None),
        (8,  "Tagline / mission:", "Hospitality. Discipline. Returns.", None),
        (9,  "Address line 1:",   "412 Mountainview Pkwy",          None),
        (10, "City / State / Zip:", "Sevierville, TN 37862",        None),
        (11, "Email (for owner contact):", "hello@smokyhosting.example", None),
        (12, "Phone:",            "(865) 555-0144",                 None),
    ]
    # Map company info to cells the Owner Statement references:
    # - B14 = Company name
    # - B15 = Tagline
    # - B16 = Address line
    # - B17 = City/State/Zip
    # - B19 = Email (referenced in statement footer)
    # Using rows 7-12 above maps to display labels but the references are 14-19.
    # Realign — put the actual company values at rows 14-19.

    for i, (row, label, demo_val, _) in enumerate(company_fields):
        # Display label
        actual_row = 14 + i
        ws.cell(row=actual_row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=actual_row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=actual_row, column=2, value=demo_val if variant == "demo" else None)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[actual_row].height = 20

    # STARTING BALANCES section (rows 21+)
    _section_band(ws, 21, "STARTING BALANCES — owner × balance carried into next month")

    # Header row 21 already used by section band. Columns header row 22? Actually
    # _section_band uses row 21. Skip 22 for the column header row.
    headers_sb = [(1, "Owner"), (2, "Starting balance")]
    for col, h in enumerate(headers_sb, start=1):
        # Note: col index is enumerate +1 but we want col 1 and 2; iteration above creates (1,"Owner"),(2,"Starting...")
        # Actually I want to fix this — use the col directly
        pass
    # Manual header row 22
    ws.cell(row=22, column=1, value="Owner").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=22, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=22, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.cell(row=22, column=2, value="Starting balance ($)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=22, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=22, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[22].height = 20

    # 30 rows owner × starting balance (rows 23-52)
    samples = OWNERS_DEMO if variant == "demo" else []
    for i in range(30):
        row = 23 + i
        if i < len(samples):
            owner_name = samples[i][0]
            balance = STARTING_BALANCES.get(owner_name, 0)
            ws.cell(row=row, column=1, value=owner_name).font = Font(
                name=FONT_BODY, size=10, color=COLOR_TEXT
            )
            cell = ws.cell(row=row, column=2, value=balance)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0.00'
        else:
            cell = ws.cell(row=row, column=1)
            apply_style(cell, input_cell_style())
            cell = ws.cell(row=row, column=2)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 18

    add_dropdown(ws, "A23:A52", "='Owners Register'!$A$6:$A$35")

    # Note row 54
    ws.merge_cells("A54:C56")
    c = ws["A54"]
    c.value = (
        "Workflow: after each month's distribution closes, copy the Owner "
        "Statement closing balance back here as next month's starting balance. "
        "Balance forward formula on the Owner Statement uses VLOOKUP into A23:B52."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(54, 57):
        ws.row_dimensions[r].height = 18


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_owners_register_tab(wb, variant)
    build_property_owner_map_tab(wb, variant)
    build_booking_log_tab(wb, variant)
    build_expense_log_tab(wb, variant)
    build_mgmt_fee_calculator_tab(wb, variant)
    build_owner_statement_tab(wb, variant)
    build_1099_worksheet_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Owner Reporting Dashboard — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "White-label owner statements + 1099 prep for property managers."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
