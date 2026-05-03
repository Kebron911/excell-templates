"""Build REV-005 Holiday + Event Pricing Calendar (v2.2 standard).

Operational-mode tool. Pre-built calendar of high-rate dates: federal
holidays, regional school breaks, local Smokies-region events. Per-event
demand-lift % drives a suggested nightly rate from the host's base rate.

Generates two files:
  templates/_masters/REV-005-holiday-event-pricing-calendar-DEMO.xlsx
  templates/_masters/REV-005-holiday-event-pricing-calendar-BLANK.xlsx
"""
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
    STATE_BAD_FILL,
)

SKU = "REV-005"
NAME = "holiday-event-pricing-calendar"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026
BASE_RATE_DEFAULT = 185.00
DEFAULT_REGION = "Southeast"

# --- Sample event data (DEMO) ---
# Tuple: (date, end_date, name, type, region, lift_pct, min_night_change, notes)

EVENTS_SAMPLES = [
    # 2026 Federal Holidays (12)
    (date(2026, 1, 1), date(2026, 1, 1), "New Year's Day",
     "Federal Holiday", "All", 25, "N",
     "Federal holiday — moderate demand, family travel."),
    (date(2026, 1, 19), date(2026, 1, 19), "Martin Luther King Jr. Day",
     "Federal Holiday", "All", 20, "Y",
     "3-day weekend; require 3-night minimum."),
    (date(2026, 2, 16), date(2026, 2, 16), "Presidents' Day",
     "Federal Holiday", "All", 25, "Y",
     "Long weekend, ski markets peak; 3-night minimum."),
    (date(2026, 5, 25), date(2026, 5, 25), "Memorial Day",
     "Federal Holiday", "All", 45, "Y",
     "Unofficial start of summer — strong demand, 3-night min."),
    (date(2026, 7, 4), date(2026, 7, 4), "Independence Day",
     "Federal Holiday", "All", 65, "Y",
     "Peak summer holiday; require 3-night minimum."),
    (date(2026, 9, 7), date(2026, 9, 7), "Labor Day",
     "Federal Holiday", "All", 80, "Y",
     "Last big summer weekend; max lift, 3-night minimum."),
    (date(2026, 10, 12), date(2026, 10, 12), "Columbus Day / Indigenous Peoples Day",
     "Federal Holiday", "All", 30, "N",
     "Fall foliage in mountain markets — moderate lift."),
    (date(2026, 11, 11), date(2026, 11, 11), "Veterans Day",
     "Federal Holiday", "All", 20, "N",
     "Mid-week observance; light lift."),
    (date(2026, 11, 26), date(2026, 11, 26), "Thanksgiving Day",
     "Federal Holiday", "All", 55, "Y",
     "Family travel peak; 3-night minimum recommended."),
    (date(2026, 12, 25), date(2026, 12, 25), "Christmas Day",
     "Federal Holiday", "All", 60, "Y",
     "Holiday week peak; consider 4-night minimum."),
    (date(2026, 12, 31), date(2026, 12, 31), "New Year's Eve",
     "Federal Holiday", "All", 80, "Y",
     "Top demand night of year; max lift + 3-night min."),
    (date(2026, 4, 5), date(2026, 4, 5), "Easter Sunday",
     "Federal Holiday", "All", 40, "Y",
     "Spring family travel; school break overlap."),

    # School breaks (4) — Southeast region defaults
    (date(2026, 3, 16), date(2026, 3, 22), "Spring Break — Southeast Schools",
     "School Break", "Southeast", 50, "Y",
     "Mid-March spring break wave; 3-night minimum."),
    (date(2026, 6, 8), date(2026, 8, 14), "Summer Break — Southeast Schools",
     "School Break", "Southeast", 35, "N",
     "Sustained baseline lift through summer."),
    (date(2026, 10, 5), date(2026, 10, 9), "Fall Break — Southeast Schools",
     "School Break", "Southeast", 40, "Y",
     "Fall break + foliage overlap; strong demand."),
    (date(2026, 12, 21), date(2026, 12, 31), "Winter Break — Southeast Schools",
     "School Break", "Southeast", 55, "Y",
     "Two-week holiday break; sustained peak."),

    # Smokies-region events (6)
    (date(2026, 4, 16), date(2026, 4, 19), "Spring Rod Run — Pigeon Forge",
     "Festival", "Southeast", 70, "Y",
     "Major car-show event; rooms book out 60+ days ahead."),
    (date(2026, 9, 17), date(2026, 9, 20), "Fall Rod Run — Pigeon Forge",
     "Festival", "Southeast", 70, "Y",
     "Larger of the two Rod Runs; peak demand."),
    (date(2026, 3, 14), date(2026, 3, 14), "Dollywood Season Opening",
     "Festival", "Southeast", 35, "N",
     "Theme park opening weekend; family demand."),
    (date(2026, 11, 7), date(2026, 11, 7), "Smoky Mountain Marathon",
     "Sporting", "Southeast", 45, "Y",
     "Runners + spectators fill Sevierville/Pigeon Forge."),
    (date(2026, 11, 13), date(2026, 12, 31), "Smoky Mountain Christmas — Dollywood",
     "Festival", "Southeast", 50, "Y",
     "Multi-week Christmas event drives sustained lift."),
    (date(2026, 8, 6), date(2026, 8, 9), "Gatlinburg Craftsmen's Fair",
     "Festival", "Southeast", 40, "Y",
     "Long-running summer arts festival; strong weekend lift."),
]

EVENT_TYPES = [
    "Federal Holiday", "School Break", "Sporting",
    "Concert", "Festival", "Convention", "Other",
]

REGIONS = [
    "All", "Northeast", "Southeast", "Midwest",
    "Southwest", "West Coast", "Mountain",
]

YN_LIST = ["Y", "N"]
STATUS_LIST = ["N", "Y"]  # rate-updated dropdown

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list, blank_length=None):
    if variant == "demo":
        return demo_list
    n = blank_length if blank_length is not None else len(demo_list)
    return [None] * n


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Holiday + Event Pricing Calendar"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Bump rates before the calendar bumps you."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: KPI cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Events next 30 days
    ws.merge_cells("A10:D10")
    c = ws["A10"]
    c.value = "EVENTS NEXT 30 DAYS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A11:D13")
    c = ws["A11"]
    c.value = (
        '=COUNTIFS(Events!A7:A206,">="&TODAY(),'
        'Events!A7:A206,"<="&TODAY()+30)'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A14:D15")
    c = ws["A14"]
    c.value = "high-rate dates inside 30 days — set rates now"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 2 (E-H): Events next 60 days
    ws.merge_cells("E10:H10")
    c = ws["E10"]
    c.value = "EVENTS NEXT 60 DAYS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E11:H13")
    c = ws["E11"]
    c.value = (
        '=COUNTIFS(Events!A7:A206,">="&TODAY(),'
        'Events!A7:A206,"<="&TODAY()+60)'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E14:H15")
    c = ws["E14"]
    c.value = "events on the 60-day horizon"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 3 (I-L): Avg lift % across upcoming 90 days
    ws.merge_cells("I10:L10")
    c = ws["I10"]
    c.value = "AVG LIFT % NEXT 90 DAYS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I11:L13")
    c = ws["I11"]
    c.value = (
        '=IFERROR(AVERAGEIFS(Events!F7:F206,'
        'Events!A7:A206,">="&TODAY(),'
        'Events!A7:A206,"<="&TODAY()+90)/100,0)'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("I14:L15")
    c = ws["I14"]
    c.value = "average demand lift on the 90-day horizon"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card borders (gold)
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 16):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: What this does card rows 17-21
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    ws.merge_cells("A18:L21")
    c = ws["A18"]
    c.value = (
        "Tracks the 12 federal holidays, regional school breaks, and "
        "local high-demand events for your market. Each event has a "
        "demand-lift % that drives a suggested nightly rate from your "
        "base. Open the Upcoming List before peaks; verify rates are "
        "bumped on Airbnb/VRBO before guests book at flat pricing."
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 4: How to use card rows 23-27
    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[23].height = 22

    ws.merge_cells("A24:L27")
    c = ws["A24"]
    c.value = (
        "  1. Set base rate + active year on Settings.\n"
        "  2. Customize Events: keep federal holidays, swap school breaks for "
        "your region, add your local festivals/concerts/games.\n"
        "  3. Each Friday, open Upcoming List — red rows are < 14 days out "
        "with rates not yet updated. Bump them, mark Status = Y."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 5: Pseudo-button nav rows 29-31
    pseudo_button(ws, "A29", "C31", "Events",
                  "'Events'!A1", variant="primary")
    pseudo_button(ws, "D29", "F31", "Year-View Calendar",
                  "'Year-View Calendar'!A1", variant="primary")
    pseudo_button(ws, "G29", "I31", "Upcoming List",
                  "'Upcoming List'!A1", variant="accent")
    pseudo_button(ws, "J29", "L31", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # ZONE 6: Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "Need full revenue coverage? Add the Revenue Bundle at "
        f"{BRAND_DOMAIN}/revenue — RevPAR, cleaning fee optimizer, "
        "break-even occupancy, holiday pricing."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(
        ws, 35,
        version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever",
    )

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Events (master register)
# ---------------------------------------------------------------------------

def build_events_tab(wb, variant):
    ws = wb.create_sheet("Events")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(
        ws, "Events",
        prev_tab="Start", next_tab="Year-View Calendar",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:I4")
    c = ws["A4"]
    c.value = (
        "Master register of high-demand dates. Suggested rate auto-calcs "
        "from base rate (Settings B7) × (1 + Lift%/100). Capacity 200 rows."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 13), ("B", 13), ("C", 36), ("D", 18),
        ("E", 14), ("F", 12), ("G", 16), ("H", 12),
        ("I", 36),
    ])

    headers = [
        "Date", "End date", "Event name", "Type",
        "Region", "Lift %", "Suggested rate", "Min-night Y/N",
        "Notes / source URL",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    samples = _val_list(variant, EVENTS_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 7 + i
        (d_start, d_end, name, etype, region,
         lift, min_night, notes) = row_data

        a = ws.cell(row=row, column=1, value=d_start)
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=row, column=2, value=d_end)
        apply_style(b, input_cell_style())
        b.number_format = "yyyy-mm-dd"

        c_cell = ws.cell(row=row, column=3, value=name)
        apply_style(c_cell, input_cell_style())

        d_cell = ws.cell(row=row, column=4, value=etype)
        apply_style(d_cell, input_cell_style())

        e_cell = ws.cell(row=row, column=5, value=region)
        apply_style(e_cell, input_cell_style())

        f_cell = ws.cell(row=row, column=6, value=lift)
        apply_style(f_cell, input_cell_style())
        f_cell.number_format = "0"
        f_cell.alignment = Alignment(horizontal="center", vertical="center")

        h_cell = ws.cell(row=row, column=8, value=min_night)
        apply_style(h_cell, input_cell_style())
        h_cell.alignment = Alignment(horizontal="center", vertical="center")

        i_cell = ws.cell(row=row, column=9, value=notes)
        apply_style(i_cell, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Apply formula + formatting to all rows 7-206
    for row in range(7, 207):
        # Suggested rate col G = base × (1 + lift/100)
        g_cell = ws.cell(
            row=row, column=7,
            value=(
                f'=IF(F{row}="","",'
                f'Settings!$B$7*(1+F{row}/100))'
            ),
        )
        apply_style(g_cell, formula_cell_style())
        g_cell.number_format = '"$"#,##0'
        g_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Apply input-cell formatting on empty rows past samples
        if row > 6 + len(samples):
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            for col in [3, 4, 5, 8, 9]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            f_cell = ws.cell(row=row, column=6)
            apply_style(f_cell, input_cell_style())
            f_cell.number_format = "0"
            f_cell.alignment = Alignment(horizontal="center", vertical="center")
            h_cell = ws.cell(row=row, column=8)
            h_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 20

    # Conditional formatting: high-lift events (>= 60%) gold-soft band
    ws.conditional_formatting.add(
        "F7:F206",
        CellIsRule(operator="greaterThanOrEqual", formula=["60"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
                   font=Font(name=FONT_BODY, size=11, bold=True,
                             color=COLOR_PRIMARY)),
    )

    # Dropdowns
    add_dropdown(ws, "D7:D206", "=Settings!$D$6:$D$15")
    add_dropdown(ws, "E7:E206", "=Settings!$E$6:$E$15")
    add_dropdown(ws, "H7:H206", '"Y,N"')

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Year-View Calendar
# ---------------------------------------------------------------------------

def build_year_view_tab(wb, variant):
    ws = wb.create_sheet("Year-View Calendar")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(
        ws, "Year-View Calendar",
        prev_tab="Events", next_tab="Upcoming List",
    )

    # Description row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = (
        "12-month grid for the active year (Settings B5). Days that fall "
        "on a registered event are gold-soft. Legend below lists each "
        "highlighted date with its event."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Calendar grid: 6 rows of months × 2 columns of months = 12 months
    # Each month occupies a 2-col span (label) × 8 rows block
    # Layout: 3 columns of months × 4 rows of months.
    # Each "month cell" = 7 columns wide (Sun-Sat) × 8 rows tall (label + days header + 6 weeks)
    # Total width = 21 cols. Stay within col A:U for printability.

    set_col_widths(ws, [(get_column_letter(c), 4) for c in range(1, 22)])

    # Build event-date set for highlight (active year)
    event_dates = {}  # date -> name
    if variant == "demo":
        for d_start, d_end, name, etype, region, lift, min_night, notes in EVENTS_SAMPLES:
            cur = d_start
            while cur <= d_end:
                if cur.year == ACTIVE_YEAR:
                    event_dates[cur] = name
                cur = date.fromordinal(cur.toordinal() + 1)

    # Render 12 months in 4 rows × 3 cols grid
    grid_top = 6
    month_width = 7   # cols per month
    month_gap_col = 0  # no gap (tight packing)
    month_height = 9  # rows per month (1 label + 1 day-header + 7 weeks max)
    month_gap_row = 1

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    gold_soft_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    parchment_alt_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    accent_thin = Side(style="thin", color=COLOR_ACCENT)

    import calendar as _cal

    for m in range(12):
        month_num = m + 1
        grid_row_pos = m // 3
        grid_col_pos = m % 3

        top_row = grid_top + grid_row_pos * (month_height + month_gap_row)
        left_col = 1 + grid_col_pos * (month_width + month_gap_col)

        # Month label
        ws.merge_cells(start_row=top_row, start_column=left_col,
                        end_row=top_row, end_column=left_col + month_width - 1)
        cell = ws.cell(row=top_row, column=left_col)
        cell.value = f"{MONTH_NAMES[m]} {ACTIVE_YEAR}"
        cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[top_row].height = 18

        # Day-of-week header
        day_header_row = top_row + 1
        for i, dn in enumerate(["S", "M", "T", "W", "T", "F", "S"]):
            cell = ws.cell(row=day_header_row, column=left_col + i, value=dn)
            cell.font = Font(name=FONT_MONO, size=8, bold=True, color=COLOR_PRIMARY)
            cell.fill = parchment_alt_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[day_header_row].height = 14

        # Days
        cal_obj = _cal.Calendar(firstweekday=6)  # Sunday-first
        weeks = cal_obj.monthdayscalendar(ACTIVE_YEAR, month_num)
        for w_idx, week in enumerate(weeks):
            row_n = day_header_row + 1 + w_idx
            ws.row_dimensions[row_n].height = 14
            for d_idx, day in enumerate(week):
                col_n = left_col + d_idx
                cell = ws.cell(row=row_n, column=col_n)
                if day == 0:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
                else:
                    cell.value = day
                    cell.font = Font(name=FONT_BODY, size=9, color=COLOR_TEXT)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    d_obj = date(ACTIVE_YEAR, month_num, day)
                    if d_obj in event_dates:
                        cell.fill = gold_soft_fill
                        cell.font = Font(name=FONT_BODY, size=9, bold=True,
                                         color=COLOR_PRIMARY)
                    else:
                        cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
                cell.border = Border(
                    left=accent_thin, right=accent_thin,
                    top=accent_thin, bottom=accent_thin,
                )

    # Legend section below the grid
    legend_top = grid_top + 4 * (month_height + month_gap_row) + 1
    ws.merge_cells(start_row=legend_top, start_column=1,
                    end_row=legend_top, end_column=21)
    c = ws.cell(row=legend_top, column=1)
    c.value = "EVENT LEGEND  —  highlighted dates above"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = gold_soft_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[legend_top].height = 22

    # List event-date entries
    if variant == "demo":
        sorted_events = sorted(event_dates.items(), key=lambda x: x[0])
        for i, (d_obj, name) in enumerate(sorted_events[:60]):
            r = legend_top + 1 + i
            cell = ws.cell(row=r, column=1, value=d_obj)
            cell.number_format = "yyyy-mm-dd"
            cell.font = Font(name=FONT_BODY, size=9, color=COLOR_TEXT)
            cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=21)
            cell2 = ws.cell(row=r, column=4, value=name)
            cell2.font = Font(name=FONT_BODY, size=9, color=COLOR_TEXT)
            cell2.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
            cell2.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[r].height = 16
    else:
        # Blank: brief instruction
        r = legend_top + 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=21)
        cell = ws.cell(row=r, column=1)
        cell.value = (
            "Add events on the Events tab — highlighted dates and a legend "
            "of date + event names will populate when you rebuild this view."
        )
        cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=2)
        ws.row_dimensions[r].height = 36

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Upcoming List
# ---------------------------------------------------------------------------

def build_upcoming_tab(wb, variant):
    ws = wb.create_sheet("Upcoming List")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(
        ws, "Upcoming List",
        prev_tab="Year-View Calendar", next_tab="Settings",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:G4")
    c = ws["A4"]
    c.value = (
        "Next 30 events sorted by date. Red rows = ≤ 14 days out AND "
        "Status = N — set the rate now before guests book at flat pricing."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 13), ("B", 12), ("C", 36), ("D", 10),
        ("E", 16), ("F", 12), ("G", 36),
    ])

    headers = [
        "Date", "Days away", "Event", "Lift %",
        "Suggested rate", "Status", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    # Build sorted upcoming list (only for demo; blank shows empty rows)
    upcoming = []
    if variant == "demo":
        today = date(2026, 5, 2)  # reference today for demo data
        for d_start, d_end, name, etype, region, lift, min_night, notes in EVENTS_SAMPLES:
            if d_start >= today:
                upcoming.append((d_start, name, lift, notes))
        upcoming.sort(key=lambda x: x[0])
        upcoming = upcoming[:30]

    for i in range(30):
        row = 7 + i
        if variant == "demo" and i < len(upcoming):
            d_obj, name, lift, notes = upcoming[i]
            a = ws.cell(row=row, column=1, value=d_obj)
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"

            # Days away formula = A - TODAY()
            b = ws.cell(row=row, column=2,
                        value=f'=IF(A{row}<>"",A{row}-TODAY(),"")')
            apply_style(b, formula_cell_style())
            b.alignment = Alignment(horizontal="center", vertical="center")
            b.number_format = "0"

            cc = ws.cell(row=row, column=3, value=name)
            apply_style(cc, input_cell_style())

            d_cell = ws.cell(row=row, column=4, value=lift)
            apply_style(d_cell, input_cell_style())
            d_cell.number_format = "0"
            d_cell.alignment = Alignment(horizontal="center", vertical="center")

            e_cell = ws.cell(row=row, column=5,
                              value=f'=IF(D{row}="","",Settings!$B$7*(1+D{row}/100))')
            apply_style(e_cell, formula_cell_style())
            e_cell.number_format = '"$"#,##0'

            f_cell = ws.cell(row=row, column=6, value="N")
            apply_style(f_cell, input_cell_style())
            f_cell.alignment = Alignment(horizontal="center", vertical="center")

            g_cell = ws.cell(row=row, column=7, value=notes)
            apply_style(g_cell, input_cell_style())
        else:
            for col in [1]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            b = ws.cell(row=row, column=2,
                        value=f'=IF(A{row}<>"",A{row}-TODAY(),"")')
            apply_style(b, formula_cell_style())
            b.alignment = Alignment(horizontal="center", vertical="center")
            b.number_format = "0"
            for col in [3, 4, 6, 7]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            d_cell = ws.cell(row=row, column=4)
            d_cell.number_format = "0"
            d_cell.alignment = Alignment(horizontal="center", vertical="center")
            f_cell = ws.cell(row=row, column=6)
            f_cell.alignment = Alignment(horizontal="center", vertical="center")
            e_cell = ws.cell(row=row, column=5,
                              value=f'=IF(D{row}="","",Settings!$B$7*(1+D{row}/100))')
            apply_style(e_cell, formula_cell_style())
            e_cell.number_format = '"$"#,##0'
        ws.row_dimensions[row].height = 20

    # Conditional formatting: red row when days <= 14 AND status = N
    red_fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    red_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    for row in range(7, 37):
        ws.conditional_formatting.add(
            f"A{row}:G{row}",
            FormulaRule(
                formula=[f'AND(ISNUMBER($B{row}),$B{row}<=14,$B{row}>=0,$F{row}="N")'],
                fill=red_fill,
                font=red_font,
            ),
        )

    # Dropdown for status col F
    add_dropdown(ws, "F7:F36", '"Y,N"')

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28),
        ("C", 4), ("D", 22), ("E", 18),
        ("F", 4),
    ])

    compact_header_band(
        ws, "Settings",
        prev_tab="Upcoming List", next_tab=None,
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Active year, base rate, and dropdown lists for Events. "
        "Edit to match your portfolio."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Active-year + base-rate + region inputs
    label_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Active year (B5)
    ws.cell(row=5, column=1, value="").fill = parchment_fill
    cell = ws.cell(row=5, column=2, value="Active year")
    cell.font = label_font
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.fill = parchment_fill
    val_cell = ws.cell(row=5, column=2)  # label is in B5; year value cell per brief is B5 — but we already used B5 for label.
    # Brief says: B5 = active year, B7 = base rate, B9 = region. Move label to A5/A7/A9.
    # Reset.
    ws.cell(row=5, column=2, value=None)

    # Per brief: B5 active year, B7 base rate, B9 region.
    # Use col D for labels to keep B exclusively for values.
    label_col_letter = "D"

    # B5 active year
    cell = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 30

    cell = ws.cell(row=5, column=4, value="Active year — drives Year-View Calendar and Upcoming List")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # B7 base rate
    cell = ws.cell(row=7, column=2, value=BASE_RATE_DEFAULT)
    apply_style(cell, input_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 30

    cell = ws.cell(row=7, column=4, value="Base nightly rate — Suggested rate = Base × (1 + Lift%)")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # B9 region
    cell = ws.cell(row=9, column=2, value=DEFAULT_REGION)
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 26
    add_dropdown(ws, "B9", '"Northeast,Southeast,Midwest,Southwest,West Coast,Mountain"')

    cell = ws.cell(row=9, column=4, value="Region — drives default school break dates")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Header strip for dropdown lists row 12
    ws.merge_cells("A12:F12")
    c = ws["A12"]
    c.value = "DROPDOWN LISTS  —  edit to match your portfolio"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[12].height = 22

    # List headers row 13
    list_headers = [
        (4, "Event types"),
        (5, "Regions"),
    ]
    # Use cols D/E for dropdown source lists (B is the value cells already)
    # Override: per build_license pattern we used cols C/D/E/F. Here let's
    # use D for types, E for regions to leave B clean.
    cell = ws.cell(row=13, column=4, value="Event types")
    cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell = ws.cell(row=13, column=5, value="Regions")
    cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[13].height = 22

    # NOTE: Events tab dropdowns reference Settings!$D$6:$D$15 + $E$6:$E$15.
    # We need source cells in those positions. Currently B5/B7/B9 are values
    # but col D row 5/7/9 hold descriptive captions. Their content does not
    # interfere with Events dropdown source range D6:D15 (rows 6-15 only).
    # Move list to col D rows 6-15? Conflict: D6, D8 are empty but D5, D7, D9
    # have descriptions. Use a different list location.
    # Switch dropdown sources to safe ranges.
    # We'll write Event types in D14:D24 and Regions in E14:E24 to avoid
    # the description rows. Then update Events tab dropdowns to point there.
    # However, Events tab build is already complete pointing to D6:D15.
    # Solution: put descriptions in a single col F instead, free up D and E.
    # Re-do labels: clear D5/D7/D9 if previously written, put them in F.
    ws.cell(row=5, column=4, value=None)
    ws.cell(row=7, column=4, value=None)
    ws.cell(row=9, column=4, value=None)
    # Re-write descriptions in col F
    cell = ws.cell(row=5, column=6, value="Active year — drives Year-View Calendar and Upcoming List")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell = ws.cell(row=7, column=6, value="Base nightly rate — Suggested rate = Base × (1 + Lift%)")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell = ws.cell(row=9, column=6, value="Region — drives default school break dates")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    set_col_widths(ws, [("F", 56)])

    # Now safely write Event types in D6:D15 and Regions in E6:E15
    # First unmerge A12:F12 so D12 is writable for the EVENT_TYPES list
    try:
        ws.unmerge_cells("A12:F12")
    except Exception:
        pass
    # Re-merge a smaller header strip A12:C12 to keep the visual treatment without blocking col D
    ws.merge_cells("A12:C12")
    for i in range(10):
        r = 6 + i
        cell = ws.cell(row=r, column=4)
        if i < len(EVENT_TYPES):
            cell.value = EVENT_TYPES[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18
        cell = ws.cell(row=r, column=5)
        if i < len(REGIONS):
            cell.value = REGIONS[i]
        apply_style(cell, input_cell_style())

    # Move list-header strip to row 12 just labels above lists
    # (already written) — but our lists span 6-15, header should be at 5.
    # The active-year/base/region values occupy B5/B7/B9 — col D5 is now a list item.
    # To keep them visually separated, add a small note at row 16.
    ws.cell(row=12, column=1, value=None)  # clear earlier merge text loc
    # Re-merge for clarity
    try:
        ws.unmerge_cells("A12:F12")
    except Exception:
        pass

    # Year-end Archive table rows 18-25 (operational pattern)
    archive_top = 18
    ws.merge_cells(start_row=archive_top, start_column=1,
                    end_row=archive_top, end_column=6)
    c = ws.cell(row=archive_top, column=1)
    c.value = "YEAR-END ARCHIVE  —  fill in Jan, then clear Events for the new year"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[archive_top].height = 22

    archive_headers = [
        (1, "Year"), (2, "Events tracked"),
        (3, "Avg lift %"), (4, "Total $ lift captured"),
    ]
    for col, label in archive_headers:
        cell = ws.cell(row=archive_top + 1, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[archive_top + 1].height = 22

    archive_years = [2026, 2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(archive_years):
        r = archive_top + 2 + i
        cell = ws.cell(row=r, column=1, value=yr)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
        for col in [2, 3, 4]:
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=r, column=3).number_format = "0%"
        ws.cell(row=r, column=4).number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    # Footer note
    note_row = archive_top + 2 + len(archive_years) + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                    end_row=note_row, end_column=6)
    c = ws.cell(row=note_row, column=1)
    c.value = (
        "Tip: each January, copy your YTD totals from Events into the row "
        "above for last year, then clear Events to repopulate with the new "
        "year's holidays + festivals."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=1)
    ws.row_dimensions[note_row].height = 36


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_events_tab(wb, variant)
    build_year_view_tab(wb, variant)
    build_upcoming_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Holiday + Event Pricing Calendar — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Pre-built calendar of high-rate dates (federal holidays, school "
        "breaks, regional events) with per-event demand lift % and "
        "auto-calculated suggested nightly rate."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
