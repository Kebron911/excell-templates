"""Build TAX-004 Schedule E Tax-Prep Workbook — wizard-style tool.

Implements templates/_briefs/TAX-004-schedule-e-tax-prep.md.

Generates BOTH:
  templates/_masters/TAX-004-schedule-e-tax-prep-DEMO.xlsx
  templates/_masters/TAX-004-schedule-e-tax-prep-BLANK.xlsx

Per-property capacity: 3 (Schedule E Part I supports 3 per page).

Tabs:
  0  Start             — wizard hero + 3-card + Quick Start + Get Started + progress
  1  Property Info     — 8 fields × 3 properties (24 inputs)
  2  Income            — 3 fields × 3 properties (9 inputs)
  3  Expenses          — 14 expense lines × 3 properties (42 inputs)
  4  Depreciation      — 5 fields × 3 properties (15 inputs)
  5  Loss Limitations  — 4 fields total + per-prop participation toggles
  6  Schedule E Map    — auto-built form mirror (no inputs)
  7  Launch            — readiness dashboard + print packet button
  8  Settings          — active tax year, classification, dropdowns

Usage:
    python build_schedule_e_tax_prep.py
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    COLOR_WHITE,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-004"
NAME = "schedule-e-tax-prep"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

VERSION_LINE = f"{SKU} · v2.2 · Free updates forever"

TAB_NAMES = [
    "Start", "Property Info", "Income", "Expenses",
    "Depreciation", "Loss Limitations", "Schedule E Map",
    "Launch", "Settings",
]
TOTAL_SECTIONS = 5  # input sections (1-5)

# Inputs per section (used for progress %)
SECTION_INPUT_COUNTS = {
    "Property Info":     24,  # 8 × 3
    "Income":             9,  # 3 × 3
    "Expenses":          42,  # 14 × 3
    "Depreciation":      15,  # 5 × 3
    "Loss Limitations":   7,  # 4 portfolio + 3 per-prop participation
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 97

# --- Sample data (DEMO) ---

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]

PROPERTY_INFO_FIELDS = [
    ("Property address",          ["123 Mountain Lane, Gatlinburg, TN 37738",
                                    "47 Creek Side Dr, Pigeon Forge, TN 37863",
                                    "888 Lakeshore Way, Townsend, TN 37882"]),
    ("Type (Schedule E code 1-8)", [3, 3, 3]),
    ("Fair-rental days",          [228, 196, 244]),
    ("Personal-use days",         [12, 8, 0]),
    ("% ownership",               [1.00, 1.00, 0.50]),
    ("Date placed in service",    [date(2022, 6, 1), date(2023, 3, 15), date(2024, 1, 12)]),
    ("EIN (if any)",              ["", "", ""]),
    ("QJV election (Y/N)",        ["N", "N", "N"]),
]

INCOME_FIELDS = [
    ("Rents received ($)",                [48200, 36100, 52400]),
    ("Royalties ($)",                     [0, 0, 0]),
    ("Refundable deposits held (memo)",   [500, 500, 750]),
]

# 14 IRS Schedule E Part I expense lines (in form order)
EXPENSE_FIELDS = [
    ("5  Advertising",              [820, 410, 1240]),
    ("6  Auto and travel",          [1860, 940, 2230]),
    ("7  Cleaning and maintenance", [6800, 5200, 7400]),
    ("8  Commissions",              [4820, 3610, 5240]),  # Airbnb host service fee 3%
    ("9  Insurance",                [1840, 1620, 2280]),
    ("10 Legal and professional",   [620, 320, 480]),
    ("11 Management fees",          [0, 0, 0]),
    ("12 Mortgage interest (banks)", [9800, 7200, 14600]),
    ("13 Other interest",           [0, 0, 0]),
    ("14 Repairs",                  [1420, 980, 2870]),
    ("15 Supplies",                 [1640, 1280, 1920]),
    ("16 Taxes",                    [2880, 2110, 3940]),
    ("17 Utilities",                [3260, 2810, 4080]),
    ("19 Other (list)",             [840, 460, 1180]),  # platform tech, software
]

DEPRECIATION_FIELDS = [
    ("Depreciable basis ($, building only)", [340000, 268000, 412000]),
    ("Date placed in service",               [date(2022, 6, 1), date(2023, 3, 15), date(2024, 1, 12)]),
    ("Recovery period (27.5 / 39 yr)",       [27.5, 27.5, 27.5]),
    ("Convention (mid-month)",               ["MM", "MM", "MM"]),
    ("Accumulated prior depreciation ($)",   [42500, 22850, 12300]),
]

LOSS_LIMITATION_PORTFOLIO = [
    ("Filing AGI ($, for $25K phase-out)",     185000),
    ("Active Participation? (Y/N)",            "Y"),
    ("Material Participation? (Y/N STR loophole)", "N"),
    ("Prior-year suspended losses ($)",        0),
]

# Per-property participation hours (informational, drives material-participation flag)
LOSS_LIMITATION_PER_PROP = [
    ("Hours of personal services this year (per prop)",
     [180, 110, 60]),
]


def add_dropdown(ws, cell_range, options):
    """Inline list dropdown."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (blank)."""
    return demo_value if variant == "demo" else None


# --- Tab builders ---

def build_start_tab(wb, variant):
    """Tab 0 — Start. Wizard hero + 3-card + Quick Start + Get Started + progress."""
    ws = wb.active
    ws.title = TAB_NAMES[0]
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — navy hero (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Schedule E Tax-Prep Workbook"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Hand your CPA a finished Schedule E."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — "What you'll build" 3-card (rows 10-19)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 20):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A11:L11")
    c = ws["A11"]; c.value = "What you'll hand to your CPA"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("📄 SCHEDULE E", "Print-ready Lines 1-26, three properties per page."),
        ("📊 PER-PROPERTY P&L", "Rents − expenses − depreciation per property."),
        ("⚖ LOSS LIMITS", "Active + material participation flagged, $25K rule applied."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (ttl, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]; c.value = ttl
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]; c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # ZONE 3 — Quick Start (rows 21-28)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(21, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A22:L22")
    c = ws["A22"]; c.value = "Quick Start — be done in 30 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 24
    quickstart_items = [
        "① Property identity (address, type, days)",
        "② Rents received per property",
        "③ All 14 Schedule E expense lines",
        "④ Depreciation basis + prior accumulated",
        "⑤ Material-participation answer + AGI",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + (i if i < 3 else i - 3)
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ZONE 4 — Get Started button (rows 30-33)
    pseudo_button(ws, "A30", "L33",
                  "GET STARTED — FILL PROPERTY INFO  →",
                  "'Property Info'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # ZONE 5 — Progress dashboard (rows 35-44)
    ws.merge_cells("A36:F36")
    c = ws["A36"]; c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Overall completion %
    ranges = [
        "'Property Info'!B8:D15",        # 8 fields × 3 props = 24
        "'Income'!B8:D10",               # 3 × 3 = 9
        "'Expenses'!B8:D21",             # 14 × 3 = 42
        "'Depreciation'!B8:D12",         # 5 × 3 = 15
        "'Loss Limitations'!B8:B11,'Loss Limitations'!B14:D14",  # 4 + 3 = 7
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    ws.merge_cells("G36:L36")
    c = ws["G36"]
    c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS},"0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Per-section status rows
    section_rows = [
        ("① Property Info",     "Property Info",     "B8:D15",                         24),
        ("② Income",            "Income",            "B8:D10",                          9),
        ("③ Expenses",          "Expenses",          "B8:D21",                        42),
        ("④ Depreciation",      "Depreciation",      "B8:D12",                        15),
        ("⑤ Loss Limitations",  "Loss Limitations",  "B8:B11",                          7),  # see custom
    ]
    LOSS_COUNTA = (
        "COUNTA('Loss Limitations'!B8:B11,'Loss Limitations'!B14:D14)"
    )
    section_counta_overrides = {"Loss Limitations": LOSS_COUNTA}
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]; c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = section_counta_overrides.get(tab, f"COUNTA('{tab}'!{range_})")
        c = ws[f"G{r}"]
        c.value = (
            f'=IF({ca}={total},"✅ Done",'
            f'IF({ca}=0,"⏳ Empty","⏳ "&{ca}&" of {total}"))'
        )
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Footer (rows 47-49)
    brand_footer(ws, 47, version_line=VERSION_LINE)

    # Print setup
    ws.print_area = "A1:L49"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def _build_three_property_tab(wb, tab_name, section_num, title, subtitle,
                               fields, variant, prev_tab, next_tab,
                               number_format=None,
                               extra_callout=None):
    """Generic builder for tabs with shape: label | P1 | P2 | P3.

    fields: list of (label, [v1, v2, v3]) tuples.
    number_format: applied to the input cells (e.g., '"$"#,##0').
    extra_callout: optional string drawn in a parchment band after the table.
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    section_header_band(ws, section_num, TOTAL_SECTIONS, title, subtitle,
                         prev_tab, next_tab)

    # Override col widths AFTER section_header_band (which uses 12 evenly).
    # We want: A wide labels, B-D = three property columns.
    set_col_widths(ws, [
        ("A", 38), ("B", 18), ("C", 18), ("D", 18),
        ("E", 6), ("F", 6), ("G", 6), ("H", 6),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction strip
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "Fill the highlighted fields for each of your properties (3 max)."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: column headers (P1/P2/P3 names)
    ws.cell(row=7, column=1, value="").fill = parchment_fill
    for col_idx, prop_name in enumerate(PROPERTIES, start=2):
        c = ws.cell(row=7, column=col_idx, value=prop_name)
        apply_style(c, header_row_style())
    ws.row_dimensions[7].height = 20

    # Row 8+: data rows
    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    for i, (label, values) in enumerate(fields):
        r = 8 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        for col_idx, v in enumerate(values, start=2):
            cell_val = _val(variant, v)
            c = ws.cell(row=r, column=col_idx, value=cell_val)
            apply_style(c, input_cell_style())
            if isinstance(v, date):
                c.number_format = "yyyy-mm-dd"
            elif number_format:
                c.number_format = number_format
        ws.row_dimensions[r].height = 18

    last_input_row = 7 + len(fields)

    # Optional callout strip
    if extra_callout:
        callout_row = last_input_row + 2
        ws.merge_cells(f"A{callout_row}:L{callout_row}")
        c = ws[f"A{callout_row}"]; c.value = extra_callout
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[callout_row].height = 32
        last_input_row = callout_row

    # Per-property total row (for tabs where it makes sense; caller can add)
    return ws, last_input_row


def build_property_info_tab(wb, variant):
    ws, _ = _build_three_property_tab(
        wb, "Property Info", 1,
        "Property Info",
        "Identity, days rented, ownership — Schedule E Lines 1-2.",
        PROPERTY_INFO_FIELDS, variant,
        prev_tab="", next_tab="Income",
        extra_callout=(
            "⚠ Personal-use days > 14 OR > 10% of rental days triggers "
            "vacation-home rules (IRC §280A) — limits deductions to rental income. "
            "Flag for your CPA if exceeded."
        ),
    )
    # Type field row 9 (was 7); QJV row 15 (was 13)
    add_dropdown(ws, "B9:D9", [1, 2, 3, 4, 5, 7, 8])
    add_dropdown(ws, "B15:D15", ["Y", "N"])  # QJV
    # Number formats on numeric rows: rows 10-12 (days, days, % ownership)
    for col in range(2, 5):
        ws.cell(row=10, column=col).number_format = "0"
        ws.cell(row=11, column=col).number_format = "0"
        ws.cell(row=12, column=col).number_format = "0.00%"


def build_income_tab(wb, variant):
    ws, last_row = _build_three_property_tab(
        wb, "Income", 2,
        "Income",
        "Rents and royalties received — Schedule E Lines 3-4.",
        INCOME_FIELDS, variant,
        prev_tab="Property Info", next_tab="Expenses",
        number_format='"$"#,##0',
        extra_callout=(
            "⚠ Refundable security deposits are NOT income (IRC §61) — "
            "list them as memo only. Forfeited deposits become income in "
            "the year of forfeiture; track separately on the Expenses tab."
        ),
    )

    # Subtotal row: Total Income (Lines 3 + 4) — rows 8-9 are rents/royalties; subtotal at row 11
    sub_row = 11
    a = ws.cell(row=sub_row, column=1,
                 value="TOTAL INCOME (Lines 3 + 4)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for col in range(2, 5):
        c = ws.cell(row=sub_row, column=col,
                     value=f"=SUM({get_column_letter(col)}8:{get_column_letter(col)}9)")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[sub_row].height = 22


def build_expenses_tab(wb, variant):
    ws, _ = _build_three_property_tab(
        wb, "Expenses", 3,
        "Expenses",
        "All 14 Schedule E Part I expense lines (5-19).",
        EXPENSE_FIELDS, variant,
        prev_tab="Income", next_tab="Depreciation",
        number_format='"$"#,##0',
    )
    # Subtotal — 14 expense lines fill rows 8-21; subtotal lands at row 23
    sub_row = 23
    a = ws.cell(row=sub_row, column=1,
                 value="TOTAL EXPENSES (Line 20)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for col in range(2, 5):
        col_letter = get_column_letter(col)
        c = ws.cell(row=sub_row, column=col,
                     value=f"=SUM({col_letter}8:{col_letter}21)")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[sub_row].height = 22


def build_depreciation_tab(wb, variant):
    ws, last_row = _build_three_property_tab(
        wb, "Depreciation", 4,
        "Depreciation",
        "Per-property basis + recovery — Schedule E Line 18.",
        DEPRECIATION_FIELDS, variant,
        prev_tab="Expenses", next_tab="Loss Limitations",
        extra_callout=(
            "ⓘ Default for residential STR is 27.5-year MACRS straight-line "
            "(non-residential = 39 yr). Building only — land is NOT depreciable. "
            "First-year and disposal-year prorate via mid-month convention. "
            "Customers electing accelerated depreciation (Section 179, bonus) "
            "should use the Section 179 Planner (TAX-009) and feed the result "
            "into Line 18 below."
        ),
    )

    # Apply formats per-row — basis row 8, date row 9, recovery row 10, accum row 12
    for col in range(2, 5):
        ws.cell(row=8, column=col).number_format = '"$"#,##0'
        ws.cell(row=9, column=col).number_format = "yyyy-mm-dd"
        ws.cell(row=10, column=col).number_format = "0.0"
        ws.cell(row=12, column=col).number_format = '"$"#,##0'

    add_dropdown(ws, "B10:D10", [27.5, 39])
    add_dropdown(ws, "B11:D11", ["MM", "MQ", "HY"])

    # Computed annual depreciation lands at row 16 (extra_callout merges row 14;
    # row 15 = blank gap; subtotal at 16). Schedule E Map references row 16.
    sub_row = 16
    a = ws.cell(row=sub_row, column=1,
                 value="ANNUAL DEPRECIATION (Schedule E Line 18)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for col in range(2, 5):
        col_letter = get_column_letter(col)
        c = ws.cell(row=sub_row, column=col,
                     value=f'=IF(AND({col_letter}8<>"",{col_letter}10<>""),'
                            f'{col_letter}8/{col_letter}10,0)')
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[sub_row].height = 22


def build_loss_limitations_tab(wb, variant):
    """Tab 5 — Loss Limitations. 4 portfolio fields + per-prop participation."""
    ws = wb.create_sheet("Loss Limitations")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 5, TOTAL_SECTIONS,
                         "Loss Limitations",
                         "Active vs. material participation — IRC §469.",
                         "Depreciation", "Schedule E Map")

    # Override widths for label/value layout
    set_col_widths(ws, [
        ("A", 44), ("B", 18), ("C", 18), ("D", 18),
        ("E", 6), ("F", 6), ("G", 6), ("H", 6),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Row 6: instruction strip
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("Portfolio-level participation drives loss treatment under "
               "IRC §469. Hours must be contemporaneously documented.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: section banner
    card_header(ws, 7, ("A", "L"), "Loss Limitations")

    # Portfolio fields (rows 8-11)
    for i, (label, val) in enumerate(LOSS_LIMITATION_PORTFOLIO):
        r = 8 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=_val(variant, val))
        apply_style(c, input_cell_style())
        if isinstance(val, (int, float)) and val > 100:
            c.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    add_dropdown(ws, "B9", ["Y", "N"])   # Active participation
    add_dropdown(ws, "B10", ["Y", "N"])  # Material participation

    # Per-property participation header (row 13)
    a = ws.cell(row=13, column=1, value="Per-property participation:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col_idx, prop in enumerate(PROPERTIES, start=2):
        c = ws.cell(row=13, column=col_idx, value=prop)
        apply_style(c, header_row_style())
    ws.row_dimensions[13].height = 20

    # Hours row (14)
    label, vals = LOSS_LIMITATION_PER_PROP[0]
    a = ws.cell(row=14, column=1, value=label)
    a.font = bold
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col_idx, v in enumerate(vals, start=2):
        c = ws.cell(row=14, column=col_idx, value=_val(variant, v))
        apply_style(c, input_cell_style())
        c.number_format = "0"
    ws.row_dimensions[14].height = 18

    # Computed flags (rows 16-19) — reference: B8=AGI, B9=Active, B10=Material
    flag_rows = [
        (16, "$25K offset eligible? (Active + AGI ≤ $100K, phases out to $150K)",
         '=IF(AND(B9="Y",B8<=100000),"Yes — full $25K",'
         'IF(AND(B9="Y",B8<150000),"Partial — phase-out range",'
         '"No — AGI > $150K or not active"))'),
        (17, "STR loophole (material participation, non-passive)?",
         '=IF(B10="Y","Yes — losses offset W-2 income (IRC §469(c)(7))",'
         '"No — losses limited to passive income or carryforward")'),
        (18, "Total per-property hours:",
         "=SUM(B14:D14)"),
        (19, "500-hour test met (single-property)?",
         '=IF(MAX(B14:D14)>=500,"Yes (one property meets it)",'
         '"No — combined-activity election may apply (consult CPA)")'),
    ]
    for r, label, formula in flag_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells(f"B{r}:D{r}")
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    # Callout (row 21)
    ws.merge_cells("A21:L21")
    c = ws["A21"]
    c.value = (
        "⚖ Passive losses you can't use this year carry forward to next year "
        "(no expiration). When you sell the property, suspended losses release "
        "in full against the gain (IRC §469(g))."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[21].height = 36

    # Footer-style nav (rows 23-24)
    pseudo_button(ws, "A23", "F24", "← Back: Depreciation",
                   "'Depreciation'!A5", variant="secondary")
    pseudo_button(ws, "G23", "L24", "Next: Schedule E Map →",
                   "'Schedule E Map'!A5", variant="secondary")
    ws.row_dimensions[23].height = 22
    ws.row_dimensions[24].height = 22


def build_schedule_e_map_tab(wb, variant):
    """Tab 6 — Schedule E Map. Auto-built mirror of Schedule E Part I.

    Pulls data from Property Info / Income / Expenses / Depreciation tabs
    into the canonical IRS line layout. No inputs on this tab.
    """
    ws = wb.create_sheet("Schedule E Map")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    compact_header_band(ws, "Schedule E Map",
                         prev_tab="Loss Limitations", next_tab="Launch")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = (
        "Print this tab and hand it to your CPA. "
        "Lines 1-26 mirror IRS Schedule E (Form 1040), Part I."
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Override widths for form-mirror layout
    set_col_widths(ws, [
        ("A", 6), ("B", 38),
        ("C", 16), ("D", 16), ("E", 16),
        ("F", 6), ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    # Row 5: column headers (Line | Field | A | B | C — mirroring Schedule E)
    headers = ["Line", "Description", "A", "B", "C"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        apply_style(c, header_row_style())
    ws.row_dimensions[5].height = 20

    # Property names (row 6)
    ws.cell(row=6, column=1, value="").fill = parchment_fill
    label = ws.cell(row=6, column=2, value="Property name")
    label.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    label.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col, prop in enumerate(PROPERTIES, start=3):
        c = ws.cell(row=6, column=col, value=prop)
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[6].height = 18

    # Row 7: empty spacer
    ws.row_dimensions[7].height = 8

    # Schedule E Part I lines — (line_num, description, formula_template)
    # Each formula template uses {col} placeholder for property column
    # (B=A, C=B, D=C). property_idx 0=B (Schedule E A), 1=C, 2=D.
    # Source columns on input tabs: B=Property 1, C=Property 2, D=Property 3.
    # Line 1a-2: from Property Info. Lines 3-4: Income. Lines 5-19: Expenses.
    # Line 18 Depreciation: from Depreciation tab. Lines 20+ computed.
    # Source row offsets follow the input-tab data layout:
    #   Property Info: rows 8-15 (8 fields, starting row 8)
    #   Income:        rows 8-10 (rents, royalties, deposits-memo)
    #   Expenses:      rows 8-21 (14 fields)
    #   Depreciation:  row 14   (annual depreciation = computed sub-row)
    line_specs = [
        # (line, label, kind, source_tab_or_formula)
        ("1a", "Property address",                    "text",    "='Property Info'!{col}8"),
        ("1b", "Type",                                "text",    "='Property Info'!{col}9"),
        ("2",  "Fair rental days",                    "num",     "='Property Info'!{col}10"),
        ("2",  "Personal use days",                   "num",     "='Property Info'!{col}11"),
        ("3",  "Rents received",                      "money",   "='Income'!{col}8"),
        ("4",  "Royalties received",                  "money",   "='Income'!{col}9"),
        ("5",  "Advertising",                         "money",   "='Expenses'!{col}8"),
        ("6",  "Auto and travel",                     "money",   "='Expenses'!{col}9"),
        ("7",  "Cleaning and maintenance",            "money",   "='Expenses'!{col}10"),
        ("8",  "Commissions",                         "money",   "='Expenses'!{col}11"),
        ("9",  "Insurance",                           "money",   "='Expenses'!{col}12"),
        ("10", "Legal and professional",              "money",   "='Expenses'!{col}13"),
        ("11", "Management fees",                     "money",   "='Expenses'!{col}14"),
        ("12", "Mortgage interest paid (banks)",      "money",   "='Expenses'!{col}15"),
        ("13", "Other interest",                      "money",   "='Expenses'!{col}16"),
        ("14", "Repairs",                             "money",   "='Expenses'!{col}17"),
        ("15", "Supplies",                            "money",   "='Expenses'!{col}18"),
        ("16", "Taxes",                               "money",   "='Expenses'!{col}19"),
        ("17", "Utilities",                           "money",   "='Expenses'!{col}20"),
        ("18", "Depreciation expense",                "money",   "='Depreciation'!{col}16"),
        ("19", "Other (list)",                        "money",   "='Expenses'!{col}21"),
    ]

    for i, (line_num, label, kind, formula_template) in enumerate(line_specs):
        r = 8 + i
        # Line column
        c = ws.cell(row=r, column=1, value=line_num)
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Description
        c = ws.cell(row=r, column=2, value=label)
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Property A/B/C (cols C/D/E in this tab; sources B/C/D on input tabs)
        for prop_idx, output_col in enumerate([3, 4, 5]):
            source_col = chr(ord("B") + prop_idx)  # B, C, D
            f = formula_template.replace("{col}", source_col)
            c = ws.cell(row=r, column=output_col, value=f)
            apply_style(c, formula_cell_style())
            if kind == "money":
                c.number_format = '"$"#,##0'
            elif kind == "num":
                c.number_format = "0"
        ws.row_dimensions[r].height = 16

    # Compute first-row + last-row of expense lines (5-19) for SUM
    # Lines 5-19 are at rows 12-26. Line 20 = sum of Lines 5-19, row 27.
    # Line 21 = Line 3 - Line 20 (income - expenses), row 28.
    # Line 26 = Total income/loss (sum of A+B+C from line 21), row 29.

    # Row 27 — Line 20: Total expenses
    r = 27
    c = ws.cell(row=r, column=1, value="20")
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c = ws.cell(row=r, column=2, value="Total expenses (Lines 5-19)")
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for output_col in [3, 4, 5]:
        col_letter = get_column_letter(output_col)
        c = ws.cell(row=r, column=output_col,
                     value=f"=SUM({col_letter}12:{col_letter}26)")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22

    # Row 28 — Line 21: Income or loss (Line 3 + 4 - Line 20)
    r = 28
    c = ws.cell(row=r, column=1, value="21")
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c = ws.cell(row=r, column=2, value="Income or (loss) [Lines 3+4 − 20]")
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    # Lines 3+4 are at rows 12 and 13 (because line_specs first money rows
    # are #4: Rents=row 12, #5: Royalties=row 13). Line 20 is at row 27.
    # Source rows for Line 3 in our tab: row 12. Line 4: row 13.
    for output_col in [3, 4, 5]:
        col_letter = get_column_letter(output_col)
        c = ws.cell(row=r, column=output_col,
                     value=f"={col_letter}12+{col_letter}13-{col_letter}27")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0;[Red]("$"#,##0)'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22

    # Row 30 — Line 26: Total income/loss (sum of A+B+C from Line 21)
    r = 30
    ws.row_dimensions[29].height = 8
    c = ws.cell(row=r, column=1, value="26")
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c = ws.cell(row=r, column=2,
                 value="Total rental real estate income or (loss) — to Form 1040 Schedule 1 Line 5")
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"C{r}:D{r}")
    c = ws.cell(row=r, column=3, value="=SUM(C28:E28)")
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.number_format = '"$"#,##0;[Red]("$"#,##0)'
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    # E30 unused (merged into C:D)
    e_cell = ws.cell(row=r, column=5, value="")
    e_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.row_dimensions[r].height = 28

    # Schedule routing footnote (row 32)
    ws.row_dimensions[31].height = 8
    ws.merge_cells("A32:L32")
    c = ws["A32"]
    c.value = (
        '=IF(\'Loss Limitations\'!B10="Y",'
        '"→ Material participation: losses non-passive, offset W-2 income.",'
        '"→ Passive: losses limited to passive income; carry forward suspended losses (Form 8582).")'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[32].height = 28

    # Print setup
    ws.print_area = "A1:L33"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_launch_tab(wb, variant):
    """Tab 7 — Launch. Readiness dashboard + print packet button."""
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero (rows 1-6, navy)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK", "'Schedule E Map'!A1",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]; c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Your Schedule E is ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Print the Schedule E Map and send it to your CPA."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # Readiness dashboard (rows 8-12)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 14):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 (A-D): Completion %
    ranges = [
        "'Property Info'!B8:D15",
        "'Income'!B8:D10",
        "'Expenses'!B8:D21",
        "'Depreciation'!B8:D12",
        "'Loss Limitations'!B8:B11,'Loss Limitations'!B14:D14",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    ws.merge_cells("A9:D9")
    c = ws["A9"]; c.value = "COMPLETION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]; c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS},"0%")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A12:D12")
    c = ws["A12"]; c.value = f"of {TOTAL_INPUTS} fields filled"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Total Schedule E line 26
    ws.merge_cells("E9:H9")
    c = ws["E9"]; c.value = "SCHEDULE E LINE 26"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]; c.value = "='Schedule E Map'!C30"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0;[Red]("$"#,##0)'
    ws.merge_cells("E12:H12")
    c = ws["E12"]; c.value = "to Form 1040 Schedule 1 Line 5"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Status
    ws.merge_cells("I9:L9")
    c = ws["I9"]; c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    # Required: Property Info rows 8-9 (address, type), Income row 8 (rents),
    # Expenses 15 (mortgage interest), Depreciation 8 (basis)
    required = [
        "'Property Info'!B8", "'Property Info'!B9",
        "'Income'!B8",
        "'Expenses'!B15", "'Depreciation'!B8",
    ]
    countblank_req = " + ".join(f'IF({r}="",1,0)' for r in required)
    c.value = (f'=IF(({countblank_req})=0,"READY",'
               f'IF(({countblank_req})<=2,"MINOR","NEEDS WORK"))')
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I12:L12")
    c = ws["I12"]; c.value = "0 missing = green"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        from openpyxl.utils.cell import column_index_from_string
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # Print packet button (rows 16-20)
    pseudo_button(ws, "A16", "L20",
                  "📄  PRINT SCHEDULE E MAP  →",
                  "'Schedule E Map'!A1", variant="primary")
    for r in range(16, 21):
        ws.row_dimensions[r].height = 24

    ws.merge_cells("A21:L21")
    c = ws["A21"]
    c.value = ("Tip: print 1 page (the Schedule E Map tab) and email it "
                "with your supporting workbook to your CPA — saves billable hours.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[21].height = 18

    # Upgrade callout (row 23)
    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = (
        "💡 Upgrade to the Tax Season Bundle ($147) at thestrledger.com/tax-bundle — "
        "Schedule E + Schedule C + Mileage + 1099 + Home Office + Section 179 + "
        "Quarterly Estimateds + Per-Diem."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[23].height = 36

    # Footer
    brand_footer(ws, 25, version_line=VERSION_LINE)

    # Print setup
    ws.print_area = "A1:L27"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Tab 8 — Settings: tax year, classification, dropdown lists."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Settings",
                         prev_tab="Launch", next_tab=None)

    # Subtitle row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · Schedule classification · references"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 36), ("B", 22), ("C", 6),
        ("D", 16), ("E", 16),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # Active tax year (row 5)
    a = ws.cell(row=5, column=1, value="Active tax year:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=5, column=2, value=2026)
    apply_style(b, input_cell_style())
    b.number_format = "0"
    ws.row_dimensions[5].height = 18

    # Schedule classification (row 6)
    a = ws.cell(row=6, column=1, value="Schedule classification:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=6, column=2, value="Schedule E (passive)")
    apply_style(b, input_cell_style())
    add_dropdown(ws, "B6", ["Schedule E (passive)", "Schedule C (active)",
                              "Ask my CPA"])
    ws.row_dimensions[6].height = 18

    # Substantial-services callout (row 7)
    ws.merge_cells("A7:E7")
    c = ws["A7"]
    c.value = (
        "⚠ Substantial-services trigger: if you provide cleaning DURING stays, "
        "meals, or concierge, IRS requires Schedule C (15.3% SE tax)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[7].height = 28

    # Year-end Archive (rows 9-15)
    ws.row_dimensions[8].height = 10
    a = ws.cell(row=9, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[9].height = 22

    ws.merge_cells("A10:E10")
    c = ws["A10"]
    c.value = ("Each January, copy total Net (Line 26) into the row for the "
               "closing year, then update the Active tax year above.")
    c.font = italic_muted
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[10].height = 28

    archive_headers = ["Year", "Schedule E Line 26", "Active/Material?"]
    for col, h in enumerate(archive_headers, start=1):
        c = ws.cell(row=11, column=col, value=h)
        apply_style(c, header_row_style())
    ws.row_dimensions[11].height = 18

    for idx, year in enumerate(range(2024, 2031), start=12):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=idx, column=2).number_format = '"$"#,##0;[Red]("$"#,##0)'
        ws.row_dimensions[idx].height = 16

    # Footer
    brand_footer(ws, 21, version_line=VERSION_LINE)


# --- Main ---

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_info_tab(wb, variant)
    build_income_tab(wb, variant)
    build_expenses_tab(wb, variant)
    build_depreciation_tab(wb, variant)
    build_loss_limitations_tab(wb, variant)
    build_schedule_e_map_tab(wb, variant)
    build_launch_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Schedule E Tax-Prep Workbook — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "IRS-compliant Schedule E (Form 1040) tax-prep workbook for STR hosts. "
        "Three properties, full Part I mapping."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
