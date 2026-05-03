"""Build FIN-004 DSCR Tracker (v2.2 standard).

Operational-mode portfolio calculator — computes Debt Service Coverage
Ratio (NOI / Annual Debt Service) per property and portfolio-wide.
Critical for portfolio hosts pursuing DSCR loans (no W-2 needed if
DSCR >= 1.20).

Five tabs: Start / Per-Property / Stress Test / Refi Eligibility /
Settings. Capacity = 10 properties.

Generates two files:
  templates/_masters/FIN-004-dscr-tracker-DEMO.xlsx
  templates/_masters/FIN-004-dscr-tracker-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
)

SKU = "FIN-004"
NAME = "dscr-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Capacity
# ---------------------------------------------------------------------------

PROP_CAPACITY = 10  # property rows on Per-Property + Stress + Refi tabs

# Stress scenarios — multiplier on annual gross revenue.
# Brief: -10% / -20% / -30% (recession-like).
STRESS_SCENARIOS = [
    ("-10% Revenue", 0.90),
    ("-20% Revenue", 0.80),
    ("-30% Revenue", 0.70),
]

# ---------------------------------------------------------------------------
# Demo data — 3 properties matching brief QA spec:
#   Smokies Ridge: DSCR 1.42 (qualifies)
#   Lakehouse:     DSCR 1.18 (marginal)
#   Creek Side:    DSCR 0.94 (negative — at risk)
#   Portfolio DSCR: ~1.21
# ---------------------------------------------------------------------------

# Tuned so DSCR = NOI / total_debt_service rounds to brief targets.
# Smokies Ridge:  Rev 90,000, OpEx 28,000 -> NOI 62,000;
#                 P&I 3,640/mo = 43,680/yr; +0 other = 43,680;
#                 DSCR = 62,000 / 43,680 = 1.4194 -> 1.42 PASS.
# Lakehouse:      Rev 78,000, OpEx 25,000 -> NOI 53,000;
#                 P&I 3,742.94/mo = 44,915.25/yr; +0 = 44,915;
#                 DSCR = 53,000 / 44,915.25 = 1.1800 -> 1.18 PASS.
# Creek Side:     Rev 52,000, OpEx 21,500 -> NOI 30,500;
#                 P&I 2,704.43/mo = 32,453.19/yr; +0 = 32,453;
#                 DSCR = 30,500 / 32,453 = 0.9398 -> 0.94 PASS.
PROPERTIES_DEMO = [
    # (name, gross_rev, op_ex, mortgage_pi_monthly, other_debt_annual,
    #  est_value, loan_balance)
    ("Smokies Ridge Cabin", 90000, 28000, 3640.00,    0, 525000, 360000),
    ("Lakehouse A",          78000, 25000, 3742.94,    0, 540000, 405000),
    ("Creek Side",           52000, 21500, 2704.43,    0, 340000, 272000),
]

# Pad to PROP_CAPACITY rows
def _pad_props(props):
    padded = list(props)
    while len(padded) < PROP_CAPACITY:
        padded.append((None, None, None, None, None, None, None))
    return padded


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "DSCR Tracker"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Which properties qualify for a DSCR loan refinance?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} - v1.0"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: KPI cards rows 10-22 (parchment)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Headline: Portfolio DSCR
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "PORTFOLIO DSCR"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 18

    ws.merge_cells("A12:L13")
    c = ws["A12"]
    # Portfolio DSCR = SUM(NOI) / SUM(Total Debt Service)
    c.value = (
        "=IFERROR(SUMIF('Per-Property'!A8:A17,\"<>\",'Per-Property'!D8:D17)/"
        "SUMIF('Per-Property'!A8:A17,\"<>\",'Per-Property'!H8:H17),0)"
    )
    c.font = Font(name=FONT_HEAD, size=48, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.00"
    ws.row_dimensions[12].height = 36
    ws.row_dimensions[13].height = 24

    # Status line
    ws.merge_cells("A14:L14")
    c = ws["A14"]
    c.value = (
        '=IF(A12>=Settings!B7,"Portfolio qualifies for DSCR loan refinance",'
        'IF(A12>=Settings!B9,"Portfolio is marginal - shop rates carefully",'
        '"Portfolio is below break-even - debt restructuring needed"))'
    )
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 22

    # Two-column KPI strip rows 16-19
    # Left: # properties >= 1.20
    ws.merge_cells("A16:F16")
    c = ws["A16"]
    c.value = "PROPERTIES QUALIFYING (DSCR >= 1.20)"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A17:F19")
    c = ws["A17"]
    c.value = (
        '=COUNTIFS(\'Per-Property\'!A8:A17,"<>",\'Per-Property\'!I8:I17,'
        '">="&Settings!B7)&" of "&COUNTIF(\'Per-Property\'!A8:A17,"<>")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Right: weakest property
    ws.merge_cells("G16:L16")
    c = ws["G16"]
    c.value = "WEAKEST PROPERTY"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G17:L18")
    c = ws["G17"]
    c.value = (
        '=IFERROR(INDEX(\'Per-Property\'!A8:A17,'
        'MATCH(MINIFS(\'Per-Property\'!I8:I17,\'Per-Property\'!A8:A17,"<>"),'
        '\'Per-Property\'!I8:I17,0)),"-")'
    )
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G19:L19")
    c = ws["G19"]
    c.value = (
        '=IFERROR("DSCR = "&TEXT(MINIFS(\'Per-Property\'!I8:I17,'
        '\'Per-Property\'!A8:A17,"<>"),"0.00"),"")'
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 28
    ws.row_dimensions[18].height = 22
    ws.row_dimensions[19].height = 18

    # Note
    ws.merge_cells("A21:L22")
    c = ws["A21"]
    c.value = (
        "DSCR = Net Operating Income / Annual Debt Service. Most DSCR "
        "lenders require >= 1.20 with LTV <= 75%. Below 1.0 means the "
        "property doesn't cover its own debt - refinance, restructure, or sell."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[21].height = 22
    ws.row_dimensions[22].height = 22

    # ZONE 3: Pseudo-button nav
    pseudo_button(ws, "A25", "F27", "->  Update Inputs",
                  "'Per-Property'!A1", variant="primary")
    pseudo_button(ws, "G25", "L27", "Stress Test",
                  "'Stress Test'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "F31", "Refi Eligibility",
                  "'Refi Eligibility'!A1", variant="secondary")
    pseudo_button(ws, "G29", "L31", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        f"Want full P&L per property? Get TAX-002 Single-Property P&L "
        f"Tracker at {BRAND_DOMAIN} - Schedule E-ready, $47."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 30

    brand_footer(ws, 35,
                 version_line=f"{SKU} - v1.0 - Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_per_property_tab(wb, variant):
    ws = wb.create_sheet("Per-Property")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 24),  # Property
        ("B", 14),  # Annual gross revenue
        ("C", 14),  # OpEx
        ("D", 14),  # NOI
        ("E", 14),  # Mortgage P&I monthly
        ("F", 14),  # Annual debt service (P&I*12)
        ("G", 14),  # Other annual debt
        ("H", 14),  # Total debt service
        ("I", 10),  # DSCR
        ("J", 16),  # Status
        ("K", 4),
        ("L", 4),
    ])

    compact_header_band(ws, "Per-Property",
                        prev_tab="Start", next_tab="Stress Test")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Yellow = edit; gray = calculated. NOI = Revenue - OpEx (no mortgage). "
        "Debt service = mortgage + HELOC + other annual debt."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 7 (table starts row 7, data rows 8-17, totals row 18)
    headers = [
        "Property",
        "Annual Gross Revenue",
        "Operating Expenses",
        "NOI",
        "Mortgage P&I (monthly)",
        "Annual Debt Service",
        "Other Annual Debt",
        "Total Debt Service",
        "DSCR",
        "Status",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=label)
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
    ws.row_dimensions[7].height = 36
    ws.freeze_panes = "A8"

    # Data rows 8-17
    props = _pad_props(PROPERTIES_DEMO if variant == "DEMO" else [])
    thin_side = Side(style="thin", color="DDDDDD")
    cell_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    for i, (name, rev, opex, pi_mo, other, _val, _bal) in enumerate(props):
        r = 8 + i

        # Col A: Property name (input)
        cell = ws.cell(row=r, column=1, value=name)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = cell_border

        # Col B: Annual gross revenue (input)
        cell = ws.cell(row=r, column=2, value=rev)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col C: Operating expenses (input, no mortgage)
        cell = ws.cell(row=r, column=3, value=opex)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col D: NOI (formula = B - C)
        cell = ws.cell(row=r, column=4,
                       value=f'=IF(A{r}="","",B{r}-C{r})')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col E: Mortgage P&I monthly (input)
        cell = ws.cell(row=r, column=5, value=pi_mo)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0.00'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col F: Annual debt service (formula = E*12)
        cell = ws.cell(row=r, column=6,
                       value=f'=IF(A{r}="","",E{r}*12)')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col G: Other annual debt (input - HELOC, etc.)
        cell = ws.cell(row=r, column=7, value=other)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col H: Total debt service (formula = F + G)
        cell = ws.cell(row=r, column=8,
                       value=f'=IF(A{r}="","",F{r}+G{r})')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col I: DSCR (formula = D / H)
        cell = ws.cell(row=r, column=9,
                       value=f'=IFERROR(IF(A{r}="","",D{r}/H{r}),"")')
        apply_style(cell, formula_cell_style())
        cell.number_format = "0.00"
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        # Col J: Status (formula)
        cell = ws.cell(row=r, column=10, value=(
            f'=IF(A{r}="","",'
            f'IF(I{r}>=Settings!$B$7,"Qualifies",'
            f'IF(I{r}>=Settings!$B$9,"Marginal","Negative")))'
        ))
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        ws.row_dimensions[r].height = 22

    # Conditional formatting on DSCR column I8:I17
    # Green-ish (parchment-alt + bold navy): >= Settings!B7 (1.20)
    # Gold: 1.0 - 1.19 (between threshold_neg and threshold_qualify)
    # Red: < Settings!B9 (1.0)
    dscr_range = "I8:I17"
    # Red — below 1.0 threshold
    ws.conditional_formatting.add(
        dscr_range,
        FormulaRule(
            formula=[f'AND($A8<>"",$I8<>"",$I8<Settings!$B$9)'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR),
        ),
    )
    # Gold — marginal (1.0 to <1.20)
    ws.conditional_formatting.add(
        dscr_range,
        FormulaRule(
            formula=[
                f'AND($A8<>"",$I8<>"",$I8>=Settings!$B$9,'
                f'$I8<Settings!$B$7)'
            ],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY),
        ),
    )
    # Green-ish — qualifies (>= 1.20). Use COLOR_PARCHMENT_ALT as the
    # green-ish accent + COLOR_ACCENT bold text (per brand_config palette
    # — no new tokens; brief permits parchment/gold for green-ish).
    ws.conditional_formatting.add(
        dscr_range,
        FormulaRule(
            formula=[f'AND($A8<>"",$I8<>"",$I8>=Settings!$B$7)'],
            fill=PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT),
        ),
    )

    # Status column conditional formatting — match the DSCR coloring
    status_range = "J8:J17"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'$J8="Negative"'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'$J8="Marginal"'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'$J8="Qualifies"'],
            fill=PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT),
        ),
    )

    # Totals row 18: portfolio totals + portfolio DSCR
    total_r = 18
    cell = ws.cell(row=total_r, column=1, value="PORTFOLIO TOTALS")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for col_idx in range(2, 11):
        c2 = ws.cell(row=total_r, column=col_idx)
        c2.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c2.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
        c2.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Sums for B, C, D, F, G, H
    for col_idx, col_letter in [(2, "B"), (3, "C"), (4, "D"),
                                  (6, "F"), (7, "G"), (8, "H")]:
        cell = ws.cell(row=total_r, column=col_idx,
                       value=f'=SUM({col_letter}8:{col_letter}17)')
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # E column: average mortgage P&I (less informative as a sum)
    cell = ws.cell(row=total_r, column=5, value="")
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    # I column: portfolio DSCR = SUM(NOI) / SUM(Total Debt Service)
    cell = ws.cell(row=total_r, column=9,
                   value=f'=IFERROR(D{total_r}/H{total_r},0)')
    cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.number_format = "0.00"
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # J column: portfolio status
    cell = ws.cell(row=total_r, column=10, value=(
        f'=IF(I{total_r}>=Settings!$B$7,"Qualifies",'
        f'IF(I{total_r}>=Settings!$B$9,"Marginal","Negative"))'
    ))
    cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[total_r].height = 28

    # Legend row 20
    ws.merge_cells("A20:L21")
    c = ws["A20"]
    c.value = (
        "Status: Qualifies = DSCR >= 1.20 (DSCR-loan eligible). "
        "Marginal = 1.0 - 1.19 (covers debt but won't pass most DSCR-loan UW). "
        "Negative = < 1.0 (property is bleeding cash)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[20].height = 18
    ws.row_dimensions[21].height = 18

    brand_footer(ws, 23,
                 version_line=f"{SKU} - v1.0 - Per-Property")

    ws.print_area = "A1:L25"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_stress_test_tab(wb, variant):
    ws = wb.create_sheet("Stress Test")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    # Layout: A property, then for each of 3 scenarios = 3 cols
    # (Stressed Revenue, NOI, DSCR), totalling 9 scenario cols + 1 label = 10.
    set_col_widths(ws, [
        ("A", 22),  # Property
        ("B", 14), ("C", 14), ("D", 9),    # Scenario 1
        ("E", 14), ("F", 14), ("G", 9),    # Scenario 2
        ("H", 14), ("I", 14), ("J", 9),    # Scenario 3
        ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Stress Test",
                        prev_tab="Per-Property", next_tab="Refi Eligibility")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "What happens to DSCR if revenue drops? Operating expenses and "
        "debt service stay constant. Watch which properties slip below 1.0."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Scenario super-header row 6 (merged across each 3-col scenario block)
    scenario_blocks = [
        ("B6:D6", STRESS_SCENARIOS[0][0]),
        ("E6:G6", STRESS_SCENARIOS[1][0]),
        ("H6:J6", STRESS_SCENARIOS[2][0]),
    ]
    for merge_range, label in scenario_blocks:
        ws.merge_cells(merge_range)
        c = ws[merge_range.split(":")[0]]
        c.value = label
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
        c.fill = PatternFill("solid", fgColor=COLOR_NAVY_TINT)
        c.alignment = Alignment(horizontal="center", vertical="center")
    # Row 6 col A header
    cell = ws.cell(row=6, column=1, value="")
    cell.fill = PatternFill("solid", fgColor=COLOR_NAVY_TINT)
    # Cols K-L spacer fill
    for c in range(11, 13):
        ws.cell(row=6, column=c).fill = PatternFill("solid", fgColor=COLOR_NAVY_TINT)
    ws.row_dimensions[6].height = 22

    # Sub-header row 7
    sub_headers = [
        (1, "Property"),
        (2, "Stressed Revenue"), (3, "Stressed NOI"), (4, "DSCR"),
        (5, "Stressed Revenue"), (6, "Stressed NOI"), (7, "DSCR"),
        (8, "Stressed Revenue"), (9, "Stressed NOI"), (10, "DSCR"),
    ]
    for col_idx, label in sub_headers:
        cell = ws.cell(row=7, column=col_idx, value=label)
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[7].height = 32
    ws.freeze_panes = "B8"

    # Data rows 8-17 — pull property name, then compute stressed metrics
    thin_side = Side(style="thin", color="DDDDDD")
    cell_border = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    for i in range(PROP_CAPACITY):
        r = 8 + i
        # Col A: pull property name from Per-Property
        cell = ws.cell(row=r, column=1,
                       value=f"=IF('Per-Property'!A{r}=\"\",\"\",'Per-Property'!A{r})")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.border = cell_border

        # For each scenario, 3 cols: stressed revenue, stressed NOI, DSCR
        for s_idx, (_label, mult) in enumerate(STRESS_SCENARIOS):
            base_col = 2 + s_idx * 3   # B=2, E=5, H=8

            # Stressed revenue = Per-Property!B * mult
            cell = ws.cell(row=r, column=base_col,
                           value=f"=IF(A{r}=\"\",\"\",'Per-Property'!B{r}*{mult})")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.border = cell_border

            # Stressed NOI = Stressed revenue - OpEx
            stressed_rev_letter = get_column_letter(base_col)
            noi_col = base_col + 1
            cell = ws.cell(row=r, column=noi_col,
                           value=(
                               f"=IF(A{r}=\"\",\"\",{stressed_rev_letter}{r}"
                               f"-'Per-Property'!C{r})"
                           ))
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.border = cell_border

            # Stressed DSCR = Stressed NOI / Total debt service
            noi_letter = get_column_letter(noi_col)
            dscr_col = base_col + 2
            cell = ws.cell(row=r, column=dscr_col, value=(
                f"=IFERROR(IF(A{r}=\"\",\"\",{noi_letter}{r}/'Per-Property'!H{r}),\"\")"
            ))
            apply_style(cell, formula_cell_style())
            cell.number_format = "0.00"
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        ws.row_dimensions[r].height = 22

    # Conditional formatting on each DSCR column (D, G, J)
    for dscr_col_letter in ["D", "G", "J"]:
        rng = f"{dscr_col_letter}8:{dscr_col_letter}17"
        # First-row anchor (row 8) for the formula
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'AND($A8<>"",{dscr_col_letter}8<>"",'
                    f'{dscr_col_letter}8<Settings!$B$9)'
                ],
                fill=PatternFill("solid", fgColor="FFCCCC"),
                font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'AND($A8<>"",{dscr_col_letter}8<>"",'
                    f'{dscr_col_letter}8>=Settings!$B$9,'
                    f'{dscr_col_letter}8<Settings!$B$7)'
                ],
                fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
                font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'AND($A8<>"",{dscr_col_letter}8<>"",'
                    f'{dscr_col_letter}8>=Settings!$B$7)'
                ],
                fill=PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT),
                font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT),
            ),
        )

    # Summary row 19 — # of properties at risk per scenario
    ws.merge_cells("A19:A19")
    cell = ws.cell(row=19, column=1, value="# AT RISK (DSCR < 1.0)")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color="F6EFE2")
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for s_idx in range(3):
        dscr_col = 4 + s_idx * 3
        dscr_letter = get_column_letter(dscr_col)
        # Span the scenario's 3 cols
        first_col_letter = get_column_letter(2 + s_idx * 3)
        last_col_letter = get_column_letter(4 + s_idx * 3)
        ws.merge_cells(f"{first_col_letter}19:{last_col_letter}19")
        c = ws.cell(row=19, column=2 + s_idx * 3,
                    value=(
                        f'=COUNTIFS(A8:A17,"<>",{dscr_letter}8:{dscr_letter}17,'
                        f'"<"&Settings!$B$9)&" of "&COUNTIF(A8:A17,"<>")'
                    ))
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[19].height = 26
    # Spacer cols K-L on row 19
    for c in range(11, 13):
        ws.cell(row=19, column=c).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    # Note row 21
    ws.merge_cells("A21:L22")
    c = ws["A21"]
    c.value = (
        "Stress test holds operating expenses and debt service constant - "
        "real downturns may also reduce OpEx (cleanings, supplies). "
        "The pessimism is intentional: a property that survives -30% revenue "
        "with no OpEx relief will survive most real-world scenarios."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[21].height = 22
    ws.row_dimensions[22].height = 22

    brand_footer(ws, 24,
                 version_line=f"{SKU} - v1.0 - Stress Test")

    ws.print_area = "A1:L26"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_refi_eligibility_tab(wb, variant):
    ws = wb.create_sheet("Refi Eligibility")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 24),  # Property
        ("B", 12),  # Current DSCR
        ("C", 16),  # Estimated value
        ("D", 16),  # Loan balance
        ("E", 12),  # Implied LTV
        ("F", 14),  # Max LTV (from settings, displayed)
        ("G", 12),  # DSCR pass?
        ("H", 12),  # LTV pass?
        ("I", 18),  # DSCR-loan eligible
        ("J", 4), ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Refi Eligibility",
                        prev_tab="Stress Test", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "DSCR loans typically require DSCR >= 1.20 AND LTV <= 75%. "
        "Enter property value + loan balance to see which qualify."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 7
    headers = [
        "Property", "Current DSCR", "Est. Property Value",
        "Current Loan Balance", "Implied LTV", "Max LTV",
        "DSCR Pass?", "LTV Pass?", "DSCR-Loan Eligible",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=label)
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[7].height = 36
    ws.freeze_panes = "A8"

    thin_side = Side(style="thin", color="DDDDDD")
    cell_border = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    props = _pad_props(PROPERTIES_DEMO if variant == "DEMO" else [])

    for i, (_name, _rev, _opex, _pi_mo, _other, val, bal) in enumerate(props):
        r = 8 + i

        # Col A: Property name (pulled from Per-Property)
        cell = ws.cell(row=r, column=1,
                       value=f"=IF('Per-Property'!A{r}=\"\",\"\",'Per-Property'!A{r})")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.border = cell_border

        # Col B: Current DSCR (pulled)
        cell = ws.cell(row=r, column=2,
                       value=f"=IFERROR(IF('Per-Property'!A{r}=\"\",\"\",'Per-Property'!I{r}),\"\")")
        apply_style(cell, formula_cell_style())
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        # Col C: Estimated property value (input)
        cell = ws.cell(row=r, column=3, value=val)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col D: Current loan balance (input)
        cell = ws.cell(row=r, column=4, value=bal)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = cell_border

        # Col E: Implied LTV (formula = D / C)
        cell = ws.cell(row=r, column=5,
                       value=f'=IFERROR(IF(OR(A{r}="",C{r}=""),"",D{r}/C{r}),"")')
        apply_style(cell, formula_cell_style())
        cell.number_format = "0.0%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        # Col F: Max LTV (from settings)
        cell = ws.cell(row=r, column=6,
                       value=f'=IF(A{r}="","",Settings!$B$11)')
        apply_style(cell, formula_cell_style())
        cell.number_format = "0%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        # Col G: DSCR pass?
        cell = ws.cell(row=r, column=7, value=(
            f'=IF(A{r}="","",IF(B{r}>=Settings!$B$7,"Pass","Fail"))'
        ))
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.border = cell_border

        # Col H: LTV pass?
        cell = ws.cell(row=r, column=8, value=(
            f'=IF(OR(A{r}="",E{r}=""),"",'
            f'IF(E{r}<=Settings!$B$11,"Pass","Fail"))'
        ))
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.border = cell_border

        # Col I: DSCR-loan eligible (both pass)
        cell = ws.cell(row=r, column=9, value=(
            f'=IF(OR(A{r}="",E{r}=""),"",'
            f'IF(AND(B{r}>=Settings!$B$7,E{r}<=Settings!$B$11),'
            f'"Eligible","Not Eligible"))'
        ))
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        cell.border = cell_border

        ws.row_dimensions[r].height = 22

    # Conditional formatting on Pass/Fail cols G + H
    for col_letter in ["G", "H"]:
        rng = f"{col_letter}8:{col_letter}17"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'${col_letter}8="Pass"'],
                fill=PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT),
                font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'${col_letter}8="Fail"'],
                fill=PatternFill("solid", fgColor="FFCCCC"),
                font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
            ),
        )

    # Eligible col I
    ws.conditional_formatting.add(
        "I8:I17",
        FormulaRule(
            formula=['$I8="Eligible"'],
            fill=PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT),
        ),
    )
    ws.conditional_formatting.add(
        "I8:I17",
        FormulaRule(
            formula=['$I8="Not Eligible"'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR),
        ),
    )

    # Summary row 19
    cell = ws.cell(row=19, column=1, value="# DSCR-LOAN ELIGIBLE")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color="F6EFE2")
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for col_idx in range(2, 9):
        c2 = ws.cell(row=19, column=col_idx)
        c2.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    cell = ws.cell(row=19, column=9, value=(
        '=COUNTIF(I8:I17,"Eligible")&" of "&COUNTIF(A8:A17,"<>")'
    ))
    cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in range(10, 13):
        ws.cell(row=19, column=c).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    ws.row_dimensions[19].height = 26

    # Note row 21
    ws.merge_cells("A21:L22")
    c = ws["A21"]
    c.value = (
        "Lender thresholds vary - many DSCR lenders accept 1.10-1.15 with LTV "
        "<= 70% or charge a rate premium. Settings tab lets you adjust the "
        "qualify threshold + max LTV to match a specific lender's overlay."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[21].height = 22
    ws.row_dimensions[22].height = 22

    brand_footer(ws, 24,
                 version_line=f"{SKU} - v1.0 - Refi Eligibility")

    ws.print_area = "A1:L26"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_NAVY_TINT

    set_col_widths(ws, [
        ("A", 36), ("B", 18),
        ("C", 4), ("D", 4), ("E", 4), ("F", 4),
        ("G", 4), ("H", 4), ("I", 4), ("J", 4),
        ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Refi Eligibility", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Adjust thresholds and active year. Yellow = edit."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Active year
    _settings_row(ws, 5, "Active year:",
                  2026 if variant == "DEMO" else None, "0")

    # DSCR threshold for Qualifies
    _settings_row(ws, 7, "DSCR threshold - Qualifies (>=):",
                  1.20, "0.00",
                  note="Most DSCR lenders use 1.20. Some accept 1.10-1.15 with overlay.")

    # DSCR threshold for Negative
    _settings_row(ws, 9, "DSCR threshold - Negative (<):",
                  1.00, "0.00",
                  note="Below 1.0 means the property doesn't cover its own debt.")

    # Max LTV for refi
    _settings_row(ws, 11, "Max LTV for DSCR-loan refi:",
                  0.75, "0%",
                  note="Standard cap is 75%. Cash-out refis often capped 70%.")

    # Property list rows 13-22 (10 properties)
    ws.merge_cells("A13:B13")
    c = ws["A13"]
    c.value = "PROPERTY LIST"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[13].height = 20

    ws.merge_cells("A14:B14")
    c = ws["A14"]
    c.value = (
        "(Reference list - edit names directly on the Per-Property tab. "
        "These mirror via formula.)"
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[14].height = 16

    for i in range(PROP_CAPACITY):
        r = 15 + i
        cell = ws.cell(row=r, column=1, value=f"Property #{i + 1}")
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        cell = ws.cell(row=r, column=2,
                       value=f"=IF('Per-Property'!A{8 + i}=\"\",\"-\",'Per-Property'!A{8 + i})")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 27,
                 version_line=f"{SKU} - v1.0 - Settings")

    ws.print_area = "A1:L29"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _settings_row(ws, row, label, value, fmt, note=None):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=row, column=2, value=value)
    apply_style(cell, input_cell_style())
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

    if note:
        ws.merge_cells(f"A{row + 1}:L{row + 1}")
        c = ws[f"A{row + 1}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row + 1].height = 14

    ws.row_dimensions[row].height = 22


# ---------------------------------------------------------------------------
# Build entrypoint
# ---------------------------------------------------------------------------

def build(variant):
    """variant: 'DEMO' or 'BLANK'"""
    wb = Workbook()
    build_start_tab(wb, variant)
    build_per_property_tab(wb, variant)
    build_stress_test_tab(wb, variant)
    build_refi_eligibility_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"DSCR Tracker ({variant}) - The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Portfolio Debt Service Coverage Ratio tracker for STR hosts - "
        "DSCR loan eligibility, stress testing, refi screening."
    )

    out = DEMO_OUT if variant == "DEMO" else BLANK_OUT
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")


def main():
    build("DEMO")
    build("BLANK")


if __name__ == "__main__":
    main()
