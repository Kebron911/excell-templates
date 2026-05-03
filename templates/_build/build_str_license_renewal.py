"""Build LGL-001 STR License Renewal Calendar (v2.2 standard).

Operational-mode tool focused on the STR-specific subset of permits
that, if lapsed, knock a listing offline: city STR permit, business
license, sales tax permit, fire/safety inspection, lottery renewal.

Sibling product to OPS-003 (broader permit tracker). Same shape,
narrower domain — copy the structural pattern, drop the multi-market
discovery + filing-log chrome.

Generates two files:
  templates/_masters/LGL-001-str-license-renewal-calendar-DEMO.xlsx
  templates/_masters/LGL-001-str-license-renewal-calendar-BLANK.xlsx
"""
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, apply_brand_header,
    brand_footer,
)

SKU = "LGL-001"
NAME = "str-license-renewal-calendar"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Active year drives the 12-Month Calendar tab (Settings!B40).
ACTIVE_YEAR = 2026

# --- Sample data (DEMO) ---
# Per brief: 12 licenses across 3 properties in 2 jurisdictions
# (Sevierville TN + Gatlinburg TN). 1 expired, 2 critical (<30d),
# 3 soon (<90d), rest active. Annual fees ~$680 portfolio-wide.
#
# Tuple: (license_no, property, jurisdiction, ltype, issued, expires,
#         fee, url, lead_days, notes)
LICENSE_SAMPLES = [
    # Smokies Ridge Cabin — Gatlinburg / Sevier (5 licenses)
    ("GAT-STR-2024-0481", "Smokies Ridge Cabin", "Gatlinburg, TN",
     "STR Permit", "2024-09-15", "2026-09-15",
     385.00, "https://gatlinburgtn.gov/short-term-rentals", 60,
     "Tier 2 vacation rental; renews annually."),
    ("SEV-BL-117820", "Smokies Ridge Cabin", "Sevier County, TN",
     "Business License", "2025-12-31", "2026-12-31",
     50.00, "https://www.seviercountytn.gov/business-license.html", 30,
     "Renews end of calendar year."),
    ("TN-SUT-44829110", "Smokies Ridge Cabin", "State of Tennessee",
     "Sales Tax", "2024-01-01", "2027-01-01",
     0.00, "https://tntap.tn.gov/", 30,
     "TN sales tax registration; no fee."),
    ("GAT-FIRE-2025-1144", "Smokies Ridge Cabin", "Gatlinburg, TN",
     "Fire/Safety", "2025-07-10", "2026-07-10",
     75.00, "https://gatlinburgtn.gov/fire-marshal", 45,
     "Annual smoke/CO + extinguisher inspection."),
    ("GAT-LOT-2024-117", "Smokies Ridge Cabin", "Gatlinburg, TN",
     "Lottery", "2025-01-15", "2026-08-15",
     25.00, "https://gatlinburgtn.gov/str-lottery", 30,
     "Annual lottery slot retention fee."),

    # Creek Side — Sevierville / Sevier County (4 licenses, 1 expired)
    ("SVL-STR-2025-0922", "Creek Side", "Sevierville, TN",
     "STR Permit", "2025-04-30", "2026-04-30",
     200.00, "https://www.seviervilletn.org/strs.html", 60,
     "EXPIRED — renew immediately to avoid listing suspension."),
    ("SVL-BL-660119", "Creek Side", "Sevierville, TN",
     "Business License", "2025-06-01", "2026-06-01",
     30.00, "https://www.seviervilletn.org/business.html", 30,
     "City business license."),
    ("SEV-FIRE-2024-0402", "Creek Side", "Sevier County, TN",
     "Fire/Safety", "2024-08-15", "2026-05-25",
     65.00, "https://www.seviercountytn.gov/fire-prevention.html", 45,
     "RED ZONE — 23 days. Schedule re-inspection now."),
    ("TN-SUT-5511980", "Creek Side", "State of Tennessee",
     "Sales Tax", "2025-01-01", "2028-01-01",
     0.00, "https://tntap.tn.gov/", 30,
     "Triennial; no fee."),

    # Lakeside Lodge — Sevierville (3 licenses)
    ("SVL-STR-2024-0114", "Lakeside Lodge", "Sevierville, TN",
     "STR Permit", "2024-08-01", "2026-06-12",
     200.00, "https://www.seviervilletn.org/strs.html", 60,
     "RED ZONE — 41 days remaining."),
    ("SVL-BL-882201", "Lakeside Lodge", "Sevierville, TN",
     "Business License", "2025-08-01", "2026-08-01",
     30.00, "https://www.seviervilletn.org/business.html", 30,
     "Annual city business license."),
    ("SEV-FIRE-2025-1190", "Lakeside Lodge", "Sevier County, TN",
     "Fire/Safety", "2025-07-30", "2026-07-30",
     65.00, "https://www.seviercountytn.gov/fire-prevention.html", 45,
     "GOLD — 89 days. Schedule with fire marshal."),
]

# Settings tab data
PROPERTIES_LIST = [
    "Smokies Ridge Cabin", "Creek Side", "Lakeside Lodge",
    "Mountain View Loft", "Pigeon Forge Retreat",
]
JURISDICTION_LIST = [
    "Gatlinburg, TN", "Sevierville, TN", "Sevier County, TN",
    "State of Tennessee", "Pigeon Forge, TN",
]
LICENSE_TYPE_LIST = [
    "STR Permit", "Business License", "Sales Tax",
    "Fire/Safety", "Lottery", "Health", "Other",
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list, blank_length=None):
    if variant == "demo":
        return demo_list
    n = blank_length if blank_length is not None else len(demo_list)
    return [None] * n


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero band rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "STR License Renewal Calendar"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Never lose a listing to an expired permit."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: KPI cards rows 10-15 — Active / Expiring 30 / Expiring 90 / YTD fees
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    def kpi_card(col_first, col_last, label, formula, num_format,
                  big_color):
        ws.merge_cells(f"{col_first}10:{col_last}10")
        c = ws[f"{col_first}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=big_color)
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{col_first}11:{col_last}13")
        c = ws[f"{col_first}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=28, bold=True, color=big_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if num_format:
            c.number_format = num_format

    # Card 1: Active licenses (status formula = ✅ Active)
    kpi_card("A", "C", "ACTIVE",
             '=COUNTIF(\'Licenses\'!H7:H106,"*Active*")',
             "0", COLOR_PRIMARY)
    ws.merge_cells("A14:C15")
    c = ws["A14"]
    c.value = "licenses currently in force"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)

    # Card 2: Expiring within 30 days
    kpi_card("D", "F", "EXPIRING <30 DAYS",
             '=COUNTIFS(\'Licenses\'!G7:G106,">=0",'
             '\'Licenses\'!G7:G106,"<30")',
             "0", COLOR_ERROR)
    ws.merge_cells("D14:F15")
    c = ws["D14"]
    c.value = "renew now — listing risk"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)

    # Card 3: Expiring 30-90 days
    kpi_card("G", "I", "EXPIRING 30-90",
             '=COUNTIFS(\'Licenses\'!G7:G106,">=30",'
             '\'Licenses\'!G7:G106,"<90")',
             "0", COLOR_ACCENT)
    ws.merge_cells("G14:I15")
    c = ws["G14"]
    c.value = "schedule renewal soon"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)

    # Card 4: Fees YTD (current calendar year, by Issued date)
    kpi_card("J", "L", "FEES YTD",
             '=SUMPRODUCT((YEAR(\'Licenses\'!E7:E106)=YEAR(TODAY()))'
             '*\'Licenses\'!I7:I106)',
             '"$"#,##0', COLOR_PRIMARY)
    ws.merge_cells("J14:L15")
    c = ws["J14"]
    c.value = "renewal fees paid this year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 16):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: Operating-while-expired warning row 17
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = (
        '=IF(COUNTIF(\'Licenses\'!H7:H106,"*Expired*")>0,'
        '"OPERATING WHILE EXPIRED — STOP. One or more licenses below are '
        'past their renewal date. Renew before the next booking starts.",'
        '"ALL CLEAR — No expired licenses in your portfolio.")'
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[17].height = 36

    ws.conditional_formatting.add(
        "A17:L17",
        FormulaRule(
            formula=['COUNTIF(\'Licenses\'!H7:H106,"*Expired*")=0'],
            fill=PatternFill("solid", fgColor="166534"),
        ),
    )

    # ZONE 4: How-to-use card rows 19-23
    ws.merge_cells("A19:L19")
    c = ws["A19"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[19].height = 22

    ws.merge_cells("A20:L23")
    c = ws["A20"]
    c.value = (
        "  1. Open Licenses — log every STR-related permit per property "
        "(STR permit, business license, sales tax, fire/safety, lottery).\n"
        "  2. Check the 12-Month Calendar tab to see fee timing across the "
        "next year — set aside cash for the high-fee months.\n"
        "  3. The Status column auto-flips to ⚠ Expired / 🔴 Critical / "
        "🟡 Soon / ✅ Active based on days-to-expiration."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)

    # ZONE 5: Pseudo-button nav rows 25-27
    pseudo_button(ws, "A25", "D27", "→ Add License",
                  "'Licenses'!A7", variant="primary")
    pseudo_button(ws, "E25", "H27", "12-Month Calendar",
                  "'12-Month Calendar'!A1", variant="primary")
    pseudo_button(ws, "I25", "L27", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    # ZONE 6: Upgrade banner row 29
    ws.merge_cells("A29:L29")
    c = ws["A29"]
    c.value = (
        "Need broader coverage across multi-jurisdiction portfolios? "
        f"Upgrade to OPS-003 License/Permit Tracker at {BRAND_DOMAIN}/upgrade."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[29].height = 36

    brand_footer(
        ws, 31,
        version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever",
    )

    ws.print_area = "A1:L33"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Licenses (master register)
# ---------------------------------------------------------------------------

def build_licenses_tab(wb, variant):
    ws = wb.create_sheet("Licenses")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(
        ws, "Licenses",
        prev_tab="Start", next_tab="12-Month Calendar",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per STR-related license. Days-to-expiration and Status "
        "auto-calculate from the Expires date — keep that column accurate."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Columns A-L per brief
    set_col_widths(ws, [
        ("A", 22), ("B", 22), ("C", 22), ("D", 18),
        ("E", 12), ("F", 12), ("G", 10), ("H", 14),
        ("I", 12), ("J", 28), ("K", 10), ("L", 30),
    ])

    headers = [
        "License #", "Property", "Jurisdiction", "Type",
        "Issued", "Expires", "Days", "Status",
        "Renewal $", "Process URL", "Lead days", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    samples = _val_list(variant, LICENSE_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 7 + i
        (license_no, prop, juris, ltype, issued, expires,
         fee, url, lead_days, notes) = row_data

        for col, val in [(1, license_no), (2, prop),
                         (3, juris), (4, ltype)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        e = ws.cell(row=row, column=5, value=_parse_date(issued))
        apply_style(e, input_cell_style())
        e.number_format = "yyyy-mm-dd"

        f_cell = ws.cell(row=row, column=6, value=_parse_date(expires))
        apply_style(f_cell, input_cell_style())
        f_cell.number_format = "yyyy-mm-dd"

        i_cell = ws.cell(row=row, column=9, value=fee)
        apply_style(i_cell, input_cell_style())
        i_cell.number_format = '"$"#,##0.00'

        j_cell = ws.cell(row=row, column=10, value=url)
        apply_style(j_cell, input_cell_style())
        j_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY,
                            underline="single")

        k_cell = ws.cell(row=row, column=11, value=lead_days)
        apply_style(k_cell, input_cell_style())
        k_cell.alignment = Alignment(horizontal="center", vertical="center")

        l_cell = ws.cell(row=row, column=12, value=notes)
        apply_style(l_cell, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Apply formulas + format to all rows 7-106 (100 capacity)
    for row in range(7, 107):
        # G: Days to expiration
        g = ws.cell(
            row=row, column=7,
            value=f'=IF(F{row}<>"",F{row}-TODAY(),"")',
        )
        apply_style(g, formula_cell_style())
        g.alignment = Alignment(horizontal="center", vertical="center")
        g.number_format = "0"

        # H: Status (per brief: Expired / Critical / Soon / Active)
        h = ws.cell(
            row=row, column=8,
            value=(
                f'=IF(F{row}="","",'
                f'IF(G{row}<0,"⚠ Expired",'
                f'IF(G{row}<30,"🔴 Critical",'
                f'IF(G{row}<90,"🟡 Soon","✅ Active"))))'
            ),
        )
        apply_style(h, formula_cell_style())
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)

        # Blank-input formatting for rows past samples
        if row > 6 + len(samples):
            for col in [5, 6]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            for col in [1, 2, 3, 4, 10, 12]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            cell9 = ws.cell(row=row, column=9)
            apply_style(cell9, input_cell_style())
            cell9.number_format = '"$"#,##0.00'
            cell11 = ws.cell(row=row, column=11)
            apply_style(cell11, input_cell_style())
            cell11.alignment = Alignment(horizontal="center",
                                           vertical="center")
            ws.row_dimensions[row].height = 20

    # Conditional formatting on Days col G:
    # < 0 expired (red), 0-29 critical (red soft), 30-89 gold soft (parchment-ish)
    ws.conditional_formatting.add(
        "G7:G106",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="FFB3B3"),
                   font=Font(name=FONT_BODY, size=10, bold=True,
                             color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "G7:G106",
        CellIsRule(operator="between", formula=["0", "29"],
                   fill=PatternFill("solid", fgColor="FFCCCC"),
                   font=Font(name=FONT_BODY, size=10, bold=True,
                             color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "G7:G106",
        CellIsRule(operator="between", formula=["30", "89"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )
    # >=90 stays parchment via input cell style; explicit rule for clarity
    ws.conditional_formatting.add(
        "G7:G106",
        CellIsRule(operator="greaterThanOrEqual", formula=["90"],
                   fill=PatternFill("solid", fgColor=COLOR_BG_LIGHT)),
    )

    # Status col H — color the whole cell by state
    ws.conditional_formatting.add(
        "H7:H106",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Expired",$H7))'],
            fill=PatternFill("solid", fgColor=COLOR_ERROR),
            font=Font(name=FONT_BODY, size=10, bold=True, color="FFFFFF"),
        ),
    )
    ws.conditional_formatting.add(
        "H7:H106",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Critical",$H7))'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        "H7:H106",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Soon",$H7))'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
            font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY),
        ),
    )
    ws.conditional_formatting.add(
        "H7:H106",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Active",$H7))'],
            fill=PatternFill("solid", fgColor=COLOR_BG_LIGHT),
            font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY),
        ),
    )

    # Dropdowns — Property B, Jurisdiction C, Type D
    add_dropdown(ws, "B7:B106", "=Settings!$B$5:$B$14")
    add_dropdown(ws, "C7:C106", "=Settings!$B$16:$B$25")
    add_dropdown(ws, "D7:D106", "=Settings!$B$27:$B$36")

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — 12-Month Calendar
# ---------------------------------------------------------------------------

def build_calendar_tab(wb, variant):
    ws = wb.create_sheet("12-Month Calendar")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(
        ws, "12-Month Calendar",
        prev_tab="Licenses", next_tab="Settings",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 16):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:O4")
    c = ws["A4"]
    c.value = (
        "Forward 12-month grid. Months are driven by Settings!B40 "
        "(active calendar year). Each cell shows the renewal fee in "
        "the month the license expires."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Columns: A License #, B Property, C Type, D-O 12 months
    set_col_widths(ws, [
        ("A", 22), ("B", 22), ("C", 18),
        ("D", 11), ("E", 11), ("F", 11), ("G", 11),
        ("H", 11), ("I", 11), ("J", 11), ("K", 11),
        ("L", 11), ("M", 11), ("N", 11), ("O", 11),
    ])

    # Header row 6 — License / Property / Type, then 12 months
    fixed_headers = ["License #", "Property", "Type"]
    for col, h in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())

    # Month headers driven by Settings!$B$40 (active year)
    # Cols D..O = months 1..12 of active year
    for m in range(12):
        col = 4 + m  # D=4
        cell = ws.cell(row=6, column=col)
        # Display "Jan 2026" style. Use DATE + TEXT.
        cell.value = (
            f'=TEXT(DATE(Settings!$B$40,{m + 1},1),"mmm yyyy")'
        )
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    # Data rows: pull from Licenses register (rows 7-106)
    # Per row: A=License#, B=Property, C=Type via direct ref;
    # months D..O: =IF(MONTH(Licenses!F7)=m, Licenses!I7, "")
    # but only when the renewal Expires falls inside the active year.
    last_row = 6 + 100  # 100 capacity
    for i in range(100):
        src = 7 + i
        row = 7 + i
        ws.cell(row=row, column=1,
                 value=f'=IF(\'Licenses\'!A{src}="","",\'Licenses\'!A{src})')
        ws.cell(row=row, column=2,
                 value=f'=IF(\'Licenses\'!B{src}="","",\'Licenses\'!B{src})')
        ws.cell(row=row, column=3,
                 value=f'=IF(\'Licenses\'!D{src}="","",\'Licenses\'!D{src})')
        for col_idx in range(1, 4):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                         indent=1)
            cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)

        for m in range(12):
            col = 4 + m
            # Show fee if Expires month == m+1 AND Expires year == active year
            cell = ws.cell(row=row, column=col)
            cell.value = (
                f'=IF(\'Licenses\'!F{src}="","",'
                f'IF(AND(MONTH(\'Licenses\'!F{src})={m + 1},'
                f'YEAR(\'Licenses\'!F{src})=Settings!$B$40),'
                f'\'Licenses\'!I{src},""))'
            )
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0;;'  # blank when 0
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Banding
        if i % 2 == 1:
            for col_idx in range(1, 16):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

        ws.row_dimensions[row].height = 18

    # Totals row at last_row + 1
    totals_row = last_row + 1
    ws.cell(row=totals_row, column=1, value="TOTAL")
    cell = ws.cell(row=totals_row, column=1)
    cell.font = Font(name=FONT_MONO, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells(start_row=totals_row, start_column=1,
                    end_row=totals_row, end_column=3)
    for col_idx in range(2, 4):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for m in range(12):
        col = 4 + m
        col_letter = get_column_letter(col)
        cell = ws.cell(row=totals_row, column=col)
        cell.value = f"=SUM({col_letter}7:{col_letter}{last_row})"
        cell.font = Font(name=FONT_BODY, size=11, bold=True,
                          color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0;;'
    ws.row_dimensions[totals_row].height = 24

    ws.freeze_panes = "D7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28), ("C", 4), ("D", 28),
    ])

    compact_header_band(
        ws, "Settings",
        prev_tab="12-Month Calendar", next_tab=None,
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Lists drive dropdowns on the Licenses tab. The active calendar "
        "year drives the 12-Month Calendar — bump it each January."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    def section_label(row, text):
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = Font(name=FONT_MONO, size=9, bold=True,
                          color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
        ws.row_dimensions[row].height = 22

    # B5:B14 — Property list (10 slots)
    section_label(5, "PROPERTIES (B5:B14)")
    # Move label to row 4 visually — re-write at row 5 anyway as header
    # then data B5..B14
    # Actually brief specifies B5-B14 = property data (no header).
    # Let's place header text at row 4 column B? Brief says B5-B14
    # is the property list. Use B5..B14 for the 10 slots, and put
    # section label at column D row 5. Simpler: section labels above the list.
    # Re-do: header in row 4 col B already taken; use rows differently.
    # The brief says: B5-B14 Property list / B16-B25 Jurisdictions /
    # B27-B36 License types / B38 Renewal alert window
    # We'll honor the brief exactly. Section labels go in column D (notes).

    # Reset what we just wrote
    ws.cell(row=5, column=2).value = None
    ws.cell(row=5, column=2).fill = PatternFill(fill_type=None)
    ws.cell(row=5, column=2).font = Font(name=FONT_BODY)
    ws.row_dimensions[5].height = 18

    # Section labels in column D
    def col_d_label(row, text):
        cell = ws.cell(row=row, column=4, value=text)
        cell.font = Font(name=FONT_MONO, size=9, bold=True,
                          color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    col_d_label(5, "← Property list (B5:B14)")
    col_d_label(16, "← Jurisdictions (B16:B25)")
    col_d_label(27, "← License types (B27:B36)")
    col_d_label(38, "← Renewal alert window (days)")
    col_d_label(40, "← Active calendar year (drives 12-Month Calendar)")

    # B5:B14 — Properties
    for i in range(10):
        r = 5 + i
        cell = ws.cell(row=r, column=2)
        if i < len(PROPERTIES_LIST):
            cell.value = PROPERTIES_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # B16:B25 — Jurisdictions
    for i in range(10):
        r = 16 + i
        cell = ws.cell(row=r, column=2)
        if i < len(JURISDICTION_LIST):
            cell.value = JURISDICTION_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # B27:B36 — License types
    for i in range(10):
        r = 27 + i
        cell = ws.cell(row=r, column=2)
        if i < len(LICENSE_TYPE_LIST):
            cell.value = LICENSE_TYPE_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # B38 — Renewal alert window (default 90 days)
    cell = ws.cell(row=38, column=2, value=90)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0"
    ws.row_dimensions[38].height = 22

    # B40 — Active calendar year
    cell = ws.cell(row=40, column=2, value=ACTIVE_YEAR)
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0"
    ws.row_dimensions[40].height = 28

    # Year-end Archive table (per skill: operational tools archive yearly)
    archive_start = 43
    ws.merge_cells(f"B{archive_start}:E{archive_start}")
    c = ws[f"B{archive_start}"]
    c.value = "YEAR-END ARCHIVE — copy KPIs each January"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[archive_start].height = 22

    arch_headers = ["Year", "Active licenses", "Total fees $",
                     "# Renewals filed"]
    for col, h in enumerate(arch_headers, start=2):
        cell = ws.cell(row=archive_start + 1, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[archive_start + 1].height = 22

    for i, yr in enumerate(range(ACTIVE_YEAR, ACTIVE_YEAR + 6)):
        r = archive_start + 2 + i
        cell = ws.cell(row=r, column=2, value=yr)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(3, 6):
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_licenses_tab(wb, variant)
    build_calendar_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "STR License Renewal Calendar — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "STR-focused license renewal calendar — tracks city STR permit, "
        "business license, sales tax, fire/safety, and lottery renewals "
        "with auto countdown + 12-month forward fee calendar."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
