"""Build ACQ-005 AirDNA Data Integrator (v2.2 standard).

Operational/calculator hybrid — paste AirDNA Rentalizer/Market Score CSV
output, layer host-specific quality multipliers, and read out an adjusted
underwriting verdict (revenue, NOI, CoC, DSCR).

Generates:
  templates/_masters/ACQ-005-airdna-data-integrator-DEMO.xlsx
  templates/_masters/ACQ-005-airdna-data-integrator-BLANK.xlsx
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    card_header, card_body_fill,
)

SKU = "ACQ-005"
NAME = "airdna-data-integrator"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (DEMO) — Smokies Ridge example from the brief
# ---------------------------------------------------------------------------

SAMPLE = {
    # Settings
    "active_year": 2026,
    "source_date": date(2026, 4, 15),
    "market_name": "Gatlinburg / Pigeon Forge, TN",
    "conservative_mult": 0.90,

    # AirDNA Annual summary
    "airdna_annual_revenue": 84000,
    "airdna_occupancy": 0.68,
    "airdna_adr": 338,
    "airdna_revpar": 230,
    "airdna_market_score": 87,
    "airdna_comparables": 42,
    "airdna_listing_url": "https://airdna.co/rentalizer/...",

    # AirDNA monthly revenue distribution (12 months, sums to ~84,000)
    # Smokies pattern: peaks in summer (Jun-Aug) and Oct (foliage), trough Jan-Feb
    "airdna_monthly": [
        4200, 4500, 6300, 7000, 7800, 9100,   # Jan-Jun
        9800, 9300, 7500, 8400, 5300, 4800,   # Jul-Dec
    ],

    # Quality multipliers (per brief)
    "qm_photos":     1.10,    # paid pro shoot
    "qm_amenities":  1.15,    # hot tub + view
    "qm_decor":      1.00,
    "qm_reviews":    0.85,    # starting from 0
    "qm_pricing":    1.00,

    # Underwriting box (operating expenses + debt service)
    "annual_opex":      28000,
    "annual_debt":      28800,
    "cash_invested":   165000,
}

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, span="A:L"):
    first, last = span.split(":")
    ws.merge_cells(f"{first}{row}:{last}{row}")
    c = ws[f"{first}{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False, output_col="C"):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    col_idx = column_index_from_string(output_col)
    cell = ws.cell(row=row, column=col_idx, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Navy hero rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "AirDNA Data Integrator"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = 'Translate "AirDNA says $87K" into what your deal will actually do.'
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # "What this does" card rows 10-15
    card_header(ws, 10, ("A", "L"), "What this does")
    body_text = (
        "Paste AirDNA Rentalizer output (annual + monthly revenue, occupancy, "
        "ADR, RevPAR, market score). The Quality Adjuster lets you mark each "
        "dimension above or below market — pro photos, hot tub, decor, review "
        "history, pricing strategy. Output: an adjusted revenue projection and "
        "a mini-underwriting verdict (NOI, CoC, DSCR) tuned to YOUR property, "
        "not the market median."
    )
    ws.merge_cells("A11:L14")
    c = ws["A11"]
    c.value = body_text
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=2)
    card_body_fill(ws, 11, 14, ("A", "L"))
    for r in range(11, 15):
        ws.row_dimensions[r].height = 18

    # "How to use" card rows 17-22
    card_header(ws, 17, ("A", "L"), "How to use — 3 steps")
    steps = [
        "1.  Open AirDNA Rentalizer for your target address. Export CSV (or copy values).",
        "2.  Paste annual + monthly figures into the AirDNA Paste tab.",
        "3.  Mark your property quality on the Quality Adjuster tab — read the verdict.",
    ]
    for i, step in enumerate(steps):
        r = 18 + i
        ws.merge_cells(f"A{r}:L{r}")
        c = ws[f"A{r}"]
        c.value = step
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[r].height = 20
    card_body_fill(ws, 18, 21, ("A", "L"))

    # Primary navigation buttons rows 24-26
    pseudo_button(ws, "A24", "C26", "1. Paste AirDNA",
                  "'AirDNA Paste'!A1", variant="primary")
    pseudo_button(ws, "D24", "F26", "2. Quality Adjuster",
                  "'Quality Adjuster'!A1", variant="primary")
    pseudo_button(ws, "G24", "I26", "3. Adjusted Underwriting",
                  "'Adjusted Underwriting'!A1", variant="accent")
    pseudo_button(ws, "J24", "L26", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    # Anti-pattern callout rows 29-32
    ws.merge_cells("A29:L29")
    c = ws["A29"]
    c.value = "⚠  IMPORTANT — AIRDNA TERMS OF SERVICE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[29].height = 20

    ws.merge_cells("A30:L32")
    c = ws["A30"]
    c.value = (
        "AirDNA is a paid product. This workbook does NOT scrape, automate, "
        "or download their data — that violates their Terms of Service. You "
        "must export or copy the data manually from your authorized AirDNA "
        "account. This template only translates what you've already paid for "
        "into your underwriting model."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=2)
    for r in range(30, 33):
        ws.row_dimensions[r].height = 18

    # Footer
    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_airdna_paste_tab(wb, variant):
    ws = wb.create_sheet("AirDNA Paste")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 32), ("C", 22),
        ("D", 4), ("E", 36),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "AirDNA Paste",
                        prev_tab="Start", next_tab="Quality Adjuster")

    # Source-date callout row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = '="Source data as of: "&TEXT(Settings!B7,"yyyy-mm-dd")&"   ·   Market: "&Settings!B9'
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[4].height = 20

    # ---- Section 1: Annual summary ----
    _section_band(ws, 6, "ANNUAL SUMMARY  (paste from AirDNA Rentalizer)")

    annual_fields = [
        (7,  "Annual revenue ($):",       _val(variant, SAMPLE["airdna_annual_revenue"]), '"$"#,##0',  "AirDNA Rentalizer 'Projected Revenue'"),
        (8,  "Occupancy (%):",            _val(variant, SAMPLE["airdna_occupancy"]),     "0.0%",      "AirDNA Rentalizer 'Projected Occupancy'"),
        (9,  "ADR ($):",                  _val(variant, SAMPLE["airdna_adr"]),           '"$"#,##0',  "AirDNA Rentalizer 'Projected ADR'"),
        (10, "RevPAR ($):",               _val(variant, SAMPLE["airdna_revpar"]),        '"$"#,##0',  "Revenue per available night"),
        (11, "Market Score (0-100):",     _val(variant, SAMPLE["airdna_market_score"]),  "0",         "AirDNA Market Score (higher = stronger)"),
        (12, "Comparables count:",        _val(variant, SAMPLE["airdna_comparables"]),   "0",         "How many comps AirDNA used"),
        (13, "Source listing URL:",       _val(variant, SAMPLE["airdna_listing_url"]),   None,        "Optional — paste AirDNA Rentalizer URL"),
    ]
    for row, label, value, fmt, note in annual_fields:
        _input_row(ws, row, label, value, fmt, note)

    # Sanity check derived: Annual rev = ADR × Occ × 365 (rough)
    _section_band(ws, 15, "SANITY CHECK")
    _output_row(ws, 16, "Implied annual rev (ADR × Occ × 365):",
                "=IFERROR(C9*C8*365,0)", '"$"#,##0')
    _output_row(ws, 17, "Variance vs AirDNA annual:",
                "=IFERROR((C7-C16)/C16,0)", "0.0%")

    # Variance conditional formatting — flag if >10% off
    variance_rule = FormulaRule(
        formula=["ABS(C17)>0.10"],
        font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
    )
    ws.conditional_formatting.add("C17", variance_rule)

    # ---- Section 2: Monthly breakdown ----
    _section_band(ws, 19, "MONTHLY REVENUE BREAKDOWN")

    # Header row 20
    headers = [("B20", "Month"), ("C20", "AirDNA Revenue"), ("E20", "% of Year")]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[20].height = 22

    # Months rows 21-32
    monthly_values = SAMPLE["airdna_monthly"] if variant == "demo" else [None] * 12
    for i, month in enumerate(MONTH_NAMES):
        r = 21 + i
        # Month label
        cell = ws.cell(row=r, column=2, value=month)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Input
        cell = ws.cell(row=r, column=3, value=monthly_values[i])
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        # % of year
        cell = ws.cell(row=r, column=5,
                       value=f"=IFERROR(C{r}/SUM($C$21:$C$32),0)")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.0%"

        # Banding for odd rows — parchment-alt
        if i % 2 == 1:
            band_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            for col in range(2, 6):
                ws.cell(row=r, column=col).fill = band_fill

        ws.row_dimensions[r].height = 17

    # Total row 33
    cell = ws.cell(row=33, column=2, value="Total monthly:")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell = ws.cell(row=33, column=3, value="=SUM(C21:C32)")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.number_format = '"$"#,##0'
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    cell = ws.cell(row=33, column=5,
                   value="=IFERROR((C33-C7)/C7,0)")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0%"

    # Variance flag if monthly total drifts >5% from annual
    var_rule = FormulaRule(
        formula=["ABS(E33)>0.05"],
        font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
    )
    ws.conditional_formatting.add("E33", var_rule)

    ws.row_dimensions[33].height = 22

    # Freeze pane
    ws.freeze_panes = "A21"

    brand_footer(ws, 36,
                 version_line=f"{SKU} · AirDNA Paste")

    ws.print_area = "A1:L38"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:3"


def build_quality_adjuster_tab(wb, variant):
    ws = wb.create_sheet("Quality Adjuster")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 32), ("C", 14), ("D", 14), ("E", 14),
        ("F", 44),
        ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Quality Adjuster",
                        prev_tab="AirDNA Paste",
                        next_tab="Adjusted Underwriting")

    # Intro callout row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "AirDNA reports the median listing in your market. Your property is NOT the median. "
        "Adjust each dimension below — 1.00 = market average, <1.00 = below market, >1.00 = above."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[4].height = 28

    # ---- Multiplier table ----
    _section_band(ws, 6, "QUALITY MULTIPLIERS  (1.00 = market median)")

    # Header row 7
    headers = [
        ("B7", "Dimension"),
        ("C7", "Min"),
        ("D7", "Your Mult"),
        ("E7", "Max"),
        ("F7", "Justification / notes"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 22

    # Data rows 8-12
    multiplier_rows = [
        ("Listing photos quality",   0.85, _val(variant, SAMPLE["qm_photos"]),    1.20,
         _val(variant, "Pro photographer hired ($450 shoot, 30 retouched images)")),
        ("Amenities (hot tub / pool / view)", 0.90, _val(variant, SAMPLE["qm_amenities"]), 1.30,
         _val(variant, "8-person hot tub on covered deck + ridge-line view")),
        ("Decor / staging",          0.85, _val(variant, SAMPLE["qm_decor"]),     1.15,
         _val(variant, "Standard cabin decor — clean, on-trend, not premium")),
        ("Reviews target (10+ at 4.8★)", 0.70, _val(variant, SAMPLE["qm_reviews"]),  1.00,
         _val(variant, "Brand-new listing — no reviews yet, will haircut Y1 occupancy")),
        ("Pricing strategy",         0.95, _val(variant, SAMPLE["qm_pricing"]),   1.10,
         _val(variant, "Default Airbnb Smart Pricing — neither premium nor discounted")),
    ]

    for i, (dim, mn, mult, mx, note) in enumerate(multiplier_rows):
        r = 8 + i
        # Dimension label
        cell = ws.cell(row=r, column=2, value=dim)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Min
        cell = ws.cell(row=r, column=3, value=mn)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # User multiplier (input)
        cell = ws.cell(row=r, column=4, value=mult)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

        # Max
        cell = ws.cell(row=r, column=5, value=mx)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Justification (input)
        cell = ws.cell(row=r, column=6, value=note)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)

        # Banding
        if i % 2 == 1:
            band_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            for col in range(2, 7):
                ws.cell(row=r, column=col).fill = band_fill

        ws.row_dimensions[r].height = 22

        # Conditional formatting on user multiplier — out-of-range = red
        oor_rule = FormulaRule(
            formula=[f"OR(D{r}<C{r},D{r}>E{r})"],
            font=Font(name=FONT_BODY, size=12, bold=True, color=COLOR_ERROR),
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
        )
        ws.conditional_formatting.add(f"D{r}", oor_rule)

    # ---- Total quality multiplier (row 14) ----
    ws.cell(row=14, column=2, value="TOTAL QUALITY MULTIPLIER:").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=14, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=14, column=4, value="=PRODUCT(D8:D12)")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.000"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[14].height = 28

    # Conservative-mode mult
    ws.cell(row=15, column=2, value="× Conservative-mode mult (Settings):").font = Font(
        name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=15, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=15, column=4, value="=Settings!B11")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.000"
    ws.row_dimensions[15].height = 18

    # Effective multiplier
    ws.cell(row=16, column=2, value="EFFECTIVE MULTIPLIER:").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=16, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=16, column=4, value="=D14*D15")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.000"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[16].height = 28

    # ---- Interpretation callout rows 18-22 ----
    _section_band(ws, 18, "INTERPRETATION")
    ws.merge_cells("B19:L22")
    c = ws["B19"]
    c.value = (
        '=IF(D16>=1.10,"Above-market property — your photos, amenities, or pricing strategy '
        'should outperform the AirDNA median. Reasonable to expect "&TEXT(D16-1,"0.0%")&" lift.",'
        'IF(D16>=0.95,"Roughly market — you should land near AirDNA''s median projection.",'
        '"Below-market property — expect "&TEXT(1-D16,"0.0%")&" haircut. Consider whether '
        'reviews ramp + decor refresh can pull you back to neutral by Year 2."))'
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=2)
    for r in range(19, 23):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 25,
                 version_line=f"{SKU} · Quality Adjuster")

    ws.print_area = "A1:L27"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_adjusted_underwriting_tab(wb, variant):
    ws = wb.create_sheet("Adjusted Underwriting")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 18),
        ("D", 18), ("E", 18),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Adjusted Underwriting",
                        prev_tab="Quality Adjuster",
                        next_tab="Settings")

    # Intro callout row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "AirDNA projection × your effective quality multiplier = adjusted underwriting. "
        "Compare AirDNA-as-paid against what you should ACTUALLY underwrite."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[4].height = 22

    # ---- Comparison table ----
    _section_band(ws, 6, "REVENUE / OCCUPANCY / ADR  —  AirDNA vs Adjusted")

    # Header row 7
    headers = [
        ("B7", "Metric"),
        ("C7", "AirDNA"),
        ("D7", "× Mult"),
        ("E7", "Adjusted"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 22

    # Multiplier reference for re-use
    eff_mult = "'Quality Adjuster'!D16"

    # Comparison rows 8-12
    comparison = [
        ("Annual revenue:",     "='AirDNA Paste'!C7",        f"={eff_mult}",      f"=C8*D8",       '"$"#,##0',  True),
        ("Occupancy:",          "='AirDNA Paste'!C8",        f"={eff_mult}",      f"=C9*D9",       "0.0%",      False),
        ("ADR:",                "='AirDNA Paste'!C9",        f"={eff_mult}",      f"=C10*D10",     '"$"#,##0',  False),
        ("RevPAR:",             "='AirDNA Paste'!C10",       f"={eff_mult}",      f"=C11*D11",     '"$"#,##0',  False),
        ("Booked nights:",      "='AirDNA Paste'!C8*365",    f"={eff_mult}",      f"=C12*D12",     "0",         False),
    ]
    for i, (label, airdna_f, mult_f, adj_f, fmt, bold) in enumerate(comparison):
        r = 8 + i
        # Label
        cell = ws.cell(row=r, column=2, value=label)
        cell.font = Font(name=FONT_BODY, size=11, bold=bold, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # AirDNA value
        cell = ws.cell(row=r, column=3, value=airdna_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt

        # Multiplier
        cell = ws.cell(row=r, column=4, value=mult_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.000"

        # Adjusted
        cell = ws.cell(row=r, column=5, value=adj_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        if bold:
            cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        # Banding
        if i % 2 == 1:
            band_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            ws.cell(row=r, column=2).fill = band_fill
            for col in range(3, 5):
                ws.cell(row=r, column=col).fill = band_fill

        ws.row_dimensions[r].height = 20

    # ---- Underwriting box ----
    _section_band(ws, 14, "MINI UNDERWRITING  —  inputs (yellow) + verdict")

    # Inputs col C (rows 15-17)
    underwriting_inputs = [
        (15, "Annual operating expenses ($):",  _val(variant, SAMPLE["annual_opex"]),   '"$"#,##0',  "Property tax + insurance + utilities + cleaning reserve + mgmt + capex"),
        (16, "Annual debt service ($):",        _val(variant, SAMPLE["annual_debt"]),   '"$"#,##0',  "P&I × 12 — pull from your loan calc or ACQ-001"),
        (17, "Total cash invested ($):",        _val(variant, SAMPLE["cash_invested"]), '"$"#,##0',  "Down + closing + rehab + furnishing"),
    ]
    for row, label, value, fmt, note in underwriting_inputs:
        _input_row(ws, row, label, value, fmt, note)

    # Outputs rows 19-22
    _output_row(ws, 19, "Adjusted NOI:",
                "=E8-C15", '"$"#,##0', emphasize=True)
    _output_row(ws, 20, "Annual cash flow:",
                "=E8-C15-C16", '"$"#,##0', emphasize=True)
    _output_row(ws, 21, "Cash-on-cash return:",
                "=IFERROR((E8-C15-C16)/C17,0)", "0.0%")
    _output_row(ws, 22, "DSCR:",
                "=IFERROR((E8-C15)/C16,0)", "0.00")

    # Conditional formatting on key metrics
    coc_good_rule = CellIsRule(
        operator="greaterThan",
        formula=["0.08"],
        font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY),
        fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
    )
    coc_bad_rule = CellIsRule(
        operator="lessThan",
        formula=["0.04"],
        font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
    )
    ws.conditional_formatting.add("C21", coc_good_rule)
    ws.conditional_formatting.add("C21", coc_bad_rule)

    dscr_good_rule = CellIsRule(
        operator="greaterThan",
        formula=["1.25"],
        font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY),
        fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
    )
    dscr_bad_rule = CellIsRule(
        operator="lessThan",
        formula=["1.05"],
        font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
    )
    ws.conditional_formatting.add("C22", dscr_good_rule)
    ws.conditional_formatting.add("C22", dscr_bad_rule)

    # ---- Headline verdict row 24 ----
    ws.merge_cells("A24:L24")
    c = ws["A24"]
    c.value = (
        '=IF(AND(C21>0.08,C22>1.25),"✅  STRONG DEAL  —  AirDNA-adjusted COC + DSCR pass",'
        'IF(OR(C21<0.04,C22<1.05),"❌  PASS  —  even AirDNA can''t save it after adjusters",'
        '"⚠  MARGINAL  —  needs a margin-of-safety sanity check"))'
    )
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[24].height = 38

    # AirDNA delta callout row 26
    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = (
        '="If you underwrote at AirDNA face value, you would have projected $"'
        '&TEXT(C8-C15-C16,"#,##0")&" cash flow. Adjusted to YOUR property: $"'
        '&TEXT(E8-C15-C16,"#,##0")&".   Δ "&TEXT(E8-C8,"+$#,##0;-$#,##0")'
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[26].height = 32

    brand_footer(ws, 29,
                 version_line=f"{SKU} · Adjusted Underwriting")

    ws.print_area = "A1:L31"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:3"


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 26), ("C", 4),
        ("D", 38),
        ("E", 8), ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Adjusted Underwriting",
                        next_tab=None)

    # Intro row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Workbook-wide settings — active year, source date, market context, sensitivity."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[4].height = 18

    # Settings cells
    _section_band(ws, 5, "ACTIVE WORKBOOK", span="A:D")

    # Active year — B5 (per brief)
    # NOTE: per brief, active year sits at B5 in the Settings tab; we leave the
    # section band on row 5 col A:D and tuck the value cell into the same row B
    # by overriding the merge below with a normal label/value pair starting row 6.
    # To stay faithful to the brief's "B5 Active year", we accept the band sits on
    # row 5 cols A-D and place the year on row 5 col B with the band fill but
    # input style on top. The merge_cells call from _section_band already merges
    # A5:D5; unmerge first.
    ws.unmerge_cells("A5:D5")
    # Restore the navy band visually, with year input as B5
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for col in range(1, 5):
        ws.cell(row=5, column=col).fill = navy_fill
    cell = ws.cell(row=5, column=1, value="ACTIVE WORKBOOK")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[5].height = 24

    # Per-row settings (B7, B9, B11 per brief — start at row 5 for active year)
    settings_rows = [
        (5,  "Active year:",                _val(variant, SAMPLE["active_year"]),         "0",              "Drives any date filtering throughout the workbook"),
        (7,  "Source data date:",           _val(variant, SAMPLE["source_date"]),         "yyyy-mm-dd",     "When did you pull AirDNA? (it goes stale fast)"),
        (9,  "AirDNA market name:",         _val(variant, SAMPLE["market_name"]),         None,             "Free text — appears on AirDNA Paste tab header"),
        (11, "Conservative-mode multiplier:", _val(variant, SAMPLE["conservative_mult"]), "0.000",          "Extra haircut applied on top of quality mult (default 0.90 = 10% margin of safety)"),
    ]
    # NOTE: brief says B5 active year, B7 source date, B9 market name, B11 conservative.
    # The B5 cell will live ON TOP of the navy band (we keep the band purely as
    # a visual anchor — the input cell repaints with input_cell_style yellow).
    for row, label, value, fmt, note in settings_rows:
        if row == 5:
            # Special case — input lives in B5 inside the navy band.
            # Place a parchment background under B-D so the input reads cleanly.
            for col in range(2, 5):
                ws.cell(row=row, column=col).fill = PatternFill(
                    "solid", fgColor=COLOR_BG_LIGHT
                )
            # Label in col A is already set above; place input in B5
            cell = ws.cell(row=row, column=2, value=value)
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fmt:
                cell.number_format = fmt
            # Note in col D
            ws.cell(row=row, column=4, value=note).font = Font(
                name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
            )
            ws.cell(row=row, column=4).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            # Override row 5 label color (since we sit on navy band, label A5 stays white)
            continue

        # Standard rows
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        cell = ws.cell(row=row, column=2, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt

        # Note col D
        ws.cell(row=row, column=4, value=note).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[row].height = 20

    # ---- Year-end archive table ----
    _section_band(ws, 14, "YEAR-END ARCHIVE  —  paste adjusted projection vs actuals", span="A:F")

    # Header row 15
    archive_headers = [
        ("A15", "Year"),
        ("B15", "AirDNA Projected"),
        ("C15", "Adjusted Projected"),
        ("D15", "Actual Revenue"),
        ("E15", "Δ vs Adjusted"),
        ("F15", "Notes"),
    ]
    for cell_ref, label in archive_headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[15].height = 22

    # Pre-fill 6 future years (active year + 0..5)
    for i in range(6):
        r = 16 + i
        # Year — formula referencing B5 active year + offset
        cell = ws.cell(row=r, column=1, value=f"=Settings!B5+{i}")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"

        # Empty input cells B-D
        for col in (2, 3, 4):
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0'

        # Δ formula
        cell = ws.cell(row=r, column=5,
                       value=f"=IFERROR((D{r}-C{r})/C{r},\"\")")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.0%"

        # Notes (input)
        cell = ws.cell(row=r, column=6)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)

        if i % 2 == 1:
            band_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = band_fill

        ws.row_dimensions[r].height = 20

    # Year-end ritual note
    ws.merge_cells("A23:F25")
    c = ws["A23"]
    c.value = (
        "Each January, paste last year's actual revenue into column D. The Δ column "
        "shows how accurate your adjusted projection was — refine your quality "
        "multipliers based on what actually happened. Over 2-3 years your underwriting "
        "calibrates to your specific style of operating."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=2)
    for r in range(23, 26):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 28,
                 version_line=f"{SKU} · Settings")

    ws.print_area = "A1:L30"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_airdna_paste_tab(wb, variant)
    build_quality_adjuster_tab(wb, variant)
    build_adjusted_underwriting_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "AirDNA Data Integrator — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Translates AirDNA Rentalizer/Market Score projections into a host-"
        "specific underwriting model with property-quality multipliers."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
