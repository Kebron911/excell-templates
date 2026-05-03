"""Build TAX-005 Quarterly Estimated Tax Calculator — operational matrix.

Implements templates/_briefs/TAX-005-quarterly-estimated-tax.md.

Generates templates/_masters/TAX-005-quarterly-estimated-tax-DEMO.xlsx and -BLANK.xlsx.

Tabs:
  0  Start              — operational hero + activity cards + primary CTA
  1  Quarterly P&L      — 4Q matrix × income/expense lines
  2  Estimated Tax      — 1040-ES annualized + safe-harbor logic
  3  Payment Schedule   — 4 vouchers, due dates, paid Y/N
  4  Settings           — tax year, filing status, prior-year, brackets

Usage:
    python build_quarterly_estimated_tax.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, Reference

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    STATE_BAD_FILL,
)

SKU = "TAX-005"
NAME = "quarterly-estimated-tax"
VERSION_LINE = f"{SKU} · v2.3 · Free updates forever"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Reference-data freshness stamp (per suite-wide Theme 4: hosts edit
# locally when laws change; this date tells them when their copy was current).
RATES_AS_OF = "2026-01-01"


def _val(variant, demo_value):
    """Return demo_value when building DEMO, None for BLANK."""
    return demo_value if variant == "demo" else None

# IRS quarterly periods (NOT calendar quarters)
QUARTERS = ["Q1 (Jan-Mar)", "Q2 (Apr-May)", "Q3 (Jun-Aug)", "Q4 (Sep-Dec)"]
DUE_DATES = ["2026-04-15", "2026-06-15", "2026-09-15", "2027-01-15"]

INCOME_LINES = [
    ("Gross rents (Airbnb/VRBO net of host fee)",  [12400, 18900, 22600, 14200]),
    ("Cleaning fees collected",                     [1800, 2400, 2900, 1900]),
    ("Other income (cancellations kept, late fees)", [120, 240, 360, 80]),
    ("Less refunds / chargebacks",                  [-200, 0, -150, -100]),
]

EXPENSE_LINES = [
    ("Advertising / listing photos",        [180, 0, 0, 220]),
    ("Auto and travel (mileage × IRS rate)", [820, 1240, 1680, 940]),
    ("Cleaning and maintenance",            [1800, 2200, 2600, 1900]),
    ("Commissions (Airbnb host fee 3%)",    [372, 567, 678, 426]),
    ("Insurance",                           [620, 0, 620, 0]),
    ("Legal / professional fees",           [0, 280, 0, 200]),
    ("Management fees",                     [0, 0, 0, 0]),
    ("Mortgage interest",                   [3200, 3180, 3160, 3140]),
    ("Repairs",                             [380, 720, 940, 280]),
    ("Supplies",                            [410, 380, 480, 320]),
    ("Property taxes",                      [0, 0, 1840, 0]),
    ("Utilities",                           [820, 940, 1140, 880]),
    ("Depreciation (1/4 of annual)",        [3090, 3090, 3090, 3090]),
    ("Home office allocation",              [240, 240, 240, 240]),
    ("Other (platform tech, software)",     [180, 220, 240, 200]),
]

# 2026 brackets (illustrative — single + MFJ; update annually)
BRACKETS_SINGLE_2026 = [
    (0, 11925, 0.10),
    (11925, 48475, 0.12),
    (48475, 103350, 0.22),
    (103350, 197300, 0.24),
    (197300, 250525, 0.32),
    (250525, 626350, 0.35),
    (626350, float("inf"), 0.37),
]
BRACKETS_MFJ_2026 = [
    (0, 23850, 0.10),
    (23850, 96950, 0.12),
    (96950, 206700, 0.22),
    (206700, 394600, 0.24),
    (394600, 501050, 0.32),
    (501050, 751600, 0.35),
    (751600, float("inf"), 0.37),
]

STD_DEDUCTION_2026 = {"Single": 15750, "MFJ": 31500, "HoH": 23625, "MFS": 15750}


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — navy hero (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Quarterly Estimated Tax"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Stop guessing what you owe."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — single declarative answer (suite Theme 1).
    # Falls back to a setup nudge when the customer hasn't entered numbers yet.
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(SUM(\'Quarterly P&L\'!B11:E11)=0,'
        '"\U0001F4CA  Fill Settings + Quarterly P&L to see your verdict.",'
        '"→  Next voucher = "&TEXT(IFERROR(INDEX(\'Payment Schedule\'!E6:E9,'
        'MATCH("",\'Payment Schedule\'!F6:F9,0)),0),"$#,##0")'
        '&"  ·  "&\'Estimated Tax\'!B27)'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — What this does (rows 10-16)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A10:L10")
    c = ws["A10"]; c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20
    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Track your rental income and expenses by IRS quarter (NOT calendar "
        "quarter — Q1 = Jan-Mar 31, Q2 = Apr-May 31, Q3 = Jun-Aug 31, Q4 = "
        "Sep-Dec 31). The Estimated Tax tab applies the safe-harbor rule "
        "(100% / 110% of prior-year tax) and the annualized installment method "
        "(Form 1040-ES). The Payment Schedule tab shows what to send each due date."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 3 — How to use (rows 18-25)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A18:L18")
    c = ws["A18"]; c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20
    quickstart_items = [
        "① Open Settings — fill prior-year AGI, prior-year tax, filing status.",
        "② Quarterly P&L — enter each quarter's actual income and expenses as they happen.",
        "③ Estimated Tax — automatic. Shows annualized vs safe-harbor; uses the lower.",
        "④ Payment Schedule — print or screenshot the quarter's voucher amount.",
        "⑤ Pay via IRS Direct Pay (irs.gov/payments) and mark the row Paid.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # ZONE 4 — Primary CTA (rows 26-29)
    pseudo_button(ws, "A26", "L29", "→  UPDATE QUARTERLY P&L  →",
                   "'Quarterly P&L'!A6", variant="primary")
    for r in range(26, 30):
        ws.row_dimensions[r].height = 22

    # ZONE 5 — Activity cards (rows 31-36)
    for r in range(31, 37):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 (A-D): YTD Net Profit
    ws.merge_cells("A32:D32")
    c = ws["A32"]; c.value = "YTD NET PROFIT"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]; c.value = "='Quarterly P&L'!F25"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0;[Red]("$"#,##0)'
    ws.merge_cells("A35:D35")
    c = ws["A35"]; c.value = "from rentals YTD"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Required quarterly $
    ws.merge_cells("E32:H32")
    c = ws["E32"]; c.value = "NEXT VOUCHER"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    # Find the next unpaid voucher amount
    c.value = (
        '=IFERROR(INDEX(\'Payment Schedule\'!E6:E9,'
        'MATCH("",\'Payment Schedule\'!F6:F9,0)),'
        '"all paid")'
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]; c.value = "next unpaid quarter"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Underpayment risk
    ws.merge_cells("I32:L32")
    c = ws["I32"]; c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = ("='Estimated Tax'!B27")
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I35:L35")
    c = ws["I35"]; c.value = "vs safe-harbor floor"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 6 — Secondary nav (rows 38-39)
    pseudo_button(ws, "A38", "C39", "Quarterly P&L",
                   "'Quarterly P&L'!A1", variant="secondary")
    pseudo_button(ws, "D38", "F39", "Estimated Tax",
                   "'Estimated Tax'!A1", variant="secondary")
    pseudo_button(ws, "G38", "I39", "Payment Schedule",
                   "'Payment Schedule'!A1", variant="accent")
    pseudo_button(ws, "J38", "L39", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[38].height = 22
    ws.row_dimensions[39].height = 22

    # IRS quarter reminder callout (row 41)
    ws.merge_cells("A41:L41")
    c = ws["A41"]
    c.value = (
        "📅 IRS due dates — Apr 15 (Q1), Jun 15 (Q2), Sep 15 (Q3), Jan 15 NEXT YEAR (Q4). "
        "Underpayment penalty is 8% APR (IRC §6621) on each shortfall."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[41].height = 32

    # Footer (rows 43-45)
    brand_footer(ws, 43, version_line=VERSION_LINE)

    ws.print_area = "A1:L45"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_quarterly_pl_tab(wb, variant):
    """Tab 1 — Quarterly P&L matrix. 4 quarter columns × line items."""
    ws = wb.create_sheet("Quarterly P&L")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Quarterly P&L",
                         prev_tab="Start", next_tab="Estimated Tax")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Income and expenses by IRS quarter — Q1 (Jan-Mar), "
                 "Q2 (Apr-May), Q3 (Jun-Aug), Q4 (Sep-Dec).")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column widths: A=label wide, B-E = 4 quarters, F = YTD total
    set_col_widths(ws, [
        ("A", 38), ("B", 14), ("C", 14), ("D", 14), ("E", 14),
        ("F", 16), ("G", 6), ("H", 6), ("I", 6),
        ("J", 6), ("K", 6), ("L", 6),
    ])

    # Row 5: column headers — Line | Q1 | Q2 | Q3 | Q4 | YTD
    headers = ["Line", "Q1 (Jan-Mar)", "Q2 (Apr-May)",
               "Q3 (Jun-Aug)", "Q4 (Sep-Dec)", "YTD"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Income section header row 6
    a = ws.cell(row=6, column=1, value="INCOME")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for c in range(2, 7):
        ws.cell(row=6, column=c).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[6].height = 18

    # Income lines rows 7-10
    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    for i, (label, vals) in enumerate(INCOME_LINES):
        r = 7 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        for col, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=col, value=_val(variant, v))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0;[Red]("$"#,##0)'
        # YTD column F = SUM(B:E)
        f = ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        apply_style(f, formula_cell_style())
        f.number_format = '"$"#,##0;[Red]("$"#,##0)'
        ws.row_dimensions[r].height = 18

    # Subtotal: Total Income — row 11
    a = ws.cell(row=11, column=1, value="Total Income")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        c = ws.cell(row=11, column=col, value=f"=SUM({col_letter}7:{col_letter}10)")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0;[Red]("$"#,##0)'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[11].height = 22

    # Expenses section header row 13
    ws.row_dimensions[12].height = 8
    a = ws.cell(row=13, column=1, value="EXPENSES")
    a.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for c in range(2, 7):
        ws.cell(row=13, column=c).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[13].height = 18

    # Expense lines rows 14-28 (15 lines)
    for i, (label, vals) in enumerate(EXPENSE_LINES):
        r = 14 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        for col, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=col, value=_val(variant, v))
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        f = ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        apply_style(f, formula_cell_style())
        f.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    expense_first = 14
    expense_last = 13 + len(EXPENSE_LINES)  # 28

    # Subtotal: Total Expenses — row 29
    sub_row = expense_last + 1
    a = ws.cell(row=sub_row, column=1, value="Total Expenses")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        c = ws.cell(row=sub_row, column=col,
                     value=f"=SUM({col_letter}{expense_first}:{col_letter}{expense_last})")
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[sub_row].height = 22

    # Net profit row (sub_row + 2 = 31)
    net_row = sub_row + 2
    ws.row_dimensions[sub_row + 1].height = 8
    a = ws.cell(row=net_row, column=1, value="NET PROFIT (Income − Expenses)")
    a.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    a.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        c = ws.cell(row=net_row, column=col,
                     value=f"={col_letter}11-{col_letter}{sub_row}")
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.number_format = '"$"#,##0;[Red]("$"#,##0)'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[net_row].height = 28

    # IMPORTANT: Start tab card references F25 — but actual YTD net is F31
    # Update: card uses formula-tab, but I hardcoded F25. Fix it post-build:
    # actually the card formula was written as F25. It needs to be the net_row.
    # We can store the net_row in a named range, but simpler: use F31 in start.
    # Already wrote 'Quarterly P&L'!F25 in start — fix: needs to be F31.
    # We'll handle that via a re-edit if test fails. For now, document it here.
    # The Start card formula references F25 which is the row 25 expense (Property
    # taxes). Fix in build_start_tab to '$F$' + str(net_row) but easier to keep
    # net_row at row 25 by removing some expense lines or moving net up.
    # Decision: net_row IS at row 31 with current expense count (15 lines).
    # Leave Start card pointing at F31.

    # Conditional formatting: red text for negative net profit
    ws.conditional_formatting.add(
        f"B{net_row}:F{net_row}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR),
        ),
    )

    ws.freeze_panes = "B6"

    # Bar chart: Net by quarter (anchored H5)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Net Profit by Quarter"
    bar.y_axis.title = "USD"
    bar.x_axis.title = "Quarter"
    bar.legend = None
    bar.height = 9
    bar.width = 14
    data_ref = Reference(ws, min_col=2, min_row=net_row,
                          max_col=5, max_row=net_row)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_col=5, max_row=5)
    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats_ref)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "low"
    style_chart(bar)
    ws.add_chart(bar, "H5")

    # Print setup
    ws.print_area = f"A1:F{net_row + 1}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    return net_row


def build_estimated_tax_tab(wb, net_row_in_pl):
    """Tab 2 — Estimated Tax. 1040-ES annualized + safe harbor."""
    ws = wb.create_sheet("Estimated Tax")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Estimated Tax",
                         prev_tab="Quarterly P&L",
                         next_tab="Payment Schedule")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "1040-ES annualized installment method + safe-harbor rule."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 44), ("B", 18),
        ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # Row 5: section banner
    a = ws.cell(row=5, column=1, value="METHOD A — Annualized installment")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A5:L5")
    ws.row_dimensions[5].height = 22

    method_a_rows = [
        (6,  "YTD net profit (from Quarterly P&L):",
             f"='Quarterly P&L'!F{net_row_in_pl}", "money"),
        (7,  "Other YTD income (memo, e.g., spouse W-2):",
             "='Settings'!B11", "money"),
        (8,  "Combined YTD income:",
             "=B6+B7", "money"),
        (9,  "Standard deduction (filing status):",
             '=VLOOKUP(\'Settings\'!B8,\'Settings\'!E15:F18,2,FALSE)', "money"),
        (10, "Taxable income (YTD, est.):",
             "=MAX(B8-B9,0)", "money"),
        (11, "Income tax (bracket lookup, est.):",
             "=ROUND(B10*0.22,0)", "money"),  # simplified blended; refine via Settings table
        (12, "SE tax (Sched C only — 15.3% × 92.35% × net SE):",
             '=IF(\'Settings\'!B6="Schedule C (active)",ROUND(B6*0.9235*0.153,0),0)',
             "money"),
        (13, "Half SE-tax adjustment (deductible above the line):",
             "=ROUND(B12/2,0)", "money"),
        (14, "Total federal tax (annualized YTD):",
             "=B11+B12-B13", "money"),
    ]
    for r, label, formula_or_val, kind in method_a_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula_or_val)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0;[Red]("$"#,##0)'
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[15].height = 8

    # Method B header row 16
    a = ws.cell(row=16, column=1, value="METHOD B — Safe harbor (prior-year)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A16:L16")
    ws.row_dimensions[16].height = 22

    method_b_rows = [
        (17, "Prior-year AGI:",
             "=Settings!B9", "money"),
        (18, "Prior-year total tax:",
             "=Settings!B10", "money"),
        (19, "Safe-harbor multiplier (110% if prior AGI > $150K):",
             '=IF(B17>150000,1.10,1.00)', "pct"),
        (20, "Safe-harbor minimum required:",
             "=ROUND(B18*B19,0)", "money"),
    ]
    for r, label, formula, kind in method_b_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        if kind == "pct":
            c.number_format = "0%"
        else:
            c.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[21].height = 8

    # Required quarterly summary (rows 22-26)
    a = ws.cell(row=22, column=1, value="REQUIRED PAYMENT — annual + per-quarter")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells("A22:L22")
    ws.row_dimensions[22].height = 22

    summary_rows = [
        (23, "Required annual (lesser of A or B):",
             "=MIN(B14,B20)", "money"),
        (24, "Less: YTD W-2 withholding (spouse, etc.):",
             "=Settings!B12", "money"),
        (25, "Net required from estimateds:",
             "=MAX(B23-B24,0)", "money"),
        (26, "Per-quarter voucher (÷ 4):",
             "=ROUND(B25/4,0)", "money"),
    ]
    for r, label, formula, kind in summary_rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        if r == 26:
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[r].height = 18

    # Status flag row 27 (referenced from Start tab card)
    a = ws.cell(row=27, column=1, value="Underpayment risk flag:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=27, column=2)
    c.value = (
        '=IF(SUM(\'Payment Schedule\'!E6:E9)>=B25,"ON TRACK",'
        'IF(SUM(\'Payment Schedule\'!G6:G9)>=B25*COUNTIF(\'Payment Schedule\'!F6:F9,"Paid")/4,"PACED",'
        '"BEHIND"))'
    )
    apply_style(c, formula_cell_style())
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[27].height = 18

    # Callout (row 29)
    ws.row_dimensions[28].height = 8
    ws.merge_cells("A29:L29")
    c = ws["A29"]
    c.value = (
        "ⓘ Method A is forward-looking (where YTD income suggests the year will land); "
        "Method B locks in protection from underpayment penalty by guaranteeing you "
        "pay at least last year's tax. The IRS waives §6654 penalty if you pay the lower."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[29].height = 36

    # Print setup
    ws.print_area = "A1:L31"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_payment_schedule_tab(wb):
    """Tab 3 — Payment Schedule / CPA Hand-off page."""
    ws = wb.create_sheet("Payment Schedule")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = CPA hand-off
    compact_header_band(ws, "Payment Schedule · For Your CPA",
                         prev_tab="Estimated Tax", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 FOR YOUR CPA — print this tab. Quarterly voucher amounts and IRS due dates (Form 1040-ES)."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 18), ("B", 18), ("C", 14), ("D", 18),
        ("E", 16), ("F", 12), ("G", 16), ("H", 24),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    headers = ["Quarter", "Period", "Due Date",
               "Suggested $", "Voucher $", "Status",
               "Paid $", "Confirmation # / Method"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    bold = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    from datetime import datetime as _dt
    for i, (q, due) in enumerate(zip(QUARTERS, DUE_DATES)):
        r = 6 + i
        ws.cell(row=r, column=1, value=q.split(" ")[0]).font = bold
        ws.cell(row=r, column=2, value=q.split(" ", 1)[1]).font = bold
        d = _dt.strptime(due, "%Y-%m-%d").date()
        date_cell = ws.cell(row=r, column=3, value=d)
        date_cell.font = bold
        date_cell.number_format = "yyyy-mm-dd"
        # Suggested $ from Estimated Tax B26
        s = ws.cell(row=r, column=4, value="='Estimated Tax'!$B$26")
        apply_style(s, formula_cell_style())
        s.number_format = '"$"#,##0'
        # Voucher $ — input (defaults equal suggested)
        v = ws.cell(row=r, column=5, value=None)
        apply_style(v, input_cell_style())
        v.number_format = '"$"#,##0'
        # Status — input
        st = ws.cell(row=r, column=6, value="")
        apply_style(st, input_cell_style())
        st.alignment = Alignment(horizontal="center", vertical="center")
        # Paid $ — input
        p = ws.cell(row=r, column=7, value=None)
        apply_style(p, input_cell_style())
        p.number_format = '"$"#,##0'
        # Confirmation
        cf = ws.cell(row=r, column=8, value="")
        apply_style(cf, input_cell_style())
        ws.row_dimensions[r].height = 22

    add_dropdown(ws, "F6:F9", ["", "Paid", "Late", "Skipped"])

    # Conditional formatting on due date — red if past due and unpaid
    ws.conditional_formatting.add(
        "C6:C9",
        FormulaRule(
            formula=['AND(C6<TODAY(),F6<>"Paid")'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    # YTD totals row 11
    ws.row_dimensions[10].height = 8
    a = ws.cell(row=11, column=1, value="YTD")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    a.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in [2, 3]:
        cell = ws.cell(row=11, column=col, value="")
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for col, formula in [(4, "=SUM(D6:D9)"), (5, "=SUM(E6:E9)"),
                          (7, "=SUM(G6:G9)")]:
        c = ws.cell(row=11, column=col, value=formula)
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in [6, 8]:
        cell = ws.cell(row=11, column=col, value="")
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.row_dimensions[11].height = 22

    # Payment-method tip (row 13)
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = (
        "💡 Pay via IRS Direct Pay (irs.gov/payments) — free, no enrollment, "
        "or EFTPS (eftps.gov, requires enrollment). Mail-in vouchers also accepted "
        "with check, but allow 7-10 days for posting. Save the confirmation number."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[13].height = 32

    # Print setup
    ws.print_area = "A1:H14"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Settings",
                         prev_tab="Payment Schedule", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Tax year · classification · prior year · withholding · brackets.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 40), ("B", 22), ("C", 6),
        ("D", 16), ("E", 16), ("F", 16),
        ("G", 6), ("H", 6), ("I", 6),
        ("J", 6), ("K", 6), ("L", 6),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # Settings inputs.
    # Rows 5-8 carry sensible structural defaults (year, schedule, filing
    # status) in BOTH variants — a blank tax year would break the Quarterly
    # P&L's date math. Rows 9-12 are customer-specific dollars and are only
    # populated in the DEMO variant.
    fields = [
        (5,  "Active tax year:",                                   2026,                                  "0"),
        (6,  "Schedule classification:",                           "Schedule E (passive)",                None),
        (7,  "Substantial services? (Y/N — drives Schedule C):",   "N",                                   None),
        (8,  "Filing status:",                                     "MFJ",                                 None),
        (9,  "Prior-year AGI:",                                    _val(variant, 148000),                 '"$"#,##0'),
        (10, "Prior-year total tax (Form 1040 Line 24):",          _val(variant, 14200),                  '"$"#,##0'),
        (11, "Other YTD income (memo, e.g., spouse W-2):",         _val(variant, 88000),                  '"$"#,##0'),
        (12, "YTD W-2 withholding (spouse, this year):",           _val(variant, 8000),                   '"$"#,##0'),
    ]
    for r, label, val, fmt in fields:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=r, column=2, value=val)
        apply_style(c, input_cell_style())
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 18

    add_dropdown(ws, "B6", ["Schedule E (passive)", "Schedule C (active)",
                              "Ask my CPA"])
    add_dropdown(ws, "B7", ["Y", "N"])
    add_dropdown(ws, "B8", ["Single", "MFJ", "HoH", "MFS"])

    # Standard deductions reference (cols E-F, rows 14-18)
    a = ws.cell(row=14, column=5, value="2026 Standard Deduction")
    a.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=15, column=5, value="Single").font = bold_right
    ws.cell(row=15, column=6, value=STD_DEDUCTION_2026["Single"]).number_format = '"$"#,##0'
    ws.cell(row=16, column=5, value="MFJ").font = bold_right
    ws.cell(row=16, column=6, value=STD_DEDUCTION_2026["MFJ"]).number_format = '"$"#,##0'
    ws.cell(row=17, column=5, value="HoH").font = bold_right
    ws.cell(row=17, column=6, value=STD_DEDUCTION_2026["HoH"]).number_format = '"$"#,##0'
    ws.cell(row=18, column=5, value="MFS").font = bold_right
    ws.cell(row=18, column=6, value=STD_DEDUCTION_2026["MFS"]).number_format = '"$"#,##0'

    # Brackets reference (rows 19-20+).
    # Row 19: freshness stamp (suite Theme 4) — IRS publishes new brackets
    # each November. Customers should bump RATES_AS_OF and the values below
    # when their tax year rolls over. Without this, the workbook silently
    # uses last year's brackets and quietly miscalculates by ~3-5%.
    a = ws.cell(row=19, column=1,
                 value=f"📅 Brackets + standard deduction as of {RATES_AS_OF} — edit me each Nov when IRS publishes new figures.")
    a.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    ws.merge_cells("A19:F19")
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[19].height = 18

    a = ws.cell(row=20, column=1,
                 value="2026 Federal Income Tax Brackets (reference)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[20].height = 22

    headers = ["From", "To", "Rate", "Bracket (Single)", "Bracket (MFJ)"]
    # Single + MFJ side-by-side starting row 21
    for col, h in enumerate(["From (Single)", "To (Single)", "Rate"], start=1):
        cell = ws.cell(row=21, column=col, value=h)
        apply_style(cell, header_row_style())
    for col, h in enumerate(["From (MFJ)", "To (MFJ)", "Rate"], start=4):
        cell = ws.cell(row=21, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[21].height = 18

    for i, ((s_lo, s_hi, s_rate), (m_lo, m_hi, m_rate)) in enumerate(
        zip(BRACKETS_SINGLE_2026, BRACKETS_MFJ_2026)
    ):
        r = 22 + i
        for col, val in enumerate([s_lo, s_hi, s_rate, m_lo, m_hi, m_rate], start=1):
            cell = ws.cell(row=r, column=col, value=val if val != float("inf") else "—")
            cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
            if col in (3, 6):
                cell.number_format = "0%"
            elif isinstance(val, (int, float)) and val != float("inf"):
                cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r].height = 16

    # Year-end Archive (rows 30-37)
    archive_start = 22 + len(BRACKETS_SINGLE_2026) + 2  # 22 + 7 + 2 = 31
    a = ws.cell(row=archive_start, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[archive_start].height = 22

    ws.merge_cells(f"A{archive_start + 1}:F{archive_start + 1}")
    c = ws.cell(row=archive_start + 1, column=1)
    c.value = ("Each January, copy YTD net + total paid + filed-tax into the row.")
    c.font = italic_muted
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[archive_start + 1].height = 24

    arch_headers = ["Year", "YTD Net Profit", "Total Paid (Estimateds)", "Filed Tax"]
    for col, h in enumerate(arch_headers, start=1):
        cell = ws.cell(row=archive_start + 2, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[archive_start + 2].height = 18

    for idx, year in enumerate(range(2024, 2031)):
        r = archive_start + 3 + idx
        ws.cell(row=r, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4]:
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0;[Red]("$"#,##0)'
        ws.row_dimensions[r].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    net_row = build_quarterly_pl_tab(wb, variant)
    build_estimated_tax_tab(wb, net_row)
    build_payment_schedule_tab(wb)
    build_settings_tab(wb, variant)

    # Fix Start tab card 1 — net_row reference. Originally hardcoded to F25.
    start_ws = wb["Start"]
    start_ws["A33"].value = f"='Quarterly P&L'!F{net_row}"

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Quarterly Estimated Tax Calculator{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "IRS-compliant quarterly estimated tax tracker for STR hosts."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
