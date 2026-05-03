"""Build REV-002 Dynamic Pricing Calculator (v2.2 standard).

Rate-engine tool — given a base rate, season multipliers, day-of-week
multipliers and event lifts, computes a suggested nightly rate for every
night of the year. Operator can override per-night, project monthly
revenue, and paste the result into Airbnb / VRBO custom pricing.

Generates two files:
  templates/_masters/REV-002-dynamic-pricing-calculator-DEMO.xlsx
  templates/_masters/REV-002-dynamic-pricing-calculator-BLANK.xlsx

Sheets (5):
  1. Start             — KPIs + nav
  2. Inputs            — base rate, season/DOW/event multipliers
  3. 365-Day Calendar  — one row per night, formulas drive Suggested rate
  4. Monthly Summary   — per-month avg/min/max + projected revenue + chart
  5. Settings          — property, active year, season name list

Critical rule: never write to a merged cell after merge.
"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "REV-002"
NAME = "dynamic-pricing-calculator"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data — Smokies Ridge Cabin profile, year 2026
# ---------------------------------------------------------------------------

SAMPLE = {
    "property_name": "Smokies Ridge Cabin",
    "active_year":   2026,
    "base_rate":     185,
    "occupancy_pct": 0.62,    # average annual occupancy assumption

    # Seasons: name | start_date | end_date | multiplier
    "seasons": [
        ("Winter",          date(2026, 1, 1),  date(2026, 3, 14),  0.75),
        ("Shoulder Spring", date(2026, 3, 15), date(2026, 5, 31),  1.00),
        ("Peak Summer",     date(2026, 6, 1),  date(2026, 8, 31),  1.45),
        ("Shoulder Fall",   date(2026, 9, 1),  date(2026, 12, 31), 1.05),
    ],

    # Day-of-week mult (Mon=0 ... Sun=6)
    "dow": [
        ("Monday",    1.00),
        ("Tuesday",   1.00),
        ("Wednesday", 1.00),
        ("Thursday",  1.05),
        ("Friday",    1.15),
        ("Saturday",  1.20),
        ("Sunday",    1.00),
    ],

    # Event lifts: date | name | multiplier | notes
    "events": [
        (date(2026, 2, 14),  "Valentine's Day",       1.30, "Romantic getaway demand"),
        (date(2026, 3, 21),  "Spring Break peak",     1.35, "Smokies family draw"),
        (date(2026, 4, 11),  "Smoky Mtn Marathon",    1.45, "Local race weekend"),
        (date(2026, 7, 4),   "July 4 Independence",   1.55, "Fireworks + holiday"),
        (date(2026, 9, 5),   "Labor Day weekend",     1.40, "Last-summer surge"),
        (date(2026, 10, 10), "Octoberfest",           1.45, "Gatlinburg festival"),
        (date(2026, 11, 26), "Thanksgiving",          1.35, "Family travel peak"),
        (date(2026, 12, 24), "Christmas Eve",         1.50, "Holiday week premium"),
    ],
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label, span_to="L"):
    ws.merge_cells(f"A{row}:{span_to}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — Hero band rows 1-9
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 10):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Dynamic Pricing Calculator"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Season, day, event — a suggested rate for every night."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — KPI block rows 11-22
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(11, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!B5="","(set property on Settings tab)",Settings!B5)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = '="PRICING YEAR  "&Settings!B7'
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # KPI strip — 4 KPIs: Avg / Peak / Low / RevPAR
    # Use simple ranges from 365-Day Calendar (column I = Final rate, rows 6..370)
    kpi_specs = [
        ("A15:C16", "AVG NIGHTLY",
         '=IFERROR(AVERAGE(\'365-Day Calendar\'!I6:I370),0)',
         '"$"#,##0'),
        ("D15:F16", "PEAK NIGHTLY",
         '=IFERROR(MAX(\'365-Day Calendar\'!I6:I370),0)',
         '"$"#,##0'),
        ("G15:I16", "LOW NIGHTLY",
         '=IFERROR(MIN(\'365-Day Calendar\'!I6:I370),0)',
         '"$"#,##0'),
        ("J15:L16", "REVPAR PROJECTION",
         '=IFERROR(AVERAGE(\'365-Day Calendar\'!I6:I370)*Inputs!C13,0)',
         '"$"#,##0'),
    ]
    for rng, label, formula, fmt in kpi_specs:
        # label row (top half of merge — so use first 1-row sub-merge for label)
        # We split each KPI into a label row + a value row to avoid the
        # "merged cell" trap. Each KPI occupies two rows (15+16 / 17+18).
        # Strategy: do label on row 15-16 merged, value on row 17-18 merged.
        pass  # placeholder; we render below explicitly per KPI

    # Render KPIs manually (label row + value row, each its own merge)
    kpi_cols = [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]
    kpi_data = [
        ("AVG NIGHTLY",
         '=IFERROR(AVERAGE(\'365-Day Calendar\'!I6:I370),0)', '"$"#,##0'),
        ("PEAK NIGHTLY",
         '=IFERROR(MAX(\'365-Day Calendar\'!I6:I370),0)', '"$"#,##0'),
        ("LOW NIGHTLY",
         '=IFERROR(MIN(\'365-Day Calendar\'!I6:I370),0)', '"$"#,##0'),
        ("REVPAR PROJ",
         '=IFERROR(AVERAGE(\'365-Day Calendar\'!I6:I370)*Inputs!C13,0)',
         '"$"#,##0'),
    ]
    for (first, last), (label, formula, fmt) in zip(kpi_cols, kpi_data):
        # Label on row 15
        ws.merge_cells(f"{first}15:{last}15")
        c = ws[f"{first}15"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Value on rows 16-18
        ws.merge_cells(f"{first}16:{last}18")
        c = ws[f"{first}16"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
    ws.row_dimensions[15].height = 16
    for r in range(16, 19):
        ws.row_dimensions[r].height = 16

    # Helper line about projections
    ws.merge_cells("A20:L20")
    c = ws["A20"]
    c.value = (
        "RevPAR projection = average suggested nightly × occupancy assumption "
        "(set on Inputs)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[20].height = 18

    # Bold caveat row 22
    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = "Re-run when base rate, peak season, or event calendar changes."
    c.font = Font(name=FONT_BODY, size=10, italic=True, bold=True,
                  color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 20

    # ZONE 3 — Pseudo-button nav (rows 25-27)
    pseudo_button(ws, "A25", "C27", "Inputs",
                  "'Inputs'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "365-Day Calendar",
                  "'365-Day Calendar'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Monthly Summary",
                  "'Monthly Summary'!A1", variant="primary")
    pseudo_button(ws, "J25", "L27", "Settings",
                  "'Settings'!A1", variant="accent")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    # Upgrade banner
    cta = (
        "Pricing across the portfolio? Get the Pricing Bundle at "
        f"{BRAND_DOMAIN}/pricing — $97."
    )
    ws.merge_cells("A30:L30")
    c = ws["A30"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[30].height = 36

    brand_footer(ws, 32,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L34"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 32), ("C", 22),
        ("D", 4), ("E", 30),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings", prev_tab="Start", next_tab="Inputs")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Property + active year. The 365-day calendar regenerates from the "
        "year cell — change it once, the whole workbook follows."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # B5 = property name, B7 = active year (per brief)
    ws.cell(row=5, column=1, value="").font = Font(name=FONT_BODY, size=11)
    ws.cell(row=5, column=2, value=_val(variant, SAMPLE["property_name"]))
    apply_style(ws.cell(row=5, column=2), input_cell_style())
    ws.cell(row=5, column=2).font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=5, column=2).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.row_dimensions[5].height = 24
    ws.cell(row=5, column=3, value="← property name").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=5, column=3).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )

    ws.cell(row=7, column=2, value=_val(variant, SAMPLE["active_year"]))
    apply_style(ws.cell(row=7, column=2), input_cell_style())
    ws.cell(row=7, column=2).font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=7, column=2).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.cell(row=7, column=2).number_format = "0"
    ws.row_dimensions[7].height = 24
    ws.cell(row=7, column=3, value="← active year (drives the calendar)").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=7, column=3).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )

    # Season name list B9-B12 (per brief)
    _section_band(ws, 9, "SEASON NAME LIST")
    season_names = [s[0] for s in SAMPLE["seasons"]]
    for i, name in enumerate(season_names):
        r = 10 + i
        ws.cell(row=r, column=2, value=_val(variant, name))
        apply_style(ws.cell(row=r, column=2), input_cell_style())
        ws.cell(row=r, column=2).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.cell(row=r, column=3, value=f"Season {i + 1}").font = Font(
            name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=r, column=3).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[r].height = 18

    # Help note
    ws.merge_cells("A16:L18")
    c = ws["A16"]
    c.value = (
        "Tip: when you roll into a new year, just update the active year "
        "above. The calendar dates rebuild and your season cutoffs (set "
        "on Inputs) follow. Event dates stay literal — re-enter them per "
        "year since holidays move."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(16, 19):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 21, version_line=f"{SKU} · Settings")


def build_inputs_tab(wb, variant):
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 32),
        ("C", 16), ("D", 16), ("E", 16),
        ("F", 14), ("G", 28),
        ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Inputs",
                        prev_tab="Settings", next_tab="365-Day Calendar")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Base rate, four seasons, day-of-week multipliers, and twelve "
        "event slots. Suggested rate = Base × Season × DOW × Event."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- BASE RATE block ----
    _section_band(ws, 6, "BASE RATE  +  OCCUPANCY ASSUMPTION")

    base_fields = [
        (7,  "Base nightly rate ($):",
         _val(variant, SAMPLE["base_rate"]),  '"$"#,##0',
         "The 'normal night' rate before any multiplier"),
        # Note: brief specifies this cell shouldn't conflict with calendar refs.
        # We'll wire the calendar to read C7 for the base.
    ]
    # Skip rows 8-12 — calendar references C7 directly. Slot occupancy at C13.
    for row, label, value, fmt, note in base_fields:
        _input_row(ws, row, label, value, fmt, note)

    # Occupancy assumption (row 13 — referenced by Start tab RevPAR formula)
    _input_row(ws, 13, "Annual occupancy assumption (%):",
               _val(variant, SAMPLE["occupancy_pct"]), "0%",
               "Used for RevPAR projection on Start tab + Monthly Summary")

    # ---- SEASON DEFINITIONS block ----
    # Brief calls for 4 seasons w/ start/end dates + multiplier.
    # We use date cells (not month/day pairs) — cleaner downstream.
    # Layout rows 15-21:
    #   Row 15 — section band
    #   Row 16 — table header (Season | Start date | End date | Multiplier | Notes)
    #   Row 17-20 — 4 season rows
    _section_band(ws, 15, "SEASON DEFINITIONS  (4 seasons, edit dates / mult)")

    season_headers = [
        (2, "Season"),
        (3, "Start date"),
        (4, "End date"),
        (5, "Multiplier"),
        (6, ""),
        (7, "Notes"),
    ]
    for col, label in season_headers:
        cell = ws.cell(row=16, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[16].height = 22

    season_notes = [
        "Slowest months — heat with discounts",
        "Wake-up demand — keep at neutral",
        "Highest demand window of the year",
        "Cooling demand but holds value",
    ]
    for i, (name, start, end, mult) in enumerate(SAMPLE["seasons"]):
        r = 17 + i
        # Col B — name (text, references Settings list optionally — keep editable)
        cell = ws.cell(row=r, column=2, value=_val(variant, name))
        apply_style(cell, input_cell_style())
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col C — start date
        cell = ws.cell(row=r, column=3, value=_val(variant, start))
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "yyyy-mm-dd"

        # Col D — end date
        cell = ws.cell(row=r, column=4, value=_val(variant, end))
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "yyyy-mm-dd"

        # Col E — multiplier
        cell = ws.cell(row=r, column=5, value=_val(variant, mult))
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Cols F-G merged — note
        ws.merge_cells(f"F{r}:G{r}")
        cell = ws[f"F{r}"]
        cell.value = season_notes[i]
        cell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[r].height = 18

    # ---- DAY-OF-WEEK block ----
    # Rows 22-30: band + 7 DOW rows
    _section_band(ws, 22, "DAY-OF-WEEK MULTIPLIERS  (Weekday code matches WEEKDAY(date,2))")

    dow_headers = [
        (2, "Day"),
        (3, "Weekday code"),
        (4, "Multiplier"),
    ]
    for col, label in dow_headers:
        cell = ws.cell(row=23, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[23].height = 22

    for i, (day_name, mult) in enumerate(SAMPLE["dow"]):
        r = 24 + i
        cell = ws.cell(row=r, column=2, value=_val(variant, day_name))
        apply_style(cell, input_cell_style())
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Weekday code: 1=Mon ... 7=Sun (Excel WEEKDAY(date,2))
        cell = ws.cell(row=r, column=3, value=i + 1)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"

        cell = ws.cell(row=r, column=4, value=_val(variant, mult))
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        ws.row_dimensions[r].height = 18

    # ---- EVENT LIFTS block ----
    # Rows 32 onward: band + 12 event slots
    _section_band(ws, 32, "EVENT LIFTS  (12 slots — add holidays, festivals, concerts)")

    event_headers = [
        (2, "Date"),
        (3, "Event name"),
        (4, "Multiplier"),
        (5, ""),
        (6, "Notes"),
    ]
    for col, label in event_headers:
        cell = ws.cell(row=33, column=col, value=label)
        apply_style(cell, header_row_style())
    # Merge label cell over E33:F33 to match data widths below
    # Avoid merging headers — header row is fine cell-by-cell. Keep as-is.
    ws.row_dimensions[33].height = 22

    EVENT_ROWS = 12
    sample_events = SAMPLE["events"]
    for i in range(EVENT_ROWS):
        r = 34 + i
        if variant == "demo" and i < len(sample_events):
            ev_date, ev_name, ev_mult, ev_notes = sample_events[i]
        else:
            ev_date = ev_name = ev_mult = ev_notes = None

        cell = ws.cell(row=r, column=2, value=ev_date)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "yyyy-mm-dd"

        cell = ws.cell(row=r, column=3, value=ev_name)
        apply_style(cell, input_cell_style())
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        cell = ws.cell(row=r, column=4, value=ev_mult)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Cols E-F merged — notes
        ws.merge_cells(f"E{r}:F{r}")
        cell = ws[f"E{r}"]
        cell.value = ev_notes
        cell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[r].height = 18

    # Footer note
    note_row = 34 + EVENT_ROWS + 1  # row 47
    ws.merge_cells(f"A{note_row}:L{note_row + 2}")
    c = ws[f"A{note_row}"]
    c.value = (
        "Note: events override season + DOW. If two events fall on the same "
        "date, the calendar uses the FIRST match in the list — order events "
        "by importance. Blank rows are ignored."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(note_row, note_row + 3):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, note_row + 4, version_line=f"{SKU} · Inputs")


def build_calendar_tab(wb, variant):
    """365-Day Calendar — one row per night, formulas drive Suggested rate.

    Layout:
      Row 1-3: navy header band
      Row 4:   subtitle
      Row 5:   table header
      Row 6 .. 370: 365 night rows

    Columns:
      A Date
      B Day-of-week (text, formula)
      C Season name (lookup)
      D Season multiplier (lookup)
      E DOW multiplier (lookup)
      F Event multiplier (lookup, default 1)
      G Suggested rate  = Base * D * E * F
      H Override (input, optional)
      I Final rate      = IF(H<>"", H, G)
    """
    ws = wb.create_sheet("365-Day Calendar")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 13),  # date
        ("B", 12),  # DOW name
        ("C", 18),  # season name
        ("D", 11),  # season mult
        ("E", 11),  # DOW mult
        ("F", 11),  # event mult
        ("G", 14),  # suggested
        ("H", 13),  # override
        ("I", 14),  # final
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "365-Day Calendar",
                        prev_tab="Inputs", next_tab="Monthly Summary")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per night. Edit the override column if you want to lock a "
        "rate; everything else recomputes from Inputs."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 5
    headers = [
        (1, "Date"),
        (2, "Day"),
        (3, "Season"),
        (4, "Season×"),
        (5, "DOW×"),
        (6, "Event×"),
        (7, "Suggested"),
        (8, "Override"),
        (9, "Final rate"),
    ]
    for col, label in headers:
        cell = ws.cell(row=5, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    # Generate 365 rows (rows 6..370). The calendar is anchored to Settings!B7
    # (active year). For row index i (0..364), date = DATE(Settings!B7,1,1)+i.
    base_rate_ref = "Inputs!$C$7"
    season_start_range = "Inputs!$C$17:$C$20"
    season_end_range = "Inputs!$D$17:$D$20"
    season_name_range = "Inputs!$B$17:$B$20"
    season_mult_range = "Inputs!$E$17:$E$20"
    dow_mult_range = "Inputs!$D$24:$D$30"
    event_date_range = "Inputs!$B$34:$B$45"
    event_mult_range = "Inputs!$D$34:$D$45"

    DAYS = 365
    for i in range(DAYS):
        r = 6 + i

        # Col A — date
        date_formula = f"=DATE(Settings!$B$7,1,1)+{i}"
        cell = ws.cell(row=r, column=1, value=date_formula)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "yyyy-mm-dd"

        # Col B — day-of-week text
        dow_formula = f'=TEXT(A{r},"ddd")'
        cell = ws.cell(row=r, column=2, value=dow_formula)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Col C — season name (find which range A{r} falls into)
        # Use INDEX+MATCH on start dates: largest start date <= A{r}
        # Equivalent: LOOKUP-style. Use INDEX/MATCH with -1 not always available.
        # Simpler: SUMPRODUCT pulling matching index, or nested IFs.
        # We'll use INDEX(MATCH) with array form:
        #   =IFERROR(INDEX(season_name, MATCH(1,(A{r}>=start)*(A{r}<=end),0)),"")
        season_name_formula = (
            f'=IFERROR(INDEX({season_name_range},'
            f'MATCH(1,(A{r}>={season_start_range})*(A{r}<={season_end_range}),0)),"")'
        )
        cell = ws.cell(row=r, column=3, value=season_name_formula)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col D — season multiplier (same lookup, return mult)
        season_mult_formula = (
            f'=IFERROR(INDEX({season_mult_range},'
            f'MATCH(1,(A{r}>={season_start_range})*(A{r}<={season_end_range}),0)),1)'
        )
        cell = ws.cell(row=r, column=4, value=season_mult_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Col E — DOW multiplier (WEEKDAY(date,2) → 1..7 → index into D24:D30)
        dow_mult_formula = f'=INDEX({dow_mult_range},WEEKDAY(A{r},2))'
        cell = ws.cell(row=r, column=5, value=dow_mult_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Col F — event multiplier (lookup A{r} in event date list, default 1)
        event_mult_formula = (
            f'=IFERROR(INDEX({event_mult_range},'
            f'MATCH(A{r},{event_date_range},0)),1)'
        )
        cell = ws.cell(row=r, column=6, value=event_mult_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Col G — suggested rate
        suggested_formula = f"={base_rate_ref}*D{r}*E{r}*F{r}"
        cell = ws.cell(row=r, column=7, value=suggested_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        # Col H — override (blank input)
        cell = ws.cell(row=r, column=8, value=None)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        # Col I — final rate
        final_formula = f'=IF(H{r}<>"",H{r},G{r})'
        cell = ws.cell(row=r, column=9, value=final_formula)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        ws.row_dimensions[r].height = 16

    # Conditional formatting on column I — final rate vs base
    # Green > base * 1.3, Gold base ± 30%, Red < base * 0.85
    last_row = 6 + DAYS - 1
    rng = f"I6:I{last_row}"

    green_fill = PatternFill("solid", fgColor="C8E6C9")
    gold_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    red_fill = PatternFill("solid", fgColor="FFD6D6")

    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f"I6>{base_rate_ref}*1.3"], fill=green_fill)
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f"I6<{base_rate_ref}*0.85"], fill=red_fill)
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[
                f"AND(I6>={base_rate_ref}*0.85,I6<={base_rate_ref}*1.3)"
            ],
            fill=gold_fill,
        )
    )

    # Freeze header
    ws.freeze_panes = "A6"

    # Print setup
    ws.print_area = f"A1:I{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:5"


def build_monthly_summary_tab(wb, variant):
    """Per-month avg/min/max + occupancy assumption + projected revenue + chart."""
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 14),
        ("C", 14), ("D", 14), ("E", 14),
        ("F", 16), ("G", 18),
        ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Monthly Summary",
                        prev_tab="365-Day Calendar", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Twelve months at a glance. Edit the occupancy column to project "
        "revenue per month — these flow into the chart."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 6
    headers = [
        (2, "Month"),
        (3, "Avg rate"),
        (4, "Min rate"),
        (5, "Max rate"),
        (6, "Occupancy %"),
        (7, "Projected revenue"),
    ]
    for col, label in headers:
        cell = ws.cell(row=6, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    # 365-Day Calendar columns: A=date, I=final rate. Rows 6..370.
    cal_dates = "'365-Day Calendar'!$A$6:$A$370"
    cal_finals = "'365-Day Calendar'!$I$6:$I$370"

    for i, month in enumerate(months):
        r = 7 + i
        month_num = i + 1

        # Col B — month name
        cell = ws.cell(row=r, column=2, value=month)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Col C — avg rate (AVERAGEIFS by month)
        avg_formula = (
            f'=IFERROR(AVERAGEIFS({cal_finals},{cal_dates},'
            f'">="&DATE(Settings!$B$7,{month_num},1),'
            f'{cal_dates},"<"&DATE(Settings!$B$7,{month_num + 1},1)),0)'
            if month_num < 12 else
            f'=IFERROR(AVERAGEIFS({cal_finals},{cal_dates},'
            f'">="&DATE(Settings!$B$7,12,1),'
            f'{cal_dates},"<="&DATE(Settings!$B$7,12,31)),0)'
        )
        cell = ws.cell(row=r, column=3, value=avg_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        # Col D — min rate (MINIFS)
        min_formula = (
            f'=IFERROR(MINIFS({cal_finals},{cal_dates},'
            f'">="&DATE(Settings!$B$7,{month_num},1),'
            f'{cal_dates},"<"&DATE(Settings!$B$7,{month_num + 1},1)),0)'
            if month_num < 12 else
            f'=IFERROR(MINIFS({cal_finals},{cal_dates},'
            f'">="&DATE(Settings!$B$7,12,1),'
            f'{cal_dates},"<="&DATE(Settings!$B$7,12,31)),0)'
        )
        cell = ws.cell(row=r, column=4, value=min_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        # Col E — max rate (MAXIFS)
        max_formula = (
            f'=IFERROR(MAXIFS({cal_finals},{cal_dates},'
            f'">="&DATE(Settings!$B$7,{month_num},1),'
            f'{cal_dates},"<"&DATE(Settings!$B$7,{month_num + 1},1)),0)'
            if month_num < 12 else
            f'=IFERROR(MAXIFS({cal_finals},{cal_dates},'
            f'">="&DATE(Settings!$B$7,12,1),'
            f'{cal_dates},"<="&DATE(Settings!$B$7,12,31)),0)'
        )
        cell = ws.cell(row=r, column=5, value=max_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        # Col F — occupancy % (input, default to Inputs!C13)
        occ_value = SAMPLE["occupancy_pct"] if variant == "demo" else None
        cell = ws.cell(row=r, column=6, value=occ_value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0%"

        # Col G — projected revenue: avg rate * days_in_month * occupancy
        # Days in month: =DAY(EOMONTH(DATE(year,month,1),0))
        days_formula = (
            f'DAY(EOMONTH(DATE(Settings!$B$7,{month_num},1),0))'
        )
        proj_formula = f"=C{r}*{days_formula}*F{r}"
        cell = ws.cell(row=r, column=7, value=proj_formula)
        apply_style(cell, formula_cell_style())
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'

        ws.row_dimensions[r].height = 18

    # Total row 19
    total_row = 19
    cell = ws.cell(row=total_row, column=2, value="ANNUAL TOTAL")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Avg of monthly avgs (weighted would be cleaner but this matches month-row table)
    cell = ws.cell(row=total_row, column=3, value=f"=AVERAGE(C7:C18)")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'

    cell = ws.cell(row=total_row, column=4, value=f"=MIN(D7:D18)")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'

    cell = ws.cell(row=total_row, column=5, value=f"=MAX(E7:E18)")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'

    cell = ws.cell(row=total_row, column=6, value=f"=AVERAGE(F7:F18)")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"

    cell = ws.cell(row=total_row, column=7, value=f"=SUM(G7:G18)")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[total_row].height = 26

    # ---- Chart: Avg rate by month ----
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "Avg suggested rate by month"
    chart.y_axis.title = "Avg nightly rate"
    chart.x_axis.title = None
    chart.height = 10
    chart.width = 18

    data = Reference(ws, min_col=3, min_row=6, max_col=3, max_row=18)
    cats = Reference(ws, min_col=2, min_row=7, max_col=2, max_row=18)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.dataLabels = DataLabelList(showVal=False)

    style_chart(chart)
    ws.add_chart(chart, "I6")

    # Insight callout below chart
    callout_row = 22
    ws.merge_cells(f"A{callout_row}:L{callout_row + 2}")
    c = ws[f"A{callout_row}"]
    c.value = (
        "Why month-level matters: occupancy assumptions vary by season — "
        "use the per-month % to get a more realistic revenue projection "
        "than a flat annual rate. Cross-check with PriceLabs or your "
        "booking history."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(callout_row, callout_row + 3):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, callout_row + 4, version_line=f"{SKU} · Monthly Summary")

    ws.print_area = f"A1:L{callout_row + 7}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_inputs_tab(wb, variant)
    build_calendar_tab(wb, variant)
    build_monthly_summary_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Dynamic Pricing Calculator — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Suggested nightly rate for every night of the year — driven by "
        "season, day-of-week, and event multipliers. Includes monthly "
        "summary and revenue projection."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
