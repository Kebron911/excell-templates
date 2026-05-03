"""Build OPS-008 Insurance Policy Tracker (v2.2 standard).

Operational-mode tool for STR insurance policies — STR dwelling, umbrella
liability, auto, equipment, cyber, AirCover backup. Surfaces expiring
policies in the next 60 days, total annual premium, claim history, and
"uninsured exposure" gaps in the Coverage Map.

Generates two files:
  templates/_masters/OPS-008-insurance-policy-tracker-DEMO.xlsx
  templates/_masters/OPS-008-insurance-policy-tracker-BLANK.xlsx
"""
from datetime import datetime, date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    apply_brand_header,
)

SKU = "OPS-008"
NAME = "insurance-policy-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

REFERENCE_AS_OF = date(2026, 5, 2)
ACTIVE_YEAR = 2026

# --- Sample data (DEMO) ---
# 6 policies across 3 properties; one expires in 22 days from REFERENCE_AS_OF.
# Tuple: (policy_no, carrier, ptype, property_, limit, deductible, premium,
#         effective, expiration, agent_name, agent_phone, notes)
POLICY_SAMPLES = [
    ("PROP-2025-44820", "Proper Insurance",
     "STR Dwelling", "Smokies Ridge Cabin",
     650000, 2500, 2840.00,
     date(2025, 9, 15), date(2026, 9, 15),
     "Jenna Holloway", "(865) 555-1108",
     "Vacation rental dwelling + business income; $1M liability rider."),
    ("PROP-2025-44912", "Proper Insurance",
     "STR Dwelling", "Creek Side",
     485000, 2500, 2210.00,
     date(2025, 4, 30), date(2026, 4, 30),
     "Jenna Holloway", "(865) 555-1108",
     "Type 2 STRP property — replacement cost coverage."),
    ("PROP-2025-44980", "Proper Insurance",
     "STR Dwelling", "Lakehouse A",
     # Expires 22 days from REFERENCE_AS_OF (2026-05-02 + 22 = 2026-05-24)
     720000, 5000, 3105.00,
     date(2025, 5, 24), REFERENCE_AS_OF + timedelta(days=22),
     "Marcus DeLeon", "(828) 555-7740",
     "RED ZONE — 22 days. Renewal quote in queue."),
    ("UMB-2026-71103", "Chubb",
     "Umbrella Liability", "Portfolio-wide",
     2000000, 0, 1480.00,
     date(2026, 1, 1), date(2027, 1, 1),
     "Diane Park", "(212) 555-2244",
     "$2M umbrella across all 3 properties — sits over Proper liability."),
    ("AUTO-2026-1144", "Geico",
     "Auto", "Portfolio-wide",
     250000, 1000, 1620.00,
     date(2026, 2, 12), date(2027, 2, 12),
     "Geico (online)", "(800) 555-1492",
     "Personal vehicle used for property visits + supply runs."),
    ("EQP-2025-9012", "Hiscox",
     "Equipment", "Smokies Ridge Cabin",
     35000, 500, 410.00,
     date(2025, 11, 1), date(2026, 11, 1),
     "Hiscox direct", "(866) 555-9988",
     "Hot tub, smart TVs, e-bikes — scheduled inland marine."),
]

# Claim Log sample data
# Tuple: (date, property, policy_no, ptype, description, amount_claimed,
#         deductible_paid, amount_paid, status, notes)
CLAIM_SAMPLES = [
    (date(2025, 11, 18), "Smokies Ridge Cabin", "PROP-2025-44820",
     "STR Dwelling",
     "Burst pipe, water damage to kitchen subfloor",
     8400.00, 2500.00, 5900.00, "Paid",
     "Filed via Proper portal; settled in 14 days."),
    (date(2026, 1, 22), "Creek Side", "PROP-2025-44912",
     "STR Dwelling",
     "Guest accidental TV damage (non-AirCover)",
     1200.00, 0.00, 0.00, "Withdrew",
     "Below deductible — withdrew claim, charged guest deposit."),
    (date(2026, 3, 5), "Lakehouse A", "PROP-2025-44980",
     "STR Dwelling",
     "Wind damage — section of roof shingles",
     6200.00, 5000.00, 1100.00, "Paid",
     "High deductible ate most of the claim."),
]

# Settings tab data
PROPERTIES_LIST = [
    "Smokies Ridge Cabin", "Creek Side", "Lakehouse A",
    "Mountain Loft", "Forest Cabin", "Lakehouse B",
    "Downtown Loft", "Beach House", "Portfolio-wide",
]
POLICY_TYPES = [
    "STR Dwelling", "Umbrella Liability", "Auto",
    "Equipment", "Cyber", "AirCover Backup", "Other",
]
CARRIERS = [
    "Proper Insurance", "Chubb", "Geico", "Hiscox",
    "State Farm", "Allstate", "Progressive", "Liberty Mutual",
    "Farmers", "USAA",
]
CLAIM_STATUS = [
    "Open", "Filed", "Paid", "Denied", "Withdrew",
]

# Coverage map columns — what we expect each property to carry
COVERAGE_COLUMNS = [
    "STR Dwelling", "Umbrella Liability", "Loss of Rents",
    "Auto", "Equipment", "Cyber",
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list, blank_length=None):
    if variant == "demo":
        return demo_list
    n = blank_length if blank_length is not None else len(demo_list)
    return [None] * n


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Insurance Policy Tracker"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Every policy on one page — gaps, renewals, claims."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: KPI cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Active policies
    ws.merge_cells("A10:D10")
    c = ws["A10"]
    c.value = "ACTIVE POLICIES"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A11:D13")
    c = ws["A11"]
    c.value = '=COUNTA(\'Policies\'!A7:A36)'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A14:D15")
    c = ws["A14"]
    c.value = "policies on the register"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 2 (E-H): Premium total
    ws.merge_cells("E10:H10")
    c = ws["E10"]
    c.value = "ANNUAL PREMIUM"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E11:H13")
    c = ws["E11"]
    c.value = '=SUM(\'Policies\'!G7:G36)'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E14:H15")
    c = ws["E14"]
    c.value = "total annual premium"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 3 (I-L): Expiring in next 60 days
    ws.merge_cells("I10:L10")
    c = ws["I10"]
    c.value = "EXPIRING ≤ 60 DAYS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I11:L13")
    c = ws["I11"]
    c.value = (
        '=COUNTIFS(\'Policies\'!J7:J36,">=0",'
        '\'Policies\'!J7:J36,"<=60")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I14:L15")
    c = ws["I14"]
    c.value = "policies due in next 60 days"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 16):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: Renewal-window banner row 17
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = (
        '=IF(COUNTIFS(\'Policies\'!J7:J36,">=0",'
        '\'Policies\'!J7:J36,"<=60")>0,'
        '"RENEWAL WINDOW OPEN — One or more policies expire in the next 60 '
        'days. Quote replacements before lapse; lapsed STR coverage means '
        'no liability protection during a guest stay.",'
        '"ALL CLEAR — No policies in the 60-day renewal window.")'
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[17].height = 36

    ws.conditional_formatting.add(
        "A17:L17",
        FormulaRule(
            formula=[
                'COUNTIFS(\'Policies\'!J7:J36,">=0",'
                '\'Policies\'!J7:J36,"<=60")=0'
            ],
            fill=PatternFill("solid", fgColor="166534"),
        ),
    )

    # ZONE 4: How-to card rows 19-23
    ws.merge_cells("A19:L19")
    c = ws["A19"]
    c.value = "HOW TO USE THIS WORKBOOK"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[19].height = 22

    ws.merge_cells("A20:L23")
    c = ws["A20"]
    c.value = (
        "1.  Log every policy on the Policies tab — STR dwelling, umbrella, "
        "auto, equipment.\n"
        "2.  Walk the Coverage Map to find gaps (a property with no liability "
        "coverage = uninsured exposure).\n"
        "3.  Use Premium Forecast for cash-flow planning around renewal months.\n"
        "4.  When you file a claim, log it on the Claim Log so renewal "
        "discussions have the history."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 5: Pseudo-button nav rows 25-27
    pseudo_button(ws, "A25", "C27", "Policies",
                  "'Policies'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "Coverage Map",
                  "'Coverage Map'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Premium Forecast",
                  "'Premium Forecast'!A1", variant="secondary")
    pseudo_button(ws, "J25", "L27", "Claim Log",
                  "'Claim Log'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "L31", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 18

    # ZONE 6: Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "Need full operations coverage? Add the Operator Bundle at "
        f"{BRAND_DOMAIN}/operator — turnover + maintenance + supply + "
        "damage claims + insurance + permits, $197."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(
        ws, 35,
        version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever",
    )

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Policies
# ---------------------------------------------------------------------------

def build_policies_tab(wb, variant):
    ws = wb.create_sheet("Policies")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(
        ws, "Policies",
        prev_tab="Start", next_tab="Coverage Map",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 15):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:N4")
    c = ws["A4"]
    c.value = (
        "One row per policy. Days-to-expiration auto-calculates; Status flips "
        "to Renew soon (<60d) or Expired (<0d) automatically."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 18), ("B", 18), ("C", 20), ("D", 20),
        ("E", 14), ("F", 12), ("G", 12),
        ("H", 13), ("I", 13), ("J", 10), ("K", 14),
        ("L", 18), ("M", 16), ("N", 30),
    ])

    headers = [
        "Policy #", "Carrier", "Type", "Property",
        "Coverage limit", "Deductible", "Annual premium",
        "Effective date", "Expiration date", "Days", "Status",
        "Agent name", "Agent phone", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    samples = _val_list(variant, POLICY_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 7 + i
        (policy_no, carrier, ptype, property_, limit, deductible,
         premium, effective, expiration, agent_name, agent_phone,
         notes) = row_data

        for col, val in [(1, policy_no), (2, carrier), (3, ptype),
                          (4, property_)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        e = ws.cell(row=row, column=5, value=limit)
        apply_style(e, input_cell_style())
        e.number_format = '"$"#,##0'

        f = ws.cell(row=row, column=6, value=deductible)
        apply_style(f, input_cell_style())
        f.number_format = '"$"#,##0'

        g = ws.cell(row=row, column=7, value=premium)
        apply_style(g, input_cell_style())
        g.number_format = '"$"#,##0.00'

        h = ws.cell(row=row, column=8, value=effective)
        apply_style(h, input_cell_style())
        h.number_format = "yyyy-mm-dd"

        i_cell = ws.cell(row=row, column=9, value=expiration)
        apply_style(i_cell, input_cell_style())
        i_cell.number_format = "yyyy-mm-dd"

        for col, val in [(12, agent_name), (13, agent_phone), (14, notes)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Apply formulas + format to all rows 7-36 (30-policy capacity)
    for row in range(7, 37):
        # Days to expiration col J
        j_cell = ws.cell(
            row=row, column=10,
            value=f'=IF(I{row}<>"",I{row}-TODAY(),"")',
        )
        apply_style(j_cell, formula_cell_style())
        j_cell.alignment = Alignment(horizontal="center", vertical="center")
        j_cell.number_format = "0"

        # Status col K
        k_cell = ws.cell(
            row=row, column=11,
            value=(
                f'=IF(I{row}="","",'
                f'IF(J{row}<0,"Expired",'
                f'IF(J{row}<60,"Renew soon","Active")))'
            ),
        )
        apply_style(k_cell, formula_cell_style())
        k_cell.alignment = Alignment(horizontal="center", vertical="center")
        k_cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)

        # Apply input style to blank rows beyond demo samples
        if row > 6 + len(samples):
            for col in [1, 2, 3, 4, 12, 13, 14]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            for col in [5, 6]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0'
            cell7 = ws.cell(row=row, column=7)
            apply_style(cell7, input_cell_style())
            cell7.number_format = '"$"#,##0.00'
            for col in [8, 9]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            ws.row_dimensions[row].height = 20

    # Conditional formatting on Days col J
    ws.conditional_formatting.add(
        "J7:J36",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="FFB3B3"),
                   font=Font(name=FONT_BODY, size=10, bold=True,
                              color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "J7:J36",
        CellIsRule(operator="between", formula=["0", "29"],
                   fill=PatternFill("solid", fgColor="FFCCCC"),
                   font=Font(name=FONT_BODY, size=10, bold=True,
                              color=COLOR_ERROR)),
    )
    ws.conditional_formatting.add(
        "J7:J36",
        CellIsRule(operator="between", formula=["30", "60"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )

    # Status col K conditional
    ws.conditional_formatting.add(
        "K7:K36",
        FormulaRule(
            formula=['$K7="Expired"'],
            fill=PatternFill("solid", fgColor="B91C1C"),
            font=Font(name=FONT_BODY, size=10, bold=True, color="FFFFFF"),
        ),
    )
    ws.conditional_formatting.add(
        "K7:K36",
        FormulaRule(
            formula=['$K7="Renew soon"'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
            font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY),
        ),
    )

    # Dropdowns
    add_dropdown(ws, "B7:B36", "=Settings!$D$6:$D$25")
    add_dropdown(ws, "C7:C36", "=Settings!$C$6:$C$15")
    add_dropdown(ws, "D7:D36", "=Settings!$B$6:$B$25")

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Coverage Map
# ---------------------------------------------------------------------------

def build_coverage_map_tab(wb, variant):
    ws = wb.create_sheet("Coverage Map")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(
        ws, "Coverage Map",
        prev_tab="Policies", next_tab="Premium Forecast",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-property coverage matrix. Each cell shows the limit "
        "if the property carries that coverage type, or GAP if there's "
        "no policy in force. Walk the GAPs before renewals."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 28

    # Property column + 6 coverage type cols + total exposure col
    col_widths = [("A", 22)]
    for i, _ in enumerate(COVERAGE_COLUMNS):
        col_widths.append((get_column_letter(2 + i), 16))
    col_widths.append((get_column_letter(2 + len(COVERAGE_COLUMNS)), 16))
    set_col_widths(ws, col_widths)

    # Header row 6
    headers = ["Property"] + COVERAGE_COLUMNS + ["Total limit"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[5].height = 6

    # Property rows 7-16 (10 capacity).
    # Pull property names from Settings, then for each (property, coverage)
    # cell, look up the policy with that property+type and return its limit
    # via SUMIFS on Policies!E (limit).
    for i in range(10):
        row = 7 + i
        # Property name pulled from Settings
        prop_cell = ws.cell(
            row=row, column=1,
            value=f'=IFERROR(IF(Settings!B{6+i}="","",Settings!B{6+i}),"")',
        )
        apply_style(prop_cell, formula_cell_style())
        prop_cell.alignment = Alignment(horizontal="left", vertical="center",
                                          indent=1)
        prop_cell.font = Font(name=FONT_BODY, size=11, bold=True,
                               color=COLOR_PRIMARY)

        # Per-coverage-type limit lookups
        for j, ctype in enumerate(COVERAGE_COLUMNS):
            col = 2 + j
            # SUMIFS on Policies!E (limit) where property=A{row} and type=ctype
            # Coverage map should also count Portfolio-wide policies that
            # apply to all properties (umbrella, auto). For simplicity:
            # SUMIFS by property+type, OR if portfolio-wide policy of that
            # type exists, take that.
            formula = (
                f'=IF(A{row}="","",'
                f'IF(SUMIFS(Policies!G7:G36,'
                f'Policies!D7:D36,A{row},'
                f'Policies!C7:C36,"{ctype}")'
                f'+SUMIFS(Policies!G7:G36,'
                f'Policies!D7:D36,"Portfolio-wide",'
                f'Policies!C7:C36,"{ctype}")=0,'
                f'"GAP",'
                f'SUMIFS(Policies!E7:E36,'
                f'Policies!D7:D36,A{row},'
                f'Policies!C7:C36,"{ctype}")'
                f'+SUMIFS(Policies!E7:E36,'
                f'Policies!D7:D36,"Portfolio-wide",'
                f'Policies!C7:C36,"{ctype}")))'
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="right", vertical="center",
                                        indent=1)
            cell.number_format = '"$"#,##0;;"GAP"'

        # Total limit col (last)
        last_col = 2 + len(COVERAGE_COLUMNS)
        first_letter = get_column_letter(2)
        last_letter = get_column_letter(last_col - 1)
        total_formula = (
            f'=IF(A{row}="","",'
            f'SUMIFS(Policies!E7:E36,'
            f'Policies!D7:D36,A{row}))'
        )
        # Note: we use raw SUMIFS so GAP cells don't break the total.
        cell = ws.cell(row=row, column=last_col, value=total_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="right", vertical="center",
                                    indent=1)
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        ws.row_dimensions[row].height = 22

    # Row striping
    for i in range(10):
        if i % 2 == 1:
            row = 7 + i
            for col in range(1, 2 + len(COVERAGE_COLUMNS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    # Highlight GAP cells in red
    last_letter = get_column_letter(1 + len(COVERAGE_COLUMNS))
    ws.conditional_formatting.add(
        f"B7:{last_letter}16",
        FormulaRule(
            formula=[f'B7="GAP"'],
            fill=PatternFill("solid", fgColor="FFB3B3"),
            font=Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR),
        ),
    )

    # Total exposure footer row 18
    total_col = 2 + len(COVERAGE_COLUMNS)
    total_letter = get_column_letter(total_col)
    ws.cell(row=18, column=1, value="Portfolio total")
    ws.cell(row=18, column=1).font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=18, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    ws.cell(row=18, column=1).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    total_cell = ws.cell(
        row=18, column=total_col,
        value=f'=SUM({total_letter}7:{total_letter}16)',
    )
    total_cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    total_cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    total_cell.alignment = Alignment(horizontal="right", vertical="center",
                                       indent=1)
    total_cell.number_format = '"$"#,##0'
    # Fill middle cells of footer
    for col in range(2, total_col):
        ws.cell(row=18, column=col).fill = PatternFill(
            "solid", fgColor=COLOR_GOLD_SOFT
        )
    ws.row_dimensions[18].height = 28

    # GAP count callout row 20
    ws.merge_cells(f"A20:{total_letter}20")
    c = ws[f"A20"]
    c.value = (
        '=IF(COUNTIF(B7:G16,"GAP")>0,'
        '"COVERAGE GAPS: "&COUNTIF(B7:G16,"GAP")'
        '&" — review the red cells above. Each gap is uninsured exposure.",'
        '"NO COVERAGE GAPS — every property carries coverage in every column.")'
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[20].height = 28
    ws.conditional_formatting.add(
        f"A20:{total_letter}20",
        FormulaRule(
            formula=['COUNTIF(B7:G16,"GAP")=0'],
            fill=PatternFill("solid", fgColor="166534"),
        ),
    )

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Premium Forecast
# ---------------------------------------------------------------------------

def build_premium_forecast_tab(wb, variant):
    ws = wb.create_sheet("Premium Forecast")
    ws.sheet_properties.tabColor = COLOR_NAVY_TINT

    compact_header_band(
        ws, "Premium Forecast",
        prev_tab="Coverage Map", next_tab="Claim Log",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "24-month renewal calendar. Premium falls in the month each policy's "
        "expiration date hits. Use the totals row for cash-flow planning."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 28

    # Active year pulled from Settings — establishes the start of the 24-month
    # window. Forecast starts at January of active year.
    # Header layout:
    #   Col A: Policy # (width 18)
    #   Col B: Carrier  (width 14)
    #   Col C: Property (width 18)
    #   Cols D+: 24 monthly columns labeled YYYY-MM
    #   Last col: Row total
    set_col_widths(ws, [
        ("A", 18), ("B", 14), ("C", 18),
    ])
    for i in range(24):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 9
    total_col_letter = get_column_letter(4 + 24)
    ws.column_dimensions[total_col_letter].width = 12

    # Build month labels via formulas referencing Settings active year
    base_headers = ["Policy #", "Carrier", "Property"]
    for col, h in enumerate(base_headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())

    # Month header row 6 cols D..AA — formula = DATE(active_year, month, 1)
    for i in range(24):
        col = 4 + i
        # Settings!B5 holds active year. Use month offset 1..24.
        month_offset = i + 1  # 1..24
        formula = (
            f'=TEXT(DATE(Settings!$B$5,{month_offset},1),"yyyy-mm")'
        )
        cell = ws.cell(row=6, column=col, value=formula)
        apply_style(cell, header_row_style())
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    text_rotation=90)

    # Total col header
    total_col = 4 + 24
    cell = ws.cell(row=6, column=total_col, value="Row total")
    apply_style(cell, header_row_style())

    ws.row_dimensions[6].height = 56
    ws.row_dimensions[5].height = 6

    # Data rows 7-36 — 30 policies max
    for i in range(30):
        row = 7 + i
        src = 7 + i  # Policies tab is also 7-indexed

        # Policy # (col A)
        a_cell = ws.cell(row=row, column=1,
                          value=f'=IF(Policies!A{src}="","",Policies!A{src})')
        apply_style(a_cell, formula_cell_style())
        a_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
        a_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)

        # Carrier (col B)
        b_cell = ws.cell(row=row, column=2,
                          value=f'=IF(Policies!B{src}="","",Policies!B{src})')
        apply_style(b_cell, formula_cell_style())
        b_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
        b_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)

        # Property (col C)
        c_cell = ws.cell(row=row, column=3,
                          value=f'=IF(Policies!D{src}="","",Policies!D{src})')
        apply_style(c_cell, formula_cell_style())
        c_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
        c_cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)

        # Monthly cells: premium if expiration falls in that month
        for j in range(24):
            col = 4 + j
            month_offset = j + 1
            # Compare YEAR(exp)*12+MONTH(exp) == active_year*12 + month_offset
            formula = (
                f'=IF(Policies!I{src}="","",'
                f'IF((YEAR(Policies!I{src})*12+MONTH(Policies!I{src}))'
                f'=(Settings!$B$5*12+{month_offset}),'
                f'Policies!G{src},""))'
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '"$"#,##0;;""'
            cell.font = Font(name=FONT_BODY, size=9, color=COLOR_TEXT)

        # Row total
        first_month_letter = get_column_letter(4)
        last_month_letter = get_column_letter(4 + 23)
        rt_cell = ws.cell(
            row=row, column=total_col,
            value=f'=IF(SUM({first_month_letter}{row}:{last_month_letter}{row})=0,'
                  f'"",SUM({first_month_letter}{row}:{last_month_letter}{row}))',
        )
        apply_style(rt_cell, formula_cell_style())
        rt_cell.alignment = Alignment(horizontal="right", vertical="center",
                                        indent=1)
        rt_cell.number_format = '"$"#,##0;;""'
        rt_cell.font = Font(name=FONT_BODY, size=10, bold=True,
                             color=COLOR_PRIMARY)

        # Row striping
        if i % 2 == 1:
            for col in range(1, total_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.fill.fgColor.rgb in (None, "00000000",
                                              "00EDEDED", "FFEDEDED"):
                    cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        ws.row_dimensions[row].height = 18

    # Totals row (row 38)
    totals_row = 38
    ws.cell(row=totals_row, column=1, value="Monthly total")
    ws.cell(row=totals_row, column=1).font = Font(
        name=FONT_HEAD, size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=totals_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.cell(row=totals_row, column=1).fill = PatternFill(
        "solid", fgColor=COLOR_PRIMARY
    )
    # Empty B and C with same fill
    for col in [2, 3]:
        ws.cell(row=totals_row, column=col).fill = PatternFill(
            "solid", fgColor=COLOR_PRIMARY
        )

    for j in range(24):
        col = 4 + j
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=totals_row, column=col,
            value=f'=IF(SUM({col_letter}7:{col_letter}36)=0,"",'
                  f'SUM({col_letter}7:{col_letter}36))',
        )
        cell.font = Font(name=FONT_BODY, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0;;""'

    # Grand total
    gt_letter = get_column_letter(total_col)
    gt_cell = ws.cell(
        row=totals_row, column=total_col,
        value=f'=SUM({gt_letter}7:{gt_letter}36)',
    )
    gt_cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    gt_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    gt_cell.alignment = Alignment(horizontal="right", vertical="center",
                                    indent=1)
    gt_cell.number_format = '"$"#,##0'
    ws.row_dimensions[totals_row].height = 26

    ws.freeze_panes = "D7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Claim Log
# ---------------------------------------------------------------------------

def build_claim_log_tab(wb, variant):
    ws = wb.create_sheet("Claim Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(
        ws, "Claim Log",
        prev_tab="Premium Forecast", next_tab="Settings",
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Every claim filed across the portfolio. Carriers ask about claim "
        "history at every renewal — keep this log accurate and ready."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 22), ("C", 18), ("D", 18),
        ("E", 36), ("F", 14), ("G", 14), ("H", 14),
        ("I", 12), ("J", 32),
    ])

    headers = [
        "Date", "Property", "Policy #", "Type",
        "Description", "$ Claimed", "$ Deductible", "$ Paid",
        "Status", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[5].height = 6

    samples = _val_list(variant, CLAIM_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 7 + i
        (date_filed, prop, policy_no, ptype, desc, claimed,
         deductible, paid, status, notes) = row_data

        a = ws.cell(row=row, column=1, value=date_filed)
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        for col, val in [(2, prop), (3, policy_no), (4, ptype), (5, desc)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        for col, val in [(6, claimed), (7, deductible), (8, paid)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0.00'

        for col, val in [(9, status), (10, notes)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Apply blank-input formatting to remaining rows (50 capacity)
    for row in range(7, 57):
        if row > 6 + len(samples):
            cell = ws.cell(row=row, column=1)
            apply_style(cell, input_cell_style())
            cell.number_format = "yyyy-mm-dd"
            for col in [2, 3, 4, 5, 9, 10]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            for col in [6, 7, 8]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0.00'
            ws.row_dimensions[row].height = 18

    # Dropdowns
    add_dropdown(ws, "B7:B56", "=Settings!$B$6:$B$25")
    add_dropdown(ws, "D7:D56", "=Settings!$C$6:$C$15")
    add_dropdown(ws, "I7:I56", "=Settings!$E$6:$E$15")

    ws.freeze_panes = "A7"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28),
        ("C", 22), ("D", 22), ("E", 16), ("F", 18),
    ])

    compact_header_band(
        ws, "Settings",
        prev_tab="Claim Log", next_tab=None,
    )

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Lists drive the dropdowns on Policies, Coverage Map, and Claim Log. "
        "Active year drives the Premium Forecast horizon."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Active year cell B5 — referenced by Premium Forecast
    ws.cell(row=5, column=1, value="").alignment = Alignment(horizontal="left")
    label = ws.cell(row=5, column=1)
    label.value = ""
    # Instead, use a clearer label arrangement:
    # Row 5 col B = active year (input). Helper labels on row 4 area already.
    ay_cell = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(ay_cell, input_cell_style())
    ay_cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    ay_cell.alignment = Alignment(horizontal="center", vertical="center")
    ay_cell.number_format = "0"

    # Label for active year
    label_cell = ws.cell(row=5, column=3, value="← Active year (drives forecast)")
    label_cell.font = Font(name=FONT_BODY, size=10, italic=True,
                            color=COLOR_MUTED)
    label_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
    label_cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Renewal alert window cell B40
    ws.cell(row=40, column=1).value = ""
    aw_cell = ws.cell(row=40, column=2, value=60)
    apply_style(aw_cell, input_cell_style())
    aw_cell.alignment = Alignment(horizontal="center", vertical="center")
    aw_cell.number_format = "0"
    aw_label = ws.cell(row=40, column=3,
                        value="← Renewal alert window (days)")
    aw_label.font = Font(name=FONT_BODY, size=10, italic=True,
                          color=COLOR_MUTED)
    aw_label.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    aw_label.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Properties col B header at row -- but we want B6:B25 as the list per
    # cross-tab dropdown formula references. Place a tiny header at row above.
    # Headers row strip — but row 5 is already the active year cell, so place
    # column headers at the section start, below the active year:
    ws.merge_cells("B7:F7")
    hdr = ws["B7"]
    hdr.value = "DROPDOWN LISTS — edit these to match your portfolio"
    hdr.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    hdr.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 22

    # Wait — dropdown formulas reference $B$6:$B$25, $C$6:$C$15, etc.
    # The "active year" sits at B5, the list-section starts at row 6.
    # We need to keep the list region as B6:B25 etc. and not collide.
    # Rework: active year stays at B5 (already placed), the strip header
    # will be placed at row 4 area. Adjust: simply leave the strip header
    # off and place per-column labels inline at row 6 as the first list
    # row. Keep cleanest approach: shift labels above to row 6 as visual
    # heading row, and lists from row 7 onward. But our dropdowns reference
    # B6:B25 — to avoid that mismatch, drop the merged header row and
    # tag each column at row 6 inside the cell (the first cell of each
    # list).
    # Simpler: remove the merged "DROPDOWN LISTS" strip and inline each
    # column header at row 6. But actually our dropdowns reference
    # $B$6:$B$25 which expects data starting at B6. We will place column
    # subtitles at row 5/6 boundaries differently: keep a tag row at row 6
    # populated with the FIRST list value, and use row 5 for active year,
    # row 4 for description. That matches license-permit-tracker where
    # row 5 had column header strips and lists started row 6.
    # Net: undo the row 7 merged header — better to restructure cleanly.
    # Remove the prior merge.
    ws.unmerge_cells("B7:F7")
    ws["B7"].value = None
    ws["B7"].fill = PatternFill(fill_type=None)
    ws["B7"].font = Font(name=FONT_BODY)

    # Column header strip at row 5 cols C-F + row 6 first data row
    # Active year remains at B5 (input). Lists span B6:B25, C6:C15, etc.
    # We'll add tiny mono labels at row 4 above each column using col cells
    # already within the description merge — instead, place column tags at
    # the bottom so the data region B6:F25 is clean.
    # Simplest pragmatic answer: write each column's first row at row 6 as
    # a real data value (matches dropdown range), and accept that the
    # "label row" lives in the strip header at row 6 itself by rendering
    # headers in row 5 cells C5..F5 with the active year still in B5.
    for col, label in [(3, "Policy types"), (4, "Carriers"),
                        (5, "Claim status")]:
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    # Properties label sits above its list at row 5 col B?  But B5 is the
    # active year. We'll add a small note at A5 with rotation OR sacrifice
    # and place "Properties" label at A6 spanning none; cleanest: place
    # the "Properties" label at row 6 col A as a single-cell tag.
    # Actually, simpler still: keep B5 the active year and use B6 as the
    # first property entry — and put a small label tag on the active-year
    # cell itself indicating "Active Year" inline. We've already done that
    # via the C5 inline label. Now add a "Properties" header inside the
    # list region by placing a non-data label row above. To avoid
    # confusion with dropdown ranges (B6:B25), we keep the labels OUT of
    # those ranges. Add a tiny strip at row 4 that says "Properties /
    # Policy types / Carriers / Claim status" — but row 4 already has
    # the description merge. Compromise: the C5/D5/E5 labels above suffice
    # to identify columns; for the Properties column, we add a header in
    # row 5 col B... but B5 is the active year. Trade-off — make B5 hold
    # both the year and a note. We'll leave the active year in B5 and put
    # the Properties label inside cell A6 as a vertical label.

    a6 = ws.cell(row=6, column=1, value="Props →")
    a6.font = Font(name=FONT_MONO, size=8, bold=True, color=COLOR_MUTED)
    a6.alignment = Alignment(horizontal="right", vertical="center")
    a6.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    ws.row_dimensions[5].height = 26

    # Properties col B rows 6-25 (20 capacity)
    for i in range(20):
        r = 6 + i
        cell = ws.cell(row=r, column=2)
        if i < len(PROPERTIES_LIST):
            cell.value = PROPERTIES_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Policy types col C rows 6-15 (10 capacity)
    for i in range(10):
        r = 6 + i
        cell = ws.cell(row=r, column=3)
        if i < len(POLICY_TYPES):
            cell.value = POLICY_TYPES[i]
        apply_style(cell, input_cell_style())

    # Carriers col D rows 6-25 (20 capacity)
    for i in range(20):
        r = 6 + i
        cell = ws.cell(row=r, column=4)
        if i < len(CARRIERS):
            cell.value = CARRIERS[i]
        apply_style(cell, input_cell_style())

    # Claim status col E rows 6-15
    for i in range(10):
        r = 6 + i
        cell = ws.cell(row=r, column=5)
        if i < len(CLAIM_STATUS):
            cell.value = CLAIM_STATUS[i]
        apply_style(cell, input_cell_style())

    # Year-end archive table — section header at row 28
    ws.merge_cells("B28:F28")
    c = ws["B28"]
    c.value = "YEAR-END ARCHIVE — fill in January, then reset the workbook"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[28].height = 22

    archive_headers = ["Year", "# Policies", "Total premium $",
                        "Claims filed", "Claims paid $"]
    for col_offset, h in enumerate(archive_headers):
        cell = ws.cell(row=29, column=2 + col_offset, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[29].height = 22

    # Pre-fill 5 future years (2026-2030)
    for i in range(5):
        r = 30 + i
        year_cell = ws.cell(row=r, column=2, value=ACTIVE_YEAR + i)
        apply_style(year_cell, input_cell_style())
        year_cell.font = Font(name=FONT_HEAD, size=11, bold=True,
                                color=COLOR_PRIMARY)
        year_cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in [3, 5]:
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0"
        for col in [4, 6]:
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 20

    # As-of stamp row 38
    ws.cell(row=38, column=2,
             value="Carrier rates verified:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    cell = ws.cell(row=38, column=3, value=REFERENCE_AS_OF)
    apply_style(cell, input_cell_style())
    cell.number_format = "yyyy-mm-dd"
    cell.font = Font(name=FONT_BODY, size=11, color=COLOR_PRIMARY)
    ws.row_dimensions[38].height = 24

    # Note
    ws.merge_cells("B39:L39")
    c = ws["B39"]
    c.value = (
        "Update this date when you re-verify carrier quotes. STR-friendly "
        "carriers (Proper, Hiscox) repricing happens annually."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=1)
    ws.row_dimensions[39].height = 30


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_policies_tab(wb, variant)
    build_coverage_map_tab(wb, variant)
    build_premium_forecast_tab(wb, variant)
    build_claim_log_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Insurance Policy Tracker — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "STR insurance policy tracker — auto-calculated days-to-expiration, "
        "coverage gap matrix, 24-month premium forecast, and claim log."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
