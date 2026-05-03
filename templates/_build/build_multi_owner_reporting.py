"""Build PAM-004 Multi-Owner Consolidated Reporting (v2.2 standard).

Operational-mode tool for Pro Pam managing multiple owner relationships.
Aggregates per-owner reporting (PAM-001's job) one level up — Pam's total
managed revenue, total commission earned, time allocated per owner,
profitability per owner relationship. Tells Pam which owner relationships
are worth keeping vs firing.

Generates two files:
  templates/_masters/PAM-004-multi-owner-consolidated-reporting-DEMO.xlsx
  templates/_masters/PAM-004-multi-owner-consolidated-reporting-BLANK.xlsx
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, LineChart, Reference

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "PAM-004"
NAME = "multi-owner-consolidated-reporting"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026

# ---------------------------------------------------------------------------
# Sample data — DEMO variant
# Brief: 4 owners, 12 properties; commission $65.5K, NOI $48K (74% margin)
# ---------------------------------------------------------------------------

# Owners — capacity 15 (rows 6-20)
OWNERS_DEMO = [
    # (id, name, email, phone, agreement_start, agreement_type,
    #  commission_pct_avg, ytd_commission_input, nps, status, notes)
    ("OWN-01", "Lakeside Holdings LLC",   "owner1@lakeside.example",
     "(865) 555-0114", "2024-01-01", "% of revenue", 0.20, 32000, 9,
     "Active", "5 properties — flagship account."),
    ("OWN-02", "Maddox Trust",            "owner2@maddox.example",
     "(865) 555-0122", "2025-06-01", "% of revenue", 0.15, 18000, 8,
     "Active", "3 properties — growing portfolio."),
    ("OWN-03", "Highland Properties LLC", "owner3@highland.example",
     "(865) 555-0138", "2025-08-01", "Hybrid (base + %)", 0.10, 11000, 7,
     "Active", "2 properties — consider commission bump at renewal."),
    ("OWN-04", "Brookline Estates",       "owner4@brookline.example",
     "(865) 555-0151", "2025-04-01", "% of revenue", 0.12, 4500, 5,
     "Churning", "2 properties — high-touch, low margin. Fire candidate."),
]

# Properties — capacity 30 (rows 6-35 on Settings)
PROPERTIES_DEMO = [
    # (property, owner)
    ("Smokies Ridge Cabin",   "Lakeside Holdings LLC"),
    ("Creek Side",            "Lakeside Holdings LLC"),
    ("Lakehouse A",           "Lakeside Holdings LLC"),
    ("Pine Vale Retreat",     "Lakeside Holdings LLC"),
    ("Cedar Lakehouse",       "Lakeside Holdings LLC"),
    ("Mountain Loft",         "Maddox Trust"),
    ("Forest Cabin",          "Maddox Trust"),
    ("Riverbend Cottage",     "Maddox Trust"),
    ("Highland Hideaway",     "Highland Properties LLC"),
    ("Smoky View Lodge",      "Highland Properties LLC"),
    ("Brookline Bungalow",    "Brookline Estates"),
    ("Birch Hollow",          "Brookline Estates"),
]

# Pam P&L — 12 months × line items
# Commission revenue is summed from owners ($65.5K). Bonus + expenses below.
PNL_INCOME_DEMO = {
    # row label -> 12 monthly amounts (Jan..Dec)
    "Commission revenue (per-owner sum)": [
        4200, 4400, 5800, 6100, 6400, 6200, 6800, 6500, 5800, 5200, 4400, 3700,
    ],
    "Bonus / one-off income": [
        0, 0, 250, 0, 0, 500, 0, 0, 0, 350, 0, 0,
    ],
}
# Sum: 65500 + 1100 = 66600 (~brief target $65.5K commission, near)
PNL_EXPENSE_DEMO = {
    "Software (PMS, accounting, comms)": [
        420, 420, 420, 420, 420, 420, 420, 420, 420, 420, 420, 420,
    ],
    "Insurance (E&O / liability)": [
        180, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180,
    ],
    "Contractor payments (non-pass-through)": [
        220, 180, 240, 260, 320, 280, 300, 290, 250, 220, 180, 160,
    ],
    "Marketing": [
        120, 120, 240, 180, 180, 180, 240, 180, 120, 120, 80, 80,
    ],
    "Office": [
        140, 140, 140, 140, 140, 140, 140, 140, 140, 140, 140, 140,
    ],
    "Mileage (deductible)": [
        85, 78, 142, 168, 195, 188, 210, 192, 168, 142, 96, 88,
    ],
    "Meals": [
        38, 42, 64, 78, 82, 78, 88, 78, 64, 52, 38, 32,
    ],
    "Other": [
        25, 25, 50, 25, 25, 75, 25, 25, 25, 50, 25, 25,
    ],
}

# Time Allocation — 300+ row capacity (1000)
ACTIVITIES = [
    "Listing mgmt", "Guest comms", "Vendor coord",
    "Reporting", "Maintenance", "Other",
]

TIME_ALLOC_DEMO = [
    # (date, owner, property, activity, hours, notes)
    # Owner A — Lakeside (180 hrs target across 5 properties)
    ("2026-01-04", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Listing mgmt", 2.5, "Q1 listing refresh"),
    ("2026-01-08", "Lakeside Holdings LLC",   "Creek Side",          "Guest comms",  3.0, "Pre-arrival batch"),
    ("2026-01-15", "Lakeside Holdings LLC",   "Lakehouse A",         "Vendor coord", 1.8, "HVAC scheduling"),
    ("2026-01-22", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Reporting",    2.2, "Owner statement prep"),
    ("2026-01-29", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Maintenance",  1.5, "Walk-through"),
    ("2026-02-03", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Guest comms",  2.8, ""),
    ("2026-02-10", "Lakeside Holdings LLC",   "Creek Side",          "Listing mgmt", 1.5, "Photo update"),
    ("2026-02-17", "Lakeside Holdings LLC",   "Lakehouse A",         "Reporting",    2.0, ""),
    ("2026-02-24", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Vendor coord", 1.2, "Cleaner onboarding"),
    ("2026-03-04", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Guest comms",  3.5, "High-volume month"),
    ("2026-03-11", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Maintenance",  2.0, "Plumbing follow-up"),
    ("2026-03-18", "Lakeside Holdings LLC",   "Creek Side",          "Reporting",    2.5, "Q1 close"),
    ("2026-03-25", "Lakeside Holdings LLC",   "Lakehouse A",         "Listing mgmt", 1.8, ""),
    ("2026-04-02", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Guest comms",  2.4, ""),
    ("2026-04-09", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Vendor coord", 1.5, ""),
    ("2026-04-16", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Reporting",    2.2, ""),
    ("2026-04-23", "Lakeside Holdings LLC",   "Creek Side",          "Maintenance",  1.8, ""),
    ("2026-05-01", "Lakeside Holdings LLC",   "Lakehouse A",         "Guest comms",  3.2, "Memorial wknd surge"),
    ("2026-05-08", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Listing mgmt", 2.0, ""),
    ("2026-05-15", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Vendor coord", 1.5, ""),
    ("2026-05-22", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Reporting",    2.0, ""),
    ("2026-05-29", "Lakeside Holdings LLC",   "Creek Side",          "Maintenance",  2.5, ""),
    ("2026-06-05", "Lakeside Holdings LLC",   "Lakehouse A",         "Guest comms",  3.0, ""),
    ("2026-06-12", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Listing mgmt", 1.8, ""),
    ("2026-06-19", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Reporting",    2.2, ""),
    ("2026-06-26", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Vendor coord", 1.5, ""),
    ("2026-07-03", "Lakeside Holdings LLC",   "Creek Side",          "Guest comms",  3.5, "July 4 surge"),
    ("2026-07-10", "Lakeside Holdings LLC",   "Lakehouse A",         "Maintenance",  2.0, ""),
    ("2026-07-17", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Reporting",    2.0, ""),
    ("2026-07-24", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Listing mgmt", 1.5, ""),
    ("2026-07-31", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Guest comms",  2.5, ""),
    ("2026-08-07", "Lakeside Holdings LLC",   "Creek Side",          "Vendor coord", 1.8, ""),
    ("2026-08-14", "Lakeside Holdings LLC",   "Lakehouse A",         "Reporting",    2.0, ""),
    ("2026-08-21", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Maintenance",  2.2, ""),
    ("2026-08-28", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Guest comms",  2.5, ""),
    ("2026-09-04", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Listing mgmt", 1.5, "Q4 prep"),
    ("2026-09-11", "Lakeside Holdings LLC",   "Creek Side",          "Reporting",    2.0, ""),
    ("2026-09-18", "Lakeside Holdings LLC",   "Lakehouse A",         "Vendor coord", 1.8, ""),
    ("2026-09-25", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Guest comms",  2.2, ""),
    ("2026-10-02", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Maintenance",  2.0, ""),
    ("2026-10-09", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Reporting",    2.5, ""),
    ("2026-10-16", "Lakeside Holdings LLC",   "Creek Side",          "Listing mgmt", 1.5, ""),
    ("2026-10-23", "Lakeside Holdings LLC",   "Lakehouse A",         "Guest comms",  2.0, ""),
    ("2026-10-30", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Vendor coord", 1.5, ""),
    ("2026-11-06", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Reporting",    2.0, ""),
    ("2026-11-13", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Maintenance",  1.8, ""),
    ("2026-11-20", "Lakeside Holdings LLC",   "Creek Side",          "Guest comms",  2.0, ""),
    ("2026-11-27", "Lakeside Holdings LLC",   "Lakehouse A",         "Listing mgmt", 1.5, ""),
    ("2026-12-04", "Lakeside Holdings LLC",   "Pine Vale Retreat",   "Reporting",    2.5, "Year-end"),
    ("2026-12-11", "Lakeside Holdings LLC",   "Cedar Lakehouse",     "Guest comms",  2.0, ""),
    ("2026-12-18", "Lakeside Holdings LLC",   "Smokies Ridge Cabin", "Vendor coord", 1.5, ""),
    ("2026-12-22", "Lakeside Holdings LLC",   "Creek Side",          "Maintenance",  2.0, "Holiday turnover"),
    # Owner B — Maddox (120 hrs target across 3 properties)
    ("2026-01-06", "Maddox Trust",            "Mountain Loft",       "Listing mgmt", 2.0, ""),
    ("2026-01-13", "Maddox Trust",            "Forest Cabin",        "Guest comms",  2.5, ""),
    ("2026-01-27", "Maddox Trust",            "Riverbend Cottage",   "Vendor coord", 1.5, ""),
    ("2026-02-05", "Maddox Trust",            "Mountain Loft",       "Reporting",    2.0, ""),
    ("2026-02-12", "Maddox Trust",            "Forest Cabin",        "Guest comms",  2.8, ""),
    ("2026-02-19", "Maddox Trust",            "Riverbend Cottage",   "Maintenance",  1.8, ""),
    ("2026-03-06", "Maddox Trust",            "Mountain Loft",       "Listing mgmt", 1.5, ""),
    ("2026-03-13", "Maddox Trust",            "Forest Cabin",        "Reporting",    2.2, ""),
    ("2026-03-20", "Maddox Trust",            "Riverbend Cottage",   "Vendor coord", 1.2, ""),
    ("2026-04-03", "Maddox Trust",            "Mountain Loft",       "Guest comms",  2.5, ""),
    ("2026-04-10", "Maddox Trust",            "Forest Cabin",        "Maintenance",  2.0, ""),
    ("2026-04-17", "Maddox Trust",            "Riverbend Cottage",   "Listing mgmt", 1.8, ""),
    ("2026-05-02", "Maddox Trust",            "Mountain Loft",       "Reporting",    2.0, ""),
    ("2026-05-09", "Maddox Trust",            "Forest Cabin",        "Guest comms",  3.0, ""),
    ("2026-05-23", "Maddox Trust",            "Riverbend Cottage",   "Vendor coord", 1.5, ""),
    ("2026-06-06", "Maddox Trust",            "Mountain Loft",       "Maintenance",  2.0, ""),
    ("2026-06-13", "Maddox Trust",            "Forest Cabin",        "Listing mgmt", 1.5, ""),
    ("2026-06-20", "Maddox Trust",            "Riverbend Cottage",   "Reporting",    2.2, ""),
    ("2026-07-04", "Maddox Trust",            "Mountain Loft",       "Guest comms",  3.0, ""),
    ("2026-07-11", "Maddox Trust",            "Forest Cabin",        "Vendor coord", 1.5, ""),
    ("2026-07-25", "Maddox Trust",            "Riverbend Cottage",   "Maintenance",  2.0, ""),
    ("2026-08-08", "Maddox Trust",            "Mountain Loft",       "Reporting",    2.0, ""),
    ("2026-08-15", "Maddox Trust",            "Forest Cabin",        "Guest comms",  2.5, ""),
    ("2026-08-22", "Maddox Trust",            "Riverbend Cottage",   "Listing mgmt", 1.5, ""),
    ("2026-09-05", "Maddox Trust",            "Mountain Loft",       "Vendor coord", 1.5, ""),
    ("2026-09-12", "Maddox Trust",            "Forest Cabin",        "Reporting",    2.0, ""),
    ("2026-09-26", "Maddox Trust",            "Riverbend Cottage",   "Guest comms",  2.2, ""),
    ("2026-10-10", "Maddox Trust",            "Mountain Loft",       "Maintenance",  2.0, ""),
    ("2026-10-17", "Maddox Trust",            "Forest Cabin",        "Listing mgmt", 1.5, ""),
    ("2026-10-31", "Maddox Trust",            "Riverbend Cottage",   "Reporting",    2.0, ""),
    ("2026-11-07", "Maddox Trust",            "Mountain Loft",       "Guest comms",  2.0, ""),
    ("2026-11-14", "Maddox Trust",            "Forest Cabin",        "Vendor coord", 1.2, ""),
    ("2026-11-28", "Maddox Trust",            "Riverbend Cottage",   "Maintenance",  1.8, ""),
    ("2026-12-05", "Maddox Trust",            "Mountain Loft",       "Reporting",    2.5, ""),
    ("2026-12-12", "Maddox Trust",            "Forest Cabin",        "Guest comms",  2.0, ""),
    ("2026-12-19", "Maddox Trust",            "Riverbend Cottage",   "Listing mgmt", 1.5, ""),
    # Owner C — Highland (95 hrs across 2 properties)
    ("2026-01-09", "Highland Properties LLC", "Highland Hideaway",   "Listing mgmt", 1.8, ""),
    ("2026-01-23", "Highland Properties LLC", "Smoky View Lodge",    "Guest comms",  2.0, ""),
    ("2026-02-06", "Highland Properties LLC", "Highland Hideaway",   "Vendor coord", 1.5, ""),
    ("2026-02-20", "Highland Properties LLC", "Smoky View Lodge",    "Reporting",    2.0, ""),
    ("2026-03-07", "Highland Properties LLC", "Highland Hideaway",   "Maintenance",  1.5, ""),
    ("2026-03-21", "Highland Properties LLC", "Smoky View Lodge",    "Guest comms",  2.2, ""),
    ("2026-04-04", "Highland Properties LLC", "Highland Hideaway",   "Listing mgmt", 1.5, ""),
    ("2026-04-18", "Highland Properties LLC", "Smoky View Lodge",    "Reporting",    2.0, ""),
    ("2026-05-02", "Highland Properties LLC", "Highland Hideaway",   "Vendor coord", 1.2, ""),
    ("2026-05-16", "Highland Properties LLC", "Smoky View Lodge",    "Guest comms",  2.5, ""),
    ("2026-06-06", "Highland Properties LLC", "Highland Hideaway",   "Maintenance",  1.8, ""),
    ("2026-06-20", "Highland Properties LLC", "Smoky View Lodge",    "Listing mgmt", 1.5, ""),
    ("2026-07-04", "Highland Properties LLC", "Highland Hideaway",   "Guest comms",  2.5, ""),
    ("2026-07-18", "Highland Properties LLC", "Smoky View Lodge",    "Reporting",    2.0, ""),
    ("2026-08-01", "Highland Properties LLC", "Highland Hideaway",   "Vendor coord", 1.5, ""),
    ("2026-08-15", "Highland Properties LLC", "Smoky View Lodge",    "Maintenance",  1.8, ""),
    ("2026-09-05", "Highland Properties LLC", "Highland Hideaway",   "Listing mgmt", 1.5, ""),
    ("2026-09-19", "Highland Properties LLC", "Smoky View Lodge",    "Guest comms",  2.0, ""),
    ("2026-10-03", "Highland Properties LLC", "Highland Hideaway",   "Reporting",    2.0, ""),
    ("2026-10-17", "Highland Properties LLC", "Smoky View Lodge",    "Vendor coord", 1.2, ""),
    ("2026-11-07", "Highland Properties LLC", "Highland Hideaway",   "Maintenance",  1.5, ""),
    ("2026-11-21", "Highland Properties LLC", "Smoky View Lodge",    "Listing mgmt", 1.5, ""),
    ("2026-12-05", "Highland Properties LLC", "Highland Hideaway",   "Guest comms",  2.0, ""),
    ("2026-12-19", "Highland Properties LLC", "Smoky View Lodge",    "Reporting",    2.5, "Year-end"),
    # Owner D — Brookline (80 hrs across 2 properties — high-touch low-margin)
    ("2026-01-11", "Brookline Estates",       "Brookline Bungalow",  "Guest comms",  3.0, "Difficult guests"),
    ("2026-01-25", "Brookline Estates",       "Birch Hollow",        "Vendor coord", 2.5, "Repeat HVAC issues"),
    ("2026-02-08", "Brookline Estates",       "Brookline Bungalow",  "Maintenance",  2.8, "Plumbing"),
    ("2026-02-22", "Brookline Estates",       "Birch Hollow",        "Guest comms",  3.2, "Multiple complaints"),
    ("2026-03-08", "Brookline Estates",       "Brookline Bungalow",  "Vendor coord", 2.0, ""),
    ("2026-03-22", "Brookline Estates",       "Birch Hollow",        "Reporting",    2.5, "Owner disputed"),
    ("2026-04-05", "Brookline Estates",       "Brookline Bungalow",  "Maintenance",  2.5, "Pest control"),
    ("2026-04-19", "Brookline Estates",       "Birch Hollow",        "Guest comms",  2.8, ""),
    ("2026-05-03", "Brookline Estates",       "Brookline Bungalow",  "Listing mgmt", 1.8, ""),
    ("2026-05-17", "Brookline Estates",       "Birch Hollow",        "Vendor coord", 2.2, ""),
    ("2026-06-07", "Brookline Estates",       "Brookline Bungalow",  "Maintenance",  3.0, "AC failure"),
    ("2026-06-21", "Brookline Estates",       "Birch Hollow",        "Guest comms",  2.5, ""),
    ("2026-07-05", "Brookline Estates",       "Brookline Bungalow",  "Reporting",    2.0, ""),
    ("2026-07-19", "Brookline Estates",       "Birch Hollow",        "Vendor coord", 1.8, ""),
    ("2026-08-02", "Brookline Estates",       "Brookline Bungalow",  "Guest comms",  2.5, ""),
    ("2026-08-16", "Brookline Estates",       "Birch Hollow",        "Maintenance",  2.2, ""),
    ("2026-09-06", "Brookline Estates",       "Brookline Bungalow",  "Listing mgmt", 1.5, ""),
    ("2026-09-20", "Brookline Estates",       "Birch Hollow",        "Reporting",    2.0, ""),
    ("2026-10-04", "Brookline Estates",       "Brookline Bungalow",  "Vendor coord", 1.8, ""),
    ("2026-10-18", "Brookline Estates",       "Birch Hollow",        "Guest comms",  2.0, ""),
    ("2026-11-08", "Brookline Estates",       "Brookline Bungalow",  "Maintenance",  2.0, ""),
    ("2026-11-22", "Brookline Estates",       "Birch Hollow",        "Listing mgmt", 1.5, ""),
    ("2026-12-06", "Brookline Estates",       "Brookline Bungalow",  "Reporting",    2.0, ""),
    ("2026-12-20", "Brookline Estates",       "Birch Hollow",        "Guest comms",  2.5, "Renewal mtg"),
]

AGREEMENT_TYPES = ["Flat $", "% of revenue", "Hybrid (base + %)", "Custom"]
STATUSES = ["Active", "Churning", "Onboarding", "Sunset"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val_list(variant, demo_list):
    return demo_list if variant == "demo" else []


def _parse_date(s):
    if not s:
        return None
    return date.fromisoformat(s)


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, max_col=12):
    last = get_column_letter(max_col)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Multi-Owner Consolidated Reporting"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Which owner relationships are worth keeping. Which ones to fire."
    c.font = Font(name=FONT_HEAD, size=13, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    cards = [
        ("A", "C", "MANAGED PROPERTIES",
         '=COUNTA(\'Settings\'!$B$23:$B$52)',
         "0", COLOR_PRIMARY, "in active portfolio"),
        ("D", "F", "TOTAL COMMISSION YTD",
         '=SUM(\'Owners\'!$I$6:$I$20)',
         '"$"#,##0', COLOR_ACCENT, "across all owners"),
        ("G", "I", "TOP OWNER",
         '=IFERROR(INDEX(\'Owners\'!$B$6:$B$20,MATCH(MAX(\'Owners\'!$I$6:$I$20),\'Owners\'!$I$6:$I$20,0)),"—")',
         "@", COLOR_PRIMARY, "highest commission YTD"),
        ("J", "L", "LOWEST NPS",
         '=IFERROR(INDEX(\'Owners\'!$B$6:$B$20,MATCH(MIN(\'Owners\'!$J$6:$J$20),\'Owners\'!$J$6:$J$20,0)),"—")',
         "@", COLOR_SECONDARY, "weakest relationship"),
    ]
    for fc, lc, label, formula, fmt, color, sub in cards:
        ws.merge_cells(f"{fc}10:{lc}10")
        c = ws[f"{fc}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=8, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{fc}11:{lc}13")
        c = ws[f"{fc}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=20, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.number_format = fmt
        ws.merge_cells(f"{fc}14:{lc}15")
        c = ws[f"{fc}14"]
        c.value = sub
        c.font = Font(name=FONT_BODY, size=8, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for fc, lc, *_ in cards:
        first_c = column_index_from_string(fc)
        last_c = column_index_from_string(lc)
        for r in range(10, 16):
            for col in range(first_c, last_c + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == first_c else existing.left,
                    right=gold_side if col == last_c else existing.right,
                )

    # Workflow callout: row 17 banner only — column hints go on row 18+
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "PRO PAM WORKFLOW"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    workflow_steps = [
        "(1)  Owners — log every owner you manage with NPS + agreement terms.",
        "(2)  Time Allocation — track every hour you spend, tagged to owner + property.",
        "(3)  Pam P&L — your business income (commission) less your expenses.",
        "(4)  Per-Owner Dashboard — see commission, hours, $/hr per owner.",
        "(5)  Owner Profitability — ranked. Bottom-3 in red. Decide who to fire.",
    ]
    for i, step in enumerate(workflow_steps):
        row = 18 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = step
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22

    # Pseudo-button nav (rows 25-27 row 1, 29-31 row 2)
    pseudo_button(ws, "A25", "C27", "Owners",
                  "'Owners'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "Per-Owner Dashboard",
                  "'Per-Owner Dashboard'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Pam P&L",
                  "'Pam P&L'!A1", variant="primary")
    pseudo_button(ws, "J25", "L27", "Time Allocation",
                  "'Time Allocation'!A1", variant="primary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "F31", "→  Owner Profitability",
                  "'Owner Profitability'!A1", variant="accent")
    pseudo_button(ws, "G29", "L31", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "Manage cleaners + commissions + multi-owner reporting in one bundle? "
        f"Get the Pro Manager Bundle at {BRAND_DOMAIN}/pro-manager — saves $1,000+ vs individual SKUs."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Owners (capacity 15)
# ---------------------------------------------------------------------------

def build_owners_tab(wb, variant):
    ws = wb.create_sheet("Owners")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Owners",
                        prev_tab="Start", next_tab="Per-Owner Dashboard")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One row per owner relationship. Capacity 15. Active year on Settings (B5)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = [
        "Owner ID", "Owner name", "Email", "Phone",
        "Properties (count)", "Agreement start", "Agreement type",
        "Commission % avg", "Total commission YTD", "NPS (1-10)",
        "Status", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 12), ("B", 26), ("C", 26), ("D", 16),
        ("E", 14), ("F", 14), ("G", 18), ("H", 14),
        ("I", 18), ("J", 12), ("K", 14), ("L", 30),
    ])

    samples = _val_list(variant, OWNERS_DEMO)
    for i in range(15):
        row = 6 + i
        if i < len(samples):
            (oid, name, email, phone, agr_start, agr_type, comm_pct,
             ytd_comm, nps, status, notes) = samples[i]
        else:
            oid = name = email = phone = agr_start = agr_type = None
            comm_pct = ytd_comm = nps = status = notes = None

        for col, val, fmt in [
            (1, oid, None),
            (2, name, None),
            (3, email, None),
            (4, phone, None),
            (6, _parse_date(agr_start) if agr_start else None, "yyyy-mm-dd"),
            (7, agr_type, None),
            (8, comm_pct, "0%"),
            (9, ytd_comm, '"$"#,##0'),
            (10, nps, "0"),
            (11, status, None),
            (12, notes, None),
        ]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt

        # Properties count — formula from Settings property list
        prop_cell = ws.cell(
            row=row, column=5,
            value=f'=IF(B{row}="","",COUNTIF(\'Settings\'!$C$23:$C$52,B{row}))',
        )
        apply_style(prop_cell, formula_cell_style())
        prop_cell.number_format = "0"

        ws.row_dimensions[row].height = 18

    add_dropdown(ws, "G6:G20", f'"{",".join(AGREEMENT_TYPES)}"')
    add_dropdown(ws, "K6:K20", f'"{",".join(STATUSES)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Per-Owner Dashboard
# ---------------------------------------------------------------------------

def build_per_owner_dashboard_tab(wb, variant):
    ws = wb.create_sheet("Per-Owner Dashboard")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Per-Owner Dashboard",
                        prev_tab="Owners", next_tab="Pam P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "One section per owner. Active year drives YTD figures (Settings!B5)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 4), ("B", 22), ("C", 16), ("D", 16),
        ("E", 16), ("F", 16), ("G", 16), ("H", 16),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    # 4 owners × ~7 rows per section = ~28 rows
    # Section starts at row 6
    section_start = 6
    for i in range(4):
        # Owner reference comes from Owners!B6, B7, ...
        owner_row = 6 + i
        block_start = section_start + i * 8

        # Banner row — DO NOT write column hints here; they go below
        ws.merge_cells(f"A{block_start}:H{block_start}")
        c = ws[f"A{block_start}"]
        c.value = f'=IFERROR("OWNER " & {i + 1} & ": " & \'Owners\'!B{owner_row}, "OWNER " & {i + 1})'
        c.font = Font(name=FONT_HEAD, size=13, bold=True, color="F6EFE2")
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[block_start].height = 22

        # Column-header row (one row BELOW the merged banner — never on banner row)
        hdr_row = block_start + 1
        hdrs = ["", "Metric", "Properties", "YTD revenue",
                "YTD commission", "YTD hours", "$/hr", "Notes"]
        for col, h in enumerate(hdrs, start=1):
            cell = ws.cell(row=hdr_row, column=col, value=h)
            cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[hdr_row].height = 18

        # Data row
        data_row = block_start + 2
        ws.cell(row=data_row, column=2, value="Totals").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=data_row, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        # Properties count
        c = ws.cell(
            row=data_row, column=3,
            value=f"=IFERROR('Owners'!E{owner_row},0)",
        )
        apply_style(c, formula_cell_style())
        c.number_format = "0"
        # YTD revenue (estimate: commission / commission %)
        c = ws.cell(
            row=data_row, column=4,
            value=f'=IFERROR(\'Owners\'!I{owner_row}/\'Owners\'!H{owner_row},0)',
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # YTD commission
        c = ws.cell(
            row=data_row, column=5,
            value=f"=IFERROR('Owners'!I{owner_row},0)",
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # YTD hours — SUMIFS Time Allocation by owner + active year
        c = ws.cell(
            row=data_row, column=6,
            value=(
                f"=IFERROR(SUMIFS('Time Allocation'!$E$6:$E$1005,"
                f"'Time Allocation'!$B$6:$B$1005,'Owners'!B{owner_row},"
                f"'Time Allocation'!$F$6:$F$1005,'Settings'!$B$5),0)"
            ),
        )
        apply_style(c, formula_cell_style())
        c.number_format = "0.0"
        # $/hr
        c = ws.cell(
            row=data_row, column=7,
            value=f"=IFERROR(E{data_row}/F{data_row},0)",
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # Notes
        c = ws.cell(
            row=data_row, column=8,
            value=f"=IFERROR('Owners'!L{owner_row},\"\")",
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[data_row].height = 24

        # Spacer row + sub-band trend chart placeholder note
        spacer_row = block_start + 3
        ws.merge_cells(f"B{spacer_row}:H{spacer_row}")
        c = ws[f"B{spacer_row}"]
        c.value = "12-month commission trend (chart on row block below)"
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[spacer_row].height = 14

    # Trend chart — single LineChart referencing total commission by month
    # Pull from Pam P&L row 6 (commission revenue) for all 12 months
    chart = LineChart()
    chart.title = "Total Commission — 12-Month Trend"
    chart.style = 2
    chart.y_axis.title = "Commission ($)"
    chart.x_axis.title = "Month"
    # Pam P&L sheet: months in row 5 cols B-M, commission row 6 cols B-M
    data_ref = Reference(wb["Pam P&L"] if "Pam P&L" in wb.sheetnames else ws,
                         min_col=2, max_col=13, min_row=6, max_row=6)
    cats_ref = Reference(wb["Pam P&L"] if "Pam P&L" in wb.sheetnames else ws,
                         min_col=2, max_col=13, min_row=5, max_row=5)
    # If Pam P&L not yet built, fallback: skip chart wiring; build later
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.series[0].tx = None  # set after build
    chart.height = 9
    chart.width = 22
    style_chart(chart)
    ws.add_chart(chart, "B40")

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Pam P&L
# ---------------------------------------------------------------------------

def build_pam_pl_tab(wb, variant):
    ws = wb.create_sheet("Pam P&L")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Pam P&L",
                        prev_tab="Per-Owner Dashboard", next_tab="Time Allocation")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Pam's business P&L — commission revenue less your expenses. "
        "Active year on Settings!B5."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Column headers row 5: Line item | Jan..Dec | YTD
    ws.cell(row=5, column=1, value="Line item")
    apply_style(ws.cell(row=5, column=1), header_row_style())
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, m in enumerate(months):
        cell = ws.cell(row=5, column=2 + i, value=m)
        apply_style(cell, header_row_style())
    cell = ws.cell(row=5, column=14, value="YTD")
    apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [("A", 36)] + [(get_column_letter(c), 9) for c in range(2, 14)] + [("N", 12)])

    # Income section — banner row 6, but we'll place data starting row 7
    # Actually the brief order: row 6 = Commission revenue (per-owner sum)
    # We'll put a section banner ABOVE income at row 6, income lines start row 7
    # Wait — Per-Owner Dashboard chart references row 6 cols B-M for commission.
    # To honor that, put commission revenue on row 6 directly (no banner above).
    # We'll simply skip the section banner; rows clearly labeled.

    # Income rows (start row 6)
    income_start = 6
    income_items = list(PNL_INCOME_DEMO.keys())
    for i, item in enumerate(income_items):
        row = income_start + i
        ws.cell(row=row, column=1, value=item).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        vals = PNL_INCOME_DEMO[item] if variant == "demo" else [None] * 12
        for m in range(12):
            col = 2 + m
            cell = ws.cell(row=row, column=col, value=vals[m])
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ytd_cell = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(ytd_cell, formula_cell_style())
        ytd_cell.number_format = '"$"#,##0'
        ws.row_dimensions[row].height = 18

    # Total income row
    total_income_row = income_start + len(income_items)
    ws.cell(row=total_income_row, column=1, value="TOTAL INCOME").font = Font(
        name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=total_income_row, column=1).fill = PatternFill(
        "solid", fgColor=COLOR_GOLD_SOFT
    )
    ws.cell(row=total_income_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    for m in range(12):
        col = 2 + m
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=total_income_row, column=col,
            value=f"=SUM({col_letter}{income_start}:{col_letter}{total_income_row - 1})",
        )
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell = ws.cell(
        row=total_income_row, column=14,
        value=f"=SUM(N{income_start}:N{total_income_row - 1})",
    )
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[total_income_row].height = 20

    # Expenses
    exp_start = total_income_row + 2
    # Banner row (single merged cell — no other writes to this row)
    _section_band(ws, exp_start - 1, "EXPENSES")

    exp_items = list(PNL_EXPENSE_DEMO.keys())
    for i, item in enumerate(exp_items):
        row = exp_start + i
        ws.cell(row=row, column=1, value=item).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        vals = PNL_EXPENSE_DEMO[item] if variant == "demo" else [None] * 12
        for m in range(12):
            col = 2 + m
            cell = ws.cell(row=row, column=col, value=vals[m])
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ytd_cell = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(ytd_cell, formula_cell_style())
        ytd_cell.number_format = '"$"#,##0'
        ws.row_dimensions[row].height = 18

    total_exp_row = exp_start + len(exp_items)
    ws.cell(row=total_exp_row, column=1, value="TOTAL EXPENSES").font = Font(
        name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=total_exp_row, column=1).fill = PatternFill(
        "solid", fgColor=COLOR_GOLD_SOFT
    )
    ws.cell(row=total_exp_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    for m in range(12):
        col = 2 + m
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=total_exp_row, column=col,
            value=f"=SUM({col_letter}{exp_start}:{col_letter}{total_exp_row - 1})",
        )
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell = ws.cell(
        row=total_exp_row, column=14,
        value=f"=SUM(N{exp_start}:N{total_exp_row - 1})",
    )
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[total_exp_row].height = 20

    # NOI row
    noi_row = total_exp_row + 2
    ws.cell(row=noi_row, column=1, value="NOI (Income − Expenses)").font = Font(
        name=FONT_HEAD, size=12, bold=True, color="F6EFE2"
    )
    ws.cell(row=noi_row, column=1).fill = PatternFill(
        "solid", fgColor=COLOR_PRIMARY
    )
    ws.cell(row=noi_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    for m in range(12):
        col = 2 + m
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=noi_row, column=col,
            value=f"={col_letter}{total_income_row}-{col_letter}{total_exp_row}",
        )
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color="F6EFE2")
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '"$"#,##0'
    cell = ws.cell(
        row=noi_row, column=14,
        value=f"=N{total_income_row}-N{total_exp_row}",
    )
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color="F6EFE2")
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[noi_row].height = 24

    # Margin row
    margin_row = noi_row + 1
    ws.cell(row=margin_row, column=1, value="NOI margin").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=margin_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    cell = ws.cell(
        row=margin_row, column=14,
        value=f"=IFERROR(N{noi_row}/N{total_income_row},0)",
    )
    apply_style(cell, formula_cell_style())
    cell.number_format = "0.0%"

    # Bar chart — NOI by month
    chart = BarChart()
    chart.type = "col"
    chart.title = "Pam NOI by Month"
    chart.y_axis.title = "NOI ($)"
    chart.x_axis.title = "Month"
    data_ref = Reference(ws, min_col=2, max_col=13, min_row=noi_row, max_row=noi_row)
    cats_ref = Reference(ws, min_col=2, max_col=13, min_row=5, max_row=5)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 24
    style_chart(chart)
    ws.add_chart(chart, f"A{margin_row + 3}")

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Time Allocation (capacity 1000)
# ---------------------------------------------------------------------------

def build_time_allocation_tab(wb, variant):
    ws = wb.create_sheet("Time Allocation")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Time Allocation",
                        prev_tab="Pam P&L", next_tab="Owner Profitability")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Append-only log of every hour you work, tagged to owner + property + activity. "
        "Capacity 1000 rows."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = ["Date", "Owner", "Property", "Activity", "Hours", "Year (auto)", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 24

    set_col_widths(ws, [
        ("A", 12), ("B", 26), ("C", 24), ("D", 16),
        ("E", 10), ("F", 12), ("G", 30),
    ])

    samples = _val_list(variant, TIME_ALLOC_DEMO)
    for i in range(1000):
        row = 6 + i
        if i < len(samples):
            (d, owner, prop, activity, hours, notes) = samples[i]
            cells = [
                (1, _parse_date(d), "yyyy-mm-dd"),
                (2, owner, None),
                (3, prop, None),
                (4, activity, None),
                (5, hours, "0.0"),
                (7, notes, None),
            ]
        else:
            cells = [
                (1, None, "yyyy-mm-dd"),
                (2, None, None),
                (3, None, None),
                (4, None, None),
                (5, None, "0.0"),
                (7, None, None),
            ]
        for col, val, fmt in cells:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            if fmt:
                cell.number_format = fmt
        # Year auto-calc
        yr_cell = ws.cell(row=row, column=6, value=f'=IF(A{row}="","",YEAR(A{row}))')
        apply_style(yr_cell, formula_cell_style())
        yr_cell.number_format = "0"
        ws.row_dimensions[row].height = 16

    add_dropdown(ws, "B6:B1005", "='Settings'!$B$7:$B$21")
    add_dropdown(ws, "C6:C1005", "='Settings'!$C$23:$C$52")
    add_dropdown(ws, "D6:D1005", f'"{",".join(ACTIVITIES)}"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 6 — Owner Profitability
# ---------------------------------------------------------------------------

def build_profitability_tab(wb, variant):
    ws = wb.create_sheet("Owner Profitability")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Owner Profitability",
                        prev_tab="Time Allocation", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-owner profitability — commission ÷ hours = $/hr. "
        "Bottom-3 highlighted red. Consider firing low-margin relationships."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    headers = ["Owner", "Commission YTD", "Hours YTD", "$/hr", "Rank", "Verdict"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 26

    set_col_widths(ws, [
        ("A", 28), ("B", 18), ("C", 14), ("D", 14), ("E", 10), ("F", 28),
    ])

    # 15 rows — formula-driven from Owners tab
    first_data = 6
    last_data = 20
    for i in range(15):
        row = first_data + i
        owner_row = 6 + i  # mirrors Owners!Bx
        # Owner name
        c = ws.cell(
            row=row, column=1,
            value=f"=IFERROR('Owners'!B{owner_row},\"\")",
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Commission YTD
        c = ws.cell(
            row=row, column=2,
            value=f"=IFERROR('Owners'!I{owner_row},0)",
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # Hours YTD
        c = ws.cell(
            row=row, column=3,
            value=(
                f"=IF('Owners'!B{owner_row}=\"\",0,"
                f"SUMIFS('Time Allocation'!$E$6:$E$1005,"
                f"'Time Allocation'!$B$6:$B$1005,'Owners'!B{owner_row},"
                f"'Time Allocation'!$F$6:$F$1005,'Settings'!$B$5))"
            ),
        )
        apply_style(c, formula_cell_style())
        c.number_format = "0.0"
        # $/hr
        c = ws.cell(
            row=row, column=4,
            value=f"=IFERROR(B{row}/C{row},0)",
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        # Rank
        c = ws.cell(
            row=row, column=5,
            value=(
                f'=IF(A{row}="","",'
                f'RANK(D{row},$D${first_data}:$D${last_data},0))'
            ),
        )
        apply_style(c, formula_cell_style())
        c.number_format = "0"
        # Verdict
        c = ws.cell(
            row=row, column=6,
            value=(
                f'=IF(A{row}="","",'
                f'IF(D{row}>=150,"Keep — strong",'
                f'IF(D{row}>=100,"OK — monitor",'
                f'IF(D{row}>=50,"Renegotiate","Fire candidate"))))'
            ),
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    # Conditional formatting — bottom-3 by $/hr (col D), only when row has owner
    # FormulaRule highlights row when rank in col E is one of the worst 3
    # of populated rows. Simpler: highlight when D-rank from bottom is <=3.
    # Use COUNTIFS to check populated count, then if RANK desc places this in
    # the bottom 3 of populated.
    # Apply to A:F across data rows.
    bottom3_formula = (
        f'AND($A6<>"",'
        f'COUNTIFS($A${first_data}:$A${last_data},"<>")-'
        f'RANK($D6,$D${first_data}:$D${last_data},1)+1<=3)'
    )
    error_fill = PatternFill("solid", fgColor=COLOR_ERROR)
    error_font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    ws.conditional_formatting.add(
        f"A{first_data}:F{last_data}",
        FormulaRule(
            formula=[bottom3_formula],
            fill=error_fill,
            font=error_font,
        ),
    )

    # Sort hint
    ws.merge_cells(f"A{last_data + 2}:F{last_data + 2}")
    c = ws[f"A{last_data + 2}"]
    c.value = (
        "Tip: select rows 6-20 → Data → Sort → by $/hr descending to see "
        "best-to-worst order. Red rows = bottom-3 — your fire candidates."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[last_data + 2].height = 22

    # Bar chart — $/hr by owner
    chart = BarChart()
    chart.type = "bar"
    chart.title = f"Profitability ($/hr) by Owner — {ACTIVE_YEAR}"
    chart.y_axis.title = "Owner"
    chart.x_axis.title = "$/hr"
    data_ref = Reference(ws, min_col=4, max_col=4, min_row=first_data, max_row=last_data)
    cats_ref = Reference(ws, min_col=1, max_col=1, min_row=first_data, max_row=last_data)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 22
    style_chart(chart)
    ws.add_chart(chart, f"A{last_data + 5}")

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 7 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_MUTED

    compact_header_band(ws, "Settings",
                        prev_tab="Owner Profitability", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Active year drives every YTD figure. Owner + property lists feed dropdowns."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 22), ("B", 28), ("C", 28),
    ])

    # Active year B5
    ws.cell(row=5, column=1, value="Active year:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=5, column=1).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=5, column=2, value=ACTIVE_YEAR if variant == "demo" else ACTIVE_YEAR)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[5].height = 24

    # Owner list section banner row 6, header row 7? — banner row only.
    _section_band(ws, 6, "OWNER LIST (B7:B21 — capacity 15)")

    # Owner names rows 7-21 (B7:B21)
    samples = OWNERS_DEMO if variant == "demo" else []
    for i in range(15):
        row = 7 + i
        ws.cell(row=row, column=1, value=f"Owner #{i + 1}").font = Font(
            name=FONT_BODY, size=10, color=COLOR_MUTED
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        if i < len(samples):
            owner_name = samples[i][1]
        else:
            owner_name = None
        cell = ws.cell(row=row, column=2, value=owner_name)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # Property list section banner row 22, props start row 23
    _section_band(ws, 22, "PROPERTY LIST (B23:C52 — capacity 30)")

    # Header row 23? Actually banner already on row 22; column header on row would
    # collide with banner if we wrote there. The banner merge is A22:L22 only —
    # row 23 is free. Use row 23 for column headers.
    # But brief says property list at B23-B37. Reconcile: section banner sits
    # ABOVE list. Move list to start row 24 to leave row 23 for headers.
    # Keeping the convention: banner row 22, header row 23, data rows 24-53.
    ws.cell(row=23, column=1, value="").font = Font(name=FONT_MONO, size=9)
    ws.cell(row=23, column=2, value="Property").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=23, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.cell(row=23, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=23, column=3, value="Owner").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=23, column=3).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.cell(row=23, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[23].height = 18

    # Hmm — brief: "B7-B21 Owner list, B23-B37 Property list".
    # We've shifted property data to rows 24+ to keep banner+header pattern.
    # The Owners + Time Alloc dropdowns reference 'Settings'!$C$23:$C$52 — fix.
    # Actually we should put property data starting at row 23 (matching brief)
    # and skip the column header. Let me revise: use the section banner only
    # as a label, and start data at row 23 with no header row.
    # Re-doing rows 23 area:
    # Clear what we wrote on row 23
    for col in range(1, 4):
        ws.cell(row=23, column=col).value = None
        ws.cell(row=23, column=col).fill = PatternFill(fill_type=None)

    # Data rows 23-52 (capacity 30): col B = property, col C = owner
    props = PROPERTIES_DEMO if variant == "demo" else []
    for i in range(30):
        row = 23 + i
        ws.cell(row=row, column=1, value=f"Property #{i + 1}").font = Font(
            name=FONT_BODY, size=10, color=COLOR_MUTED
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        if i < len(props):
            prop_name, owner = props[i]
        else:
            prop_name = owner = None
        b_cell = ws.cell(row=row, column=2, value=prop_name)
        apply_style(b_cell, input_cell_style())
        c_cell = ws.cell(row=row, column=3, value=owner)
        apply_style(c_cell, input_cell_style())
        ws.row_dimensions[row].height = 18

    # Owner dropdown for property→owner column
    add_dropdown(ws, "C23:C52", "='Settings'!$B$7:$B$21")

    # Note row 54
    ws.merge_cells("A54:C56")
    c = ws["A54"]
    c.value = (
        "Tip: keep property names exactly consistent across Time Allocation "
        "and Owners — formulas use exact-match SUMIFS."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(54, 57):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 58,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_owners_tab(wb, variant)
    # Pam P&L must exist before Per-Owner Dashboard chart references it
    build_pam_pl_tab(wb, variant)
    # Move Pam P&L to position 4 after dashboards build — but easier to swap order.
    # Simpler: build Per-Owner Dashboard AFTER Pam P&L, then reorder sheets.
    build_per_owner_dashboard_tab(wb, variant)
    build_time_allocation_tab(wb, variant)
    build_profitability_tab(wb, variant)
    build_settings_tab(wb, variant)

    # Reorder sheets to brief: Start, Owners, Per-Owner Dashboard, Pam P&L,
    # Time Allocation, Owner Profitability, Settings
    desired = [
        "Start", "Owners", "Per-Owner Dashboard", "Pam P&L",
        "Time Allocation", "Owner Profitability", "Settings",
    ]
    wb._sheets = [wb[name] for name in desired]

    wb.properties.title = "Multi-Owner Consolidated Reporting — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Pro Pam multi-owner P&L + profitability ranking + time allocation."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
