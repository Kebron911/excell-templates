"""Build ACQ-010 Seller-Finance Offer Calculator (v2.2 standard).

Operational/calculator tool — models a seller-finance acquisition offer
versus a conventional 25%-down DSCR loan, surfaces side-by-side cash
deltas, and prints a ready-to-send draft term sheet.

Generates:
  templates/_masters/ACQ-010-seller-finance-offer-DEMO.xlsx
  templates/_masters/ACQ-010-seller-finance-offer-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    apply_brand_header,
)

SKU = "ACQ-010"
NAME = "seller-finance-offer"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (per brief — Smokies Ridge cabin, $485K seller-finance)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Term Inputs
    "purchase_price":       485000,
    "earnest_money":        5000,
    "down_pct":             0.10,    # 10% down on seller finance
    "seller_rate":          0.065,   # 6.5%
    "amort_years":          30,
    "balloon_years":        7,
    "io_months":            0,       # interest-only period (months)
    "prepay_penalty":       "None after Year 3 (3-2-1 step-down)",
    "closing_costs":        3500,    # seller-finance closing (lower than conventional)
    "closing_date":         "2026-07-15",

    # Term Sheet parties
    "buyer_name":           "Daniel Harrison",
    "buyer_entity":         "Smokies Ridge Holdings LLC",
    "seller_name":          "Robert & Mary Whitfield",
    "property_address":     "123 Mountain Lane, Gatlinburg, TN 37738",

    # Settings
    "property_name":        "Smokies Ridge Cabin",
    "conv_rate":            0.0725,  # current market DSCR rate
    "conv_term":            30,
    "conv_down_pct":        0.25,
    "conv_closing_pct":     0.03,    # ~3% of purchase
    "ann_noi":              42000,   # annual NOI for CoC calc
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False, col=3):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=col, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet 1 — Start (KPI dashboard)
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Navy hero rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Seller-Finance Offer Calculator"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "What does this seller-finance deal really cost vs going to a bank?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Property name (rows 10-11)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!B5="","(set property name on Settings tab)",Settings!B5)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # 4 KPI cards row 13-19 (col groups: A-C, D-F, G-I, J-L)
    kpi_specs = [
        ("A", "C", "MONTHLY P&I",      "='Side-by-Side'!C12", '"$"#,##0'),
        ("D", "F", "TOTAL COST OF LOAN","='Side-by-Side'!C16", '"$"#,##0'),
        ("G", "I", "VS-CONV SAVINGS",   "='Side-by-Side'!F18", '"$"#,##0'),
        ("J", "L", "CASH-ON-CASH",      "='Side-by-Side'!C19", '0.0%'),
    ]
    for first, last, label, formula, fmt in kpi_specs:
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        ws.merge_cells(f"{first}14:{last}16")
        c = ws[f"{first}14"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

    ws.row_dimensions[13].height = 18
    for r in range(14, 17):
        ws.row_dimensions[r].height = 16

    # Verdict-style takeaway row 18
    ws.merge_cells("A18:L19")
    c = ws["A18"]
    c.value = (
        '=IF(\'Side-by-Side\'!F18>0,'
        '"✅  Seller finance saves $"&TEXT(\'Side-by-Side\'!F18,"#,##0")&" cash up front · "&TEXT(\'Side-by-Side\'!F19,"0.0%")&" CoC lift",'
        '"⚠  Seller terms net more expensive than conventional — review Side-by-Side")'
    )
    c.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(18, 20):
        ws.row_dimensions[r].height = 20

    # Pseudo-button nav rows 21-23
    pseudo_button(ws, "A21", "C23", "→ Term Inputs",
                  "'Term Inputs'!A1", variant="primary")
    pseudo_button(ws, "D21", "F23", "Side-by-Side",
                  "'Side-by-Side'!A1", variant="primary")
    pseudo_button(ws, "G21", "I23", "Term Sheet",
                  "'Term Sheet'!A1", variant="accent")
    pseudo_button(ws, "J21", "L23", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(21, 24):
        ws.row_dimensions[r].height = 22

    # Upgrade banner
    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = (
        "💡  Negotiating multiple seller-finance offers? Pair with the "
        f"STR Deal Analyzer at {BRAND_DOMAIN} for full underwriting + DSCR stress test."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[26].height = 36

    brand_footer(ws, 28,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L30"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Term Inputs
# ---------------------------------------------------------------------------

def build_term_inputs_tab(wb, variant):
    ws = wb.create_sheet("Term Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 20),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Term Inputs",
                        prev_tab="Start", next_tab="Side-by-Side")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "The seller-finance offer terms you're modeling — edit yellow cells."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "PRICE & DOWN PAYMENT")
    fields = [
        (7,  "Purchase price ($):",        _val(variant, SAMPLE["purchase_price"]), '"$"#,##0', ""),
        (8,  "Earnest money ($):",         _val(variant, SAMPLE["earnest_money"]),  '"$"#,##0', "Held in escrow at signing"),
        (9,  "Seller-finance down (%):",   _val(variant, SAMPLE["down_pct"]),       "0.0%",    "Typical 5-15% for seller carry"),
        (11, "Seller note rate (%):",      _val(variant, SAMPLE["seller_rate"]),    "0.000%",  "Fixed; often below market on motivated sellers"),
        (12, "Amortization (years):",      _val(variant, SAMPLE["amort_years"]),    "0",       "30-yr amortization standard"),
        (13, "Balloon (years):",           _val(variant, SAMPLE["balloon_years"]),  "0",       "0 = fully amortizing; 5-7 typical balloon"),
        (14, "Interest-only period (months):", _val(variant, SAMPLE["io_months"]),  "0",       "0 = no IO; 6-24 months reduces Y1 cash burn"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 16, "PREPAYMENT & CLOSING")
    fields2 = [
        (17, "Prepayment penalty terms:",  _val(variant, SAMPLE["prepay_penalty"]), None, "Free-text — pasted into Term Sheet"),
        (18, "Closing costs ($):",         _val(variant, SAMPLE["closing_costs"]),  '"$"#,##0', "Seller-finance often 1-2% (no lender fees)"),
        (19, "Target closing date:",       _val(variant, SAMPLE["closing_date"]),   None, "YYYY-MM-DD format"),
    ]
    for row, label, value, fmt, note in fields2:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 21, "DERIVED")
    _output_row(ws, 22, "Down payment ($):", "=C7*C9", '"$"#,##0')
    _output_row(ws, 23, "Seller carry-back ($):", "=C7*(1-C9)", '"$"#,##0', emphasize=True)
    _output_row(ws, 24, "Total cash to close ($):", "=C7*C9+C18", '"$"#,##0')

    brand_footer(ws, 27, version_line=f"{SKU} · Term Inputs")

    ws.print_area = "A1:L29"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Side-by-Side
# ---------------------------------------------------------------------------

def build_side_by_side_tab(wb, variant):
    ws = wb.create_sheet("Side-by-Side")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 18), ("D", 18), ("E", 4),
        ("F", 18), ("G", 18),
        ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Side-by-Side",
                        prev_tab="Term Inputs", next_tab="Term Sheet")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Seller finance vs conventional 25%-down DSCR loan — row-by-row deltas."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Two-column header row 6
    ws.cell(row=6, column=2, value="Metric")
    apply_style(ws.cell(row=6, column=2), header_row_style())

    ws.cell(row=6, column=3, value="Seller Finance")
    apply_style(ws.cell(row=6, column=3), header_row_style())
    ws.cell(row=6, column=3).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    ws.cell(row=6, column=4, value="Conventional 25%")
    apply_style(ws.cell(row=6, column=4), header_row_style())

    ws.cell(row=6, column=6, value="Δ Cash Saved")
    apply_style(ws.cell(row=6, column=6), header_row_style())
    ws.cell(row=6, column=6).fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    ws.cell(row=6, column=6).font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    ws.cell(row=6, column=7, value="Δ Annual CF")
    apply_style(ws.cell(row=6, column=7), header_row_style())
    ws.cell(row=6, column=7).fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    ws.cell(row=6, column=7).font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[6].height = 24

    # ---- Formula building blocks ----
    # Seller finance side
    sf_price        = "'Term Inputs'!C7"
    sf_down_pct     = "'Term Inputs'!C9"
    sf_rate         = "'Term Inputs'!C11"
    sf_amort        = "'Term Inputs'!C12"
    sf_balloon      = "'Term Inputs'!C13"
    sf_closing      = "'Term Inputs'!C18"
    sf_loan         = f"({sf_price}*(1-{sf_down_pct}))"
    sf_down         = f"({sf_price}*{sf_down_pct})"
    sf_pi           = f"IFERROR(PMT({sf_rate}/12,{sf_amort}*12,-{sf_loan}),0)"
    sf_annual       = f"({sf_pi}*12)"
    sf_total_cash   = f"({sf_down}+{sf_closing})"
    # Total interest over loan life: sum of (P&I × months_paid) - principal_paid_at_balloon
    # If balloon > 0: months_paid = balloon × 12; balance owed = remaining principal
    # Total cost of loan = (P&I × months_paid) + balloon_balance + down + closing
    # Months in payment phase
    sf_months       = f"IF({sf_balloon}>0,{sf_balloon}*12,{sf_amort}*12)"
    # Remaining balance at balloon (if any) — using FV of loan
    sf_balloon_bal  = f"IFERROR(-FV({sf_rate}/12,{sf_months},{sf_pi},{sf_loan}),0)"
    sf_total_cost   = f"({sf_pi}*{sf_months}+{sf_balloon_bal}+{sf_down}+{sf_closing})"
    sf_total_int    = f"({sf_pi}*{sf_months}-({sf_loan}-{sf_balloon_bal}))"

    # Conventional side
    conv_down_pct   = "Settings!B11"   # 25%
    conv_closing_pct = "Settings!B13"  # 3%
    conv_rate       = "Settings!B7"
    conv_term       = "Settings!B9"
    conv_loan       = f"({sf_price}*(1-{conv_down_pct}))"
    conv_down       = f"({sf_price}*{conv_down_pct})"
    conv_closing    = f"({sf_price}*{conv_closing_pct})"
    conv_pi         = f"IFERROR(PMT({conv_rate}/12,{conv_term}*12,-{conv_loan}),0)"
    conv_annual     = f"({conv_pi}*12)"
    conv_total_cash = f"({conv_down}+{conv_closing})"
    conv_months     = f"({conv_term}*12)"
    conv_total_cost = f"({conv_pi}*{conv_months}+{conv_down}+{conv_closing})"
    conv_total_int  = f"({conv_pi}*{conv_months}-{conv_loan})"

    # Annual NOI from settings
    ann_noi         = "Settings!B15"

    # Cash flow = NOI - debt service
    sf_cf           = f"({ann_noi}-{sf_annual})"
    conv_cf         = f"({ann_noi}-{conv_annual})"

    # CoC
    sf_coc          = f"({sf_cf}/{sf_total_cash})"
    conv_coc        = f"({conv_cf}/{conv_total_cash})"

    # ---- Rows 7-19 ----
    rows = [
        # (row, label, sf_formula, conv_formula, fmt, delta_kind)
        # delta_kind: "cash" = conv-sf for cash savings; "cf" = sf-conv; "" = no delta
        (7,  "Down payment",                f"={sf_down}",        f"={conv_down}",       '"$"#,##0', "cash"),
        (8,  "Closing costs",               f"={sf_closing}",     f"={conv_closing}",    '"$"#,##0', "cash"),
        (9,  "Total cash in",               f"={sf_total_cash}",  f"={conv_total_cash}", '"$"#,##0', "cash"),
        (10, "Loan amount",                 f"={sf_loan}",        f"={conv_loan}",       '"$"#,##0', ""),
        (11, "Interest rate",               f"={sf_rate}",        f"={conv_rate}",       "0.000%",   ""),
        (12, "Monthly P&I",                 f"={sf_pi}",          f"={conv_pi}",         '"$"#,##0.00', "cf_monthly"),
        (13, "Annual debt service",         f"={sf_annual}",      f"={conv_annual}",     '"$"#,##0', "cf"),
        (14, "Term / Balloon (years)",      f"=IF({sf_balloon}>0,{sf_balloon}&\"-yr balloon\",{sf_amort}&\"-yr fully amort\")",
                                            f"={conv_term}&\"-yr fully amort\"", None, ""),
        (15, "Balloon balance owed",        f"={sf_balloon_bal}", "=0",                  '"$"#,##0', ""),
        (16, "Total cost of loan (life)",   f"={sf_total_cost}",  f"={conv_total_cost}", '"$"#,##0', ""),
        (17, "Total interest paid",         f"={sf_total_int}",   f"={conv_total_int}",  '"$"#,##0', ""),
        (18, "Annual cash flow (NOI − DS)", f"={sf_cf}",          f"={conv_cf}",         '"$"#,##0', ""),
        (19, "Cash-on-cash",                f"={sf_coc}",         f"={conv_coc}",        "0.0%",     ""),
    ]

    for row, label, sf_f, conv_f, fmt, delta_kind in rows:
        # Label col B
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # SF col C
        cell = ws.cell(row=row, column=3, value=sf_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        # Conv col D
        cell = ws.cell(row=row, column=4, value=conv_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt

        # Delta cols F-G
        if delta_kind == "cash":
            # Conv − SF (positive = SF saves cash)
            cell = ws.cell(row=row, column=6, value=f"=D{row}-C{row}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
            cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ACCENT)
        elif delta_kind == "cf":
            # SF − Conv (positive = SF more cash flow)
            cell = ws.cell(row=row, column=7, value=f"=C{row}-D{row}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
            cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ACCENT)
        elif delta_kind == "cf_monthly":
            cell = ws.cell(row=row, column=7, value=f"=(C{row}-D{row})*12")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0;[Red]-"$"#,##0'

        ws.row_dimensions[row].height = 20

    # Headline summary band — row 21
    ws.merge_cells("B21:G21")
    c = ws["B21"]
    c.value = (
        f'=IF(F9>0,'
        f'"Seller finance saves $"&TEXT(F9,"#,##0")&" cash up front. "&'
        f'IF(C18>D18,"And nets +$"&TEXT(C18-D18,"#,##0")&" more annual cash flow.",'
        f'"But nets −$"&TEXT(D18-C18,"#,##0")&" less annual cash flow.")&" CoC: "&TEXT(C19,"0.0%")&" vs "&TEXT(D19,"0.0%")&".",'
        f'"Conventional is cheaper here — seller terms above market.")'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[21].height = 36

    # Effective cost callout row 23
    ws.merge_cells("B23:G25")
    c = ws["B23"]
    c.value = (
        "Effective cost note: if seller note rate exceeds market (Settings!B7), "
        "the absolute monthly P&I may be higher despite lower cash in. The CoC "
        "metric is the honest tiebreaker — cash flow earned per dollar invested. "
        "If you're cash-constrained, low-cash deals often win even with higher rate."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    for r in range(23, 26):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 27, version_line=f"{SKU} · Side-by-Side")

    ws.print_area = "A1:G29"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — Term Sheet (formula-concatenated print-ready doc)
# ---------------------------------------------------------------------------

def build_term_sheet_tab(wb, variant):
    ws = wb.create_sheet("Term Sheet")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 22), ("C", 22), ("D", 22), ("E", 22),
        ("F", 4), ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Term Sheet",
                        prev_tab="Side-by-Side", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Print-ready draft term sheet — generated from Term Inputs + Settings. Edit parties below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # PARTIES — yellow inputs (rows 6-9)
    _section_band(ws, 6, "PARTIES (edit these — flow into the document below)")

    parties = [
        (7, "Buyer name:",     _val(variant, SAMPLE["buyer_name"])),
        (8, "Buyer entity:",   _val(variant, SAMPLE["buyer_entity"])),
        (9, "Seller name(s):", _val(variant, SAMPLE["seller_name"])),
    ]
    for row, label, value in parties:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        ws.merge_cells(f"C{row}:E{row}")
        cell = ws.cell(row=row, column=3, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    _section_band(ws, 11, "DRAFT TERM SHEET — copy/paste or print")

    # Document title row 13
    ws.merge_cells("B13:E13")
    c = ws["B13"]
    c.value = "SELLER-FINANCED PURCHASE — TERM SHEET"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 26

    ws.merge_cells("B14:E14")
    c = ws["B14"]
    c.value = '="Drafted: "&TEXT(TODAY(),"mmmm d, yyyy")'
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 16

    # Document body — formula concatenation. Each row is one section.
    doc_rows = [
        # (row, label_text, formula_or_static)
        (16, "Property:",
         '="Property: "&Settings!B5&" — "&Settings!B17'),
        (17, "Buyer:",
         '="Buyer: "&C7&IF(C8="",""," (acting through "&C8&")")'),
        (18, "Seller:",
         '="Seller: "&C9'),
        (20, "1. PURCHASE PRICE",
         '="1. Purchase Price: $"&TEXT(\'Term Inputs\'!C7,"#,##0")&" USD."'),
        (21, "2. EARNEST MONEY",
         '="2. Earnest Money: $"&TEXT(\'Term Inputs\'!C8,"#,##0")&" deposited within three (3) business days of mutual execution, held in escrow and credited at Closing."'),
        (22, "3. DOWN PAYMENT",
         '="3. Down Payment: $"&TEXT(\'Term Inputs\'!C7*\'Term Inputs\'!C9,"#,##0")&" ("&TEXT(\'Term Inputs\'!C9,"0.0%")&" of Purchase Price), payable at Closing."'),
        (23, "4. SELLER NOTE",
         '="4. Seller Carry-Back Note: Principal of $"&TEXT(\'Term Inputs\'!C7*(1-\'Term Inputs\'!C9),"#,##0")&" at "&TEXT(\'Term Inputs\'!C11,"0.000%")&" fixed annual interest, amortized over "&\'Term Inputs\'!C12&" years, "&IF(\'Term Inputs\'!C13>0,"with a balloon payment of all unpaid principal and accrued interest due at the end of Year "&\'Term Inputs\'!C13&".","fully amortizing with no balloon.")'),
        (24, "5. INTEREST-ONLY",
         '=IF(\'Term Inputs\'!C14>0,"5. Interest-Only Period: First "&\'Term Inputs\'!C14&" months interest-only, then converting to fully amortizing payments per Section 4.","5. Payment Schedule: Standard amortizing monthly payments of approximately $"&TEXT(IFERROR(PMT(\'Term Inputs\'!C11/12,\'Term Inputs\'!C12*12,-(\'Term Inputs\'!C7*(1-\'Term Inputs\'!C9))),0),"#,##0.00")&" beginning the first day of the month following Closing.")'),
        (25, "6. PREPAYMENT",
         '="6. Prepayment: "&\'Term Inputs\'!C17&". Buyer may prepay all or any portion of principal subject to the foregoing terms."'),
        (26, "7. SECURITY",
         '="7. Security: Note secured by a first-position Deed of Trust / Mortgage on the Property recorded at Closing."'),
        (27, "8. CLOSING DATE",
         '="8. Closing Date: On or before "&\'Term Inputs\'!C19&", or such later date as the parties mutually agree in writing."'),
        (28, "9. CLOSING COSTS",
         '="9. Closing Costs: Each party bears their customary share. Buyer estimates total closing costs of $"&TEXT(\'Term Inputs\'!C18,"#,##0")&"."'),
        (29, "10. CONTINGENCIES",
         '="10. Contingencies: This offer is contingent upon (a) Buyer\'s satisfactory inspection of the Property within fifteen (15) days, (b) clear and marketable title, (c) Buyer\'s confirmation of permitted short-term-rental use, and (d) any zoning, HOA, or insurance verifications customary in the jurisdiction."'),
        (30, "11. ASSIGNMENT",
         '="11. Assignment: Buyer may assign this agreement to a wholly-owned entity ("&IF(C8="","TBD",C8)&") prior to Closing without further consent."'),
        (31, "12. DEFAULT",
         '="12. Default & Cure: Standard 30-day cure period; on uncured default, Seller\'s remedies are limited to those stated in the Note and Deed of Trust."'),
        (32, "13. NON-BINDING",
         '="13. This Term Sheet is non-binding and intended to summarize the principal economic terms. The parties shall execute a definitive Purchase & Sale Agreement and Promissory Note reflecting these terms."'),
    ]

    for row, _label, formula in doc_rows:
        ws.merge_cells(f"B{row}:E{row}")
        c = ws[f"B{row}"]
        c.value = formula
        if row in (16, 17, 18):
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        elif row in (20, 23, 25, 27):
            c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT, bold=True)
        else:
            c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 30 if row >= 20 else 18

    # Signature block
    ws.merge_cells("B34:E34")
    c = ws["B34"]
    c.value = "SIGNATURES"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[34].height = 22

    sig_pairs = [
        (36, "Buyer:", '=C7'),
        (37, "Date:", ""),
        (39, "Seller:", '=C9'),
        (40, "Date:", ""),
    ]
    for row, label, val in sig_pairs:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        ws.merge_cells(f"C{row}:E{row}")
        cell = ws.cell(row=row, column=3, value=val if val else None)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
        # Bottom border for signature line
        side = Side(style="thin", color=COLOR_PRIMARY)
        for col in range(3, 6):
            ws.cell(row=row, column=col).border = Border(bottom=side)
        ws.row_dimensions[row].height = 22

    brand_footer(ws, 43, version_line=f"{SKU} · Term Sheet · NON-BINDING DRAFT")

    ws.print_area = "B13:E42"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.7, bottom=0.7)


# ---------------------------------------------------------------------------
# Sheet 5 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_MUTED

    set_col_widths(ws, [
        ("A", 4), ("B", 24), ("C", 4), ("D", 38),
        ("E", 8), ("F", 8), ("G", 8), ("H", 8),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Term Sheet", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Property identity + conventional-loan benchmarks for the Side-by-Side comparison."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Settings rows — note: Side-by-Side references B7, B9, B11, B13, B15, B17
    # And Start references B5
    settings = [
        (5,  "Property name:",            _val(variant, SAMPLE["property_name"]), None,        ""),
        (7,  "Conventional rate (%):",    _val(variant, SAMPLE["conv_rate"]),     "0.000%",    "Today's market DSCR rate (default 7.25%)"),
        (9,  "Conventional term (yrs):",  _val(variant, SAMPLE["conv_term"]),     "0",         "30 standard"),
        (11, "Conventional down (%):",    _val(variant, SAMPLE["conv_down_pct"]), "0.0%",      "25% standard for DSCR"),
        (13, "Conventional closing (%):", _val(variant, SAMPLE["conv_closing_pct"]), "0.0%",   "Approx 3% of purchase"),
        (15, "Annual NOI (after opex, $):", _val(variant, SAMPLE["ann_noi"]),     '"$"#,##0',  "From your operating P&L — drives CoC"),
        (17, "Property address:",         _val(variant, SAMPLE["property_address"]), None,     "Used in Term Sheet"),
    ]
    for row, label, value, fmt, note in settings:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=row, column=2)  # label cell already set; now place input in col B+1? No — Settings uses col B for input
        # Actually we want the input in column B per the Side-by-Side references (B5, B7, ...)
        # Put label in col A merged or shift labels. Let's redo: label A, input B.
        # But col widths above assigned A=4 (narrow). Re-architect: put label in B label-row above, input in B value-row?
        # Simpler: keep label col B as text label, put the editable VALUE also in col B by overwriting? That conflicts.
        # Cleanest: write label into D (the "note" column), input into B.
        pass

    # Actually rebuild — clearer approach: input cell in col B (referenced as Settings!B5 etc.),
    # label in col D (free-text), note in col D continuation. Reset widths so B is wide.
    # Reset col widths for clarity:
    set_col_widths(ws, [
        ("A", 4), ("B", 28), ("C", 4),
        ("D", 30), ("E", 36),
        ("F", 8), ("G", 8), ("H", 8),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    # Clear any prior label writes in col B (rows 5,7,9,11,13,15,17)
    for r in (5, 7, 9, 11, 13, 15, 17):
        ws.cell(row=r, column=2).value = None
        ws.cell(row=r, column=2).font = Font(name=FONT_BODY, size=11)

    for row, label, value, fmt, note in settings:
        # Input cell — col B
        cell = ws.cell(row=row, column=2, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt

        # Label — col D
        lcell = ws.cell(row=row, column=4, value=label)
        lcell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        lcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Note — col E
        if note:
            ncell = ws.cell(row=row, column=5, value=note)
            ncell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
            ncell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[row].height = 18

    # Year-end archive table (Year | Closed? | Final P&I | Notes) — small ritual table
    _section_band(ws, 20, "DEAL ARCHIVE — log each seller-finance offer outcome")

    arch_headers = ["Year", "Property", "Outcome", "Final monthly P&I", "Notes"]
    for i, h in enumerate(arch_headers):
        cell = ws.cell(row=21, column=2 + i, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[21].height = 22

    # Pre-stamp 2026-2030
    for i, yr in enumerate(range(2026, 2031)):
        r = 22 + i
        cell = ws.cell(row=r, column=2, value=yr)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(3, 7):
            blank = ws.cell(row=r, column=col)
            apply_style(blank, input_cell_style())
            blank.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 30, version_line=f"{SKU} · Settings")

    ws.print_area = "A1:F32"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_term_inputs_tab(wb, variant)
    build_side_by_side_tab(wb, variant)
    build_term_sheet_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Seller-Finance Offer Calculator — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Models seller-finance acquisition terms — monthly P&I, total cost, "
        "side-by-side vs conventional 25%-down DSCR, and a print-ready draft term sheet."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
