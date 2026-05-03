"""Build STR-003 Refi-or-Sell Decision Matrix (v2.2 standard).

Compares 3 paths for an existing STR property:
  1. Hold + cash-out refi
  2. Sell outright (and reinvest proceeds)
  3. Sell + 1031 exchange into a larger asset

Outputs 5-year wealth under each path; gold-soft fills the winning card
on the Start tab scoreboard.

Generates:
  templates/_masters/STR-003-refi-or-sell-decision-matrix-DEMO.xlsx
  templates/_masters/STR-003-refi-or-sell-decision-matrix-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
)

SKU = "STR-003"
NAME = "refi-or-sell-decision-matrix"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (per brief — Lakehouse scenario, Path 3 wins)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Settings
    "property_name":   "Lakehouse",
    "hold_period":     5,
    "value_growth":    0.03,
    "reinvest_rate":   0.07,
    "refi_rate":       0.0725,
    "ltcg_rate":       0.15,
    "recapture_rate":  0.25,

    # Property Inputs
    "current_value":   720000,
    "current_loan":    469000,
    "current_pi":      2950,    # current monthly P&I
    "annual_noi":      48000,
    "original_basis":  310000,
    "years_held":      8,
    "accum_dep":       130000,

    # Path 1 Refi
    "refi_ltv":        0.75,
    "refi_term":       30,
    "refi_closing":    11000,

    # Path 2 Sell
    "selling_pct":     0.075,    # 6% commission + 1.5% closing

    # Path 3 1031
    "replacement_value": 1100000,
    "replacement_noi":   72000,
    "replacement_loan":  770000,    # 70% LTV typical
    "replacement_rate":  0.0725,
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Hero band rows 1-9
    for r in range(1, 10):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Refi-or-Sell Decision Matrix"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Three paths. One year-5 number each. The biggest wins."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Property name band row 11
    for r in range(10, 13):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!C5="","(enter property on Settings tab)",Settings!C5)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # 3-card scoreboard rows 14-22
    # Cards span: Path1 = A:D, Path2 = E:H, Path3 = I:L
    # Each card has: header (1 row), value (3 rows tall), caveat (1 row)
    # Highest year-5 wealth gets gold-soft fill via conditional Excel formula
    # We can't conditionally fill a merged cell at write-time, so we use
    # the IS-WINNER formula pattern: write three separate cards, each
    # whose body fill is set programmatically based on demo data.
    # For BLANK, we leave neutral parchment; for DEMO we highlight the
    # winning card. We then add an Excel CF rule that overlays gold-soft
    # when that card's value equals MAX of the three.

    # Card headers row 14
    card_headers = [
        ("A14", "D14", "PATH 1 · HOLD + REFI", COLOR_PRIMARY),
        ("E14", "H14", "PATH 2 · SELL", COLOR_SECONDARY),
        ("I14", "L14", "PATH 3 · SELL + 1031", COLOR_ACCENT),
    ]
    for tl, br, label, color in card_headers:
        ws.merge_cells(f"{tl}:{br}")
        c = ws[tl]
        c.value = label
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 22

    # "5-YEAR WEALTH" subtitle row 15
    for col_range in [("A15", "D15"), ("E15", "H15"), ("I15", "L15")]:
        tl, br = col_range
        ws.merge_cells(f"{tl}:{br}")
        c = ws[tl]
        c.value = "5-YEAR WEALTH"
        c.font = Font(name=FONT_MONO, size=8, color=COLOR_MUTED)
        c.fill = parchment_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[15].height = 14

    # Card body — value rows 16-19 (4 rows tall each)
    # Pull from each path tab's emphasized total cell
    card_values = [
        ("A16", "D19", "='Path 1 Refi'!C30", COLOR_PRIMARY),
        ("E16", "H19", "='Path 2 Sell'!C25", COLOR_SECONDARY),
        ("I16", "L19", "='Path 3 1031'!C30", COLOR_ACCENT),
    ]
    # Compute winner: card with MAX of the three values gets gold-soft fill.
    # We resolve at build-time: read the variant. For DEMO we know Path 3 wins.
    # For BLANK, all are 0 / blank, so default no highlight.
    # However, brief says "gold-soft fill on highest year-5 wealth" — this
    # should be data-driven, not build-time. We achieve that with a static
    # gold-soft fill on a separate "winner indicator" row that uses an IF
    # formula. Cell fills can't be IF-driven without conditional formatting,
    # so use Excel CF rule on each card-value range.

    from openpyxl.formatting.rule import FormulaRule

    for tl, br, formula, font_color in card_values:
        ws.merge_cells(f"{tl}:{br}")
        c = ws[tl]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=24, bold=True, color=font_color)
        c.fill = parchment_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '"$"#,##0'

    # Conditional formatting: gold-soft fill when card value equals MAX
    gold_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    # Build expressions referencing the same cells the merged cells use
    p1_ref = "'Path 1 Refi'!$C$30"
    p2_ref = "'Path 2 Sell'!$C$25"
    p3_ref = "'Path 3 1031'!$C$30"

    # Apply CF to each card's value range — fires when that card's wealth
    # equals MAX of the three (and total is positive).
    ws.conditional_formatting.add(
        "A16:D19",
        FormulaRule(
            formula=[f"AND({p1_ref}>0,{p1_ref}=MAX({p1_ref},{p2_ref},{p3_ref}))"],
            fill=gold_fill,
        ),
    )
    ws.conditional_formatting.add(
        "E16:H19",
        FormulaRule(
            formula=[f"AND({p2_ref}>0,{p2_ref}=MAX({p1_ref},{p2_ref},{p3_ref}))"],
            fill=gold_fill,
        ),
    )
    ws.conditional_formatting.add(
        "I16:L19",
        FormulaRule(
            formula=[f"AND({p3_ref}>0,{p3_ref}=MAX({p1_ref},{p2_ref},{p3_ref}))"],
            fill=gold_fill,
        ),
    )

    for r in range(16, 20):
        ws.row_dimensions[r].height = 18

    # Card caveat / sub-line row 20
    card_subs = [
        ("A20", "D20", '="Cash out: "&TEXT(\'Path 1 Refi\'!C12,"$#,##0")'),
        ("E20", "H20", '="Net cash: "&TEXT(\'Path 2 Sell\'!C13,"$#,##0")'),
        ("I20", "L20", '="Replacement: "&TEXT(\'Path 3 1031\'!C13,"$#,##0")'),
    ]
    for tl, br, formula in card_subs:
        ws.merge_cells(f"{tl}:{br}")
        c = ws[tl]
        c.value = formula
        c.font = Font(name=FONT_MONO, size=9, color=COLOR_MUTED)
        c.fill = parchment_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[20].height = 16

    # Row 21-22: spacer parchment
    for r in range(21, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill
        ws.row_dimensions[r].height = 8

    # Caveat banner row 23
    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = (
        "Scenario projections only. Actual outcome depends on market timing, "
        "guest reality, and financing availability when you act."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[23].height = 22

    # Nav buttons rows 25-27
    pseudo_button(ws, "A25", "C27", "Property Inputs",
                  "'Property Inputs'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "Path 1: Refi",
                  "'Path 1 Refi'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Path 2: Sell",
                  "'Path 2 Sell'!A1", variant="primary")
    pseudo_button(ws, "J25", "L27", "Path 3: 1031",
                  "'Path 3 1031'!A1", variant="primary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "F31", "→  Settings & Assumptions",
                  "'Settings'!A1", variant="accent")
    pseudo_button(ws, "G29", "L31", f"More tools at {BRAND_DOMAIN}",
                  "'Start'!A1", variant="secondary",
                  external_link=f"https://{BRAND_DOMAIN}")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    brand_footer(ws, 33,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L35"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_property_inputs_tab(wb, variant):
    ws = wb.create_sheet("Property Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Property Inputs",
                        prev_tab="Start", next_tab="Path 1 Refi")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "What you own today — value, loan, NOI, basis, depreciation"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "CURRENT POSITION")

    fields = [
        (7,  "Current market value ($):",       _val(variant, SAMPLE["current_value"]),   '"$"#,##0',  "From STR-002 valuation, broker BPO, or recent comp"),
        (8,  "Current loan balance ($):",       _val(variant, SAMPLE["current_loan"]),    '"$"#,##0',  "Today's principal owed"),
        (9,  "Current monthly P&I ($):",        _val(variant, SAMPLE["current_pi"]),      '"$"#,##0.00', "Existing payment (P&I only, exclude tax/ins escrow)"),
        (10, "Annual NOI ($):",                 _val(variant, SAMPLE["annual_noi"]),      '"$"#,##0',  "Net operating income — revenue minus opex (no debt)"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 12, "TAX BASIS")

    tax_fields = [
        (13, "Original cost basis ($):",        _val(variant, SAMPLE["original_basis"]),  '"$"#,##0',  "Purchase price + capitalized improvements"),
        (14, "Years held:",                     _val(variant, SAMPLE["years_held"]),      "0",         ""),
        (15, "Accumulated depreciation ($):",   _val(variant, SAMPLE["accum_dep"]),       '"$"#,##0',  "From your tax returns (Form 4562 / Sched E)"),
    ]
    for row, label, value, fmt, note in tax_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 17, "DERIVED — IF SOLD TODAY")

    _output_row(ws, 18, "Adjusted basis:", "=C13-C15", '"$"#,##0')
    _output_row(ws, 19, "Realized gain (sale @ current value):",
                "=C7*(1-Settings!C19)-C18", '"$"#,##0')
    _output_row(ws, 20, "Depreciation recapture portion:",
                "=MIN(C15,MAX(0,C19))", '"$"#,##0')
    _output_row(ws, 21, "Capital gain portion:",
                "=MAX(0,C19-C20)", '"$"#,##0')
    _output_row(ws, 22, "Recapture tax (25%):",
                "=C20*Settings!C17", '"$"#,##0')
    _output_row(ws, 23, "Capital gain tax (LTCG):",
                "=C21*Settings!C15", '"$"#,##0')
    _output_row(ws, 24, "Total tax if sold today:",
                "=C22+C23", '"$"#,##0', emphasize=True)

    brand_footer(ws, 27, version_line=f"{SKU} · Property Inputs")


def build_path1_refi_tab(wb, variant):
    ws = wb.create_sheet("Path 1 Refi")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Path 1: Hold + Refi",
                        prev_tab="Property Inputs", next_tab="Path 2 Sell")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Cash-out refi at current rates. Wealth = cash out + 5-yr cash flow + final equity."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "REFI INPUTS")
    fields = [
        (7,  "New LTV (% of value):",       _val(variant, SAMPLE["refi_ltv"]),     "0.0%",       "Cash-out STR loans typically max 70-75% LTV"),
        (8,  "Refi term (years):",          _val(variant, SAMPLE["refi_term"]),    "0",          "30 typical"),
        (9,  "Refi closing costs ($):",     _val(variant, SAMPLE["refi_closing"]), '"$"#,##0',   "2-3% of loan amount typical"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 11, "REFI MECHANICS")
    _output_row(ws, 12, "New loan amount:",
                "='Property Inputs'!C7*C7", '"$"#,##0')
    # Cash out = new loan - current loan - closing costs
    _output_row(ws, 13, "Cash out at close:",
                "=C12-'Property Inputs'!C8-C9", '"$"#,##0', emphasize=True)
    # New monthly P&I via Excel PMT
    _output_row(ws, 14, "New monthly P&I (Excel PMT):",
                "=IFERROR(PMT(Settings!C13/12,C8*12,-C12),0)", '"$"#,##0.00')
    _output_row(ws, 15, "New annual debt service:",
                "=C14*12", '"$"#,##0')
    _output_row(ws, 16, "Cash flow change (NOI - new DS):",
                "='Property Inputs'!C10-C15", '"$"#,##0')
    _output_row(ws, 17, "Old annual debt service (for ref):",
                "='Property Inputs'!C9*12", '"$"#,##0')

    _section_band(ws, 19, "5-YEAR PROJECTION")
    _output_row(ws, 20, "Annual cash flow (Y1, post-refi):",
                "=C16", '"$"#,##0')
    _output_row(ws, 21, "5-year cumulative cash flow:",
                "=C20*Settings!C7", '"$"#,##0')
    # Property value at end of hold period: V*(1+g)^n
    _output_row(ws, 22, "Property value at year-5:",
                "='Property Inputs'!C7*(1+Settings!C9)^Settings!C7", '"$"#,##0')
    # Loan balance at year-5: =FV-style amortization remaining
    # Remaining balance after n years: PV*(1+r)^n - PMT*((1+r)^n - 1)/r,
    # Excel: =-FV(rate, periods_paid, pmt, -loan_amount)
    _output_row(ws, 23, "New loan balance at year-5:",
                "=-FV(Settings!C13/12,Settings!C7*12,C14,-C12)", '"$"#,##0')
    _output_row(ws, 24, "Final equity at year-5:",
                "=C22-C23", '"$"#,##0')

    _section_band(ws, 27, "TOTAL 5-YEAR WEALTH")
    # Total wealth = cash out + cumulative cash flow + final equity
    _output_row(ws, 30, "Total wealth at year-5:",
                "=C13+C21+C24", '"$"#,##0', emphasize=True)

    # Methodology note
    ws.merge_cells("A33:L34")
    c = ws["A33"]
    c.value = (
        "Method: Hold the property, refinance to extract tax-free cash. "
        "5-yr wealth = cash extracted + cumulative levered cash flow + "
        "final equity (value × growth − amortized new loan)."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(33, 35):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 36, version_line=f"{SKU} · Path 1 Refi")


def build_path2_sell_tab(wb, variant):
    ws = wb.create_sheet("Path 2 Sell")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Path 2: Sell",
                        prev_tab="Path 1 Refi", next_tab="Path 3 1031")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Sell outright. Reinvest net cash at the assumed rate. Wealth = future value at year-5."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "SALE MECHANICS")
    _output_row(ws, 7, "Gross sale price (current value):",
                "='Property Inputs'!C7", '"$"#,##0')
    _output_row(ws, 8, "Selling costs (commission + closing):",
                "=C7*Settings!C19", '"$"#,##0')
    _output_row(ws, 9, "Net sale proceeds:",
                "=C7-C8", '"$"#,##0')
    _output_row(ws, 10, "Mortgage payoff:",
                "='Property Inputs'!C8", '"$"#,##0')
    _output_row(ws, 11, "Cash before tax:",
                "=C9-C10", '"$"#,##0')
    _output_row(ws, 12, "Total tax (recapture + LTCG):",
                "='Property Inputs'!C24", '"$"#,##0')
    _output_row(ws, 13, "Net cash freed up:",
                "=C11-C12", '"$"#,##0', emphasize=True)

    _section_band(ws, 15, "REINVESTMENT")
    _input_row(ws, 16, "Reinvestment rate (override Settings):",
               _val(variant, SAMPLE["reinvest_rate"]), "0.0%",
               "Default pulls from Settings; override for sensitivity")
    _output_row(ws, 17, "Effective reinvest rate:",
                '=IF(C16="",Settings!C11,C16)', "0.0%")
    _output_row(ws, 18, "Years compounding:",
                "=Settings!C7", "0")
    # FV at year n: PV * (1+r)^n
    _output_row(ws, 19, "Future value at year-5:",
                "=C13*(1+C17)^C18", '"$"#,##0')

    _section_band(ws, 22, "TOTAL 5-YEAR WEALTH")
    _output_row(ws, 25, "Total wealth at year-5:",
                "=C19", '"$"#,##0', emphasize=True)

    ws.merge_cells("A28:L29")
    c = ws["A28"]
    c.value = (
        "Method: Sell at current value, pay selling costs and taxes "
        "(recapture at 25% + capital gain at LTCG rate), reinvest net "
        "cash at the assumed rate for the hold period."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(28, 30):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 31, version_line=f"{SKU} · Path 2 Sell")


def build_path3_1031_tab(wb, variant):
    ws = wb.create_sheet("Path 3 1031")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Path 3: Sell + 1031",
                        prev_tab="Path 2 Sell", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Sell + 1031 into a larger asset. Defer all gain. Wealth = replacement equity + 5-yr cash flow."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "REPLACEMENT PROPERTY INPUTS")
    fields = [
        (7,  "Replacement property value ($):",   _val(variant, SAMPLE["replacement_value"]), '"$"#,##0', "Equal or greater for full gain deferral"),
        (8,  "Replacement annual NOI ($):",       _val(variant, SAMPLE["replacement_noi"]),   '"$"#,##0', "Larger asset = larger NOI typically"),
        (9,  "Replacement loan amount ($):",      _val(variant, SAMPLE["replacement_loan"]),  '"$"#,##0', "70% LTV typical for STR"),
        (10, "Replacement loan rate:",            _val(variant, SAMPLE["replacement_rate"]),  "0.000%",   "Current acquisition rate"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 12, "1031 MECHANICS")
    # Cash needed for replacement = value - new loan
    _output_row(ws, 13, "Equity into replacement:",
                "=C7-C9", '"$"#,##0', emphasize=True)
    # Cash from sale (no taxes paid in 1031): net proceeds - payoff
    _output_row(ws, 14, "Cash from sale (deferred):",
                "='Path 2 Sell'!C9-'Path 2 Sell'!C10", '"$"#,##0')
    # Carryover basis = adjusted basis + (replacement - relinquished)
    _output_row(ws, 15, "Carryover basis (replacement):",
                "='Property Inputs'!C18+(C7-'Property Inputs'!C7)", '"$"#,##0')

    _section_band(ws, 17, "REPLACEMENT OPERATIONS")
    # Replacement monthly P&I via Excel PMT
    _output_row(ws, 18, "Replacement monthly P&I (Excel PMT):",
                "=IFERROR(PMT(C10/12,30*12,-C9),0)", '"$"#,##0.00')
    _output_row(ws, 19, "Replacement annual debt service:",
                "=C18*12", '"$"#,##0')
    _output_row(ws, 20, "Annual cash flow (NOI − DS):",
                "=C8-C19", '"$"#,##0')
    _output_row(ws, 21, "5-year cumulative cash flow:",
                "=C20*Settings!C7", '"$"#,##0')

    _section_band(ws, 23, "EXIT EQUITY AT YEAR-5")
    _output_row(ws, 24, "Replacement value at year-5:",
                "=C7*(1+Settings!C9)^Settings!C7", '"$"#,##0')
    _output_row(ws, 25, "Replacement loan balance at year-5:",
                "=-FV(C10/12,Settings!C7*12,C18,-C9)", '"$"#,##0')
    _output_row(ws, 26, "Final equity at year-5:",
                "=C24-C25", '"$"#,##0')

    _section_band(ws, 28, "TOTAL 5-YEAR WEALTH")
    # Total = final equity + cumulative cash flow
    _output_row(ws, 30, "Total wealth at year-5:",
                "=C26+C21", '"$"#,##0', emphasize=True)

    ws.merge_cells("A33:L34")
    c = ws["A33"]
    c.value = (
        "Method: 1031 exchange defers all gain. Equity rolls into a "
        "larger asset; depreciation basis carries over. 5-yr wealth = "
        "year-5 equity + cumulative cash flow on the replacement."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(33, 35):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 36, version_line=f"{SKU} · Path 3 1031")


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Path 3 1031", next_tab="Start",
                        next_label="HOME →")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Property name + assumptions used by all three path tabs."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Inputs aligned to brief-specified rows (B5 name, B7 hold, B9 growth,
    # B11 reinvest, B13 refi rate, B15 LTCG, B17 recapture, B19 selling pct).
    _input_row(ws, 5, "Property name:",  _val(variant, SAMPLE["property_name"]),
               None, "Shows on Start tab hero band")
    _input_row(ws, 7, "Hold / projection period (years):",
               _val(variant, SAMPLE["hold_period"]), "0",
               "Default 5 yrs — matches brief")
    _input_row(ws, 9, "Property value growth (%/yr):",
               _val(variant, SAMPLE["value_growth"]), "0.0%",
               "Annual appreciation assumption (3% conservative default)")
    _input_row(ws, 11, "Reinvestment rate (Path 2):",
               _val(variant, SAMPLE["reinvest_rate"]), "0.0%",
               "Where do sale proceeds go? S&P ≈ 7-9%, another deal ≈ 8-12%")
    _input_row(ws, 13, "Refi rate (Path 1):",
               _val(variant, SAMPLE["refi_rate"]), "0.000%",
               "Today's cash-out STR rate")
    _input_row(ws, 15, "Long-term capital gains rate:",
               _val(variant, SAMPLE["ltcg_rate"]), "0.0%",
               "15% or 20% federal — check your bracket + state add-on")
    _input_row(ws, 17, "Depreciation recapture rate:",
               _val(variant, SAMPLE["recapture_rate"]), "0.0%",
               "Federal flat 25% on accumulated depreciation")
    _input_row(ws, 19, "Selling costs (% of sale):",
               _val(variant, SAMPLE["selling_pct"]), "0.0%",
               "6% commission + 1.5% closing typical (= 7.5%)")

    # Suppress section band header — Settings stays clean
    brand_footer(ws, 22, version_line=f"{SKU} · Settings")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_inputs_tab(wb, variant)
    build_path1_refi_tab(wb, variant)
    build_path2_sell_tab(wb, variant)
    build_path3_1031_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Refi-or-Sell Decision Matrix — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Three-path 5-year wealth comparison for an existing STR property: "
        "hold + cash-out refi vs. sell + reinvest vs. sell + 1031 exchange. "
        "Surfaces tax consequences, transaction costs, and ongoing cash flow."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
