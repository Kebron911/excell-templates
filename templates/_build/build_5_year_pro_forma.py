"""Build ACQ-008 5-Year Pro-Forma (v2.2 standard).

Operational mode — 5-year property-level pro-forma matrix with revenue
growth, expense inflation, debt amortization, depreciation, and
cumulative cash flow. Outputs IRR, equity multiple, and exit value at
year 5 across 3 scenarios (Conservative / Base / Aggressive).

Generates two files:
  templates/_masters/ACQ-008-5-year-pro-forma-DEMO.xlsx   (sample data)
  templates/_masters/ACQ-008-5-year-pro-forma-BLANK.xlsx  (empty inputs)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
)

SKU = "ACQ-008"
NAME = "5-year-pro-forma"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (per brief — $485K Smokies cabin)
# ---------------------------------------------------------------------------

SAMPLE = {
    "property_name":   "Smokies Ridge Cabin",
    "active_year":     2026,
    "hold_period":     5,
    "discount_rate":   0.10,
    "tax_rate":        0.28,

    # Purchase / financing
    "purchase":        485000,
    "down_pct":        0.25,
    "rate":            0.0725,
    "term_years":      30,
    "closing_costs":   9000,
    "rehab":           35000,
    "furnishing":      25000,

    # Year-1 revenue / expense baseline
    "y1_revenue":      96000,
    "vacancy_pct":     0.05,

    # Year-1 operating expenses (line items)
    "property_tax":    5800,
    "insurance":       3400,
    "hoa":             0,
    "utilities":       2400,
    "internet":        960,
    "software":        480,
    "marketing":       1200,
    "cleaning":        14000,
    "supplies":        2400,
    "mgmt":            0,
    "maintenance":     3600,
    "other_opex":      1800,

    # Growth assumptions
    "rev_growth":      0.04,
    "exp_inflation":   0.03,
    "capex_pct":       0.03,

    # Exit
    "exit_cap":        0.075,
    "selling_cost":    0.06,
}

# Land allocation (cost basis less land); 27.5-yr SL depreciation
LAND_PCT = 0.20
DEPREC_LIFE = 27.5


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False, col=3):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=row, column=col, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Tab 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "5-Year Pro-Forma"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Five-year cash flow, IRR, and exit math — what real underwriters use."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Property name banner + KPI band (rows 10-22)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!B7="","(enter property on Settings tab)",Settings!B7)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # KPI labels row 13
    ws.merge_cells("A13:D13")
    c = ws["A13"]
    c.value = "YEAR-5 NOI"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E13:H13")
    c = ws["E13"]
    c.value = "5-YEAR IRR"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I13:L13")
    c = ws["I13"]
    c.value = "EQUITY MULTIPLE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # KPI values row 14-15
    ws.merge_cells("A14:D15")
    c = ws["A14"]
    c.value = "='5-Year Pro-Forma'!G16"  # NOI Y5
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("E14:H15")
    c = ws["E14"]
    c.value = "='Exit Analysis'!C20"  # IRR
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0%"

    ws.merge_cells("I14:L15")
    c = ws["I14"]
    c.value = "='Exit Analysis'!C19"  # Equity multiple
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.00\"x\""

    ws.row_dimensions[14].height = 28
    ws.row_dimensions[15].height = 18

    # Secondary metric — Year-5 sale price + total cash returned
    ws.merge_cells("A17:F17")
    c = ws["A17"]
    c.value = "YEAR-5 SALE PRICE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G17:L17")
    c = ws["G17"]
    c.value = "TOTAL CASH RETURNED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 18

    ws.merge_cells("A18:F19")
    c = ws["A18"]
    c.value = "='Exit Analysis'!C8"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("G18:L19")
    c = ws["G18"]
    c.value = "='Exit Analysis'!C17"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.row_dimensions[18].height = 24
    ws.row_dimensions[19].height = 16

    # Nav buttons rows 24-26
    pseudo_button(ws, "A24", "C26", "Inputs",
                  "'Inputs'!A1", variant="primary")
    pseudo_button(ws, "D24", "F26", "5-Year Pro-Forma",
                  "'5-Year Pro-Forma'!A1", variant="primary")
    pseudo_button(ws, "G24", "I26", "Scenarios",
                  "'Scenarios'!A1", variant="primary")
    pseudo_button(ws, "J24", "L26", "Exit Analysis",
                  "'Exit Analysis'!A1", variant="primary")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A28", "F30", "→  Inputs (start here)",
                  "'Inputs'!A1", variant="accent")
    pseudo_button(ws, "G28", "L30", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(28, 31):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        f"💡  Underwriting a portfolio? Get the Acquisition Bundle at "
        f"{BRAND_DOMAIN} — deal analyzer, cost-to-launch, arbitrage finder, "
        f"and this pro-forma together."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 2 — Inputs
# ---------------------------------------------------------------------------

def build_inputs_tab(wb, variant):
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Inputs",
                        prev_tab="Start", next_tab="5-Year Pro-Forma")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "All assumptions live here. Yellow = editable."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # --- Purchase + financing ---
    _section_band(ws, 6, "PURCHASE & FINANCING")
    purchase_fields = [
        (7,  "Purchase price ($):",       _val(variant, SAMPLE["purchase"]),       '"$"#,##0', ""),
        (8,  "Down payment (%):",         _val(variant, SAMPLE["down_pct"]),       "0.0%",    "20-25% typical for STR"),
        (9,  "Interest rate (%):",        _val(variant, SAMPLE["rate"]),           "0.000%",  ""),
        (10, "Term (years):",             _val(variant, SAMPLE["term_years"]),     "0",       "30 typical"),
        (11, "Closing costs ($):",        _val(variant, SAMPLE["closing_costs"]),  '"$"#,##0', "2-3% of purchase typical"),
        (12, "Rehab budget ($):",         _val(variant, SAMPLE["rehab"]),          '"$"#,##0', ""),
        (13, "Furnishing budget ($):",    _val(variant, SAMPLE["furnishing"]),     '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in purchase_fields:
        _input_row(ws, row, label, value, fmt, note)

    # --- Year-1 revenue + expenses ---
    _section_band(ws, 15, "YEAR-1 REVENUE & EXPENSES")
    rev_exp_fields = [
        (16, "Year-1 gross revenue ($):", _val(variant, SAMPLE["y1_revenue"]),     '"$"#,##0', "ADR × occ × 365"),
        (17, "Vacancy / credit loss (%):", _val(variant, SAMPLE["vacancy_pct"]),   "0.0%",    "5% typical for STR cancellations"),
        (18, "Property tax (annual):",    _val(variant, SAMPLE["property_tax"]),   '"$"#,##0', ""),
        (19, "Insurance (annual):",       _val(variant, SAMPLE["insurance"]),      '"$"#,##0', "STR rider included"),
        (20, "HOA / Condo (annual):",     _val(variant, SAMPLE["hoa"]),            '"$"#,##0', ""),
        (21, "Utilities (annual):",       _val(variant, SAMPLE["utilities"]),      '"$"#,##0', ""),
        (22, "Internet (annual):",        _val(variant, SAMPLE["internet"]),       '"$"#,##0', ""),
        (23, "Software / PMS (annual):",  _val(variant, SAMPLE["software"]),       '"$"#,##0', ""),
        (24, "Marketing (annual):",       _val(variant, SAMPLE["marketing"]),      '"$"#,##0', ""),
        (25, "Cleaning (annual):",        _val(variant, SAMPLE["cleaning"]),       '"$"#,##0', "Total Y1 turnover labor"),
        (26, "Supplies (annual):",        _val(variant, SAMPLE["supplies"]),       '"$"#,##0', ""),
        (27, "Management fee (annual):",  _val(variant, SAMPLE["mgmt"]),           '"$"#,##0', "0 if self-managed"),
        (28, "Maintenance (annual):",     _val(variant, SAMPLE["maintenance"]),    '"$"#,##0', ""),
        (29, "Other operating ($):",      _val(variant, SAMPLE["other_opex"]),     '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in rev_exp_fields:
        _input_row(ws, row, label, value, fmt, note)

    # --- Growth assumptions ---
    _section_band(ws, 31, "GROWTH ASSUMPTIONS")
    growth_fields = [
        (32, "Annual revenue growth (%):", _val(variant, SAMPLE["rev_growth"]),     "0.0%", "Typical 3-5%"),
        (33, "Annual expense inflation (%):", _val(variant, SAMPLE["exp_inflation"]), "0.0%", "Typical 3%"),
        (34, "Capital reserve (% of revenue):", _val(variant, SAMPLE["capex_pct"]),  "0.0%", "Typical 3%"),
    ]
    for row, label, value, fmt, note in growth_fields:
        _input_row(ws, row, label, value, fmt, note)

    # --- Exit assumptions ---
    _section_band(ws, 36, "EXIT ASSUMPTIONS")
    exit_fields = [
        (37, "Exit cap rate (%):",        _val(variant, SAMPLE["exit_cap"]),       "0.000%", "Entry cap or +50bps"),
        (38, "Selling cost (%):",         _val(variant, SAMPLE["selling_cost"]),   "0.0%",   "5-7% typical"),
    ]
    for row, label, value, fmt, note in exit_fields:
        _input_row(ws, row, label, value, fmt, note)

    # --- Derived totals ---
    _section_band(ws, 40, "DERIVED")
    _output_row(ws, 41, "Loan amount:",
                "=C7*(1-C8)", '"$"#,##0')
    _output_row(ws, 42, "Annual debt service (P&I):",
                "=IFERROR(PMT(C9/12,C10*12,-C41)*12,0)", '"$"#,##0')
    _output_row(ws, 43, "Total cash to close:",
                "=C7*C8+C11+C12+C13", '"$"#,##0', emphasize=True)
    _output_row(ws, 44, "Annual depreciation (27.5-yr SL):",
                f"=(C7+C12+C13)*(1-{LAND_PCT})/{DEPREC_LIFE}", '"$"#,##0')
    _output_row(ws, 45, "Total Y1 operating expenses:",
                "=SUM(C18:C29)", '"$"#,##0')

    brand_footer(ws, 47, version_line=f"{SKU} · Inputs")


# ---------------------------------------------------------------------------
# Tab 3 — 5-Year Pro-Forma
# ---------------------------------------------------------------------------

def build_pro_forma_tab(wb, variant):
    ws = wb.create_sheet("5-Year Pro-Forma")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14),
        ("H", 4),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "5-Year Pro-Forma",
                        prev_tab="Inputs", next_tab="Scenarios")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Five-year matrix. Revenue grows at Inputs!C32; expenses inflate at Inputs!C33."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Year headers row 6
    ws.cell(row=6, column=2, value="Line item").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE
    )
    apply_style(ws.cell(row=6, column=2), header_row_style())
    ws.cell(row=6, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for i, col in enumerate(["C", "D", "E", "F", "G"], start=1):
        cell = ws[f"{col}6"]
        cell.value = f"Year {i}"
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # Helper to write a matrix row
    def matrix_row(row, label, col_formulas, fmt='"$"#,##0', emphasize=False, band=False):
        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.font = Font(
            name=FONT_HEAD if emphasize else FONT_BODY,
            size=12 if emphasize else 11,
            bold=emphasize or band,
            color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
        )
        label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if band:
            label_cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        for i, formula in enumerate(col_formulas):
            cell = ws.cell(row=row, column=3 + i, value=formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.number_format = fmt
            if emphasize:
                cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
                cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            elif band:
                cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        ws.row_dimensions[row].height = 22 if emphasize else 18

    # Year offsets: Y1 idx 0, Y5 idx 4
    # Revenue formulas: gross potential = Y1 rev × (1+g)^(yr-1)
    rev_formulas = [
        f"=Inputs!C16*(1+Inputs!C32)^{i}" for i in range(5)
    ]
    # Vacancy = revenue × vacancy %
    vac_formulas = [
        f"=-C{8+0}*Inputs!C17" if False else None
        for i in range(5)
    ]
    # Easier: build directly referencing the row number we'll write to
    # Row 8 = gross revenue, row 9 = vacancy, row 10 = effective gross income

    # --- Row 7: section band Revenue ---
    matrix_row(7, "REVENUE", ["", "", "", "", ""], band=True)

    # Row 8: gross potential revenue
    matrix_row(8, "Gross potential revenue",
               [f"=Inputs!$C$16*(1+Inputs!$C$32)^{i}" for i in range(5)])

    # Row 9: vacancy / credit loss
    matrix_row(9, "  Less: vacancy / credit loss",
               [f"=-{col}8*Inputs!$C$17" for col in ["C", "D", "E", "F", "G"]])

    # Row 10: effective gross income
    matrix_row(10, "Effective gross income",
               [f"={col}8+{col}9" for col in ["C", "D", "E", "F", "G"]])

    # Row 11: section band Expenses
    matrix_row(11, "OPERATING EXPENSES", ["", "", "", "", ""], band=True)

    # Row 12: total operating expenses (Y1 sum × inflation)
    matrix_row(12, "  Operating expenses (inflated)",
               [f"=-Inputs!$C$45*(1+Inputs!$C$33)^{i}" for i in range(5)])

    # Row 13: capital reserves (% of EGI)
    matrix_row(13, "  Capital reserves (% of revenue)",
               [f"=-{col}10*Inputs!$C$34" for col in ["C", "D", "E", "F", "G"]])

    # Row 14: total expenses
    matrix_row(14, "Total expenses",
               [f"={col}12+{col}13" for col in ["C", "D", "E", "F", "G"]])

    # Row 16: NOI (skip 15 for breathing room, actually keep tight)
    matrix_row(16, "NET OPERATING INCOME (NOI)",
               [f"={col}10+{col}14" for col in ["C", "D", "E", "F", "G"]],
               emphasize=True)

    # Row 17: debt service
    matrix_row(17, "  Less: debt service (P&I)",
               [f"=-Inputs!$C$42" for _ in range(5)])

    # Row 18: cash flow before tax
    matrix_row(18, "CASH FLOW BEFORE TAX",
               [f"={col}16+{col}17" for col in ["C", "D", "E", "F", "G"]],
               emphasize=True)

    # Row 20: tax section band
    matrix_row(20, "TAX TREATMENT", ["", "", "", "", ""], band=True)

    # Row 21: depreciation
    matrix_row(21, "  Depreciation (27.5-yr SL)",
               [f"=-Inputs!$C$44" for _ in range(5)])

    # Row 22: mortgage interest approx (Y1 = loan × rate, then approximate amortization)
    # For simplicity use IPMT for each year
    matrix_row(22, "  Mortgage interest (IPMT)",
               [f"=IFERROR(IPMT(Inputs!$C$9/12,({i}*12)+1,Inputs!$C$10*12,-Inputs!$C$41)*-12,0)"
                if False else
                f"=IFERROR(CUMIPMT(Inputs!$C$9/12,Inputs!$C$10*12,Inputs!$C$41,({i-1}*12)+1,{i}*12,0),0)"
                for i in range(1, 6)])

    # Row 23: taxable income = NOI + depreciation (negative) + mortgage interest (negative)
    # Note: row 21 and row 22 are already negative; NOI is positive
    matrix_row(23, "Taxable income",
               [f"={col}16+{col}21+{col}22" for col in ["C", "D", "E", "F", "G"]])

    # Row 24: tax owed (only if taxable income > 0)
    matrix_row(24, "  Tax owed (Settings!B13 × max(0, taxable))",
               [f"=-MAX(0,{col}23)*Settings!$B$13" for col in ["C", "D", "E", "F", "G"]])

    # Row 25: after-tax cash flow
    matrix_row(25, "AFTER-TAX CASH FLOW",
               [f"={col}18+{col}24" for col in ["C", "D", "E", "F", "G"]],
               emphasize=True)

    # Row 27: cumulative
    matrix_row(27, "Cumulative cash flow",
               ["=C25",
                "=C27+D25",
                "=D27+E25",
                "=E27+F25",
                "=F27+G25"])

    brand_footer(ws, 30, version_line=f"{SKU} · 5-Year Pro-Forma")


# ---------------------------------------------------------------------------
# Tab 4 — Scenarios
# ---------------------------------------------------------------------------

def build_scenarios_tab(wb, variant):
    ws = wb.create_sheet("Scenarios")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 16), ("D", 16), ("E", 16),
        ("F", 4), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Scenarios",
                        prev_tab="5-Year Pro-Forma", next_tab="Exit Analysis")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Three side-by-side projections. Adjust the assumption rows; the math flows."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Headers row 6
    headers = [
        ("B6", "Metric"),
        ("C6", "Conservative"),
        ("D6", "Base"),
        ("E6", "Aggressive"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # Assumption rows 7-9 (editable)
    assumption_rows = [
        (7, "Revenue growth (%):",       0.01, 0.04, 0.06, "0.0%"),
        (8, "Expense inflation (%):",    0.04, 0.03, 0.02, "0.0%"),
        (9, "Vacancy / occupancy hit (%):", 0.10, 0.05, 0.02, "0.0%"),
    ]
    for row, label, c_val, base_val, agg_val, fmt in assumption_rows:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col, value in [(3, c_val), (4, base_val), (5, agg_val)]:
            cell = ws.cell(row=row, column=col, value=_val(variant, value))
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = fmt
        ws.row_dimensions[row].height = 18

    # --- Helpers per scenario column ---
    # Each column references its own growth/inflation/vacancy in C7:E9
    # Y5 revenue = Inputs!C16 × (1 - vacancy) × (1 + growth)^4
    # Y5 expenses = Inputs!C45 × (1 + inflation)^4
    # Y5 NOI = Y5 rev - Y5 exp - capex
    # Sum 5-yr cash flow + sale proceeds → total return

    def per_col(col):
        g = f"{col}$7"
        infl = f"{col}$8"
        vac = f"{col}$9"
        # Sum of revenue Y1..Y5 with vacancy
        rev_sum = (
            f"SUMPRODUCT(Inputs!$C$16*(1-{vac})*(1+{g})^"
            f"{{0;1;2;3;4}})"
        )
        exp_sum = (
            f"SUMPRODUCT(Inputs!$C$45*(1+{infl})^{{0;1;2;3;4}})"
        )
        capex_sum = (
            f"SUMPRODUCT(Inputs!$C$16*(1-{vac})*(1+{g})^"
            f"{{0;1;2;3;4}}*Inputs!$C$34)"
        )
        debt_sum = "Inputs!$C$42*5"
        # Y5 NOI for exit
        y5_rev = f"Inputs!$C$16*(1-{vac})*(1+{g})^4"
        y5_exp = f"Inputs!$C$45*(1+{infl})^4"
        y5_noi = f"({y5_rev}-{y5_exp}-{y5_rev}*Inputs!$C$34)"
        sale = f"({y5_noi}/Inputs!$C$37)*(1-Inputs!$C$38)"
        # Approximate remaining loan balance at end of Y5 with CUMPRINC
        remaining = (
            "(Inputs!$C$41+CUMPRINC(Inputs!$C$9/12,Inputs!$C$10*12,"
            "Inputs!$C$41,1,60,0))"
        )
        net_sale = f"({sale}-{remaining})"
        cumulative_cf = f"({rev_sum}-{exp_sum}-{capex_sum}-{debt_sum})"
        total_return = f"({cumulative_cf}+{net_sale})"
        equity_mult = f"({total_return}/Inputs!$C$43)"
        return {
            "rev_sum": rev_sum,
            "exp_sum": exp_sum,
            "capex_sum": capex_sum,
            "debt_sum": debt_sum,
            "y5_noi": y5_noi,
            "sale": sale,
            "net_sale": net_sale,
            "cumulative_cf": cumulative_cf,
            "total_return": total_return,
            "equity_mult": equity_mult,
            "remaining": remaining,
        }

    summary_rows = [
        (11, "5-yr gross revenue", "rev_sum", '"$"#,##0', False),
        (12, "5-yr operating expenses", "exp_sum", '"$"#,##0', False),
        (13, "5-yr capital reserves", "capex_sum", '"$"#,##0', False),
        (14, "5-yr debt service", "debt_sum", '"$"#,##0', False),
        (15, "5-yr cumulative cash flow", "cumulative_cf", '"$"#,##0', True),
        (17, "Year-5 NOI", "y5_noi", '"$"#,##0', False),
        (18, "Year-5 sale price", "sale", '"$"#,##0', False),
        (19, "Less: loan balance at Y5", "remaining", '"$"#,##0', False),
        (20, "Net sale proceeds", "net_sale", '"$"#,##0', True),
        (22, "Total cash returned (CF + sale)", "total_return", '"$"#,##0', True),
        (23, "Equity multiple", "equity_mult", "0.00\"x\"", True),
    ]

    for row, label, key, fmt, emphasize in summary_rows:
        label_font = Font(
            name=FONT_HEAD if emphasize else FONT_BODY,
            size=12 if emphasize else 11,
            bold=True,
            color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
        )
        ws.cell(row=row, column=2, value=label).font = label_font
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for i, col_letter in enumerate(["C", "D", "E"]):
            formula = f"={per_col(col_letter)[key]}"
            cell = ws.cell(row=row, column=3 + i, value=formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = fmt
            if emphasize:
                cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
                cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[row].height = 22 if emphasize else 18

    # Section bands at row 10 and 16
    _section_band(ws, 10, "5-YEAR FINANCIAL SUMMARY")
    _section_band(ws, 16, "EXIT METRICS (YEAR 5 SALE)")
    _section_band(ws, 21, "TOTAL RETURN")

    brand_footer(ws, 26, version_line=f"{SKU} · Scenarios")


# ---------------------------------------------------------------------------
# Tab 5 — Exit Analysis
# ---------------------------------------------------------------------------

def build_exit_tab(wb, variant):
    ws = wb.create_sheet("Exit Analysis")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Exit Analysis",
                        prev_tab="Scenarios", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Year-5 sale math: NOI ÷ exit cap = sale price. Less loan balance, less selling costs."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "EXIT VALUATION")

    _output_row(ws, 7, "Year-5 NOI:",
                "='5-Year Pro-Forma'!G16", '"$"#,##0')
    _output_row(ws, 8, "Year-5 sale price (NOI ÷ exit cap):",
                "=C7/Inputs!C37", '"$"#,##0', emphasize=True)
    _output_row(ws, 9, "  Less: selling costs:",
                "=-C8*Inputs!C38", '"$"#,##0')
    _output_row(ws, 10, "  Less: remaining loan balance:",
                "=-(Inputs!C41+CUMPRINC(Inputs!C9/12,Inputs!C10*12,Inputs!C41,1,60,0))",
                '"$"#,##0')
    _output_row(ws, 11, "Net sale proceeds:",
                "=C8+C9+C10", '"$"#,##0', emphasize=True)

    _section_band(ws, 13, "TOTAL RETURN & IRR")

    _output_row(ws, 14, "Total cash invested:",
                "=Inputs!C43", '"$"#,##0')
    _output_row(ws, 15, "Cumulative 5-yr cash flow:",
                "='5-Year Pro-Forma'!G27", '"$"#,##0')
    _output_row(ws, 16, "Net sale proceeds:",
                "=C11", '"$"#,##0')
    _output_row(ws, 17, "Total cash returned:",
                "=C15+C16", '"$"#,##0', emphasize=True)
    _output_row(ws, 18, "Total profit:",
                "=C17-C14", '"$"#,##0')
    _output_row(ws, 19, "Equity multiple (× cash invested):",
                "=C17/C14", "0.00\"x\"", emphasize=True)
    # IRR uses Excel's IRR over years 0..5. Year 0 = -cash invested.
    # Years 1-4 = '5-Year Pro-Forma'!C25..F25 (after-tax CF)
    # Year 5 = G25 + net sale proceeds (C11)
    _output_row(ws, 20, "Internal Rate of Return (IRR):",
                "=IRR(CHOOSE({1;2;3;4;5;6},"
                "-C14,"
                "'5-Year Pro-Forma'!C25,"
                "'5-Year Pro-Forma'!D25,"
                "'5-Year Pro-Forma'!E25,"
                "'5-Year Pro-Forma'!F25,"
                "'5-Year Pro-Forma'!G25+C11))",
                "0.0%", emphasize=True)

    # Methodology footnote
    ws.merge_cells("A23:L25")
    c = ws["A23"]
    c.value = (
        "IRR uses Excel's IRR() across 6 periods: year-0 negative cash investment, "
        "years 1-4 after-tax cash flow, year 5 after-tax cash flow plus net sale proceeds. "
        "Loan balance computed via CUMPRINC over 60 monthly payments. "
        "Equity multiple = total cash returned ÷ total cash invested."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(23, 26):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 27, version_line=f"{SKU} · Exit Analysis")


# ---------------------------------------------------------------------------
# Tab 6 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 28),
        ("C", 4), ("D", 38),
        ("E", 8), ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Exit Analysis", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Workbook-level settings. Property name shows on the Start tab banner."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    fields = [
        (5,  "Active year:",        _val(variant, SAMPLE["active_year"]),    "0",     "Year 1 calendar year"),
        (7,  "Property name:",      _val(variant, SAMPLE["property_name"]),  None,    "Shown on Start banner"),
        (9,  "Hold period (years):", _val(variant, SAMPLE["hold_period"]),   "0",     "Default 5"),
        (11, "Discount rate (NPV):", _val(variant, SAMPLE["discount_rate"]), "0.0%",  "Default 10%"),
        (13, "Blended tax rate:",   _val(variant, SAMPLE["tax_rate"]),       "0.0%",  "Used for taxable-income calc"),
    ]
    for row, label, value, fmt, note in fields:
        # Note shifts to col D (after C used for input)
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=row, column=2 if False else 2)
        # Place input in col B's neighbor — use B for label, B-col 2; place input in column 2... wait
        # We'll override: input in column 2 conflicts. Reset to col 4 (D-style)
        # Actually column 2 already has the label. Use col D (4) for input as documented in brief (B5 etc).
        # Brief says B5 = active year. So input goes in column 2. But col 2 is the label.
        # Reinterpret brief: B5 means cell B5 = active year value. Then label in col A.
        # Adjust — clear what we wrote and redo
        pass
    # Redo Settings rows per brief layout: B5/B7/B9/B11/B13 hold values, A column hosts labels
    # Clear previous label writes
    for row, *_ in fields:
        ws.cell(row=row, column=2).value = None

    set_col_widths(ws, [
        ("A", 28), ("B", 22),
        ("C", 4), ("D", 38),
        ("E", 8), ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    for row, label, value, fmt, note in fields:
        # Label in col A (right-aligned)
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        lbl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        # Input in col B
        cell = ws.cell(row=row, column=2, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt
        # Note in col D
        if note:
            ws.merge_cells(f"D{row}:L{row}")
            n = ws[f"D{row}"]
            n.value = note
            n.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
            n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    brand_footer(ws, 17, version_line=f"{SKU} · Settings")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_inputs_tab(wb, variant)
    build_pro_forma_tab(wb, variant)
    build_scenarios_tab(wb, variant)
    build_exit_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "5-Year Pro-Forma — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Five-year property-level pro-forma with revenue growth, expense "
        "inflation, debt amortization, depreciation, IRR, equity multiple, "
        "and exit analysis across three scenarios."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
