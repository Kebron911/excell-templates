"""Build TAX-003 1099-NEC Contractor Tracker Excel file (v2.2 standard).

Operational-mode register: customer logs payments to vendors year-round,
files 1099-NECs by January 31. Implements the v2.2 visual + interaction
language set by the Welcome Book v2.2 and Mileage Log v2.3.

Per skill str-tax-context.md: this tracker is the exception that does NOT
need an Active tax year cell — vendor relationships span multiple years
and the YTD totals are deliberately rolling. Dates are still parsed to
real datetime objects (good hygiene; lets users build their own date
pivots without surprises).

Generates templates/_masters/TAX-003-1099-nec-tracker.xlsx.
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    COLOR_FORMULA_TINT, STATE_BAD_FILL, STATE_GOOD_FILL, STATE_WARN_FILL,
    STATE_INFO_FILL,
)

SKU = "TAX-003"
NAME = "1099-nec-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None

# --- Sample data ---

# v2.3 added Entity Type, TIN-Match Result, COI Expiry columns.
# Entity Type drives the corporate-exemption rule: payments to S-Corp/C-Corp
# are 1099-NEC EXEMPT (except attorneys). LLC depends on tax election —
# customer should ask. Individual/Sole-Prop/Partnership = 1099 required.
CONTRACTORS = [
    ("Sarah Smokies Clean",  "Smokies Clean LLC",  "XX-XXXXXXX", "147 Pine St",  "Gatlinburg",   "TN", "37738", "sarah@smokiesclean.com",  "(865) 555-0145", "Cleaning services",          "Yes", "2025-12-15", "LLC (single-member)", "Match", "2026-12-31"),
    ("Bob Handyman",         "Bob's Ridge Repair", "XX-XXXXXXX", "20 Ridge Rd",  "Sevierville",  "TN", "37862", "bob@ridgerepair.com",     "(865) 555-0198", "General handyman + repairs", "Yes", "2025-11-20", "Sole Proprietor",      "Match", "2026-08-15"),
    ("Lens Photography",     "Lens Co",            "—",          "1 Market St",  "Knoxville",    "TN", "37902", "hello@lensco.com",        "(865) 555-0211", "Listing photography",        "No",  "",            "Unknown",              "Not run", ""),
    ("Joe Landscape",        "Joe's Lawn Care",    "XX-XXXXXXX", "55 Oak Ln",    "Pigeon Forge", "TN", "37863", "joe@joeslawn.com",        "(865) 555-0155", "Lawn + landscape",           "No",  "",            "Sole Proprietor",      "Not run", "2026-04-01"),
    ("Quick Plumbing",       "Quick Plumbing LLC", "XX-XXXXXXX", "88 Water Way", "Sevierville",  "TN", "37862", "info@quickplumbing.com",  "(865) 555-0166", "Emergency plumbing",         "No",  "",            "S-Corp",               "Match", "2027-01-31"),
]


def _build_payments():
    payments = []
    # Sarah — 24 turnovers, 2 per month Jan-Dec, $400 Venmo Smokies Ridge
    for month in range(1, 13):
        for day in (7, 21):
            payments.append((
                f"2026-{month:02d}-{day:02d}", "Sarah Smokies Clean", 400,
                "Venmo", "Smokies Ridge", "Turnover", ""
            ))
    payments += [
        ("2026-01-15", "Bob Handyman", 280, "Zelle", "Smokies Ridge", "Kitchen faucet repair",    ""),
        ("2026-02-22", "Bob Handyman", 150, "Venmo", "Creek Side",    "Door lock replacement",    ""),
        ("2026-04-10", "Bob Handyman", 220, "Check", "Lakehouse A",   "Deck board replacement",   ""),
        ("2026-05-28", "Bob Handyman", 180, "Venmo", "Smokies Ridge", "Water heater fitting",     ""),
        ("2026-07-14", "Bob Handyman", 170, "Venmo", "Creek Side",    "Misc fixes",               ""),
    ]
    payments.append(("2026-03-10", "Lens Photography", 500, "ACH", "Lakehouse A", "Listing photos", "One-time project"))
    joe_dates = [
        "2026-04-05", "2026-04-19", "2026-05-03", "2026-05-17", "2026-06-07",
        "2026-06-21", "2026-07-05", "2026-07-19", "2026-08-16", "2026-09-13",
    ]
    for d in joe_dates:
        payments.append((d, "Joe Landscape", 80, "Zelle", "Smokies Ridge", "Weekly lawn", ""))
    payments += [
        ("2026-06-12", "Quick Plumbing", 300, "Check", "Creek Side",  "Emergency leak",     ""),
        ("2026-09-05", "Quick Plumbing", 300, "Check", "Lakehouse A", "Water heater call",  ""),
    ]
    return payments


PAYMENTS = _build_payments()

PENALTIES = [
    ("Filed ≤30 days late:",    "$60 per form"),
    ("Filed 31 days–Aug 1:",    "$130 per form"),
    ("Filed after Aug 1:",      "$330 per form"),
    ("Intentional disregard:",  "$660 per form"),
]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational-mode hero + cards + activity dashboard)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "1099-NEC Contractor Tracker"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Catch the $600 line before January 31 does."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Anyone you pay $600+ in a calendar year for business services requires "
        "a 1099-NEC issued by January 31 of the following year. Penalties run "
        "$60–$660 per missed form (IRS Pub 1220). This tracker logs every payment, "
        "flags vendors who cross the threshold in real time, and surfaces who's "
        "missing a W-9 — so January is a reporting day, not a scramble."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Contractors tab: add every vendor you pay — name, EIN/SSN, address, W-9 status.",
        "② Payment Log: log every payment AS IT HAPPENS (date, amount, method) — not at year-end.",
        "③ 1099 Prep Dashboard: watch who crosses $600 in real time; status flips automatically.",
        "④ Monthly: scan Dashboard for anyone YES who's missing a W-9 → request the W-9 immediately.",
        "⑤ December: confirm every YES vendor has a W-9 on file before year-end.",
        "⑥ January: file 1099-NECs (QuickBooks, Track1099, or your CPA) using Dashboard data.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "ADD A CONTRACTOR" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  ADD A NEW CONTRACTOR  (OPEN CONTRACTORS)",
                   "'Contractors'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Total Vendors Tracked
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "VENDORS TRACKED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = "=COUNTA(Contractors!A6:A55)"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "rows on the Contractors tab"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Vendors Over $600
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "VENDORS OVER $600"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = '=COUNTIF(\'1099 Prep Dashboard\'!C6:C55,"YES")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "1099-NEC required"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Total $ Paid
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "TOTAL $ PAID"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = "=SUM('Payment Log'!C6:C2005)"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "all contractor payments"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Payment Log",
                   "'Payment Log'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "1099 Prep Dashboard",
                   "'1099 Prep Dashboard'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Print Dashboard",
                   "'1099 Prep Dashboard'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: January-31 deadline callout (rows 42-43) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📅 JAN 31 DEADLINE: file 1099-NECs for every YES vendor with both the "
        "IRS and the recipient. Missing a W-9? Request it NOW — vendors slow-walk "
        "in January. 📄 Print the Dashboard for your CPA or filing service."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Threshold reminder (row 44) ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "⚠ The $600 threshold applies to total YTD payments per vendor for "
        "business services. Reimbursements for materials don't count if "
        "separately invoiced. When in doubt, issue the form."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    add_upgrade_banner(ws, 46)

    brand_footer(ws, 48,
                 version_line="TAX-003 · v2.2 · Free updates forever")

    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_contractors_tab(wb, variant):
    """Sheet 1 — Contractors (one row per vendor, W-9 tracking)."""
    ws = wb.create_sheet("Contractors")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Contractors",
                         prev_tab="Start", next_tab="Payment Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Add every vendor you pay — name, EIN/SSN, address, W-9 status"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 22), ("B", 22), ("C", 14), ("D", 28), ("E", 14),
        ("F", 5),  ("G", 9),  ("H", 24), ("I", 14), ("J", 22),
        ("K", 7),  ("L", 12), ("M", 18), ("N", 10), ("O", 12),
    ])

    headers = [
        "Name", "Business Name", "EIN/SSN", "Address", "City",
        "State", "Zip", "Email", "Phone", "Services",
        "W-9?", "W-9 Date",
        "Entity Type",        # v2.3 — drives corp-exemption rule
        "TIN Match",          # v2.3 — IRS TIN Matching service result
        "COI Expiry",         # v2.3 — Certificate of Insurance (liability)
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Rows 6+: sample contractors (DEMO only)
    contractor_rows = CONTRACTORS if variant == "demo" else []
    for row_idx, contractor in enumerate(contractor_rows, start=6):
        for col_idx, value in enumerate(contractor, start=1):
            # W-9 Date (12) and COI Expiry (15) are dates — parse strings
            if col_idx in (12, 15):
                value = _parse_date(value)
            cell = ws.cell(row=row_idx, column=col_idx,
                            value=value if value not in ("", None) else None)
            apply_style(cell, input_cell_style())
            if col_idx in (12, 15) and value:
                cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[row_idx].height = 16

    # Rows 11-55: blank capacity
    for row_idx in range(11, 56):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx in (12, 15):
                cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[row_idx].height = 16

    add_dropdown(ws, "K6:K55", '"Yes,No"')
    add_dropdown(ws, "M6:M55",
                  '"Individual,Sole Proprietor,Partnership,LLC (single-member),'
                  'LLC (multi-member),S-Corp,C-Corp,Attorney (always 1099),Unknown"')
    add_dropdown(ws, "N6:N55", '"Match,Mismatch,Not run"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_log_tab(wb, variant):
    """Sheet 2 — Payment Log (one row per payment, 2000 capacity rows)."""
    ws = wb.create_sheet("Payment Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Payment Log",
                         prev_tab="Contractors", next_tab="1099 Prep Dashboard")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Log every payment AS IT HAPPENS — not at year-end (IRS contemporaneous standard)"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 26), ("C", 12), ("D", 16),
        ("E", 20), ("F", 28), ("G", 22), ("H", 12), ("I", 12),
    ])

    # v2.3 added Backup Withhold $ (mandatory 24% when no W-9) and the
    # derived 1099-NEC Eligible? flag — Credit Card / Venmo / PayPal payments
    # are reported by the processor on 1099-K, NOT 1099-NEC. Including them
    # would double-tax the contractor.
    headers = [
        "Date", "Contractor", "Amount", "Payment Method", "Property",
        "Description", "Notes",
        "Backup Withhold $",
        "1099-NEC Eligible?",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    payment_rows = PAYMENTS if variant == "demo" else []
    for i, payment in enumerate(payment_rows, start=6):
        date_val, contractor, amount, method, prop, desc, notes = payment

        a = ws.cell(row=i, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=contractor)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=amount)
        apply_style(c, input_cell_style())
        c.number_format = '"$"#,##0.00'

        d = ws.cell(row=i, column=4, value=method)
        apply_style(d, input_cell_style())

        e = ws.cell(row=i, column=5, value=prop)
        apply_style(e, input_cell_style())

        f = ws.cell(row=i, column=6, value=desc)
        apply_style(f, input_cell_style())

        g = ws.cell(row=i, column=7, value=notes if notes else None)
        apply_style(g, input_cell_style())

        h = ws.cell(row=i, column=8)  # Backup Withhold $ — input
        apply_style(h, input_cell_style())
        h.number_format = '"$"#,##0.00'

        # 1099-NEC Eligible? formula — exclude TPSO / card payments
        i_cell = ws.cell(
            row=i, column=9,
            value=f'=IF(D{i}="","",IF(OR(D{i}="Credit Card",D{i}="Venmo",D{i}="PayPal"),"No","Yes"))'
        )
        apply_style(i_cell, formula_cell_style())
        i_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 16

    last_data_row = len(payment_rows) + 5
    for row_idx in range(last_data_row + 1, 2006):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx in (3, 8):
                cell.number_format = '"$"#,##0.00'
        # Capacity-row 1099-NEC Eligible formula
        i_cell = ws.cell(
            row=row_idx, column=9,
            value=f'=IF(D{row_idx}="","",IF(OR(D{row_idx}="Credit Card",D{row_idx}="Venmo",D{row_idx}="PayPal"),"No","Yes"))'
        )
        apply_style(i_cell, formula_cell_style())
        i_cell.alignment = Alignment(horizontal="center", vertical="center")

    add_dropdown(ws, "B6:B2005", "=Contractors!$A$6:$A$55")
    add_dropdown(ws, "D6:D2005",
                  '"Venmo,Zelle,Check,Cash,ACH,Credit Card,PayPal,Other"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_dashboard_tab(wb):
    """Sheet 3 — 1099 Prep Dashboard / CPA Hand-off page."""
    ws = wb.create_sheet("1099 Prep Dashboard")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = CPA hand-off

    compact_header_band(ws, "1099 Prep Dashboard · For Your CPA",
                         prev_tab="Payment Log", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 FOR YOUR CPA — print this tab. Watches who crosses the $600 threshold in real time."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 28), ("B", 16), ("C", 18), ("D", 16), ("E", 18),
    ])

    headers = ["Contractor", "YTD Paid", "1099 Required?", "W-9 on File?", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Rows 6-55: formula rows (50 vendor capacity)
    for i in range(6, 56):
        a = ws.cell(
            row=i, column=1,
            value=f'=IF(Contractors!A{i}="","",Contractors!A{i})',
        )
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")

        # v2.3 — filter on 1099-NEC Eligible? = "Yes" so credit card / Venmo /
        # PayPal payments (which the processor reports on 1099-K) don't get
        # double-counted. Including them would double-tax the contractor.
        b = ws.cell(
            row=i, column=2,
            value=(
                f'=IF(A{i}="","",'
                f'SUMIFS(\'Payment Log\'!$C:$C,'
                f'\'Payment Log\'!$B:$B,A{i},'
                f'\'Payment Log\'!$I:$I,"Yes"))'
            ),
        )
        apply_style(b, formula_cell_style())
        b.number_format = '"$"#,##0.00'

        # v2.3 — corp-exemption logic: S-Corp and C-Corp are 1099-NEC EXEMPT
        # except for attorneys (always required, regardless of entity).
        # Pulls Entity Type from Contractors col M via VLOOKUP.
        c = ws.cell(
            row=i, column=3,
            value=(
                f'=IF(A{i}="","",'
                f'IF(B{i}<Settings!$B$5,"no",'
                f'IF(IFERROR(VLOOKUP(A{i},Contractors!$A$6:$M$55,13,FALSE),"")="Attorney (always 1099)","YES",'
                f'IF(OR(IFERROR(VLOOKUP(A{i},Contractors!$A$6:$M$55,13,FALSE),"")="S-Corp",'
                f'IFERROR(VLOOKUP(A{i},Contractors!$A$6:$M$55,13,FALSE),"")="C-Corp"),'
                f'"exempt (corp)","YES"))))'
            ),
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        d = ws.cell(
            row=i, column=4,
            value=(
                f'=IF(A{i}="","",'
                f'IFERROR(VLOOKUP(A{i},Contractors!$A$6:$L$55,11,FALSE),"?"))'
            ),
        )
        apply_style(d, formula_cell_style())
        d.alignment = Alignment(horizontal="center", vertical="center")

        e = ws.cell(
            row=i, column=5,
            value=(
                f'=IF(A{i}="","",'
                f'IF(C{i}="YES",'
                f'IF(D{i}="Yes","✓ Ready","⚠ Need W-9"),'
                f'"n/a"))'
            ),
        )
        apply_style(e, formula_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 16

    # Conditional formatting (preserved from v1)
    ws.conditional_formatting.add(
        "C6:C55",
        FormulaRule(
            formula=['C6="YES"'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True),
        ),
    )
    ws.conditional_formatting.add(
        "C6:C55",
        FormulaRule(
            formula=['C6="no"'],
            fill=PatternFill("solid", fgColor=COLOR_FORMULA_TINT),
        ),
    )
    # v2.3 — corp-exemption status (over $600 but S-Corp/C-Corp): info-grade
    ws.conditional_formatting.add(
        "C6:C55",
        FormulaRule(
            formula=['C6="exempt (corp)"'],
            fill=PatternFill("solid", fgColor=STATE_INFO_FILL),
        ),
    )
    ws.conditional_formatting.add(
        "E6:E55",
        FormulaRule(
            formula=['E6="✓ Ready"'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
        ),
    )
    ws.conditional_formatting.add(
        "E6:E55",
        FormulaRule(
            formula=['E6="⚠ Need W-9"'],
            fill=PatternFill("solid", fgColor=STATE_WARN_FILL),
        ),
    )

    ws.row_dimensions[59].height = 8

    hdr60 = ws.cell(row=60, column=1, value="Summary")
    hdr60.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[60].height = 22

    summary_rows = [
        (61, "Contractors requiring 1099:", '=COUNTIF(C6:C55,"YES")', "0"),
        (62, "Of those, ready (W-9 on file):", '=COUNTIF(E6:E55,"✓ Ready")', "0"),
        (63, "Of those, need W-9:", '=COUNTIF(E6:E55,"⚠ Need W-9")', "0"),
        (64, "Total 1099-NEC $ volume:", '=SUMIFS(B6:B55,C6:C55,"YES")', '"$"#,##0.00'),
        # v2.3 additions
        (65, "Exempt (corp / 1099-K processor):", '=COUNTIF(C6:C55,"exempt (corp)")', "0"),
        (66, "Backup withholding remitted YTD:", "=SUM('Payment Log'!H6:H2005)", '"$"#,##0.00'),
    ]
    for row_num, label, formula, num_fmt in summary_rows:
        a = ws.cell(row=row_num, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")

        b = ws.cell(row=row_num, column=2, value=formula)
        b.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        b.number_format = num_fmt
        b.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row_num].height = 18

    # --- Chart: Donut — $ paid by contractor (top of dashboard) ---
    # Anchored G5 so it doesn't overlap the data table cols A-E.
    donut = DoughnutChart()
    donut.title = "$ Paid by Contractor (YTD)"
    donut.height = 9
    donut.width = 11
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=2, min_row=5, max_row=20, max_col=2)
    donut_cats = Reference(ws, min_col=1, min_row=6, max_row=20)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=False, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, "G5")

    ws.freeze_panes = "A6"

    ws.print_area = "A1:E64"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    """Sheet 4 — Settings (IRS threshold, penalty schedule, year-end archive)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="1099 Prep Dashboard", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "IRS threshold · penalty schedule · year-end archive"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 36), ("B", 16), ("C", 14), ("D", 16)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- IRS threshold (row 5) — referenced by Dashboard formulas ---
    a5 = ws.cell(row=5, column=1, value="IRS 1099-NEC threshold ($):")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")

    b5 = ws.cell(row=5, column=2, value=600)
    apply_style(b5, input_cell_style())
    b5.number_format = '"$"#,##0'
    ws.row_dimensions[5].height = 18

    # Row 6: explainer
    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = (
        "Threshold is locked at $600 by IRS rule. Don't change unless the IRS "
        "publishes a new threshold (rare) — every Dashboard formula filters on this cell."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 28

    ws.row_dimensions[7].height = 8

    # --- Penalty schedule (rows 8-13) ---
    sect8 = ws.cell(row=8, column=1, value="IRS 1099 penalty schedule (reference)")
    sect8.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[8].height = 22

    pen_help = ws.cell(row=9, column=1,
                        value="Per missed/late form. Don't be the host paying $660 a head.")
    pen_help.font = italic_muted
    ws.row_dimensions[9].height = 16

    for i, (label, amount) in enumerate(PENALTIES, start=10):
        a = ws.cell(row=i, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")

        b = ws.cell(row=i, column=2, value=amount)
        b.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        b.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[i].height = 16

    ws.row_dimensions[14].height = 10

    # --- Year-end Archive (rows 15-23) ---
    sect15 = ws.cell(row=15, column=1, value="Year-end Archive")
    sect15.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[15].height = 22

    ws.merge_cells("A16:D16")
    c16 = ws["A16"]
    c16.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then clear the Payment Log for the new year (keep Contractors)."
    )
    c16.font = italic_muted
    c16.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[16].height = 28

    archive_headers = ["Year", "Vendors Tracked", "Vendors > $600", "Total $ Paid"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=17, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[17].height = 18

    for idx, year in enumerate(range(2024, 2031), start=18):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0"
        cell = ws.cell(row=idx, column=4)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_contractors_tab(wb, variant)
    build_log_tab(wb, variant)
    build_dashboard_tab(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"1099-NEC Contractor Tracker{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = "1099-NEC threshold tracker for STR hosts paying contractors (v2.3)."

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
