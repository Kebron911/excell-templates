"""Build LGL-003 Guest Screening Log + Ban List Excel file.

Operational-mode register: per-booking screening checklist + portfolio-wide
ban list. Records what the host verified before accepting a booking and
aggregates "do not host" list across properties — so a problem guest banned
at property A can't book property B.

Forks the v2.3 register-with-flags pattern from build_1099_nec_tracker.py.

Generates:
  templates/_masters/LGL-003-guest-screening-log-DEMO.xlsx
  templates/_masters/LGL-003-guest-screening-log-BLANK.xlsx
"""
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
)

SKU = "LGL-003"
NAME = "guest-screening-log"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"


def _parse_date(s):
    if not s:
        return None
    if isinstance(s, date):
        return s
    return datetime.strptime(s, "%Y-%m-%d").date()


# --- Sample data (DEMO) ---

# 25 screenings Jan-Mar 2026: 22 approved + 3 declined
# Schema: booking_date, guest_name, property, platform, profile_age_months,
#         verified_id, prior_reviews, avg_rating, trip_purpose, local,
#         screening_service, decision, notes
SCREENINGS = [
    ("2026-01-04", "Megan Albright",   "Smokies Ridge", "Airbnb",  18, "Yes", 12, 4.9, "Family vacation",       "No",  "SuperHog",      "Approved", ""),
    ("2026-01-08", "Tyler Brooks",     "Creek Side",    "VRBO",    24, "Yes",  8, 4.8, "Anniversary trip",      "No",  "Autohost",      "Approved", ""),
    ("2026-01-11", "Jordan Cho",       "Lakehouse A",   "Airbnb",   6, "Yes",  3, 5.0, "Weekend getaway",       "No",  "SuperHog",      "Approved", ""),
    ("2026-01-14", "Patrick Diaz",     "Smokies Ridge", "Booking",  1, "No",   0, 0.0, "Just visiting",         "Yes", "Manual",        "Declined", "Local + no profile age"),
    ("2026-01-17", "Lauren Edwards",   "Creek Side",    "Airbnb",   9, "Yes",  5, 4.7, "Hiking trip",           "No",  "SuperHog",      "Approved", ""),
    ("2026-01-21", "Marcus Foster",    "Lakehouse A",   "Direct",  36, "Yes", 22, 4.9, "Family reunion",        "No",  "Manual",        "Approved", "Repeat guest"),
    ("2026-01-24", "Nadia Greene",     "Smokies Ridge", "Airbnb",  14, "Yes",  9, 4.8, "Romantic weekend",      "No",  "Autohost",      "Approved", ""),
    ("2026-01-28", "Omar Hayes",       "Creek Side",    "VRBO",    20, "Yes", 11, 4.9, "Birthday celebration",  "No",  "SuperHog",      "Approved", ""),
    ("2026-02-02", "Priya Iyer",       "Lakehouse A",   "Airbnb",   8, "Yes",  4, 4.6, "Family vacation",       "No",  "SuperHog",      "Approved", ""),
    ("2026-02-05", "Quinn Jackson",    "Smokies Ridge", "Airbnb",   3, "No",   1, 5.0, "Vague",                 "Yes", "Manual",        "Declined", "Profile thin + local"),
    ("2026-02-09", "Riley King",       "Creek Side",    "Airbnb",  16, "Yes",  7, 4.8, "Hiking + brewery tour", "No",  "Autohost",      "Approved", ""),
    ("2026-02-12", "Sasha Liu",        "Lakehouse A",   "VRBO",    22, "Yes", 14, 4.9, "Retreat",               "No",  "SuperHog",      "Approved", ""),
    ("2026-02-15", "Tomas Martinez",   "Smokies Ridge", "Booking",  4, "Yes",  2, 4.5, "Vacation",              "No",  "Manual",        "Approved", ""),
    ("2026-02-19", "Uma Nakamura",     "Creek Side",    "Airbnb",  30, "Yes", 18, 4.9, "Anniversary",           "No",  "SuperHog",      "Approved", ""),
    ("2026-02-23", "Victor Ortiz",     "Lakehouse A",   "Direct",  12, "Yes",  6, 4.8, "Family vacation",       "No",  "Manual",        "Approved", ""),
    ("2026-02-26", "Wendy Patel",      "Smokies Ridge", "Airbnb",   2, "No",   0, 0.0, "Party (red flag)",      "No",  "SuperHog",      "Declined", "Suspicious — declined per house rules"),
    ("2026-03-01", "Xander Quinn",     "Creek Side",    "Airbnb",  10, "Yes",  5, 4.7, "Hiking",                "No",  "Autohost",      "Approved", ""),
    ("2026-03-05", "Yara Reyes",       "Lakehouse A",   "Airbnb",  15, "Yes",  8, 4.9, "Family",                "No",  "SuperHog",      "Approved", ""),
    ("2026-03-08", "Zachary Stone",    "Smokies Ridge", "VRBO",    25, "Yes", 12, 4.8, "Anniversary",           "No",  "SuperHog",      "Approved", ""),
    ("2026-03-12", "Aiden Thompson",   "Creek Side",    "Airbnb",   7, "Yes",  3, 4.6, "Weekend",               "No",  "Autohost",      "Approved", ""),
    ("2026-03-15", "Bella Underwood",  "Lakehouse A",   "Booking",  5, "Yes",  2, 4.5, "Birthday",              "No",  "Manual",        "Approved", ""),
    ("2026-03-19", "Caleb Vega",       "Smokies Ridge", "Airbnb",  19, "Yes", 10, 4.8, "Family",                "No",  "SuperHog",      "Approved", ""),
    ("2026-03-22", "Daria Wong",       "Creek Side",    "VRBO",    28, "Yes", 16, 4.9, "Retreat",               "No",  "SuperHog",      "Approved", ""),
    ("2026-03-26", "Ethan Xu",         "Lakehouse A",   "Direct",  44, "Yes", 28, 5.0, "Family reunion",        "No",  "Manual",        "Approved", "Repeat guest"),
    ("2026-03-29", "Fiona Yates",      "Smokies Ridge", "Airbnb",  11, "Yes",  5, 4.7, "Hiking + relaxation",   "No",  "Autohost",      "Approved", ""),
]

# 2 banned guests (per brief)
BANNED = [
    ("Wendy Patel",      "https://airbnb.com/users/show/00000",  "Threw unauthorized party despite no-party rule. Neighbor noise complaint, $400 cleanup.", "2026-02-27", "Daniel H.", "Smokies Ridge", 5, "Yes", "Photos + neighbor email saved to Google Drive folder"),
    ("Patrick Diaz",     "https://airbnb.com/users/show/00001",  "Smoked indoors, lied about it. Cigarette burns on couch + smell remediation $650.",       "2026-01-15", "Daniel H.", "Smokies Ridge", 4, "Yes", "Smoke detector log + cleaner photos"),
]

# Decline scripts (4 per brief)
DECLINE_SCRIPTS = [
    (
        "1. Local guest decline",
        "Hi {Guest},\n\n"
        "Thanks for your interest in {Property}! Unfortunately, our house rules don't permit local "
        "bookings (within 30 miles of the property) — we've found this policy helps us maintain the "
        "quiet, respectful atmosphere our neighbors and other guests appreciate.\n\n"
        "If you're looking for a special-occasion stay outside our area, we'd be happy to host you. "
        "Otherwise, please consider a hotel or local short-term rental that allows local stays.\n\n"
        "Wishing you the best,\n{Host}",
    ),
    (
        "2. Insufficient profile decline (verification request)",
        "Hi {Guest},\n\n"
        "Thanks for reaching out about {Property}. Before I can confirm your booking, I'd like to "
        "verify a couple of details since your profile is fairly new:\n\n"
        "  • Can you complete Airbnb's ID verification (it takes ~2 minutes)?\n"
        "  • Could you tell me a bit more about your trip — who's coming, the purpose, and arrival time?\n\n"
        "Once I have those, I'll confirm right away. This helps me protect both my home and my neighbors, "
        "and gives you a smooth check-in.\n\n"
        "Thanks for understanding,\n{Host}",
    ),
    (
        "3. Suspicious-pattern decline (no reason given)",
        "Hi {Guest},\n\n"
        "Thanks for considering {Property}. Unfortunately, I'm not able to accept this booking. "
        "I'd encourage you to use Airbnb's search filters to find a property that's a better match for "
        "your trip.\n\n"
        "Wishing you the best,\n{Host}\n\n"
        "(Note to self: do NOT give a reason. Platform terms allow hosts to decline without explanation; "
        "stating a reason can create discrimination-claim exposure. Redirect to platform search.)",
    ),
    (
        "4. Repeat-offender / banned guest decline",
        "Hi {Guest},\n\n"
        "Thanks for reaching out. Unfortunately, I'm not able to accept this booking.\n\n"
        "Best,\n{Host}\n\n"
        "(Note to self: previously banned guest. Brief reply only. Do NOT reference the prior incident "
        "or the ban list. Document in Screening Log as 'Declined — banned guest' and consider reporting "
        "the inquiry to the platform if behavior escalates.)",
    ),
]

# Settings — property list, screening services, decline reasons
PROPERTY_LIST = [
    "Smokies Ridge", "Creek Side", "Lakehouse A", "", "", "", "", "", "", "",
]
SCREENING_SERVICES = [
    "SuperHog", "Autohost", "TurnoverBnB", "Manual", "None", "", "", "", "", "",
]
DECLINE_REASONS = [
    "Local guest (within 30mi)",
    "Insufficient profile (no reviews + new account)",
    "Profile flags (party/event references)",
    "Suspicious messaging pattern",
    "Banned guest (portfolio ban list)",
    "Group size mismatch",
    "Date conflict (booked elsewhere)",
    "House rule conflict (pets/smoking/etc.)",
    "Other (see notes)",
    "",
]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational hero + KPIs + nav)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Guest Screening Log + Ban List"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Screen the booking before it screens you."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "A two-part shield against the booking that ruins your week. The Screening Log "
        "captures what you verified before saying yes — profile age, ID, prior reviews, "
        "screening service used, decision. The Ban List aggregates problem guests across "
        "your portfolio, so a guest banned at one property can't book another. Together "
        "they make every booking a deliberate decision, not a reflex."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Every new booking inquiry: log it in Screening Log BEFORE accepting (date, profile age, ID, reviews, decision).",
        "② Watch the row-color flags: gold = local guest (risk), red = thin profile + no reviews (auto-decline candidate).",
        "③ If you must decline: open Decline Scripts, copy the right one, send via platform — no custom reasons.",
        "④ Incident-after-stay: add the guest to Ban List with reason, severity (1-5), and evidence kept.",
        "⑤ Repeat inquirer: scan Ban List before accepting any booking — the platform doesn't share host-side bans.",
        "⑥ Settings tab: edit your property list, screening services, and decline reasons to match your portfolio.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "LOG A NEW SCREENING" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  LOG A NEW SCREENING  (OPEN SCREENING LOG)",
                   "'Screening Log'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Bookings Screened YTD
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "BOOKINGS SCREENED YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = (
        f"=COUNTIFS('Screening Log'!A6:A305,\">=\"&DATE(Settings!$F$5,1,1),"
        f"'Screening Log'!A6:A305,\"<=\"&DATE(Settings!$F$5,12,31))"
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "rows in active year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Banned Guests
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "BANNED GUESTS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = "=COUNTA('Ban List'!A6:A55)"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "do-not-host portfolio-wide"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Declined Count
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "DECLINED COUNT"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = '=COUNTIF(\'Screening Log\'!L6:L305,"Declined")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "🚫 bookings declined YTD"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Screening Log",
                   "'Screening Log'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Ban List",
                   "'Ban List'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "Decline Scripts",
                   "'Decline Scripts'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Risk callout (rows 42-43) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "⚠ THE RISK SCREEN: profile age <2 months AND zero prior reviews = "
        "auto-decline candidate. Local guests within 30 miles = elevated risk "
        "(party bookings). When in doubt, request a video call before accepting."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Privacy reminder (row 44) ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "🔒 Ban List is for YOUR portfolio only. Sharing externally raises "
        "legal/defamation risk — consult counsel before discussing with other hosts."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    brand_footer(ws, 46,
                 version_line=f"{SKU} · v1.0 · Free updates forever")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_screening_log_tab(wb, variant):
    """Sheet 1 — Screening Log (capacity 300)."""
    ws = wb.create_sheet("Screening Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Screening Log",
                         prev_tab="Start", next_tab="Ban List")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Per-booking screening — log BEFORE accepting (gold row = local guest, red row = thin profile)"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12),  # Booking date
        ("B", 22),  # Guest name
        ("C", 16),  # Property
        ("D", 12),  # Platform
        ("E", 9),   # Profile age (months)
        ("F", 8),   # Verified ID
        ("G", 9),   # Prior reviews
        ("H", 9),   # Avg rating
        ("I", 22),  # Trip purpose
        ("J", 7),   # Local
        ("K", 16),  # Screening service
        ("L", 11),  # Decision
        ("M", 28),  # Notes
    ])

    headers = [
        "Booking Date", "Guest Name", "Property", "Platform",
        "Profile Age (mo)", "Verified ID?", "Prior Reviews",
        "Avg Rating", "Trip Purpose", "Local?",
        "Screening Service", "Decision", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28
    for col in range(1, 14):
        ws.cell(row=5, column=col).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    rows = SCREENINGS if variant == "demo" else []
    for i, screening in enumerate(rows, start=6):
        (booking_date, guest, prop, platform, age, verified_id,
         reviews, rating, purpose, local, service, decision, notes) = screening

        a = ws.cell(row=i, column=1, value=_parse_date(booking_date))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=guest)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=prop)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=platform)
        apply_style(d, input_cell_style())

        e = ws.cell(row=i, column=5, value=age)
        apply_style(e, input_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")

        f = ws.cell(row=i, column=6, value=verified_id)
        apply_style(f, input_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        g = ws.cell(row=i, column=7, value=reviews)
        apply_style(g, input_cell_style())
        g.alignment = Alignment(horizontal="center", vertical="center")

        h = ws.cell(row=i, column=8, value=rating if rating else None)
        apply_style(h, input_cell_style())
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.number_format = "0.0"

        i_cell = ws.cell(row=i, column=9, value=purpose)
        apply_style(i_cell, input_cell_style())

        j = ws.cell(row=i, column=10, value=local)
        apply_style(j, input_cell_style())
        j.alignment = Alignment(horizontal="center", vertical="center")

        k = ws.cell(row=i, column=11, value=service)
        apply_style(k, input_cell_style())

        l_cell = ws.cell(row=i, column=12, value=decision)
        apply_style(l_cell, input_cell_style())
        l_cell.alignment = Alignment(horizontal="center", vertical="center")

        m = ws.cell(row=i, column=13, value=notes if notes else None)
        apply_style(m, input_cell_style())

        ws.row_dimensions[i].height = 16

    # Capacity rows up through row 305 (300 entries)
    last_data_row = len(rows) + 5
    for row_idx in range(last_data_row + 1, 306):
        for col_idx in range(1, 14):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            elif col_idx == 8:
                cell.number_format = "0.0"
            if col_idx in (5, 6, 7, 8, 10, 12):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 16

    # --- Dropdowns ---
    add_dropdown(ws, "C6:C305", "=Settings!$B$5:$B$14")
    add_dropdown(ws, "D6:D305", '"Airbnb,VRBO,Booking,Direct,Other"')
    add_dropdown(ws, "F6:F305", '"Yes,No"')
    add_dropdown(ws, "J6:J305", '"Yes,No"')
    add_dropdown(ws, "K6:K305", "=Settings!$B$16:$B$25")
    add_dropdown(ws, "L6:L305", '"Approved,Declined,Pending"')

    # --- Conditional formatting ---
    # Gold tint if local
    ws.conditional_formatting.add(
        "A6:M305",
        FormulaRule(
            formula=['$J6="Yes"'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
        ),
    )
    # Red tint if profile age < 2 AND prior reviews = 0 (auto-decline candidate)
    ws.conditional_formatting.add(
        "A6:M305",
        FormulaRule(
            formula=['AND(ISNUMBER($E6),$E6<2,ISNUMBER($G6),$G6=0)'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
        ),
    )
    # Decision col color: green Approved, red Declined
    ws.conditional_formatting.add(
        "L6:L305",
        FormulaRule(
            formula=['$L6="Approved"'],
            fill=PatternFill("solid", fgColor="C7EFCF"),
            font=Font(bold=True, color="0F5132"),
        ),
    )
    ws.conditional_formatting.add(
        "L6:L305",
        FormulaRule(
            formula=['$L6="Declined"'],
            fill=PatternFill("solid", fgColor="F8D7DA"),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        "L6:L305",
        FormulaRule(
            formula=['$L6="Pending"'],
            fill=PatternFill("solid", fgColor="FFF3BF"),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_ban_list_tab(wb, variant):
    """Sheet 2 — Ban List (capacity 50)."""
    ws = wb.create_sheet("Ban List")
    ws.sheet_properties.tabColor = COLOR_ERROR  # red flag tab

    compact_header_band(ws, "Ban List",
                         prev_tab="Screening Log", next_tab="Decline Scripts")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = (
        "🔒 PRIVACY: maintain for YOUR portfolio only. Sharing externally raises "
        "legal/defamation risk — consult counsel."
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 22),  # Guest name
        ("B", 32),  # Profile link
        ("C", 36),  # Reason banned
        ("D", 13),  # Banned date
        ("E", 16),  # Banned by host
        ("F", 18),  # Property
        ("G", 9),   # Severity
        ("H", 11),  # Evidence
        ("I", 32),  # Notes
    ])

    headers = [
        "Guest Name", "Profile Platform Link", "Reason Banned",
        "Banned Date", "Banned By Host", "Property Where Incident",
        "Severity (1-5)", "Evidence Kept?", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28
    for col in range(1, 10):
        ws.cell(row=5, column=col).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    rows = BANNED if variant == "demo" else []
    for i, banned in enumerate(rows, start=6):
        (name, link, reason, banned_date, host, prop,
         severity, evidence, notes) = banned

        a = ws.cell(row=i, column=1, value=name)
        apply_style(a, input_cell_style())

        b = ws.cell(row=i, column=2, value=link)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=reason)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=_parse_date(banned_date))
        apply_style(d, input_cell_style())
        d.number_format = "yyyy-mm-dd"

        e = ws.cell(row=i, column=5, value=host)
        apply_style(e, input_cell_style())

        f = ws.cell(row=i, column=6, value=prop)
        apply_style(f, input_cell_style())

        g = ws.cell(row=i, column=7, value=severity)
        apply_style(g, input_cell_style())
        g.alignment = Alignment(horizontal="center", vertical="center")

        h = ws.cell(row=i, column=8, value=evidence)
        apply_style(h, input_cell_style())
        h.alignment = Alignment(horizontal="center", vertical="center")

        i_cell = ws.cell(row=i, column=9, value=notes if notes else None)
        apply_style(i_cell, input_cell_style())

        ws.row_dimensions[i].height = 22

    last_data_row = len(rows) + 5
    for row_idx in range(last_data_row + 1, 56):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 4:
                cell.number_format = "yyyy-mm-dd"
            if col_idx in (7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    add_dropdown(ws, "F6:F55", "=Settings!$B$5:$B$14")
    add_dropdown(ws, "G6:G55", '"1,2,3,4,5"')
    add_dropdown(ws, "H6:H55", '"Yes,No"')

    # Severity color scale (1=mild beige, 5=red)
    ws.conditional_formatting.add(
        "G6:G55",
        FormulaRule(
            formula=['$G6>=4'],
            fill=PatternFill("solid", fgColor="F8D7DA"),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        "G6:G55",
        FormulaRule(
            formula=['AND($G6>=2,$G6<=3)'],
            fill=PatternFill("solid", fgColor="FFF3BF"),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_decline_scripts_tab(wb):
    """Sheet 3 — Decline Scripts (4 copy-paste templates)."""
    ws = wb.create_sheet("Decline Scripts")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = reference

    compact_header_band(ws, "Decline Scripts",
                         prev_tab="Ban List", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Copy-paste scripts for declining a booking. Replace {Guest}, {Property}, {Host} with your details."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    row = 6
    for title, body in DECLINE_SCRIPTS:
        # Title strip — gold band
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = title.upper()
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22
        row += 1

        # Body — parchment block, ~10 rows tall
        body_rows = 9
        ws.merge_cells(f"A{row}:L{row + body_rows - 1}")
        c = ws[f"A{row}"]
        c.value = body
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="top",
                                 wrap_text=True, indent=2)
        # Apply parchment to all body cells in merged range
        for r in range(row, row + body_rows):
            for col in range(1, 13):
                ws.cell(row=r, column=col).fill = PatternFill(
                    "solid", fgColor=COLOR_BG_LIGHT
                )
            ws.row_dimensions[r].height = 16
        row += body_rows

        # Spacer
        row += 1

    # Footer reminder
    ws.merge_cells(f"A{row}:L{row + 1}")
    c = ws[f"A{row}"]
    c.value = (
        "⚠ NEVER state a discriminatory reason (race, religion, family status, etc.). "
        "Platforms allow you to decline without explanation — use scripts 3 and 4 when in doubt."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 22

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Decline Scripts", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active year · property list · screening services · decline reasons"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 28), ("B", 32), ("C", 4), ("D", 28), ("E", 4),
        ("F", 14), ("G", 28),
    ])

    bold_label = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    section_head = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)

    # --- Active Year (cell F5) — referenced by Start tab YTD formulas ---
    # Placed at F5 to avoid conflict with B5:B14 property list dropdown range.
    ws.merge_cells("D5:E5")
    de5 = ws["D5"]
    de5.value = "Active Year:"
    de5.font = bold_label
    de5.alignment = Alignment(horizontal="right", vertical="center")

    f5 = ws.cell(row=5, column=6, value=2026)
    apply_style(f5, input_cell_style())
    f5.number_format = "0"
    f5.font = Font(name=FONT_BODY, size=14, bold=True, color=COLOR_PRIMARY)
    f5.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 24

    ws.merge_cells("D6:G6")
    c6 = ws["D6"]
    c6.value = (
        "YTD KPIs on Start tab filter on this year. Bump every January."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 22

    # --- Property list (B5:B14) — referenced by dropdowns in Screening Log col C and Ban List col F ---
    a5 = ws.cell(row=5, column=1, value="Properties:")
    a5.font = bold_label
    a5.alignment = Alignment(horizontal="right", vertical="top")

    for i, prop in enumerate(PROPERTY_LIST, start=5):
        cell = ws.cell(row=i, column=2, value=prop if prop else None)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[i].height = 16

    # Help text under property list
    ws.merge_cells("A15:B15")
    c15 = ws["A15"]
    c15.value = "↑ Edit to match your portfolio (up to 10 properties)"
    c15.font = italic_muted
    c15.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[15].height = 18

    # --- Screening services (B16:B25) ---
    a16 = ws.cell(row=16, column=1, value="Screening Services:")
    a16.font = bold_label
    a16.alignment = Alignment(horizontal="right", vertical="top")

    for i, svc in enumerate(SCREENING_SERVICES, start=16):
        cell = ws.cell(row=i, column=2, value=svc if svc else None)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[i].height = 16

    ws.merge_cells("A26:B26")
    c26 = ws["A26"]
    c26.value = "↑ Third-party guest screening tools (or 'Manual' if you DIY)"
    c26.font = italic_muted
    c26.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[26].height = 18

    # --- Decline reasons (B27:B36) ---
    a27 = ws.cell(row=27, column=1, value="Decline Reasons:")
    a27.font = bold_label
    a27.alignment = Alignment(horizontal="right", vertical="top")

    for i, reason in enumerate(DECLINE_REASONS, start=27):
        cell = ws.cell(row=i, column=2, value=reason if reason else None)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[i].height = 16

    ws.merge_cells("A37:B37")
    c37 = ws["A37"]
    c37.value = "↑ Reference list — pair with Decline Scripts when responding"
    c37.font = italic_muted
    c37.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[37].height = 18

    # --- Year-end Archive (rows 40-48) ---
    sect = ws.cell(row=40, column=1, value="Year-end Archive")
    sect.font = section_head
    ws.row_dimensions[40].height = 22

    ws.merge_cells("A41:G41")
    c41 = ws["A41"]
    c41.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then clear the Screening Log for the new year (keep Ban List always)."
    )
    c41.font = italic_muted
    c41.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[41].height = 28

    archive_headers = ["Year", "Bookings Screened", "Approved", "Declined", "Banned (cumulative)"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=42, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[42].height = 18

    for idx, year in enumerate(range(2024, 2031), start=43):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0"
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_screening_log_tab(wb, variant)
    build_ban_list_tab(wb, variant)
    build_decline_scripts_tab(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Guest Screening Log + Ban List{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Per-booking screening checklist + portfolio-wide ban list for STR hosts (v1.0)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
