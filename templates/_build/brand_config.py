"""Shared brand config for The STR Ledger Excel templates.

Colors, fonts, and helper functions used across all product builds.
Single source of truth — if brand palette changes, update here only.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# --- Brand palette (authoritative: brand/brand-decisions.md §4.1) ---
BRAND_NAME = "The STR Ledger"
BRAND_DOMAIN = "thestrledger.com"
BRAND_EMAIL = "hello@thestrledger.com"

# Authoritative: brand/brand-decisions.md §4.1
COLOR_PRIMARY = "12304E"      # Harbor Navy — authority, hero bg, primary type
COLOR_SECONDARY = "B5725E"    # Clay Rose — warmth, secondary accent (use sparingly)
COLOR_ACCENT = "C9A24B"       # Muted Gold — period-mark, rules, highlights
COLOR_TEXT = "2B2B2B"         # Graphite — body copy (softer than pure black)
COLOR_BG_LIGHT = "F6EFE2"     # Parchment — paper, warmed white, default bg
COLOR_MUTED = "6B7280"        # Slate — captions/footnotes (non-brand utility)
COLOR_ERROR = "B91C1C"        # Red — threshold warnings (non-brand utility)

# Tints/shades (brand-decisions.md §4.3) for UI states
COLOR_NAVY_TINT = "2A4867"       # hover, secondary navy
COLOR_NAVY_SHADE = "0A1F35"      # pressed, shadow
COLOR_PARCHMENT_ALT = "EFE5D0"   # alt-row banding, subtle card
COLOR_GOLD_SOFT = "E2C884"       # disabled/muted accent

# --- Utility neutrals (text on dark fills, body emphasis, low-chrome backgrounds) ---
COLOR_WHITE = "FFFFFF"           # text/icons on navy/gold fills
COLOR_INK = "222222"             # bold body text emphasis (softer than near-black)
COLOR_NEAR_BLACK = "1A1A1A"      # deepest text emphasis, headers
COLOR_GRAY_LIGHT = "CCCCCC"      # divider lines, low-emphasis borders
COLOR_INPUT_TINT = "FFF7D6"      # input-cell yellow tint
COLOR_FORMULA_TINT = "EDEDED"    # calculated-cell gray tint

# --- Semantic state fills (conditional formatting — bad/good/warn) ---
STATE_BAD_FILL = "FFCCCC"        # light red wash — over budget, deadline missed, alert
STATE_GOOD_FILL = "C7EFCF"       # light green wash — on track, paid, healthy
STATE_WARN_FILL = "FFF3BF"       # light yellow wash — caution, due soon, marginal
STATE_BAD_TEXT = "7F1D1D"        # dark red text on bad-state cells
STATE_GOOD_TEXT = "166534"       # dark green text on good-state cells
STATE_INFO_FILL = "E0E7FF"       # light blue wash — informational, lead/pipeline

# --- Fonts ---
FONT_HEAD = "Georgia"         # serif for headings
FONT_BODY = "Calibri"         # sans for body (default Excel-safe)
FONT_MONO = "Consolas"

# --- Helper: apply branding to a worksheet ---
def apply_brand_header(ws, title, subtitle=""):
    """Place a branded title block on rows 1-3 of the worksheet."""
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    ws["A3"] = f"{BRAND_NAME} · {BRAND_DOMAIN}"
    ws["A3"].font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14

def header_row_style():
    return {
        "font": Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor=COLOR_PRIMARY),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin", color=COLOR_PRIMARY),
            right=Side(style="thin", color=COLOR_PRIMARY),
            top=Side(style="thin", color=COLOR_PRIMARY),
            bottom=Side(style="thin", color=COLOR_PRIMARY),
        ),
    }

def input_cell_style():
    """Yellow-tinted fill indicates a cell the user should edit."""
    return {
        "fill": PatternFill("solid", fgColor="FFF7D6"),
        "font": Font(name=FONT_BODY, size=11, color=COLOR_TEXT),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

def formula_cell_style():
    """Gray-tinted fill indicates a calculated cell (do not edit)."""
    return {
        "fill": PatternFill("solid", fgColor="EDEDED"),
        "font": Font(name=FONT_BODY, size=11, color=COLOR_TEXT, italic=True),
        "alignment": Alignment(horizontal="right", vertical="center"),
    }

def set_col_widths(ws, widths):
    """widths is a list of (col_letter_or_int, width) pairs.

    Accepts either a column letter ("A") or a 1-based integer index.
    """
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def apply_style(cell, style_dict):
    """Apply a style dict (returned by the *_style() helpers below) to a cell.

    Callers: apply_style(ws.cell(row=5, column=1), input_cell_style())
    """
    for attr, value in style_dict.items():
        setattr(cell, attr, value)

def add_upgrade_banner(ws, row):
    """Place a prominent upgrade CTA on the given row."""
    cell = ws.cell(row=row, column=1, value=(
        f"💡 Upgrade to the Full version at {BRAND_DOMAIN}/upgrade "
        f"— multi-property, depreciation, multi-LLC support."
    ))
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[row].height = 30


# --- Wizard/tool helpers (v2 additions) ---

def pseudo_button(ws, top_left, bottom_right, text, hyperlink_target,
                  variant="primary", external_link=None):
    """Render a merged-cells block that looks and acts like a button.

    Args:
        ws: worksheet
        top_left: cell reference string, e.g. "A10"
        bottom_right: cell reference string, e.g. "L13"
        text: button label
        hyperlink_target: in-workbook reference, e.g. "'Property'!A5".
            Ignored when external_link is set.
        variant: "primary" (navy fill, parchment text) /
                 "secondary" (parchment fill, navy text, gold border) /
                 "accent" (gold fill, navy text)
        external_link: optional external path/URL (e.g.
            "welcome-book-renderer.html"). When truthy, the HYPERLINK
            formula uses this value directly (no "#" prefix) and
            hyperlink_target is ignored. When None (default), falls
            back to the in-workbook "#{hyperlink_target}" behavior.

    The top-left cell holds the formula-driven hyperlink; the merge
    spans to bottom-right; border + fill applied to the merged range.
    """
    palettes = {
        "primary": {
            "fill": COLOR_PRIMARY,
            "font_color": "F6EFE2",  # parchment
            "border_color": COLOR_ACCENT,  # gold border on navy btn
            "border_thick": True,
        },
        "secondary": {
            "fill": COLOR_BG_LIGHT,  # parchment
            "font_color": COLOR_PRIMARY,
            "border_color": COLOR_ACCENT,
            "border_thick": False,
        },
        "accent": {
            "fill": COLOR_ACCENT,  # gold
            "font_color": COLOR_PRIMARY,
            "border_color": COLOR_PRIMARY,
            "border_thick": True,
        },
    }
    p = palettes.get(variant, palettes["primary"])

    # Merge the range
    merge_range = f"{top_left}:{bottom_right}"
    ws.merge_cells(merge_range)

    # Top-left cell holds the hyperlink
    cell = ws[top_left]
    # Escape any double quotes in text (Excel uses "" to represent a literal " inside a quoted string)
    escaped_text = text.replace('"', '""')
    if external_link:
        cell.value = f'=HYPERLINK("{external_link}", "{escaped_text}")'
    else:
        cell.value = f'=HYPERLINK("#{hyperlink_target}", "{escaped_text}")'
    cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=p["font_color"])
    cell.fill = PatternFill("solid", fgColor=p["fill"])
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

    # Apply border to every cell in the merged range
    border_style = "medium" if p["border_thick"] else "thin"
    side = Side(style=border_style, color=p["border_color"])
    border = Border(left=side, right=side, top=side, bottom=side)

    # Parse top_left + bottom_right to iterate cells
    tl_col_letter, tl_row = coordinate_from_string(top_left)
    br_col_letter, br_row = coordinate_from_string(bottom_right)
    tl_col = column_index_from_string(tl_col_letter)
    br_col = column_index_from_string(br_col_letter)
    for r in range(tl_row, br_row + 1):
        for c in range(tl_col, br_col + 1):
            ws.cell(row=r, column=c).border = border


def card_header(ws, row, col_range, text, height=20):
    """Render a card header strip — 1 row tall, gold-soft fill, Consolas
    tracked uppercase label.

    Args:
        ws: worksheet
        row: row number (1-based)
        col_range: (first_col_letter, last_col_letter), e.g. ("A", "L")
        text: label text (will be rendered uppercase)
        height: row height in points
    """
    first, last = col_range
    merge_range = f"{first}{row}:{last}{row}"
    ws.merge_cells(merge_range)
    cell = ws[f"{first}{row}"]
    cell.value = text.upper()
    cell.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor="E2C884")  # gold-soft
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height


def card_body_fill(ws, row_start, row_end, col_range, border=True):
    """Apply parchment fill + optional gold border to a card body range.

    Args:
        ws: worksheet
        row_start: first body row (row after card header)
        row_end: last body row
        col_range: (first_col_letter, last_col_letter)
        border: if True, draw thin gold border around the whole range
    """
    first, last = col_range
    first_col = column_index_from_string(first)
    last_col = column_index_from_string(last)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(row_start, row_end + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            # Skip cells that already have a meaningful fill (input-cell yellow, or any
            # explicit color that isn't the default empty). Explicit rgb check is more
            # robust than the default-value check alone.
            existing_rgb = None
            if cell.fill.fgColor:
                existing_rgb = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else None
            # Apply parchment only if cell has no meaningful fill yet
            if existing_rgb in (None, "00000000", "None"):
                cell.fill = parchment_fill
    if border:
        side = Side(style="thin", color=COLOR_ACCENT)
        # Top row
        for c in range(first_col, last_col + 1):
            ws.cell(row=row_start, column=c).border = Border(
                top=side,
                left=side if c == first_col else ws.cell(row=row_start, column=c).border.left,
                right=side if c == last_col else ws.cell(row=row_start, column=c).border.right,
            )
        # Bottom row
        for c in range(first_col, last_col + 1):
            existing = ws.cell(row=row_end, column=c).border
            ws.cell(row=row_end, column=c).border = Border(
                bottom=side,
                left=side if c == first_col else existing.left,
                right=side if c == last_col else existing.right,
                top=existing.top,
            )
        # Left + right columns (middle rows)
        for r in range(row_start + 1, row_end):
            existing_l = ws.cell(row=r, column=first_col).border
            ws.cell(row=r, column=first_col).border = Border(
                left=side, top=existing_l.top, bottom=existing_l.bottom,
                right=existing_l.right,
            )
            existing_r = ws.cell(row=r, column=last_col).border
            ws.cell(row=r, column=last_col).border = Border(
                right=side, top=existing_r.top, bottom=existing_r.bottom,
                left=existing_r.left,
            )


def section_header_band(ws, section_num, total_sections, title, subtitle,
                         prev_tab, next_tab):
    """Render the navy step-header band at the top of an input tab.

    Rows 1-5:
    - Row 2: [← BACK] | SECTION N OF M | [NEXT →]
    - Row 4: title (large, Georgia bold parchment)
    - Row 5: subtitle (italic muted gold)

    Args:
        ws: worksheet
        section_num: 1-based section number
        total_sections: total number of input sections (8 for Welcome Book)
        title: section title
        subtitle: section subtitle
        prev_tab: name of previous tab for BACK button (empty string on section 1)
        next_tab: name of next tab for NEXT button (empty string on section N)
    """
    # Set row heights
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 36
    ws.row_dimensions[5].height = 18

    # Fill rows 1-5 cols A-L with navy
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 6):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Row 2 navigation
    if prev_tab:
        pseudo_button(ws, "A2", "C2", "← BACK",
                      f"'{prev_tab}'!A5", variant="primary")
    else:
        # Section 1: BACK goes to Start
        pseudo_button(ws, "A2", "C2", "← START",
                      "'Start'!A1", variant="primary")
    # Center label
    ws.merge_cells("D2:I2")
    c = ws["D2"]
    c.value = f"SECTION {section_num} OF {total_sections}"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    if next_tab:
        pseudo_button(ws, "J2", "L2", "NEXT →",
                      f"'{next_tab}'!A5", variant="accent")
    else:
        pseudo_button(ws, "J2", "L2", "LAUNCH →",
                      "'Launch'!A1", variant="accent")

    # Row 4 title
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = title
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill

    # Row 5 subtitle
    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = subtitle
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill


def compact_header_band(ws, title, prev_tab=None, next_tab=None,
                         prev_label=None, next_label=None):
    """Render a 3-row navy header strip for operational-mode tabs.

    Lighter than section_header_band — no SECTION N OF M chrome, no
    big serif title row. Suited to data-table tabs where the
    table itself is the centerpiece.

    Layout:
      Row 1: navy spacer (height 8)
      Row 2: [← BACK] | TITLE (mono, right-aligned) | [NEXT →]
      Row 3: navy spacer (height 6)
    Existing content below row 3 (typically a row-4 spacer then a
    header row at row 5) is left untouched.

    Args:
        ws: worksheet
        title: tab name shown in mono uppercase at the right
        prev_tab: name of tab the BACK button should jump to
            (e.g., "Start"). Pass None to omit the BACK button.
        next_tab: name of tab the NEXT button should jump to.
            Pass None to omit the NEXT button.
        prev_label: optional override for BACK button label
            (default: "← BACK")
        next_label: optional override for NEXT button label
            (default: "NEXT →")
    """
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 6

    for r in range(1, 4):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    if prev_tab:
        pseudo_button(
            ws, "A2", "C2",
            prev_label or "← BACK",
            f"'{prev_tab}'!A1",
            variant="primary",
        )

    # Center label slot — narrows when both buttons are present
    if prev_tab and next_tab:
        label_range = "D2:I2"
    elif prev_tab and not next_tab:
        label_range = "D2:L2"
    elif next_tab and not prev_tab:
        label_range = "A2:I2"
    else:
        label_range = "A2:L2"
    ws.merge_cells(label_range)
    label_cell = ws[label_range.split(":")[0]]
    label_cell.value = title.upper()
    label_cell.font = Font(name=FONT_MONO, size=10, bold=True,
                            color=COLOR_ACCENT)
    label_cell.alignment = Alignment(
        horizontal="right" if next_tab or not prev_tab else "right",
        vertical="center", indent=2,
    )
    label_cell.fill = navy_fill

    if next_tab:
        pseudo_button(
            ws, "J2", "L2",
            next_label or "NEXT →",
            f"'{next_tab}'!A1",
            variant="accent",
        )


def brand_footer(ws, start_row, version_line=None,
                  contact_line=None):
    """Render the gold-rule + email + version footer.

    Three rows starting at start_row:
      start_row     : gold thin-rule top border across A-L
      start_row + 1 : contact line (graphite, slate-muted)
      start_row + 2 : version line (italic gold)

    Args:
        ws: worksheet
        start_row: first row of the footer (the gold-rule row)
        version_line: defaults to "Free updates forever · v2.2".
            Override to include a release date or SKU stamp.
        contact_line: defaults to "Questions? <BRAND_EMAIL>".
    """
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=start_row, column=c).border = Border(top=gold_side)

    ws.merge_cells(f"A{start_row + 1}:L{start_row + 1}")
    c = ws[f"A{start_row + 1}"]
    c.value = contact_line or f"Questions? {BRAND_EMAIL}"
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row + 1].height = 18

    ws.merge_cells(f"A{start_row + 2}:L{start_row + 2}")
    c = ws[f"A{start_row + 2}"]
    c.value = version_line or "Free updates forever · v2.2"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row + 2].height = 18


# Chart palette — used in order for multi-series / pie slices.
# Stays within the brand tokens; never use raw hex in a build script.
CHART_PALETTE = [
    COLOR_PRIMARY,      # Harbor Navy
    COLOR_ACCENT,       # Muted Gold
    COLOR_SECONDARY,    # Clay Rose
    COLOR_NAVY_TINT,    # Navy hover
    COLOR_GOLD_SOFT,    # Soft gold
    COLOR_NAVY_SHADE,   # Navy pressed
    COLOR_PARCHMENT_ALT,# Parchment alt
]


def style_chart(chart, single_color=None):
    """Apply brand styling to an openpyxl chart in-place.

    For bar/column/line charts: tints all series in CHART_PALETTE order
    (or single_color if given). For pie/doughnut charts: sets per-slice
    colors via DataPoint records — single_color is ignored since pies
    need varied slices.

    Args:
        chart: openpyxl chart object (BarChart, PieChart, etc.)
        single_color: hex string (no #) to use for every series.
            None = rotate through CHART_PALETTE.

    Sets brand-consistent: navy axis lines, no chart border, gold gridlines
    (Excel keeps default if we don't touch them, which is fine).
    """
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import ColorChoice
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.marker import DataPoint

    chart_type = type(chart).__name__

    if chart_type in ("PieChart", "DoughnutChart", "PieChart3D"):
        # Pies/donuts: one series, color each slice differently
        if chart.series:
            series = chart.series[0]
            for i, color in enumerate(CHART_PALETTE):
                dp = DataPoint(idx=i)
                dp.graphicalProperties = GraphicalProperties(
                    solidFill=color
                )
                series.dPt.append(dp)
    else:
        # Bar/column/line: each series gets one color
        for i, series in enumerate(chart.series):
            color = single_color or CHART_PALETTE[i % len(CHART_PALETTE)]
            series.graphicalProperties = GraphicalProperties(
                solidFill=color
            )
            # Thin navy line on the series border (clean editorial look)
            series.graphicalProperties.line = LineProperties(
                solidFill=COLOR_NAVY_SHADE, w=4500
            )
