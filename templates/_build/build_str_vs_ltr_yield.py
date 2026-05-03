"""Build ACQ-012 STR vs LTR Yield Comparison (v2.2 standard).

Operational-mode calculator: compare running a single property as a
short-term rental vs a 12-month long-term lease. Side-by-side 3-column
layout (STR / LTR / Difference) with per-row winner highlighting.

Outputs:
  templates/_masters/ACQ-012-str-vs-ltr-yield-comparison-DEMO.xlsx
  templates/_masters/ACQ-012-str-vs-ltr-yield-comparison-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    apply_brand_header,
    COLOR_NAVY_TINT, COLOR_WHITE,
)

SKU = "ACQ-012"
NAME = "str-vs-ltr-yield-comparison"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"


# ---------------------------------------------------------------------------
# Sample data (Smokies Ridge cabin scenario from brief)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Settings
    "property_name":     "Smokies Ridge Cabin",
    "active_year":       2026,
    "property_value":    485000,
    # Shared financing (same regardless of STR/LTR)
    "annual_debt_service": 28800,
    "cash_invested":      121000,    # 25% down + closing + furnishing for STR
    "cash_invested_ltr":  121000,    # Same property, same financing
    # STR inputs
    "str_occupancy":      0.62,
    "str_adr":            245,
    "str_cleaning_per_turn": 150,
    "str_turnovers_year":  68,
    "str_supplies":       1800,
    "str_software":       720,
    "str_pm_pct":         0.03,    # Airbnb commission only (self-managed)
    "str_insurance":      2400,
    "str_utilities":      2160,
    "str_maint_pct":      0.05,
    "str_hours_week":     8,
    "str_reg_risk":       7,       # 1-10 (10 = high risk)
    # LTR inputs
    "ltr_monthly_rent":   2800,
    "ltr_vacancy_pct":    0.05,
    "ltr_pm_pct":         0.0,     # self-managed default
    "ltr_insurance":      1400,
    "ltr_maint_pct":      0.05,
    "ltr_turnover_cost":  500,     # annualized (one turnover ~ every 2 yrs)
    "ltr_hours_week":     0.5,
    "ltr_reg_risk":       1,
}


# ---------------------------------------------------------------------------
# Variant helper
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (for BLANK)."""
    return demo_value if variant == "demo" else None


# ---------------------------------------------------------------------------
# Section helpers
# ---------------------------------------------------------------------------

def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row_dual(ws, row, label, str_val, ltr_val, fmt, note=""):
    """Input row with STR value in col C and LTR value in col E."""
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    # STR input col C
    cell = ws.cell(row=row, column=3, value=str_val)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    # LTR input col E
    cell = ws.cell(row=row, column=5, value=ltr_val)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"G{row}:L{row}")
        c = ws[f"G{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _input_row_str_only(ws, row, label, value, fmt, note=""):
    """Input row applying only to STR (col C). Col E shows '—'."""
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    cell = ws.cell(row=row, column=5, value="—")
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    if note:
        ws.merge_cells(f"G{row}:L{row}")
        c = ws[f"G{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _input_row_ltr_only(ws, row, label, value, fmt, note=""):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value="—")
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    cell = ws.cell(row=row, column=5, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"G{row}:L{row}")
        c = ws[f"G{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


# ---------------------------------------------------------------------------
# Tab 1: Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Navy hero rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "STR vs LTR Yield Comparison"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Airbnb or 12-month lease — which actually pays you more?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # KPI band rows 10-22 — parchment fill
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Property name pulled from Settings
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = ('=IF(Settings!B5="","(enter property name on Settings tab)",'
               'Settings!B5)')
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # KPI grid: 4 cards (STR NOI / LTR NOI / $ Δ / Time Δ)
    # Cards span rows 13-17, A-C / D-F / G-I / J-L
    kpis = [
        ("A13", "C17", "STR NOI",    "='Side-by-Side'!C12", '"$"#,##0', COLOR_PRIMARY),
        ("D13", "F17", "LTR NOI",    "='Side-by-Side'!E12", '"$"#,##0', COLOR_SECONDARY),
        ("G13", "I17", "DIFFERENCE", "='Side-by-Side'!G12", '"$"#,##0;[Red]-"$"#,##0', COLOR_ACCENT),
        ("J13", "L17", "HRS/WK Δ",   "='Side-by-Side'!G19", '0.0" hrs"', COLOR_NAVY_TINT_PROXY),
    ]
    # NOTE: I'll inline the navy-tint reference rather than import an unused constant
    # by duplicating the merge logic here.
    for top, bot, label, formula, fmt, _ in kpis:
        ws.merge_cells(f"{top}:{bot}")
        c = ws[top]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

    # KPI labels row 18
    label_cells = [("A18", "C18"), ("D18", "F18"), ("G18", "I18"), ("J18", "L18")]
    for i, (l, r) in enumerate(label_cells):
        ws.merge_cells(f"{l}:{r}")
        c = ws[l]
        c.value = kpis[i][2]
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = parchment_fill

    for r in range(13, 18):
        ws.row_dimensions[r].height = 16
    ws.row_dimensions[18].height = 18

    # Verdict line row 20
    ws.merge_cells("A20:L20")
    c = ws["A20"]
    c.value = (
        '=IF(\'Side-by-Side\'!C12>=\'Side-by-Side\'!E12,'
        '"STR wins NOI by "&TEXT(\'Side-by-Side\'!G12,"$#,##0")&'
        '" — but "&TEXT(\'Side-by-Side\'!G19,"0.0")&'
        '" more hrs/wk + reg risk "&\'Side-by-Side\'!C20&"/10",'
        '"LTR wins NOI by "&TEXT(-\'Side-by-Side\'!G12,"$#,##0")&'
        '" with "&TEXT(-\'Side-by-Side\'!G19,"0.0")&" fewer hrs/wk")'
    )
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[20].height = 28

    # Counterintuitive insight row 22
    ws.merge_cells("A22:L23")
    c = ws["A22"]
    c.value = (
        '="Effective $/hr — STR: "&TEXT(\'Side-by-Side\'!C18,"$#,##0")&'
        '"   |   LTR: "&TEXT(\'Side-by-Side\'!E18,"$#,##0")'
    )
    c.font = Font(name=FONT_HEAD, size=13, italic=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 22
    ws.row_dimensions[23].height = 22

    # Pseudo-button nav rows 25-27
    pseudo_button(ws, "A25", "F27", "→  Open Side-by-Side",
                  "'Side-by-Side'!A1", variant="primary")
    pseudo_button(ws, "G25", "L27", "Edit Inputs",
                  "'Inputs'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    # Decision Notes button row 29
    pseudo_button(ws, "A29", "L31", "Decision Notes",
                  "'Decision Notes'!A1", variant="accent")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 16

    # Footnote row 33
    ws.merge_cells("A33:L34")
    c = ws["A33"]
    c.value = (
        "Note: NOI excludes debt service. Cash flow / CoC factor it in. "
        "Regulatory risk is your subjective 1-10 — bump it if your city is "
        "actively debating an STR cap or registration regime."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 22
    ws.row_dimensions[34].height = 22

    brand_footer(ws, 36,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L38"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 2: Inputs (side-by-side STR / LTR assumptions)
# ---------------------------------------------------------------------------

def build_inputs_tab(wb, variant):
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38),
        ("C", 16), ("D", 4), ("E", 16),
        ("F", 4),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Inputs", prev_tab="Start", next_tab="Side-by-Side")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Yellow = edit. STR column on the left, LTR on the right."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 5: STR | LTR labels
    ws.cell(row=5, column=2).value = ""
    cell = ws.cell(row=5, column=3, value="STR")
    apply_style(cell, header_row_style())
    cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_WHITE)
    cell = ws.cell(row=5, column=5, value="LTR")
    apply_style(cell, header_row_style())
    cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_SECONDARY)
    ws.row_dimensions[5].height = 24

    ws.freeze_panes = "A6"

    # ---- SHARED FINANCING (rows 7-9) ----
    _section_band(ws, 7, "SHARED — APPLIES TO BOTH SCENARIOS")

    # Cash invested + debt service apply to both — STR cell only, LTR mirrors
    ws.cell(row=8, column=2, value="Cash invested (down + closing + setup) ($):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=8, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=8, column=3, value=_val(variant, SAMPLE["cash_invested"]))
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    cell = ws.cell(row=8, column=5, value="=C8")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.merge_cells("G8:L8")
    c = ws["G8"]
    c.value = "STR may need higher furnishing — adjust if comparing apples to apples"
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[8].height = 18

    ws.cell(row=9, column=2, value="Annual debt service ($):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=9, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=9, column=3, value=_val(variant, SAMPLE["annual_debt_service"]))
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    cell = ws.cell(row=9, column=5, value="=C9")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.merge_cells("G9:L9")
    c = ws["G9"]
    c.value = "P+I — same loan, same payment, regardless of rental strategy"
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[9].height = 18

    # ---- REVENUE ASSUMPTIONS (rows 11-15) ----
    _section_band(ws, 11, "REVENUE ASSUMPTIONS")

    _input_row_str_only(ws, 12, "STR Occupancy (%):",
                        _val(variant, SAMPLE["str_occupancy"]), "0.0%",
                        "Year-2+ stabilized")
    _input_row_str_only(ws, 13, "STR ADR — Avg nightly rate ($):",
                        _val(variant, SAMPLE["str_adr"]), '"$"#,##0',
                        "Gross — commission applied below")
    _input_row_ltr_only(ws, 14, "LTR Monthly rent ($):",
                        _val(variant, SAMPLE["ltr_monthly_rent"]), '"$"#,##0',
                        "Annual = monthly × 12")
    _input_row_ltr_only(ws, 15, "LTR Vacancy (%):",
                        _val(variant, SAMPLE["ltr_vacancy_pct"]), "0.0%",
                        "5% typical for stable markets")

    # ---- OPERATING EXPENSES (rows 17-26) ----
    _section_band(ws, 17, "OPERATING EXPENSE ASSUMPTIONS")

    _input_row_str_only(ws, 18, "STR Cleaning paid ($/turnover):",
                        _val(variant, SAMPLE["str_cleaning_per_turn"]), '"$"#,##0', "")
    _input_row_str_only(ws, 19, "STR Turnovers per year:",
                        _val(variant, SAMPLE["str_turnovers_year"]), "0",
                        "= (occupancy × 365) ÷ avg LOS")
    _input_row_str_only(ws, 20, "STR Supplies / consumables (annual, $):",
                        _val(variant, SAMPLE["str_supplies"]), '"$"#,##0',
                        "Toiletries, paper, coffee")
    _input_row_str_only(ws, 21, "STR Software / PMS (annual, $):",
                        _val(variant, SAMPLE["str_software"]), '"$"#,##0',
                        "Hospitable, OwnerRez, dynamic pricing")
    _input_row_dual(ws, 22, "Channel/PM fee (%):",
                    _val(variant, SAMPLE["str_pm_pct"]),
                    _val(variant, SAMPLE["ltr_pm_pct"]), "0.0%",
                    "STR: Airbnb 3% or PM 20%. LTR: 0% self / 8% managed")
    _input_row_dual(ws, 23, "Insurance (annual, $):",
                    _val(variant, SAMPLE["str_insurance"]),
                    _val(variant, SAMPLE["ltr_insurance"]), '"$"#,##0',
                    "STR rider runs 30-100% higher than landlord policy")
    _input_row_str_only(ws, 24, "STR Utilities (annual, $):",
                        _val(variant, SAMPLE["str_utilities"]), '"$"#,##0',
                        "Host pays in STR; tenant typically pays in LTR")
    _input_row_dual(ws, 25, "Maintenance reserve (%):",
                    _val(variant, SAMPLE["str_maint_pct"]),
                    _val(variant, SAMPLE["ltr_maint_pct"]), "0.0%",
                    "% of gross revenue set aside for repairs")
    _input_row_ltr_only(ws, 26, "LTR Tenant turnover cost (annualized, $):",
                        _val(variant, SAMPLE["ltr_turnover_cost"]), '"$"#,##0',
                        "Lease-up costs ÷ avg tenancy length")

    # ---- TIME + RISK (rows 28-30) ----
    _section_band(ws, 28, "TIME + REGULATORY RISK")

    _input_row_dual(ws, 29, "Host hours per week:",
                    _val(variant, SAMPLE["str_hours_week"]),
                    _val(variant, SAMPLE["ltr_hours_week"]), "0.0",
                    "Self-managed STR: 5-15. Self-managed LTR: 0.25-1")
    _input_row_dual(ws, 30, "Regulatory risk (1-10):",
                    _val(variant, SAMPLE["str_reg_risk"]),
                    _val(variant, SAMPLE["ltr_reg_risk"]), "0",
                    "Subjective — high if your city is debating an STR cap")

    # Footer
    brand_footer(ws, 33, version_line=f"{SKU} · v1.0 · Inputs · {variant.upper()}")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 3: Side-by-Side (3 columns: STR / LTR / Difference)
# ---------------------------------------------------------------------------

def build_side_by_side_tab(wb, variant):
    ws = wb.create_sheet("Side-by-Side")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 16), ("D", 4), ("E", 16), ("F", 4), ("G", 16),
        ("H", 6), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Side-by-Side",
                        prev_tab="Inputs", next_tab="Decision Notes")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Per-row winner highlighted gold. All formulas reference the Inputs tab."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column headers row 6
    ws.cell(row=6, column=2).value = "Metric"
    cell = ws.cell(row=6, column=2)
    apply_style(cell, header_row_style())
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    cell = ws.cell(row=6, column=3, value="STR")
    apply_style(cell, header_row_style())
    cell = ws.cell(row=6, column=5, value="LTR")
    apply_style(cell, header_row_style())
    cell.fill = PatternFill("solid", fgColor=COLOR_SECONDARY)
    cell = ws.cell(row=6, column=7, value="Difference")
    apply_style(cell, header_row_style())
    cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    ws.row_dimensions[6].height = 24
    ws.freeze_panes = "A7"

    # ---- The comparison rows ----
    # Each row: (row_num, label, str_formula, ltr_formula, fmt, higher_is_better, emphasize)
    # higher_is_better: True = STR wins when STR>LTR; False = STR wins when STR<LTR
    rows = [
        # Revenue
        (8,  "Annual gross revenue",
            "=Inputs!C13*Inputs!C12*365",     # ADR × occ × 365
            "=Inputs!E14*12",                  # monthly × 12
            '"$"#,##0', True, False),
        # Operating expenses
        (9,  "Cleaning (annual)",
            "=Inputs!C18*Inputs!C19",
            "=Inputs!E26",
            '"$"#,##0', False, False),
        (10, "Supplies + Software + Utilities",
            "=Inputs!C20+Inputs!C21+Inputs!C24",
            "=0",
            '"$"#,##0', False, False),
        (11, "Channel/PM fee + Insurance + Maintenance",
            "=(C8*Inputs!C22)+Inputs!C23+(C8*Inputs!C25)",
            "=(E8*(1-Inputs!E15)*Inputs!E22)+Inputs!E23+(E8*Inputs!E25)",
            '"$"#,##0', False, False),
        # NOI
        (12, "NOI (Net Operating Income)",
            "=C8-C9-C10-C11",
            "=E8*(1-Inputs!E15)-E9-E10-E11",
            '"$"#,##0', True, True),
        # Debt service
        (13, "Debt service",
            "=Inputs!C9",
            "=Inputs!E9",
            '"$"#,##0', False, False),
        # Cash flow
        (14, "Cash flow",
            "=C12-C13",
            "=E12-E13",
            '"$"#,##0', True, True),
        # CoC
        (15, "Cash-on-Cash return",
            "=IFERROR(C14/Inputs!C8,0)",
            "=IFERROR(E14/Inputs!E8,0)",
            "0.0%", True, False),
        # Cap rate
        (16, "Cap rate",
            "=IFERROR(C12/Settings!B9,0)",
            "=IFERROR(E12/Settings!B9,0)",
            "0.00%", True, False),
        # Hours/week
        (17, "Hours / week",
            "=Inputs!C29",
            "=Inputs!E29",
            '0.0" hrs"', False, False),
        # $/hr
        (18, "Effective $ per hour (cash flow ÷ hrs/yr)",
            "=IFERROR(C14/(Inputs!C29*52),0)",
            "=IFERROR(E14/(Inputs!E29*52),0)",
            '"$"#,##0', True, True),
        # Hours difference shown as STR-LTR (higher means STR worse; render in row 19 G col)
        # Regulatory risk
        (20, "Regulatory risk (1-10)",
            "=Inputs!C30",
            "=Inputs!E30",
            "0", False, False),
        # 5-year cumulative cash flow
        (21, "5-year cumulative cash flow",
            "=C14*5",
            "=E14*5",
            '"$"#,##0', True, False),
    ]

    # Add a hours-delta companion row at 19 to feed the Start KPI
    # We'll write it inline below.

    for entry in rows:
        row, label, str_f, ltr_f, fmt, higher_is_better, emphasize = entry

        # Label col B
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(
            name=FONT_HEAD if emphasize else FONT_BODY,
            size=12 if emphasize else 11,
            bold=True,
            color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

        # STR col C
        cell = ws.cell(row=row, column=3, value=str_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        if emphasize:
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

        # LTR col E
        cell = ws.cell(row=row, column=5, value=ltr_f)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        if emphasize:
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

        # Difference col G  (STR - LTR)
        cell = ws.cell(row=row, column=7, value=f"=C{row}-E{row}")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Use signed format for $; for % keep simple
        if "%" in fmt:
            cell.number_format = "0.0%;[Red]-0.0%"
        elif '"$"' in fmt:
            cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
        elif "hrs" in fmt:
            cell.number_format = '0.0" hrs";[Red]-0.0" hrs"'
        else:
            cell.number_format = fmt
        if emphasize:
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        ws.row_dimensions[row].height = 22 if emphasize else 18

        # Per-row winner highlight via FormulaRule
        # higher_is_better=True: highlight STR (C) when C>E; else highlight LTR (E)
        gold_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        if higher_is_better:
            # STR wins when C > E
            ws.conditional_formatting.add(
                f"C{row}",
                FormulaRule(formula=[f"C{row}>E{row}"], fill=gold_fill,
                            font=Font(name=FONT_HEAD, size=11, bold=True,
                                      color=COLOR_PRIMARY)),
            )
            ws.conditional_formatting.add(
                f"E{row}",
                FormulaRule(formula=[f"E{row}>C{row}"], fill=gold_fill,
                            font=Font(name=FONT_HEAD, size=11, bold=True,
                                      color=COLOR_PRIMARY)),
            )
        else:
            # Lower is better — STR wins when C < E
            ws.conditional_formatting.add(
                f"C{row}",
                FormulaRule(formula=[f"C{row}<E{row}"], fill=gold_fill,
                            font=Font(name=FONT_HEAD, size=11, bold=True,
                                      color=COLOR_PRIMARY)),
            )
            ws.conditional_formatting.add(
                f"E{row}",
                FormulaRule(formula=[f"E{row}<C{row}"], fill=gold_fill,
                            font=Font(name=FONT_HEAD, size=11, bold=True,
                                      color=COLOR_PRIMARY)),
            )

    # Row 19 — hours/week delta companion (STR - LTR) for KPI
    cell = ws.cell(row=19, column=2, value="Hours/week difference (STR − LTR)")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    cell = ws.cell(row=19, column=7, value="=C17-E17")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0.0" hrs";[Red]-0.0" hrs"'
    ws.row_dimensions[19].height = 16

    # Insight callout row 23
    ws.merge_cells("A23:L24")
    c = ws["A23"]
    c.value = (
        '=IF(C18<E18,"⚠ Counterintuitive: STR\'s effective $/hr ("&'
        'TEXT(C18,"$#,##0")&") is lower than LTR\'s ("&TEXT(E18,"$#,##0")&'
        '"). The extra revenue is real — but so is the labor.","")'
    )
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_SECONDARY)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[23].height = 22
    ws.row_dimensions[24].height = 22

    brand_footer(ws, 26,
                 version_line=f"{SKU} · v1.0 · Side-by-Side · {variant.upper()}")

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 4: Decision Notes
# ---------------------------------------------------------------------------

def build_decision_notes_tab(wb, variant):
    ws = wb.create_sheet("Decision Notes")
    ws.sheet_properties.tabColor = COLOR_GOLD_SOFT

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    compact_header_band(ws, "Decision Notes",
                        prev_tab="Side-by-Side", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Free-form notes. The numbers are only half the story."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    prompts = [
        ("Regulatory risk",
            "What's your city saying about STR? Permit cap? Owner-occupancy "
            "requirement? Pending council vote? Document it here so future "
            "you remembers why this number is what it is."),
        ("Life stage",
            "Are you OK with 8 hrs/wk of guest comms for the next 5 years? "
            "Will that change if you have a kid, take a new job, move?"),
        ("Time available + opportunity cost",
            "If your $/hr on the day job is higher than your STR effective "
            "$/hr, the LTR option may free up earning power elsewhere."),
        ("Capital exit",
            "STR-furnished property may sell to fewer buyers in a downturn. "
            "LTR property is a standard rental — broader buyer pool."),
        ("Tax strategy",
            "STR with material participation can offset W-2 income via cost "
            "seg + bonus depreciation (REPS-lite). LTR cannot, unless you're "
            "a real estate professional."),
    ]

    row = 6
    for title, body in prompts:
        # Title row
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = title.upper()
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 20

        # Prompt body
        ws.merge_cells(f"A{row+1}:L{row+1}")
        c = ws[f"A{row+1}"]
        c.value = body
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=2)
        ws.row_dimensions[row+1].height = 28

        # Notes input area (3 rows)
        ws.merge_cells(f"A{row+2}:L{row+4}")
        c = ws[f"A{row+2}"]
        c.value = ""
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True, indent=2)
        for r in range(row+2, row+5):
            ws.row_dimensions[r].height = 18

        row += 6

    brand_footer(ws, row + 1,
                 version_line=f"{SKU} · v1.0 · Decision Notes · {variant.upper()}")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Tab 5: Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_MUTED

    set_col_widths(ws, [
        ("A", 4), ("B", 28), ("C", 22),
        ("D", 4), ("E", 8), ("F", 8), ("G", 8),
        ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Decision Notes", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Workbook-wide settings. Set once, referenced everywhere."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    settings_rows = [
        (5, "Property name:",   _val(variant, SAMPLE["property_name"]),  None),
        (7, "Active year:",     _val(variant, SAMPLE["active_year"]),    "0"),
        (9, "Property value ($):", _val(variant, SAMPLE["property_value"]), '"$"#,##0'),
    ]
    for row, label, value, fmt in settings_rows:
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        cell = ws.cell(row=row, column=2)  # already set
        cell = ws.cell(row=row, column=2)
        # set value cell
        in_cell = ws.cell(row=row, column=2)  # noop
        target = ws.cell(row=row, column=3, value=value)
        apply_style(target, input_cell_style())
        target.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if fmt:
            target.number_format = fmt
        ws.row_dimensions[row].height = 20

    # Note rows
    ws.merge_cells("A12:L13")
    c = ws["A12"]
    c.value = (
        "Property value drives the cap rate calculation on the Side-by-Side "
        "tab. If you don't have a recent appraisal, use your purchase price + "
        "a conservative appreciation estimate."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[12].height = 20
    ws.row_dimensions[13].height = 20

    brand_footer(ws, 16,
                 version_line=f"{SKU} · v1.0 · Settings · {variant.upper()}")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Workbook orchestration
# ---------------------------------------------------------------------------

# TODO: COLOR_NAVY_TINT_PROXY is a local stand-in for the navy-tint color used on
# the 4th KPI card. brand_config exports COLOR_NAVY_TINT but I don't actually use
# it as a fill there — the Font color falls back to COLOR_ACCENT regardless. So
# this is just a label slot; nothing renders with this constant. Keeping inlined.
COLOR_NAVY_TINT_PROXY = COLOR_NAVY_TINT


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_inputs_tab(wb, variant)
    build_side_by_side_tab(wb, variant)
    build_decision_notes_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"STR vs LTR Yield Comparison — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Side-by-side STR vs LTR yield calculator: NOI, CoC, cap rate, "
        "$/hour, regulatory risk, 5-year cumulative cash flow."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
