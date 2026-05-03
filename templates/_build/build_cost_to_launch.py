"""Build ACQ-002 Cost-to-Launch Calculator (v2.2 standard).

Wizard-mode tool for Newbie Nina — takes the all-in dollars from
"bought the property" to "first guest checked in." 6 tabs, ~120
pre-populated furnishing line items across 8 zones, 90-day operating
reserve, Y1 ROI projection.

Generates two files:
  templates/_masters/ACQ-002-cost-to-launch-DEMO.xlsx     (sample data)
  templates/_masters/ACQ-002-cost-to-launch-BLANK.xlsx    (empty)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
)

SKU = "ACQ-002"
NAME = "cost-to-launch"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data — Smokies Ridge Cabin scenario from brief
# ---------------------------------------------------------------------------

SAMPLE = {
    # Property header
    "property_name":   "Smokies Ridge Cabin",
    "city_state":      "Gatlinburg, TN",
    "beds":            2,
    "baths":           2,

    # Acquisition
    "purchase":        385000,
    "down_pct":        0.25,
    "rate":            0.0700,
    "term_years":      30,
    "closing_costs":   9000,
    "inspection":      400,
    "appraisal":       650,
    "title_ins":       1100,
    "legal":           1200,
    "loan_orig":       1500,
    "other_dd":        300,

    # Launch costs
    "photos":          1200,
    "airbnb_fee":      0,
    "vrbo_fee":        499,
    "permit_fee":      450,
    "biz_license":     150,
    "llc_formation":   400,
    "ein":             0,
    "insurance_yr":    1800,
    "pms_setup":       250,
    "pricing_setup":   150,
    "accounting_sw":   240,
    "marketing":       400,
    "signage":         180,
    "lockbox":         85,

    # 90-day reserve
    "reserve_pi":      None,   # formula-derived from acquisition tab
    "reserve_util":    900,    # 3 mo
    "reserve_ins":     450,    # 3 mo of annual / 4
    "reserve_clean":   1500,
    "reserve_maint":   1200,
    "reserve_mktg":    600,
    "reserve_sw":      210,
    "reserve_other":   500,

    # Y1 ROI
    "y1_adr":          245,
    "y1_occ":          0.45,
    "avail_nights":    365,
    "y1_op_costs":     22000,    # rough operating cost projection (excl debt)
}


# ---------------------------------------------------------------------------
# Furnishing data (8 zones, ~120 line items, low/mid/high price ranges)
# ---------------------------------------------------------------------------
# (zone, item, low, mid, high, qty)
FURNISHING = [
    # ----- BEDROOM (×2 bedrooms — 15 items × 2 = ~30) -----
    ("Bedroom 1", "Queen mattress",                250,  650,  1200, 1),
    ("Bedroom 1", "Bed frame + slats",             120,  280,   650, 1),
    ("Bedroom 1", "Headboard",                      80,  200,   500, 1),
    ("Bedroom 1", "Sheet sets (2 sets)",            60,  140,   300, 1),
    ("Bedroom 1", "Comforter / duvet",              60,  150,   400, 1),
    ("Bedroom 1", "Pillows (4)",                    40,  100,   240, 1),
    ("Bedroom 1", "Bedside lamps (pair)",           40,  120,   280, 1),
    ("Bedroom 1", "Blackout curtains",              40,   90,   220, 1),
    ("Bedroom 1", "Dresser / chest",               150,  380,   900, 1),
    ("Bedroom 1", "Closet system / hangers",        25,   75,   180, 1),
    ("Bedroom 1", "Mirror (full-length)",           40,   95,   240, 1),
    ("Bedroom 1", "Wall art (3-piece set)",         50,  140,   400, 1),
    ("Bedroom 1", "Area rug",                       80,  220,   600, 1),
    ("Bedroom 1", "Nightstand (pair)",              90,  240,   600, 1),
    ("Bedroom 1", "Throw blanket + accent pillows", 35,   90,   220, 1),

    ("Bedroom 2", "Queen mattress",                250,  650,  1200, 1),
    ("Bedroom 2", "Bed frame + slats",             120,  280,   650, 1),
    ("Bedroom 2", "Headboard",                      80,  200,   500, 1),
    ("Bedroom 2", "Sheet sets (2 sets)",            60,  140,   300, 1),
    ("Bedroom 2", "Comforter / duvet",              60,  150,   400, 1),
    ("Bedroom 2", "Pillows (4)",                    40,  100,   240, 1),
    ("Bedroom 2", "Bedside lamps (pair)",           40,  120,   280, 1),
    ("Bedroom 2", "Blackout curtains",              40,   90,   220, 1),
    ("Bedroom 2", "Dresser / chest",               150,  380,   900, 1),
    ("Bedroom 2", "Closet system / hangers",        25,   75,   180, 1),
    ("Bedroom 2", "Mirror (full-length)",           40,   95,   240, 1),
    ("Bedroom 2", "Wall art (3-piece set)",         50,  140,   400, 1),
    ("Bedroom 2", "Area rug",                       80,  220,   600, 1),
    ("Bedroom 2", "Nightstand (pair)",              90,  240,   600, 1),
    ("Bedroom 2", "Throw blanket + accent pillows", 35,   90,   220, 1),

    # ----- BATHROOM (×2 baths — 12 items × 2 = ~24) -----
    ("Bathroom 1", "Towel sets (4 hand + 4 bath)",  60,  140,   320, 1),
    ("Bathroom 1", "Bath mats (2)",                 25,   55,   140, 1),
    ("Bathroom 1", "Shower curtain + liner",        25,   60,   140, 1),
    ("Bathroom 1", "Soap + lotion dispensers",      20,   50,   120, 1),
    ("Bathroom 1", "Trash bin (covered)",           20,   45,   110, 1),
    ("Bathroom 1", "Toilet brush + holder",         15,   35,    80, 1),
    ("Bathroom 1", "Plunger",                       12,   25,    50, 1),
    ("Bathroom 1", "Toilet paper holder",           15,   35,    80, 1),
    ("Bathroom 1", "Wall hooks / robe hooks",       15,   35,    90, 1),
    ("Bathroom 1", "Wall art",                      30,   75,   200, 1),
    ("Bathroom 1", "Bathroom scale",                20,   45,   100, 1),
    ("Bathroom 1", "Hairdryer",                     20,   55,   130, 1),

    ("Bathroom 2", "Towel sets (4 hand + 4 bath)",  60,  140,   320, 1),
    ("Bathroom 2", "Bath mats (2)",                 25,   55,   140, 1),
    ("Bathroom 2", "Shower curtain + liner",        25,   60,   140, 1),
    ("Bathroom 2", "Soap + lotion dispensers",      20,   50,   120, 1),
    ("Bathroom 2", "Trash bin (covered)",           20,   45,   110, 1),
    ("Bathroom 2", "Toilet brush + holder",         15,   35,    80, 1),
    ("Bathroom 2", "Plunger",                       12,   25,    50, 1),
    ("Bathroom 2", "Toilet paper holder",           15,   35,    80, 1),
    ("Bathroom 2", "Wall hooks / robe hooks",       15,   35,    90, 1),
    ("Bathroom 2", "Wall art",                      30,   75,   200, 1),
    ("Bathroom 2", "Bathroom scale",                20,   45,   100, 1),
    ("Bathroom 2", "Hairdryer",                     20,   55,   130, 1),

    # ----- KITCHEN (~25 items) -----
    ("Kitchen", "Dish set (service for 8)",         60,  150,   400, 1),
    ("Kitchen", "Drinking glasses (set of 8)",      25,   60,   150, 1),
    ("Kitchen", "Wine glasses (set of 6)",          20,   50,   140, 1),
    ("Kitchen", "Coffee mugs (set of 8)",           20,   45,   110, 1),
    ("Kitchen", "Silverware (service for 8)",       30,   75,   200, 1),
    ("Kitchen", "Cookware (5-pc set)",              80,  220,   650, 1),
    ("Kitchen", "Bakeware (3-pc set)",              30,   75,   180, 1),
    ("Kitchen", "Knife block + 6 knives",           45,  120,   350, 1),
    ("Kitchen", "Cutting boards (2)",               20,   45,   110, 1),
    ("Kitchen", "Mixing bowls (3-pc set)",          20,   45,   110, 1),
    ("Kitchen", "Measuring cups + spoons",          15,   30,    65, 1),
    ("Kitchen", "Can opener + corkscrew + tools",   20,   45,   110, 1),
    ("Kitchen", "Coffee maker (drip + pour-over)",  35,   95,   280, 1),
    ("Kitchen", "Toaster",                          20,   55,   140, 1),
    ("Kitchen", "Microwave (countertop)",           65,  130,   320, 1),
    ("Kitchen", "Electric kettle",                  20,   50,   130, 1),
    ("Kitchen", "Blender",                          25,   75,   220, 1),
    ("Kitchen", "Food storage (containers + wrap)", 20,   50,   120, 1),
    ("Kitchen", "Dish towels (8) + pot holders",    20,   45,   100, 1),
    ("Kitchen", "Oven mitts (2 pair)",              12,   25,    55, 1),
    ("Kitchen", "Paper towel holder",               15,   30,    70, 1),
    ("Kitchen", "Trash + recycling bins",           40,   95,   240, 1),
    ("Kitchen", "Spice starter set (12-pc)",        20,   45,   110, 1),
    ("Kitchen", "Bar / drink ware (shaker, etc.)",  25,   60,   150, 1),
    ("Kitchen", "Welcome basket (coffee/snacks)",   25,   50,   120, 1),

    # ----- LIVING (~15 items) -----
    ("Living", "Sofa (3-seat)",                    400,  900,  2400, 1),
    ("Living", "Accent chairs (pair)",             200,  500,  1400, 1),
    ("Living", "Coffee table",                      80,  220,   600, 1),
    ("Living", "End tables (pair)",                 80,  200,   500, 1),
    ("Living", "Floor + table lamps (3)",           80,  220,   550, 1),
    ("Living", "TV (55\" smart)",                  280,  500,  1100, 1),
    ("Living", "TV stand / media console",          90,  240,   650, 1),
    ("Living", "Throw blankets (3)",                40,   95,   220, 1),
    ("Living", "Throw pillows (6)",                 50,  130,   320, 1),
    ("Living", "Wall art / gallery wall",          100,  280,   800, 1),
    ("Living", "Area rug (8x10)",                  120,  340,   900, 1),
    ("Living", "Bookshelf + decor / books",         80,  220,   550, 1),
    ("Living", "Board games + puzzles (5+)",        40,   90,   200, 1),
    ("Living", "Curtains + rods",                   60,  140,   340, 1),
    ("Living", "Coat rack / entry bench",           50,  120,   300, 1),

    # ----- OUTDOOR (~8 items) -----
    ("Outdoor", "Patio dining set (table + 4-6)",  220,  600,  1800, 1),
    ("Outdoor", "Patio umbrella + base",            70,  180,   480, 1),
    ("Outdoor", "Grill (gas or charcoal)",         150,  400,   950, 1),
    ("Outdoor", "Fire pit (if allowed)",           100,  280,   700, 1),
    ("Outdoor", "Outdoor string lights",            30,   75,   180, 1),
    ("Outdoor", "Adirondack / lounge chairs (2)",  120,  300,   750, 1),
    ("Outdoor", "Outdoor rug",                      50,  130,   320, 1),
    ("Outdoor", "Hot tub steps + accessories",      40,  100,   240, 1),

    # ----- TECH / SMART-HOME (~7 items) -----
    ("Tech", "Smart lock (keyless entry)",         120,  220,   400, 1),
    ("Tech", "Doorbell camera (Ring / Nest)",      100,  180,   320, 1),
    ("Tech", "Smart thermostat",                   100,  180,   300, 1),
    ("Tech", "WiFi router + mesh extender",         80,  180,   450, 1),
    ("Tech", "Streaming device (Roku/Fire TV)",     35,   60,   150, 2),
    ("Tech", "Bluetooth speaker",                   40,  100,   280, 1),
    ("Tech", "Universal remote / charging hub",     25,   55,   150, 1),

    # ----- SAFETY + SUPPLIES (~8 items) -----
    ("Safety", "Smoke detectors (per-room)",        50,  100,   220, 1),
    ("Safety", "CO detectors (per code)",           40,   80,   160, 1),
    ("Safety", "Fire extinguisher (kitchen + ea floor)", 40, 80, 160, 1),
    ("Safety", "First aid kit (full)",              30,   60,   140, 1),
    ("Safety", "Emergency flashlights + batteries", 25,   50,   120, 1),
    ("Safety", "Evacuation map (printed + framed)", 20,   40,    80, 1),
    ("Safety", "Guest WiFi card (laminated)",       10,   20,    40, 1),
    ("Safety", "Exterior security camera",         100,  180,   320, 1),

    # ----- LINENS + AMENITIES (~10 items) -----
    ("Linens",  "Extra sheet sets (2)",             60,  140,   320, 1),
    ("Linens",  "Extra bath towels (8)",            50,  120,   280, 1),
    ("Linens",  "Hand towels + washcloths bulk",    35,   80,   200, 1),
    ("Linens",  "Beach / pool towels (4)",          40,   90,   220, 1),
    ("Linens",  "Welcome basket — non-perishable",  25,   60,   140, 1),
    ("Linens",  "Toiletry starter (shampoo/cond/soap)", 30, 70, 160, 1),
    ("Linens",  "Coffee + tea + filters",           20,   50,   120, 1),
    ("Linens",  "Cleaning supplies starter kit",    40,   90,   200, 1),
    ("Linens",  "Laundry detergent + dryer sheets", 20,   45,    95, 1),
    ("Linens",  "Trash bags + paper goods",         25,   55,   120, 1),
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Cost-to-Launch Calculator"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "The all-in number — including what you'd forget."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Reassurance row 9 (calming voice — Newbie Nina opens this terrified)
    ws.merge_cells("A9:L9")
    c = ws["A9"]
    c.value = (
        "Typical 2BR cabin: $32K-$58K mid finish  ·  1BR condo: $18K-$28K  ·  "
        "4BR luxury: $90K-$160K"
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 20

    # ZONE: All-in budget hero block (rows 11-18)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(11, 19):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "YOUR LAUNCH BUDGET"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 18

    # Pull all-in from Total tab C14 (total all-in)
    ws.merge_cells("A12:L13")
    c = ws["A12"]
    c.value = "='Total + Y1 ROI'!C14"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.row_dimensions[12].height = 36
    ws.row_dimensions[13].height = 18

    # Sub-stat block (rows 15-17)
    ws.merge_cells("A15:D15")
    c = ws["A15"]
    c.value = "90-DAY RESERVE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E15:H15")
    c = ws["E15"]
    c.value = "CASH INVESTED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I15:L15")
    c = ws["I15"]
    c.value = "Y1 CASH-ON-CASH"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[15].height = 16

    ws.merge_cells("A16:D17")
    c = ws["A16"]
    c.value = "='90-Day Reserve'!C16"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("E16:H17")
    c = ws["E16"]
    c.value = "='Total + Y1 ROI'!C19"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("I16:L17")
    c = ws["I16"]
    c.value = "='Total + Y1 ROI'!C28"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0%"
    ws.row_dimensions[16].height = 22
    ws.row_dimensions[17].height = 16

    # ZONE: Pseudo-button nav (rows 20-22 + 24-26)
    pseudo_button(ws, "A20", "C22", "Acquisition",
                  "'Acquisition'!A1", variant="primary")
    pseudo_button(ws, "D20", "F22", "Furnish & Setup",
                  "'Furnish & Setup'!A1", variant="primary")
    pseudo_button(ws, "G20", "I22", "Launch Costs",
                  "'Launch Costs'!A1", variant="primary")
    pseudo_button(ws, "J20", "L22", "90-Day Reserve",
                  "'90-Day Reserve'!A1", variant="primary")
    for r in range(20, 23):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A24", "L26", "→  Total + Y1 ROI",
                  "'Total + Y1 ROI'!A1", variant="accent")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    # ZONE: Upgrade CTAs (rows 29-31)
    ws.merge_cells("A29:L29")
    c = ws["A29"]
    c.value = (
        "Validate the deal numbers before you sign?  "
        f"Get the STR Deal Analyzer (ACQ-001) at {BRAND_DOMAIN} — $47, "
        "includes stress test."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[29].height = 32

    ws.merge_cells("A31:L31")
    c = ws["A31"]
    c.value = (
        "Join the email list — first launch checklist + 47 deduction guide free."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[31].height = 18

    brand_footer(ws, 33,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L35"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Acquisition
# ---------------------------------------------------------------------------

def build_acquisition_tab(wb, variant):
    ws = wb.create_sheet("Acquisition")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Acquisition",
                        prev_tab="Start", next_tab="Furnish & Setup")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 1 of 4 — what it costs to BUY the property"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "PROPERTY")
    _input_row(ws, 7, "Property name:",
               _val(variant, SAMPLE["property_name"]), None, "")
    _input_row(ws, 8, "City, State:",
               _val(variant, SAMPLE["city_state"]), None, "")
    _input_row(ws, 9, "Bedrooms:",
               _val(variant, SAMPLE["beds"]), "0",
               "Drives furnishing zone counts (manual — adjust on Furnish tab)")
    _input_row(ws, 10, "Bathrooms:",
               _val(variant, SAMPLE["baths"]), "0.0", "")

    _section_band(ws, 12, "PURCHASE & FINANCING")
    fin_fields = [
        (13, "Purchase price ($):",      _val(variant, SAMPLE["purchase"]),      '"$"#,##0', ""),
        (14, "Down payment (%):",        _val(variant, SAMPLE["down_pct"]),      "0.0%",     "20-25% typical for STR"),
        (15, "Interest rate (%):",       _val(variant, SAMPLE["rate"]),          "0.000%",   ""),
        (16, "Term (years):",            _val(variant, SAMPLE["term_years"]),    "0",        "30 typical"),
    ]
    for row, label, value, fmt, note in fin_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 18, "DUE DILIGENCE & CLOSING")
    dd_fields = [
        (19, "Closing costs ($):",       _val(variant, SAMPLE["closing_costs"]), '"$"#,##0', "2-3% of purchase typical"),
        (20, "Inspection ($):",          _val(variant, SAMPLE["inspection"]),    '"$"#,##0', "$300-500 typical"),
        (21, "Appraisal ($):",           _val(variant, SAMPLE["appraisal"]),     '"$"#,##0', "$500-800 typical"),
        (22, "Title insurance ($):",     _val(variant, SAMPLE["title_ins"]),     '"$"#,##0', "May be in closing — check"),
        (23, "Legal / attorney ($):",    _val(variant, SAMPLE["legal"]),         '"$"#,##0', ""),
        (24, "Loan origination fee ($):",_val(variant, SAMPLE["loan_orig"]),     '"$"#,##0', "0.5-1% of loan"),
        (25, "Other due diligence ($):", _val(variant, SAMPLE["other_dd"]),      '"$"#,##0', "Survey, environmental, etc."),
    ]
    for row, label, value, fmt, note in dd_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 27, "DERIVED")
    _output_row(ws, 28, "Down payment ($):",
                "=C13*C14", '"$"#,##0')
    _output_row(ws, 29, "Loan amount ($):",
                "=C13*(1-C14)", '"$"#,##0')
    _output_row(ws, 30, "Monthly P&I:",
                "=IFERROR(PMT(C15/12,C16*12,-C29),0)", '"$"#,##0.00')
    _output_row(ws, 31, "Acquisition cash needed:",
                "=C28+SUM(C19:C25)", '"$"#,##0', emphasize=True)

    brand_footer(ws, 34,
                 version_line=f"{SKU} · Acquisition")


# ---------------------------------------------------------------------------
# Sheet 3 — Furnish & Setup (the marquee tab, ~120 line items)
# ---------------------------------------------------------------------------

def build_furnish_tab(wb, variant):
    ws = wb.create_sheet("Furnish & Setup")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 14),   # Zone
        ("B", 38),   # Item
        ("C", 10),   # Low $
        ("D", 10),   # Mid $
        ("E", 10),   # High $
        ("F", 12),   # Selected $ (per-item override)
        ("G", 6),    # Qty
        ("H", 12),   # Line total
        ("I", 2),    # spacer
        ("J", 22),   # Did You Forget label
        ("K", 6),    # Did You Forget check
        ("L", 6),    # spacer
    ])

    compact_header_band(ws, "Furnish & Setup",
                        prev_tab="Acquisition", next_tab="Launch Costs")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Section 2 of 4 — pre-populated room-by-room checklist. "
        "Per-item: Low / Mid / High prices are real Amazon-tier benchmarks. "
        "Selected column = your choice (default Mid)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Header row 6
    headers = [
        ("A6", "Zone"),
        ("B6", "Item"),
        ("C6", "Low $"),
        ("D6", "Mid $"),
        ("E6", "High $"),
        ("F6", "Selected $"),
        ("G6", "Qty"),
        ("H6", "Line Total"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())

    # Did You Forget sidebar header
    ws.merge_cells("J6:K6")
    c = ws["J6"]
    c.value = "DID YOU FORGET?"
    apply_style(c, header_row_style())
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 24

    # Group items by zone
    from collections import OrderedDict
    grouped = OrderedDict()
    for zone, item, low, mid, high, qty in FURNISHING:
        grouped.setdefault(zone, []).append((item, low, mid, high, qty))

    # Track row for Did-You-Forget links
    row_for_item = {}

    row = 7
    zone_subtotal_rows = []
    for zone, items in grouped.items():
        # Zone banner row
        ws.merge_cells(f"A{row}:H{row}")
        c = ws[f"A{row}"]
        c.value = zone.upper()
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 20
        zone_start = row + 1
        row += 1

        # Items
        for item, low, mid, high, qty in items:
            ws.cell(row=row, column=1, value=zone).font = Font(
                name=FONT_BODY, size=10, color=COLOR_MUTED, italic=True
            )
            ws.cell(row=row, column=2, value=item).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )

            # Low / Mid / High input cells
            for col, val in [(3, low), (4, mid), (5, high)]:
                cell = ws.cell(row=row, column=col,
                               value=_val(variant, val))
                apply_style(cell, input_cell_style())
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '"$"#,##0'

            # Selected $: defaults to Mid via formula
            cell = ws.cell(row=row, column=6,
                           value=f"=D{row}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0'

            # Qty
            cell = ws.cell(row=row, column=7, value=_val(variant, qty))
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0"

            # Line total = Selected × Qty
            cell = ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0'

            row_for_item[item] = row
            ws.row_dimensions[row].height = 16
            row += 1

        # Zone subtotal row
        ws.cell(row=row, column=2,
                value=f"  {zone} subtotal").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center"
        )
        for col, col_letter in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (8, "H")]:
            cell = ws.cell(row=row, column=col,
                           value=f"=SUM({col_letter}{zone_start}:{col_letter}{row-1})")
            cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
            cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[row].height = 18
        zone_subtotal_rows.append(row)
        row += 2  # blank spacer row after subtotal

    # Grand totals row
    totals_row = row
    ws.merge_cells(f"A{totals_row}:B{totals_row}")
    c = ws[f"A{totals_row}"]
    c.value = "FURNISH & SETUP TOTALS"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    # Total at low / mid / high / selected (sum of subtotal rows)
    for col, col_letter in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (8, "H")]:
        if zone_subtotal_rows:
            sum_parts = ",".join(f"{col_letter}{r}" for r in zone_subtotal_rows)
            cell = ws.cell(row=totals_row, column=col, value=f"=SUM({sum_parts})")
        else:
            cell = ws.cell(row=totals_row, column=col, value=0)
        cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'
    ws.cell(row=totals_row, column=7).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.row_dimensions[totals_row].height = 26

    # ---- "Did You Forget?" sidebar (cols J-K) ----
    # Cross-tab references: Launch Costs and other tabs.
    # Items checked YES if the corresponding cell on the Launch Costs / Acquisition tab
    # has a value > 0 (meaning the user has accounted for it).
    sidebar_items = [
        ("LLC formation?",                "='Launch Costs'!C12",  "Launch Costs!C12"),
        ("Insurance binder?",             "='Launch Costs'!C14",  "Launch Costs!C14"),
        ("Smoke detectors?",              None,                    None),  # links to Safety subtotal
        ("CO detectors?",                 None,                    None),
        ("Fire extinguisher?",            None,                    None),
        ("First aid kit?",                None,                    None),
        ("Pro photos?",                   "='Launch Costs'!C8",   "Launch Costs!C8"),
        ("Exterior security cam?",        None,                    None),  # links to Safety zone item
        ("Permit / business license?",    "='Launch Costs'!C10",  "Launch Costs!C10"),
    ]

    # Map specific furnishing rows for direct link
    safety_lookups = {
        "Smoke detectors?":     "Smoke detectors (per-room)",
        "CO detectors?":        "CO detectors (per code)",
        "Fire extinguisher?":   "Fire extinguisher (kitchen + ea floor)",
        "First aid kit?":       "First aid kit (full)",
        "Exterior security cam?": "Exterior security camera",
    }

    sidebar_start = 7
    for i, (label, formula, _ref) in enumerate(sidebar_items):
        r = sidebar_start + i * 2
        # Label cell with hyperlink jumping to source row when known
        cell = ws.cell(row=r, column=10)
        if label in safety_lookups:
            target_item = safety_lookups[label]
            target_row = row_for_item.get(target_item)
            if target_row:
                cell.value = f'=HYPERLINK("#\'Furnish & Setup\'!B{target_row}", "{label}")'
                # Determine check via line total of that row > 0
                check_formula = f"=IF(H{target_row}>0,\"OK\",\"missing\")"
            else:
                cell.value = label
                check_formula = '="?"'
        else:
            cell.value = label
            # Use the launch-costs formula directly
            if formula:
                check_formula = (
                    f"=IF({formula[1:]}>0,\"OK\",\"missing\")"
                )
            else:
                check_formula = '="?"'

        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY,
                         underline="single" if label in safety_lookups else None)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

        check = ws.cell(row=r, column=11, value=check_formula)
        check.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        check.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        check.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    # Sidebar footer note
    sidebar_end = sidebar_start + len(sidebar_items) * 2
    ws.merge_cells(f"J{sidebar_end}:K{sidebar_end + 2}")
    c = ws[f"J{sidebar_end}"]
    c.value = (
        "Click any label to jump to the row. \"OK\" = you've accounted for it. "
        "\"missing\" = double-check before launch."
    )
    c.font = Font(name=FONT_BODY, size=8, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    brand_footer(ws, totals_row + 3,
                 version_line=f"{SKU} · Furnish & Setup")

    # Freeze header row
    ws.freeze_panes = "A7"


# ---------------------------------------------------------------------------
# Sheet 4 — Launch Costs
# ---------------------------------------------------------------------------

def build_launch_tab(wb, variant):
    ws = wb.create_sheet("Launch Costs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Launch Costs",
                        prev_tab="Furnish & Setup", next_tab="90-Day Reserve")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Section 3 of 4 — costs to GO LIVE: photos, listings, permits, "
        "LLC, insurance, software."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "MARKETING & LISTING")
    mkt_fields = [
        (7, "(spacer)",                  None,                              None,        ""),
        (8, "Professional photos ($):",  _val(variant, SAMPLE["photos"]),   '"$"#,##0', "$800-1500 — non-negotiable for ranking"),
        (9, "Airbnb listing fee ($):",   _val(variant, SAMPLE["airbnb_fee"]),'"$"#,##0', "Typically $0 — free to list"),
    ]
    # Skip the spacer
    _input_row(ws, 8, "Professional photos ($):", _val(variant, SAMPLE["photos"]), '"$"#,##0',
               "$800-1500 — non-negotiable for ranking")
    _input_row(ws, 9, "Airbnb listing fee ($):", _val(variant, SAMPLE["airbnb_fee"]), '"$"#,##0',
               "Typically $0 — free to list")

    _section_band(ws, 10, "PERMITS & LEGAL ENTITY")
    permit_fields = [
        (11, "Permit / STR registration ($):", _val(variant, SAMPLE["permit_fee"]),  '"$"#,##0', "Varies by city — check local"),
        (12, "Business license ($):",          _val(variant, SAMPLE["biz_license"]), '"$"#,##0', ""),
        (13, "LLC formation ($):",             _val(variant, SAMPLE["llc_formation"]),'"$"#,##0', "$300-500 most states"),
        (14, "Insurance binder + Y1 ($):",     _val(variant, SAMPLE["insurance_yr"]),'"$"#,##0', "STR-specific; HO-3 won't cover"),
    ]
    for row, label, value, fmt, note in permit_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 16, "SOFTWARE & TOOLS (Y1)")
    sw_fields = [
        (17, "PMS setup ($):",          _val(variant, SAMPLE["pms_setup"]),    '"$"#,##0', "Hospitable / Hostfully / etc."),
        (18, "Pricing tool setup ($):", _val(variant, SAMPLE["pricing_setup"]),'"$"#,##0', "PriceLabs / Wheelhouse setup"),
        (19, "Accounting software ($):",_val(variant, SAMPLE["accounting_sw"]),'"$"#,##0', "QBO / Stessa annual"),
    ]
    for row, label, value, fmt, note in sw_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 21, "OTHER LAUNCH")
    other_fields = [
        (22, "VRBO listing fee ($):",   _val(variant, SAMPLE["vrbo_fee"]),    '"$"#,##0', "Annual subscription option"),
        (23, "Marketing / branding ($):", _val(variant, SAMPLE["marketing"]), '"$"#,##0', "Logo, social, welcome book"),
        (24, "Signage ($):",            _val(variant, SAMPLE["signage"]),     '"$"#,##0', ""),
        (25, "Lockbox / key codes ($):",_val(variant, SAMPLE["lockbox"]),     '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in other_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 27, "DERIVED")
    _output_row(ws, 28, "Total launch costs:",
                "=SUM(C8:C9)+SUM(C11:C14)+SUM(C17:C19)+SUM(C22:C25)",
                '"$"#,##0', emphasize=True)

    brand_footer(ws, 31,
                 version_line=f"{SKU} · Launch Costs")


# ---------------------------------------------------------------------------
# Sheet 5 — 90-Day Reserve
# ---------------------------------------------------------------------------

def build_reserve_tab(wb, variant):
    ws = wb.create_sheet("90-Day Reserve")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "90-Day Reserve",
                        prev_tab="Launch Costs", next_tab="Total + Y1 ROI")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Section 4 of 4 — operating runway. The difference between "
        "\"launched\" and \"sustained.\" Non-negotiable."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "90-DAY OPERATING RESERVE")

    # P&I × 3 — pulled from Acquisition tab
    ws.cell(row=7, column=2, value="Mortgage P&I × 3 mo:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=7, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=7, column=3, value="=Acquisition!C30*3")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.merge_cells("E7:L7")
    c = ws["E7"]
    c.value = "Auto-calculated from Acquisition tab P&I × 3 months"
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 18

    reserve_fields = [
        (8,  "Utilities × 3 mo ($):",          _val(variant, SAMPLE["reserve_util"]),  '"$"#,##0', "$200-400/mo typical"),
        (9,  "Insurance reserve ($):",         _val(variant, SAMPLE["reserve_ins"]),   '"$"#,##0', "Annual / 4 (3 mo)"),
        (10, "Cleaning reserve ($):",          _val(variant, SAMPLE["reserve_clean"]), '"$"#,##0', "Pre-bookings until cash flow normalizes"),
        (11, "Maintenance reserve ($):",       _val(variant, SAMPLE["reserve_maint"]), '"$"#,##0', "Surprise repairs in Y1"),
        (12, "Marketing reserve ($):",         _val(variant, SAMPLE["reserve_mktg"]),  '"$"#,##0', "Boost listing first 90 days"),
        (13, "Software subscriptions × 3 mo:", _val(variant, SAMPLE["reserve_sw"]),    '"$"#,##0', "PMS + pricing + accounting"),
        (14, "Other contingency ($):",         _val(variant, SAMPLE["reserve_other"]), '"$"#,##0', "HOA dues, surprises"),
    ]
    for row, label, value, fmt, note in reserve_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 15, "TOTAL RESERVE")
    _output_row(ws, 16, "90-day reserve:",
                "=SUM(C7:C14)", '"$"#,##0', emphasize=True)

    # Reassurance / warning row
    ws.merge_cells("A18:L19")
    c = ws["A18"]
    c.value = (
        "Why this matters: The #1 reason new STR investors fail isn't a bad "
        "deal — it's running out of cash before bookings stabilize. Three "
        "months of fixed costs in reserve = the difference between launched "
        "and sustained. Don't skip this."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[18].height = 22
    ws.row_dimensions[19].height = 22

    brand_footer(ws, 22,
                 version_line=f"{SKU} · 90-Day Reserve")


# ---------------------------------------------------------------------------
# Sheet 6 — Total + Y1 ROI
# ---------------------------------------------------------------------------

def build_total_tab(wb, variant):
    ws = wb.create_sheet("Total + Y1 ROI")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 20),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Total + Y1 ROI",
                        prev_tab="90-Day Reserve", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Your all-in budget + Y1 cash-on-cash return projection."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- ALL-IN ROLLUP (rows 6-14) ----
    _section_band(ws, 6, "ALL-IN BUDGET")

    # Row 7: Acquisition (purchase + DD/closing — NOT down payment alone, full)
    rollup_rows = [
        (7,  "Acquisition (full, incl. purchase price):",
              "=Acquisition!C13+SUM(Acquisition!C19:C25)"),
        (8,  "Furnish & Setup (selected total):",
              # Furnish totals row varies; reference SUM of zone-subtotal column F.
              # Simplest: pull the H-grand-total via cell reference to known row.
              "='Furnish & Setup'!H{furnish_total_row}"),
        (9,  "Launch Costs:",
              "='Launch Costs'!C28"),
        (10, "90-Day Reserve:",
              "='90-Day Reserve'!C16"),
    ]
    # We'll compute the furnish total row dynamically below by querying ws — but
    # that's hard before save. Use a robust SUM-by-formula instead: sum all H
    # cells from row 7 down by sheet reference. Since the totals row is the
    # last bold row, we can use SUMIF on the zone-name column. Simpler: hard-link.
    # Practical approach: use a known formula pattern — sum items from rows 7 to
    # ~250 of column H in Furnish, then divide by 2 (zones counted twice with
    # subtotals + grand total)? No — instead, we'll pre-compute the totals row
    # number once we know it. Inject after build.
    #
    # Cleaner fix: compute furnish_total_row before this function uses it.
    # We'll patch the formula in main() after building Furnish, OR just recompute
    # via a cell reference. The Furnish builder appended a totals row at a known
    # offset that depends on FURNISHING + zones. Simpler: read it from the wb
    # after Furnish builder runs.
    #
    # For maintainability and to avoid coupling, we use a safe SUMIFS-style:
    # the Furnish grand total row labels col A as merged "FURNISH & SETUP TOTALS"
    # — but col A is merged into B. We'll instead use a named formula approach:
    # sum F-column items only where zone name in col A is in the known zone
    # list. But that's fragile. Best: pass furnish_total_row in via wb attr.

    # We will fill these rows in main() after build_furnish_tab returns the row.
    # For now, write placeholder that we patch later.
    for r, label, formula in rollup_rows:
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    # Total all-in (row 14, big gold)
    ws.cell(row=14, column=2, value="TOTAL ALL-IN:").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=14, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=14, column=3, value="=SUM(C7:C10)")
    cell.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[14].height = 32

    # ---- CASH INVESTED (rows 17-19) ----
    _section_band(ws, 16, "CASH INVESTED  (excludes mortgage principal)")

    ws.cell(row=17, column=2, value="Down payment + DD/closing:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=17, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=17, column=3, value="=Acquisition!C31")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'

    ws.cell(row=18, column=2, value="+ Furnish + Launch + Reserve:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=18, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=18, column=3, value="=C8+C9+C10")
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'

    _output_row(ws, 19, "Cash invested (out of pocket):",
                "=C17+C18", '"$"#,##0', emphasize=True)

    # ---- Y1 ROI INPUTS (rows 22-25) ----
    _section_band(ws, 21, "Y1 REVENUE PROJECTION  (your inputs)")
    roi_inputs = [
        (22, "Y1 ADR — avg nightly rate ($):", _val(variant, SAMPLE["y1_adr"]),     '"$"#,##0', "Pull from AirDNA / comps; haircut for Y1"),
        (23, "Y1 occupancy (%):",              _val(variant, SAMPLE["y1_occ"]),     "0.0%",     "40-50% typical for Y1; 65% stabilized"),
        (24, "Available nights/yr:",           _val(variant, SAMPLE["avail_nights"]),"0",        "365 default; less if owner-blocked"),
        (25, "Y1 operating costs ($):",        _val(variant, SAMPLE["y1_op_costs"]),'"$"#,##0', "Excl. mortgage; rough projection"),
    ]
    for row, label, value, fmt, note in roi_inputs:
        _input_row(ws, row, label, value, fmt, note)

    # ---- Y1 ROI OUTPUTS (rows 27-31) ----
    _section_band(ws, 26, "Y1 OUTPUTS")
    _output_row(ws, 27, "Y1 gross revenue:",
                "=C22*C23*C24", '"$"#,##0')
    _output_row(ws, 28, "Y1 cash-on-cash %:",
                "=IFERROR((C27-C25-Acquisition!C30*12)/C19,0)", "0.0%", emphasize=True)
    _output_row(ws, 29, "Y1 net cash flow:",
                "=C27-C25-Acquisition!C30*12", '"$"#,##0')
    _output_row(ws, 30, "Payback (months, if positive):",
                "=IFERROR(IF(C29>0,C19/(C29/12),0),0)", "0.0")
    _output_row(ws, 31, "Break-even occupancy:",
                "=IFERROR((C25+Acquisition!C30*12)/(C22*C24),0)", "0.0%")

    brand_footer(ws, 34,
                 version_line=f"{SKU} · Total + Y1 ROI")

    # Print area
    ws.print_area = "A1:L36"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_acquisition_tab(wb, variant)
    build_furnish_tab(wb, variant)
    build_launch_tab(wb, variant)
    build_reserve_tab(wb, variant)
    build_total_tab(wb, variant)

    # Patch the Furnish total reference in Total tab.
    # Find the FURNISH & SETUP TOTALS row by scanning col A on Furnish tab.
    furnish_ws = wb["Furnish & Setup"]
    furnish_total_row = None
    for r in range(1, furnish_ws.max_row + 1):
        v = furnish_ws.cell(row=r, column=1).value
        if isinstance(v, str) and "FURNISH & SETUP TOTALS" in v:
            furnish_total_row = r
            break
    if furnish_total_row is None:
        # Fallback: pick last row with a formula in col H
        for r in range(furnish_ws.max_row, 0, -1):
            v = furnish_ws.cell(row=r, column=8).value
            if isinstance(v, str) and v.startswith("=SUM("):
                furnish_total_row = r
                break

    total_ws = wb["Total + Y1 ROI"]
    if furnish_total_row:
        total_ws["C8"].value = f"='Furnish & Setup'!H{furnish_total_row}"
    else:
        total_ws["C8"].value = 0

    wb.properties.title = "Cost-to-Launch Calculator — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Cost-to-launch wizard for first-time STR hosts. 6 tabs, ~120 "
        "pre-populated furnishing line items, 90-day operating reserve, "
        "Y1 cash-on-cash projection."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
