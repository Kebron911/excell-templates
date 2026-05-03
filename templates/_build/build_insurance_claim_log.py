"""Build LGL-004 Insurance Claim Log (v2.2 standard).

Operational-mode tool tracking claims filed against the host's UNDERLYING
STR insurance policy (storm, fire, theft, water, liability, vandalism).
This is distinct from OPS-002 Damage Claim + AirCover Log, which covers
platform claims (Airbnb AirCover / VRBO / Booking.com). This SKU is for
the carrier-side policy: HO-3, DP-3, commercial STR endorsements, etc.

5 tabs:
  1. Start           — KPIs (open claims, $ at risk, recovery YTD) + nav
  2. Claims Log      — master register, capacity 50
  3. Active Claim    — single-claim deep-dive worksheet
  4. Annual Summary  — recovery rate, frequency, premium-vs-recovery, breakdowns
  5. Settings        — active year, property list, carrier list, status list

Generates two files:
  templates/_masters/LGL-004-insurance-claim-log-DEMO.xlsx
  templates/_masters/LGL-004-insurance-claim-log-BLANK.xlsx
"""
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style, header_row_style,
    set_col_widths, apply_style,
    pseudo_button, compact_header_band, apply_brand_header,
    brand_footer,
    COLOR_GRAY_LIGHT, COLOR_INPUT_TINT, COLOR_WHITE, STATE_BAD_FILL,
)

SKU = "LGL-004"
NAME = "insurance-claim-log"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR_DEFAULT = 2026  # written into Settings!$B$5; all formulas read it

# ---------------------------------------------------------------------------
# Sample data (DEMO)
# ---------------------------------------------------------------------------

# Brief: 3 claims across 2 properties — 2 wind/storm 2024-2025, 1 water 2026 in-progress
# Columns: claim_no, date_loss, property, carrier, policy_no, type, description,
#          est_damages, deductible, submitted, recovered, date_filed, date_closed,
#          status, notes
CLAIMS_LOG_SAMPLES = [
    ("CLM-2024-001", "2024-09-12", "Smokies Ridge Cabin", "Foremost Signature",
     "FS-9923441", "Wind/Storm",
     "Tropical depression downed two pines onto roof — shingles + decking damage; minor interior leak in master.",
     14800, 2500, 14800, 11200,
     "2024-09-15", "2024-10-28",
     "Closed",
     "Adjuster site visit day 4. Roof contractor estimate aligned with carrier scope. ACV check after depreciation."),
    ("CLM-2025-002", "2025-03-18", "Creek Side", "Proper Insurance",
     "PRO-44519", "Wind/Storm",
     "Hail event — full roof replacement approved + screened porch screening + 2 windows.",
     22400, 5000, 22400, 17900,
     "2025-03-19", "2025-05-21",
     "Closed",
     "Proper STR endorsement covered loss-of-rent for 18 days at booked-rate. Recovery includes $3,200 LOR."),
    ("CLM-2026-001", "2026-02-04", "Smokies Ridge Cabin", "Foremost Signature",
     "FS-9923441", "Water",
     "Supply line burst behind washer mid-stay; water damage to laundry, hallway, and master subfloor.",
     9200, 2500, 9200, 0,
     "2026-02-05", None,
     "Estimate received",
     "Adjuster confirmed covered peril (sudden/accidental discharge). Awaiting approval of mitigation invoice."),
]

# Active Claim Detail demo = the 2026 Q1 water-line incident (in progress)
ACTIVE_DEMO = {
    "loss": {
        "claim_no": "CLM-2026-001",
        "date_of_loss": "2026-02-04",
        "time_of_loss": "Approx. 02:30 PM",
        "property": "Smokies Ridge Cabin",
        "carrier": "Foremost Signature",
        "policy_no": "FS-9923441",
        "type": "Water",
        "severity": 4,
        "narrative": (
            "Mid-stay guest reported water pooling under washer in laundry room. "
            "Burst supply line. Cleaner shut off main within ~20 min of guest call. "
            "Water tracked under hardwood in hallway and into master closet. "
            "Restoration vendor (ServePro) extracted same-day, set drying equipment "
            "for 4 days. Subfloor in master shows soft spot ~6 sq ft."
        ),
    },
    "damages": [
        # (item, replacement_cost, depreciation, ACV)
        ("Hallway hardwood (38 sqft)",   1280.00, 320.00, 960.00),
        ("Master closet flooring",        780.00, 195.00, 585.00),
        ("Subfloor patch + level",        420.00,   0.00, 420.00),
        ("Laundry vinyl + base trim",     560.00, 140.00, 420.00),
        ("ServePro mitigation invoice",  3450.00,   0.00, 3450.00),
        ("Drying equipment rental",       890.00,   0.00, 890.00),
        ("Loss-of-rent (4 cancelled nights)", 1820.00, 0.00, 1820.00),
    ],
    "evidence": [
        ("Photos of damage (pre-mitigation)",         "Yes"),
        ("Photos of damage (post-mitigation)",        "Yes"),
        ("ServePro mitigation invoice",                "Yes"),
        ("Plumber repair receipt (supply line)",       "Yes"),
        ("Contractor estimate(s) for restoration",     "Yes"),
        ("Original purchase receipts (flooring)",      "Partial"),
        ("Police/fire report (if applicable)",         "—"),
        ("Weather report (if storm event)",            "—"),
        ("Witness statements (cleaner, vendor)",       "Yes"),
        ("Loss-of-rent calculation worksheet",         "Yes"),
    ],
    "comms": [
        ("2026-02-04 03:14 PM", "Phone",   "Out",  "FNOL call to Foremost. Claim # assigned: CLM-2026-001."),
        ("2026-02-05 09:02 AM", "Email",   "Out",  "Sent initial photos + ServePro invoice."),
        ("2026-02-06 11:40 AM", "Phone",   "In",   "Adjuster Pat Mendoza assigned. Site visit scheduled 2/9."),
        ("2026-02-09 02:30 PM", "In-person","Out", "Adjuster site visit. Documented all damage areas."),
        ("2026-02-12 10:15 AM", "Email",   "In",   "Carrier scope of loss received — $9,200 estimate."),
        ("2026-02-15 04:48 PM", "Email",   "Out",  "Submitted mitigation + flooring estimates for review."),
    ],
    "timeline": [
        ("FNOL filed",             "2026-02-05"),
        ("Adjuster assigned",      "2026-02-06"),
        ("Adjuster site visit",    "2026-02-09"),
        ("Carrier estimate received", "2026-02-12"),
        ("Approval / scope agreed", None),
        ("ACV check issued",       None),
        ("Repairs completed",      None),
        ("Final settlement",       None),
    ],
    "outcome": {
        "amount_submitted": 9200.00,
        "amount_recovered": 0.00,
        "deductible":       2500.00,
        "denial_reason":    "—",
        "lessons": (
            "Add quarterly supply-line check to turnover SOP. Replace original "
            "rubber washer hoses with braided stainless on every property — "
            "$30 each, prevents this entire claim category."
        ),
    },
}

# Settings
PROPERTIES_LIST = [
    "Smokies Ridge Cabin", "Creek Side", "Lakehouse A",
    "Lakehouse B", "Mountain View", "Downtown Loft",
]
CARRIERS_LIST = [
    "Foremost Signature", "Proper Insurance", "Allstate",
    "State Farm", "Farmers", "USAA", "Travelers",
    "American Modern", "CBIZ Vacation Rental", "Other",
]
STATUS_LIST = [
    "FNOL", "Adjuster assigned", "Estimate received",
    "Approved", "Settled", "Denied", "Closed",
]
TYPE_LIST = [
    "Wind/Storm", "Fire", "Water", "Theft",
    "Liability", "Vandalism", "Other",
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _val_list(variant, demo_list, blank_length=None):
    if variant == "demo":
        return demo_list
    n = blank_length if blank_length is not None else len(demo_list)
    return [None] * n


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 1 — Start: navy hero + KPI cards + nav buttons."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero band rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Insurance Claim Log"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Storm, fire, theft, water — every claim filed against your STR policy."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: KPI cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Open claims
    ws.merge_cells("A10:D10")
    c = ws["A10"]
    c.value = "OPEN CLAIMS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A11:D13")
    c = ws["A11"]
    c.value = (
        '=COUNTIFS(\'Claims Log\'!P6:P55,"<>Settled",'
        '\'Claims Log\'!P6:P55,"<>Closed",'
        '\'Claims Log\'!P6:P55,"<>Denied",'
        '\'Claims Log\'!P6:P55,"<>",'
        '\'Claims Log\'!A6:A55,"<>")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A14:D15")
    c = ws["A14"]
    c.value = "claims still in flight"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 2 (E-H): $ at risk
    ws.merge_cells("E10:H10")
    c = ws["E10"]
    c.value = "$ AT RISK"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E11:H13")
    c = ws["E11"]
    c.value = (
        '=SUMIFS(\'Claims Log\'!J6:J55,'
        '\'Claims Log\'!P6:P55,"<>Settled",'
        '\'Claims Log\'!P6:P55,"<>Closed",'
        '\'Claims Log\'!P6:P55,"<>Denied",'
        '\'Claims Log\'!A6:A55,"<>")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E14:H15")
    c = ws["E14"]
    c.value = "submitted, not yet recovered"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 3 (I-L): Recovery YTD
    ws.merge_cells("I10:L10")
    c = ws["I10"]
    c.value = "RECOVERY YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I11:L13")
    c = ws["I11"]
    c.value = (
        '=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*'
        '\'Claims Log\'!K6:K55)'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I14:L15")
    c = ws["I14"]
    c.value = "recovered this year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 16):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: What this does card rows 17-21 (parchment-alt)
    for r in range(17, 22):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = PatternFill(
                "solid", fgColor=COLOR_PARCHMENT_ALT
            )

    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    ws.merge_cells("A18:L21")
    c = ws["A18"]
    c.value = (
        "Tracks claims filed against your underlying STR insurance carrier "
        "(HO-3, DP-3, Proper, Foremost, etc.) — separate from platform claims like "
        "Airbnb AirCover. Records every claim from FNOL through settlement, "
        "captures recovery $ vs deductible, and surfaces lessons-learned so each "
        "claim makes the next one easier."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=2
    )

    # ZONE 4: Pseudo-buttons rows 23-27
    pseudo_button(ws, "A23", "F26", "→ NEW CLAIM",
                  "'Claims Log'!A6", variant="primary")
    pseudo_button(ws, "G23", "L26", "Active Claim Detail",
                  "'Active Claim'!A1", variant="secondary")
    for r in range(23, 27):
        ws.row_dimensions[r].height = 20

    pseudo_button(ws, "A28", "F30", "Annual Summary",
                  "'Annual Summary'!A1", variant="secondary")
    pseudo_button(ws, "G28", "L30", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(28, 31):
        ws.row_dimensions[r].height = 20

    # ZONE 5: Upgrade banner row 32
    ws.merge_cells("A32:L32")
    c = ws["A32"]
    c.value = (
        "💡  Need full operations coverage? Add the Operator Bundle at "
        f"{BRAND_DOMAIN}/operator — turnover + maintenance + claims + insurance."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[32].height = 32

    brand_footer(ws, 34,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L36"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_claims_log_tab(wb, variant):
    """Sheet 2 — Claims Log master register, capacity 50 rows."""
    ws = wb.create_sheet("Claims Log")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Claims Log",
                        prev_tab="Start", next_tab="Active Claim")

    # Row 4 description
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 18):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:Q4")
    c = ws["A4"]
    c.value = (
        "Master register — every claim filed against the underlying carrier. "
        "Net to host (col L) = Recovered − Deductible. Days to close (col O) "
        "auto-calculates once you set Date closed."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Claim #", "Date of loss", "Property", "Carrier",
        "Policy #", "Type", "Loss description", "$ Est. damages",
        "Deductible", "$ Submitted", "$ Recovered", "Net to host",
        "Date filed", "Date closed", "Days to close", "Status", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 14), ("B", 12), ("C", 22), ("D", 20),
        ("E", 14), ("F", 12), ("G", 38), ("H", 13),
        ("I", 11), ("J", 12), ("K", 12), ("L", 12),
        ("M", 12), ("N", 12), ("O", 11), ("P", 18),
        ("Q", 32),
    ])

    samples = _val_list(variant, CLAIMS_LOG_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 6 + i
        (claim_no, date_loss, prop, carrier, policy_no, ctype, desc,
         est, ded, sub, rec, date_filed, date_closed, status, notes) = row_data

        # Text fields
        for col, val in [(1, claim_no), (3, prop), (4, carrier),
                         (5, policy_no), (6, ctype), (7, desc),
                         (16, status), (17, notes)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        # Date fields
        for col, val in [(2, date_loss), (13, date_filed), (14, date_closed)]:
            cell = ws.cell(row=row, column=col, value=_parse_date(val))
            apply_style(cell, input_cell_style())
            cell.number_format = "yyyy-mm-dd"

        # Money fields
        for col, val in [(8, est), (9, ded), (10, sub), (11, rec)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'

        ws.row_dimensions[row].height = 22

    # Apply formulas + blank input style on rows 6-55 (capacity 50)
    for row in range(6, 56):
        # Net to host (col L) = Recovered − Deductible
        l_cell = ws.cell(
            row=row, column=12,
            value=f'=IF(K{row}="","",K{row}-I{row})',
        )
        apply_style(l_cell, formula_cell_style())
        l_cell.number_format = '"$"#,##0'

        # Days to close (col O) = N − M
        o_cell = ws.cell(
            row=row, column=15,
            value=f'=IF(AND(M{row}<>"",N{row}<>""),N{row}-M{row},"")',
        )
        apply_style(o_cell, formula_cell_style())
        o_cell.number_format = "0"
        o_cell.alignment = Alignment(horizontal="center", vertical="center")

        # For rows beyond samples: apply input style + number format on input cols
        if row > 5 + len(samples):
            for col in [1, 3, 4, 5, 6, 7, 16, 17]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            for col in [2, 13, 14]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            for col in [8, 9, 10, 11]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0'

    # Conditional formatting on Status (col P): red if Denied
    ws.conditional_formatting.add(
        "P6:P55",
        CellIsRule(operator="equal", formula=['"Denied"'],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    # Gold-soft if FNOL/Adjuster assigned/Estimate received
    ws.conditional_formatting.add(
        "P6:P55",
        CellIsRule(operator="equal", formula=['"FNOL"'],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )
    ws.conditional_formatting.add(
        "P6:P55",
        CellIsRule(operator="equal", formula=['"Adjuster assigned"'],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )
    ws.conditional_formatting.add(
        "P6:P55",
        CellIsRule(operator="equal", formula=['"Estimate received"'],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )

    # Dropdowns
    add_dropdown(ws, "C6:C55", "=Settings!$B$7:$B$16")  # Property
    add_dropdown(ws, "D6:D55", "=Settings!$B$18:$B$27")  # Carrier
    add_dropdown(ws, "F6:F55", '"' + ",".join(TYPE_LIST) + '"')  # Type
    add_dropdown(ws, "P6:P55", "=Settings!$B$29:$B$38")  # Status

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_active_claim_tab(wb, variant):
    """Sheet 3 — Active Claim Detail per-claim worksheet."""
    ws = wb.create_sheet("Active Claim")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 26), ("C", 14), ("D", 14),
        ("E", 14), ("F", 14), ("G", 12), ("H", 8),
        ("I", 8), ("J", 8), ("K", 8), ("L", 10),
    ])

    compact_header_band(ws, "Active Claim Detail",
                        prev_tab="Claims Log", next_tab="Annual Summary")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-claim deep dive — fill from top to bottom. When the claim closes, "
        "copy the outcome row up to the Claims Log so the dashboard reflects it. "
        "Reuse this tab for the next claim (or duplicate it to keep history)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    demo = variant == "demo"
    pdata = ACTIVE_DEMO if demo else None

    def _section_header(row, text):
        ws.merge_cells(f"A{row}:L{row}")
        cc = ws[f"A{row}"]
        cc.value = text
        cc.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        cc.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    # ===== §1 LOSS DETAILS rows 6-19 =====
    _section_header(6, "§1  LOSS DETAILS")

    loss_fields = [
        ("Claim #:",        "claim_no",       7,  None),
        ("Date of loss:",   "date_of_loss",   8,  "yyyy-mm-dd"),
        ("Time of loss:",   "time_of_loss",   9,  None),
        ("Property:",       "property",       10, None),
        ("Carrier:",        "carrier",        11, None),
        ("Policy #:",       "policy_no",      12, None),
        ("Type of loss:",   "type",           13, None),
        ("Severity (1-5):", "severity",       14, None),
    ]
    for label, key, row, num_fmt in loss_fields:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        ws.merge_cells(f"C{row}:G{row}")
        cell = ws[f"C{row}"]
        if demo and pdata:
            val = pdata["loss"][key]
            if key == "date_of_loss":
                cell.value = _parse_date(val)
            else:
                cell.value = val
        apply_style(cell, input_cell_style())
        if num_fmt:
            cell.number_format = num_fmt
        ws.row_dimensions[row].height = 18

    # Dropdowns inside the loss block
    add_dropdown(ws, "C10", "=Settings!$B$7:$B$16")
    add_dropdown(ws, "C11", "=Settings!$B$18:$B$27")
    add_dropdown(ws, "C13", '"' + ",".join(TYPE_LIST) + '"')

    # Narrative rows 16-19
    ws.cell(row=16, column=2, value="Narrative:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C16:L19")
    cell = ws["C16"]
    if demo and pdata:
        cell.value = pdata["loss"]["narrative"]
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=1
    )
    for r in range(16, 20):
        ws.row_dimensions[r].height = 20

    # ===== §2 DAMAGE INVENTORY rows 21-34 =====
    _section_header(21, "§2  DAMAGE INVENTORY")

    dmg_headers = ["Item description", "Replacement cost", "Depreciation", "ACV"]
    dmg_cols = [("B", "F"), ("G", "H"), ("I", "J"), ("K", "L")]
    for (cstart, cend), h in zip(dmg_cols, dmg_headers):
        ws.merge_cells(f"{cstart}22:{cend}22")
        cell = ws[f"{cstart}22"]
        cell.value = h
        apply_style(cell, header_row_style())
    ws.row_dimensions[22].height = 20

    dmg_demo = pdata["damages"] if demo and pdata else []
    for i in range(10):
        r = 23 + i
        sample = dmg_demo[i] if i < len(dmg_demo) else (None, None, None, None)
        item, cost, depr, acv = sample

        ws.merge_cells(f"B{r}:F{r}")
        cell = ws[f"B{r}"]
        cell.value = item
        apply_style(cell, input_cell_style())

        ws.merge_cells(f"G{r}:H{r}")
        cell = ws[f"G{r}"]
        cell.value = cost
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0.00'

        ws.merge_cells(f"I{r}:J{r}")
        cell = ws[f"I{r}"]
        cell.value = depr
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0.00'

        ws.merge_cells(f"K{r}:L{r}")
        cell = ws[f"K{r}"]
        cell.value = acv
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[r].height = 18

    # Total row 33
    ws.cell(row=33, column=2, value="Total ACV:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("K33:L33")
    cell = ws["K33"]
    cell.value = "=SUM(K23:L32)"
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[33].height = 22

    # ===== §3 EVIDENCE CHECKLIST rows 35-46 =====
    _section_header(35, "§3  EVIDENCE CHECKLIST")

    default_evidence_labels = [
        "Photos of damage (pre-mitigation)",
        "Photos of damage (post-mitigation)",
        "Mitigation invoice (water-extraction, board-up, tarp)",
        "Repair / contractor receipts",
        "Contractor estimate(s) for restoration",
        "Original purchase receipts (damaged items)",
        "Police / fire report (if applicable)",
        "Weather report (if storm event)",
        "Witness statements (cleaner, vendor, neighbor)",
        "Loss-of-rent calculation worksheet",
    ]
    evidence_demo = pdata["evidence"] if demo and pdata else None
    for i in range(10):
        r = 36 + i
        if demo and evidence_demo and i < len(evidence_demo):
            label, mark = evidence_demo[i]
        else:
            label, mark = default_evidence_labels[i], None

        # Yes/Partial/— cell col B
        cb = ws.cell(row=r, column=2, value=mark)
        cb.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        cb.fill = PatternFill("solid", fgColor=COLOR_INPUT_TINT)
        cb.alignment = Alignment(horizontal="center", vertical="center")
        cb.border = Border(
            left=Side(style="thin", color=COLOR_GRAY_LIGHT),
            right=Side(style="thin", color=COLOR_GRAY_LIGHT),
            top=Side(style="thin", color=COLOR_GRAY_LIGHT),
            bottom=Side(style="thin", color=COLOR_GRAY_LIGHT),
        )
        add_dropdown(ws, f"B{r}", '"Yes,Partial,No,—"')

        # Label cols C-L
        ws.merge_cells(f"C{r}:L{r}")
        cell = ws[f"C{r}"]
        cell.value = label
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[r].height = 18

    # ===== §4 COMMUNICATIONS LOG rows 48-58 =====
    _section_header(48, "§4  COMMUNICATIONS LOG")

    comm_headers = ["Date / time", "Channel", "Direction", "Summary"]
    comm_cols = [("B", "D"), ("E", "F"), ("G", "G"), ("H", "L")]
    for (cstart, cend), h in zip(comm_cols, comm_headers):
        ws.merge_cells(f"{cstart}49:{cend}49")
        cell = ws[f"{cstart}49"]
        cell.value = h
        apply_style(cell, header_row_style())
    ws.row_dimensions[49].height = 20

    comms_demo = pdata["comms"] if demo and pdata else []
    for i in range(10):
        r = 50 + i
        sample = comms_demo[i] if i < len(comms_demo) else (None, None, None, None)
        date_t, channel, direction, summary = sample

        ws.merge_cells(f"B{r}:D{r}")
        cell = ws[f"B{r}"]
        cell.value = date_t
        apply_style(cell, input_cell_style())

        ws.merge_cells(f"E{r}:F{r}")
        cell = ws[f"E{r}"]
        cell.value = channel
        apply_style(cell, input_cell_style())
        add_dropdown(ws, f"E{r}", '"Phone,Email,In-person,Text/SMS,Portal,Letter"')

        cell = ws.cell(row=r, column=7, value=direction)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        add_dropdown(ws, f"G{r}", '"In,Out"')

        ws.merge_cells(f"H{r}:L{r}")
        cell = ws[f"H{r}"]
        cell.value = summary
        apply_style(cell, input_cell_style())

        ws.row_dimensions[r].height = 18

    # ===== §5 TIMELINE rows 61-70 =====
    _section_header(61, "§5  TIMELINE  —  FNOL → SETTLEMENT")

    timeline_milestones = [
        "FNOL filed",
        "Adjuster assigned",
        "Adjuster site visit",
        "Carrier estimate received",
        "Approval / scope agreed",
        "ACV check issued",
        "Repairs completed",
        "Final settlement",
    ]
    timeline_demo = pdata["timeline"] if demo and pdata else None
    for i, milestone in enumerate(timeline_milestones):
        r = 62 + i
        ws.cell(row=r, column=2, value=milestone).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        ws.merge_cells(f"C{r}:F{r}")
        cell = ws[f"C{r}"]
        if demo and timeline_demo:
            _, val = timeline_demo[i]
            if val:
                cell.value = _parse_date(val)
        apply_style(cell, input_cell_style())
        cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[r].height = 18

    # ===== §6 OUTCOME rows 71-79 =====
    _section_header(71, "§6  OUTCOME + LESSONS LEARNED")

    outcome_fields = [
        ("$ Submitted:",   "amount_submitted", '"$"#,##0.00'),
        ("$ Recovered:",   "amount_recovered", '"$"#,##0.00'),
        ("Deductible:",    "deductible",       '"$"#,##0.00'),
        ("Denial reason:", "denial_reason",    None),
    ]
    for i, (label, key, num_fmt) in enumerate(outcome_fields):
        r = 72 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        ws.merge_cells(f"C{r}:L{r}")
        cell = ws[f"C{r}"]
        if demo and pdata:
            cell.value = pdata["outcome"][key]
        apply_style(cell, input_cell_style())
        if num_fmt:
            cell.number_format = num_fmt
        ws.row_dimensions[r].height = 18

    # Net to host (formula)
    ws.cell(row=76, column=2, value="Net to host:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C76:L76")
    cell = ws["C76"]
    cell.value = "=C73-C74"
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[76].height = 20

    ws.cell(row=78, column=2, value="Lessons learned:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C78:L80")
    cell = ws["C78"]
    if demo and pdata:
        cell.value = pdata["outcome"]["lessons"]
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=1
    )
    for r in range(78, 81):
        ws.row_dimensions[r].height = 20

    # Footer reminder row 82
    ws.merge_cells("A82:L82")
    c = ws["A82"]
    c.value = (
        "When the claim closes — copy the key outcome values up to the matching "
        "row on the Claims Log tab so the Start dashboard reflects the result."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[82].height = 28

    ws.print_area = "A1:L82"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_annual_summary_tab(wb, variant):
    """Sheet 4 — Annual Summary."""
    ws = wb.create_sheet("Annual Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 18), ("D", 18),
        ("E", 6), ("F", 6), ("G", 6), ("H", 6),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "Annual Summary",
                        prev_tab="Active Claim", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Year-end summary — recovery rate, claim frequency, premium-vs-recovery "
        "ratio, and breakdowns by category and property. The active year is "
        "set in Settings → cell B5."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Active year display
    ws.cell(row=6, column=2, value="Active year:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    cell = ws.cell(row=6, column=3, value="=Settings!$B$5")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 24

    # ===== TOTALS rows 8-15 =====
    ws.merge_cells("A8:L8")
    c = ws["A8"]
    c.value = "TOTALS"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[8].height = 28

    totals = [
        ("Claim count:",
         '=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*(\'Claims Log\'!A6:A55<>""))',
         "0"),
        ("Total $ submitted:",
         '=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*\'Claims Log\'!J6:J55)',
         '"$"#,##0'),
        ("Total $ recovered:",
         '=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*\'Claims Log\'!K6:K55)',
         '"$"#,##0'),
        ("Recovery rate:",
         '=IFERROR(C11/C10,0)',
         "0%"),
        ("Avg days to close:",
         '=IFERROR(AVERAGEIFS(\'Claims Log\'!O6:O55,'
         '\'Claims Log\'!B6:B55,">="&DATE(Settings!$B$5,1,1),'
         '\'Claims Log\'!B6:B55,"<="&DATE(Settings!$B$5,12,31),'
         '\'Claims Log\'!N6:N55,"<>"),0)',
         "0"),
        ("Annual premium paid:",
         '=Settings!$B$40',
         '"$"#,##0'),
        ("Insurance ROI (recovery ÷ premium):",
         '=IFERROR(C11/C14,0)',
         "0%"),
    ]
    for i, (label, formula, fmt) in enumerate(totals):
        r = 9 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = fmt
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 20

    # ===== BY CATEGORY rows 17-25 =====
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "BY CLAIM TYPE"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 28

    cat_headers = ["Type", "$ Submitted", "$ Recovered"]
    for col, h in enumerate(cat_headers, start=2):
        cell = ws.cell(row=18, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[18].height = 22

    for i, ctype in enumerate(TYPE_LIST):
        r = 19 + i
        ws.cell(row=r, column=2, value=ctype).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        sub_cell = ws.cell(
            row=r, column=3,
            value=(
                f'=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*'
                f'(\'Claims Log\'!F6:F55=B{r})*\'Claims Log\'!J6:J55)'
            ),
        )
        apply_style(sub_cell, formula_cell_style())
        sub_cell.number_format = '"$"#,##0'

        rec_cell = ws.cell(
            row=r, column=4,
            value=(
                f'=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*'
                f'(\'Claims Log\'!F6:F55=B{r})*\'Claims Log\'!K6:K55)'
            ),
        )
        apply_style(rec_cell, formula_cell_style())
        rec_cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    # ===== BY PROPERTY rows 27-37 =====
    ws.merge_cells("A27:L27")
    c = ws["A27"]
    c.value = "BY PROPERTY"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[27].height = 28

    prop_headers = ["Property", "$ Submitted", "$ Recovered"]
    for col, h in enumerate(prop_headers, start=2):
        cell = ws.cell(row=28, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[28].height = 22

    for i in range(10):
        r = 29 + i
        if i < len(PROPERTIES_LIST):
            prop = PROPERTIES_LIST[i]
            ws.cell(row=r, column=2, value=prop).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )
            sub_cell = ws.cell(
                row=r, column=3,
                value=(
                    f'=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*'
                    f'(\'Claims Log\'!C6:C55=B{r})*\'Claims Log\'!J6:J55)'
                ),
            )
            apply_style(sub_cell, formula_cell_style())
            sub_cell.number_format = '"$"#,##0'

            rec_cell = ws.cell(
                row=r, column=4,
                value=(
                    f'=SUMPRODUCT((YEAR(\'Claims Log\'!B6:B55)=Settings!$B$5)*'
                    f'(\'Claims Log\'!C6:C55=B{r})*\'Claims Log\'!K6:K55)'
                ),
            )
            apply_style(rec_cell, formula_cell_style())
            rec_cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    # Pseudo-button back to log
    pseudo_button(ws, "B41", "D43", "← Back to Claims Log",
                  "'Claims Log'!A1", variant="secondary")
    for r in range(41, 44):
        ws.row_dimensions[r].height = 18

    ws.print_area = "A1:D45"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 5 — Settings: active year + property/carrier/status lists."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 30),
        ("C", 4), ("D", 28), ("E", 28),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Annual Summary", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Edit these lists to match your portfolio + carriers. They drive "
        "the dropdowns and the year-scoped formulas on every other tab."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Active year — cell B5 (per brief, cell B5 holds the active year value)
    cell = ws.cell(row=5, column=2, value=ACTIVE_YEAR_DEFAULT)
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0"

    ws.cell(row=5, column=4,
            value="← Active year (drives all annual formulas)").font = Font(
        name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED
    )
    ws.row_dimensions[5].height = 26

    # Property list header row 6
    cell = ws.cell(row=6, column=2, value="PROPERTIES")
    cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[6].height = 20

    # Properties B7-B16 (10 capacity)
    for i in range(10):
        r = 7 + i
        cell = ws.cell(row=r, column=2)
        if i < len(PROPERTIES_LIST):
            cell.value = PROPERTIES_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Carrier list header row 17
    cell = ws.cell(row=17, column=2, value="CARRIERS")
    cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[17].height = 20

    # Carriers B18-B27 (10 capacity)
    for i in range(10):
        r = 18 + i
        cell = ws.cell(row=r, column=2)
        if i < len(CARRIERS_LIST):
            cell.value = CARRIERS_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Status list header row 28
    cell = ws.cell(row=28, column=2, value="CLAIM STATUSES")
    cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[28].height = 20

    # Statuses B29-B38 (10 capacity)
    for i in range(10):
        r = 29 + i
        cell = ws.cell(row=r, column=2)
        if i < len(STATUS_LIST):
            cell.value = STATUS_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Annual premium row 40
    ws.cell(row=39, column=2, value="ANNUAL PREMIUM").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=39, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.cell(row=39, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[39].height = 20

    cell = ws.cell(row=40, column=2, value=8400 if variant == "demo" else None)
    apply_style(cell, input_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[40].height = 22

    ws.cell(row=40, column=4,
            value="← Total premium paid this year (for ROI calc)").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )

    # Year-end archive table
    ws.merge_cells("A43:L43")
    c = ws["A43"]
    c.value = "YEAR-END ARCHIVE  —  fill each January from the Annual Summary"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[43].height = 24

    arch_headers = ["Year", "Claims", "$ Submitted", "$ Recovered"]
    for col, h in enumerate(arch_headers, start=2):
        cell = ws.cell(row=44, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[44].height = 20

    # 6 future years pre-stamped
    for i in range(6):
        r = 45 + i
        year = ACTIVE_YEAR_DEFAULT + i
        cell = ws.cell(row=r, column=2, value=year)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)
        for col in [3, 4, 5]:
            ac = ws.cell(row=r, column=col)
            apply_style(ac, input_cell_style())
            if col in (4, 5):
                ac.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_claims_log_tab(wb, variant)
    build_active_claim_tab(wb, variant)
    build_annual_summary_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Insurance Claim Log — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Carrier-side insurance claim register for STR hosts — wind/storm, "
        "fire, water, theft, liability. Tracks FNOL through settlement, "
        "recovery vs deductible, and annual ROI vs premium paid."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
