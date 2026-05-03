"""Build FIN-005 Year-Over-Year Comparison Workbook (v2.3 standard).

Operational-mode matrix tool. Forks build_pl_single_property.py and extends
the 12-col monthly matrix to a 24-col PY+CY matrix (12 prior-year months,
12 current-year months) with Total + delta columns.

Five tabs:
  1. Start            — Hero, KPI cards (revenue YoY %, NOI YoY %, top mover,
                        top problem), nav buttons
  2. Per Property     — Three property sections, each with Revenue + Expenses
                        PY/CY matrices and per-row delta + % delta
  3. Portfolio Roll-up — Big numbers (revenue / expenses / NOI PY vs CY) plus
                         stacked-bar PY-vs-CY by property and per-month line
                         chart (CY vs PY)
  4. Top Movers       — Top-3 favorable + top-3 unfavorable categories,
                        sentiment marker
  5. Settings         — Property list, current year (B5), prior year (B7)

DEMO data (per brief): 3 properties (Smokies Ridge Cabin, Lakehouse Retreat,
Coastal Cottage), 2025 vs 2026 (Jan-Mar partial CY), portfolio revenue
$282K → $298K (+5.7%), insurance refinance favorable, HVAC failure unfavorable.

BLANK variant: same scaffolding, empty inputs, formulas intact.

Critical rule: never YEAR(TODAY()). Every year reference resolves to
Settings!$B$5 (current) or Settings!$B$7 (prior).
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    COLOR_WHITE, STATE_BAD_FILL, STATE_GOOD_FILL,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "FIN-005"
NAME = "yoy-comparison-workbook"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

CURRENT_YEAR_DEFAULT = 2026
PRIOR_YEAR_DEFAULT = 2025

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# 6 line items per matrix — single revenue row, 5 expense category rollups.
# Keeps the 24-col grid readable in landscape print.
REVENUE_ROWS = ["Rents + cleaning fees collected"]

EXPENSE_ROWS = [
    "Cleaning + maintenance",
    "Insurance",
    "Mortgage interest",
    "Repairs",
    "Utilities",
    "Management fees",
    "Supplies",
    "Other",
]

# 3 demo properties — names match the brief's narrative
# (insurance refinance favorable; Lakehouse HVAC compressor failure unfavorable).
PROPERTIES = [
    "Smokies Ridge Cabin",
    "Lakehouse Retreat",
    "Coastal Cottage",
]

# Demo PY (2025) revenue per property — full year of bookings
# Per-property arrays are 12 monthly figures (Jan..Dec).
DEMO_PY_REVENUE = {
    "Smokies Ridge Cabin":  [ 8400, 7600, 9800, 11200, 12400, 13800,
                             14600, 13200, 11400, 10200, 9100, 8800],
    "Lakehouse Retreat":    [ 5200, 4900, 6800,  9200, 11800, 13400,
                             14800, 13600, 10200,  7400, 5800, 5400],
    "Coastal Cottage":      [ 6800, 6400, 7900,  9400, 10800, 12200,
                             13400, 12600, 10100,  8400, 7200, 7100],
}

# Demo CY (2026) revenue per property — Jan..Mar real, Apr..Dec zeros
# (the workbook the host opens in early CY to compare partial vs prior).
DEMO_CY_REVENUE = {
    "Smokies Ridge Cabin":  [ 9100, 8400, 10600, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Lakehouse Retreat":    [ 5600, 5300,  7400, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Coastal Cottage":      [ 7400, 7000,  8600, 0, 0, 0, 0, 0, 0, 0, 0, 0],
}

# Demo PY expenses — per property, per category, 12 monthly figures
DEMO_PY_EXPENSES = {
    "Smokies Ridge Cabin": {
        "Cleaning + maintenance": [450, 450, 600, 750, 900, 1050, 1200, 1050, 900, 750, 600, 450],
        "Insurance":              [340, 340, 340, 340, 340, 340,  340,  340, 340, 340, 340, 340],
        "Mortgage interest":      [400, 400, 400, 400, 400, 400,  400,  400, 400, 400, 400, 400],
        "Repairs":                [  0, 200,   0, 150,   0, 320,    0,  180,   0, 240,   0,  90],
        "Utilities":              [210, 230, 220, 200, 240, 260,  290,  280, 220, 210, 240, 260],
        "Management fees":        [  0,   0,   0,   0,   0,   0,    0,    0,   0,   0,   0,   0],
        "Supplies":               [120, 100, 140, 160, 180, 220,  240,  220, 180, 140, 110, 110],
        "Other":                  [240, 200, 280, 320, 360, 420,  480,  440, 360, 280, 200, 220],
    },
    "Lakehouse Retreat": {
        "Cleaning + maintenance": [300, 300, 450, 600, 850, 1050, 1200, 1050, 750, 450, 350, 300],
        "Insurance":              [380, 380, 380, 380, 380, 380,  380,  380, 380, 380, 380, 380],
        "Mortgage interest":      [340, 340, 340, 340, 340, 340,  340,  340, 340, 340, 340, 340],
        # Aug 2025 HVAC compressor failure — $3,200 — surfaces as top problem
        "Repairs":                [  0,   0, 100,   0, 220,   0, 3200,  150,   0, 100,   0,   0],
        "Utilities":              [180, 200, 210, 220, 240, 280,  320,  300, 240, 200, 190, 190],
        "Management fees":        [260, 245, 340, 460, 590, 670,  740,  680, 510, 370, 290, 270],
        "Supplies":               [ 80,  80, 110, 130, 150, 180,  200,  180, 140, 100,  80,  80],
        "Other":                  [180, 160, 210, 270, 330, 380,  430,  390, 290, 220, 180, 170],
    },
    "Coastal Cottage": {
        "Cleaning + maintenance": [400, 380, 540, 680, 820, 960, 1080, 980, 760, 580, 460, 440],
        "Insurance":              [310, 310, 310, 310, 310, 310,  310, 310, 310, 310, 310, 310],
        "Mortgage interest":      [360, 360, 360, 360, 360, 360,  360, 360, 360, 360, 360, 360],
        "Repairs":                [ 80,   0, 140,   0, 200,   0,  220,   0,  90,   0, 110,   0],
        "Utilities":              [195, 210, 215, 210, 230, 260,  290, 280, 230, 210, 215, 220],
        "Management fees":        [340, 320, 395, 470, 540, 610,  670, 630, 505, 420, 360, 355],
        "Supplies":               [110,  90, 130, 150, 170, 200,  220, 200, 165, 120, 110, 100],
        "Other":                  [205, 195, 240, 285, 325, 365,  400, 380, 305, 250, 215, 215],
    },
}

# Demo CY expenses (Jan..Mar real, rest zero)
# Insurance dropped per property (~$150/mo savings) — the favorable mover.
DEMO_CY_EXPENSES = {
    "Smokies Ridge Cabin": {
        "Cleaning + maintenance": [480, 470, 640, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Insurance":              [190, 190, 190, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Mortgage interest":      [400, 400, 400, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Repairs":                [  0, 220,   0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Utilities":              [220, 240, 230, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Management fees":        [  0,   0,   0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Supplies":               [125, 110, 145, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Other":                  [250, 215, 295, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    },
    "Lakehouse Retreat": {
        "Cleaning + maintenance": [320, 320, 470, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Insurance":              [230, 230, 230, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Mortgage interest":      [340, 340, 340, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Repairs":                [  0,   0, 110, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Utilities":              [195, 210, 220, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Management fees":        [280, 265, 370, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Supplies":               [ 85,  85, 115, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Other":                  [185, 165, 220, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    },
    "Coastal Cottage": {
        "Cleaning + maintenance": [420, 400, 580, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Insurance":              [180, 180, 180, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Mortgage interest":      [360, 360, 360, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Repairs":                [ 90,   0, 150, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Utilities":              [205, 220, 225, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Management fees":        [365, 345, 425, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Supplies":               [115,  95, 140, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        "Other":                  [215, 205, 255, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    },
}


def _val(variant, demo_value):
    """Return demo_value for DEMO build, None for BLANK."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (hero + 4 KPI cards + nav)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Year-Over-Year Comparison"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Did this year actually go better — or did it just feel that way?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT — single-headline answer driven by Portfolio Roll-up totals.
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(\'Portfolio Roll-up\'!B7=0,'
        '"\U0001F4CA  Set your years on Settings, then enter PY + CY data to see your verdict.",'
        'IF((\'Portfolio Roll-up\'!B11-\'Portfolio Roll-up\'!B7)/\'Portfolio Roll-up\'!B7>=0,'
        '"✅  Revenue YoY = "&TEXT((\'Portfolio Roll-up\'!B11-\'Portfolio Roll-up\'!B7)/\'Portfolio Roll-up\'!B7,"+0.0%")'
        '&"  —  CY pacing ahead of PY.",'
        '"⚠  Revenue YoY = "&TEXT((\'Portfolio Roll-up\'!B11-\'Portfolio Roll-up\'!B7)/\'Portfolio Roll-up\'!B7,"+0.0%")'
        '&"  —  CY pacing behind PY; review Top Movers."))'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Side-by-side prior-year vs current-year P&L for up to three properties. "
        "Per-property 24-col matrix (12 PY months / 12 CY months) with $ delta + "
        "% delta. Portfolio rollup surfaces revenue YoY %, NOI YoY %, and the "
        "top-3 favorable + top-3 unfavorable category movers — the variance-analysis "
        "conversation a CPA would charge $300/hr to walk you through."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: 4 KPI cards (rows 18-23) ---
    # Card layout matches the brief: revenue YoY %, NOI YoY %, top mover,
    # top problem. Values pull from Portfolio Roll-up + Top Movers.
    for r in range(18, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1: Revenue YoY %
    ws.merge_cells("A18:C18")
    c = ws["A18"]
    c.value = "REVENUE YoY %"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A19:C20")
    c = ws["A19"]
    c.value = (
        "=IFERROR((\'Portfolio Roll-up\'!B11-\'Portfolio Roll-up\'!B7)"
        "/\'Portfolio Roll-up\'!B7,0)"
    )
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "+0.0%;-0.0%;0.0%"
    ws.merge_cells("A21:C21")
    c = ws["A21"]
    c.value = "portfolio gross"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2: NOI YoY %
    ws.merge_cells("D18:F18")
    c = ws["D18"]
    c.value = "NOI YoY %"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("D19:F20")
    c = ws["D19"]
    c.value = (
        "=IFERROR((\'Portfolio Roll-up\'!B19-\'Portfolio Roll-up\'!B15)"
        "/ABS(\'Portfolio Roll-up\'!B15),0)"
    )
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "+0.0%;-0.0%;0.0%"
    ws.merge_cells("D21:F21")
    c = ws["D21"]
    c.value = "net operating income"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3: Top mover (favorable)
    ws.merge_cells("G18:I18")
    c = ws["G18"]
    c.value = "TOP MOVER (+)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("G19:I20")
    c = ws["G19"]
    c.value = "=IFERROR('Top Movers'!B6 & \" / \" & 'Top Movers'!C6, \"—\")"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("G21:I21")
    c = ws["G21"]
    c.value = "=IFERROR(TEXT('Top Movers'!F6,\"$#,##0\"),\"\")"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 4: Top problem (unfavorable)
    ws.merge_cells("J18:L18")
    c = ws["J18"]
    c.value = "TOP PROBLEM (−)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("J19:L20")
    c = ws["J19"]
    c.value = "=IFERROR('Top Movers'!B14 & \" / \" & 'Top Movers'!C14, \"—\")"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("J21:L21")
    c = ws["J21"]
    c.value = "=IFERROR(TEXT('Top Movers'!F14,\"$#,##0\"),\"\")"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Gold borders on the 4 KPI cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(18, 22):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 18 else existing.top,
                    bottom=gold_side if r == 21 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 4: Primary "OPEN PER PROPERTY" button (rows 25-28) ---
    pseudo_button(ws, "A25", "L28",
                   "→  OPEN PER-PROPERTY COMPARISON",
                   "'Per Property'!A1", variant="primary")
    for r in range(25, 29):
        ws.row_dimensions[r].height = 20

    # --- ZONE 5: Secondary nav (rows 30-31) ---
    pseudo_button(ws, "A30", "C31", "Portfolio Roll-up",
                   "'Portfolio Roll-up'!A1", variant="secondary")
    pseudo_button(ws, "D30", "F31", "Top Movers",
                   "'Top Movers'!A1", variant="accent")
    pseudo_button(ws, "G30", "I31", "Settings",
                   "'Settings'!A1", variant="secondary")
    pseudo_button(ws, "J30", "L31", "\U0001F4C4 Print Pack",
                   "'Portfolio Roll-up'!A1", variant="secondary")
    ws.row_dimensions[30].height = 22
    ws.row_dimensions[31].height = 22

    # --- ZONE 6: Year-end ritual callout (row 33) ---
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "\U0001F4C5  YEAR-END RITUAL: paste prior-year totals from your P&L "
        "tracker into the PY columns, current-year YTD into CY, and let the "
        "deltas tell the story. Bump Settings B5/B7 each January."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 32

    # --- ZONE 7: Upgrade banner (row 35) ---
    add_upgrade_banner(ws, 35)

    # --- ZONE 8: Footer (rows 37-39) ---
    brand_footer(ws, 37,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:L39"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def _write_24col_header(ws, header_row):
    """Write the 24-col PY+CY matrix header at the given row.

    Layout:
      A: Category | B-M: Jan PY..Dec PY | N: Total PY | O-Z: Jan CY..Dec CY |
      AA: Total CY | AB: $ Δ | AC: % Δ
    """
    # Year-spanner row (header_row): "PRIOR YEAR" over B-N, "CURRENT YEAR" over O-AA
    ws.merge_cells(start_row=header_row, start_column=2,
                    end_row=header_row, end_column=14)
    c = ws.cell(row=header_row, column=2,
                 value="=\"PRIOR YEAR (\"&Settings!$B$7&\")\"")
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_NAVY_TINT_HEX())
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=header_row, start_column=15,
                    end_row=header_row, end_column=27)
    c = ws.cell(row=header_row, column=15,
                 value="=\"CURRENT YEAR (\"&Settings!$B$5&\")\"")
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=header_row, start_column=28,
                    end_row=header_row, end_column=29)
    c = ws.cell(row=header_row, column=28, value="DELTA")
    c.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Detail header row (header_row + 1)
    detail = header_row + 1
    headers = (["Category"] + MONTHS + ["Total PY"]
               + MONTHS + ["Total CY", "$ Δ", "% Δ"])
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=detail, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[header_row].height = 18
    ws.row_dimensions[detail].height = 20


def COLOR_NAVY_TINT_HEX():
    """Return the navy tint token. Inlined to keep raw-hex out of code."""
    # TODO: import COLOR_NAVY_TINT directly when refactoring brand_config import block
    from brand_config import COLOR_NAVY_TINT
    return COLOR_NAVY_TINT


def _write_property_section(ws, prop_idx, prop_name, start_row, variant):
    """Render one property block: title band, Revenue matrix, Expenses matrix.

    Returns the row immediately AFTER the section's last row.
    Layout per property:
      start_row    : property name band (gold-soft fill)
      start_row+1  : year-spanner header
      start_row+2  : detail header (Category, Jan..Dec PY, Total PY, Jan..Dec CY, Total CY, $Δ, %Δ)
      start_row+3  : "REVENUE" sub-header
      start_row+4..start_row+4+N-1 : revenue rows (single row in this template)
      ... revenue total
      blank
      "EXPENSES" sub-header
      expense rows
      expense total
      NET row
    """
    # Property name band — gold-soft, mono, all caps. Spans full 24 cols + delta.
    ws.merge_cells(start_row=start_row, start_column=1,
                    end_row=start_row, end_column=29)
    c = ws.cell(row=start_row, column=1,
                 value=f"PROPERTY {prop_idx}: {prop_name.upper()}")
    c.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[start_row].height = 24

    # Year-spanner + detail header at start_row+1, +2
    _write_24col_header(ws, start_row + 1)

    # REVENUE sub-header at start_row + 3
    rev_hdr = start_row + 3
    ws.cell(row=rev_hdr, column=1, value="REVENUE").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[rev_hdr].height = 20

    # Revenue rows (only one in this template — total rents + cleaning fees)
    rev_data_start = rev_hdr + 1
    for i, label in enumerate(REVENUE_ROWS):
        row = rev_data_start + i

        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")

        # PY months (cols 2..13 = B..M)
        for m in range(12):
            cell = ws.cell(row=row, column=2 + m)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
            if variant == "demo":
                cell.value = DEMO_PY_REVENUE[prop_name][m]

        # PY total (col 14 = N)
        n_col = get_column_letter(14)
        cell = ws.cell(row=row, column=14,
                        value=f"=SUM(B{row}:M{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'

        # CY months (cols 15..26 = O..Z)
        for m in range(12):
            cell = ws.cell(row=row, column=15 + m)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
            if variant == "demo":
                cell.value = DEMO_CY_REVENUE[prop_name][m]

        # CY total (col 27 = AA)
        cell = ws.cell(row=row, column=27,
                        value=f"=SUM(O{row}:Z{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'

        # $ delta (col 28 = AB)
        cell = ws.cell(row=row, column=28,
                        value=f"=AA{row}-N{row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0'

        # % delta (col 29 = AC)
        cell = ws.cell(row=row, column=29,
                        value=f"=IFERROR((AA{row}-N{row})/N{row},0)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;[Red]-0.0%;0.0%"

        ws.row_dimensions[row].height = 16

    rev_last = rev_data_start + len(REVENUE_ROWS) - 1
    rev_total_row = rev_last + 1

    # Revenue subtotal row (single rev row, but keep the band consistent)
    a = ws.cell(row=rev_total_row, column=1, value="Total revenue")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    for col in list(range(2, 28)) + [28]:
        col_letter = get_column_letter(col)
        cell = ws.cell(row=rev_total_row, column=col,
                        value=f"=SUM({col_letter}{rev_data_start}:{col_letter}{rev_last})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0' if col != 28 else '"$"#,##0;[Red]-"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=rev_total_row, column=29,
                    value=f"=IFERROR((AA{rev_total_row}-N{rev_total_row})/N{rev_total_row},0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
    cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)
    ws.row_dimensions[rev_total_row].height = 18

    # Spacer
    spacer_1 = rev_total_row + 1
    ws.row_dimensions[spacer_1].height = 8

    # EXPENSES sub-header
    exp_hdr = spacer_1 + 1
    ws.cell(row=exp_hdr, column=1, value="EXPENSES").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[exp_hdr].height = 20

    # Expense rows
    exp_data_start = exp_hdr + 1
    for i, label in enumerate(EXPENSE_ROWS):
        row = exp_data_start + i

        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")

        for m in range(12):
            cell = ws.cell(row=row, column=2 + m)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
            if variant == "demo":
                cell.value = DEMO_PY_EXPENSES[prop_name][label][m]

        cell = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'

        for m in range(12):
            cell = ws.cell(row=row, column=15 + m)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
            if variant == "demo":
                cell.value = DEMO_CY_EXPENSES[prop_name][label][m]

        cell = ws.cell(row=row, column=27, value=f"=SUM(O{row}:Z{row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'

        cell = ws.cell(row=row, column=28, value=f"=AA{row}-N{row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0'

        cell = ws.cell(row=row, column=29,
                        value=f"=IFERROR((AA{row}-N{row})/N{row},0)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;[Red]-0.0%;0.0%"

        ws.row_dimensions[row].height = 16

    exp_last = exp_data_start + len(EXPENSE_ROWS) - 1
    exp_total_row = exp_last + 1

    a = ws.cell(row=exp_total_row, column=1, value="Total expenses")
    a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    a.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 29):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=exp_total_row, column=col,
                        value=f"=SUM({col_letter}{exp_data_start}:{col_letter}{exp_last})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0' if col != 28 else '"$"#,##0;[Red]-"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_ERROR)
    cell = ws.cell(row=exp_total_row, column=29,
                    value=f"=IFERROR((AA{exp_total_row}-N{exp_total_row})/N{exp_total_row},0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
    cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_ERROR)
    ws.row_dimensions[exp_total_row].height = 18

    # NET row
    net_row = exp_total_row + 2
    ws.row_dimensions[exp_total_row + 1].height = 8

    a = ws.cell(row=net_row, column=1, value="NET INCOME")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 28):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=net_row, column=col,
                        value=f"={col_letter}{rev_total_row}-{col_letter}{exp_total_row}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)
    # $ delta
    cell = ws.cell(row=net_row, column=28,
                    value=f"=AA{net_row}-N{net_row}")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)
    # % delta
    cell = ws.cell(row=net_row, column=29,
                    value=f"=IFERROR((AA{net_row}-N{net_row})/ABS(N{net_row}),0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
    cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)
    ws.row_dimensions[net_row].height = 22

    # Conditional formatting: red fill on negative deltas, green on positive
    ws.conditional_formatting.add(
        f"AB{rev_data_start}:AB{net_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                    fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        f"AB{rev_data_start}:AB{net_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                    fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )

    return net_row + 2  # 1 blank row after net


def build_per_property_tab(wb, variant):
    """Sheet 1 — Per Property (3 stacked sections, 24-col PY+CY matrix each)."""
    ws = wb.create_sheet("Per Property")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Per Property",
                         prev_tab="Start", next_tab="Portfolio Roll-up")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 30):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:AC4")
    c4 = ws["A4"]
    c4.value = (
        "Each property: 12 PY months (Jan..Dec) + 12 CY months + Total PY / "
        "Total CY / $ Δ / % Δ. Year labels driven by Settings B5 + B7."
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column widths: A wide for category, B-AA narrow for monthly cells,
    # AB-AC slightly wider for totals/deltas.
    widths = [("A", 32)]
    for i in range(2, 28):  # B..AA — 26 cols
        widths.append((get_column_letter(i), 9))
    widths.append(("AB", 11))
    widths.append(("AC", 9))
    set_col_widths(ws, widths)

    # Build 3 property sections, stacked vertically
    next_row = 6
    for idx, prop in enumerate(PROPERTIES, start=1):
        next_row = _write_property_section(ws, idx, prop, next_row, variant)
        # Insert spacer row between sections
        next_row += 1

    ws.freeze_panes = "B6"

    ws.print_area = f"A1:AC{next_row}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LEDGER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)


def build_portfolio_rollup(wb):
    """Sheet 2 — Portfolio Roll-up (big numbers + 2 charts)."""
    ws = wb.create_sheet("Portfolio Roll-up")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Portfolio Roll-up",
                         prev_tab="Per Property", next_tab="Top Movers")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Portfolio totals — PY vs CY across all properties on the Per Property tab"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 16), ("C", 16),
                        ("D", 14), ("E", 12), ("F", 12), ("G", 12),
                        ("H", 12), ("I", 12), ("J", 12), ("K", 12), ("L", 12)])

    bold_left = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    left_align = Alignment(horizontal="left", vertical="center")

    # The Per Property tab has 3 sections. Each section's revenue total row
    # is at a known offset: section starts at row 6, section spans
    # 4 (band+header+detail+REVENUE) + 1 rev row = 5 → rev_total = row+5,
    # then 2 (spacer + EXPENSES hdr) + 8 expense rows = +10 → exp_total = row+15,
    # then +2 → net = row+17.  Section length = 18 rows + 1 spacer = 19.
    # Section 1: rev_total=11, exp_total=21, net=23
    # Section 2: rev_total=30, exp_total=40, net=42
    # Section 3: rev_total=49, exp_total=59, net=61
    rev_totals = ["'Per Property'!N11", "'Per Property'!N30", "'Per Property'!N49"]
    rev_totals_cy = ["'Per Property'!AA11", "'Per Property'!AA30", "'Per Property'!AA49"]
    exp_totals = ["'Per Property'!N21", "'Per Property'!N40", "'Per Property'!N59"]
    exp_totals_cy = ["'Per Property'!AA21", "'Per Property'!AA40", "'Per Property'!AA59"]

    # Headers row 5
    ws.cell(row=5, column=1, value="Metric").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    h_py = ws.cell(row=5, column=2, value="=\"Prior (\"&Settings!$B$7&\")\"")
    apply_style(h_py, header_row_style())
    h_cy = ws.cell(row=5, column=3, value="=\"Current (\"&Settings!$B$5&\")\"")
    apply_style(h_cy, header_row_style())
    h_d = ws.cell(row=5, column=4, value="$ Δ")
    apply_style(h_d, header_row_style())
    h_p = ws.cell(row=5, column=5, value="% Δ")
    apply_style(h_p, header_row_style())
    ws.row_dimensions[5].height = 20

    # Row 6: spacer / section heading "REVENUE"
    ws.cell(row=6, column=1, value="REVENUE").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[6].height = 18

    # Per-property revenue rows (rows 7-9), portfolio total at row 11
    # Card 1 in Start uses B11 / B7 → portfolio rev PY at row 7,
    #   total at row 11. So layout:
    # Row 7: portfolio revenue PY total in B7. We have to be careful.
    # Plan: row 7 is the "Portfolio revenue" PY+CY summary.
    a = ws.cell(row=7, column=1, value="Portfolio revenue (all properties)")
    a.font = bold_left
    a.alignment = left_align
    cell_b = ws.cell(row=7, column=2,
                      value=f"={'+'.join(rev_totals)}")
    apply_style(cell_b, formula_cell_style())
    cell_b.number_format = '"$"#,##0'
    cell_b.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell_c = ws.cell(row=7, column=3, value=f"={'+'.join(rev_totals_cy)}")
    apply_style(cell_c, formula_cell_style())
    cell_c.number_format = '"$"#,##0'
    cell_c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell_d = ws.cell(row=7, column=4, value="=C7-B7")
    apply_style(cell_d, formula_cell_style())
    cell_d.number_format = '"$"#,##0;[Red]-"$"#,##0'
    cell_e = ws.cell(row=7, column=5, value="=IFERROR((C7-B7)/B7,0)")
    apply_style(cell_e, formula_cell_style())
    cell_e.number_format = "+0.0%;[Red]-0.0%;0.0%"
    ws.row_dimensions[7].height = 24

    # Per-property breakdown rows 8-10
    for i, prop in enumerate(PROPERTIES):
        r = 8 + i
        ws.cell(row=r, column=1, value=f"  {prop}").font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell = ws.cell(row=r, column=2, value=f"={rev_totals[i]}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=r, column=3, value=f"={rev_totals_cy[i]}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cell = ws.cell(row=r, column=5, value=f"=IFERROR((C{r}-B{r})/B{r},0)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
        ws.row_dimensions[r].height = 16

    # Row 11: portfolio revenue total — referenced by Start KPIs (B11, B7)
    # Start tab Card 1 expects: B7 (PY) and B11 (CY). So row 11 must be the CY
    # value. We've already used B7 as the PY portfolio rev in row 7. Use B11 for
    # CY portfolio rev (mirrored from C7).
    a = ws.cell(row=11, column=1, value="Portfolio revenue — CY pacing")
    a.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    a.alignment = left_align
    cell = ws.cell(row=11, column=2, value="=C7")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    ws.row_dimensions[11].height = 16

    # Row 12: spacer
    ws.row_dimensions[12].height = 8

    # Row 13: EXPENSES heading
    ws.cell(row=13, column=1, value="EXPENSES").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[13].height = 18

    # Row 14: Portfolio expenses PY+CY
    a = ws.cell(row=14, column=1, value="Portfolio expenses (all properties)")
    a.font = bold_left
    a.alignment = left_align
    cell = ws.cell(row=14, column=2, value=f"={'+'.join(exp_totals)}")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ERROR)
    cell = ws.cell(row=14, column=3, value=f"={'+'.join(exp_totals_cy)}")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ERROR)
    cell = ws.cell(row=14, column=4, value="=C14-B14")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
    cell = ws.cell(row=14, column=5, value="=IFERROR((C14-B14)/B14,0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
    ws.row_dimensions[14].height = 24

    # Row 15: NOI total — Start Card 2 expects B15 (PY NOI), B19 (CY NOI).
    a = ws.cell(row=15, column=1, value="NOI — Net operating income (PY)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = left_align
    cell = ws.cell(row=15, column=2, value="=B7-B14")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=15, column=3, value="=C7-C14")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell = ws.cell(row=15, column=4, value="=C15-B15")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
    cell = ws.cell(row=15, column=5, value="=IFERROR((C15-B15)/ABS(B15),0)")
    apply_style(cell, formula_cell_style())
    cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
    ws.row_dimensions[15].height = 24

    # Row 16-18: per-property NOI breakdown
    for i, prop in enumerate(PROPERTIES):
        r = 16 + i
        ws.cell(row=r, column=1, value=f"  {prop}").font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell = ws.cell(row=r, column=2,
                        value=f"={rev_totals[i]}-{exp_totals[i]}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=r, column=3,
                        value=f"={rev_totals_cy[i]}-{exp_totals_cy[i]}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cell = ws.cell(row=r, column=5,
                        value=f"=IFERROR((C{r}-B{r})/ABS(B{r}),0)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
        ws.row_dimensions[r].height = 16

    # Row 19: NOI CY mirror — referenced by Start Card 2.
    a = ws.cell(row=19, column=1, value="NOI — CY pacing")
    a.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    a.alignment = left_align
    cell = ws.cell(row=19, column=2, value="=C15")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED)
    ws.row_dimensions[19].height = 16

    # ----- Per-month portfolio sums for charts (rows 22-26) -----
    # Row 22: month labels
    ws.cell(row=22, column=1, value="MONTHLY (portfolio)").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[22].height = 20
    # Row 23: header band — "Month", PY1..PY12, CY1..CY12
    ws.cell(row=23, column=1, value="Month").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)
    for m, lbl in enumerate(MONTHS):
        cell = ws.cell(row=23, column=2 + m, value=lbl)
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center")

    # Row 24: PY revenue per month (sum of three properties' Jan PY..Dec PY rows)
    # Per-property revenue row offsets: section 1 row 10, sec 2 row 29, sec 3 row 48.
    # PY months are cols B..M (cols 2..13).
    a = ws.cell(row=24, column=1,
                 value="=\"Revenue PY (\"&Settings!$B$7&\")\"")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_PRIMARY)
    rev_property_rows = [10, 29, 48]
    for m in range(12):
        col_letter = get_column_letter(2 + m)
        terms = "+".join(f"'Per Property'!{col_letter}{r}" for r in rev_property_rows)
        cell = ws.cell(row=24, column=2 + m, value=f"={terms}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
    ws.row_dimensions[24].height = 16

    # Row 25: CY revenue per month (cols O..Z = 15..26 in Per Property)
    a = ws.cell(row=25, column=1,
                 value="=\"Revenue CY (\"&Settings!$B$5&\")\"")
    a.font = Font(name=FONT_BODY, size=11, color=COLOR_ACCENT)
    for m in range(12):
        col_letter = get_column_letter(15 + m)
        terms = "+".join(f"'Per Property'!{col_letter}{r}" for r in rev_property_rows)
        cell = ws.cell(row=25, column=2 + m, value=f"={terms}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
    ws.row_dimensions[25].height = 16

    # ----- Stacked-bar chart: PY vs CY per property (anchor G5) -----
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Portfolio Revenue: Prior vs Current Year"
    bar.style = 2
    bar.y_axis.title = "Revenue"
    bar.x_axis.title = "Property"
    bar.height = 9
    bar.width = 16
    # Data: B7:C10 — Portfolio total + 3 properties (rows 7..10), cols B-C
    data = Reference(ws, min_col=2, max_col=3, min_row=7, max_row=10)
    cats = Reference(ws, min_col=1, max_col=1, min_row=8, max_row=10)
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(cats)
    style_chart(bar)
    ws.add_chart(bar, "G5")

    # ----- Per-month line chart: CY vs PY (anchor G22) -----
    line = LineChart()
    line.title = "Per-month revenue: CY vs PY"
    line.style = 2
    line.y_axis.title = "Revenue"
    line.x_axis.title = "Month"
    line.height = 9
    line.width = 16
    line_data = Reference(ws, min_col=1, max_col=13, min_row=24, max_row=25)
    line.add_data(line_data, titles_from_data=True, from_rows=True)
    line_cats = Reference(ws, min_col=2, max_col=13, min_row=23, max_row=23)
    line.set_categories(line_cats)
    style_chart(line)
    ws.add_chart(line, "G22")

    ws.print_area = "A1:L40"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_top_movers(wb):
    """Sheet 3 — Top Movers (top-3 favorable + top-3 unfavorable categories)."""
    ws = wb.create_sheet("Top Movers")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Top Movers",
                         prev_tab="Portfolio Roll-up", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = (
        "Top-3 favorable + top-3 unfavorable expense moves across all properties. "
        "Driven by SMALL/LARGE on the per-property delta column."
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6), ("B", 28), ("C", 22), ("D", 14), ("E", 14),
        ("F", 14), ("G", 12), ("H", 8),
    ])

    # ----- Hidden helper block: enumerate (Category, Property, PY, CY, $Δ, %Δ) for all 24 expense rows -----
    # Layout in helpers: rows 30..53 (24 rows) — 3 properties × 8 categories.
    # Per-property expense rows on Per Property tab:
    #   Section 1 starts at row 6: expense rows 13..20 (exp_data_start=13, exp_last=20)
    #   Section 2 starts at row 25: expense rows 32..39
    #   Section 3 starts at row 44: expense rows 51..58
    helper_start = 30
    section_offsets = [(0, 13), (1, 32), (2, 51)]  # (prop_index, exp_data_start)

    helper_row = helper_start
    for prop_idx, exp_start in section_offsets:
        prop_name = PROPERTIES[prop_idx]
        for cat_idx, cat in enumerate(EXPENSE_ROWS):
            src_row = exp_start + cat_idx
            ws.cell(row=helper_row, column=1, value=cat).font = Font(
                name=FONT_BODY, size=9, color=COLOR_MUTED)
            ws.cell(row=helper_row, column=2, value=prop_name).font = Font(
                name=FONT_BODY, size=9, color=COLOR_MUTED)
            cell = ws.cell(row=helper_row, column=3,
                            value=f"='Per Property'!N{src_row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell = ws.cell(row=helper_row, column=4,
                            value=f"='Per Property'!AA{src_row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'
            cell = ws.cell(row=helper_row, column=5,
                            value=f"=D{helper_row}-C{helper_row}")
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
            helper_row += 1

    helper_end = helper_row - 1  # row 53

    # Hide the helper rows so the customer doesn't see them
    for r in range(helper_start, helper_end + 1):
        ws.row_dimensions[r].hidden = True

    # ----- Top-3 favorable (smallest deltas — most negative = expense decrease) -----
    fav_hdr = ws.cell(row=5, column=1, value="✅")
    fav_hdr.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    fav_title = ws.cell(row=5, column=2, value="TOP-3 FAVORABLE (expense decreases)")
    fav_title.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[5].height = 22

    headers = ["Rank", "Category", "Property", "PY", "CY", "$ Δ", "% Δ"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h) if col == 1 else None
        # Skipped — header rendered after data writes below for clarity
    # Rewrite headers cleanly above the data block (row 5 is the band; use row 6
    # as the first data row to match Start KPI references B6 / F6).
    # Refactor: put data rows 6..8 (top 3 favorable), then row 13 header for
    # unfavorable, then 14..16 data.
    # Actually Start KPI Card 3 expects B6, C6, F6. So row 6 = #1 favorable.

    # Top-3 favorable rows (row 6, 7, 8): pick the 3 smallest deltas (most negative)
    helper_delta_range = f"E{helper_start}:E{helper_end}"
    helper_cat_range = f"A{helper_start}:A{helper_end}"
    helper_prop_range = f"B{helper_start}:B{helper_end}"
    helper_py_range = f"C{helper_start}:C{helper_end}"
    helper_cy_range = f"D{helper_start}:D{helper_end}"

    for rank in range(1, 4):
        row = 5 + rank  # 6, 7, 8
        # Rank
        cell_rank = ws.cell(row=row, column=1, value=rank)
        cell_rank.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        cell_rank.alignment = Alignment(horizontal="center")
        # Category — INDEX/MATCH on the rank-th smallest delta
        # SMALL returns the rank-th smallest (most negative for fav)
        smallest = f"SMALL({helper_delta_range},{rank})"
        match = f"MATCH({smallest},{helper_delta_range},0)"
        cell = ws.cell(row=row, column=2,
                        value=f"=INDEX({helper_cat_range},{match})")
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell = ws.cell(row=row, column=3,
                        value=f"=INDEX({helper_prop_range},{match})")
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell = ws.cell(row=row, column=4,
                        value=f"=INDEX({helper_py_range},{match})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=row, column=5,
                        value=f"=INDEX({helper_cy_range},{match})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=row, column=6, value=f"={smallest}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_PRIMARY)
        cell = ws.cell(row=row, column=7,
                        value=f"=IFERROR(F{row}/D{row},0)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
        cell = ws.cell(row=row, column=8, value="↓")
        cell.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 18

    # Spacer
    ws.row_dimensions[10].height = 12

    # ----- Top-3 unfavorable (largest deltas — most positive = expense increase) -----
    unfav_hdr = ws.cell(row=12, column=1, value="⚠")
    unfav_hdr.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ERROR)
    unfav_title = ws.cell(row=12, column=2, value="TOP-3 UNFAVORABLE (expense increases)")
    unfav_title.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_ERROR)
    ws.row_dimensions[12].height = 22

    for rank in range(1, 4):
        row = 13 + rank  # 14, 15, 16  (Start Card 4 expects B14/C14/F14 for #1 problem)
        cell_rank = ws.cell(row=row, column=1, value=rank)
        cell_rank.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ERROR)
        cell_rank.alignment = Alignment(horizontal="center")
        largest = f"LARGE({helper_delta_range},{rank})"
        match = f"MATCH({largest},{helper_delta_range},0)"
        cell = ws.cell(row=row, column=2,
                        value=f"=INDEX({helper_cat_range},{match})")
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell = ws.cell(row=row, column=3,
                        value=f"=INDEX({helper_prop_range},{match})")
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell = ws.cell(row=row, column=4,
                        value=f"=INDEX({helper_py_range},{match})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=row, column=5,
                        value=f"=INDEX({helper_cy_range},{match})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=row, column=6, value=f"={largest}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True, color=COLOR_ERROR)
        cell = ws.cell(row=row, column=7,
                        value=f"=IFERROR(F{row}/D{row},0)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "+0.0%;[Red]-0.0%;0.0%"
        cell = ws.cell(row=row, column=8, value="↑")
        cell.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_ERROR)
        cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 18

    # Caveat at row 18
    ws.merge_cells("A18:H18")
    c = ws["A18"]
    c.value = (
        "ℹ  Movers ranked across categories × properties (24 rows). "
        "Ties broken arbitrarily by Excel — if two rows show identical deltas, "
        "MATCH returns the first."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[18].height = 28

    ws.print_area = "A1:H20"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    """Sheet 4 — Settings (Current year B5, Prior year B7, property list B9-B18)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Top Movers", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = (
        "Current year drives every CY label · Prior year drives every PY label · "
        "property list shown for reference"
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 16), ("C", 14)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # B5: Current year (REQUIRED — referenced everywhere as Settings!$B$5)
    a5 = ws.cell(row=5, column=1, value="Current year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=CURRENT_YEAR_DEFAULT)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    # Row 6: explainer for current year
    ws.merge_cells("A6:C6")
    c6 = ws["A6"]
    c6.value = (
        "Bump this each January. Drives every CY column header, the "
        "Start verdict cell, and every $Δ + %Δ on the Per Property tab."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 28

    # B7: Prior year (REQUIRED — referenced as Settings!$B$7)
    a7 = ws.cell(row=7, column=1, value="Prior year:")
    a7.font = bold_right
    a7.alignment = Alignment(horizontal="right", vertical="center")
    b7 = ws.cell(row=7, column=2, value=PRIOR_YEAR_DEFAULT)
    apply_style(b7, input_cell_style())
    b7.number_format = "0"
    ws.row_dimensions[7].height = 18

    # Row 8: spacer / explainer
    ws.merge_cells("A8:C8")
    c8 = ws["A8"]
    c8.value = (
        "Usually current_year - 1. Drives every PY column header on Per Property "
        "and the PY series in the portfolio charts."
    )
    c8.font = italic_muted
    c8.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[8].height = 28

    # B9-B18: Property list (10 rows of capacity; 3 used in DEMO)
    sect = ws.cell(row=9, column=1, value="Property list (B9..B18)")
    sect.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[9].height = 20

    for i in range(10):
        row = 9 + i  # rows 9..18 — but row 9 is the header. Actually brief says
        # "B9-B18 Property list", so put names in B9..B18 directly.
    # Re-read brief: "B9-B18 Property list". So B9 first slot.
    # Overwrite the header — put it in A9 instead, names in B9..B18.
    ws.cell(row=9, column=1).value = ""  # clear header label
    # Use rows 9..18 for property names; column A holds slot label.
    for i in range(10):
        row = 9 + i
        ws.cell(row=row, column=1, value=f"Property {i + 1}:").font = bold_right
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="right", vertical="center")
        b = ws.cell(row=row, column=2)
        apply_style(b, input_cell_style())
        if i < len(PROPERTIES):
            b.value = PROPERTIES[i]
        ws.row_dimensions[row].height = 16

    # Row 19: explainer for property list
    ws.merge_cells("A19:C19")
    c19 = ws["A19"]
    c19.value = (
        "Up to 10 property names — the Per Property tab uses the first 3. "
        "Edit Per Property section banners directly to swap properties."
    )
    c19.font = italic_muted
    c19.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[19].height = 28

    ws.print_area = "A1:C22"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Main build function
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_per_property_tab(wb, variant)
    build_portfolio_rollup(wb)
    build_top_movers(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Year-Over-Year Comparison Workbook{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Side-by-side prior-year vs current-year P&L for up to 3 STR properties. "
        "Surfaces revenue YoY %, NOI YoY %, top-3 favorable + unfavorable category movers."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
