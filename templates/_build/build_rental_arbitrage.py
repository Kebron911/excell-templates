"""Build ACQ-003 Rental Arbitrage Analyzer (v2.2 standard).

Wizard-mode tool for aspiring arbitrage operators — model the lease vs.
revenue spread, stress test the Year-2 rent escalator, validate an MTR
fallback, and produce a 1-page Landlord Pitch Page (white-label).

Generates two files:
  templates/_masters/ACQ-003-rental-arbitrage-analyzer-DEMO.xlsx
  templates/_masters/ACQ-003-rental-arbitrage-analyzer-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_NEAR_BLACK, COLOR_WHITE,
    STATE_BAD_FILL, STATE_GOOD_FILL, STATE_WARN_FILL,
)

SKU = "ACQ-003"
NAME = "rental-arbitrage-analyzer"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (per brief QA section — 2BR urban apartment)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Lease Terms
    "property_address":   "428 Linden Ave, Unit 3B",
    "city_state_zip":     "Denver, CO 80218",
    "unit_type":          "Apartment",
    "beds":               2,
    "baths":              1.5,
    "sqft":               980,
    "monthly_rent":       2400,
    "security_deposit":   4800,
    "lease_term_months":  12,
    "lease_start":        "2026-06-01",
    "sublet_allowed":     "Conditional",
    "str_zone_allowed":   "Yes",
    "insurance_required": "Yes",
    "pet_allowed":        "No",
    "renewal_escalator":  0.05,
    "early_term":         "60-day notice + 2 months' rent",
    "landlord_name":      "Kestrel Property Group, LLC",
    "landlord_contact":   "Maria Chen · maria@kestrelpg.com · (303) 555-0142",

    # Setup Costs (low/mid/high) — small sample, ~80-line model summarized to categories
    # Stored as (low, mid, high) per category
    "setup_kitchen":       (1200, 2000, 3200),
    "setup_bedding":       (1500, 2400, 3600),
    "setup_bath":          (400,  650,  900),
    "setup_living":        (1800, 2800, 4200),
    "setup_decor":         (600,  1100, 1800),
    "setup_tech":          (700,  1100, 1700),
    "setup_security":      (200,  450,  700),
    "setup_supplies":      (300,  500,  800),
    "setup_cleaning":      (150,  300,  500),
    "setup_smalltools":    (200,  400,  700),
    "setup_outdoor":       (300,  600,  1000),
    "setup_consumables":   (250,  450,  700),
    "setup_signage":       (100,  250,  400),
    "setup_buffer":        (500,  800,  1300),

    # Revenue Projection
    "adr":              185,
    "stabilized_occ":   0.70,
    "y1_occ":           0.42,
    "los":              2.8,
    "cleaning_fee":     145,
    "platform_pct":     0.15,
    "seasonality":      1.00,

    # Operating Math (monthly)
    "utilities":        220,
    "internet":         80,
    "software":         40,
    "cleaning_paid":    135,
    "supplies_per_to":  18,
    "marketing":        45,
    "insurance":        85,
    "other":            40,

    # MTR Fallback
    "mtr_rent":         3400,
    "mtr_occ":          0.90,
    "mtr_utilities":    240,
    "mtr_insurance":    85,
    "mtr_minor_clean":  90,

    # Landlord pitch page
    "tenant_name":      "Apex Stay Co.",
    "tenant_llc":       "Apex Stay Holdings LLC",
    "tenant_contact":   "Daniel Harrison · daniel@apexstay.co · (720) 555-0188",
    "tenant_business":  "Professionally managed mid-term + short-term housing operator.",
    "ref_1":            "James Whitman · Pearl St Properties · (303) 555-0119",
    "ref_2":            "Anita Reyes · Highland Holdings · (720) 555-0177",
    "prior_history":    "3 prior master-lease units in Denver metro since 2024 — zero late payments, all renewed.",
}

UNIT_TYPES = ["Apartment", "Condo", "Townhouse", "Single-family", "Duplex unit", "Other"]
YES_NO = ["Yes", "No"]
YES_NO_COND = ["Yes", "No", "Conditional"]
YES_NO_UNSURE = ["Yes", "No", "Unsure"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero band (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Rental Arbitrage Analyzer"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Will the numbers actually work — through Year 2?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: Property identifier + verdict (rows 10-22)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    # Pull property address from Lease Terms!C7 (value cell, not label B7)
    c.value = '=IF(\'Lease Terms\'!C7="","(enter property on Lease Terms tab)",\'Lease Terms\'!C7)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # VERDICT label
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "VERDICT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # Verdict readout pulls from 'Stress Test'!A4 (verdict block top-left)
    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = "='Stress Test'!B22"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 18

    # Y1 + Stabilized side-by-side cash flow (rows 17-19)
    ws.merge_cells("A17:F17")
    c = ws["A17"]
    c.value = "Y1 MONTHLY CASH FLOW"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G17:L17")
    c = ws["G17"]
    c.value = "STABILIZED MONTHLY CASH FLOW"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 18

    ws.merge_cells("A18:F19")
    c = ws["A18"]
    c.value = "='Operating Math'!C22"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("G18:L19")
    c = ws["G18"]
    c.value = "='Operating Math'!D22"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.row_dimensions[18].height = 28
    ws.row_dimensions[19].height = 18

    # Stress test + MTR fallback indicators (row 21-22)
    ws.merge_cells("A21:F21")
    c = ws["A21"]
    c.value = "='Stress Test'!B23"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G21:L21")
    c = ws["G21"]
    c.value = "='MTR Fallback'!B16"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[21].height = 24

    # ZONE 3: Pseudo-button nav (rows 25-27)
    pseudo_button(ws, "A25", "C27", "Lease Terms",
                  "'Lease Terms'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "Setup Costs",
                  "'Setup Costs'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Revenue",
                  "'Revenue Projection'!A1", variant="primary")
    pseudo_button(ws, "J25", "L27", "Operating",
                  "'Operating Math'!A1", variant="primary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "D31", "→  Stress Test",
                  "'Stress Test'!A1", variant="accent")
    pseudo_button(ws, "E29", "H31", "→  MTR Fallback",
                  "'MTR Fallback'!A1", variant="accent")
    pseudo_button(ws, "I29", "L31", "→  Landlord Pitch Page",
                  "'Landlord Pitch'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 34
    cta = (
        f"Want to underwrite owned properties too? Get the STR Deal Analyzer "
        f"(ACQ-001) at {BRAND_DOMAIN} — $47."
    )
    ws.merge_cells("A34:L34")
    c = ws["A34"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[34].height = 30

    brand_footer(ws, 36,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L38"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Lease Terms
# ---------------------------------------------------------------------------

def build_lease_terms_tab(wb, variant):
    ws = wb.create_sheet("Lease Terms")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 32), ("C", 22),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Lease Terms",
                        prev_tab="Start", next_tab="Setup Costs")

    # Section banner row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 1 of 4 — what the landlord is offering"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # PROPERTY
    _section_band(ws, 6, "PROPERTY")
    fields_prop = [
        (7,  "Property address:",   _val(variant, SAMPLE["property_address"]),  None,    ""),
        (8,  "City, State, Zip:",   _val(variant, SAMPLE["city_state_zip"]),    None,    ""),
        (9,  "Unit type:",          _val(variant, SAMPLE["unit_type"]),         None,    "Dropdown"),
        (10, "Bedrooms:",           _val(variant, SAMPLE["beds"]),              "0",     ""),
        (11, "Bathrooms:",          _val(variant, SAMPLE["baths"]),             "0.0",   ""),
        (12, "Square feet:",        _val(variant, SAMPLE["sqft"]),              "#,##0", ""),
    ]
    for row, label, value, fmt, note in fields_prop:
        _input_row(ws, row, label, value, fmt, note)
    add_dropdown(ws, "C9", f'"{",".join(UNIT_TYPES)}"')

    # LEASE FINANCIALS
    _section_band(ws, 14, "LEASE FINANCIALS")
    fields_lease = [
        (15, "Monthly rent ($):",        _val(variant, SAMPLE["monthly_rent"]),     '"$"#,##0', "What the landlord wants you to pay"),
        (16, "Security deposit ($):",    _val(variant, SAMPLE["security_deposit"]), '"$"#,##0', "Typically 1-2 months' rent"),
        (17, "Lease term (months):",     _val(variant, SAMPLE["lease_term_months"]), "0",       "12 typical"),
        (18, "Lease start date:",        _val(variant, SAMPLE["lease_start"]),       None,      ""),
        (19, "Renewal escalator (%):",   _val(variant, SAMPLE["renewal_escalator"]), "0.0%",    "Year-2+ rent hike clause. Default 5%."),
    ]
    for row, label, value, fmt, note in fields_lease:
        _input_row(ws, row, label, value, fmt, note)

    # RESTRICTIONS — these are the deal-killers
    _section_band(ws, 21, "RESTRICTIONS  (the deal-killers)")
    fields_restrict = [
        (22, "Sublet allowed?",        _val(variant, SAMPLE["sublet_allowed"]),     None, "Yes / No / Conditional. NO = STR illegal under this lease."),
        (23, "STR-zone allowed?",      _val(variant, SAMPLE["str_zone_allowed"]),   None, "City zoning permits short-term rentals?"),
        (24, "Insurance addendum req.?", _val(variant, SAMPLE["insurance_required"]), None, "Most landlord-friendly. Plan to carry $1M liability."),
        (25, "Pets allowed?",          _val(variant, SAMPLE["pet_allowed"]),        None, "Yes / No"),
    ]
    for row, label, value, fmt, note in fields_restrict:
        _input_row(ws, row, label, value, fmt, note)
    add_dropdown(ws, "C22", f'"{",".join(YES_NO_COND)}"')
    add_dropdown(ws, "C23", f'"{",".join(YES_NO_UNSURE)}"')
    add_dropdown(ws, "C24", f'"{",".join(YES_NO)}"')
    add_dropdown(ws, "C25", f'"{",".join(YES_NO)}"')

    # Restriction red-flag row 27
    ws.merge_cells("A27:L27")
    c = ws["A27"]
    c.value = (
        '=IF(OR(C22="No",C23="No"),'
        '"RED FLAG: Sublet or STR-zone is NO — this property cannot legally run as STR. The Verdict will lock to PASS.",'
        'IF(OR(C22="Conditional",C23="Unsure"),'
        '"YELLOW FLAG: Get the sublet allowance and STR-zoning in WRITING from the landlord and city before signing.",'
        '"All restrictions clear — STR-eligible under this lease."))'
    )
    c.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
    ws.row_dimensions[27].height = 30

    # EARLY-TERM + LANDLORD INFO
    _section_band(ws, 29, "EARLY-TERM CLAUSE  +  LANDLORD CONTACT")
    fields_lord = [
        (30, "Early-termination clause:", _val(variant, SAMPLE["early_term"]),       None, "Free text. Most arbitrage operators want a 60-day exit."),
        (31, "Landlord name / entity:",   _val(variant, SAMPLE["landlord_name"]),     None, ""),
        (32, "Landlord contact:",         _val(variant, SAMPLE["landlord_contact"]),  None, "Name · email · phone"),
    ]
    for row, label, value, fmt, note in fields_lord:
        _input_row(ws, row, label, value, fmt, note)

    brand_footer(ws, 35, version_line=f"{SKU} · Lease Terms")


# ---------------------------------------------------------------------------
# Sheet 3 — Setup Costs
# ---------------------------------------------------------------------------

def build_setup_costs_tab(wb, variant):
    ws = wb.create_sheet("Setup Costs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 30),
        ("C", 14), ("D", 14), ("E", 14),
        ("F", 6), ("G", 32),
        ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Setup Costs",
                        prev_tab="Lease Terms", next_tab="Revenue Projection")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Section 2 of 4 — one-time furnish + tech + supplies. Smaller than owned-property "
        "setup (no major appliances — landlord supplies those)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Header row 6
    headers = [
        ("B6", "Category"),
        ("C6", "Low ($)"),
        ("D6", "Mid ($)"),
        ("E6", "High ($)"),
        ("G6", "Note"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Categories
    categories = [
        ("Kitchen — cookware/plates/utensils",  SAMPLE["setup_kitchen"],     "Coffee maker, knives, full kit for 6"),
        ("Bedding — mattresses/linens/towels",  SAMPLE["setup_bedding"],     "Hotel-grade for premium ADR"),
        ("Bath — toilet brush/scale/shower",    SAMPLE["setup_bath"],        "Soft launch consumables included"),
        ("Living — sofa/coffee table/TV",       SAMPLE["setup_living"],      "Sleeper sofa raises occupancy"),
        ("Decor — art/rug/lamps",               SAMPLE["setup_decor"],       "Photogenic = better listing photos"),
        ("Tech — smart locks/Wi-Fi/cameras",    SAMPLE["setup_tech"],        "Schlage / eero / Ring Outdoor"),
        ("Security — exterior cameras/sensors", SAMPLE["setup_security"],    "Required by some city STR ordinances"),
        ("Cleaning supplies (start kit)",       SAMPLE["setup_supplies"],    "Mop / vacuum / starter chemicals"),
        ("Initial deep clean",                  SAMPLE["setup_cleaning"],    "Pro deep-clean before listing"),
        ("Small tools / fasteners / hooks",     SAMPLE["setup_smalltools"],  "Drill, level, picture hangers"),
        ("Outdoor — patio set / firepit",       SAMPLE["setup_outdoor"],     "If unit has outdoor space"),
        ("Consumables — soap/coffee/paper",     SAMPLE["setup_consumables"], "First 3 months of restock"),
        ("Signage / lockbox / welcome book",    SAMPLE["setup_signage"],     "Includes printed welcome book"),
        ("Buffer (5-8% of mid total)",          SAMPLE["setup_buffer"],      "Always overruns. Always."),
    ]

    start_row = 7
    for i, (label, costs, note) in enumerate(categories):
        r = start_row + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        for col_idx, cost in zip((3, 4, 5), costs):
            cell = ws.cell(row=r, column=col_idx, value=_val(variant, cost))
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0'
        ws.cell(row=r, column=7, value=note).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=r, column=7).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[r].height = 18

    # Totals row
    total_row = start_row + len(categories) + 1
    ws.cell(row=total_row, column=2, value="TOTAL SETUP COST").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=total_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    last_cat_row = start_row + len(categories) - 1
    for col_idx, color in zip((3, 4, 5), (COLOR_PRIMARY, COLOR_ACCENT, COLOR_SECONDARY)):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_row, column=col_idx,
                        value=f"=SUM({col_letter}{start_row}:{col_letter}{last_cat_row})")
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'
        cell.border = Border(
            top=Side(style="medium", color=COLOR_ACCENT),
            bottom=Side(style="medium", color=COLOR_ACCENT),
        )
    ws.row_dimensions[total_row].height = 26

    # Cash to launch (deposit + setup mid)
    cash_row = total_row + 2
    ws.cell(row=cash_row, column=2, value="Cash to launch (deposit + Mid setup):").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=cash_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=cash_row, column=4,
                    value=f"='Lease Terms'!C16+D{total_row}")
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[cash_row].height = 24

    brand_footer(ws, cash_row + 3, version_line=f"{SKU} · Setup Costs")


# ---------------------------------------------------------------------------
# Sheet 4 — Revenue Projection
# ---------------------------------------------------------------------------

def build_revenue_tab(wb, variant):
    ws = wb.create_sheet("Revenue Projection")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Revenue Projection",
                        prev_tab="Setup Costs", next_tab="Operating Math")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 3 of 4 — paste from AirDNA / Rabbu / KeyData. Haircut Y1 hard."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "MARKET ASSUMPTIONS")
    fields = [
        (7,  "ADR — avg nightly rate ($):",    _val(variant, SAMPLE["adr"]),            '"$"#,##0', "Paste from AirDNA"),
        (8,  "Stabilized occupancy (%):",       _val(variant, SAMPLE["stabilized_occ"]), "0.0%",     "Year 2-3 expected"),
        (9,  "Year 1 occupancy (%):",            _val(variant, SAMPLE["y1_occ"]),         "0.0%",     "Default 60% of stabilized — arbitrage Y1 is harder"),
        (10, "Avg length of stay (nights):",     _val(variant, SAMPLE["los"]),            "0.0",      "1.5-7 typical"),
        (11, "Cleaning fee charged ($):",        _val(variant, SAMPLE["cleaning_fee"]),   '"$"#,##0', "What guests pay (separate from cleaning paid)"),
        (12, "Platform commission (%):",         _val(variant, SAMPLE["platform_pct"]),   "0.0%",     "Airbnb 15%, VRBO 8%, direct 0-3%"),
        (13, "Seasonality factor (multiplier):", _val(variant, SAMPLE["seasonality"]),    "0.00",     "1.00 = neutral; 1.10 = peak market; 0.90 = off-peak"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    # Derived
    _section_band(ws, 15, "DERIVED  (monthly)")
    # Formulas — referencing cells on this tab. C7=ADR, C8=stab occ, C9=Y1 occ, C10=LOS, C11=cleaning fee, C12=platform pct, C13=season
    # Booked nights (Y1 monthly) = C9 * 30.42 (avg days/mo) * C13
    _output_row(ws, 16, "Y1 booked nights / mo:",
                "=C9*30.42*C13", "0.0")
    _output_row(ws, 17, "Stabilized booked nights / mo:",
                "=C8*30.42*C13", "0.0")
    _output_row(ws, 18, "Y1 turnovers / mo:",
                "=C16/C10", "0.0")
    _output_row(ws, 19, "Stabilized turnovers / mo:",
                "=C17/C10", "0.0")
    _output_row(ws, 20, "Y1 gross revenue / mo:",
                "=C16*C7+C18*C11", '"$"#,##0')
    _output_row(ws, 21, "Stabilized gross revenue / mo:",
                "=C17*C7+C19*C11", '"$"#,##0')
    _output_row(ws, 22, "Y1 net revenue / mo (after platform):",
                "=C20*(1-C12)", '"$"#,##0', emphasize=True)
    _output_row(ws, 23, "Stabilized net revenue / mo:",
                "=C21*(1-C12)", '"$"#,##0', emphasize=True)

    # AirDNA paste callout
    ws.merge_cells("A26:L26")
    c = ws["A26"]
    c.value = "PASTE FROM AIRDNA / RABBU / KEYDATA"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[26].height = 20

    ws.merge_cells("A27:L29")
    c = ws["A27"]
    c.value = (
        "Pull market RevPAR, ADR, and occupancy from AirDNA Markets. Look for the "
        "same beds/baths/unit type. Then haircut: arbitrage Y1 occupancy is roughly "
        "60% of stabilized, not 80%. Ramp time is real — listings need ~50 reviews "
        "before they crack the algorithm."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    for r in range(27, 30):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 32, version_line=f"{SKU} · Revenue")


# ---------------------------------------------------------------------------
# Sheet 5 — Operating Math
# ---------------------------------------------------------------------------

def build_operating_math_tab(wb, variant):
    ws = wb.create_sheet("Operating Math")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 16), ("D", 16),
        ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Operating Math",
                        prev_tab="Revenue Projection", next_tab="Stress Test")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 4 of 4 — monthly cash flow. Y1 vs Stabilized side-by-side."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 6
    headers = [
        ("B6", "Line item"),
        ("C6", "Year 1 / mo"),
        ("D6", "Stabilized / mo"),
        ("E6", "Note"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Inputs — operating monthly costs (single column input C, but we will replicate to D)
    _section_band(ws, 7, "MONTHLY OPERATING COSTS")
    op_fields = [
        (8,  "Rent (from Lease Terms):",  "='Lease Terms'!C15", '"$"#,##0', "Auto-pulled"),
        (9,  "Utilities ($):",             _val(variant, SAMPLE["utilities"]), '"$"#,##0', "Electric/gas/water/trash"),
        (10, "Internet ($):",              _val(variant, SAMPLE["internet"]),  '"$"#,##0', ""),
        (11, "Software / PMS ($):",        _val(variant, SAMPLE["software"]),  '"$"#,##0', "Hospitable, Hostfully, etc."),
        (12, "Cleaning paid ($/turnover):", _val(variant, SAMPLE["cleaning_paid"]), '"$"#,##0', "What you pay your cleaner"),
        (13, "Supplies ($/turnover):",     _val(variant, SAMPLE["supplies_per_to"]), '"$"#,##0', "Soap, paper, coffee restock"),
        (14, "Marketing ($):",             _val(variant, SAMPLE["marketing"]),  '"$"#,##0', "Photos refresh, Pinterest ads"),
        (15, "Insurance ($):",             _val(variant, SAMPLE["insurance"]),  '"$"#,##0', "Proper / Slice STR-specific"),
        (16, "Other ($):",                 _val(variant, SAMPLE["other"]),      '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in op_fields:
        # Custom row layout — input C, mirror to D
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        if isinstance(value, str) and value.startswith("="):
            # Formula — set on both C and D
            cell_c = ws.cell(row=row, column=3, value=value)
            apply_style(cell_c, formula_cell_style())
            cell_c.alignment = Alignment(horizontal="center", vertical="center")
            cell_c.number_format = fmt
            cell_d = ws.cell(row=row, column=4, value=f"=C{row}")
            apply_style(cell_d, formula_cell_style())
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            cell_d.number_format = fmt
        else:
            cell_c = ws.cell(row=row, column=3, value=value)
            apply_style(cell_c, input_cell_style())
            cell_c.alignment = Alignment(horizontal="center", vertical="center")
            cell_c.number_format = fmt
            # Stabilized column mirrors Y1 by default (user can override)
            cell_d = ws.cell(row=row, column=4, value=value)
            apply_style(cell_d, input_cell_style())
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            cell_d.number_format = fmt
        ws.cell(row=row, column=5, value=note).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=row, column=5).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[row].height = 18

    # ROLL-UP rows 18-22
    _section_band(ws, 18, "ROLL-UP")
    # Total operating cost = rent + utilities + internet + software + marketing + insurance + other + (cleaning+supplies)*turnovers
    # Y1 turnovers: 'Revenue Projection'!C18; Stab turnovers: 'Revenue Projection'!C19
    y1_op = (
        "=C8+C9+C10+C11+C14+C15+C16"
        "+(C12+C13)*'Revenue Projection'!C18"
    )
    stab_op = (
        "=D8+D9+D10+D11+D14+D15+D16"
        "+(D12+D13)*'Revenue Projection'!C19"
    )
    ws.cell(row=19, column=2, value="Total operating cost / mo:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=19, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell_c = ws.cell(row=19, column=3, value=y1_op)
    apply_style(cell_c, formula_cell_style())
    cell_c.alignment = Alignment(horizontal="center", vertical="center")
    cell_c.number_format = '"$"#,##0'
    cell_d = ws.cell(row=19, column=4, value=stab_op)
    apply_style(cell_d, formula_cell_style())
    cell_d.alignment = Alignment(horizontal="center", vertical="center")
    cell_d.number_format = '"$"#,##0'
    ws.row_dimensions[19].height = 20

    # Net revenue (from Revenue Projection)
    ws.cell(row=20, column=2, value="Net revenue / mo:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=20, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell_c = ws.cell(row=20, column=3, value="='Revenue Projection'!C22")
    apply_style(cell_c, formula_cell_style())
    cell_c.alignment = Alignment(horizontal="center", vertical="center")
    cell_c.number_format = '"$"#,##0'
    cell_d = ws.cell(row=20, column=4, value="='Revenue Projection'!C23")
    apply_style(cell_d, formula_cell_style())
    cell_d.alignment = Alignment(horizontal="center", vertical="center")
    cell_d.number_format = '"$"#,##0'
    ws.row_dimensions[20].height = 20

    # MONTHLY CASH FLOW (the key output)
    ws.cell(row=22, column=2, value="MONTHLY CASH FLOW:").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=22, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell_c = ws.cell(row=22, column=3, value="=C20-C19")
    cell_c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    cell_c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell_c.alignment = Alignment(horizontal="center", vertical="center")
    cell_c.number_format = '"$"#,##0'
    cell_d = ws.cell(row=22, column=4, value="=D20-D19")
    cell_d.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    cell_d.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell_d.alignment = Alignment(horizontal="center", vertical="center")
    cell_d.number_format = '"$"#,##0'
    ws.row_dimensions[22].height = 28

    # Annual rollup
    ws.cell(row=23, column=2, value="Annual cash flow:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=23, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell_c = ws.cell(row=23, column=3, value="=C22*12")
    apply_style(cell_c, formula_cell_style())
    cell_c.alignment = Alignment(horizontal="center", vertical="center")
    cell_c.number_format = '"$"#,##0'
    cell_d = ws.cell(row=23, column=4, value="=D22*12")
    apply_style(cell_d, formula_cell_style())
    cell_d.alignment = Alignment(horizontal="center", vertical="center")
    cell_d.number_format = '"$"#,##0'
    ws.row_dimensions[23].height = 18

    # Conditional formatting on monthly cash flow row 22
    green_fill = PatternFill("solid", fgColor=STATE_GOOD_FILL)
    yellow_fill = PatternFill("solid", fgColor=STATE_WARN_FILL)
    red_fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    for col_letter in ("C", "D"):
        ws.conditional_formatting.add(
            f"{col_letter}22",
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=red_fill)
        )
        ws.conditional_formatting.add(
            f"{col_letter}22",
            CellIsRule(operator="between", formula=["0", "300"], stopIfTrue=False, fill=yellow_fill)
        )
        ws.conditional_formatting.add(
            f"{col_letter}22",
            CellIsRule(operator="greaterThan", formula=["300"], stopIfTrue=False, fill=green_fill)
        )

    brand_footer(ws, 26, version_line=f"{SKU} · Operating Math")


# ---------------------------------------------------------------------------
# Sheet 6 — Stress Test
# ---------------------------------------------------------------------------

def build_stress_test_tab(wb, variant):
    ws = wb.create_sheet("Stress Test")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 18), ("D", 18), ("E", 18),
        ("F", 6), ("G", 30),
        ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Stress Test",
                        prev_tab="Operating Math", next_tab="MTR Fallback")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "The arbitrage killer is the Year-2 rent escalator. Most gurus skip it. "
        "This tab bakes a 5% escalator into Year 2 and a 7%+ shock case."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Headers row 6
    headers = [
        ("B6", "Metric"),
        ("C6", "Base (Y1)"),
        ("D6", "Y2 + 5% escalator"),
        ("E6", "Worst case"),
        ("G6", "Method"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # Build formulas using value cells from Operating Math + Revenue Projection
    # Y1 cash flow direct
    base_cf = "'Operating Math'!C22"
    # Y2 + 5% escalator: use stabilized cash flow (Operating Math!D22) and subtract the rent delta.
    # Rent delta = Lease Terms!C15 * Lease Terms!C19 (escalator pct)
    y2_cf = "('Operating Math'!D22 - 'Lease Terms'!C15*'Lease Terms'!C19)"
    # Worst: stabilized rent +7%, occupancy -10pp, supplies +15%
    # = stabilized net revenue at (stab_occ - 0.10) - stabilized opex - rent*7% - supplies*15%
    # Simplified: stabilized cf - rent*0.07 - rent_increment - occ_revenue_loss - supplies_15
    # We approximate worst case = D22 - rent*0.07 - (rev/occ * 0.10) - (supplies bump)
    # Cleaner: derive worst-case revenue and worst-case opex directly
    worst_rev = (
        # adjusted occ = stab - 0.10
        "((('Revenue Projection'!C8-0.10)*30.42*'Revenue Projection'!C13)*'Revenue Projection'!C7"
        " + (('Revenue Projection'!C8-0.10)*30.42*'Revenue Projection'!C13/'Revenue Projection'!C10)*'Revenue Projection'!C11)"
        "*(1-'Revenue Projection'!C12)"
    )
    worst_opex = (
        # rent up 7%, supplies up 15%, otherwise stabilized
        "('Lease Terms'!C15*1.07 + 'Operating Math'!D9 + 'Operating Math'!D10 + 'Operating Math'!D11 + 'Operating Math'!D14 + 'Operating Math'!D15 + 'Operating Math'!D16"
        " + ('Operating Math'!D12 + 'Operating Math'!D13*1.15)"
        "*((('Revenue Projection'!C8-0.10)*30.42*'Revenue Projection'!C13)/'Revenue Projection'!C10))"
    )
    worst_cf = f"({worst_rev} - {worst_opex})"

    # Restriction lock: if sublet=No or zone=No, force ❌ across the board
    # We don't override the cash flow numbers, but we do override the verdict.

    # Metric rows 7-12
    metrics = [
        ("Monthly cash flow:",
         f"={base_cf}", f"={y2_cf}", f"={worst_cf}",
         '"$"#,##0',
         "Net rev − total opex"),
        ("Annual cash flow:",
         f"={base_cf}*12", f"={y2_cf}*12", f"={worst_cf}*12",
         '"$"#,##0',
         "Monthly × 12"),
        ("Effective rent paid:",
         "='Lease Terms'!C15", "='Lease Terms'!C15*(1+'Lease Terms'!C19)", "='Lease Terms'!C15*1.07",
         '"$"#,##0',
         "Year-2 escalator applied"),
        ("Net revenue / mo:",
         "='Revenue Projection'!C22", "='Revenue Projection'!C23", f"={worst_rev}",
         '"$"#,##0',
         "After platform commission"),
        ("Cash margin:",
         f"=IFERROR({base_cf}/'Revenue Projection'!C22,0)",
         f"=IFERROR({y2_cf}/'Revenue Projection'!C23,0)",
         f"=IFERROR({worst_cf}/{worst_rev},0)",
         "0.0%",
         "Cash flow ÷ net revenue"),
    ]
    for i, (label, fy1, fy2, fworst, fmt, note) in enumerate(metrics):
        r = 7 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11,
            bold=(label.startswith("Monthly cash flow")),
            color=COLOR_TEXT,
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col_idx, formula in [(3, fy1), (4, fy2), (5, fworst)]:
            cell = ws.cell(row=r, column=col_idx, value=formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = fmt
            if label.startswith("Monthly cash flow"):
                cell.font = Font(name=FONT_HEAD, size=12, bold=True,
                                  color=(COLOR_PRIMARY if col_idx == 3 else
                                         COLOR_ACCENT if col_idx == 4 else
                                         COLOR_SECONDARY))
        ws.cell(row=r, column=7, value=note).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=r, column=7).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[r].height = 20

    # Per-scenario verdict row 13
    ws.cell(row=13, column=2, value="Verdict:").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=13, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    for col_idx, cf in [(3, base_cf), (4, y2_cf), (5, worst_cf)]:
        formula = (
            f'=IF({cf}>300,"DEAL",'
            f'IF({cf}>0,"MARGINAL","FAIL"))'
        )
        cell = ws.cell(row=13, column=col_idx, value=formula)
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 26

    # Conditional formatting on cash flow row 7
    green_fill = PatternFill("solid", fgColor=STATE_GOOD_FILL)
    yellow_fill = PatternFill("solid", fgColor=STATE_WARN_FILL)
    red_fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    for col_letter in ("C", "D", "E"):
        ws.conditional_formatting.add(
            f"{col_letter}7",
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=red_fill)
        )
        ws.conditional_formatting.add(
            f"{col_letter}7",
            CellIsRule(operator="between", formula=["0", "300"], stopIfTrue=False, fill=yellow_fill)
        )
        ws.conditional_formatting.add(
            f"{col_letter}7",
            CellIsRule(operator="greaterThan", formula=["300"], stopIfTrue=False, fill=green_fill)
        )

    # Methodology footnote
    ws.merge_cells("A16:L19")
    c = ws["A16"]
    c.value = (
        "Stress test method: Year-2 column applies the renewal escalator from Lease Terms "
        "directly to rent, holds occupancy at stabilized. Worst-case applies a 7% rent shock, "
        "drops occupancy 10 percentage points, and adds 15% to supplies cost. If the worst-case "
        "column shows DEAL or MARGINAL — your underwriting has a real safety margin. If it shows "
        "FAIL — the deal works only if your assumptions hold perfectly. Most don't."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    for r in range(16, 20):
        ws.row_dimensions[r].height = 18

    # OVERALL VERDICT (rows 21-23) — referenced by Start tab
    ws.merge_cells("B21:E21")
    c = ws["B21"]
    c.value = "OVERALL VERDICT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[21].height = 18

    # B22 — verdict string
    # Restriction lock: sublet=No or STR-zone=No → forced PASS
    # Otherwise base on Y1 + worst-case
    ws.merge_cells("B22:E22")
    c = ws["B22"]
    c.value = (
        f'=IF(OR(\'Lease Terms\'!C22="No",\'Lease Terms\'!C23="No"),'
        f'"PASS — illegal under current lease",'
        f'IF(AND({base_cf}>0,{worst_cf}>0),"DEAL",'
        f'IF({base_cf}>0,"MARGINAL — survives Y1 but fails worst case",'
        f'"PASS — Y1 cash flow negative")))'
    )
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[22].height = 36

    # B23 — stress test summary line referenced by Start
    ws.merge_cells("B23:E23")
    c = ws["B23"]
    c.value = (
        f'=IF({worst_cf}>0,"Stress survives ($"&TEXT({worst_cf},"#,##0")&"/mo worst)",'
        f'IF({worst_cf}>-200,"Stress marginal (−$"&TEXT(ABS({worst_cf}),"#,##0")&"/mo)",'
        f'"Stress fails (−$"&TEXT(ABS({worst_cf}),"#,##0")&"/mo)"))'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    ws.row_dimensions[23].height = 24

    brand_footer(ws, 26, version_line=f"{SKU} · Stress Test")


# ---------------------------------------------------------------------------
# Sheet 7 — MTR Fallback
# ---------------------------------------------------------------------------

def build_mtr_fallback_tab(wb, variant):
    ws = wb.create_sheet("MTR Fallback")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "MTR Fallback",
                        prev_tab="Stress Test", next_tab="Landlord Pitch")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Plan B: if the STR play fails, can mid-term-rental (corporate housing, "
        "travel nurses, insurance displacement) cover the rent? Real arbitrage "
        "operators have an exit. Gurus don't."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 28

    _section_band(ws, 6, "MTR INPUTS  (paste from Furnished Finder)")
    fields = [
        (7,  "MTR rent / mo ($):",       _val(variant, SAMPLE["mtr_rent"]),       '"$"#,##0', "30+ day stays. Furnished Finder 'rent' field."),
        (8,  "MTR occupancy (%):",        _val(variant, SAMPLE["mtr_occ"]),        "0.0%",     "Once contracted, ~90% on signed term"),
        (9,  "Utilities ($/mo):",         _val(variant, SAMPLE["mtr_utilities"]),  '"$"#,##0', "Sometimes pass-through to MTR tenant"),
        (10, "Insurance ($/mo):",         _val(variant, SAMPLE["mtr_insurance"]),  '"$"#,##0', ""),
        (11, "Minor cleaning ($/mo):",    _val(variant, SAMPLE["mtr_minor_clean"]), '"$"#,##0', "Quarterly turnover ÷ 3"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    # Derived
    _section_band(ws, 13, "MTR CASH FLOW")
    # Net MTR revenue = mtr_rent * mtr_occ (no platform fee, no cleaning fee per stay)
    # MTR opex = utilities + insurance + minor cleaning
    # Cash flow = net rev - rent - opex
    _output_row(ws, 14, "Net MTR revenue / mo:",
                "=C7*C8", '"$"#,##0')
    _output_row(ws, 15, "MTR cash flow / mo:",
                "=C14-'Lease Terms'!C15-C9-C10-C11", '"$"#,##0', emphasize=True)

    # Verdict B16 — referenced by Start tab
    ws.merge_cells("B16:L16")
    c = ws["B16"]
    c.value = (
        '=IF(C15>0,"MTR covers rent: $"&TEXT(C15,"#,##0")&"/mo cushion",'
        'IF(C15>-300,"MTR shortfall: $"&TEXT(ABS(C15),"#,##0")&"/mo (manageable)",'
        '"MTR cannot cover rent: $"&TEXT(ABS(C15),"#,##0")&"/mo gap"))'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 28

    # Conditional formatting on row 15 cell C15
    green_fill = PatternFill("solid", fgColor=STATE_GOOD_FILL)
    yellow_fill = PatternFill("solid", fgColor=STATE_WARN_FILL)
    red_fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    ws.conditional_formatting.add(
        "C15",
        CellIsRule(operator="lessThan", formula=["-300"], stopIfTrue=False, fill=red_fill)
    )
    ws.conditional_formatting.add(
        "C15",
        CellIsRule(operator="between", formula=["-300", "0"], stopIfTrue=False, fill=yellow_fill)
    )
    ws.conditional_formatting.add(
        "C15",
        CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, fill=green_fill)
    )

    # Methodology footnote
    ws.merge_cells("A19:L22")
    c = ws["A19"]
    c.value = (
        "Mid-term rentals (30+ days) skip Airbnb's platform fee, skip per-stay cleaning, "
        "and stabilize at ~90% occupancy once a tenant signs. Travel-nurse contracts run "
        "13 weeks; corporate housing 60-180 days; insurance displacement (people whose "
        "homes burned/flooded) often 6+ months. Pull MTR rents from Furnished Finder for "
        "your zip code. If MTR covers rent + a small cushion — your downside is bounded. "
        "If MTR shortfall is large — DO NOT sign this lease. The escape hatch isn't there."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    for r in range(19, 23):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 25, version_line=f"{SKU} · MTR Fallback")


# ---------------------------------------------------------------------------
# Sheet 8 — Landlord Pitch Page (white-label)
# ---------------------------------------------------------------------------

def build_landlord_pitch_tab(wb, variant):
    ws = wb.create_sheet("Landlord Pitch")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    # Print area: A1:G50 portrait letter — wider columns since white-label
    set_col_widths(ws, [
        ("A", 3), ("B", 22), ("C", 28), ("D", 22), ("E", 22), ("F", 18), ("G", 3),
        ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    # NOTE: This tab is white-label — NO STR Ledger header band, NO compact_header_band.
    # The operator drops in their own logo. We provide ONLY a small back-button.

    # Tiny back button row 1 — small navy chip in column A-B only
    pseudo_button(ws, "A1", "B1", "← BACK",
                  "'Start'!A1", variant="primary")
    ws.row_dimensions[1].height = 20

    # Logo placeholder — row 3
    ws.merge_cells("B3:F4")
    c = ws["B3"]
    c.value = "[ YOUR LOGO HERE ]"
    c.font = Font(name=FONT_MONO, size=12, color=COLOR_MUTED, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        top=Side(style="dashed", color=COLOR_MUTED),
        bottom=Side(style="dashed", color=COLOR_MUTED),
        left=Side(style="dashed", color=COLOR_MUTED),
        right=Side(style="dashed", color=COLOR_MUTED),
    )
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24

    # Title — row 6
    ws.merge_cells("B6:F6")
    c = ws["B6"]
    c.value = "Master-Lease Proposal"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_NEAR_BLACK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 36

    ws.merge_cells("B7:F7")
    c = ws["B7"]
    c.value = '="Date prepared: "&TEXT(TODAY(),"mmmm d, yyyy")'
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 18

    # Section divider helper for white-label
    def pitch_section(row, label):
        ws.merge_cells(f"B{row}:F{row}")
        cell = ws[f"B{row}"]
        cell.value = label
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_NEAR_BLACK)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=COLOR_NEAR_BLACK))
        ws.row_dimensions[row].height = 22

    # PROPOSED TENANT
    pitch_section(9, "PROPOSED TENANT")
    tenant_rows = [
        (10, "Operator:",        _val(variant, SAMPLE["tenant_name"])),
        (11, "Entity / LLC:",    _val(variant, SAMPLE["tenant_llc"])),
        (12, "Contact:",         _val(variant, SAMPLE["tenant_contact"])),
        (13, "Business:",        _val(variant, SAMPLE["tenant_business"])),
    ]
    for row, label, value in tenant_rows:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.merge_cells(f"C{row}:F{row}")
        cell = ws.cell(row=row, column=3, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    # PROPERTY
    pitch_section(15, "PROPERTY")
    prop_rows = [
        (16, "Address:",     "='Lease Terms'!C7"),
        (17, "Unit type:",   "='Lease Terms'!C9&\" · \"&'Lease Terms'!C10&\" bd / \"&'Lease Terms'!C11&\" ba\""),
    ]
    for row, label, formula in prop_rows:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.merge_cells(f"C{row}:F{row}")
        cell = ws.cell(row=row, column=3, value=formula)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_NEAR_BLACK)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    # PROPOSED TERMS
    pitch_section(19, "PROPOSED TERMS")
    term_rows = [
        (20, "Monthly rent:",        "='Lease Terms'!C15", '"$"#,##0'),
        (21, "Security deposit:",    "='Lease Terms'!C16", '"$"#,##0'),
        (22, "Lease term:",          "='Lease Terms'!C17&\" months\"", None),
        (23, "Start date:",          "='Lease Terms'!C18", None),
        (24, "Renewal escalator:",   "='Lease Terms'!C19", "0.0%"),
        (25, "Tenant pays:",         "Utilities · internet · insurance · cleaning · repairs <$200", None),
        (26, "Landlord pays:",       "Property tax · HOA · structural repairs · major systems", None),
    ]
    for row, label, formula, *fmt in term_rows:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.merge_cells(f"C{row}:F{row}")
        cell = ws.cell(row=row, column=3, value=formula)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_NEAR_BLACK)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if fmt and fmt[0]:
            cell.number_format = fmt[0]
        ws.row_dimensions[row].height = 18

    # WHY THIS WORKS FOR YOU
    pitch_section(28, "WHY THIS WORKS FOR YOU")
    why_points = [
        (29, "Guaranteed rent — paid on the 1st whether or not the unit is booked. You collect either way."),
        (30, "Zero tenant turnover — one professional operator, not a revolving door of renters."),
        (31, "Premium maintenance — we treat this unit like our reputation depends on it (because it does)."),
        (32, "Tenant carries $1M liability + STR rider — your asset is over-insured, not under."),
        (33, "Indemnification clause — guests' actions are our liability, never yours."),
    ]
    for row, text in why_points:
        ws.merge_cells(f"B{row}:F{row}")
        c = ws[f"B{row}"]
        c.value = "  +  " + text
        c.font = Font(name=FONT_BODY, size=10, color=COLOR_NEAR_BLACK)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20

    # REFERENCES + HISTORY
    pitch_section(35, "REFERENCES  +  PRIOR HISTORY")
    ref_rows = [
        (36, "Reference 1:",   _val(variant, SAMPLE["ref_1"])),
        (37, "Reference 2:",   _val(variant, SAMPLE["ref_2"])),
        (38, "Prior history:", _val(variant, SAMPLE["prior_history"])),
    ]
    for row, label, value in ref_rows:
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.merge_cells(f"C{row}:F{row}")
        cell = ws.cell(row=row, column=3, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[row].height = 22

    # SIGNATURES
    pitch_section(41, "SIGNATURES")
    # Tenant
    ws.cell(row=43, column=2, value="Tenant signature:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
    )
    ws.merge_cells("C43:E43")
    sig_cell = ws["C43"]
    sig_cell.border = Border(bottom=Side(style="thin", color=COLOR_NEAR_BLACK))
    ws.cell(row=43, column=6, value="Date:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
    )
    ws.row_dimensions[43].height = 28

    ws.cell(row=45, column=2, value="Landlord signature:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
    )
    ws.merge_cells("C45:E45")
    sig_cell = ws["C45"]
    sig_cell.border = Border(bottom=Side(style="thin", color=COLOR_NEAR_BLACK))
    ws.cell(row=45, column=6, value="Date:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_NEAR_BLACK
    )
    ws.row_dimensions[45].height = 28

    # Disclaimer footer
    ws.merge_cells("B48:F49")
    c = ws["B48"]
    c.value = (
        "This proposal is non-binding and subject to a formal master-lease agreement. "
        "Both parties should consult an attorney before signing."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[48].height = 14
    ws.row_dimensions[49].height = 14

    # Print setup — white-label, portrait letter
    ws.print_area = "A1:G50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_lease_terms_tab(wb, variant)
    build_setup_costs_tab(wb, variant)
    build_revenue_tab(wb, variant)
    build_operating_math_tab(wb, variant)
    build_stress_test_tab(wb, variant)
    build_mtr_fallback_tab(wb, variant)
    build_landlord_pitch_tab(wb, variant)

    wb.properties.title = "Rental Arbitrage Analyzer — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Rental arbitrage underwriting workbook — lease vs. revenue spread, "
        "Year-2 rent-escalator stress test, MTR fallback, landlord pitch page."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
