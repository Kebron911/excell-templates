"""Build REV-006 PriceLabs / Wheelhouse / Beyond ROI Comparison (v2.2 standard).

A/B test framework for paid pricing tools. Host runs the tool for one
period and runs without it for another (matched length). Workbook
computes whether the tool earned its monthly fee.

Forked from build_break_even_occupancy.py (single-page calc) — extended
to two stacked 60-row daily grids + a side-by-side ROI verdict tab.

Generates two files:
  templates/_masters/REV-006-pricing-tool-roi-comparison-DEMO.xlsx
  templates/_masters/REV-006-pricing-tool-roi-comparison-BLANK.xlsx
"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
)

SKU = "REV-006"
NAME = "pricing-tool-roi-comparison"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample DEMO data — Smokies Ridge cabin, PriceLabs Aug-Oct 2025
# ---------------------------------------------------------------------------

DEMO_SETUP = {
    "tool_name":         "PriceLabs",
    "monthly_fee":       20.00,
    "per_listing_fee":   20.00,
    "baseline_start":    date(2025, 6, 1),
    "baseline_end":      date(2025, 7, 30),  # 60 nights
    "test_start":        date(2025, 8, 1),
    "test_end":          date(2025, 9, 29),  # 60 nights
    "lead_time_avg":     21,                  # days
    "property_name":     "Smokies Ridge Cabin",
    "active_year":       2025,
    "cleaning_per_turn": 95.00,
}

TOOL_OPTIONS = [
    "PriceLabs", "Wheelhouse", "Beyond", "AirDNA Smart Rates", "Other",
]

YN_LIST = ["Y", "N"]


# Two 60-day grids — pseudo-random but deterministic so DEMO is reproducible.
# Baseline: 71% occ, ~$192 ADR, no tool.
# Test:     76% occ, ~$208 ADR, PriceLabs active.
def _build_demo_grids():
    baseline = []
    test = []

    base_start = DEMO_SETUP["baseline_start"]
    test_start = DEMO_SETUP["test_start"]

    # Baseline: 60 days. Pattern: weekends booked more, $185-200 rate range.
    base_pattern = [
        # (rate, booked_yn) per day; 60 entries; ~71% occupancy
        (185, "Y"), (185, "Y"), (185, "N"), (185, "N"), (190, "Y"),
        (200, "Y"), (200, "Y"), (185, "N"), (185, "Y"), (185, "Y"),
        (185, "N"), (185, "N"), (190, "Y"), (200, "Y"), (200, "Y"),
        (185, "Y"), (185, "Y"), (185, "N"), (185, "Y"), (190, "Y"),
        (200, "Y"), (200, "Y"), (185, "N"), (185, "N"), (185, "Y"),
        (185, "Y"), (190, "Y"), (200, "Y"), (200, "Y"), (185, "Y"),
        (185, "N"), (185, "N"), (185, "Y"), (185, "Y"), (190, "Y"),
        (200, "Y"), (200, "Y"), (185, "Y"), (185, "Y"), (185, "N"),
        (185, "Y"), (190, "Y"), (200, "Y"), (200, "Y"), (185, "N"),
        (185, "N"), (185, "Y"), (185, "Y"), (190, "Y"), (200, "Y"),
        (200, "Y"), (185, "Y"), (185, "N"), (185, "N"), (185, "Y"),
        (190, "Y"), (200, "Y"), (200, "Y"), (185, "Y"), (185, "N"),
    ]
    for i, (rate, booked) in enumerate(base_pattern):
        d = base_start + timedelta(days=i)
        baseline.append((d, rate, booked))

    # Test: 60 days. PriceLabs lifts rate on high-demand nights, reduces
    # rate to fill mid-week → ~76% occ, ~$208 ADR.
    test_pattern = [
        (195, "Y"), (195, "Y"), (175, "Y"), (175, "N"), (215, "Y"),
        (235, "Y"), (235, "Y"), (175, "Y"), (195, "Y"), (195, "Y"),
        (175, "N"), (175, "Y"), (215, "Y"), (235, "Y"), (235, "Y"),
        (195, "Y"), (195, "Y"), (175, "Y"), (175, "N"), (215, "Y"),
        (235, "Y"), (235, "Y"), (175, "Y"), (175, "N"), (195, "Y"),
        (195, "Y"), (215, "Y"), (235, "Y"), (235, "Y"), (195, "Y"),
        (175, "N"), (175, "Y"), (195, "Y"), (195, "Y"), (215, "Y"),
        (235, "Y"), (235, "Y"), (195, "Y"), (195, "Y"), (175, "Y"),
        (195, "Y"), (215, "Y"), (235, "Y"), (235, "Y"), (175, "N"),
        (175, "Y"), (195, "Y"), (195, "Y"), (215, "Y"), (235, "Y"),
        (235, "Y"), (195, "Y"), (175, "N"), (175, "Y"), (195, "Y"),
        (215, "Y"), (235, "Y"), (235, "Y"), (195, "Y"), (175, "N"),
    ]
    for i, (rate, booked) in enumerate(test_pattern):
        d = test_start + timedelta(days=i)
        test.append((d, rate, booked))

    return baseline, test


DEMO_BASELINE, DEMO_TEST = _build_demo_grids()

DEMO_NOTES_BASELINE = {
    0:  "Set rate manually; flat pricing.",
    6:  "Mid-June weekend held strong.",
    20: "Walk-in booking; could have priced higher.",
    34: "Slow Tuesday — left $185 flat.",
    50: "Weekend filled at base rate.",
}

DEMO_NOTES_TEST = {
    0:  "PriceLabs auto-set; trusting the algo.",
    4:  "Tool bumped weekend to $215.",
    13: "Big lift to $235 — booked anyway.",
    27: "Good price drop filled mid-week.",
    44: "Tool dropped rate to $175 to fill.",
    56: "Algo nailed the weekend bump.",
}


# ---------------------------------------------------------------------------
# Variant helpers
# ---------------------------------------------------------------------------

def _v(variant, demo_value, blank_value=None):
    return demo_value if variant == "demo" else blank_value


def _add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Pricing Tool ROI Comparison"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Did PriceLabs / Wheelhouse / Beyond actually earn its fee?"
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: KPI hero block rows 10-22
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Property name (pulled from Settings)
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!B5="","(enter property on Settings tab)",Settings!B5)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # Headline label
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "ROI OF PAID PRICING"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # Big lift $ number
    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = '=IFERROR("$"&TEXT(\'ROI Verdict\'!C20,"#,##0")&" / month","—")'
    c.font = Font(name=FONT_HEAD, size=48, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 24

    # Lift % subtitle
    ws.merge_cells("A16:L16")
    c = ws["A16"]
    c.value = (
        '=IFERROR("= "&TEXT(\'ROI Verdict\'!C21,"+0.0%;-0.0%")'
        '&" lift on monthly NET (after tool fee)","")'
    )
    c.font = Font(name=FONT_BODY, size=12, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 22

    # Verdict cell — the brief-mandated formula
    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = (
        '=IF(\'ROI Verdict\'!C20>0, '
        '"✅ Tool earned its fee — keep", '
        '"🛑 Tool didn\'t pay for itself — drop")'
    )
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 30

    # Footnote
    ws.merge_cells("A20:L21")
    c = ws["A20"]
    c.value = (
        "Note: lift assumes the two periods are otherwise comparable. Run the "
        "test for ≥60 days to absorb booking-lead-time noise. Adjust for "
        "seasonality if baseline and test span very different demand windows."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[20].height = 22
    ws.row_dimensions[21].height = 22

    # ZONE 3: Pseudo-button nav rows 24-26
    pseudo_button(ws, "A24", "D26", "→  A/B Test Setup",
                  "'A_B Test Setup'!A1", variant="primary")
    pseudo_button(ws, "E24", "H26", "Daily Performance",
                  "'Daily Performance'!A1", variant="secondary")
    pseudo_button(ws, "I24", "L26", "ROI Verdict",
                  "'ROI Verdict'!A1", variant="accent")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    # Upgrade banner row 29
    ws.merge_cells("A29:L29")
    c = ws["A29"]
    c.value = (
        "💡  Pair this with REV-005 Holiday + Event Pricing Calendar — "
        f"pre-built high-rate dates for your market. {BRAND_DOMAIN}"
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[29].height = 36

    brand_footer(ws, 31,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L33"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — A/B Test Setup
# ---------------------------------------------------------------------------

def build_setup_tab(wb, variant):
    ws = wb.create_sheet("A_B Test Setup")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 20),
        ("D", 4), ("E", 50),
        ("F", 6), ("G", 6), ("H", 6), ("I", 6),
        ("J", 6), ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "A/B Test Setup",
                        prev_tab="Start", next_tab="Daily Performance")

    # Intro band row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Define your two test periods (matched length, ≥60 days each). "
        "Yellow = edit; gray = calculated."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    ws.freeze_panes = "A5"

    # SECTION: TOOL ---------------------------------------------------------
    _section_band(ws, 5, "TOOL UNDER TEST")

    _input_row(ws, 6, "B", "C", "Tool tested:",
               _v(variant, DEMO_SETUP["tool_name"]),
               None,
               "Dropdown: PriceLabs / Wheelhouse / Beyond / AirDNA / Other")
    _add_dropdown(ws, "C6", '"PriceLabs,Wheelhouse,Beyond,AirDNA Smart Rates,Other"')

    _input_row(ws, 7, "B", "C", "Monthly fee ($):",
               _v(variant, DEMO_SETUP["monthly_fee"]),
               '"$"#,##0.00',
               "What the tool charges per month (flat or per listing)")

    _input_row(ws, 8, "B", "C", "Per-listing fee ($/listing/mo):",
               _v(variant, DEMO_SETUP["per_listing_fee"]),
               '"$"#,##0.00',
               "Some tools charge per listing; some flat. Use the larger.")

    # SECTION: PERIODS ------------------------------------------------------
    _section_band(ws, 10, "TEST PERIODS")

    _input_row(ws, 11, "B", "C", "Baseline start date:",
               _v(variant, DEMO_SETUP["baseline_start"]),
               "yyyy-mm-dd",
               "Period without the tool (or before you started using it)")

    _input_row(ws, 12, "B", "C", "Baseline end date:",
               _v(variant, DEMO_SETUP["baseline_end"]),
               "yyyy-mm-dd",
               "Match length to test period (60 days minimum)")

    _input_row(ws, 13, "B", "C", "Test start date:",
               _v(variant, DEMO_SETUP["test_start"]),
               "yyyy-mm-dd",
               "Period with the tool active")

    _input_row(ws, 14, "B", "C", "Test end date:",
               _v(variant, DEMO_SETUP["test_end"]),
               "yyyy-mm-dd",
               "Match length to baseline (60 days minimum)")

    # Period-length check
    _output_row(ws, 15, "B", "C", "Baseline length (days):",
                "=IFERROR(C12-C11+1,0)", "0",
                "Auto-calculated from start/end")

    _output_row(ws, 16, "B", "C", "Test length (days):",
                "=IFERROR(C14-C13+1,0)", "0",
                "Auto-calculated; should match baseline")

    # SECTION: ASSUMPTIONS --------------------------------------------------
    _section_band(ws, 18, "ASSUMPTIONS")

    _input_row(ws, 19, "B", "C", "Reservation lead time (days, avg):",
               _v(variant, DEMO_SETUP["lead_time_avg"]),
               "0",
               "Why the test needs ≥60 days — bookings made before the tool "
               "kicked in are still in the test window.")

    _input_row(ws, 20, "B", "C", "Cleaning cost per turnover ($):",
               _v(variant, DEMO_SETUP["cleaning_per_turn"]),
               '"$"#,##0.00',
               "Used in NET calc on ROI Verdict tab. Charge to guest is "
               "separate (pulled from daily ADR rows).")

    # SECTION: SETUP CHECK --------------------------------------------------
    _section_band(ws, 22, "SETUP CHECK")

    ws.merge_cells("A23:L26")
    c = ws["A23"]
    c.value = (
        "✓ Use the same property and same season for both periods if you can.\n"
        "✓ 60 days is the minimum — booking lead time means the first 2-3 "
        "weeks of the test still reflect baseline pricing decisions.\n"
        "✓ If you can't run baseline in real time, look up your historical "
        "data on the same calendar dates one year prior.\n"
        "✓ Don't change anything else (listing photos, minimum stay, "
        "cleaning fee) during the test — keep variables isolated."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=2)
    for r in range(23, 27):
        ws.row_dimensions[r].height = 22

    brand_footer(ws, 28, version_line=f"{SKU} · v1.0 · A/B Test Setup")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 3 — Daily Performance (two stacked 60-row grids)
# ---------------------------------------------------------------------------

def build_daily_tab(wb, variant):
    ws = wb.create_sheet("Daily Performance")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4),
        ("B", 13), ("C", 12), ("D", 11), ("E", 12), ("F", 32),
        ("G", 4),
        ("H", 13), ("I", 12), ("J", 11), ("K", 12), ("L", 32),
    ])

    compact_header_band(ws, "Daily Performance",
                        prev_tab="A_B Test Setup", next_tab="ROI Verdict")

    # Intro row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Two 60-row grids — left = baseline (no tool), right = test (tool active). "
        "Booked = Y/N. ADR is what guest paid that night. Row 71 = totals."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # --- Grid headers row 6 (banner) and row 7 (column labels) ------------
    # BASELINE banner B6:F6
    ws.merge_cells("B6:F6")
    c = ws["B6"]
    c.value = "BASELINE  —  no tool"
    c.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # TEST banner H6:L6
    ws.merge_cells("H6:L6")
    c = ws["H6"]
    c.value = "TEST  —  tool active"
    c.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers row 7
    base_headers = [(2, "Date"), (3, "Set rate"), (4, "Booked Y/N"),
                    (5, "ADR"), (6, "Notes")]
    test_headers = [(8, "Date"), (9, "Set rate"), (10, "Booked Y/N"),
                    (11, "ADR"), (12, "Notes")]
    for col, label in base_headers + test_headers:
        cell = ws.cell(row=7, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 22

    # --- Data rows 8 to 67 (60 rows) -------------------------------------
    DATA_TOP = 8
    DATA_BOT = 67  # 8..67 = 60 rows

    thin = Side(style="thin", color=COLOR_PARCHMENT_ALT)
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for i in range(60):
        r = DATA_TOP + i

        # --- Baseline columns B-F ---
        # B: date
        cell = ws.cell(row=r, column=2)
        if variant == "demo" and i < len(DEMO_BASELINE):
            cell.value = DEMO_BASELINE[i][0]
        apply_style(cell, input_cell_style())
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # C: set rate
        cell = ws.cell(row=r, column=3)
        if variant == "demo" and i < len(DEMO_BASELINE):
            cell.value = DEMO_BASELINE[i][1]
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # D: booked Y/N
        cell = ws.cell(row=r, column=4)
        if variant == "demo" and i < len(DEMO_BASELINE):
            cell.value = DEMO_BASELINE[i][2]
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # E: ADR (formula = rate if booked else blank)
        cell = ws.cell(row=r, column=5)
        cell.value = f'=IF(D{r}="Y",C{r},"")'
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # F: notes
        cell = ws.cell(row=r, column=6)
        if variant == "demo" and i in DEMO_NOTES_BASELINE:
            cell.value = DEMO_NOTES_BASELINE[i]
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT, italic=True)

        # --- Test columns H-L ---
        cell = ws.cell(row=r, column=8)
        if variant == "demo" and i < len(DEMO_TEST):
            cell.value = DEMO_TEST[i][0]
        apply_style(cell, input_cell_style())
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = ws.cell(row=r, column=9)
        if variant == "demo" and i < len(DEMO_TEST):
            cell.value = DEMO_TEST[i][1]
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = ws.cell(row=r, column=10)
        if variant == "demo" and i < len(DEMO_TEST):
            cell.value = DEMO_TEST[i][2]
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = ws.cell(row=r, column=11)
        cell.value = f'=IF(J{r}="Y",I{r},"")'
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = ws.cell(row=r, column=12)
        if variant == "demo" and i in DEMO_NOTES_TEST:
            cell.value = DEMO_NOTES_TEST[i]
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT, italic=True)

        # alt-row banding
        if i % 2 == 1:
            alt = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
            for col in [2, 3, 4, 5, 6, 8, 9, 10, 11, 12]:
                ws.cell(row=r, column=col).fill = alt

        ws.row_dimensions[r].height = 16

    # Dropdowns Y/N for booked columns
    _add_dropdown(ws, f"D{DATA_TOP}:D{DATA_BOT}", '"Y,N"')
    _add_dropdown(ws, f"J{DATA_TOP}:J{DATA_BOT}", '"Y,N"')

    # --- Conditional formatting: booked rows tinted parchment-alt ----------
    # Already alt-row banded; add subtle highlight when D=Y / J=Y
    booked_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.conditional_formatting.add(
        f"B{DATA_TOP}:F{DATA_BOT}",
        FormulaRule(formula=[f'$D{DATA_TOP}="Y"'],
                     fill=booked_fill, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        f"H{DATA_TOP}:L{DATA_BOT}",
        FormulaRule(formula=[f'$J{DATA_TOP}="Y"'],
                     fill=booked_fill, stopIfTrue=False),
    )

    # --- Totals row 69 (occupancy / ADR / RevPAR) -------------------------
    TOT = 69
    # Spacer row 68
    ws.row_dimensions[68].height = 8

    # Baseline totals
    cell = ws.cell(row=TOT, column=2, value="TOTALS")
    cell.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_BG_LIGHT)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # C69: Occupancy %
    cell = ws.cell(row=TOT, column=3,
                   value=f'=IFERROR(COUNTIF(D{DATA_TOP}:D{DATA_BOT},"Y")'
                          f'/COUNTA(B{DATA_TOP}:B{DATA_BOT}),0)')
    apply_style(cell, formula_cell_style())
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    # D69: avg ADR (booked nights only)
    cell = ws.cell(row=TOT, column=4,
                   value=f'=IFERROR(AVERAGEIF(D{DATA_TOP}:D{DATA_BOT},'
                          f'"Y",E{DATA_TOP}:E{DATA_BOT}),0)')
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    # E69: RevPAR = sum(ADR)/COUNTA(date)
    cell = ws.cell(row=TOT, column=5,
                   value=f'=IFERROR(SUM(E{DATA_TOP}:E{DATA_BOT})'
                          f'/COUNTA(B{DATA_TOP}:B{DATA_BOT}),0)')
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    # F69: label
    cell = ws.cell(row=TOT, column=6, value="← occ / ADR / RevPAR")
    cell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Test totals
    cell = ws.cell(row=TOT, column=8, value="TOTALS")
    cell.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = ws.cell(row=TOT, column=9,
                   value=f'=IFERROR(COUNTIF(J{DATA_TOP}:J{DATA_BOT},"Y")'
                          f'/COUNTA(H{DATA_TOP}:H{DATA_BOT}),0)')
    apply_style(cell, formula_cell_style())
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    cell = ws.cell(row=TOT, column=10,
                   value=f'=IFERROR(AVERAGEIF(J{DATA_TOP}:J{DATA_BOT},'
                          f'"Y",K{DATA_TOP}:K{DATA_BOT}),0)')
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    cell = ws.cell(row=TOT, column=11,
                   value=f'=IFERROR(SUM(K{DATA_TOP}:K{DATA_BOT})'
                          f'/COUNTA(H{DATA_TOP}:H{DATA_BOT}),0)')
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_PRIMARY)

    cell = ws.cell(row=TOT, column=12, value="← occ / ADR / RevPAR")
    cell.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[TOT].height = 22

    brand_footer(ws, TOT + 2,
                 version_line=f"{SKU} · v1.0 · Daily Performance")

    ws.freeze_panes = "A8"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 4 — ROI Verdict
# ---------------------------------------------------------------------------

def build_verdict_tab(wb, variant):
    ws = wb.create_sheet("ROI Verdict")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 18), ("D", 18),
        ("E", 6), ("F", 30),
        ("G", 6), ("H", 6), ("I", 6),
        ("J", 6), ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "ROI Verdict",
                        prev_tab="Daily Performance", next_tab="Settings")

    # Intro row 4
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Side-by-side: ADR / RevPAR / occupancy / gross / NET (after fees). "
        "Lift > 0 means the tool earned its keep."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row 6
    headers = [
        (2, "Metric"),
        (3, "Baseline (no tool)"),
        (4, "Test (tool active)"),
        (6, "Notes"),
    ]
    for col, label in headers:
        cell = ws.cell(row=6, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # --- Metric rows ------------------------------------------------------
    # Helper for metric row writes — never write to merged cell after merge.
    def metric(row, label, base_formula, test_formula, fmt, note=None,
               emphasize=False):
        # Label col B
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(
            name=FONT_HEAD if emphasize else FONT_BODY,
            size=12 if emphasize else 11,
            bold=True,
            color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
        )
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

        # Baseline col C
        cell = ws.cell(row=row, column=3, value=base_formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if emphasize:
            cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        # Test col D
        cell = ws.cell(row=row, column=4, value=test_formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if emphasize:
            cell.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_ACCENT)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        # Note col F (write FIRST, then merge — avoids writing to merged cell)
        if note:
            ws.cell(row=row, column=6, value=note).font = Font(
                name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
            )
            ws.cell(row=row, column=6).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            ws.merge_cells(start_row=row, start_column=6,
                            end_row=row, end_column=12)

        ws.row_dimensions[row].height = 22 if emphasize else 18

    # Daily Performance ranges:
    #   Baseline date B8:B67, booked D8:D67, ADR formula E8:E67, set rate C8:C67
    #   Test     date H8:H67, booked J8:J67, ADR formula K8:K67, set rate I8:I67
    DP = "'Daily Performance'!"

    metric(7,
           "Occupancy:",
           f'=IFERROR(COUNTIF({DP}D8:D67,"Y")/COUNTA({DP}B8:B67),0)',
           f'=IFERROR(COUNTIF({DP}J8:J67,"Y")/COUNTA({DP}H8:H67),0)',
           "0.0%",
           note="Booked nights ÷ available nights in the period")

    metric(8,
           "ADR (avg booked night, $):",
           f'=IFERROR(AVERAGEIF({DP}D8:D67,"Y",{DP}E8:E67),0)',
           f'=IFERROR(AVERAGEIF({DP}J8:J67,"Y",{DP}K8:K67),0)',
           '"$"#,##0',
           note="Average daily rate on booked nights")

    metric(9,
           "RevPAR (revenue per available night, $):",
           f'=IFERROR(SUM({DP}E8:E67)/COUNTA({DP}B8:B67),0)',
           f'=IFERROR(SUM({DP}K8:K67)/COUNTA({DP}H8:H67),0)',
           '"$"#,##0',
           note="Sum of ADR ÷ available nights — true revenue density")

    metric(10,
           "Booked nights:",
           f'=COUNTIF({DP}D8:D67,"Y")',
           f'=COUNTIF({DP}J8:J67,"Y")',
           "0",
           note="Count of nights where Booked=Y")

    metric(11,
           "Gross revenue ($):",
           f'=SUM({DP}E8:E67)',
           f'=SUM({DP}K8:K67)',
           '"$"#,##0',
           note="Sum of nightly ADR — gross booking revenue")

    metric(12,
           "Cleaning costs ($):",
           f'=COUNTIF({DP}D8:D67,"Y")*\'A_B Test Setup\'!C20',
           f'=COUNTIF({DP}J8:J67,"Y")*\'A_B Test Setup\'!C20',
           '"$"#,##0',
           note="Booked nights × cleaning per turnover (from Setup tab)")

    metric(13,
           "Tool fees ($):",
           "=0",
           "=IFERROR(('A_B Test Setup'!C16/30)*'A_B Test Setup'!C7,0)",
           '"$"#,##0',
           note="(Test length ÷ 30) × monthly fee — pro-rated per period")

    metric(14,
           "NET revenue ($):",
           "=C11-C12-C13",
           "=D11-D12-D13",
           '"$"#,##0',
           note="Gross − cleaning − tool fees",
           emphasize=True)

    # Spacer row 16 + period-normalized monthly NET row 17
    ws.row_dimensions[16].height = 8

    metric(17,
           "Monthly NET equivalent ($):",
           "=IFERROR(C14/('A_B Test Setup'!C15/30),0)",
           "=IFERROR(D14/('A_B Test Setup'!C16/30),0)",
           '"$"#,##0',
           note="NET normalized to per-month for fair comparison",
           emphasize=True)

    # --- Lift summary block rows 19-22 ------------------------------------
    # Row 19 is the section header
    _section_band(ws, 19, "LIFT  —  did the tool pay for itself?")

    # Row 20: Lift $ (the brief reference for verdict)
    cell = ws.cell(row=20, column=2, value="Lift $/month:")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    # Lift $ in C20 — defined name "Lift" referenced in verdict formula
    cell = ws.cell(row=20, column=3, value="=D17-C17")
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.number_format = '"$"#,##0;[Red]"-$"#,##0'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        top=Side(style="thin", color=COLOR_ACCENT),
        bottom=Side(style="thin", color=COLOR_ACCENT),
        left=Side(style="thin", color=COLOR_ACCENT),
        right=Side(style="thin", color=COLOR_ACCENT),
    )

    # D20 left intentionally blank (single-value metric)
    ws.cell(row=20, column=4).fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Note F20:L20
    ws.cell(row=20, column=6, value="Test monthly NET − Baseline monthly NET").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=20, column=6).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.merge_cells("F20:L20")
    ws.row_dimensions[20].height = 24

    # Row 21: Lift %
    cell = ws.cell(row=21, column=2, value="Lift % (vs baseline):")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    cell = ws.cell(row=21, column=3, value="=IFERROR((D17-C17)/C17,0)")
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.number_format = "+0.0%;-0.0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        top=Side(style="thin", color=COLOR_ACCENT),
        bottom=Side(style="thin", color=COLOR_ACCENT),
        left=Side(style="thin", color=COLOR_ACCENT),
        right=Side(style="thin", color=COLOR_ACCENT),
    )
    ws.cell(row=21, column=4).fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    ws.cell(row=21, column=6, value="Lift $ as % of baseline NET").font = Font(
        name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED
    )
    ws.cell(row=21, column=6).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.merge_cells("F21:L21")
    ws.row_dimensions[21].height = 22

    # Row 23: Verdict cell — brief-mandated formula references C20 ("Lift")
    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = (
        '=IF(C20>0, '
        '"✅ Tool earned its fee — keep", '
        '"🛑 Tool didn\'t pay for itself — drop")'
    )
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 40

    # Conditional formatting on lift cell — red if negative
    ws.conditional_formatting.add(
        "C20",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT),
                   font=Font(name=FONT_HEAD, size=14, bold=True,
                              color=COLOR_ERROR)),
    )

    # Recommendation paragraph row 25-27
    ws.merge_cells("A25:L27")
    c = ws["A25"]
    c.value = (
        "Decision rule: a +5% lift on a $5,000/month NET = $250/month — "
        "easily covers a $20-50/month tool. A flat or negative lift means "
        "either the tool isn't suited to your market or the periods weren't "
        "comparable (different season, different lead-time exposure). "
        "If undecided, run another 60-day window in the opposite season."
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(25, 28):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 29, version_line=f"{SKU} · v1.0 · ROI Verdict")

    ws.print_area = "A1:L31"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 6),
        ("D", 50),
        ("E", 6), ("F", 6), ("G", 6), ("H", 6),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="ROI Verdict", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Property + active year. The Start tab pulls these for the headline."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Property name
    cell = ws.cell(row=5, column=1, value="Property:")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=5, column=2,
                   value=_v(variant, DEMO_SETUP["property_name"]))
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 22

    cell = ws.cell(row=5, column=4,
                   value="Property nickname — appears on the Start tab headline")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Active year
    cell = ws.cell(row=7, column=1, value="Active tax year:")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    cell = ws.cell(row=7, column=2,
                   value=_v(variant, DEMO_SETUP["active_year"]))
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    cell = ws.cell(row=7, column=4,
                   value="Year the test ran in — used in archive notes")
    cell.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Tip box
    ws.merge_cells("A10:L13")
    c = ws["A10"]
    c.value = (
        "Tip: archive each finished test before starting a new one. Save "
        "this file as 'REV-006-2025-PriceLabs-Aug.xlsx' (or similar), then "
        "duplicate the BLANK template for the next test. Over time you'll "
        "build a comparison library — PriceLabs vs. Wheelhouse vs. Beyond — "
        "specific to your property and market."
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=2)
    for r in range(10, 14):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 15, version_line=f"{SKU} · v1.0 · Settings")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Helpers (forked from build_break_even_occupancy)
# ---------------------------------------------------------------------------

def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label_col, value_col, label, value, fmt, note):
    """Write label, input, and note. Note is merged AFTER write (E:L)."""
    ws.cell(row=row, column=column_index_from_string(label_col),
            value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=column_index_from_string(label_col)).alignment = (
        Alignment(horizontal="right", vertical="center", indent=1)
    )

    cell = ws.cell(row=row, column=column_index_from_string(value_col),
                   value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        # Write note FIRST, then merge — never write to merged cell after merge.
        ws.cell(row=row, column=5, value=note).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=row, column=5).alignment = Alignment(
            horizontal="left", vertical="center", indent=1, wrap_text=True
        )
        ws.merge_cells(start_row=row, start_column=5,
                        end_row=row, end_column=12)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label_col, value_col, label, formula, fmt, note,
                emphasize=False):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=column_index_from_string(label_col),
            value=label).font = label_font
    ws.cell(row=row, column=column_index_from_string(label_col)).alignment = (
        Alignment(horizontal="right", vertical="center", indent=1)
    )

    cell = ws.cell(row=row, column=column_index_from_string(value_col),
                   value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    if note:
        ws.cell(row=row, column=5, value=note).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=row, column=5).alignment = Alignment(
            horizontal="left", vertical="center", indent=1, wrap_text=True
        )
        ws.merge_cells(start_row=row, start_column=5,
                        end_row=row, end_column=12)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_setup_tab(wb, variant)
    build_daily_tab(wb, variant)
    build_verdict_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Pricing Tool ROI Comparison — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "A/B test framework: did PriceLabs / Wheelhouse / Beyond actually "
        "earn their monthly fee? 60-day baseline vs. 60-day test, lift "
        "calc, keep/drop verdict."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
