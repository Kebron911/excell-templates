"""Build ACQ-006 Rehab Budget + ROI Projection (v2.2 standard).

Operational-mode tool: line-item rehab register (capacity 100) with ROI
projection. 5 tabs — Start, Budget Detail, Cost Allocation, ROI Projection,
Settings.

Generates two files:
  templates/_masters/ACQ-006-rehab-budget-roi-DEMO.xlsx     (sample data)
  templates/_masters/ACQ-006-rehab-budget-roi-BLANK.xlsx    (empty)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    apply_brand_header,
    style_chart,
    COLOR_FORMULA_TINT, COLOR_WHITE,
)

SKU = "ACQ-006"
NAME = "rehab-budget-roi"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Active rehab year — written into Settings, not derived via YEAR(TODAY()).
ACTIVE_REHAB_YEAR = 2026

# Categories (drive the dropdown + Cost Allocation roll-up).
CATEGORIES = [
    "Demolition", "Plumbing", "Electrical", "HVAC", "Drywall", "Flooring",
    "Cabinets", "Countertops", "Appliances", "Painting", "Fixtures",
    "Furniture", "Decor", "Outdoor", "Permits", "Misc",
]

PAYMENT_STATUSES = [
    "Estimate only", "Deposit paid", "In progress", "Complete", "Paid in full",
]

CAPEX_REPAIR = ["Capex", "Repair"]

LINE_CAPACITY = 100  # rows of register

# ---------------------------------------------------------------------------
# Sample data — Smokies Ridge cabin, full kitchen + 2 bath rehab
# (per brief: 25 line items, $48K budget, $51K actual ~+6.3%)
# Each: (Category, Description, Estimate, Actual, Contractor, Status, CapexRepair, Notes)
# ---------------------------------------------------------------------------

SAMPLE_LINES = [
    ("Demolition",   "Kitchen demo + haul-off",        1800,  2100, "ABC Demo Co",       "Paid in full",  "Repair", ""),
    ("Demolition",   "Bath 1 + 2 demo",                1400,  1500, "ABC Demo Co",       "Paid in full",  "Repair", ""),
    ("Plumbing",     "Bath 1 rough-in + fixtures",     3200,  3400, "Smokies Plumbing",  "Paid in full",  "Capex",  ""),
    ("Plumbing",     "Bath 2 rough-in + fixtures",     3000,  3250, "Smokies Plumbing",  "Paid in full",  "Capex",  ""),
    ("Plumbing",     "Kitchen sink + supply lines",    1100,  1200, "Smokies Plumbing",  "Paid in full",  "Capex",  ""),
    ("Electrical",   "Kitchen circuits + outlets",     2400,  2350, "Bright Volt",       "Paid in full",  "Capex",  ""),
    ("Electrical",   "Bath GFCI + lighting",           1600,  1650, "Bright Volt",       "Paid in full",  "Capex",  ""),
    ("HVAC",         "Mini-split for upstairs",        2800,  3100, "Mountain HVAC",     "Paid in full",  "Capex",  "Added zone"),
    ("Drywall",      "Kitchen + 2 baths drywall",      2200,  2300, "Hill Drywall",      "Paid in full",  "Repair", ""),
    ("Flooring",     "LVP throughout (1100 sqft)",     4400,  4600, "Floor Pros",        "Paid in full",  "Capex",  ""),
    ("Flooring",     "Tile in baths",                  1800,  1900, "Floor Pros",        "Paid in full",  "Capex",  ""),
    ("Cabinets",     "Kitchen cabinets (semi-custom)", 5200,  5500, "Cabinets-N-More",   "Paid in full",  "Capex",  ""),
    ("Cabinets",     "Bath vanities (×2)",             1800,  1950, "Cabinets-N-More",   "Paid in full",  "Capex",  ""),
    ("Countertops",  "Kitchen quartz",                 2400,  2600, "Stone Stop",        "Paid in full",  "Capex",  ""),
    ("Countertops",  "Bath vanity tops",                900,   950, "Stone Stop",        "Paid in full",  "Capex",  ""),
    ("Appliances",   "Range / fridge / dish / micro",  3600,  3800, "Best Buy",          "Paid in full",  "Capex",  ""),
    ("Painting",     "Interior repaint (full)",        2400,  2550, "Color Crew",        "Paid in full",  "Repair", ""),
    ("Fixtures",     "Lighting fixtures package",      1100,  1200, "Wayfair",           "Paid in full",  "Capex",  ""),
    ("Fixtures",     "Plumbing fixtures finish",        800,   900, "Build.com",         "Paid in full",  "Capex",  ""),
    ("Furniture",    "Living + bed sets",              2800,  2900, "Self",              "Paid in full",  "Capex",  ""),
    ("Decor",        "Wall art + accents",              900,   950, "Self",              "Paid in full",  "Repair", ""),
    ("Outdoor",      "Hot tub install pad + steps",    1800,  1850, "Hot Tub LLC",       "Paid in full",  "Capex",  ""),
    ("Outdoor",      "Deck stain + railings",          1100,  1200, "Color Crew",        "Paid in full",  "Repair", ""),
    ("Permits",      "STR permit + electrical permit",  450,   480, "City of Gatlinburg","Paid in full",  "Repair", ""),
    ("Misc",         "Cleaning + final detail",         900,   860, "Sparkle STR",       "Paid in full",  "Repair", ""),
]

# Sanity check at build-time
assert len(SAMPLE_LINES) == 25, f"expected 25 sample lines, got {len(SAMPLE_LINES)}"

SAMPLE_SETTINGS = {
    "property_name":  "Smokies Ridge Cabin",
    "rehab_start":    "2026-02-01",
    "rehab_end":      "2026-04-15",
    "contingency_pct": 0.15,
    # ROI inputs
    "pre_rehab_adr":  145,
    "post_rehab_adr": 185,
    "occ_lift_pct":   0.08,   # 8 percentage-point lift
    "post_rehab_occ": 0.55,   # used for 5-yr cumulative
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _add_dropdown(ws, cell_range, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 10) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Rehab Budget + ROI"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Track every line. Project the lift. Know your payback."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ----- KPI strip (rows 10-17) — Budget · Spent · Variance % · ROI -----
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 18):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "PROJECT AT A GLANCE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[10].height = 18

    # 4 KPI cards: A-C, D-F, G-I, J-L
    kpi_specs = [
        ("A", "C", "BUDGET (EST)",       "='Cost Allocation'!C26",  '"$"#,##0',  COLOR_PRIMARY),
        ("D", "F", "SPENT (ACTUAL)",     "='Cost Allocation'!D26",  '"$"#,##0',  COLOR_PRIMARY),
        ("G", "I", "VARIANCE %",         "=IFERROR(('Cost Allocation'!D26-'Cost Allocation'!C26)/'Cost Allocation'!C26,0)", "0.0%", COLOR_SECONDARY),
        ("J", "L", "5-YR ROI",           "='ROI Projection'!C20",   "0.0%",      COLOR_ACCENT),
    ]
    ws.row_dimensions[11].height = 16
    ws.row_dimensions[12].height = 26
    ws.row_dimensions[13].height = 18
    for first, last, label, formula, fmt, label_color in kpi_specs:
        ws.merge_cells(f"{first}11:{last}11")
        c = ws[f"{first}11"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=label_color)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{first}12:{last}13")
        c = ws[f"{first}12"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=18, bold=True, color=label_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

    # ----- "What this does" card -----
    ws.merge_cells("A15:L17")
    c = ws["A15"]
    c.value = (
        "Log every line item — estimate vs actual, contractor, payment status, "
        "capex/repair tag. The Cost Allocation tab rolls it up by category and "
        "shows where dollars went. ROI Projection turns rehab spend into a "
        "year-1 revenue lift and a payback-period number you can defend."
    )
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[15].height = 16
    ws.row_dimensions[16].height = 16
    ws.row_dimensions[17].height = 16

    # ----- Pseudo-button nav (rows 19-21) -----
    pseudo_button(ws, "A19", "C21", "Budget Detail",
                  "'Budget Detail'!A1", variant="primary")
    pseudo_button(ws, "D19", "F21", "Cost Allocation",
                  "'Cost Allocation'!A1", variant="primary")
    pseudo_button(ws, "G19", "I21", "ROI Projection",
                  "'ROI Projection'!A1", variant="primary")
    pseudo_button(ws, "J19", "L21", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(19, 22):
        ws.row_dimensions[r].height = 22

    # Primary CTA (rows 23-25) — log a line
    pseudo_button(ws, "A23", "L25", "→  ADD A REHAB LINE",
                  "'Budget Detail'!B7", variant="accent")
    for r in range(23, 26):
        ws.row_dimensions[r].height = 22

    # ----- "How to use" card -----
    ws.merge_cells("A27:L30")
    c = ws["A27"]
    c.value = (
        "1) Set property + dates on Settings.   "
        "2) Add lines on Budget Detail (estimate first; actual once paid).   "
        "3) Watch Variance % — red >+15% means you're over.   "
        "4) Cost Allocation auto-rolls by category, splits Capex/Repair for tax."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    for r in range(27, 31):
        ws.row_dimensions[r].height = 16

    # Upgrade banner
    ws.merge_cells("A32:L32")
    c = ws["A32"]
    c.value = (
        "Capex vs Repair classification matters for tax. Pair this with "
        "Cost Segregation DIY (TAX-010) to claim accelerated depreciation."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[32].height = 28

    brand_footer(ws, 34,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L36"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Budget Detail (the main register, 100 rows capacity)
# ---------------------------------------------------------------------------

# Columns: A Line# | B Category | C Description | D Estimate | E Actual
#          F Variance | G Variance % | H Contractor | I Status | J Capex/Repair | K Notes
DETAIL_HEADERS = [
    ("A", "Line #",       6),
    ("B", "Category",     14),
    ("C", "Description",  34),
    ("D", "Estimate",     12),
    ("E", "Actual",       12),
    ("F", "Variance",     12),
    ("G", "Var %",        9),
    ("H", "Contractor",   22),
    ("I", "Status",       18),
    ("J", "Capex/Repair", 13),
    ("K", "Notes",        24),
]

DETAIL_HEADER_ROW = 5
DETAIL_FIRST_DATA_ROW = 6


def build_detail_tab(wb, variant):
    ws = wb.create_sheet("Budget Detail")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [(letter, w) for letter, _, w in DETAIL_HEADERS])

    compact_header_band(ws, "Budget Detail",
                         prev_tab="Start", next_tab="Cost Allocation")

    # Row 4 spacer / explainer
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 12):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:K4")
    c = ws["A4"]
    c.value = (
        "Log every rehab line. Variance is auto-calculated. "
        "Capex/Repair drives tax treatment downstream."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Header row (row 5)
    for letter, label, _w in DETAIL_HEADERS:
        cell = ws[f"{letter}{DETAIL_HEADER_ROW}"]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[DETAIL_HEADER_ROW].height = 22

    # Data rows: 100 capacity
    last_row = DETAIL_FIRST_DATA_ROW + LINE_CAPACITY - 1
    parchment_alt_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    for i in range(LINE_CAPACITY):
        r = DETAIL_FIRST_DATA_ROW + i
        line_num = i + 1

        # Sample data on first 25 rows (DEMO only)
        if variant == "demo" and i < len(SAMPLE_LINES):
            cat, desc, est, actual, contractor, status, cap, notes = SAMPLE_LINES[i]
        else:
            cat = desc = contractor = status = cap = notes = None
            est = actual = None

        # A — Line # (auto, formula stamped)
        cell = ws.cell(row=r, column=1, value=line_num)
        cell.font = Font(name=FONT_MONO, size=10, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # B — Category (input, dropdown)
        cell = ws.cell(row=r, column=2, value=cat)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # C — Description (input)
        cell = ws.cell(row=r, column=3, value=desc)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # D — Estimate (input)
        cell = ws.cell(row=r, column=4, value=est)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.number_format = '"$"#,##0'

        # E — Actual (input)
        cell = ws.cell(row=r, column=5, value=actual)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.number_format = '"$"#,##0'

        # F — Variance (formula =E-D, only when D is filled)
        cell = ws.cell(row=r, column=6,
                       value=f'=IF(AND(ISNUMBER(D{r}),ISNUMBER(E{r})),E{r}-D{r},"")')
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.number_format = '"$"#,##0;[Red]"-$"#,##0'

        # G — Variance %
        cell = ws.cell(row=r, column=7,
                       value=f'=IF(AND(ISNUMBER(D{r}),ISNUMBER(E{r}),D{r}<>0),(E{r}-D{r})/D{r},"")')
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.0%"

        # H — Contractor
        cell = ws.cell(row=r, column=8, value=contractor)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # I — Status (dropdown)
        cell = ws.cell(row=r, column=9, value=status)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # J — Capex / Repair (dropdown)
        cell = ws.cell(row=r, column=10, value=cap)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # K — Notes
        cell = ws.cell(row=r, column=11, value=notes)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[r].height = 16

        # Alternating-row banding subtle override on odd data rows.
        if i % 2 == 1:
            for col in range(1, 12):
                cell = ws.cell(row=r, column=col)
                # Only override if the cell uses the default formula gray —
                # leave the input-yellow cells alone.
                fg = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                if fg and str(fg).upper().endswith(COLOR_FORMULA_TINT):
                    cell.fill = parchment_alt_fill

    # Totals row (just below capacity)
    totals_row = last_row + 1
    ws.cell(row=totals_row, column=3, value="TOTALS").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT
    )
    ws.cell(row=totals_row, column=3).fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    ws.cell(row=totals_row, column=3).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    for col_letter, col_num in [("D", 4), ("E", 5), ("F", 6)]:
        cell = ws.cell(row=totals_row, column=col_num,
                       value=f'=SUM({col_letter}{DETAIL_FIRST_DATA_ROW}:{col_letter}{last_row})')
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.number_format = '"$"#,##0'
    # Variance % roll-up
    cell = ws.cell(row=totals_row, column=7,
                   value=f'=IFERROR((E{totals_row}-D{totals_row})/D{totals_row},0)')
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0%"
    # Fill the rest of the totals row with navy for visual continuity
    for col in (1, 2, 8, 9, 10, 11):
        ws.cell(row=totals_row, column=col).fill = PatternFill(
            "solid", fgColor=COLOR_PRIMARY
        )
    ws.row_dimensions[totals_row].height = 24

    # Conditional formatting on Variance % column (G6:G last_row)
    var_range = f"G{DETAIL_FIRST_DATA_ROW}:G{last_row}"
    # Red if > 15%
    ws.conditional_formatting.add(
        var_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(G{DETAIL_FIRST_DATA_ROW}),G{DETAIL_FIRST_DATA_ROW}>0.15)'],
            fill=PatternFill("solid", fgColor=COLOR_ERROR),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE),
        ),
    )
    # Gold if 5-15%
    ws.conditional_formatting.add(
        var_range,
        FormulaRule(
            formula=[
                f'AND(ISNUMBER(G{DETAIL_FIRST_DATA_ROW}),'
                f'G{DETAIL_FIRST_DATA_ROW}>0.05,G{DETAIL_FIRST_DATA_ROW}<=0.15)'
            ],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY),
        ),
    )
    # Parchment if <= 5% positive (or any below)
    ws.conditional_formatting.add(
        var_range,
        FormulaRule(
            formula=[
                f'AND(ISNUMBER(G{DETAIL_FIRST_DATA_ROW}),'
                f'G{DETAIL_FIRST_DATA_ROW}<=0.05)'
            ],
            fill=PatternFill("solid", fgColor=COLOR_BG_LIGHT),
        ),
    )

    # Dropdowns: Category (B), Status (I), Capex/Repair (J)
    _add_dropdown(ws, f"B{DETAIL_FIRST_DATA_ROW}:B{last_row}", CATEGORIES)
    _add_dropdown(ws, f"I{DETAIL_FIRST_DATA_ROW}:I{last_row}", PAYMENT_STATUSES)
    _add_dropdown(ws, f"J{DETAIL_FIRST_DATA_ROW}:J{last_row}", CAPEX_REPAIR)

    # Freeze header row
    ws.freeze_panes = f"A{DETAIL_FIRST_DATA_ROW}"

    # Print setup — landscape because 11 columns
    ws.print_area = f"A1:K{totals_row}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return totals_row


# ---------------------------------------------------------------------------
# Sheet 3 — Cost Allocation (per-category roll-up + capex/repair split + pie)
# ---------------------------------------------------------------------------

def build_allocation_tab(wb, variant):
    ws = wb.create_sheet("Cost Allocation")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 18), ("C", 14), ("D", 14), ("E", 14), ("F", 10),
        ("G", 4),
        ("H", 18), ("I", 14), ("J", 14),
        ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Cost Allocation",
                         prev_tab="Budget Detail", next_tab="ROI Projection")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "By-category roll-up · Capex vs Repair split · Spend by category pie"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- Per-category table (B6:F25, 16 cats) ----
    # Header row 6
    cat_headers = [("B6", "Category"), ("C6", "Estimate"),
                   ("D6", "Actual"), ("E6", "Variance"), ("F6", "Var %")]
    for cell_ref, label in cat_headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Each category row references Budget Detail SUMIFS
    detail_first = DETAIL_FIRST_DATA_ROW
    detail_last = DETAIL_FIRST_DATA_ROW + LINE_CAPACITY - 1
    detail_b_range = f"'Budget Detail'!B{detail_first}:B{detail_last}"
    detail_d_range = f"'Budget Detail'!D{detail_first}:D{detail_last}"
    detail_e_range = f"'Budget Detail'!E{detail_first}:E{detail_last}"

    cat_first_row = 7
    for i, cat in enumerate(CATEGORIES):
        r = cat_first_row + i
        # Category label
        cell = ws.cell(row=r, column=2, value=cat)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Estimate sumif
        cell = ws.cell(row=r, column=3,
                       value=f'=SUMIFS({detail_d_range},{detail_b_range},B{r})')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Actual sumif
        cell = ws.cell(row=r, column=4,
                       value=f'=SUMIFS({detail_e_range},{detail_b_range},B{r})')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Variance
        cell = ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0;[Red]"-$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Var %
        cell = ws.cell(row=r, column=6,
                       value=f'=IFERROR((D{r}-C{r})/C{r},0)')
        apply_style(cell, formula_cell_style())
        cell.number_format = "0.0%"
        cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 16

    cat_last_row = cat_first_row + len(CATEGORIES) - 1  # = 22

    # Totals row at row 26 (cat_last_row + 4 leaves space)
    totals_row = cat_last_row + 4
    cell = ws.cell(row=totals_row, column=2, value="TOTAL")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in [3, 4, 5]:
        col_letter = get_column_letter(col)
        cell = ws.cell(row=totals_row, column=col,
                       value=f"=SUM({col_letter}{cat_first_row}:{col_letter}{cat_last_row})")
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.number_format = '"$"#,##0'
    cell = ws.cell(row=totals_row, column=6,
                   value=f"=IFERROR((D{totals_row}-C{totals_row})/C{totals_row},0)")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0%"
    ws.row_dimensions[totals_row].height = 24

    # ---- Capex vs Repair split (H6:J9) ----
    headers_cr = [("H6", "Classification"), ("I6", "Estimate"), ("J6", "Actual")]
    for cell_ref, label in headers_cr:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    detail_j_range = f"'Budget Detail'!J{detail_first}:J{detail_last}"

    for i, cls in enumerate(CAPEX_REPAIR):
        r = 7 + i
        cell = ws.cell(row=r, column=8, value=cls)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        cell = ws.cell(row=r, column=9,
                       value=f'=SUMIFS({detail_d_range},{detail_j_range},H{r})')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        cell = ws.cell(row=r, column=10,
                       value=f'=SUMIFS({detail_e_range},{detail_j_range},H{r})')
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r].height = 16

    # Capex/Repair Total row
    cr_total_row = 9
    cell = ws.cell(row=cr_total_row, column=8, value="TOTAL")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_BG_LIGHT)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for col in [9, 10]:
        col_letter = get_column_letter(col)
        cell = ws.cell(row=cr_total_row, column=col,
                       value=f"=SUM({col_letter}7:{col_letter}8)")
        cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.number_format = '"$"#,##0'
    ws.row_dimensions[cr_total_row].height = 22

    # Tax-treatment note (H11:J14)
    ws.merge_cells("H11:J14")
    c = ws["H11"]
    c.value = (
        "Capex = depreciated over time (typically 27.5 yr res / 5-15 yr "
        "via cost seg). Repair = deductible in the year incurred. "
        "Hand this split to your CPA."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)

    # ---- Pie chart "Spend by category" ----
    pie = PieChart()
    pie.title = "Spend by Category (Actual)"
    pie.height = 9
    pie.width = 13
    # Data: actual column for each cat. Categories: cat name column
    data_ref = Reference(ws, min_col=4, min_row=6,
                          max_row=cat_last_row, max_col=4)  # include header row
    cats_ref = Reference(ws, min_col=2, min_row=cat_first_row,
                          max_row=cat_last_row)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.dataLabels = DataLabelList(showCatName=True, showPercent=True)
    style_chart(pie)
    ws.add_chart(pie, "H16")

    # Print setup
    ws.print_area = f"A1:L{totals_row + 2}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return totals_row


# ---------------------------------------------------------------------------
# Sheet 4 — ROI Projection
# ---------------------------------------------------------------------------

def build_roi_tab(wb, variant, allocation_total_row):
    ws = wb.create_sheet("ROI Projection")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "ROI Projection",
                         prev_tab="Cost Allocation", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Pre vs post-rehab nightly rate · year-1 lift · payback period · 5-yr cumulative"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ----- INPUTS section -----
    def _section_band(row, label):
        ws.merge_cells(f"A{row}:L{row}")
        cc = ws[f"A{row}"]
        cc.value = label
        cc.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
        cc.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22

    def _input_row(row, label, value, fmt, note=""):
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=row, column=3, value=value)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt
        if note:
            ws.merge_cells(f"E{row}:L{row}")
            cc = ws[f"E{row}"]
            cc.value = note
            cc.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
            cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    def _output_row(row, label, formula, fmt, emphasize=False):
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_HEAD if emphasize else FONT_BODY,
            size=12 if emphasize else 11,
            bold=True,
            color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=row, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        if emphasize:
            cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        ws.row_dimensions[row].height = 22 if emphasize else 18

    _section_band(6, "REVENUE INPUTS")
    _input_row(7, "Pre-rehab nightly rate ($):",
               _val(variant, SAMPLE_SETTINGS["pre_rehab_adr"]),
               '"$"#,##0', "ADR before this rehab — your baseline")
    _input_row(8, "Post-rehab nightly rate ($):",
               _val(variant, SAMPLE_SETTINGS["post_rehab_adr"]),
               '"$"#,##0', "Defensible projection — pull from comp set")
    _input_row(9, "Occupancy lift (% points):",
               _val(variant, SAMPLE_SETTINGS["occ_lift_pct"]),
               "0.0%", "Extra occupancy nightly — e.g., +8% = 8 pts more")
    _input_row(10, "Post-rehab occupancy (%):",
               _val(variant, SAMPLE_SETTINGS["post_rehab_occ"]),
               "0.0%", "Used for 5-yr cumulative lift forecast")

    _section_band(12, "OUTPUTS")
    # Pull total rehab spend from Cost Allocation totals row, col D (Actual).
    total_spend_ref = f"'Cost Allocation'!D{allocation_total_row}"
    _output_row(13, "Total rehab spend (actual):",
                f"={total_spend_ref}", '"$"#,##0')
    # Year-1 revenue lift = (post - pre) * 365 * occ_lift  (per brief formula)
    _output_row(14, "Year-1 revenue lift:",
                "=(C8-C7)*365*C9", '"$"#,##0', emphasize=True)
    # Payback months
    _output_row(15, "Payback (months):",
                f"=IFERROR(({total_spend_ref})/(((C8-C7)*365*C9)/12),0)",
                "0.0")
    # Annual revenue lift, alternative: post_rate*post_occ*365 - pre_rate*(post_occ-occ_lift)*365
    # Using brief's interpretation: stick with year-1 formula for cumulative.
    _output_row(16, "Year-1 lift (alt: rate × occ delta):",
                "=(C8*C10-C7*(C10-C9))*365", '"$"#,##0')

    _section_band(18, "5-YEAR PROJECTION")
    _output_row(19, "5-year cumulative lift:",
                "=C14*5", '"$"#,##0')
    # 5-yr ROI = cumulative lift / spend
    _output_row(20, "5-year ROI on rehab:",
                f"=IFERROR((C14*5)/{total_spend_ref},0)",
                "0.0%", emphasize=True)
    _output_row(21, "Net 5-yr profit (lift − spend):",
                f"=C14*5-{total_spend_ref}", '"$"#,##0')

    # Reassurance note
    ws.merge_cells("A24:L26")
    c = ws["A24"]
    c.value = (
        "Why this matters: rehab dollars are an investment, not an expense. "
        "If your post-rehab ADR + occupancy lift can't pay back the spend in "
        "60 months, the rehab is over-scoped. Right-size before you sign "
        "contractor estimates."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[24].height = 18
    ws.row_dimensions[25].height = 18
    ws.row_dimensions[26].height = 18

    brand_footer(ws, 29,
                 version_line=f"{SKU} · ROI Projection")

    ws.print_area = "A1:L31"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 5 — Settings
# ---------------------------------------------------------------------------

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    set_col_widths(ws, [
        ("A", 30), ("B", 24), ("C", 14), ("D", 14), ("E", 14), ("F", 14),
    ])

    compact_header_band(ws, "Settings",
                         prev_tab="ROI Projection", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = "Property identity · rehab dates · contingency · year-end archive"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # Row 5 — Property name (B5)
    a = ws.cell(row=5, column=1, value="Property name:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=5, column=2, value=_val(variant, SAMPLE_SETTINGS["property_name"]))
    apply_style(b, input_cell_style())
    ws.row_dimensions[5].height = 20

    # Row 6 — Active rehab year (used instead of YEAR(TODAY()))
    a = ws.cell(row=6, column=1, value="Active rehab year:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=6, column=2, value=ACTIVE_REHAB_YEAR)
    apply_style(b, input_cell_style())
    b.number_format = "0"
    ws.row_dimensions[6].height = 18
    ws.merge_cells("C6:F6")
    c6 = ws["C6"]
    c6.value = "Bump in January for the next rehab. Drives any year-filtered logic."
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Row 7 — Rehab start date (B7)
    a = ws.cell(row=7, column=1, value="Rehab start date:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=7, column=2,
                 value=_val(variant, SAMPLE_SETTINGS["rehab_start"]))
    apply_style(b, input_cell_style())
    b.number_format = "yyyy-mm-dd"
    ws.row_dimensions[7].height = 18

    # Row 9 — Rehab end date target (B9)
    a = ws.cell(row=9, column=1, value="Rehab end date (target):")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=9, column=2,
                 value=_val(variant, SAMPLE_SETTINGS["rehab_end"]))
    apply_style(b, input_cell_style())
    b.number_format = "yyyy-mm-dd"
    ws.row_dimensions[9].height = 18

    # Row 11 — Contingency % (B11)
    a = ws.cell(row=11, column=1, value="Contingency %:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=11, column=2,
                 value=_val(variant, SAMPLE_SETTINGS["contingency_pct"])
                 if variant == "demo" else 0.15)
    apply_style(b, input_cell_style())
    b.number_format = "0.0%"
    ws.row_dimensions[11].height = 18
    ws.merge_cells("C11:F11")
    c11 = ws["C11"]
    c11.value = "Default 15%. Reserve held back for surprises."
    c11.font = italic_muted
    c11.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Row 13 — Contingency $ (computed: estimate × pct)
    a = ws.cell(row=13, column=1, value="Contingency $ (auto):")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(row=13, column=2,
                   value="='Cost Allocation'!C26*B11")
    apply_style(cell, formula_cell_style())
    cell.number_format = '"$"#,##0'
    ws.row_dimensions[13].height = 20

    # ----- Year-end Archive table -----
    ws.row_dimensions[15].height = 10
    sect = ws.cell(row=16, column=1, value="Year-end Archive")
    sect.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[16].height = 22

    ws.merge_cells("A17:F17")
    c17 = ws["A17"]
    c17.value = (
        "Each January, copy the YTD totals from the Cost Allocation tab into "
        "the row for the closing year, then clear Budget Detail for the next "
        "rehab project."
    )
    c17.font = italic_muted
    c17.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[17].height = 24

    archive_headers = ["Year", "Property", "Total $", "5-yr ROI %", "Notes"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=18, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[18].height = 18

    for idx, year in enumerate(range(2024, 2031), start=19):
        cell = ws.cell(row=idx, column=1, value=year)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=idx, column=3).number_format = '"$"#,##0'
        ws.cell(row=idx, column=4).number_format = "0.0%"
        ws.row_dimensions[idx].height = 16

    brand_footer(ws, 28, version_line=f"{SKU} · Settings")

    ws.print_area = "A1:F30"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_detail_tab(wb, variant)
    allocation_total_row = build_allocation_tab(wb, variant)
    build_roi_tab(wb, variant, allocation_total_row)
    build_settings_tab(wb, variant)

    wb.properties.title = "Rehab Budget + ROI — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Rehab line-item register with embedded ROI projection. "
        "Per-line estimate vs actual, contractor, payment status, capex/repair tag. "
        "Cost allocation by category, capex/repair split, year-1 lift, payback."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
