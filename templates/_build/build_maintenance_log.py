"""Build OPS-006 Maintenance Log + Vendor CRM (v2.2 visual standard).

Operational-mode dual-shape: a chronological work-order log (TAX-001 flat-log
pattern) plus a vendor register (TAX-003 register pattern). One source of
truth for repairs, recurring services, and preferred vendors across the
portfolio.

Implements templates/_briefs/OPS-006-maintenance-log-vendor-crm.md.
Generates DEMO + BLANK masters in templates/_masters/.

Usage:
    python build_maintenance_log.py
"""
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "OPS-006"
NAME = "maintenance-log-vendor-crm"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026

# --- Constants ---

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]

CATEGORIES = [
    "Plumbing", "HVAC", "Electrical", "Appliance",
    "Cosmetic", "Pest", "Landscape", "Other",
]

STATUSES = ["Open", "Scheduled", "Done", "Deferred"]

CAPEX_REPAIR = ["Repair", "Capex"]

TRADES = [
    "Plumber", "HVAC", "Electrician", "Handyman",
    "Cleaner", "Pest control", "Landscaper", "Locksmith",
    "Roofer", "Other",
]

# Work Orders sample data — 25 rows Jan-Mar 2026, mix of statuses.
# Tuple: (date_opened, property, category, issue, vendor, date_completed,
#         cost, receipt, status, capex_repair, notes)
WORK_ORDERS = [
    ("2026-01-04", "Smokies Ridge", "Plumbing",   "Kitchen faucet drip",            "Quick Plumbing",   "2026-01-06", 145, "Y", "Done",      "Repair", "Replaced cartridge"),
    ("2026-01-08", "Creek Side",    "HVAC",       "Furnace not igniting",           "Mountain HVAC",    "2026-01-09", 320, "Y", "Done",      "Repair", "Replaced igniter"),
    ("2026-01-12", "Lakehouse A",   "Appliance",  "Dishwasher won't drain",         "Bob Handyman",     "2026-01-15", 180, "Y", "Done",      "Repair", "Cleared trap"),
    ("2026-01-15", "Smokies Ridge", "Pest",       "Quarterly pest spray",           "BugOff Pest",      "2026-01-15", 95,  "Y", "Done",      "Repair", "Q1 service"),
    ("2026-01-20", "Creek Side",    "Cosmetic",   "Touch-up paint guest bedroom",   "Bob Handyman",     "2026-01-22", 220, "",  "Done",      "Repair", ""),
    ("2026-01-25", "Lakehouse A",   "Electrical", "Hot tub breaker tripping",       "Spark Electric",   "2026-01-28", 410, "Y", "Done",      "Repair", "Wiring damaged by rodent"),
    ("2026-02-02", "Smokies Ridge", "Landscape",  "Storm debris cleanup",           "Joe Landscape",    "2026-02-04", 180, "",  "Done",      "Repair", ""),
    ("2026-02-06", "Creek Side",    "Plumbing",   "Hot tub leak inspection",        "Quick Plumbing",   "2026-02-07", 160, "Y", "Done",      "Repair", "Tightened union"),
    ("2026-02-10", "Lakehouse A",   "HVAC",       "Mini-split blowing warm",        "Mountain HVAC",    "2026-02-12", 285, "Y", "Done",      "Repair", "Recharged refrigerant"),
    ("2026-02-14", "Smokies Ridge", "Appliance",  "Fridge ice maker out",           "Bob Handyman",     "",           0,   "",  "Scheduled", "Repair", "Parts on order"),
    ("2026-02-18", "Creek Side",    "Cosmetic",   "Replace damaged blinds",         "Bob Handyman",     "2026-02-22", 140, "Y", "Done",      "Repair", "Living room"),
    ("2026-02-22", "Lakehouse A",   "Pest",       "Rodent in attic",                "BugOff Pest",      "2026-02-25", 225, "Y", "Done",      "Repair", "Set traps + entry seal"),
    ("2026-02-28", "Smokies Ridge", "Plumbing",   "Toilet running master bath",     "Quick Plumbing",   "",           0,   "",  "Open",      "Repair", "Reported by guest"),
    ("2026-03-02", "Creek Side",    "Electrical", "Outlet not working kitchen",     "Spark Electric",   "2026-03-04", 165, "Y", "Done",      "Repair", "GFCI replaced"),
    ("2026-03-05", "Lakehouse A",   "Landscape",  "Mulch + shrub trim",             "Joe Landscape",    "2026-03-08", 380, "",  "Done",      "Repair", "Spring refresh"),
    ("2026-03-09", "Smokies Ridge", "HVAC",       "Spring tune-up",                 "Mountain HVAC",    "2026-03-12", 175, "Y", "Done",      "Repair", "Annual recurring"),
    ("2026-03-14", "Creek Side",    "Appliance",  "Microwave replacement",          "Bob Handyman",     "2026-03-16", 320, "Y", "Done",      "Capex",  "Out of warranty"),
    ("2026-03-18", "Lakehouse A",   "Cosmetic",   "Deck stain refresh",             "Bob Handyman",     "",           0,   "",  "Scheduled", "Capex",  "Waiting on dry weekend"),
    ("2026-03-20", "Smokies Ridge", "Electrical", "Smart-lock battery + reset",     "Lock-N-Key",       "2026-03-20", 65,  "Y", "Done",      "Repair", ""),
    ("2026-03-22", "Creek Side",    "Plumbing",   "Slow shower drain",              "Quick Plumbing",   "",           0,   "",  "Open",      "Repair", "Guest reported"),
    ("2026-03-24", "Lakehouse A",   "HVAC",       "Filter replacement",             "Mountain HVAC",    "2026-03-24", 45,  "Y", "Done",      "Repair", "Quarterly"),
    ("2026-03-26", "Smokies Ridge", "Pest",       "Carpenter ant treatment",        "BugOff Pest",      "2026-03-28", 175, "Y", "Done",      "Repair", "Around deck posts"),
    ("2026-03-28", "Creek Side",    "Cosmetic",   "Replace bath caulking",          "Bob Handyman",     "",           0,   "",  "Deferred",  "Repair", "Postponed to Q2"),
    ("2026-03-30", "Lakehouse A",   "Other",      "Replace welcome mat + signage",  "",                 "2026-03-30", 85,  "Y", "Done",      "Repair", "Self-installed"),
    ("2026-03-31", "Smokies Ridge", "Landscape",  "Gutter clean",                   "Joe Landscape",    "",           0,   "",  "Scheduled", "Repair", "Spring service"),
]

# Vendor register — 8 preferred vendors.
# Tuple: (name, trade, phone, email, license, insured, hourly, area,
#         last_used, rating, notes)
VENDORS = [
    ("Quick Plumbing",   "Plumber",      "(865) 555-0166", "info@quickplumbing.com",   "TN-PL-12847", "Y", 95,  "Sevierville · Gatlinburg",         "2026-03-22", 5, "Emergency 24/7"),
    ("Mountain HVAC",    "HVAC",         "(865) 555-0218", "service@mtnhvac.com",      "TN-HV-09122", "Y", 110, "All Smokies area",                 "2026-03-24", 5, "Maintenance contract option"),
    ("Bob Handyman",     "Handyman",     "(865) 555-0198", "bob@ridgerepair.com",      "—",           "Y", 65,  "Sevierville · Pigeon Forge",       "2026-03-28", 4, "Best for small repairs"),
    ("Spark Electric",   "Electrician",  "(865) 555-0341", "dispatch@sparkelectric.co","TN-EL-04518", "Y", 125, "Knoxville · Sevierville",          "2026-03-04", 5, "Slow on weekends"),
    ("BugOff Pest",      "Pest control", "(865) 555-0277", "schedule@bugoff.co",       "TN-PC-22019", "Y", 0,   "All Smokies area",                 "2026-03-28", 4, "Quarterly contract $95/visit"),
    ("Joe Landscape",    "Landscaper",   "(865) 555-0155", "joe@joeslawn.com",         "—",           "N", 50,  "Pigeon Forge · Gatlinburg",        "2026-03-31", 4, "No COI yet — request"),
    ("Lock-N-Key",       "Locksmith",    "(865) 555-0399", "help@locknkey.co",         "TN-LK-00821", "Y", 85,  "Knoxville metro",                  "2026-03-20", 5, "24/7 emergency"),
    ("Smokies Clean",    "Cleaner",      "(865) 555-0145", "sarah@smokiesclean.com",   "—",           "Y", 35,  "Gatlinburg · Sevierville",         "2026-03-30", 5, "Primary turnover"),
]

# Recurring services — 8 rows pre-stocked.
# Tuple: (service, property, frequency_days, last_done, vendor, status)
RECURRING = [
    ("HVAC spring tune-up",    "Smokies Ridge", 365, "2025-04-15", "Mountain HVAC",  "Active"),
    ("HVAC fall tune-up",      "Smokies Ridge", 365, "2025-10-08", "Mountain HVAC",  "Active"),
    ("Gutter clean",           "Creek Side",    180, "2025-10-12", "Joe Landscape",  "Active"),
    ("Pest spray (quarterly)", "All",           90,  "2026-01-15", "BugOff Pest",    "Active"),
    ("Septic pump (3-yr)",     "Lakehouse A",   1095,"2024-06-01", "Quick Plumbing", "Active"),
    ("Dryer vent clean",       "Smokies Ridge", 365, "2025-05-20", "Bob Handyman",   "Active"),
    ("Deck stain refresh",     "Lakehouse A",   730, "2024-04-10", "Bob Handyman",   "Active"),
    ("Smoke detector test",    "All",           180, "2025-12-01", "Self",           "Active"),
]


# --- Helpers ---

def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational hero + cards + activity dashboard)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- Navy hero (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Maintenance Log + Vendor CRM"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "The next emergency call is one tab away."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- "What this does" card (rows 10-16) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "One source of truth for repairs, maintenance, and recurring services across "
        "every property. The Work Orders log is your chronological record of every "
        "ticket; the Vendors register is your preferred-vendor CRM (plumber, HVAC, "
        "handyman); Recurring Services tracks the annual rituals (HVAC tune-ups, "
        "pest, septic). Annual Cost Report breaks $ by property + category for tax."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- "How to use" card (rows 18-25) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart = [
        "① Log every repair the moment it's reported — Status = Open until done.",
        "② Pull the Vendor name from the Vendors dropdown — that wires up YTD spend.",
        "③ Tag Capex vs Repair so the Annual Cost Report keeps your CPA happy.",
        "④ Recurring Services flags anything coming due in 14 days — review weekly.",
        "⑤ Open + > 14 days old auto-highlights red. Close it or escalate.",
    ]
    for i, item in enumerate(quickstart):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- Primary button (rows 26-29) ---
    pseudo_button(ws, "A26", "L29",
                   "→  ADD A NEW WORK ORDER  (OPEN LOG)",
                   "'Work Orders'!A6", variant="primary")
    for r in range(26, 30):
        ws.row_dimensions[r].height = 22

    # --- Activity at-a-glance (rows 31-37) ---
    for r in range(31, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Open tickets
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "OPEN TICKETS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = '=COUNTIF(\'Work Orders\'!J6:J305,"Open")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "still unresolved"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): YTD spend
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "YTD SPEND"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = (
        "=SUMIFS('Work Orders'!H6:H305,"
        "'Work Orders'!A6:A305,\">=\"&DATE(Settings!$B$5,1,1),"
        "'Work Orders'!A6:A305,\"<\"&DATE(Settings!$B$5+1,1,1))"
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "active tax year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Overdue
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "OVERDUE > 14d"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    # Open tickets where today minus date opened > 14
    c.value = (
        '=SUMPRODUCT(((\'Work Orders\'!J6:J305="Open")*'
        '(TODAY()-\'Work Orders\'!A6:A305>14)*'
        '(\'Work Orders\'!A6:A305<>"")))'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "needs follow-up"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Gold borders around 3 cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- Secondary nav (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Vendors",
                   "'Vendors'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Recurring",
                   "'Recurring Services'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Annual Cost Report",
                   "'Annual Cost Report'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- Tax / year-end callout ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📄 TAX TIME: print Annual Cost Report (File → Print → 'Print Active Sheet'). "
        "📅 YEAR END: bump Settings → Active tax year + clear closed tickets."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # Upgrade banner + footer
    add_upgrade_banner(ws, 44)
    brand_footer(ws, 46,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_work_orders_tab(wb, variant):
    """Sheet 1 — Work Orders (chronological log, capacity 300)."""
    ws = wb.create_sheet("Work Orders")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Work Orders",
                         prev_tab="Start", next_tab="Vendors")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per ticket — close it when the work is done"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Date Opened",          # A
        "Property",             # B
        "Category",             # C
        "Issue Summary",        # D
        "Vendor",               # E
        "Date Completed",       # F
        "Days Open",            # G  (formula)
        "$ Cost",               # H
        "Receipt",              # I
        "Status",               # J
        "Capex/Repair",         # K
        "Notes",                # L
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 12), ("B", 16), ("C", 13), ("D", 28),
        ("E", 18), ("F", 13), ("G", 9),  ("H", 11),
        ("I", 9),  ("J", 11), ("K", 12), ("L", 26),
    ])

    samples = WORK_ORDERS if variant == "demo" else []

    for i in range(6, 306):  # capacity 300
        sample_idx = i - 6
        sample = samples[sample_idx] if sample_idx < len(samples) else None

        if sample:
            (d_open, prop, cat, issue, vendor, d_done,
             cost, receipt, status, capex, notes) = sample

            a = ws.cell(row=i, column=1, value=_parse_date(d_open))
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"

            b = ws.cell(row=i, column=2, value=prop)
            apply_style(b, input_cell_style())

            c = ws.cell(row=i, column=3, value=cat)
            apply_style(c, input_cell_style())

            d = ws.cell(row=i, column=4, value=issue)
            apply_style(d, input_cell_style())

            e = ws.cell(row=i, column=5, value=vendor if vendor else None)
            apply_style(e, input_cell_style())

            f_val = _parse_date(d_done) if d_done else None
            f = ws.cell(row=i, column=6, value=f_val)
            apply_style(f, input_cell_style())
            f.number_format = "yyyy-mm-dd"

            h = ws.cell(row=i, column=8, value=cost if cost else None)
            apply_style(h, input_cell_style())
            h.number_format = '"$"#,##0.00'

            ic = ws.cell(row=i, column=9, value=receipt if receipt else None)
            apply_style(ic, input_cell_style())
            ic.alignment = Alignment(horizontal="center", vertical="center")

            j = ws.cell(row=i, column=10, value=status)
            apply_style(j, input_cell_style())

            k = ws.cell(row=i, column=11, value=capex if capex else None)
            apply_style(k, input_cell_style())

            l = ws.cell(row=i, column=12, value=notes if notes else None)
            apply_style(l, input_cell_style())
        else:
            # BLANK / unused capacity — still pre-style cells for UX
            for col in [1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12]:
                cell = ws.cell(row=i, column=col)
                apply_style(cell, input_cell_style())
            ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
            ws.cell(row=i, column=6).number_format = "yyyy-mm-dd"
            ws.cell(row=i, column=8).number_format = '"$"#,##0.00'
            ws.cell(row=i, column=9).alignment = Alignment(horizontal="center", vertical="center")

        # Days Open formula (col G):
        # If completed: F - A. Else if opened: TODAY()-A. Else "".
        g_cell = ws.cell(
            row=i, column=7,
            value=(
                f'=IF(A{i}="","",'
                f'IF(F{i}<>"",F{i}-A{i},TODAY()-A{i}))'
            ),
        )
        g_cell.number_format = "0"
        if sample:
            apply_style(g_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    # Dropdowns
    add_dropdown(ws, "B6:B305", "=Settings!$B$7:$B$16")
    add_dropdown(ws, "C6:C305", "=Settings!$B$27:$B$36")
    add_dropdown(ws, "E6:E305", "=Vendors!$A$6:$A$55")
    add_dropdown(ws, "I6:I305", '"Y,N"')
    add_dropdown(ws, "J6:J305", "=Settings!$B$18:$B$25")
    add_dropdown(ws, "K6:K305", '"Repair,Capex"')

    # Conditional formatting: row red where Status=Open AND Days open > 14.
    ws.conditional_formatting.add(
        "A6:L305",
        FormulaRule(
            formula=['AND($J6="Open",$G6>14)'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(color=COLOR_ERROR, bold=True),
        ),
    )

    # Freeze pane on header row
    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_vendors_tab(wb, variant):
    """Sheet 2 — Vendors (preferred-vendor CRM, capacity 50)."""
    ws = wb.create_sheet("Vendors")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Vendors",
                         prev_tab="Work Orders", next_tab="Recurring Services")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Preferred-vendor CRM — your call sheet for the next emergency"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Vendor Name",         # A
        "Trade",               # B
        "Phone",               # C
        "Email",               # D
        "License #",           # E
        "Insured?",            # F
        "Hourly Rate",         # G
        "Service Area",        # H
        "Last Used",           # I
        "YTD Spend",           # J  (formula)
        "Rating 1-5",          # K
        "Notes",               # L
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 22), ("B", 13), ("C", 16), ("D", 26),
        ("E", 15), ("F", 9),  ("G", 11), ("H", 22),
        ("I", 12), ("J", 12), ("K", 8),  ("L", 22),
    ])

    samples = VENDORS if variant == "demo" else []

    for i in range(6, 56):  # 50 capacity
        sample_idx = i - 6
        sample = samples[sample_idx] if sample_idx < len(samples) else None

        if sample:
            (name, trade, phone, email, lic, insured, hourly,
             area, last_used, rating, notes) = sample

            for col, val, fmt in [
                (1, name, None),
                (2, trade, None),
                (3, phone, None),
                (4, email, None),
                (5, lic, None),
                (6, insured, None),
                (7, hourly if hourly else None, '"$"#,##0'),
                (8, area, None),
                (9, _parse_date(last_used) if last_used else None, "yyyy-mm-dd"),
                (11, rating, "0"),
                (12, notes if notes else None, None),
            ]:
                cell = ws.cell(row=i, column=col, value=val)
                apply_style(cell, input_cell_style())
                if fmt:
                    cell.number_format = fmt
                if col == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12]:
                cell = ws.cell(row=i, column=col)
                apply_style(cell, input_cell_style())
            ws.cell(row=i, column=7).number_format = '"$"#,##0'
            ws.cell(row=i, column=9).number_format = "yyyy-mm-dd"
            ws.cell(row=i, column=11).number_format = "0"
            ws.cell(row=i, column=6).alignment = Alignment(horizontal="center", vertical="center")

        # YTD Spend formula — sums Work Orders!H where Vendor matches A{i}
        # AND date opened falls in Settings!$B$5 active tax year.
        j_cell = ws.cell(
            row=i, column=10,
            value=(
                f'=IF(A{i}="","",SUMIFS(\'Work Orders\'!$H$6:$H$305,'
                f'\'Work Orders\'!$E$6:$E$305,A{i},'
                f'\'Work Orders\'!$A$6:$A$305,">="&DATE(Settings!$B$5,1,1),'
                f'\'Work Orders\'!$A$6:$A$305,"<"&DATE(Settings!$B$5+1,1,1)))'
            ),
        )
        j_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(j_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    # Dropdowns
    add_dropdown(ws, "B6:B55",
                  '"Plumber,HVAC,Electrician,Handyman,Cleaner,'
                  'Pest control,Landscaper,Locksmith,Roofer,Other"')
    add_dropdown(ws, "F6:F55", '"Y,N"')
    add_dropdown(ws, "K6:K55", '"1,2,3,4,5"')

    # Conditional formatting — Insured? = N flagged red
    ws.conditional_formatting.add(
        "F6:F55",
        FormulaRule(
            formula=['F6="N"'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(color=COLOR_ERROR, bold=True),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_recurring_tab(wb, variant):
    """Sheet 3 — Recurring Services (HVAC tune-ups, gutter cleans, pest, etc)."""
    ws = wb.create_sheet("Recurring Services")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Recurring Services",
                         prev_tab="Vendors",
                         next_tab="Annual Cost Report")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Annual rituals — Next Due auto-calculates from Last Done + Frequency"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Service",          # A
        "Property",         # B
        "Frequency (days)", # C
        "Last Done",        # D
        "Next Due",         # E (formula)
        "Days Until",       # F (formula)
        "Vendor",           # G
        "Status",           # H
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 26), ("B", 16), ("C", 16), ("D", 13),
        ("E", 13), ("F", 12), ("G", 18), ("H", 12),
    ])

    samples = RECURRING if variant == "demo" else []

    for i in range(6, 31):  # capacity 25
        sample_idx = i - 6
        sample = samples[sample_idx] if sample_idx < len(samples) else None

        if sample:
            service, prop, freq, last_done, vendor, status = sample
            ws.cell(row=i, column=1, value=service)
            ws.cell(row=i, column=2, value=prop)
            ws.cell(row=i, column=3, value=freq)
            ws.cell(row=i, column=4, value=_parse_date(last_done))
            ws.cell(row=i, column=7, value=vendor)
            ws.cell(row=i, column=8, value=status)
        else:
            pass  # leave blank

        for col in [1, 2, 3, 4, 7, 8]:
            cell = ws.cell(row=i, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=i, column=3).number_format = "0"
        ws.cell(row=i, column=4).number_format = "yyyy-mm-dd"

        # Next Due formula: D + C
        e_cell = ws.cell(
            row=i, column=5,
            value=f'=IF(AND(D{i}<>"",C{i}<>""),D{i}+C{i},"")',
        )
        e_cell.number_format = "yyyy-mm-dd"
        apply_style(e_cell, formula_cell_style())

        # Days Until: E - TODAY()
        f_cell = ws.cell(
            row=i, column=6,
            value=f'=IF(E{i}="","",E{i}-TODAY())',
        )
        f_cell.number_format = "0"
        apply_style(f_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    add_dropdown(ws, "B6:B30",
                  '"Smokies Ridge,Creek Side,Lakehouse A,All"')
    add_dropdown(ws, "H6:H30", '"Active,Paused,Done"')

    # Conditional formatting on Days Until
    # Red: ≤ 0  (overdue)
    ws.conditional_formatting.add(
        "F6:F30",
        FormulaRule(
            formula=['AND(F6<>"",F6<=0)'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(color=COLOR_ERROR, bold=True),
        ),
    )
    # Gold-soft: 0 < x ≤ 14 (due soon)
    ws.conditional_formatting.add(
        "F6:F30",
        FormulaRule(
            formula=['AND(F6>0,F6<=14)'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_annual_report_tab(wb):
    """Sheet 4 — Annual Cost Report (per-property + per-category breakdown)."""
    ws = wb.create_sheet("Annual Cost Report")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Annual Cost Report · For Your CPA",
                         prev_tab="Recurring Services",
                         next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = (
        "📄 FOR YOUR CPA — print this tab. Filter by Settings → Active tax year."
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 28), ("B", 16), ("C", 14),
    ])

    bold_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # --- Top KPIs (rows 5-9) ---
    year_filter_template = (
        "SUMIFS('Work Orders'!$H$6:$H$305,"
        "'Work Orders'!$A$6:$A$305,\">=\"&DATE(Settings!$B$5,1,1),"
        "'Work Orders'!$A$6:$A$305,\"<\"&DATE(Settings!$B$5+1,1,1)"
    )

    kpis = [
        ("YTD Total Spend ($):",
         f"={year_filter_template})",
         '"$"#,##0.00'),
        ("Total Tickets:",
         '=COUNTIFS(\'Work Orders\'!$A$6:$A$305,">="&DATE(Settings!$B$5,1,1),'
         '\'Work Orders\'!$A$6:$A$305,"<"&DATE(Settings!$B$5+1,1,1))',
         "0"),
        ("Capex Spend ($):",
         f"={year_filter_template},"
         '\'Work Orders\'!$K$6:$K$305,"Capex")',
         '"$"#,##0.00'),
        ("Repair Spend ($):",
         f"={year_filter_template},"
         '\'Work Orders\'!$K$6:$K$305,"Repair")',
         '"$"#,##0.00'),
        ("Open Tickets (any year):",
         '=COUNTIF(\'Work Orders\'!$J$6:$J$305,"Open")',
         "0"),
    ]
    for i, (label, formula, fmt) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=label).font = bold_font
        b_cell = ws.cell(row=i, column=2, value=formula)
        apply_style(b_cell, formula_cell_style())
        if fmt:
            b_cell.number_format = fmt
        ws.row_dimensions[i].height = 16

    ws.row_dimensions[10].height = 10

    # --- $ by Category (rows 11-12 hdr, 13+) ---
    h11 = ws.cell(row=11, column=1, value="$ Spend by Category")
    h11.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[11].height = 20

    for col, h in enumerate(["Category", "$ Spend", "Tickets"], start=1):
        cell = ws.cell(row=12, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[12].height = 18

    cat_start = 13
    for idx, cat in enumerate(CATEGORIES, start=cat_start):
        ws.cell(row=idx, column=1, value=cat).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        b_cell = ws.cell(
            row=idx, column=2,
            value=(
                f"={year_filter_template},"
                f"'Work Orders'!$C$6:$C$305,A{idx})"
            ),
        )
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = '"$"#,##0.00'
        c_cell = ws.cell(
            row=idx, column=3,
            value=(
                f'=COUNTIFS(\'Work Orders\'!$A$6:$A$305,">="&DATE(Settings!$B$5,1,1),'
                f'\'Work Orders\'!$A$6:$A$305,"<"&DATE(Settings!$B$5+1,1,1),'
                f"'Work Orders'!$C$6:$C$305,A{idx})"
            ),
        )
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = "0"
        ws.row_dimensions[idx].height = 16

    cat_end = cat_start + len(CATEGORIES) - 1

    # --- $ by Property ---
    prop_hdr_row = cat_end + 2
    ws.row_dimensions[cat_end + 1].height = 10

    h_prop = ws.cell(row=prop_hdr_row, column=1, value="$ Spend by Property")
    h_prop.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[prop_hdr_row].height = 20

    for col, h in enumerate(["Property", "$ Spend", "Tickets"], start=1):
        cell = ws.cell(row=prop_hdr_row + 1, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[prop_hdr_row + 1].height = 18

    prop_start = prop_hdr_row + 2
    for idx, prop in enumerate(PROPERTIES, start=prop_start):
        ws.cell(row=idx, column=1, value=prop).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        b_cell = ws.cell(
            row=idx, column=2,
            value=(
                f"={year_filter_template},"
                f"'Work Orders'!$B$6:$B$305,A{idx})"
            ),
        )
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = '"$"#,##0.00'
        c_cell = ws.cell(
            row=idx, column=3,
            value=(
                f'=COUNTIFS(\'Work Orders\'!$A$6:$A$305,">="&DATE(Settings!$B$5,1,1),'
                f'\'Work Orders\'!$A$6:$A$305,"<"&DATE(Settings!$B$5+1,1,1),'
                f"'Work Orders'!$B$6:$B$305,A{idx})"
            ),
        )
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = "0"
        ws.row_dimensions[idx].height = 16

    prop_end = prop_start + len(PROPERTIES) - 1

    # --- Donut chart: $ by Category (anchored E5) ---
    donut = DoughnutChart()
    donut.title = "$ Spend by Category"
    donut.height = 9
    donut.width = 11
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=2, min_row=12,
                            max_row=cat_end, max_col=2)
    donut_cats = Reference(ws, min_col=1, min_row=cat_start, max_row=cat_end)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=False, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, "E5")

    # --- Column chart: $ by Property ---
    col_chart = BarChart()
    col_chart.type = "col"
    col_chart.title = "$ Spend by Property"
    col_chart.y_axis.title = "$ Spend"
    col_chart.x_axis.title = "Property"
    col_chart.legend = None
    col_chart.height = 7
    col_chart.width = 11
    prop_data = Reference(ws, min_col=2, min_row=prop_hdr_row + 1,
                           max_row=prop_end, max_col=2)
    prop_cats = Reference(ws, min_col=1, min_row=prop_start, max_row=prop_end)
    col_chart.add_data(prop_data, titles_from_data=True)
    col_chart.set_categories(prop_cats)
    col_chart.x_axis.delete = False
    col_chart.y_axis.delete = False
    col_chart.x_axis.tickLblPos = "low"
    style_chart(col_chart)
    ws.add_chart(col_chart, f"E{prop_hdr_row}")

    # Print setup
    ws.print_area = f"A1:L{prop_end + 2}"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    """Sheet 5 — Settings: active year + dropdown source lists + archive."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Annual Cost Report", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · property + status + category lists · year-end archive"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 32), ("B", 22), ("C", 6), ("D", 22), ("E", 22), ("F", 22),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active tax year (B5) — drives every $-by-year filter ---
    a5 = ws.cell(row=5, column=1, value="Active tax year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 20

    # Header for Property list (B7-B16)
    a6 = ws.cell(row=6, column=2, value="Property list (10 max):")
    a6.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[6].height = 18

    for idx in range(7, 17):
        prop_idx = idx - 7
        val = PROPERTIES[prop_idx] if prop_idx < len(PROPERTIES) else None
        cell = ws.cell(row=idx, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # Header for Status list (B18-B25)
    a17 = ws.cell(row=17, column=2, value="Status list:")
    a17.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[17].height = 18

    for idx in range(18, 26):
        st_idx = idx - 18
        val = STATUSES[st_idx] if st_idx < len(STATUSES) else None
        cell = ws.cell(row=idx, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # Header for Category list (B27-B36)
    a26 = ws.cell(row=26, column=2, value="Category list:")
    a26.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[26].height = 18

    for idx in range(27, 37):
        cat_idx = idx - 27
        val = CATEGORIES[cat_idx] if cat_idx < len(CATEGORIES) else None
        cell = ws.cell(row=idx, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # --- Year-end Archive (rows 5-13 in cols D-F) ---
    d5 = ws.cell(row=5, column=4, value="Year-end Archive")
    d5.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)

    ws.merge_cells("D6:F6")
    c6 = ws["D6"]
    c6.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then optionally clear closed Work Orders for the new year."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[6].height = 30

    archive_headers = ["Year", "Total $ Spent", "# Tickets"]
    for col, h in enumerate(archive_headers, start=4):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 18

    for idx, year in enumerate(range(2024, 2031), start=8):
        ws.cell(row=idx, column=4, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [5, 6]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=idx, column=5).number_format = '"$"#,##0.00'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_work_orders_tab(wb, variant)
    build_vendors_tab(wb, variant)
    build_recurring_tab(wb, variant)
    build_annual_report_tab(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Maintenance Log + Vendor CRM{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Maintenance work-order log + preferred-vendor CRM for STR portfolios."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
