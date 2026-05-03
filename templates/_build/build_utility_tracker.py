"""Build OPS-007 Utility Usage + Trend Tracker Excel files (v2.2 standard).

Operational-mode tracker: customer logs every utility bill (electric, gas,
water, internet, trash, propane) per property. Surfaces $/night, YoY trend,
and "anomaly" flags when a month's bill jumps >25% vs prior year — catches
leaks, unmonitored loads, and pricing changes.

Forked from build_pl_single_property.py (12-column matrix pattern).

Generates BOTH DEMO and BLANK variants from shared code (per the suite-wide
customer-eye theme: pre-populated demo data converts 5x better than empty
cells, but customers still need a clean copy to drop their own data into).
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, Reference

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "OPS-007"
NAME = "utility-usage-tracker"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

SERVICES = ["Electric", "Gas", "Water", "Internet", "Trash", "Propane"]

PROPERTIES_DEMO = [
    "Smokies Ridge Cabin",
    "Pigeon Forge Lodge",
    "Sevierville Chalet",
]

ACTIVE_TAX_YEAR = 2026

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Bills capacity
BILLS_CAPACITY = 500
BILLS_FIRST_ROW = 6
BILLS_LAST_ROW = BILLS_FIRST_ROW + BILLS_CAPACITY - 1   # = 505


# ---------------------------------------------------------------------------
# Sample data (DEMO)
# Full 2025 + Jan-Mar 2026 across 3 properties × 6 services.
# One seeded anomaly: Property 2 March 2026 water = 2x prior year (fake leak).
# ---------------------------------------------------------------------------

def _seed_bills():
    """Generate ~84 bill rows for the DEMO variant."""
    # base monthly bill amount + usage by service
    base = {
        # service: ($/mo, usage/mo, units)
        "Electric": (135, 1100, "kWh"),
        "Gas":      (60,  45,   "therms"),
        "Water":    (55,  3500, "gal"),
        "Internet": (89,  None, "GB"),
        "Trash":    (35,  None, "pickups"),
        "Propane":  (110, 80,   "gal"),
    }
    # per-property multipliers
    prop_mult = {
        "Smokies Ridge Cabin":   1.00,
        "Pigeon Forge Lodge":    1.20,
        "Sevierville Chalet":    0.90,
    }
    # mild seasonality factor by month for utility variability
    season = {1: 1.20, 2: 1.15, 3: 1.05, 4: 0.95, 5: 0.90, 6: 1.00,
              7: 1.10, 8: 1.10, 9: 0.95, 10: 0.90, 11: 1.05, 12: 1.20}

    rows = []
    # 2025 full year + 2026 Jan-Mar
    plan = [(2025, m) for m in range(1, 13)] + [(2026, m) for m in range(1, 4)]

    for prop in PROPERTIES_DEMO:
        for (yr, mo) in plan:
            for svc, (amt, usage, units) in base.items():
                pm = prop_mult[prop]
                sf = season[mo]
                bill_amt = round(amt * pm * sf, 2)
                if usage is None:
                    use_val = ""
                else:
                    use_val = round(usage * pm * sf, 1)

                # Period start = 1st, end = 28th of same month
                period_start = f"{yr:04d}-{mo:02d}-01"
                period_end = f"{yr:04d}-{mo:02d}-28"
                bill_date = f"{yr:04d}-{mo:02d}-15"
                notes = ""

                # Seed the anomaly: Property 2 March 2026 water = 2x prior year
                if (prop == "Pigeon Forge Lodge" and svc == "Water"
                        and yr == 2026 and mo == 3):
                    bill_amt = round(bill_amt * 2.1, 2)
                    if use_val != "":
                        use_val = round(use_val * 2.1, 1)
                    notes = "Possible leak — verify"

                rows.append((
                    bill_date, prop, svc, period_start, period_end,
                    bill_amt, use_val, units, notes,
                ))
    return rows


SAMPLE_BILLS = _seed_bills()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    return datetime.strptime(s, "%Y-%m-%d").date()


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (KPI hero + cards + nav)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Utility Usage + Trend"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Catch a leak before it costs a month."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Verdict cell (anomaly count) — single headline answer
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(COUNTA(\'Bills Log\'!B6:B505)=0,'
        '"\U0001F4CA  Log your first bill to see your verdict.",'
        'IF(\'Trend & Anomalies\'!$B$3=0,'
        '"✅  No anomalies detected — usage is steady year-over-year.",'
        '"⚠  "&\'Trend & Anomalies\'!$B$3&" anomalies detected — see Trend & Anomalies tab."))'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"OPS-007 · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Track every utility bill (electric, gas, water, internet, trash, propane) per "
        "property in a 12-column matrix. The Trend & Anomalies tab compares each month "
        "to the prior year and flags any service whose bill jumped >25% — your early "
        "warning system for leaks, runaway HVAC, and price hikes you didn't see coming."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Settings: list your properties + confirm active year.",
        "② Bills Log: log every utility bill — date, property, service, $, usage, notes.",
        "③ Property Matrix: per-property 12-month roll-up by service (auto).",
        "④ Trend & Anomalies: scan the ⚠ flags — anything >25% YoY needs a look.",
        "⑤ Bump 'Active tax year' on Settings each January so YoY checks stay aligned.",
        "⑥ Log bills the day they arrive — backfilling kills the trend signal.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 20

    # --- ZONE 4: Primary "LOG BILL" button ---
    pseudo_button(ws, "A27", "L30",
                   "→  LOG A BILL",
                   "'Bills Log'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: 3 KPI cards ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): YTD $
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "YTD UTILITY $"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = (
        '=SUMIFS(\'Bills Log\'!F:F,'
        '\'Bills Log\'!A:A,">="&DATE(Settings!$B$5,1,1),'
        '\'Bills Log\'!A:A,"<"&DATE(Settings!$B$5+1,1,1))'
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "all properties combined"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Avg $/month
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "AVG $/MONTH"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = (
        '=IFERROR(SUMIFS(\'Bills Log\'!F:F,'
        '\'Bills Log\'!A:A,">="&DATE(Settings!$B$5,1,1),'
        '\'Bills Log\'!A:A,"<"&DATE(Settings!$B$5+1,1,1))/12,0)'
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "across all services"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Anomaly count
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "ANOMALIES"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = "='Trend & Anomalies'!$B$3"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0"
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "YoY jumps >25%"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Gold borders around the 3 cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav ---
    pseudo_button(ws, "A39", "C40", "Bills Log",
                   "'Bills Log'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Property Matrix",
                   "'Property Matrix'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "⚠ Trend & Anomalies",
                   "'Trend & Anomalies'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Anomaly explainer ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "⚠ ANOMALY = a month's bill jumped >25% vs the same month last year. "
        "Common causes: leak, AC/heat running while vacant, rate hike, "
        "or a misread meter. Check the Trend & Anomalies tab monthly."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Upgrade banner ---
    add_upgrade_banner(ws, 44)

    # --- ZONE 9: Footer ---
    brand_footer(ws, 46,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_bills_log(wb, variant):
    """Sheet 1 — Bills Log (flat log of every utility bill)."""
    ws = wb.create_sheet("Bills Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Bills Log",
                         prev_tab="Start", next_tab="Property Matrix")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per utility bill — date, property, service, $, usage"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 24), ("C", 14), ("D", 12), ("E", 12),
        ("F", 12), ("G", 12), ("H", 10), ("I", 12), ("J", 28),
    ])

    headers = [
        "Bill Date", "Property", "Service",
        "Period Start", "Period End",
        "$ Amount", "Usage", "Units", "$/Unit", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    sample_rows = SAMPLE_BILLS if variant == "demo" else []

    for i, r in enumerate(sample_rows, start=BILLS_FIRST_ROW):
        bill_date, prop, svc, ps, pe, amt, usage, units, notes = r

        a = ws.cell(row=i, column=1, value=_parse_date(bill_date))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=prop)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=svc)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=_parse_date(ps))
        apply_style(d, input_cell_style())
        d.number_format = "yyyy-mm-dd"

        e = ws.cell(row=i, column=5, value=_parse_date(pe))
        apply_style(e, input_cell_style())
        e.number_format = "yyyy-mm-dd"

        f = ws.cell(row=i, column=6, value=amt)
        apply_style(f, input_cell_style())
        f.number_format = '"$"#,##0.00'

        g = ws.cell(row=i, column=7, value=usage if usage != "" else None)
        apply_style(g, input_cell_style())
        g.number_format = "#,##0.0"

        h_cell = ws.cell(row=i, column=8, value=units)
        apply_style(h_cell, input_cell_style())
        h_cell.alignment = Alignment(horizontal="center", vertical="center")

        # $/unit formula — divides $ by usage; IFERROR keeps blank rows clean
        i_cell = ws.cell(row=i, column=9, value=f'=IFERROR(F{i}/G{i},"")')
        apply_style(i_cell, formula_cell_style())
        i_cell.number_format = '"$"#,##0.0000'

        j_cell = ws.cell(row=i, column=10, value=notes if notes else None)
        apply_style(j_cell, input_cell_style())

        ws.row_dimensions[i].height = 16

    # Blank capacity rows
    last_data_row = len(sample_rows) + (BILLS_FIRST_ROW - 1)
    for row_idx in range(last_data_row + 1, BILLS_LAST_ROW + 1):
        for col_idx in range(1, 11):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == 9:
                cell.value = f'=IFERROR(F{row_idx}/G{row_idx},"")'
                apply_style(cell, formula_cell_style())
                cell.number_format = '"$"#,##0.0000'
            else:
                apply_style(cell, input_cell_style())
                if col_idx in (1, 4, 5):
                    cell.number_format = "yyyy-mm-dd"
                if col_idx == 6:
                    cell.number_format = '"$"#,##0.00'
                if col_idx == 7:
                    cell.number_format = "#,##0.0"
                if col_idx == 8:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dropdowns
    add_dropdown(ws, f"B{BILLS_FIRST_ROW}:B{BILLS_LAST_ROW}",
                 "=Settings!$B$7:$B$16")
    add_dropdown(ws, f"C{BILLS_FIRST_ROW}:C{BILLS_LAST_ROW}",
                 "=Settings!$B$18:$B$25")

    ws.freeze_panes = f"A{BILLS_FIRST_ROW}"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def _build_property_matrix_section(ws, start_row, prop_index, prop_ref):
    """Build a per-property block on the Property Matrix tab.

    Layout per property block (rows relative to start_row):
      r0     : compact header band (gold-soft, prop name)
      r1     : column headers ("Service", Jan..Dec, Total, Avg)
      r2..r7 : 6 service rows with SUMIFS per month
      r8     : Total row (sum of services)
      r9     : Avg/month row
      blank gap row before next property block
    Returns the row after the block (where a chart can anchor / next block starts).
    """
    # Property header strip (merged row, gold-soft)
    hdr_row = start_row
    ws.merge_cells(start_row=hdr_row, start_column=1,
                   end_row=hdr_row, end_column=15)
    hdr = ws.cell(row=hdr_row, column=1)
    # Use the property reference (cell-link) so renaming on Settings cascades.
    hdr.value = f'="Property: "&{prop_ref}'
    hdr.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    hdr.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[hdr_row].height = 22

    # Column headers
    col_hdr_row = hdr_row + 1
    headers = ["Service"] + MONTHS + ["Total", "Avg"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=col_hdr_row, column=col_idx, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[col_hdr_row].height = 18

    # Service rows
    for s_idx, svc in enumerate(SERVICES):
        r = col_hdr_row + 1 + s_idx
        a = ws.cell(row=r, column=1, value=svc)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 16

        for m in range(1, 13):
            col = m + 1
            if m < 12:
                end_year_expr = "Settings!$B$5"
                end_month = m + 1
            else:
                end_year_expr = "Settings!$B$5+1"
                end_month = 1

            formula = (
                f"=SUMIFS('Bills Log'!$F$6:$F${BILLS_LAST_ROW},"
                f"'Bills Log'!$A$6:$A${BILLS_LAST_ROW},\">=\"&DATE(Settings!$B$5,{m},1),"
                f"'Bills Log'!$A$6:$A${BILLS_LAST_ROW},\"<\"&DATE({end_year_expr},{end_month},1),"
                f"'Bills Log'!$B$6:$B${BILLS_LAST_ROW},{prop_ref},"
                f"'Bills Log'!$C$6:$C${BILLS_LAST_ROW},$A{r})"
            )
            cell = ws.cell(row=r, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'

        total_cell = ws.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
        apply_style(total_cell, formula_cell_style())
        total_cell.number_format = '"$"#,##0'
        total_cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True,
                                color=COLOR_PRIMARY)

        avg_cell = ws.cell(row=r, column=15, value=f"=IFERROR(N{r}/12,0)")
        apply_style(avg_cell, formula_cell_style())
        avg_cell.number_format = '"$"#,##0'

    last_svc_row = col_hdr_row + len(SERVICES)

    # Total row
    tot_row = last_svc_row + 1
    tlabel = ws.cell(row=tot_row, column=1, value="TOTAL")
    tlabel.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    tlabel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[tot_row].height = 18

    for col in range(2, 16):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                        value=f"=SUM({col_letter}{col_hdr_row+1}:{col_letter}{last_svc_row})")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=11, bold=True, italic=True,
                          color=COLOR_PRIMARY)

    # Per-property bar chart (Total $ by month)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"$ by Month — Property {prop_index + 1}"
    chart.y_axis.title = "$"
    chart.x_axis.title = "Month"
    chart.height = 6
    chart.width = 18
    data_ref = Reference(ws, min_col=2, min_row=tot_row,
                          max_col=13, max_row=tot_row)
    cat_ref = Reference(ws, min_col=2, min_row=col_hdr_row,
                         max_col=13, max_row=col_hdr_row)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    style_chart(chart)
    # Anchor chart 2 rows below the total row
    anchor = f"A{tot_row + 2}"
    ws.add_chart(chart, anchor)

    # Reserve 12 rows for the chart, then a gap row
    next_row = tot_row + 14
    return next_row


def build_property_matrix(wb):
    """Sheet 2 — Property Matrix (per-property 12-month roll-ups + bar charts)."""
    ws = wb.create_sheet("Property Matrix")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Property Matrix",
                         prev_tab="Bills Log", next_tab="Trend & Anomalies")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Per-property 12-month roll-up — auto-rolled from Bills Log"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column widths: A=service label (24), B-M=months (10), N=total (12), O=avg (10)
    widths = [("A", 24)]
    for c in range(2, 14):
        widths.append((get_column_letter(c), 10))
    widths.append(("N", 12))
    widths.append(("O", 10))
    set_col_widths(ws, widths)

    # Build 3 property sections using cell links into Settings B7:B9
    row = 6
    for i in range(3):
        prop_ref = f"Settings!$B${7 + i}"
        row = _build_property_matrix_section(ws, row, i, prop_ref)

    ws.freeze_panes = "B6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_trend_anomalies(wb):
    """Sheet 3 — Trend & Anomalies (per-property × per-service YoY check)."""
    ws = wb.create_sheet("Trend & Anomalies")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Trend & Anomalies",
                         prev_tab="Property Matrix", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill

    # Anomaly count anchor cell at $B$3 (referenced by Start tab)
    # NOTE: $B$3 sits inside the navy header band (rows 1-3). The header
    # band wrote a fill to row 3, but compact_header_band did not merge or
    # write the value of B3 — so we can safely set it here.
    a3 = ws.cell(row=3, column=1, value="Anomalies:")
    a3.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    a3.alignment = Alignment(horizontal="right", vertical="center")
    a3.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)

    # B3 holds the COUNTIF over the anomaly-flag column (populated below)
    # — we set the formula after we know the flag column range.

    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = (
        "Year-over-year check — anomaly if a month's bill jumped >25% vs prior year"
    )
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 24), ("B", 14), ("C", 8),
        ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 22),
    ])

    # Column headers (row 5)
    headers = [
        "Property", "Service", "Month",
        "Current Yr $", "Prior Yr $", "YoY %", "Δ $", "Anomaly",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Active-year reference (drives YoY)
    # current = Settings!$B$5 ; prior = Settings!$B$5 - 1
    # For each property × service × month combo, build a row.
    # 3 properties × 6 services × 12 months = 216 rows.

    start_row = 6
    row_idx = start_row

    for p_i in range(3):
        prop_ref = f"Settings!$B${7 + p_i}"
        for svc in SERVICES:
            for m in range(1, 13):
                if m < 12:
                    end_month = m + 1
                    end_year_cur = "Settings!$B$5"
                    end_year_pri = "(Settings!$B$5-1)"
                else:
                    end_month = 1
                    end_year_cur = "Settings!$B$5+1"
                    end_year_pri = "Settings!$B$5"

                # A: property (cell link)
                a = ws.cell(row=row_idx, column=1, value=f"={prop_ref}")
                a.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
                a.alignment = Alignment(horizontal="left", vertical="center", indent=1)

                # B: service literal
                b = ws.cell(row=row_idx, column=2, value=svc)
                b.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
                b.alignment = Alignment(horizontal="left", vertical="center", indent=1)

                # C: month label
                c = ws.cell(row=row_idx, column=3, value=MONTHS[m - 1])
                c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
                c.alignment = Alignment(horizontal="center", vertical="center")

                # D: current-year $ for this prop × svc × month
                cur_formula = (
                    f"=SUMIFS('Bills Log'!$F$6:$F${BILLS_LAST_ROW},"
                    f"'Bills Log'!$A$6:$A${BILLS_LAST_ROW},\">=\"&DATE(Settings!$B$5,{m},1),"
                    f"'Bills Log'!$A$6:$A${BILLS_LAST_ROW},\"<\"&DATE({end_year_cur},{end_month},1),"
                    f"'Bills Log'!$B$6:$B${BILLS_LAST_ROW},A{row_idx},"
                    f"'Bills Log'!$C$6:$C${BILLS_LAST_ROW},B{row_idx})"
                )
                d = ws.cell(row=row_idx, column=4, value=cur_formula)
                apply_style(d, formula_cell_style())
                d.number_format = '"$"#,##0'

                # E: prior-year $ for same prop × svc × month
                pri_formula = (
                    f"=SUMIFS('Bills Log'!$F$6:$F${BILLS_LAST_ROW},"
                    f"'Bills Log'!$A$6:$A${BILLS_LAST_ROW},\">=\"&DATE(Settings!$B$5-1,{m},1),"
                    f"'Bills Log'!$A$6:$A${BILLS_LAST_ROW},\"<\"&DATE({end_year_pri},{end_month},1),"
                    f"'Bills Log'!$B$6:$B${BILLS_LAST_ROW},A{row_idx},"
                    f"'Bills Log'!$C$6:$C${BILLS_LAST_ROW},B{row_idx})"
                )
                e = ws.cell(row=row_idx, column=5, value=pri_formula)
                apply_style(e, formula_cell_style())
                e.number_format = '"$"#,##0'

                # F: YoY %
                f = ws.cell(
                    row=row_idx, column=6,
                    value=f'=IFERROR(IF(E{row_idx}=0,"",D{row_idx}/E{row_idx}-1),"")'
                )
                apply_style(f, formula_cell_style())
                f.number_format = "0.0%"

                # G: Δ $
                g = ws.cell(
                    row=row_idx, column=7,
                    value=f"=D{row_idx}-E{row_idx}"
                )
                apply_style(g, formula_cell_style())
                g.number_format = '"$"#,##0;[Red]"-$"#,##0'

                # H: anomaly flag — exact formula from brief
                #   =IF(ABS(current/prior-1)>0.25, "⚠ "&ROUND((current/prior-1)*100,0)&"%", "")
                # Wrapped in IFERROR so prior=0 (no prior-yr data) returns blank.
                h_cell = ws.cell(
                    row=row_idx, column=8,
                    value=(
                        f'=IFERROR(IF(ABS(D{row_idx}/E{row_idx}-1)>0.25,'
                        f'"⚠ "&ROUND((D{row_idx}/E{row_idx}-1)*100,0)&"%",""),"")'
                    )
                )
                h_cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_ERROR)
                h_cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                h_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

                ws.row_dimensions[row_idx].height = 14
                row_idx += 1

    last_data_row = row_idx - 1

    # Now wire the anomaly-count cell at B3
    b3 = ws.cell(row=3, column=2,
                  value=f'=COUNTIF(H{start_row}:H{last_data_row},"⚠*")')
    b3.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    b3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    b3.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    b3.number_format = "0"

    # Conditional formatting: highlight anomaly rows red
    ws.conditional_formatting.add(
        f"H{start_row}:H{last_data_row}",
        CellIsRule(
            operator="notEqual",
            formula=['""'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
        ),
    )

    # Auto-filter so the customer can sort by YoY % to see worst offenders
    ws.auto_filter.ref = f"A5:H{last_data_row}"
    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 4 — Settings (active year, properties, services)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Trend & Anomalies", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · Property list · Service list"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 30), ("B", 28), ("C", 14), ("D", 16),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active tax year (B5) — drives every YoY check ---
    a5 = ws.cell(row=5, column=1, value="Active tax year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=ACTIVE_TAX_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = (
        "Bump this each January. The Trend & Anomalies tab compares this year "
        "to (this year - 1). Without it, dashboards return $0 when the system "
        "clock doesn't match logged dates."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 32

    # --- Property list (B7:B16, 10 slots) ---
    sect_a = ws.cell(row=7, column=1, value="Properties:")
    sect_a.font = bold_right
    sect_a.alignment = Alignment(horizontal="right", vertical="top")

    prop_values = PROPERTIES_DEMO if variant == "demo" else []
    for slot in range(10):
        row = 7 + slot
        cell = ws.cell(row=row, column=2)
        if slot < len(prop_values):
            cell.value = prop_values[slot]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    # --- Service list (B18:B25, 8 slots) ---
    sect_b = ws.cell(row=17, column=1, value="Services:")
    sect_b.font = bold_right
    sect_b.alignment = Alignment(horizontal="right", vertical="top")
    ws.row_dimensions[17].height = 16

    # Default service list: 6 named + 2 spare slots
    default_services = SERVICES + ["Other", ""]
    for slot in range(8):
        row = 18 + slot
        cell = ws.cell(row=row, column=2,
                        value=default_services[slot] if default_services[slot] else None)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    # Helper note
    ws.merge_cells("A27:D28")
    c27 = ws["A27"]
    c27.value = (
        "The Bills Log Property and Service dropdowns read directly from these "
        "two ranges (B7:B16 and B18:B25). Renaming a property here cascades to "
        "the matrix automatically."
    )
    c27.font = italic_muted
    c27.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[27].height = 18
    ws.row_dimensions[28].height = 18


# ---------------------------------------------------------------------------
# Main build function
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_bills_log(wb, variant)
    build_property_matrix(wb)
    build_trend_anomalies(wb)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Utility Usage + Trend Tracker{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Per-property utility tracker with YoY anomaly detection (v2.2)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
