"""Build TAX-006 Home Office Deduction Allocator — wizard tool.

Implements templates/_briefs/TAX-006-home-office-allocator.md.

Generates BOTH:
  templates/_masters/TAX-006-home-office-allocator-DEMO.xlsx
  templates/_masters/TAX-006-home-office-allocator-BLANK.xlsx

Tabs:
  0  Start             — wizard hero + 3-card + QS + Get Started + progress
  1  Eligibility       — regular+exclusive use, principal place of business
  2  Space Allocation  — home sqft, office sqft, business %
  3  Method Election   — Simplified vs Actual side-by-side
  4  Actual Expenses   — direct + indirect + depreciation
  5  STR Routing       — Schedule E vs C, where it lands
  6  Form 8829 Map     — auto-built form mirror
  7  Launch            — readiness + print packet
  8  Settings          — tax year, classification

Usage:
    python build_home_office_allocator.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    COLOR_WHITE,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-006"
NAME = "home-office-allocator"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
VERSION_LINE = f"{SKU} · v2.2 · Free updates forever"

TOTAL_SECTIONS = 5

# Inputs per section (used for progress %)
SECTION_INPUT_COUNTS = {
    "Eligibility":      5,
    "Space Allocation": 4,
    "Method Election":  1,
    "Actual Expenses": 10,
    "STR Routing":      1,
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 21

# DEMO data
SAMPLE = {
    "Eligibility": {
        "regular_use":         "Y",
        "exclusive_use":       "Y",
        "principal_place":     "Y",
        "storage_inventory":   "N",
        "day_care":            "N",
    },
    "Space Allocation": {
        "total_home_sqft":     2400,
        "office_sqft":         110,
        "rooms_total":         9,
        "rooms_business":      1,
    },
    "Method Election": {
        "method":              "Actual",
    },
    "Actual Expenses": {
        # Direct (only-office)
        "direct_repairs":      0,
        # Indirect (whole house, business-use % applied)
        "utilities_electric":  2400,
        "utilities_gas":       980,
        "utilities_water":     420,
        "internet":            720,
        "insurance":           1840,
        "repairs":             620,
        "real_estate_tax":     4800,
        "mortgage_interest":   9800,
        # Depreciation
        "home_basis_building": 312000,  # building only
    },
    "STR Routing": {
        "schedule":            "Schedule E (passive)",
    },
}


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def build_input_tab(wb, tab_name, section_num, title, subtitle, fields,
                     prev_tab, next_tab):
    """Build a flat single-column input tab.

    fields: list of (label, sample_value, kind, options_or_None) tuples.
            kind: 'text' | 'money' | 'sqft' | 'pct' | 'count' | 'yn'

    Returns (ws, last_input_row). Caller owns everything after last_input_row
    (computed rows, callouts, footer nav).
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    section_header_band(ws, section_num, TOTAL_SECTIONS, title, subtitle,
                         prev_tab, next_tab)

    # Label-narrow / input-wide layout
    set_col_widths(ws, [
        ("A", 38),
        ("B", 18), ("C", 8), ("D", 8), ("E", 8), ("F", 8),
        ("G", 8), ("H", 8), ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "Fill the highlighted fields below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: card header
    card_header(ws, 7, ("A", "L"), title)

    # Rows 8+: fields
    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    for i, (label, value, kind, options) in enumerate(fields):
        r = 8 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if kind in ("money", "sqft", "pct", "count"):
            c = ws.cell(row=r, column=2, value=value)
            apply_style(c, input_cell_style())
            if kind == "money":
                c.number_format = '"$"#,##0'
            elif kind == "sqft":
                c.number_format = '0" sqft"'
            elif kind == "pct":
                c.number_format = "0.00%"
            elif kind == "count":
                c.number_format = "0"
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        else:
            ws.merge_cells(f"B{r}:L{r}")
            c = ws[f"B{r}"]
            c.value = value
            apply_style(c, input_cell_style())
            c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
        if options:
            add_dropdown(ws, f"B{r}", options)
        ws.row_dimensions[r].height = 22

    last_input_row = 7 + len(fields)
    return ws, last_input_row


def append_callout(ws, row, text):
    """Append a gold-soft callout strip on the given row (full width)."""
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 36


def append_footer_nav(ws, row, prev_tab, next_tab, default_next="Form 8829 Map"):
    """Append BACK / NEXT secondary buttons on rows row and row+1."""
    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = f"Next: {default_next} →"
        next_target = f"'{default_next}'!A1"
    pseudo_button(ws, f"A{row}", f"F{row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{row}", f"L{row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 22


def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 hero
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Home Office Allocator"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Make your laundry room pay you back."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — single declarative answer (suite Theme 1).
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(IFERROR(\'Launch\'!A10,0)>0,'
        '"✅  Home-office deduction = "&TEXT(\'Launch\'!A10,"$#,##0")'
        '&"  ·  "&\'Launch\'!E10&" method.",'
        '"\U0001F4CA  Fill Eligibility + Space Allocation to see your deduction.")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — 3-card "what you'll get"
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A11:L11")
    c = ws["A11"]; c.value = "What you'll get"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("📐 BUSINESS-USE %", "Your office sqft as a % of total home — drives every indirect expense."),
        ("⚖ METHOD COMPARE", "Simplified ($5/sqft, max $1,500) vs Actual side-by-side, recommend the better one."),
        ("📄 FORM 8829", "Print-ready Form 8829 mirror — Schedule E line 18a or Schedule C line 30."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (ttl, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]; c.value = ttl
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]; c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # ZONE 3 — Quick Start
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A23:L23")
    c = ws["A23"]; c.value = "Quick Start — be done in 10 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24
    quickstart_items = [
        "① Eligibility: regular + exclusive + principal place tests",
        "② Measure: total home sqft, office sqft",
        "③ Method: Simplified or Actual",
        "④ Actual expenses: utilities, insurance, mortgage int., depreciation",
        "⑤ Routing: Schedule E or Schedule C",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + (i if i < 3 else i - 3)
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ZONE 4 — Get Started
    pseudo_button(ws, "A30", "L33",
                   "GET STARTED — CHECK ELIGIBILITY  →",
                   "'Eligibility'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # ZONE 5 — Progress dashboard
    ws.merge_cells("A36:F36")
    c = ws["A36"]; c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ranges = [
        "'Eligibility'!B8:B12",      # 5
        "'Space Allocation'!B8:B11", # 4
        "'Method Election'!B8",      # 1
        "'Actual Expenses'!B8:B17",  # 10
        "'STR Routing'!B8",          # 1
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    ws.merge_cells("G36:L36")
    c = ws["G36"]
    c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS},"0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    section_rows = [
        ("① Eligibility",       "Eligibility",       "B8:B12",  5),
        ("② Space Allocation",  "Space Allocation",  "B8:B11",  4),
        ("③ Method Election",   "Method Election",   "B8",      1),
        ("④ Actual Expenses",   "Actual Expenses",   "B8:B17", 10),
        ("⑤ STR Routing",       "STR Routing",       "B8",      1),
    ]
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]; c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (f'=IF({ca}={total},"✅ Done",'
                   f'IF({ca}=0,"⏳ Empty","⏳ "&{ca}&" of {total}"))')
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    brand_footer(ws, 47, version_line=VERSION_LINE)

    ws.print_area = "A1:L49"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_eligibility_tab(wb, variant):
    s = SAMPLE["Eligibility"] if variant == "demo" else {}
    fields = [
        ("Office space used REGULARLY for STR business?", _val(variant, s.get("regular_use")), "yn", ["Y", "N"]),
        ("Office space used EXCLUSIVELY for business?",   _val(variant, s.get("exclusive_use")), "yn", ["Y", "N"]),
        ("Is this your PRINCIPAL PLACE of STR business?", _val(variant, s.get("principal_place")), "yn", ["Y", "N"]),
        ("Used for storage of inventory/samples (alt qualifier)?", _val(variant, s.get("storage_inventory")), "yn", ["Y", "N"]),
        ("Used for day-care services (alt qualifier)?",    _val(variant, s.get("day_care")), "yn", ["Y", "N"]),
    ]
    ws, last = build_input_tab(
        wb, "Eligibility", 1,
        "Eligibility", "IRC §280A — pass all three core tests, or the deduction is zero.",
        fields, prev_tab="", next_tab="Space Allocation",
    )
    card_body_fill(ws, 8, last, ("A", "L"), border=True)
    append_callout(ws, last + 2, (
        "ⓘ Three-test rule (IRC §280A): the office must be used (1) regularly, "
        "(2) exclusively, and (3) as your principal place of business OR a place "
        "where you meet clients/customers OR a separate structure. STR managers "
        "typically qualify under 'principal place of business' if substantial "
        "admin work happens there and there's no other fixed location."
    ))
    append_footer_nav(ws, last + 4, prev_tab="", next_tab="Space Allocation")


def build_space_allocation_tab(wb, variant):
    s = SAMPLE["Space Allocation"] if variant == "demo" else {}
    fields = [
        ("Total home square footage:",   _val(variant, s.get("total_home_sqft")), "sqft", None),
        ("Office square footage (regular + exclusive):", _val(variant, s.get("office_sqft")), "sqft", None),
        ("Total rooms in home (alt method):",          _val(variant, s.get("rooms_total")), "count", None),
        ("Rooms used for business (alt method):",      _val(variant, s.get("rooms_business")), "count", None),
    ]
    ws, last = build_input_tab(
        wb, "Space Allocation", 2,
        "Space Allocation", "Square footage drives the indirect-expense allocation.",
        fields, prev_tab="Eligibility", next_tab="Method Election",
    )
    # Business-use % at row last+2 (Settings tab refers to it as B13)
    biz_row = last + 2  # 11 + 2 = 13 (Form 8829 Map references row 13)
    a = ws.cell(row=biz_row, column=1, value="Business-use % (computed):")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=biz_row, column=2, value='=IF(B8>0,B9/B8,0)')
    apply_style(c, formula_cell_style())
    c.number_format = "0.00%"
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    ws.row_dimensions[biz_row].height = 24

    card_body_fill(ws, 8, last, ("A", "L"), border=True)
    append_callout(ws, biz_row + 2, (
        "Business-use % above = office sqft ÷ total home sqft. The IRS "
        "also accepts rooms-based allocation if all rooms are roughly equal in size — "
        "less common for STR. When in doubt, use sqft."
    ))
    append_footer_nav(ws, biz_row + 4,
                      prev_tab="Eligibility", next_tab="Method Election")


def build_method_election_tab(wb, variant):
    s = SAMPLE["Method Election"] if variant == "demo" else {}
    fields = [
        ("Method:", _val(variant, s.get("method")), "yn",
         ["Simplified", "Actual"]),
    ]
    ws, last = build_input_tab(
        wb, "Method Election", 3,
        "Method Election", "Simplified ($5/sqft cap $1,500) vs Actual.",
        fields, prev_tab="Space Allocation", next_tab="Actual Expenses",
    )
    card_body_fill(ws, 8, last, ("A", "L"), border=True)

    # Side-by-side comparison table at last+3 (skipping last+1 callout, last+2 spacer)
    set_col_widths(ws, [
        ("A", 38), ("B", 18), ("C", 18), ("D", 18),
        ("E", 6), ("F", 6), ("G", 6), ("H", 6),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    cmp_first = last + 3  # = 11 (1 input row + 2 spacer)
    headers = ["", "Simplified", "Actual", "Difference"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=cmp_first, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[cmp_first].height = 20

    rows_def = [
        ("Deduction this year",
         "=MIN('Space Allocation'!B9,300)*5",
         '=IF(B8="Actual",\'Actual Expenses\'!B22,0)',
         f"=C{cmp_first + 1}-B{cmp_first + 1}"),
        ("Depreciation recapture risk?",
         '"None"', '"Yes (Sec 1250)"', '""'),
        ("Loss carryover allowed?",
         '"No"', '"Yes"', '""'),
        ("Recommended?",
         f'=IF(B{cmp_first + 1}>=C{cmp_first + 1},"✅ Pick this","")',
         f'=IF(C{cmp_first + 1}>B{cmp_first + 1},"✅ Pick this","")',
         '""'),
    ]
    for i, (label, simpl, actual, diff) in enumerate(rows_def):
        r = cmp_first + 1 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        for col, formula in [(2, simpl), (3, actual), (4, diff)]:
            c = ws.cell(row=r, column=col, value=formula)
            apply_style(c, formula_cell_style())
            if i == 0:
                c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 20

    # Highlight the dollar row
    for col in range(2, 5):
        cell = ws.cell(row=cmp_first + 1, column=col)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

    last_used = cmp_first + 4  # 4 comparison rows
    append_callout(ws, last_used + 2, (
        "⚠ Actual-method depreciation locks the home's office-portion into "
        "Section 1250 unrecaptured-gain treatment if you sell — taxed at up to 25% "
        "on the depreciation recapture. Simplified avoids this. If you plan to sell "
        "the home within 5 years, the math often favors Simplified."
    ))
    append_footer_nav(ws, last_used + 4,
                      prev_tab="Space Allocation", next_tab="Actual Expenses")


def build_actual_expenses_tab(wb, variant):
    s = SAMPLE["Actual Expenses"] if variant == "demo" else {}
    fields = [
        ("Direct repairs (only-office, 100% deductible):", _val(variant, s.get("direct_repairs")), "money", None),
        ("Indirect — Electric ($/yr):",      _val(variant, s.get("utilities_electric")), "money", None),
        ("Indirect — Gas ($/yr):",           _val(variant, s.get("utilities_gas")), "money", None),
        ("Indirect — Water ($/yr):",         _val(variant, s.get("utilities_water")), "money", None),
        ("Indirect — Internet ($/yr):",      _val(variant, s.get("internet")), "money", None),
        ("Indirect — Insurance ($/yr):",     _val(variant, s.get("insurance")), "money", None),
        ("Indirect — Repairs (whole house):", _val(variant, s.get("repairs")), "money", None),
        ("Indirect — Real estate tax:",      _val(variant, s.get("real_estate_tax")), "money", None),
        ("Indirect — Mortgage interest:",    _val(variant, s.get("mortgage_interest")), "money", None),
        ("Home depreciable basis (building only):", _val(variant, s.get("home_basis_building")), "money", None),
    ]
    ws, last = build_input_tab(
        wb, "Actual Expenses", 4,
        "Actual Expenses", "Direct + indirect + depreciation. Indirect is allocated by business-use %.",
        fields, prev_tab="Method Election", next_tab="STR Routing",
    )
    card_body_fill(ws, 8, last, ("A", "L"), border=True)

    # Computed totals — Form 8829 Map / Method Election compare reference B22.
    # 10 inputs at rows 8-17, last=17. Computed rows at 19-22.
    biz_pct = "'Space Allocation'!B13"
    rows = [
        (19, "Total indirect (sum):", "=SUM(B9:B16)"),
        (20, "Indirect × business-use % (Form 8829 Line 22):",
         f"=B19*{biz_pct}"),
        (21, "Annual depreciation (basis × biz % ÷ 39):",
         f"=B17*{biz_pct}/39"),
        (22, "Direct + Indirect × biz% + Depreciation:",
         "=B8+B20+B21"),
    ]
    for r, label, formula in rows:
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if r == 22:
            a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        if r == 22:
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        ws.row_dimensions[r].height = 20

    append_callout(ws, 24, (
        "ⓘ Direct expenses (only the office) are 100% deductible. "
        "Indirect expenses (whole house) are deductible at the business-use % from "
        "the Space Allocation tab. Home depreciation = (basis × business-use %) ÷ 39 yrs."
    ))
    append_footer_nav(ws, 26, prev_tab="Method Election", next_tab="STR Routing")


def build_str_routing_tab(wb, variant):
    s = SAMPLE["STR Routing"] if variant == "demo" else {}
    fields = [
        ("Schedule classification:", _val(variant, s.get("schedule")), "yn",
         ["Schedule E (passive)", "Schedule C (active)"]),
    ]
    ws, last = build_input_tab(
        wb, "STR Routing", 5,
        "STR Routing", "Where the deduction lands on your tax return.",
        fields, prev_tab="Actual Expenses", next_tab="Form 8829 Map",
    )
    card_body_fill(ws, 8, last, ("A", "L"), border=True)

    # Routing display at last+2 (= row 10)
    routing_row = last + 2
    a = ws.cell(row=routing_row, column=1, value="Form line for your deduction:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells(f"B{routing_row}:L{routing_row}")
    c = ws.cell(row=routing_row, column=2)
    c.value = (
        '=IF(B8="Schedule E (passive)",'
        '"Schedule E Line 18a (Other) — describe as ""Home office""",'
        'IF(B8="Schedule C (active)",'
        '"Schedule C Line 30 (via Form 8829)","Set classification above"))'
    )
    apply_style(c, formula_cell_style())
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[routing_row].height = 24

    append_callout(ws, routing_row + 2, (
        "Schedule E filers: deduction lands on Line 18a (Other) of Schedule E. "
        "Schedule C filers: deduction lands on Form 8829 → Line 30 of Schedule C. "
        "Both methods (Simplified and Actual) work for both Schedule E and C."
    ))
    append_footer_nav(ws, routing_row + 4,
                      prev_tab="Actual Expenses", next_tab="Form 8829 Map")


def build_form_8829_map_tab(wb, variant):
    """Tab 6 — Form 8829 mirror. Output only."""
    ws = wb.create_sheet("Form 8829 Map")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Form 8829 Map",
                         prev_tab="STR Routing", next_tab="Launch")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = ("Print this tab — mirror of Form 8829 (Expenses for Business "
                 "Use of Your Home).")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6), ("B", 50), ("C", 16),
        ("D", 6), ("E", 6), ("F", 6), ("G", 6),
        ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    # Row 5 column headers
    headers = ["Line", "Description", "Amount"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Form 8829 lines (simplified mapping, not all 36 — focus on the customer-facing ones)
    lines = [
        ("Part I", "BUSINESS-USE %", None, "header"),
        ("1",      "Office area (sqft)",                     "='Space Allocation'!B9", "sqft"),
        ("2",      "Total home area (sqft)",                 "='Space Allocation'!B8", "sqft"),
        ("3",      "Business-use percentage",                "='Space Allocation'!B13", "pct"),
        ("Part II", "EXPENSES",                              None, "header"),
        ("8",      "Direct expenses",                        "='Actual Expenses'!B8", "money"),
        ("9-13",   "Indirect (utilities + ins. + repairs + RE tax + mortgage int.)",
                   "='Actual Expenses'!B19", "money"),
        ("14",     "Indirect × business-use %",              "='Actual Expenses'!B20", "money"),
        ("15",     "Add lines 8 + 14 (operating exp.)",      "=C12+C14", "money"),  # row 15 is line 8 (row 12) + line 14 (row 14)
        ("Part III", "DEPRECIATION",                         None, "header"),
        ("36",     "Annual depreciation (39-yr SL × biz %)", "='Actual Expenses'!B21", "money"),
        ("Part IV", "TOTAL",                                 None, "header"),
        ("",       "TOTAL ACTUAL DEDUCTION",                 "='Actual Expenses'!B22", "total"),
        ("",       "TOTAL SIMPLIFIED DEDUCTION",
                   "=MIN('Space Allocation'!B9,300)*5", "total"),
        ("",       "ELECTED METHOD",
                   '=\'Method Election\'!B8', "text"),
        ("",       "FINAL DEDUCTION",
                   '=IF(\'Method Election\'!B8="Simplified",MIN(\'Space Allocation\'!B9,300)*5,IF(\'Method Election\'!B8="Actual",\'Actual Expenses\'!B22,0))',
                   "total"),
    ]

    r = 6
    for line_num, label, formula, kind in lines:
        if kind == "header":
            ws.merge_cells(f"A{r}:C{r}")
            c = ws.cell(row=r, column=1, value=f"  {line_num} — {label}")
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
            c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 22
        else:
            ws.cell(row=r, column=1, value=line_num).font = Font(
                name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
            )
            ws.cell(row=r, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=r, column=2, value=label).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )
            ws.cell(row=r, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            c = ws.cell(row=r, column=3, value=formula)
            apply_style(c, formula_cell_style())
            if kind == "money" or kind == "total":
                c.number_format = '"$"#,##0'
            elif kind == "sqft":
                c.number_format = '0" sqft"'
            elif kind == "pct":
                c.number_format = "0.00%"
            if kind == "total":
                c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                ws.cell(row=r, column=2).font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            ws.row_dimensions[r].height = 18
        r += 1

    # Print setup
    ws.print_area = f"A1:C{r}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_launch_tab(wb, variant):
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK", "'Form 8829 Map'!A1",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]; c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Your home-office deduction"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Print the Form 8829 Map and route per your filing schedule."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # Dashboard cards
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 14):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 — Final deduction $
    ws.merge_cells("A9:D9")
    c = ws["A9"]; c.value = "FINAL DEDUCTION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]
    c.value = (
        '=IF(\'Method Election\'!B8="Simplified",MIN(\'Space Allocation\'!B9,300)*5,'
        'IF(\'Method Election\'!B8="Actual",\'Actual Expenses\'!B22,0))'
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("A12:D12")
    c = ws["A12"]; c.value = "this tax year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 — Method
    ws.merge_cells("E9:H9")
    c = ws["E9"]; c.value = "METHOD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]
    c.value = '=\'Method Election\'!B8'
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E12:H12")
    c = ws["E12"]; c.value = "elected for this year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 — Schedule routing
    ws.merge_cells("I9:L9")
    c = ws["I9"]; c.value = "SCHEDULE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    c.value = '=\'STR Routing\'!B8'
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("I12:L12")
    c = ws["I12"]; c.value = "where it lands"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # Print button
    pseudo_button(ws, "A16", "L20",
                   "📄  PRINT FORM 8829 MAP  →",
                   "'Form 8829 Map'!A1", variant="primary")
    for r in range(16, 21):
        ws.row_dimensions[r].height = 24

    # Upgrade callout
    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = (
        "💡 Upgrade to the Tax Season Bundle ($147) at thestrledger.com/tax-bundle."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[23].height = 36

    brand_footer(ws, 25, version_line=VERSION_LINE)

    ws.print_area = "A1:L27"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Settings", prev_tab="Launch", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · references"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 36), ("B", 22), ("C", 6),
        ("D", 8), ("E", 8), ("F", 8), ("G", 8), ("H", 8),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a = ws.cell(row=5, column=1, value="Active tax year:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=5, column=2, value=2026)
    apply_style(b, input_cell_style())
    b.number_format = "0"
    ws.row_dimensions[5].height = 18

    # Reference: simplified-method rate
    a = ws.cell(row=6, column=1, value="Simplified-method rate ($/sqft):")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=6, column=2, value=5)
    apply_style(b, input_cell_style())
    b.number_format = '"$"0.00'
    ws.row_dimensions[6].height = 18

    a = ws.cell(row=7, column=1, value="Simplified-method sqft cap:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=7, column=2, value=300)
    apply_style(b, input_cell_style())
    b.number_format = '0" sqft"'
    ws.row_dimensions[7].height = 18

    a = ws.cell(row=8, column=1, value="Home depreciation life (yr):")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=8, column=2, value=39)
    apply_style(b, input_cell_style())
    ws.row_dimensions[8].height = 18

    # Year-end Archive
    ws.row_dimensions[10].height = 8
    a = ws.cell(row=11, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[11].height = 22

    headers = ["Year", "Method", "Deduction ($)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[12].height = 18

    for idx, year in enumerate(range(2024, 2031), start=13):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=idx, column=3).number_format = '"$"#,##0'
        ws.row_dimensions[idx].height = 16

    add_dropdown(ws, "B13:B19", ["Simplified", "Actual"])

    brand_footer(ws, 22, version_line=VERSION_LINE)


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_eligibility_tab(wb, variant)
    build_space_allocation_tab(wb, variant)
    build_method_election_tab(wb, variant)
    build_actual_expenses_tab(wb, variant)
    build_str_routing_tab(wb, variant)
    build_form_8829_map_tab(wb, variant)
    build_launch_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Home Office Deduction Allocator — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "IRS-compliant home-office deduction allocator (Simplified vs Actual, Form 8829)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
