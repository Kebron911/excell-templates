"""Build OPS-005 Supply Inventory + Par-Level Restock Calculator (v2.2 standard).

Operational-mode register: customer logs consumables per property with a
par level (minimum on-hand) and a current count, then flags everything
below par on a single restock list. Eliminates the "ran out of TP between
guests" failure mode.

Per skill operational-mode.md: Settings!$B$5 holds the active year for
informational purposes (no SUMIFS depends on it — restock is real-time
state). Forks register pattern from build_1099_nec_tracker.py.

Generates templates/_masters/OPS-005-supply-inventory-par-level-{DEMO,BLANK}.xlsx.
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    STATE_BAD_FILL,
)

SKU = "OPS-005"
NAME = "supply-inventory-par-level"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# --- Sample data (DEMO) -----------------------------------------------------

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]
VENDORS = ["Costco", "Amazon", "Sam's Club", "Local"]
CATEGORIES = ["Linens", "Bath", "Kitchen", "Cleaning", "Paper", "Decor", "Other"]

# Tuple shape: (SKU, Item, Category, Property, Par, OnHand, Vendor, UnitCost,
#               ReorderQty, LastRestocked, Notes)
INVENTORY = [
    # Smokies Ridge — 12 items, 3 below par
    ("LIN-001", "Queen sheet set (white)",      "Linens",   "Smokies Ridge", 4,  3, "Amazon",     32.00, 2,  "2026-03-15", "100% cotton, 400TC"),
    ("LIN-002", "King sheet set (white)",       "Linens",   "Smokies Ridge", 2,  2, "Amazon",     38.00, 1,  "2026-03-15", ""),
    ("LIN-003", "Bath towel (white, 27x54)",    "Bath",     "Smokies Ridge", 12, 14, "Costco",    6.50,  6,  "2026-04-02", "Hospitality grade"),
    ("LIN-004", "Hand towel (white)",           "Bath",     "Smokies Ridge", 8,  9, "Costco",    3.25,  6,  "2026-04-02", ""),
    ("LIN-005", "Washcloth (white)",            "Bath",     "Smokies Ridge", 12, 8, "Costco",    1.85,  12, "2026-02-10", "Below par"),
    ("PAP-001", "TP 24-pack (Charmin)",         "Paper",    "Smokies Ridge", 2,  1, "Costco",    24.99, 1,  "2026-04-15", "Below par — order now"),
    ("PAP-002", "Paper towel 12-pack",          "Paper",    "Smokies Ridge", 2,  3, "Sam's Club", 19.99, 1,  "2026-04-15", ""),
    ("CLN-001", "Dawn dish soap 3-pack",        "Cleaning", "Smokies Ridge", 1,  2, "Costco",    11.99, 1,  "2026-03-22", ""),
    ("CLN-002", "Lysol wipes 4-pack",           "Cleaning", "Smokies Ridge", 1,  1, "Sam's Club", 14.99, 1,  "2026-04-15", ""),
    ("KIT-001", "K-cups variety 80-pack",       "Kitchen",  "Smokies Ridge", 1,  0, "Costco",    36.99, 1,  "2026-03-01", "Below par — out!"),
    ("KIT-002", "Coffee filter #4 (200ct)",     "Kitchen",  "Smokies Ridge", 1,  2, "Local",      4.99,  1, "2026-02-20", ""),
    ("BTH-001", "Toiletry kit (shampoo+body)",  "Bath",     "Smokies Ridge", 6,  10, "Amazon",   18.00, 6,  "2026-03-15", "Bulk hospitality set"),

    # Creek Side — 12 items, 3 below par
    ("LIN-006", "Queen sheet set (white)",      "Linens",   "Creek Side", 4,  4, "Amazon",     32.00, 2,  "2026-03-30", ""),
    ("LIN-007", "Bath towel (white, 27x54)",    "Bath",     "Creek Side", 10, 7, "Costco",     6.50,  6,  "2026-02-15", "Below par"),
    ("LIN-008", "Hand towel (white)",           "Bath",     "Creek Side", 6,  6, "Costco",     3.25,  6,  "2026-03-30", ""),
    ("LIN-009", "Washcloth (white)",            "Bath",     "Creek Side", 10, 11, "Costco",    1.85,  12, "2026-03-30", ""),
    ("PAP-003", "TP 24-pack (Charmin)",         "Paper",    "Creek Side", 2,  2, "Costco",    24.99,  1, "2026-04-10", ""),
    ("PAP-004", "Paper towel 12-pack",          "Paper",    "Creek Side", 1,  0, "Sam's Club", 19.99, 1, "2026-02-25", "Below par — out!"),
    ("CLN-003", "Dawn dish soap 3-pack",        "Cleaning", "Creek Side", 1,  1, "Costco",    11.99,  1, "2026-04-10", ""),
    ("CLN-004", "Glass cleaner refill",         "Cleaning", "Creek Side", 1,  0, "Local",      5.99,  2, "2026-01-15", "Below par"),
    ("KIT-003", "K-cups variety 80-pack",       "Kitchen",  "Creek Side", 1,  1, "Costco",    36.99,  1, "2026-04-01", ""),
    ("KIT-004", "Dish sponge 6-pack",           "Kitchen",  "Creek Side", 1,  2, "Sam's Club", 7.99,  1, "2026-03-15", ""),
    ("BTH-002", "Toiletry kit (shampoo+body)",  "Bath",     "Creek Side", 4,  6, "Amazon",    18.00,  6, "2026-03-30", ""),
    ("DEC-001", "Welcome candle (cinnamon)",    "Decor",    "Creek Side", 2,  3, "Local",      8.99,  2, "2026-02-10", ""),

    # Lakehouse A — 11 items, 2 below par
    ("LIN-010", "King sheet set (white)",       "Linens",   "Lakehouse A", 3,  3, "Amazon",     38.00, 1,  "2026-03-20", ""),
    ("LIN-011", "Bath towel (white, 27x54)",    "Bath",     "Lakehouse A", 14, 12, "Costco",    6.50,  6,  "2026-03-20", "Below par"),
    ("LIN-012", "Beach towel (oversized)",      "Bath",     "Lakehouse A", 6,  7, "Costco",    11.99, 6,  "2026-04-05", "Lake-only set"),
    ("PAP-005", "TP 24-pack (Charmin)",         "Paper",    "Lakehouse A", 2,  3, "Costco",    24.99, 1,  "2026-04-12", ""),
    ("PAP-006", "Paper towel 12-pack",          "Paper",    "Lakehouse A", 2,  2, "Sam's Club", 19.99, 1, "2026-04-12", ""),
    ("CLN-005", "Dawn dish soap 3-pack",        "Cleaning", "Lakehouse A", 1,  1, "Costco",    11.99, 1,  "2026-04-12", ""),
    ("CLN-006", "Lysol wipes 4-pack",           "Cleaning", "Lakehouse A", 2,  1, "Sam's Club", 14.99, 1, "2026-03-15", "Below par"),
    ("KIT-005", "K-cups variety 80-pack",       "Kitchen",  "Lakehouse A", 1,  1, "Costco",    36.99, 1,  "2026-04-01", ""),
    ("KIT-006", "Dish sponge 6-pack",           "Kitchen",  "Lakehouse A", 1,  1, "Sam's Club", 7.99,  1, "2026-04-12", ""),
    ("BTH-003", "Toiletry kit (shampoo+body)",  "Bath",     "Lakehouse A", 6,  8, "Amazon",    18.00, 6,  "2026-03-20", ""),
    ("DEC-002", "Welcome candle (lakeside)",    "Decor",    "Lakehouse A", 2,  2, "Local",      8.99,  2, "2026-04-01", ""),
]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational hero + below-par alert + activity cards)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Supply Inventory + Par-Level Restock"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Never run out of TP between guests again."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Tracks consumables (linens, toiletries, cleaning supplies, paper goods) "
        "per property with a par level — the minimum quantity you keep on hand. "
        "When on-hand drops below par, the item shows up on the Restock List with "
        "the shortfall and dollar cost. Vendor Quick Order pages bundle every "
        "below-par item into copy-paste shopping lists for Costco, Amazon, "
        "Sam's Club, and local stores."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Settings: edit the property list, vendor list, and category list to match your portfolio.",
        "② Inventory tab: add every consumable — SKU, item, property, par level, current count, vendor, unit cost.",
        "③ After each turnover: update the On-hand column. Shortfall and $ to restock recalculate automatically.",
        "④ Restock List: glance at the count of items below par and the total restock cost — sorted by shortfall.",
        "⑤ Vendor Quick Order: open the page for the vendor you're shopping at, copy the list, paste into the cart.",
        "⑥ When supplies arrive: bump the On-hand count and update Last Restocked date — par alerts clear automatically.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "OPEN INVENTORY" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  OPEN INVENTORY  (LOG ON-HAND COUNTS)",
                   "'Inventory'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Items Tracked
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "ITEMS TRACKED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = "=COUNTA(Inventory!A6:A205)"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "rows on the Inventory tab"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Items Below Par (RED)
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "ITEMS BELOW PAR"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = '=COUNTIF(Inventory!G6:G205,">0")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "need restock now"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Total $ to Restock
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "$ TO RESTOCK"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = "=SUM(Inventory!K6:K205)"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "shortfall × unit cost"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "→ Restock List",
                   "'Restock List'!A1", variant="accent")
    pseudo_button(ws, "D39", "F40", "Vendor Quick Order",
                   "'Vendor Quick Order'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "Inventory",
                   "'Inventory'!A1", variant="secondary")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Operational tip callout (rows 42-43) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "🧺 PAR-LEVEL RULE: par = (linen sets per turnover × turnovers between "
        "restocks) + 1 buffer set. For a 4-bed Queen, 2 turnovers/wk, weekly "
        "Costco run → par = (4 × 2) + 4 = 12 sheet sets. Build it in once, "
        "stop scrambling forever."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Reminder strip (row 44) ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "⚠ Update On-hand counts after every turnover. Stale counts = surprise "
        "shortages. The Restock List is only as honest as the column to its left."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    brand_footer(ws, 46,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_inventory_tab(wb, variant):
    """Sheet 1 — Inventory (master register, 200 rows capacity)."""
    ws = wb.create_sheet("Inventory")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Inventory",
                         prev_tab="Start", next_tab="Restock List")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 14):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:M4")
    c4 = ws["A4"]
    c4.value = "Master register — SKU, item, property, par, on-hand, vendor, unit cost"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 11), ("B", 28), ("C", 12), ("D", 18), ("E", 7),
        ("F", 9),  ("G", 10), ("H", 13), ("I", 11), ("J", 10),
        ("K", 13), ("L", 14), ("M", 28),
    ])

    headers = [
        "SKU", "Item", "Category", "Property",
        "Par", "On-hand", "Shortfall",
        "Vendor", "Unit cost", "Reorder qty",
        "$ to restock", "Last restocked", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    inventory_rows = INVENTORY if variant == "demo" else []
    last_data_row = 5 + len(inventory_rows)

    for row_idx, item in enumerate(inventory_rows, start=6):
        (sku, name, cat, prop, par, on_hand, vendor,
         unit_cost, reorder_qty, last_restocked, notes) = item

        a = ws.cell(row=row_idx, column=1, value=sku)
        apply_style(a, input_cell_style())

        b = ws.cell(row=row_idx, column=2, value=name)
        apply_style(b, input_cell_style())

        c = ws.cell(row=row_idx, column=3, value=cat)
        apply_style(c, input_cell_style())

        d = ws.cell(row=row_idx, column=4, value=prop)
        apply_style(d, input_cell_style())

        e = ws.cell(row=row_idx, column=5, value=par)
        apply_style(e, input_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")

        f = ws.cell(row=row_idx, column=6, value=on_hand)
        apply_style(f, input_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        # Shortfall = MAX(0, par - on_hand) — formula
        g = ws.cell(
            row=row_idx, column=7,
            value=f"=IF(OR(E{row_idx}=\"\",F{row_idx}=\"\"),\"\",MAX(0,E{row_idx}-F{row_idx}))",
        )
        apply_style(g, formula_cell_style())
        g.alignment = Alignment(horizontal="center", vertical="center")

        h = ws.cell(row=row_idx, column=8, value=vendor)
        apply_style(h, input_cell_style())

        i = ws.cell(row=row_idx, column=9, value=unit_cost)
        apply_style(i, input_cell_style())
        i.number_format = '"$"#,##0.00'

        j = ws.cell(row=row_idx, column=10, value=reorder_qty)
        apply_style(j, input_cell_style())
        j.alignment = Alignment(horizontal="center", vertical="center")

        # $ to restock = shortfall × unit cost
        k = ws.cell(
            row=row_idx, column=11,
            value=f"=IF(OR(G{row_idx}=\"\",I{row_idx}=\"\"),\"\",G{row_idx}*I{row_idx})",
        )
        apply_style(k, formula_cell_style())
        k.number_format = '"$"#,##0.00'

        l = ws.cell(row=row_idx, column=12, value=_parse_date(last_restocked))
        apply_style(l, input_cell_style())
        l.number_format = "yyyy-mm-dd"

        m = ws.cell(row=row_idx, column=13, value=notes if notes else None)
        apply_style(m, input_cell_style())

        ws.row_dimensions[row_idx].height = 16

    # Capacity rows (up to row 205 = 200-row capacity)
    for row_idx in range(max(last_data_row + 1, 6), 206):
        for col_idx in range(1, 14):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx in (7, 11):
                # formula columns
                if col_idx == 7:
                    cell.value = (
                        f"=IF(OR(E{row_idx}=\"\",F{row_idx}=\"\"),\"\","
                        f"MAX(0,E{row_idx}-F{row_idx}))"
                    )
                else:
                    cell.value = (
                        f"=IF(OR(G{row_idx}=\"\",I{row_idx}=\"\"),\"\","
                        f"G{row_idx}*I{row_idx})"
                    )
                apply_style(cell, formula_cell_style())
                if col_idx == 7:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.number_format = '"$"#,##0.00'
            else:
                apply_style(cell, input_cell_style())
                if col_idx == 9:
                    cell.number_format = '"$"#,##0.00'
                if col_idx == 12:
                    cell.number_format = "yyyy-mm-dd"
                if col_idx in (5, 6, 10):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 16

    # Dropdowns
    add_dropdown(ws, "C6:C205", f'"{",".join(CATEGORIES)}"')
    add_dropdown(ws, "D6:D205", "=Settings!$B$7:$B$16")
    add_dropdown(ws, "H6:H205", "=Settings!$B$18:$B$27")

    # Conditional formatting: highlight rows where Shortfall > 0 (red tint on G)
    ws.conditional_formatting.add(
        "G6:G205",
        FormulaRule(
            formula=['AND(ISNUMBER(G6),G6>0)'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    # Strong red where on-hand = 0 AND par > 0 (out of stock)
    ws.conditional_formatting.add(
        "F6:F205",
        FormulaRule(
            formula=['AND(ISNUMBER(F6),F6=0,ISNUMBER(E6),E6>0)'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_restock_tab(wb):
    """Sheet 2 — Restock List (filtered: only items where Shortfall > 0)."""
    ws = wb.create_sheet("Restock List")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Restock List · What to Buy Now",
                         prev_tab="Inventory", next_tab="Vendor Quick Order")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 PRINT THIS LIST — items where on-hand < par, sorted by shortfall (largest first)."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 11), ("B", 28), ("C", 18), ("D", 12),
        ("E", 9), ("F", 10), ("G", 11), ("H", 13), ("I", 13),
    ])

    # --- KPI strip (row 5: BIG number "Items below par: N") ---
    ws.merge_cells("A5:I5")
    c5 = ws["A5"]
    c5.value = (
        '="Items below par: "&COUNTIF(Inventory!G6:G205,">0")'
        '&"   ·   $ to restock total: $"&TEXT(SUM(Inventory!K6:K205),"#,##0.00")'
    )
    c5.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_ERROR)
    c5.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c5.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 32

    # --- Filtered table header (row 7) ---
    headers = [
        "SKU", "Item", "Property", "Vendor",
        "Par", "On-hand", "Shortfall", "Unit cost", "$ to restock",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 20

    # --- Filtered rows (8-107, capacity 100) ---
    # INDEX/SMALL pattern: pull rows where Inventory!G > 0, sorted by shortfall desc.
    # Helper: AGGREGATE returns the k-th largest non-error value.
    # We use AGGREGATE(15, 6, ...) which is SMALL ignoring errors. To sort by
    # shortfall descending, we use LARGE on (G * 1000 - row_offset) trick — but
    # Excel-friendly approach: sort by shortfall desc using LARGE on shortfall
    # then break ties by row order.
    #
    # Approach: compute a ranking key = Inventory!G * 1000 + (1000 - row_offset)
    # so larger shortfall ranks higher; tie-break preserves entry order.
    # Then INDEX/MATCH on that key.
    #
    # Simpler & more robust: use AGGREGATE to pull the row index of the n-th
    # largest shortfall (with row-offset tiebreaker), then INDEX everything.
    for out_row in range(8, 108):
        n = out_row - 7  # 1-based rank
        # Row index (within Inventory!A6:A205) of the n-th largest shortfall.
        # Formula: AGGREGATE(14, 6, (Inventory!G6:G205 - ROW(...)*0.0001)
        #                      / (Inventory!G6:G205 > 0), n)
        # Result is the synthetic key; we need the actual row offset.
        # Use a row-aware AGGREGATE LARGE on (G + ROW-based tiebreaker).
        # Simpler approach using array math (Excel 365 not assumed; AGGREGATE
        # supports array forms in legacy Excel).
        rank_key = (
            f'IFERROR(AGGREGATE(14,6,'
            f'(Inventory!$G$6:$G$205+(205-ROW(Inventory!$G$6:$G$205))/10000)'
            f'/(Inventory!$G$6:$G$205>0),{n}),"")'
        )
        # Translate rank_key back to a row offset within G6:G205:
        # We use MATCH on the synthetic key.
        match_expr = (
            f'IFERROR(MATCH({rank_key},'
            f'(Inventory!$G$6:$G$205+(205-ROW(Inventory!$G$6:$G$205))/10000),0),"")'
        )

        # Helper cell? No — we inline via LET-free pattern: re-compute match
        # in each output formula. Compact, no helper column.
        def col_formula(inventory_col):
            return (
                f'=IFERROR(IF({rank_key}="","",'
                f'INDEX(Inventory!${inventory_col}$6:${inventory_col}$205,'
                f'{match_expr})),"")'
            )

        # SKU (Inventory col A)
        a = ws.cell(row=out_row, column=1, value=col_formula("A"))
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")
        # Item (B)
        b = ws.cell(row=out_row, column=2, value=col_formula("B"))
        apply_style(b, formula_cell_style())
        b.alignment = Alignment(horizontal="left", vertical="center")
        # Property (D)
        c = ws.cell(row=out_row, column=3, value=col_formula("D"))
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center")
        # Vendor (H)
        d = ws.cell(row=out_row, column=4, value=col_formula("H"))
        apply_style(d, formula_cell_style())
        d.alignment = Alignment(horizontal="left", vertical="center")
        # Par (E)
        e = ws.cell(row=out_row, column=5, value=col_formula("E"))
        apply_style(e, formula_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")
        # On-hand (F)
        f = ws.cell(row=out_row, column=6, value=col_formula("F"))
        apply_style(f, formula_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")
        # Shortfall (G)
        g = ws.cell(row=out_row, column=7, value=col_formula("G"))
        apply_style(g, formula_cell_style())
        g.alignment = Alignment(horizontal="center", vertical="center")
        # Unit cost (I)
        i = ws.cell(row=out_row, column=8, value=col_formula("I"))
        apply_style(i, formula_cell_style())
        i.number_format = '"$"#,##0.00'
        # $ to restock (K)
        k = ws.cell(row=out_row, column=9, value=col_formula("K"))
        apply_style(k, formula_cell_style())
        k.number_format = '"$"#,##0.00'

        ws.row_dimensions[out_row].height = 16

    # Conditional formatting: red row fill where Shortfall ≥ Par (zero on hand)
    ws.conditional_formatting.add(
        "A8:I107",
        FormulaRule(
            formula=['AND(ISNUMBER($G8),ISNUMBER($E8),$G8>=$E8)'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
        ),
    )

    ws.freeze_panes = "A8"

    ws.print_area = "A1:I60"
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_vendor_quick_order_tab(wb):
    """Sheet 3 — Vendor Quick Order (4 sections, one per vendor)."""
    ws = wb.create_sheet("Vendor Quick Order")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Vendor Quick Order",
                         prev_tab="Restock List", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 8):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:G4")
    c4 = ws["A4"]
    c4.value = "Copy-paste shopping lists per vendor — only items below par."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 11), ("B", 32), ("C", 18), ("D", 11),
        ("E", 11), ("F", 12), ("G", 13),
    ])

    # 4 vendor sections — each gets its own header strip and a 20-row table
    vendors = ["Costco", "Amazon", "Sam's Club", "Local"]
    section_start = 6
    section_height = 25  # 1 title + 1 header + 20 data + 3 spacer

    for v_idx, vendor in enumerate(vendors):
        base_row = section_start + (v_idx * section_height)
        title_row = base_row
        header_row = base_row + 2
        data_start = base_row + 3
        data_end = base_row + 22  # 20 data rows

        # Vendor section title strip (gold-soft)
        ws.merge_cells(f"A{title_row}:G{title_row}")
        title_cell = ws[f"A{title_row}"]
        title_cell.value = (
            f'="{vendor.upper()}  ·  "&COUNTIFS(Inventory!H6:H205,"{vendor}",'
            f'Inventory!G6:G205,">0")&" item(s) below par"'
        )
        title_cell.font = Font(name=FONT_MONO, size=11, bold=True,
                                color=COLOR_PRIMARY)
        title_cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                          indent=2)
        ws.row_dimensions[title_row].height = 22

        # Spacer
        ws.row_dimensions[title_row + 1].height = 6

        # Table header
        headers = ["SKU", "Item", "Property", "Par", "On-hand",
                   "Reorder qty", "Subtotal $"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            apply_style(cell, header_row_style())
        ws.row_dimensions[header_row].height = 18

        # Data rows: filter to vendor + shortfall>0, ranked by shortfall desc
        for out_row in range(data_start, data_end + 1):
            n = out_row - data_start + 1  # 1-based rank within this vendor

            # Filter: vendor matches AND shortfall>0
            # rank_key = G + (205-ROW)/10000 only where vendor=X and G>0
            rank_key = (
                f'IFERROR(AGGREGATE(14,6,'
                f'(Inventory!$G$6:$G$205+(205-ROW(Inventory!$G$6:$G$205))/10000)'
                f'/((Inventory!$G$6:$G$205>0)*(Inventory!$H$6:$H$205="{vendor}")),'
                f'{n}),"")'
            )
            match_expr = (
                f'IFERROR(MATCH({rank_key},'
                f'(Inventory!$G$6:$G$205+(205-ROW(Inventory!$G$6:$G$205))/10000),0),"")'
            )

            def col_formula(inventory_col, rk=rank_key, mx=match_expr):
                return (
                    f'=IFERROR(IF({rk}="","",'
                    f'INDEX(Inventory!${inventory_col}$6:${inventory_col}$205,'
                    f'{mx})),"")'
                )

            # SKU
            cell = ws.cell(row=out_row, column=1, value=col_formula("A"))
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Item
            cell = ws.cell(row=out_row, column=2, value=col_formula("B"))
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Property
            cell = ws.cell(row=out_row, column=3, value=col_formula("D"))
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Par
            cell = ws.cell(row=out_row, column=4, value=col_formula("E"))
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # On-hand
            cell = ws.cell(row=out_row, column=5, value=col_formula("F"))
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Reorder qty
            cell = ws.cell(row=out_row, column=6, value=col_formula("J"))
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Subtotal $ (= reorder qty × unit cost)
            subtotal_formula = (
                f'=IFERROR(IF({rank_key}="","",'
                f'INDEX(Inventory!$J$6:$J$205,{match_expr})*'
                f'INDEX(Inventory!$I$6:$I$205,{match_expr})),"")'
            )
            cell = ws.cell(row=out_row, column=7, value=subtotal_formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0.00'

            ws.row_dimensions[out_row].height = 14

        # Vendor total row
        total_row = data_end + 1
        ws.merge_cells(f"A{total_row}:F{total_row}")
        tlabel = ws[f"A{total_row}"]
        tlabel.value = f"{vendor} order total:"
        tlabel.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        tlabel.alignment = Alignment(horizontal="right", vertical="center",
                                      indent=1)
        tlabel.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

        ttotal = ws.cell(
            row=total_row, column=7,
            value=(
                f'=SUMPRODUCT((Inventory!H6:H205="{vendor}")*'
                f'(Inventory!G6:G205>0)*Inventory!J6:J205*Inventory!I6:I205)'
            ),
        )
        ttotal.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ttotal.alignment = Alignment(horizontal="right", vertical="center")
        ttotal.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        ttotal.number_format = '"$"#,##0.00'
        ws.row_dimensions[total_row].height = 20

    last_row = section_start + (len(vendors) * section_height)
    ws.print_area = f"A1:G{last_row}"
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 4 — Settings (active year + property/vendor/category lists)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Vendor Quick Order", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active year · property list · vendor list · category list"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 24), ("C", 4), ("D", 32), ("E", 24)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active year (row 5) ---
    a5 = ws.cell(row=5, column=1, value="Active year (informational):")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")

    b5 = ws.cell(row=5, column=2, value=2026)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    b5.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 18

    ws.merge_cells("A6:E6")
    c6 = ws["A6"]
    c6.value = (
        "Inventory state is real-time (not year-locked) — this cell is "
        "informational and used in archive labels."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 24

    # --- Property list (rows 7-16) ---
    sect7 = ws.cell(row=7, column=1, value="PROPERTIES")
    sect7.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[7].height = 22

    demo_props = PROPERTIES if variant == "demo" else []
    for idx in range(10):
        row = 7 + idx + 1  # 8..17 — but spec says B7-B16 so use 7..16
        # Actually spec calls out B7-B16 (10 cells). Place property list there.

    # Per brief: B7-B16 Property list (10).
    # Treat row 7 as a section title visually but place data at rows 7..16.
    # Reconcile: keep title at row 7 styling, but data starts at row 8.
    # Update: rewrite cleanly — put title at row 7, leave data rows 8-17,
    # then update Inventory dropdown ref to $B$8:$B$17.
    # For brief compliance (B7-B16), instead use D column for the title row.
    # Cleaner: title goes into A7, data into B7:B16. So no merged title.

    # Reset row 7 (title we just wrote) — keep but move data to B7:B16.
    sect7.value = ""
    ws.row_dimensions[7].height = 18

    a7 = ws.cell(row=7, column=1, value="Property list (used in Inventory dropdown):")
    a7.font = bold_right
    a7.alignment = Alignment(horizontal="right", vertical="center")

    for idx in range(10):
        row = 7 + idx
        cell = ws.cell(row=row, column=2)
        if variant == "demo" and idx < len(demo_props):
            cell.value = demo_props[idx]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    # --- Vendor list (rows 18-27) ---
    a18 = ws.cell(row=18, column=1, value="Vendor list (used in Inventory dropdown):")
    a18.font = bold_right
    a18.alignment = Alignment(horizontal="right", vertical="center")

    demo_vendors = VENDORS if variant == "demo" else []
    for idx in range(10):
        row = 18 + idx
        cell = ws.cell(row=row, column=2)
        if variant == "demo" and idx < len(demo_vendors):
            cell.value = demo_vendors[idx]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[28].height = 8

    # --- Category list (rows 29-35; informational) ---
    a29 = ws.cell(row=29, column=1, value="Category list (locked — used in Inventory dropdown):")
    a29.font = bold_right
    a29.alignment = Alignment(horizontal="right", vertical="center")

    for idx, cat in enumerate(CATEGORIES):
        row = 29 + idx
        cell = ws.cell(row=row, column=2, value=cat)
        apply_style(cell, formula_cell_style())  # locked-ish appearance
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16

    # --- Year-end Archive (rows 38-46) ---
    sect38 = ws.cell(row=38, column=1, value="Year-end Archive")
    sect38.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[38].height = 22

    ws.merge_cells("A39:E39")
    c39 = ws["A39"]
    c39.value = (
        "Each January, copy YTD totals into the row for the closing year. "
        "Inventory state itself is real-time and rolls forward."
    )
    c39.font = italic_muted
    c39.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[39].height = 24

    archive_headers = ["Year", "Items Tracked", "$ Spent on Restocks", "Notes"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=40, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[40].height = 18

    for idx, year in enumerate(range(2024, 2031), start=41):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0"
        cell = ws.cell(row=idx, column=3)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        cell = ws.cell(row=idx, column=4)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16


# Update Inventory dropdown reference helper — the tab references
# Settings!$B$7:$B$16 and $B$18:$B$27 directly, which matches what we wrote.

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_inventory_tab(wb, variant)
    build_restock_tab(wb)
    build_vendor_quick_order_tab(wb)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Supply Inventory + Par-Level Restock{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Per-property consumables register with par-level restock alerts (v2.2)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
