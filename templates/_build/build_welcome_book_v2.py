"""Build GST-001 Airbnb Welcome Book v2 — wizard-style tool.

Ships two files in the same Etsy download:
- GST-001-welcome-book-DEMO.xlsx (all 80 inputs filled, sample property)
- GST-001-welcome-book-BLANK.xlsx (all inputs cleared)

Implements the design at:
  docs/superpowers/specs/2026-04-23-welcome-book-v2-tool-redesign.md

12 tabs:
  0  Start            — hero + 3-card + quickstart + Get Started + progress
  1  Property         — identity inputs (8 fields)
  2  Arrival          — address, entry, parking (7 fields + 4 static bullets)
  3  WiFi + Tech      — network, streaming, smart devices (8 fields)
  4  House Rules      — policies with 4 dropdowns (7 fields)
  5  Local Guide      — 20-row table x 4 input cols (80 fields)
  6  Trash            — pickup + maintenance (7 fields)
  7  Departure        — checkout checklist (6 fields + formula)
  8  Emergency        — 911 block + contacts (9 fields + 1 hardcoded + 1 formula)
  9  Review & Print   — readiness dashboard + pseudo-button + 3-page live preview
  10 Bonus            — pre-written Airbnb listing copy (200 words)
  11 × Host Notes     — private host-only, quarantined with red warning

Usage:
    python build_welcome_book_v2.py
    # generates BOTH DEMO and BLANK files
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    # Constants
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_NAVY_TINT, COLOR_NAVY_SHADE, COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    # Existing helpers
    input_cell_style, formula_cell_style, header_row_style,
    set_col_widths, apply_style,
    # v2 helpers (Task 1)
    pseudo_button, card_header, card_body_fill, section_header_band,
)

BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / "GST-001-welcome-book-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / "GST-001-welcome-book-BLANK.xlsx"

# --- Tab names (used for navigation + tabColor) ---

TAB_NAMES = [
    "Start",              # 0
    "Property",           # 1
    "Arrival",            # 2
    "WiFi + Tech",        # 3
    "House Rules",        # 4
    "Local Guide",        # 5
    "Trash",              # 6
    "Departure",          # 7
    "Emergency",          # 8
    "Review & Print",     # 9
    "Bonus",              # 10
    "× Host Notes",       # 11
]

# Input counts per section (for progress formula)
SECTION_INPUT_COUNTS = {
    "Property": 8,
    "Arrival": 7,
    "WiFi + Tech": 8,
    "House Rules": 7,
    "Local Guide": 80,   # 20 rows × 4 cols
    "Trash": 7,
    "Departure": 6,
    "Emergency": 9,
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 132

# --- Sample data (DEMO variant) ---

SAMPLE = {
    "Property": {
        "name": "Smokies Ridge Cabin",
        "host": "Daniel",
        "host_phone": "+1 (555) 555-0199",
        "check_in": "2026-05-10",
        "check_out": "2026-05-15",
        "property_type": "Cabin",
        "max_guests": 6,
        "address": "123 Mountain Lane, Gatlinburg, TN 37738",
    },
    "Arrival": {
        "address": "123 Mountain Lane, Gatlinburg, TN 37738",
        "entry_method": "Smart lock",
        "door_code": "4321",
        "parking": "Gravel drive on the right — there's room for 2 cars. Do not block the mailbox.",
        "route": "If GPS sends you up the fire road, ignore it. Stay on Ridge Way past the red barn.",
        "arrival_window": "After 3 PM",
        "early_option": "Coffee shops in the Local Guide tab are the best bet; the lockbox won't open before 3 PM.",
    },
    "WiFi": {
        "ssid": "SmokiesRidge_Guest",
        "password": "welcome2024",
        "backup_ssid": "",
        "tv_streaming": "Netflix: already signed in. Hulu: guest@smokiesridge.com / stay2024",
        "smart_lock_note": "Same as entry — 4321",
        "thermostat": "Nest on the hallway wall. Please keep between 65-78°F. Auto-schedule resumes after you leave.",
        "tv_controls": "The black Roku remote is the one to use. HDMI1 is Fire TV.",
        "wifi_support": "Host first (see Emergency tab). If no answer — Spectrum: 1-833-267-6094",
    },
    "Rules": {
        "quiet_hours": "10 PM – 7 AM",
        "max_guests": 6,
        "smoking": "No smoking",
        "pets": "No pets",
        "events": "No events",
        "shoes": "Remove at door",
        "custom_rules": "• Hot tub closes at 10 PM\n• Please don't move furniture\n• If you break something, tell us — we'd rather fix it than discover it later",
    },
    # Local Guide samples — category + first-of-each-row values (blank rows marked "")
    "Local": [
        ("Coffee", "Mountain Grind", "0.8 mi", "(865) 555-0100", "Best espresso in town; small pastry case fills fast"),
        ("Coffee", "", "", "", ""),
        ("Restaurant", "The Cast Iron", "2.1 mi", "(865) 555-0118", "Sunday brunch is chef's-kiss — reserve ahead"),
        ("Restaurant", "Ridge BBQ", "1.4 mi", "(865) 555-0122", "Brisket sells out by 7 PM Fri/Sat"),
        ("Restaurant", "", "", "", ""),
        ("Grocery", "Food City", "2.8 mi", "(865) 555-0133", "Open until 10 PM"),
        ("Grocery", "", "", "", ""),
        ("Takeout", "Smoky Pizza Co", "1.9 mi", "(865) 555-0144", "Delivers to the cabin"),
        ("Takeout", "", "", "", ""),
        ("Pharmacy", "Walgreens", "3.5 mi", "(865) 555-0155", "24-hr location"),
        ("Gas station", "Shell on Hwy 321", "0.5 mi", "", "Cheapest in the valley"),
        ("Hospital/Urgent care", "LeConte Medical Center", "8.1 mi", "(865) 446-7000", "ER 24/7"),
        ("Coffee alt", "", "", "", ""),
        ("Outdoor/Hike", "Cades Cove Loop", "14 mi", "", "11-mile scenic loop; allow 3 hrs"),
        ("Outdoor/Hike", "", "", "", ""),
        ("Kid-friendly", "Dollywood", "9 mi", "(865) 428-9488", "Worth a full day"),
        ("Kid-friendly", "", "", "", ""),
        ("Date night", "The Wine Cellar", "6.2 mi", "(865) 555-0166", "Prix fixe Thu/Fri/Sat"),
        ("Bar/Nightlife", "Mill & Main", "2.4 mi", "(865) 555-0177", "Live music Fri/Sat after 9"),
        ("Emergency (non-911)", "Sevier County Sheriff", "", "(865) 436-5181", "Non-emergency line"),
    ],
    "Trash": {
        "pickup_day": "Thursday",
        "bin_location": "Around the right side of the house, next to the shed",
        "recycling_accepted": "Cardboard, plastic #1-2, aluminum. No glass.",
        "sorting_rules": "Green = recycling, black = trash. When in doubt — trash.",
        "pickup_location": "To the curb by 7 AM Thursday. Bring back Friday.",
        "thermostat_range": "65-78°F — auto-schedule handles the rest",
        "power_outage": "Breaker panel is in the laundry room. Sevier Electric: (865) 453-2887. Text host if it's longer than an hour.",
    },
    "Departure": {
        "checkout_time": "11:00 AM",
        "linen_location": "hallway laundry basket",
        "trash_spot": "curb (Thursday) or dumpster on-site",
        "thermostat_setting": "72°F",
        "key_return": "in lockbox, reset to 0000",
        "custom_tasks": "• Throw any leftover food\n• Text the host when on the road — we'll release your deposit faster",
    },
    "Emergency": {
        "hospital_name": "LeConte Medical Center",
        "hospital_phone": "(865) 446-7000",
        "hospital_address": "742 Middle Creek Rd, Sevierville, TN 37862",
        "urgent_care_name": "FastMed Urgent Care",
        "urgent_care_phone": "(865) 428-1020",
        "police_non_emergency": "(865) 436-5181",
        "vet": "Mountain Vet ER — (865) 329-1905",
        "utility": "Sevier Electric: (865) 453-2887",
    },
    "Host": {
        "cleaner_name": "Sarah @ Smokies Clean",
        "cleaner_phone": "(865) 555-0145",
        "handyman": "Bob — (865) 555-0198",
        "plumber": "Ridge Plumbing — (865) 555-0177",
        "pool_service": "HotTub Haven — quarterly service",
        "insurance": "ABC-1234567 (Proper Insurance)",
        "wifi_admin": "admin / router-pw-here",
        "smart_lock_master": "9999 — not for guest use",
        "safe_codes": "Under-sink safe: 1234\nGarage keypad: 5678",
        "private_notes": "Hot tub chemistry is touchy — do not tell guests to refill. If hot tub is cloudy, call HotTub Haven.",
    },
}


def add_dropdown(ws, cell_range, options):
    """Add an in-cell dropdown with the given list of options."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _sample_or_blank(variant, value):
    """Return value if variant='demo', else empty string."""
    return value if variant == "demo" else ""


from dataclasses import dataclass, field


@dataclass
class Card:
    """One input card inside an input tab."""
    header: str                                   # e.g., "ENTRY & PARKING"
    # Each row: (label, value_or_empty, options). If options is a list,
    # a dropdown is attached. value can be a Python str/int/date or a
    # formula string starting with "=".
    rows: list = field(default_factory=list)
    # Static content rows (not inputs) — rendered as plain text full-width
    static: list = field(default_factory=list)
    # Height per input row (default 24; use 48 for wrap-text multiline)
    row_height: int = 24


def build_input_tab(wb, section_num, tab_name, title, subtitle, cards,
                    input_count, prev_tab, next_tab):
    """Render an input tab using the card-grouped layout.

    Args:
        wb: workbook
        section_num: 1-based section index (used in "SECTION N OF 8")
        tab_name: name of this tab
        title: big title shown in the step header band
        subtitle: italic subtitle shown below title
        cards: list of Card objects
        input_count: for print_area height calculation
        prev_tab: name of tab to go back to (or "" for Start)
        next_tab: name of tab to go next to (or "" for Review & Print)
    """
    ws = wb[tab_name]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Row 1-5: step header band
    section_header_band(ws, section_num, 8, title, subtitle, prev_tab, next_tab)

    # Row 6: blank spacer (parchment)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.row_dimensions[6].height = 6
    for c in range(1, 13):
        ws.cell(row=6, column=c).fill = parchment_fill

    # Row 7: instruction strip
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = ("Fill the highlighted fields below. The 'Review & Print' tab "
               "(last) shows what your guest will see.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    # Freeze panes: rows 1-7 stay visible
    ws.freeze_panes = "A8"

    # Render cards starting at row 9
    current_row = 9
    for card in cards:
        # Card header (1 row)
        card_header(ws, current_row, ("A", "L"), card.header)
        header_row = current_row
        current_row += 1

        # Input rows
        body_start = current_row
        for row_spec in card.rows:
            if len(row_spec) == 2:
                label, value = row_spec
                options = None
            elif len(row_spec) == 3:
                label, value, options = row_spec
            else:
                raise ValueError(f"Card row must be 2 or 3 tuple, got {row_spec}")

            # Label (cols A-C, merged)
            ws.merge_cells(f"A{current_row}:C{current_row}")
            lc = ws[f"A{current_row}"]
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center",
                                      indent=1)
            # Input (cols D-L, merged)
            ws.merge_cells(f"D{current_row}:L{current_row}")
            ic = ws[f"D{current_row}"]
            if value != "" and value is not None:
                if isinstance(value, str) and value.startswith("="):
                    ic.value = value
                    apply_style(ic, formula_cell_style())
                else:
                    ic.value = value
                    apply_style(ic, input_cell_style())
                    ic.alignment = Alignment(horizontal="left",
                                              vertical="center",
                                              wrap_text=True, indent=1)
            else:
                apply_style(ic, input_cell_style())

            # Dropdown
            if options:
                add_dropdown(ws, f"D{current_row}", options)

            ws.row_dimensions[current_row].height = card.row_height
            current_row += 1

        # Static content rows (after inputs)
        for static_text in card.static:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            sc = ws[f"A{current_row}"]
            sc.value = static_text
            sc.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
            sc.alignment = Alignment(horizontal="left", vertical="center",
                                      wrap_text=True, indent=2)
            sc.fill = parchment_fill
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        body_end = current_row - 1
        # Apply card body border + fill
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)

        # 1-row spacer between cards (parchment)
        for c in range(1, 13):
            ws.cell(row=current_row, column=c).fill = parchment_fill
        ws.row_dimensions[current_row].height = 12
        current_row += 1

    # Section footer
    footer_row = current_row + 1
    # Gold thin-rule divider
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=current_row, column=c).border = Border(top=gold_side)
        ws.cell(row=current_row, column=c).fill = parchment_fill

    # Back/Next pseudo-buttons (secondary variant)
    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = "Next: Review & Print →"
        next_target = "'Review & Print'!A1"

    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22


# --- Per-tab builders (stubs; filled in Tasks 3-8) ---

def build_start_tab(wb, variant):
    """Tab 0 — Start.

    Zones:
      1. Hero band (rows 1-8, navy)
      2. What you'll build — 3-card grid (rows 10-20)
      3. Quick Start card (rows 22-28, parchment-alt)
      4. Get Started full-width button (rows 30-33)
      5. Progress dashboard (rows 35-46)
      6. Footer (rows 48-52)
    """
    ws = wb.active
    ws.title = TAB_NAMES[0]
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    # Columns A:L = 8 units each
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: HERO BAND (rows 1-8, navy) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Row 2: small brand name top-left
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"{BRAND_NAME}"
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # Row 4: big title
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Welcome Book"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    # Row 5: subtitle (campaign tagline for guest-experience SKUs)
    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Welcome books that earn 5-stars."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 7: tiny SKU tag
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "GST-001 · v2.0"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: WHAT YOU'LL BUILD (rows 10-20, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Row 11: section heading
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "What you'll build in the next 15 minutes"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # Rows 13-18: 3 cards, cols A-D / E-H / I-L
    cards = [
        ("📖 PRINT", "Spiral-bound binder for the kitchen counter."),
        ("📱 QR CODE", "Scan on arrival — view on the guest's phone."),
        ("✉ EMAIL PDF", "Send ahead of check-in. One page, every tab."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (title, desc) in enumerate(cards):
        first, last = col_groups[idx]
        # Card header (row 13)
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]
        c.value = title
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        # Card body (rows 14-18)
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]
        c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # --- ZONE 3: QUICK START CARD (rows 22-28) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = "Quick Start — be done in 5 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24

    quickstart_items = [
        "① Property name + host phone",
        "② WiFi network + password",
        "③ Trash pickup day",
        "④ Checkout time",
        "⑤ Hospital + 911 note",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + i if i < 3 else 24 + (i - 3)  # 2 cols of items
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # --- ZONE 4: GET STARTED BUTTON (rows 30-33) ---
    pseudo_button(ws, "A30", "L33",
                   "GET STARTED — FILL YOUR PROPERTY INFO  →",
                   "'Property'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: PROGRESS DASHBOARD (rows 35-46) ---
    ws.merge_cells("A36:F36")
    c = ws["A36"]
    c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Overall completion % formula
    ws.merge_cells("G36:L36")
    c = ws["G36"]
    # Build COUNTA sum over all 8 input-tab ranges
    ranges = [
        "'Property'!B5:B12",    # 8 fields
        "'Arrival'!B5:B11",     # 7 fields
        "'WiFi + Tech'!B5:B12", # 8 fields
        "'House Rules'!B5:B11", # 7 fields
        "'Local Guide'!B6:E25", # 80 cells (20 rows × 4 cols)
        "'Trash'!B5:B11",       # 7 fields
        "'Departure'!B5:B10",   # 6 fields
        "'Emergency'!B5:B13",   # 9 fields
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    c.value = f"=TEXT(({counta_sum})/{TOTAL_INPUTS}, \"0%\") & \" complete\""
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Per-section status rows 38-45
    section_rows = [
        ("① Property Info",   "Property",    "B5:B12",   8),
        ("② Arrival",          "Arrival",     "B5:B11",   7),
        ("③ WiFi + Tech",      "WiFi + Tech", "B5:B12",   8),
        ("④ House Rules",      "House Rules", "B5:B11",   7),
        ("⑤ Local Guide",      "Local Guide", "B6:E25",  80),
        ("⑥ Trash",            "Trash",       "B5:B11",   7),
        ("⑦ Departure",        "Departure",   "B5:B10",   6),
        ("⑧ Emergency",        "Emergency",   "B5:B13",   9),
    ]
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        # Label (cols A-F)
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        # Status (cols G-J) — formula
        ws.merge_cells(f"G{r}:J{r}")
        c = ws[f"G{r}"]
        c.value = (
            f'=IF(COUNTA(\'{tab}\'!{range_})={total},"✅ Done",'
            f'IF(COUNTA(\'{tab}\'!{range_})=0,"⏳ Empty",'
            f'"⏳ "&COUNTA(\'{tab}\'!{range_})&" of {total}"))'
        )
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        # Arrow hyperlink (cols K-L)
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # --- ZONE 6: FOOTER (rows 48-52) ---
    # Row 49: gold thin-rule divider
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=49, column=c).border = Border(top=gold_side)

    # Row 50: contact
    ws.merge_cells("A50:L50")
    c = ws["A50"]
    c.value = f"Questions? {BRAND_EMAIL}"
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 51: version + updates promise
    ws.merge_cells("A51:L51")
    c = ws["A51"]
    c.value = "Free updates forever · v2.0 · Released 2026-05"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Print area: Start tab can be printed as cover sheet
    ws.print_area = "A1:L52"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_property_tab(wb, variant):
    """Tab 1 — Property. Identity + stay dates."""
    wb.create_sheet(TAB_NAMES[1])
    s = SAMPLE["Property"] if variant == "demo" else {}
    cards = [
        Card(
            header="Identity",
            rows=[
                ("Property name:", s.get("name", "")),
                ("Host first name:", s.get("host", "")),
            ],
        ),
        Card(
            header="Guest Details",
            rows=[
                ("Host phone (text preferred):", s.get("host_phone", "")),
                ("Check-in date:", s.get("check_in", "")),
                ("Check-out date:", s.get("check_out", "")),
                ("Property type:", s.get("property_type", ""),
                 ["Single-family", "Cabin", "Condo", "Apartment",
                  "Multi-family", "Glamping", "Other"]),
                ("Max guests:", s.get("max_guests", "")),
            ],
        ),
        Card(
            header="Setup",
            rows=[
                ("Full address:", s.get("address", "")),
            ],
            static=[
                "First time? Go to Start \u2192 Quick Start for the 5-minute path.",
            ],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=1, tab_name="Property",
        title="Property Info",
        subtitle="Who you are and where they're staying.",
        cards=cards, input_count=8,
        prev_tab="", next_tab="Arrival",
    )


def build_arrival_tab(wb, variant):
    """Tab 2 — Arrival."""
    wb.create_sheet(TAB_NAMES[2])
    s = SAMPLE["Arrival"] if variant == "demo" else {}
    cards = [
        Card(
            header="Entry & Parking",
            rows=[
                ("Full address:", s.get("address", "")),
                ("Entry method:", s.get("entry_method", ""),
                 ["Smart lock", "Key lockbox", "In-person", "Other"]),
                ("Door/lock code:", s.get("door_code", "")),
                ("Parking instructions:", s.get("parking", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Route & Timing",
            rows=[
                ("Best route (if tricky):", s.get("route", "")),
                ("Arrival time window:", s.get("arrival_window", "")),
                ("If you arrive early:", s.get("early_option", "")),
            ],
            row_height=36,
        ),
        Card(
            header="First 5 Minutes Inside",
            rows=[],
            static=[
                "1. Crank the thermostat to your comfort \u2014 WiFi tab has controls.",
                "2. Connect to WiFi \u2014 network + password on the WiFi tab.",
                "3. Check the fridge for welcome items.",
                "4. If anything's broken, text the host immediately.",
            ],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=2, tab_name="Arrival",
        title="Arrival & Check-in",
        subtitle="Everything your guest needs on day one.",
        cards=cards, input_count=7,
        prev_tab="Property", next_tab="WiFi + Tech",
    )


def build_wifi_tab(wb, variant):
    """Tab 3 — WiFi + Tech."""
    wb.create_sheet(TAB_NAMES[3])
    s = SAMPLE["WiFi"] if variant == "demo" else {}
    cards = [
        Card(
            header="Network",
            rows=[
                ("WiFi network name:", s.get("ssid", "")),
                ("WiFi password:", s.get("password", "")),
                ("Backup network (if any):", s.get("backup_ssid", "")),
            ],
        ),
        Card(
            header="Entertainment",
            rows=[
                ("TV streaming \u2014 service + login:", s.get("tv_streaming", "")),
                ("How to adjust TV volume/inputs:", s.get("tv_controls", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Smart Devices",
            rows=[
                ("Smart lock code (if different):", s.get("smart_lock_note", "")),
                ("Smart thermostat notes:", s.get("thermostat", "")),
                ("Who to call if WiFi fails:", s.get("wifi_support", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=3, tab_name="WiFi + Tech",
        title="WiFi & Technology",
        subtitle="So nobody has to text you about the wifi password.",
        cards=cards, input_count=8,
        prev_tab="Arrival", next_tab="House Rules",
    )


def build_rules_tab(wb, variant):
    """Tab 4 — House Rules."""
    wb.create_sheet(TAB_NAMES[4])
    s = SAMPLE["Rules"] if variant == "demo" else {}
    cards = [
        Card(
            header="The Basics",
            rows=[
                ("Quiet hours:", s.get("quiet_hours", "")),
                ("Maximum guests:", s.get("max_guests", "")),
            ],
        ),
        Card(
            header="Policies",
            rows=[
                ("Smoking:", s.get("smoking", ""),
                 ["No smoking", "No smoking inside", "Smoking OK in designated area"]),
                ("Pets:", s.get("pets", ""),
                 ["No pets", "Pets OK with deposit", "Pets OK no deposit"]),
                ("Events/parties:", s.get("events", ""),
                 ["No events", "Small gatherings OK with notice", "OK"]),
                ("Shoes inside:", s.get("shoes", ""),
                 ["Remove at door", "OK", "Preferred removed"]),
            ],
        ),
        Card(
            header="Custom Rules",
            rows=[
                ("Your rules (one per line):", s.get("custom_rules", "")),
            ],
            row_height=90,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=4, tab_name="House Rules",
        title="House Rules",
        subtitle="Short, clear, and why each one exists.",
        cards=cards, input_count=7,
        prev_tab="WiFi + Tech", next_tab="Local Guide",
    )


def build_local_tab(wb, variant):
    """Tab 5 — Local Guide. 20 category rows x 4 input cols = 80 inputs.

    Uses a custom layout (not the generic card template) because Local
    Guide is a wide table, not a label/input pair structure.
    """
    wb.create_sheet(TAB_NAMES[5])
    ws = wb["Local Guide"]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [
        ("A", 22), ("B", 28), ("C", 10), ("D", 16), ("E", 45),
    ] + [(get_column_letter(c), 2) for c in range(6, 13)])

    # Step header band
    section_header_band(ws, 5, 8, "Local Guide",
                         "What we'd tell a friend visiting.",
                         "House Rules", "Trash")

    # Row 6 spacer + Row 7 instruction
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in (6, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "Fill the rows you want to share. Skip any you don't."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    # Row 8: blank
    # Row 9: column headers
    headers = ["Category", "Name", "Distance", "Phone", "Why we love it"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=9, column=i, value=h)
        apply_style(c, header_row_style())

    # Rows 10-29: 20 category rows (hardcoded category column, input B-E)
    local_data = SAMPLE["Local"] if variant == "demo" else [
        (cat, "", "", "", "") for cat, *_ in SAMPLE["Local"]
    ]
    for i, (cat, name, dist, phone, notes) in enumerate(local_data):
        r = 10 + i
        ws.cell(row=r, column=1, value=cat).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1)
        for col, val in enumerate([name, dist, phone, notes], start=2):
            c = ws.cell(row=r, column=col, value=val)
            apply_style(c, input_cell_style())
            c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 28

    # Footer
    footer_row = 31
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=footer_row, column=c).border = Border(top=gold_side)
        ws.cell(row=footer_row, column=c).fill = parchment
    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   "\u2190 Back: House Rules", "'House Rules'!A5",
                   variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   "Next: Trash \u2192", "'Trash'!A5", variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22

    # Print area: landscape, 2 pages
    ws.print_area = f"A1:E{footer_row + 2}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_trash_tab(wb, variant):
    """Tab 6 — Trash + Maintenance."""
    wb.create_sheet(TAB_NAMES[6])
    s = SAMPLE["Trash"] if variant == "demo" else {}
    cards = [
        Card(
            header="Pickup & Bins",
            rows=[
                ("Trash pickup day:", s.get("pickup_day", ""),
                 ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                  "Saturday", "Sunday", "No pickup \u2014 dumpster on-site"]),
                ("Bin location:", s.get("bin_location", "")),
                ("Recycling accepted:", s.get("recycling_accepted", "")),
                ("What goes in which bin:", s.get("sorting_rules", "")),
            ],
            row_height=30,
        ),
        Card(
            header="Operational",
            rows=[
                ("Where to put bins on pickup morning:", s.get("pickup_location", "")),
                ("HVAC/thermostat range to leave:", s.get("thermostat_range", "")),
                ("If the power goes out:", s.get("power_outage", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=6, tab_name="Trash",
        title="Trash, Recycling & Maintenance",
        subtitle="The stuff nobody likes to ask about.",
        cards=cards, input_count=7,
        prev_tab="Local Guide", next_tab="Departure",
    )


def build_departure_tab(wb, variant):
    """Tab 7 — Departure."""
    wb.create_sheet(TAB_NAMES[7])
    s = SAMPLE["Departure"] if variant == "demo" else {}
    cards = [
        Card(
            header="When",
            rows=[
                ("Checkout time:", s.get("checkout_time", "")),
                ("Checkout day:",
                 '=IF(Property!B8<>"", TEXT(Property!B8,"dddd, mmmm d"), "See Property tab")'),
            ],
        ),
        Card(
            header="Before You Leave Checklist",
            rows=[
                ("\u2610 Strip bed linens and leave in:", s.get("linen_location", "")),
                ("\u2610 Take trash + recycling to:", s.get("trash_spot", "")),
                ("\u2610 Turn thermostat to:", s.get("thermostat_setting", "")),
                ("\u2610 Leave key:", s.get("key_return", "")),
            ],
            static=[
                "\u2610 Run the dishwasher",
                "\u2610 Lock all doors + windows",
            ],
        ),
        Card(
            header="Custom Checkout Tasks",
            rows=[
                ("Your tasks (one per line):", s.get("custom_tasks", "")),
            ],
            row_height=60,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=7, tab_name="Departure",
        title="Checkout",
        subtitle="What to do before you drive away.",
        cards=cards, input_count=6,
        prev_tab="Trash", next_tab="Emergency",
    )


def build_emergency_tab(wb, variant):
    """Tab 8 — Emergency. 911 block + contacts."""
    wb.create_sheet(TAB_NAMES[8])
    s = SAMPLE["Emergency"] if variant == "demo" else {}
    ws = wb["Emergency"]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Header band
    section_header_band(ws, 8, 8, "Emergency Contacts",
                         "Keep this one nearby.",
                         "Departure", "")

    # Row 6 spacer
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=6, column=c).fill = parchment
    ws.row_dimensions[6].height = 6

    # Row 7: big red 911 block (merged A:L, red tint fill, Georgia 22pt bold red)
    ws.merge_cells("A7:L8")
    c = ws["A7"]
    c.value = "IN AN EMERGENCY \u2014 CALL 911"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor="FFE8E8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        top=Side(style="medium", color=COLOR_ERROR),
        bottom=Side(style="medium", color=COLOR_ERROR),
        left=Side(style="medium", color=COLOR_ERROR),
        right=Side(style="medium", color=COLOR_ERROR),
    )
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 30

    # Freeze panes (rows 1-8 stay visible)
    ws.freeze_panes = "A9"

    # Row 9: instruction strip
    ws.merge_cells("A9:L9")
    c = ws["A9"]
    c.value = "911 first. Then the contacts below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 22

    # Cards starting row 11
    current_row = 11
    for hdr, rows in [
        ("Hospital", [
            ("Nearest hospital:", s.get("hospital_name", "")),
            ("Hospital phone:", s.get("hospital_phone", "")),
            ("Hospital address:", s.get("hospital_address", "")),
        ]),
        ("Urgent Care & Police", [
            ("Urgent care:", s.get("urgent_care_name", "")),
            ("Urgent care phone:", s.get("urgent_care_phone", "")),
            ("Non-emergency police:", s.get("police_non_emergency", "")),
        ]),
        ("Other", [
            ("Poison control:", "1-800-222-1222"),  # hardcoded
            ("24-hr vet (if pets):", s.get("vet", "")),
            ("Utility outage reporting:", s.get("utility", "")),
            ("Host phone (call/text):", "=Property!B7"),  # formula
        ]),
    ]:
        card_header(ws, current_row, ("A", "L"), hdr)
        current_row += 1
        body_start = current_row
        for label, value in rows:
            ws.merge_cells(f"A{current_row}:C{current_row}")
            lc = ws[f"A{current_row}"]
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            ws.merge_cells(f"D{current_row}:L{current_row}")
            ic = ws[f"D{current_row}"]
            if isinstance(value, str) and value.startswith("="):
                ic.value = value
                apply_style(ic, formula_cell_style())
            elif value == "1-800-222-1222":
                # hardcoded — not an input
                ic.value = value
                ic.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
                ic.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ic.fill = parchment
            else:
                ic.value = value
                apply_style(ic, input_cell_style())
                ic.alignment = Alignment(horizontal="left", vertical="center",
                                          wrap_text=True, indent=1)
            ws.row_dimensions[current_row].height = 24
            current_row += 1
        body_end = current_row - 1
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)
        # Spacer
        for c in range(1, 13):
            ws.cell(row=current_row, column=c).fill = parchment
        ws.row_dimensions[current_row].height = 12
        current_row += 1

    # Footer
    footer_row = current_row + 1
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=footer_row, column=c).border = Border(top=gold_side)
        ws.cell(row=footer_row, column=c).fill = parchment
    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   "\u2190 Back: Departure", "'Departure'!A5", variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   "Review & Print \u2192", "'Review & Print'!A1",
                   variant="accent")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22

    ws.print_area = f"A1:L{footer_row + 2}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_review_print_tab(wb, variant):
    """Tab 9 — Review & Print. The payoff tab with live 3-page preview."""
    ws = wb.create_sheet(TAB_NAMES[9])
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Review header (rows 1-6, navy) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Row 2: back + label
    pseudo_button(ws, "A2", "C2", "← BACK", "'Emergency'!A5", variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]
    c.value = "REVIEW & PRINT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28

    # Row 4: title
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Your welcome book is ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42

    # Row 5: subtitle
    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Preview below. Hit Print when happy."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # --- ZONE 2: Readiness dashboard (rows 8-14, parchment) ---
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 15):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # 3 cards: A-D (Completion), E-H (Red Flags), I-L (Status)
    # Card 1: Completion %
    ws.merge_cells("A9:D9")
    c = ws["A9"]
    c.value = "COMPLETION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]
    ranges = [
        "'Property'!B5:B12", "'Arrival'!B5:B11", "'WiFi + Tech'!B5:B12",
        "'House Rules'!B5:B11", "'Local Guide'!B10:E29", "'Trash'!B5:B11",
        "'Departure'!B5:B10", "'Emergency'!B5:B13",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    c.value = f"=TEXT(({counta_sum})/{TOTAL_INPUTS}, \"0%\")"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A12:D12")
    c = ws["A12"]
    c.value = f"of {TOTAL_INPUTS} fields filled"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2: Red flags (required fields missing)
    # Define required fields: Property!B5, B6, B7 + WiFi!B5, B6 + Trash!B5 +
    # Departure!B5 + Emergency!B7, B8, B10. 10 required fields.
    required = [
        "'Property'!B5", "'Property'!B6", "'Property'!B7",
        "'WiFi + Tech'!B5", "'WiFi + Tech'!B6",
        "'Trash'!B5",
        "'Departure'!B5",
        "'Emergency'!B7", "'Emergency'!B8", "'Emergency'!B10",
    ]
    countblank_req = " + ".join(f'IF({r}="",1,0)' for r in required)
    ws.merge_cells("E9:H9")
    c = ws["E9"]
    c.value = "RED FLAGS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]
    c.value = f"={countblank_req}"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E12:H12")
    c = ws["E12"]
    c.value = "required fields empty"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3: Status (READY / MINOR / NEEDS WORK)
    ws.merge_cells("I9:L9")
    c = ws["I9"]
    c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    c.value = f'=IF(({countblank_req})=0,"READY",IF(({countblank_req})<=2,"MINOR","NEEDS WORK"))'
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I12:L12")
    c = ws["I12"]
    c.value = "0 = green · 1-2 = yellow · 3+ = red"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders (all 3 cards)
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc, lc_ = column_index_from_string(first), column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                new_border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )
                ws.cell(row=r, column=col).border = new_border

    # --- ZONE 3: Generate PDF pseudo-button (rows 16-21) ---
    pseudo_button(ws, "A16", "L21",
                   "📄  GENERATE YOUR PDF\nPress Ctrl+P → choose 'Save as PDF'",
                   "'Review & Print'!A24",  # no-op scroll anchor
                   variant="primary")
    for r in range(16, 22):
        ws.row_dimensions[r].height = 28

    # Row 22: fine print
    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = "Print area is pre-set to letter portrait. × Host Notes tab excluded."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 4: Live preview (rows 24-80, 3 pages) ---
    # PAGE 1 — rows 24-45: hero + property + arrival + WiFi
    # Hero band
    for r in (24, 25):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A24:L25")
    c = ws["A24"]
    c.value = '=CONCATENATE("Welcome to ", Property!B5)'
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[24].height = 30
    ws.row_dimensions[25].height = 30

    # Property info block (rows 27-35)
    def _preview_row(row, label, source_formula, bold=False, size=11):
        ws.merge_cells(f"A{row}:D{row}")
        lc = ws[f"A{row}"]
        lc.value = label
        lc.font = Font(name=FONT_BODY, size=size, bold=True, color=COLOR_PRIMARY)
        lc.alignment = Alignment(horizontal="right", indent=1)
        ws.merge_cells(f"E{row}:L{row}")
        vc = ws[f"E{row}"]
        # Wrap source cell to render "(not set)" if empty
        vc.value = f'=IF({source_formula}="","(not set)",{source_formula})'
        vc.font = Font(name=FONT_BODY, size=size,
                       bold=bold, color=COLOR_TEXT)
        vc.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)

    _preview_row(27, "Host:",            "Property!B6")
    _preview_row(28, "Host phone:",      "Property!B7")
    _preview_row(29, "Check-in:",        "Property!B8")
    _preview_row(30, "Check-out:",       "Property!B9")
    _preview_row(32, "Address:",         "Arrival!B5")
    _preview_row(33, "Entry method:",    "Arrival!B6")
    _preview_row(34, "Door/lock code:",  "Arrival!B7")
    _preview_row(35, "Parking:",         "Arrival!B8", size=10)

    # WiFi — rows 37-40 with BIG text for credentials
    _preview_row(37, "",                 '""', size=12)  # section header spacer
    ws.merge_cells("A38:L38")
    c = ws["A38"]
    c.value = "WiFi"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    _preview_row(39, "Network:", "'WiFi + Tech'!B5", bold=True, size=16)
    _preview_row(40, "Password:", "'WiFi + Tech'!B6", bold=True, size=16)

    # Page break at row 46
    ws.row_breaks.append(Break(id=46))

    # PAGE 2 — rows 47-65: rules + local guide top 10
    ws.merge_cells("A48:L48")
    c = ws["A48"]
    c.value = "House Rules"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[48].height = 28
    _preview_row(49, "Quiet hours:",      "'House Rules'!B5")
    _preview_row(50, "Max guests:",       "'House Rules'!B6")
    _preview_row(51, "Smoking:",          "'House Rules'!B7")
    _preview_row(52, "Pets:",             "'House Rules'!B8")
    _preview_row(53, "Events:",           "'House Rules'!B9")
    _preview_row(54, "Shoes:",            "'House Rules'!B10")

    # Local Guide top 10 — rows 56-65
    ws.merge_cells("A56:L56")
    c = ws["A56"]
    c.value = "Local Guide — Our Top 10"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[56].height = 28
    for i in range(10):
        r = 57 + i
        src_row = 10 + i
        # Category
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].value = f"='Local Guide'!A{src_row}"
        ws[f"A{r}"].font = Font(name=FONT_BODY, size=10, bold=True,
                                  color=COLOR_PRIMARY)
        # Name
        ws.merge_cells(f"C{r}:F{r}")
        ws[f"C{r}"].value = f"=IF('Local Guide'!B{src_row}=\"\",\"\",'Local Guide'!B{src_row})"
        ws[f"C{r}"].font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        # Distance + phone
        ws.merge_cells(f"G{r}:H{r}")
        ws[f"G{r}"].value = f"=IF('Local Guide'!C{src_row}=\"\",\"\",'Local Guide'!C{src_row})"
        ws[f"G{r}"].font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
        # Notes
        ws.merge_cells(f"I{r}:L{r}")
        ws[f"I{r}"].value = f"=IF('Local Guide'!E{src_row}=\"\",\"\",'Local Guide'!E{src_row})"
        ws[f"I{r}"].font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)

    # Page break at row 66
    ws.row_breaks.append(Break(id=66))

    # PAGE 3 — rows 67-80: trash + departure + emergency
    ws.merge_cells("A68:L68")
    c = ws["A68"]
    c.value = "Trash & Maintenance"
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    _preview_row(69, "Pickup day:",       "'Trash'!B5")
    _preview_row(70, "Bin location:",     "'Trash'!B6")

    ws.merge_cells("A72:L72")
    c = ws["A72"]
    c.value = "Checkout"
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    _preview_row(73, "Time:",             "Departure!B5")
    _preview_row(74, "Linens:",           "Departure!B7")
    _preview_row(75, "Key return:",       "Departure!B10")

    # Emergency — red-bordered block rows 77-80
    ws.merge_cells("A77:L77")
    c = ws["A77"]
    c.value = "Emergency — Call 911 First"
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor="FFE8E8")
    c.alignment = Alignment(horizontal="center")
    _preview_row(78, "Hospital:",         "Emergency!B7")
    _preview_row(79, "Hospital phone:",   "Emergency!B8")
    _preview_row(80, "Host phone:",       "Property!B7", bold=True, size=14)

    # Print area: A24:L80 (3 pages)
    ws.print_area = "A24:L80"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 3
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_bonus_tab(wb, variant):
    """Tab 10 — Bonus: Listing Copy. Implemented in Task 7."""
    ws = wb.create_sheet(TAB_NAMES[10])
    ws.sheet_properties.tabColor = COLOR_GOLD_SOFT


def build_host_notes_tab(wb, variant):
    """Tab 11 — × Host Notes. Implemented in Task 8."""
    ws = wb.create_sheet(TAB_NAMES[11])
    ws.sheet_properties.tabColor = COLOR_SECONDARY


def build_workbook(out_path, variant):
    """Build a workbook with all 12 tabs.

    Args:
        out_path: pathlib.Path where the .xlsx will be saved
        variant: "demo" (filled with SAMPLE) or "blank" (all inputs empty)
    """
    assert variant in ("demo", "blank"), f"Unknown variant: {variant}"

    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_tab(wb, variant)
    build_arrival_tab(wb, variant)
    build_wifi_tab(wb, variant)
    build_rules_tab(wb, variant)
    build_local_tab(wb, variant)
    build_trash_tab(wb, variant)
    build_departure_tab(wb, variant)
    build_emergency_tab(wb, variant)
    build_review_print_tab(wb, variant)
    build_bonus_tab(wb, variant)
    build_host_notes_tab(wb, variant)

    suffix = " (DEMO)" if variant == "demo" else ""
    wb.properties.title = f"Airbnb Welcome Book{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Guest-facing welcome book tool for Airbnb/VRBO hosts — wizard UI, "
        "card-grouped inputs, live 3-page PDF preview."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, variant="demo")
    build_workbook(BLANK_OUT, variant="blank")


if __name__ == "__main__":
    main()
