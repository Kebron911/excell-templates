"""Build TAX-013 Depreciation Tracker — operational register.

Implements templates/_briefs/TAX-013-depreciation-tracker.md.

Master register of every depreciable asset across the portfolio, with
MACRS class assignment, year-by-year depreciation schedule (Year 1 →
Year 27) and current-year deduction line ready to drop into
Schedule C / E. Connects to TAX-009 (Section 179) and TAX-010 (cost-seg).

Generates BOTH:
  templates/_masters/TAX-013-depreciation-tracker-DEMO.xlsx
  templates/_masters/TAX-013-depreciation-tracker-BLANK.xlsx

Tabs (6):
  0  Start                — KPIs + Add Asset CTA
  1  Asset Register       — master register (capacity 100 assets)
  2  Schedule by Year     — 27-year matrix (assets × years)
  3  Current Year Summary — per-asset Y_active deduction + Schedule C/E line map
  4  Form 4562 Worksheet  — Part I + II + III + Part V (listed property)
  5  Settings             — active year, properties, MACRS class table

Usage:
    python build_depreciation_tracker.py
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    COLOR_WHITE,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-013"
NAME = "depreciation-tracker"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
VERSION_LINE = f"{SKU} · v2.3 · Free updates forever"

ACTIVE_YEAR_DEFAULT = 2026
ASSET_CAPACITY = 100
SCHEDULE_YEARS = 27  # 27 columns covers 27.5-yr residential rental tail
HEADER_ROW = 7       # asset register header row
DATA_START_ROW = 8   # first asset data row

# ----------------------------------------------------------------------------
# IRS MACRS half-year-convention percentage tables (rounded to IRS Pub 946 Tbl A-1/A-2/A-3)
# Format: per-class list of Year-1..Year-N percentages.  Sum ≈ 100%.
# ----------------------------------------------------------------------------
MACRS_TABLES = {
    # 5-year, 200% DB, half-year (IRS Pub 946 Table A-1)
    "5":   [0.2000, 0.3200, 0.1920, 0.1152, 0.1152, 0.0576] + [0.0000] * (SCHEDULE_YEARS - 6),
    # 7-year, 200% DB, half-year (Table A-1)
    "7":   [0.1429, 0.2449, 0.1749, 0.1249, 0.0893, 0.0892, 0.0893, 0.0446] + [0.0000] * (SCHEDULE_YEARS - 8),
    # 15-year, 150% DB, half-year (Table A-1)
    "15":  [0.0500, 0.0950, 0.0855, 0.0770, 0.0693, 0.0623, 0.0590, 0.0590,
            0.0591, 0.0590, 0.0591, 0.0590, 0.0591, 0.0590, 0.0591, 0.0295] + [0.0000] * (SCHEDULE_YEARS - 16),
    # 27.5-year residential rental, SL, mid-month (mid-year approximation:
    # Y1 ≈ 3.485%, full years ≈ 3.636%, final year ≈ 1.97%).
    "27.5": ([0.03485] + [0.03636] * 26 + [0.01970])[:SCHEDULE_YEARS],
    # 39-year nonresidential, SL, mid-month (approximation:
    # Y1 ≈ 2.461%, full years ≈ 2.564%).  Showing first 27 of 40 years.
    "39":  ([0.02461] + [0.02564] * (SCHEDULE_YEARS - 1)),
    # Land — non-depreciable
    "Land": [0.0000] * SCHEDULE_YEARS,
}

MACRS_CLASS_OPTIONS = ["5", "7", "15", "27.5", "39", "Land"]

# ----------------------------------------------------------------------------
# DEMO sample (8 assets — see brief)
# ----------------------------------------------------------------------------
SAMPLE_PROPERTIES = [
    "Smokies Ridge Cabin",
    "Creek Side Lodge",
    "Lakehouse A",
]

SAMPLE_ASSETS = [
    # (name, property, placed_in_service, cost, class, convention, method,
    #  s179, bonus, salvage, disposition, biz_use_pct, vehicle?, notes)
    ("Smokies Ridge — building",        "Smokies Ridge Cabin", date(2024, 6, 15), 250_000, "27.5", "MM", "SL",         0,      0, 0, None, 1.00, "N", "Cost basis = purchase − land allocation"),
    ("Creek Side Lodge — building",     "Creek Side Lodge",    date(2024, 6, 15), 250_000, "27.5", "MM", "SL",         0,      0, 0, None, 1.00, "N", "Cost basis = purchase − land allocation"),
    ("Lakehouse A — building",          "Lakehouse A",         date(2024, 6, 15), 250_000, "27.5", "MM", "SL",         0,      0, 0, None, 1.00, "N", "Cost basis = purchase − land allocation"),
    ("Whole-home furniture set",        "Smokies Ridge Cabin", date(2024, 6, 20),  18_000, "7",    "HY", "200% DB",    0,  7_200, 0, None, 1.00, "N", "Bonus 40% taken Y1"),
    ("Stainless appliance package",     "Creek Side Lodge",    date(2024, 7,  5),   2_400, "5",    "HY", "200% DB",    0,    960, 0, None, 1.00, "N", "Bonus 40% taken Y1"),
    ("Smart TV + mounts package",       "Lakehouse A",         date(2024, 8, 12),   1_600, "5",    "HY", "200% DB",    0,    640, 0, None, 1.00, "N", "Bonus 40% taken Y1"),
    ("Ford F-150 (heavy truck >6,000lb)", "Smokies Ridge Cabin", date(2024, 1, 15), 42_000, "5",   "HY", "200% DB", 30_500,  4_600, 0, None, 0.78, "Y", "Listed property — Form 4562 Part V"),
    ("Driveway extension + parking",    "Smokies Ridge Cabin", date(2024, 9,  1),  12_000, "15",   "HY", "150% DB",    0,  4_800, 0, None, 1.00, "N", "Land improvement"),
]


# ----------------------------------------------------------------------------
# Local helpers
# ----------------------------------------------------------------------------

def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def append_callout(ws, row, text):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 36


# ============================================================================
# Tab 0 — Start
# ============================================================================

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- Hero (rows 1-8, navy) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Depreciation Tracker"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Every depreciable asset, every year, one ledger."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Verdict row 6
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(IFERROR(\'Current Year Summary\'!F'
        f'{DATA_START_ROW + ASSET_CAPACITY + 1},0)>0,'
        '"✅  Current-year depreciation = "&'
        f'TEXT(\'Current Year Summary\'!F{DATA_START_ROW + ASSET_CAPACITY + 1},"$#,##0")'
        '&"  ·  Assets tracked = "&'
        f'COUNTA(\'Asset Register\'!A{DATA_START_ROW}:A{DATA_START_ROW + ASSET_CAPACITY - 1}),'
        '"\U0001F4CA  Add an asset on the Asset Register to see your deduction.")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- KPI cards (rows 10-18) ---
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 19):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    ws.merge_cells("A11:L11")
    c = ws["A11"]; c.value = "Portfolio at a glance"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    kpis = [
        ("ASSETS TRACKED",
         f'=COUNTA(\'Asset Register\'!A{DATA_START_ROW}:A{DATA_START_ROW + ASSET_CAPACITY - 1})',
         "0", COLOR_PRIMARY),
        ("CURRENT-YEAR DEPRECIATION",
         f"='Current Year Summary'!F{DATA_START_ROW + ASSET_CAPACITY + 1}",
         '"$"#,##0', COLOR_ACCENT),
        ("ACCUMULATED THROUGH PRIOR YEAR",
         f"='Current Year Summary'!G{DATA_START_ROW + ASSET_CAPACITY + 1}",
         '"$"#,##0', COLOR_PRIMARY),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (label, formula, fmt, color) in enumerate(kpis):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]; c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}17")
        c = ws[f"{first}14"]; c.value = formula
        c.font = Font(name=FONT_HEAD, size=24, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        ws.merge_cells(f"{first}18:{last}18")
        ws[f"{first}18"].border = Border(bottom=Side(style="thin", color=COLOR_ACCENT))
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # Year-N drop-off callout
    ws.merge_cells("A20:L20")
    c = ws["A20"]; c.value = "Quick Start"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[20].height = 24

    qs_items = [
        "① Settings: confirm your active tax year and property list",
        "② Asset Register: enter every depreciable asset (cost, class, date)",
        "③ Schedule by Year: see the full life-of-asset depreciation arc",
        "④ Current Year Summary: drop the totals onto Schedule C line 13 / Schedule E line 18",
        "⑤ Form 4562 Worksheet: print + hand to your CPA",
    ]
    for i, item in enumerate(qs_items):
        row = 22 + i
        ws.merge_cells(f"B{row}:L{row}")
        c = ws[f"B{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    # Primary CTA
    pseudo_button(ws, "A29", "L32",
                   "ADD AN ASSET — OPEN REGISTER  →",
                   "'Asset Register'!A5", variant="primary")
    for r in range(29, 33):
        ws.row_dimensions[r].height = 22

    # Secondary buttons
    pseudo_button(ws, "A34", "F35",
                   "📅  Schedule by Year",
                   "'Schedule by Year'!A1", variant="secondary")
    pseudo_button(ws, "G34", "L35",
                   "📄  Form 4562 Worksheet",
                   "'Form 4562 Worksheet'!A1", variant="secondary")
    ws.row_dimensions[34].height = 20
    ws.row_dimensions[35].height = 20

    append_callout(ws, 37, (
        "ⓘ  Connects with TAX-009 (Section 179 Planner) and TAX-010 (Cost "
        "Segregation DIY).  §179 + Bonus amounts entered on the Asset "
        "Register reduce adjusted basis before the year-by-year MACRS "
        "schedule runs."
    ))

    brand_footer(ws, 40, version_line=VERSION_LINE)

    ws.print_area = "A1:L42"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ============================================================================
# Tab 1 — Asset Register
# ============================================================================

def build_asset_register_tab(wb, variant):
    ws = wb.create_sheet("Asset Register")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 15)])
    compact_header_band(ws, "Asset Register",
                         prev_tab="Start", next_tab="Schedule by Year")

    # Subtitle row 4
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A4:N4")
    c = ws["A4"]
    c.value = (
        f"Master register — capacity {ASSET_CAPACITY} assets.  "
        f"Buildings = 27.5 (residential) / 39 (commercial), furniture = 7, "
        f"appliances/tech/vehicles = 5, land improvements = 15.  Land is "
        f"non-depreciable."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[4].height = 32

    # Column widths optimized for the 14-col asset register
    set_col_widths(ws, [
        ("A", 32),  # Asset name
        ("B", 18),  # Property
        ("C", 13),  # Date placed in service
        ("D", 12),  # Cost basis
        ("E", 7),   # MACRS class
        ("F", 6),   # Convention
        ("G", 11),  # Method
        ("H", 10),  # §179 taken
        ("I", 10),  # Bonus
        ("J", 12),  # Adjusted basis (formula)
        ("K", 8),   # Recovery period (formula)
        ("L", 9),   # Salvage
        ("M", 11),  # Disposition date
        ("N", 24),  # Notes
    ])

    # Header row
    headers = [
        "Asset name", "Property", "Placed in service",
        "Cost basis", "MACRS yr", "Conv.", "Method",
        "§179 taken", "Bonus", "Adj. basis", "Recovery",
        "Salvage", "Disposition", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[HEADER_ROW].height = 22

    # Row 6 mini-instruction band
    ws.merge_cells("A6:N6")
    c = ws["A6"]
    c.value = "Yellow cells = your input.  Gray cells = formulas (do not edit)."
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = parchment
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[6].height = 16

    # Data rows
    for i in range(ASSET_CAPACITY):
        r = DATA_START_ROW + i

        if variant == "demo" and i < len(SAMPLE_ASSETS):
            (name, prop, dt, cost, klass, conv, method,
             s179, bonus, salvage, disp, biz_pct, is_vehicle, notes) = SAMPLE_ASSETS[i]
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=prop)
            ws.cell(row=r, column=3, value=dt)
            ws.cell(row=r, column=4, value=cost)
            ws.cell(row=r, column=5, value=klass)
            ws.cell(row=r, column=6, value=conv)
            ws.cell(row=r, column=7, value=method)
            ws.cell(row=r, column=8, value=s179)
            ws.cell(row=r, column=9, value=bonus)
            ws.cell(row=r, column=12, value=salvage)
            ws.cell(row=r, column=13, value=disp)
            ws.cell(row=r, column=14, value=notes)

        # Style input cells (A-I, L, M, N)
        for col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 9, 12, 13, 14]:
            apply_style(ws.cell(row=r, column=col_idx), input_cell_style())

        # Formula columns: J (adjusted basis), K (recovery period)
        c_adj = ws.cell(row=r, column=10,
                         value=f"=IF(D{r}=\"\",\"\",MAX(D{r}-IFERROR(H{r},0)-IFERROR(I{r},0),0))")
        apply_style(c_adj, formula_cell_style())
        c_adj.number_format = '"$"#,##0'

        c_rec = ws.cell(row=r, column=11,
                         value=f"=IF(E{r}=\"\",\"\",IFERROR(VLOOKUP(E{r},Settings!$A$22:$B$27,2,FALSE),\"\"))")
        apply_style(c_rec, formula_cell_style())
        c_rec.alignment = Alignment(horizontal="center", vertical="center")

        # Formats
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4).number_format = '"$"#,##0'
        ws.cell(row=r, column=8).number_format = '"$"#,##0'
        ws.cell(row=r, column=9).number_format = '"$"#,##0'
        ws.cell(row=r, column=12).number_format = '"$"#,##0'
        ws.cell(row=r, column=13).number_format = "yyyy-mm-dd"

        ws.row_dimensions[r].height = 17

    last_row = DATA_START_ROW + ASSET_CAPACITY - 1

    # Dropdowns
    add_dropdown(ws, f"B{DATA_START_ROW}:B{last_row}", SAMPLE_PROPERTIES)
    add_dropdown(ws, f"E{DATA_START_ROW}:E{last_row}", MACRS_CLASS_OPTIONS)
    add_dropdown(ws, f"F{DATA_START_ROW}:F{last_row}", ["HY", "MQ", "MM"])
    add_dropdown(ws, f"G{DATA_START_ROW}:G{last_row}", ["200% DB", "150% DB", "SL"])

    # Totals row
    tot_row = last_row + 1
    a = ws.cell(row=tot_row, column=1, value="Portfolio totals:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for col, label_col in [(4, "D"), (8, "H"), (9, "I"), (10, "J")]:
        c = ws.cell(row=tot_row, column=col,
                     value=f"=SUM({label_col}{DATA_START_ROW}:{label_col}{last_row})")
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[tot_row].height = 24

    append_callout(ws, tot_row + 2, (
        "ⓘ  Adjusted basis = Cost − §179 − Bonus.  This is the amount "
        "the year-by-year MACRS schedule depreciates over the recovery "
        "period.  If you elect §179 or bonus, the standard percentage "
        "tables apply only to the residual basis, not the full cost."
    ))


# ============================================================================
# Tab 2 — Schedule by Year (matrix)
# ============================================================================

def build_schedule_by_year_tab(wb, variant):
    ws = wb.create_sheet("Schedule by Year")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Schedule by Year",
                         prev_tab="Asset Register",
                         next_tab="Current Year Summary")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        f"Asset × Year matrix.  Each cell = (Adjusted basis × MACRS Y_n %) "
        f"from the Settings lookup table.  Bottom row = portfolio total per year."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[4].height = 22

    # Col widths — A wide for asset, B compact for class, then 27 narrow year cols
    widths = [("A", 32), ("B", 8), ("C", 12)]
    # year cols D..AD (27 columns) — letters: D..Z (23) then AA..AD (4) = 27
    year_cols = []
    for i in range(SCHEDULE_YEARS):
        col_idx = 4 + i  # D = 4
        year_cols.append(get_column_letter(col_idx))
    for letter in year_cols:
        widths.append((letter, 11))
    set_col_widths(ws, widths)

    # Header row 7
    headers = ["Asset", "Class", "Adj. basis"] + [f"Year {i+1}" for i in range(SCHEDULE_YEARS)]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[HEADER_ROW].height = 22

    # Data rows: mirror Asset Register, then per-year formula
    for i in range(ASSET_CAPACITY):
        r = DATA_START_ROW + i
        ar_row = DATA_START_ROW + i  # same row index

        # A: asset name mirror
        c = ws.cell(row=r, column=1, value=f"='Asset Register'!A{ar_row}")
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # B: class mirror
        c = ws.cell(row=r, column=2, value=f"='Asset Register'!E{ar_row}")
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        # C: adjusted basis mirror
        c = ws.cell(row=r, column=3, value=f"='Asset Register'!J{ar_row}")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'

        # Year columns: depreciation = adj basis × MACRS_pct(class, year_n)
        # Settings!A22:AC27 holds the MACRS table:
        #   A = class label, B = recovery period, C..AC = Year-1..Year-27 percentages.
        for y in range(SCHEDULE_YEARS):
            col_idx = 4 + y
            year_letter = get_column_letter(col_idx)
            # INDEX/MATCH against Settings MACRS table; col 3 = Year 1
            formula = (
                f'=IF(OR($B{r}="",$B{r}="Land",$C{r}=""),"",'
                f'IFERROR($C{r}*INDEX(Settings!$C$22:$AC$27,'
                f'MATCH($B{r},Settings!$A$22:$A$27,0),{y+1}),0))'
            )
            cc = ws.cell(row=r, column=col_idx, value=formula)
            apply_style(cc, formula_cell_style())
            cc.number_format = '"$"#,##0'
            cc.font = Font(name=FONT_BODY, size=9, color=COLOR_TEXT)

        ws.row_dimensions[r].height = 16

    last_row = DATA_START_ROW + ASSET_CAPACITY - 1

    # Totals row — sum each year column
    tot_row = last_row + 1
    a = ws.cell(row=tot_row, column=1, value="Portfolio total / year:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells(f"A{tot_row}:C{tot_row}")
    for y in range(SCHEDULE_YEARS):
        col_idx = 4 + y
        letter = get_column_letter(col_idx)
        c = ws.cell(row=tot_row, column=col_idx,
                     value=f"=SUM({letter}{DATA_START_ROW}:{letter}{last_row})")
        c.font = Font(name=FONT_HEAD, size=10, bold=True, color=COLOR_ACCENT)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
    ws.row_dimensions[tot_row].height = 22

    # Freeze panes — keep asset name + class + basis visible while scrolling years
    ws.freeze_panes = "D8"

    append_callout(ws, tot_row + 2, (
        "ⓘ  Year columns reflect the IRS half-year (HY) and mid-month "
        "(MM) percentage tables loaded in Settings.  For mid-quarter "
        "(MQ) convention or short tax years, adjust manually — this "
        "ledger uses the standard tables."
    ))


# ============================================================================
# Tab 3 — Current Year Summary
# ============================================================================

def build_current_year_summary_tab(wb, variant):
    ws = wb.create_sheet("Current Year Summary")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Current Year Summary",
                         prev_tab="Schedule by Year",
                         next_tab="Form 4562 Worksheet")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Current-year deduction per asset.  Total flows to Schedule C "
        "line 13 (active business) or Schedule E line 18 (passive rental)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 32),  # Asset
        ("B", 8),   # Class
        ("C", 13),  # Placed in service
        ("D", 13),  # Original basis
        ("E", 13),  # Adjusted basis
        ("F", 14),  # Current-year depreciation
        ("G", 16),  # Accumulated through prior year
        ("H", 18),  # Schedule line mapping
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    headers = [
        "Asset", "Class", "Placed in svc",
        "Original basis", "Adj. basis",
        "Current-year deprec.", "Accum. through prior yr",
        "Schedule line",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[HEADER_ROW].height = 22

    # Per-asset rows
    for i in range(ASSET_CAPACITY):
        r = DATA_START_ROW + i
        ar = DATA_START_ROW + i
        sb = DATA_START_ROW + i  # Schedule by Year row

        # A — asset name
        c = ws.cell(row=r, column=1, value=f"='Asset Register'!A{ar}")
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # B — class
        c = ws.cell(row=r, column=2, value=f"='Asset Register'!E{ar}")
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        # C — placed in service
        c = ws.cell(row=r, column=3, value=f"='Asset Register'!C{ar}")
        apply_style(c, formula_cell_style())
        c.number_format = "yyyy-mm-dd"

        # D — original cost basis
        c = ws.cell(row=r, column=4, value=f"='Asset Register'!D{ar}")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'

        # E — adjusted basis
        c = ws.cell(row=r, column=5, value=f"='Asset Register'!J{ar}")
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'

        # F — current-year depreciation
        # year_n = active_year - YEAR(placed_in_service) + 1; clamp to 1..27
        # then use INDEX into Schedule by Year row's year columns (D..AD).
        year_n_expr = (
            f'IF(\'Asset Register\'!C{ar}="",0,'
            f'MAX(MIN(Settings!$B$5-YEAR(\'Asset Register\'!C{ar})+1,{SCHEDULE_YEARS}),0))'
        )
        # If year_n is 0 or asset disposed before active year, returns 0
        formula_f = (
            f'=IFERROR(IF({year_n_expr}<1,0,'
            f'IF(AND(\'Asset Register\'!M{ar}<>"",'
            f'YEAR(\'Asset Register\'!M{ar})<Settings!$B$5),0,'
            f'INDEX(\'Schedule by Year\'!$D{sb}:$AD{sb},1,{year_n_expr}))),0)'
        )
        c = ws.cell(row=r, column=6, value=formula_f)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # G — accumulated through prior year (sum Year 1 .. Year_n-1)
        formula_g = (
            f'=IFERROR(IF({year_n_expr}<=1,0,'
            f'SUM(OFFSET(\'Schedule by Year\'!$D{sb},0,0,1,{year_n_expr}-1))),0)'
        )
        c = ws.cell(row=r, column=7, value=formula_g)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'

        # H — schedule line mapping (formula references Settings B6 schedule classification)
        c = ws.cell(row=r, column=8,
                     value='=IF(Settings!$B$6="Schedule C (active)","Sch C, line 13",'
                           'IF(Settings!$B$6="Schedule E (passive)","Sch E, line 18",'
                           '"Ask CPA"))')
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 17

    last_row = DATA_START_ROW + ASSET_CAPACITY - 1

    # Total row — current-year + accumulated
    tot_row = last_row + 1
    a = ws.cell(row=tot_row, column=1, value="Portfolio total →")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells(f"A{tot_row}:E{tot_row}")
    for col, letter in [(6, "F"), (7, "G")]:
        c = ws.cell(row=tot_row, column=col,
                     value=f"=SUM({letter}{DATA_START_ROW}:{letter}{last_row})")
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.number_format = '"$"#,##0'
    # H gets a label
    c = ws.cell(row=tot_row, column=8, value="↓ drop on form")
    c.font = Font(name=FONT_MONO, size=9, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[tot_row].height = 26

    append_callout(ws, tot_row + 2, (
        "ⓘ  Schedule line mapping reads from Settings!B6.  Active "
        "business hosts (substantial services — daily cleaning, meals, "
        "concierge) file Schedule C; most STR landlords file Schedule "
        "E.  When in doubt, ask your CPA — the choice drives whether "
        "self-employment tax (15.3%) applies."
    ))


# ============================================================================
# Tab 4 — Form 4562 Worksheet
# ============================================================================

def build_form_4562_worksheet_tab(wb, variant):
    ws = wb.create_sheet("Form 4562 Worksheet")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    compact_header_band(ws, "Form 4562 Worksheet",
                         prev_tab="Current Year Summary",
                         next_tab="Settings")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Mirror of IRS Form 4562 — Parts I (§179), II (Bonus), III (MACRS), "
        "and Part V (Listed Property — vehicles).  Print + hand to your CPA."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 8), ("B", 50), ("C", 18),
        ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    # Section banner helper
    def banner(row, text):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def line(row, num, label, formula, kind="money"):
        ws.cell(row=row, column=1, value=num).font = Font(
            name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        c = ws.cell(row=row, column=3, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0' if kind in ("money", "total") else "0"
        if kind == "total":
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            ws.cell(row=row, column=2).font = Font(
                name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
            )
        ws.row_dimensions[row].height = 18

    last_ar_row = DATA_START_ROW + ASSET_CAPACITY - 1
    cur_tot_row = DATA_START_ROW + ASSET_CAPACITY + 1  # Current Year Summary total row

    r = 6
    banner(r, "PART I — §179 ELECTION TO EXPENSE")
    r += 1
    line(r, "1", "Total §179 elected (sum from Asset Register col H):",
         f"=SUM('Asset Register'!H{DATA_START_ROW}:H{last_ar_row})", "total"); r += 1

    r += 1
    banner(r, "PART II — SPECIAL DEPRECIATION (BONUS)")
    r += 1
    line(r, "14", "Total Bonus depreciation (sum from Asset Register col I):",
         f"=SUM('Asset Register'!I{DATA_START_ROW}:I{last_ar_row})", "total"); r += 1

    r += 1
    banner(r, "PART III — MACRS DEPRECIATION (current year)")
    r += 1
    line(r, "17", "MACRS deduction for assets placed in service prior years:",
         f"='Current Year Summary'!F{cur_tot_row}", "money"); r += 1
    line(r, "19a-i", "MACRS deduction for current-year placements (subset of 17):",
         f'=SUMIFS(\'Current Year Summary\'!F{DATA_START_ROW}:F{last_ar_row},'
         f"'Asset Register'!C{DATA_START_ROW}:C{last_ar_row},\">=\"&"
         f'DATE(Settings!$B$5,1,1))', "money"); r += 1
    line(r, "TOT", "GRAND TOTAL CURRENT-YEAR DEPRECIATION:",
         f"='Current Year Summary'!F{cur_tot_row}+SUM('Asset Register'!I{DATA_START_ROW}:I{last_ar_row})",
         "total"); r += 1

    r += 1
    banner(r, "PART V — LISTED PROPERTY (VEHICLES) — §280F")
    r += 2

    # Listed-property worksheet — vehicle table
    listed_headers = [
        "Vehicle (Asset name)", "Date placed in service",
        "Total miles (Y)", "Business miles (Y)", "Commuting miles (Y)",
        "Other personal (Y)", "Business-use %", "Evidence kept?",
    ]
    for col, h in enumerate(listed_headers, start=1):
        col_letter = get_column_letter(col + 1)  # start at column B
        cell = ws.cell(row=r, column=col + 1, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[r].height = 28
    listed_header_row = r
    r += 1

    # 5 rows for listed-property entries (vehicles flagged in Asset Register)
    listed_capacity = 5
    sample_vehicles = [
        ("Ford F-150 (heavy truck >6,000lb)", date(2024, 1, 15),
         12_500, 9_750, 1_500, 1_250, 0.78, "Y"),
    ] if variant == "demo" else []

    for i in range(listed_capacity):
        for col_idx in range(2, 10):
            apply_style(ws.cell(row=r, column=col_idx), input_cell_style())
        if i < len(sample_vehicles):
            (vname, vdate, total_mi, biz_mi, comm_mi, other_mi, biz_pct, ev) = sample_vehicles[i]
            ws.cell(row=r, column=2, value=vname)
            ws.cell(row=r, column=3, value=vdate)
            ws.cell(row=r, column=4, value=total_mi)
            ws.cell(row=r, column=5, value=biz_mi)
            ws.cell(row=r, column=6, value=comm_mi)
            ws.cell(row=r, column=7, value=other_mi)
            ws.cell(row=r, column=8, value=biz_pct)
            ws.cell(row=r, column=9, value=ev)
        # Formats
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        for cc in (4, 5, 6, 7):
            ws.cell(row=r, column=cc).number_format = "#,##0"
        ws.cell(row=r, column=8).number_format = "0.00%"
        # Evidence dropdown
        ws.row_dimensions[r].height = 17
        r += 1

    add_dropdown(ws, f"I{listed_header_row+1}:I{listed_header_row+listed_capacity}", ["Y", "N"])

    r += 1
    append_callout(ws, r, (
        "ⓘ  Listed-property §280F rule: vehicles must be used > 50% for "
        "qualified business use to claim §179 or bonus depreciation in "
        "Year 1.  If business-use % drops to ≤ 50% in a later year, "
        "prior accelerated depreciation can be recaptured as ordinary "
        "income.  Investment use (e.g., bank trips for STR deposits) is "
        "deductible but does NOT count toward the 50% test."
    ))
    r += 2

    append_callout(ws, r, (
        "ⓘ  Pull total / business / commuting miles from your TAX-001 "
        "Mileage Log if you maintain one.  Otherwise enter manual "
        "year-end totals.  Evidence-kept = log, calendar entries, "
        "receipts, contractor agreements supporting the business-use "
        "percentage."
    ))

    ws.print_area = f"A1:I{r + 2}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ============================================================================
# Tab 5 — Settings
# ============================================================================

def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 32)])
    compact_header_band(ws, "Settings",
                         prev_tab="Form 4562 Worksheet", next_tab=None)

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Active tax year · property list · Schedule C/E classification · "
        "MACRS percentage table (IRS Pub 946 Tables A-1 / A-6)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[4].height = 22

    set_col_widths(ws, [
        ("A", 18), ("B", 16),
    ])
    # Year columns C..AC = 27 cols, narrow
    for i in range(SCHEDULE_YEARS):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 7

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # B5 — Active tax year
    a = ws.cell(row=5, column=1, value="Active tax year:")
    a.font = bold
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=5, column=2, value=ACTIVE_YEAR_DEFAULT)
    apply_style(c, input_cell_style())
    c.number_format = "0"
    ws.row_dimensions[5].height = 22

    # B6 — Schedule classification
    a = ws.cell(row=6, column=1, value="Schedule classification:")
    a.font = bold
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=6, column=2,
                 value=_val(variant, "Schedule E (passive)"))
    apply_style(c, input_cell_style())
    add_dropdown(ws, "B6",
                  ["Schedule E (passive)", "Schedule C (active)", "Ask my CPA"])
    ws.row_dimensions[6].height = 22

    # Property list (B7..B16)
    a = ws.cell(row=7, column=1, value="Property list:")
    a.font = bold
    a.alignment = Alignment(horizontal="right", vertical="top", indent=1)
    for i in range(10):
        r = 7 + i
        c = ws.cell(row=r, column=2)
        if variant == "demo" and i < len(SAMPLE_PROPERTIES):
            c.value = SAMPLE_PROPERTIES[i]
        apply_style(c, input_cell_style())

    # Freshness stamp
    ws.merge_cells("A18:F18")
    c = ws["A18"]
    c.value = (
        "📅  MACRS percentage tables current as of IRS Pub 946 (2024 rev).  "
        "Tables A-1 / A-6 — half-year + mid-month conventions.  Update if IRS revises."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[18].height = 28

    # MACRS class table — header at row 21, data 22-27
    a = ws.cell(row=20, column=1, value="MACRS Percentage Table")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[20].height = 22

    # Row 21 — column headers
    headers = ["Class", "Recovery (yr)"] + [f"Y{i+1}" for i in range(SCHEDULE_YEARS)]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=21, column=col, value=h)
        apply_style(cell, header_row_style())
        cell.font = Font(name=FONT_BODY, size=9, bold=True, color=COLOR_WHITE)
    ws.row_dimensions[21].height = 22

    # Rows 22-27 — one per MACRS class
    class_rows = [
        ("5",    "5",    MACRS_TABLES["5"]),
        ("7",    "7",    MACRS_TABLES["7"]),
        ("15",   "15",   MACRS_TABLES["15"]),
        ("27.5", "27.5", MACRS_TABLES["27.5"]),
        ("39",   "39",   MACRS_TABLES["39"]),
        ("Land", "n/a",  MACRS_TABLES["Land"]),
    ]
    for i, (klass, rec, pcts) in enumerate(class_rows):
        r = 22 + i
        ws.cell(row=r, column=1, value=klass).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="center", vertical="center")
        ws.cell(row=r, column=2, value=rec).alignment = Alignment(
            horizontal="center", vertical="center")
        ws.cell(row=r, column=2).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        for y, pct in enumerate(pcts):
            cc = ws.cell(row=r, column=3 + y, value=pct)
            cc.number_format = "0.00%"
            cc.font = Font(name=FONT_BODY, size=8, color=COLOR_TEXT)
            cc.alignment = Alignment(horizontal="center", vertical="center")
            # Alt-row banding
            if i % 2 == 1:
                cc.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        ws.row_dimensions[r].height = 16

    # Year-end Archive
    ws.row_dimensions[29].height = 8
    a = ws.cell(row=30, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[30].height = 22

    archive_headers = ["Year", "Assets tracked", "Current-year deprec.",
                        "Accumulated", "§179 used", "Bonus used"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=31, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[31].height = 18
    for idx, year in enumerate(range(2024, 2031), start=32):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in range(2, 7):
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            if col >= 3:
                cell.number_format = '"$"#,##0'
            else:
                cell.number_format = "0"
        ws.row_dimensions[idx].height = 16

    brand_footer(ws, 41, version_line=VERSION_LINE)


# ============================================================================
# Workbook orchestrator
# ============================================================================

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_asset_register_tab(wb, variant)
    build_schedule_by_year_tab(wb, variant)
    build_current_year_summary_tab(wb, variant)
    build_form_4562_worksheet_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = f"Depreciation Tracker — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Master depreciation register (5/7/15/27.5/39-yr MACRS) for STR hosts. "
        "Year-by-year schedule + Form 4562 worksheet + Part V listed property."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
