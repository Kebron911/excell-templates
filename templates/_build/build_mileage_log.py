"""Build TAX-001 STR Mileage Log Excel file (v2.2 visual standard).

Implements templates/_briefs/TAX-001-mileage-log-spec.md with the v2.2
visual + interaction language set by the Welcome Book v2.2 — operational
mode (recurring data entry; no completion %).

Generates templates/_masters/TAX-001-mileage-log.xlsx.

Usage:
    python build_mileage_log.py

Dependencies: openpyxl (see requirements.txt).
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    STATE_BAD_FILL, STATE_GOOD_FILL,
    STATE_WARN_FILL,
)

SKU = "TAX-001"
NAME = "mileage-log"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# --- Constants ---

IRS_RATE_2026 = 0.725  # IRS Notice 2026-10 — 72.5¢/mi business use, +2.5¢ from 2025
RATES_AS_OF = "2026-01-01"  # Suite Theme 4 — bump when IRS publishes next Notice


def _val(variant, demo_value):
    """Return demo_value when building DEMO, None for BLANK."""
    return demo_value if variant == "demo" else None

PURPOSES = [
    "Property inspection",
    "Turnover (laundry/clean)",
    "Supplies run",
    "Guest transport",
    "Repairs",
    "Contractor walkthrough",
    "Emergency response",
    "Meeting cleaner / staff",
    "Bank / deposit",
    "Mailing / post office",
    "CPA / attorney meeting",
    "Other (describe in Notes)",
]

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]

# Sample rows 6-25 — 20 trips Jan-Mar 2026
# v2.3.1: tuple grew to 12 fields with Logged On (audit-defense — proves
# contemporaneous entry) and Parking + Tolls $ (separately deductible
# under Standard Mileage Rate, IRS Pub 463).
# Tuple: (trip_date, property, destination, purpose, start_odo, end_odo,
#         typed_miles, notes, duration_hr, receipt_y_n, logged_on, parking_tolls)
MILEAGE_SAMPLES = [
    ("2026-01-05", "Smokies Ridge", "Home Depot run",                        "Supplies run",                  45120, 45144, "", "Hot tub chemicals",                  1.5, "Y", "2026-01-05", 0),
    ("2026-01-08", "Smokies Ridge", "Property inspection visit",             "Property inspection",           45150, 45195, "", "Quarterly inspection",               2.5, "",  "2026-01-09", 0),
    ("2026-01-15", "Creek Side",    "Knoxville airport guest pickup",        "Guest transport",               45200, 45292, "", "Airbnb Plus comped ride",            3.0, "",  "2026-01-15", 14),
    ("2026-01-20", "Smokies Ridge", "Meet new cleaner onsite",               "Meeting cleaner / staff",       45300, 45345, "", "Onboard Jamie",                      2.0, "",  "2026-01-21", 0),
    ("2026-02-02", "Creek Side",    "Turnover walkthrough",                  "Turnover (laundry/clean)",      45400, 45490, "", "",                                    4.0, "Y", "2026-02-02", 0),
    ("2026-02-07", "Lakehouse A",   "Supplies + inspection",                 "Supplies run",                  45500, 45565, "", "",                                    2.5, "Y", "2026-02-08", 0),
    ("2026-02-10", "Smokies Ridge", "Burst pipe emergency",                  "Emergency response",            45600, 45648, "", "Plumber met me",                     3.5, "Y", "2026-02-10", 0),
    ("2026-02-14", "Creek Side",    "Hot tub service",                       "Contractor walkthrough",        45700, 45790, "", "",                                    2.0, "",  "2026-02-15", 0),
    ("2026-02-18", "Lakehouse A",   "Quarterly inspection",                  "Property inspection",           45800, 45867, "", "",                                    2.5, "",  "2026-02-19", 0),
    ("2026-02-22", "Smokies Ridge", "Guest pickup airport",                  "Guest transport",               45900, 45992, "", "",                                    3.0, "",  "2026-02-22", 18),
    ("2026-02-28", "Creek Side",    "Cleaner check-in",                      "Meeting cleaner / staff",       46050, 46095, "", "",                                    1.5, "",  "2026-03-01", 0),
    ("2026-03-05", "Lakehouse A",   "Lowe's run",                            "Supplies run",                  46150, 46178, "", "Bought new kettle",                  1.0, "Y", "2026-03-05", 0),
    ("2026-03-10", "Smokies Ridge", "Turnover issue callback",               "Turnover (laundry/clean)",      46200, 46245, "", "Guest complaint follow-up",          2.0, "",  "2026-03-11", 0),
    ("2026-03-14", "Creek Side",    "Handyman meeting",                      "Contractor walkthrough",        46300, 46349, "", "",                                    1.5, "",  "2026-03-15", 0),
    ("2026-03-18", "Smokies Ridge", "Bank deposit run",                      "Bank / deposit",                46400, 46420, "", "Q1 guest deposits",                  0.75, "Y", "2026-03-18", 0),
    ("2026-03-22", "Lakehouse A",   "Guest transport — late flight",         "Guest transport",               46500, 46587, "", "",                                    3.5, "",  "2026-03-22", 22),
    ("2026-03-25", "Creek Side",    "Supplies — Costco",                     "Supplies run",                  46650, 46695, "", "",                                    1.5, "Y", "2026-03-26", 0),
    ("2026-03-28", "Smokies Ridge", "Mailing key set to next guest",         "Mailing / post office",         46800, 46815, "", "Lockbox malfunction backup",         0.5, "Y", "2026-03-28", 0),
    ("2026-03-30", "Lakehouse A",   "CPA tax-prep meeting",                  "CPA / attorney meeting",        46900, 46928, "", "Q1 review (Re: Smith CPA)",          1.5, "Y", "2026-03-31", 0),
    ("2026-03-31", "Smokies Ridge", "Month-end inspection",                  "Property inspection",           46950, 46995, "", "",                                    2.0, "",  "2026-04-01", 0),
]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def add_dropdown(ws, cell_range, formula1):
    """Add a list data validation to the given cell range."""
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational-mode hero + cards + activity dashboard)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "STR Mileage Log"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Close your year before April does."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "The IRS requires 4 things on every business-mileage entry: date, "
        "destination, business purpose, and miles driven (Pub 463). This log "
        "captures all 4 plus odometer readings, property allocation, and "
        "calculated $ deduction at the current rate. Print the Monthly "
        "Summary and YTD Dashboard tabs for your CPA at year-end."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Click the navy button below to open the Mileage Log — log one row per trip.",
        "② Use odometer columns (Start + End) OR type miles directly. Formula prefers odometer.",
        "③ Fill 'Duration (hr)' too — it counts toward 100/500-hour material-participation tests.",
        "④ Column K shows ✓ when an entry meets the IRS 4-element rule; ⚠ flags missing fields.",
        "⑤ Log within 7 days of each trip — auditors disallow records reconstructed at year-end.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "ADD A NEW TRIP" button (rows 26-29) ---
    pseudo_button(ws, "A26", "L29",
                   "→  ADD A NEW TRIP  (OPEN MILEAGE LOG)",
                   "'Mileage Log'!A6", variant="primary")
    for r in range(26, 30):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 31-37) ---
    for r in range(31, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Trips Logged
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "TRIPS LOGGED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = "=COUNTA('Mileage Log'!A6:A2005)"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "trips this year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): YTD Deduction
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "YTD DEDUCTION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = "='Monthly Summary'!C19"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "at current IRS rate"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Last Entry
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "LAST ENTRY"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = '=IFERROR(TEXT(MAX(\'Mileage Log\'!A6:A2005),"yyyy-mm-dd"),"—")'
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "most recent trip date"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Gold borders around the 3 cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Monthly Summary",
                   "'Monthly Summary'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "YTD Dashboard",
                   "'YTD Dashboard'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Print Packet",
                   "'Monthly Summary'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Tax-time / year-end checkpoint callout (rows 42-43) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📄 TAX TIME: print Monthly Summary + YTD Dashboard "
        "(File → Print → 'Print Active Sheets'). "
        "📅 YEAR END: copy YTD totals into Settings → Year-end Archive before Jan 1."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: IRS rate maintenance reminder (row 44) ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        f"⚠ IRS rate updates January 1 each year. Current 2026: ${IRS_RATE_2026:.3f}/mi (Notice 2026-10). "
        "Update Settings → B5 each January (irs.gov Publication 463)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    # --- ZONE 9: Upgrade banner (row 46) — preserved from v1 ---
    add_upgrade_banner(ws, 46)

    # --- ZONE 10: Footer (rows 48-50) ---
    brand_footer(ws, 48,
                 version_line="TAX-001 · v2.2 · Free updates forever")

    # Print setup
    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_log_tab(wb, variant):
    """Sheet 1 — Mileage Log (one row per trip, 2000 capacity rows)."""
    ws = wb.create_sheet("Mileage Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    # v2.2 compact navy header (rows 1-3) replaces apply_brand_header
    compact_header_band(ws, "Mileage Log",
                         prev_tab="Start", next_tab="Monthly Summary")

    # Row 4: subtitle/spacer (parchment fill, italic)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per trip — odometer or typed miles"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Row 5: styled headers
    # v2.3   added Duration + Receipt (material-participation hours, audit cross-ref)
    # v2.3.1 adds Logged On (audit: contemporaneous-record proof) and
    #        Parking + Tolls $ (separately deductible under Standard Mileage Rate)
    headers = [
        "Trip Date",
        "Property",
        "Destination",
        "Business Purpose",
        "Start Odo",
        "End Odo",
        "Miles (typed alt)",
        "Calculated Miles",
        "$ Deduction",
        "Notes",
        "✔ IRS",
        "Duration (hr)",
        "📎 Receipt",
        "Logged On",         # v2.3.1 — when this row was entered (audit defense)
        "Parking + Tolls $", # v2.3.1 — separately deductible per Pub 463
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Col widths — tightened to fit 15 cols on landscape Letter
    set_col_widths(ws, [
        ("A", 11), ("B", 16), ("C", 20), ("D", 18),
        ("E", 8), ("F", 8), ("G", 8), ("H", 9),
        ("I", 11), ("J", 18), ("K", 10), ("L", 9), ("M", 8),
        ("N", 11), ("O", 10),
    ])

    # Rows 6-2005: data + formulas. DEMO populates rows 6-25 with sample
    # trips; BLANK leaves all rows empty (formulas still wired so the customer
    # gets ✓/⚠ feedback the moment they enter a row).
    samples = MILEAGE_SAMPLES if variant == "demo" else []
    for i in range(6, 2006):
        sample_idx = i - 6
        sample = samples[sample_idx] if sample_idx < len(samples) else None

        if sample:
            (date_val, prop, dest, purpose,
             start_odo, end_odo, typed_miles, notes,
             duration_hr, receipt_yn, logged_on, parking_tolls) = sample

            date_obj = datetime.strptime(date_val, "%Y-%m-%d").date()
            a = ws.cell(row=i, column=1, value=date_obj)
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"

            b = ws.cell(row=i, column=2, value=prop)
            apply_style(b, input_cell_style())

            c = ws.cell(row=i, column=3, value=dest)
            apply_style(c, input_cell_style())

            d = ws.cell(row=i, column=4, value=purpose)
            apply_style(d, input_cell_style())

            e = ws.cell(row=i, column=5, value=start_odo)
            apply_style(e, input_cell_style())

            f = ws.cell(row=i, column=6, value=end_odo)
            apply_style(f, input_cell_style())

            g_val = typed_miles if typed_miles != "" else None
            g = ws.cell(row=i, column=7, value=g_val)
            apply_style(g, input_cell_style())

            j = ws.cell(row=i, column=10, value=notes if notes else None)
            apply_style(j, input_cell_style())

            l = ws.cell(row=i, column=12, value=duration_hr if duration_hr else None)
            apply_style(l, input_cell_style())
            l.number_format = "0.00"

            m = ws.cell(row=i, column=13, value=receipt_yn if receipt_yn else None)
            apply_style(m, input_cell_style())
            m.alignment = Alignment(horizontal="center", vertical="center")

            # v2.3.1 — Logged On (date entered) for contemporaneous-record proof
            n_val = datetime.strptime(logged_on, "%Y-%m-%d").date() if logged_on else None
            n = ws.cell(row=i, column=14, value=n_val)
            apply_style(n, input_cell_style())
            n.number_format = "yyyy-mm-dd"

            # v2.3.1 — Parking + Tolls $ (separately deductible)
            o = ws.cell(row=i, column=15, value=parking_tolls if parking_tolls else None)
            apply_style(o, input_cell_style())
            o.number_format = '"$"#,##0.00'
        else:
            l = ws.cell(row=i, column=12)
            apply_style(l, input_cell_style())
            l.number_format = "0.00"
            m = ws.cell(row=i, column=13)
            apply_style(m, input_cell_style())
            m.alignment = Alignment(horizontal="center", vertical="center")
            n = ws.cell(row=i, column=14)
            apply_style(n, input_cell_style())
            n.number_format = "yyyy-mm-dd"
            o = ws.cell(row=i, column=15)
            apply_style(o, input_cell_style())
            o.number_format = '"$"#,##0.00'

        h_cell = ws.cell(
            row=i, column=8,
            value=f"=IF(AND(E{i}<>\"\",F{i}<>\"\"), F{i}-E{i}, IF(G{i}<>\"\", G{i}, 0))",
        )
        h_cell.number_format = "0"
        if sample:
            apply_style(h_cell, formula_cell_style())

        i_cell = ws.cell(
            row=i, column=9,
            value=f"=H{i}*Settings!$B$5",
        )
        i_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(i_cell, formula_cell_style())

        k_cell = ws.cell(
            row=i, column=11,
            value=f'=IF(AND(A{i}<>"",C{i}<>"",D{i}<>"",H{i}>0),"✓","⚠ missing")',
        )
        if sample:
            apply_style(k_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    # Dropdowns (preserved + receipt Y/N)
    add_dropdown(ws, "B6:B2005", "=Settings!$E$11:$E$30")
    add_dropdown(ws, "D6:D2005", "=Settings!$G$11:$G$30")
    add_dropdown(ws, "M6:M2005", '"Y,N"')

    # Conditional formatting on K6:K2005 (preserved)
    ws.conditional_formatting.add(
        "K6:K2005",
        FormulaRule(
            formula=['K6="⚠ missing"'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
        ),
    )
    ws.conditional_formatting.add(
        "K6:K2005",
        FormulaRule(
            formula=['K6="✓"'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
        ),
    )

    # Freeze on row 6 so the navy header + table header stay visible
    ws.freeze_panes = "A6"

    # Print setup: landscape (13 columns), fit to width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_monthly_tab(wb):
    """Sheet 2 — Monthly Summary (auto-calculated from Mileage Log)."""
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Monthly Summary",
                         prev_tab="Mileage Log", next_tab="YTD Dashboard")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Auto-calculated from the Mileage Log"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Row 5: headers (preserved)
    headers = ["Month", "Miles", "$ Deduction", "Trip Count"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [("A", 14), ("B", 14), ("C", 16), ("D", 14)])

    # Rows 6-17: monthly SUMIFS / COUNTIFS (preserved from v1)
    for i in range(6, 18):
        month_num = i - 5
        next_month = month_num + 1

        ws.cell(row=i, column=1, value=MONTHS[month_num - 1]).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )

        # Year filter sourced from Settings!$B$6 (Active tax year), not
        # YEAR(TODAY()) — see Settings tab. This decouples the dashboard
        # from Excel's real-world clock so sample/historical years still
        # aggregate correctly.
        if month_num < 12:
            b_formula = (
                f"=SUMIFS('Mileage Log'!$H:$H,"
                f"'Mileage Log'!$A:$A,\">=\"&DATE(Settings!$B$6,{month_num},1),"
                f"'Mileage Log'!$A:$A,\"<\"&DATE(Settings!$B$6,{next_month},1))"
            )
        else:
            b_formula = (
                f"=SUMIFS('Mileage Log'!$H:$H,"
                f"'Mileage Log'!$A:$A,\">=\"&DATE(Settings!$B$6,{month_num},1),"
                f"'Mileage Log'!$A:$A,\"<\"&DATE(Settings!$B$6+1,1,1))"
            )
        b_cell = ws.cell(row=i, column=2, value=b_formula)
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = "0"

        c_cell = ws.cell(row=i, column=3, value=f"=B{i}*Settings!$B$5")
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = '"$"#,##0.00'

        if month_num < 12:
            d_formula = (
                f"=COUNTIFS('Mileage Log'!$A:$A,\">=\"&DATE(Settings!$B$6,{month_num},1),"
                f"'Mileage Log'!$A:$A,\"<\"&DATE(Settings!$B$6,{next_month},1))"
            )
        else:
            d_formula = (
                f"=COUNTIFS('Mileage Log'!$A:$A,\">=\"&DATE(Settings!$B$6,{month_num},1),"
                f"'Mileage Log'!$A:$A,\"<\"&DATE(Settings!$B$6+1,1,1))"
            )
        d_cell = ws.cell(row=i, column=4, value=d_formula)
        apply_style(d_cell, formula_cell_style())

        ws.row_dimensions[i].height = 16

    ws.row_dimensions[18].height = 8

    # Row 19: YTD totals (preserved with original gold tint)
    ytd_gold = PatternFill("solid", fgColor=STATE_WARN_FILL)
    ytd_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    a19 = ws.cell(row=19, column=1, value="YTD Total")
    a19.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a19.fill = ytd_gold

    b19 = ws.cell(row=19, column=2, value="=SUM(B6:B17)")
    b19.font = ytd_font
    b19.fill = ytd_gold
    b19.number_format = "0"

    c19 = ws.cell(row=19, column=3, value="=SUM(C6:C17)")
    c19.font = ytd_font
    c19.fill = ytd_gold
    c19.number_format = '"$"#,##0.00'

    d19 = ws.cell(row=19, column=4, value="=SUM(D6:D17)")
    d19.font = ytd_font
    d19.fill = ytd_gold

    ws.row_dimensions[19].height = 20

    ws.freeze_panes = "A6"

    # --- Chart: Miles by Month (column chart, anchored F5) ---
    bar = BarChart()
    bar.type = "col"
    bar.title = "Miles by Month"
    bar.y_axis.title = "Miles"
    bar.x_axis.title = "Month"
    bar.legend = None
    bar.height = 9
    bar.width = 18
    miles_ref = Reference(ws, min_col=2, min_row=5, max_row=17, max_col=2)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=17)
    bar.add_data(miles_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    # Force category tick labels to render below the bars. Without these,
    # openpyxl's default chart style preset hides the X-axis category
    # text — bars appear unlabelled.
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "low"
    style_chart(bar)
    ws.add_chart(bar, "F5")

    # Print setup — print area covers data + chart for tax-time
    ws.print_area = "A1:N22"
    ws.print_title_rows = "1:5"  # nav + headers repeat on each printed page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_ytd_tab(wb):
    """Sheet 3 — YTD Dashboard (totals + breakdown by purpose and property).

    Doubles as the CPA hand-off page: tab is gold so a CPA spots it on
    open, subtitle reads "📄 For your CPA — print this tab", print area
    is set below.
    """
    ws = wb.create_sheet("YTD Dashboard")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "YTD Dashboard · For Your CPA",
                         prev_tab="Monthly Summary", next_tab="Settings")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 FOR YOUR CPA — print this tab. Totals + breakdown by purpose and property."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 36), ("B", 18), ("C", 14)])

    bold_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # --- YTD top-line KPIs (rows 5-8) ---
    # v2.3.1 adds Parking & Tolls KPI — separately deductible per Pub 463
    kpis = [
        ("YTD Total Miles:",                       "='Monthly Summary'!B19",            "0"),
        ("YTD Total Deduction ($):",               "='Monthly Summary'!C19",            '"$"#,##0.00'),
        ("YTD Trips:",                             "='Monthly Summary'!D19",            None),
        ("YTD Hours (material participation):",    "=SUM('Mileage Log'!L6:L2005)",      "0.00"),
        ("YTD Parking + Tolls ($):",               "=SUM('Mileage Log'!O6:O2005)",      '"$"#,##0.00'),
    ]
    for i, (label, formula, fmt) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=label).font = bold_font
        b_cell = ws.cell(row=i, column=2, value=formula)
        apply_style(b_cell, formula_cell_style())
        if fmt:
            b_cell.number_format = fmt
        ws.row_dimensions[i].height = 16

    ws.row_dimensions[10].height = 10

    # --- Miles by Business Purpose (rows 11-12 headers, 13+ data) ---
    hdr11 = ws.cell(row=11, column=1, value="Miles by Business Purpose")
    hdr11.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[11].height = 20

    for col, h in enumerate(["Purpose", "Miles", "$"], start=1):
        cell = ws.cell(row=12, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[12].height = 18

    purpose_start_row = 13
    for idx, purpose in enumerate(PURPOSES, start=purpose_start_row):
        ws.cell(row=idx, column=1, value=purpose).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        b_cell = ws.cell(
            row=idx, column=2,
            value=f"=SUMIFS('Mileage Log'!$H:$H,'Mileage Log'!$D:$D,A{idx})",
        )
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = "0"
        c_cell = ws.cell(row=idx, column=3, value=f"=B{idx}*Settings!$B$5")
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[idx].height = 16

    purpose_end_row = purpose_start_row + len(PURPOSES) - 1

    # --- Miles by Property (after purposes table) ---
    property_header_row = purpose_end_row + 2
    ws.row_dimensions[purpose_end_row + 1].height = 10

    hdr_prop = ws.cell(row=property_header_row, column=1,
                        value="Miles by Property")
    hdr_prop.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[property_header_row].height = 20

    for col, h in enumerate(["Property", "Miles", "$"], start=1):
        cell = ws.cell(row=property_header_row + 1, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[property_header_row + 1].height = 18

    property_start_row = property_header_row + 2
    last_property_row = property_start_row + len(PROPERTIES) - 1
    for idx, prop in enumerate(PROPERTIES, start=property_start_row):
        ws.cell(row=idx, column=1, value=prop).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        b_cell = ws.cell(
            row=idx, column=2,
            value=f"=SUMIFS('Mileage Log'!$H:$H,'Mileage Log'!$B:$B,A{idx})",
        )
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = "0"
        c_cell = ws.cell(row=idx, column=3, value=f"=B{idx}*Settings!$B$5")
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[idx].height = 16

    # --- Schedule routing footnote (after property breakdown) ---
    # Merged A:E only (not full L) so the column chart at F+ doesn't
    # cover it. Charts go in cols F-L; data + footnote stay in A-E.
    ws.row_dimensions[last_property_row + 1].height = 10
    footnote_row = last_property_row + 2
    ws.merge_cells(f"A{footnote_row}:E{footnote_row}")
    c = ws.cell(row=footnote_row, column=1)
    c.value = (
        '=IF(Settings!$B$15="Schedule E (passive)",'
        '"→ Schedule E (Form 1040), Line 6.",'
        'IF(Settings!$B$15="Schedule C (active)",'
        '"→ Schedule C, Line 9 + Form 4562 Part V.",'
        '"→ Schedule routing not set — see Settings."))'
    )
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[footnote_row].height = 28

    # --- Chart 1: Donut — $ by Business Purpose ---
    # Anchored F5 (top-right, beside YTD KPIs) so it doesn't overlap the
    # data tables in cols A-C. Sized to fit cols F-L approximately.
    donut = DoughnutChart()
    donut.title = "$ Deduction by Purpose"
    donut.height = 9
    donut.width = 11
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=3, min_row=12,
                            max_row=purpose_end_row, max_col=3)
    donut_cats = Reference(ws, min_col=1, min_row=purpose_start_row,
                            max_row=purpose_end_row)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=False, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, "F5")

    # --- Chart 2: Column — Miles by Property ---
    # Anchored F{property_header_row} so it lines up vertically with the
    # property data in cols A-C, but lives in cols F-L (no overlap).
    col_chart = BarChart()
    col_chart.type = "col"
    col_chart.title = "Miles by Property"
    col_chart.y_axis.title = "Miles"
    col_chart.x_axis.title = "Property"
    col_chart.legend = None
    col_chart.height = 7
    col_chart.width = 11
    prop_data = Reference(ws, min_col=2,
                           min_row=property_header_row + 1,
                           max_row=last_property_row, max_col=2)
    prop_cats = Reference(ws, min_col=1,
                           min_row=property_start_row,
                           max_row=last_property_row)
    col_chart.add_data(prop_data, titles_from_data=True)
    col_chart.set_categories(prop_cats)
    col_chart.x_axis.delete = False
    col_chart.y_axis.delete = False
    col_chart.x_axis.tickLblPos = "low"
    style_chart(col_chart)
    ws.add_chart(col_chart, f"F{property_header_row}")

    # Print setup — covers totals + both breakdowns + footnote + charts
    ws.print_area = f"A1:N{footnote_row + 2}"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    """Sheet 4 — Settings: IRS rate, tax classification, Form 4562 Part V,
    vehicles, dropdown sources, frequent destinations, year-end archive."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="YTD Dashboard", next_tab=None)

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "IRS rate · tax classification · Form 4562 · vehicles · sources"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 36), ("B", 18), ("C", 14), ("D", 16),
        ("E", 22), ("F", 6), ("G", 26),
        ("I", 20), ("J", 26), ("K", 12),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- IRS rate (row 5) + Active tax year (row 6) ---
    a5 = ws.cell(row=5, column=1, value="IRS Rate (per mile, business):")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=IRS_RATE_2026)
    apply_style(b5, input_cell_style())
    b5.number_format = '"$"0.000'
    # Freshness stamp (suite Theme 4): IRS publishes the new business mileage
    # rate each December via Notice. Customer should bump B5 + RATES_AS_OF
    # when their tax year rolls over. Without this, mileage deductions silently
    # use last year's rate.
    ws.merge_cells("C5:G5")
    c5_note = ws["C5"]
    c5_note.value = (
        f"📅 Rate as of {RATES_AS_OF} — IRS publishes the next year's rate "
        f"each Dec; edit B5 then."
    )
    c5_note.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c5_note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 20

    # Active tax year drives Monthly Summary SUMIFS — without this,
    # date-filtered formulas return 0 when Excel's clock doesn't match
    # the year of logged trips. Customer updates this each January.
    a6 = ws.cell(row=6, column=1, value="Active tax year:")
    a6.font = bold_right
    a6.alignment = Alignment(horizontal="right", vertical="center")
    b6 = ws.cell(row=6, column=2, value=2026)
    apply_style(b6, input_cell_style())
    b6.number_format = "0"
    ws.row_dimensions[6].height = 18

    # --- Historical rates (rows 7-10) ---
    hist_hdr = ws.cell(row=7, column=1, value="Historical rates for reference:")
    hist_hdr.font = italic_muted
    ws.row_dimensions[7].height = 16

    # IRS Notice 2026-10 set 2026 at 72.5¢. 2025 was 70¢, 2024 67¢, 2023 65.5¢.
    historical = [(8, "2023:", 0.655), (9, "2024:", 0.67), (10, "2025:", 0.70)]
    for row, label, val in historical:
        ws.cell(row=row, column=1, value=label).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT)
        rc = ws.cell(row=row, column=2, value=val)
        rc.number_format = '"$"0.000'
        rc.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
        ws.row_dimensions[row].height = 16

    # Property + Purpose dropdown headers stay at col E, G row 10
    e10 = ws.cell(row=10, column=5, value="Property list")
    e10.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    g10 = ws.cell(row=10, column=7, value="Purpose list")
    g10.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    for idx, prop in enumerate(PROPERTIES, start=11):
        cell = ws.cell(row=idx, column=5, value=prop)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    for idx, purpose in enumerate(PURPOSES, start=11):
        cell = ws.cell(row=idx, column=7, value=purpose)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # --- Tax Classification & Recordkeeping (rows 12-17) ---
    ws.row_dimensions[11].height = 10

    sect12 = ws.cell(row=12, column=1, value="Tax Classification & Recordkeeping")
    sect12.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[12].height = 22

    # Row 13: Home-office toggle
    a13 = ws.cell(row=13, column=1,
                   value="Home office is principal place of business:")
    a13.font = bold_right
    a13.alignment = Alignment(horizontal="right", vertical="center")
    b13 = ws.cell(row=13, column=2, value="Not yet established")
    apply_style(b13, input_cell_style())
    add_dropdown(ws, "B13", '"Yes,No,Not yet established"')
    ws.row_dimensions[13].height = 18

    # Row 14: home-office explainer
    ws.merge_cells("A14:G14")
    c14 = ws["A14"]
    c14.value = (
        "If 'Yes', trips from home → STR are deductible business travel. "
        "If 'No', they're non-deductible commuting (track separately for "
        "Form 4562 Part V below)."
    )
    c14.font = italic_muted
    c14.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[14].height = 24

    # Row 15: Schedule classification
    a15 = ws.cell(row=15, column=1, value="Schedule classification:")
    a15.font = bold_right
    a15.alignment = Alignment(horizontal="right", vertical="center")
    b15 = ws.cell(row=15, column=2, value="Ask my CPA")
    apply_style(b15, input_cell_style())
    add_dropdown(ws, "B15",
                  '"Schedule E (passive),Schedule C (active),Ask my CPA"')
    ws.row_dimensions[15].height = 18

    # Row 16: substantial-services callout
    ws.merge_cells("A16:G16")
    c16 = ws["A16"]
    c16.value = (
        "⚠ Substantial-services trigger: if you provide cleaning DURING "
        "stays, meals, or concierge, IRS requires Schedule C "
        "(15.3% self-employment tax applies)."
    )
    c16.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c16.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[16].height = 28

    ws.row_dimensions[17].height = 8

    # --- Form 4562 Part V — Vehicle Information (rows 18-26) ---
    sect18 = ws.cell(row=18, column=1,
                      value="Form 4562 Part V — Vehicle Information")
    sect18.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[18].height = 22

    ws.merge_cells("A19:G19")
    c19 = ws["A19"]
    c19.value = (
        "Required if you depreciate the vehicle (actual-expense method or "
        "Section 179). Hand these numbers to your CPA at year-end."
    )
    c19.font = italic_muted
    c19.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[19].height = 22

    f4562_rows = [
        (20, "Year-start odometer:",       None,                                       True,  "0"),
        (21, "Year-end odometer:",         None,                                       True,  "0"),
        (22, "Total miles (all use):",     "=IF(AND(B20<>\"\",B21<>\"\"),B21-B20,0)",  False, "0"),
        (23, "Total business miles (logged):", "=SUM('Mileage Log'!H6:H2005)",          False, "0"),
        (24, "Total commuting miles:",     None,                                       True,  "0"),
        (25, "Total personal miles:",      "=IF(B22>0,B22-B23-B24,0)",                 False, "0"),
        (26, "Business use %:",            '=IF(B22>0,B23/B22,0)',                      False, "0.0%"),
    ]
    for row, label, formula, is_input, fmt in f4562_rows:
        a = ws.cell(row=row, column=1, value=label)
        a.font = bold_right
        a.alignment = Alignment(horizontal="right", vertical="center")
        b = ws.cell(row=row, column=2, value=formula)
        if is_input:
            apply_style(b, input_cell_style())
        else:
            apply_style(b, formula_cell_style())
        b.number_format = fmt
        ws.row_dimensions[row].height = 16

    # v2.3.1 — Listed-property 50% rule alert (IRC §280F)
    # If business use ≤50%, no Section 179 / accelerated depreciation; prior-year
    # benefits may be subject to recapture as ordinary income. Conditional
    # formatting hot-spots B26 red when this threshold is breached.
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        "B26",
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["0.5"],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    ws.merge_cells("A27:G27")
    c27 = ws["A27"]
    c27.value = (
        "⚠ If Business Use % drops to 50% or below, IRS disallows Section 179 / "
        "accelerated depreciation and may recapture prior-year deductions as "
        "ordinary income (IRC §280F)."
    )
    c27.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c27.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[27].height = 28

    # --- Vehicles & Method Election (rows 28-33) ---
    sect28 = ws.cell(row=28, column=1, value="Vehicles & Method Election")
    sect28.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[28].height = 22

    ws.merge_cells("A29:G29")
    c29 = ws["A29"]
    c29.value = (
        "List each vehicle once. The Method (Standard mileage / Actual "
        "expenses) locks in Year 1 for owned cars; leased cars must keep "
        "Standard for the full lease (Pub 463)."
    )
    c29.font = italic_muted
    c29.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.row_dimensions[29].height = 28

    veh_headers = ["Nickname", "Year placed in service",
                    "Owned/Leased", "Method", "Notes"]
    for col, h in enumerate(veh_headers, start=1):
        cell = ws.cell(row=30, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[30].height = 18

    # 3 sample/blank vehicle rows
    sample_vehicles = [
        ("Truck (F-150)", 2026, "Owned",  "Standard mileage", "Primary STR vehicle"),
        ("",              "",   "",       "",                  ""),
        ("",              "",   "",       "",                  ""),
    ]
    for idx, vehicle in enumerate(sample_vehicles, start=31):
        for col, val in enumerate(vehicle, start=1):
            cell = ws.cell(row=idx, column=col,
                            value=val if val != "" else None)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # Dropdowns for Owned/Leased (col C) and Method (col D), rows 31-33
    add_dropdown(ws, "C31:C33", '"Owned,Leased"')
    add_dropdown(ws, "D31:D33", '"Standard mileage,Actual expenses"')

    ws.row_dimensions[34].height = 8

    # --- Year-end Archive (rows 35-43) ---
    sect35 = ws.cell(row=35, column=1, value="Year-end Archive")
    sect35.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[35].height = 22

    ws.merge_cells("A36:D36")
    c36 = ws["A36"]
    c36.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then clear the Mileage Log for the new year."
    )
    c36.font = italic_muted
    c36.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[36].height = 30

    archive_headers = ["Year", "Total Miles", "Total $", "# Trips"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=37, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[37].height = 18

    for idx, year in enumerate(range(2024, 2031), start=38):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=idx, column=3).number_format = '"$"#,##0.00'
        ws.row_dimensions[idx].height = 16

    # --- Frequent Destinations reference table (cols I-K, rows 7-26) ---
    # Customer fills these in once; uses for typical-miles lookup when
    # typing by hand. Keeps drives consistent for IRS audit defense.
    i7 = ws.cell(row=7, column=9, value="Frequent destinations:")
    i7.font = italic_muted
    ws.row_dimensions[7].height = 16

    for col, h in zip([9, 10, 11], ["From", "To", "Typical mi"]):
        cell = ws.cell(row=10, column=col, value=h)
        apply_style(cell, header_row_style())

    sample_destinations = [
        ("Smokies Ridge", "Home Depot",          24),
        ("Smokies Ridge", "Knoxville airport",   46),
        ("Smokies Ridge", "Lowe's",              22),
        ("Smokies Ridge", "Costco",              52),
        ("Creek Side",    "Home Depot",          18),
        ("Creek Side",    "Knoxville airport",   45),
        ("Lakehouse A",   "Home Depot",          14),
        ("Lakehouse A",   "Costco",              33),
    ]
    for idx, (frm, to, mi) in enumerate(sample_destinations, start=11):
        i_cell = ws.cell(row=idx, column=9, value=frm)
        apply_style(i_cell, input_cell_style())
        j_cell = ws.cell(row=idx, column=10, value=to)
        apply_style(j_cell, input_cell_style())
        k_cell = ws.cell(row=idx, column=11, value=mi)
        apply_style(k_cell, input_cell_style())
        k_cell.number_format = "0"
        ws.row_dimensions[idx].height = 16

    # 7 more empty rows for customer to extend (rows 19-25)
    for idx in range(19, 26):
        for col in [9, 10, 11]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_log_tab(wb, variant)
    build_monthly_tab(wb)
    build_ytd_tab(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"STR Mileage Log{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = "IRS-compliant mileage log for STR hosts (Schedule C/E)."

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
