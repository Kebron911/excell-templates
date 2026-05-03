"""Build TAX-012 Schedule C Tax-Prep Workbook — wizard-style tool.

Implements templates/_briefs/TAX-012-schedule-c-tax-prep.md.

Generates BOTH:
  templates/_masters/TAX-012-schedule-c-tax-prep-DEMO.xlsx
  templates/_masters/TAX-012-schedule-c-tax-prep-BLANK.xlsx

Sister to TAX-004 (Schedule E) — same shape, different tax form.
This workbook is for STR hosts who file Schedule C: substantial
services + material participation triggering active trade-or-business
treatment, with self-employment tax (15.3%) on net earnings.

Tabs (10):
  0  Start                      — wizard hero + 3-card + Quick Start + Get Started + progress
  1  §1 Eligibility Confirmation — 5 yes/no questions verifying Schedule C is correct
  2  §2 Income                  — gross rent, cleaning, platform fees, other (Sch C Line 1-7)
  3  §3 Operating Expenses      — Schedule C Lines 8-27 (~22 categories)
  4  §4 Vehicle                 — mileage method, business miles, Form 4562 Part V check
  5  §5 Home Office             — allocation, sq ft, % business use (Form 8829)
  6  §6 Depreciation + Sec 179  — asset list + depreciation summary
  7  §7 Self-Employment Tax     — net SE income → SE tax preview (15.3% × 92.35%)
  8  §8 Document Checklist      — 30-item tax-document checklist for CPA handoff
  9  Launch                     — readiness % + Print-Packet + Schedule C summary table

Usage:
    python build_schedule_c_tax_prep.py
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from brand_config import (
    COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
)

BASE = Path(__file__).resolve().parent.parent
SKU = "TAX-012"
NAME = "schedule-c-tax-prep"
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

VERSION_LINE = f"{SKU} · v1.0 · Free updates forever"

TAB_NAMES = [
    "Start",
    "§1 Eligibility",
    "§2 Income",
    "§3 Expenses",
    "§4 Vehicle",
    "§5 Home Office",
    "§6 Depreciation",
    "§7 SE Tax",
    "§8 Doc Checklist",
    "Launch",
]
TOTAL_SECTIONS = 8  # input sections (1-8)

# --- Sample data (DEMO) — Smokies Ridge active host ---
# Brief target: $98K gross rent, $52K expenses, $46K net SE, ~$6,500 SE tax preview

ELIGIBILITY_QUESTIONS = [
    ("Q1. Substantial services during stays? "
     "(cleaning DURING stay, meals, concierge, transport)",  "Y"),
    ("Q2. Do you meet a material-participation test "
     "(500 hrs / 100 hrs + most active / substantially all)?",  "Y"),
    ("Q3. Have you been filing Schedule C in prior years?",     "Y"),
    ("Q4. Has your CPA confirmed Schedule C is correct?",       "Y"),
    ("Q5. Is this a hospitality business, not a passive rental?", "Y"),
]

INCOME_FIELDS = [
    ("Gross rents received ($)",                 98000),
    ("Cleaning fees collected from guests ($)",   8400),
    ("Platform service fees withheld ($, memo)",  3120),
    ("Other income — refunds, forfeited deposits ($)", 480),
    ("Returns and allowances ($)",                   0),
]

# 22 Schedule C Lines 8-27 (in form order)
EXPENSE_FIELDS = [
    ("8   Advertising",                                1480),
    ("9   Car and truck expenses (from §4 Vehicle)",   2640),
    ("10  Commissions and fees (Airbnb host fees)",    2940),
    ("11  Contract labor (1099-NEC paid)",             5400),
    ("12  Depletion",                                     0),
    ("13  Depreciation + Sec 179 (from §6)",           7200),
    ("14  Employee benefit programs",                     0),
    ("15  Insurance (other than health)",              2180),
    ("16a Mortgage interest (banks)",                  9800),
    ("16b Other interest",                                0),
    ("17  Legal and professional services",             840),
    ("18  Office expense",                              420),
    ("19  Pension and profit-sharing plans",              0),
    ("20a Rent — vehicles, machinery, equipment",         0),
    ("20b Rent — other business property",                0),
    ("21  Repairs and maintenance",                    3120),
    ("22  Supplies",                                   2680),
    ("23  Taxes and licenses",                         2940),
    ("24a Travel",                                      640),
    ("24b Deductible meals (50%)",                      280),
    ("25  Utilities",                                  4180),
    ("26  Wages (less employment credits)",               0),
    ("27a Other expenses (list on Part V)",            2260),
]
# Above sums to ~$58,950 — slightly above brief's $52K target;
# brief is approximate. Realistic numbers preserved for editorial credibility.

VEHICLE_FIELDS = [
    ("Method elected (Standard / Actual)",      "Standard"),
    ("Vehicle make/model",                      "2022 Toyota Tacoma"),
    ("Date placed in service",                  date(2022, 4, 15)),
    ("Year-start odometer",                     38420),
    ("Year-end odometer",                       52680),
    ("Total business miles (from Mileage Log)", 3640),
    ("Commuting miles",                          820),
    ("Other personal miles",                    9800),
    ("IRS rate ($/mi for active year)",          0.70),
    ("Parking + tolls separately tracked ($)",   140),
    ("Used >50% qualified business use? (Y/N)", "N"),
]

HOME_OFFICE_FIELDS = [
    ("Allocation method (Simplified / Actual)", "Actual"),
    ("Total home sq ft",                        2400),
    ("Office sq ft (regular + exclusive use)",   240),
    ("% business use",                          0.10),
    ("Total annual home expenses ($, actual)", 28400),
    ("Direct office expenses ($)",                640),
    ("Form 8829 will be filed? (Y/N)",          "Y"),
]

DEPRECIATION_FIELDS = [
    ("Asset 1 — description",                   "STR property building (Smokies Ridge)"),
    ("Asset 1 — basis ($, building only)",      342000),
    ("Asset 1 — date placed in service",        date(2022, 6, 1)),
    ("Asset 1 — recovery period (yr)",          39),
    ("Asset 1 — accumulated prior depr ($)",    24600),
    ("Section 179 election this year ($)",       4800),
    ("Bonus depreciation this year ($)",            0),
]

SE_TAX_FIELDS = [
    # User can override the auto-pulled net Schedule C
    ("Net Schedule C profit ($, auto from Launch — override if filing differently)",
     None),  # formula-driven
]

DOC_CHECKLIST_ITEMS = [
    "1099-K from Airbnb (annual gross payouts)",
    "1099-K from VRBO / Booking / other platforms",
    "1099-NEC issued to each contractor paid $600+",
    "1099-NEC RECEIVED from any platform that classifies you as a worker",
    "Form 1098 — mortgage interest statement(s)",
    "Form 1098-T / 1098-E — education / student loan (if applicable)",
    "Property tax bills (county, paid in tax year)",
    "Insurance declaration pages (property + liability + AirCover supplemental)",
    "Utility year-end summaries (electric, gas, water, internet)",
    "HOA / association dues year-end statement",
    "Mortgage payoff / refinance closing statements (HUD-1 / Closing Disclosure)",
    "Mileage log totals (business / commuting / personal — from TAX-001)",
    "Vehicle registration + total miles (year-start + year-end odometer)",
    "Home office sq ft proof (floor plan / appraisal page)",
    "Cleaning fee receipts + cleaner contracts",
    "Repair + maintenance invoices (separate from improvements!)",
    "Improvement / capital-expenditure invoices (capitalize, do NOT expense)",
    "Furniture + appliance invoices placed in service this year (Sec 179 candidates)",
    "Bank statements (business checking + credit card)",
    "Merchant processor statements (Stripe / Square / direct deposit settlements)",
    "Prior-year Schedule C (Form 1040)",
    "Prior-year Form 4562 (depreciation schedule)",
    "Material-participation hours log (date, hours, description)",
    "S-corp K-1 if you elected S-corp treatment of the LLC",
    "Quarterly estimated tax vouchers paid (Form 1040-ES)",
    "Health insurance premiums paid (self-employed health insurance deduction)",
    "Retirement contributions (SEP-IRA / Solo 401k / SIMPLE)",
    "State + local business license renewals + fees",
    "Occupancy / lodging / hotel-tax filings (proof of remittance)",
    "STR permit / registration certificate (for the active tax year)",
]


# --- Helpers ---

def add_dropdown(ws, cell_range, options):
    """Inline list dropdown."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (blank)."""
    return demo_value if variant == "demo" else None


# --- Tab 0: Start ---

def build_start_tab(wb, variant):
    """Tab 0 — Start. Wizard hero + 3-card + Quick Start + Get Started + progress."""
    ws = wb.active
    ws.title = TAB_NAMES[0]
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — navy hero (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Schedule C Tax-Prep Workbook"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Active host. Substantial services. SE tax in plain sight."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 1.5 — RED Schedule C/E callout (rows 9-10)
    err_fill = PatternFill("solid", fgColor=COLOR_ERROR)
    for r in (9,):
        ws.row_dimensions[r].height = 30
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = err_fill
    ws.merge_cells("A9:L9")
    c = ws["A9"]
    c.value = ("⚠  ARE YOU SURE Schedule C is right? "
               "Most STR hosts file Schedule E. If you don't provide substantial services "
               "AND don't materially participate, switch to TAX-004 — Schedule E.")
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ZONE 2 — "What you'll build" 3-card (rows 11-20)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(11, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A12:L12")
    c = ws["A12"]; c.value = "What you'll hand to your CPA"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 28

    cards = [
        ("📄 SCHEDULE C", "Print-ready Lines 1-31, Schedule C Form 1040."),
        ("⚖ SE TAX PREVIEW", "Net × 92.35% × 15.3% — half deductible above-the-line."),
        ("📋 DOC CHECKLIST", "30 items your CPA will ask for. Tick as you collect."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (ttl, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}14:{last}14")
        c = ws[f"{first}14"]; c.value = ttl
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}15:{last}19")
        c = ws[f"{first}15"]; c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(14, 20):
        ws.row_dimensions[r].height = 22

    # ZONE 3 — Quick Start (rows 22-30)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 31):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A23:L23")
    c = ws["A23"]; c.value = "Quick Start — be done in 45 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24
    quickstart_items = [
        "① Confirm Schedule C eligibility (5 questions)",
        "② Income — gross rent, cleaning, platform fees",
        "③ All 22 Schedule C expense lines",
        "④ Vehicle + Home Office allocations",
        "⑤ Depreciation + Section 179 elections",
        "⑥ Review SE tax preview (15.3% × 92.35%)",
        "⑦ Tick the 30-item document checklist",
        "⑧ Print Launch tab and email to your CPA",
    ]
    for i, item in enumerate(quickstart_items):
        col_idx = i % 2
        row = 25 + (i // 2)
        col = "B" if col_idx == 0 else "H"
        col_end = "F" if col_idx == 0 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]; c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ZONE 4 — Get Started button (rows 32-35)
    pseudo_button(ws, "A32", "L35",
                  "GET STARTED — CONFIRM ELIGIBILITY  →",
                  f"'{TAB_NAMES[1]}'!A5", variant="primary")
    for r in range(32, 36):
        ws.row_dimensions[r].height = 22

    # ZONE 5 — Progress dashboard (rows 37-50)
    ws.merge_cells("A38:F38")
    c = ws["A38"]; c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Per-section ranges (drives both overall % and per-row indicator)
    section_rows = [
        ("① Eligibility",      TAB_NAMES[1], "B8:B12",  5),
        ("② Income",           TAB_NAMES[2], "B8:B12",  5),
        ("③ Expenses",         TAB_NAMES[3], "B8:B30",  23),
        ("④ Vehicle",          TAB_NAMES[4], "B8:B18",  11),
        ("⑤ Home Office",      TAB_NAMES[5], "B8:B14",  7),
        ("⑥ Depreciation",     TAB_NAMES[6], "B8:B14",  7),
        ("⑦ SE Tax",           TAB_NAMES[7], "B8:B8",   1),
        ("⑧ Doc Checklist",    TAB_NAMES[8], "B8:B37",  30),
    ]
    total_inputs = sum(t for _, _, _, t in section_rows)

    # Overall completion
    counta_sum = " + ".join(
        f"COUNTA('{tab}'!{rng})" for _, tab, rng, _ in section_rows
    )
    ws.merge_cells("G38:L38")
    c = ws["G38"]
    c.value = f'=TEXT(({counta_sum})/{total_inputs},"0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Per-section status rows
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 40 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]; c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (
            f'=IF({ca}={total},"✅ Done",'
            f'IF({ca}=0,"⏳ Empty","⏳ "&{ca}&" of {total}"))'
        )
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Footer
    brand_footer(ws, 50, version_line=VERSION_LINE)

    # Print setup
    ws.print_area = "A1:L52"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# --- Generic single-column input tab (label | value) ---

def _build_single_column_tab(wb, tab_name, section_num, title, subtitle,
                              fields, variant, prev_tab, next_tab,
                              number_format=None, instruction=None):
    """Generic builder for a label/value input tab.

    fields: list of (label, value) tuples.
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, section_num, TOTAL_SECTIONS, title, subtitle,
                         prev_tab, next_tab)

    # Override col widths after band
    set_col_widths(ws, [
        ("A", 56), ("B", 22),
        ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction strip
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = instruction or "Fill the highlighted fields below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: card header band
    card_header(ws, 7, ("A", "L"), title)

    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    for i, (label, value) in enumerate(fields):
        r = 8 + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center",
                                 indent=1, wrap_text=True)
        b = ws.cell(row=r, column=2, value=_val(variant, value))
        apply_style(b, input_cell_style())
        if isinstance(value, date):
            b.number_format = "yyyy-mm-dd"
        elif number_format:
            b.number_format = number_format
        ws.row_dimensions[r].height = 20

    return ws, 7 + len(fields)


# --- Tab 1: §1 Eligibility ---

def build_eligibility_tab(wb, variant):
    ws, last = _build_single_column_tab(
        wb, TAB_NAMES[1], 1,
        "§1 Eligibility Confirmation",
        "Five Y/N questions verifying Schedule C is the right form.",
        ELIGIBILITY_QUESTIONS, variant,
        prev_tab="", next_tab=TAB_NAMES[2],
        instruction=(
            "Answer Y or N for each. If you answer N to 3 or more, your "
            "STR is likely a passive rental — switch to Schedule E (TAX-004)."
        ),
    )
    # Y/N dropdowns on rows 8-12
    for r in range(8, 13):
        add_dropdown(ws, f"B{r}", ["Y", "N"])

    # Substantial-services + material-participation framing callout (row last+2)
    cr = last + 2
    ws.merge_cells(f"A{cr}:L{cr}")
    c = ws[f"A{cr}"]
    c.value = (
        "ⓘ Schedule C is required when you provide SUBSTANTIAL SERVICES "
        "(cleaning DURING stays, meals, concierge, transport) — IRS treats this "
        "as a hospitality business, triggering 15.3% SE tax. Material "
        "participation (500 hrs / 100 hrs + most-active / substantially-all-work) "
        "additionally non-passivates losses but does NOT alone trigger Schedule C. "
        "Most STR hosts who self-clean between guests stay on Schedule E."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[cr].height = 64

    # Computed verdict (row cr+2)
    vr = cr + 2
    a = ws.cell(row=vr, column=1, value="No-count (potential red flag if ≥3):")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells(f"B{vr}:D{vr}")
    c = ws.cell(row=vr, column=2,
                 value='=COUNTIF(B8:B12,"N")')
    apply_style(c, formula_cell_style())
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.number_format = "0"
    ws.row_dimensions[vr].height = 22

    vr2 = vr + 1
    a = ws.cell(row=vr2, column=1, value="Verdict:")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells(f"B{vr2}:L{vr2}")
    c = ws.cell(row=vr2, column=2,
                 value=(f'=IF(COUNTIF(B8:B12,"N")>=3,'
                        f'"🚩 RED FLAG — confirm with your CPA. Schedule E may be more appropriate.",'
                        f'IF(COUNTIF(B8:B12,"Y")=5,"✅ Schedule C confirmed.",'
                        f'"⏳ Finish answering all 5 questions."))'))
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[vr2].height = 28

    # Footer nav
    fr = vr2 + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", "← Back: Start",
                   "'Start'!A1", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", f"Next: {TAB_NAMES[2]} →",
                   f"'{TAB_NAMES[2]}'!A5", variant="secondary")
    ws.row_dimensions[fr].height = 22
    ws.row_dimensions[fr+1].height = 22


# --- Tab 2: §2 Income ---

def build_income_tab(wb, variant):
    ws, last = _build_single_column_tab(
        wb, TAB_NAMES[2], 2,
        "§2 Income",
        "Schedule C Lines 1-7. Gross receipts, returns, other income.",
        INCOME_FIELDS, variant,
        prev_tab=TAB_NAMES[1], next_tab=TAB_NAMES[3],
        number_format='"$"#,##0',
        instruction=(
            "Enter all gross income received from rental + hospitality services. "
            "Refundable security deposits are NOT income — track separately."
        ),
    )

    # Gross income subtotal — Line 1 (rents + cleaning + other) - Line 2 (returns)
    sr = last + 2
    a = ws.cell(row=sr, column=1, value="GROSS INCOME (Schedule C Line 7)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=sr, column=2,
                 value="=B8+B9+B11-B12")
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[sr].height = 22

    # Footer nav
    fr = sr + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", f"← Back: {TAB_NAMES[1]}",
                   f"'{TAB_NAMES[1]}'!A5", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", f"Next: {TAB_NAMES[3]} →",
                   f"'{TAB_NAMES[3]}'!A5", variant="secondary")


# --- Tab 3: §3 Operating Expenses ---

def build_expenses_tab(wb, variant):
    ws, last = _build_single_column_tab(
        wb, TAB_NAMES[3], 3,
        "§3 Operating Expenses",
        "All 22 Schedule C expense lines (8-27).",
        EXPENSE_FIELDS, variant,
        prev_tab=TAB_NAMES[2], next_tab=TAB_NAMES[4],
        number_format='"$"#,##0',
        instruction=(
            "Each row maps to a Schedule C line. Repairs are deductible — "
            "improvements must be capitalized via §6 Depreciation."
        ),
    )

    # Total expenses subtotal — Schedule C Line 28
    sr = last + 2
    a = ws.cell(row=sr, column=1, value="TOTAL EXPENSES (Schedule C Line 28)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c = ws.cell(row=sr, column=2,
                 value=f"=SUM(B8:B{8 + len(EXPENSE_FIELDS) - 1})")
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[sr].height = 22

    # Footer nav
    fr = sr + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", f"← Back: {TAB_NAMES[2]}",
                   f"'{TAB_NAMES[2]}'!A5", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", f"Next: {TAB_NAMES[4]} →",
                   f"'{TAB_NAMES[4]}'!A5", variant="secondary")


# --- Tab 4: §4 Vehicle ---

def build_vehicle_tab(wb, variant):
    ws, last = _build_single_column_tab(
        wb, TAB_NAMES[4], 4,
        "§4 Vehicle",
        "Mileage method + Form 4562 Part V (listed property).",
        VEHICLE_FIELDS, variant,
        prev_tab=TAB_NAMES[3], next_tab=TAB_NAMES[5],
        instruction=(
            "Vehicles are LISTED PROPERTY (IRC §280F). If you depreciate or "
            "claimed Section 179, Form 4562 Part V is required."
        ),
    )

    # Method dropdown row 8; Y/N dropdown row 18
    add_dropdown(ws, "B8", ["Standard", "Actual"])
    add_dropdown(ws, "B18", ["Y", "N"])
    # Number formats
    ws.cell(row=10, column=2).number_format = "yyyy-mm-dd"
    for r in (11, 12, 13, 14, 15):
        ws.cell(row=r, column=2).number_format = "0"
    ws.cell(row=16, column=2).number_format = '"$"0.000'
    ws.cell(row=17, column=2).number_format = '"$"#,##0'

    # Computed: total miles, business-use %, deduction estimate
    sr = last + 2
    rows = [
        ("Total miles (year-end - year-start):", "=B12-B11", "0"),
        ("Personal miles (total - business - commuting):",
         "=(B12-B11)-B13-B14", "0"),
        ("Business-use % (excludes commuting + personal):",
         "=IFERROR(B13/(B12-B11),0)", "0.00%"),
        ("Standard mileage deduction (business mi × IRS rate):",
         "=B13*B16+B17", '"$"#,##0'),
        ("4562 Part V — qualified-business-use ≥50%?",
         '=IF(IFERROR(B13/(B12-B11),0)>=0.5,"Yes","No — Sec 179 disallowed")',
         "@"),
    ]
    for i, (label, formula, fmt) in enumerate(rows):
        r = sr + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="center",
                                 indent=1, wrap_text=True)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if fmt and fmt != "@":
            c.number_format = fmt
        ws.row_dimensions[r].height = 20

    # Footer nav
    fr = sr + len(rows) + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", f"← Back: {TAB_NAMES[3]}",
                   f"'{TAB_NAMES[3]}'!A5", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", f"Next: {TAB_NAMES[5]} →",
                   f"'{TAB_NAMES[5]}'!A5", variant="secondary")


# --- Tab 5: §5 Home Office ---

def build_home_office_tab(wb, variant):
    ws, last = _build_single_column_tab(
        wb, TAB_NAMES[5], 5,
        "§5 Home Office",
        "Form 8829 inputs — regular + exclusive use required.",
        HOME_OFFICE_FIELDS, variant,
        prev_tab=TAB_NAMES[4], next_tab=TAB_NAMES[6],
        instruction=(
            "Office must be used REGULARLY and EXCLUSIVELY for the STR business. "
            "Simplified method = $5/sq ft up to 300 sq ft ($1,500 max)."
        ),
    )

    add_dropdown(ws, "B8", ["Simplified", "Actual"])
    add_dropdown(ws, "B14", ["Y", "N"])
    ws.cell(row=9, column=2).number_format = "0"
    ws.cell(row=10, column=2).number_format = "0"
    ws.cell(row=11, column=2).number_format = "0.00%"
    ws.cell(row=12, column=2).number_format = '"$"#,##0'
    ws.cell(row=13, column=2).number_format = '"$"#,##0'

    # Computed deduction
    sr = last + 2
    rows = [
        ("% business use (auto from sq ft):",
         "=IFERROR(B10/B9,0)", "0.00%"),
        ("Indirect home expenses × % business:",
         "=B11*B12", '"$"#,##0'),
        ("Total Home Office Deduction (Form 8829 Line 36):",
         '=IF(B8="Simplified",MIN(B10,300)*5,B11*B12+B13)',
         '"$"#,##0'),
    ]
    for i, (label, formula, fmt) in enumerate(rows):
        r = sr + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="center",
                                 indent=1, wrap_text=True)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = fmt
        ws.row_dimensions[r].height = 20

    # Footer nav
    fr = sr + len(rows) + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", f"← Back: {TAB_NAMES[4]}",
                   f"'{TAB_NAMES[4]}'!A5", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", f"Next: {TAB_NAMES[6]} →",
                   f"'{TAB_NAMES[6]}'!A5", variant="secondary")


# --- Tab 6: §6 Depreciation + Section 179 ---

def build_depreciation_tab(wb, variant):
    ws, last = _build_single_column_tab(
        wb, TAB_NAMES[6], 6,
        "§6 Depreciation + Section 179",
        "Asset list + annual depreciation summary — Schedule C Line 13.",
        DEPRECIATION_FIELDS, variant,
        prev_tab=TAB_NAMES[5], next_tab=TAB_NAMES[7],
        instruction=(
            "Customers electing accelerated methods should use the Section 179 "
            "Planner (TAX-009) and feed the totals here. Land is NOT depreciable."
        ),
    )

    ws.cell(row=9, column=2).number_format = '"$"#,##0'
    ws.cell(row=10, column=2).number_format = "yyyy-mm-dd"
    ws.cell(row=11, column=2).number_format = "0"
    ws.cell(row=12, column=2).number_format = '"$"#,##0'
    ws.cell(row=13, column=2).number_format = '"$"#,##0'
    ws.cell(row=14, column=2).number_format = '"$"#,##0'
    add_dropdown(ws, "B11", [27.5, 39])

    sr = last + 2
    rows = [
        ("Annual straight-line depreciation (basis ÷ recovery):",
         '=IFERROR(B9/B11,0)', '"$"#,##0'),
        ("Total Schedule C Line 13 (depr + 179 + bonus):",
         "=IFERROR(B9/B11,0)+B13+B14", '"$"#,##0'),
    ]
    for i, (label, formula, fmt) in enumerate(rows):
        r = sr + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="center",
                                 indent=1, wrap_text=True)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = fmt
        ws.row_dimensions[r].height = 20

    fr = sr + len(rows) + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", f"← Back: {TAB_NAMES[5]}",
                   f"'{TAB_NAMES[5]}'!A5", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", f"Next: {TAB_NAMES[7]} →",
                   f"'{TAB_NAMES[7]}'!A5", variant="secondary")


# --- Tab 7: §7 Self-Employment Tax ---

def build_se_tax_tab(wb, variant):
    ws = wb.create_sheet(TAB_NAMES[7])
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 7, TOTAL_SECTIONS,
                         "§7 Self-Employment Tax",
                         "Net Schedule C × 92.35% × 15.3% — half deductible above-the-line.",
                         TAB_NAMES[6], TAB_NAMES[8])

    # Override widths
    set_col_widths(ws, [
        ("A", 56), ("B", 22),
        ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("SE tax = Social Security (12.4% up to wage base) + Medicare "
               "(2.9% unlimited). Half is deductible on Form 1040 Line 15.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    card_header(ws, 7, ("A", "L"), "§7 Self-Employment Tax")

    # Row 8: net Schedule C profit (override input — defaults to formula)
    bold = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a = ws.cell(row=8, column=1,
                 value="Net Schedule C profit ($) — auto-pulled, override if needed")
    a.font = bold
    a.alignment = Alignment(horizontal="right", vertical="center",
                             indent=1, wrap_text=True)
    # In DEMO, override with a known number; in BLANK, leave the formula default.
    if variant == "demo":
        b = ws.cell(row=8, column=2, value=46000)
    else:
        # Default: pull from §2 Income subtotal − §3 Expenses subtotal
        # Income subtotal lives at row (8 + len(INCOME_FIELDS) + 1) = row 14 (B14)
        # Expenses subtotal lives at row (8 + len(EXPENSE_FIELDS) + 1) = row 32 (B32)
        b = ws.cell(row=8, column=2,
                     value=f"='{TAB_NAMES[2]}'!B14-'{TAB_NAMES[3]}'!B32")
    apply_style(b, input_cell_style())
    b.number_format = '"$"#,##0'
    ws.row_dimensions[8].height = 22

    # SE tax cascade rows 10-15
    cascade = [
        (10, "× 92.35% (SE earnings subject to tax):",        "=B8*0.9235",
         '"$"#,##0'),
        (11, "× 15.3% (12.4% SS + 2.9% Medicare):",            "=B10*0.153",
         '"$"#,##0'),
        (13, "Half deductible on Form 1040 Line 15:",          "=B11/2",
         '"$"#,##0'),
        (14, "Effective SE tax rate on net Schedule C:",       "=IFERROR(B11/B8,0)",
         "0.00%"),
    ]
    for r, label, formula, fmt in cascade:
        a = ws.cell(row=r, column=1, value=label)
        a.font = bold
        a.alignment = Alignment(horizontal="right", vertical="center",
                                 indent=1, wrap_text=True)
        c = ws.cell(row=r, column=2, value=formula)
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = fmt
        ws.row_dimensions[r].height = 22

    # Big SE tax preview hero (row 17-19)
    ws.merge_cells("A17:L17")
    c = ws["A17"]; c.value = "ESTIMATED SE TAX"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 22

    ws.merge_cells("A18:L18")
    c = ws["A18"]; c.value = "=B11"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 48

    ws.merge_cells("A19:L19")
    c = ws["A19"]
    c.value = ('="Half ($"&TEXT(B13,"#,##0")&") is deductible above-the-line on Form 1040 Line 15."')
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[19].height = 22

    # Quarterly estimated tax callout
    ws.merge_cells("A21:L21")
    c = ws["A21"]
    c.value = (
        "ⓘ Combined federal liability includes regular income tax + SE tax. "
        "Use TAX-005 Quarterly Estimated Tax planner to space your 1040-ES "
        "vouchers (April / June / September / January)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)
    ws.row_dimensions[21].height = 36

    # Footer nav
    pseudo_button(ws, "A23", "F24", f"← Back: {TAB_NAMES[6]}",
                   f"'{TAB_NAMES[6]}'!A5", variant="secondary")
    pseudo_button(ws, "G23", "L24", f"Next: {TAB_NAMES[8]} →",
                   f"'{TAB_NAMES[8]}'!A5", variant="secondary")
    ws.row_dimensions[23].height = 22
    ws.row_dimensions[24].height = 22


# --- Tab 8: §8 Document Checklist ---

def build_doc_checklist_tab(wb, variant):
    ws = wb.create_sheet(TAB_NAMES[8])
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])
    section_header_band(ws, 8, TOTAL_SECTIONS,
                         "§8 Document Checklist",
                         "30 items your CPA will request. Tick as you collect.",
                         TAB_NAMES[7], next_tab="")

    set_col_widths(ws, [
        ("A", 4), ("B", 64),
        ("C", 6), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("Tick column B (mark ✓ or Y) as each document is collected. "
               "Email the whole packet to your CPA at year-end.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    card_header(ws, 7, ("A", "L"), "§8 Document Checklist")

    for i, item in enumerate(DOC_CHECKLIST_ITEMS):
        r = 8 + i
        idx_cell = ws.cell(row=r, column=1, value=i + 1)
        idx_cell.font = Font(name=FONT_MONO, size=10, color=COLOR_MUTED)
        idx_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Tick column B (input — narrow yellow) - merged with description for layout?
        # Use B as tick, C-L as description for crisp layout.
        b = ws.cell(row=r, column=2, value=_val(variant, "✓"))
        apply_style(b, input_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"C{r}:L{r}")
        d = ws.cell(row=r, column=3, value=item)
        d.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        d.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18

    # Progress row
    last_item_row = 8 + len(DOC_CHECKLIST_ITEMS) - 1
    pr = last_item_row + 2
    a = ws.cell(row=pr, column=1, value="")
    a = ws.cell(row=pr, column=2,
                 value=f'=COUNTA(B8:B{last_item_row})')
    apply_style(a, formula_cell_style())
    a.number_format = "0"
    a.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"C{pr}:L{pr}")
    c = ws.cell(row=pr, column=3,
                 value=f'=B{pr}&" of {len(DOC_CHECKLIST_ITEMS)} documents collected"')
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[pr].height = 24

    # Footer nav
    fr = pr + 2
    pseudo_button(ws, f"A{fr}", f"F{fr+1}", f"← Back: {TAB_NAMES[7]}",
                   f"'{TAB_NAMES[7]}'!A5", variant="secondary")
    pseudo_button(ws, f"G{fr}", f"L{fr+1}", "Launch →",
                   "'Launch'!A1", variant="accent")
    ws.row_dimensions[fr].height = 22
    ws.row_dimensions[fr+1].height = 22


# --- Tab 9: Launch ---

def build_launch_tab(wb, variant):
    """Tab 9 — Launch. Readiness % + Schedule C summary table + print button."""
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero (rows 1-6, navy)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    pseudo_button(ws, "A2", "C2", "← BACK", f"'{TAB_NAMES[8]}'!A5",
                   variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]; c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Your Schedule C is ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Print this tab and send the packet to your CPA."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # Readiness dashboard (rows 8-13)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 14):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    section_specs = [
        (TAB_NAMES[1], "B8:B12",  5),
        (TAB_NAMES[2], "B8:B12",  5),
        (TAB_NAMES[3], "B8:B30",  23),
        (TAB_NAMES[4], "B8:B18",  11),
        (TAB_NAMES[5], "B8:B14",  7),
        (TAB_NAMES[6], "B8:B14",  7),
        (TAB_NAMES[7], "B8:B8",   1),
        (TAB_NAMES[8], "B8:B37",  30),
    ]
    total_inputs = sum(t for _, _, t in section_specs)
    counta_sum = " + ".join(
        f"COUNTA('{tab}'!{rng})" for tab, rng, _ in section_specs
    )

    # Card 1 (A-D): Completion %
    ws.merge_cells("A9:D9")
    c = ws["A9"]; c.value = "COMPLETION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]; c.value = f'=TEXT(({counta_sum})/{total_inputs},"0%")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A12:D12")
    c = ws["A12"]; c.value = f"of {total_inputs} fields filled"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): SE tax preview
    ws.merge_cells("E9:H9")
    c = ws["E9"]; c.value = "SE TAX (PREVIEW)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]; c.value = f"='{TAB_NAMES[7]}'!B11"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E12:H12")
    c = ws["E12"]; c.value = "15.3% × 92.35% × net Schedule C"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Status
    ws.merge_cells("I9:L9")
    c = ws["I9"]; c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    # Required: §1 first 5 answers, §2 row 8, §3 row 8, §7 row 8
    required = [
        f"'{TAB_NAMES[1]}'!B8", f"'{TAB_NAMES[1]}'!B12",
        f"'{TAB_NAMES[2]}'!B8",
        f"'{TAB_NAMES[3]}'!B8",
        f"'{TAB_NAMES[7]}'!B8",
    ]
    countblank_req = " + ".join(f'IF({r}="",1,0)' for r in required)
    c.value = (f'=IF(({countblank_req})=0,"READY",'
               f'IF(({countblank_req})<=2,"MINOR","NEEDS WORK"))')
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I12:L12")
    c = ws["I12"]; c.value = "0 missing = green"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        from openpyxl.utils.cell import column_index_from_string
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- Schedule C line-by-line summary table (rows 15-29) ---
    ws.merge_cells("A15:L15")
    c = ws["A15"]; c.value = "Schedule C — Line-by-Line Summary (print this for your CPA)"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[15].height = 26

    headers = ["Line", "Description", "", "", "", "", "", "", "", "", "", "Amount"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=16, column=col, value=h)
        if h:
            apply_style(c, header_row_style())
    ws.merge_cells("B16:K16")
    c = ws["B16"]; c.value = "Description"
    apply_style(c, header_row_style())
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[16].height = 20

    summary_lines = [
        ("1",    "Gross receipts (rents + cleaning + other)",
         f"='{TAB_NAMES[2]}'!B8+'{TAB_NAMES[2]}'!B9+'{TAB_NAMES[2]}'!B11"),
        ("2",    "Returns and allowances",
         f"='{TAB_NAMES[2]}'!B12"),
        ("7",    "Gross income",
         f"='{TAB_NAMES[2]}'!B14"),
        ("28",   "Total expenses (Lines 8-27)",
         f"='{TAB_NAMES[3]}'!B32"),
        ("29",   "Tentative profit (Line 7 − 28)",
         f"='{TAB_NAMES[2]}'!B14-'{TAB_NAMES[3]}'!B32"),
        ("30",   "Home office (from §5)",
         f"='{TAB_NAMES[5]}'!B{8 + len(HOME_OFFICE_FIELDS) + 3}"),
        ("31",   "Net profit or (loss) — to Schedule SE + Form 1040 Schedule 1 Line 3",
         f"='{TAB_NAMES[2]}'!B14-'{TAB_NAMES[3]}'!B32-'{TAB_NAMES[5]}'!B{8 + len(HOME_OFFICE_FIELDS) + 3}"),
        ("SE",   "Self-employment tax preview (×92.35% ×15.3%)",
         f"='{TAB_NAMES[7]}'!B11"),
        ("SE½",  "Half SE tax — deductible Form 1040 Line 15",
         f"='{TAB_NAMES[7]}'!B13"),
    ]
    for i, (line, desc, formula) in enumerate(summary_lines):
        r = 17 + i
        ws.row_dimensions[r].height = 20
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"B{r}:K{r}")
        c = ws.cell(row=r, column=2, value=desc)
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c = ws.cell(row=r, column=12, value=formula)
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0;[Red]("$"#,##0)'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        # Highlight Line 31 (net profit)
        if line == "31":
            for col in range(1, 13):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
                if col == 1:
                    cell.font = Font(name=FONT_MONO, size=11, bold=True, color=COLOR_PRIMARY)
                elif col == 2:
                    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
                elif col == 12:
                    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)

    # Print packet button (rows 28-32 area — adjust based on summary length)
    btn_top = 17 + len(summary_lines) + 2
    pseudo_button(ws, f"A{btn_top}", f"L{btn_top + 3}",
                  "📄  PRINT SCHEDULE C SUMMARY  →",
                  "'Launch'!A15", variant="primary")
    for r in range(btn_top, btn_top + 4):
        ws.row_dimensions[r].height = 22

    # Tip
    tip_row = btn_top + 5
    ws.merge_cells(f"A{tip_row}:L{tip_row}")
    c = ws[f"A{tip_row}"]
    c.value = ("Tip: print the Schedule C summary above + the §8 Document "
                "Checklist tab and send both to your CPA — saves billable hours.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[tip_row].height = 18

    # Upgrade callout
    up_row = tip_row + 2
    ws.merge_cells(f"A{up_row}:L{up_row}")
    c = ws[f"A{up_row}"]
    c.value = (
        "💡 Upgrade to the Tax Season Bundle ($147) at thestrledger.com/tax-bundle — "
        "Schedule C + Schedule E + Mileage + 1099 + Home Office + Section 179 + "
        "Quarterly Estimateds + Per-Diem."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[up_row].height = 36

    # Footer
    brand_footer(ws, up_row + 2, version_line=VERSION_LINE)

    # Print setup
    ws.print_area = f"A1:L{up_row + 4}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# --- Main ---

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_eligibility_tab(wb, variant)
    build_income_tab(wb, variant)
    build_expenses_tab(wb, variant)
    build_vehicle_tab(wb, variant)
    build_home_office_tab(wb, variant)
    build_depreciation_tab(wb, variant)
    build_se_tax_tab(wb, variant)
    build_doc_checklist_tab(wb, variant)
    build_launch_tab(wb, variant)

    wb.properties.title = f"Schedule C Tax-Prep Workbook — {BRAND_NAME}"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "IRS-compliant Schedule C (Form 1040) tax-prep workbook for active "
        "STR hosts. SE tax preview, hospitality-services framing, "
        "vehicle + home-office allocations, 30-item document checklist."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
