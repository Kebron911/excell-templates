"""Build OPS-004 Cleaning Cost per Turnover Tracker (v2.2 visual standard).

Implements templates/_briefs/OPS-004-cleaning-cost-per-turnover.md.
Operational mode (recurring data entry): Start hero + flat Log + Summary
rollup + Settings (active tax year, property/cleaner roster, year-end
archive). Mirrors the build_mileage_log.py pattern.

Generates:
  templates/_masters/OPS-004-cleaning-cost-per-turnover-DEMO.xlsx
  templates/_masters/OPS-004-cleaning-cost-per-turnover-BLANK.xlsx

Usage:
    python build_cleaning_cost_per_turnover.py
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, Reference

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    STATE_WARN_FILL,
)

SKU = "OPS-004"
NAME = "cleaning-cost-per-turnover"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# --- Constants ---

DEFAULT_HOURLY_RATE = 35.00
ACTIVE_TAX_YEAR = 2026

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]
CLEANERS = ["Jamie M.", "Carla R.", "Pete H.", "BrightHome Co."]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Sample rows 6-25 — 20 turnover events Jan-Mar 2026 across 3 properties
# Tuple: (date, property, cleaner, booking_id, hours, hourly_rate,
#         flat_fee, supplies, nights, notes)
TURNOVER_SAMPLES = [
    ("2026-01-04", "Smokies Ridge", "Jamie M.",      "ABNB-1041", 2.5, 35.0,  0,    18.50, 3,  "Standard turn"),
    ("2026-01-08", "Creek Side",    "Carla R.",      "ABNB-1042", 3.0, 35.0,  0,    24.00, 4,  "Hot tub deep clean"),
    ("2026-01-12", "Smokies Ridge", "Jamie M.",      "VRBO-2210", 2.0, 35.0,  0,    12.75, 2,  ""),
    ("2026-01-15", "Lakehouse A",   "BrightHome Co.","ABNB-1051",   0,    0,  120,  0,     5,  "Flat-fee contract"),
    ("2026-01-19", "Creek Side",    "Carla R.",      "ABNB-1059", 3.5, 35.0,  0,    32.00, 5,  "Pet hair extra"),
    ("2026-01-22", "Smokies Ridge", "Jamie M.",      "VRBO-2231", 2.5, 35.0,  0,    15.50, 3,  ""),
    ("2026-01-27", "Lakehouse A",   "BrightHome Co.","ABNB-1064",   0,    0,  120,  9.25,  4,  ""),
    ("2026-02-02", "Creek Side",    "Pete H.",       "ABNB-1071", 2.5, 40.0,  0,    18.00, 3,  "Carla unavailable"),
    ("2026-02-05", "Smokies Ridge", "Jamie M.",      "VRBO-2255", 2.0, 35.0,  0,    14.25, 2,  ""),
    ("2026-02-09", "Lakehouse A",   "BrightHome Co.","ABNB-1080",   0,    0,  130,  0,     6,  "Rate bump"),
    ("2026-02-13", "Creek Side",    "Carla R.",      "ABNB-1085", 3.0, 35.0,  0,    22.50, 4,  ""),
    ("2026-02-16", "Smokies Ridge", "Jamie M.",      "ABNB-1089", 2.5, 35.0,  0,    19.75, 3,  "Valentine's quick turn"),
    ("2026-02-20", "Lakehouse A",   "BrightHome Co.","VRBO-2280",   0,    0,  120,  0,     4,  ""),
    ("2026-02-24", "Creek Side",    "Carla R.",      "ABNB-1099", 3.5, 35.0,  0,    28.00, 5,  "Stained linens replaced"),
    ("2026-02-28", "Smokies Ridge", "Jamie M.",      "ABNB-1104", 2.0, 35.0,  0,    11.50, 2,  ""),
    ("2026-03-04", "Lakehouse A",   "BrightHome Co.","ABNB-1110",   0,    0,  120,  16.00, 5,  "Extra supplies bill"),
    ("2026-03-08", "Creek Side",    "Pete H.",       "VRBO-2305", 2.5, 40.0,  0,    18.00, 3,  ""),
    ("2026-03-12", "Smokies Ridge", "Jamie M.",      "ABNB-1119", 2.5, 35.0,  0,    16.50, 3,  ""),
    ("2026-03-17", "Lakehouse A",   "BrightHome Co.","ABNB-1128",   0,    0,  130,  0,     7,  "Long stay turn"),
    ("2026-03-22", "Creek Side",    "Carla R.",      "ABNB-1135", 3.0, 35.0,  0,    21.00, 4,  ""),
]


def add_dropdown(ws, cell_range, formula1):
    """Add a list data validation to the given cell range."""
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational hero + cards + KPI dashboard)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Cleaning Cost per Turnover"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Price the cleaning fee like you mean it."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.2 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Logs every turnover with hours, supplies, and flat fees so you "
        "see the real cost per turnover and per night of stay. The "
        "Summary tab rolls cost up by property and by month — feed those "
        "numbers into your cleaning-fee pricing each season so you stop "
        "subsidizing guests with thin margins."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Click the navy button below to open the Log — one row per turnover event.",
        "② Use Hours + Hourly Rate for hourly cleaners, OR Flat Fee for contracted services.",
        "③ Total $ auto-calculates as (Hours × Rate) + Flat Fee + Supplies.",
        "④ Enter Nights of stay just ended — gets you $ per night, the number Airbnb pricing cares about.",
        "⑤ Visit the Summary tab monthly to see per-property and per-month rollups.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "ADD A NEW TURNOVER" button (rows 26-29) ---
    pseudo_button(ws, "A26", "L29",
                   "→  ADD A NEW TURNOVER  (OPEN LOG)",
                   "'Log'!A6", variant="primary")
    for r in range(26, 30):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: KPI cards (rows 31-37) ---
    for r in range(31, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Total Turnovers YTD
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "TURNOVERS YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = "='Summary'!B6"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0"
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "events logged this year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Avg $/Turnover
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "AVG $/TURNOVER"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = "='Summary'!B8"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "true cost per event"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Last Entry
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "LAST ENTRY"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = '=IFERROR(TEXT(MAX(\'Log\'!A6:A305),"yyyy-mm-dd"),"—")'
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "most recent turnover date"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Gold borders around the 3 cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav (rows 39-40) ---
    pseudo_button(ws, "A39", "D40", "Summary",
                   "'Summary'!A1", variant="secondary")
    pseudo_button(ws, "E39", "H40", "📄 Print Packet",
                   "'Summary'!A1", variant="accent")
    pseudo_button(ws, "I39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Year-end callout (row 42) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📄 TAX TIME: print the Summary tab (File → Print → 'Print Active Sheets'). "
        "📅 YEAR END: copy YTD totals into Settings → Year-end Archive before Jan 1."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Upgrade banner (row 44) ---
    add_upgrade_banner(ws, 44)

    # --- ZONE 9: Footer (rows 46-48) ---
    brand_footer(ws, 46,
                 version_line=f"{SKU} · v2.2 · Free updates forever")

    # Print setup
    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_log_tab(wb, variant):
    """Sheet 1 — Log: one row per turnover, capacity 300 rows (6-305)."""
    ws = wb.create_sheet("Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Cleaning Log",
                         prev_tab="Start", next_tab="Summary")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per turnover — Hours×Rate or Flat Fee, plus supplies"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Row 5: headers
    headers = [
        "Date",                # A
        "Property",            # B
        "Cleaner",             # C
        "Booking ID",          # D
        "Hours",               # E
        "Hourly Rate",         # F
        "Flat Fee",            # G
        "Supplies $",          # H
        "Total $",             # I (formula)
        "Nights of stay",      # J
        "$ per night",         # K (formula)
        "Notes",               # L
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Col widths — fits 12 cols on landscape Letter
    set_col_widths(ws, [
        ("A", 11), ("B", 16), ("C", 16), ("D", 12),
        ("E", 8), ("F", 10), ("G", 10), ("H", 11),
        ("I", 11), ("J", 10), ("K", 11), ("L", 22),
    ])

    samples = TURNOVER_SAMPLES if variant == "demo" else []
    for i in range(6, 306):
        sample_idx = i - 6
        sample = samples[sample_idx] if sample_idx < len(samples) else None

        if sample:
            (date_val, prop, cleaner, booking_id, hours, hourly_rate,
             flat_fee, supplies, nights, notes) = sample

            date_obj = datetime.strptime(date_val, "%Y-%m-%d").date()
            a = ws.cell(row=i, column=1, value=date_obj)
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"

            b = ws.cell(row=i, column=2, value=prop)
            apply_style(b, input_cell_style())

            c = ws.cell(row=i, column=3, value=cleaner)
            apply_style(c, input_cell_style())

            d = ws.cell(row=i, column=4, value=booking_id)
            apply_style(d, input_cell_style())

            e = ws.cell(row=i, column=5, value=hours if hours else None)
            apply_style(e, input_cell_style())
            e.number_format = "0.00"

            f = ws.cell(row=i, column=6, value=hourly_rate if hourly_rate else None)
            apply_style(f, input_cell_style())
            f.number_format = '"$"#,##0.00'

            g = ws.cell(row=i, column=7, value=flat_fee if flat_fee else None)
            apply_style(g, input_cell_style())
            g.number_format = '"$"#,##0.00'

            h = ws.cell(row=i, column=8, value=supplies if supplies else None)
            apply_style(h, input_cell_style())
            h.number_format = '"$"#,##0.00'

            j = ws.cell(row=i, column=10, value=nights if nights else None)
            apply_style(j, input_cell_style())
            j.number_format = "0"

            l = ws.cell(row=i, column=12, value=notes if notes else None)
            apply_style(l, input_cell_style())
        else:
            # Empty rows still get formats set so user input looks right
            a = ws.cell(row=i, column=1)
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"
            for col_idx in [2, 3, 4]:
                cell = ws.cell(row=i, column=col_idx)
                apply_style(cell, input_cell_style())
            e = ws.cell(row=i, column=5)
            apply_style(e, input_cell_style())
            e.number_format = "0.00"
            for col_idx in [6, 7, 8]:
                cell = ws.cell(row=i, column=col_idx)
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0.00'
            j = ws.cell(row=i, column=10)
            apply_style(j, input_cell_style())
            j.number_format = "0"
            l = ws.cell(row=i, column=12)
            apply_style(l, input_cell_style())

        # I: Total $ = (E*F) + G + H, treating blanks as 0
        i_cell = ws.cell(
            row=i, column=9,
            value=(
                f"=IF(AND(E{i}=\"\",F{i}=\"\",G{i}=\"\",H{i}=\"\"),\"\","
                f"IFERROR(E{i}*F{i},0)+IFERROR(G{i},0)+IFERROR(H{i},0))"
            ),
        )
        i_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(i_cell, formula_cell_style())

        # K: $ per night = I/J, blank-safe
        k_cell = ws.cell(
            row=i, column=11,
            value=f'=IF(OR(J{i}="",J{i}=0,I{i}=""),"",I{i}/J{i})',
        )
        k_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(k_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    # Dropdowns
    add_dropdown(ws, "B6:B305", "=Settings!$B$9:$B$18")
    add_dropdown(ws, "C6:C305", "=Settings!$B$20:$B$29")

    # Conditional formatting: highlight unusually high $/turnover (>$150)
    ws.conditional_formatting.add(
        "I6:I305",
        FormulaRule(
            formula=['AND(I6<>"",I6>150)'],
            fill=PatternFill("solid", fgColor=STATE_WARN_FILL),
        ),
    )

    ws.freeze_panes = "A6"

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_summary_tab(wb):
    """Sheet 2 — Summary: KPIs + per-property + per-month + bar chart."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Summary · For Your CPA",
                         prev_tab="Log", next_tab="Settings")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 FOR YOUR CPA — print this tab. Per-property and per-month rollups."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 26), ("B", 14), ("C", 16), ("D", 14), ("E", 14),
    ])

    bold_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)

    # --- YTD KPIs (rows 5-9) ---
    # Year filter via Settings!$B$5 (active tax year), not YEAR(TODAY())
    kpis = [
        ("YTD Turnovers:",
         "=COUNTIFS('Log'!$A$6:$A$305,\">=\"&DATE(Settings!$B$5,1,1),"
         "'Log'!$A$6:$A$305,\"<\"&DATE(Settings!$B$5+1,1,1))",
         "0"),
        ("YTD Total Spent:",
         "=SUMIFS('Log'!$I$6:$I$305,'Log'!$A$6:$A$305,"
         "\">=\"&DATE(Settings!$B$5,1,1),'Log'!$A$6:$A$305,"
         "\"<\"&DATE(Settings!$B$5+1,1,1))",
         '"$"#,##0.00'),
        ("Avg $/Turnover:",
         "=IFERROR(B7/B6,0)",
         '"$"#,##0.00'),
        ("Avg $/Night:",
         "=IFERROR(SUMIFS('Log'!$I$6:$I$305,'Log'!$A$6:$A$305,"
         "\">=\"&DATE(Settings!$B$5,1,1),'Log'!$A$6:$A$305,"
         "\"<\"&DATE(Settings!$B$5+1,1,1)) / "
         "SUMIFS('Log'!$J$6:$J$305,'Log'!$A$6:$A$305,"
         "\">=\"&DATE(Settings!$B$5,1,1),'Log'!$A$6:$A$305,"
         "\"<\"&DATE(Settings!$B$5+1,1,1)),0)",
         '"$"#,##0.00'),
        ("Most Expensive Month:",
         "=IFERROR(INDEX(A23:A34,MATCH(MAX(C23:C34),C23:C34,0)),\"—\")",
         None),
    ]
    for i, (label, formula, fmt) in enumerate(kpis, start=6):
        a = ws.cell(row=i, column=1, value=label)
        a.font = bold_font
        a.alignment = Alignment(horizontal="right", vertical="center")
        b = ws.cell(row=i, column=2, value=formula)
        apply_style(b, formula_cell_style())
        if fmt:
            b.number_format = fmt
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[11].height = 8

    # --- Per-Property table (rows 12-18) ---
    hdr = ws.cell(row=12, column=1, value="By Property")
    hdr.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[12].height = 22

    prop_headers = ["Property", "Turnovers", "Total $", "Avg $/Turn"]
    for col, h in enumerate(prop_headers, start=1):
        cell = ws.cell(row=13, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[13].height = 18

    prop_start = 14
    for idx, prop in enumerate(PROPERTIES, start=prop_start):
        ws.cell(row=idx, column=1, value=prop).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        # Turnovers = COUNTIFS for property + active year
        b_cell = ws.cell(
            row=idx, column=2,
            value=(
                f"=COUNTIFS('Log'!$B$6:$B$305,A{idx},"
                f"'Log'!$A$6:$A$305,\">=\"&DATE(Settings!$B$5,1,1),"
                f"'Log'!$A$6:$A$305,\"<\"&DATE(Settings!$B$5+1,1,1))"
            ),
        )
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = "0"

        c_cell = ws.cell(
            row=idx, column=3,
            value=(
                f"=SUMIFS('Log'!$I$6:$I$305,'Log'!$B$6:$B$305,A{idx},"
                f"'Log'!$A$6:$A$305,\">=\"&DATE(Settings!$B$5,1,1),"
                f"'Log'!$A$6:$A$305,\"<\"&DATE(Settings!$B$5+1,1,1))"
            ),
        )
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = '"$"#,##0.00'

        d_cell = ws.cell(row=idx, column=4, value=f"=IFERROR(C{idx}/B{idx},0)")
        apply_style(d_cell, formula_cell_style())
        d_cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[idx].height = 16

    prop_end = prop_start + len(PROPERTIES) - 1

    ws.row_dimensions[prop_end + 1].height = 8

    # --- Per-Month table (rows 21-34) ---
    month_section_row = 21
    hdr_m = ws.cell(row=month_section_row, column=1, value="By Month")
    hdr_m.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[month_section_row].height = 22

    month_hdr_row = 22
    month_headers = ["Month", "Turnovers", "Total $", "Avg $/Turn"]
    for col, h in enumerate(month_headers, start=1):
        cell = ws.cell(row=month_hdr_row, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[month_hdr_row].height = 18

    month_start = 23
    for i in range(0, 12):
        row = month_start + i
        month_num = i + 1
        next_month = month_num + 1

        ws.cell(row=row, column=1, value=MONTHS[i]).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )

        if month_num < 12:
            count_formula = (
                f"=COUNTIFS('Log'!$A$6:$A$305,"
                f"\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Log'!$A$6:$A$305,"
                f"\"<\"&DATE(Settings!$B$5,{next_month},1))"
            )
            sum_formula = (
                f"=SUMIFS('Log'!$I$6:$I$305,'Log'!$A$6:$A$305,"
                f"\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Log'!$A$6:$A$305,"
                f"\"<\"&DATE(Settings!$B$5,{next_month},1))"
            )
        else:
            count_formula = (
                f"=COUNTIFS('Log'!$A$6:$A$305,"
                f"\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Log'!$A$6:$A$305,"
                f"\"<\"&DATE(Settings!$B$5+1,1,1))"
            )
            sum_formula = (
                f"=SUMIFS('Log'!$I$6:$I$305,'Log'!$A$6:$A$305,"
                f"\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Log'!$A$6:$A$305,"
                f"\"<\"&DATE(Settings!$B$5+1,1,1))"
            )

        b_cell = ws.cell(row=row, column=2, value=count_formula)
        apply_style(b_cell, formula_cell_style())
        b_cell.number_format = "0"

        c_cell = ws.cell(row=row, column=3, value=sum_formula)
        apply_style(c_cell, formula_cell_style())
        c_cell.number_format = '"$"#,##0.00'

        d_cell = ws.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)")
        apply_style(d_cell, formula_cell_style())
        d_cell.number_format = '"$"#,##0.00'

        ws.row_dimensions[row].height = 16

    # --- Bar chart: $ by Month, anchored F12 ---
    bar = BarChart()
    bar.type = "col"
    bar.title = "$ by Month"
    bar.y_axis.title = "Total $"
    bar.x_axis.title = "Month"
    bar.legend = None
    bar.height = 9
    bar.width = 16
    data_ref = Reference(ws, min_col=3, min_row=month_hdr_row,
                         max_row=month_start + 11, max_col=3)
    cats_ref = Reference(ws, min_col=1, min_row=month_start,
                         max_row=month_start + 11)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "low"
    style_chart(bar)
    ws.add_chart(bar, "F12")

    # Print setup
    ws.print_area = "A1:N36"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    """Sheet 3 — Settings: active year, default rate, property + cleaner
    rosters, year-end archive."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Summary", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · default rate · property/cleaner rosters · year-end archive"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 30), ("B", 22), ("C", 14), ("D", 14), ("E", 14), ("F", 14),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active tax year (row 5) ---
    a5 = ws.cell(row=5, column=1, value="Active tax year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=ACTIVE_TAX_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 20

    # Row 6: explainer
    ws.merge_cells("A6:F6")
    c6 = ws["A6"]
    c6.value = (
        "Drives all date-filtered SUMIFS on the Summary tab. Bump this each "
        "January so historical years stay archived but separate."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 22

    # --- Default hourly rate (row 7) ---
    a7 = ws.cell(row=7, column=1, value="Default hourly rate:")
    a7.font = bold_right
    a7.alignment = Alignment(horizontal="right", vertical="center")
    b7 = ws.cell(row=7, column=2, value=DEFAULT_HOURLY_RATE)
    apply_style(b7, input_cell_style())
    b7.number_format = '"$"#,##0.00'
    ws.row_dimensions[7].height = 18

    # --- Property roster (rows 9-18) ---
    a8 = ws.cell(row=8, column=1, value="Property roster (up to 10):")
    a8.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[8].height = 20

    for idx in range(9, 19):
        prop_idx = idx - 9
        val = PROPERTIES[prop_idx] if prop_idx < len(PROPERTIES) else None
        cell = ws.cell(row=idx, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # --- Cleaner roster (rows 19-29) ---
    a19 = ws.cell(row=19, column=1, value="Cleaner roster (up to 10):")
    a19.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[19].height = 20

    for idx in range(20, 30):
        cleaner_idx = idx - 20
        val = CLEANERS[cleaner_idx] if cleaner_idx < len(CLEANERS) else None
        cell = ws.cell(row=idx, column=2, value=val)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # --- Year-end archive (rows 38-50) ---
    ws.row_dimensions[37].height = 10
    sect = ws.cell(row=38, column=1, value="Year-end Archive")
    sect.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[38].height = 22

    ws.merge_cells("A39:F39")
    c39 = ws["A39"]
    c39.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then clear the Log for the new year."
    )
    c39.font = italic_muted
    c39.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[39].height = 24

    archive_headers = ["Year", "Turnovers", "Total $", "Avg $/Turn", "Avg $/Night"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=40, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[40].height = 18

    for idx, year in enumerate(range(2024, 2031), start=41):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
        ws.cell(row=idx, column=2).number_format = "0"
        ws.cell(row=idx, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=idx, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=idx, column=5).number_format = '"$"#,##0.00'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_log_tab(wb, variant)
    build_summary_tab(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Cleaning Cost per Turnover{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Track turnover cleaning cost — true cost per turnover and per night "
        "to inform cleaning-fee pricing."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
