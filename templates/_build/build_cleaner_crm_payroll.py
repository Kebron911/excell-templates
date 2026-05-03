"""Build PAM-002 Cleaner CRM + Payroll Excel file (v2.3 standard).

Operational-mode hybrid: register (Cleaners CRM) + log (Turnover Assignments)
+ derived dashboards (Payroll Run / Performance / 1099-NEC Year-End).

For Pro Pam — managing 10+ properties with a cleaning team. Combines cleaner
contact + capacity tracking, per-turnover assignment log, payroll calculator
(hourly + per-turnover + bonus), performance scorecard, and year-end 1099-NEC
totals that drop straight into TAX-003.

Dates are real datetime.date objects with yyyy-mm-dd format. Settings B5
holds the active tax year — formulas filter on it instead of YEAR(TODAY()).

Generates templates/_masters/PAM-002-cleaner-crm-payroll-{DEMO,BLANK}.xlsx.
"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    STATE_BAD_FILL, STATE_GOOD_FILL, STATE_WARN_FILL,
)

SKU = "PAM-002"
NAME = "cleaner-crm-payroll"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026

# --- Sample data (DEMO) ---

# 6 cleaners; 4 will cross the $600 1099 threshold based on volume below.
# Columns: ID, Name, Phone, Email, Address, EIN/SSN last4, W-9?,
#          Pay rate type, Hourly rate, Per-turnover flat, Service area,
#          Certifications, Active?, Avg rating (formula-overwritten), Notes
CLEANERS = [
    ("C-001", "Sarah Mitchell",   "(865) 555-0145", "sarah@smokiesclean.com",  "147 Pine St, Gatlinburg, TN 37738",   "4421", "Yes", "Per-turnover", 0,  120, "Gatlinburg, Pigeon Forge",       "Insured, bonded, COVID-19", "Y", None, "Top performer — keep on rotation"),
    ("C-002", "Maria Gonzalez",   "(865) 555-0192", "maria@cleanmtn.com",      "82 Ridge Ln, Pigeon Forge, TN 37863", "8810", "Yes", "Hourly",       28, 0,   "Pigeon Forge, Sevierville",      "Insured",                    "Y", None, "Reliable, slow on linens"),
    ("C-003", "Jenny Park",       "(865) 555-0211", "jenny@parkclean.com",     "9 Oak Dr, Sevierville, TN 37862",     "5573", "Yes", "Hybrid",       22, 90,  "Sevierville, Knoxville",         "Insured, bonded",            "Y", None, "Backup for Sarah"),
    ("C-004", "Tom Bradley",      "(865) 555-0166", "tom@bradleyturn.com",     "55 Lake Rd, Gatlinburg, TN 37738",    "2298", "No",  "Per-turnover", 0,  110, "Gatlinburg",                     "Bonded",                     "Y", None, "Need W-9 — request before next pay"),
    ("C-005", "Lisa Chen",        "(865) 555-0177", "lisa@chenclean.net",      "21 Cherry St, Knoxville, TN 37902",   "6634", "Yes", "Per-turnover", 0,  115, "Knoxville",                      "Insured, bonded, eco",       "Y", None, ""),
    ("C-006", "Backup Helper",    "(865) 555-0188", "—",                        "Various",                              "—",    "No",  "Hourly",       20, 0,   "Local only",                     "—",                          "N", None, "Inactive — used Q1 only"),
]

PROPERTIES = [
    "Smokies Ridge",
    "Creek Side",
    "Lakehouse A",
    "Mountain View",
    "Downtown Loft",
]


def _build_turnovers():
    """Build 240 turnover rows Jan-Mar 2026 across 3 properties / 6 cleaners.

    Distribution roughly:
      Sarah   : 65 turnovers (top performer, 4.9 avg rating)
      Maria   : 50 turnovers, hourly
      Jenny   : 45 turnovers, hybrid
      Tom     : 40 turnovers, per-turnover (no W-9 — flagged)
      Lisa    : 35 turnovers
      Backup  : 5 turnovers (Q1 fill-in)
    """
    rows = []
    start = date(ACTIVE_YEAR, 1, 1)
    # Build a deterministic schedule
    cleaner_schedule = [
        ("Sarah Mitchell",  65, 4.9, 3.5, "Per-turnover", 120, 0,   "Smokies Ridge"),
        ("Maria Gonzalez",  50, 4.6, 4.5, "Hourly",       0,   28,  "Creek Side"),
        ("Jenny Park",      45, 4.7, 4.0, "Hybrid",       90,  22,  "Lakehouse A"),
        ("Tom Bradley",     40, 4.3, 3.8, "Per-turnover", 110, 0,   "Smokies Ridge"),
        ("Lisa Chen",       35, 4.5, 4.0, "Per-turnover", 115, 0,   "Mountain View"),
        ("Backup Helper",   5,  4.0, 5.0, "Hourly",       0,   20,  "Downtown Loft"),
    ]
    day_offset = 0
    for cleaner, count, _avg_rating, avg_hours, _rate_type, flat, hourly, default_prop in cleaner_schedule:
        for i in range(count):
            d = start + timedelta(days=day_offset % 89)  # spread Jan-Mar
            day_offset += 1
            # Vary property
            prop = PROPERTIES[(i + day_offset) % 3]
            hours = round(avg_hours + ((i % 5) - 2) * 0.25, 2)
            bonus = 25 if (i % 12 == 0 and cleaner == "Sarah Mitchell") else 0
            # Inspection rating: jitter around avg
            rating = max(1, min(5, round(_avg_rating + ((i % 7) - 3) * 0.1, 1)))
            status = "Inspected" if i < count - 2 else "Done"
            rows.append((d, prop, cleaner, hours, flat, bonus, status, rating, ""))
    return rows


TURNOVERS = _build_turnovers()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (KPIs + nav cards)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Cleaner CRM + Payroll"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Your cleaning team — paid right, rated honestly, 1099-ready."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Once you cross 5+ properties, your cleaning team becomes a small business "
        "of its own. This workbook tracks every cleaner (CRM), logs every turnover "
        "(hours, flat fee, bonus, inspection rating), runs payroll in one click, "
        "and rolls totals into a 1099-NEC year-end summary that drops straight "
        "into TAX-003. Stop paying late, stop rehiring drama, and walk into "
        "January with clean books."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: KPI cards (rows 18-23) ---
    for r in range(18, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-C): Active cleaners
    ws.merge_cells("A18:C18")
    c = ws["A18"]
    c.value = "ACTIVE CLEANERS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A19:C20")
    c = ws["A19"]
    c.value = '=COUNTIF(Cleaners!M6:M55,"Y")'
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A21:C21")
    c = ws["A21"]
    c.value = "on roster"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (D-F): Turnovers YTD
    ws.merge_cells("D18:F18")
    c = ws["D18"]
    c.value = "TURNOVERS YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("D19:F20")
    c = ws["D19"]
    c.value = (
        f"=SUMPRODUCT((YEAR('Turnover Assignment'!A6:A1005)=Settings!$B$5)*"
        f"('Turnover Assignment'!A6:A1005<>\"\"))"
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D21:F21")
    c = ws["D21"]
    c.value = "in active year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (G-I): $ Paid YTD
    ws.merge_cells("G18:I18")
    c = ws["G18"]
    c.value = "$ PAID YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("G19:I20")
    c = ws["G19"]
    c.value = (
        f"=SUMPRODUCT((YEAR('Turnover Assignment'!A6:A1005)=Settings!$B$5)*"
        f"'Turnover Assignment'!K6:K1005)"
    )
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("G21:I21")
    c = ws["G21"]
    c.value = "all cleaner pay"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 4 (J-L): Top performer
    ws.merge_cells("J18:L18")
    c = ws["J18"]
    c.value = "TOP PERFORMER"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("J19:L20")
    c = ws["J19"]
    c.value = (
        '=IFERROR(INDEX(Performance!A6:A55,'
        'MATCH(MAX(Performance!C6:C55),Performance!C6:C55,0)),"—")'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("J21:L21")
    c = ws["J21"]
    c.value = "by avg rating"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(18, 22):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 18 else existing.top,
                    bottom=gold_side if r == 21 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 4: Primary "OPEN CLEANERS" button (rows 25-28) ---
    pseudo_button(ws, "A25", "L28",
                   "→  ADD A CLEANER  (OPEN CLEANERS CRM)",
                   "'Cleaners'!A6", variant="primary")
    for r in range(25, 29):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Secondary nav buttons (rows 30-31) ---
    pseudo_button(ws, "A30", "C31", "Turnover Assignment",
                   "'Turnover Assignment'!A1", variant="secondary")
    pseudo_button(ws, "D30", "F31", "Payroll Run",
                   "'Payroll Run'!A1", variant="secondary")
    pseudo_button(ws, "G30", "I31", "Performance",
                   "'Performance'!A1", variant="secondary")
    pseudo_button(ws, "J30", "L31", "1099-NEC Year-End",
                   "'1099-NEC Year-End'!A1", variant="accent")
    ws.row_dimensions[30].height = 22
    ws.row_dimensions[31].height = 22

    # --- ZONE 6: How to use callout (rows 33-39) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(33, 40):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[33].height = 20

    quickstart_items = [
        "① Cleaners tab: enter every team member — pay rate, W-9 status, certifications, service area.",
        "② Turnover Assignment: log every cleaning AS IT HAPPENS — hours, flat fee, bonus, rating.",
        "③ Payroll Run: set the period start/end → totals roll up per cleaner. Print and pay.",
        "④ Performance: review monthly — flag top performers, decide who to coach or cut.",
        "⑤ 1099-NEC Year-End: in January, copy totals into TAX-003 1099-NEC tracker.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 34 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 20

    # --- ZONE 7: $600 reminder ---
    ws.merge_cells("A41:L41")
    c = ws["A41"]
    c.value = (
        "⚠ Pay any cleaner $600+ in a calendar year and you owe them a 1099-NEC "
        "by January 31 (IRS Pub 1220, $60–$660 penalty per missed form). "
        "Year-End tab tracks who crosses the line."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[41].height = 32

    add_upgrade_banner(ws, 43)

    brand_footer(ws, 45,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_cleaners_tab(wb, variant):
    """Sheet 1 — Cleaners CRM (capacity 30, register pattern from TAX-003)."""
    ws = wb.create_sheet("Cleaners")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Cleaners CRM",
                         prev_tab="Start", next_tab="Turnover Assignment")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Add every cleaner — pay rate, W-9 status, certifications, service area"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 9),  ("B", 22), ("C", 16), ("D", 26), ("E", 32),
        ("F", 9),  ("G", 7),  ("H", 14), ("I", 11), ("J", 12),
        ("K", 24), ("L", 28), ("M", 8),  ("N", 11), ("O", 28),
    ])

    headers = [
        "Cleaner ID", "Name", "Phone", "Email", "Address",
        "EIN/SSN last4", "W-9?", "Pay rate type", "Hourly $", "Per-turnover $",
        "Service area", "Certifications", "Active?", "Avg rating", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # DEMO data fills first 6 rows; capacity to row 35 (30 slots)
    cleaner_rows = CLEANERS if variant == "demo" else []
    for row_idx, cleaner in enumerate(cleaner_rows, start=6):
        for col_idx, value in enumerate(cleaner, start=1):
            cell = ws.cell(row=row_idx, column=col_idx,
                            value=value if value not in ("", None) else None)
            apply_style(cell, input_cell_style())
            if col_idx == 9:  # hourly $
                cell.number_format = '"$"#,##0.00'
            if col_idx == 10:  # per-turnover $
                cell.number_format = '"$"#,##0.00'
            if col_idx == 14:  # avg rating — formula overwrite below
                pass
        # Avg rating formula pulls from Performance
        rating_cell = ws.cell(
            row=row_idx, column=14,
            value=(
                f'=IFERROR(VLOOKUP(B{row_idx},Performance!$A$6:$D$55,4,FALSE),"")'
            ),
        )
        apply_style(rating_cell, formula_cell_style())
        rating_cell.number_format = "0.00"
        rating_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    # Capacity rows
    blank_start = len(cleaner_rows) + 6
    for row_idx in range(blank_start, 36):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx in (9, 10):
                cell.number_format = '"$"#,##0.00'
        rating_cell = ws.cell(
            row=row_idx, column=14,
            value=(
                f'=IFERROR(VLOOKUP(B{row_idx},Performance!$A$6:$D$55,4,FALSE),"")'
            ),
        )
        apply_style(rating_cell, formula_cell_style())
        rating_cell.number_format = "0.00"
        rating_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    add_dropdown(ws, "G6:G35", '"Yes,No"')
    add_dropdown(ws, "H6:H35", '"Hourly,Per-turnover,Hybrid"')
    add_dropdown(ws, "M6:M35", '"Y,N"')

    # Conditional fmt — flag missing W-9
    ws.conditional_formatting.add(
        "G6:G35",
        FormulaRule(
            formula=['G6="No"'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    # Inactive cleaners — gray out name
    ws.conditional_formatting.add(
        "B6:B35",
        FormulaRule(
            formula=['$M6="N"'],
            font=Font(italic=True, color=COLOR_MUTED),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_turnover_log_tab(wb, variant):
    """Sheet 2 — Turnover Assignment log (capacity 1000)."""
    ws = wb.create_sheet("Turnover Assignment")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Turnover Assignment",
                         prev_tab="Cleaners", next_tab="Payroll Run")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Log every turnover AS IT HAPPENS — hours, flat fee, bonus, inspection rating"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 18), ("C", 22), ("D", 9), ("E", 12),
        ("F", 10), ("G", 13), ("H", 13), ("I", 14), ("J", 9), ("K", 12),
    ])

    headers = [
        "Date", "Property", "Cleaner", "Hours", "Per-turn $",
        "Bonus $", "Hourly $ (lkup)", "Status", "Inspection rating", "Notes", "Total payment",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    turnover_rows = TURNOVERS if variant == "demo" else []
    for i, t in enumerate(turnover_rows, start=6):
        d, prop, cleaner, hours, flat, bonus, status, rating, notes = t

        a = ws.cell(row=i, column=1, value=d)
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=prop)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=cleaner)
        apply_style(c, input_cell_style())

        d_cell = ws.cell(row=i, column=4, value=hours)
        apply_style(d_cell, input_cell_style())
        d_cell.number_format = "0.00"

        e = ws.cell(row=i, column=5, value=flat if flat else None)
        apply_style(e, input_cell_style())
        e.number_format = '"$"#,##0.00'

        f = ws.cell(row=i, column=6, value=bonus if bonus else None)
        apply_style(f, input_cell_style())
        f.number_format = '"$"#,##0.00'

        # G — hourly lookup from Cleaners (formula)
        g = ws.cell(
            row=i, column=7,
            value=(
                f'=IFERROR(VLOOKUP(C{i},Cleaners!$B$6:$I$35,8,FALSE),0)'
            ),
        )
        apply_style(g, formula_cell_style())
        g.number_format = '"$"#,##0.00'

        h = ws.cell(row=i, column=8, value=status)
        apply_style(h, input_cell_style())

        i_cell = ws.cell(row=i, column=9, value=rating)
        apply_style(i_cell, input_cell_style())
        i_cell.alignment = Alignment(horizontal="center", vertical="center")
        i_cell.number_format = "0.0"

        j = ws.cell(row=i, column=10, value=notes if notes else None)
        apply_style(j, input_cell_style())

        # K — total payment formula (hours * hourly + flat + bonus)
        k = ws.cell(
            row=i, column=11,
            value=(
                f'=IF(C{i}="","",'
                f'IFERROR(D{i},0)*G{i}+IFERROR(E{i},0)+IFERROR(F{i},0))'
            ),
        )
        apply_style(k, formula_cell_style())
        k.number_format = '"$"#,##0.00'

        ws.row_dimensions[i].height = 16

    last_data_row = len(turnover_rows) + 5
    for row_idx in range(last_data_row + 1, 1006):
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx in (5, 6):
                cell.number_format = '"$"#,##0.00'
            if col_idx == 4:
                cell.number_format = "0.00"
            if col_idx == 9:
                cell.number_format = "0.0"
        # G — hourly lookup
        g = ws.cell(
            row=row_idx, column=7,
            value=(
                f'=IFERROR(VLOOKUP(C{row_idx},Cleaners!$B$6:$I$35,8,FALSE),0)'
            ),
        )
        apply_style(g, formula_cell_style())
        g.number_format = '"$"#,##0.00'
        # K — total payment
        k = ws.cell(
            row=row_idx, column=11,
            value=(
                f'=IF(C{row_idx}="","",'
                f'IFERROR(D{row_idx},0)*G{row_idx}+IFERROR(E{row_idx},0)+IFERROR(F{row_idx},0))'
            ),
        )
        apply_style(k, formula_cell_style())
        k.number_format = '"$"#,##0.00'

    add_dropdown(ws, "B6:B1005", "=Settings!$B$8:$B$17")
    add_dropdown(ws, "C6:C1005", "=Cleaners!$B$6:$B$35")
    add_dropdown(ws, "H6:H1005",
                  '"Scheduled,In progress,Done,Inspected"')

    # Status conditional fmt
    ws.conditional_formatting.add(
        "H6:H1005",
        FormulaRule(
            formula=['H6="Inspected"'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
        ),
    )
    ws.conditional_formatting.add(
        "H6:H1005",
        FormulaRule(
            formula=['H6="In progress"'],
            fill=PatternFill("solid", fgColor=STATE_WARN_FILL),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_payroll_run_tab(wb, variant):
    """Sheet 3 — Payroll Run (period-based payroll calc per cleaner)."""
    ws = wb.create_sheet("Payroll Run")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Payroll Run",
                         prev_tab="Turnover Assignment", next_tab="Performance")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Set period dates → totals roll up per cleaner. Print this tab to pay."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 24), ("B", 12), ("C", 11), ("D", 12),
        ("E", 12), ("F", 12), ("G", 12),
    ])

    # --- Period inputs (rows 5-7) ---
    label_p1 = ws.cell(row=5, column=1, value="Period start:")
    label_p1.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    label_p1.alignment = Alignment(horizontal="right", vertical="center")
    p1 = ws.cell(row=5, column=2,
                  value=date(ACTIVE_YEAR, 1, 1) if variant == "demo" else None)
    apply_style(p1, input_cell_style())
    p1.number_format = "yyyy-mm-dd"

    label_p2 = ws.cell(row=6, column=1, value="Period end:")
    label_p2.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    label_p2.alignment = Alignment(horizontal="right", vertical="center")
    p2 = ws.cell(row=6, column=2,
                  value=date(ACTIVE_YEAR, 1, 14) if variant == "demo" else None)
    apply_style(p2, input_cell_style())
    p2.number_format = "yyyy-mm-dd"

    ws.row_dimensions[7].height = 8

    # --- Header row 8 ---
    headers = [
        "Cleaner", "Turnovers", "Hours", "Hourly $",
        "Per-turn $", "Bonuses", "Gross pay",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[8].height = 20

    # --- Per-cleaner rows 9-38 (30 slots) ---
    for i in range(9, 39):
        cleaner_row = i - 3  # Cleaners row 6 -> row 9
        a = ws.cell(
            row=i, column=1,
            value=f'=IF(Cleaners!B{cleaner_row}="","",Cleaners!B{cleaner_row})',
        )
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")

        # B Turnovers in period
        b = ws.cell(
            row=i, column=2,
            value=(
                f'=IF(A{i}="","",'
                f'COUNTIFS(\'Turnover Assignment\'!$C$6:$C$1005,A{i},'
                f'\'Turnover Assignment\'!$A$6:$A$1005,">="&$B$5,'
                f'\'Turnover Assignment\'!$A$6:$A$1005,"<="&$B$6))'
            ),
        )
        apply_style(b, formula_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.number_format = "0"

        # C Hours
        c = ws.cell(
            row=i, column=3,
            value=(
                f'=IF(A{i}="","",'
                f'SUMIFS(\'Turnover Assignment\'!$D$6:$D$1005,'
                f'\'Turnover Assignment\'!$C$6:$C$1005,A{i},'
                f'\'Turnover Assignment\'!$A$6:$A$1005,">="&$B$5,'
                f'\'Turnover Assignment\'!$A$6:$A$1005,"<="&$B$6))'
            ),
        )
        apply_style(c, formula_cell_style())
        c.number_format = "0.00"

        # D Hourly $ earned
        d = ws.cell(
            row=i, column=4,
            value=(
                f'=IF(A{i}="","",'
                f'C{i}*IFERROR(VLOOKUP(A{i},Cleaners!$B$6:$I$35,8,FALSE),0))'
            ),
        )
        apply_style(d, formula_cell_style())
        d.number_format = '"$"#,##0.00'

        # E Per-turnover $ earned
        e = ws.cell(
            row=i, column=5,
            value=(
                f'=IF(A{i}="","",'
                f'SUMIFS(\'Turnover Assignment\'!$E$6:$E$1005,'
                f'\'Turnover Assignment\'!$C$6:$C$1005,A{i},'
                f'\'Turnover Assignment\'!$A$6:$A$1005,">="&$B$5,'
                f'\'Turnover Assignment\'!$A$6:$A$1005,"<="&$B$6))'
            ),
        )
        apply_style(e, formula_cell_style())
        e.number_format = '"$"#,##0.00'

        # F Bonuses
        f = ws.cell(
            row=i, column=6,
            value=(
                f'=IF(A{i}="","",'
                f'SUMIFS(\'Turnover Assignment\'!$F$6:$F$1005,'
                f'\'Turnover Assignment\'!$C$6:$C$1005,A{i},'
                f'\'Turnover Assignment\'!$A$6:$A$1005,">="&$B$5,'
                f'\'Turnover Assignment\'!$A$6:$A$1005,"<="&$B$6))'
            ),
        )
        apply_style(f, formula_cell_style())
        f.number_format = '"$"#,##0.00'

        # G Gross pay
        g = ws.cell(
            row=i, column=7,
            value=f'=IF(A{i}="","",D{i}+E{i}+F{i})',
        )
        apply_style(g, formula_cell_style())
        g.number_format = '"$"#,##0.00'
        g.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        ws.row_dimensions[i].height = 16

    # --- Totals row 40 ---
    ws.row_dimensions[39].height = 6
    tot_label = ws.cell(row=40, column=1, value="TOTALS")
    tot_label.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    tot_label.alignment = Alignment(horizontal="right", vertical="center")
    tot_label.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for col, formula in [
        (2, "=SUM(B9:B38)"),
        (3, "=SUM(C9:C38)"),
        (4, "=SUM(D9:D38)"),
        (5, "=SUM(E9:E38)"),
        (6, "=SUM(F9:F38)"),
        (7, "=SUM(G9:G38)"),
    ]:
        cell = ws.cell(row=40, column=col, value=formula)
        cell.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if col == 2:
            cell.number_format = "0"
        elif col == 3:
            cell.number_format = "0.00"
        else:
            cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[40].height = 22

    ws.freeze_panes = "A9"

    ws.print_area = "A1:G42"
    ws.print_title_rows = "1:8"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_performance_tab(wb):
    """Sheet 4 — Performance scorecard per cleaner."""
    ws = wb.create_sheet("Performance")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Performance",
                         prev_tab="Payroll Run", next_tab="1099-NEC Year-End")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Per-cleaner scorecard — review monthly to coach, flag, or cut"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 22), ("B", 14), ("C", 13), ("D", 14),
        ("E", 13), ("F", 22),
    ])

    headers = [
        "Cleaner", "Turnovers YTD", "Avg rating", "Avg hrs/turn",
        "Complaints", "Performance flag",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    for i in range(6, 56):
        cleaner_row = i  # Cleaners row 6 -> row 6
        a = ws.cell(
            row=i, column=1,
            value=f'=IF(Cleaners!B{cleaner_row}="","",Cleaners!B{cleaner_row})',
        )
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")

        # B Turnovers YTD — filter on Settings active year
        b = ws.cell(
            row=i, column=2,
            value=(
                f'=IF(A{i}="","",'
                f'SUMPRODUCT((\'Turnover Assignment\'!$C$6:$C$1005=A{i})*'
                f'(YEAR(\'Turnover Assignment\'!$A$6:$A$1005)=Settings!$B$5)))'
            ),
        )
        apply_style(b, formula_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.number_format = "0"

        # C Avg rating
        c = ws.cell(
            row=i, column=3,
            value=(
                f'=IF(A{i}="","",'
                f'IFERROR(AVERAGEIFS(\'Turnover Assignment\'!$I$6:$I$1005,'
                f'\'Turnover Assignment\'!$C$6:$C$1005,A{i},'
                f'\'Turnover Assignment\'!$I$6:$I$1005,">0"),""))'
            ),
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.00"

        # D Avg hrs/turn
        d = ws.cell(
            row=i, column=4,
            value=(
                f'=IF(A{i}="","",'
                f'IFERROR(AVERAGEIFS(\'Turnover Assignment\'!$D$6:$D$1005,'
                f'\'Turnover Assignment\'!$C$6:$C$1005,A{i},'
                f'\'Turnover Assignment\'!$D$6:$D$1005,">0"),""))'
            ),
        )
        apply_style(d, formula_cell_style())
        d.number_format = "0.00"

        # E Complaints — manual input
        e = ws.cell(row=i, column=5)
        apply_style(e, input_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")
        e.number_format = "0"

        # F Performance flag
        f = ws.cell(
            row=i, column=6,
            value=(
                f'=IF(A{i}="","",'
                f'IF(C{i}="","Insufficient data",'
                f'IF(AND(C{i}>=4.7,IFERROR(E{i},0)<=1),"Top performer",'
                f'IF(OR(C{i}<4.0,IFERROR(E{i},0)>=3),"Needs review","Standard"))))'
            ),
        )
        apply_style(f, formula_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 16

    # Conditional fmt
    ws.conditional_formatting.add(
        "F6:F55",
        FormulaRule(
            formula=['F6="Top performer"'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
            font=Font(bold=True, color=COLOR_PRIMARY),
        ),
    )
    ws.conditional_formatting.add(
        "F6:F55",
        FormulaRule(
            formula=['F6="Needs review"'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    # --- Bar chart: turnovers per cleaner (anchored H5) ---
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Turnovers YTD by Cleaner"
    chart.height = 9
    chart.width = 14
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=5, max_row=20, max_col=2)
    cats = Reference(ws, min_col=1, min_row=6, max_row=20)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    style_chart(chart)
    ws.add_chart(chart, "H5")

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_year_end_tab(wb):
    """Sheet 5 — 1099-NEC Year-End summary (drops into TAX-003)."""
    ws = wb.create_sheet("1099-NEC Year-End")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "1099-NEC Year-End · For Your CPA",
                         prev_tab="Performance", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 Year-end totals per cleaner — drops directly into TAX-003 1099-NEC tracker"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 22), ("B", 14), ("C", 18), ("D", 14),
        ("E", 18), ("F", 22),
    ])

    headers = [
        "Cleaner", "Total paid YTD", "1099 required?", "W-9 on file?",
        "Mailing address ready?", "Status",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    for i in range(6, 36):  # 30-row capacity
        cleaner_row = i  # Cleaners row 6 -> row 6
        a = ws.cell(
            row=i, column=1,
            value=f'=IF(Cleaners!B{cleaner_row}="","",Cleaners!B{cleaner_row})',
        )
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")

        # B Total paid YTD — filter on active year
        b = ws.cell(
            row=i, column=2,
            value=(
                f'=IF(A{i}="","",'
                f'SUMPRODUCT((\'Turnover Assignment\'!$C$6:$C$1005=A{i})*'
                f'(YEAR(\'Turnover Assignment\'!$A$6:$A$1005)=Settings!$B$5)*'
                f'\'Turnover Assignment\'!$K$6:$K$1005))'
            ),
        )
        apply_style(b, formula_cell_style())
        b.number_format = '"$"#,##0.00'

        # C 1099 required? — threshold from Settings B20
        c = ws.cell(
            row=i, column=3,
            value=(
                f'=IF(A{i}="","",'
                f'IF(B{i}>=Settings!$B$20,"✓ Required","Below threshold"))'
            ),
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        # D W-9 on file? — VLOOKUP Cleaners col G (W-9?)
        d = ws.cell(
            row=i, column=4,
            value=(
                f'=IF(A{i}="","",'
                f'IFERROR(VLOOKUP(A{i},Cleaners!$B$6:$G$35,6,FALSE),"?"))'
            ),
        )
        apply_style(d, formula_cell_style())
        d.alignment = Alignment(horizontal="center", vertical="center")

        # E Mailing address ready? — Cleaners col E (Address)
        e = ws.cell(
            row=i, column=5,
            value=(
                f'=IF(A{i}="","",'
                f'IF(IFERROR(VLOOKUP(A{i},Cleaners!$B$6:$E$35,4,FALSE),"")="","⚠ Missing","✓ On file"))'
            ),
        )
        apply_style(e, formula_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")

        # F Status
        f = ws.cell(
            row=i, column=6,
            value=(
                f'=IF(A{i}="","",'
                f'IF(C{i}="✓ Required",'
                f'IF(AND(D{i}="Yes",LEFT(E{i},1)="✓"),"✓ Ready to file","⚠ Missing W-9 or address"),'
                f'"n/a"))'
            ),
        )
        apply_style(f, formula_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 16

    # Conditional fmt
    ws.conditional_formatting.add(
        "C6:C35",
        FormulaRule(
            formula=['C6="✓ Required"'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True),
        ),
    )
    ws.conditional_formatting.add(
        "F6:F35",
        FormulaRule(
            formula=['F6="✓ Ready to file"'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
        ),
    )
    ws.conditional_formatting.add(
        "F6:F35",
        FormulaRule(
            formula=['LEFT(F6,1)="⚠"'],
            fill=PatternFill("solid", fgColor=STATE_WARN_FILL),
        ),
    )

    # Summary block (rows 38-42)
    ws.row_dimensions[37].height = 8
    hdr = ws.cell(row=38, column=1, value="Summary")
    hdr.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[38].height = 22

    summary_rows = [
        (39, "Cleaners requiring 1099:", '=COUNTIF(C6:C35,"✓ Required")', "0"),
        (40, "Of those, ready to file:", '=COUNTIF(F6:F35,"✓ Ready to file")', "0"),
        (41, "Of those, missing W-9 or address:",
            '=COUNTIF(C6:C35,"✓ Required")-COUNTIF(F6:F35,"✓ Ready to file")', "0"),
        (42, "Total 1099-NEC $ volume:",
            '=SUMIFS(B6:B35,C6:C35,"✓ Required")', '"$"#,##0.00'),
    ]
    for row_num, label, formula, num_fmt in summary_rows:
        a = ws.cell(row=row_num, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center")

        b = ws.cell(row=row_num, column=2, value=formula)
        b.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        b.number_format = num_fmt
        b.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_num].height = 18

    # Donut chart: $ paid by cleaner (anchored H5)
    donut = DoughnutChart()
    donut.title = "$ Paid by Cleaner (YTD)"
    donut.height = 9
    donut.width = 11
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=2, min_row=5, max_row=20, max_col=2)
    donut_cats = Reference(ws, min_col=1, min_row=6, max_row=20)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=False, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, "H5")

    ws.freeze_panes = "A6"

    ws.print_area = "A1:F44"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 6 — Settings (active year, property list, cadence, threshold)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="1099-NEC Year-End", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active year · property list · pay cadence · 1099 threshold"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 36), ("B", 22), ("C", 14), ("D", 16)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active tax year (B5) — referenced by all date filters ---
    a5 = ws.cell(row=5, column=1, value="Active tax year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    # Row 6: explainer
    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = (
        "All YTD formulas filter on this year. Update once each January after "
        "you archive the prior year's data."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 24

    # --- Property list (B8:B17, 10 slots) ---
    a7 = ws.cell(row=7, column=1, value="Properties (used in Turnover dropdown):")
    a7.font = bold_right
    a7.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[7].height = 18

    prop_list = PROPERTIES if variant == "demo" else []
    for i in range(8, 18):
        cell = ws.cell(row=i, column=2)
        if i - 8 < len(prop_list):
            cell.value = prop_list[i - 8]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[i].height = 16

    # --- Pay cadence (B18) ---
    a18 = ws.cell(row=18, column=1, value="Payment cadence:")
    a18.font = bold_right
    a18.alignment = Alignment(horizontal="right", vertical="center")
    b18 = ws.cell(row=18, column=2, value="Bi-weekly")
    apply_style(b18, input_cell_style())
    add_dropdown(ws, "B18", '"Weekly,Bi-weekly,Monthly"')
    ws.row_dimensions[18].height = 18

    ws.row_dimensions[19].height = 8

    # --- 1099 threshold (B20) ---
    a20 = ws.cell(row=20, column=1, value="1099-NEC threshold ($):")
    a20.font = bold_right
    a20.alignment = Alignment(horizontal="right", vertical="center")
    b20 = ws.cell(row=20, column=2, value=600)
    apply_style(b20, input_cell_style())
    b20.number_format = '"$"#,##0'
    ws.row_dimensions[20].height = 18

    # Row 21: explainer
    ws.merge_cells("A21:D21")
    c21 = ws["A21"]
    c21.value = (
        "$600 is locked by IRS rule. Don't change unless the IRS publishes a "
        "new threshold (rare) — Year-End formulas filter on this cell."
    )
    c21.font = italic_muted
    c21.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[21].height = 28

    ws.row_dimensions[22].height = 10

    # --- Year-end archive (rows 23+) ---
    sect = ws.cell(row=23, column=1, value="Year-end Archive")
    sect.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[23].height = 22

    ws.merge_cells("A24:D24")
    c24 = ws["A24"]
    c24.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "then clear the Turnover Assignment log for the new year (keep Cleaners)."
    )
    c24.font = italic_muted
    c24.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[24].height = 28

    archive_headers = ["Year", "Active cleaners", "Turnovers", "Total $ paid"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=25, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[25].height = 18

    for idx, year in enumerate(range(2024, 2031), start=26):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0"
        cell = ws.cell(row=idx, column=4)
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_cleaners_tab(wb, variant)
    build_turnover_log_tab(wb, variant)
    build_payroll_run_tab(wb, variant)
    build_performance_tab(wb)
    build_year_end_tab(wb)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Cleaner CRM + Payroll{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Cleaner CRM + per-turnover log + payroll calc + performance + "
        "1099-NEC year-end summary for STR hosts with cleaning teams (v2.3)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
