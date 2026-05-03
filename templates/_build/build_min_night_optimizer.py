"""Build REV-004 Minimum-Night-Stay Optimizer (v2.2 standard).

Compares 5 MNS policies (1n, 2n, 3n, 5n, 7n) for the same property on the
same page: gross revenue, turnover costs, host labor, NET, NET/night.
Highlights the winning column gold-soft via FormulaRule. Bar chart of NET
revenue by MNS scenario.

Generates:
  templates/_masters/REV-004-min-night-stay-optimizer-DEMO.xlsx
  templates/_masters/REV-004-min-night-stay-optimizer-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "REV-004"
NAME = "min-night-stay-optimizer"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# ---------------------------------------------------------------------------
# Sample data (DEMO) — Smokies Ridge baseline per brief
# ---------------------------------------------------------------------------

SAMPLE = {
    "property_name":   "Smokies Ridge Cabin",
    "active_year":     2026,
    "base_rate":       185,
    "cleaning_cost":   90,
    "supply_restock":  25,
    "host_hours":      1.5,
    "host_hourly":     40,
    # Per scenario: (mns_nights, occupancy, adr)
    # 1n / 2n / 3n / 5n / 7n
    "scenarios": [
        (1, 0.78, 185),
        (2, 0.76, 185),
        (3, 0.70, 192),
        (5, 0.58, 200),
        (7, 0.45, 210),
    ],
}

# Scenario column layout on Comparison tab — C, D, E, F, G
SCEN_COLS = ["C", "D", "E", "F", "G"]
SCEN_LABELS = ["1-night MNS", "2-night MNS", "3-night MNS",
               "5-night MNS", "7-night MNS"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (BLANK)."""
    return demo_value if variant == "demo" else None


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note=""):
    """Standard label-in-B / input-in-C / note-in-E:L row."""
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — Hero band rows 1-9
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 10):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Minimum-Night-Stay Optimizer"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Does 1-night really beat 3-night? Run the math."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — The question + recommendation block rows 11-22
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(11, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Settings!C5="","(set property on Settings tab)",Settings!C5)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # Question prompt
    ws.merge_cells("A13:L14")
    c = ws["A13"]
    c.value = (
        "Most hosts assume a 1-night minimum maximizes revenue.\n"
        "Turnover cost often makes 3-night MNS net better. Let's see."
    )
    c.font = Font(name=FONT_BODY, size=12, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    ws.row_dimensions[13].height = 22
    ws.row_dimensions[14].height = 22

    # Recommended MNS readout — pulled from Comparison tab
    ws.merge_cells("A16:L16")
    c = ws["A16"]
    c.value = "RECOMMENDED MINIMUM-NIGHT-STAY"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 18

    ws.merge_cells("A17:L18")
    c = ws["A17"]
    # Comparison!C24:G24 holds NET revenue per scenario; row 7 holds MNS labels
    c.value = (
        '=INDEX(Comparison!C7:G7,MATCH(MAX(Comparison!C24:G24),'
        'Comparison!C24:G24,0))&"-night minimum"'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 32
    ws.row_dimensions[18].height = 22

    ws.merge_cells("A20:L20")
    c = ws["A20"]
    c.value = (
        '="Annual NET: "&TEXT(MAX(Comparison!C24:G24),"$#,##0")'
        '&"   ·   vs 1-night NET: "&TEXT(Comparison!C24,"$#,##0")'
    )
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[20].height = 22

    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = (
        '="Lift over 1-night minimum: "&'
        'TEXT(MAX(Comparison!C24:G24)-Comparison!C24,"+$#,##0;-$#,##0;$0")&'
        '" / year"'
    )
    c.font = Font(name=FONT_BODY, size=12, italic=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 20

    # ZONE 3 — Pseudo-button nav rows 25-27
    pseudo_button(ws, "A25", "D27", "→  Inputs",
                  "'Inputs'!A1", variant="primary")
    pseudo_button(ws, "E25", "H27", "→  Run Comparison",
                  "'Comparison'!A1", variant="accent")
    pseudo_button(ws, "I25", "L27", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    # ZONE 4 — Upgrade banner row 30
    ws.merge_cells("A30:L30")
    c = ws["A30"]
    c.value = (
        "Optimizing pricing across scenarios? Get the Pricing Bundle at "
        f"{BRAND_DOMAIN}/pricing — cleaning fee + MNS + base-rate together."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    ws.row_dimensions[30].height = 36

    brand_footer(ws, 32,
                 version_line=f"{SKU} · v1.0 · Free updates forever")

    ws.print_area = "A1:L34"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_inputs_tab(wb, variant):
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 40), ("C", 16),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Inputs",
                        prev_tab="Start", next_tab="Comparison")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Yellow = edit. Set base costs once, then tune per-scenario "
        "occupancy + ADR by MNS."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- BASE PROPERTY (rows 6-13) ----
    _section_band(ws, 6, "BASE PROPERTY  &  COSTS")

    base_fields = [
        (7,  "Base nightly rate ($):",
         _val(variant, SAMPLE["base_rate"]),       '"$"#,##0',
         "Reference rate — per-scenario ADR overrides below"),
        (8,  "Cleaning cost per turnover ($):",
         _val(variant, SAMPLE["cleaning_cost"]),   '"$"#,##0',
         "What you pay your cleaner each turnover"),
        (9,  "Supply restock per turnover ($):",
         _val(variant, SAMPLE["supply_restock"]),  '"$"#,##0',
         "Toiletries, paper, coffee, consumables"),
        (10, "Host hours per turnover:",
         _val(variant, SAMPLE["host_hours"]),      "0.0",
         "Coordination, inspection, supply runs"),
        (11, "Host hourly opportunity cost ($):",
         _val(variant, SAMPLE["host_hourly"]),     '"$"#,##0',
         "What your time is worth — be honest"),
    ]
    for row, label, value, fmt, note in base_fields:
        _input_row(ws, row, label, value, fmt, note)

    # ---- PER-SCENARIO INPUTS (rows 15-22) ----
    _section_band(ws, 15, "PER-SCENARIO  —  OCCUPANCY  &  ADR BY MNS")

    # Header row 16: Scenario | Occupancy | ADR | (note)
    ws.cell(row=16, column=2, value="MNS scenario")
    ws.cell(row=16, column=3, value="Expected occ.")
    ws.cell(row=16, column=4, value="Expected ADR")
    ws.merge_cells("E16:L16")
    ws.cell(row=16, column=5, value="Note")
    for col in range(2, 6):
        cell = ws.cell(row=16, column=col)
        apply_style(cell, header_row_style())
    ws.row_dimensions[16].height = 24

    # Rows 17-21: 5 scenarios
    notes_per_row = [
        "1-night = max nightly availability, max turnovers",
        "Filters out single-night party bookers",
        "Sweet spot for most STR markets",
        "Long-weekend bookings only",
        "Week-long stays — lowest ops cost, lowest occupancy",
    ]
    for i, (mns, occ, adr) in enumerate(SAMPLE["scenarios"]):
        r = 17 + i
        # Label
        cell = ws.cell(row=r, column=2, value=f"{mns}-night MNS")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        # Occupancy input (col C)
        cell = ws.cell(row=r, column=3, value=_val(variant, occ))
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.0%"
        # ADR input (col D)
        cell = ws.cell(row=r, column=4, value=_val(variant, adr))
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0'
        # Note (E:L merged) — set value BEFORE merge to avoid touching merged cells
        ws.cell(row=r, column=5, value=notes_per_row[i]).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=r, column=5).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.merge_cells(f"E{r}:L{r}")
        ws.row_dimensions[r].height = 20

    # Footnote 24
    ws.merge_cells("A24:L26")
    c = ws["A24"]
    c.value = (
        "Why ADR rises with longer MNS: longer minimums filter to "
        "vacationers (less price-sensitive) and avoid Airbnb's length-of-stay "
        "discount territory. Conservative estimate: +$5-15/night per "
        "additional MNS step. Pull real numbers from PriceLabs/Wheelhouse."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(24, 27):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, 28, version_line=f"{SKU} · Inputs")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_comparison_tab(wb, variant):
    """5 scenarios side-by-side. Cols C/D/E/F/G = 1n/2n/3n/5n/7n."""
    ws = wb.create_sheet("Comparison")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14),
        ("H", 4),
        ("I", 8), ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Comparison",
                        prev_tab="Inputs", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Five MNS policies side-by-side. Gold-soft column = highest NET. "
        "Numbers update from Inputs."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- HEADER row 6 ----
    ws.cell(row=6, column=2, value="Metric")
    apply_style(ws.cell(row=6, column=2), header_row_style())
    for i, label in enumerate(SCEN_LABELS):
        col = 3 + i  # C..G
        cell = ws.cell(row=6, column=col, value=label)
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 26

    # ---- ROW 7: MNS as numeric (used by formulas + chart category labels) ----
    ws.cell(row=7, column=2, value="Min nights per booking").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=7, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    for i, (mns, _, _) in enumerate(SAMPLE["scenarios"]):
        col = 3 + i
        cell = ws.cell(row=7, column=col, value=mns)
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    ws.row_dimensions[7].height = 22

    # Inputs cell map for formula reuse:
    #   Inputs!C7  = base nightly rate (unused below — ADR overrides)
    #   Inputs!C8  = cleaning cost per turnover
    #   Inputs!C9  = supply restock per turnover
    #   Inputs!C10 = host hours per turnover
    #   Inputs!C11 = host hourly cost
    # Per-scenario Inputs rows 17..21:
    #   Inputs!C{17+i} = occupancy
    #   Inputs!D{17+i} = ADR

    # Helper to place a formula row keyed off scenario index.
    # IMPORTANT: never touches merged cells.
    def _formula_row(ws, row, label, formula_template, fmt, bold=False,
                     fill_emphasize=False):
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(
            name=FONT_HEAD if bold else FONT_BODY,
            size=12 if bold else 11,
            bold=True,
            color=COLOR_PRIMARY if bold else COLOR_TEXT,
        )
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        for i in range(5):
            col = 3 + i
            input_row = 17 + i  # row on Inputs tab
            formula = formula_template.format(ir=input_row, sc=SCEN_COLS[i])
            c = ws.cell(row=row, column=col, value=formula)
            apply_style(c, formula_cell_style())
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = fmt
            if bold:
                c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
            if fill_emphasize:
                c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

        ws.row_dimensions[row].height = 22 if bold else 18

    # ---- DERIVED ROWS ----
    # Avg nights per booking = MNS (row 7 numeric value at the column itself)
    # Bookings/year = 365 × occ / avg_nights
    # Turnovers/year = bookings/year (each booking = 1 turnover for simplicity)
    # Gross revenue = bookings × avg_nights × ADR  ==  365 × occ × ADR
    # Cleaning cost = turnovers × Inputs!C8
    # Supply cost   = turnovers × Inputs!C9
    # Host labor    = turnovers × Inputs!C10 × Inputs!C11
    # NET = Gross - cleaning - supply - host_labor
    # NET/night = NET / (365 × occ)

    # Row 8 — avg nights per booking (= MNS)
    _formula_row(ws, 8, "Avg nights per booking",
                 "={sc}7", "0.0")

    # Row 9 — bookings / year
    _formula_row(ws, 9, "Bookings / year",
                 "=IFERROR((365*Inputs!C{ir})/{sc}7,0)", "0")

    # Row 10 — turnovers / year (= bookings/year for 1-booking-per-stay model)
    _formula_row(ws, 10, "Turnovers / year",
                 "={sc}9", "0")

    # Row 11 — occupied nights / year
    _formula_row(ws, 11, "Occupied nights / year",
                 "=365*Inputs!C{ir}", "0")

    # Row 13 — Gross revenue  (emphasize section)
    _section_band(ws, 12, "REVENUE  &  COSTS")
    _formula_row(ws, 13, "Gross revenue",
                 "=365*Inputs!C{ir}*Inputs!D{ir}", '"$"#,##0')

    # Row 14 — Cleaning cost total
    _formula_row(ws, 14, "Cleaning cost (total)",
                 "={sc}10*Inputs!$C$8", '"$"#,##0')

    # Row 15 — Supply restock cost
    _formula_row(ws, 15, "Supply restock cost",
                 "={sc}10*Inputs!$C$9", '"$"#,##0')

    # Row 16 — Host labor cost
    _formula_row(ws, 16, "Host labor cost",
                 "={sc}10*Inputs!$C$10*Inputs!$C$11", '"$"#,##0')

    # Row 17 — Total operating costs
    _formula_row(ws, 17, "Total operating costs",
                 "=SUM({sc}14:{sc}16)", '"$"#,##0',
                 fill_emphasize=True)

    # ---- NET BLOCK ----
    _section_band(ws, 23, "NET RESULT")

    # Row 24 — NET revenue (EMPHASIZED — winner highlighted via FormulaRule)
    _formula_row(ws, 24, "ANNUAL NET REVENUE",
                 "={sc}13-{sc}17", '"$"#,##0', bold=True)

    # Row 25 — NET / night
    _formula_row(ws, 25, "NET / occupied night",
                 "=IFERROR({sc}24/{sc}11,0)", '"$"#,##0.00')

    # Row 26 — Δ vs 1-night
    cell = ws.cell(row=26, column=2, value="Δ vs 1-night MNS")
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_SECONDARY)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for i in range(5):
        col = 3 + i
        if i == 0:
            c = ws.cell(row=26, column=col, value='="—"')
            c.font = Font(name=FONT_BODY, size=11, color=COLOR_MUTED)
        else:
            c = ws.cell(row=26, column=col, value=f"={SCEN_COLS[i]}24-C24")
            apply_style(c, formula_cell_style())
            c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
            c.number_format = '"+"$#,##0;"-"$#,##0;"$0"'
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[26].height = 20

    # ---- WINNER HIGHLIGHT — gold-soft on the column with max NET ----
    # Apply a FormulaRule per metric row (8..17, 24..26) on each scenario column
    # so the entire winning column lights up gold-soft.
    gold_soft_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    highlight_rows = list(range(8, 12)) + [13, 14, 15, 16, 17, 24, 25]
    for i, col_letter in enumerate(SCEN_COLS):
        # Range = col{first_row}:col{last_row} for this scenario column
        target_range = f"{col_letter}8:{col_letter}26"
        # Match formula: this column's NET (row 24) equals max NET across C24:G24
        rule_formula = f"{col_letter}$24=MAX($C$24:$G$24)"
        ws.conditional_formatting.add(
            target_range,
            FormulaRule(formula=[rule_formula], fill=gold_soft_fill)
        )
        # Header (row 6 + row 7) winner emphasis with bolder gold
        gold_strong_fill = PatternFill("solid", fgColor=COLOR_ACCENT)
        ws.conditional_formatting.add(
            f"{col_letter}6:{col_letter}7",
            FormulaRule(
                formula=[rule_formula],
                fill=gold_strong_fill,
                font=Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2"),
            )
        )

    # ---- BAR CHART — NET revenue by MNS ----
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "NET revenue by MNS"
    chart.y_axis.title = "Annual NET ($)"
    chart.x_axis.title = "Minimum-night-stay"
    chart.height = 9
    chart.width = 18

    # Data = row 24, cols C:G (single series, no header)
    data_ref = Reference(ws, min_col=3, max_col=7, min_row=24, max_row=24)
    cats_ref = Reference(ws, min_col=3, max_col=7, min_row=7, max_row=7)
    chart.add_data(data_ref, titles_from_data=False, from_rows=True)
    chart.set_categories(cats_ref)
    if chart.series:
        chart.series[0].tx = None  # no series title
    chart.legend = None
    chart.dataLabels = DataLabelList(showVal=True)

    style_chart(chart, single_color=COLOR_PRIMARY)

    # Anchor below comparison table
    ws.add_chart(chart, "B29")

    # ---- INSIGHT footnote ----
    insight_row = 50
    ws.merge_cells(f"A{insight_row}:L{insight_row + 2}")
    c = ws[f"A{insight_row}"]
    c.value = (
        "Why this matters: every turnover costs cleaning + supplies + your "
        "hours. A 1-night minimum maximizes booking velocity but multiplies "
        "turnover cost. Most properties net more at a 2- or 3-night minimum "
        "even with lower occupancy. Re-run when your cleaner rate or "
        "occupancy data shifts."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=2)
    for r in range(insight_row, insight_row + 3):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, insight_row + 4,
                 version_line=f"{SKU} · Comparison")

    ws.print_area = f"A1:L{insight_row + 7}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 32), ("C", 28),
        ("D", 4), ("E", 36),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Comparison", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Property selector + active year. Used in the Start tab headline."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _input_row(ws, 5, "Property:",
               _val(variant, SAMPLE["property_name"]), None,
               "Shown on Start tab headline")
    _input_row(ws, 7, "Active year:",
               _val(variant, SAMPLE["active_year"]), "0",
               "Tax/reporting year for this analysis")

    brand_footer(ws, 12, version_line=f"{SKU} · Settings")

    ws.print_area = "A1:L15"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_inputs_tab(wb, variant)
    build_comparison_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Minimum-Night-Stay Optimizer — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Compare 1n / 2n / 3n / 5n / 7n minimum-night-stay policies for an "
        "STR property side-by-side. Surfaces the truth that turnover costs "
        "often make a 2- or 3-night minimum net more than 1-night."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
