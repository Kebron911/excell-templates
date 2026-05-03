"""Build MKT-002 Review Response Tracker + Template Library (v2.3 standard).

Operational-mode register: customer logs every guest review (rating, content,
sentiment), tracks whether/when the host responded, captures lessons-learned,
and provides 8 copy-paste reply templates organized by review tone.

Cuts review-response time from 15 min to 3 min and prevents the
"I forgot to reply to that 3-star" failure mode that hurts ranking.

Generates:
- templates/_masters/MKT-002-review-response-tracker-DEMO.xlsx
- templates/_masters/MKT-002-review-response-tracker-BLANK.xlsx
"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    card_header, card_body_fill, apply_brand_header,
    STATE_BAD_FILL, STATE_GOOD_FILL,
)

SKU = "MKT-002"
NAME = "review-response-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"


# --- Sample data (DEMO) ---
# 28 reviews Q1 2026, avg ~4.7, 3 properties, 1 critical 3*, 1 hostile 1*

PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]
CHANNELS = ["Airbnb", "VRBO", "Booking", "Direct"]
SENTIMENTS = ["Glowing", "Positive", "Mixed", "Critical", "Hostile"]
TAGS = [
    "cleaning", "wifi", "noise", "comms", "checkin", "amenities",
    "location", "value", "linens", "comfort",
]


def _r(date_received, prop, channel, guest, booking, stars, public, private,
       sentiment, response_sent, response_date, tags, notes):
    """Helper: build a review tuple."""
    return (date_received, prop, channel, guest, booking, stars, public,
            private, sentiment, response_sent, response_date, tags, notes)


REVIEWS = [
    # January
    _r(date(2026, 1, 4),  "Smokies Ridge", "Airbnb",  "Megan",  "HM4F2K", 5,
       "Spotless cabin, hot tub was perfect after a snowy hike. Will be back!",
       "", "Glowing", "Y", date(2026, 1, 5),  "cleaning, comms", ""),
    _r(date(2026, 1, 8),  "Creek Side",    "Airbnb",  "David",  "HM4G7L", 5,
       "Communication was instant, check-in seamless. Beautiful views.",
       "", "Glowing", "Y", date(2026, 1, 9),  "comms, location", ""),
    _r(date(2026, 1, 11), "Lakehouse A",   "VRBO",    "Sarah",  "VR2218", 4,
       "Great place — wifi was a little slow for video calls though.",
       "Wifi router seemed old.", "Positive", "Y", date(2026, 1, 12), "wifi", ""),
    _r(date(2026, 1, 15), "Smokies Ridge", "Airbnb",  "Tom",    "HM4J9P", 5,
       "Linens felt like a hotel. Loved it.",
       "", "Glowing", "Y", date(2026, 1, 16), "linens, cleaning", ""),
    _r(date(2026, 1, 19), "Creek Side",    "Booking", "Rachel", "BK7791", 5,
       "Perfect for a quiet weekend. Spotless.",
       "", "Glowing", "Y", date(2026, 1, 20), "cleaning", ""),
    _r(date(2026, 1, 23), "Lakehouse A",   "Airbnb",  "Marcus", "HM4M2X", 5,
       "Host was amazing — answered every question within minutes.",
       "", "Glowing", "Y", date(2026, 1, 24), "comms", ""),
    _r(date(2026, 1, 27), "Smokies Ridge", "Airbnb",  "Lin",    "HM4P4D", 4,
       "Cute cabin. Driveway was slick — needed more salt.",
       "", "Positive", "Y", date(2026, 1, 28), "amenities", ""),
    _r(date(2026, 1, 30), "Creek Side",    "Direct",  "Jenny",  "DIR-014", 5,
       "Booked direct, got a great rate. Cabin was perfect.",
       "", "Glowing", "Y", date(2026, 1, 31), "value, cleaning", ""),

    # February
    _r(date(2026, 2, 3),  "Lakehouse A",   "VRBO",    "Greg",   "VR2244", 5,
       "Best lake view in the area. Worth every penny.",
       "", "Glowing", "Y", date(2026, 2, 4),  "location", ""),
    _r(date(2026, 2, 7),  "Smokies Ridge", "Airbnb",  "Priya",  "HM4S7T", 5,
       "Hot tub, fireplace, cozy linens. 10/10.",
       "", "Glowing", "Y", date(2026, 2, 8),  "comfort, linens", ""),
    _r(date(2026, 2, 10), "Creek Side",    "Airbnb",  "Eli",    "HM4T1A", 3,
       "Wifi was unusable. I work remote and had to drive into town.",
       "Speed test = 4 Mbps down. Spec says 100.", "Critical", "Y",
       date(2026, 2, 11), "wifi", "ESCALATED — replaced router 2/14"),
    _r(date(2026, 2, 14), "Lakehouse A",   "Airbnb",  "Anna",   "HM4U2B", 5,
       "Beautiful, well-stocked. Will recommend.",
       "", "Glowing", "Y", date(2026, 2, 15), "amenities, cleaning", ""),
    _r(date(2026, 2, 17), "Smokies Ridge", "Booking", "Carlos", "BK7822", 4,
       "Lovely. A bit of road noise at night but minor.",
       "", "Positive", "Y", date(2026, 2, 18), "noise", ""),
    _r(date(2026, 2, 20), "Creek Side",    "Airbnb",  "Hannah", "HM4W6E", 5,
       "Spotless. Felt new even though it isn't.",
       "", "Glowing", "Y", date(2026, 2, 21), "cleaning", ""),
    _r(date(2026, 2, 24), "Lakehouse A",   "Airbnb",  "Kev",    "HM4X3F", 5,
       "Easy check-in, great host. Already planning the next trip.",
       "", "Glowing", "Y", date(2026, 2, 25), "checkin, comms", ""),
    _r(date(2026, 2, 27), "Smokies Ridge", "Airbnb",  "Maria",  "HM4Y8G", 4,
       "Loved the cabin. Wifi could be faster.",
       "", "Positive", "Y", date(2026, 2, 28), "wifi", ""),

    # March
    _r(date(2026, 3, 2),  "Creek Side",    "VRBO",    "Tony",   "VR2301", 5,
       "Communication was outstanding. Place was immaculate.",
       "", "Glowing", "Y", date(2026, 3, 3),  "comms, cleaning", ""),
    _r(date(2026, 3, 5),  "Lakehouse A",   "Airbnb",  "Jess",   "HM5B2H", 5,
       "Stayed for 5 nights — never wanted to leave.",
       "", "Glowing", "Y", date(2026, 3, 6),  "comfort", ""),
    _r(date(2026, 3, 9),  "Smokies Ridge", "Airbnb",  "Brad",   "HM5D5J", 5,
       "Better than the photos. Hot tub was 102 — perfect.",
       "", "Glowing", "Y", date(2026, 3, 10), "amenities", ""),
    _r(date(2026, 3, 12), "Creek Side",    "Airbnb",  "Nora",   "HM5E1K", 1,
       "Smelled like mold the entire stay. Health hazard.",
       "Photos taken — looks like ordinary cabin smell, not mold. "
       "Cleaner reported no issues.",
       "Hostile", "Y", date(2026, 3, 14), "cleaning",
       "False mold claim — disputed via Airbnb. No refund issued."),
    _r(date(2026, 3, 15), "Lakehouse A",   "Booking", "Will",   "BK7901", 5,
       "Top-notch in every way.",
       "", "Glowing", "Y", date(2026, 3, 16), "cleaning, comms", ""),
    _r(date(2026, 3, 18), "Smokies Ridge", "Airbnb",  "Tara",   "HM5G7L", 5,
       "Cleaning was hotel-grade. Linens smelled like fresh laundry.",
       "", "Glowing", "Y", date(2026, 3, 19), "cleaning, linens", ""),
    _r(date(2026, 3, 21), "Creek Side",    "Airbnb",  "Dan",    "HM5H9M", 4,
       "Great spot. Wifi a bit slow but workable now.",
       "", "Positive", "Y", date(2026, 3, 22), "wifi", ""),
    _r(date(2026, 3, 24), "Lakehouse A",   "Direct",  "Liz",    "DIR-022", 5,
       "Booked direct — best decision. Spotless place, easy host.",
       "", "Glowing", "Y", date(2026, 3, 25), "value, comms", ""),
    _r(date(2026, 3, 26), "Smokies Ridge", "Airbnb",  "Kim",    "HM5K4N", 5,
       "Quiet, clean, gorgeous. Loved it.",
       "", "Glowing", "Y", date(2026, 3, 27), "cleaning, location", ""),
    _r(date(2026, 3, 28), "Creek Side",    "VRBO",    "Sam",    "VR2334", 5,
       "Amazing host. Place was perfect.",
       "", "Glowing", "N", None, "comms", "Reminder: respond by 4/4"),
    _r(date(2026, 3, 29), "Lakehouse A",   "Airbnb",  "Owen",   "HM5L5P", 5,
       "Will absolutely book again next year.",
       "", "Glowing", "Y", date(2026, 3, 30), "comfort, location", ""),
    _r(date(2026, 3, 30), "Smokies Ridge", "Airbnb",  "Ruth",   "HM5N1Q", 4,
       "Cute place. Light fixture in bathroom was loose.",
       "Reported privately — owner thanked us.",
       "Positive", "N", None, "amenities",
       "Reminder: respond by 4/6"),
]


# --- Reply templates (8) ---
# Each: (id, title, when_to_use, suggested_subject, body)

REPLY_TEMPLATES = [
    (
        "TPL-01",
        "5★ glowing — gracious + invite-back",
        "Use for unambiguously glowing 5-star reviews where the guest mentioned "
        "specifics (cleanliness, view, communication). Match their warmth.",
        "Thank you so much, [GUEST FIRST NAME]!",
        (
            "[GUEST FIRST NAME], thank you so much for the kind words — they "
            "genuinely made our day.\n\n"
            "We're thrilled the [SPECIFIC DETAIL — hot tub / view / linens] "
            "lived up to expectations. Hosts like to think the small touches "
            "matter, and reviews like yours confirm it.\n\n"
            "You'd be welcome back any time. If you ever return to "
            "[CITY/AREA], reach out directly and we'll take care of you.\n\n"
            "Safe travels,\n[HOST NAME]"
        ),
    ),
    (
        "TPL-02",
        "5★ generic positive — gracious + specific callback",
        "Use for short 5-star reviews with little detail. Pull a specific from "
        "the booking (length of stay, season, party size) so the reply doesn't "
        "feel canned.",
        "Thank you, [GUEST FIRST NAME]!",
        (
            "[GUEST FIRST NAME], thank you for the 5 stars — we really "
            "appreciate it.\n\n"
            "It was a pleasure hosting you for [LENGTH] in [MONTH]. "
            "[SEASON DETAIL — hope you got to see the leaves change / glad "
            "the snow held off / etc.].\n\n"
            "We'd love to host you again. Safe travels,\n[HOST NAME]"
        ),
    ),
    (
        "TPL-03",
        "4★ mixed-positive — appreciate + soft acknowledgment",
        "Use for 4-star reviews with mostly-positive content and one small "
        "complaint. Acknowledge the issue briefly, don't over-apologize.",
        "Thanks for the honest feedback, [GUEST FIRST NAME]",
        (
            "[GUEST FIRST NAME], thank you for staying with us and for the "
            "thoughtful review.\n\n"
            "We're really glad you enjoyed [POSITIVE DETAIL]. You're right "
            "about [ISSUE] — that's good feedback, and we [FIX TAKEN OR "
            "PLANNED].\n\n"
            "Hope to host you again. — [HOST NAME]"
        ),
    ),
    (
        "TPL-04",
        "3★ mixed — acknowledge + improvement note + invite back",
        "Use for 3-star reviews where the complaint is real and fixable. "
        "Show ownership without grovelling. Future guests read this.",
        "Thank you for the feedback, [GUEST FIRST NAME]",
        (
            "[GUEST FIRST NAME], thank you for the candid review. Honest "
            "feedback is the only way we improve.\n\n"
            "You're absolutely right about [ISSUE]. We [SPECIFIC ACTION — "
            "replaced the router / brought in a deep cleaner / scheduled "
            "the repair]. It shouldn't have been an issue during your "
            "stay, and we own that.\n\n"
            "We'd appreciate the chance to host you again under better "
            "circumstances. — [HOST NAME]"
        ),
    ),
    (
        "TPL-05",
        "2★ negative-justified — acknowledge + corrective action",
        "Use when the negative review is fair. Apologize once, state what "
        "changed, do not over-explain. Future guests need to see ownership.",
        "Thank you for the feedback, [GUEST FIRST NAME]",
        (
            "[GUEST FIRST NAME], I'm sorry your stay fell short. The "
            "issues you described — [ISSUE 1] and [ISSUE 2] — are not the "
            "experience we want any guest to have.\n\n"
            "Here is what changed since you left: [SPECIFIC FIX 1]. "
            "[SPECIFIC FIX 2].\n\n"
            "Thank you for taking the time to write — it's how we get "
            "better. — [HOST NAME]"
        ),
    ),
    (
        "TPL-06",
        "2★ negative-unjustified — calm correction + facts",
        "Use when a negative review misstates facts. Stay calm, state "
        "facts, no defensiveness. Future guests judge tone, not who's "
        "right.",
        "Thank you for the feedback, [GUEST FIRST NAME]",
        (
            "[GUEST FIRST NAME], thank you for the review. We always "
            "appreciate honest feedback, and we'd like to add some "
            "context for future guests.\n\n"
            "Regarding [POINT]: [FACTUAL CORRECTION — receipts / logs / "
            "what was actually communicated]. We were happy to [WHAT YOU "
            "DID] when you raised it during the stay.\n\n"
            "We hope your travels going forward are better. "
            "— [HOST NAME]"
        ),
    ),
    (
        "TPL-07",
        "1★ hostile/false — measured factual correction (no apology)",
        "Use only when the review is hostile, false, or alleges something "
        "serious (mold, pests, safety) without merit. No apology — that "
        "reads as admission. Stick to documented facts.",
        "Response to [GUEST FIRST NAME]'s review",
        (
            "[GUEST FIRST NAME] — we want to address the claims in this "
            "review for any future guests reading.\n\n"
            "[CLAIM 1] is not accurate. [DOCUMENTED FACT — cleaner "
            "checklist on date / inspection log / photo timestamp].\n\n"
            "We attempted to address concerns during the stay by "
            "[WHAT YOU OFFERED]; the offer was [DECLINED / NOT "
            "RESPONDED TO].\n\n"
            "We host hundreds of guests a year and stand by the home and "
            "the cleaning standard. Future guests are welcome to read the "
            "[N] reviews above this one. — [HOST NAME]"
        ),
    ),
    (
        "TPL-08",
        "Late-checkout / damage-claim guest review (delicate)",
        "Use when a guest who was late-checking-out or had a damage claim "
        "leaves a defensive negative review. Tone is the whole game — "
        "professional, not petty.",
        "Thank you for staying, [GUEST FIRST NAME]",
        (
            "[GUEST FIRST NAME], thank you for staying with us.\n\n"
            "We're sorry the end of your stay didn't go smoothly. For "
            "transparency to future guests: our checkout time is "
            "[CHECKOUT TIME], clearly stated in the listing and the "
            "house manual, and [WHAT HAPPENED — late checkout fee / "
            "damage claim] followed our standard policy.\n\n"
            "We wish you safe travels and would welcome you back under "
            "different terms. — [HOST NAME]"
        ),
    ),
]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational hero + KPIs + nav)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Review Response Tracker"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Reply with grace in 3 minutes — not 15."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Logs every guest review (rating, content, sentiment), tracks whether "
        "you responded and how fast, and stores 8 reply templates organized by "
        "review tone. Cuts response time from 15 minutes to 3, and prevents the "
        "\"I forgot to reply to that 3-star\" failure mode that quietly drags "
        "your ranking. Lessons-Learned auto-aggregates the tags you assign so "
        "themes (wifi, cleaning, comms) surface before they sink an average."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Reviews Log: log every review the day it lands — date, stars, public text, sentiment.",
        "② Reply Templates: pick the template that matches the tone — copy, swap [BRACKETS], paste.",
        "③ Tag each review (cleaning, wifi, noise, comms) so Lessons Learned can find patterns.",
        "④ Mark Response sent? = Y and date stamped — Days-to-respond auto-calculates.",
        "⑤ Lessons Learned: scan the top tags monthly. The recurring ones are property fixes.",
        "⑥ Settings: update property + channel lists once a year. Active year drives KPIs.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "ADD A NEW REVIEW" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  ADD A NEW REVIEW  (OPEN REVIEWS LOG)",
                   "'Reviews Log'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 3 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Average rating YTD
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "AVG RATING (ACTIVE YR)"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = (
        "=IFERROR(AVERAGEIFS('Reviews Log'!F6:F305,"
        "'Reviews Log'!A6:A305,\">=\"&DATE(Settings!$B$5,1,1),"
        "'Reviews Log'!A6:A305,\"<=\"&DATE(Settings!$B$5,12,31)),0)"
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.00"
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "stars · all properties"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Reviews this year
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "REVIEWS LOGGED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = (
        "=COUNTIFS('Reviews Log'!A6:A305,\">=\"&DATE(Settings!$B$5,1,1),"
        "'Reviews Log'!A6:A305,\"<=\"&DATE(Settings!$B$5,12,31))"
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "in active year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Response rate %
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "RESPONSE RATE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = (
        "=IFERROR(COUNTIFS('Reviews Log'!J6:J305,\"Y\","
        "'Reviews Log'!A6:A305,\">=\"&DATE(Settings!$B$5,1,1),"
        "'Reviews Log'!A6:A305,\"<=\"&DATE(Settings!$B$5,12,31))/"
        "COUNTIFS('Reviews Log'!A6:A305,\">=\"&DATE(Settings!$B$5,1,1),"
        "'Reviews Log'!A6:A305,\"<=\"&DATE(Settings!$B$5,12,31)),0)"
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "responses / reviews"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav — 4 buttons (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Reply Templates",
                   "'Reply Templates'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Lessons Learned",
                   "'Lessons Learned'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Print Lessons",
                   "'Lessons Learned'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Reminder callout (row 42) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "⏱ Respond within 48 hours. Future guests scroll past your replies "
        "before they book — silence reads louder than the review itself."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Tone reminder (row 44) ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "⚠ Never apologize on a 1-star false claim — that reads as admission. "
        "Use TPL-07 for hostile reviews and let the documented facts speak."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    add_upgrade_banner(ws, 46)

    brand_footer(ws, 48,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_reviews_log_tab(wb, variant):
    """Sheet 1 — Reviews Log (one row per review, 300 capacity)."""
    ws = wb.create_sheet("Reviews Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Reviews Log",
                         prev_tab="Start", next_tab="Reply Templates")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Log every review the day it arrives — stars + sentiment + tags drive everything else"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # 14 columns: A-N
    set_col_widths(ws, [
        ("A", 12),  # Date received
        ("B", 16),  # Property
        ("C", 11),  # Channel
        ("D", 14),  # Guest first name
        ("E", 12),  # Booking ID
        ("F", 7),   # Stars
        ("G", 40),  # Public review text
        ("H", 28),  # Private feedback
        ("I", 11),  # Sentiment
        ("J", 8),   # Response sent?
        ("K", 12),  # Response date
        ("L", 9),   # Days to respond
        ("M", 22),  # Tags
        ("N", 22),  # Notes
    ])

    headers = [
        "Date Received", "Property", "Channel", "Guest", "Booking ID",
        "Stars", "Public Review Text", "Private Feedback", "Sentiment",
        "Sent?", "Response Date", "Days", "Tags", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    review_rows = REVIEWS if variant == "demo" else []

    for i, rev in enumerate(review_rows, start=6):
        (date_recv, prop, channel, guest, booking, stars, public, private,
         sentiment, sent, resp_date, tags, notes) = rev

        a = ws.cell(row=i, column=1, value=date_recv)
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=prop)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=channel)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=guest)
        apply_style(d, input_cell_style())

        e = ws.cell(row=i, column=5, value=booking)
        apply_style(e, input_cell_style())

        f = ws.cell(row=i, column=6, value=stars)
        apply_style(f, input_cell_style())
        f.alignment = Alignment(horizontal="center", vertical="center")

        g = ws.cell(row=i, column=7, value=public)
        apply_style(g, input_cell_style())
        g.alignment = Alignment(horizontal="left", vertical="top",
                                  wrap_text=True)

        h = ws.cell(row=i, column=8, value=private if private else None)
        apply_style(h, input_cell_style())
        h.alignment = Alignment(horizontal="left", vertical="top",
                                  wrap_text=True)

        i_cell = ws.cell(row=i, column=9, value=sentiment)
        apply_style(i_cell, input_cell_style())

        j = ws.cell(row=i, column=10, value=sent)
        apply_style(j, input_cell_style())
        j.alignment = Alignment(horizontal="center", vertical="center")

        k = ws.cell(row=i, column=11, value=resp_date)
        apply_style(k, input_cell_style())
        k.number_format = "yyyy-mm-dd"

        # Days to respond formula
        l_cell = ws.cell(
            row=i, column=12,
            value=f'=IF(OR(A{i}="",K{i}=""),"",K{i}-A{i})',
        )
        apply_style(l_cell, formula_cell_style())
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        l_cell.number_format = "0"

        m = ws.cell(row=i, column=13, value=tags)
        apply_style(m, input_cell_style())

        n = ws.cell(row=i, column=14, value=notes if notes else None)
        apply_style(n, input_cell_style())
        n.alignment = Alignment(horizontal="left", vertical="top",
                                  wrap_text=True)

        ws.row_dimensions[i].height = 32

    # Capacity rows up to 305 (300 reviews)
    last_data_row = len(review_rows) + 5
    for row_idx in range(last_data_row + 1, 306):
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1 or col_idx == 11:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center",
                                             vertical="center")
            if col_idx == 10:
                cell.alignment = Alignment(horizontal="center",
                                             vertical="center")
            if col_idx in (7, 8, 14):
                cell.alignment = Alignment(horizontal="left",
                                             vertical="top", wrap_text=True)
        # Days-to-respond formula
        l_cell = ws.cell(
            row=row_idx, column=12,
            value=f'=IF(OR(A{row_idx}="",K{row_idx}=""),"",K{row_idx}-A{row_idx})',
        )
        apply_style(l_cell, formula_cell_style())
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        l_cell.number_format = "0"
        ws.row_dimensions[row_idx].height = 18

    # Dropdowns
    add_dropdown(ws, "B6:B305", "=Settings!$B$7:$B$16")
    add_dropdown(ws, "C6:C305", "=Settings!$B$18:$B$25")
    add_dropdown(ws, "I6:I305", '"Glowing,Positive,Mixed,Critical,Hostile"')
    add_dropdown(ws, "J6:J305", '"Y,N"')

    # Conditional formatting on Stars column F
    # 5 stars green
    ws.conditional_formatting.add(
        "F6:F305",
        FormulaRule(
            formula=["F6=5"],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
            font=Font(bold=True, color=COLOR_TEXT),
        ),
    )
    # 4 stars parchment
    ws.conditional_formatting.add(
        "F6:F305",
        FormulaRule(
            formula=["F6=4"],
            fill=PatternFill("solid", fgColor=COLOR_BG_LIGHT),
            font=Font(color=COLOR_TEXT),
        ),
    )
    # 3 stars gold
    ws.conditional_formatting.add(
        "F6:F305",
        FormulaRule(
            formula=["F6=3"],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
            font=Font(bold=True, color=COLOR_TEXT),
        ),
    )
    # 1-2 stars red
    ws.conditional_formatting.add(
        "F6:F305",
        FormulaRule(
            formula=["AND(F6>=1,F6<=2)"],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    # J column: red if N AND >7 days since received (use TODAY)
    ws.conditional_formatting.add(
        "J6:J305",
        FormulaRule(
            formula=['AND(J6="N",A6<>"",TODAY()-A6>7)'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_reply_templates_tab(wb):
    """Sheet 2 — Reply Templates (8 cards: title + when + subject + body)."""
    ws = wb.create_sheet("Reply Templates")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    compact_header_band(ws, "Reply Templates",
                         prev_tab="Reviews Log", next_tab="Lessons Learned")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "8 templates · pick the tone · copy · swap [BRACKETS] · paste"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6),   # ID column
        ("B", 14), ("C", 14), ("D", 14), ("E", 14), ("F", 14),
        ("G", 14), ("H", 14), ("I", 14), ("J", 14), ("K", 14),
        ("L", 8),
    ])

    # Each template card spans rows: 1 header + 1 when-to-use + 1 subject + N body lines
    # Body height = 7 rows (~6-12 lines wrapped).
    # Card layout:
    #   Row R     : card_header — "TPL-01 · Title"
    #   Row R+1   : "WHEN TO USE: ..." (parchment)
    #   Row R+2   : "SUGGESTED SUBJECT: ..." (parchment)
    #   Rows R+3 .. R+9 : body (parchment, wrapped)
    #   Row R+10  : spacer (height 6)

    current_row = 6
    for tpl_id, title, when, subject, body in REPLY_TEMPLATES:
        # Card header
        card_header(ws, current_row, ("A", "L"),
                     f"{tpl_id}  ·  {title}", height=22)

        when_row = current_row + 1
        subj_row = current_row + 2
        body_start = current_row + 3
        body_end = current_row + 9

        # When-to-use row
        ws.merge_cells(f"A{when_row}:L{when_row}")
        c = ws[f"A{when_row}"]
        c.value = f"WHEN TO USE  ·  {when}"
        c.font = Font(name=FONT_BODY, size=10, italic=True,
                       color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)
        ws.row_dimensions[when_row].height = 32

        # Subject row
        ws.merge_cells(f"A{subj_row}:L{subj_row}")
        c = ws[f"A{subj_row}"]
        c.value = f"SUGGESTED SUBJECT  ·  {subject}"
        c.font = Font(name=FONT_MONO, size=10, bold=True,
                       color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)
        ws.row_dimensions[subj_row].height = 18

        # Body row (merged, parchment, multi-line)
        ws.merge_cells(f"A{body_start}:L{body_end}")
        c = ws[f"A{body_start}"]
        c.value = body
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top",
                                  wrap_text=True, indent=2)
        for r in range(body_start, body_end + 1):
            ws.row_dimensions[r].height = 18

        # Apply parchment fill + gold border via card_body_fill
        card_body_fill(ws, when_row, body_end, ("A", "L"), border=True)

        # Spacer row
        ws.row_dimensions[body_end + 1].height = 8

        current_row = body_end + 2

    # Footer
    brand_footer(ws, current_row + 1,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = f"A1:L{current_row + 4}"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_lessons_tab(wb):
    """Sheet 3 — Lessons Learned (auto top-tags + manual action items)."""
    ws = wb.create_sheet("Lessons Learned")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Lessons Learned",
                         prev_tab="Reply Templates", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Themes that recur · which tags drag the score · what to fix on the property"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 18), ("B", 14), ("C", 14), ("D", 14),
        ("E", 4),
        ("F", 30), ("G", 14), ("H", 16),
    ])

    # --- Section 1: Tag frequency table (rows 6-17) ---
    sect_a = ws.cell(row=6, column=1, value="Tag Frequency · Active Year")
    sect_a.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[6].height = 22

    # Headers row 7
    tag_headers = ["Tag", "Mentions", "Avg Stars", "Drag?"]
    for col, h in enumerate(tag_headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 18

    # Rows 8-17: pull from Settings tag suggestions B27:B36
    for i in range(8, 18):
        settings_idx = 27 + (i - 8)  # Settings B27..B36

        a = ws.cell(
            row=i, column=1,
            value=f"=IF(Settings!B{settings_idx}=\"\",\"\",Settings!B{settings_idx})",
        )
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")

        b = ws.cell(
            row=i, column=2,
            value=(
                f'=IF(A{i}="","",'
                f'SUMPRODUCT((\'Reviews Log\'!$A$6:$A$305>=DATE(Settings!$B$5,1,1))'
                f'*(\'Reviews Log\'!$A$6:$A$305<=DATE(Settings!$B$5,12,31))'
                f'*ISNUMBER(SEARCH(A{i},\'Reviews Log\'!$M$6:$M$305))))'
            ),
        )
        apply_style(b, formula_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.number_format = "0"

        c = ws.cell(
            row=i, column=3,
            value=(
                f'=IFERROR(SUMPRODUCT('
                f'(\'Reviews Log\'!$A$6:$A$305>=DATE(Settings!$B$5,1,1))'
                f'*(\'Reviews Log\'!$A$6:$A$305<=DATE(Settings!$B$5,12,31))'
                f'*ISNUMBER(SEARCH(A{i},\'Reviews Log\'!$M$6:$M$305))'
                f'*\'Reviews Log\'!$F$6:$F$305)/B{i},"")'
            ),
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.00"

        d = ws.cell(
            row=i, column=4,
            value=f'=IF(OR(A{i}="",B{i}=0),"",IF(C{i}<4.5,"⚠ drag","ok"))',
        )
        apply_style(d, formula_cell_style())
        d.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 16

    # Conditional fmt: Drag column red
    ws.conditional_formatting.add(
        "D8:D17",
        FormulaRule(
            formula=['D8="⚠ drag"'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        "D8:D17",
        FormulaRule(
            formula=['D8="ok"'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
        ),
    )

    # --- Section 2: Action Items table (right side, rows 6-25) ---
    sect_b = ws.cell(row=6, column=6, value="Action Items · Property Fixes")
    sect_b.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)

    action_headers = ["Issue / Theme", "Priority", "Status"]
    for offset, h in enumerate(action_headers):
        cell = ws.cell(row=7, column=6 + offset, value=h)
        apply_style(cell, header_row_style())

    # Pre-fill 2 sample action items in DEMO-friendly way (left blank in BLANK
    # too — they're placeholders the user types into either way).
    sample_actions = [
        ("Replace wifi router at Lakehouse A + Creek Side", "High", "Done"),
        ("Add salt bin to Smokies Ridge driveway by Nov", "Med", "Open"),
    ]
    for offset, (issue, prio, status) in enumerate(sample_actions):
        r = 8 + offset
        a = ws.cell(row=r, column=6, value=issue)
        apply_style(a, input_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
        b = ws.cell(row=r, column=7, value=prio)
        apply_style(b, input_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=r, column=8, value=status)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22

    # Capacity rows 10-25
    for r in range(10, 26):
        a = ws.cell(row=r, column=6)
        apply_style(a, input_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
        b = ws.cell(row=r, column=7)
        apply_style(b, input_cell_style())
        b.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=r, column=8)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    add_dropdown(ws, "G8:G25", '"High,Med,Low"')
    add_dropdown(ws, "H8:H25", '"Open,In Progress,Done,Deferred"')

    # Footer
    brand_footer(ws, 28,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:H30"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 4 — Settings (active year, properties, channels, tags, archive)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Lessons Learned", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active year · property list · channel list · tag suggestions · year-end archive"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 22), ("C", 16), ("D", 16), ("E", 16)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Active year (row 5) — drives KPIs ---
    a5 = ws.cell(row=5, column=1, value="Active year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")

    b5 = ws.cell(row=5, column=2, value=2026)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    ws.merge_cells("A6:E6")
    c6 = ws["A6"]
    c6.value = (
        "Drives the Start tab KPIs and Lessons Learned tag aggregation. "
        "Bump in January after archiving last year's totals below."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 24

    # --- Properties (rows 7-16) ---
    a7 = ws.cell(row=7, column=1, value="Properties (used by Reviews Log dropdown):")
    a7.font = bold_right
    a7.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[7].height = 18

    prop_seed = PROPERTIES if variant == "demo" else []
    for i in range(7, 17):
        cell = ws.cell(row=i, column=2)
        apply_style(cell, input_cell_style())
        if i - 7 < len(prop_seed):
            cell.value = prop_seed[i - 7]
        ws.row_dimensions[i].height = 16

    # --- Channels (rows 18-25) ---
    a18 = ws.cell(row=18, column=1, value="Channels (used by Reviews Log dropdown):")
    a18.font = bold_right
    a18.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[18].height = 18

    chan_seed = CHANNELS  # both variants get channels — fixed list
    for i in range(18, 26):
        cell = ws.cell(row=i, column=2)
        apply_style(cell, input_cell_style())
        if i - 18 < len(chan_seed):
            cell.value = chan_seed[i - 18]
        ws.row_dimensions[i].height = 16

    # --- Tag suggestions (rows 27-36) ---
    a27 = ws.cell(row=27, column=1, value="Tag suggestions (used by Lessons Learned):")
    a27.font = bold_right
    a27.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[27].height = 18

    # Tags seed both DEMO and BLANK so Lessons Learned tab is functional
    # immediately — customer can edit/extend.
    for i in range(27, 37):
        cell = ws.cell(row=i, column=2)
        apply_style(cell, input_cell_style())
        if i - 27 < len(TAGS):
            cell.value = TAGS[i - 27]
        ws.row_dimensions[i].height = 16

    # --- Year-end Archive (rows 38-46) ---
    sect38 = ws.cell(row=38, column=1, value="Year-end Archive")
    sect38.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[38].height = 22

    ws.merge_cells("A39:E39")
    c39 = ws["A39"]
    c39.value = (
        "Each January, copy YTD totals into the row for the closing year, "
        "bump Active year (B5), and clear the Reviews Log for the new year."
    )
    c39.font = italic_muted
    c39.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[39].height = 28

    archive_headers = ["Year", "Reviews", "Avg Stars", "Response %", "5★ Count"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=40, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[40].height = 18

    for idx, year in enumerate(range(2024, 2031), start=41):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        cell = ws.cell(row=idx, column=2)
        apply_style(cell, input_cell_style())
        cell.number_format = "0"
        cell = ws.cell(row=idx, column=3)
        apply_style(cell, input_cell_style())
        cell.number_format = "0.00"
        cell = ws.cell(row=idx, column=4)
        apply_style(cell, input_cell_style())
        cell.number_format = "0%"
        cell = ws.cell(row=idx, column=5)
        apply_style(cell, input_cell_style())
        cell.number_format = "0"
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_reviews_log_tab(wb, variant)
    build_reply_templates_tab(wb)
    build_lessons_tab(wb)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Review Response Tracker{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Review tracker + 8 reply-template library for STR hosts. "
        "Cuts response time 15min -> 3min (v2.3)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
