"""Build MKT-001 Listing SEO Audit (v2.2 standard).

Wizard-mode tool — host scores ONE Airbnb listing across 40 weighted
criteria, gets a prioritized fix list, and schedules a re-audit. Methodology
page + Top-5 Comp Compare differentiate Full ($47) from Lite ($27 / 25 crits).

Generates three files:
  templates/_masters/MKT-001-listing-seo-audit-DEMO.xlsx     (Full + sample)
  templates/_masters/MKT-001-listing-seo-audit-BLANK.xlsx    (Full + empty)
  templates/_lite/MKT-001-listing-seo-audit-lite.xlsx        (Lite, 4 tabs, sample)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
)

SKU = "MKT-001"
NAME = "listing-seo-audit"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
LITE_OUT = BASE / "_lite" / f"{SKU}-{NAME}-lite.xlsx"


# ---------------------------------------------------------------------------
# 40-criteria rubric (category, weight%, items[])
# Each item: (id, criterion text, points-max, suggested fix text, demo Y/N/Partial)
# ---------------------------------------------------------------------------

# Category weight applied per the brief: A=15, B=20, C=30, D=20, E=15.
# Implementation: each criterion in a category contributes equal share of
# that category's weighted score. Within Audit Score we score 0/5/10 per
# criterion (No/Partial/Yes) and then weight at the category level.

CRITERIA = [
    # Category A — Title (5 criteria, 15% weight)
    ("A", "Title",
     [
         (1,  "Title length 35-50 chars (sweet spot)",
              "Trim to 50 chars max. Cut weak words first: 'beautiful', 'amazing', 'cozy'.",
              "No"),     # demo: 51 chars
         (2,  "Includes city or neighborhood",
              "Add city or neighborhood near the front of the title.",
              "Partial"),
         (3,  "Includes 1-2 highest-value amenities (hot tub / view / walkable)",
              "Promote your top 1-2 amenities into the title. Hot tub > view > walkable.",
              "Yes"),
         (4,  "No ALL CAPS, no emojis, no clickbait",
              "Strip caps and emojis. Airbnb's algorithm and guests both penalize them.",
              "Yes"),
         (5,  "Differentiating phrase or experience hook",
              "Add a 2-3 word hook: 'sunset deck', 'creekside cabin', 'walk to brewery'.",
              "Partial"),
     ]),
    # Category B — Description (8 criteria, 20% weight)
    ("B", "Description",
     [
         (6,  "Opens with a hook line (NOT 'Welcome to my home!')",
              "Replace generic welcome with one sentence that paints a scene or promise.",
              "Yes"),
         (7,  "Length between 150-500 words",
              "Aim for 220-350 words. Less = thin. More = guests skim and miss key info.",
              "Yes"),
         (8,  "Lists what makes property unique in first 3 sentences",
              "Move your unique edge (view, hot tub, location) into the opening 3 sentences.",
              "Partial"),
         (9,  "Mentions surrounding area + walkability or drive times",
              "Add 1-2 lines on nearby attractions with drive minutes ('15 min to Dollywood').",
              "Yes"),
         (10, "Has dedicated bedrooms section with sleeping config",
              "Add a 'Sleeping arrangements' block: 'Bedroom 1: King · Bedroom 2: Queen + bunk'.",
              "Yes"),
         (11, "Has dedicated amenities section (formatted, scannable)",
              "Add a bullet list: 'Hot tub · Fast WiFi · EV charger · Workspace'.",
              "Yes"),
         (12, "Includes house rules (positive framing)",
              "Reframe rules as guest-friendly: 'Quiet hours after 10pm to keep neighbors happy'.",
              "Partial"),
         (13, "Closes with hospitality / response promise",
              "End with: 'Message us anytime — we typically reply within an hour.'",
              "No"),
     ]),
    # Category C — Photos (12 criteria, 30% weight)
    ("C", "Photos",
     [
         (14, "At least 25 photos posted",
              "Add 6+ photos to reach 25. Twilight exterior, kitchen detail, bedroom #2 are common gaps.",
              "No"),     # demo: 19 photos
         (15, "Cover photo is exterior or hero interior shot",
              "Swap cover to exterior or hero living shot. Bedroom covers are conversion killers.",
              "No"),     # demo: bedroom cover
         (16, "Bedroom photo for every bedroom",
              "Add at least one wide-angle shot per bedroom. Show the bed clearly.",
              "Yes"),
         (17, "Bathroom photo for every bathroom",
              "Add bathroom shots. Highlight cleanliness, towels, walk-in shower if you have one.",
              "Partial"),
         (18, "Kitchen photo (wide + detail)",
              "Add a wide kitchen shot plus 1 detail (coffee setup, island, range).",
              "Yes"),
         (19, "Living / common space photo",
              "Show the main hangout space with sofa, lighting, and TV in frame.",
              "Yes"),
         (20, "Outdoor space photo (porch, deck, yard, view)",
              "Add an outdoor shot. Hot tub at twilight or a deck view drives bookings.",
              "Yes"),
         (21, "Twilight or sunset photo",
              "Add 1-2 twilight shots. Hire a pro or shoot 30 min after sunset, lights ON.",
              "No"),
         (22, "Professional quality (sharp, well-lit, level horizon)",
              "Pay $250-500 for a 1-hour STR photographer. Highest-ROI marketing dollar you spend.",
              "Partial"),
         (23, "Photos sequenced as guest journey (arrival → living → bedrooms → bath → kitchen → outdoor)",
              "Drag photos into journey order in the Airbnb editor. Cover, then living, then bedrooms.",
              "No"),
         (24, "Captions on every photo",
              "Caption each photo: 'King bed with mountain view', not 'Bedroom 1'.",
              "No"),
         (25, "No phone-camera-mirror shots, no clutter visible",
              "Remove any mirror-selfie shots. Hide cords, remotes, and personal items before shooting.",
              "Yes"),
     ]),
    # Category D — Amenities (8 criteria, 20% weight)
    ("D", "Amenities",
     [
         (26, "At least 40 amenities checked",
              "Walk through the Airbnb amenity list and check everything you actually offer.",
              "Partial"),  # demo: 38
         (27, "Top-tier amenities listed (hot tub / pool / view / fireplace)",
              "If you have any of these, make sure they're checked AND mentioned in title/description.",
              "Yes"),
         (28, "Fast WiFi listed with speed (≥ 100 Mbps)",
              "Add the speed. Run a speedtest, screenshot result, paste in description.",
              "Partial"),
         (29, "Workspace / desk listed",
              "Add a desk + chair to a corner. 30%+ of weekday bookings filter for workspace.",
              "Yes"),
         (30, "Self check-in OR clear arrival instructions",
              "Add a smart lock or lockbox. Self check-in unlocks Instant Book guest segments.",
              "Yes"),
         (31, "Pet policy clearly stated",
              "State explicitly: 'Pet-friendly — 2 pets max, $25 fee per stay.'",
              "Yes"),
         (32, "Family-friendly amenities (if family-friendly listing)",
              "Add: high chair, pack-n-play, board games, baby gates. Family searches filter for these.",
              "Partial"),
         (33, "Accessibility features listed",
              "Audit accessibility: step-free entry, grab bars, wide doorways. Tag what applies.",
              "No"),
     ]),
    # Category E — Trust signals (7 criteria, 15% weight)
    ("E", "Trust signals (Airbnb algorithm)",
     [
         (34, "Response time < 1 hour",
              "Turn on Airbnb push notifications. Auto-reply messages count toward response time.",
              "No"),     # demo: 4hr
         (35, "Response rate ≥ 95%",
              "Reply to every inquiry within 24 hours, even with a 'one moment' holding message.",
              "Yes"),
         (36, "Instant Book is ON",
              "Turn on Instant Book with verified-guest + good-reviews requirement. Biggest single ranking lift.",
              "No"),
         (37, "Superhost status",
              "Maintain 4.8+ rating, 90%+ response, 1%+ cancel rate, 10+ stays per year.",
              "Yes"),
         (38, "Cancellation policy = Flexible or Moderate",
              "Switch to Moderate. Strict policies suppress ranking and conversion both.",
              "No"),     # demo: Strict
         (39, "20+ reviews with 4.8+ avg",
              "Until you hit 20 reviews, optimize EVERY stay for a 5-star: pre-arrival message, mid-stay check-in, post-stay thank-you.",
              "Partial"),
         (40, "Verified ID + government ID requirement set",
              "In Settings, require government ID for booking. Reduces party / bad-guest risk.",
              "Yes"),
     ]),
]

CATEGORY_WEIGHTS = {"A": 15, "B": 20, "C": 30, "D": 20, "E": 15}

# Lite version: trim each category to most-impactful subset = 25 criteria total.
# A=4, B=5, C=8, D=4, E=4 ~ same proportional weight.
LITE_IDS = {
    1, 2, 3, 5,                              # A: 4 of 5
    6, 7, 8, 10, 13,                         # B: 5 of 8
    14, 15, 16, 18, 19, 21, 22, 23,          # C: 8 of 12 (highest-impact)
    26, 27, 28, 30,                          # D: 4 of 8
    34, 36, 37, 38,                          # E: 4 of 7
}


# Sample listing data (from brief QA section)
SAMPLE = {
    "listing_name":      "Smokies Ridge Cabin — Hot Tub, Mountain Views, Pet-Friendly",
    "listing_url":       "https://airbnb.com/h/smokies-ridge-cabin",
    "title_text":        "Smokies Ridge Cabin — Hot Tub, Mountain Views, Pet-Friendly",
    "title_chars":       60,    # actually 60 chars; just over the 50 cap
    "description":       (
        "Wake up to fog rolling off the Smokies from your private hot tub. "
        "Smokies Ridge Cabin is a 3-bedroom mountain retreat 15 minutes from "
        "downtown Gatlinburg and 25 from Dollywood. Sleeping: King master, "
        "Queen + bunk room, Queen guest. Amenities include fast WiFi (200 Mbps), "
        "EV charger, fireplace, fenced yard for dogs, and a fully-stocked "
        "kitchen. Quiet hours after 10pm — we love our neighbors. Pet-friendly "
        "with a $25 fee. Self check-in via smart lock. We respond within an hour."
    ),
    "desc_words":        220,
    "photo_count":       19,
    "cover_photo":       "Bedroom",
    "bedroom_photos":    3,
    "bathroom_photos":   1,
    "kitchen_photos":    2,
    "outdoor_photos":    3,
    "photos_pro":        "No",
    "twilight_photos":   "No",
    "photos_sequenced":  "No",
    "amenity_count":     38,
    "has_hot_tub":       "Yes",
    "has_pool":          "No",
    "has_fireplace":     "Yes",
    "has_fast_wifi":     "Yes",
    "has_workspace":     "Yes",
    "has_ev_charger":    "Yes",
    "pet_friendly":      "Yes",
    "kid_friendly":      "Yes",
    "self_checkin":      "Yes",
    "smart_lock":        "Yes",
    "response_rate":     0.97,
    "response_time":     "1-12hr",
    "instant_book":      "No",
    "superhost":         "Yes",
    "review_count":      24,
    "avg_rating":        4.82,
    "cancellation":      "Strict",
    "reaudit_date":      "2026-08-01",
}

# Sample comp data — 5 competitor listings
COMP_NAMES = [
    "Mtn View Lodge — Hot Tub & Views",
    "Cozy Creekside Cabin",
    "Sunset Ridge Retreat",
    "Bear Den Hideaway",
    "Foothills Family Cabin",
]

# For the comp grid: each comp gets a Y/N/Partial across all 40 criteria
# (just demo data — diverse to show how grid surfaces gaps).
import random
random.seed(42)
COMP_SCORES = {}
for comp in COMP_NAMES:
    COMP_SCORES[comp] = [random.choice(["Yes", "Yes", "Yes", "Partial", "No"]) for _ in range(40)]

# Sample re-audit history (4 prior audits)
SAMPLE_HISTORY = [
    ("2025-08-15", 52, "Pre-pro-photos baseline."),
    ("2025-11-02", 61, "Added pro photos — big jump in Photos category."),
    ("2026-02-10", 65, "Description rewrite + amenities expansion."),
    ("2026-05-01", 67, "Today's audit."),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label, fill_color=None):
    """Navy section band with serif white label."""
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=fill_color or COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _category_band(ws, row, letter, name, weight_pct, col_range="A:L"):
    """Gold-soft category band on Audit Score sheet."""
    first, last = col_range.split(":")
    ws.merge_cells(f"{first}{row}:{last}{row}")
    c = ws[f"{first}{row}"]
    c.value = f"CATEGORY {letter} — {name.upper()}  ·  {weight_pct}% WEIGHT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


# ---------------------------------------------------------------------------
# Sheet 1 — Start
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant, is_lite):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    title = "Listing SEO Audit"
    if is_lite:
        title += " — Lite"
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = title
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Score, fix list, re-audit date — your listing diagnosed."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    label_suffix = "LITE" if is_lite else variant.upper()
    c.value = f"{SKU} · v1.0 · {label_suffix}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: Listing name + score (rows 10-22)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(\'Listing Inputs\'!C7="","(enter listing name on Listing Inputs)",\'Listing Inputs\'!C7)'
    c.font = Font(name=FONT_HEAD, size=16, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 26

    # SCORE label
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "OVERALL SCORE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # Score readout — pulls from Audit Score!C2 (rolled-up percentage)
    ws.merge_cells("A14:L16")
    c = ws["A14"]
    c.value = '=TEXT(\'Audit Score\'!C2,"0")&" / 100"'
    c.font = Font(name=FONT_HEAD, size=42, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 24
    ws.row_dimensions[16].height = 18

    # Verdict band based on score
    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = (
        '=IF(\'Audit Score\'!C2>=90,"PARCHMENT — listing is dialed in",'
        'IF(\'Audit Score\'!C2>=80,"GOLD-SOFT — strong listing, polish remaining",'
        'IF(\'Audit Score\'!C2>=60,"GOLD — typical listing; clear room to grow",'
        '"RED — major fixes needed before paid traffic helps")))'
    )
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[18].height = 26

    # 5-category subscores (row 20) - small mono labels + scores
    cat_labels = [("A20", "TITLE"), ("C20", "DESC"), ("E20", "PHOTOS"),
                  ("G20", "AMEN"), ("I20", "TRUST"), ("K20", "TOTAL")]
    for cell_ref, txt in cat_labels:
        cell = ws[cell_ref]
        cell.value = txt
        cell.font = Font(name=FONT_MONO, size=8, bold=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Subscores row 21 — pull from Audit Score weighted category subtotals
    subscore_refs = [
        ("B21", "='Audit Score'!E5"),       # Cat A weighted score
        ("D21", "='Audit Score'!E6"),       # Cat B
        ("F21", "='Audit Score'!E7"),       # Cat C
        ("H21", "='Audit Score'!E8"),       # Cat D
        ("J21", "='Audit Score'!E9"),       # Cat E
        ("L21", "='Audit Score'!C11"),      # total
    ]
    for cell_ref, formula in subscore_refs:
        cell = ws[cell_ref]
        cell.value = formula
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
    ws.row_dimensions[20].height = 16
    ws.row_dimensions[21].height = 22

    # ZONE 3: Top 3 fixes (rows 24-30)
    ws.merge_cells("A24:L24")
    c = ws["A24"]
    c.value = "TOP 3 FIXES TO PRIORITIZE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[24].height = 22

    # Three rows: rank | criterion | fix
    # Pull from Audit Score sorted by lowest score × highest weight
    # We hard-link to dedicated rows on Audit Score (a small "top 3" helper block).
    for i in range(3):
        r = 25 + i
        # Rank
        ws.cell(row=r, column=1, value=f"{i+1}.").font = Font(
            name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT
        )
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        # Criterion text — pulls from Audit Score top-3 helper block
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]
        c.value = f"='Audit Score'!H{20+i}"   # criterion text
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        # Specific fix text
        ws.merge_cells(f"F{r}:L{r}")
        c = ws[f"F{r}"]
        c.value = f"='Audit Score'!I{20+i}"
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 32
        for col in range(1, 13):
            ws.cell(row=r, column=col).fill = parchment_fill

    # Re-audit date (row 29)
    ws.cell(row=29, column=2, value="Next re-audit by:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=29, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    ws.cell(row=29, column=2).fill = parchment_fill
    cell = ws.cell(row=29, column=3, value="='Re-audit Tracker'!C7")
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "yyyy-mm-dd"
    cell.fill = parchment_fill
    ws.merge_cells("E29:L29")
    c = ws["E29"]
    c.value = "Set on the Re-audit Tracker tab. Re-run after material listing changes."
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = parchment_fill
    ws.row_dimensions[29].height = 18

    # ZONE 4: Pseudo-button nav
    pseudo_button(ws, "A32", "C34", "Listing Inputs",
                  "'Listing Inputs'!A1", variant="primary")
    pseudo_button(ws, "D32", "F34", "Audit Score",
                  "'Audit Score'!A1", variant="primary")
    if not is_lite:
        pseudo_button(ws, "G32", "I34", "Comp Compare",
                      "'Comp Compare'!A1", variant="primary")
        pseudo_button(ws, "J32", "L34", "Methodology",
                      "'Methodology'!A1", variant="secondary")
    else:
        pseudo_button(ws, "G32", "I34", "Re-audit Tracker",
                      "'Re-audit Tracker'!A1", variant="primary")
        pseudo_button(ws, "J32", "L34", "Upgrade to Full →",
                      "Start!A1", variant="accent",
                      external_link=f"https://{BRAND_DOMAIN}/seo-audit-full")
    for r in range(32, 35):
        ws.row_dimensions[r].height = 22

    if not is_lite:
        pseudo_button(ws, "A36", "L38", "→  Re-audit Tracker",
                      "'Re-audit Tracker'!A1", variant="accent")
        for r in range(36, 39):
            ws.row_dimensions[r].height = 22

    # ZONE 5: Upgrade banner
    cta_row = 40 if not is_lite else 36
    if is_lite:
        cta = (
            "💡  Want comp-compare + public methodology page (40 criteria)? "
            f"Upgrade to Full at {BRAND_DOMAIN}/seo-audit-full — $47."
        )
    else:
        cta = (
            "💡  Auditing your whole portfolio? Get the Marketing Bundle at "
            f"{BRAND_DOMAIN} — SEO audit + review tracker + referral CRM, $97."
        )
    ws.merge_cells(f"A{cta_row}:L{cta_row}")
    c = ws[f"A{cta_row}"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[cta_row].height = 36

    footer_row = cta_row + 2
    brand_footer(ws, footer_row,
                 version_line=f"{SKU} · v1.0 · {label_suffix} · Free updates forever")

    ws.print_area = f"A1:L{footer_row + 4}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Sheet 2 — Listing Inputs
# ---------------------------------------------------------------------------

def build_listing_inputs_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Listing Inputs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 32), ("C", 24),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Listing Inputs",
                        prev_tab="Start", next_tab="Audit Score")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Paste your live listing details here. The Audit Score tab reads from these inputs."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # ---- LISTING IDENTITY (rows 6-9) ----
    _section_band(ws, 6, "LISTING IDENTITY")
    _input_row(ws, 7, "Listing name:",
               _val(variant, SAMPLE["listing_name"]), None,
               "How you'll reference this listing internally")
    _input_row(ws, 8, "Listing URL:",
               _val(variant, SAMPLE["listing_url"]), None,
               "airbnb.com/h/... (paste the public link)")

    # ---- TITLE BLOCK (rows 10-12) ----
    _section_band(ws, 10, "TITLE")
    _input_row(ws, 11, "Title text (paste):",
               _val(variant, SAMPLE["title_text"]), None,
               "Airbnb caps titles at 50 characters")
    _input_row(ws, 12, "Title char count:",
               _val(variant, SAMPLE["title_chars"]), "0",
               "Should be 35-50 — under is thin, over gets truncated")

    # ---- DESCRIPTION BLOCK (rows 14-17) ----
    _section_band(ws, 14, "DESCRIPTION")
    # Multi-line description in merged C15:L17
    ws.cell(row=15, column=2, value="Description:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=15, column=2).alignment = Alignment(
        horizontal="right", vertical="top", indent=1
    )
    ws.merge_cells("C15:L17")
    apply_style(ws["C15"], input_cell_style())
    ws["C15"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws["C15"].value = _val(variant, SAMPLE["description"])
    for r in range(15, 18):
        ws.row_dimensions[r].height = 32

    _input_row(ws, 18, "Description word count:",
               _val(variant, SAMPLE["desc_words"]), "0",
               "Sweet spot: 150-500 words")

    # ---- PHOTOS BLOCK (rows 20-29) ----
    _section_band(ws, 20, "PHOTOS")
    photos_fields = [
        (21, "Total photo count:",            _val(variant, SAMPLE["photo_count"]),     "0",   "Aim for ≥ 25"),
        (22, "Cover photo type:",             _val(variant, SAMPLE["cover_photo"]),     None, "Dropdown — should be Exterior or hero Living"),
        (23, "Bedroom photos count:",         _val(variant, SAMPLE["bedroom_photos"]),  "0",   "≥ 1 per bedroom"),
        (24, "Bathroom photos count:",        _val(variant, SAMPLE["bathroom_photos"]), "0",   "≥ 1 per bathroom"),
        (25, "Kitchen photos count:",         _val(variant, SAMPLE["kitchen_photos"]),  "0",   "Wide + 1 detail shot"),
        (26, "Outdoor photos count:",         _val(variant, SAMPLE["outdoor_photos"]),  "0",   "Porch, deck, yard, view"),
        (27, "Photos professional?",          _val(variant, SAMPLE["photos_pro"]),      None, "Yes/No — paid photographer or pro-grade DSLR"),
        (28, "Twilight / sunset photos?",     _val(variant, SAMPLE["twilight_photos"]), None, "Yes/No — at least 1-2"),
        (29, "Sequenced as guest journey?",   _val(variant, SAMPLE["photos_sequenced"]), None, "Yes/No — arrival → living → bedrooms → bath → kitchen → outdoor"),
    ]
    for row, label, value, fmt, note in photos_fields:
        _input_row(ws, row, label, value, fmt, note)
    add_dropdown(ws, "C22", '"Exterior wide,Interior living,Bedroom,Kitchen,Outdoor amenity,Other"')
    add_dropdown(ws, "C27", '"Yes,No"')
    add_dropdown(ws, "C28", '"Yes,No"')
    add_dropdown(ws, "C29", '"Yes,No"')

    # ---- AMENITIES BLOCK (rows 31-41) ----
    _section_band(ws, 31, "AMENITIES")
    amen_fields = [
        (32, "Total amenities listed:",       _val(variant, SAMPLE["amenity_count"]), "0",   "Walk through Airbnb's full list — aim for ≥ 40"),
        (33, "Hot tub?",                       _val(variant, SAMPLE["has_hot_tub"]),     None, ""),
        (34, "Pool?",                          _val(variant, SAMPLE["has_pool"]),         None, ""),
        (35, "Fireplace?",                     _val(variant, SAMPLE["has_fireplace"]),    None, ""),
        (36, "Fast WiFi (≥ 100 Mbps)?",        _val(variant, SAMPLE["has_fast_wifi"]),    None, ""),
        (37, "Workspace / desk?",              _val(variant, SAMPLE["has_workspace"]),    None, ""),
        (38, "EV charger?",                    _val(variant, SAMPLE["has_ev_charger"]),   None, ""),
        (39, "Pet-friendly?",                  _val(variant, SAMPLE["pet_friendly"]),     None, ""),
        (40, "Kid-friendly?",                  _val(variant, SAMPLE["kid_friendly"]),     None, ""),
        (41, "Self check-in?",                 _val(variant, SAMPLE["self_checkin"]),     None, ""),
    ]
    for row, label, value, fmt, note in amen_fields:
        _input_row(ws, row, label, value, fmt, note)
    for r in range(33, 42):
        add_dropdown(ws, f"C{r}", '"Yes,No"')

    # ---- TRUST SIGNALS BLOCK (rows 43-50) ----
    _section_band(ws, 43, "TRUST SIGNALS")
    trust_fields = [
        (44, "Response rate (%):",            _val(variant, SAMPLE["response_rate"]),  "0%",  "Airbnb dashboard shows this"),
        (45, "Response time:",                 _val(variant, SAMPLE["response_time"]),   None, "Dropdown — <1hr is the sweet spot"),
        (46, "Instant Book on?",               _val(variant, SAMPLE["instant_book"]),    None, "Yes/No"),
        (47, "Superhost?",                     _val(variant, SAMPLE["superhost"]),       None, "Yes/No"),
        (48, "Total reviews:",                 _val(variant, SAMPLE["review_count"]),    "0",   ""),
        (49, "Average rating:",                _val(variant, SAMPLE["avg_rating"]),       "0.00", ""),
        (50, "Cancellation policy:",           _val(variant, SAMPLE["cancellation"]),     None, "Dropdown"),
    ]
    for row, label, value, fmt, note in trust_fields:
        _input_row(ws, row, label, value, fmt, note)
    add_dropdown(ws, "C45", '"<1hr,1-12hr,12-24hr,>24hr"')
    add_dropdown(ws, "C46", '"Yes,No"')
    add_dropdown(ws, "C47", '"Yes,No"')
    add_dropdown(ws, "C50", '"Flexible,Moderate,Strict,Long-term"')

    brand_footer(ws, 53, version_line=f"{SKU} · Listing Inputs")


# ---------------------------------------------------------------------------
# Sheet 3 — Audit Score (40 criteria, weighted)
# ---------------------------------------------------------------------------

def build_audit_score_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Audit Score")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    # Columns:
    # A = ID, B = Criterion text, C = Met? (Y/N/Partial), D = points, E = (cat band)
    # F = "Specific fix" text
    # H/I (rows 20-22) = top-3 helper block (criterion text / fix)
    set_col_widths(ws, [
        ("A", 5), ("B", 56), ("C", 10), ("D", 9), ("E", 9),
        ("F", 56),
        ("G", 4), ("H", 38), ("I", 50),
        ("J", 6), ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "Audit Score",
                        prev_tab="Listing Inputs",
                        next_tab="Comp Compare" if not is_lite else "Re-audit Tracker")

    # ---- Big rollup score block (row 11, below totals) ----
    # Row 2 is inside the merged header band, so the rollup lives at row 11.
    # Start tab pulls from C11.
    ws.cell(row=11, column=2, value="OVERALL SCORE (0-100):").font = Font(
        name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=11, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    ws.row_dimensions[11].height = 28

    # Filter criteria for Lite variant
    filtered = []
    for letter, name, items in CRITERIA:
        kept = [it for it in items if (not is_lite) or it[0] in LITE_IDS]
        if kept:
            filtered.append((letter, name, kept))

    # Plan layout starting from row 25 (after header rollup block + 5 cat rows + helper).
    # Layout structure:
    #   row 5-9:   one row per category — F=weighted-score, E=points-earned, D=points-max
    #   row 11-13: rollup formula references / spacers
    #   row 14:    table header for the criteria table
    #   row 20-22: TOP 3 HELPER BLOCK (referenced from Start tab)
    #   row 25+:   criteria table with category bands

    # ---- Category subscore table (rows 5-9) ----
    ws.cell(row=4, column=2, value="CATEGORY SUBSCORES").font = Font(
        name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=4, column=2).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    cat_header_cells = [
        ("B4", "CATEGORY"), ("C4", "MAX PTS"),
        ("D4", "EARNED"), ("E4", "WEIGHTED"),
    ]
    for cell_ref, txt in cat_header_cells:
        cell = ws[cell_ref]
        cell.value = txt
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    # We'll fill cat subscore rows after we know criteria-row ranges.
    cat_subscore_rows = {}  # letter -> row

    # ---- TOP 3 HELPER BLOCK (rows 19-22) ----
    ws.cell(row=19, column=8, value="TOP 3 LOWEST-SCORING (CRITERION + FIX)").font = Font(
        name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=19, column=8).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.row_dimensions[19].height = 18
    # rows 20, 21, 22 will be filled via array formulas below

    # ---- Criteria table header (row 24) ----
    table_headers = [
        ("A24", "ID"), ("B24", "Criterion"), ("C24", "Met?"),
        ("D24", "Pts"), ("E24", "Max"), ("F24", "Specific fix"),
    ]
    for cell_ref, txt in table_headers:
        cell = ws[cell_ref]
        cell.value = txt
        apply_style(cell, header_row_style())
    ws.row_dimensions[24].height = 22

    # ---- Build criteria rows starting at row 25 ----
    current_row = 25
    cat_first_data_row = {}
    cat_last_data_row = {}

    for letter, name, items in filtered:
        weight = CATEGORY_WEIGHTS[letter]
        _category_band(ws, current_row, letter, name, weight, col_range="A:F")
        current_row += 1

        cat_first_data_row[letter] = current_row
        for (id_num, crit_text, fix_text, demo_met) in items:
            # A: ID
            ws.cell(row=current_row, column=1, value=id_num).font = Font(
                name=FONT_MONO, size=10, color=COLOR_MUTED
            )
            ws.cell(row=current_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            # B: Criterion text
            ws.cell(row=current_row, column=2, value=crit_text).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )
            ws.cell(row=current_row, column=2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True, indent=1
            )
            # C: Met? (input)
            met_value = demo_met if variant == "demo" else None
            cell = ws.cell(row=current_row, column=3, value=met_value)
            apply_style(cell, input_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # D: Points formula (Yes=10, Partial=5, No=0)
            cell = ws.cell(row=current_row, column=4,
                           value=f'=IF(C{current_row}="Yes",10,IF(C{current_row}="Partial",5,IF(C{current_row}="No",0,0)))')
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0"
            # E: Max points (always 10)
            cell = ws.cell(row=current_row, column=5, value=10)
            cell.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # F: Specific fix
            ws.cell(row=current_row, column=6, value=fix_text).font = Font(
                name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT
            )
            ws.cell(row=current_row, column=6).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True, indent=1
            )

            ws.row_dimensions[current_row].height = 30
            current_row += 1

        cat_last_data_row[letter] = current_row - 1

        # Subtotal row
        first = cat_first_data_row[letter]
        last = cat_last_data_row[letter]
        ws.cell(row=current_row, column=2,
                value=f"Category {letter} subtotal:").font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
        )
        ws.cell(row=current_row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        # Earned
        cell = ws.cell(row=current_row, column=4, value=f"=SUM(D{first}:D{last})")
        apply_style(cell, formula_cell_style())
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
        # Max
        cell = ws.cell(row=current_row, column=5, value=f"=SUM(E{first}:E{last})")
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
        ws.row_dimensions[current_row].height = 22

        cat_subscore_rows[letter] = current_row
        current_row += 2

    # ---- Fill category subscore table (rows 5-9) ----
    cat_order = ["A", "B", "C", "D", "E"]
    for i, letter in enumerate(cat_order):
        if letter not in cat_subscore_rows:
            # category not in Lite — still show 0/empty
            continue
        srow = 5 + i
        sub_row = cat_subscore_rows[letter]
        # Find category name
        name = next(n for L, n, _ in CRITERIA if L == letter)
        weight = CATEGORY_WEIGHTS[letter]

        ws.cell(row=srow, column=2,
                value=f"Category {letter} — {name}  ({weight}% weight)").font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        ws.cell(row=srow, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        # Max pts (col C) — sum of E in category range
        cell = ws.cell(row=srow, column=3, value=f"=E{sub_row}")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
        # Earned (col D)
        cell = ws.cell(row=srow, column=4, value=f"=D{sub_row}")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
        # Weighted score (col E) = (earned/max) * weight
        cell = ws.cell(row=srow, column=5,
                       value=f'=IF(C{srow}=0,0,(D{srow}/C{srow})*{weight})')
        apply_style(cell, formula_cell_style())
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.0"
        ws.row_dimensions[srow].height = 20

    # ---- Total row (row 10) — overall % ----
    total_weights = sum(CATEGORY_WEIGHTS[L] for L in cat_order if L in cat_subscore_rows)
    ws.cell(row=10, column=2, value=f"TOTAL  (max possible {total_weights})").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=10, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=10, column=5, value=f"=SUM(E5:E9)")
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0"
    ws.row_dimensions[10].height = 26

    # The OVERALL SCORE in C11 is the total scaled to 0-100.
    ws.cell(row=11, column=3, value=f"=E10/{total_weights}*100")
    ws.cell(row=11, column=3).font = Font(
        name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT
    )
    ws.cell(row=11, column=3).fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.cell(row=11, column=3).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.cell(row=11, column=3).number_format = "0"

    # Conditional formatting for overall score cell:
    # Red <60.
    ws.conditional_formatting.add(
        "C11",
        CellIsRule(operator="lessThan", formula=["60"],
                   fill=PatternFill("solid", fgColor=COLOR_ERROR),
                   font=Font(color=COLOR_WHITE, bold=True))
    )

    # Conditional formatting on Met? cells across the whole criteria range.
    # Red fill on "No", gold-soft on "Partial", parchment on "Yes"
    # We loop over each category data range.
    for letter in cat_order:
        if letter not in cat_first_data_row:
            continue
        first = cat_first_data_row[letter]
        last = cat_last_data_row[letter]
        rng = f"C{first}:C{last}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'$C{first}="No"'],
                fill=PatternFill("solid", fgColor=COLOR_ERROR),
                font=Font(color=COLOR_WHITE, bold=True),
                stopIfTrue=False,
            )
        )

    # ---- Top-3 helper block (rows 20-22) ----
    # Identify the three lowest-scoring criteria across all categories,
    # weighted by category weight × points-missed.
    # Excel approach: build a "loss" formula in column J for each criterion row,
    # then use LARGE / INDEX-MATCH at rows 20-22.
    # Add hidden-ish loss column in J (col 10) for each criterion row.
    # We use: loss = (max - earned) * category_weight / 10
    # Map criterion row -> category weight
    crit_rows = []
    for letter in cat_order:
        if letter not in cat_first_data_row:
            continue
        weight = CATEGORY_WEIGHTS[letter]
        for r in range(cat_first_data_row[letter], cat_last_data_row[letter] + 1):
            ws.cell(row=r, column=10, value=f"=(E{r}-D{r})*{weight}/10")
            ws.cell(row=r, column=10).font = Font(
                name=FONT_BODY, size=8, color=COLOR_MUTED
            )
            ws.cell(row=r, column=10).number_format = "0.00"
            crit_rows.append(r)

    # Hide column J visually (set narrow) and column K/L
    ws.column_dimensions["J"].width = 6
    ws.column_dimensions["J"].hidden = True

    # Build a single contiguous loss range covering the criteria rows.
    # Since rows have category bands and subtotal rows interleaved, we need
    # an array of separate ranges OR we accept the spread (subtotal rows have
    # no value in J because we didn't set them — but they live in the range).
    # Simplest robust approach: include from first_data_row of cat A through
    # last_data_row of cat E, then in helper compute use IFERROR to skip blanks.
    first_crit = cat_first_data_row["A"]
    last_crit = cat_last_data_row[cat_order[-1] if cat_order[-1] in cat_last_data_row else "E"]
    loss_range = f"J{first_crit}:J{last_crit}"
    crit_text_range = f"B{first_crit}:B{last_crit}"
    fix_range = f"F{first_crit}:F{last_crit}"

    for i in range(3):
        helper_row = 20 + i
        # Rank the i-th largest loss value
        rank_formula_loss = f'=LARGE({loss_range},{i+1})'
        # Find the criterion text matching that loss (use INDEX/MATCH)
        # We add a small disambiguator i*0.0001 so duplicate losses don't collide.
        # Since we can't easily mutate the J column with row-aware tiebreaks here,
        # we accept that duplicates show the first match — fine for a Top-3.
        crit_formula = (
            f'=IFERROR(INDEX({crit_text_range},'
            f'MATCH(LARGE({loss_range},{i+1}),{loss_range},0)),"")'
        )
        fix_formula = (
            f'=IFERROR(INDEX({fix_range},'
            f'MATCH(LARGE({loss_range},{i+1}),{loss_range},0)),"")'
        )
        # column H = criterion text
        ws.cell(row=helper_row, column=8, value=crit_formula).font = Font(
            name=FONT_BODY, size=10, color=COLOR_TEXT
        )
        ws.cell(row=helper_row, column=8).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        # column I = specific fix
        ws.cell(row=helper_row, column=9, value=fix_formula).font = Font(
            name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED
        )
        ws.cell(row=helper_row, column=9).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.row_dimensions[helper_row].height = 22

    # Data validation on Met? column for all criterion rows
    for letter in cat_order:
        if letter not in cat_first_data_row:
            continue
        first = cat_first_data_row[letter]
        last = cat_last_data_row[letter]
        add_dropdown(ws, f"C{first}:C{last}", '"Yes,Partial,No"')

    # Final total-rollup row at very bottom of criteria
    final_total_row = current_row
    ws.cell(row=final_total_row, column=2,
            value="OVERALL  (sum of weighted category scores)").font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=final_total_row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=final_total_row, column=4, value=f"=SUM(D5:D9)")
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0"
    cell = ws.cell(row=final_total_row, column=5, value=f"=E10")
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.0"
    ws.row_dimensions[final_total_row].height = 26

    brand_footer(ws, final_total_row + 3, version_line=f"{SKU} · Audit Score")


# ---------------------------------------------------------------------------
# Sheet 4 — Top-5 Comp Compare (Full only)
# ---------------------------------------------------------------------------

def build_comp_compare_tab(wb, variant):
    ws = wb.create_sheet("Comp Compare")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 5), ("B", 50),
        ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 12), ("H", 12),
        ("I", 6), ("J", 6), ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "Comp Compare",
                        prev_tab="Audit Score", next_tab="Methodology")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Paste 5 competitor listings as you'd score them. The grid surfaces where comps beat you."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Comp name row (row 6)
    ws.cell(row=6, column=2, value="Comp →").font = Font(
        name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY
    )
    ws.cell(row=6, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    # Mine in C6
    cell = ws.cell(row=6, column=3,
                   value="='Listing Inputs'!C7" if variant == "demo" else "Mine")
    cell.font = Font(name=FONT_HEAD, size=11, bold=True, color=COLOR_ACCENT)
    cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Comp 1-5 in D-H
    for i in range(5):
        col = 4 + i
        cell = ws.cell(row=6, column=col,
                       value=_val(variant, COMP_NAMES[i]))
        apply_style(cell, input_cell_style())
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[6].height = 36

    # Header row (row 7)
    headers = [("A7", "ID"), ("B7", "Criterion"), ("C7", "Mine"),
               ("D7", "Comp 1"), ("E7", "Comp 2"), ("F7", "Comp 3"),
               ("G7", "Comp 4"), ("H7", "Comp 5")]
    for cell_ref, txt in headers:
        cell = ws[cell_ref]
        cell.value = txt
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 22

    # Criterion rows starting at row 9 (with category bands)
    current_row = 9
    crit_index = 0
    for letter, name, items in CRITERIA:
        weight = CATEGORY_WEIGHTS[letter]
        # Category band across A-H
        ws.merge_cells(f"A{current_row}:H{current_row}")
        c = ws[f"A{current_row}"]
        c.value = f"CATEGORY {letter} — {name.upper()}  ·  {weight}% WEIGHT"
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for (id_num, crit_text, fix_text, demo_met) in items:
            ws.cell(row=current_row, column=1, value=id_num).font = Font(
                name=FONT_MONO, size=9, color=COLOR_MUTED
            )
            ws.cell(row=current_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=current_row, column=2, value=crit_text).font = Font(
                name=FONT_BODY, size=10, color=COLOR_TEXT
            )
            ws.cell(row=current_row, column=2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True, indent=1
            )
            # Mine — pull from Audit Score
            # Match by ID via INDEX/MATCH on column A of Audit Score
            mine_formula = (
                f'=IFERROR(INDEX(\'Audit Score\'!$C:$C,'
                f'MATCH({id_num},\'Audit Score\'!$A:$A,0)),"")'
            )
            cell = ws.cell(row=current_row, column=3, value=mine_formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            # Comp 1-5
            for j in range(5):
                col = 4 + j
                comp_value = COMP_SCORES[COMP_NAMES[j]][crit_index] if variant == "demo" else None
                cell = ws.cell(row=current_row, column=col, value=comp_value)
                apply_style(cell, input_cell_style())
                cell.alignment = Alignment(horizontal="center", vertical="center")
                add_dropdown(ws, f"{get_column_letter(col)}{current_row}",
                             '"Yes,Partial,No"')

            ws.row_dimensions[current_row].height = 22
            current_row += 1
            crit_index += 1

    # Conditional formatting — highlight comp cells that beat me (Yes when I'm No/Partial)
    # We use a per-row formula rule on each comp column.
    crit_data_first = 10  # first criterion row (after first cat band at row 9)
    crit_data_last = current_row - 1
    for col_letter in ["D", "E", "F", "G", "H"]:
        # Highlight when comp = "Yes" AND mine != "Yes"
        ws.conditional_formatting.add(
            f"{col_letter}{crit_data_first}:{col_letter}{crit_data_last}",
            FormulaRule(
                formula=[f'AND(${col_letter}{crit_data_first}="Yes",$C{crit_data_first}<>"Yes")'],
                fill=PatternFill("solid", fgColor=COLOR_ERROR),
                font=Font(color=COLOR_WHITE, bold=True),
                stopIfTrue=False,
            )
        )

    # Footnote
    ws.merge_cells(f"A{current_row + 1}:H{current_row + 2}")
    c = ws[f"A{current_row + 1}"]
    c.value = (
        "Red cells = comp beats you on that criterion. Score each comp the same way "
        "you scored your own listing. Pick comps that are roughly the same beds/baths/lot."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(current_row + 1, current_row + 3):
        ws.row_dimensions[r].height = 16

    brand_footer(ws, current_row + 5, version_line=f"{SKU} · Comp Compare")


# ---------------------------------------------------------------------------
# Sheet 5 — Methodology (Full only)
# ---------------------------------------------------------------------------

def build_methodology_tab(wb, variant):
    ws = wb.create_sheet("Methodology")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 5), ("B", 50), ("C", 9),
        ("D", 50), ("E", 28), ("F", 14),
        ("G", 4), ("H", 4), ("I", 4),
        ("J", 4), ("K", 4), ("L", 4),
    ])

    compact_header_band(ws, "Methodology",
                        prev_tab="Comp Compare", next_tab="Re-audit Tracker")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = "How we score — published rubric. Public-facing. Last updated 2026-05-01."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 20

    # Header row (row 6)
    headers = [("A6", "ID"), ("B6", "Criterion"), ("C6", "Weight%"),
               ("D6", "Why it matters"), ("E6", "Source"), ("F6", "Last verified")]
    for cell_ref, txt in headers:
        cell = ws[cell_ref]
        cell.value = txt
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Source / why-it-matters mapping (concise, evidence-cited)
    WHY = {
        1:  ("Airbnb truncates titles past 50 chars on mobile search results — your hook gets cut.", "Airbnb published"),
        2:  ("Location keywords match guest search queries directly. ~70% of searches include geo.", "Airbnb published"),
        3:  ("Top-tier amenities (hot tub, view) are the #1 booking decision driver.", "Hospitable 2024 study"),
        4:  ("Caps and emojis trigger algorithm-spam heuristics + reduce perceived professionalism.", "Host community consensus"),
        5:  ("Differentiating phrases capture the 'experience' signal Airbnb's algorithm rewards.", "AirDNA Marketscore"),
        6:  ("Generic openings lose the first-3-second skim test. Hook = retention.", "Hospitable 2024 study"),
        7:  ("<150 words = thin signal. >500 words = guests skim and miss key info.", "Host community consensus"),
        8:  ("Guests decide in 3 sentences. Bury the lead and you lose them.", "Conversion best practice"),
        9:  ("Walkability/drive times answer the silent question every guest has.", "Airbnb host blog"),
        10: ("Sleeping config is the #1 question guests message about.", "Host community consensus"),
        11: ("Scannable amenity list reduces messages and improves conversion.", "Hospitable 2024 study"),
        12: ("Positive-framed rules signal hospitality. Negative framing kills bookings.", "Host community consensus"),
        13: ("Response promise sets expectations and improves message-to-book rate.", "Conversion best practice"),
        14: ("Listings with 25+ photos out-book sub-25 by ~30% on average.", "AirDNA 2023 study"),
        15: ("Cover photo is the click-through trigger. Bedroom covers underperform exterior 2:1.", "AirDNA 2023 study"),
        16: ("Guests filter by bedroom count — they expect to see each one.", "Airbnb published"),
        17: ("Bathroom photos signal cleanliness, the #1 review factor.", "Airbnb review data"),
        18: ("Kitchen is the second-most-photographed room in winning listings.", "AirDNA 2023 study"),
        19: ("Living photo establishes the 'hangout' value — drives long-stay bookings.", "Conversion best practice"),
        20: ("Outdoor space signals seasonal value, hot tubs convert highest.", "Hospitable 2024 study"),
        21: ("Twilight photos beat daytime photos on engagement by ~40%.", "Hospitable 2024 study"),
        22: ("Pro photos pay back in 30-60 days. Highest ROI marketing spend.", "STR investor consensus"),
        23: ("Guest journey order matches mental model: arrive → live → sleep → eat → outside.", "UX best practice"),
        24: ("Captioned photos rank higher in Airbnb search — accessibility signal.", "Airbnb published"),
        25: ("Mirror-selfies and clutter signal amateur listing → lower trust.", "Host community consensus"),
        26: ("Amenity count is a direct algorithm completeness signal.", "Airbnb published"),
        27: ("Top-tier amenities have explicit search filters — must be checked to appear.", "Airbnb published"),
        28: ("WiFi speed in description converts remote workers (30%+ weekday demand).", "AirDNA workforce data"),
        29: ("Workspace amenity has its own search filter — easy ranking lift.", "Airbnb published"),
        30: ("Self check-in unlocks Instant Book guest segments with separate filter.", "Airbnb published"),
        31: ("Pet policy clarity reduces wrong-fit bookings and bad reviews.", "Host community consensus"),
        32: ("Family filters drive 40%+ of weekend bookings in family markets.", "AirDNA family data"),
        33: ("Accessibility tags expand qualified-buyer pool with no inventory change.", "Airbnb published"),
        34: ("Response time <1hr is the #1 published Airbnb ranking factor.", "Airbnb published"),
        35: ("Response rate <90% suppresses ranking algorithmically.", "Airbnb published"),
        36: ("Instant Book is the single largest published ranking lift available.", "Airbnb published"),
        37: ("Superhost gets a separate badge + dedicated search filter.", "Airbnb published"),
        38: ("Strict cancellation policies reduce conversion 15-20%.", "Hospitable 2024 study"),
        39: ("4.8+ avg with 20+ reviews is the threshold for top-of-search visibility.", "Airbnb published"),
        40: ("ID-verified guests reduce party / damage risk + improve filter-pool size.", "Airbnb published"),
    }

    current_row = 7
    for letter, name, items in CRITERIA:
        weight = CATEGORY_WEIGHTS[letter]
        n_items = len(items)
        per_item_weight = weight / n_items
        # Category band
        ws.merge_cells(f"A{current_row}:F{current_row}")
        c = ws[f"A{current_row}"]
        c.value = f"CATEGORY {letter} — {name.upper()}  ·  {weight}% TOTAL  ·  {per_item_weight:.2f}% PER CRITERION"
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for (id_num, crit_text, fix_text, demo_met) in items:
            why_text, source = WHY.get(id_num, ("Evidence under review.", "Pending"))
            ws.cell(row=current_row, column=1, value=id_num).font = Font(
                name=FONT_MONO, size=9, color=COLOR_MUTED
            )
            ws.cell(row=current_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=current_row, column=2, value=crit_text).font = Font(
                name=FONT_BODY, size=10, color=COLOR_TEXT
            )
            ws.cell(row=current_row, column=2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True, indent=1
            )
            ws.cell(row=current_row, column=3,
                    value=round(per_item_weight, 2)).font = Font(
                name=FONT_BODY, size=10, color=COLOR_PRIMARY, bold=True
            )
            ws.cell(row=current_row, column=3).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=current_row, column=3).number_format = "0.00"
            ws.cell(row=current_row, column=4, value=why_text).font = Font(
                name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT
            )
            ws.cell(row=current_row, column=4).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True, indent=1
            )
            ws.cell(row=current_row, column=5, value=source).font = Font(
                name=FONT_MONO, size=9, color=COLOR_MUTED
            )
            ws.cell(row=current_row, column=5).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            ws.cell(row=current_row, column=6, value="2026-05-01").font = Font(
                name=FONT_BODY, size=9, color=COLOR_MUTED
            )
            ws.cell(row=current_row, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.row_dimensions[current_row].height = 32
            current_row += 1

    # Footer note
    ws.merge_cells(f"A{current_row + 1}:F{current_row + 3}")
    c = ws[f"A{current_row + 1}"]
    c.value = (
        "Methodology updated when Airbnb publishes algorithm changes. Re-verified quarterly. "
        f"Sources: Airbnb published guidelines · Hospitable host studies · AirDNA Marketscore · STR investor community consensus. "
        f"Questions on methodology? {BRAND_EMAIL}."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(current_row + 1, current_row + 4):
        ws.row_dimensions[r].height = 14

    brand_footer(ws, current_row + 5,
                 version_line=f"{SKU} · Methodology · v1.0 · 2026-05-01")


# ---------------------------------------------------------------------------
# Sheet 6 — Re-audit Tracker
# ---------------------------------------------------------------------------

def build_reaudit_tracker_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Re-audit Tracker")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 28), ("C", 16),
        ("D", 14), ("E", 50),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Re-audit Tracker",
                        prev_tab="Methodology" if not is_lite else "Audit Score",
                        next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Schedule your next audit and log every prior re-run. Track score progression."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Next re-audit date input (row 6-7)
    _section_band(ws, 6, "NEXT RE-AUDIT")
    ws.cell(row=7, column=2, value="Re-audit by:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=7, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=7, column=3, value=_val(variant, SAMPLE["reaudit_date"]))
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "yyyy-mm-dd"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    ws.merge_cells("E7:L7")
    c = ws["E7"]
    c.value = "Re-run after pro photos refresh, description rewrite, or pricing/ADR change."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 24

    # Days remaining (row 8)
    ws.cell(row=8, column=2, value="Days until next audit:").font = Font(
        name=FONT_BODY, size=11, color=COLOR_TEXT
    )
    ws.cell(row=8, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    cell = ws.cell(row=8, column=3,
                   value='=IF(C7="","",C7-TODAY())')
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0"

    # Conditional formatting on days-remaining: red if past due (negative), gold soon (<30)
    ws.conditional_formatting.add(
        "C8",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=COLOR_ERROR),
                   font=Font(color=COLOR_WHITE, bold=True))
    )
    ws.conditional_formatting.add(
        "C8",
        CellIsRule(operator="between", formula=["0", "30"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
                   font=Font(color=COLOR_PRIMARY, bold=True))
    )

    # ---- Audit history (rows 11-17) ----
    _section_band(ws, 11, "AUDIT HISTORY  (5 most recent)")
    headers = [("B12", "Audit date"), ("C12", "Score"),
               ("D12", "Δ vs prior"), ("E12", "Notes / what changed")]
    for cell_ref, txt in headers:
        cell = ws[cell_ref]
        cell.value = txt
        apply_style(cell, header_row_style())
    ws.row_dimensions[12].height = 22

    # 5 history rows starting at row 13
    history = SAMPLE_HISTORY if variant == "demo" else [(None, None, "")] * 4
    for i in range(5):
        r = 13 + i
        if i < len(history):
            date_val, score_val, note_val = history[i] if variant == "demo" and i < len(history) else (None, None, None)
        else:
            date_val, score_val, note_val = (None, None, None)

        # Date
        cell = ws.cell(row=r, column=2, value=date_val)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "yyyy-mm-dd"
        # Score
        cell = ws.cell(row=r, column=3, value=score_val)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0"
        # Delta — formula referencing prior row
        if i == 0:
            delta_formula = '=""'
        else:
            delta_formula = f'=IF(OR(C{r}="",C{r-1}=""),"",C{r}-C{r-1})'
        cell = ws.cell(row=r, column=4, value=delta_formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "+0;-0;0"
        # Note
        cell = ws.cell(row=r, column=5, value=note_val)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[r].height = 20

    # Conditional formatting on delta column — green positive, red negative
    ws.conditional_formatting.add(
        "D13:D17",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
                   font=Font(color=COLOR_PRIMARY, bold=True))
    )
    ws.conditional_formatting.add(
        "D13:D17",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=COLOR_ERROR),
                   font=Font(color=COLOR_WHITE, bold=True))
    )

    # Cadence guidance (rows 20-23)
    _section_band(ws, 20, "WHEN TO RE-AUDIT")
    cadence_items = [
        ("Quarterly minimum.",
         "STR markets shift season-to-season. A Q1 score is stale by Q3."),
        ("After every material listing change.",
         "New photos, rewritten description, amenity changes, pricing changes."),
        ("After Airbnb algorithm announcements.",
         "Methodology page updates after Airbnb publishes changes — re-run to verify."),
    ]
    for i, (head, body) in enumerate(cadence_items):
        r = 21 + i
        ws.cell(row=r, column=2, value=head).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.merge_cells(f"C{r}:L{r}")
        c = ws[f"C{r}"]
        c.value = body
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 26, version_line=f"{SKU} · Re-audit Tracker")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant, is_lite):
    wb = Workbook()
    build_start_tab(wb, variant, is_lite)
    build_listing_inputs_tab(wb, variant, is_lite)
    build_audit_score_tab(wb, variant, is_lite)
    if not is_lite:
        build_comp_compare_tab(wb, variant)
        build_methodology_tab(wb, variant)
    build_reaudit_tracker_tab(wb, variant, is_lite)

    wb.properties.title = "Listing SEO Audit — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "40-criteria Airbnb listing audit with weighted scoring, prioritized "
        "fix list, methodology page, comp compare, and re-audit cadence tracker."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo", is_lite=False)
    build_workbook(BLANK_OUT, "blank", is_lite=False)
    build_workbook(LITE_OUT, "demo", is_lite=True)


if __name__ == "__main__":
    main()
