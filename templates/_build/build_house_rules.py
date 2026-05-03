"""Build GST-002 House Rules Builder — wizard-style tool.

Walks the host through 6 sections of house rules and outputs a printable,
listing-ready rules document via a formula-driven Rules Output tab.

Ships two files:
  templates/_masters/GST-002-house-rules-builder-DEMO.xlsx
  templates/_masters/GST-002-house-rules-builder-BLANK.xlsx

Tabs (9):
  0  Start                  — wizard hero
  1  §1 Property + Capacity — address, capacity, sleeping, parking
  2  §2 Quiet Hours + Parties
  3  §3 Smoking + Substances
  4  §4 Pets
  5  §5 Damage + Liability
  6  §6 Check-out + Penalties
  7  Launch                 — readiness + BUILD RULES DOC button
  8  Rules Output           — formula-concatenated, print-ready rules
"""
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_NAVY_TINT, COLOR_NAVY_SHADE, COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    pseudo_button, card_header, card_body_fill, section_header_band,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
)

BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / "GST-002-house-rules-builder-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / "GST-002-house-rules-builder-BLANK.xlsx"

# --- Tab structure ---

TAB_PROP = "§1 Property + Capacity"
TAB_QUIET = "§2 Quiet Hours + Parties"
TAB_SMOKE = "§3 Smoking + Substances"
TAB_PETS = "§4 Pets"
TAB_DAMAGE = "§5 Damage + Liability"
TAB_CHECKOUT = "§6 Check-out + Penalties"

TAB_NAMES = [
    "Start",
    TAB_PROP,
    TAB_QUIET,
    TAB_SMOKE,
    TAB_PETS,
    TAB_DAMAGE,
    TAB_CHECKOUT,
    "Launch",
    "Rules Output",
]

# Inputs per tab (col B starting at row 8) for progress formula
SECTION_INPUT_COUNTS = {
    TAB_PROP:     6,   # property name, address, max guests, bedrooms, sleeping, parking
    TAB_QUIET:    6,   # quiet start, quiet end, parties policy, gathering size, guests-of-guests, outdoor music
    TAB_SMOKE:    5,   # smoking inside, smoking outside, vaping, weed, alcohol notes
    TAB_PETS:     5,   # pets allowed, pet count, weight limit, pet fee, pet rules
    TAB_DAMAGE:   6,   # furniture moving, pool/hot tub, fireplace, kids policy, max occupancy, damage policy
    TAB_CHECKOUT: 6,   # checkout time, late checkout fee, missing-item fee, smoking fine, key return, deposit policy
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 34
TOTAL_SECTIONS = 6

# --- Sample data (DEMO variant) ---

SAMPLE = {
    "Property": {
        "name":       "Smokies Ridge Cabin",
        "address":    "123 Mountain Lane, Gatlinburg, TN 37738",
        "max_guests": 10,
        "bedrooms":   "4 BR / 3 BA",
        "sleeping":   "1 King, 2 Queens, 2 Twin bunk-beds, 1 Queen sleeper-sofa",
        "parking":    "Gravel drive — fits 3 cars. No street parking after 8 PM.",
    },
    "Quiet": {
        "quiet_start":   "10:00 PM",
        "quiet_end":     "8:00 AM",
        "parties":       "No parties or events of any kind",
        "gathering":     "Day visitors OK until 9 PM (max 4 above booked guest count)",
        "guests_guests": "Must be approved in advance via message",
        "outdoor_music": "No outdoor amplified music after sunset",
    },
    "Smoke": {
        "inside":  "No smoking, vaping, or cannabis use anywhere inside",
        "outside": "Smoking permitted only on the gravel turnaround — butts in metal can",
        "vaping":  "Outdoor only — same rules as smoking",
        "weed":    "Outdoor only where state-legal; never inside",
        "alcohol": "Adults only; please drink responsibly. No kegs.",
    },
    "Pets": {
        "allowed":      "Yes — dogs only",
        "count":        "Max 2 dogs",
        "weight":       "Each dog under 50 lb",
        "fee":          "$50 per pet per stay (non-refundable)",
        "rules":        "Crate when unattended; no pets on furniture; clean up yard waste daily",
    },
    "Damage": {
        "furniture":  "Please do not move beds or living-room furniture",
        "hot_tub":    "Hot tub closes at 10 PM; no glass; shower before entering; max 6 adults",
        "fireplace":  "Wood-burning fireplace is OFF for guest stays — do not light",
        "kids":       "Children must be supervised in loft and on the deck at all times",
        "occupancy":  "Maximum 10 guests overnight (matches Airbnb listing)",
        "damage":     "Damage above $250 will be reported to Airbnb resolution / VRBO claim",
    },
    "Checkout": {
        "time":        "10:00 AM",
        "late_fee":    "$50 per hour past 10 AM (must request in advance)",
        "missing":     "$25 + replacement cost per missing item",
        "smoke_fine":  "$250 deep-clean fee if smoking detected indoors",
        "keys":        "Smart-lock — no key return needed; do not change the code",
        "deposit":     "Damage deposit refunded within 7 days of check-out, post-inspection",
    },
}


# --- Local helpers ---

def add_dropdown(ws, cell_range, options):
    """Attach an in-cell dropdown."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


@dataclass
class Card:
    """One input card inside an input tab."""
    header: str
    rows: list = field(default_factory=list)
    static: list = field(default_factory=list)
    row_height: int = 24


def build_input_tab(wb, section_num, tab_name, title, subtitle, cards,
                    prev_tab, next_tab):
    """Render an input tab using the v2.1 flattened layout (forked from
    welcome book wizard reference).
    """
    ws = wb[tab_name]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    section_header_band(ws, section_num, TOTAL_SECTIONS, title, subtitle,
                        prev_tab, next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction strip
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("Fill the highlighted fields below. The 'Launch' tab "
               "(last) builds your print-ready rules document.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: section banner
    card_header(ws, 7, ("A", "L"), title)

    ws.freeze_panes = "A8"

    current_row = 8
    is_first_card = True
    body_start = current_row

    for card in cards:
        if not is_first_card and card.rows:
            gold_side = Side(style="thin", color=COLOR_ACCENT)
            for col in range(1, 13):
                existing = ws.cell(row=current_row, column=col).border
                ws.cell(row=current_row, column=col).border = Border(
                    top=gold_side,
                    bottom=existing.bottom,
                    left=existing.left,
                    right=existing.right,
                )

        for row_spec in card.rows:
            if len(row_spec) == 2:
                label, value = row_spec
                options = None
            elif len(row_spec) == 3:
                label, value, options = row_spec
            else:
                raise ValueError(f"Card row must be 2/3 tuple, got {row_spec}")

            lc = ws.cell(row=current_row, column=1)
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center",
                                     indent=1)

            ws.merge_cells(f"B{current_row}:L{current_row}")
            ic = ws[f"B{current_row}"]
            if value != "" and value is not None:
                if isinstance(value, str) and value.startswith("="):
                    ic.value = value
                    apply_style(ic, formula_cell_style())
                else:
                    ic.value = value
                    apply_style(ic, input_cell_style())
                    ic.alignment = Alignment(horizontal="left",
                                             vertical="center",
                                             wrap_text=True, indent=1)
            else:
                apply_style(ic, input_cell_style())

            if options:
                add_dropdown(ws, f"B{current_row}", options)

            ws.row_dimensions[current_row].height = card.row_height
            current_row += 1

        for static_text in card.static:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            sc = ws[f"A{current_row}"]
            sc.value = static_text
            sc.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
            sc.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=2)
            sc.fill = parchment_fill
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        is_first_card = False

    body_end = current_row - 1
    if body_end >= body_start:
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)

    # Section footer with secondary nav
    footer_row = current_row + 1
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=current_row, column=c).border = Border(top=gold_side)
        ws.cell(row=current_row, column=c).fill = parchment_fill

    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = "Next: Launch →"
        next_target = "'Launch'!A1"

    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                  prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                  next_label, next_target, variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22


# --- Start tab ---

def build_start_tab(wb, variant):
    """Tab 0 — Start. Six-zone wizard hero."""
    ws = wb.active
    ws.title = TAB_NAMES[0]
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: HERO BAND (rows 1-8, navy)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "House Rules Builder"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Specific, defensible rules — not 4 generic bullets."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "GST-002 · v2.0"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: WHAT YOU'LL BUILD (rows 10-20, parchment)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "What you'll build in the next 12 minutes"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    cards = [
        ("LISTING-READY", "Paste straight into Airbnb / VRBO 'House rules' field."),
        ("PRINTABLE PDF", "One-page rules sheet for the property binder."),
        ("DEFENSIBLE", "25-40 specific rules — not 4 generic bullets."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (title, desc) in enumerate(cards):
        first, last = col_groups[idx]
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]
        c.value = title
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]
        c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # ZONE 3: QUICK START (rows 22-28)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = "Quick Start — be done in 10 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24

    quickstart_items = [
        "① Property name + capacity",
        "② Quiet hours window",
        "③ Smoking policy",
        "④ Pet policy",
        "⑤ Check-out time + fees",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + i if i < 3 else 24 + (i - 3)
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ZONE 4: GET STARTED (rows 30-33)
    pseudo_button(ws, "A30", "L33",
                  "GET STARTED — FILL §1 PROPERTY + CAPACITY  →",
                  f"'{TAB_PROP}'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # ZONE 5: PROGRESS DASHBOARD (rows 35-43)
    ws.merge_cells("A36:F36")
    c = ws["A36"]
    c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ranges = [
        f"'{TAB_PROP}'!B8:B13",
        f"'{TAB_QUIET}'!B8:B13",
        f"'{TAB_SMOKE}'!B8:B12",
        f"'{TAB_PETS}'!B8:B12",
        f"'{TAB_DAMAGE}'!B8:B13",
        f"'{TAB_CHECKOUT}'!B8:B13",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)

    ws.merge_cells("G36:L36")
    c = ws["G36"]
    c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS}, "0%") & " complete"'
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    section_rows = [
        ("① Property + Capacity",     TAB_PROP,     "B8:B13", 6),
        ("② Quiet Hours + Parties",   TAB_QUIET,    "B8:B13", 6),
        ("③ Smoking + Substances",    TAB_SMOKE,    "B8:B12", 5),
        ("④ Pets",                    TAB_PETS,     "B8:B12", 5),
        ("⑤ Damage + Liability",      TAB_DAMAGE,   "B8:B13", 6),
        ("⑥ Check-out + Penalties",   TAB_CHECKOUT, "B8:B13", 6),
    ]
    for i, (label, tab, range_, total) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(f"G{r}:J{r}")
        ca = f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (
            f'=IF({ca}={total},"✅ Done",'
            f'IF({ca}=0,"⏳ Empty",'
            f'"⏳ "&{ca}&" of {total}"))'
        )
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # ZONE 6: FOOTER (rows 46-48)
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=46, column=c).border = Border(top=gold_side)

    ws.merge_cells("A47:L47")
    c = ws["A47"]
    c.value = f"Questions? {BRAND_EMAIL}"
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A48:L48")
    c = ws["A48"]
    c.value = "Free updates forever · v2.0 · Released 2026-05"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.print_area = "A1:L48"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# --- Section input tabs ---

def build_property_tab(wb, variant):
    wb.create_sheet(TAB_PROP)
    s = SAMPLE["Property"] if variant == "demo" else {}
    cards = [
        Card(
            header="Identity",
            rows=[
                ("Property name:", s.get("name", "")),
                ("Full address:", s.get("address", "")),
            ],
            row_height=28,
        ),
        Card(
            header="Capacity",
            rows=[
                ("Max overnight guests:", s.get("max_guests", "")),
                ("Bedrooms / bathrooms:", s.get("bedrooms", "")),
                ("Sleeping arrangements:", s.get("sleeping", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Parking",
            rows=[
                ("Parking instructions:", s.get("parking", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=1, tab_name=TAB_PROP,
        title="§1 Property + Capacity",
        subtitle="The basics that frame every other rule.",
        cards=cards, prev_tab="", next_tab=TAB_QUIET,
    )


def build_quiet_tab(wb, variant):
    wb.create_sheet(TAB_QUIET)
    s = SAMPLE["Quiet"] if variant == "demo" else {}
    cards = [
        Card(
            header="Quiet Hours",
            rows=[
                ("Quiet hours start (e.g., 10:00 PM):", s.get("quiet_start", "")),
                ("Quiet hours end (e.g., 8:00 AM):", s.get("quiet_end", "")),
            ],
        ),
        Card(
            header="Parties + Visitors",
            rows=[
                ("Parties / events policy:", s.get("parties", ""),
                 ["No parties or events of any kind",
                  "Small gatherings OK with prior approval",
                  "Events allowed (with cleaning fee)",
                  "Custom — see notes"]),
                ("Day visitors / gatherings:", s.get("gathering", "")),
                ("Guests of guests rule:", s.get("guests_guests", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Outdoor Sound",
            rows=[
                ("Outdoor music after dark:", s.get("outdoor_music", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=2, tab_name=TAB_QUIET,
        title="§2 Quiet Hours + Parties",
        subtitle="Set the noise floor — neighbors are listening.",
        cards=cards, prev_tab=TAB_PROP, next_tab=TAB_SMOKE,
    )


def build_smoke_tab(wb, variant):
    wb.create_sheet(TAB_SMOKE)
    s = SAMPLE["Smoke"] if variant == "demo" else {}
    cards = [
        Card(
            header="Smoking + Vaping",
            rows=[
                ("Smoking — inside the home:", s.get("inside", ""),
                 ["No smoking, vaping, or cannabis use anywhere inside",
                  "Smoking allowed in designated room only",
                  "Custom — see notes"]),
                ("Smoking — outside / designated area:", s.get("outside", "")),
                ("Vaping policy:", s.get("vaping", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Cannabis + Alcohol",
            rows=[
                ("Cannabis / weed policy:", s.get("weed", "")),
                ("Alcohol notes:", s.get("alcohol", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=3, tab_name=TAB_SMOKE,
        title="§3 Smoking + Substances",
        subtitle="Smoke smell is the #1 cause of $250+ cleaning claims.",
        cards=cards, prev_tab=TAB_QUIET, next_tab=TAB_PETS,
    )


def build_pets_tab(wb, variant):
    wb.create_sheet(TAB_PETS)
    s = SAMPLE["Pets"] if variant == "demo" else {}
    cards = [
        Card(
            header="Pet Policy",
            rows=[
                ("Pets allowed?", s.get("allowed", ""),
                 ["No pets — service animals only",
                  "Yes — dogs only",
                  "Yes — dogs and cats",
                  "Yes — case-by-case approval"]),
                ("Number of pets allowed:", s.get("count", "")),
                ("Weight limit per pet:", s.get("weight", "")),
            ],
        ),
        Card(
            header="Pet Fees + Rules",
            rows=[
                ("Pet fee:", s.get("fee", "")),
                ("Pet ground rules:", s.get("rules", "")),
            ],
            row_height=36,
            static=[
                "Service animals are exempt from pet fees and pet bans by law (ADA/FHA).",
            ],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=4, tab_name=TAB_PETS,
        title="§4 Pets",
        subtitle="Specific pet rules avoid surprise damage and disputes.",
        cards=cards, prev_tab=TAB_SMOKE, next_tab=TAB_DAMAGE,
    )


def build_damage_tab(wb, variant):
    wb.create_sheet(TAB_DAMAGE)
    s = SAMPLE["Damage"] if variant == "demo" else {}
    cards = [
        Card(
            header="Furniture + Amenities",
            rows=[
                ("Moving furniture:", s.get("furniture", "")),
                ("Pool / hot tub rules:", s.get("hot_tub", "")),
                ("Fireplace rules:", s.get("fireplace", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Supervision + Liability",
            rows=[
                ("Kids / minors policy:", s.get("kids", "")),
                ("Maximum occupancy (overnight):", s.get("occupancy", "")),
                ("Damage / claim policy:", s.get("damage", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=5, tab_name=TAB_DAMAGE,
        title="§5 Damage + Liability",
        subtitle="The rules that hold up in an AirCover or VRBO claim.",
        cards=cards, prev_tab=TAB_PETS, next_tab=TAB_CHECKOUT,
    )


def build_checkout_tab(wb, variant):
    wb.create_sheet(TAB_CHECKOUT)
    s = SAMPLE["Checkout"] if variant == "demo" else {}
    cards = [
        Card(
            header="Check-out Time",
            rows=[
                ("Check-out time:", s.get("time", "")),
                ("Late check-out fee:", s.get("late_fee", "")),
            ],
        ),
        Card(
            header="Penalties",
            rows=[
                ("Missing-item fee:", s.get("missing", "")),
                ("Smoking violation fine:", s.get("smoke_fine", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Keys + Deposit",
            rows=[
                ("Key / lock return:", s.get("keys", "")),
                ("Damage deposit policy:", s.get("deposit", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=6, tab_name=TAB_CHECKOUT,
        title="§6 Check-out + Penalties",
        subtitle="What happens if guests don't follow the rules.",
        cards=cards, prev_tab=TAB_DAMAGE, next_tab="",
    )


# --- Launch tab ---

def build_launch_tab(wb, variant):
    """Tab 7 — Launch. Readiness dashboard + BUILD RULES DOC button."""
    ws = wb.create_sheet("Launch")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Header (rows 1-6, navy)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    pseudo_button(ws, "A2", "C2", f"← BACK: {TAB_CHECKOUT}",
                  f"'{TAB_CHECKOUT}'!A5", variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]
    c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Your house rules are ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Build the print-ready doc, then paste into your listing."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # ZONE 2: Readiness dashboard (rows 8-14)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 15):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    ranges = [
        f"'{TAB_PROP}'!B8:B13",
        f"'{TAB_QUIET}'!B8:B13",
        f"'{TAB_SMOKE}'!B8:B12",
        f"'{TAB_PETS}'!B8:B12",
        f"'{TAB_DAMAGE}'!B8:B13",
        f"'{TAB_CHECKOUT}'!B8:B13",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)

    # Card 1: Completion
    ws.merge_cells("A9:D9")
    c = ws["A9"]
    c.value = "COMPLETION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]
    c.value = f'=TEXT(({counta_sum})/{TOTAL_INPUTS}, "0%")'
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A12:D12")
    c = ws["A12"]
    c.value = f"of {TOTAL_INPUTS} fields filled"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2: Red Flags — required fields empty
    required = [
        f"'{TAB_PROP}'!B8",       # property name
        f"'{TAB_PROP}'!B10",      # max guests
        f"'{TAB_QUIET}'!B8",      # quiet start
        f"'{TAB_QUIET}'!B9",      # quiet end
        f"'{TAB_QUIET}'!B10",     # parties policy
        f"'{TAB_SMOKE}'!B8",      # smoking inside
        f"'{TAB_PETS}'!B8",       # pets allowed
        f"'{TAB_CHECKOUT}'!B8",   # checkout time
    ]
    countblank_req = " + ".join(f'IF({r}="",1,0)' for r in required)

    ws.merge_cells("E9:H9")
    c = ws["E9"]
    c.value = "RED FLAGS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]
    c.value = f"={countblank_req}"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E12:H12")
    c = ws["E12"]
    c.value = "required fields empty"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3: Status
    ws.merge_cells("I9:L9")
    c = ws["I9"]
    c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    c.value = (f'=IF(({countblank_req})=0,"READY",'
               f'IF(({countblank_req})<=2,"MINOR","NEEDS WORK"))')
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I12:L12")
    c = ws["I12"]
    c.value = "0 = green · 1-2 = yellow · 3+ = red"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: BUILD RULES DOC button (rows 16-21)
    pseudo_button(ws, "A16", "L21",
                  "📄  BUILD RULES DOC →",
                  "'Rules Output'!A1",
                  variant="primary")
    for r in range(16, 22):
        ws.row_dimensions[r].height = 28

    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = ("Opens the Rules Output tab — select all, copy, paste into "
               "Airbnb / VRBO 'House rules' field. Or Ctrl+P for the printable PDF.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 18

    ws.print_area = "A1:L22"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# --- Rules Output tab (formula-driven) ---

def build_rules_output_tab(wb, variant):
    """Tab 8 — Rules Output. Formula-concatenated, print-ready rules.

    All cells are formulas pulling from the input tabs. The tab is laid out
    as a single-column document with section headers and bullet lines. Bullet
    lines collapse to empty when their source input is blank, so partially
    filled workbooks render a partial doc.
    """
    ws = wb.create_sheet("Rules Output")
    ws.sheet_properties.tabColor = COLOR_GOLD_SOFT
    set_col_widths(ws, [("A", 6), ("B", 90)])

    # Header band (gold-soft — this is a reward / output tab)
    gold_soft_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for r in range(1, 6):
        for c in range(1, 3):
            ws.cell(row=r, column=c).fill = gold_soft_fill

    # Row 2: back button (col A) + label (col B)
    pseudo_button(ws, "A2", "A2", "←", "'Launch'!A1", variant="secondary")
    c = ws["B2"]
    c.value = "RULES OUTPUT — copy, print, or paste into your listing"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A4:B4")
    c = ws["A4"]
    c.value = '=IF(\'§1 Property + Capacity\'!B8="","House Rules",\'§1 Property + Capacity\'!B8 & " — House Rules")'
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A5:B5")
    c = ws["A5"]
    c.value = ('=IF(\'§1 Property + Capacity\'!B9="","",'
               '"Address: " & \'§1 Property + Capacity\'!B9)')
    c.font = Font(name=FONT_HEAD, size=11, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 18

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    current_row = 7

    # Helper: render a section header (col A spacer, col B header)
    def section_hdr(row, text):
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = parchment
        cell = ws.cell(row=row, column=2)
        cell.value = text
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=COLOR_ACCENT))
        ws.row_dimensions[row].height = 24

    # Helper: render a bullet line (formula concatenation)
    def bullet(row, formula):
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = parchment
        bc = ws.cell(row=row, column=1)
        bc.value = ('=IF(' + _strip_eq(formula) + '="","","•")')
        bc.font = Font(name=FONT_BODY, size=11, color=COLOR_ACCENT)
        bc.alignment = Alignment(horizontal="center", vertical="top")
        tc = ws.cell(row=row, column=2)
        tc.value = formula
        tc.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        tc.alignment = Alignment(horizontal="left", vertical="top",
                                 wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 30

    # === SECTION: Property + Capacity ===
    section_hdr(current_row, "Property + Capacity")
    current_row += 1
    bullet(current_row,
           '=IF(\'§1 Property + Capacity\'!B10="","","Maximum overnight guests: " & \'§1 Property + Capacity\'!B10)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§1 Property + Capacity\'!B11="","","Bedrooms / bathrooms: " & \'§1 Property + Capacity\'!B11)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§1 Property + Capacity\'!B12="","","Sleeping arrangements: " & \'§1 Property + Capacity\'!B12)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§1 Property + Capacity\'!B13="","","Parking: " & \'§1 Property + Capacity\'!B13)')
    current_row += 2

    # === SECTION: Quiet Hours + Parties ===
    section_hdr(current_row, "Quiet Hours + Parties")
    current_row += 1
    bullet(current_row,
           '=IF(OR(\'§2 Quiet Hours + Parties\'!B8="",\'§2 Quiet Hours + Parties\'!B9=""),"",'
           '"Quiet hours: " & \'§2 Quiet Hours + Parties\'!B8 & " to " & \'§2 Quiet Hours + Parties\'!B9 & " — sound carries through walls.")')
    current_row += 1
    bullet(current_row,
           '=IF(\'§2 Quiet Hours + Parties\'!B10="","","Parties / events: " & \'§2 Quiet Hours + Parties\'!B10)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§2 Quiet Hours + Parties\'!B11="","","Day visitors / gatherings: " & \'§2 Quiet Hours + Parties\'!B11)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§2 Quiet Hours + Parties\'!B12="","","Guests of guests: " & \'§2 Quiet Hours + Parties\'!B12)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§2 Quiet Hours + Parties\'!B13="","","Outdoor music: " & \'§2 Quiet Hours + Parties\'!B13)')
    current_row += 2

    # === SECTION: Smoking + Substances ===
    section_hdr(current_row, "Smoking + Substances")
    current_row += 1
    bullet(current_row,
           '=IF(\'§3 Smoking + Substances\'!B8="","","Smoking — inside: " & \'§3 Smoking + Substances\'!B8)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§3 Smoking + Substances\'!B9="","","Smoking — outside: " & \'§3 Smoking + Substances\'!B9)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§3 Smoking + Substances\'!B10="","","Vaping: " & \'§3 Smoking + Substances\'!B10)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§3 Smoking + Substances\'!B11="","","Cannabis: " & \'§3 Smoking + Substances\'!B11)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§3 Smoking + Substances\'!B12="","","Alcohol: " & \'§3 Smoking + Substances\'!B12)')
    current_row += 2

    # === SECTION: Pets ===
    section_hdr(current_row, "Pets")
    current_row += 1
    bullet(current_row,
           '=IF(\'§4 Pets\'!B8="","","Pet policy: " & \'§4 Pets\'!B8)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§4 Pets\'!B9="","","Number allowed: " & \'§4 Pets\'!B9)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§4 Pets\'!B10="","","Weight limit: " & \'§4 Pets\'!B10)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§4 Pets\'!B11="","","Pet fee: " & \'§4 Pets\'!B11)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§4 Pets\'!B12="","","Ground rules: " & \'§4 Pets\'!B12)')
    current_row += 1
    # Static service-animal disclaimer (always shows)
    for c in range(1, 3):
        ws.cell(row=current_row, column=c).fill = parchment
    sc = ws.cell(row=current_row, column=2)
    sc.value = "Service animals are exempt from pet fees and pet bans by law (ADA/FHA)."
    sc.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    sc.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    # === SECTION: Damage + Liability ===
    section_hdr(current_row, "Damage + Liability")
    current_row += 1
    bullet(current_row,
           '=IF(\'§5 Damage + Liability\'!B8="","","Furniture: " & \'§5 Damage + Liability\'!B8)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§5 Damage + Liability\'!B9="","","Pool / hot tub: " & \'§5 Damage + Liability\'!B9)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§5 Damage + Liability\'!B10="","","Fireplace: " & \'§5 Damage + Liability\'!B10)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§5 Damage + Liability\'!B11="","","Children: " & \'§5 Damage + Liability\'!B11)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§5 Damage + Liability\'!B12="","","Maximum occupancy: " & \'§5 Damage + Liability\'!B12)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§5 Damage + Liability\'!B13="","","Damage policy: " & \'§5 Damage + Liability\'!B13)')
    current_row += 2

    # === SECTION: Check-out + Penalties ===
    section_hdr(current_row, "Check-out + Penalties")
    current_row += 1
    bullet(current_row,
           '=IF(\'§6 Check-out + Penalties\'!B8="","","Check-out time: " & \'§6 Check-out + Penalties\'!B8)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§6 Check-out + Penalties\'!B9="","","Late check-out fee: " & \'§6 Check-out + Penalties\'!B9)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§6 Check-out + Penalties\'!B10="","","Missing items: " & \'§6 Check-out + Penalties\'!B10)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§6 Check-out + Penalties\'!B11="","","Smoking violation: " & \'§6 Check-out + Penalties\'!B11)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§6 Check-out + Penalties\'!B12="","","Keys / lock: " & \'§6 Check-out + Penalties\'!B12)')
    current_row += 1
    bullet(current_row,
           '=IF(\'§6 Check-out + Penalties\'!B13="","","Damage deposit: " & \'§6 Check-out + Penalties\'!B13)')
    current_row += 2

    # Footer
    footer_row = current_row
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 3):
        ws.cell(row=footer_row, column=c).border = Border(top=gold_side)
        ws.cell(row=footer_row, column=c).fill = parchment
    ws.merge_cells(f"A{footer_row + 1}:B{footer_row + 1}")
    c = ws[f"A{footer_row + 1}"]
    c.value = (f"Generated by {BRAND_NAME} House Rules Builder · "
               f"{BRAND_DOMAIN}")
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = parchment
    ws.row_dimensions[footer_row + 1].height = 18

    # Print setup — fit to one page width, allow paging vertically
    ws.print_area = f"A1:B{footer_row + 1}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # let it page naturally
    ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6)


def _strip_eq(formula):
    """Strip leading '=' from a formula string for nesting inside IF()."""
    return formula[1:] if formula.startswith("=") else formula


# --- Build orchestrator ---

def build_workbook(out_path, variant):
    """Build a workbook with all 9 tabs."""
    assert variant in ("demo", "blank"), f"Unknown variant: {variant}"

    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_tab(wb, variant)
    build_quiet_tab(wb, variant)
    build_smoke_tab(wb, variant)
    build_pets_tab(wb, variant)
    build_damage_tab(wb, variant)
    build_checkout_tab(wb, variant)
    build_launch_tab(wb, variant)
    build_rules_output_tab(wb, variant)

    suffix = " (DEMO)" if variant == "demo" else ""
    wb.properties.title = f"House Rules Builder{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Wizard tool for STR hosts — produces a print-ready, listing-ready "
        "house-rules document from 6 sections of structured inputs."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, variant="demo")
    build_workbook(BLANK_OUT, variant="blank")


if __name__ == "__main__":
    main()
