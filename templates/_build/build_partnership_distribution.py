"""Build FIN-007 Partnership Distribution Tracker Excel file (v2.3 standard).

Operational-mode register (forked from build_1099_nec_tracker.py) with a
matrix sub-pattern for the per-partner P&L allocation. For STR partnerships
(LLC taxed as partnership). Tracks capital contributions, profit/loss
allocation by ownership %, distributions paid, and ending capital account
per partner — the K-1 Item L numbers a managing partner hands the tax
preparer in February.

Generates:
  templates/_masters/FIN-007-partnership-distribution-tracker-DEMO.xlsx
  templates/_masters/FIN-007-partnership-distribution-tracker-BLANK.xlsx
"""
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "FIN-007"
NAME = "partnership-distribution-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026
PARTNERS_CAP = 8         # roster capacity
CONTRIB_CAP = 100        # contributions log capacity
DISTRIB_CAP = 200        # distributions log capacity

# --- Sample partners (DEMO) ---
# 2-partner LLC: Daniel 60% / Partner 40%. Beginning capital $80K / $40K.
PARTNERS = [
    (1, "Daniel Harrison",  "1234", 0.60, 80000, "Managing partner"),
    (2, "Jordan Reeves",    "5678", 0.40, 40000, "Capital partner"),
]

# --- Sample contributions (DEMO) ---
CONTRIBUTIONS = [
    ("2026-01-15", "Daniel Harrison",  5000, "Cash",                  "Working capital top-up"),
    ("2026-03-22", "Jordan Reeves",    2500, "Cash",                  "Furnishing budget"),
    ("2026-06-10", "Daniel Harrison",  3000, "Property contribution", "Tools + outdoor furniture"),
]

# --- Sample distributions (DEMO) ---
# Daniel total $18K, Jordan total $12K — quarterly cash distributions.
DISTRIBUTIONS = [
    ("2026-03-31", "Daniel Harrison", 4500, "Cash",             "Smokies Ridge LLC", "Q1 distribution"),
    ("2026-03-31", "Jordan Reeves",   3000, "Cash",             "Smokies Ridge LLC", "Q1 distribution"),
    ("2026-06-30", "Daniel Harrison", 4500, "Cash",             "Smokies Ridge LLC", "Q2 distribution"),
    ("2026-06-30", "Jordan Reeves",   3000, "Cash",             "Smokies Ridge LLC", "Q2 distribution"),
    ("2026-09-30", "Daniel Harrison", 4500, "Cash",             "Smokies Ridge LLC", "Q3 distribution"),
    ("2026-09-30", "Jordan Reeves",   3000, "Cash",             "Smokies Ridge LLC", "Q3 distribution"),
    ("2026-12-15", "Daniel Harrison", 4500, "Tax distribution", "Smokies Ridge LLC", "Q4 + tax distribution"),
    ("2026-12-15", "Jordan Reeves",   3000, "Tax distribution", "Smokies Ridge LLC", "Q4 + tax distribution"),
]

# 2026 partnership net income $36K → $21.6K / $14.4K allocated
DEMO_NET_INCOME = 36000
DEMO_GUARANTEED_PAYMENTS = 0  # special allocation row

# --- Settings demo values ---
DEMO_PARTNERSHIP_NAME = "Smokies Ridge Holdings LLC"
DEMO_EIN = "XX-XXXXXXX"
DEMO_ALLOCATION_METHOD = "Pro-rata by ownership"


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational-mode hero + KPI cards + nav)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Partnership Distribution Tracker"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "K-1 Item L numbers, ready for your tax preparer in February."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "For STR partnerships (LLC taxed as a partnership, not single-member). "
        "Tracks capital contributions, profit/loss allocation by ownership %, "
        "distributions paid, and ending capital account per partner. The five "
        "K-1 Item L numbers — beginning capital, contributions, profit allocated, "
        "distributions, ending capital — drop straight onto Schedule K-1 for each "
        "partner. Ownership % checksum flags the moment the roster drifts off 100%."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Settings: set the active tax year, partnership name, EIN, and allocation method.",
        "② Partners: enter every partner — name, ownership %, beginning capital. % MUST sum to 100.",
        "③ Capital Contributions: log each new contribution as it happens (date, partner, amount, type).",
        "④ Distributions: log every distribution paid (date, partner, amount, type, source entity).",
        "⑤ P&L Allocation: enter annual partnership net income — per-partner allocation auto-computes.",
        "⑥ February: hand the Partners tab + P&L Allocation to your tax preparer for K-1s.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "ADD A DISTRIBUTION" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  ADD A DISTRIBUTION  (OPEN DISTRIBUTIONS)",
                   "'Distributions'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: KPI cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Partners count
    ws.merge_cells("A32:D32")
    c = ws["A32"]
    c.value = "PARTNERS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]
    c.value = f"=COUNTA(Partners!B6:B{5 + PARTNERS_CAP})"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]
    c.value = "rows on the Partners tab"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Total distributions YTD
    ws.merge_cells("E32:H32")
    c = ws["E32"]
    c.value = "DISTRIBUTIONS YTD"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = f"=SUM(Distributions!C6:C{5 + DISTRIB_CAP})"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]
    c.value = "all partners, current year"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Total ending capital
    ws.merge_cells("I32:L32")
    c = ws["I32"]
    c.value = "TOTAL ENDING CAPITAL"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = f"=SUM(Partners!I6:I{5 + PARTNERS_CAP})"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("I35:L35")
    c = ws["I35"]
    c.value = "sum of partner capital accounts"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Partners",
                   "'Partners'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Capital Contributions",
                   "'Capital Contributions'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "P&L Allocation",
                   "'P&L Allocation'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: K-1 callout ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "📅 K-1 PREP: in February, the Partners tab shows the five Item L numbers "
        "for each partner — beginning capital, contributions, profit allocated, "
        "distributions, ending capital. Hand it to your tax preparer."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Ownership % warning ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "⚠ Ownership % across all partners must equal 100%. The Partners tab "
        "shows a checksum — fix it before relying on the P&L allocation."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    add_upgrade_banner(ws, 46)

    brand_footer(ws, 48,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_partners_tab(wb, variant):
    """Sheet 1 — Partners (one row per partner; ownership % checksum)."""
    ws = wb.create_sheet("Partners")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Partners",
                         prev_tab="Start",
                         next_tab="Capital Contributions")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Partner roster (capacity 8) — % must sum to 100. Five K-1 Item L numbers per row."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 6),  ("B", 26), ("C", 14), ("D", 12),
        ("E", 18), ("F", 18), ("G", 18), ("H", 18), ("I", 18), ("J", 28),
    ])

    headers = [
        "Partner #", "Partner name", "Tax ID last4", "Ownership %",
        "Beginning capital (Jan 1)",
        "Contributions YTD",
        "Profit allocated YTD",
        "Distributions YTD",
        "Ending capital",
        "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    # Row 6 .. 5+PARTNERS_CAP — partner data rows
    partner_rows = PARTNERS if variant == "demo" else []
    for offset in range(PARTNERS_CAP):
        r = 6 + offset
        if offset < len(partner_rows):
            num, pname, last4, own_pct, beg_cap, notes = partner_rows[offset]
        else:
            num = offset + 1
            pname = None
            last4 = None
            own_pct = None
            beg_cap = None
            notes = None

        # A: Partner #
        a = ws.cell(row=r, column=1, value=num)
        apply_style(a, input_cell_style())
        a.alignment = Alignment(horizontal="center", vertical="center")

        # B: Partner name
        b = ws.cell(row=r, column=2, value=pname)
        apply_style(b, input_cell_style())

        # C: Tax ID last4
        c = ws.cell(row=r, column=3, value=last4)
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        # D: Ownership % (input as decimal — formatted as %)
        d = ws.cell(row=r, column=4, value=own_pct)
        apply_style(d, input_cell_style())
        d.number_format = "0.00%"
        d.alignment = Alignment(horizontal="right", vertical="center")

        # E: Beginning capital
        e = ws.cell(row=r, column=5, value=beg_cap)
        apply_style(e, input_cell_style())
        e.number_format = '"$"#,##0.00'

        # F: Contributions YTD = SUMIFS(Capital Contributions!C, B, partner-name)
        f = ws.cell(
            row=r, column=6,
            value=(
                f'=IF(B{r}="","",'
                f"SUMIFS('Capital Contributions'!$C$6:$C${5 + CONTRIB_CAP},"
                f"'Capital Contributions'!$B$6:$B${5 + CONTRIB_CAP},B{r}))"
            ),
        )
        apply_style(f, formula_cell_style())
        f.number_format = '"$"#,##0.00'

        # G: Profit allocated YTD — pulled from P&L Allocation matrix col D
        # P&L Allocation rows mirror Partners rows (row r is partner offset).
        g = ws.cell(
            row=r, column=7,
            value=(
                f'=IF(B{r}="","",'
                f"IFERROR(VLOOKUP(B{r},'P&L Allocation'!$A$8:$D${7 + PARTNERS_CAP},4,FALSE),0))"
            ),
        )
        apply_style(g, formula_cell_style())
        g.number_format = '"$"#,##0.00'

        # H: Distributions YTD
        h = ws.cell(
            row=r, column=8,
            value=(
                f'=IF(B{r}="","",'
                f"SUMIFS(Distributions!$C$6:$C${5 + DISTRIB_CAP},"
                f"Distributions!$B$6:$B${5 + DISTRIB_CAP},B{r}))"
            ),
        )
        apply_style(h, formula_cell_style())
        h.number_format = '"$"#,##0.00'

        # I: Ending capital = E + F + G - H
        i_cell = ws.cell(
            row=r, column=9,
            value=f'=IF(B{r}="","",E{r}+F{r}+G{r}-H{r})',
        )
        apply_style(i_cell, formula_cell_style())
        i_cell.number_format = '"$"#,##0.00'
        i_cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # J: Notes
        j = ws.cell(row=r, column=10, value=notes)
        apply_style(j, input_cell_style())

        ws.row_dimensions[r].height = 18

    # --- Checksum row ---
    checksum_row = 6 + PARTNERS_CAP + 1   # spacer at 6+CAP, checksum at 6+CAP+1
    ws.row_dimensions[checksum_row - 1].height = 8

    label_cell = ws.cell(row=checksum_row, column=2,
                          value="Ownership % checksum (must equal 100%)")
    label_cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=checksum_row, start_column=1,
                   end_row=checksum_row, end_column=3)

    checksum_cell = ws.cell(
        row=checksum_row, column=4,
        value=f"=SUM(D6:D{5 + PARTNERS_CAP})",
    )
    apply_style(checksum_cell, formula_cell_style())
    checksum_cell.number_format = "0.00%"
    checksum_cell.alignment = Alignment(horizontal="right", vertical="center")
    checksum_cell.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_PRIMARY)

    flag_cell = ws.cell(
        row=checksum_row, column=5,
        value=f'=IF(ROUND(D{checksum_row},4)=1,"✓ OK","⚠ FIX — must equal 100%")',
    )
    apply_style(flag_cell, formula_cell_style())
    flag_cell.alignment = Alignment(horizontal="left", vertical="center")
    flag_cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR)
    ws.merge_cells(start_row=checksum_row, start_column=5,
                   end_row=checksum_row, end_column=10)
    ws.row_dimensions[checksum_row].height = 22

    # Conditional formatting on the checksum cell — red flag if != 100%
    ws.conditional_formatting.add(
        f"D{checksum_row}",
        FormulaRule(
            formula=[f"ROUND(D{checksum_row},4)<>1"],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        f"D{checksum_row}",
        FormulaRule(
            formula=[f"ROUND(D{checksum_row},4)=1"],
            fill=PatternFill("solid", fgColor="C7EFCF"),
            font=Font(bold=True, color=COLOR_PRIMARY),
        ),
    )
    # Flag column: same logic
    ws.conditional_formatting.add(
        f"E{checksum_row}",
        FormulaRule(
            formula=[f"ROUND(D{checksum_row},4)<>1"],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    # --- K-1 Item L summary (compact restatement under checksum) ---
    k1_header_row = checksum_row + 2
    ws.row_dimensions[k1_header_row - 1].height = 8

    hdr = ws.cell(row=k1_header_row, column=1,
                   value="K-1 Item L summary — these five numbers per partner")
    hdr.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.merge_cells(start_row=k1_header_row, start_column=1,
                   end_row=k1_header_row, end_column=10)
    ws.row_dimensions[k1_header_row].height = 22

    k1_help = ws.cell(
        row=k1_header_row + 1, column=1,
        value=("Beginning capital · Contributions · Profit allocated · "
               "Distributions · Ending capital → drop straight onto Schedule K-1."),
    )
    k1_help.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    ws.merge_cells(start_row=k1_header_row + 1, start_column=1,
                   end_row=k1_header_row + 1, end_column=10)
    ws.row_dimensions[k1_header_row + 1].height = 16

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_contributions_tab(wb, variant):
    """Sheet 2 — Capital Contributions log (capacity 100)."""
    ws = wb.create_sheet("Capital Contributions")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Capital Contributions",
                         prev_tab="Partners",
                         next_tab="Distributions")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Log every contribution AS IT HAPPENS — date, partner, amount, type, notes."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 26), ("C", 14), ("D", 22), ("E", 38),
    ])

    headers = ["Date", "Partner", "Amount", "Type", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    contrib_rows = CONTRIBUTIONS if variant == "demo" else []
    for i, row in enumerate(contrib_rows, start=6):
        date_val, partner, amount, ctype, notes = row

        a = ws.cell(row=i, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=partner)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=amount)
        apply_style(c, input_cell_style())
        c.number_format = '"$"#,##0.00'

        d = ws.cell(row=i, column=4, value=ctype)
        apply_style(d, input_cell_style())

        e = ws.cell(row=i, column=5, value=notes)
        apply_style(e, input_cell_style())

        ws.row_dimensions[i].height = 16

    last_data_row = len(contrib_rows) + 5
    for row_idx in range(last_data_row + 1, 6 + CONTRIB_CAP):
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 3:
                cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[row_idx].height = 16

    add_dropdown(ws, f"B6:B{5 + CONTRIB_CAP}",
                  f"=Partners!$B$6:$B${5 + PARTNERS_CAP}")
    add_dropdown(ws, f"D6:D{5 + CONTRIB_CAP}",
                  '"Cash,Property contribution,Services"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_distributions_tab(wb, variant):
    """Sheet 3 — Distributions log (capacity 200)."""
    ws = wb.create_sheet("Distributions")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Distributions",
                         prev_tab="Capital Contributions",
                         next_tab="P&L Allocation")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Log every distribution paid — date, partner, amount, type, source entity."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 12), ("B", 26), ("C", 14), ("D", 20),
        ("E", 26), ("F", 32),
    ])

    headers = ["Date", "Partner", "Amount", "Type", "Source entity", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    distrib_rows = DISTRIBUTIONS if variant == "demo" else []
    for i, row in enumerate(distrib_rows, start=6):
        date_val, partner, amount, dtype, src, notes = row

        a = ws.cell(row=i, column=1, value=_parse_date(date_val))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        b = ws.cell(row=i, column=2, value=partner)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=amount)
        apply_style(c, input_cell_style())
        c.number_format = '"$"#,##0.00'

        d = ws.cell(row=i, column=4, value=dtype)
        apply_style(d, input_cell_style())

        e = ws.cell(row=i, column=5, value=src)
        apply_style(e, input_cell_style())

        f = ws.cell(row=i, column=6, value=notes)
        apply_style(f, input_cell_style())

        ws.row_dimensions[i].height = 16

    last_data_row = len(distrib_rows) + 5
    for row_idx in range(last_data_row + 1, 6 + DISTRIB_CAP):
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx == 1:
                cell.number_format = "yyyy-mm-dd"
            if col_idx == 3:
                cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[row_idx].height = 16

    add_dropdown(ws, f"B6:B{5 + DISTRIB_CAP}",
                  f"=Partners!$B$6:$B${5 + PARTNERS_CAP}")
    add_dropdown(ws, f"D6:D{5 + DISTRIB_CAP}",
                  '"Cash,Property,In-kind,Tax distribution,Liquidating"')

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_pl_allocation_tab(wb, variant):
    """Sheet 4 — P&L Allocation (matrix sub-pattern: net income × ownership %)."""
    ws = wb.create_sheet("P&L Allocation")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = CPA hand-off

    compact_header_band(ws, "P&L Allocation · For Your CPA",
                         prev_tab="Distributions",
                         next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 FOR YOUR CPA — annual partnership net income allocated by ownership %."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 26), ("B", 14), ("C", 18), ("D", 18), ("E", 22),
    ])

    # --- Inputs band (rows 5-6) ---
    in1_label = ws.cell(row=5, column=1, value="Annual partnership net income:")
    in1_label.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    in1_label.alignment = Alignment(horizontal="right", vertical="center")
    in1 = ws.cell(row=5, column=2,
                   value=_val(variant, DEMO_NET_INCOME))
    apply_style(in1, input_cell_style())
    in1.number_format = '"$"#,##0.00'
    ws.row_dimensions[5].height = 20

    in2_label = ws.cell(row=6, column=1, value="Special allocations (e.g., guaranteed payments):")
    in2_label.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    in2_label.alignment = Alignment(horizontal="right", vertical="center")
    in2 = ws.cell(row=6, column=2,
                   value=_val(variant, DEMO_GUARANTEED_PAYMENTS))
    apply_style(in2, input_cell_style())
    in2.number_format = '"$"#,##0.00'
    ws.row_dimensions[6].height = 20

    # --- Header row 7 ---
    headers = [
        "Partner",
        "Ownership %",
        "Pro-rata share",
        "Final allocated income",
        "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[7].height = 22

    # --- Matrix rows 8 .. 7+PARTNERS_CAP ---
    for offset in range(PARTNERS_CAP):
        r = 8 + offset
        partners_row = 6 + offset

        # A: Partner name (mirror Partners!B)
        a = ws.cell(
            row=r, column=1,
            value=f'=IF(Partners!B{partners_row}="","",Partners!B{partners_row})',
        )
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center")

        # B: Ownership %
        b = ws.cell(
            row=r, column=2,
            value=f'=IF(Partners!B{partners_row}="","",Partners!D{partners_row})',
        )
        apply_style(b, formula_cell_style())
        b.number_format = "0.00%"

        # C: Pro-rata share = ownership % × (net income - special)
        c = ws.cell(
            row=r, column=3,
            value=(
                f'=IF(A{r}="","",'
                f'($B$5-$B$6)*B{r})'
            ),
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0.00'

        # D: Final allocated income — mirror C unless special allocation logic
        # (kept simple: pro-rata share IS the K-1 number for line 1).
        d = ws.cell(
            row=r, column=4,
            value=f'=IF(A{r}="","",C{r})',
        )
        apply_style(d, formula_cell_style())
        d.number_format = '"$"#,##0.00'
        d.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # E: Notes (input)
        e = ws.cell(row=r, column=5)
        apply_style(e, input_cell_style())

        ws.row_dimensions[r].height = 18

    # --- Totals row ---
    totals_row = 8 + PARTNERS_CAP
    label = ws.cell(row=totals_row, column=1, value="TOTAL")
    label.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    label.alignment = Alignment(horizontal="right", vertical="center")

    pct_total = ws.cell(
        row=totals_row, column=2,
        value=f"=SUM(B8:B{7 + PARTNERS_CAP})",
    )
    apply_style(pct_total, formula_cell_style())
    pct_total.number_format = "0.00%"
    pct_total.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    share_total = ws.cell(
        row=totals_row, column=3,
        value=f"=SUM(C8:C{7 + PARTNERS_CAP})",
    )
    apply_style(share_total, formula_cell_style())
    share_total.number_format = '"$"#,##0.00'
    share_total.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    final_total = ws.cell(
        row=totals_row, column=4,
        value=f"=SUM(D8:D{7 + PARTNERS_CAP})",
    )
    apply_style(final_total, formula_cell_style())
    final_total.number_format = '"$"#,##0.00'
    final_total.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

    ws.row_dimensions[totals_row].height = 20

    # Conditional flag: total ownership != 100% → red
    ws.conditional_formatting.add(
        f"B{totals_row}",
        FormulaRule(
            formula=[f"ROUND(B{totals_row},4)<>1"],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    # --- Donut chart: allocation share by partner ---
    chart_anchor_row = totals_row + 3
    donut = DoughnutChart()
    donut.title = "Profit allocation share"
    donut.height = 9
    donut.width = 11
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=4, min_row=7,
                            max_row=7 + PARTNERS_CAP, max_col=4)
    donut_cats = Reference(ws, min_col=1, min_row=8,
                            max_row=7 + PARTNERS_CAP)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=False, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, f"G7")

    ws.freeze_panes = "A8"

    ws.print_area = f"A1:E{totals_row + 1}"
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 5 — Settings (active year, partnership info, allocation method)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="P&L Allocation", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active tax year · partnership info · allocation method"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 36), ("B", 32), ("C", 14), ("D", 16)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # B5 — Active tax year (per project rule)
    a5 = ws.cell(row=5, column=1, value="Active tax year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=_val(variant, ACTIVE_YEAR))
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    b5.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 20

    # Row 6: explainer
    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = (
        "Active year drives every YTD calculation and the K-1 Item L summary. "
        "Update each January when starting a new tax year."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 28

    # B7 — Partnership name
    a7 = ws.cell(row=7, column=1, value="Partnership name:")
    a7.font = bold_right
    a7.alignment = Alignment(horizontal="right", vertical="center")
    b7 = ws.cell(row=7, column=2, value=_val(variant, DEMO_PARTNERSHIP_NAME))
    apply_style(b7, input_cell_style())
    ws.row_dimensions[7].height = 20

    # B9 — EIN
    a9 = ws.cell(row=9, column=1, value="EIN:")
    a9.font = bold_right
    a9.alignment = Alignment(horizontal="right", vertical="center")
    b9 = ws.cell(row=9, column=2, value=_val(variant, DEMO_EIN))
    apply_style(b9, input_cell_style())
    ws.row_dimensions[9].height = 20

    # B11 — Allocation method
    a11 = ws.cell(row=11, column=1, value="Allocation method:")
    a11.font = bold_right
    a11.alignment = Alignment(horizontal="right", vertical="center")
    b11 = ws.cell(row=11, column=2,
                   value=_val(variant, DEMO_ALLOCATION_METHOD))
    apply_style(b11, input_cell_style())
    add_dropdown(ws, "B11",
                  '"Pro-rata by ownership,Special allocation per agreement"')
    ws.row_dimensions[11].height = 20

    # B13 — Partner count (formula)
    a13 = ws.cell(row=13, column=1, value="Partner count (auto):")
    a13.font = bold_right
    a13.alignment = Alignment(horizontal="right", vertical="center")
    b13 = ws.cell(row=13, column=2,
                   value=f"=COUNTA(Partners!B6:B{5 + PARTNERS_CAP})")
    apply_style(b13, formula_cell_style())
    b13.number_format = "0"
    b13.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[13].height = 20

    ws.row_dimensions[14].height = 10

    # --- K-1 Prep summary section (rows 15+) ---
    sect15 = ws.cell(row=15, column=1, value="K-1 Item L summary (per partner)")
    sect15.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[15].height = 22

    ws.merge_cells("A16:D16")
    c16 = ws["A16"]
    c16.value = (
        "These five numbers per partner go directly to K-1 Item L. "
        "See the Partners tab for the live values."
    )
    c16.font = italic_muted
    c16.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[16].height = 28

    summary_labels = [
        (17, "Beginning capital", "Partners!E"),
        (18, "Contributions", "Partners!F"),
        (19, "Profit allocated", "Partners!G"),
        (20, "Distributions", "Partners!H"),
        (21, "Ending capital", "Partners!I"),
    ]
    for row_num, label, col_ref in summary_labels:
        a = ws.cell(row=row_num, column=1, value=label + " (total):")
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="center")
        b = ws.cell(
            row=row_num, column=2,
            value=f"=SUM({col_ref}6:{col_ref}{5 + PARTNERS_CAP})",
        )
        b.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        b.number_format = '"$"#,##0.00'
        b.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 18


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_partners_tab(wb, variant)
    build_contributions_tab(wb, variant)
    build_distributions_tab(wb, variant)
    build_pl_allocation_tab(wb, variant)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Partnership Distribution Tracker{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Partnership distribution tracker: capital contributions, "
        "P&L allocation by ownership %, distributions, and K-1 Item L "
        "ending-capital figures for STR LLCs taxed as partnerships (v2.3)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
