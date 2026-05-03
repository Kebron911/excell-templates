"""Build MKT-003 Referral Source + Repeat Guest CRM Excel file (v2.3 standard).

Operational-mode register: host logs each guest with channel-of-origin and
referral lineage so the year-end "thanks + come back" outreach can target
the highest-LTV, repeat-friendly guests — driving direct bookings and
eliminating channel fees.

Forks the build_1099_nec_tracker.py register/CRM pattern.

Generates:
  templates/_masters/MKT-003-referral-source-repeat-guest-crm-DEMO.xlsx
  templates/_masters/MKT-003-referral-source-repeat-guest-crm-BLANK.xlsx
"""
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "MKT-003"
NAME = "referral-source-repeat-guest-crm"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Capacity for guest CRM register
GUEST_CAPACITY = 500
LAST_GUEST_ROW = 5 + GUEST_CAPACITY  # header row 5, data rows 6..505


# --- Sample data (DEMO variant) ---
# 35 guests over 2 years across 3 properties, 7 repeats, 2 referral chains,
# top LTV $4,820, channel mix 64% Airbnb / 18% direct / 12% VRBO / 6% referral.
PROPERTIES = ["Smokies Ridge", "Creek Side", "Lakehouse A"]
CHANNELS = ["Airbnb", "VRBO", "Booking", "Direct", "Referral"]


def _d(y, m, day):
    return date(y, m, day)


# (first_booking_date, first, last, email, phone, property, channel, referrer,
#  stays, total_nights, total_paid, avg_rating, last_stay_date, notes)
GUESTS = [
    # Top LTV — 4 stays = $4,820
    (_d(2024, 6, 12), "Marcus", "Reynolds", "marcus.r@example.com", "(404) 555-0182", "Smokies Ridge", "Airbnb", "", 4, 17, 4820, 5.0, _d(2026, 3, 8), "Anniversary repeat — wife loves the firepit"),
    (_d(2024, 7, 4), "Jennifer", "Chen", "jchen@example.com", "(312) 555-0244", "Lakehouse A", "Airbnb", "", 3, 14, 3960, 5.0, _d(2026, 1, 22), "Annual family reunion booker"),
    (_d(2024, 8, 19), "David", "Patel", "dpatel@example.com", "(617) 555-0311", "Creek Side", "VRBO", "", 3, 12, 3210, 4.7, _d(2025, 11, 14), "Likes the hot tub access"),
    (_d(2024, 9, 2), "Sarah", "Kim", "sarah.kim@example.com", "(213) 555-0455", "Smokies Ridge", "Direct", "", 3, 10, 2870, 5.0, _d(2026, 2, 18), "Found us via blog — bookmarked direct site"),
    (_d(2024, 10, 11), "Michael", "Brennan", "mbrennan@example.com", "(303) 555-0177", "Smokies Ridge", "Airbnb", "", 2, 8, 2240, 4.5, _d(2025, 10, 5), "Brought brother's family second time"),
    # Referral chain 1: Marcus → Lopez → Bauer (3-deep)
    (_d(2025, 1, 14), "Elena", "Lopez", "elena.l@example.com", "(602) 555-0289", "Lakehouse A", "Referral", "Marcus Reynolds", 2, 7, 1980, 5.0, _d(2026, 4, 12), "Marcus's college friend"),
    (_d(2025, 5, 22), "Tomas", "Bauer", "tbauer@example.com", "(305) 555-0366", "Lakehouse A", "Referral", "Elena Lopez", 1, 4, 1240, 4.8, _d(2025, 5, 26), "Booked after Elena's recommendation"),
    # Referral chain 2: Jennifer → Wong
    (_d(2025, 8, 7), "Rebecca", "Wong", "rwong@example.com", "(415) 555-0422", "Lakehouse A", "Referral", "Jennifer Chen", 1, 5, 1480, 5.0, _d(2025, 8, 12), "Jennifer's coworker"),
    # Repeats
    (_d(2025, 2, 18), "James", "OConnor", "joconnor@example.com", "(617) 555-0533", "Creek Side", "Airbnb", "", 2, 9, 2560, 4.9, _d(2026, 2, 14), "Valentine's weekend repeater"),
    (_d(2025, 3, 6), "Priya", "Sharma", "psharma@example.com", "(469) 555-0177", "Smokies Ridge", "Direct", "", 2, 6, 1740, 5.0, _d(2026, 1, 9), "Booked direct after first Airbnb stay"),
    # One-time stays — Airbnb majority
    (_d(2025, 1, 5), "Robert", "Anderson", "", "(206) 555-0211", "Smokies Ridge", "Airbnb", "", 1, 3, 880, 5.0, _d(2025, 1, 8), ""),
    (_d(2025, 1, 18), "Linda", "Garcia", "lgarcia@example.com", "", "Creek Side", "Airbnb", "", 1, 2, 620, 4.5, _d(2025, 1, 20), ""),
    (_d(2025, 2, 3), "Thomas", "Wright", "twright@example.com", "(704) 555-0444", "Lakehouse A", "Airbnb", "", 1, 4, 1180, 4.8, _d(2025, 2, 7), ""),
    (_d(2025, 2, 21), "Amanda", "Davis", "", "(919) 555-0588", "Smokies Ridge", "Airbnb", "", 1, 3, 920, 5.0, _d(2025, 2, 24), ""),
    (_d(2025, 3, 14), "Christopher", "Lee", "clee@example.com", "(773) 555-0166", "Creek Side", "Airbnb", "", 1, 2, 580, 4.0, _d(2025, 3, 16), "Complained about Wi-Fi"),
    (_d(2025, 4, 2), "Nicole", "Thompson", "nthompson@example.com", "(503) 555-0277", "Smokies Ridge", "Airbnb", "", 1, 3, 940, 5.0, _d(2025, 4, 5), ""),
    (_d(2025, 4, 19), "Brian", "Walker", "bwalker@example.com", "(801) 555-0322", "Lakehouse A", "Airbnb", "", 1, 5, 1420, 4.7, _d(2025, 4, 24), ""),
    (_d(2025, 5, 8), "Megan", "Hall", "", "(615) 555-0411", "Creek Side", "Airbnb", "", 1, 2, 600, 5.0, _d(2025, 5, 10), ""),
    (_d(2025, 6, 1), "Kevin", "Young", "kyoung@example.com", "(404) 555-0455", "Smokies Ridge", "Airbnb", "", 1, 4, 1240, 4.8, _d(2025, 6, 5), ""),
    (_d(2025, 6, 18), "Rachel", "King", "rking@example.com", "", "Lakehouse A", "Airbnb", "", 1, 3, 980, 5.0, _d(2025, 6, 21), ""),
    (_d(2025, 7, 9), "Steven", "Wright", "swright@example.com", "(813) 555-0577", "Creek Side", "Airbnb", "", 1, 2, 660, 4.5, _d(2025, 7, 11), ""),
    (_d(2025, 7, 25), "Jessica", "Lopez", "jlopez@example.com", "(956) 555-0388", "Smokies Ridge", "VRBO", "", 1, 3, 920, 5.0, _d(2025, 7, 28), ""),
    (_d(2025, 8, 14), "Daniel", "Hill", "dhill@example.com", "(704) 555-0644", "Lakehouse A", "VRBO", "", 1, 4, 1280, 4.7, _d(2025, 8, 18), ""),
    (_d(2025, 9, 6), "Ashley", "Scott", "ascott@example.com", "(919) 555-0712", "Creek Side", "Airbnb", "", 1, 2, 640, 5.0, _d(2025, 9, 8), ""),
    (_d(2025, 9, 22), "Matthew", "Green", "mgreen@example.com", "(404) 555-0822", "Smokies Ridge", "Direct", "", 1, 3, 880, 5.0, _d(2025, 9, 25), "Found via Pinterest"),
    (_d(2025, 10, 11), "Olivia", "Adams", "oadams@example.com", "", "Lakehouse A", "Airbnb", "", 1, 4, 1280, 4.8, _d(2025, 10, 15), ""),
    (_d(2025, 11, 2), "Andrew", "Baker", "abaker@example.com", "(303) 555-0911", "Smokies Ridge", "Airbnb", "", 1, 3, 940, 5.0, _d(2025, 11, 5), ""),
    (_d(2025, 12, 12), "Samantha", "Carter", "scarter@example.com", "(212) 555-0188", "Creek Side", "VRBO", "", 1, 2, 720, 4.5, _d(2025, 12, 14), ""),
    (_d(2025, 12, 27), "Joshua", "Mitchell", "jmitchell@example.com", "(617) 555-0233", "Smokies Ridge", "Airbnb", "", 1, 4, 1320, 5.0, _d(2025, 12, 31), "NYE booking"),
    (_d(2026, 1, 8), "Lauren", "Perez", "lperez@example.com", "(305) 555-0444", "Lakehouse A", "Airbnb", "", 1, 3, 980, 4.7, _d(2026, 1, 11), ""),
    (_d(2026, 1, 24), "Ryan", "Roberts", "rroberts@example.com", "(206) 555-0566", "Creek Side", "Airbnb", "", 1, 2, 660, 5.0, _d(2026, 1, 26), ""),
    (_d(2026, 2, 9), "Hannah", "Turner", "hturner@example.com", "", "Smokies Ridge", "Direct", "", 1, 3, 920, 5.0, _d(2026, 2, 12), "Direct from blog"),
    (_d(2026, 2, 28), "Tyler", "Phillips", "tphillips@example.com", "(469) 555-0699", "Lakehouse A", "Airbnb", "", 1, 4, 1280, 4.8, _d(2026, 3, 4), ""),
    (_d(2026, 3, 17), "Brittany", "Campbell", "bcampbell@example.com", "(615) 555-0744", "Creek Side", "Airbnb", "", 1, 2, 640, 5.0, _d(2026, 3, 19), ""),
    (_d(2026, 4, 4), "Jordan", "Parker", "jparker@example.com", "(404) 555-0833", "Smokies Ridge", "Airbnb", "", 1, 3, 960, 5.0, _d(2026, 4, 7), ""),
]


REFERRAL_SOURCES = [
    "Past guest — Marcus Reynolds",
    "Past guest — Jennifer Chen",
    "Past guest — Sarah Kim",
    "Pinterest",
    "Instagram",
    "TikTok",
    "Blog post",
    "Word of mouth",
    "Local concierge",
    "Other",
]


# --- Outreach script copy (4 templates) ---
OUTREACH_SCRIPTS = [
    {
        "title": "1. Thanks for staying — post-checkout day 2-3",
        "subject": "Thank you, {first_name} — and a small ask",
        "body": (
            "Hi {first_name},\n\n"
            "It was a pleasure hosting you at {property} this past weekend. "
            "We hope the place gave you the rest you came for — and that the "
            "view from the deck delivered.\n\n"
            "If you have a minute, a 5-star review on {channel} would mean the "
            "world to us. We're a small operation and every review compounds.\n\n"
            "If anything wasn't perfect, please reply directly — I'd rather "
            "hear it from you than read about it later.\n\n"
            "Thanks again, and come back soon.\n\n"
            "— {host_first}"
        ),
    },
    {
        "title": "2. Year-end 'we miss you' — repeat-eligible guests",
        "subject": "It's been a year, {first_name} — back for another?",
        "body": (
            "Hi {first_name},\n\n"
            "It's been almost a year since you stayed with us at {property}, "
            "and I just wanted to reach out personally. The place has had a "
            "few upgrades since you were last here — new mattresses, the "
            "firepit area got a redesign, and we added a coffee bar.\n\n"
            "Spring dates are starting to fill in. If you're thinking about "
            "another trip, I'd love to host you again. Reply to this and I'll "
            "block the dates for you direct — no platform fees on your end.\n\n"
            "Hope you're well.\n\n"
            "— {host_first}"
        ),
    },
    {
        "title": "3. Direct-book invitation — after second Airbnb stay ($50 off)",
        "subject": "{first_name} — book direct next time and save $50",
        "body": (
            "Hi {first_name},\n\n"
            "Thanks again for staying with us at {property}. Since you've "
            "been with us twice now, I'd love to invite you to book direct "
            "next time you're heading our way.\n\n"
            "Booking direct gets you:\n"
            "  • $50 off your stay (use code REPEAT50)\n"
            "  • No Airbnb service fees\n"
            "  • First pick of dates before they open on the platforms\n"
            "  • A direct line to me if anything comes up\n\n"
            "Our direct site: {brand_domain}\n\n"
            "Same place, same hosts, better deal.\n\n"
            "— {host_first}"
        ),
    },
    {
        "title": "4. Referral request — after a 5-star review",
        "subject": "{first_name} — would you tell a friend about us?",
        "body": (
            "Hi {first_name},\n\n"
            "Thank you so much for the 5-star review — it genuinely makes a "
            "difference for a small operation like ours.\n\n"
            "If you have a friend or family member who might enjoy {property} "
            "as much as you did, I'd love to extend the same hospitality to "
            "them. Send them my way and I'll give them $75 off their first "
            "stay — and a $50 credit on your next visit as a thank-you.\n\n"
            "Just reply with their name and I'll take it from there.\n\n"
            "Grateful for you.\n\n"
            "— {host_first}"
        ),
    },
]


# --- Helpers ---

def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (operational-mode hero + KPIs + primary CTA)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Referral & Repeat Guest CRM"
    c.font = Font(name=FONT_HEAD, size=34, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 46

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Your top guests are worth more than your next listing photo."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: "What this does" card (rows 10-16, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20

    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Tracks where every guest came from (Airbnb / VRBO / Direct / Referral) "
        "and surfaces who keeps coming back. The repeat 20% of your guests will "
        "drive 80% of your direct bookings — if you can find them. This CRM "
        "scores guest LTV, maps referral chains, and feeds 4 outreach scripts "
        "you can copy-paste each year-end to convert OTA bookings into direct ones."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 3: "How to use" card (rows 18-25, parchment-alt) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20

    quickstart_items = [
        "① Guest CRM tab: log every guest as soon as the booking confirms — name, channel, property.",
        "② After each stay: update # Stays, Total nights, Total $ paid, Avg rating.",
        "③ Repeat & Referral tab: see your top-15 LTV guests and the channel mix donut.",
        "④ Year-end: open Outreach Templates, pick a script, paste into your email tool.",
        "⑤ Tag a referrer? Put their name in column H — referral chains build themselves.",
        "⑥ Settings: keep property + channel lists current so dropdowns stay clean.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # --- ZONE 4: Primary "ADD A GUEST" button (rows 27-30) ---
    pseudo_button(ws, "A27", "L30",
                   "→  ADD A NEW GUEST  (OPEN GUEST CRM)",
                   "'Guest CRM'!A6", variant="primary")
    for r in range(27, 31):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Activity at-a-glance — 4 cards (rows 32-37) ---
    for r in range(32, 38):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-C): Total Guests
    ws.merge_cells("A32:C32")
    c = ws["A32"]
    c.value = "TOTAL GUESTS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:C34")
    c = ws["A33"]
    c.value = f"=COUNTA('Guest CRM'!B6:B{LAST_GUEST_ROW})"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:C35")
    c = ws["A35"]
    c.value = "rows in the CRM"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (D-F): Repeat %
    ws.merge_cells("D32:F32")
    c = ws["D32"]
    c.value = "REPEAT %"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("D33:F34")
    c = ws["D33"]
    c.value = (
        f"=IFERROR(COUNTIF('Guest CRM'!I6:I{LAST_GUEST_ROW},\">1\")"
        f"/COUNTA('Guest CRM'!B6:B{LAST_GUEST_ROW}),0)"
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("D35:F35")
    c = ws["D35"]
    c.value = "guests with 2+ stays"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (G-I): Direct-book %
    ws.merge_cells("G32:I32")
    c = ws["G32"]
    c.value = "DIRECT-BOOK %"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("G33:I34")
    c = ws["G33"]
    c.value = (
        f"=IFERROR(COUNTIF('Guest CRM'!G6:G{LAST_GUEST_ROW},\"Direct\")"
        f"/COUNTA('Guest CRM'!B6:B{LAST_GUEST_ROW}),0)"
    )
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("G35:I35")
    c = ws["G35"]
    c.value = "fee-free bookings"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 4 (J-L): Top LTV $
    ws.merge_cells("J32:L32")
    c = ws["J32"]
    c.value = "TOP LTV"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("J33:L34")
    c = ws["J33"]
    c.value = f"=IFERROR(MAX('Guest CRM'!K6:K{LAST_GUEST_ROW}),0)"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("J35:L35")
    c = ws["J35"]
    c.value = "highest single-guest revenue"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "C"), ("D", "F"), ("G", "I"), ("J", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # --- ZONE 6: Secondary nav (rows 39-40) ---
    pseudo_button(ws, "A39", "C40", "Guest CRM",
                   "'Guest CRM'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Repeat & Referral",
                   "'Repeat & Referral'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Outreach Templates",
                   "'Outreach Templates'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # --- ZONE 7: Direct-book payoff callout (row 42) ---
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "💰 EVERY DIRECT BOOKING from a repeat guest saves you 14-20% in OTA "
        "fees. A guest with $1,200 LTV booked direct = ~$200 back in your "
        "pocket per stay. The 4 outreach scripts on the next tab convert."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[42].height = 36

    # --- ZONE 8: Reminder (row 44) ---
    ws.merge_cells("A44:L44")
    c = ws["A44"]
    c.value = (
        "⚠ OTAs (Airbnb/VRBO) restrict contacting guests outside their platform "
        "DURING the booking. Outreach is fine AFTER checkout. Capture email at "
        "stay (welcome book card) and you're free to use it post-stay."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[44].height = 28

    add_upgrade_banner(ws, 46)

    brand_footer(ws, 48,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:L50"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_guest_crm_tab(wb, variant):
    """Sheet 1 — Guest CRM register (capacity 500)."""
    ws = wb.create_sheet("Guest CRM")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Guest CRM",
                         prev_tab="Start", next_tab="Repeat & Referral")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Master register — log every guest, channel of origin, and referral lineage"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 13),  # First booking date
        ("B", 14),  # First name
        ("C", 14),  # Last name
        ("D", 26),  # Email
        ("E", 16),  # Phone
        ("F", 16),  # Property
        ("G", 12),  # Channel
        ("H", 22),  # Referrer
        ("I", 8),   # # Stays
        ("J", 9),   # Total nights
        ("K", 12),  # Total $ paid
        ("L", 9),   # Avg rating
        ("M", 13),  # Last stay date
        ("N", 13),  # Repeat?
        ("O", 30),  # Notes
    ])

    headers = [
        "First Booking Date", "First Name", "Last Name",
        "Email", "Phone",
        "Property", "Channel", "Referrer (if any)",
        "# Stays", "Total Nights", "Total $ Paid", "Avg Rating",
        "Last Stay Date", "Repeat?", "Notes / Preferences",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 22

    # Rows 6+: sample guests (DEMO only)
    guest_rows = GUESTS if variant == "demo" else []

    for row_idx, guest in enumerate(guest_rows, start=6):
        (first_date, first, last, email, phone, prop, channel,
         referrer, stays, nights, paid, rating, last_date, notes) = guest

        # A — First booking date
        a = ws.cell(row=row_idx, column=1, value=first_date)
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        # B-C names
        b = ws.cell(row=row_idx, column=2, value=first)
        apply_style(b, input_cell_style())
        c = ws.cell(row=row_idx, column=3, value=last)
        apply_style(c, input_cell_style())

        # D email, E phone
        d = ws.cell(row=row_idx, column=4, value=email if email else None)
        apply_style(d, input_cell_style())
        e = ws.cell(row=row_idx, column=5, value=phone if phone else None)
        apply_style(e, input_cell_style())

        # F property, G channel, H referrer
        f = ws.cell(row=row_idx, column=6, value=prop)
        apply_style(f, input_cell_style())
        g = ws.cell(row=row_idx, column=7, value=channel)
        apply_style(g, input_cell_style())
        h = ws.cell(row=row_idx, column=8, value=referrer if referrer else None)
        apply_style(h, input_cell_style())

        # I stays, J nights, K paid, L rating
        i = ws.cell(row=row_idx, column=9, value=stays)
        apply_style(i, input_cell_style())
        i.alignment = Alignment(horizontal="center", vertical="center")
        i.number_format = "0"

        j = ws.cell(row=row_idx, column=10, value=nights)
        apply_style(j, input_cell_style())
        j.alignment = Alignment(horizontal="center", vertical="center")
        j.number_format = "0"

        k = ws.cell(row=row_idx, column=11, value=paid)
        apply_style(k, input_cell_style())
        k.number_format = '"$"#,##0'

        l = ws.cell(row=row_idx, column=12, value=rating)
        apply_style(l, input_cell_style())
        l.alignment = Alignment(horizontal="center", vertical="center")
        l.number_format = "0.0"

        # M last stay date
        m = ws.cell(row=row_idx, column=13, value=last_date)
        apply_style(m, input_cell_style())
        m.number_format = "yyyy-mm-dd"

        # N Repeat? — formula
        n = ws.cell(
            row=row_idx, column=14,
            value=f'=IF(I{row_idx}="","",IF(I{row_idx}>1,"✓ Repeat",""))',
        )
        apply_style(n, formula_cell_style())
        n.alignment = Alignment(horizontal="center", vertical="center")

        # O notes
        o = ws.cell(row=row_idx, column=15, value=notes if notes else None)
        apply_style(o, input_cell_style())

        ws.row_dimensions[row_idx].height = 16

    # Capacity rows — blank input cells + formula for col N
    last_data_row = len(guest_rows) + 5
    for row_idx in range(last_data_row + 1, LAST_GUEST_ROW + 1):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == 14:
                # Repeat formula
                cell.value = (
                    f'=IF(I{row_idx}="","",'
                    f'IF(I{row_idx}>1,"✓ Repeat",""))'
                )
                apply_style(cell, formula_cell_style())
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                apply_style(cell, input_cell_style())
                if col_idx in (1, 13):
                    cell.number_format = "yyyy-mm-dd"
                elif col_idx in (9, 10):
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 11:
                    cell.number_format = '"$"#,##0'
                elif col_idx == 12:
                    cell.number_format = "0.0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 16

    # Conditional formatting — repeat-row gold-soft tint
    ws.conditional_formatting.add(
        f"A6:O{LAST_GUEST_ROW}",
        FormulaRule(
            formula=[f'$N6="✓ Repeat"'],
            fill=PatternFill("solid", fgColor="FFF3BF"),
        ),
    )

    # Data validation — channel + property + referrer dropdowns
    add_dropdown(ws, f"F6:F{LAST_GUEST_ROW}",
                  "=Settings!$B$7:$B$16")
    add_dropdown(ws, f"G6:G{LAST_GUEST_ROW}",
                  "=Settings!$B$18:$B$22")

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_repeat_referral_tab(wb):
    """Sheet 2 — Repeat & Referral KPIs + top-15 + channel mix donut."""
    ws = wb.create_sheet("Repeat & Referral")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = analysis tab

    compact_header_band(ws, "Repeat & Referral",
                         prev_tab="Guest CRM", next_tab="Outreach Templates")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Top-15 LTV guests · channel mix · referral lineage"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 22), ("B", 18), ("C", 9), ("D", 12), ("E", 12),
        ("F", 4),
        ("G", 14), ("H", 12), ("I", 12),
    ])

    # --- KPI strip (row 5-7) ---
    kpi_labels = [
        ("A5", "TOTAL GUESTS",
         f"=COUNTA('Guest CRM'!B6:B{LAST_GUEST_ROW})", "0"),
        ("B5", "REPEAT GUESTS",
         f"=COUNTIF('Guest CRM'!I6:I{LAST_GUEST_ROW},\">1\")", "0"),
        ("C5", "REPEAT %",
         f"=IFERROR(COUNTIF('Guest CRM'!I6:I{LAST_GUEST_ROW},\">1\")"
         f"/COUNTA('Guest CRM'!B6:B{LAST_GUEST_ROW}),0)", "0%"),
        ("D5", "DIRECT-BOOK %",
         f"=IFERROR(COUNTIF('Guest CRM'!G6:G{LAST_GUEST_ROW},\"Direct\")"
         f"/COUNTA('Guest CRM'!B6:B{LAST_GUEST_ROW}),0)", "0%"),
        ("E5", "AVG LTV",
         f"=IFERROR(AVERAGE('Guest CRM'!K6:K{LAST_GUEST_ROW}),0)",
         '"$"#,##0'),
    ]
    for cell_ref, label, _, _ in kpi_labels:
        c = ws[cell_ref]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for cell_ref, _, formula, fmt in kpi_labels:
        col_letter = cell_ref[0]
        v = ws[f"{col_letter}6"]
        v.value = formula
        v.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_PRIMARY)
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.number_format = fmt
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 28

    # KPI border
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for col_letter in ["A", "B", "C", "D", "E"]:
        col = column_index_from_string(col_letter)
        for r in (5, 6):
            existing = ws.cell(row=r, column=col).border
            ws.cell(row=r, column=col).border = Border(
                top=gold_side if r == 5 else existing.top,
                bottom=gold_side if r == 6 else existing.bottom,
                left=gold_side, right=gold_side,
            )

    # --- Top-15 by Total $ paid (rows 9-25) ---
    ws.row_dimensions[8].height = 10

    ws.merge_cells("A9:E9")
    c = ws["A9"]
    c.value = "TOP 15 GUESTS BY LIFETIME REVENUE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[9].height = 20

    top_headers = ["Rank", "Guest", "Stays", "Total $", "Repeat?"]
    for col, h in enumerate(top_headers, start=1):
        cell = ws.cell(row=10, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[10].height = 18

    # Use LARGE() to pull top 15 by Total $ (col K), then INDEX/MATCH back
    # to get the guest name + repeat flag.
    for i in range(15):
        row = 11 + i
        rank = i + 1

        # Rank
        a = ws.cell(row=row, column=1, value=rank)
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # Total $ first (so other columns can ref it)
        d_cell = ws.cell(
            row=row, column=4,
            value=(
                f"=IFERROR(LARGE('Guest CRM'!$K$6:$K${LAST_GUEST_ROW},"
                f"{rank}),\"\")"
            ),
        )
        apply_style(d_cell, formula_cell_style())
        d_cell.number_format = '"$"#,##0'
        d_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Guest name (first + last) — INDEX/MATCH on the Total $ value
        b = ws.cell(
            row=row, column=2,
            value=(
                f'=IF(D{row}="","",'
                f"INDEX('Guest CRM'!$B$6:$B${LAST_GUEST_ROW},"
                f"MATCH(D{row},'Guest CRM'!$K$6:$K${LAST_GUEST_ROW},0))"
                f'&" "&'
                f"INDEX('Guest CRM'!$C$6:$C${LAST_GUEST_ROW},"
                f"MATCH(D{row},'Guest CRM'!$K$6:$K${LAST_GUEST_ROW},0)))"
            ),
        )
        apply_style(b, formula_cell_style())
        b.alignment = Alignment(horizontal="left", vertical="center")

        # Stays
        c = ws.cell(
            row=row, column=3,
            value=(
                f'=IF(D{row}="","",'
                f"INDEX('Guest CRM'!$I$6:$I${LAST_GUEST_ROW},"
                f"MATCH(D{row},'Guest CRM'!$K$6:$K${LAST_GUEST_ROW},0)))"
            ),
        )
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Repeat?
        e = ws.cell(
            row=row, column=5,
            value=(
                f'=IF(C{row}="","",IF(C{row}>1,"✓","-"))'
            ),
        )
        apply_style(e, formula_cell_style())
        e.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 16

    # Conditional fmt — repeats highlighted gold-soft
    ws.conditional_formatting.add(
        "A11:E25",
        FormulaRule(
            formula=['$E11="✓"'],
            fill=PatternFill("solid", fgColor="FFF3BF"),
        ),
    )

    # --- Channel mix breakdown (rows 9-15, cols G-I) ---
    ws.merge_cells("G9:I9")
    c = ws["G9"]
    c.value = "CHANNEL MIX"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    chan_headers = ["Channel", "Guests", "% Mix"]
    for col_idx, h in enumerate(chan_headers, start=7):
        cell = ws.cell(row=10, column=col_idx, value=h)
        apply_style(cell, header_row_style())

    channel_list = ["Airbnb", "VRBO", "Booking", "Direct", "Referral"]
    for i, chan in enumerate(channel_list):
        row = 11 + i

        g = ws.cell(row=row, column=7, value=chan)
        g.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        g.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        h = ws.cell(
            row=row, column=8,
            value=f'=COUNTIF(\'Guest CRM\'!G6:G{LAST_GUEST_ROW},"{chan}")',
        )
        apply_style(h, formula_cell_style())
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.number_format = "0"

        i_cell = ws.cell(
            row=row, column=9,
            value=(
                f'=IFERROR(H{row}/COUNTA(\'Guest CRM\'!B6:B{LAST_GUEST_ROW}),0)'
            ),
        )
        apply_style(i_cell, formula_cell_style())
        i_cell.alignment = Alignment(horizontal="right", vertical="center")
        i_cell.number_format = "0%"

        ws.row_dimensions[row].height = 16

    # --- Channel mix donut chart (anchored G17) ---
    donut = DoughnutChart()
    donut.title = "Channel Mix — Where Guests Came From"
    donut.height = 9
    donut.width = 12
    donut.holeSize = 50
    donut_data = Reference(ws, min_col=8, min_row=10, max_row=15, max_col=8)
    donut_cats = Reference(ws, min_col=7, min_row=11, max_row=15)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_cats)
    donut.dataLabels = DataLabelList(showCatName=True, showPercent=True)
    style_chart(donut)
    ws.add_chart(donut, "G17")

    # --- Referral graph (rows 27-45) ---
    ws.row_dimensions[26].height = 10

    ws.merge_cells("A27:E27")
    c = ws["A27"]
    c.value = "REFERRAL LINEAGE — WHO REFERRED WHOM"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[27].height = 20

    ref_headers = ["#", "Referrer", "→", "Referred Guest", "Their LTV"]
    for col, h in enumerate(ref_headers, start=1):
        cell = ws.cell(row=28, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[28].height = 18

    # Up to 15 referral rows — pulled by SMALL row-position match on
    # Channel = "Referral" in the CRM. Manual fallback: customer types names.
    for i in range(15):
        row = 29 + i
        rank = i + 1
        # SMALL trick to pull the i-th row index where channel is "Referral"
        # then INDEX into B (first), C (last), H (referrer), K (LTV).
        helper = (
            f"IFERROR(SMALL("
            f"IF('Guest CRM'!$G$6:$G${LAST_GUEST_ROW}=\"Referral\","
            f"ROW('Guest CRM'!$G$6:$G${LAST_GUEST_ROW})),{rank}),\"\")"
        )

        a = ws.cell(row=row, column=1, value=rank)
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="center", vertical="center")

        b = ws.cell(
            row=row, column=2,
            value=(
                f'{{=IFERROR(INDEX(\'Guest CRM\'!$H:$H,{helper}),"")}}'
            ),
        )
        # Use array-formula via simple version (openpyxl writes literal string;
        # users on modern Excel get dynamic arrays).
        b.value = (
            f'=IFERROR(INDEX(\'Guest CRM\'!$H$1:$H${LAST_GUEST_ROW},'
            f'AGGREGATE(15,6,'
            f'ROW(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW})/'
            f'(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW}="Referral"),{rank})),"")'
        )
        apply_style(b, formula_cell_style())
        b.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_arrow = ws.cell(row=row, column=3, value="→")
        c_arrow.font = Font(name=FONT_BODY, size=12, bold=True, color=COLOR_ACCENT)
        c_arrow.alignment = Alignment(horizontal="center", vertical="center")

        d = ws.cell(
            row=row, column=4,
            value=(
                f'=IFERROR(INDEX(\'Guest CRM\'!$B$1:$B${LAST_GUEST_ROW},'
                f'AGGREGATE(15,6,'
                f'ROW(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW})/'
                f'(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW}="Referral"),{rank}))'
                f'&" "&'
                f'INDEX(\'Guest CRM\'!$C$1:$C${LAST_GUEST_ROW},'
                f'AGGREGATE(15,6,'
                f'ROW(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW})/'
                f'(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW}="Referral"),{rank})),"")'
            ),
        )
        apply_style(d, formula_cell_style())
        d.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        e = ws.cell(
            row=row, column=5,
            value=(
                f'=IFERROR(INDEX(\'Guest CRM\'!$K$1:$K${LAST_GUEST_ROW},'
                f'AGGREGATE(15,6,'
                f'ROW(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW})/'
                f'(\'Guest CRM\'!$G$6:$G${LAST_GUEST_ROW}="Referral"),{rank})),"")'
            ),
        )
        apply_style(e, formula_cell_style())
        e.alignment = Alignment(horizontal="right", vertical="center")
        e.number_format = '"$"#,##0'

        ws.row_dimensions[row].height = 16

    # --- Footer ---
    brand_footer(ws, 47,
                 version_line=f"{SKU} · v2.3 · Repeat & Referral analysis")

    ws.print_area = "A1:I49"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_outreach_tab(wb):
    """Sheet 3 — 4 copy-paste outreach scripts."""
    ws = wb.create_sheet("Outreach Templates")
    ws.sheet_properties.tabColor = COLOR_SECONDARY  # clay rose = communication

    compact_header_band(ws, "Outreach Templates",
                         prev_tab="Repeat & Referral", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Copy-paste scripts — replace {placeholders} with the guest's name and your details"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 4), ("B", 14), ("C", 80),
    ])

    # Each script gets: title row (gold-soft band) + subject row + body row
    cur_row = 6
    for i, script in enumerate(OUTREACH_SCRIPTS):
        # Title band
        ws.merge_cells(f"A{cur_row}:C{cur_row}")
        c = ws[f"A{cur_row}"]
        c.value = script["title"].upper()
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Subject row
        a = ws.cell(row=cur_row, column=2, value="Subject:")
        a.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="top")

        b = ws.cell(row=cur_row, column=3, value=script["subject"])
        b.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_PRIMARY)
        b.alignment = Alignment(horizontal="left", vertical="top",
                                 wrap_text=True, indent=1)
        b.fill = parchment_fill
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        # Body row
        a = ws.cell(row=cur_row, column=2, value="Body:")
        a.font = Font(name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="right", vertical="top")

        b = ws.cell(row=cur_row, column=3, value=script["body"])
        b.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        b.alignment = Alignment(horizontal="left", vertical="top",
                                 wrap_text=True, indent=1)
        b.fill = parchment_fill

        # Body row height — proportional to body length
        body_lines = script["body"].count("\n") + 1
        ws.row_dimensions[cur_row].height = max(140, body_lines * 16)
        cur_row += 1

        # Spacer
        ws.row_dimensions[cur_row].height = 12
        cur_row += 1

    # Placeholder reference card
    ws.merge_cells(f"A{cur_row}:C{cur_row}")
    c = ws[f"A{cur_row}"]
    c.value = "PLACEHOLDERS — REPLACE BEFORE SENDING"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[cur_row].height = 22
    cur_row += 1

    placeholders = [
        ("{first_name}", "Guest's first name (column B of Guest CRM)"),
        ("{property}", "Property they stayed at (column F)"),
        ("{channel}", "Channel they booked through — Airbnb / VRBO / etc"),
        ("{host_first}", "Your first name (the host signing the email)"),
        ("{brand_domain}", "Your direct-booking site URL"),
    ]
    for placeholder, desc in placeholders:
        a = ws.cell(row=cur_row, column=2, value=placeholder)
        a.font = Font(name=FONT_MONO, size=10, color=COLOR_PRIMARY)
        a.alignment = Alignment(horizontal="right", vertical="center")

        b = ws.cell(row=cur_row, column=3, value=desc)
        b.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        b.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    cur_row += 2
    brand_footer(ws, cur_row,
                 version_line=f"{SKU} · v2.3 · Outreach scripts")

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 4 — Settings (active year, properties, channels, referral sources)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Outreach Templates", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active year · property list · channel list · referral-source list"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 30), ("B", 28), ("C", 14), ("D", 16)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # --- Row 5: Active year ---
    a5 = ws.cell(row=5, column=1, value="Active year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")

    b5 = ws.cell(row=5, column=2, value=2026)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    b5.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 18

    ws.merge_cells("A6:D6")
    c6 = ws["A6"]
    c6.value = (
        "Used for year-end outreach campaigns (the 'we miss you' script "
        "filters guests whose last_stay_date < B5)."
    )
    c6.font = italic_muted
    c6.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[6].height = 24

    # --- Rows 7-16: Property list ---
    a7 = ws.cell(row=7, column=1, value="Property list (CRM dropdown):")
    a7.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a7.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 22

    sample_properties = (
        PROPERTIES + [""] * 7 if variant == "demo" else [""] * 10
    )
    for i, prop in enumerate(sample_properties[:10]):
        cell = ws.cell(row=7 + i, column=2, value=prop if prop else None)
        apply_style(cell, input_cell_style())

    # --- Rows 18-22: Channel list ---
    a18 = ws.cell(row=17, column=1, value="Channel list (CRM dropdown):")
    a18.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a18.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[17].height = 22

    for i, chan in enumerate(CHANNELS):
        cell = ws.cell(row=18 + i, column=2, value=chan)
        apply_style(cell, input_cell_style())

    # --- Rows 23-32: Referral source list ---
    a23 = ws.cell(row=23, column=1, value="Referral sources (reference):")
    a23.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a23.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[23].height = 22

    sources = (
        REFERRAL_SOURCES if variant == "demo"
        else ["Pinterest", "Instagram", "TikTok", "Blog post",
              "Word of mouth", "Local concierge", "Other", "", "", ""]
    )
    for i, src in enumerate(sources[:10]):
        cell = ws.cell(row=24 + i, column=2, value=src if src else None)
        apply_style(cell, input_cell_style())

    # --- Year-end Archive (rows 35+) ---
    sect35 = ws.cell(row=35, column=1, value="Year-end Archive")
    sect35.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[35].height = 22

    ws.merge_cells("A36:D36")
    c36 = ws["A36"]
    c36.value = (
        "Each January, copy the YTD KPIs into the closing-year row before you "
        "purge old guests. Repeat-rate trend tells you if your direct-book "
        "playbook is working."
    )
    c36.font = italic_muted
    c36.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    ws.row_dimensions[36].height = 28

    archive_headers = ["Year", "Total Guests", "Repeat %", "Direct-Book %"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=37, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[37].height = 18

    for idx, year in enumerate(range(2024, 2031), start=38):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0"
        for col in [3, 4]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0%"
        ws.row_dimensions[idx].height = 16

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_guest_crm_tab(wb, variant)
    build_repeat_referral_tab(wb)
    build_outreach_tab(wb)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Referral & Repeat Guest CRM{suffix} — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Channel-of-origin + repeat-stay CRM for STR hosts — "
        "drives direct bookings via targeted year-end outreach (v2.3)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
