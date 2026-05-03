"""Build LGL-002 Transient Occupancy Tax (TOT) Filing Calendar — operational tool.

Implements templates/_briefs/LGL-002-tot-filing-calendar.md.

Generates:
  templates/_masters/LGL-002-tot-filing-calendar-DEMO.xlsx
  templates/_masters/LGL-002-tot-filing-calendar-BLANK.xlsx

Tabs:
  0  Start            — operational hero + KPI cards + primary CTA
  1  Revenue Matrix   — 12-month per-property gross + cleaning input grid
  2  TOT Calculator   — per-property × month tax calc with platform-remit split
  3  Filing List      — rolling 12-month forward view sorted by due date
  4  Settings         — active tax year + property + jurisdiction lookup tables

Usage:
    python build_tot_filing_calendar.py
"""
from datetime import date
from pathlib import Path
from calendar import monthrange

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, compact_header_band, brand_footer,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
)

SKU = "LGL-002"
NAME = "tot-filing-calendar"
VERSION_LINE = f"{SKU} · v2.3 · Free updates forever"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

# Reference-data freshness stamp — TOT rates change. Customer should bump
# this and the Settings table whenever they refresh from their jurisdiction.
RATES_AS_OF = "2026-01-01"

ACTIVE_YEAR = 2026

MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

# DEMO data: 3 properties, 2 in Sevier County TN (Airbnb remits),
# 1 in Asheville NC (host self-files at 6%).
PROPERTIES = [
    # (name, jurisdiction)
    ("Smoky Mtn Cabin",   "Sevier County TN"),
    ("Pigeon Forge Lodge", "Sevier County TN"),
    ("Asheville Bungalow", "Asheville NC"),
]

# DEMO monthly gross rent per property (12 months)
DEMO_GROSS = {
    "Smoky Mtn Cabin":     [4200, 4400, 5800, 5200, 4900, 5400,
                              6800, 6200, 5100, 4800, 4200, 5600],
    "Pigeon Forge Lodge":  [3800, 3900, 5100, 4700, 4400, 4800,
                              6200, 5700, 4600, 4200, 3700, 4900],
    "Asheville Bungalow":  [2400, 2600, 3400, 3800, 4100, 4400,
                              4900, 4700, 3900, 3400, 2900, 3700],
}
# DEMO monthly cleaning fees per property
DEMO_CLEAN = {
    "Smoky Mtn Cabin":     [380, 410, 540, 480, 460, 510,
                              640, 580, 470, 440, 390, 520],
    "Pigeon Forge Lodge":  [340, 360, 470, 430, 410, 440,
                              580, 530, 420, 390, 340, 450],
    "Asheville Bungalow":  [220, 240, 320, 350, 380, 410,
                              460, 440, 360, 320, 270, 340],
}

# Jurisdiction registry: name | TOT rate | filing freq | due day | platform-remit | URL
JURISDICTIONS = [
    # (name, rate, frequency, due_day, platform_remit, url)
    ("Sevier County TN",   0.03,  "Monthly", 20, "Y",
     "https://www.seviercountytn.gov"),
    ("Asheville NC",       0.06,  "Monthly", 20, "N",
     "https://www.buncombecounty.org/governing/depts/tax"),
    ("Nashville TN",       0.06,  "Monthly", 20, "Y",
     "https://www.nashville.gov/finance"),
    ("New Orleans LA",     0.0675, "Monthly", 20, "N",
     "https://www.nola.gov/finance"),
    ("San Diego CA",       0.105,  "Monthly", 1,  "N",
     "https://www.sandiego.gov/treasurer"),
    ("Austin TX",          0.09,  "Quarterly", 20, "N",
     "https://www.austintexas.gov/department/hotel-occupancy-tax"),
    ("Park City UT",       0.0395, "Monthly", 30, "Y",
     "https://parkcity.org"),
]


def _val(variant, demo_value):
    """Return demo_value when building DEMO, None for BLANK."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(str(o) for o in options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1 — navy hero (rows 1-8)
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "TOT Filing Calendar"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color="F6EFE2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Never miss an occupancy-tax filing again."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6: VERDICT cell — next deadline + $ owed
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = (
        '=IF(SUM(\'TOT Calculator\'!H7:H42)=0,'
        '"\U0001F4CA  Fill Settings + Revenue Matrix to see your verdict.",'
        '"→  Next filing = "&TEXT(MIN(\'Filing List\'!C6:C41),"mmm d")'
        '&"  ·  Host owes YTD = "&TEXT(SUM(\'TOT Calculator\'!H7:H42),"$#,##0"))'
    )
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = navy_fill
    ws.row_dimensions[6].height = 32

    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2 — What this does (rows 10-16)
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A10:L10")
    c = ws["A10"]; c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20
    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "Aggregates monthly STR revenue by property and jurisdiction, computes "
        "transient occupancy / lodging tax due, and surfaces filing deadlines. "
        "Splits 'platform-remitted' (Airbnb files for you) from 'host self-files' "
        "so you only chase the filings that are actually yours. Missing a TOT "
        "filing = penalty + listing-suspension risk in many jurisdictions."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # ZONE 3 — How to use (rows 18-25)
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 26):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A18:L18")
    c = ws["A18"]; c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20
    quickstart_items = [
        "① Settings — set active tax year, list your properties + jurisdictions, confirm rates.",
        "② Revenue Matrix — paste or type each property's monthly gross rent + cleaning.",
        "③ TOT Calculator — auto-computes tax owed; platform-remitted rows show $0 host owes.",
        "④ Filing List — rolling deadlines, sorted ascending; red <7 days, gold 7-30.",
        "⑤ Mark Filed + Confirmation # after each submission.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # ZONE 4 — Primary CTA (rows 26-29)
    pseudo_button(ws, "A26", "L29", "→  GO TO FILING LIST  →",
                   "'Filing List'!A6", variant="primary")
    for r in range(26, 30):
        ws.row_dimensions[r].height = 22

    # ZONE 5 — Activity cards (rows 31-36)
    for r in range(31, 37):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 (A-D): TOT YTD owed
    ws.merge_cells("A32:D32")
    c = ws["A32"]; c.value = "TOT YTD HOST OWES"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]; c.value = "=SUM('TOT Calculator'!H7:H42)"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0;[Red]("$"#,##0)'
    ws.merge_cells("A35:D35")
    c = ws["A35"]; c.value = "self-filed only (excl. platform)"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 (E-H): Next deadline date
    ws.merge_cells("E32:H32")
    c = ws["E32"]; c.value = "NEXT DEADLINE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]
    c.value = (
        '=IFERROR(TEXT(MINIFS(\'Filing List\'!C6:C41,'
        '\'Filing List\'!E6:E41,"Pending"),"mmm d, yyyy"),"all clear")'
    )
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E35:H35")
    c = ws["E35"]; c.value = "earliest pending filing"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 (I-L): Pending filing count
    ws.merge_cells("I32:L32")
    c = ws["I32"]; c.value = "PENDING FILINGS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = '=COUNTIF(\'Filing List\'!E6:E41,"Pending")'
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0"
    ws.merge_cells("I35:L35")
    c = ws["I35"]; c.value = "rolling 12-month view"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 6 — Secondary nav (rows 38-39)
    pseudo_button(ws, "A38", "C39", "Revenue Matrix",
                   "'Revenue Matrix'!A1", variant="secondary")
    pseudo_button(ws, "D38", "F39", "TOT Calculator",
                   "'TOT Calculator'!A1", variant="secondary")
    pseudo_button(ws, "G38", "I39", "Filing List",
                   "'Filing List'!A1", variant="accent")
    pseudo_button(ws, "J38", "L39", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[38].height = 22
    ws.row_dimensions[39].height = 22

    # Legal-disclaimer callout (row 41)
    ws.merge_cells("A41:L41")
    c = ws["A41"]
    c.value = (
        "⚖  TOT rules vary widely by city/county. This calendar is a working "
        "tool, not legal advice. Confirm rate + filing rules with your "
        f"jurisdiction; rates table dated {RATES_AS_OF} — verify before filing."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[41].height = 36

    # Footer (rows 43-45)
    brand_footer(ws, 43, version_line=VERSION_LINE)

    ws.print_area = "A1:L45"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_revenue_matrix_tab(wb, variant):
    """Tab 1 — Revenue Matrix: properties × 12 months (gross + cleaning)."""
    ws = wb.create_sheet("Revenue Matrix")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Revenue Matrix",
                         prev_tab="Start", next_tab="TOT Calculator")

    # Row 4: subtitle
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:N4")
    c4 = ws["A4"]
    c4.value = ("Per-property monthly gross rent + cleaning fees. "
                  "Two rows per property (G = Gross, C = Cleaning). "
                  "Paste from your PMS export or type values manually.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 26), ("B", 6),
        ("C", 11), ("D", 11), ("E", 11), ("F", 11),
        ("G", 11), ("H", 11), ("I", 11), ("J", 11),
        ("K", 11), ("L", 11), ("M", 11), ("N", 11),
        ("O", 13),
    ])

    # Row 5: header — Property | Type | Jan...Dec | YTD
    headers = ["Property", "Type"] + MONTHS + ["YTD"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    bold_label = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    type_font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_MUTED)

    # 2 rows per property: Gross (G) + Cleaning (C). Banded fill on alternate
    # property pairs.
    for i, (prop, _juris) in enumerate(PROPERTIES):
        gross_row = 6 + i * 2
        clean_row = gross_row + 1
        band_fill = (PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
                       if i % 2 == 0 else None)

        # Property name spans both rows
        ws.merge_cells(start_row=gross_row, start_column=1,
                        end_row=clean_row, end_column=1)
        a = ws.cell(row=gross_row, column=1, value=prop)
        a.font = bold_label
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if band_fill:
            for r in (gross_row, clean_row):
                a2 = ws.cell(row=r, column=1)
                a2.fill = band_fill

        # Type labels
        t1 = ws.cell(row=gross_row, column=2, value="G")
        t1.font = type_font
        t1.alignment = Alignment(horizontal="center", vertical="center")
        t2 = ws.cell(row=clean_row, column=2, value="C")
        t2.font = type_font
        t2.alignment = Alignment(horizontal="center", vertical="center")
        if band_fill:
            t1.fill = band_fill; t2.fill = band_fill

        # 12 months of inputs
        gross_vals = DEMO_GROSS.get(prop, [0] * 12)
        clean_vals = DEMO_CLEAN.get(prop, [0] * 12)
        for m in range(12):
            col = 3 + m
            g = ws.cell(row=gross_row, column=col,
                          value=_val(variant, gross_vals[m]))
            apply_style(g, input_cell_style())
            g.number_format = '"$"#,##0'
            c = ws.cell(row=clean_row, column=col,
                          value=_val(variant, clean_vals[m]))
            apply_style(c, input_cell_style())
            c.number_format = '"$"#,##0'

        # YTD column O = SUM(C:N)
        for r in (gross_row, clean_row):
            f = ws.cell(row=r, column=15,
                          value=f"=SUM(C{r}:N{r})")
            apply_style(f, formula_cell_style())
            f.number_format = '"$"#,##0'

        ws.row_dimensions[gross_row].height = 18
        ws.row_dimensions[clean_row].height = 18

    last_row = 5 + len(PROPERTIES) * 2

    # Portfolio total row
    total_row = last_row + 2
    ws.row_dimensions[last_row + 1].height = 8
    a = ws.cell(row=total_row, column=1, value="PORTFOLIO TOTAL (Gross + Cleaning)")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    a.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells(start_row=total_row, start_column=1,
                    end_row=total_row, end_column=2)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                     value=f"=SUM({col_letter}6:{col_letter}{last_row})")
        c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[total_row].height = 24

    ws.freeze_panes = "C6"

    # Print setup
    ws.print_area = f"A1:O{total_row + 1}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_tot_calculator_tab(wb, variant):
    """Tab 2 — TOT Calculator: per-property × month tax calc."""
    ws = wb.create_sheet("TOT Calculator")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "TOT Calculator",
                         prev_tab="Revenue Matrix", next_tab="Filing List")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:H4")
    c4 = ws["A4"]
    c4.value = ("Per property × month: taxable revenue × jurisdiction TOT rate. "
                  "Platform-remitted rows zero out (Airbnb files for you).")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 26), ("B", 22), ("C", 10),
        ("D", 14), ("E", 14), ("F", 11),
        ("G", 13), ("H", 14),
    ])

    headers = ["Property", "Jurisdiction", "Month",
                "Gross Rent", "Cleaning", "Rate",
                "Platform Files?", "$ Host Owes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    bold = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)

    # Per property: 12 monthly rows (one per month) — 3 props × 12 = 36 rows
    # Rows 7..42
    for p_idx, (prop, juris) in enumerate(PROPERTIES):
        # Find this property's rows in Revenue Matrix:
        # gross row = 6 + p_idx * 2; cleaning row = gross + 1
        gross_rm_row = 6 + p_idx * 2
        clean_rm_row = gross_rm_row + 1

        for m in range(12):
            r = 7 + p_idx * 12 + m
            band_fill = (PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
                            if (p_idx + m) % 2 == 0 else None)

            # Col A: Property
            a = ws.cell(row=r, column=1, value=prop)
            a.font = bold
            a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if band_fill:
                a.fill = band_fill

            # Col B: Jurisdiction (input — defaults to mapped value, blank in BLANK)
            b = ws.cell(row=r, column=2, value=_val(variant, juris))
            apply_style(b, input_cell_style())

            # Col C: Month
            mc = ws.cell(row=r, column=3, value=MONTHS[m])
            mc.font = bold
            mc.alignment = Alignment(horizontal="center", vertical="center")
            if band_fill:
                mc.fill = band_fill

            # Col D: Gross from Revenue Matrix
            rm_col = get_column_letter(3 + m)  # C..N
            d = ws.cell(row=r, column=4,
                          value=f"='Revenue Matrix'!{rm_col}{gross_rm_row}")
            apply_style(d, formula_cell_style())
            d.number_format = '"$"#,##0'

            # Col E: Cleaning from Revenue Matrix
            e = ws.cell(row=r, column=5,
                          value=f"='Revenue Matrix'!{rm_col}{clean_rm_row}")
            apply_style(e, formula_cell_style())
            e.number_format = '"$"#,##0'

            # Col F: Rate (VLOOKUP against Settings jurisdictions table)
            f_ = ws.cell(row=r, column=6,
                           value=f'=IFERROR(VLOOKUP(B{r},Settings!$B$10:$G$25,2,FALSE),0)')
            apply_style(f_, formula_cell_style())
            f_.number_format = "0.00%"

            # Col G: Platform Files? (VLOOKUP)
            g = ws.cell(row=r, column=7,
                          value=f'=IFERROR(VLOOKUP(B{r},Settings!$B$10:$G$25,5,FALSE),"?")')
            apply_style(g, formula_cell_style())
            g.alignment = Alignment(horizontal="center", vertical="center")

            # Col H: $ Host Owes — if platform files, $0; else (gross+cleaning) * rate
            h = ws.cell(row=r, column=8,
                          value=f'=IF(G{r}="Y",0,ROUND((D{r}+E{r})*F{r},2))')
            apply_style(h, formula_cell_style())
            h.number_format = '"$"#,##0.00'

            ws.row_dimensions[r].height = 16

    last_row = 7 + len(PROPERTIES) * 12 - 1  # 42

    # Total row
    total_row = last_row + 2
    ws.row_dimensions[last_row + 1].height = 8
    a = ws.cell(row=total_row, column=1, value="TOTAL — $ HOST OWES YTD")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    a.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells(start_row=total_row, start_column=1,
                    end_row=total_row, end_column=7)
    for col in range(2, 8):
        ws.cell(row=total_row, column=col).fill = (
            PatternFill("solid", fgColor=COLOR_PRIMARY)
        )
    c = ws.cell(row=total_row, column=8,
                  value=f"=SUM(H7:H{last_row})")
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.number_format = '"$"#,##0.00'
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[total_row].height = 24

    # Conditional formatting: gold for platform-remitted rows (informational)
    ws.conditional_formatting.add(
        f"A7:H{last_row}",
        FormulaRule(
            formula=[f'$G7="Y"'],
            font=Font(name=FONT_BODY, size=11, italic=True, color=COLOR_MUTED),
        ),
    )

    ws.freeze_panes = "A6"

    # Print setup
    ws.print_area = f"A1:H{total_row + 1}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def _due_date_for(year, month_idx, due_day):
    """Filing for `month_idx` (0-11) of `year` is due on `due_day` of the
    NEXT month. Clamp due_day to the month's last day if needed.
    """
    fy = year
    fm = month_idx + 2  # next month, 1-based
    if fm > 12:
        fm -= 12
        fy += 1
    last_day = monthrange(fy, fm)[1]
    return date(fy, fm, min(due_day, last_day))


def build_filing_list_tab(wb, variant):
    """Tab 3 — Filing List: rolling 12-month forward view sorted by deadline."""
    ws = wb.create_sheet("Filing List")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = action surface
    compact_header_band(ws, "Filing List · Upcoming Deadlines",
                         prev_tab="TOT Calculator", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:H4")
    c4 = ws["A4"]
    c4.value = ("📅 Rolling 12-month forward view per property × period. "
                  "Red = <7 days, gold = 7-30 days, parchment = >30 days.")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 26), ("B", 16), ("C", 14),
        ("D", 14), ("E", 12), ("F", 22),
        ("G", 6), ("H", 6),
    ])

    headers = ["Property", "Period", "Due Date",
                "$ Owed", "Status", "Confirmation #"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    bold = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)

    # Build rows: 3 properties × 12 periods (matching TOT Calculator order).
    # Period label = "<MonName> <Year>" (the revenue month being reported).
    # Due date computed via _due_date_for using each jurisdiction's due_day.
    juris_due_day = {j[0]: j[3] for j in JURISDICTIONS}

    for p_idx, (prop, juris) in enumerate(PROPERTIES):
        for m in range(12):
            r = 6 + p_idx * 12 + m
            band_fill = (PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
                            if (p_idx + m) % 2 == 0 else None)

            # Property
            a = ws.cell(row=r, column=1, value=prop)
            a.font = bold
            a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if band_fill:
                a.fill = band_fill

            # Period label
            pc = ws.cell(row=r, column=2, value=f"{MONTHS[m]} {ACTIVE_YEAR}")
            pc.font = bold
            pc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if band_fill:
                pc.fill = band_fill

            # Due date — datetime.date object
            due_day = juris_due_day.get(juris, 20)
            d_obj = _due_date_for(ACTIVE_YEAR, m, due_day)
            dc = ws.cell(row=r, column=3, value=d_obj)
            dc.font = bold
            dc.number_format = "yyyy-mm-dd"
            dc.alignment = Alignment(horizontal="center", vertical="center")

            # $ Owed — pull from TOT Calculator H column for the matching row
            calc_row = 7 + p_idx * 12 + m
            o = ws.cell(row=r, column=4,
                          value=f"='TOT Calculator'!H{calc_row}")
            apply_style(o, formula_cell_style())
            o.number_format = '"$"#,##0.00'

            # Status — input dropdown (default Pending in DEMO, blank in BLANK)
            s = ws.cell(row=r, column=5, value=_val(variant, "Pending"))
            apply_style(s, input_cell_style())
            s.alignment = Alignment(horizontal="center", vertical="center")

            # Confirmation # — input
            cf = ws.cell(row=r, column=6, value=None)
            apply_style(cf, input_cell_style())

            ws.row_dimensions[r].height = 18

    last_row = 5 + len(PROPERTIES) * 12  # 41

    add_dropdown(ws, f"E6:E{last_row}", ["Pending", "Filed", "Paid", "N/A"])

    # Conditional formatting on the WHOLE row band by due date
    # Rule 1: <7 days from today AND status != Filed/Paid → red
    ws.conditional_formatting.add(
        f"A6:F{last_row}",
        FormulaRule(
            formula=[f'AND($C6-TODAY()<7,$C6>=TODAY(),$E6<>"Filed",$E6<>"Paid",$E6<>"N/A")'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
        ),
    )
    # Rule 2: 7-30 days, not done → gold-soft
    ws.conditional_formatting.add(
        f"A6:F{last_row}",
        FormulaRule(
            formula=[f'AND($C6-TODAY()>=7,$C6-TODAY()<=30,$E6<>"Filed",$E6<>"Paid",$E6<>"N/A")'],
            fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT),
        ),
    )
    # Rule 3: Past due, not done → strong red
    ws.conditional_formatting.add(
        f"A6:F{last_row}",
        FormulaRule(
            formula=[f'AND($C6<TODAY(),$E6<>"Filed",$E6<>"Paid",$E6<>"N/A")'],
            fill=PatternFill("solid", fgColor="F8B4B4"),
            font=Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ERROR),
        ),
    )

    # Enable autofilter so user can sort by due date / filter pending
    ws.auto_filter.ref = f"A5:F{last_row}"
    ws.freeze_panes = "A6"

    # Print setup
    ws.print_area = f"A1:F{last_row + 1}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Settings",
                         prev_tab="Filing List", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:G4")
    c4 = ws["A4"]
    c4.value = ("Active tax year · property list · jurisdiction registry "
                  "(rate, frequency, due day, platform-remit Y/N).")
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 32), ("B", 26), ("C", 10),
        ("D", 13), ("E", 11), ("F", 13),
        ("G", 38),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    # B5 active tax year — STRUCTURAL, present in both variants (rule 4).
    a = ws.cell(row=5, column=1, value="Active tax year:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    yc = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(yc, input_cell_style())
    yc.number_format = "0"
    ws.row_dimensions[5].height = 20

    # Property list section (rows 7-16) — B7..B16 reserved per brief
    a = ws.cell(row=7, column=1, value="Properties")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A7:B7")
    ws.row_dimensions[7].height = 20

    for i in range(10):
        r = 8 + i  # 8..17 (we'll list properties starting row 8)
        # Slot label
        lbl = ws.cell(row=r, column=1, value=f"Property {i + 1}:")
        lbl.font = bold_right
        lbl.alignment = Alignment(horizontal="right", vertical="center")
        # Slot value
        prop_val = (PROPERTIES[i][0]
                       if (variant == "demo" and i < len(PROPERTIES))
                       else None)
        c = ws.cell(row=r, column=2, value=prop_val)
        apply_style(c, input_cell_style())
        ws.row_dimensions[r].height = 16

    # Jurisdiction list (rows 19-26) — B19..B25 per spec range
    a = ws.cell(row=19, column=1, value="Jurisdictions")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A19:G19")
    ws.row_dimensions[19].height = 20

    # Header row at row 20
    j_headers = ["#", "Jurisdiction", "TOT Rate", "Frequency",
                  "Due Day", "Platform Remits?", "Filing URL"]
    for col, h in enumerate(j_headers, start=1):
        cell = ws.cell(row=20, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[20].height = 20

    # Note: TOT Calculator's VLOOKUP references Settings!$B$10:$G$25 (range
    # spans 16 rows). Jurisdiction data starts at row 21 here, so adjust
    # references accordingly. Actually we want B$21:G$X — fix Calculator
    # references at the end. For now, populate jurisdictions starting row 21.

    for i, (jname, rate, freq, due_day, remit, url) in enumerate(JURISDICTIONS):
        r = 21 + i
        idx = ws.cell(row=r, column=1, value=i + 1)
        idx.font = Font(name=FONT_MONO, size=10, color=COLOR_MUTED)
        idx.alignment = Alignment(horizontal="center", vertical="center")
        # Always fill structural lookup rows (jurisdiction registry is
        # reference data, not customer-specific dollars).
        nc = ws.cell(row=r, column=2, value=jname)
        apply_style(nc, input_cell_style())
        rc = ws.cell(row=r, column=3, value=rate)
        apply_style(rc, input_cell_style())
        rc.number_format = "0.00%"
        fc = ws.cell(row=r, column=4, value=freq)
        apply_style(fc, input_cell_style())
        fc.alignment = Alignment(horizontal="center", vertical="center")
        dc = ws.cell(row=r, column=5, value=due_day)
        apply_style(dc, input_cell_style())
        dc.alignment = Alignment(horizontal="center", vertical="center")
        rmc = ws.cell(row=r, column=6, value=remit)
        apply_style(rmc, input_cell_style())
        rmc.alignment = Alignment(horizontal="center", vertical="center")
        uc = ws.cell(row=r, column=7, value=url)
        apply_style(uc, input_cell_style())
        ws.row_dimensions[r].height = 16

    juris_last = 20 + len(JURISDICTIONS)
    add_dropdown(ws, f"D21:D{juris_last}", ["Monthly", "Quarterly", "Annual"])
    add_dropdown(ws, f"F21:F{juris_last}", ["Y", "N"])

    # Freshness stamp + legal disclaimer (row juris_last + 2)
    note_row = juris_last + 2
    ws.merge_cells(f"A{note_row}:G{note_row}")
    a = ws.cell(row=note_row, column=1)
    a.value = (
        f"📅 Rates table dated {RATES_AS_OF}. TOT rules change frequently — "
        "verify rate, due day, and platform-remit status with your jurisdiction "
        "before filing. This workbook is a working tool, not legal advice."
    )
    a.font = italic_muted
    a.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    a.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    ws.row_dimensions[note_row].height = 36

    # Year-end Archive (suite-wide pattern)
    archive_start = note_row + 2
    a = ws.cell(row=archive_start, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[archive_start].height = 22

    ws.merge_cells(f"A{archive_start + 1}:G{archive_start + 1}")
    c = ws.cell(row=archive_start + 1, column=1)
    c.value = ("Each January, copy YTD totals into the row for the year you "
                 "just closed, then clear Revenue Matrix for the new year.")
    c.font = italic_muted
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[archive_start + 1].height = 24

    arch_headers = ["Year", "Gross Revenue", "Cleaning",
                      "$ Host Owes", "$ Platform Filed"]
    for col, h in enumerate(arch_headers, start=1):
        cell = ws.cell(row=archive_start + 2, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[archive_start + 2].height = 18

    for idx, yr in enumerate(range(2024, 2031)):
        r = archive_start + 3 + idx
        ws.cell(row=r, column=1, value=yr).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=r, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 16

    # Print setup
    ws.print_area = f"A1:G{archive_start + 9}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return juris_last


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_revenue_matrix_tab(wb, variant)
    build_tot_calculator_tab(wb, variant)
    build_filing_list_tab(wb, variant)
    juris_last = build_settings_tab(wb, variant)

    # Patch TOT Calculator VLOOKUP ranges to point at the actual Settings
    # jurisdiction range (rows 21..juris_last). Default range used during
    # initial write was $B$10:$G$25 — overwrite to the correct one.
    calc_ws = wb["TOT Calculator"]
    new_range = f"Settings!$B$21:$G${juris_last}"
    for r in range(7, 7 + len(PROPERTIES) * 12):
        # F column: rate lookup (col 2)
        calc_ws.cell(row=r, column=6).value = (
            f'=IFERROR(VLOOKUP(B{r},{new_range},2,FALSE),0)'
        )
        # G column: platform remit lookup (col 5)
        calc_ws.cell(row=r, column=7).value = (
            f'=IFERROR(VLOOKUP(B{r},{new_range},5,FALSE),"?")'
        )

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"TOT Filing Calendar{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Per-property monthly STR occupancy tax tracker with platform-remit "
        "split and rolling deadline view."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
