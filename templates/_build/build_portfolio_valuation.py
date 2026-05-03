"""Build STR-002 Portfolio Valuation Model (v2.2 standard).

Operational tool — host runs annually OR before refi/sale conversation.
Computes current market value of an STR portfolio using 3 methods averaged
(cap rate x NOI, comparable $/BR, and gross rent multiplier), then surfaces
total equity, debt-to-value (LTV), and "what would a buyer pay" estimate.

Forks the structure from build_str_deal_analyzer.py (NOI/cap pattern,
helpers, page layout) and the dashboard/chart approach from
build_owner_reporting_dashboard.py.

Generates:
  templates/_masters/STR-002-portfolio-valuation-model-DEMO.xlsx
  templates/_masters/STR-002-portfolio-valuation-model-BLANK.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
)

SKU = "STR-002"
NAME = "portfolio-valuation-model"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026
PROPERTY_CAP = 10  # portfolio capacity

# ---------------------------------------------------------------------------
# Sample data (per brief: 3 properties, $1.74M value, $1.10M debt, $634K eq)
# ---------------------------------------------------------------------------

SAMPLE_PROPS = [
    # name, market_type, noi, comp_per_br, beds, baths,
    # gross_rent, grm, loan_balance
    ("Smokies Ridge",  "Mountain resort STR", 42000, 195000, 3, 2, 96000,  6.04, 364000),
    ("Lakehouse",      "Beach STR",           48000, 240000, 3, 2.5, 110000, 6.55, 469000),
    ("Creek Side",     "Rural / cabin",       34000, 150000, 3, 2, 78000,  5.64, 270000),
]

# Cap-rate reference table — editable lookup keyed off market_type
SAMPLE_CAP_RATES = [
    ("Mountain resort STR", 0.075),
    ("Beach STR",           0.070),
    ("Urban STR",           0.090),
    ("Rural / cabin",       0.080),
]

MARKET_TYPES = [m for m, _ in SAMPLE_CAP_RATES]


# ---------------------------------------------------------------------------
# Utilities (mirrored from build_str_deal_analyzer.py)
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color="F6EFE2")
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _kpi_card(ws, top_left, bottom_right, label, formula, fmt, accent=False):
    """Render a parchment KPI card with a small label and a big number.

    CRITICAL: write label/formula FIRST, then merge separately. We never
    write to a merged cell after merging.
    """
    tl_col_letter, tl_row = top_left[0], int(top_left[1:])
    br_col_letter, br_row = bottom_right[0], int(bottom_right[1:])

    # Merge label strip (top row of card) and big-number block (rows below)
    label_range = f"{top_left}:{br_col_letter}{tl_row}"
    number_range = f"{tl_col_letter}{tl_row + 1}:{bottom_right}"

    ws[top_left] = label.upper()
    ws[top_left].font = Font(name=FONT_MONO, size=9, bold=True,
                              color=COLOR_PRIMARY)
    ws[top_left].alignment = Alignment(horizontal="center", vertical="center")
    ws[top_left].fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.merge_cells(label_range)

    big_cell_ref = f"{tl_col_letter}{tl_row + 1}"
    ws[big_cell_ref] = formula
    ws[big_cell_ref].number_format = fmt
    ws[big_cell_ref].font = Font(
        name=FONT_HEAD, size=22, bold=True,
        color=COLOR_ACCENT if accent else COLOR_PRIMARY,
    )
    ws[big_cell_ref].fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    ws[big_cell_ref].alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(number_range)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Navy hero band rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Brand wordmark row 2 (write BEFORE merging)
    ws["A2"] = BRAND_NAME
    ws["A2"].font = Font(name=FONT_HEAD, size=14, color="F6EFE2")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A2:F2")

    # Title row 4 (write BEFORE merging)
    ws["A4"] = "Portfolio Valuation Model"
    ws["A4"].font = Font(name=FONT_HEAD, size=30, bold=True, color="F6EFE2")
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44
    ws.merge_cells("A4:L4")

    ws["A5"] = "Three valuation methods. One number a buyer would pay."
    ws["A5"].font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A5:L5")

    ws["A7"] = f"{SKU} · v1.0 · {variant.upper()}"
    ws["A7"].font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A7:L7")

    # Parchment headline zone rows 10-22
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # KPI cards: portfolio value, equity, LTV
    # Card 1: Market value (rows 11-14, cols A-D)
    _kpi_card(ws, "A11", "D14", "Portfolio market value",
              "='Equity Summary'!C7", '"$"#,##0', accent=False)
    # Card 2: Equity (rows 11-14, cols E-H)
    _kpi_card(ws, "E11", "H14", "Total equity",
              "='Equity Summary'!C9", '"$"#,##0', accent=True)
    # Card 3: Portfolio LTV (rows 11-14, cols I-L)
    _kpi_card(ws, "I11", "L14", "Portfolio LTV",
              "='Equity Summary'!C10", "0.0%", accent=False)
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 28
    ws.row_dimensions[13].height = 16
    ws.row_dimensions[14].height = 16

    # Subhead row 16 (write BEFORE merging)
    ws["A16"] = ('="As of "&TEXT(Settings!B5,"yyyy")'
                  '&" · "&COUNTIF(Settings!B7:B16,"<>")&" properties tracked"')
    ws["A16"].font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_MUTED)
    ws["A16"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 22
    ws.merge_cells("A16:L16")

    # Buyer estimate row 18-20 (write BEFORE merge)
    ws["A18"] = "WHAT A BUYER WOULD PAY"
    ws["A18"].font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    ws["A18"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 18
    ws.merge_cells("A18:L18")

    ws["A19"] = "='Equity Summary'!C7"
    ws["A19"].font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_ACCENT)
    ws["A19"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A19"].number_format = '"$"#,##0'
    ws.row_dimensions[19].height = 36
    ws.merge_cells("A19:L20")

    # Pseudo-button nav rows 24-26
    pseudo_button(ws, "A24", "C26", "Per Property",
                  "'Per Property'!A1", variant="primary")
    pseudo_button(ws, "D24", "F26", "Cap Rate Reference",
                  "'Cap Rate Reference'!A1", variant="primary")
    pseudo_button(ws, "G24", "I26", "Equity Summary",
                  "'Equity Summary'!A1", variant="accent")
    pseudo_button(ws, "J24", "L26", "Settings",
                  "'Settings'!A1", variant="primary")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    # CTA banner row 29 (write BEFORE merging)
    cta = (
        "💡  Run this annually — and any time you're talking refi or sale. "
        f"Free updates forever at {BRAND_DOMAIN}."
    )
    ws["A29"] = cta
    ws["A29"].font = Font(name=FONT_BODY, size=11, bold=True, color="FFFFFF")
    ws["A29"].fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    ws["A29"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[29].height = 36
    ws.merge_cells("A29:L29")

    brand_footer(ws, 31,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L33"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_per_property_tab(wb, variant):
    ws = wb.create_sheet("Per Property")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    # Columns A..O (15 cols)
    set_col_widths(ws, [
        ("A", 22), ("B", 14), ("C", 10), ("D", 14),
        ("E", 14), ("F", 7), ("G", 7), ("H", 14),
        ("I", 14), ("J", 8), ("K", 14), ("L", 14),
        ("M", 14), ("N", 14), ("O", 9),
    ])

    compact_header_band(ws, "Per Property",
                        prev_tab="Start", next_tab="Cap Rate Reference")

    # Subtitle row 4 (write BEFORE merging)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws["A4"] = "Three methods averaged into final value column · capacity 10 properties"
    ws["A4"].font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:L4")

    # ---- Header row 6 ----
    headers = [
        ("A", "Property"),
        ("B", "NOI YTD"),
        ("C", "Cap rate"),
        ("D", "M1: Cap×NOI"),
        ("E", "Comp $/BR"),
        ("F", "Beds"),
        ("G", "Baths"),
        ("H", "M2: Comp value"),
        ("I", "Annual gross rent"),
        ("J", "GRM"),
        ("K", "M3: GRM×Rent"),
        ("L", "Average value"),
        ("M", "Loan balance"),
        ("N", "Equity"),
        ("O", "LTV"),
    ]
    for col, label in headers:
        cell = ws[f"{col}6"]
        cell.value = label
        apply_style(cell, header_row_style())
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
    ws.row_dimensions[6].height = 32

    # ---- Data rows 7..(7+PROPERTY_CAP-1) = 7..16 ----
    first_data = 7
    last_data = first_data + PROPERTY_CAP - 1  # 16

    for i in range(PROPERTY_CAP):
        r = first_data + i
        sample = SAMPLE_PROPS[i] if i < len(SAMPLE_PROPS) else None

        # A: Property name — pull from Settings (single source of truth)
        ws.cell(row=r, column=1).value = f"=Settings!B{7 + i}"
        apply_style(ws.cell(row=r, column=1), formula_cell_style())
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )

        # B: NOI YTD (input)
        c = ws.cell(row=r, column=2)
        c.value = sample[2] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '"$"#,##0'

        # C: Cap rate — VLOOKUP off Settings market type to Cap Rate Reference
        # Property's market type lives at Settings!C7+i (column C in Settings)
        ws.cell(row=r, column=3).value = (
            f"=IFERROR(VLOOKUP(Settings!C{7 + i},"
            f"'Cap Rate Reference'!A7:B16,2,FALSE),0)"
        )
        apply_style(ws.cell(row=r, column=3), formula_cell_style())
        ws.cell(row=r, column=3).number_format = "0.00%"
        ws.cell(row=r, column=3).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # D: Method 1 = NOI / cap
        ws.cell(row=r, column=4).value = (
            f'=IF(OR(B{r}="",C{r}=0,C{r}=""),"",B{r}/C{r})'
        )
        apply_style(ws.cell(row=r, column=4), formula_cell_style())
        ws.cell(row=r, column=4).number_format = '"$"#,##0'
        ws.cell(row=r, column=4).alignment = Alignment(
            horizontal="right", vertical="center"
        )

        # E: Comp $/BR (input)
        c = ws.cell(row=r, column=5)
        c.value = sample[3] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '"$"#,##0'

        # F: Beds (input)
        c = ws.cell(row=r, column=6)
        c.value = sample[4] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0"

        # G: Baths (input)
        c = ws.cell(row=r, column=7)
        c.value = sample[5] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.0"

        # H: Method 2 = Comp $/BR × Beds + bath premium
        # Per brief column H "Method 2: Comp value (formula =E*F adjusted)"
        # We use beds as primary driver, +5% per extra half-bath beyond beds
        ws.cell(row=r, column=8).value = (
            f'=IF(OR(E{r}="",F{r}=""),"",E{r}*F{r}*(1+MAX(0,G{r}-F{r})*0.05))'
        )
        apply_style(ws.cell(row=r, column=8), formula_cell_style())
        ws.cell(row=r, column=8).number_format = '"$"#,##0'
        ws.cell(row=r, column=8).alignment = Alignment(
            horizontal="right", vertical="center"
        )

        # I: Annual gross rent (input)
        c = ws.cell(row=r, column=9)
        c.value = sample[6] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '"$"#,##0'

        # J: GRM (input)
        c = ws.cell(row=r, column=10)
        c.value = sample[7] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.00"

        # K: Method 3 = Gross rent × GRM
        ws.cell(row=r, column=11).value = (
            f'=IF(OR(I{r}="",J{r}=""),"",I{r}*J{r})'
        )
        apply_style(ws.cell(row=r, column=11), formula_cell_style())
        ws.cell(row=r, column=11).number_format = '"$"#,##0'
        ws.cell(row=r, column=11).alignment = Alignment(
            horizontal="right", vertical="center"
        )

        # L: Average value (avg of D, H, K — only counts non-blanks)
        ws.cell(row=r, column=12).value = (
            f'=IFERROR(AVERAGE(IF(ISNUMBER(D{r}),D{r}),'
            f'IF(ISNUMBER(H{r}),H{r}),IF(ISNUMBER(K{r}),K{r})),"")'
        )
        apply_style(ws.cell(row=r, column=12), formula_cell_style())
        ws.cell(row=r, column=12).number_format = '"$"#,##0'
        ws.cell(row=r, column=12).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY
        )
        ws.cell(row=r, column=12).alignment = Alignment(
            horizontal="right", vertical="center"
        )
        ws.cell(row=r, column=12).fill = PatternFill(
            "solid", fgColor=COLOR_GOLD_SOFT
        )

        # M: Loan balance (input)
        c = ws.cell(row=r, column=13)
        c.value = sample[8] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '"$"#,##0'

        # N: Equity = L - M
        ws.cell(row=r, column=14).value = (
            f'=IF(OR(L{r}="",M{r}=""),"",L{r}-M{r})'
        )
        apply_style(ws.cell(row=r, column=14), formula_cell_style())
        ws.cell(row=r, column=14).number_format = '"$"#,##0'
        ws.cell(row=r, column=14).alignment = Alignment(
            horizontal="right", vertical="center"
        )

        # O: LTV = M / L
        ws.cell(row=r, column=15).value = (
            f'=IF(OR(L{r}="",L{r}=0,M{r}=""),"",M{r}/L{r})'
        )
        apply_style(ws.cell(row=r, column=15), formula_cell_style())
        ws.cell(row=r, column=15).number_format = "0.0%"
        ws.cell(row=r, column=15).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        ws.row_dimensions[r].height = 20

    # ---- Totals row ----
    totals_row = last_data + 1  # 17
    ws.cell(row=totals_row, column=1).value = "PORTFOLIO TOTAL"
    ws.cell(row=totals_row, column=1).font = Font(
        name=FONT_HEAD, size=12, bold=True, color="F6EFE2"
    )
    ws.cell(row=totals_row, column=1).fill = PatternFill(
        "solid", fgColor=COLOR_PRIMARY
    )
    ws.cell(row=totals_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )

    # Fill remaining cols with navy (write BEFORE any merge — none here)
    for col in range(2, 16):
        ws.cell(row=totals_row, column=col).fill = PatternFill(
            "solid", fgColor=COLOR_PRIMARY
        )

    # L total: sum avg value
    ws.cell(row=totals_row, column=12).value = (
        f'=SUMIF(L{first_data}:L{last_data},"<>")'
    )
    ws.cell(row=totals_row, column=12).number_format = '"$"#,##0'
    ws.cell(row=totals_row, column=12).font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT
    )
    ws.cell(row=totals_row, column=12).alignment = Alignment(
        horizontal="right", vertical="center"
    )

    # M total: sum loan balance
    ws.cell(row=totals_row, column=13).value = (
        f'=SUM(M{first_data}:M{last_data})'
    )
    ws.cell(row=totals_row, column=13).number_format = '"$"#,##0'
    ws.cell(row=totals_row, column=13).font = Font(
        name=FONT_BODY, size=11, bold=True, color="F6EFE2"
    )
    ws.cell(row=totals_row, column=13).alignment = Alignment(
        horizontal="right", vertical="center"
    )

    # N total: sum equity
    ws.cell(row=totals_row, column=14).value = (
        f'=SUMIF(N{first_data}:N{last_data},"<>")'
    )
    ws.cell(row=totals_row, column=14).number_format = '"$"#,##0'
    ws.cell(row=totals_row, column=14).font = Font(
        name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT
    )
    ws.cell(row=totals_row, column=14).alignment = Alignment(
        horizontal="right", vertical="center"
    )

    # O total: portfolio LTV = M_total / L_total
    ws.cell(row=totals_row, column=15).value = (
        f'=IFERROR(M{totals_row}/L{totals_row},0)'
    )
    ws.cell(row=totals_row, column=15).number_format = "0.0%"
    ws.cell(row=totals_row, column=15).font = Font(
        name=FONT_BODY, size=11, bold=True, color="F6EFE2"
    )
    ws.cell(row=totals_row, column=15).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.row_dimensions[totals_row].height = 24

    ws.freeze_panes = "B7"

    brand_footer(ws, totals_row + 3,
                 version_line=f"{SKU} · Per Property")

    ws.print_area = f"A1:O{totals_row + 5}"
    ws.print_title_rows = "1:6"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_cap_rate_reference_tab(wb, variant):
    ws = wb.create_sheet("Cap Rate Reference")
    ws.sheet_properties.tabColor = COLOR_GOLD_SOFT

    set_col_widths(ws, [
        ("A", 28), ("B", 14), ("C", 4),
        ("D", 12), ("E", 12),
    ] + [(get_column_letter(c), 8) for c in range(6, 13)])

    compact_header_band(ws, "Cap Rate Reference",
                        prev_tab="Per Property", next_tab="Equity Summary")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws["A4"] = ("Edit ranges per your actual market — Per Property tab "
                "VLOOKUPs against this table.")
    ws["A4"].font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:L4")

    # Header row 6
    for col_letter, label in [("A", "Market type"), ("B", "Cap rate"),
                                ("D", "Range low"), ("E", "Range high")]:
        cell = ws[f"{col_letter}6"]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Data rows 7..16 — capacity 10 market types
    range_hints = {
        "Mountain resort STR": (0.06, 0.08),
        "Beach STR":           (0.07, 0.09),
        "Urban STR":           (0.08, 0.10),
        "Rural / cabin":       (0.08, 0.12),
    }
    for i in range(10):
        r = 7 + i
        if i < len(SAMPLE_CAP_RATES):
            mname, mrate = SAMPLE_CAP_RATES[i]
            mlo, mhi = range_hints.get(mname, (None, None))
        else:
            mname, mrate, mlo, mhi = None, None, None, None

        # A: market name (input)
        c = ws.cell(row=r, column=1)
        c.value = mname if variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # B: cap rate (input)
        c = ws.cell(row=r, column=2)
        c.value = mrate if variant == "demo" else None
        apply_style(c, input_cell_style())
        c.number_format = "0.00%"
        c.alignment = Alignment(horizontal="center", vertical="center")

        # D-E: typical range hint (input, optional)
        c = ws.cell(row=r, column=4)
        c.value = mlo if variant == "demo" else None
        apply_style(c, input_cell_style())
        c.number_format = "0.00%"
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=r, column=5)
        c.value = mhi if variant == "demo" else None
        apply_style(c, input_cell_style())
        c.number_format = "0.00%"
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 20

    # Methodology callout row 19+
    ws["A19"] = "WHY THESE RANGES?"
    ws["A19"].font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    ws["A19"].fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws["A19"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[19].height = 20
    ws.merge_cells("A19:L19")

    ws["A20"] = (
        "STR cap rates run higher than long-term rentals because operating "
        "risk is higher and revenue is more volatile. Mountain/beach resort "
        "markets compress to 6-8% (institutional appetite); urban and rural "
        "stay in the 8-12% zone (thinner buyer pool). When in doubt, ask a "
        "local STR-savvy commercial broker for a recent comp sale and "
        "back-solve the cap from their NOI."
    )
    ws["A20"].font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    ws["A20"].fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    ws["A20"].alignment = Alignment(horizontal="left", vertical="top",
                                      wrap_text=True, indent=2)
    for r in range(20, 24):
        ws.row_dimensions[r].height = 18
    ws.merge_cells("A20:L23")

    brand_footer(ws, 26,
                 version_line=f"{SKU} · Cap Rate Reference")


def build_equity_summary_tab(wb, variant):
    ws = wb.create_sheet("Equity Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 30), ("C", 18),
        ("D", 4), ("E", 22), ("F", 16),
    ] + [(get_column_letter(c), 8) for c in range(7, 13)])

    compact_header_band(ws, "Equity Summary",
                        prev_tab="Cap Rate Reference", next_tab="Settings")

    # Subtitle row 4 (write BEFORE merging)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws["A4"] = "Roll-up of value, debt, equity, and LTV — plus per-property equity pie chart."
    ws["A4"].font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:L4")

    # ---- Roll-up labels (col B) and values (col C) ----
    _section_band(ws, 6, "PORTFOLIO ROLL-UP")

    rollup = [
        (7, "Total portfolio market value:", "='Per Property'!L17", '"$"#,##0', True),
        (8, "Total loan balances:",          "='Per Property'!M17", '"$"#,##0', False),
        (9, "Total equity:",                 "='Per Property'!N17", '"$"#,##0', True),
        (10, "Portfolio LTV:",               "='Per Property'!O17", "0.0%",     False),
    ]
    for row, label, formula, fmt, emphasize in rollup:
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = Font(
            name=FONT_HEAD if emphasize else FONT_BODY,
            size=12 if emphasize else 11,
            bold=True,
            color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )

        c = ws.cell(row=row, column=3, value=formula)
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        if emphasize:
            c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

        ws.row_dimensions[row].height = 22 if emphasize else 18

    # ---- Per-property equity pie chart data block (cols E-F) ----
    # Row 6 is occupied by the section band (merged A6:L6) — put headers on row 12 instead,
    # below the 4-row roll-up that ends at row 10.
    pie_header_row = 12
    ws.cell(row=pie_header_row, column=5, value="Property")
    apply_style(ws.cell(row=pie_header_row, column=5), header_row_style())
    ws.cell(row=pie_header_row, column=6, value="Equity")
    apply_style(ws.cell(row=pie_header_row, column=6), header_row_style())
    ws.row_dimensions[pie_header_row].height = 24

    first_data = pie_header_row + 1
    last_data = first_data + PROPERTY_CAP - 1

    for i in range(PROPERTY_CAP):
        r = first_data + i
        # Property name from Settings (E col)
        c = ws.cell(row=r, column=5)
        c.value = f"=Settings!B{7 + i}"
        apply_style(c, formula_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Equity pulled from Per Property!N (F col)
        c = ws.cell(row=r, column=6)
        c.value = (
            f"=IFERROR(IF('Per Property'!N{r}=\"\",0,'Per Property'!N{r}),0)"
        )
        apply_style(c, formula_cell_style())
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 18

    # ---- Pie chart: Equity by Property ----
    pie = PieChart()
    pie.title = "Equity by Property"
    pie.height = 9
    pie.width = 13
    pie_data = Reference(ws, min_col=6, min_row=6, max_row=last_data, max_col=6)
    pie_cats = Reference(ws, min_col=5, min_row=first_data, max_row=last_data)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList(showCatName=True, showPercent=True)
    style_chart(pie)
    ws.add_chart(pie, "H6")

    # Methodology footnote
    ws["A20"] = "VALUATION METHODS"
    ws["A20"].font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    ws["A20"].fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws["A20"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[20].height = 20
    ws.merge_cells("A20:L20")

    ws["A21"] = (
        "Method 1 — Income (Cap × NOI): NOI ÷ market cap rate. The number a "
        "buyer with a calculator runs.   Method 2 — Sales comparison "
        "($/BR × beds, half-bath premium): the number a residential agent runs.   "
        "Method 3 — Gross Rent Multiplier (rent × GRM): the back-of-the-envelope "
        "an experienced STR investor runs. The three averaged is your defensible "
        "valuation — when one method is wildly off, dig in before quoting it."
    )
    ws["A21"].font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    ws["A21"].fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    ws["A21"].alignment = Alignment(horizontal="left", vertical="top",
                                      wrap_text=True, indent=2)
    for r in range(21, 25):
        ws.row_dimensions[r].height = 18
    ws.merge_cells("A21:L24")

    brand_footer(ws, 27,
                 version_line=f"{SKU} · Equity Summary")

    ws.print_area = "A1:L29"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 26), ("C", 26),
        ("D", 4), ("E", 38),
    ] + [(get_column_letter(c), 8) for c in range(6, 13)])

    compact_header_band(ws, "Settings",
                        prev_tab="Equity Summary", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws["A4"] = "Property roster, valuation date, and market types — single source of truth."
    ws["A4"].font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:L4")

    # B5 active year
    ws["B5"].value = "Valuation date:"
    ws["B5"].font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    ws["B5"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws["C5"]
    c.value = ACTIVE_YEAR if variant == "demo" else None
    apply_style(c, input_cell_style())
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0"
    ws.row_dimensions[5].height = 22

    # Header row 6 for property roster
    for col, label in [("B", "Property name"), ("C", "Market type")]:
        cell = ws[f"{col}6"]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 22

    # Property roster B7:B16 + market type C7:C16
    for i in range(PROPERTY_CAP):
        r = 7 + i
        sample = SAMPLE_PROPS[i] if i < len(SAMPLE_PROPS) else None

        # B: name
        c = ws.cell(row=r, column=2)
        c.value = sample[0] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # C: market type (dropdown)
        c = ws.cell(row=r, column=3)
        c.value = sample[1] if sample and variant == "demo" else None
        apply_style(c, input_cell_style())
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[r].height = 20

    # Dropdown for market type column C7:C16 — references Cap Rate Reference!A7:A16
    add_dropdown(ws, "C7:C16", "='Cap Rate Reference'!$A$7:$A$16")

    # Methodology / settings notes
    ws["A19"] = "HOW TO USE"
    ws["A19"].font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    ws["A19"].fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws["A19"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[19].height = 20
    ws.merge_cells("A19:L19")

    ws["A20"] = (
        "1) Set valuation year above.   2) List each property and pick its "
        "market type from the dropdown (drives cap rate via Cap Rate Reference).   "
        "3) Customize the Cap Rate Reference table for your actual sub-market — "
        "Gatlinburg's cap is not Aspen's.   4) Fill NOI, comps, rent, GRM, and "
        "loan balance on the Per Property tab.   5) Read the answer on Start "
        "and Equity Summary."
    )
    ws["A20"].font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    ws["A20"].fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    ws["A20"].alignment = Alignment(horizontal="left", vertical="top",
                                      wrap_text=True, indent=2)
    for r in range(20, 25):
        ws.row_dimensions[r].height = 18
    ws.merge_cells("A20:L24")

    brand_footer(ws, 27,
                 version_line=f"{SKU} · Settings")


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    # Build dependent tabs in order — Cap Rate Reference before Per Property
    # would normally matter for VLOOKUPs at runtime, but openpyxl doesn't
    # evaluate; Excel resolves on open regardless of sheet order.
    build_per_property_tab(wb, variant)
    build_cap_rate_reference_tab(wb, variant)
    build_equity_summary_tab(wb, variant)
    build_settings_tab(wb, variant)

    wb.properties.title = "Portfolio Valuation Model — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Portfolio market-value workbook using 3 valuation methods averaged "
        "(cap rate x NOI, comparable sales per BR, GRM x rent), with equity, "
        "LTV, and per-property equity pie chart."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
