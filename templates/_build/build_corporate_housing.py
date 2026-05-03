"""Build SPC-002 Corporate Housing / Travel-Nurse Tracker (v2.3 standard).

Operational-mode register + matrix for mid-term rental hosts (30+ day
stays) on Furnished Finder, Blueground, AnyPlace, or directly with
corporate clients / travel nurses.

Different model than nightly STR:
- Monthly rent (not nightly) — MRR matters more than ADR
- Longer stays (30 days to 12 months) — gap days are the killer metric
- Recovery deposits, utility reimbursement options
- Pipeline-driven (inquiry → application → approved → signed)

Forks structure from build_pl_single_property.py (matrix tab) and
build_1099_nec_tracker.py (register + dashboard pattern).

Generates:
- SPC-002-corporate-housing-travel-nurse-tracker-DEMO.xlsx
- SPC-002-corporate-housing-travel-nurse-tracker-BLANK.xlsx
"""
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, add_upgrade_banner, apply_style,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    COLOR_FORMULA_TINT, STATE_GOOD_FILL, STATE_WARN_FILL,
    STATE_BAD_FILL,
    STATE_INFO_FILL,
)

SKU = "SPC-002"
NAME = "corporate-housing-travel-nurse-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

ACTIVE_YEAR = 2026

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

PROFESSIONS = [
    "Travel nurse", "Corporate exec", "Locum doctor",
    "Insurance adjuster", "Construction", "Film crew", "Other",
]
CHANNELS = ["Furnished Finder", "Blueground", "AnyPlace",
            "Direct", "Referral", "Other"]
LEASE_TYPES = ["3-month", "6-month", "Month-to-month",
               "Corporate negotiated", "Travel nurse contract", "Other"]
STATUSES = ["Lead", "Application", "Approved",
            "Active", "Move-out", "Past"]

# Demo property list — 3 properties (2 Furnished Finder, 1 corporate-direct)
DEMO_PROPERTIES = [
    "Smokies Ridge — Unit A",
    "Smokies Ridge — Unit B",
    "Knoxville Corporate Loft",
]

# 8 demo tenants — Jan-Mar 2026 dominant; 4 active, 2 past, 2 future-signed
# Columns: tenant_id, name, email, phone, employer, profession,
#          property, channel, lease_start, lease_end, monthly_rent,
#          deposit, utilities_incl, status, notes
DEMO_TENANTS = [
    ("CH-001", "Sarah Williams", "swilliams@nursegrp.com", "(865) 555-0110",
     "MedStaff Travel Nurses", "Travel nurse",
     "Smokies Ridge — Unit A", "Furnished Finder",
     "2026-01-05", "2026-04-04", 3200, 500, "Y",
     "Active", "13-week ICU contract — UTMC Knoxville"),
    ("CH-002", "James Patel", "jpatel@boldco.com", "(865) 555-0118",
     "BoldCo Engineering", "Corporate exec",
     "Knoxville Corporate Loft", "Direct",
     "2026-01-15", "2026-07-14", 3400, 1000, "N",
     "Active", "6-month relocation; co invoiced direct"),
    ("CH-003", "Maria Lopez", "mlopez@travelmed.io", "(865) 555-0142",
     "TravelMed Agency", "Travel nurse",
     "Smokies Ridge — Unit B", "Furnished Finder",
     "2026-02-01", "2026-04-30", 2800, 500, "Y",
     "Active", "13-week ER contract"),
    ("CH-004", "Daniel O'Connor", "doconnor@locumhealth.com", "(865) 555-0156",
     "LocumHealth Physicians", "Locum doctor",
     "Smokies Ridge — Unit A", "Furnished Finder",
     "2026-04-10", "2026-07-09", 3300, 500, "Y",
     "Approved", "Signed lease; arrives after Sarah departs"),
    ("CH-005", "Rebecca Chen", "rchen@allstate.com", "(865) 555-0163",
     "Allstate Claims", "Insurance adjuster",
     "Smokies Ridge — Unit B", "Referral",
     "2026-05-15", "2026-08-14", 2950, 500, "Y",
     "Approved", "Storm-deployment claims; future-signed"),
    ("CH-006", "Tom Henderson", "thenderson@hendersonbuilds.com", "(615) 555-0178",
     "Henderson Construction", "Construction",
     "Knoxville Corporate Loft", "Direct",
     "2025-09-01", "2025-12-31", 3200, 1000, "N",
     "Past", "I-40 bridge project — completed"),
    ("CH-007", "Lena Brooks", "lbrooks@filmworks.tv", "(310) 555-0185",
     "FilmWorks Productions", "Film crew",
     "Smokies Ridge — Unit A", "Furnished Finder",
     "2025-10-01", "2025-12-31", 3100, 500, "Y",
     "Past", "Documentary shoot"),
    ("CH-008", "Kevin Park", "kpark@nursegrp.com", "(865) 555-0192",
     "MedStaff Travel Nurses", "Travel nurse",
     "Smokies Ridge — Unit B", "Furnished Finder",
     "2026-05-15", "2026-08-14", 2800, 500, "Y",
     "Lead", "Inquiry only — application pending"),
]

# Pipeline funnel — by channel
# (channel, inquiries, applications, approved, signed)
DEMO_PIPELINE = [
    ("Furnished Finder", 20, 12, 8,  7),
    ("Blueground",        4,  2, 1,  1),
    ("Direct",            5,  4, 3,  2),
    ("Referral",          3,  2, 2,  1),
    ("Other",             2,  1, 0,  0),
]


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 0 — Start (KPI hero + nav cards)."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Navy hero band (rows 1-8) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Corporate Housing & Travel-Nurse Tracker"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Mid-term hosting is a different game — gap days are the killer metric."
    c.font = Font(name=FONT_HEAD, size=13, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: KPI cards (rows 10-17) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 18):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # 4 KPIs across 12 cols → 3 cols each
    # Card 1: ACTIVE TENANTS (A-C)
    # Card 2: MRR (D-F)
    # Card 3: OCCUPANCY % (G-I)
    # Card 4: GAP DAYS (J-L)
    kpi_cards = [
        ("A", "C", "ACTIVE TENANTS", COLOR_PRIMARY,
         '=COUNTIF(Tenants!N6:N105,"Active")', "0",
         "currently in residence"),
        ("D", "F", "MRR (PORTFOLIO)", COLOR_ACCENT,
         '=SUMIFS(Tenants!K6:K105,Tenants!N6:N105,"Active")', '"$"#,##0',
         "monthly recurring revenue"),
        ("G", "I", "TENANTS YTD", COLOR_NAVY_TINT,
         '=COUNTA(Tenants!A6:A105)', "0",
         "all leases logged"),
        ("J", "L", "FUTURE-SIGNED", COLOR_SECONDARY,
         '=COUNTIF(Tenants!N6:N105,"Approved")', "0",
         "approved but not yet active"),
    ]

    for first, last, label, color, formula, fmt, sub in kpi_cards:
        ws.merge_cells(f"{first}10:{last}10")
        c = ws[f"{first}10"]
        c.value = label
        c.font = Font(name=FONT_MONO, size=9, bold=True, color=color)
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{first}11:{last}13")
        c = ws[f"{first}11"]
        c.value = formula
        c.font = Font(name=FONT_HEAD, size=28, bold=True, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        ws.merge_cells(f"{first}14:{last}14")
        c = ws[f"{first}14"]
        c.value = sub
        c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last, *_ in kpi_cards:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 15):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 14 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    ws.row_dimensions[15].height = 8

    # --- ZONE 3: WHAT THIS DOES card (rows 17-22) ---
    for r in range(17, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 20

    ws.merge_cells("A18:L22")
    c = ws["A18"]
    c.value = (
        "Mid-term hosts (30-day+ stays on Furnished Finder, Blueground, or "
        "direct with corporate clients/travel nurses) need to track tenant "
        "pipeline, monthly recurring revenue, and the gap days that quietly "
        "eat margin between leases. This workbook is a CRM + revenue matrix "
        "+ pipeline funnel — purpose-built for the mid-term model."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # --- ZONE 4: Primary nav button (rows 24-27) ---
    pseudo_button(ws, "A24", "L27",
                   "→  ADD A TENANT  (OPEN TENANTS)",
                   "'Tenants'!A6", variant="primary")
    for r in range(24, 28):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: Secondary nav (rows 29-30) ---
    pseudo_button(ws, "A29", "C30", "Lease Calendar",
                   "'Lease Calendar'!A1", variant="secondary")
    pseudo_button(ws, "D29", "F30", "Monthly Revenue",
                   "'Monthly Revenue'!A1", variant="secondary")
    pseudo_button(ws, "G29", "I30", "Pipeline",
                   "'Pipeline'!A1", variant="accent")
    pseudo_button(ws, "J29", "L30", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[29].height = 22
    ws.row_dimensions[30].height = 22

    # --- ZONE 6: Mid-term hosting callout (rows 32-33) ---
    ws.merge_cells("A32:L32")
    c = ws["A32"]
    c.value = (
        "💡 MID-TERM RULE OF THUMB: a 1-week gap on a $3,000/mo unit costs "
        "$700 in lost rent. Future-signed leases beat backfilling. Watch the "
        "Pipeline tab — your inquiry-to-signed conversion is your true "
        "occupancy lever."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[32].height = 36

    add_upgrade_banner(ws, 35)
    brand_footer(ws, 37,
                 version_line=f"{SKU} · v2.3 · Free updates forever")

    ws.print_area = "A1:L40"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_tenants_tab(wb, variant):
    """Sheet 1 — Tenants register (CRM, capacity 100)."""
    ws = wb.create_sheet("Tenants")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Tenants",
                         prev_tab="Start", next_tab="Lease Calendar")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per tenant — past, current, future. Pipeline (Lead → Past) tracked in Status."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 10), ("B", 22), ("C", 26), ("D", 14), ("E", 22),
        ("F", 16), ("G", 26), ("H", 18), ("I", 12), ("J", 12),
        ("K", 11), ("L", 12), ("M", 9),  ("N", 13), ("O", 32),
    ])

    headers = [
        "Tenant ID", "Name", "Email", "Phone", "Employer / Agency",
        "Profession", "Property Assigned", "Channel",
        "Lease Start", "Lease End", "Monthly Rent", "Deposit",
        "Util Incl?", "Status", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    tenant_rows = DEMO_TENANTS if variant == "demo" else []
    for i, t in enumerate(tenant_rows, start=6):
        (tid, name, email, phone, employer, prof, prop, channel,
         ls, le, rent, deposit, util, status, notes) = t

        a = ws.cell(row=i, column=1, value=tid)
        apply_style(a, input_cell_style())

        b = ws.cell(row=i, column=2, value=name)
        apply_style(b, input_cell_style())

        c = ws.cell(row=i, column=3, value=email)
        apply_style(c, input_cell_style())

        d = ws.cell(row=i, column=4, value=phone)
        apply_style(d, input_cell_style())

        e = ws.cell(row=i, column=5, value=employer)
        apply_style(e, input_cell_style())

        f = ws.cell(row=i, column=6, value=prof)
        apply_style(f, input_cell_style())

        g = ws.cell(row=i, column=7, value=prop)
        apply_style(g, input_cell_style())

        h = ws.cell(row=i, column=8, value=channel)
        apply_style(h, input_cell_style())

        ic = ws.cell(row=i, column=9, value=_parse_date(ls))
        apply_style(ic, input_cell_style())
        ic.number_format = "yyyy-mm-dd"

        jc = ws.cell(row=i, column=10, value=_parse_date(le))
        apply_style(jc, input_cell_style())
        jc.number_format = "yyyy-mm-dd"

        kc = ws.cell(row=i, column=11, value=rent)
        apply_style(kc, input_cell_style())
        kc.number_format = '"$"#,##0'

        lc = ws.cell(row=i, column=12, value=deposit)
        apply_style(lc, input_cell_style())
        lc.number_format = '"$"#,##0'

        mc = ws.cell(row=i, column=13, value=util)
        apply_style(mc, input_cell_style())
        mc.alignment = Alignment(horizontal="center", vertical="center")

        nc = ws.cell(row=i, column=14, value=status)
        apply_style(nc, input_cell_style())
        nc.alignment = Alignment(horizontal="center", vertical="center")

        oc = ws.cell(row=i, column=15, value=notes if notes else None)
        apply_style(oc, input_cell_style())

        ws.row_dimensions[i].height = 16

    # Capacity rows up to row 105 (100 tenants total)
    last_data_row = len(tenant_rows) + 5
    for row_idx in range(last_data_row + 1, 106):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, input_cell_style())
            if col_idx in (9, 10):
                cell.number_format = "yyyy-mm-dd"
            if col_idx in (11, 12):
                cell.number_format = '"$"#,##0'
            if col_idx in (13, 14):
                cell.alignment = Alignment(horizontal="center",
                                            vertical="center")
        ws.row_dimensions[row_idx].height = 16

    add_dropdown(ws, "F6:F105", '"' + ",".join(PROFESSIONS) + '"')
    add_dropdown(ws, "G6:G105", "=Settings!$B$7:$B$16")
    add_dropdown(ws, "H6:H105", '"' + ",".join(CHANNELS) + '"')
    add_dropdown(ws, "M6:M105", '"Y,N"')
    add_dropdown(ws, "N6:N105", '"' + ",".join(STATUSES) + '"')

    # Conditional formatting on Status
    status_colors = [
        ("Active",      STATE_GOOD_FILL),
        ("Approved",    STATE_WARN_FILL),
        ("Lead",        STATE_INFO_FILL),
        ("Application", STATE_WARN_FILL),
        ("Move-out",    STATE_BAD_FILL),
        ("Past",        COLOR_FORMULA_TINT),
    ]
    for status, color in status_colors:
        ws.conditional_formatting.add(
            "N6:N105",
            FormulaRule(
                formula=[f'N6="{status}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_lease_calendar_tab(wb, variant):
    """Sheet 2 — Lease Calendar (12-month Gantt grid: properties × months)."""
    ws = wb.create_sheet("Lease Calendar")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Lease Calendar",
                         prev_tab="Tenants", next_tab="Monthly Revenue")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "12-month Gantt grid — rows = properties, columns = months. Cells colored by tenant status."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Column widths: A = property name, B-M = months
    set_col_widths(ws, [("A", 32)] + [(get_column_letter(2 + i), 14)
                                      for i in range(12)] + [("N", 12)])

    # Header row 5: Property | Jan | Feb | ... | Dec | YTD Active
    header = ["Property"] + MONTHS + ["Active Months"]
    for col, h in enumerate(header, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # Property rows — pull from Settings B7:B16 (10-property capacity)
    # Each cell shows tenant name during lease period (concatenated if overlap).
    # Formula matches month start to tenant lease range:
    #   Lookup: any tenant whose lease overlaps month → return name
    # Uses TEXTJOIN with array IF over Tenants!I6:J105.
    # Empty cells = gap days (highlight red via conditional formatting).
    for prop_idx in range(10):
        row = 6 + prop_idx
        ws.row_dimensions[row].height = 22

        # Column A — property name from Settings
        a = ws.cell(row=row, column=1,
                    value=f"=IF(Settings!$B${7 + prop_idx}=\"\",\"\","
                          f"Settings!$B${7 + prop_idx})")
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        # Columns B-M — month cells
        for m in range(1, 13):
            col = m + 1
            # Month start: DATE(Settings!B5, m, 1)
            # Month end:   DATE(Settings!B5, m+1, 1) - 1
            # A tenant occupies this month if:
            #   tenant_property = property name AND
            #   lease_start <= month_end AND lease_end >= month_start
            # Return first matching tenant name (TEXTJOIN handles overlap)
            formula = (
                f'=IFERROR(TEXTJOIN(" / ",TRUE,'
                f'IF((Tenants!$G$6:$G$105=$A{row})*'
                f'(Tenants!$I$6:$I$105<=DATE(Settings!$B$5,{m + 1 if m < 12 else 13},'
                f'{1 if m < 12 else 1})-{1 if m < 12 else 1})*'
                f'(Tenants!$J$6:$J$105>=DATE(Settings!$B$5,{m},1)),'
                f'Tenants!$B$6:$B$105,"")),"")'
            )
            # Simpler robust version: use SUMPRODUCT + INDEX for first match
            formula = (
                f'=IF($A{row}="","",'
                f'IFERROR(INDEX(Tenants!$B$6:$B$105,'
                f'MATCH(1,'
                f'(Tenants!$G$6:$G$105=$A{row})*'
                f'(Tenants!$I$6:$I$105<=DATE(Settings!$B$5,{m},EOMONTH(DATE(Settings!$B$5,{m},1),0)-DATE(Settings!$B$5,{m},1)+1))*'
                f'(Tenants!$J$6:$J$105>=DATE(Settings!$B$5,{m},1)),0)),""))'
            )
            # Cleanest correct version using boolean array MATCH:
            #   end of month m = DATE(year, m+1, 1) - 1
            if m < 12:
                month_end_expr = f"DATE(Settings!$B$5,{m + 1},1)-1"
            else:
                month_end_expr = f"DATE(Settings!$B$5+1,1,1)-1"
            month_start_expr = f"DATE(Settings!$B$5,{m},1)"
            formula = (
                f'=IF($A{row}="","",'
                f'IFERROR(INDEX(Tenants!$B$6:$B$105,'
                f'MATCH(1,'
                f'(Tenants!$G$6:$G$105=$A{row})*'
                f'(Tenants!$I$6:$I$105<={month_end_expr})*'
                f'(Tenants!$J$6:$J$105>={month_start_expr}),0)),""))'
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.font = Font(name=FONT_BODY, size=9, color=COLOR_TEXT)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                         wrap_text=True)
            cell.number_format = "@"

        # Column N — count of months occupied
        ytd = ws.cell(row=row, column=14,
                      value=f'=COUNTIF(B{row}:M{row},"?*")')
        apply_style(ytd, formula_cell_style())
        ytd.font = Font(name=FONT_BODY, size=11, bold=True, italic=True,
                         color=COLOR_PRIMARY)
        ytd.alignment = Alignment(horizontal="center", vertical="center")

    # Conditional formatting:
    # - Cell with text (occupied) → green tint
    # - Empty cell with property name in col A → red tint (gap day)
    grid_range = "B6:M15"
    # Occupied cells (any text)
    ws.conditional_formatting.add(
        grid_range,
        FormulaRule(
            formula=['LEN(B6)>0'],
            fill=PatternFill("solid", fgColor=STATE_GOOD_FILL),
            font=Font(name=FONT_BODY, size=9, bold=True, color=COLOR_PRIMARY),
        ),
    )
    # Empty cells where property exists → gap (red)
    ws.conditional_formatting.add(
        grid_range,
        FormulaRule(
            formula=['AND(LEN(B6)=0,LEN($A6)>0)'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
        ),
    )

    # Legend rows below the grid
    legend_row = 17
    ws.row_dimensions[legend_row].height = 8
    ws.merge_cells(f"A{legend_row + 1}:M{legend_row + 1}")
    c = ws.cell(row=legend_row + 1, column=1)
    c.value = (
        "Legend: GREEN cell = occupied month (tenant name shown)   ·   "
        "RED cell = gap month (no tenant covers this month — backfill or "
        "raise on Pipeline tab)"
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    ws.row_dimensions[legend_row + 1].height = 24

    # Active-year footnote
    ws.merge_cells(f"A{legend_row + 3}:M{legend_row + 3}")
    c = ws.cell(row=legend_row + 3, column=1)
    c.value = (
        '="Showing year "&Settings!$B$5&". Change Active Year on Settings to '
        'switch the calendar to a different year."'
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "B6"

    ws.print_area = f"A1:N{legend_row + 4}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_monthly_revenue_tab(wb):
    """Sheet 3 — Monthly Revenue (12-col MRR matrix per property + chart)."""
    ws = wb.create_sheet("Monthly Revenue")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Monthly Revenue",
                         prev_tab="Lease Calendar", next_tab="Pipeline")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Per-property monthly rent collected — bump Active Year on Settings to switch."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32)] + [(get_column_letter(2 + i), 11)
                                      for i in range(12)] + [("N", 13)])

    header = ["Property"] + MONTHS + ["YTD"]
    for col, h in enumerate(header, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    # 10 property rows
    for prop_idx in range(10):
        row = 6 + prop_idx
        ws.row_dimensions[row].height = 18

        a = ws.cell(row=row, column=1,
                    value=f"=IF(Settings!$B${7 + prop_idx}=\"\",\"\","
                          f"Settings!$B${7 + prop_idx})")
        apply_style(a, formula_cell_style())
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)

        for m in range(1, 13):
            col = m + 1
            if m < 12:
                month_end = f"DATE(Settings!$B$5,{m + 1},1)-1"
            else:
                month_end = f"DATE(Settings!$B$5+1,1,1)-1"
            month_start = f"DATE(Settings!$B$5,{m},1)"
            # Sum monthly_rent (col K) for any tenant whose lease overlaps
            # this month and whose property = this row's property.
            formula = (
                f'=IF($A{row}="","",'
                f'SUMPRODUCT('
                f'(Tenants!$G$6:$G$105=$A{row})*'
                f'(Tenants!$I$6:$I$105<={month_end})*'
                f'(Tenants!$J$6:$J$105>={month_start})*'
                f'Tenants!$K$6:$K$105))'
            )
            cell = ws.cell(row=row, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.number_format = '"$"#,##0'

        ytd = ws.cell(row=row, column=14,
                      value=f'=IF($A{row}="","",SUM(B{row}:M{row}))')
        apply_style(ytd, formula_cell_style())
        ytd.number_format = '"$"#,##0'
        ytd.font = Font(name=FONT_BODY, size=11, bold=True, italic=True,
                         color=COLOR_PRIMARY)

    # Portfolio total row
    tot_row = 17
    ws.row_dimensions[16].height = 8
    a = ws.cell(row=tot_row, column=1, value="PORTFOLIO TOTAL")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tot_row].height = 22

    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                        value=f"=SUM({col_letter}6:{col_letter}15)")
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0'
        cell.font = Font(name=FONT_BODY, size=12, bold=True, italic=True,
                          color=COLOR_PRIMARY)

    # --- Bar chart: MRR by month (anchored row 19) ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Portfolio MRR by Month"
    chart.y_axis.title = "Monthly Rent ($)"
    chart.x_axis.title = "Month"
    chart.height = 9
    chart.width = 18
    chart.legend = None
    data = Reference(ws, min_col=2, max_col=13,
                      min_row=tot_row, max_row=tot_row)
    cats = Reference(ws, min_col=2, max_col=13, min_row=5, max_row=5)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    style_chart(chart)
    ws.add_chart(chart, "A19")

    ws.freeze_panes = "B6"

    ws.print_area = f"A1:N{tot_row + 1}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_pipeline_tab(wb, variant):
    """Sheet 4 — Pipeline funnel (inquiry → application → approved → signed)."""
    ws = wb.create_sheet("Pipeline")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = analytics

    compact_header_band(ws, "Pipeline",
                         prev_tab="Monthly Revenue", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Inquiry-to-signed conversion funnel by channel — find your strongest acquisition lever."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 22), ("B", 13), ("C", 14), ("D", 13), ("E", 13),
        ("F", 14), ("G", 14), ("H", 14), ("I", 14),
    ])

    headers = [
        "Channel", "Inquiries", "Applications", "Approved", "Signed",
        "Inq → App %", "App → Apv %", "Apv → Sign %", "Inq → Sign %",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    pipeline_rows = DEMO_PIPELINE if variant == "demo" else [
        (ch, None, None, None, None) for ch in CHANNELS
    ]

    for i, p in enumerate(pipeline_rows, start=6):
        channel, inq, app, apv, sign = p
        ws.row_dimensions[i].height = 18

        a = ws.cell(row=i, column=1, value=channel)
        apply_style(a, input_cell_style())
        a.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)

        b = ws.cell(row=i, column=2, value=inq)
        apply_style(b, input_cell_style())
        b.number_format = "0"
        b.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=i, column=3, value=app)
        apply_style(c, input_cell_style())
        c.number_format = "0"
        c.alignment = Alignment(horizontal="center", vertical="center")

        d = ws.cell(row=i, column=4, value=apv)
        apply_style(d, input_cell_style())
        d.number_format = "0"
        d.alignment = Alignment(horizontal="center", vertical="center")

        e = ws.cell(row=i, column=5, value=sign)
        apply_style(e, input_cell_style())
        e.number_format = "0"
        e.alignment = Alignment(horizontal="center", vertical="center")

        # Conversion %
        f = ws.cell(row=i, column=6, value=f'=IFERROR(C{i}/B{i},0)')
        apply_style(f, formula_cell_style())
        f.number_format = "0%"
        f.alignment = Alignment(horizontal="center", vertical="center")

        g = ws.cell(row=i, column=7, value=f'=IFERROR(D{i}/C{i},0)')
        apply_style(g, formula_cell_style())
        g.number_format = "0%"
        g.alignment = Alignment(horizontal="center", vertical="center")

        h = ws.cell(row=i, column=8, value=f'=IFERROR(E{i}/D{i},0)')
        apply_style(h, formula_cell_style())
        h.number_format = "0%"
        h.alignment = Alignment(horizontal="center", vertical="center")

        i_cell = ws.cell(row=i, column=9, value=f'=IFERROR(E{i}/B{i},0)')
        apply_style(i_cell, formula_cell_style())
        i_cell.number_format = "0%"
        i_cell.alignment = Alignment(horizontal="center", vertical="center")
        i_cell.font = Font(name=FONT_BODY, size=11, bold=True,
                            color=COLOR_PRIMARY)

    # Total row (row 12)
    tot_row = 12
    ws.row_dimensions[11].height = 8
    a = ws.cell(row=tot_row, column=1, value="TOTAL")
    a.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[tot_row].height = 22

    for col_idx, col_letter in enumerate(["B", "C", "D", "E"], start=2):
        cell = ws.cell(row=tot_row, column=col_idx,
                        value=f"=SUM({col_letter}6:{col_letter}10)")
        apply_style(cell, formula_cell_style())
        cell.number_format = "0"
        cell.font = Font(name=FONT_BODY, size=12, bold=True, italic=True,
                          color=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    f = ws.cell(row=tot_row, column=6,
                 value=f'=IFERROR(C{tot_row}/B{tot_row},0)')
    apply_style(f, formula_cell_style())
    f.number_format = "0%"
    f.alignment = Alignment(horizontal="center", vertical="center")
    f.font = Font(name=FONT_BODY, size=12, bold=True, italic=True,
                   color=COLOR_PRIMARY)

    g = ws.cell(row=tot_row, column=7,
                 value=f'=IFERROR(D{tot_row}/C{tot_row},0)')
    apply_style(g, formula_cell_style())
    g.number_format = "0%"
    g.alignment = Alignment(horizontal="center", vertical="center")
    g.font = Font(name=FONT_BODY, size=12, bold=True, italic=True,
                   color=COLOR_PRIMARY)

    h = ws.cell(row=tot_row, column=8,
                 value=f'=IFERROR(E{tot_row}/D{tot_row},0)')
    apply_style(h, formula_cell_style())
    h.number_format = "0%"
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.font = Font(name=FONT_BODY, size=12, bold=True, italic=True,
                   color=COLOR_PRIMARY)

    i_cell = ws.cell(row=tot_row, column=9,
                      value=f'=IFERROR(E{tot_row}/B{tot_row},0)')
    apply_style(i_cell, formula_cell_style())
    i_cell.number_format = "0%"
    i_cell.alignment = Alignment(horizontal="center", vertical="center")
    i_cell.font = Font(name=FONT_BODY, size=12, bold=True, italic=True,
                        color=COLOR_PRIMARY)

    # --- Funnel chart (BarChart, horizontal totals showing stages) ---
    # Build a small staging block at A14:B17 that the chart can consume.
    ws.row_dimensions[13].height = 8

    stage_header = ws.cell(row=14, column=1, value="Funnel Stage (totals)")
    stage_header.font = Font(name=FONT_HEAD, size=12, bold=True,
                               color=COLOR_PRIMARY)
    ws.row_dimensions[14].height = 22

    stages = [
        ("Inquiries",     f'=B{tot_row}'),
        ("Applications",  f'=C{tot_row}'),
        ("Approved",      f'=D{tot_row}'),
        ("Signed",        f'=E{tot_row}'),
    ]
    for idx, (label, formula) in enumerate(stages, start=15):
        a = ws.cell(row=idx, column=1, value=label)
        a.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)

        b = ws.cell(row=idx, column=2, value=formula)
        apply_style(b, formula_cell_style())
        b.number_format = "0"
        b.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[idx].height = 18

    funnel_chart = BarChart()
    funnel_chart.type = "bar"  # horizontal bars = funnel
    funnel_chart.title = "Pipeline Funnel (All Channels)"
    funnel_chart.height = 8
    funnel_chart.width = 14
    funnel_chart.legend = None
    fdata = Reference(ws, min_col=2, max_col=2, min_row=15, max_row=18)
    fcats = Reference(ws, min_col=1, max_col=1, min_row=15, max_row=18)
    funnel_chart.add_data(fdata, titles_from_data=False)
    funnel_chart.set_categories(fcats)
    funnel_chart.dataLabels = DataLabelList(showVal=True)
    style_chart(funnel_chart)
    ws.add_chart(funnel_chart, "D14")

    # Conditional format on Inq → Sign % (col I) — green if >= 25%, red < 10%
    ws.conditional_formatting.add(
        "I6:I10",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.25"],
                    fill=PatternFill("solid", fgColor=STATE_GOOD_FILL)),
    )
    ws.conditional_formatting.add(
        "I6:I10",
        CellIsRule(operator="lessThan", formula=["0.10"],
                    fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )

    ws.print_area = "A1:I22"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 5 — Settings (active year, properties, channels, lease types)."""
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    compact_header_band(ws, "Settings",
                         prev_tab="Pipeline", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Active year · property list · channel list · lease type list"
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [("A", 32), ("B", 32), ("C", 6), ("D", 16)])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True,
                          color=COLOR_MUTED)

    # --- Active year (row 5) — drives Lease Calendar + Monthly Revenue ---
    a5 = ws.cell(row=5, column=1, value="Active year:")
    a5.font = bold_right
    a5.alignment = Alignment(horizontal="right", vertical="center")
    b5 = ws.cell(row=5, column=2, value=ACTIVE_YEAR)
    apply_style(b5, input_cell_style())
    b5.number_format = "0"
    ws.row_dimensions[5].height = 18

    ws.row_dimensions[6].height = 6

    # --- Property list (rows 7-16) ---
    sect_label = ws.cell(row=7, column=1, value="Property list (10 max)")
    sect_label.font = Font(name=FONT_HEAD, size=12, bold=True,
                             color=COLOR_PRIMARY)
    sect_label.alignment = Alignment(horizontal="right", vertical="center")

    demo_props = DEMO_PROPERTIES if variant == "demo" else []
    for prop_idx in range(10):
        row = 7 + prop_idx
        cell = ws.cell(row=row, column=2)
        if prop_idx < len(demo_props):
            cell.value = demo_props[prop_idx]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[17].height = 6

    # --- Channel list (rows 18-25) ---
    sect2 = ws.cell(row=18, column=1, value="Channel list")
    sect2.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    sect2.alignment = Alignment(horizontal="right", vertical="center")

    for idx, ch in enumerate(CHANNELS):
        row = 18 + idx
        cell = ws.cell(row=row, column=2, value=ch)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[26].height = 6

    # --- Lease type list (rows 27-34) ---
    sect3 = ws.cell(row=27, column=1, value="Lease type list")
    sect3.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    sect3.alignment = Alignment(horizontal="right", vertical="center")

    for idx, lt in enumerate(LEASE_TYPES):
        row = 27 + idx
        cell = ws.cell(row=row, column=2, value=lt)
        cell.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[35].height = 8

    # --- Year-end archive ---
    sect4 = ws.cell(row=36, column=1, value="Year-end archive")
    sect4.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    sect4.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[36].height = 22

    ws.merge_cells("A37:D37")
    c37 = ws["A37"]
    c37.value = (
        "Each January, log YTD totals for the closing year, then bump "
        "Active year above and clear past tenants from the Tenants tab "
        "(or archive separately)."
    )
    c37.font = italic_muted
    c37.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[37].height = 28

    archive_headers = ["Year", "Tenants Hosted", "Total MRR Collected",
                        "Gap Days"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=38, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[38].height = 18

    for idx, year in enumerate(range(2024, 2031), start=39):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            cell.number_format = "0" if col != 3 else '"$"#,##0'
        ws.row_dimensions[idx].height = 16


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_tenants_tab(wb, variant)
    build_lease_calendar_tab(wb, variant)
    build_monthly_revenue_tab(wb)
    build_pipeline_tab(wb, variant)
    build_settings_tab(wb, variant)

    suffix = f" ({variant.upper()})"
    wb.properties.title = (
        f"Corporate Housing & Travel-Nurse Tracker{suffix} — The STR Ledger"
    )
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Mid-term rental (30+ day) tenant CRM, lease calendar, MRR matrix, "
        "and pipeline funnel for Furnished Finder / Blueground / corporate-direct hosts (v2.3)."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
