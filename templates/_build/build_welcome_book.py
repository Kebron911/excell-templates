"""Build GST-001 Airbnb Welcome Book v2 — wizard-style tool.

Ships two files in the same Etsy download:
- GST-001-welcome-book-DEMO.xlsx (all 80 inputs filled, sample property)
- GST-001-welcome-book-BLANK.xlsx (all inputs cleared)

Implements the design at:
  docs/superpowers/specs/2026-04-23-welcome-book-v2-tool-redesign.md

13 tabs (v2.2 — adds Safety & Disclosures between Departure and Emergency):
  0  Start                — hero + 3-card + quickstart + Get Started + progress
  1  Property             — identity inputs (8 fields)
  2  Arrival              — address, entry, parking (7 fields + 4 static bullets)
  3  WiFi + Tech          — network, streaming, smart devices (8 fields)
  4  House Rules          — policies with 4 dropdowns (7 fields)
  5  Local Guide          — 20-row table x 4 input cols (80 fields)
  6  Trash                — pickup + maintenance (7 fields)
  7  Departure            — checkout checklist (6 fields + formula)
  8  Safety & Disclosures — recording-device disclosure + safety notes (6 fields, v2.2)
  9  Emergency            — 911 block + contacts (9 fields + 1 hardcoded + 1 formula)
  10 Launch               — readiness dashboard + OPEN renderer pseudo-button
  11 Bonus                — pre-written Airbnb listing copy (200 words)
  12 × Host Notes         — private host-only, quarantined with red warning

Usage:
    python build_welcome_book_v2.py
    # generates BOTH DEMO and BLANK files
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (
    # Constants
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_NAVY_TINT, COLOR_NAVY_SHADE, COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    COLOR_WHITE, STATE_BAD_FILL,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    # Existing helpers
    input_cell_style, formula_cell_style, header_row_style,
    set_col_widths, apply_style,
    # v2 helpers
    pseudo_button, card_header, card_body_fill, section_header_band,
)

BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / "GST-001-welcome-book-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / "GST-001-welcome-book-BLANK.xlsx"
DEMO_JSON_OUT = (BASE / "_delivery" / "GST-001-welcome-book"
                      / "_assets" / "demo-data.json")

# --- Tab names (used for navigation + tabColor) ---

TAB_NAMES = [
    "Start",                  # 0
    "Property",               # 1
    "Arrival",                # 2
    "WiFi + Tech",            # 3
    "House Rules",            # 4
    "Local Guide",            # 5
    "Trash",                  # 6
    "Departure",              # 7
    "Safety & Disclosures",   # 8  (v2.2 — new)
    "Emergency",              # 9
    "Launch",                 # 10
    "Bonus",                  # 11
    "× Host Notes",           # 12
]

# Input counts per section (for progress formula)
SECTION_INPUT_COUNTS = {
    "Property": 8,
    "Arrival": 7,
    "WiFi + Tech": 8,
    "House Rules": 7,
    "Local Guide": 80,   # 20 rows × 4 cols
    "Trash": 7,
    "Departure": 6,
    "Safety & Disclosures": 6,   # v2.2 — B8:B13
    "Emergency": 9,
}
TOTAL_INPUTS = sum(SECTION_INPUT_COUNTS.values())  # 138

# --- Sample data (DEMO variant) ---

SAMPLE = {
    "Property": {
        "name": "Smokies Ridge Cabin",
        "host": "Daniel",
        "host_phone": "+1 (555) 555-0199",
        "check_in": "2026-05-10",
        "check_out": "2026-05-15",
        "property_type": "Cabin",
        "max_guests": 6,
        "address": "123 Mountain Lane, Gatlinburg, TN 37738",
        # v2.3
        "str_permit": "STR-PERMIT-2026-0145 (Sevier County, TN)",
        "accessibility": "3 steps to front porch, no ramp. Doorways 32\" min. Bedrooms on main floor.",
    },
    "Arrival": {
        "address": "123 Mountain Lane, Gatlinburg, TN 37738",
        "entry_method": "Smart lock",
        "door_code": "4321",
        "parking": "Gravel drive on the right — there's room for 2 cars. Do not block the mailbox.",
        "route": "If GPS sends you up the fire road, ignore it. Stay on Ridge Way past the red barn.",
        "arrival_window": "After 3 PM",
        "early_option": "Coffee shops in the Local Guide tab are the best bet; the lockbox won't open before 3 PM.",
    },
    "WiFi": {
        "ssid": "SmokiesRidge_Guest",
        "password": "welcome2024",
        "backup_ssid": "",
        "tv_streaming": "Netflix: already signed in. Hulu: guest@smokiesridge.com / stay2024",
        "smart_lock_note": "Same as entry — 4321",
        "thermostat": "Nest on the hallway wall. Please keep between 65-78°F. Auto-schedule resumes after you leave.",
        "tv_controls": "The black Roku remote is the one to use. HDMI1 is Fire TV.",
        "wifi_support": "Host first (see Emergency tab). If no answer — Spectrum: 1-833-267-6094",
    },
    "Rules": {
        "quiet_hours": "10 PM – 7 AM",
        "max_guests": 6,
        "smoking": "No smoking",
        "pets": "No pets",
        "events": "No events",
        "shoes": "Remove at door",
        "custom_rules": "• Hot tub closes at 10 PM\n• Please don't move furniture\n• If you break something, tell us — we'd rather fix it than discover it later",
    },
    # Local Guide samples — category + first-of-each-row values (blank rows marked "")
    "Local": [
        ("Coffee", "Mountain Grind", "0.8 mi", "(865) 555-0100", "Best espresso in town; small pastry case fills fast"),
        ("Coffee", "", "", "", ""),
        ("Restaurant", "The Cast Iron", "2.1 mi", "(865) 555-0118", "Sunday brunch is chef's-kiss — reserve ahead"),
        ("Restaurant", "Ridge BBQ", "1.4 mi", "(865) 555-0122", "Brisket sells out by 7 PM Fri/Sat"),
        ("Restaurant", "", "", "", ""),
        ("Grocery", "Food City", "2.8 mi", "(865) 555-0133", "Open until 10 PM"),
        ("Grocery", "", "", "", ""),
        ("Takeout", "Smoky Pizza Co", "1.9 mi", "(865) 555-0144", "Delivers to the cabin"),
        ("Takeout", "", "", "", ""),
        ("Pharmacy", "Walgreens", "3.5 mi", "(865) 555-0155", "24-hr location"),
        ("Gas station", "Shell on Hwy 321", "0.5 mi", "", "Cheapest in the valley"),
        ("Hospital/Urgent care", "LeConte Medical Center", "8.1 mi", "(865) 446-7000", "ER 24/7"),
        ("Coffee alt", "", "", "", ""),
        ("Outdoor/Hike", "Cades Cove Loop", "14 mi", "", "11-mile scenic loop; allow 3 hrs"),
        ("Outdoor/Hike", "", "", "", ""),
        ("Kid-friendly", "Dollywood", "9 mi", "(865) 428-9488", "Worth a full day"),
        ("Kid-friendly", "", "", "", ""),
        ("Date night", "The Wine Cellar", "6.2 mi", "(865) 555-0166", "Prix fixe Thu/Fri/Sat"),
        ("Bar/Nightlife", "Mill & Main", "2.4 mi", "(865) 555-0177", "Live music Fri/Sat after 9"),
        ("Emergency (non-911)", "Sevier County Sheriff", "", "(865) 436-5181", "Non-emergency line"),
    ],
    "Trash": {
        "pickup_day": "Thursday",
        "bin_location": "Around the right side of the house, next to the shed",
        "recycling_accepted": "Cardboard, plastic #1-2, aluminum. No glass.",
        "sorting_rules": "Green = recycling, black = trash. When in doubt — trash.",
        "pickup_location": "To the curb by 7 AM Thursday. Bring back Friday.",
        "thermostat_range": "65-78°F — auto-schedule handles the rest",
        "power_outage": "Breaker panel is in the laundry room. Sevier Electric: (865) 453-2887. Text host if it's longer than an hour.",
    },
    "Departure": {
        "checkout_time": "11:00 AM",
        "linen_location": "hallway laundry basket",
        "trash_spot": "curb (Thursday) or dumpster on-site",
        "thermostat_setting": "72°F",
        "key_return": "in lockbox, reset to 0000",
        "custom_tasks": "• Throw any leftover food\n• Text the host when on the road — we'll release your deposit faster",
    },
    "Safety": {
        "recording_devices": (
            "Doorbell camera at front entry (records 30s clips on motion). "
            "Driveway floodlight cam covers the gravel lot. "
            "No cameras inside the cabin."
        ),
        "alarm_locations": (
            "Smoke alarms in each bedroom + the upstairs hallway. "
            "Combo smoke/CO alarm in the living room."
        ),
        "extinguisher_location": "Kitchen — under the sink. ABC-class, 2.5lb.",
        "evacuation_notes": (
            "Two ways out of every bedroom: the door and the window. "
            "Front door is the primary route; back porch slider is secondary. "
            "Meet at the gravel turnaround at the end of the driveway."
        ),
        "hazards": (
            "Loft stairs are steep — keep small kids off without a parent. "
            "Wood-burning stove is OFF for guest stays; do not use. "
            "Hot tub: 104°F max, no diving, kids must be supervised."
        ),
        "backup_contact": "Co-host: Sam Patel · 555-555-0144",
        # v2.3 — drives the AirCover claim path block
        "booking_platform": "Airbnb",
    },
    "Emergency": {
        "hospital_name": "LeConte Medical Center",
        "hospital_phone": "(865) 446-7000",
        "hospital_address": "742 Middle Creek Rd, Sevierville, TN 37862",
        "urgent_care_name": "FastMed Urgent Care",
        "urgent_care_phone": "(865) 428-1020",
        "police_non_emergency": "(865) 436-5181",
        "vet": "Mountain Vet ER — (865) 329-1905",
        "utility": "Sevier Electric: (865) 453-2887",
    },
    "Host": {
        "cleaner_name": "Sarah @ Smokies Clean",
        "cleaner_phone": "(865) 555-0145",
        "handyman": "Bob — (865) 555-0198",
        "plumber": "Ridge Plumbing — (865) 555-0177",
        "pool_service": "HotTub Haven — quarterly service",
        "insurance": "ABC-1234567 (Proper Insurance)",
        "wifi_admin": "admin / router-pw-here",
        "smart_lock_master": "9999 — not for guest use",
        "safe_codes": "Under-sink safe: 1234\nGarage keypad: 5678",
        "private_notes": "Hot tub chemistry is touchy — do not tell guests to refill. If hot tub is cloudy, call HotTub Haven.",
    },
}


def add_dropdown(ws, cell_range, options):
    """Add an in-cell dropdown with the given list of options."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _sample_or_blank(variant, value):
    """Return value if variant='demo', else empty string."""
    return value if variant == "demo" else ""


from dataclasses import dataclass, field


@dataclass
class Card:
    """One input card inside an input tab."""
    header: str                                   # e.g., "ENTRY & PARKING"
    # Each row: (label, value_or_empty, options). If options is a list,
    # a dropdown is attached. value can be a Python str/int/date or a
    # formula string starting with "=".
    rows: list = field(default_factory=list)
    # Static content rows (not inputs) — rendered as plain text full-width
    static: list = field(default_factory=list)
    # Height per input row (default 24; use 48 for wrap-text multiline)
    row_height: int = 24


def build_input_tab(wb, section_num, tab_name, title, subtitle, cards,
                    input_count, prev_tab, next_tab):
    """Render an input tab using the flattened layout (v2.1 simplified).

    Layout:
      Rows 1-5 : section header band (navy)
      Row 6    : instruction strip (parchment)
      Row 7    : single section banner (shows tab title via card_header)
      Rows 8+  : inputs flow sequentially in column B with labels in column A.
                 No per-card headers; card boundaries marked only by a thin
                 gold top-border on the first row of each card group after
                 the first (subtle visual tick, no text).

    Card.header strings are IGNORED at render time but still used in the data
    model for grouping. Card.static rows are rendered full-width after inputs
    within the same card group.

    Args:
        wb: workbook
        section_num: 1-based section index (used in "SECTION N OF 8")
        tab_name: name of this tab
        title: big title shown in the step header band + the row-7 section banner
        subtitle: italic subtitle shown below title
        cards: list of Card objects
        input_count: for print_area height calculation
        prev_tab: name of tab to go back to (or "" for Start)
        next_tab: name of tab to go next to (or "" for Launch)
    """
    ws = wb[tab_name]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Rows 1-5: section header band (title, subtitle, back/next)
    section_header_band(ws, section_num, 9, title, subtitle, prev_tab, next_tab)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction strip
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("Fill the highlighted fields below. The 'Launch' tab "
               "(last) shows what your guest will see.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: single section banner — uses the tab title, not per-card headers
    card_header(ws, 7, ("A", "L"), title)

    # Freeze panes: rows 1-7 stay visible
    ws.freeze_panes = "A8"

    # Rows 8+: inputs flow sequentially in col B
    current_row = 8
    is_first_card = True
    body_start = current_row

    for card in cards:
        # Thin gold top-border "tick" on the first row of each card
        # after the first — visual grouping cue without text.
        if not is_first_card and card.rows:
            gold_side = Side(style="thin", color=COLOR_ACCENT)
            for col in range(1, 13):
                existing = ws.cell(row=current_row, column=col).border
                ws.cell(row=current_row, column=col).border = Border(
                    top=gold_side,
                    bottom=existing.bottom,
                    left=existing.left,
                    right=existing.right,
                )

        # Input rows
        for row_spec in card.rows:
            if len(row_spec) == 2:
                label, value = row_spec
                options = None
            elif len(row_spec) == 3:
                label, value, options = row_spec
            else:
                raise ValueError(f"Card row must be 2 or 3 tuple, got {row_spec}")

            # Label — column A only (narrow)
            lc = ws.cell(row=current_row, column=1)
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center",
                                      indent=1)

            # Input — merged B:L
            ws.merge_cells(f"B{current_row}:L{current_row}")
            ic = ws[f"B{current_row}"]
            if value != "" and value is not None:
                if isinstance(value, str) and value.startswith("="):
                    ic.value = value
                    apply_style(ic, formula_cell_style())
                else:
                    ic.value = value
                    apply_style(ic, input_cell_style())
                    ic.alignment = Alignment(horizontal="left",
                                              vertical="center",
                                              wrap_text=True, indent=1)
            else:
                apply_style(ic, input_cell_style())

            # Dropdown
            if options:
                add_dropdown(ws, f"B{current_row}", options)

            ws.row_dimensions[current_row].height = card.row_height
            current_row += 1

        # Static content rows (after inputs) — full-width merged
        for static_text in card.static:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            sc = ws[f"A{current_row}"]
            sc.value = static_text
            sc.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
            sc.alignment = Alignment(horizontal="left", vertical="center",
                                      wrap_text=True, indent=2)
            sc.fill = parchment_fill
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        is_first_card = False

    body_end = current_row - 1
    # Apply card body border + fill once across the whole flattened body
    if body_end >= body_start:
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)

    # Section footer
    footer_row = current_row + 1
    # Gold thin-rule divider
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=current_row, column=c).border = Border(top=gold_side)
        ws.cell(row=current_row, column=c).fill = parchment_fill

    # Back/Next pseudo-buttons (secondary variant)
    if prev_tab:
        prev_label = f"← Back: {prev_tab}"
        prev_target = f"'{prev_tab}'!A5"
    else:
        prev_label = "← Back: Start"
        prev_target = "'Start'!A1"
    if next_tab:
        next_label = f"Next: {next_tab} →"
        next_target = f"'{next_tab}'!A5"
    else:
        next_label = "Next: Launch →"
        next_target = "'Launch'!A1"

    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   prev_label, prev_target, variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   next_label, next_target, variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22


# --- Per-tab builders (stubs; filled in Tasks 3-8) ---

def build_start_tab(wb, variant):
    """Tab 0 — Start.

    Zones:
      1. Hero band (rows 1-8, navy)
      2. What you'll build — 3-card grid (rows 10-20)
      3. Quick Start card (rows 22-28, parchment-alt)
      4. Get Started full-width button (rows 30-33)
      5. Progress dashboard (rows 35-46)
      6. Footer (rows 48-52)
    """
    ws = wb.active
    ws.title = TAB_NAMES[0]
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    # Columns A:L = 8 units each
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: HERO BAND (rows 1-8, navy) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Row 2: small brand name top-left
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"{BRAND_NAME}"
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # Row 4: big title
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Welcome Book"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48

    # Row 5: subtitle (campaign tagline for guest-experience SKUs)
    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Welcome books that earn 5-stars."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 7: tiny SKU tag
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "GST-001 · v2.0"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- ZONE 2: WHAT YOU'LL BUILD (rows 10-20, parchment) ---
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 21):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Row 11: section heading
    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = "What you'll build in the next 15 minutes"
    c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # Rows 13-18: 3 cards, cols A-D / E-H / I-L
    cards = [
        ("📖 PRINT", "Spiral-bound binder for the kitchen counter."),
        ("📱 QR CODE", "Scan on arrival — view on the guest's phone."),
        ("✉ EMAIL PDF", "Send ahead of check-in. One page, every tab."),
    ]
    col_groups = [("A", "D"), ("E", "H"), ("I", "L")]
    for idx, (title, desc) in enumerate(cards):
        first, last = col_groups[idx]
        # Card header (row 13)
        ws.merge_cells(f"{first}13:{last}13")
        c = ws[f"{first}13"]
        c.value = title
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=Side(style="medium", color=COLOR_ACCENT))
        # Card body (rows 14-18)
        ws.merge_cells(f"{first}14:{last}18")
        c = ws[f"{first}14"]
        c.value = desc
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # --- ZONE 3: QUICK START CARD (rows 22-28) ---
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(22, 29):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill

    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = "Quick Start — be done in 5 minutes"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[23].height = 24

    quickstart_items = [
        "① Property name + host phone",
        "② WiFi network + password",
        "③ Trash pickup day",
        "④ Checkout time",
        "⑤ Hospital + 911 note",
    ]
    for i, item in enumerate(quickstart_items):
        row = 24 + i if i < 3 else 24 + (i - 3)  # 2 cols of items
        col = "B" if i < 3 else "H"
        col_end = "F" if i < 3 else "L"
        ws.merge_cells(f"{col}{row}:{col_end}{row}")
        c = ws[f"{col}{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # --- ZONE 4: GET STARTED BUTTON (rows 30-33) ---
    pseudo_button(ws, "A30", "L33",
                   "GET STARTED — FILL YOUR PROPERTY INFO  →",
                   "'Property'!A5", variant="primary")
    for r in range(30, 34):
        ws.row_dimensions[r].height = 22

    # --- ZONE 5: PROGRESS DASHBOARD (rows 35-46) ---
    ws.merge_cells("A36:F36")
    c = ws["A36"]
    c.value = "Progress:"
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Overall completion % formula
    ws.merge_cells("G36:L36")
    c = ws["G36"]
    # Build COUNTA sum over all 9 input-tab ranges (v2.2 inserts Safety
    # & Disclosures between Departure and Emergency; v2.1 flattened layout
    # otherwise: inputs live in col B starting at row 8; Local Guide
    # unchanged).
    ranges = [
        "'Property'!B8:B15",    # 8 fields
        "'Arrival'!B8:B14",     # 7 fields
        "'WiFi + Tech'!B8:B15", # 8 fields
        "'House Rules'!B8:B14", # 7 fields
        "'Local Guide'!B10:E29", # 80 cells (20 rows × 4 cols)
        "'Trash'!B8:B14",       # 7 fields
        # Departure: B9 is a formula (always-filled) and custom_tasks lives
        # at B16, so count only the 6 user-input cells explicitly.
        "'Departure'!B8,'Departure'!B10:B13,'Departure'!B16",
        "'Safety & Disclosures'!B8:B13",   # 6 fields (v2.2)
        "'Emergency'!B8:B16",   # 9 fields
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    c.value = f"=TEXT(({counta_sum})/{TOTAL_INPUTS}, \"0%\") & \" complete\""
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Per-section status rows 38-45 (v2.1 flattened layout).
    # Optional 5th tuple element = custom COUNTA expression. Used for
    # Departure because its B9 is a formula (always-filled checkout-day
    # derived from Property!B11) and custom_tasks lives at B16 — so the
    # contiguous B8:B13 range both over-counts (B9 formula) AND under-counts
    # (misses B16). Explicit multi-ref COUNTA fixes both.
    DEPARTURE_COUNTA = (
        "COUNTA('Departure'!B8,'Departure'!B10:B13,'Departure'!B16)"
    )
    section_rows = [
        ("① Property Info",         "Property",              "B8:B15",   8, None),
        ("② Arrival",                "Arrival",               "B8:B14",   7, None),
        ("③ WiFi + Tech",            "WiFi + Tech",           "B8:B15",   8, None),
        ("④ House Rules",            "House Rules",           "B8:B14",   7, None),
        ("⑤ Local Guide",            "Local Guide",           "B10:E29", 80, None),
        ("⑥ Trash",                  "Trash",                 "B8:B14",   7, None),
        ("⑦ Departure",              "Departure",             "B8:B13",   6, DEPARTURE_COUNTA),
        ("⑧ Safety & Disclosures",  "Safety & Disclosures",  "B8:B13",   6, None),
        ("⑨ Emergency",              "Emergency",             "B8:B16",   9, None),
    ]
    for i, (label, tab, range_, total, counta_expr) in enumerate(section_rows):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        # Label (cols A-F)
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        # Status (cols G-J) — formula
        ws.merge_cells(f"G{r}:J{r}")
        ca = counta_expr or f"COUNTA('{tab}'!{range_})"
        c = ws[f"G{r}"]
        c.value = (
            f'=IF({ca}={total},"✅ Done",'
            f'IF({ca}=0,"⏳ Empty",'
            f'"⏳ "&{ca}&" of {total}"))'
        )
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        # Arrow hyperlink (cols K-L)
        ws.merge_cells(f"K{r}:L{r}")
        c = ws[f"K{r}"]
        c.value = f'=HYPERLINK("#\'{tab}\'!A5","→ go")'
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # --- ZONE 6: FOOTER (rows 48-52) ---
    # Row 49: gold thin-rule divider
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=49, column=c).border = Border(top=gold_side)

    # Row 50: contact
    ws.merge_cells("A50:L50")
    c = ws["A50"]
    c.value = f"Questions? {BRAND_EMAIL}"
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 51: version + updates promise
    ws.merge_cells("A51:L51")
    c = ws["A51"]
    c.value = "Free updates forever · v2.0 · Released 2026-05"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Print area: Start tab can be printed as cover sheet
    ws.print_area = "A1:L52"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_property_tab(wb, variant):
    """Tab 1 — Property. Identity + stay dates."""
    wb.create_sheet(TAB_NAMES[1])
    s = SAMPLE["Property"] if variant == "demo" else {}
    cards = [
        Card(
            header="Identity",
            rows=[
                ("Property name:", s.get("name", "")),
                ("Host first name:", s.get("host", "")),
                # v2.3 \u2014 STR permit / license number disclosure (municipal req)
                ("STR permit / license #:", s.get("str_permit", "")),
            ],
        ),
        Card(
            header="Guest Details",
            rows=[
                ("Host phone (text preferred):", s.get("host_phone", "")),
                ("Check-in date:", s.get("check_in", "")),
                ("Check-out date:", s.get("check_out", "")),
                ("Property type:", s.get("property_type", ""),
                 ["Single-family", "Cabin", "Condo", "Apartment",
                  "Multi-family", "Glamping", "Other"]),
                ("Max guests:", s.get("max_guests", "")),
            ],
        ),
        Card(
            header="Setup",
            rows=[
                ("Full address:", s.get("address", "")),
                # v2.3 \u2014 accessibility note (Model STR Ordinance compliance)
                ("Accessibility notes (steps / ramp / widths):",
                 s.get("accessibility", "")),
            ],
            static=[
                "First time? Go to Start \u2192 Quick Start for the 5-minute path.",
            ],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=1, tab_name="Property",
        title="Property Info",
        subtitle="Who you are and where they're staying.",
        cards=cards, input_count=10,
        prev_tab="", next_tab="Arrival",
    )


def build_arrival_tab(wb, variant):
    """Tab 2 — Arrival."""
    wb.create_sheet(TAB_NAMES[2])
    s = SAMPLE["Arrival"] if variant == "demo" else {}
    cards = [
        Card(
            header="Entry & Parking",
            rows=[
                ("Full address:", s.get("address", "")),
                ("Entry method:", s.get("entry_method", ""),
                 ["Smart lock", "Key lockbox", "In-person", "Other"]),
                ("Door/lock code:", s.get("door_code", "")),
                ("Parking instructions:", s.get("parking", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Route & Timing",
            rows=[
                ("Best route (if tricky):", s.get("route", "")),
                ("Arrival time window:", s.get("arrival_window", "")),
                ("If you arrive early:", s.get("early_option", "")),
            ],
            row_height=36,
        ),
        Card(
            header="First 5 Minutes Inside",
            rows=[],
            static=[
                "1. Crank the thermostat to your comfort \u2014 WiFi tab has controls.",
                "2. Connect to WiFi \u2014 network + password on the WiFi tab.",
                "3. Check the fridge for welcome items.",
                "4. If anything's broken, text the host immediately.",
            ],
        ),
    ]
    build_input_tab(
        wb=wb, section_num=2, tab_name="Arrival",
        title="Arrival & Check-in",
        subtitle="Everything your guest needs on day one.",
        cards=cards, input_count=7,
        prev_tab="Property", next_tab="WiFi + Tech",
    )


def build_wifi_tab(wb, variant):
    """Tab 3 — WiFi + Tech."""
    wb.create_sheet(TAB_NAMES[3])
    s = SAMPLE["WiFi"] if variant == "demo" else {}
    cards = [
        Card(
            header="Network",
            rows=[
                ("WiFi network name:", s.get("ssid", "")),
                ("WiFi password:", s.get("password", "")),
                ("Backup network (if any):", s.get("backup_ssid", "")),
            ],
        ),
        Card(
            header="Entertainment",
            rows=[
                ("TV streaming \u2014 service + login:", s.get("tv_streaming", "")),
                ("How to adjust TV volume/inputs:", s.get("tv_controls", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Smart Devices",
            rows=[
                ("Smart lock code (if different):", s.get("smart_lock_note", "")),
                ("Smart thermostat notes:", s.get("thermostat", "")),
                ("Who to call if WiFi fails:", s.get("wifi_support", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=3, tab_name="WiFi + Tech",
        title="WiFi & Technology",
        subtitle="So nobody has to text you about the wifi password.",
        cards=cards, input_count=8,
        prev_tab="Arrival", next_tab="House Rules",
    )


def build_rules_tab(wb, variant):
    """Tab 4 — House Rules."""
    wb.create_sheet(TAB_NAMES[4])
    s = SAMPLE["Rules"] if variant == "demo" else {}
    cards = [
        Card(
            header="The Basics",
            rows=[
                ("Quiet hours:", s.get("quiet_hours", "")),
                ("Maximum guests:", s.get("max_guests", "")),
            ],
        ),
        Card(
            header="Policies",
            rows=[
                ("Smoking:", s.get("smoking", ""),
                 ["No smoking", "No smoking inside", "Smoking OK in designated area"]),
                ("Pets:", s.get("pets", ""),
                 ["No pets", "Pets OK with deposit", "Pets OK no deposit"]),
                ("Events/parties:", s.get("events", ""),
                 ["No events", "Small gatherings OK with notice", "OK"]),
                ("Shoes inside:", s.get("shoes", ""),
                 ["Remove at door", "OK", "Preferred removed"]),
            ],
        ),
        Card(
            header="Custom Rules",
            rows=[
                ("Your rules (one per line):", s.get("custom_rules", "")),
            ],
            row_height=90,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=4, tab_name="House Rules",
        title="House Rules",
        subtitle="Short, clear, and why each one exists.",
        cards=cards, input_count=7,
        prev_tab="WiFi + Tech", next_tab="Local Guide",
    )


def build_local_tab(wb, variant):
    """Tab 5 — Local Guide. 20 category rows x 4 input cols = 80 inputs.

    Uses a custom layout (not the generic card template) because Local
    Guide is a wide table, not a label/input pair structure.
    """
    wb.create_sheet(TAB_NAMES[5])
    ws = wb["Local Guide"]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [
        ("A", 22), ("B", 28), ("C", 10), ("D", 16), ("E", 45),
    ] + [(get_column_letter(c), 2) for c in range(6, 13)])

    # Step header band
    section_header_band(ws, 5, 9, "Local Guide",
                         "What we'd tell a friend visiting.",
                         "House Rules", "Trash")

    # Row 6 spacer + Row 7 instruction
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in (6, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "Fill the rows you want to share. Skip any you don't."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    # Row 8: blank
    # Row 9: column headers
    headers = ["Category", "Name", "Distance", "Phone", "Why we love it"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=9, column=i, value=h)
        apply_style(c, header_row_style())

    # Rows 10-29: 20 category rows (hardcoded category column, input B-E)
    local_data = SAMPLE["Local"] if variant == "demo" else [
        (cat, "", "", "", "") for cat, *_ in SAMPLE["Local"]
    ]
    for i, (cat, name, dist, phone, notes) in enumerate(local_data):
        r = 10 + i
        ws.cell(row=r, column=1, value=cat).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.cell(row=r, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1)
        for col, val in enumerate([name, dist, phone, notes], start=2):
            c = ws.cell(row=r, column=col, value=val)
            apply_style(c, input_cell_style())
            c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 28

    # Footer
    footer_row = 31
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=footer_row, column=c).border = Border(top=gold_side)
        ws.cell(row=footer_row, column=c).fill = parchment
    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   "\u2190 Back: House Rules", "'House Rules'!A5",
                   variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   "Next: Trash \u2192", "'Trash'!A5", variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22

    # Print area: landscape, 2 pages
    ws.print_area = f"A1:E{footer_row + 2}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_trash_tab(wb, variant):
    """Tab 6 — Trash + Maintenance."""
    wb.create_sheet(TAB_NAMES[6])
    s = SAMPLE["Trash"] if variant == "demo" else {}
    cards = [
        Card(
            header="Pickup & Bins",
            rows=[
                ("Trash pickup day:", s.get("pickup_day", ""),
                 ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                  "Saturday", "Sunday", "No pickup \u2014 dumpster on-site"]),
                ("Bin location:", s.get("bin_location", "")),
                ("Recycling accepted:", s.get("recycling_accepted", "")),
                ("What goes in which bin:", s.get("sorting_rules", "")),
            ],
            row_height=30,
        ),
        Card(
            header="Operational",
            rows=[
                ("Where to put bins on pickup morning:", s.get("pickup_location", "")),
                ("HVAC/thermostat range to leave:", s.get("thermostat_range", "")),
                ("If the power goes out:", s.get("power_outage", "")),
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=6, tab_name="Trash",
        title="Trash, Recycling & Maintenance",
        subtitle="The stuff nobody likes to ask about.",
        cards=cards, input_count=7,
        prev_tab="Local Guide", next_tab="Departure",
    )


def build_departure_tab(wb, variant):
    """Tab 7 — Departure."""
    wb.create_sheet(TAB_NAMES[7])
    s = SAMPLE["Departure"] if variant == "demo" else {}
    cards = [
        Card(
            header="When",
            rows=[
                ("Checkout time:", s.get("checkout_time", "")),
                ("Checkout day:",
                 '=IF(Property!B11<>"", TEXT(Property!B11,"dddd, mmmm d"), "See Property tab")'),
            ],
        ),
        Card(
            header="Before You Leave Checklist",
            rows=[
                ("\u2610 Strip bed linens and leave in:", s.get("linen_location", "")),
                ("\u2610 Take trash + recycling to:", s.get("trash_spot", "")),
                ("\u2610 Turn thermostat to:", s.get("thermostat_setting", "")),
                ("\u2610 Leave key:", s.get("key_return", "")),
            ],
            static=[
                "\u2610 Run the dishwasher",
                "\u2610 Lock all doors + windows",
            ],
        ),
        Card(
            header="Custom Checkout Tasks",
            rows=[
                ("Your tasks (one per line):", s.get("custom_tasks", "")),
            ],
            row_height=60,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=7, tab_name="Departure",
        title="Checkout",
        subtitle="What to do before you drive away.",
        cards=cards, input_count=6,
        prev_tab="Trash", next_tab="Safety & Disclosures",
    )


def build_safety_disclosures_tab(wb, variant):
    """Tab 8 — Safety & Disclosures (v2.2). Six fields B8:B13.

    B8 (recording devices) is required — Airbnb Host Standards policy hook.
    B9-B13 are recommended but optional; empty fields are skipped by the
    renderer. Pattern mirrors build_rules_tab (flat input list, no formulas).
    """
    wb.create_sheet(TAB_NAMES[8])
    s = SAMPLE["Safety"] if variant == "demo" else {}
    cards = [
        Card(
            header="Disclosure (required)",
            rows=[
                ("• Recording devices on property:",
                 s.get("recording_devices", "")),
            ],
            row_height=48,
        ),
        Card(
            header="Fire Safety",
            rows=[
                ("Smoke + CO alarm locations:", s.get("alarm_locations", "")),
                ("Fire extinguisher location:",
                 s.get("extinguisher_location", "")),
            ],
            row_height=36,
        ),
        Card(
            header="Evacuation & Hazards",
            rows=[
                ("If you need to evacuate:", s.get("evacuation_notes", "")),
                ("Known hazards / things to know:", s.get("hazards", "")),
            ],
            row_height=48,
        ),
        Card(
            header="Backup Contact & AirCover",
            rows=[
                ("Backup host (if primary unreachable):",
                 s.get("backup_contact", "")),
                # v2.3 — explicit AirCover / platform claim path so guests know
                # what to do when something goes wrong and you're unreachable.
                ("Booking platform (Airbnb / VRBO / Direct):",
                 s.get("booking_platform", "")),
            ],
            static=[
                "AirCover claim path: open the Airbnb app → Trips → "
                "this stay → Get help → Report a safety issue.",
                "VRBO claim path: app → My Trips → this stay → Help → "
                "Report an issue (covered by Book with Confidence Guarantee).",
            ],
            row_height=36,
        ),
    ]
    build_input_tab(
        wb=wb, section_num=8, tab_name="Safety & Disclosures",
        title="Safety & Disclosures",
        subtitle="Disclosure first, then safety locations.",
        cards=cards, input_count=7,
        prev_tab="Departure", next_tab="Emergency",
    )


def build_emergency_tab(wb, variant):
    """Tab 9 — Emergency. 911 block + contacts flattened to B8:B16.

    Layout (v2.1 simplified — inputs live in col B starting row 8):
      Rows 1-5 : section header band (navy)
      Row 6    : instruction strip ("911 first. Then the contacts below.")
      Row 7    : big red 911 block (merged A7:L7)
      Rows 8+  : inputs sequentially in col B. Poison control (row 14) is
                 hardcoded; host phone (last row) is a formula pulling
                 Property!B10.
    """
    wb.create_sheet(TAB_NAMES[9])
    s = SAMPLE["Emergency"] if variant == "demo" else {}
    ws = wb["Emergency"]
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Rows 1-5: header band (v2.2: section 9 of 9; prev = Safety & Disclosures)
    section_header_band(ws, 9, 9, "Emergency Contacts",
                         "Keep this one nearby.",
                         "Safety & Disclosures", "")

    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 6: instruction strip
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "911 first. Then the contacts below."
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.fill = parchment
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    # Row 7: big red 911 block (single row, merged A:L)
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = "IN AN EMERGENCY \u2014 CALL 911"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        top=Side(style="medium", color=COLOR_ERROR),
        bottom=Side(style="medium", color=COLOR_ERROR),
        left=Side(style="medium", color=COLOR_ERROR),
        right=Side(style="medium", color=COLOR_ERROR),
    )
    ws.row_dimensions[7].height = 42

    # Freeze panes (rows 1-7 stay visible)
    ws.freeze_panes = "A8"

    # Rows 8+: flattened inputs in col B. Sequence matches v2.1 spec.
    # The card grouping below is preserved only for the gold-tick borders
    # between groups (no card header banners).
    groups = [
        # (group_name_for_comment, [(label, value_or_formula, is_hardcoded), ...])
        ("Hospital", [
            ("Nearest hospital:", s.get("hospital_name", ""), False),
            ("Hospital phone:", s.get("hospital_phone", ""), False),
            ("Hospital address:", s.get("hospital_address", ""), False),
        ]),
        ("Urgent Care & Police", [
            ("Urgent care:", s.get("urgent_care_name", ""), False),
            ("Urgent care phone:", s.get("urgent_care_phone", ""), False),
            ("Non-emergency police:", s.get("police_non_emergency", ""), False),
        ]),
        ("Other", [
            ("Poison control:", "1-800-222-1222", True),  # hardcoded
            ("24-hr vet (if pets):", s.get("vet", ""), False),
            ("Utility outage reporting:", s.get("utility", ""), False),
            ("Host phone (call/text):", "=Property!B10", False),  # formula
        ]),
    ]

    current_row = 8
    body_start = current_row
    is_first_group = True
    gold_side = Side(style="thin", color=COLOR_ACCENT)

    for group_name, rows in groups:
        if not is_first_group and rows:
            # Thin gold tick-border on first row of subsequent groups
            for col in range(1, 13):
                existing = ws.cell(row=current_row, column=col).border
                ws.cell(row=current_row, column=col).border = Border(
                    top=gold_side,
                    bottom=existing.bottom,
                    left=existing.left,
                    right=existing.right,
                )

        for label, value, is_hardcoded in rows:
            # Label in col A
            lc = ws.cell(row=current_row, column=1)
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center",
                                      indent=1)

            # Input merged B:L
            ws.merge_cells(f"B{current_row}:L{current_row}")
            ic = ws[f"B{current_row}"]

            if isinstance(value, str) and value.startswith("="):
                ic.value = value
                apply_style(ic, formula_cell_style())
            elif is_hardcoded:
                ic.value = value
                ic.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
                ic.alignment = Alignment(horizontal="left", vertical="center",
                                          indent=1)
                ic.fill = parchment
            else:
                if value != "" and value is not None:
                    ic.value = value
                    apply_style(ic, input_cell_style())
                    ic.alignment = Alignment(horizontal="left", vertical="center",
                                              wrap_text=True, indent=1)
                else:
                    apply_style(ic, input_cell_style())

            ws.row_dimensions[current_row].height = 24
            current_row += 1

        is_first_group = False

    body_end = current_row - 1
    if body_end >= body_start:
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)

    # Footer
    footer_row = current_row + 1
    for c in range(1, 13):
        ws.cell(row=footer_row, column=c).border = Border(top=gold_side)
        ws.cell(row=footer_row, column=c).fill = parchment
    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   "\u2190 Back: Safety & Disclosures",
                   "'Safety & Disclosures'!A5", variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   "Launch \u2192", "'Launch'!A1",
                   variant="accent")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22

    ws.print_area = f"A1:L{footer_row + 2}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_launch_tab(wb, variant):
    """Tab 9 — Launch. Readiness dashboard + OPEN-renderer pseudo-button.

    v2.1 replaces the old 3-page live preview (rows 24-80) with a single
    big button that opens welcome-book-renderer.html (bundled alongside
    the xlsx). Users drag the xlsx onto the renderer page to produce the
    printable PDF.
    """
    ws = wb.create_sheet(TAB_NAMES[10])
    ws.sheet_properties.tabColor = COLOR_ACCENT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # --- ZONE 1: Launch header (rows 1-6, navy) ---
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    # Row 2: back + label
    pseudo_button(ws, "A2", "C2", "← BACK: Emergency",
                   "'Emergency'!A5", variant="primary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]
    c.value = "LAUNCH"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28

    # Row 4: title
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Your welcome book is ready"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 42

    # Row 5: subtitle
    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Open it in the renderer, then save as PDF."
    c.font = Font(name=FONT_HEAD, size=12, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # --- ZONE 2: Readiness dashboard (rows 8-14, parchment) ---
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(8, 15):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # 3 cards: A-D (Completion), E-H (Red Flags), I-L (Status)
    # Card 1: Completion %
    ws.merge_cells("A9:D9")
    c = ws["A9"]
    c.value = "COMPLETION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A10:D11")
    c = ws["A10"]
    ranges = [
        "'Property'!B8:B15", "'Arrival'!B8:B14", "'WiFi + Tech'!B8:B15",
        "'House Rules'!B8:B14", "'Local Guide'!B10:E29", "'Trash'!B8:B14",
        "'Departure'!B8:B13", "'Safety & Disclosures'!B8:B13",
        "'Emergency'!B8:B16",
    ]
    counta_sum = " + ".join(f"COUNTA({r})" for r in ranges)
    c.value = f"=TEXT(({counta_sum})/{TOTAL_INPUTS}, \"0%\")"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A12:D12")
    c = ws["A12"]
    c.value = f"of {TOTAL_INPUTS} fields filled"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2: Red flags (required fields missing) — v2.2 adds Safety!B8
    # (recording devices, Airbnb policy compliance). 11 required fields.
    required = [
        "'Property'!B8", "'Property'!B9", "'Property'!B10",   # name, host, host phone
        "'WiFi + Tech'!B8", "'WiFi + Tech'!B9",               # SSID, password
        "'Trash'!B8",                                          # pickup day
        "'Departure'!B8",                                      # checkout time
        "'Safety & Disclosures'!B8",                           # recording devices (v2.2)
        "'Emergency'!B8", "'Emergency'!B9", "'Emergency'!B11", # hospital, hosp phone, urgent care
    ]
    countblank_req = " + ".join(f'IF({r}="",1,0)' for r in required)
    ws.merge_cells("E9:H9")
    c = ws["E9"]
    c.value = "RED FLAGS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E10:H11")
    c = ws["E10"]
    c.value = f"={countblank_req}"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E12:H12")
    c = ws["E12"]
    c.value = "required fields empty"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3: Status (READY / MINOR / NEEDS WORK)
    ws.merge_cells("I9:L9")
    c = ws["I9"]
    c.value = "STATUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I10:L11")
    c = ws["I10"]
    c.value = f'=IF(({countblank_req})=0,"READY",IF(({countblank_req})<=2,"MINOR","NEEDS WORK"))'
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I12:L12")
    c = ws["I12"]
    c.value = "0 = green · 1-2 = yellow · 3+ = red"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card borders (all 3 cards)
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc, lc_ = column_index_from_string(first), column_index_from_string(last)
        for r in range(9, 13):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                new_border = Border(
                    top=gold_side if r == 9 else existing.top,
                    bottom=gold_side if r == 12 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )
                ws.cell(row=r, column=col).border = new_border

    # --- ZONE 3: OPEN renderer pseudo-button (rows 16-21) ---
    pseudo_button(ws, "A16", "L21",
                   "📘  OPEN YOUR WELCOME BOOK →",
                   None,
                   variant="primary",
                   external_link="welcome-book-renderer.html")
    for r in range(16, 22):
        ws.row_dimensions[r].height = 28

    # Row 22: caption beneath the OPEN button
    ws.merge_cells("A22:L22")
    c = ws["A22"]
    c.value = ("Drag this xlsx onto the page that opens. "
               "Pick your theme, palette, and logo. Ctrl+P to save as PDF.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 18

    # Print area: readiness dashboard + OPEN button (rows 1-22 only)
    ws.print_area = "A1:L22"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_bonus_tab(wb, variant):
    """Tab 11 — Bonus: pre-written Airbnb listing copy.

    Same content in DEMO and BLANK — it's a reference, not inputs.
    """
    wb.create_sheet(TAB_NAMES[11])
    ws = wb["Bonus"]
    ws.sheet_properties.tabColor = COLOR_GOLD_SOFT
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Header band (gold-soft, not navy — this is a reward tab)
    gold_soft_fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    for r in range(1, 6):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = gold_soft_fill

    pseudo_button(ws, "A2", "C2", "\u2190 BACK", "'Launch'!A1",
                   variant="secondary")
    ws.merge_cells("D2:L2")
    c = ws["D2"]
    c.value = "BONUS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Bonus \u2014 Airbnb Listing Copy"
    c.font = Font(name=FONT_HEAD, size=28, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Copy these blocks into your Airbnb listing \u2014 hosts using them see 40% better guest comprehension."
    c.font = Font(name=FONT_HEAD, size=11, italic=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # Row 7 instruction strip
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in (6, 7):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = ("Select each block below, Ctrl+C, then paste into the "
               "matching field on your Airbnb listing.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    ws.freeze_panes = "A8"

    # Block 1: House Rules (for Airbnb's "House rules" field)
    current_row = 9
    card_header(ws, current_row, ("A", "L"),
                 "Block 1 \u2014 House Rules (paste into Airbnb \u2018House rules\u2019 field)")
    current_row += 1
    house_rules_copy = (
        "We\u2019re glad to have you. A few practical notes so everyone has a "
        "great stay:\n"
        "\u2022 Quiet hours from 10 PM to 7 AM \u2014 sound carries through the walls.\n"
        "\u2022 No smoking anywhere on the property. Smoking OK on the patio.\n"
        "\u2022 Max 6 guests \u2014 if that changes, let us know.\n"
        "\u2022 Please keep shoes at the door. Dust out, mountain in.\n"
        "\u2022 If something breaks, tell us \u2014 we\u2019d rather fix it than discover it "
        "later.\n"
        "\u2022 Check-in after 3 PM. Checkout by 11 AM. Same-day turnovers are "
        "tight \u2014 thanks for leaving on time."
    )
    ws.merge_cells(f"A{current_row}:L{current_row}")
    c = ws[f"A{current_row}"]
    c.value = house_rules_copy
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True,
                             indent=2)
    ws.row_dimensions[current_row].height = 130
    card_body_fill(ws, current_row, current_row, ("A", "L"), border=True)
    current_row += 2  # spacer

    # Block 2: Check-in instructions
    card_header(ws, current_row, ("A", "L"),
                 "Block 2 \u2014 Check-in Instructions (paste into \u2018Check-in\u2019 field)")
    current_row += 1
    checkin_copy = (
        "Day-of check-in:\n\n"
        "1. The door uses a smart lock. Your code is 4321 (reset after "
        "checkout). \n"
        "2. Pull into the gravel drive on the right \u2014 there\u2019s parking for 2 "
        "cars. Don\u2019t block the mailbox.\n"
        "3. WiFi is SmokiesRidge_Guest / welcome2024. Password is also on "
        "the counter.\n"
        "4. If you arrive before 3 PM, the Local Guide tab has a great coffee "
        "shop 0.8 mi away.\n\n"
        "Host is Daniel \u2014 text me at (555) 555-0199 anytime. Seriously, we\u2019d "
        "rather hear from you than the morning-after emergency."
    )
    ws.merge_cells(f"A{current_row}:L{current_row}")
    c = ws[f"A{current_row}"]
    c.value = checkin_copy
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True,
                             indent=2)
    ws.row_dimensions[current_row].height = 180
    card_body_fill(ws, current_row, current_row, ("A", "L"), border=True)
    current_row += 2

    # Block 3: Before-you-leave reminder (for Airbnb messaging Day N-1)
    card_header(ws, current_row, ("A", "L"),
                 "Block 3 \u2014 Checkout Reminder (send Airbnb message Day N-1)")
    current_row += 1
    checkout_copy = (
        "Hope your stay\u2019s been great. Quick reminders for tomorrow:\n\n"
        "\u2022 Checkout is 11 AM sharp \u2014 we have a same-day turnover.\n"
        "\u2022 Leave linens in the hallway laundry basket.\n"
        "\u2022 Run the dishwasher.\n"
        "\u2022 Take trash + recycling to the curb (Thursday) or dumpster on-site.\n"
        "\u2022 Thermostat to 72\u00b0F before you lock up.\n"
        "\u2022 Return the door code to 0000.\n\n"
        "Text me when you\u2019re on the road \u2014 I\u2019ll release your deposit faster."
    )
    ws.merge_cells(f"A{current_row}:L{current_row}")
    c = ws[f"A{current_row}"]
    c.value = checkout_copy
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True,
                             indent=2)
    ws.row_dimensions[current_row].height = 160
    card_body_fill(ws, current_row, current_row, ("A", "L"), border=True)
    current_row += 2

    # Footer
    footer_row = current_row
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for c in range(1, 13):
        ws.cell(row=footer_row, column=c).border = Border(top=gold_side)
        ws.cell(row=footer_row, column=c).fill = parchment
    pseudo_button(ws, f"A{footer_row}", f"F{footer_row + 1}",
                   "\u2190 Back: Launch", "'Launch'!A1",
                   variant="secondary")
    pseudo_button(ws, f"G{footer_row}", f"L{footer_row + 1}",
                   "Host Notes (private) \u2192", "'\u00d7 Host Notes'!A1",
                   variant="secondary")
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[footer_row + 1].height = 22

    ws.print_area = f"A1:L{footer_row + 2}"


def build_host_notes_tab(wb, variant):
    """Tab 12 — × Host Notes. Private host-only, quarantined."""
    s = SAMPLE["Host"] if variant == "demo" else {}
    wb.create_sheet(TAB_NAMES[12])
    ws = wb["× Host Notes"]
    ws.sheet_properties.tabColor = COLOR_SECONDARY  # Clay Rose
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Rows 1-5: big red warning block
    red_tint_fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
    for r in range(1, 6):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = red_tint_fill

    ws.merge_cells("A1:L2")
    c = ws["A1"]
    c.value = "⚠  PRIVATE — HIDE BEFORE SHARING"
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        top=Side(style="medium", color=COLOR_ERROR),
        bottom=Side(style="medium", color=COLOR_ERROR),
        left=Side(style="medium", color=COLOR_ERROR),
        right=Side(style="medium", color=COLOR_ERROR),
    )
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A3:L3")
    c = ws["A3"]
    c.value = ("Right-click this tab → 'Hide' before saving the workbook as "
               "a PDF or sharing with guests.")
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = ("This tab's Print Area is deliberately set to A1:A1 so if you "
               "accidentally print it, you get a near-empty page, not your "
               "private notes.")
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 18

    ws.freeze_panes = "A6"

    # Cards
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    current_row = 7
    for hdr, rows in [
        ("Cleaners & Handyman", [
            ("Cleaner name:", s.get("cleaner_name", "")),
            ("Cleaner phone:", s.get("cleaner_phone", "")),
            ("Handyman:", s.get("handyman", "")),
            ("Plumber:", s.get("plumber", "")),
        ]),
        ("Property Systems", [
            ("Pool/hot tub service:", s.get("pool_service", "")),
            ("Insurance policy #:", s.get("insurance", "")),
            ("WiFi admin password (router):", s.get("wifi_admin", "")),
            ("Smart lock master code:", s.get("smart_lock_master", "")),
            ("Passcodes for safes/boxes:", s.get("safe_codes", "")),
        ]),
        ("Private Notes (never share)", [
            ("Things to NOT tell the guest:", s.get("private_notes", "")),
        ]),
    ]:
        card_header(ws, current_row, ("A", "L"), hdr)
        current_row += 1
        body_start = current_row
        for label, value in rows:
            ws.merge_cells(f"A{current_row}:C{current_row}")
            lc = ws[f"A{current_row}"]
            lc.value = label
            lc.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
            lc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            ws.merge_cells(f"D{current_row}:L{current_row}")
            ic = ws[f"D{current_row}"]
            ic.value = value
            apply_style(ic, input_cell_style())
            ic.alignment = Alignment(horizontal="left", vertical="center",
                                      wrap_text=True, indent=1)
            ws.row_dimensions[current_row].height = (36 if label.endswith(":")
                and len(str(value)) > 40 else 24)
            current_row += 1
        body_end = current_row - 1
        card_body_fill(ws, body_start, body_end, ("A", "L"), border=True)
        for c in range(1, 13):
            ws.cell(row=current_row, column=c).fill = parchment
        ws.row_dimensions[current_row].height = 12
        current_row += 1

    # Final reminder row
    ws.merge_cells(f"A{current_row + 1}:L{current_row + 2}")
    c = ws[f"A{current_row + 1}"]
    c.value = ("Right-click this tab → Hide before saving or sharing. "
               "Excel: RIGHT-click tab → Hide. Google Sheets: right-click → Hide sheet.")
    c.font = Font(name=FONT_HEAD, size=12, bold=True, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = red_tint_fill
    ws.row_dimensions[current_row + 1].height = 20
    ws.row_dimensions[current_row + 2].height = 20

    # Print area: A1:A1 ONLY. Protects against accidental print.
    ws.print_area = "A1:A1"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER



def build_workbook(out_path, variant):
    """Build a workbook with all 12 tabs.

    Args:
        out_path: pathlib.Path where the .xlsx will be saved
        variant: "demo" (filled with SAMPLE) or "blank" (all inputs empty)
    """
    assert variant in ("demo", "blank"), f"Unknown variant: {variant}"

    wb = Workbook()
    build_start_tab(wb, variant)
    build_property_tab(wb, variant)
    build_arrival_tab(wb, variant)
    build_wifi_tab(wb, variant)
    build_rules_tab(wb, variant)
    build_local_tab(wb, variant)
    build_trash_tab(wb, variant)
    build_departure_tab(wb, variant)
    build_safety_disclosures_tab(wb, variant)
    build_emergency_tab(wb, variant)
    build_launch_tab(wb, variant)
    build_bonus_tab(wb, variant)
    build_host_notes_tab(wb, variant)

    suffix = " (DEMO)" if variant == "demo" else ""
    wb.properties.title = f"Airbnb Welcome Book{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = (
        "Guest-facing welcome book tool for Airbnb/VRBO hosts — wizard UI, "
        "card-grouped inputs, live 3-page PDF preview."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def export_demo_json():
    """Export SAMPLE constants to a JSON file the renderer loads for demo mode.

    Schema mirrors the data contract §5 of the v2.1 spec — keyed by tab name,
    value is a dict mapping cell addresses (e.g. "B8") to the sample value.
    Local Guide is a list (table layout, not key/value).

    All addresses use the v2.1 flattened layout: first input at B8 per tab.
    Cell addresses verified against the actual DEMO xlsx output:
      - Departure: custom_tasks lands at B16 (after formula B9, 4 checklist
        rows B10-B13, and 2 static checklist rows). B14-B15 are static.
      - Emergency: poison control hardcoded at B14, host phone formula at
        B17 (excluded — renderer reads Property!B10 directly).
    """
    out = {
        "Property": {
            "B8":  SAMPLE["Property"]["name"],
            "B9":  SAMPLE["Property"]["host"],
            "B10": SAMPLE["Property"]["host_phone"],
            "B11": SAMPLE["Property"]["check_in"],
            "B12": SAMPLE["Property"]["check_out"],
            "B13": SAMPLE["Property"]["property_type"],
            "B14": SAMPLE["Property"]["max_guests"],
            "B15": SAMPLE["Property"]["address"],
        },
        "Arrival": {
            "B8":  SAMPLE["Arrival"]["address"],
            "B9":  SAMPLE["Arrival"]["entry_method"],
            "B10": SAMPLE["Arrival"]["door_code"],
            "B11": SAMPLE["Arrival"]["parking"],
            "B12": SAMPLE["Arrival"]["route"],
            "B13": SAMPLE["Arrival"]["arrival_window"],
            "B14": SAMPLE["Arrival"]["early_option"],
        },
        "WiFi + Tech": {
            "B8":  SAMPLE["WiFi"]["ssid"],
            "B9":  SAMPLE["WiFi"]["password"],
            "B10": SAMPLE["WiFi"]["backup_ssid"],
            "B11": SAMPLE["WiFi"]["tv_streaming"],
            "B12": SAMPLE["WiFi"]["smart_lock_note"],
            "B13": SAMPLE["WiFi"]["thermostat"],
            "B14": SAMPLE["WiFi"]["tv_controls"],
            "B15": SAMPLE["WiFi"]["wifi_support"],
        },
        "House Rules": {
            "B8":  SAMPLE["Rules"]["quiet_hours"],
            "B9":  SAMPLE["Rules"]["max_guests"],
            "B10": SAMPLE["Rules"]["smoking"],
            "B11": SAMPLE["Rules"]["pets"],
            "B12": SAMPLE["Rules"]["events"],
            "B13": SAMPLE["Rules"]["shoes"],
            "B14": SAMPLE["Rules"]["custom_rules"],
        },
        "Local Guide": [
            {"cat": cat, "name": name, "dist": dist, "phone": phone,
             "notes": notes}
            for (cat, name, dist, phone, notes) in SAMPLE["Local"]
        ],
        "Trash": {
            "B8":  SAMPLE["Trash"]["pickup_day"],
            "B9":  SAMPLE["Trash"]["bin_location"],
            "B10": SAMPLE["Trash"]["recycling_accepted"],
            "B11": SAMPLE["Trash"]["sorting_rules"],
            "B12": SAMPLE["Trash"]["pickup_location"],
            "B13": SAMPLE["Trash"]["thermostat_range"],
            "B14": SAMPLE["Trash"]["power_outage"],
        },
        "Departure": {
            "B8":  SAMPLE["Departure"]["checkout_time"],
            # B9 is a derived formula in the xlsx (checkout day).
            # For demo JSON we use a static plausible value matching the
            # sample check_in date (2026-05-10 → check_out 2026-05-15).
            "B9":  "Friday, May 15",
            "B10": SAMPLE["Departure"]["linen_location"],
            "B11": SAMPLE["Departure"]["trash_spot"],
            "B12": SAMPLE["Departure"]["thermostat_setting"],
            "B13": SAMPLE["Departure"]["key_return"],
            # B14-B15 are static checklist rows ("Run the dishwasher",
            # "Lock all doors + windows") — no input data.
            "B16": SAMPLE["Departure"]["custom_tasks"],
        },
        "Safety & Disclosures": {
            "B8":  SAMPLE["Safety"]["recording_devices"],
            "B9":  SAMPLE["Safety"]["alarm_locations"],
            "B10": SAMPLE["Safety"]["extinguisher_location"],
            "B11": SAMPLE["Safety"]["evacuation_notes"],
            "B12": SAMPLE["Safety"]["hazards"],
            "B13": SAMPLE["Safety"]["backup_contact"],
        },
        "Emergency": {
            "B8":  SAMPLE["Emergency"]["hospital_name"],
            "B9":  SAMPLE["Emergency"]["hospital_phone"],
            "B10": SAMPLE["Emergency"]["hospital_address"],
            "B11": SAMPLE["Emergency"]["urgent_care_name"],
            "B12": SAMPLE["Emergency"]["urgent_care_phone"],
            "B13": SAMPLE["Emergency"]["police_non_emergency"],
            "B14": "1-800-222-1222",  # poison control, hardcoded in xlsx
            "B15": SAMPLE["Emergency"]["vet"],
            "B16": SAMPLE["Emergency"]["utility"],
            # B17 is a formula (=Property!B10) — renderer reads host phone
            # from Property!B10 directly, so we don't include it here.
        },
    }

    DEMO_JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(DEMO_JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"Wrote demo JSON: {DEMO_JSON_OUT}")


def main():
    build_workbook(DEMO_OUT, variant="demo")
    build_workbook(BLANK_OUT, variant="blank")
    export_demo_json()


if __name__ == "__main__":
    main()
