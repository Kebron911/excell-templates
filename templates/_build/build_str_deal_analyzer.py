"""Build ACQ-001 STR Deal Analyzer (v2.2 standard).

Wizard-mode tool — single property under consideration, one fill, one
verdict. Stress test + Y1/Stabilized toggle + CPA Summary differentiate
Full ($47) from Lite ($27).

Generates three files:
  templates/_masters/ACQ-001-str-deal-analyzer-DEMO.xlsx     (Full + sample data)
  templates/_masters/ACQ-001-str-deal-analyzer-BLANK.xlsx    (Full + empty)
  templates/_lite/ACQ-001-str-deal-analyzer-lite.xlsx        (Lite, 5 tabs, sample data)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_WHITE,
)

SKU = "ACQ-001"
NAME = "str-deal-analyzer"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
LITE_OUT = BASE / "_lite" / f"{SKU}-{NAME}-lite.xlsx"

# ---------------------------------------------------------------------------
# Sample data (per brief QA section — 3-bed Gatlinburg cabin)
# ---------------------------------------------------------------------------

SAMPLE = {
    # Property
    "property_name":   "Smokies Ridge Cabin",
    "address":         "123 Mountain Lane",
    "city_state_zip":  "Gatlinburg, TN 37738",
    "prop_type":       "Cabin",
    "beds":            3,
    "baths":           2,
    "sqft":            1850,
    "year_built":      2014,
    "lot_type":        "Rural",
    "amenity_hot_tub": "Yes",
    "amenity_view":    "Yes",

    # Market
    "adr":             285,
    "comp_adr":        268,
    "stabilized_occ":  0.68,
    "y1_occ":          0.44,    # ~65% of stabilized per brief default
    "los":             3.4,
    "platform_pct":    0.15,

    # Costs (annual)
    "property_tax":    5800,
    "insurance":       3400,
    "hoa":             0,
    "utilities":       2400,
    "internet":        960,
    "software":        480,
    "marketing":       1200,
    "cleaning_paid":   185,    # per turnover
    "supplies":        35,     # per turnover
    "mgmt_pct":        0.0,
    "maint_pct":       0.05,
    "capex_pct":       0.03,

    # Finance
    "purchase":        485000,
    "down_pct":        0.25,
    "rate":            0.0700,
    "term_years":      30,
    "closing_costs":   9000,
    "rehab":           35000,
    "furnishing":      25000,
}

PROPERTY_TYPES = ["Single-family", "Cabin", "Condo", "Multi-unit", "Glamping", "Other"]
LOT_TYPES = ["Urban", "Suburban", "Rural"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None."""
    return demo_value if variant == "demo" else None


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _section_band(ws, row, label):
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24


def _input_row(ws, row, label, value, fmt=None, note="", input_col="C"):
    ws.cell(row=row, column=2, value=label).font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    col_idx = column_index_from_string(input_col)
    cell = ws.cell(row=row, column=col_idx, value=value)
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt

    if note:
        # Note in cols E-L
        ws.merge_cells(f"E{row}:L{row}")
        c = ws[f"E{row}"]
        c.value = note
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[row].height = 18


def _output_row(ws, row, label, formula, fmt, emphasize=False):
    label_font = Font(
        name=FONT_HEAD if emphasize else FONT_BODY,
        size=12 if emphasize else 11,
        bold=True,
        color=COLOR_PRIMARY if emphasize else COLOR_TEXT,
    )
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )

    cell = ws.cell(row=row, column=3, value=formula)
    apply_style(cell, formula_cell_style())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    if emphasize:
        cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_ACCENT)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)

    ws.row_dimensions[row].height = 22 if emphasize else 18


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant, is_lite):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    title = "STR Deal Analyzer"
    if is_lite:
        title += " — Lite"
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = title
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "Yes-or-no on a property in 10 minutes."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    label_suffix = "LITE" if is_lite else variant.upper()
    c.value = f"{SKU} · v1.0 · {label_suffix}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: Property name + verdict block (rows 10-22)
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 23):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    ws.merge_cells("A11:L11")
    c = ws["A11"]
    c.value = '=IF(Property!C7="","(enter property on Property tab)",Property!C7)'
    c.font = Font(name=FONT_HEAD, size=18, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 28

    # VERDICT label
    ws.merge_cells("A13:L13")
    c = ws["A13"]
    c.value = "VERDICT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 18

    # Verdict readout — pulls from Verdict tab merged cell A17
    ws.merge_cells("A14:L15")
    c = ws["A14"]
    c.value = "=Verdict!A17"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 24

    # Y1 + Stabilized side-by-side cash flow (rows 17-19)
    ws.merge_cells("A17:F17")
    c = ws["A17"]
    c.value = "Y1 ANNUAL CASH FLOW"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G17:L17")
    c = ws["G17"]
    c.value = "STABILIZED CASH FLOW"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 18

    ws.merge_cells("A18:F19")
    c = ws["A18"]
    c.value = "=Verdict!C10"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'

    ws.merge_cells("G18:L19")
    c = ws["G18"]
    c.value = "=Verdict!D10"
    c.font = Font(name=FONT_HEAD, size=24, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.row_dimensions[18].height = 28
    ws.row_dimensions[19].height = 18

    # Stress test verdict (Full only)
    if not is_lite:
        ws.merge_cells("A21:L21")
        c = ws["A21"]
        c.value = "=Verdict!A22"
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[21].height = 24

    # ZONE 3: Pseudo-button nav (rows 24-26)
    pseudo_button(ws, "A24", "C26", "Property",
                  "'Property'!A1", variant="primary")
    pseudo_button(ws, "D24", "F26", "Market",
                  "'Market'!A1", variant="primary")
    pseudo_button(ws, "G24", "I26", "Costs",
                  "'Costs'!A1", variant="primary")
    pseudo_button(ws, "J24", "L26", "Finance",
                  "'Finance'!A1", variant="primary")
    for r in range(24, 27):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A28", "F30", "→  Verdict + Stress Test",
                  "'Verdict'!A1", variant="accent")
    if not is_lite:
        pseudo_button(ws, "G28", "L30", "CPA / Partner Summary",
                      "'CPA Summary'!A1", variant="secondary")
    else:
        # Upgrade prompt button for Lite
        pseudo_button(ws, "G28", "L30", "Upgrade to Full →",
                      "'Start'!A33", variant="accent",
                      external_link=f"https://{BRAND_DOMAIN}/deal-analyzer-full")
    for r in range(28, 31):
        ws.row_dimensions[r].height = 22

    # ZONE 4: Upgrade banner (row 33)
    if is_lite:
        cta = (
            "💡  Need stress testing + Y1/stabilized toggle + CPA-ready "
            f"summary? Upgrade to Full at {BRAND_DOMAIN}/deal-analyzer-full — $47."
        )
    else:
        cta = (
            "💡  Underwriting multiple properties? Get the 3-Property "
            f"Side-by-Side at {BRAND_DOMAIN} — $67, or grab the Acquisition "
            "Bundle for $147."
        )
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = cta
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {label_suffix} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_property_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Property")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28), ("C", 22),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Property",
                        prev_tab="Start", next_tab="Market")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 1 of 4 — what you're underwriting"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "PROPERTY")

    fields = [
        (7,  "Property name:",     _val(variant, SAMPLE["property_name"]),   None,           ""),
        (8,  "Street address:",    _val(variant, SAMPLE["address"]),         None,           ""),
        (9,  "City, State, Zip:",  _val(variant, SAMPLE["city_state_zip"]),  None,           ""),
        (10, "Property type:",     _val(variant, SAMPLE["prop_type"]),       None,           "Dropdown"),
        (11, "Bedrooms:",          _val(variant, SAMPLE["beds"]),            "0",            ""),
        (12, "Bathrooms:",         _val(variant, SAMPLE["baths"]),           "0.0",          ""),
        (13, "Square feet:",       _val(variant, SAMPLE["sqft"]),            "#,##0",        ""),
        (14, "Year built:",        _val(variant, SAMPLE["year_built"]),      "0",            ""),
        (15, "Lot type:",          _val(variant, SAMPLE["lot_type"]),        None,           "Dropdown — drives ADR adjustment intuition"),
        (17, "Hot tub?",           _val(variant, SAMPLE["amenity_hot_tub"]), None,           "Yes/No"),
        (18, "View premium?",      _val(variant, SAMPLE["amenity_view"]),    None,           "Yes/No"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    add_dropdown(ws, "C10", f'"{",".join(PROPERTY_TYPES)}"')
    add_dropdown(ws, "C15", f'"{",".join(LOT_TYPES)}"')
    add_dropdown(ws, "C17", '"Yes,No"')
    add_dropdown(ws, "C18", '"Yes,No"')

    brand_footer(ws, 21,
                 version_line=f"{SKU} · Property")


def build_market_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Market")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Market",
                        prev_tab="Property", next_tab="Costs")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 2 of 4 — ADR, occupancy, comp benchmarks"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "MARKET ASSUMPTIONS")

    fields = [
        (7,  "ADR — Avg nightly rate ($):",       _val(variant, SAMPLE["adr"]),            '"$"#,##0', "Paste from AirDNA / verify with comps"),
        (8,  "Comp ADR (3-comp average, $):",      _val(variant, SAMPLE["comp_adr"]),       '"$"#,##0', "Sanity check on your ADR assumption"),
        (9,  "Stabilized occupancy (%):",          _val(variant, SAMPLE["stabilized_occ"]), "0.0%",    "Year 2-3 expected"),
        (10, "Year 1 occupancy (%):",              _val(variant, SAMPLE["y1_occ"]),         "0.0%",    "Default ≈ 65% of stabilized (BiggerPockets consensus)"),
        (11, "Avg length of stay (nights):",       _val(variant, SAMPLE["los"]),            "0.0",     "1.5-7 typical"),
        (12, "Platform commission (%):",            _val(variant, SAMPLE["platform_pct"]),   "0.0%",    "Airbnb 15%, VRBO 8%, direct 0-3%"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    # AirDNA paste callout rows 14-17
    ws.merge_cells("A14:L14")
    c = ws["A14"]
    c.value = "📋  PASTE FROM AIRDNA / RABBU / KEYDATA"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[14].height = 20

    ws.merge_cells("A15:L17")
    c = ws["A15"]
    c.value = (
        "Pull market RevPAR, ADR, and occupancy from AirDNA Markets (or Rabbu / "
        "KeyData free comp). Look for the same beds/baths/lot type. Then haircut: "
        "if AirDNA shows median ADR, use the 25th-percentile ADR for Y1 here. "
        "AirDNA optimism is the #1 reason new STR investors get burned — the "
        "Verdict tab's stress test lives at AirDNA's 25th-percentile assumptions."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    for r in range(15, 18):
        ws.row_dimensions[r].height = 18

    brand_footer(ws, 20,
                 version_line=f"{SKU} · Market")


def build_costs_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Costs")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Costs",
                        prev_tab="Market", next_tab="Finance")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 3 of 4 — operating costs (mirrors TAX-002 / FIN-002 structure)"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "ANNUAL FIXED COSTS")
    fixed_fields = [
        (7,  "Property tax (annual, $):",   _val(variant, SAMPLE["property_tax"]), '"$"#,##0', ""),
        (8,  "Insurance (annual, $):",      _val(variant, SAMPLE["insurance"]),    '"$"#,##0', "STR rider included"),
        (9,  "HOA / Condo (annual, $):",    _val(variant, SAMPLE["hoa"]),          '"$"#,##0', ""),
        (10, "Utilities (annual, $):",      _val(variant, SAMPLE["utilities"]),    '"$"#,##0', "Electric/gas/water/trash combined"),
        (11, "Internet (annual, $):",       _val(variant, SAMPLE["internet"]),     '"$"#,##0', ""),
        (12, "Software / PMS (annual, $):", _val(variant, SAMPLE["software"]),     '"$"#,##0', ""),
        (13, "Marketing (annual, $):",      _val(variant, SAMPLE["marketing"]),    '"$"#,##0', "Photos refresh, listing fees"),
    ]
    for row, label, value, fmt, note in fixed_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 15, "PER-TURNOVER VARIABLE")
    var_fields = [
        (16, "Cleaning paid ($/turnover):",  _val(variant, SAMPLE["cleaning_paid"]), '"$"#,##0', ""),
        (17, "Supplies ($/turnover):",       _val(variant, SAMPLE["supplies"]),      '"$"#,##0', ""),
    ]
    for row, label, value, fmt, note in var_fields:
        _input_row(ws, row, label, value, fmt, note)

    _section_band(ws, 19, "RESERVES & MGMT (% OF REVENUE)")
    pct_fields = [
        (20, "Management fee (%):",          _val(variant, SAMPLE["mgmt_pct"]),  "0.0%", "0% if self-managed; 15-25% with a PM"),
        (21, "Maintenance reserve (%):",     _val(variant, SAMPLE["maint_pct"]), "0.0%", "5% typical"),
        (22, "CapEx reserve (%):",           _val(variant, SAMPLE["capex_pct"]), "0.0%", "3% typical"),
    ]
    for row, label, value, fmt, note in pct_fields:
        _input_row(ws, row, label, value, fmt, note)

    brand_footer(ws, 25,
                 version_line=f"{SKU} · Costs")


def build_finance_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Finance")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 38), ("C", 18),
        ("D", 4), ("E", 38),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Finance",
                        prev_tab="Costs", next_tab="Verdict")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Section 4 of 4 — purchase, loan, closing, rehab, furnishing"
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    _section_band(ws, 6, "FINANCING")

    fields = [
        (7,  "Purchase price ($):",      _val(variant, SAMPLE["purchase"]),       '"$"#,##0', ""),
        (8,  "Down payment (%):",        _val(variant, SAMPLE["down_pct"]),       "0.0%",    "20-25% typical for STR"),
        (9,  "Interest rate (%):",       _val(variant, SAMPLE["rate"]),           "0.000%",  ""),
        (10, "Term (years):",            _val(variant, SAMPLE["term_years"]),     "0",       "30 typical"),
        (11, "Closing costs ($):",       _val(variant, SAMPLE["closing_costs"]),  '"$"#,##0', "2-3% of purchase typical"),
        (12, "Rehab budget ($):",        _val(variant, SAMPLE["rehab"]),          '"$"#,##0', ""),
        (13, "Furnishing budget ($):",   _val(variant, SAMPLE["furnishing"]),     '"$"#,##0', "Cross-reference ACQ-002"),
    ]
    for row, label, value, fmt, note in fields:
        _input_row(ws, row, label, value, fmt, note)

    # Derived values
    _section_band(ws, 15, "DERIVED")
    _output_row(ws, 16, "Loan amount:",
                "=C7*(1-C8)", '"$"#,##0')
    _output_row(ws, 17, "Monthly P&I:",
                "=IFERROR(PMT(C9/12,C10*12,-C16),0)", '"$"#,##0.00')
    _output_row(ws, 18, "Annual debt service (P&I):",
                "=C17*12", '"$"#,##0')
    _output_row(ws, 19, "Annual mortgage interest (Y1 approx):",
                "=C16*C9", '"$"#,##0')
    _output_row(ws, 20, "Total cash to close:",
                "=C7*C8+C11+C12+C13", '"$"#,##0', emphasize=True)

    brand_footer(ws, 23,
                 version_line=f"{SKU} · Finance")


def build_verdict_tab(wb, variant, is_lite):
    ws = wb.create_sheet("Verdict")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 16),
        ("D", 16), ("E", 16),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "Verdict",
                        prev_tab="Finance",
                        next_tab="CPA Summary" if not is_lite else None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Toggle Y1 vs Stabilized below. The Stress Test column compares your "
        "deal against AirDNA's pessimistic case." if not is_lite else
        "Y1 quick verdict. Upgrade to Full for stress testing + stabilized comparison."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # ---- Headers row 6 ----
    headers = [("B6", "Metric"), ("C6", "Year 1"), ("D6", "Stabilized")]
    if not is_lite:
        headers.append(("E6", "Stress Test"))
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        apply_style(cell, header_row_style())
    ws.row_dimensions[6].height = 24

    # ---- Common formula building blocks (use named refs for clarity) ----
    # The formulas below reference Market/Costs/Finance directly.
    # Y1 occupancy = Market!C10; Stabilized = Market!C9
    # Stress = Market!C9 * 0.85 AND ADR -10% AND expenses +20%

    # Revenue formulas
    # Booked nights = occ × 365 (assume 365 available)
    # Gross revenue = booked_nights × ADR + booked_turnovers × cleaning_collected
    #   NOTE: Deal Analyzer uses the brief's simplified revenue model — ADR × occ × 365
    #   plus cleaning fees. We don't know cleaning_collected here; assume cleaning_paid
    #   passed through (cleaning fee = cleaning paid for simplicity in this MVP).

    # Y1 revenue (net of platform commission Market!C12)
    rev_y1 = "((Market!C7)*(Market!C10)*365*(1-Market!C12))"
    # Stabilized revenue
    rev_stab = "((Market!C7)*(Market!C9)*365*(1-Market!C12))"
    # Stress: ADR -10%, occupancy = stabilized × 0.85, expenses +20%
    rev_stress = "((Market!C7*0.9)*(Market!C9*0.85)*365*(1-Market!C12))"

    # Operating expenses (annual fixed + variable per turnover × turnovers + reserves)
    # Turnovers = booked_nights ÷ LOS
    # Variable annual = turnovers × (cleaning_paid + supplies)
    # Fixed annual = sum of Costs!C7:C13
    fixed_costs = "(SUM(Costs!C7:C13))"
    variable_per_turnover = "(Costs!C16+Costs!C17)"
    # Mgmt + maint + capex as % of revenue
    pct_costs_y1 = f"({rev_y1}*(Costs!C20+Costs!C21+Costs!C22))"
    pct_costs_stab = f"({rev_stab}*(Costs!C20+Costs!C21+Costs!C22))"
    pct_costs_stress = f"({rev_stress}*(Costs!C20+Costs!C21+Costs!C22))"
    # Variable cost annual
    var_cost_y1 = f"((Market!C10*365/Market!C11)*{variable_per_turnover})"
    var_cost_stab = f"((Market!C9*365/Market!C11)*{variable_per_turnover})"
    var_cost_stress = f"((Market!C9*0.85*365/Market!C11)*{variable_per_turnover}*1.2)"

    # NOI = revenue - opex
    noi_y1 = f"({rev_y1}-{fixed_costs}-{pct_costs_y1}-{var_cost_y1})"
    noi_stab = f"({rev_stab}-{fixed_costs}-{pct_costs_stab}-{var_cost_stab})"
    noi_stress = f"({rev_stress}-{fixed_costs}*1.2-{pct_costs_stress}-{var_cost_stress})"

    # Debt service
    debt_service = "(Finance!C18)"

    # Cash flow = NOI - debt service (for Y1, debt service is full P&I)
    cf_y1 = f"({noi_y1}-{debt_service})"
    cf_stab = f"({noi_stab}-{debt_service})"
    cf_stress = f"({noi_stress}-{debt_service})"

    # Cash invested = down + closing + rehab + furnishing
    cash_invested = "(Finance!C7*Finance!C8+Finance!C11+Finance!C12+Finance!C13)"

    # Cash-on-cash %
    coc_y1 = f"({cf_y1}/{cash_invested})"
    coc_stab = f"({cf_stab}/{cash_invested})"
    coc_stress = f"({cf_stress}/{cash_invested})"

    # DSCR = NOI / debt service
    dscr_y1 = f"({noi_y1}/{debt_service})"
    dscr_stab = f"({noi_stab}/{debt_service})"
    dscr_stress = f"({noi_stress}/{debt_service})"

    # Cap rate = NOI / purchase price
    cap_y1 = f"({noi_y1}/Finance!C7)"
    cap_stab = f"({noi_stab}/Finance!C7)"

    # ---- Metric rows 7-15 ----
    metric_rows = [
        ("Gross revenue:",       rev_y1,    rev_stab,    rev_stress,    '"$"#,##0'),
        ("NOI (operating profit):", noi_y1,  noi_stab,    noi_stress,    '"$"#,##0'),
        ("Annual debt service:", debt_service, debt_service, debt_service, '"$"#,##0'),
        ("ANNUAL CASH FLOW:",    cf_y1,     cf_stab,     cf_stress,     '"$"#,##0'),
        ("Cash invested:",       cash_invested, cash_invested, cash_invested, '"$"#,##0'),
        ("Cash-on-cash:",        coc_y1,    coc_stab,    coc_stress,    "0.0%"),
        ("DSCR:",                dscr_y1,   dscr_stab,   dscr_stress,   "0.00"),
        ("Cap rate:",            cap_y1,    cap_stab,    cap_stab,      "0.0%"),
    ]
    for i, (label, fy1, fstab, fstress, fmt) in enumerate(metric_rows):
        r = 7 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11,
            bold=(label in ("ANNUAL CASH FLOW:",)),
            color=COLOR_TEXT,
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )

        cell = ws.cell(row=r, column=3, value=f"={fy1}")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        if label == "ANNUAL CASH FLOW:":
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)

        cell = ws.cell(row=r, column=4, value=f"={fstab}")
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = fmt
        if label == "ANNUAL CASH FLOW:":
            cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_ACCENT)

        if not is_lite:
            cell = ws.cell(row=r, column=5, value=f"={fstress}")
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = fmt
            if label == "ANNUAL CASH FLOW:":
                cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_SECONDARY)

        ws.row_dimensions[r].height = 20

    # ---- VERDICT formula row 17 ----
    # Verdict logic per brief:
    # ✅ DEAL — stabilized COC > 8% AND DSCR > 1.25
    # ⚠ MARGINAL — stabilized COC 4-8% OR DSCR 1.05-1.25
    # ❌ PASS — stabilized COC < 4% OR DSCR < 1.05
    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = (
        f'=IF(AND({coc_stab}>0.08,{dscr_stab}>1.25),"✅  DEAL",'
        f'IF(OR({coc_stab}<0.04,{dscr_stab}<1.05),"❌  PASS",'
        f'"⚠  MARGINAL"))'
    )
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[17].height = 36

    # The Start tab references Verdict!A17 (merged-cell verdict string),
    # Verdict!C10 (Y1 cash flow), and Verdict!D10 (stabilized cash flow).

    # ---- Stress test summary (Full only) row 22 ----
    if not is_lite:
        ws.merge_cells("A22:L22")
        c = ws["A22"]
        c.value = (
            f'=IF({cf_stress}>0,"✅  Survives stress test ($"&TEXT({cf_stress},"#,##0")&" CF)",'
            f'IF({cf_stress}>-5000,"⚠  Marginal under stress (−$"&TEXT(ABS({cf_stress}),"#,##0")&")",'
            f'"❌  Fails under stress ($"&TEXT({cf_stress},"#,##0")&")"))'
        )
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        ws.row_dimensions[22].height = 28

        # Stress test methodology footnote
        ws.merge_cells("A24:L26")
        c = ws["A24"]
        c.value = (
            "Stress test method: ADR -10%, occupancy haircut to 85% of stabilized, "
            "operating expenses +20%. Roughly equivalent to AirDNA's 25th-percentile "
            "projection — what real underwriters use. If your deal fails this stress, "
            "you're underwriting too aggressively or the property is marginal."
        )
        c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in range(24, 27):
            ws.row_dimensions[r].height = 16

    brand_footer(ws, 28,
                 version_line=f"{SKU} · Verdict")


def build_cpa_summary_tab(wb):
    """Sheet 7 — CPA / Partner 1-page printable summary (Full only)."""
    ws = wb.create_sheet("CPA Summary")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT

    set_col_widths(ws, [
        ("A", 4), ("B", 36), ("C", 18),
        ("D", 18), ("E", 18),
        ("F", 8), ("G", 8), ("H", 8), ("I", 8),
        ("J", 8), ("K", 8), ("L", 8),
    ])

    compact_header_band(ws, "CPA / Partner Summary",
                        prev_tab="Verdict", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill

    # Header
    ws.merge_cells("A4:L5")
    c = ws["A4"]
    c.value = "Deal Underwriting Summary"
    c.font = Font(name=FONT_HEAD, size=20, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 16

    # Property name + analysis date
    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = '=Property!C7&"   ·   "&Property!C9'
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    ws.merge_cells("A8:L8")
    c = ws["A8"]
    c.value = '="Analysis date: "&TEXT(TODAY(),"yyyy-mm-dd")'
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 16

    # KEY TERMS section
    _section_band(ws, 10, "KEY TERMS")

    terms = [
        ("Purchase price:",  "=Finance!C7",       '"$"#,##0'),
        ("Down payment:",    "=Finance!C7*Finance!C8", '"$"#,##0'),
        ("Loan amount:",     "=Finance!C16",      '"$"#,##0'),
        ("Interest rate:",   "=Finance!C9",       "0.000%"),
        ("Term:",            '=Finance!C10&" years"', None),
        ("Total cash to close:", "=Finance!C20",  '"$"#,##0'),
    ]
    for i, (label, formula, fmt) in enumerate(terms):
        r = 11 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt
        ws.row_dimensions[r].height = 18

    # 5-YEAR PRO-FORMA
    _section_band(ws, 19, "5-YEAR PRO-FORMA  (3% appreciation, stabilized after Y2)")

    headers = ["", "Year 1", "Year 2", "Year 3", "Year 5"]
    pro_cols = ["B", "C", "D", "E"]
    ws.cell(row=20, column=2, value="").font = Font(size=11, bold=True)
    ws.cell(row=20, column=3, value="Year 1").font = Font(size=11, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=20, column=4, value="Year 2 Stab").font = Font(size=11, bold=True, color=COLOR_ACCENT)
    ws.cell(row=20, column=5, value="Year 5 Stab").font = Font(size=11, bold=True, color=COLOR_ACCENT)
    for col in (3, 4, 5):
        ws.cell(row=20, column=col).alignment = Alignment(horizontal="center")
    ws.row_dimensions[20].height = 22

    pro_rows = [
        ("Revenue (gross):",
         "=Market!C7*Market!C10*365",
         "=Market!C7*Market!C9*365",
         "=Market!C7*Market!C9*365*1.06"),  # 3%/yr × 2 yrs ≈ 6%
        ("Operating expenses:",
         "=SUM(Costs!C7:C13)+(Market!C7*Market!C10*365)*(Costs!C20+Costs!C21+Costs!C22)+(Market!C10*365/Market!C11)*(Costs!C16+Costs!C17)",
         "=SUM(Costs!C7:C13)+(Market!C7*Market!C9*365)*(Costs!C20+Costs!C21+Costs!C22)+(Market!C9*365/Market!C11)*(Costs!C16+Costs!C17)",
         "=SUM(Costs!C7:C13)*1.10+(Market!C7*Market!C9*365*1.06)*(Costs!C20+Costs!C21+Costs!C22)+(Market!C9*365/Market!C11)*(Costs!C16+Costs!C17)*1.10"),
        ("NOI:",
         "=C21-C22",
         "=D21-D22",
         "=E21-E22"),
        ("Debt service:",
         "=Finance!C18",
         "=Finance!C18",
         "=Finance!C18"),
        ("Cash flow:",
         "=C23-C24",
         "=D23-D24",
         "=E23-E24"),
        ("Cumulative cash:",
         "=C25",
         "=C25+D25",
         "=C25+D25+(D25*3)"),
    ]
    for i, (label, fy1, fy2, fy5) in enumerate(pro_rows):
        r = 21 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=(label in ("NOI:", "Cash flow:")),
            color=COLOR_TEXT,
        )
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        for col, formula in [(3, fy1), (4, fy2), (5, fy5)]:
            cell = ws.cell(row=r, column=col, value=formula)
            apply_style(cell, formula_cell_style())
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 18

    # VERDICT section
    _section_band(ws, 28, "VERDICT  +  STRESS TEST")

    ws.merge_cells("B29:E29")
    c = ws["B29"]
    c.value = "=Verdict!A17"
    c.font = Font(name=FONT_HEAD, size=18, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    ws.row_dimensions[29].height = 30

    ws.merge_cells("B30:E30")
    c = ws["B30"]
    c.value = "=Verdict!A22"
    c.font = Font(name=FONT_BODY, size=11, italic=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[30].height = 22

    # NOTES section (manual)
    _section_band(ws, 32, "UNDERWRITING NOTES")
    ws.merge_cells("B33:E37")
    c = ws["B33"]
    apply_style(c, input_cell_style())
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(33, 38):
        ws.row_dimensions[r].height = 18

    # Print area + page setup
    ws.print_area = "A1:E40"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant, is_lite):
    wb = Workbook()
    build_start_tab(wb, variant, is_lite)
    build_property_tab(wb, variant, is_lite)
    build_market_tab(wb, variant, is_lite)
    build_costs_tab(wb, variant, is_lite)
    build_finance_tab(wb, variant, is_lite)
    build_verdict_tab(wb, variant, is_lite)
    if not is_lite:
        build_cpa_summary_tab(wb)

    wb.properties.title = "STR Deal Analyzer — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Single-property STR underwriting workbook with stress test, "
        "Y1/stabilized toggle, and CPA/partner printable summary."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo", is_lite=False)
    build_workbook(BLANK_OUT, "blank", is_lite=False)
    build_workbook(LITE_OUT, "demo", is_lite=True)


if __name__ == "__main__":
    main()
