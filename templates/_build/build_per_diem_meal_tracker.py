"""Build TAX-007 Per-Diem Meal Deduction Tracker — operational log.

Implements templates/_briefs/TAX-007-per-diem-meal-tracker.md.

Generates templates/_masters/TAX-007-per-diem-meal-tracker-DEMO.xlsx and -BLANK.xlsx.

Tabs:
  0  Start             — operational hero + activity cards + primary CTA
  1  Meals Log         — one row per meal (Actual method)
  2  Per-Diem Log      — daily allowance (M&IE × days × city rate)
  3  Monthly Summary   — totals + breakdown
  4  Settings          — tax year · M&IE rates · classification

Usage:
    python build_per_diem_meal_tracker.py
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

from brand_config import (COLOR_PRIMARY, COLOR_ACCENT, COLOR_TEXT, COLOR_MUTED,
    COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_EMAIL,
    pseudo_button, compact_header_band, brand_footer, style_chart,
    set_col_widths, apply_style, input_cell_style, formula_cell_style,
    header_row_style,
    STATE_BAD_FILL,
    STATE_WARN_FILL,
)

SKU = "TAX-007"
NAME = "per-diem-meal-tracker"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"
VERSION_LINE = f"{SKU} · v2.3 · Free updates forever"
RATES_AS_OF = "2026-01-01"  # GSA publishes new M&IE rates each Oct


def _val(variant, demo_value):
    return demo_value if variant == "demo" else None

# 2026 GSA M&IE rates (illustrative defaults — editable in Settings)
M_IE_HIGH = 80   # high-cost cities (NYC, SF, etc.)
M_IE_STD = 59    # standard CONUS

PURPOSES = [
    "Property inspection meal",
    "Cleaner / staff meal",
    "Contractor walkthrough lunch",
    "CPA / attorney meeting meal",
    "Co-host coordination",
    "Owner/investor meeting",
    "Continuing education event",
    "Other (describe in Notes)",
]

DEDUCTIBLE_PERCENTS = [50, 80, 100, 0]

# Sample meals (Jan-Mar 2026)
MEALS_SAMPLES = [
    # (date, property, location, restaurant, attendees, purpose, amount, tip, receipt, method, dedpct)
    ("2026-01-08", "Smokies Ridge", "Gatlinburg, TN", "Mama's Cafe",       "Self + Jamie (cleaner)", "Cleaner / staff meal",         42, 8, "Y", "Actual", 50),
    ("2026-01-15", "Creek Side",    "Pigeon Forge, TN", "Pizza Inn",       "Self",                    "Property inspection meal",     14, 2, "Y", "Actual", 50),
    ("2026-01-22", "Smokies Ledge", "Knoxville, TN", "Aubrey's",           "Self + Smith CPA",        "CPA / attorney meeting meal", 64, 12, "Y", "Actual", 50),
    ("2026-02-02", "Lakehouse A",   "Townsend, TN", "Riverside Diner",     "Self + crew (3)",         "Cleaner / staff meal",        58, 10, "Y", "Actual", 100),  # employer-provided
    ("2026-02-10", "Smokies Ridge", "Sevierville, TN", "Cracker Barrel",   "Self",                    "Property inspection meal",    18, 3, "",  "Actual", 50),
    ("2026-02-14", "Creek Side",    "Pigeon Forge, TN", "Local Goat",      "Self + handyman",         "Contractor walkthrough lunch", 36, 7, "Y", "Actual", 50),
    ("2026-02-22", "Smokies Ridge", "Gatlinburg, TN", "The Park Grill",    "Self + 2 owners",         "Owner/investor meeting",      120, 24, "Y", "Actual", 50),
    ("2026-03-05", "Lakehouse A",   "Townsend, TN", "Burger Master",       "Self",                    "Property inspection meal",    12, 2, "Y", "Actual", 50),
    ("2026-03-10", "Smokies Ridge", "Gatlinburg, TN", "Mountain Lodge",    "Self + cleaner",          "Cleaner / staff meal",        38, 7, "Y", "Actual", 50),
    ("2026-03-18", "Creek Side",    "Pigeon Forge, TN", "Mellow Mushroom", "Self + handyman + plumber", "Contractor walkthrough lunch", 76, 14, "Y", "Actual", 50),
    ("2026-03-25", "Smokies Ridge", "Sevierville, TN", "Five Oaks BBQ",    "Self + Smith CPA",        "CPA / attorney meeting meal", 48, 9, "Y", "Actual", 50),
]

# Sample per-diem trips
PER_DIEM_SAMPLES = [
    # (date_start, location, full_days, partial_days, m_ie_rate)
    ("2026-02-08", "Nashville, TN (continuing-ed conf)",  2, 1, M_IE_STD),
    ("2026-03-15", "Atlanta, GA (STR Summit)",             2, 1, M_IE_HIGH),
]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def build_start_tab(wb, variant):
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # Hero
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A4:L4")
    c = ws["A4"]; c.value = "Per-Diem Meal Tracker"
    c.font = Font(name=FONT_HEAD, size=36, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 48
    ws.merge_cells("A5:L5")
    c = ws["A5"]; c.value = "Track meals deductibly. Don't lose another receipt."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A7:L7")
    c = ws["A7"]; c.value = f"{SKU} · v2.3 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # What this does
    parchment = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 17):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment
    ws.merge_cells("A10:L10")
    c = ws["A10"]; c.value = "WHAT THIS DOES"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[10].height = 20
    ws.merge_cells("A11:L15")
    c = ws["A11"]
    c.value = (
        "IRS Pub 463 requires 5 elements per meal: date, amount, business "
        "purpose, attendees, location. This log captures all 5 plus the "
        "deductible % (50% standard, 100% employer-provided, 0% entertainment). "
        "Two methods supported — Actual (track every meal) or Per-Diem (M&IE "
        "rate × travel days). Print the Monthly Summary at year-end for your CPA."
    )
    c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=2)

    # How to use
    qs_fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    for r in range(18, 25):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = qs_fill
    ws.merge_cells("A18:L18")
    c = ws["A18"]; c.value = "HOW TO USE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[18].height = 20
    quickstart_items = [
        "① Click below to open Meals Log — log one row per meal as it happens.",
        "② Pick the deductible % per row: 50% standard, 100% staff, 0% entertainment.",
        "③ Travel out of town overnight? Use Per-Diem Log instead — no receipts needed.",
        "④ Log within 7 days — auditors disallow records reconstructed at year-end.",
    ]
    for i, item in enumerate(quickstart_items):
        row = 19 + i
        ws.merge_cells(f"A{row}:L{row}")
        c = ws[f"A{row}"]
        c.value = item
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=2)
        ws.row_dimensions[row].height = 22

    # Primary CTA
    pseudo_button(ws, "A26", "L29", "→  ADD A NEW MEAL  (OPEN MEALS LOG)",
                   "'Meals Log'!A6", variant="primary")
    for r in range(26, 30):
        ws.row_dimensions[r].height = 22

    # Activity cards
    for r in range(31, 37):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment

    # Card 1 — Meals logged
    ws.merge_cells("A32:D32")
    c = ws["A32"]; c.value = "MEALS LOGGED"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A33:D34")
    c = ws["A33"]; c.value = "=COUNTA('Meals Log'!A6:A2005)"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A35:D35")
    c = ws["A35"]; c.value = "rows in log"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 2 — YTD deduction
    ws.merge_cells("E32:H32")
    c = ws["E32"]; c.value = "YTD DEDUCTION"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E33:H34")
    c = ws["E33"]; c.value = "='Monthly Summary'!C19"
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E35:H35")
    c = ws["E35"]; c.value = "post-% applied"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Card 3 — Last entry
    ws.merge_cells("I32:L32")
    c = ws["I32"]; c.value = "LAST ENTRY"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I33:L34")
    c = ws["I33"]
    c.value = '=IFERROR(TEXT(MAX(\'Meals Log\'!A6:A2005),"yyyy-mm-dd"),"—")'
    c.font = Font(name=FONT_HEAD, size=22, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("I35:L35")
    c = ws["I35"]; c.value = "most recent meal"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(32, 36):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 32 else existing.top,
                    bottom=gold_side if r == 35 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # Secondary nav
    pseudo_button(ws, "A39", "C40", "Meals Log",
                   "'Meals Log'!A1", variant="secondary")
    pseudo_button(ws, "D39", "F40", "Per-Diem Log",
                   "'Per-Diem Log'!A1", variant="secondary")
    pseudo_button(ws, "G39", "I40", "📄 Monthly Summary",
                   "'Monthly Summary'!A1", variant="accent")
    pseudo_button(ws, "J39", "L40", "Settings",
                   "'Settings'!A1", variant="secondary")
    ws.row_dimensions[39].height = 22
    ws.row_dimensions[40].height = 22

    # Audit-defense reminder
    ws.merge_cells("A42:L42")
    c = ws["A42"]
    c.value = (
        "⚠ Pub 463: meals require 5 elements per row — date, amount, "
        "business purpose, attendees, location. Receipts required for any "
        "single meal ≥ $75. Reconstructed records are the #1 reason meal "
        "deductions get disallowed."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[42].height = 32

    brand_footer(ws, 44, version_line=VERSION_LINE)

    ws.print_area = "A1:L46"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_meals_log_tab(wb, variant):
    ws = wb.create_sheet("Meals Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Meals Log",
                         prev_tab="Start", next_tab="Per-Diem Log")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "One row per meal — Pub 463 5-element rule (date · amount · purpose · attendees · location)."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    # Row 5: headers
    headers = [
        "Date", "Property", "Location (city)", "Restaurant",
        "Attendees", "Business Purpose", "Amount $", "Tip $",
        "Total $", "📎 Receipt", "Method", "Ded %", "Ded $",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 11), ("B", 16), ("C", 16), ("D", 18),
        ("E", 22), ("F", 22), ("G", 9), ("H", 7),
        ("I", 9), ("J", 9), ("K", 11), ("L", 7), ("M", 10),
    ])

    # Rows 6-2005 — DEMO populates samples; BLANK starts empty
    meals = MEALS_SAMPLES if variant == "demo" else []
    for i in range(6, 2006):
        sample_idx = i - 6
        sample = meals[sample_idx] if sample_idx < len(meals) else None

        if sample:
            (date_val, prop, location, restaurant, attendees, purpose,
             amount, tip, receipt, method, dedpct) = sample
            d = datetime.strptime(date_val, "%Y-%m-%d").date()
            a = ws.cell(row=i, column=1, value=d)
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"

            for col_idx, val in enumerate(
                [prop, location, restaurant, attendees, purpose], start=2
            ):
                c = ws.cell(row=i, column=col_idx, value=val)
                apply_style(c, input_cell_style())

            g = ws.cell(row=i, column=7, value=amount)
            apply_style(g, input_cell_style())
            g.number_format = '"$"#,##0.00'

            h = ws.cell(row=i, column=8, value=tip)
            apply_style(h, input_cell_style())
            h.number_format = '"$"#,##0.00'

            j = ws.cell(row=i, column=10, value=receipt if receipt else None)
            apply_style(j, input_cell_style())
            j.alignment = Alignment(horizontal="center", vertical="center")

            k = ws.cell(row=i, column=11, value=method)
            apply_style(k, input_cell_style())

            l = ws.cell(row=i, column=12, value=dedpct)
            apply_style(l, input_cell_style())
            l.number_format = "0"
            l.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Blank rows — apply input styles
            for col_idx in range(1, 9):
                c = ws.cell(row=i, column=col_idx)
                apply_style(c, input_cell_style())
                if col_idx == 1:
                    c.number_format = "yyyy-mm-dd"
                elif col_idx in (7, 8):
                    c.number_format = '"$"#,##0.00'
            for col_idx in [10, 11, 12]:
                c = ws.cell(row=i, column=col_idx)
                apply_style(c, input_cell_style())
                if col_idx == 12:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # I = Total $ (G+H)
        i_cell = ws.cell(row=i, column=9, value=f"=IFERROR(G{i}+H{i},0)")
        i_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(i_cell, formula_cell_style())

        # M = Ded $ (Total × Ded%/100)
        m_cell = ws.cell(row=i, column=13, value=f"=IFERROR(I{i}*L{i}/100,0)")
        m_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(m_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    # Dropdowns
    add_dropdown(ws, "F6:F2005", "=Settings!$E$11:$E$25")  # purposes
    add_dropdown(ws, "J6:J2005", '"Y,N"')                   # receipt
    add_dropdown(ws, "K6:K2005", '"Actual,Per-diem,Reimbursed"')
    add_dropdown(ws, "L6:L2005", '"50,80,100,0"')

    # Conditional formatting — red on entertainment (0%) rows; yellow on missing receipt for >$75
    ws.conditional_formatting.add(
        "L6:L2005",
        FormulaRule(
            formula=['L6=0'],
            fill=PatternFill("solid", fgColor=STATE_BAD_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )
    ws.conditional_formatting.add(
        "J6:J2005",
        FormulaRule(
            formula=['AND(I6>=75,J6<>"Y")'],
            fill=PatternFill("solid", fgColor=STATE_WARN_FILL),
            font=Font(bold=True, color=COLOR_ERROR),
        ),
    )

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_per_diem_log_tab(wb, variant):
    ws = wb.create_sheet("Per-Diem Log")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Per-Diem Log",
                         prev_tab="Meals Log", next_tab="Monthly Summary")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Daily M&IE allowance for overnight business travel (no receipts < $75)."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = ["Trip Start", "Destination",
               "# Full Days", "# Partial Days (75% rate)", "M&IE $/day",
               "Subtotal $", "Ded % (50% std)", "Ded $"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [
        ("A", 12), ("B", 32),
        ("C", 12), ("D", 22), ("E", 12),
        ("F", 14), ("G", 14), ("H", 12),
    ])

    trips = PER_DIEM_SAMPLES if variant == "demo" else []
    for i in range(6, 106):  # 100-row capacity
        sample_idx = i - 6
        sample = trips[sample_idx] if sample_idx < len(trips) else None

        if sample:
            (date_val, location, full_days, partial_days, m_ie) = sample
            d = datetime.strptime(date_val, "%Y-%m-%d").date()
            a = ws.cell(row=i, column=1, value=d)
            apply_style(a, input_cell_style())
            a.number_format = "yyyy-mm-dd"

            b = ws.cell(row=i, column=2, value=location)
            apply_style(b, input_cell_style())

            c_ = ws.cell(row=i, column=3, value=full_days)
            apply_style(c_, input_cell_style())
            c_.number_format = "0"

            d_ = ws.cell(row=i, column=4, value=partial_days)
            apply_style(d_, input_cell_style())
            d_.number_format = "0"

            e = ws.cell(row=i, column=5, value=m_ie)
            apply_style(e, input_cell_style())
            e.number_format = '"$"#,##0'
        else:
            for col_idx in range(1, 6):
                c = ws.cell(row=i, column=col_idx)
                apply_style(c, input_cell_style())
                if col_idx == 1:
                    c.number_format = "yyyy-mm-dd"
                elif col_idx == 5:
                    c.number_format = '"$"#,##0'

        # F = Subtotal: full_days × M&IE + partial_days × M&IE × 75%
        f_cell = ws.cell(row=i, column=6, value=f"=IFERROR(C{i}*E{i}+D{i}*E{i}*0.75,0)")
        f_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(f_cell, formula_cell_style())

        # G = Ded % (default 50)
        g_cell = ws.cell(row=i, column=7, value=50 if sample else None)
        if sample:
            apply_style(g_cell, input_cell_style())
        else:
            apply_style(g_cell, input_cell_style())
        g_cell.number_format = "0"
        g_cell.alignment = Alignment(horizontal="center", vertical="center")

        # H = Ded $
        h_cell = ws.cell(row=i, column=8, value=f"=IFERROR(F{i}*G{i}/100,0)")
        h_cell.number_format = '"$"#,##0.00'
        if sample:
            apply_style(h_cell, formula_cell_style())

        if sample:
            ws.row_dimensions[i].height = 16

    add_dropdown(ws, "G6:G105", '"50,100"')

    ws.freeze_panes = "A6"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_monthly_summary_tab(wb):
    """Tab — Monthly Summary / CPA Hand-off page."""
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = COLOR_ACCENT  # gold = CPA hand-off
    compact_header_band(ws, "Monthly Summary · For Your CPA",
                         prev_tab="Per-Diem Log", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "📄 FOR YOUR CPA — print this tab. Auto-aggregated from Meals Log + Per-Diem Log."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = ["Month", "Meals $ (Actual)", "Deductible $ (combined)",
               "# Meals"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 20

    set_col_widths(ws, [("A", 14), ("B", 18), ("C", 22), ("D", 12)])

    for i in range(6, 18):
        month_num = i - 5
        next_month = month_num + 1

        ws.cell(row=i, column=1, value=MONTHS[month_num - 1]).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )

        if month_num < 12:
            actual_b = (
                f"=SUMIFS('Meals Log'!$I:$I,"
                f"'Meals Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Meals Log'!$A:$A,\"<\"&DATE(Settings!$B$5,{next_month},1))"
            )
            ded_meals = (
                f"SUMIFS('Meals Log'!$M:$M,"
                f"'Meals Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Meals Log'!$A:$A,\"<\"&DATE(Settings!$B$5,{next_month},1))"
            )
            ded_pd = (
                f"SUMIFS('Per-Diem Log'!$H:$H,"
                f"'Per-Diem Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Per-Diem Log'!$A:$A,\"<\"&DATE(Settings!$B$5,{next_month},1))"
            )
            count_d = (
                f"=COUNTIFS('Meals Log'!$A:$A,\">=\"&DATE(Settings!$B$5,{month_num},1),"
                f"'Meals Log'!$A:$A,\"<\"&DATE(Settings!$B$5,{next_month},1))"
            )
        else:
            actual_b = (
                f"=SUMIFS('Meals Log'!$I:$I,"
                f"'Meals Log'!$A:$A,\">=\"&DATE(Settings!$B$5,12,1),"
                f"'Meals Log'!$A:$A,\"<\"&DATE(Settings!$B$5+1,1,1))"
            )
            ded_meals = (
                f"SUMIFS('Meals Log'!$M:$M,"
                f"'Meals Log'!$A:$A,\">=\"&DATE(Settings!$B$5,12,1),"
                f"'Meals Log'!$A:$A,\"<\"&DATE(Settings!$B$5+1,1,1))"
            )
            ded_pd = (
                f"SUMIFS('Per-Diem Log'!$H:$H,"
                f"'Per-Diem Log'!$A:$A,\">=\"&DATE(Settings!$B$5,12,1),"
                f"'Per-Diem Log'!$A:$A,\"<\"&DATE(Settings!$B$5+1,1,1))"
            )
            count_d = (
                f"=COUNTIFS('Meals Log'!$A:$A,\">=\"&DATE(Settings!$B$5,12,1),"
                f"'Meals Log'!$A:$A,\"<\"&DATE(Settings!$B$5+1,1,1))"
            )

        b = ws.cell(row=i, column=2, value=actual_b)
        apply_style(b, formula_cell_style())
        b.number_format = '"$"#,##0.00'

        c_combined = ws.cell(row=i, column=3, value=f"={ded_meals}+{ded_pd}")
        apply_style(c_combined, formula_cell_style())
        c_combined.number_format = '"$"#,##0.00'

        d_count = ws.cell(row=i, column=4, value=count_d)
        apply_style(d_count, formula_cell_style())

        ws.row_dimensions[i].height = 16

    ws.row_dimensions[18].height = 8

    # YTD row 19
    ytd_gold = PatternFill("solid", fgColor=STATE_WARN_FILL)
    ytd_font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    a19 = ws.cell(row=19, column=1, value="YTD Total")
    a19.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    a19.fill = ytd_gold
    for col in [2, 3, 4]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=19, column=col, value=f"=SUM({col_letter}6:{col_letter}17)")
        c.font = ytd_font
        c.fill = ytd_gold
        if col in (2, 3):
            c.number_format = '"$"#,##0.00'
    ws.row_dimensions[19].height = 20

    ws.freeze_panes = "A6"

    # Bar chart — deductible $ by month
    bar = BarChart()
    bar.type = "col"
    bar.title = "Deductible Meal $ by Month"
    bar.legend = None
    bar.height = 9
    bar.width = 14
    data_ref = Reference(ws, min_col=3, min_row=5, max_row=17, max_col=3)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=17)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "low"
    style_chart(bar)
    ws.add_chart(bar, "F5")

    ws.print_area = "A1:N22"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT
    compact_header_band(ws, "Settings",
                         prev_tab="Monthly Summary", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value = "Tax year · M&IE rates · purpose list · year-end archive."
    c4.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    set_col_widths(ws, [
        ("A", 36), ("B", 18),
        ("C", 6),
        ("D", 16), ("E", 28),
        ("F", 6),
        ("G", 16), ("H", 14),
    ])

    bold_right = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT)
    italic_muted = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)

    a = ws.cell(row=5, column=1, value="Active tax year:")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=5, column=2, value=2026)
    apply_style(b, input_cell_style())
    b.number_format = "0"
    ws.row_dimensions[5].height = 18

    a = ws.cell(row=6, column=1, value="GSA M&IE — High-cost city ($/day):")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=6, column=2, value=M_IE_HIGH)
    apply_style(b, input_cell_style())
    b.number_format = '"$"#,##0'
    ws.row_dimensions[6].height = 18

    a = ws.cell(row=7, column=1, value="GSA M&IE — Standard CONUS ($/day):")
    a.font = bold_right
    a.alignment = Alignment(horizontal="right", vertical="center")
    b = ws.cell(row=7, column=2, value=M_IE_STD)
    apply_style(b, input_cell_style())
    b.number_format = '"$"#,##0'
    ws.row_dimensions[7].height = 18

    # Freshness stamp (suite Theme 4) — GSA publishes new M&IE rates every Oct
    # for the federal fiscal year starting Oct 1. Customer should bump B6+B7
    # then; without this, deductions silently use last year's rate.
    ws.merge_cells("A8:E8")
    c8_note = ws["A8"]
    c8_note.value = (
        f"📅 GSA M&IE rates as of {RATES_AS_OF} — GSA publishes new rates "
        f"each October (gsa.gov/perdiem); edit B6 + B7 then."
    )
    c8_note.font = italic_muted
    c8_note.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)
    ws.row_dimensions[8].height = 20

    # Purposes column (E10 header, E11+ rows for dropdown source)
    e10 = ws.cell(row=10, column=5, value="Purpose list (Meals Log dropdown source)")
    e10.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    for idx, purpose in enumerate(PURPOSES, start=11):
        cell = ws.cell(row=idx, column=5, value=purpose)
        apply_style(cell, input_cell_style())
        ws.row_dimensions[idx].height = 16

    # Substantial-services callout
    ws.merge_cells("A9:E9")
    c = ws["A9"]
    c.value = (
        "⚠ Schedule C filers: Some meals are 100% (employer-provided to staff "
        "while traveling, on-premises, etc.). Schedule E filers: meal deductions "
        "are still allowed but more limited — confirm with your CPA."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[9].height = 28

    # Year-end Archive
    a = ws.cell(row=20, column=1, value="Year-end Archive")
    a.font = Font(name=FONT_HEAD, size=13, bold=True, color=COLOR_PRIMARY)
    ws.row_dimensions[20].height = 22
    ws.merge_cells("A21:E21")
    c = ws["A21"]
    c.value = ("Each January, copy YTD totals into the row for the closing "
               "year before resetting the log.")
    c.font = italic_muted
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[21].height = 24

    archive_headers = ["Year", "Meal $ (Actual)", "Deductible $", "# Meals"]
    for col, h in enumerate(archive_headers, start=1):
        cell = ws.cell(row=22, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[22].height = 18
    for idx, year in enumerate(range(2024, 2031), start=23):
        ws.cell(row=idx, column=1, value=year).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        for col in [2, 3, 4]:
            cell = ws.cell(row=idx, column=col)
            apply_style(cell, input_cell_style())
            if col in (2, 3):
                cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[idx].height = 16


def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_meals_log_tab(wb, variant)
    build_per_diem_log_tab(wb, variant)
    build_monthly_summary_tab(wb)
    build_settings_tab(wb)

    suffix = f" ({variant.upper()})"
    wb.properties.title = f"Per-Diem Meal Tracker{suffix} — The STR Ledger"
    wb.properties.creator = BRAND_NAME
    wb.properties.company = BRAND_NAME
    wb.properties.description = "IRS Pub 463-compliant meal deduction tracker for STR hosts."

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path.name}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
