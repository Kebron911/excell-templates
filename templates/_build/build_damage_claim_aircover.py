"""Build OPS-002 Damage Claim + AirCover Log (v2.2 standard).

Operational-mode tool for STR damage claims. Tracks every incident from
discovery → evidence gathering → submission → resolution, with a
deadline countdown keyed to checkout date (Airbnb's standard 14-day
window). Forces evidence completeness before submission and provides
5 copy-paste message scripts for the panicked-Sunday-evening claim.

Generates two files:
  templates/_masters/OPS-002-damage-claim-aircover-log-DEMO.xlsx
  templates/_masters/OPS-002-damage-claim-aircover-log-BLANK.xlsx
"""
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from brand_config import (COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_TEXT,
    COLOR_MUTED, COLOR_BG_LIGHT, COLOR_ERROR,
    COLOR_PARCHMENT_ALT, COLOR_GOLD_SOFT, COLOR_NAVY_TINT,
    FONT_HEAD, FONT_BODY, FONT_MONO,
    BRAND_NAME, BRAND_DOMAIN, BRAND_EMAIL,
    input_cell_style, formula_cell_style,
    header_row_style, set_col_widths, apply_style,
    pseudo_button, compact_header_band, brand_footer,
    COLOR_GRAY_LIGHT, COLOR_INPUT_TINT, COLOR_WHITE, STATE_BAD_FILL, STATE_BAD_TEXT, STATE_GOOD_FILL,
    STATE_GOOD_TEXT,
)

SKU = "OPS-002"
NAME = "damage-claim-aircover-log"
BASE = Path(__file__).resolve().parent.parent
DEMO_OUT = BASE / "_masters" / f"{SKU}-{NAME}-DEMO.xlsx"
BLANK_OUT = BASE / "_masters" / f"{SKU}-{NAME}-BLANK.xlsx"

REFERENCE_AS_OF = "2026-05-02"
FILING_WINDOW_DEFAULT = 14  # Airbnb's standard 14-day-from-checkout window

# --- Sample data (DEMO) ---

# Per brief QA section: 3 claims Jan-Mar 2026
CLAIMS_REGISTER_SAMPLES = [
    # (date_disc, property, guest, booking, channel, checkout, damage, category, claimed, recovered, status, notes)
    ("2026-01-15", "Smokies Ridge Cabin", "Jordan M.",       "HMABCD123", "Airbnb",
     "2026-01-15", "Coffee table cracked diagonal — guest moved it on uneven floor",
     "Furniture", 340, 340, "Paid",
     "Resolved AirCover in 9 days; replacement table sourced from Wayfair."),
    ("2026-02-22", "Creek Side", "Taylor R.",                "HMEFGH456", "Airbnb",
     "2026-02-21", "Bedside lamp shattered + linen stain on duvet (red wine)",
     "Decor", 180, 90, "Paid",
     "Partial payout — AirCover argued lamp was 'normal wear'; duvet covered."),
    ("2026-03-08", "Lakehouse A", "Devin K.",                "HMIJKL789", "Airbnb",
     "2026-03-06", "TV cracked — guest claimed pre-existing; we have arrival photos showing intact",
     "Appliance", 720, 0, "Disputed",
     "Open. Arrival-day photos timestamped + sent. Awaiting Airbnb decision."),
]

# Claim Packet sample = the Jan 15 coffee table claim (fully populated)
PACKET_DEMO = {
    "incident": {
        "discovery_date": "2026-01-15",
        "discovery_time": "10:42 AM",
        "discovered_by": "Cleaner — Sarah (Smokies Clean)",
        "severity": 3,
        "category": "Furniture",
        "narrative": (
            "Cleaner found the coffee table cracked diagonally across the top "
            "after guest checkout. Floor in living area is uneven by ~½ inch; "
            "guest had moved the table to set up a board game. Damage is "
            "structural — table is no longer usable."
        ),
    },
    "damage_list": [
        # (item, replacement_cost, receipt, photo)
        ("West Elm Mid-Century Coffee Table (oak)", 340.00, "Yes", "Yes"),
    ],
    "evidence": [
        # (item label, checked? Y/N)
        ("Before-stay photo of item (intact)",        "✓"),
        ("After-stay photo of damage",                "✓"),
        ("Replacement receipt or invoice",            "✓"),
        ("Written guest message acknowledging",       "✓"),
        ("Second witness if applicable",              "—"),
        ("Police report number (if applicable)",      "—"),
        ("Date-stamped photos (metadata visible)",    "✓"),
        ("Original receipt of damaged item",          "✓"),
        ("Communication log with guest in writing",   "✓"),
        ("Cleaner statement or text message",         "✓"),
        ("Photos of contributing context (uneven floor)", "✓"),
        ("Inventory of items in room before stay",    "✓"),
    ],
    "comms": [
        # (date, channel, direction, summary)
        ("2026-01-15 11:02 AM", "Airbnb msg", "Out",
         "Notified guest, asked what happened, included photo."),
        ("2026-01-15 02:18 PM", "Airbnb msg", "In",
         "Guest acknowledged moving table, apologized, declined direct payment."),
        ("2026-01-15 03:30 PM", "Airbnb msg", "Out",
         "Opened AirCover claim through Resolution Center."),
        ("2026-01-18 09:15 AM", "Airbnb msg", "In",
         "AirCover requested replacement receipt + serial of damaged item."),
        ("2026-01-18 10:40 AM", "Airbnb msg", "Out",
         "Sent receipt scan + product page screenshot."),
        ("2026-01-24 01:22 PM", "Airbnb email", "In",
         "AirCover approved $340 full payout."),
    ],
    "timeline": [
        # (milestone, date)
        ("Discovery",                "2026-01-15"),
        ("Guest contact opened",     "2026-01-15"),
        ("Platform claim escalated", "2026-01-15"),
        ("Submission complete",      "2026-01-18"),
        ("Platform response",        "2026-01-18"),
        ("Resolution",               "2026-01-24"),
    ],
    "outcome": {
        "amount_requested": 340.00,
        "amount_received": 340.00,
        "denial_reason": "—",
        "lessons": (
            "Cleaner texts photos of any room with movable heavy furniture "
            "WITHIN 30 minutes of arrival from now on. Reduces 'pre-existing' "
            "disputes."
        ),
    },
}

# Settings tab data
PROPERTIES_LIST = [
    "Smokies Ridge Cabin", "Creek Side", "Lakehouse A",
    "Lakehouse B", "Mountain View", "Downtown Loft",
]
STATUS_LIST = [
    "Discovered", "Evidence gathering", "Submitted",
    "Disputed", "Paid", "Closed unpaid", "Other",
]
CATEGORY_LIST = [
    "Furniture", "Appliance", "Decor", "Linen", "Structural", "Other",
]

# Coverage Reference tab content — 3 platform sections
COVERAGE_REFERENCE = {
    "Airbnb AirCover for Hosts": {
        "url": "https://www.airbnb.com/aircover-for-hosts",
        "covered": [
            "Damage to your home, furnishings, or personal belongings caused by guest",
            "Pet damage (whether or not pets were disclosed)",
            "Income lost from cancellations required by guest damage",
            "Deep cleaning cost from extreme mess (smoke, biohazard)",
            "Damage from guests' guests / unauthorized parties",
            "Auto / boat / equipment damage at the listing",
            "Up to $3M USD in primary host damage protection",
            "$1M USD in liability for guest injury or guest property damage",
        ],
        "not_covered": [
            "Normal wear and tear (worn carpet, sun-faded fabric, scuffs)",
            "Pre-existing damage (date-stamp your before-stay photos)",
            "Damage caused by host or host's family",
            "Cash, securities, deeds, or collectibles above stated limits",
            "Acts of war, nuclear/radiological events, intentional damage by host",
        ],
        "filing_window": (
            "Within 14 days of guest checkout OR before next guest checks in, "
            "whichever is sooner. The 'next-guest' constraint catches many hosts off guard."
        ),
        "evidence_required": (
            "Photos of damage, replacement receipts/invoices, written communication "
            "with guest, original receipts of damaged items if requested."
        ),
        "dispute_path": (
            "Resolution Center → AirCover claim → if denied, request review by Airbnb "
            "Trust & Safety. Final escalation: state DOI complaint (rare)."
        ),
    },
    "VRBO Property Damage Protection": {
        "url": "https://help.vrbo.com/articles/Property-Damage-Protection",
        "covered": [
            "Accidental damage by guest up to property value",
            "Comes auto-included on most VRBO bookings (verify per booking)",
            "Replaces traditional damage deposits in most jurisdictions",
        ],
        "not_covered": [
            "Intentional damage (file a separate claim with criminal report)",
            "Pre-existing damage",
            "Normal wear and tear",
            "Damage from guests' pets unless pets were authorized + paid for",
        ],
        "filing_window": (
            "Within 60 days of guest departure for most accidental damage. "
            "Verify the specific booking's terms — coverage varies by package."
        ),
        "evidence_required": (
            "Damage photos, repair estimate or receipt, booking details. "
            "VRBO may request a 3-day notice to guest before filing."
        ),
        "dispute_path": (
            "VRBO Customer Support → escalate to Trust & Safety. Owners with "
            "Premier Partner status get faster review."
        ),
    },
    "Booking.com Partner Liability Insurance": {
        "url": "https://partner.booking.com/en-us/help",
        "covered": [
            "Guest property damage ($/€ amount varies by region)",
            "Third-party liability at the property",
            "Coverage included automatically — no opt-in",
        ],
        "not_covered": [
            "Intentional damage by host or host's family",
            "Pre-existing damage",
            "Normal wear and tear",
            "Loss of income from cancellations (separate process)",
        ],
        "filing_window": (
            "Generally 30 days from guest checkout. Confirm via your "
            "Extranet → Inbox → 'Report a problem with this guest' flow."
        ),
        "evidence_required": (
            "Photos, receipts, damage description, guest booking reference."
        ),
        "dispute_path": (
            "Booking.com Partner Support → if denied, escalate via Account "
            "Manager (Preferred Partner+ tier)."
        ),
    },
}

# Message Scripts tab — 5 scripts
MESSAGE_SCRIPTS = [
    {
        "title": "Script 1 — First contact, propose direct resolution",
        "when": (
            "Within 2-4 hours of damage discovery. Goal: settle direct so "
            "you skip the platform process and the guest avoids a claim "
            "history. Offer ~70% of replacement cost as a friction-killer."
        ),
        "subject": "Heads up about [item] in your stay",
        "body": (
            "Hi [Guest first name],\n\n"
            "Hope your trip home was easy. Quick heads-up — when our cleaner "
            "arrived this morning, we found [specific damage to specific "
            "item]. We're sure it was an accident.\n\n"
            "Easiest path for both of us: would you be open to sending "
            "$[70% of replacement cost] via Airbnb's Resolution Center "
            "directly? That covers the replacement and we close it out — "
            "no formal claim, no marks on your account.\n\n"
            "Photos attached for context. Let me know what you think.\n\n"
            "Thanks,\n[Your name]"
        ),
        "placeholders": "[Guest first name] · [specific damage to specific item] · [70% of replacement cost] · [Your name]",
    },
    {
        "title": "Script 2 — Formal claim opening",
        "when": (
            "If guest doesn't respond to Script 1 within 24 hours, OR if "
            "guest declined to pay direct. File the platform claim. "
            "Tone: factual, not adversarial."
        ),
        "subject": "Filing AirCover claim — [Booking ID]",
        "body": (
            "Hi [Guest first name],\n\n"
            "Following up on the [damage to item] from your stay — since we "
            "weren't able to resolve directly, I'm filing an AirCover claim "
            "for $[full replacement cost].\n\n"
            "I've attached: before-stay photos showing intact, after-stay "
            "photos showing damage, and the replacement receipt.\n\n"
            "Airbnb may reach out for your side of the story. No action "
            "needed from you here unless they ask.\n\n"
            "Best,\n[Your name]"
        ),
        "placeholders": "[Guest first name] · [damage to item] · [Booking ID] · [full replacement cost] · [Your name]",
    },
    {
        "title": "Script 3 — Evidence-request response (when AirCover asks for more)",
        "when": (
            "AirCover or platform support asks for additional evidence. "
            "Respond within 48 hours — slow responses correlate with denials. "
            "Be specific, attach everything in one message."
        ),
        "subject": "RE: AirCover claim [claim ID] — additional evidence",
        "body": (
            "Hi [Agent name],\n\n"
            "Thanks for the follow-up. Attaching the additional materials "
            "you requested:\n\n"
            "1. [evidence type 1] — [filename]\n"
            "2. [evidence type 2] — [filename]\n"
            "3. [evidence type 3] — [filename]\n\n"
            "For context: [brief 2-sentence narrative — what happened, why "
            "this evidence is dispositive]. Happy to provide anything else "
            "useful.\n\n"
            "[Your name] · [Listing name]"
        ),
        "placeholders": "[Agent name] · [claim ID] · [evidence type 1/2/3] · [filename] · [Your name] · [Listing name]",
    },
    {
        "title": "Script 4 — Dispute response (when platform denies/partial-pays)",
        "when": (
            "Platform partially pays or denies. Don't escalate emotionally. "
            "Reference specific policy language. Request review by a different "
            "agent. ~30% of disputed claims flip on second review."
        ),
        "subject": "Request for second review — claim [claim ID]",
        "body": (
            "Hi [Agent name],\n\n"
            "Following up on the partial decision on claim [claim ID] "
            "($[awarded] of $[requested]). Respectfully requesting a second "
            "review based on the following:\n\n"
            "• [Policy language X applies because Y]\n"
            "• [Evidence already submitted shows Z]\n"
            "• [Comparable claim resolved differently — if you have one]\n\n"
            "I want to resolve this fairly without escalating further. "
            "Could you assign a second agent to review?\n\n"
            "[Your name]"
        ),
        "placeholders": "[Agent name] · [claim ID] · [awarded amount] · [requested amount] · [Policy language X / evidence Z]",
    },
    {
        "title": "Script 5 — Post-resolution review management",
        "when": (
            "After claim closes, before review windows expire. Decide whether "
            "to leave a public review. Honest > harsh. Mention damage in a "
            "way that warns future hosts without sounding bitter."
        ),
        "subject": "(internal note — NOT sent to guest; private review only)",
        "body": (
            "[Guest first name] communicated promptly and respectfully. "
            "Stay was within house rules and quiet hours were observed. "
            "[Damage to specific item] occurred during the stay — claim "
            "was [paid in full / partial / disputed] through AirCover. "
            "Thumbs-up may be appropriate depending on responsiveness; "
            "use the private feedback to flag the damage for future hosts."
        ),
        "placeholders": "[Guest first name] · [Damage to specific item] · [paid in full / partial / disputed]",
    },
]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _val(variant, demo_value):
    """Return demo_value if variant == 'demo', else None (blank workbook)."""
    return demo_value if variant == "demo" else None


def _val_list(variant, demo_list, blank_length=None):
    """Return demo_list if variant == 'demo', else a list of None of same length."""
    if variant == "demo":
        return demo_list
    n = blank_length if blank_length is not None else len(demo_list)
    return [None] * n


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def add_dropdown(ws, cell_range, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_start_tab(wb, variant):
    """Sheet 1 — Start tab with at-a-glance + deadline-watch banner."""
    ws = wb.active
    ws.title = "Start"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [(get_column_letter(c), 8) for c in range(1, 13)])

    # ZONE 1: Navy hero band rows 1-8
    navy_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = navy_fill

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = BRAND_NAME
    c.font = Font(name=FONT_HEAD, size=14, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "Damage Claim + AirCover Log"
    c.font = Font(name=FONT_HEAD, size=30, bold=True, color=COLOR_BG_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = "When a guest breaks something, this is the calm move."
    c.font = Font(name=FONT_HEAD, size=14, italic=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A7:L7")
    c = ws["A7"]
    c.value = f"{SKU} · v1.0 · {variant.upper()}"
    c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ZONE 2: At-a-glance metric cards — 3 cards rows 10-15
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for r in range(10, 16):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = parchment_fill

    # Card 1 (A-D): Open Claims
    ws.merge_cells("A10:D10")
    c = ws["A10"]
    c.value = "OPEN CLAIMS"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A11:D13")
    c = ws["A11"]
    c.value = (
        '=COUNTIFS(\'Claims Register\'!K6:K205,"<>Paid",'
        '\'Claims Register\'!K6:K205,"<>Closed unpaid",'
        '\'Claims Register\'!A6:A205,"<>")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A14:D15")
    c = ws["A14"]
    c.value = "claims still in flight"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 2 (E-H): $ at risk
    ws.merge_cells("E10:H10")
    c = ws["E10"]
    c.value = "$ AT RISK"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("E11:H13")
    c = ws["E11"]
    c.value = (
        '=SUMIFS(\'Claims Register\'!I6:I205,'
        '\'Claims Register\'!K6:K205,"<>Paid",'
        '\'Claims Register\'!K6:K205,"<>Closed unpaid",'
        '\'Claims Register\'!A6:A205,"<>")'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '"$"#,##0'
    ws.merge_cells("E14:H15")
    c = ws["E14"]
    c.value = "claimed, not yet recovered"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Card 3 (I-L): Recovery Rate YTD
    ws.merge_cells("I10:L10")
    c = ws["I10"]
    c.value = "RECOVERY RATE"
    c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("I11:L13")
    c = ws["I11"]
    c.value = (
        '=IFERROR(SUM(\'Claims Register\'!J6:J205)'
        '/SUM(\'Claims Register\'!I6:I205),0)'
    )
    c.font = Font(name=FONT_HEAD, size=32, bold=True, color=COLOR_SECONDARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    ws.merge_cells("I14:L15")
    c = ws["I14"]
    c.value = "$ recovered ÷ $ claimed"
    c.font = Font(name=FONT_BODY, size=9, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Borders on cards
    gold_side = Side(style="thin", color=COLOR_ACCENT)
    for first, last in [("A", "D"), ("E", "H"), ("I", "L")]:
        fc = column_index_from_string(first)
        lc_ = column_index_from_string(last)
        for r in range(10, 16):
            for col in range(fc, lc_ + 1):
                existing = ws.cell(row=r, column=col).border
                ws.cell(row=r, column=col).border = Border(
                    top=gold_side if r == 10 else existing.top,
                    bottom=gold_side if r == 15 else existing.bottom,
                    left=gold_side if col == fc else existing.left,
                    right=gold_side if col == lc_ else existing.right,
                )

    # ZONE 3: Deadline watch banner rows 17-23
    for r in range(17, 24):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)

    ws.merge_cells("A17:L17")
    c = ws["A17"]
    c.value = "⚠  DEADLINE WATCH  —  CLAIMS WITH < 7 DAYS LEFT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[17].height = 22

    # 5 rows of formula-driven priority list (small INDEX/SMALL pattern,
    # but to keep the formula readable we list the top-5 days-to-deadline
    # values directly with conditional row visibility).
    ws.merge_cells("A18:L18")
    c = ws["A18"]
    c.value = (
        "Open the Claims Register tab — rows with red 'Days to deadline' "
        "values are listed there in priority order. The countdown is keyed "
        "to the guest's checkout date plus your filing-window default "
        "(Settings → 14 days for Airbnb)."
    )
    c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[18].height = 36

    ws.merge_cells("A20:L23")
    c = ws["A20"]
    c.value = (
        "Why this matters: AirCover's standard window is 14 days from guest "
        "checkout — OR before the next guest checks in, whichever is sooner. "
        "Most denied claims are denied for missing the window, not for weak "
        "evidence. Open the Coverage Reference tab for the current policy "
        "(verified " + REFERENCE_AS_OF + ")."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)

    # ZONE 4: Pseudo-button nav rows 25-27
    pseudo_button(ws, "A25", "C27", "Claims Register",
                  "'Claims Register'!A1", variant="primary")
    pseudo_button(ws, "D25", "F27", "New Claim Packet",
                  "'Claim Packet'!A1", variant="primary")
    pseudo_button(ws, "G25", "I27", "Coverage Reference",
                  "'Coverage Reference'!A1", variant="secondary")
    pseudo_button(ws, "J25", "L27", "Message Scripts",
                  "'Message Scripts'!A1", variant="secondary")
    for r in range(25, 28):
        ws.row_dimensions[r].height = 22

    pseudo_button(ws, "A29", "F31", "Annual Damage P&L",
                  "'Annual Damage P&L'!A1", variant="secondary")
    pseudo_button(ws, "G29", "L31", "Settings",
                  "'Settings'!A1", variant="secondary")
    for r in range(29, 32):
        ws.row_dimensions[r].height = 22

    # ZONE 5: Upgrade banner row 33
    ws.merge_cells("A33:L33")
    c = ws["A33"]
    c.value = (
        "💡  Need full operations coverage? Add the Operator Bundle at "
        f"{BRAND_DOMAIN}/operator — turnover + maintenance + supply + "
        "damage claims + insurance + permits, $197."
    )
    c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_WHITE)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[33].height = 36

    brand_footer(ws, 35,
                 version_line=f"{SKU} · v1.0 · {variant.upper()} · Free updates forever")

    ws.print_area = "A1:L37"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_claims_register_tab(wb, variant):
    """Sheet 2 — Claims Register flat log."""
    ws = wb.create_sheet("Claims Register")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    compact_header_band(ws, "Claims Register",
                        prev_tab="Start", next_tab="Claim Packet")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:N4")
    c = ws["A4"]
    c.value = (
        "Master log — all claims. Days-to-deadline is auto-calculated from "
        "checkout date + filing window (Settings → 14 days default)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    headers = [
        "Date discovered", "Property", "Guest", "Booking ID",
        "Channel", "Checkout", "Damage description", "Category",
        "$ Claimed", "$ Recovered", "Status", "Days to deadline",
        "Recovery %", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[5].height = 28

    set_col_widths(ws, [
        ("A", 12), ("B", 22), ("C", 18), ("D", 14),
        ("E", 11), ("F", 12), ("G", 38), ("H", 12),
        ("I", 11), ("J", 12), ("K", 16), ("L", 11),
        ("M", 10), ("N", 28),
    ])

    samples = _val_list(variant, CLAIMS_REGISTER_SAMPLES, blank_length=0)
    for i, row_data in enumerate(samples):
        row = 6 + i
        (date_disc, prop, guest, booking, channel, checkout,
         damage, category, claimed, recovered, status, notes) = row_data

        a = ws.cell(row=row, column=1, value=_parse_date(date_disc))
        apply_style(a, input_cell_style())
        a.number_format = "yyyy-mm-dd"

        for col, val in [(2, prop), (3, guest), (4, booking), (5, channel)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        f = ws.cell(row=row, column=6, value=_parse_date(checkout))
        apply_style(f, input_cell_style())
        f.number_format = "yyyy-mm-dd"

        for col, val in [(7, damage), (8, category)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())

        for col, val in [(9, claimed), (10, recovered)]:
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, input_cell_style())
            cell.number_format = '"$"#,##0'

        k = ws.cell(row=row, column=11, value=status)
        apply_style(k, input_cell_style())

        n = ws.cell(row=row, column=14, value=notes)
        apply_style(n, input_cell_style())

        ws.row_dimensions[row].height = 22

    # Apply formulas + format to all rows 6-205
    for row in range(6, 206):
        # Days-to-deadline formula col L
        l_cell = ws.cell(
            row=row, column=12,
            value=(
                f'=IF(AND(F{row}<>"",K{row}<>"Paid",K{row}<>"Closed unpaid"),'
                f'Settings!$B$25-(TODAY()-F{row}),"")'
            ),
        )
        apply_style(l_cell, formula_cell_style())
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        l_cell.number_format = "0"

        # Recovery % formula col M
        m_cell = ws.cell(
            row=row, column=13,
            value=f'=IF(I{row}>0,J{row}/I{row},"")'
        )
        apply_style(m_cell, formula_cell_style())
        m_cell.number_format = "0%"

        # For rows that have no sample (blank rows or blank variant), apply input style
        if row > 5 + len(samples):
            for col in [1, 6]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = "yyyy-mm-dd"
            for col in [2, 3, 4, 5, 7, 8, 11, 14]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
            for col in [9, 10]:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, input_cell_style())
                cell.number_format = '"$"#,##0'

    # Conditional formatting on Days-to-deadline (col L)
    # Red <3, gold-soft 3-7, parchment >7
    ws.conditional_formatting.add(
        "L6:L205",
        CellIsRule(operator="lessThan", formula=["3"],
                   fill=PatternFill("solid", fgColor=STATE_BAD_FILL)),
    )
    ws.conditional_formatting.add(
        "L6:L205",
        CellIsRule(operator="between", formula=["3", "7"],
                   fill=PatternFill("solid", fgColor=COLOR_GOLD_SOFT)),
    )

    # Dropdowns
    add_dropdown(ws, "B6:B205", "=Settings!$B$5:$B$24")
    add_dropdown(ws, "E6:E205", '"Airbnb,VRBO,Booking.com,Direct,Other"')
    add_dropdown(ws, "H6:H205", "=Settings!$D$5:$D$14")
    add_dropdown(ws, "K6:K205", "=Settings!$C$5:$C$14")

    ws.freeze_panes = "A6"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_claim_packet_tab(wb, variant):
    """Sheet 3 — Single Claim Packet form."""
    ws = wb.create_sheet("Claim Packet")
    ws.sheet_properties.tabColor = COLOR_SECONDARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28), ("C", 18), ("D", 14),
        ("E", 14), ("F", 14), ("G", 14), ("H", 8),
        ("I", 8), ("J", 8), ("K", 8), ("L", 10),
    ])

    compact_header_band(ws, "Claim Packet",
                        prev_tab="Claims Register", next_tab="Coverage Reference")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Per-claim worksheet — fill from top to bottom. Reuse for next claim "
        "(or duplicate this tab if you want to keep history)."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18

    demo = variant == "demo"
    pdata = PACKET_DEMO if demo else None

    # ===== §1 INCIDENT (rows 6-15) =====
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = "§1  INCIDENT"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[6].height = 22

    incident_fields = [
        ("Discovery date:",  "B7",  "discovery_date",  "yyyy-mm-dd"),
        ("Discovery time:",  "B8",  "discovery_time",  None),
        ("Discovered by:",   "B9",  "discovered_by",   None),
        ("Severity (1-5):",  "B10", "severity",        None),
        ("Category:",        "B11", "category",        None),
    ]
    for label, cell_ref, key, num_fmt in incident_fields:
        row = int("".join(filter(str.isdigit, cell_ref)))
        ws.cell(row=row, column=2, value=label).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        # Inputs in cols C-G merged
        ws.merge_cells(f"C{row}:G{row}")
        cell = ws[f"C{row}"]
        if demo and pdata:
            val = pdata["incident"][key]
            if key == "discovery_date":
                cell.value = _parse_date(val)
            else:
                cell.value = val
        apply_style(cell, input_cell_style())
        if num_fmt:
            cell.number_format = num_fmt
        ws.row_dimensions[row].height = 18

    # Narrative row 12-14 (multi-line wrap)
    ws.cell(row=12, column=2, value="Narrative:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C12:L14")
    cell = ws["C12"]
    if demo and pdata:
        cell.value = pdata["incident"]["narrative"]
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(12, 15):
        ws.row_dimensions[r].height = 24

    # Dropdown for category col C row 11
    add_dropdown(ws, "C11", "=Settings!$D$5:$D$14")

    # ===== §2 DAMAGE LIST (rows 16-29) =====
    ws.merge_cells("A16:L16")
    c = ws["A16"]
    c.value = "§2  DAMAGE LIST"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[16].height = 22

    # Headers row 17
    damage_headers = ["Item description", "Replacement cost", "Receipt?", "Photo?"]
    damage_cols = [("B", "F"), ("G", "H"), ("I", "J"), ("K", "L")]
    for (cstart, cend), h in zip(damage_cols, damage_headers):
        ws.merge_cells(f"{cstart}17:{cend}17")
        cell = ws[f"{cstart}17"]
        cell.value = h
        apply_style(cell, header_row_style())
    ws.row_dimensions[17].height = 20

    # 10 data rows 18-27
    damage_demo = pdata["damage_list"] if demo and pdata else []
    for i in range(10):
        r = 18 + i
        sample = damage_demo[i] if i < len(damage_demo) else (None, None, None, None)
        item, cost, receipt, photo = sample

        ws.merge_cells(f"B{r}:F{r}")
        cell = ws[f"B{r}"]
        cell.value = item
        apply_style(cell, input_cell_style())

        ws.merge_cells(f"G{r}:H{r}")
        cell = ws[f"G{r}"]
        cell.value = cost
        apply_style(cell, input_cell_style())
        cell.number_format = '"$"#,##0.00'

        ws.merge_cells(f"I{r}:J{r}")
        cell = ws[f"I{r}"]
        cell.value = receipt
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        add_dropdown(ws, f"I{r}", '"Yes,No"')

        ws.merge_cells(f"K{r}:L{r}")
        cell = ws[f"K{r}"]
        cell.value = photo
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        add_dropdown(ws, f"K{r}", '"Yes,No"')

        ws.row_dimensions[r].height = 18

    # Total row 29
    ws.cell(row=29, column=2, value="Total replacement cost:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("G29:H29")
    cell = ws["G29"]
    cell.value = "=SUM(G18:H27)"
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
    cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[29].height = 22

    # ===== §3 EVIDENCE CHECKLIST (rows 31-45) =====
    ws.merge_cells("A31:L31")
    c = ws["A31"]
    c.value = "§3  EVIDENCE CHECKLIST  —  AT LEAST 75% TO BE 'SUBMIT-READY'"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[31].height = 22

    # 12 evidence rows
    evidence_demo = pdata["evidence"] if demo and pdata else None
    default_labels = [
        "Before-stay photo of item (intact)",
        "After-stay photo of damage",
        "Replacement receipt or invoice",
        "Written guest message acknowledging",
        "Second witness if applicable",
        "Police report number (if applicable)",
        "Date-stamped photos (metadata visible)",
        "Original receipt of damaged item",
        "Communication log with guest in writing",
        "Cleaner statement or text message",
        "Photos of contributing context",
        "Inventory of items in room before stay",
    ]
    for i in range(12):
        r = 32 + i
        if demo and evidence_demo:
            label, mark = evidence_demo[i]
        else:
            label, mark = default_labels[i], None

        # Checkbox col B
        cb = ws.cell(row=r, column=2, value=mark)
        cb.font = Font(name=FONT_BODY, size=14, color=COLOR_PRIMARY)
        cb.fill = PatternFill("solid", fgColor=COLOR_INPUT_TINT)
        cb.alignment = Alignment(horizontal="center", vertical="center")
        cb.border = Border(
            left=Side(style="thin", color=COLOR_GRAY_LIGHT),
            right=Side(style="thin", color=COLOR_GRAY_LIGHT),
            top=Side(style="thin", color=COLOR_GRAY_LIGHT),
            bottom=Side(style="thin", color=COLOR_GRAY_LIGHT),
        )

        # Label cols C-L
        ws.merge_cells(f"C{r}:L{r}")
        cell = ws[f"C{r}"]
        cell.value = label
        cell.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[r].height = 18

    # Completeness score row 45
    ws.cell(row=45, column=2, value="Completeness score:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C45:F45")
    cell = ws["C45"]
    cell.value = '=COUNTIF(B32:B43,"✓")/12'
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.number_format = "0%"

    # Submit-ready gate row 46
    ws.cell(row=46, column=2, value="Submit-ready?").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C46:F46")
    cell = ws["C46"]
    cell.value = '=IF(C45>=0.75,"✅ Submit-ready","⚠ Gather more evidence")'
    apply_style(cell, formula_cell_style())
    cell.font = Font(name=FONT_HEAD, size=12, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[45].height = 22
    ws.row_dimensions[46].height = 22

    # ===== §4 COMMUNICATIONS LOG (rows 48-58) =====
    ws.merge_cells("A48:L48")
    c = ws["A48"]
    c.value = "§4  COMMUNICATIONS LOG"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[48].height = 22

    comm_headers = ["Date / time", "Channel", "Direction", "Summary"]
    comm_cols = [("B", "D"), ("E", "F"), ("G", "G"), ("H", "L")]
    for (cstart, cend), h in zip(comm_cols, comm_headers):
        ws.merge_cells(f"{cstart}49:{cend}49")
        cell = ws[f"{cstart}49"]
        cell.value = h
        apply_style(cell, header_row_style())
    ws.row_dimensions[49].height = 20

    comms_demo = pdata["comms"] if demo and pdata else []
    for i in range(8):
        r = 50 + i
        sample = comms_demo[i] if i < len(comms_demo) else (None, None, None, None)
        date_t, channel, direction, summary = sample

        ws.merge_cells(f"B{r}:D{r}")
        cell = ws[f"B{r}"]
        cell.value = date_t
        apply_style(cell, input_cell_style())

        ws.merge_cells(f"E{r}:F{r}")
        cell = ws[f"E{r}"]
        cell.value = channel
        apply_style(cell, input_cell_style())

        cell = ws.cell(row=r, column=7, value=direction)
        apply_style(cell, input_cell_style())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        add_dropdown(ws, f"G{r}", '"In,Out"')

        ws.merge_cells(f"H{r}:L{r}")
        cell = ws[f"H{r}"]
        cell.value = summary
        apply_style(cell, input_cell_style())

        ws.row_dimensions[r].height = 18

    # ===== §5 TIMELINE (rows 59-67) =====
    ws.merge_cells("A59:L59")
    c = ws["A59"]
    c.value = "§5  TIMELINE"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[59].height = 22

    timeline_demo = pdata["timeline"] if demo and pdata else None
    timeline_milestones = [
        "Discovery", "Guest contact opened", "Platform claim escalated",
        "Submission complete", "Platform response", "Resolution",
    ]
    for i, milestone in enumerate(timeline_milestones):
        r = 60 + i
        ws.cell(row=r, column=2, value=milestone).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        ws.merge_cells(f"C{r}:F{r}")
        cell = ws[f"C{r}"]
        if demo and timeline_demo:
            cell.value = _parse_date(timeline_demo[i][1])
        apply_style(cell, input_cell_style())
        cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[r].height = 18

    # ===== §6 OUTCOME (rows 67-75) =====
    ws.merge_cells("A67:L67")
    c = ws["A67"]
    c.value = "§6  OUTCOME"
    c.font = Font(name=FONT_MONO, size=10, bold=True, color=COLOR_PRIMARY)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[67].height = 22

    outcome_fields = [
        ("$ Requested:",   "amount_requested"),
        ("$ Received:",    "amount_received"),
        ("Denial reason:", "denial_reason"),
    ]
    for i, (label, key) in enumerate(outcome_fields):
        r = 68 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
        )
        ws.merge_cells(f"C{r}:L{r}")
        cell = ws[f"C{r}"]
        if demo and pdata:
            cell.value = pdata["outcome"][key]
        apply_style(cell, input_cell_style())
        if key in ("amount_requested", "amount_received"):
            cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 18

    ws.cell(row=71, column=2, value="Lessons learned:").font = Font(
        name=FONT_BODY, size=10, bold=True, color=COLOR_TEXT
    )
    ws.merge_cells("C71:L73")
    cell = ws["C71"]
    if demo and pdata:
        cell.value = pdata["outcome"]["lessons"]
    apply_style(cell, input_cell_style())
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for r in range(71, 74):
        ws.row_dimensions[r].height = 22

    # Footer reminder
    ws.merge_cells("A75:L75")
    c = ws["A75"]
    c.value = (
        "When the claim closes — copy the outcome row up to the Claims Register "
        "so the at-a-glance dashboard reflects it. Then reuse this packet for "
        "the next claim."
    )
    c.font = Font(name=FONT_BODY, size=9, italic=True, color=COLOR_MUTED)
    c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[75].height = 28

    ws.print_area = "A1:L75"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def build_coverage_reference_tab(wb, variant):
    """Sheet 4 — Coverage Reference (3 platform sections)."""
    ws = wb.create_sheet("Coverage Reference")
    ws.sheet_properties.tabColor = COLOR_PARCHMENT_ALT[:6] if False else COLOR_BG_LIGHT

    set_col_widths(ws, [
        ("A", 4), ("B", 26),
        ("C", 12), ("D", 12), ("E", 12), ("F", 12),
        ("G", 12), ("H", 12), ("I", 12), ("J", 12),
        ("K", 12), ("L", 12),
    ])

    compact_header_band(ws, "Coverage Reference",
                        prev_tab="Claim Packet", next_tab="Message Scripts")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        f"Policy snapshots — verified {REFERENCE_AS_OF}. Verify the current "
        "policy at each platform's help center before relying on this for a claim."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    current_row = 6
    for platform_name, data in COVERAGE_REFERENCE.items():
        # Platform banner
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = platform_name
        c.font = Font(name=FONT_HEAD, size=16, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # URL row
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = f"Reference: {data['url']}"
        c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # spacer
        current_row += 1

        # Covered
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = "✓  COVERED"
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=STATE_GOOD_TEXT)
        c.fill = PatternFill("solid", fgColor=STATE_GOOD_FILL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for item in data["covered"]:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            c = ws[f"A{current_row}"]
            c.value = f"  •  {item}"
            c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1

        # Not covered
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = "✗  NOT COVERED"
        c.font = Font(name=FONT_MONO, size=10, bold=True, color=STATE_BAD_TEXT)
        c.fill = PatternFill("solid", fgColor=STATE_BAD_FILL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for item in data["not_covered"]:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            c = ws[f"A{current_row}"]
            c.value = f"  •  {item}"
            c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1

        # Filing window + Evidence + Dispute
        meta_fields = [
            ("FILING WINDOW",  data["filing_window"]),
            ("EVIDENCE REQUIRED", data["evidence_required"]),
            ("DISPUTE PATH",   data["dispute_path"]),
        ]
        for label, content in meta_fields:
            ws.merge_cells(f"A{current_row}:L{current_row}")
            c = ws[f"A{current_row}"]
            c.value = label
            c.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
            c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            ws.merge_cells(f"A{current_row}:L{current_row + 1}")
            c = ws[f"A{current_row}"]
            c.value = content
            c.font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
            ws.row_dimensions[current_row].height = 22
            ws.row_dimensions[current_row + 1].height = 22
            current_row += 2

        # spacer between platforms
        current_row += 2

    # Final disclaimer
    ws.merge_cells(f"A{current_row}:L{current_row + 2}")
    c = ws[f"A{current_row}"]
    c.value = (
        "⚠  Claim policies change frequently. Treat this tab as a starting "
        "point, not as legal advice. For high-value claims (>$500), confirm "
        "the current policy directly with the platform before submission."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_ERROR)
    c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(current_row, current_row + 3):
        ws.row_dimensions[r].height = 18


def build_message_scripts_tab(wb, variant):
    """Sheet 5 — Message Scripts (5 copy-paste scripts)."""
    ws = wb.create_sheet("Message Scripts")
    ws.sheet_properties.tabColor = COLOR_BG_LIGHT

    set_col_widths(ws, [
        ("A", 4), ("B", 26),
        ("C", 12), ("D", 12), ("E", 12), ("F", 12),
        ("G", 12), ("H", 12), ("I", 12), ("J", 12),
        ("K", 12), ("L", 12),
    ])

    compact_header_band(ws, "Message Scripts",
                        prev_tab="Coverage Reference", next_tab="Annual Damage P&L")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Five copy-paste scripts for the most-common claim moments. "
        "Replace the [BRACKETED] fields with your specifics."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    current_row = 6
    for script in MESSAGE_SCRIPTS:
        # Title
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = script["title"]
        c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
        c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # When to send
        ws.merge_cells(f"A{current_row}:L{current_row + 1}")
        c = ws[f"A{current_row}"]
        c.value = "When to send:  " + script["when"]
        c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
        c.fill = PatternFill("solid", fgColor=COLOR_PARCHMENT_ALT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
        ws.row_dimensions[current_row].height = 22
        ws.row_dimensions[current_row + 1].height = 22
        current_row += 2

        # Subject line
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = "Subject:  " + script["subject"]
        c.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Body
        body_lines = script["body"].count("\n") + 1
        body_height = max(body_lines * 14, 80)
        ws.merge_cells(f"A{current_row}:L{current_row + max(body_lines - 1, 4)}")
        c = ws[f"A{current_row}"]
        c.value = script["body"]
        c.font = Font(name=FONT_BODY, size=11, color=COLOR_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
        for r in range(current_row, current_row + max(body_lines, 5)):
            ws.row_dimensions[r].height = 16
        current_row += max(body_lines, 5)

        # Placeholders
        ws.merge_cells(f"A{current_row}:L{current_row}")
        c = ws[f"A{current_row}"]
        c.value = "Customize:  " + script["placeholders"]
        c.font = Font(name=FONT_MONO, size=9, color=COLOR_ACCENT)
        c.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Spacer
        current_row += 2


def build_annual_damage_pl_tab(wb, variant):
    """Sheet 6 — Annual Damage P&L."""
    ws = wb.create_sheet("Annual Damage P&L")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    set_col_widths(ws, [
        ("A", 4), ("B", 36),
        ("C", 18), ("D", 6), ("E", 6), ("F", 6),
        ("G", 6), ("H", 6), ("I", 6), ("J", 6),
        ("K", 6), ("L", 6),
    ])

    compact_header_band(ws, "Annual Damage P&L",
                        prev_tab="Message Scripts", next_tab="Settings")

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Year-end summary — claims filed, recovery rate, top loss categories, "
        "by-property breakdown. Set the year below; everything else auto-pulls."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # Year selector
    ws.cell(row=6, column=2, value="Tax year:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    cell = ws.cell(row=6, column=3, value="=YEAR(TODAY())")
    apply_style(cell, input_cell_style())
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 24

    # TOTALS section row 8
    ws.merge_cells("A8:L8")
    c = ws["A8"]
    c.value = "TOTALS"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[8].height = 28

    totals = [
        ("Total $ claimed:",
         '=SUMPRODUCT((YEAR(\'Claims Register\'!A6:A205)=C6)*\'Claims Register\'!I6:I205)',
         '"$"#,##0.00'),
        ("Total $ recovered:",
         '=SUMPRODUCT((YEAR(\'Claims Register\'!A6:A205)=C6)*\'Claims Register\'!J6:J205)',
         '"$"#,##0.00'),
        ("Recovery rate:",
         '=IFERROR(C10/C9,0)',
         "0%"),
        ("Claim count:",
         '=SUMPRODUCT((YEAR(\'Claims Register\'!A6:A205)=C6)*(\'Claims Register\'!A6:A205<>""))',
         "0"),
        ("Avg claim size:",
         '=IFERROR(C9/C12,0)',
         '"$"#,##0.00'),
    ]
    for i, (label, formula, fmt) in enumerate(totals):
        r = 9 + i
        ws.cell(row=r, column=2, value=label).font = Font(
            name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
        )
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, formula_cell_style())
        cell.number_format = fmt
        cell.font = Font(name=FONT_BODY, size=11, bold=True, color=COLOR_PRIMARY)
        ws.row_dimensions[r].height = 20

    # TOP LOSS CATEGORIES row 15
    ws.merge_cells("A15:L15")
    c = ws["A15"]
    c.value = "TOP LOSS CATEGORIES"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[15].height = 28

    # Category breakdown rows 16-21 (6 categories)
    for i, category in enumerate(CATEGORY_LIST):
        r = 16 + i
        ws.cell(row=r, column=2, value=category).font = Font(
            name=FONT_BODY, size=11, color=COLOR_TEXT
        )
        cell = ws.cell(
            row=r, column=3,
            value=(
                f'=SUMPRODUCT((YEAR(\'Claims Register\'!A6:A205)=C6)*'
                f'(\'Claims Register\'!H6:H205=B{r})*\'Claims Register\'!I6:I205)'
            ),
        )
        apply_style(cell, formula_cell_style())
        cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 18

    # BY PROPERTY row 23
    ws.merge_cells("A23:L23")
    c = ws["A23"]
    c.value = "BY PROPERTY"
    c.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_BG_LIGHT)
    c.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[23].height = 28

    # Property header row 24
    headers = ["Property", "$ Claimed", "$ Recovered"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=24, column=col, value=h)
        apply_style(cell, header_row_style())
    ws.row_dimensions[24].height = 22

    # Property rows 25-34 (10 capacity)
    for i in range(10):
        r = 25 + i
        if i < len(PROPERTIES_LIST):
            prop = PROPERTIES_LIST[i]
            ws.cell(row=r, column=2, value=prop).font = Font(
                name=FONT_BODY, size=11, color=COLOR_TEXT
            )

            claimed_cell = ws.cell(
                row=r, column=3,
                value=(
                    f'=SUMPRODUCT((YEAR(\'Claims Register\'!A6:A205)=C6)*'
                    f'(\'Claims Register\'!B6:B205=B{r})*\'Claims Register\'!I6:I205)'
                ),
            )
            apply_style(claimed_cell, formula_cell_style())
            claimed_cell.number_format = '"$"#,##0.00'

            recovered_cell = ws.cell(
                row=r, column=4,
                value=(
                    f'=SUMPRODUCT((YEAR(\'Claims Register\'!A6:A205)=C6)*'
                    f'(\'Claims Register\'!B6:B205=B{r})*\'Claims Register\'!J6:J205)'
                ),
            )
            apply_style(recovered_cell, formula_cell_style())
            recovered_cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 18

    # Print area
    ws.print_area = "A1:D36"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)


def build_settings_tab(wb, variant):
    """Sheet 7 — Settings shell. Real layout is in _fix_settings_layout()
    so dropdown references stay in one place.
    """
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    set_col_widths(ws, [
        ("A", 4), ("B", 28),
        ("C", 28), ("D", 28),
    ])

    compact_header_band(ws, "Settings",
                        prev_tab="Annual Damage P&L", next_tab=None)

    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
    for c in range(1, 13):
        ws.cell(row=4, column=c).fill = parchment_fill
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Edit these lists to match your portfolio + workflow. They drive the "
        "dropdowns on the Claims Register and Claim Packet tabs."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_TEXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22


def _fix_settings_layout(ws):
    """Lay out Settings tab rows after the description band properly."""
    parchment_fill = PatternFill("solid", fgColor=COLOR_BG_LIGHT)

    # Row 5 — column headers (italic muted)
    headers = [(2, "Properties"), (3, "Claim status"), (4, "Damage category")]
    for col, label in headers:
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(name=FONT_MONO, size=9, bold=True, color=COLOR_PRIMARY)
        cell.fill = PatternFill("solid", fgColor=COLOR_GOLD_SOFT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 20

    # Properties col B rows 6-25 (20 capacity)
    for i in range(20):
        r = 6 + i
        cell = ws.cell(row=r, column=2)
        if i < len(PROPERTIES_LIST):
            cell.value = PROPERTIES_LIST[i]
        apply_style(cell, input_cell_style())
        ws.row_dimensions[r].height = 18

    # Statuses col C rows 6-15 (10 capacity)
    for i in range(10):
        r = 6 + i
        cell = ws.cell(row=r, column=3)
        if i < len(STATUS_LIST):
            cell.value = STATUS_LIST[i]
        apply_style(cell, input_cell_style())

    # Categories col D rows 6-15 (10 capacity)
    for i in range(10):
        r = 6 + i
        cell = ws.cell(row=r, column=4)
        if i < len(CATEGORY_LIST):
            cell.value = CATEGORY_LIST[i]
        apply_style(cell, input_cell_style())

    # Filing-window default cell — row 27 col C
    ws.cell(row=27, column=2, value="Filing window (days):").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    cell = ws.cell(row=27, column=3, value=FILING_WINDOW_DEFAULT)
    apply_style(cell, input_cell_style())
    cell.number_format = "0"
    cell.font = Font(name=FONT_HEAD, size=14, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[27].height = 24

    # Note on filing window
    ws.merge_cells("B28:L28")
    c = ws["B28"]
    c.value = (
        "Default 14 days = Airbnb's standard filing window from guest checkout. "
        "VRBO is generally 60 days, Booking.com generally 30. Set this to match "
        "your most-common channel or the strictest one."
    )
    c.font = Font(name=FONT_BODY, size=10, italic=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[28].height = 36

    # Reference as-of stamp
    ws.cell(row=30, column=2, value="Coverage Reference verified:").font = Font(
        name=FONT_BODY, size=11, bold=True, color=COLOR_TEXT
    )
    cell = ws.cell(row=30, column=3, value=_parse_date(REFERENCE_AS_OF))
    apply_style(cell, input_cell_style())
    cell.number_format = "yyyy-mm-dd"
    cell.font = Font(name=FONT_BODY, size=11, color=COLOR_PRIMARY)
    ws.row_dimensions[30].height = 24


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(out_path, variant):
    wb = Workbook()
    build_start_tab(wb, variant)
    build_claims_register_tab(wb, variant)
    build_claim_packet_tab(wb, variant)
    build_coverage_reference_tab(wb, variant)
    build_message_scripts_tab(wb, variant)
    build_annual_damage_pl_tab(wb, variant)
    build_settings_tab(wb, variant)
    # Apply Settings layout fixup (the unused build_settings_tab body left a
    # mostly-empty Settings sheet; the real layout is here so dropdown
    # references are correct.)
    _fix_settings_layout(wb["Settings"])

    # Filing-window cell reference fix — Claims Register formulas reference
    # Settings!$B$25 in the brief, but our Settings layout puts the value at C27.
    # Update Claims Register Days-to-deadline formulas to reference $C$27.
    ws_reg = wb["Claims Register"]
    for row in range(6, 206):
        ws_reg.cell(row=row, column=12).value = (
            f'=IF(AND(F{row}<>"",K{row}<>"Paid",K{row}<>"Closed unpaid"),'
            f'Settings!$C$27-(TODAY()-F{row}),"")'
        )

    # Update dropdown source ranges on Claims Register to match Settings layout (rows 6-25 / 6-15)
    # Properties col B → Settings!$B$6:$B$25; Status col K → $C$6:$C$15; Category col H → $D$6:$D$15
    # Re-add data validations with corrected refs (the old ones still exist; openpyxl will keep both,
    # but Excel uses the most recent).
    add_dropdown(ws_reg, "B6:B205", "=Settings!$B$6:$B$25")
    add_dropdown(ws_reg, "H6:H205", "=Settings!$D$6:$D$15")
    add_dropdown(ws_reg, "K6:K205", "=Settings!$C$6:$C$15")

    # Also the Claim Packet category dropdown
    ws_pkt = wb["Claim Packet"]
    add_dropdown(ws_pkt, "C11", "=Settings!$D$6:$D$15")

    wb.properties.title = "Damage Claim + AirCover Log — The STR Ledger"
    wb.properties.creator = "The STR Ledger"
    wb.properties.company = "The STR Ledger"
    wb.properties.description = (
        "Damage claim management workspace for STR hosts — deadline countdown, "
        "evidence completeness, message scripts, AirCover/VRBO/Booking reference."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    build_workbook(DEMO_OUT, "demo")
    build_workbook(BLANK_OUT, "blank")


if __name__ == "__main__":
    main()
